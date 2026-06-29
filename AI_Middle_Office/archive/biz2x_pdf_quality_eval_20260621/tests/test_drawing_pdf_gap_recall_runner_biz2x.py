from __future__ import annotations

import json
import sys
from pathlib import Path

from openpyxl import load_workbook

from app.services.drawing_pdf_gap_recall_runner import (
    run_gap_recall_plan,
    write_gap_recall_run_outputs,
)
from scripts import biz2x_pdf_gap_recall_run


def _recall_plan(tmp_path: Path) -> dict[str, object]:
    image = tmp_path / "whole.png"
    image.write_bytes(b"fake image bytes")
    return {
        "plan_rows": [
            {
                "task_no": 1,
                "gap_no": 1,
                "gap_priority": "P1_missing_core",
                "answer_item_name": "拆除不锈钢玻璃门",
                "recommended_pass": "door_window_demolition",
                "source_file": "drawing.pdf",
                "page": 1,
                "tile_id": "p001_whole",
                "tile_type": "whole_page_preview",
                "image_path": str(image),
            },
            {
                "task_no": 2,
                "gap_no": 2,
                "gap_priority": "P1_missing_core",
                "answer_item_name": "拆除单开实木门",
                "recommended_pass": "door_window_demolition",
                "source_file": "drawing.pdf",
                "page": 1,
                "tile_id": "p001_whole",
                "tile_type": "whole_page_preview",
                "image_path": str(image),
            },
            {
                "task_no": 3,
                "gap_no": 3,
                "gap_priority": "P2_missing_mep",
                "answer_item_name": "水表供货及安装",
                "recommended_pass": "fixture_valve_schedule",
                "source_file": "drawing.pdf",
                "page": 1,
                "tile_id": "p001_whole",
                "tile_type": "whole_page_preview",
                "image_path": str(image),
            },
        ]
    }


def test_gap_recall_runner_dry_run_dedupes_visual_calls(tmp_path: Path):
    report = run_gap_recall_plan(_recall_plan(tmp_path), execute=False)

    assert report["summary"]["execute"] is False
    assert report["summary"]["plan_task_count"] == 3
    assert report["summary"]["unique_visual_call_count"] == 2
    assert report["summary"]["covered_gap_count"] == 3
    assert report["summary"]["status_counts"] == {"planned_dry_run": 2}
    door_call = next(row for row in report["call_rows"] if row["recommended_pass"] == "door_window_demolition")
    assert door_call["covered_task_nos"] == "1；2"
    assert door_call["covered_gap_nos"] == "1；2"
    assert "拆除不锈钢玻璃门" in door_call["covered_answer_items"]
    assert report["evidence_rows"] == []


def test_gap_recall_runner_execute_uses_injected_vision_client(tmp_path: Path):
    calls = []

    async def fake_vision_client(base64_image, mime_type, **kwargs):
        calls.append({"base64_image": base64_image, "mime_type": mime_type, **kwargs})
        return {
            "raw_content": "{}",
            "evidence_items": [
                {
                    "evidence_role": "construction_note",
                    "discipline": "decoration",
                    "text": "拆除不锈钢玻璃门",
                    "item_hint": "拆除不锈钢玻璃门",
                    "spec_or_method": "含门套、门扇及五金",
                    "suggested_unit": "套",
                    "confidence": 0.86,
                    "needs_manual_review": True,
                    "reason": "门窗拆除说明可见",
                }
            ],
        }

    report = run_gap_recall_plan(_recall_plan(tmp_path), execute=True, max_calls=1, vision_client=fake_vision_client)

    assert len(calls) == 1
    assert calls[0]["prompt_mode"] == "door_window_demolition"
    assert calls[0]["tile_context"]["gap_nos"] == [1, 2]
    assert report["summary"]["status_counts"] == {"success": 1}
    assert report["summary"]["evidence_count"] == 1
    row = report["evidence_rows"][0]
    assert row["source_kind"] == "pdf_gap_recall_llm"
    assert row["vision_pass"] == "door_window_demolition"
    assert row["item_hint"] == "拆除不锈钢玻璃门"


def test_gap_recall_runner_writes_outputs(tmp_path: Path):
    report = run_gap_recall_plan(_recall_plan(tmp_path), execute=False)

    outputs = write_gap_recall_run_outputs(report, tmp_path / "out", stem="run")

    assert Path(outputs["json"]).exists()
    assert Path(outputs["call_csv"]).exists()
    assert Path(outputs["evidence_csv"]).exists()
    assert Path(outputs["markdown"]).exists()
    assert Path(outputs["xlsx"]).exists()
    workbook = load_workbook(outputs["xlsx"])
    assert workbook.sheetnames == ["run_summary", "visual_calls", "recall_evidence"]
    assert workbook["visual_calls"]["A2"].value == 1


def test_gap_recall_runner_cli_dry_run(tmp_path: Path, monkeypatch, capsys):
    recall_plan_json = tmp_path / "recall_plan.json"
    recall_plan_json.write_text(json.dumps(_recall_plan(tmp_path), ensure_ascii=False), encoding="utf-8")
    output_dir = tmp_path / "cli_out"
    monkeypatch.setattr(
        sys,
        "argv",
        [
            "biz2x_pdf_gap_recall_run.py",
            "--recall-plan-json",
            str(recall_plan_json),
            "--output-dir",
            str(output_dir),
            "--max-calls",
            "1",
            "--timestamp",
            "fixed",
        ],
    )

    assert biz2x_pdf_gap_recall_run.main() == 0
    payload = json.loads(capsys.readouterr().out)

    assert payload["ok"] is True
    assert payload["summary"]["execute"] is False
    assert payload["summary"]["unique_visual_call_count"] == 1
    assert Path(payload["outputs"]["xlsx"]).exists()
