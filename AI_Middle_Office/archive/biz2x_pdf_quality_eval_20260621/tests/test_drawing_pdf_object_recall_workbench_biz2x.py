from __future__ import annotations

import json
import sys
from pathlib import Path

from openpyxl import load_workbook

from app.services.drawing_pdf_external_recall_template_status import build_external_recall_template_status
from app.services.drawing_pdf_gap_recall_importer import (
    build_gap_recall_external_import_report,
    load_external_recall_results,
)
from app.services.drawing_pdf_object_recall_workbench import (
    build_object_recall_workbench,
    write_object_recall_workbench_outputs,
)
from scripts import biz2x_pdf_object_recall_workbench


def test_object_recall_workbench_links_plan_images_without_answer_leakage(tmp_path: Path):
    image = tmp_path / "page.png"
    image.write_bytes(b"fake")
    report = build_object_recall_workbench(
        _object_pack(),
        recall_plans=[_recall_plan(image)],
    )

    assert report["summary"]["object_recall_task_count"] == 1
    assert report["summary"]["image_link_count"] == 1
    assert report["summary"]["importable_row_count"] == 0
    assert report["summary"]["answer_only_count"] == 1
    assert report["summary"]["answer_columns_count_as_evidence"] is False
    row = report["workbench_rows"][0]
    assert row["image_path"] == str(image)
    assert row["image_exists"] is True
    assert row["image_source"] == "recall_plan_target"
    assert row["fill_status"] == "answer_only_reference"
    assert row["ready_for_import"] == "false"


def test_object_recall_workbench_uses_image_root_evidence_id(tmp_path: Path):
    image_root = tmp_path / "images"
    image_root.mkdir()
    image = image_root / "0001_target_R03-PDFEV-000013.png"
    image.write_bytes(b"fake")

    report = build_object_recall_workbench(_object_pack(), image_roots=[image_root])

    row = report["workbench_rows"][0]
    assert row["image_path"] == str(image)
    assert row["image_source"] == "image_root_evidence_id"
    assert report["summary"]["image_link_count"] == 1


def test_object_recall_workbench_falls_back_to_source_page_whole_image(tmp_path: Path):
    image_root = tmp_path / "images"
    image_root.mkdir()
    image = image_root / "03.整理完毕1017信达资产职工餐厅施工图_cad2pdf_1-1.png"
    image.write_bytes(b"fake")
    pack = _object_pack()
    pack["recall_rows"][0]["candidate_source_files"] = (
        "03.整理完毕1017信达资产职工餐厅施工图(cad2pdf)(1).pdf；"
        "04.信达资产职工餐厅水电施工图(cad2pdf).pdf"
    )
    pack["recall_rows"][0]["current_candidate_evidence_ids"] = ""

    report = build_object_recall_workbench(pack, image_roots=[image_root])

    row = report["workbench_rows"][0]
    assert row["image_path"] == str(image)
    assert row["image_source"] == "image_root_source_page_whole"
    assert report["summary"]["image_link_count"] == 1


def test_object_recall_workbench_preserves_filled_evidence_status(tmp_path: Path):
    pack = _object_pack()
    pack["recall_rows"][0]["evidence_item_hint"] = "stainless glass door demolition"
    pack["recall_rows"][0]["evidence_spec_or_method"] = "remove frame leaf hardware"
    pack["recall_rows"][0]["evidence_suggested_unit"] = "set"
    pack["recall_rows"][0]["evidence_text"] = "door schedule note"

    report = build_object_recall_workbench(pack)

    assert report["summary"]["importable_row_count"] == 1
    assert report["workbench_rows"][0]["fill_status"] == "importable"
    assert report["workbench_rows"][0]["ready_for_import"] == "true"


def test_object_recall_workbench_uses_recommended_pass_fallback_image(tmp_path: Path):
    image = tmp_path / "finish_schedule.png"
    image.write_bytes(b"fake")
    pack = _object_pack()
    row = pack["recall_rows"][0]
    row["recommended_pass"] = "finish_schedule"
    row["object_class"] = "finish_wall"
    row["candidate_source_files"] = ""
    row["evidence_pages"] = ""
    row["evidence_tiles"] = ""
    row["current_candidate_evidence_ids"] = ""

    report = build_object_recall_workbench(
        pack,
        fallback_images={"finish_schedule": image},
    )

    workbench_row = report["workbench_rows"][0]
    assert workbench_row["image_path"] == str(image)
    assert workbench_row["image_exists"] is True
    assert workbench_row["image_source"] == "fallback_image:finish_schedule"
    assert workbench_row["source_file"] == str(image)
    assert workbench_row["page"] == "fallback"
    assert workbench_row["tile_id"] == "fallback_image:finish_schedule"
    assert workbench_row["fill_status"] == "answer_only_reference"
    assert workbench_row["ready_for_import"] == "false"
    assert report["summary"]["image_link_count"] == 1
    assert report["summary"]["importable_row_count"] == 0


def test_object_recall_workbench_uses_task_specific_image_before_fallback(tmp_path: Path):
    task_image = tmp_path / "task_1_zoom.png"
    fallback_image = tmp_path / "finish_schedule.png"
    task_image.write_bytes(b"task image")
    fallback_image.write_bytes(b"fallback image")
    pack = _object_pack()
    row = pack["recall_rows"][0]
    row["recommended_pass"] = "finish_schedule"
    row["object_class"] = "finish_wall"
    row["candidate_source_files"] = ""
    row["evidence_pages"] = ""
    row["evidence_tiles"] = ""
    row["current_candidate_evidence_ids"] = ""

    report = build_object_recall_workbench(
        pack,
        fallback_images={"finish_schedule": fallback_image},
        task_images={"1": task_image},
    )

    workbench_row = report["workbench_rows"][0]
    assert workbench_row["image_path"] == str(task_image)
    assert workbench_row["image_exists"] is True
    assert workbench_row["image_source"] == "task_image:1"
    assert workbench_row["source_file"] == str(task_image)
    assert workbench_row["page"] == "task_image"
    assert workbench_row["tile_id"] == "task_image:1"
    assert workbench_row["fill_status"] == "answer_only_reference"
    assert report["summary"]["image_source_counts"] == {"task_image:1": 1}


def test_object_recall_workbench_writes_outputs(tmp_path: Path):
    image = tmp_path / "page.png"
    image.write_bytes(b"fake")
    report = build_object_recall_workbench(_object_pack(), recall_plans=[_recall_plan(image)])

    outputs = write_object_recall_workbench_outputs(report, tmp_path / "out", stem="workbench")

    assert Path(outputs["json"]).exists()
    assert Path(outputs["csv"]).exists()
    assert Path(outputs["markdown"]).exists()
    workbook = load_workbook(outputs["xlsx"])
    assert workbook.sheetnames == ["summary", "readme", "object_recall_workbench", "class_summary", "pass_summary", "source_page_summary"]
    sheet = workbook["object_recall_workbench"]
    assert sheet["H2"].value == "open_image"
    assert sheet["H2"].hyperlink.target == str(image)
    assert sheet["Z2"].value == "answer_only_reference"


def test_filled_object_recall_workbench_can_be_imported_as_external_results(tmp_path: Path):
    image = tmp_path / "page.png"
    image.write_bytes(b"fake")
    report = build_object_recall_workbench(_object_pack(), recall_plans=[_recall_plan(image)])
    outputs = write_object_recall_workbench_outputs(report, tmp_path / "out", stem="workbench")
    workbook = load_workbook(outputs["xlsx"])
    sheet = workbook["object_recall_workbench"]
    headers = {cell.value: cell.column for cell in sheet[1]}
    sheet.cell(row=2, column=headers["evidence_item_hint"]).value = "stainless glass door demolition"
    sheet.cell(row=2, column=headers["evidence_spec_or_method"]).value = "remove frame leaf hardware"
    sheet.cell(row=2, column=headers["evidence_suggested_unit"]).value = "set"
    sheet.cell(row=2, column=headers["evidence_text"]).value = "door schedule note"
    workbook.save(outputs["xlsx"])
    workbook.close()

    external_results = load_external_recall_results(outputs["xlsx"])
    status = build_external_recall_template_status(external_results)
    import_report = build_gap_recall_external_import_report(
        external_results,
        recall_plan=_recall_plan(image),
        source_name="filled-workbench",
    )

    assert status["summary"]["importable_row_count"] == 1
    assert status["summary"]["ready_for_external_import"] is True
    assert import_report["summary"]["evidence_count"] == 1
    assert import_report["summary"]["validation_status_counts"] == {"imported": 1}
    row = import_report["evidence_rows"][0]
    assert row["source_file"] == "03.pdf"
    assert row["page"] == "1"
    assert row["tile_id"] == "p001_whole"
    assert row["vision_pass"] == "door_window_demolition"
    assert row["item_hint"] == "stainless glass door demolition"
    assert row["suggested_unit"] == "set"


def test_object_recall_workbench_cli(tmp_path: Path, monkeypatch, capsys):
    image = tmp_path / "page.png"
    image.write_bytes(b"fake")
    object_json = tmp_path / "object.json"
    plan_json = tmp_path / "plan.json"
    object_json.write_text(json.dumps(_object_pack(), ensure_ascii=False), encoding="utf-8")
    plan_json.write_text(json.dumps(_recall_plan(image), ensure_ascii=False), encoding="utf-8")
    output_dir = tmp_path / "cli_out"
    monkeypatch.setattr(
        sys,
        "argv",
        [
            "biz2x_pdf_object_recall_workbench.py",
            "--object-recall-json",
            str(object_json),
            "--recall-plan-json",
            str(plan_json),
            "--output-dir",
            str(output_dir),
            "--timestamp",
            "fixed",
        ],
    )

    assert biz2x_pdf_object_recall_workbench.main() == 0
    payload = json.loads(capsys.readouterr().out)

    assert payload["ok"] is True
    assert payload["summary"]["image_link_count"] == 1
    assert Path(payload["outputs"]["xlsx"]).exists()


def _object_pack() -> dict[str, object]:
    return {
        "summary": {"object_recall_task_count": 1},
        "recall_rows": [
            {
                "task_no": 1,
                "gap_priority": "P1_object_missing",
                "object_class": "door_window_demolition",
                "recommended_pass": "door_window_demolition",
                "candidate_source_files": "03.pdf",
                "evidence_pages": "1",
                "evidence_tiles": "p001_whole",
                "target_item_name": "stainless glass door demolition",
                "target_feature": "remove stainless glass door",
                "target_unit": "set",
                "target_object_terms": "stainless glass door",
                "required_evidence_keywords": "door; demolition",
                "review_instruction": "find real drawing evidence",
                "current_candidate_item_name": "floor demolition",
                "current_candidate_unit": "m2",
                "current_candidate_evidence_ids": "R03-PDFEV-000013",
                "evidence_item_hint": "",
                "evidence_spec_or_method": "",
                "evidence_suggested_unit": "",
                "evidence_text": "",
            }
        ],
    }


def _recall_plan(image: Path) -> dict[str, object]:
    return {
        "plan_rows": [
            {
                "task_no": 1,
                "gap_no": 1,
                "answer_item_name": "stainless glass door demolition",
                "recommended_pass": "door_window_demolition",
                "source_file": "03.pdf",
                "page": 1,
                "tile_id": "p001_whole",
                "tile_type": "whole_page_preview",
                "evidence_id": "R03-PDFEV-000013",
                "image_path": str(image),
            }
        ]
    }
