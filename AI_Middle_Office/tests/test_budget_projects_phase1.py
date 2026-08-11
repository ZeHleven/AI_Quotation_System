from __future__ import annotations

import uuid
from io import BytesIO

import pytest
import xlwt
from openpyxl import Workbook
from sqlalchemy import event

from app.core.config import settings
from app.core.database import SessionLocal, engine
from app.core.security import get_password_hash
from app.models.budget_project import (
    BUDGET_IMPORT_STATUS_ACTIVE,
    BUDGET_IMPORT_STATUS_SUPERSEDED,
    BudgetProjectImportBatch,
    BudgetProjectImportLifecycleEvent,
    BudgetProjectImportRevision,
    BudgetProjectProfile,
)
from app.models.project_progress import Project
from app.models.user import User, UserRole
from app.services.budget_projects import (
    MAX_IMPORT_BYTES,
    _apply_workbook_semantics,
    _guarded_sequence_columns,
    _workbook_summary_multipliers,
)
from app.services.rbac import get_available_modules
from app.services.requirement_standardizer import standardize_requirement_excel_bytes


PASSWORD = "secret123"
XLSX_CONTENT_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
XLS_CONTENT_TYPE = "application/vnd.ms-excel"


def _set_flag(name: str, value: bool) -> bool:
    old_value = getattr(settings, name)
    object.__setattr__(settings, name, value)
    return old_value


def _create_user(*, role: str = "user", roles: list[str] | None = None) -> User:
    db = SessionLocal()
    try:
        user = User(
            username=f"budget_p1_{uuid.uuid4().hex[:12]}",
            hashed_password=get_password_hash(PASSWORD),
            role=role,
            role_version=1,
            quota=20,
            is_active=True,
        )
        db.add(user)
        db.flush()
        for assigned_role in roles or []:
            db.add(UserRole(user_id=user.id, role=assigned_role, created_by=None, note="budget phase 1 test"))
        db.commit()
        db.refresh(user)
        return user
    finally:
        db.close()


def _login(client, user: User) -> dict[str, str]:
    response = client.post("/api/v1/auth/login", data={"username": user.username, "password": PASSWORD})
    assert response.status_code == 200, response.text
    return {"Authorization": f"Bearer {response.json()['access_token']}"}


def _replace_roles(user_id: int, roles: list[str], *, legacy_role: str = "none") -> User:
    db = SessionLocal()
    try:
        user = db.query(User).filter(User.id == user_id).one()
        db.query(UserRole).filter(UserRole.user_id == user_id).delete(synchronize_session=False)
        user.role = legacy_role
        user.role_version = int(user.role_version or 1) + 1
        db.flush()
        for assigned_role in roles:
            db.add(UserRole(user_id=user_id, role=assigned_role, created_by=None, note="budget role replacement"))
        db.commit()
        db.refresh(user)
        return user
    finally:
        db.close()


def _workbook_bytes(rows: list[list[object]], *, title: str = "清单") -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = title
    for row in rows:
        sheet.append(row)
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def _legacy_workbook_bytes(rows: list[list[object]], *, title: str = "清单") -> bytes:
    workbook = xlwt.Workbook()
    sheet = workbook.add_sheet(title)
    for row_index, row in enumerate(rows):
        for column_index, value in enumerate(row):
            sheet.write(row_index, column_index, value)
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def _quantity_cases_workbook() -> bytes:
    return _workbook_bytes(
        [
            ["序号", "项目名称", "规格", "工程量", "单位"],
            [1, "项目一", "规格A", 1, "m"],
            [2, "项目二", "规格B", 2, "m"],
            [3, "项目三", "规格C", None, "m"],
            [4, "项目四", "规格D", 0, "m"],
            [5, "项目五", "规格E", "abc", "m"],
            [6, "项目六", "规格F", "1-2", "m"],
            [7, "项目七", "规格G", -1, "m"],
            [8, "项目八", "规格H", "999999999999999999999999999999", "m"],
            [9, "项目九", "规格I", "0.12345678901234567890123456789012345", "m"],
            [10, "项目十", "规格J", "0.0000004", "m"],
        ]
    )


def _sequence_only_workbook() -> bytes:
    return _workbook_bytes(
        [
            ["序号", "项目名称", "单位"],
            [1, "项目一", "m"],
            [2, "项目二", "m"],
            [3, "项目三", "m"],
        ]
    )


def _headerless_sequence_workbook() -> bytes:
    return _workbook_bytes(
        [
            ["章节说明"],
            ["1.", "拆除", "m"],
            ["2、", "砌筑", "m"],
            ["3.", "涂料", "m"],
        ],
        title="无表头",
    )


def _headerless_dual_sequence_workbook() -> bytes:
    return _workbook_bytes(
        [
            ["1.", "项目一", 1, "m"],
            ["2.", "项目二", 2, "m"],
            ["3.", "项目三", 3, "m"],
        ],
        title="双序列",
    )


def _repeated_header_and_raw_unit_workbook() -> bytes:
    return _workbook_bytes(
        [
            ["序号", "项目名称", "项目特征", "工程量", "单位"],
            ["序号", "项目名称", "项目特征", "工程量", "单位"],
            [1, "土方开挖", "人工开挖", 5, "m³"],
            [2, "土方回填", "分层夯实", 3, "m³"],
        ],
        title="装修",
    )


def _workbook_with_legacy_price_column() -> bytes:
    return _workbook_bytes(
        [
            ["序号", "项目名称", "项目特征", "工程量", "单位", "综合单价"],
            [1, "墙面涂料", "两遍乳胶漆", 12, "m²", 88.5],
        ],
        title="装修工程量清单",
    )


def _formula_budget_model_workbook() -> bytes:
    workbook = Workbook()
    workbook.remove(workbook.active)

    note = workbook.create_sheet("编制说明")
    note.append(["说明", "仅作为报价说明"])

    rules = workbook.create_sheet("计算规则")
    rules.append(["精装修计算规则"])
    rules.append(["扣除门窗洞口面积，不作为报价明细"])

    summary = workbook.create_sheet("汇总表")
    summary.append([""])
    summary.append(["序号", "栋号", "户型", "户数（套）", "面积", "面积合计", "安装部分（元）"])
    summary.append([1, "5栋", "A1", 44, 80, "=D3*E3", "='安装部分清单（户内）'!F6"])

    install = workbook.create_sheet("安装部分清单（户内）")
    install.append(["安装部分清单报价表"])
    install.append(["工程名称：测试项目"])
    install.append(["序号", "分部分项名称", "项目特征/做法要求", "单位", "工程量", "合计", "综合单价"])
    install.append([1, "PPR给水管 DN15", "管道安装", "m", 2, "=G4*E4", 0])
    install.append([2, "LED灯盘", "灯具安装", "套", 3, "=G5*E5", 0])
    install.append(["", "A1户型小计", "", "", "", "=SUM(F4:F5)", ""])

    brand = workbook.create_sheet("主材品牌表")
    brand.append(["材料名称", "品牌"])
    brand.append(["PPR管", "品牌A"])

    loss = workbook.create_sheet("瓷砖损耗")
    loss.append(["区域", "损耗率"])
    loss.append(["墙面", "3%"])

    backup = workbook.create_sheet("备用清单")
    backup.append(["说明"])
    backup.append(["可选项，默认不进入报价"])

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def _create_project(client, headers: dict[str, str], name: str = "预算项目一期") -> dict:
    response = client.post(
        "/api/v1/admin/budget-projects",
        headers=headers,
        json={"name": name, "client_name": "测试客户", "owner_department": "成本部"},
    )
    assert response.status_code == 200, response.text
    return response.json()["data"]


def _upload(client, headers: dict[str, str], project_id: int, content: bytes, filename: str = "清单.xlsx"):
    return client.post(
        f"/api/v1/admin/budget-projects/{project_id}/imports",
        headers=headers,
        files={"file": (filename, content, XLSX_CONTENT_TYPE)},
    )


def test_budget_project_feature_flag_and_role_module(client):
    quote_user = _create_user(roles=["quote_user"])
    headers = _login(client, quote_user)
    old_flag = _set_flag("feature_budget_projects", False)
    try:
        response = client.get("/api/v1/admin/budget-projects", headers=headers)
        db = SessionLocal()
        try:
            attached_user = db.query(User).filter(User.id == quote_user.id).one()
            module = next(item for item in get_available_modules(attached_user) if item["key"] == "budget_projects")
        finally:
            db.close()
    finally:
        _set_flag("feature_budget_projects", old_flag)
    assert response.status_code == 404
    assert response.json()["detail"] == "NOT_FOUND"
    assert module["path"] == "/admin/budget-projects"
    assert module["status"] == "pending"


def test_budget_project_upload_accepts_legacy_xls(client):
    quote_user = _create_user(roles=["quote_user"])
    headers = _login(client, quote_user)
    old_flag = _set_flag("feature_budget_projects", True)
    try:
        project = _create_project(client, headers)
        content = _legacy_workbook_bytes(
            [
                ["序号", "项目名称", "项目特征", "工程量", "单位"],
                [1, "一楼墙面拆除", "含垃圾清运", 18, "㎡"],
                [2, "二楼墙面拆除", "含垃圾清运", 20, "㎡"],
            ]
        )
        response = client.post(
            f"/api/v1/admin/budget-projects/{project['id']}/imports",
            headers=headers,
            files={"file": ("旧版清单.xls", content, XLS_CONTENT_TYPE)},
        )
    finally:
        _set_flag("feature_budget_projects", old_flag)

    assert response.status_code == 200, response.text
    data = response.json()["data"]
    assert data["source_file"]["filename"] == "旧版清单.xls"
    assert data["summary"]["standard_item_count"] == 2


def test_budget_quantity_safe_zero_keeps_items_and_does_not_compare_other_column_values(client):
    quote_user = _create_user(roles=["quote_user"])
    headers = _login(client, quote_user)
    old_flag = _set_flag("feature_budget_projects", True)
    try:
        project = _create_project(client, headers)
        uploaded = _upload(client, headers, project["id"], _quantity_cases_workbook())
        assert uploaded.status_code == 200, uploaded.text
        batch = uploaded.json()["data"]
        assert batch["summary"]["standard_item_count"] == 10
        assert batch["summary"]["valid_quantity_count"] == 2
        assert batch["summary"]["invalid_quantity_count"] == 8

        rows_response = client.get(
            f"/api/v1/admin/budget-projects/imports/{batch['batch_uuid']}/rows",
            headers=headers,
            params={"standard_items_only": True, "page_size": 100},
        )
        assert rows_response.status_code == 200, rows_response.text
        body = rows_response.json()
        assert body["total"] == 10
        assert body["page"] == 1
        assert body["page_size"] == 100
        rows = body["data"]
        assert [(row["raw_quantity"], row["quantity_status"], row["calculation_quantity"]) for row in rows] == [
            ("1", "valid", 1.0),
            ("2", "valid", 2.0),
            ("", "missing", 0.0),
            ("0", "zero", 0.0),
            ("abc", "non_numeric", 0.0),
            ("1-2", "abnormal", 0.0),
            ("-1", "abnormal", 0.0),
            ("999999999999999999999999999999", "abnormal", 0.0),
            ("0.12345678901234567890123456789012345", "unsupported_precision", 0.0),
            ("0.0000004", "precision_underflow", 0.0),
        ]
        # Sequence=quantity equality on another column is legitimate and must not
        # turn the mapped 工程量 into a sequence-column false positive.
        assert rows[0]["quantity_reason"] == "VALID_SOURCE_QUANTITY"
        assert rows[1]["quantity_reason"] == "VALID_SOURCE_QUANTITY"
        assert rows[2]["quantity_reason"] == "EMPTY_QUANTITY"
        assert rows[-1]["quantity_reason"] == "BELOW_SUPPORTED_PRECISION"
        assert all(row["item_name"] for row in rows)
    finally:
        _set_flag("feature_budget_projects", old_flag)


def test_budget_workbook_semantics_excludes_reference_sheets_and_applies_summary_multiplier():
    content = _formula_budget_model_workbook()
    preview = standardize_requirement_excel_bytes(content, filename="formula-budget.xlsx")
    parsed = _apply_workbook_semantics(preview, content)

    roles = {item["sheet_name"]: item["sheet_role"] for item in parsed["sheet_mappings"]}
    assert roles["编制说明"] == "metadata"
    assert roles["计算规则"] == "calculation_rule"
    assert roles["汇总表"] == "summary_analysis"
    assert roles["主材品牌表"] == "material_reference"
    assert roles["瓷砖损耗"] == "loss_reference"
    assert roles["备用清单"] == "optional_backup"

    reference_rows = [
        row
        for row in parsed["rows"]
        if row.get("source_sheet") in {"编制说明", "计算规则"} and row.get("row_type") == "reference_row"
    ]
    assert [(row["source_sheet"], row["raw_row_index"], row["item_name"]) for row in reference_rows] == [
        ("编制说明", 1, "说明 仅作为报价说明"),
        ("计算规则", 1, "精装修计算规则"),
        ("计算规则", 2, "扣除门窗洞口面积，不作为报价明细"),
    ]
    assert all(row.get("quantity") is None for row in reference_rows)

    install_mapping = next(
        item for item in parsed["sheet_mappings"] if item["sheet_name"] == "安装部分清单（户内）"
    )
    assert install_mapping["field_mapping"]["B"] == "item_name"
    assert install_mapping["field_mapping"]["C"] == "spec"
    assert install_mapping["field_mapping"]["D"] == "unit"
    assert install_mapping["field_mapping"]["E"] == "quantity"
    assert install_mapping["field_mapping"]["F"] == "ignore"
    assert install_mapping["field_mapping"]["G"] == "ignore"

    bill_rows = [
        row
        for row in parsed["rows"]
        if row.get("source_sheet") == "安装部分清单（户内）" and row.get("row_type") == "data_row"
    ]
    assert [(row["item_name"], row["quantity"], row.get("budget_summary_multiplier")) for row in bill_rows[:2]] == [
        ("PPR给水管 DN15", 2, "44"),
        ("LED灯盘", 3, "44"),
    ]


def test_summary_multiplier_accumulates_same_sheet_building_aliases():
    workbook = Workbook()
    summary = workbook.active
    summary.title = "\u6c47\u603b\u8868"
    detail = workbook.create_sheet("\u5b89\u88c5\u90e8\u5206\u6e05\u5355\uff08\u6237\u5185\uff09")
    summary.append(["\u5e8f\u53f7", "\u680b\u53f7", "\u6237\u578b", "\u6237\u6570", "\u5408\u4ef7"])
    summary.append([1, "5#", "A1", 44, f"='{detail.title}'!F4"])
    # The second building intentionally reuses the first building subtotal.
    summary.append([2, "6#", "A1", 44, "=E2"])
    summary.append(["", "\u5408\u8ba1", "", "=SUM(D2:D3)", "=SUM(E2:E3)"])
    detail.append(["\u5e8f\u53f7", "\u9879\u76ee\u540d\u79f0", "\u5355\u4f4d", "\u5de5\u7a0b\u91cf", "\u5355\u4ef7", "\u5408\u4ef7"])
    detail.append([1, "\u7ba1\u9053\u5b89\u88c5", "m", 2, 0, "=D2*E2"])
    detail.append([2, "\u706f\u5177\u5b89\u88c5", "\u5957", 3, 0, "=D3*E3"])
    detail.append(["", "A1\u6237\u578b\u5c0f\u8ba1", "", "", "", "=SUM(F2:F3)"])
    output = BytesIO()
    workbook.save(output)

    multipliers = _workbook_summary_multipliers(output.getvalue())

    assert str(multipliers[(detail.title, 2)]["multiplier"]) == "88"
    assert str(multipliers[(detail.title, 3)]["multiplier"]) == "88"
    source_values = {
        entry["value"]
        for source in multipliers[(detail.title, 2)]["sources"]
        for entry in source["summary_scope"]
        if entry["label"] in {"\u680b\u53f7", "\u6237\u6570"}
    }
    assert {"5#", "6#", "44"}.issubset(source_values)


def test_reference_sheet_rows_survive_manual_remap(client):
    quote_user = _create_user(roles=["quote_user"])
    headers = _login(client, quote_user)
    old_flag = _set_flag("feature_budget_projects", True)
    try:
        project = _create_project(client, headers, "参考信息重新映射")
        uploaded = _upload(client, headers, project["id"], _formula_budget_model_workbook())
        assert uploaded.status_code == 200, uploaded.text
        batch = uploaded.json()["data"]
        mappings = [
            {
                "sheet_name": mapping["sheet_name"],
                "field_mapping": mapping["applied_field_mapping"],
            }
            for mapping in batch["sheet_mappings"]
        ]

        remapped = client.post(
            f"/api/v1/admin/budget-projects/imports/{batch['batch_uuid']}/remap",
            headers=headers,
            json={
                "expected_remap_revision": batch["remap_revision"],
                "sheet_mappings": mappings,
            },
        )
        assert remapped.status_code == 200, remapped.text

        expected_rows = {"编制说明": 1, "计算规则": 2}
        for sheet_name, expected_count in expected_rows.items():
            rows_response = client.get(
                f"/api/v1/admin/budget-projects/imports/{batch['batch_uuid']}/rows",
                headers=headers,
                params={"source_sheet": sheet_name, "page_size": 100},
            )
            assert rows_response.status_code == 200, rows_response.text
            rows = rows_response.json()["data"]
            assert len(rows) == expected_count
            assert all(row["row_type"] == "reference_row" for row in rows)
            assert all(row["is_standard_item"] is False for row in rows)
            assert all(row["quantity_status"] == "not_applicable" for row in rows)

        bill_rows_response = client.get(
            f"/api/v1/admin/budget-projects/imports/{batch['batch_uuid']}/rows",
            headers=headers,
            params={"source_sheet": "安装部分清单（户内）", "standard_items_only": True, "page_size": 100},
        )
        assert bill_rows_response.status_code == 200, bill_rows_response.text
        bill_rows = bill_rows_response.json()["data"]
        assert {row["standard_row"].get("budget_summary_multiplier") for row in bill_rows} == {"44"}
    finally:
        _set_flag("feature_budget_projects", old_flag)


def test_sequence_column_and_headerless_numbering_cannot_be_remapped_to_quantity(client):
    quote_user = _create_user(roles=["quote_user"])
    headers = _login(client, quote_user)
    old_flag = _set_flag("feature_budget_projects", True)
    try:
        project = _create_project(client, headers, "序号防护")
        uploaded = _upload(client, headers, project["id"], _sequence_only_workbook())
        assert uploaded.status_code == 200, uploaded.text
        batch_uuid = uploaded.json()["data"]["batch_uuid"]
        remap = client.post(
            f"/api/v1/admin/budget-projects/imports/{batch_uuid}/remap",
            headers=headers,
            json={
                "sheet_mappings": [
                    {
                        "sheet_name": "清单",
                        "field_mapping": {"A": "quantity", "B": "item_name", "C": "unit"},
                    }
                ]
            },
        )
        assert remap.status_code == 422
        assert remap.json()["detail"]["code"] == "SEQUENCE_COLUMN_CANNOT_BE_QUANTITY"

        headerless_project = _create_project(client, headers, "无表头序号防护")
        headerless = _upload(client, headers, headerless_project["id"], _headerless_sequence_workbook())
        assert headerless.status_code == 200, headerless.text
        headerless_uuid = headerless.json()["data"]["batch_uuid"]
        headerless_remap = client.post(
            f"/api/v1/admin/budget-projects/imports/{headerless_uuid}/remap",
            headers=headers,
            json={
                "sheet_mappings": [
                    {
                        "sheet_name": "无表头",
                        "field_mapping": {"A": "quantity", "B": "item_name", "C": "unit"},
                    }
                ]
            },
        )
        assert headerless_remap.status_code == 422, headerless_remap.text
        assert headerless_remap.json()["detail"]["code"] == "SEQUENCE_COLUMN_CANNOT_BE_QUANTITY"

        dual_project = _create_project(client, headers, "双序列只保护最左列")
        dual = _upload(client, headers, dual_project["id"], _headerless_dual_sequence_workbook())
        assert dual.status_code == 200, dual.text
        dual_uuid = dual.json()["data"]["batch_uuid"]
        dual_remap = client.post(
            f"/api/v1/admin/budget-projects/imports/{dual_uuid}/remap",
            headers=headers,
            json={
                "sheet_mappings": [
                    {
                        "sheet_name": "双序列",
                        "field_mapping": {"A": "ignore", "B": "item_name", "C": "quantity", "D": "unit"},
                    }
                ]
            },
        )
        assert dual_remap.status_code == 200, dual_remap.text
        dual_rows = client.get(
            f"/api/v1/admin/budget-projects/imports/{dual_uuid}/rows",
            headers=headers,
            params={"standard_items_only": True, "page_size": 100},
        )
        assert dual_rows.status_code == 200
        assert [row["quantity_status"] for row in dual_rows.json()["data"]] == ["valid", "valid", "valid"]
        assert [row["calculation_quantity"] for row in dual_rows.json()["data"]] == [1.0, 2.0, 3.0]
    finally:
        _set_flag("feature_budget_projects", old_flag)


def test_headerless_sequence_guard_ignores_section_rows_and_accepts_number_punctuation():
    preview = standardize_requirement_excel_bytes(_headerless_sequence_workbook(), filename="headerless.xlsx")
    assert "A" in _guarded_sequence_columns(preview, "无表头")


def test_repeated_header_is_traceable_but_not_counted_and_raw_mapped_unit_is_recovered(client):
    quote_user = _create_user(roles=["quote_user"])
    headers = _login(client, quote_user)
    old_flag = _set_flag("feature_budget_projects", True)
    try:
        project = _create_project(client, headers, "重复表头与单位回填")
        uploaded = _upload(client, headers, project["id"], _repeated_header_and_raw_unit_workbook())
        assert uploaded.status_code == 200, uploaded.text
        batch = uploaded.json()["data"]
        assert batch["summary"]["standard_item_count"] == 2
        assert batch["summary"]["valid_quantity_count"] == 2

        rows_response = client.get(
            f"/api/v1/admin/budget-projects/imports/{batch['batch_uuid']}/rows",
            headers=headers,
            params={"page_size": 100},
        )
        assert rows_response.status_code == 200, rows_response.text
        rows = rows_response.json()["data"]
        repeated = next(row for row in rows if row["raw_row_index"] == 2)
        assert repeated["row_type"] == "repeated_header"
        assert repeated["is_standard_item"] is False
        assert repeated["item_name"] == "项目名称"
        assert "BUDGET_REPEATED_HEADER_ROW" in repeated["warnings"]

        standard_rows = [row for row in rows if row["is_standard_item"]]
        assert [row["raw_row_index"] for row in standard_rows] == [3, 4]
        assert [row["unit"] for row in standard_rows] == ["m³", "m³"]
        assert all("BUDGET_UNIT_RECOVERED_FROM_MAPPED_CELL" in row["warnings"] for row in standard_rows)
        assert all(row["standard_row"]["budget_unit_source"] == "mapped_raw_cell" for row in standard_rows)
    finally:
        _set_flag("feature_budget_projects", old_flag)


def test_applied_mapping_round_trips_with_legacy_price_column_locked_to_ignore(client):
    quote_user = _create_user(roles=["quote_user"])
    headers = _login(client, quote_user)
    old_flag = _set_flag("feature_budget_projects", True)
    try:
        project = _create_project(client, headers, "价格列映射往返")
        uploaded = _upload(client, headers, project["id"], _workbook_with_legacy_price_column())
        assert uploaded.status_code == 200, uploaded.text
        batch = uploaded.json()["data"]
        mapping = batch["sheet_mappings"][0]
        assert mapping["sheet_name"] == "装修工程量清单"
        assert mapping["applied_field_mapping"]["F"] == "ignore"
        price_column = next(item for item in mapping["current_columns"] if item["column"] == "F")
        assert price_column["is_price"] is True
        assert price_column["locked_ignore"] is True

        round_trip = client.post(
            f"/api/v1/admin/budget-projects/imports/{batch['batch_uuid']}/remap",
            headers=headers,
            json={
                "sheet_mappings": [
                    {
                        "sheet_name": mapping["sheet_name"],
                        "field_mapping": mapping["applied_field_mapping"],
                    }
                ]
            },
        )
        assert round_trip.status_code == 200, round_trip.text
        round_trip_mapping = round_trip.json()["data"]["sheet_mappings"][0]
        assert round_trip_mapping["applied_field_mapping"]["F"] == "ignore"
        assert round_trip.json()["data"]["summary"]["standard_item_count"] == 1

        forbidden_mapping = dict(round_trip_mapping["applied_field_mapping"])
        forbidden_mapping["F"] = "quantity"
        locked = client.post(
            f"/api/v1/admin/budget-projects/imports/{batch['batch_uuid']}/remap",
            headers=headers,
            json={
                "sheet_mappings": [
                    {
                        "sheet_name": mapping["sheet_name"],
                        "field_mapping": forbidden_mapping,
                    }
                ]
            },
        )
        assert locked.status_code == 422
        assert locked.json()["detail"]["code"] == "BUDGET_LOCKED_IGNORE_COLUMN"
        assert locked.json()["detail"]["columns"] == ["F"]
    finally:
        _set_flag("feature_budget_projects", old_flag)


def test_duplicate_uploads_create_independent_batches_and_never_query_frozen_cost_sources(client):
    quote_user = _create_user(roles=["quote_user"])
    headers = _login(client, quote_user)
    old_flag = _set_flag("feature_budget_projects", True)
    statements: list[str] = []

    def capture_statement(_conn, _cursor, statement, _parameters, _context, _executemany):
        statements.append(str(statement).lower())

    try:
        project = _create_project(client, headers, "重复上传")
        content = _quantity_cases_workbook()
        event.listen(engine, "before_cursor_execute", capture_statement)
        try:
            first = _upload(client, headers, project["id"], content)
            second = _upload(client, headers, project["id"], content)
        finally:
            event.remove(engine, "before_cursor_execute", capture_statement)
        assert first.status_code == 200, first.text
        assert second.status_code == 200, second.text
        first_batch = first.json()["data"]
        second_batch = second.json()["data"]
        assert first_batch["batch_uuid"] != second_batch["batch_uuid"]
        assert first_batch["source_file"]["sha256"] == second_batch["source_file"]["sha256"]
        assert first_batch["source_file"]["storage_mode"] == "metadata_only"
        assert first_batch["source_file"]["download_available"] is False

        imports = client.get(
            f"/api/v1/admin/budget-projects/{project['id']}/imports",
            headers=headers,
        )
        assert imports.status_code == 200
        assert imports.json()["total"] == 2
        forbidden_tables = (
            "enterprise_quota",
            "cost_items",
            "cost_measurements",
            "cost_measurement_lines",
            "project_cost_import",
        )
        captured_sql = "\n".join(statements)
        assert all(table not in captured_sql for table in forbidden_tables)
    finally:
        _set_flag("feature_budget_projects", old_flag)


def test_duplicate_sheet_mapping_is_422_and_archived_batch_cannot_be_remapped(client):
    quote_user = _create_user(roles=["quote_user"])
    headers = _login(client, quote_user)
    old_flag = _set_flag("feature_budget_projects", True)
    try:
        project = _create_project(client, headers, "归档与重复映射")
        uploaded = _upload(client, headers, project["id"], _quantity_cases_workbook())
        assert uploaded.status_code == 200
        batch_uuid = uploaded.json()["data"]["batch_uuid"]
        mapping = {"A": "ignore", "B": "item_name", "C": "spec", "D": "quantity", "E": "unit"}
        duplicate = client.post(
            f"/api/v1/admin/budget-projects/imports/{batch_uuid}/remap",
            headers=headers,
            json={
                "sheet_mappings": [
                    {"sheet_name": "清单", "field_mapping": mapping},
                    {"sheet_name": "清单", "field_mapping": mapping},
                ]
            },
        )
        assert duplicate.status_code == 422
        assert duplicate.json()["detail"]["code"] == "DUPLICATE_BUDGET_IMPORT_SHEET_MAPPING"

        unknown_column = client.post(
            f"/api/v1/admin/budget-projects/imports/{batch_uuid}/remap",
            headers=headers,
            json={"sheet_mappings": [{"sheet_name": "清单", "field_mapping": {"ZZZ": "quantity"}}]},
        )
        assert unknown_column.status_code == 422
        assert unknown_column.json()["detail"]["code"] == "UNKNOWN_BUDGET_MAPPING_COLUMN"

        unknown_field = client.post(
            f"/api/v1/admin/budget-projects/imports/{batch_uuid}/remap",
            headers=headers,
            json={"sheet_mappings": [{"sheet_name": "清单", "field_mapping": {"B": "total_price"}}]},
        )
        assert unknown_field.status_code == 422
        assert unknown_field.json()["detail"]["code"] == "UNKNOWN_BUDGET_MAPPING_FIELD"

        archived = client.patch(
            f"/api/v1/admin/budget-projects/{project['id']}/archive",
            headers=headers,
            json={"reason": "phase 1 test"},
        )
        assert archived.status_code == 200
        assert archived.json()["data"]["workspace_status"] == "archived"
        remap = client.post(
            f"/api/v1/admin/budget-projects/imports/{batch_uuid}/remap",
            headers=headers,
            json={"sheet_mappings": [{"sheet_name": "清单", "field_mapping": mapping}]},
        )
        assert remap.status_code == 409
        assert remap.json()["detail"] == "BUDGET_PROJECT_ARCHIVED"
        update = client.patch(
            f"/api/v1/admin/budget-projects/{project['id']}",
            headers=headers,
            json={"name": "归档后不允许更新"},
        )
        assert update.status_code == 409
        archived_upload = _upload(client, headers, project["id"], _quantity_cases_workbook())
        assert archived_upload.status_code == 409
    finally:
        _set_flag("feature_budget_projects", old_flag)


def test_import_rejects_empty_corrupt_oversized_and_sheet_dimension_limits(client):
    quote_user = _create_user(roles=["quote_user"])
    headers = _login(client, quote_user)
    old_flag = _set_flag("feature_budget_projects", True)
    try:
        project = _create_project(client, headers, "上传边界")
        empty = _upload(client, headers, project["id"], b"")
        assert empty.status_code == 422
        assert empty.json()["detail"] == "BUDGET_IMPORT_FILE_EMPTY"

        corrupt = _upload(client, headers, project["id"], b"not-an-xlsx")
        assert corrupt.status_code == 400

        oversized = _upload(client, headers, project["id"], b"x" * (MAX_IMPORT_BYTES + 1))
        assert oversized.status_code == 413
        assert oversized.json()["detail"] == "BUDGET_IMPORT_FILE_TOO_LARGE"

        formatted_wide = Workbook()
        formatted_sheet = formatted_wide.active
        formatted_sheet.title = "编制说明"
        formatted_sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=254)
        formatted_sheet["A1"] = "编制说明"
        formatted_sheet.append(["项目名称", "工程量", "单位"])
        formatted_sheet.append(["墙面抹灰", 12, "㎡"])
        formatted_buffer = BytesIO()
        formatted_wide.save(formatted_buffer)
        formatted_upload = _upload(client, headers, project["id"], formatted_buffer.getvalue())
        assert formatted_upload.status_code == 200, formatted_upload.text

        too_many_rows = _workbook_bytes(
            [["项目名称", "工程量", "单位"]]
            + [[f"项目{index}", index, "m"] for index in range(1, 801)],
            title="801行",
        )
        row_limit = _upload(client, headers, project["id"], too_many_rows)
        assert row_limit.status_code == 422
        assert row_limit.json()["detail"]["code"] == "BUDGET_IMPORT_WORKBOOK_LIMIT_EXCEEDED"

        too_many_columns = _workbook_bytes([[f"列{index}" for index in range(1, 302)]], title="301列")
        column_limit = _upload(client, headers, project["id"], too_many_columns)
        assert column_limit.status_code == 422
        assert column_limit.json()["detail"]["code"] == "BUDGET_IMPORT_WORKBOOK_LIMIT_EXCEEDED"
        assert column_limit.json()["detail"]["max_columns_per_sheet"] == 300
    finally:
        _set_flag("feature_budget_projects", old_flag)


def test_budget_project_idor_and_pure_quote_scope_are_owner_only(client):
    owner = _create_user(roles=["quote_user"])
    outsider = _create_user(roles=["quote_user"])
    admin = _create_user(role="admin", roles=["admin"])
    owner_headers = _login(client, owner)
    outsider_headers = _login(client, outsider)
    admin_headers = _login(client, admin)
    old_flag = _set_flag("feature_budget_projects", True)
    try:
        project = _create_project(client, owner_headers, "IDOR保护")
        uploaded = _upload(client, owner_headers, project["id"], _quantity_cases_workbook())
        assert uploaded.status_code == 200
        batch_uuid = uploaded.json()["data"]["batch_uuid"]

        assert client.get(f"/api/v1/admin/budget-projects/{project['id']}", headers=outsider_headers).status_code == 404
        assert client.patch(
            f"/api/v1/admin/budget-projects/{project['id']}",
            headers=outsider_headers,
            json={"name": "越权"},
        ).status_code == 404
        assert _upload(client, outsider_headers, project["id"], _quantity_cases_workbook()).status_code == 404
        assert client.get(
            f"/api/v1/admin/budget-projects/imports/{batch_uuid}", headers=outsider_headers
        ).status_code == 404
        assert client.post(
            f"/api/v1/admin/budget-projects/imports/{batch_uuid}/remap",
            headers=outsider_headers,
            json={"sheet_mappings": [{"sheet_name": "清单", "field_mapping": {"B": "item_name"}}]},
        ).status_code == 404

        external = _create_project(client, admin_headers, "仅被指定为经理")
        db = SessionLocal()
        try:
            external_model = db.query(Project).filter(Project.id == external["id"]).one()
            external_model.project_manager_id = outsider.id
            db.commit()
        finally:
            db.close()
        own_list = client.get("/api/v1/admin/budget-projects", headers=outsider_headers)
        assert own_list.status_code == 200
        assert all(item["id"] != external["id"] for item in own_list.json()["data"])
        assert client.get(
            f"/api/v1/admin/budget-projects/{external['id']}", headers=outsider_headers
        ).status_code == 404
    finally:
        _set_flag("feature_budget_projects", old_flag)


def test_role_downgrade_blocks_all_budget_mutations_but_keeps_owned_read(client):
    owner = _create_user(roles=["quote_user"])
    owner_headers = _login(client, owner)
    old_flag = _set_flag("feature_budget_projects", True)
    try:
        project = _create_project(client, owner_headers, "降权写保护")
        uploaded = _upload(client, owner_headers, project["id"], _quantity_cases_workbook())
        assert uploaded.status_code == 200
        batch_uuid = uploaded.json()["data"]["batch_uuid"]
        downgraded = _replace_roles(owner.id, ["project_member"])
        downgraded_headers = _login(client, downgraded)

        assert client.get(
            f"/api/v1/admin/budget-projects/{project['id']}", headers=downgraded_headers
        ).status_code == 200
        assert client.patch(
            f"/api/v1/admin/budget-projects/{project['id']}",
            headers=downgraded_headers,
            json={"name": "不应成功"},
        ).status_code == 403
        assert client.patch(
            f"/api/v1/admin/budget-projects/{project['id']}/archive",
            headers=downgraded_headers,
            json={"reason": "不应成功"},
        ).status_code == 403
        assert _upload(
            client, downgraded_headers, project["id"], _quantity_cases_workbook()
        ).status_code == 403
        assert client.post(
            f"/api/v1/admin/budget-projects/imports/{batch_uuid}/remap",
            headers=downgraded_headers,
            json={
                "sheet_mappings": [
                    {
                        "sheet_name": "清单",
                        "field_mapping": {"A": "ignore", "B": "item_name", "C": "spec", "D": "quantity", "E": "unit"},
                    }
                ]
            },
        ).status_code == 403
    finally:
        _set_flag("feature_budget_projects", old_flag)


@pytest.mark.parametrize(
    ("role", "assigned_roles", "expected_list", "expected_create"),
    [
        ("user", ["quote_user"], 200, 200),
        ("user", ["project_manager"], 200, 200),
        ("user", ["project_member"], 200, 403),
        ("viewer", [], 200, 403),
        ("none", [], 403, 403),
    ],
)
def test_budget_project_role_matrix(client, role, assigned_roles, expected_list, expected_create):
    user = _create_user(role=role, roles=assigned_roles)
    headers = _login(client, user)
    old_flag = _set_flag("feature_budget_projects", True)
    try:
        listed = client.get("/api/v1/admin/budget-projects", headers=headers)
        created = client.post(
            "/api/v1/admin/budget-projects",
            headers=headers,
            json={"name": f"role-{role}-{uuid.uuid4().hex[:6]}"},
        )
    finally:
        _set_flag("feature_budget_projects", old_flag)
    assert listed.status_code == expected_list, listed.text
    assert created.status_code == expected_create, created.text


def test_budget_import_revision_lifecycle_active_pointer_and_optimistic_remap(client):
    owner = _create_user(roles=["quote_user"])
    headers = _login(client, owner)
    old_flag = _set_flag("feature_budget_projects", True)
    try:
        project = _create_project(client, headers, "导入修订与启用状态机")
        first_upload = _upload(client, headers, project["id"], _quantity_cases_workbook())
        assert first_upload.status_code == 200, first_upload.text
        first = first_upload.json()["data"]
        assert first["status"] == "parsed"
        assert first["remap_revision"] == 0
        assert first["revision_count"] == 1
        assert first["current_revision_id"]
        assert first["confirmed_revision_id"] is None
        assert first["capabilities"] == {
            "can_remap": True,
            "can_confirm": True,
            "can_activate": False,
        }

        initial_revision = client.get(
            f"/api/v1/admin/budget-projects/imports/{first['batch_uuid']}/revisions/{first['current_revision_id']}",
            headers=headers,
        )
        assert initial_revision.status_code == 200, initial_revision.text
        initial_snapshot = initial_revision.json()["data"]
        assert initial_snapshot["revision_number"] == 0
        assert initial_snapshot["revision_kind"] == "initial"

        mapping = first["sheet_mappings"][0]["applied_field_mapping"]
        stale = client.post(
            f"/api/v1/admin/budget-projects/imports/{first['batch_uuid']}/remap",
            headers=headers,
            json={
                "expected_remap_revision": 99,
                "sheet_mappings": [{"sheet_name": "清单", "field_mapping": mapping}],
            },
        )
        assert stale.status_code == 409, stale.text
        assert stale.json()["detail"]["code"] == "BUDGET_IMPORT_REVISION_CONFLICT"

        remapped = client.post(
            f"/api/v1/admin/budget-projects/imports/{first['batch_uuid']}/remap",
            headers=headers,
            json={
                "expected_remap_revision": 0,
                "sheet_mappings": [{"sheet_name": "清单", "field_mapping": mapping}],
            },
        )
        assert remapped.status_code == 200, remapped.text
        remapped_data = remapped.json()["data"]
        assert remapped_data["remap_revision"] == 1
        assert remapped_data["revision_count"] == 2
        assert remapped_data["current_revision"]["revision_kind"] == "remap"

        unchanged_initial = client.get(
            f"/api/v1/admin/budget-projects/imports/{first['batch_uuid']}/revisions/{first['current_revision_id']}",
            headers=headers,
        )
        assert unchanged_initial.status_code == 200
        assert unchanged_initial.json()["data"] == initial_snapshot

        revisions = client.get(
            f"/api/v1/admin/budget-projects/imports/{first['batch_uuid']}/revisions",
            headers=headers,
        )
        assert revisions.status_code == 200
        assert revisions.json()["total"] == 2
        assert [item["revision_number"] for item in revisions.json()["data"]] == [1, 0]

        confirmed = client.post(
            f"/api/v1/admin/budget-projects/imports/{first['batch_uuid']}/confirm",
            headers=headers,
        )
        assert confirmed.status_code == 200, confirmed.text
        confirmed_data = confirmed.json()["data"]
        assert confirmed_data["status"] == "confirmed"
        assert confirmed_data["confirmed_revision_id"] == confirmed_data["current_revision_id"]
        assert confirmed_data["capabilities"] == {
            "can_remap": False,
            "can_confirm": False,
            "can_activate": True,
        }
        frozen = client.post(
            f"/api/v1/admin/budget-projects/imports/{first['batch_uuid']}/remap",
            headers=headers,
            json={
                "expected_remap_revision": 1,
                "sheet_mappings": [{"sheet_name": "清单", "field_mapping": mapping}],
            },
        )
        assert frozen.status_code == 409
        assert frozen.json()["detail"]["code"] == "BUDGET_IMPORT_REMAP_FROZEN"

        activated = client.post(
            f"/api/v1/admin/budget-projects/imports/{first['batch_uuid']}/activate",
            headers=headers,
        )
        assert activated.status_code == 200, activated.text
        active_first = activated.json()["data"]
        assert active_first["status"] == "active"
        assert active_first["is_active"] is True
        active_import = client.get(
            f"/api/v1/admin/budget-projects/{project['id']}/active-import",
            headers=headers,
        )
        assert active_import.status_code == 200, active_import.text
        active_data = active_import.json()["data"]
        assert active_data["id"] == first["id"]
        assert active_data["active_revision"]["id"] == active_first["confirmed_revision_id"]
        assert active_data["active_revision"]["snapshot_sha256"] == active_first["confirmed_revision"]["snapshot_sha256"]

        second_upload = _upload(client, headers, project["id"], _quantity_cases_workbook())
        assert second_upload.status_code == 200
        second = second_upload.json()["data"]
        assert client.post(
            f"/api/v1/admin/budget-projects/imports/{second['batch_uuid']}/confirm",
            headers=headers,
        ).status_code == 200
        second_active = client.post(
            f"/api/v1/admin/budget-projects/imports/{second['batch_uuid']}/activate",
            headers=headers,
        )
        assert second_active.status_code == 200, second_active.text

        first_after_switch = client.get(
            f"/api/v1/admin/budget-projects/imports/{first['batch_uuid']}",
            headers=headers,
        ).json()["data"]
        assert first_after_switch["status"] == "superseded"
        assert first_after_switch["is_active"] is False
        reactivated = client.post(
            f"/api/v1/admin/budget-projects/imports/{first['batch_uuid']}/activate",
            headers=headers,
        )
        assert reactivated.status_code == 200, reactivated.text
        assert reactivated.json()["data"]["status"] == "active"
        second_after_reactivate = client.get(
            f"/api/v1/admin/budget-projects/imports/{second['batch_uuid']}",
            headers=headers,
        ).json()["data"]
        assert second_after_reactivate["status"] == "superseded"

        db = SessionLocal()
        try:
            profile = db.query(BudgetProjectProfile).filter_by(project_id=project["id"]).one()
            first_model = db.query(BudgetProjectImportBatch).filter_by(id=first["id"]).one()
            second_model = db.query(BudgetProjectImportBatch).filter_by(id=second["id"]).one()
            assert profile.active_import_batch_id == first_model.id
            assert profile.active_import_revision_id == first_model.confirmed_revision_id
            assert first_model.status == BUDGET_IMPORT_STATUS_ACTIVE
            assert second_model.status == BUDGET_IMPORT_STATUS_SUPERSEDED
            assert db.query(BudgetProjectImportBatch).filter_by(
                project_id=project["id"], status=BUDGET_IMPORT_STATUS_ACTIVE
            ).count() == 1
            assert db.query(BudgetProjectImportRevision).filter_by(batch_id=first_model.id).count() == 2
            assert [
                item.event_type
                for item in db.query(BudgetProjectImportLifecycleEvent)
                .filter_by(batch_id=first_model.id)
                .order_by(BudgetProjectImportLifecycleEvent.id)
                .all()
            ] == ["confirmed", "activated", "superseded", "reactivated"]
        finally:
            db.close()
    finally:
        _set_flag("feature_budget_projects", old_flag)


def test_budget_active_import_rejects_cross_project_batch_and_revision_pointers(client):
    owner = _create_user(roles=["quote_user"])
    headers = _login(client, owner)
    old_flag = _set_flag("feature_budget_projects", True)
    try:
        first_project = _create_project(client, headers, "指针归属项目甲")
        second_project = _create_project(client, headers, "指针归属项目乙")
        batches = []
        for project in (first_project, second_project):
            uploaded = _upload(client, headers, project["id"], _quantity_cases_workbook())
            assert uploaded.status_code == 200
            batch = uploaded.json()["data"]
            assert client.post(
                f"/api/v1/admin/budget-projects/imports/{batch['batch_uuid']}/confirm",
                headers=headers,
            ).status_code == 200
            activated = client.post(
                f"/api/v1/admin/budget-projects/imports/{batch['batch_uuid']}/activate",
                headers=headers,
            )
            assert activated.status_code == 200
            batches.append(activated.json()["data"])

        db = SessionLocal()
        try:
            first_profile = db.query(BudgetProjectProfile).filter_by(
                project_id=first_project["id"]
            ).one()
            first_profile.active_import_batch_id = batches[1]["id"]
            first_profile.active_import_revision_id = batches[1]["confirmed_revision_id"]
            db.commit()
        finally:
            db.close()
        invalid_pointer = client.get(
            f"/api/v1/admin/budget-projects/{first_project['id']}/active-import",
            headers=headers,
        )
        assert invalid_pointer.status_code == 409
        assert invalid_pointer.json()["detail"] == "BUDGET_ACTIVE_IMPORT_POINTER_INVALID"

        db = SessionLocal()
        try:
            first_profile = db.query(BudgetProjectProfile).filter_by(
                project_id=first_project["id"]
            ).one()
            first_batch = db.query(BudgetProjectImportBatch).filter_by(id=batches[0]["id"]).one()
            first_profile.active_import_batch_id = None
            first_profile.active_import_revision_id = None
            first_batch.status = "confirmed"
            first_batch.confirmed_revision_id = batches[1]["confirmed_revision_id"]
            db.commit()
        finally:
            db.close()
        invalid_revision = client.post(
            f"/api/v1/admin/budget-projects/imports/{batches[0]['batch_uuid']}/activate",
            headers=headers,
        )
        assert invalid_revision.status_code == 409
        assert invalid_revision.json()["detail"]["code"] == "BUDGET_IMPORT_REVISION_OWNERSHIP_INVALID"
    finally:
        _set_flag("feature_budget_projects", old_flag)


def test_project_can_remap_when_older_parsed_batch_is_not_latest(client):
    owner = _create_user(roles=["quote_user"])
    headers = _login(client, owner)
    old_flag = _set_flag("feature_budget_projects", True)
    try:
        project = _create_project(client, headers, "旧解析批次仍可重映射")
        older = _upload(client, headers, project["id"], _quantity_cases_workbook()).json()["data"]
        newer = _upload(client, headers, project["id"], _quantity_cases_workbook()).json()["data"]
        assert client.post(
            f"/api/v1/admin/budget-projects/imports/{newer['batch_uuid']}/confirm",
            headers=headers,
        ).status_code == 200
        assert client.post(
            f"/api/v1/admin/budget-projects/imports/{newer['batch_uuid']}/activate",
            headers=headers,
        ).status_code == 200

        detail = client.get(
            f"/api/v1/admin/budget-projects/{project['id']}", headers=headers
        )
        assert detail.status_code == 200
        assert detail.json()["data"]["latest_import"]["id"] == newer["id"]
        assert detail.json()["data"]["capabilities"]["can_remap"] is True
        older_detail = client.get(
            f"/api/v1/admin/budget-projects/imports/{older['batch_uuid']}",
            headers=headers,
        )
        assert older_detail.status_code == 200
        assert older_detail.json()["data"]["capabilities"]["can_remap"] is True
    finally:
        _set_flag("feature_budget_projects", old_flag)
