from __future__ import annotations

import csv
import json
from datetime import datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any, Mapping

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


CONFIRMATION_SHEET_NAME = "人工确认补量"
FEATURE_SHEET_NAME = "项目特征明细"
EVIDENCE_SHEET_NAME = "工程量证据明细"
GUIDE_SHEET_NAME = "填写说明"
FINAL_SHEET_NAME = "最终四字段清单"

ADOPT_COLUMN = "是否采用（业务填写：是/否）"
REVIEW_COLUMN = "核验结论（业务填写：通过/有问题）"
MANUAL_QUANTITY_COLUMN = "人工工程量（业务填写）"
MANUAL_UNIT_COLUMN = "确认单位（业务填写）"
MANUAL_FEATURE_COLUMN = "项目特征（业务确认/可修改）"
MANUAL_NAME_COLUMN = "项目名称（业务确认/可修改）"
QUANTITY_SOURCE_COLUMN = "工程量来源说明（业务填写）"
ISSUE_COLUMN = "问题说明（业务填写）"

CONFIRMATION_HEADERS = [
    "确认行号",
    ADOPT_COLUMN,
    REVIEW_COLUMN,
    MANUAL_QUANTITY_COLUMN,
    MANUAL_UNIT_COLUMN,
    QUANTITY_SOURCE_COLUMN,
    MANUAL_NAME_COLUMN,
    MANUAL_FEATURE_COLUMN,
    ISSUE_COLUMN,
    "候选编号",
    "标准项目编码",
    "标准项目名称",
    "标准单位",
    "工程量状态",
    "待补量原因",
    "工程量规则类型",
    "标准工程量计算规则",
    "建议工程量",
    "建议单位",
    "工程量证据摘要",
    "图纸识别名称",
    "图纸识别规格或做法",
    "来源文件",
    "来源行号",
    "匹配置信度",
    "项目特征缺失字段",
]

FINAL_HEADERS = ["项目名称", "项目特征", "单位", "工程量"]

_PLACEHOLDER_MARKERS = ("待确认", "待补", "缺失", "missing_needs_manual_review")


def build_drawing_confirmation_pack(
    standard_match_report: Mapping[str, Any],
    quantity_evidence_report: Mapping[str, Any],
) -> dict[str, Any]:
    feature_lookup = _build_feature_lookup(standard_match_report)
    match_lookup = _build_match_lookup(standard_match_report)
    rows: list[dict[str, Any]] = []
    feature_rows: list[dict[str, Any]] = []

    for index, quantity_row in enumerate(quantity_evidence_report.get("quantity_candidates") or [], start=1):
        key = _candidate_lookup_key(quantity_row)
        match_row = match_lookup.get(key, {})
        features = feature_lookup.get(key, [])
        feature_text = _build_feature_text(features)
        missing_fields = [
            _clean_text(item.get("field_name"))
            for item in features
            if _clean_text(item.get("status")) == "missing_needs_manual_review"
        ]
        confirmation_row = {
            "确认行号": f"BIZ2x6-{index:04d}",
            ADOPT_COLUMN: "待确认",
            REVIEW_COLUMN: "",
            MANUAL_QUANTITY_COLUMN: quantity_row.get("suggested_quantity") or "",
            MANUAL_UNIT_COLUMN: quantity_row.get("suggested_unit") or _first_unit(quantity_row.get("unit_options")),
            QUANTITY_SOURCE_COLUMN: "",
            MANUAL_NAME_COLUMN: quantity_row.get("standard_item_name", ""),
            MANUAL_FEATURE_COLUMN: feature_text,
            ISSUE_COLUMN: "",
            "候选编号": quantity_row.get("candidate_key", ""),
            "标准项目编码": quantity_row.get("standard_item_code", ""),
            "标准项目名称": quantity_row.get("standard_item_name", ""),
            "标准单位": "、".join(quantity_row.get("unit_options") or []),
            "工程量状态": quantity_row.get("quantity_status", ""),
            "待补量原因": quantity_row.get("quantity_block_reason", ""),
            "工程量规则类型": quantity_row.get("quantity_formula_type", ""),
            "标准工程量计算规则": quantity_row.get("quantity_rule_text", ""),
            "建议工程量": quantity_row.get("suggested_quantity", ""),
            "建议单位": quantity_row.get("suggested_unit", ""),
            "工程量证据摘要": quantity_row.get("evidence_summary", ""),
            "图纸识别名称": quantity_row.get("source_name", ""),
            "图纸识别规格或做法": quantity_row.get("source_spec_or_method", ""),
            "来源文件": quantity_row.get("source_file", ""),
            "来源行号": quantity_row.get("source_row_number", ""),
            "匹配置信度": match_row.get("match_confidence", ""),
            "项目特征缺失字段": "、".join(missing_fields),
        }
        rows.append(confirmation_row)
        for feature in features:
            feature_rows.append(
                {
                    "确认行号": confirmation_row["确认行号"],
                    "候选编号": confirmation_row["候选编号"],
                    "标准项目编码": confirmation_row["标准项目编码"],
                    "标准项目名称": confirmation_row["标准项目名称"],
                    "项目特征字段": feature.get("field_name", ""),
                    "候选填充值": feature.get("candidate_value", ""),
                    "状态": feature.get("status", ""),
                    "置信度": feature.get("confidence", ""),
                    "证据文本": feature.get("evidence_text", ""),
                }
            )

    evidence_rows = _build_evidence_rows(quantity_evidence_report)
    return {
        "ok": True,
        "phase": "BIZ-2x-6-manual-confirmation-pack",
        "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "safe_for_final_quantity_list": False,
        "summary": {
            "confirmation_row_count": len(rows),
            "feature_detail_count": len(feature_rows),
            "evidence_detail_count": len(evidence_rows),
            "final_export_requires_manual_confirmation": True,
        },
        "confirmation_rows": rows,
        "feature_rows": feature_rows,
        "evidence_rows": evidence_rows,
    }


def validate_confirmation_rows(rows: list[Mapping[str, Any]]) -> dict[str, Any]:
    issues: list[dict[str, Any]] = []
    final_rows: list[dict[str, Any]] = []
    skipped_count = 0

    for row_index, row in enumerate(rows, start=1):
        adopt = _clean_text(row.get(ADOPT_COLUMN))
        if adopt not in {"是", "Y", "y", "yes", "YES"}:
            skipped_count += 1
            continue
        row_id = _clean_text(row.get("确认行号")) or str(row_index)
        row_issues: list[str] = []
        if _clean_text(row.get(REVIEW_COLUMN)) != "通过":
            row_issues.append("核验结论必须填写“通过”")
        project_name = _clean_text(row.get(MANUAL_NAME_COLUMN))
        feature_text = _clean_text(row.get(MANUAL_FEATURE_COLUMN))
        unit = _clean_text(row.get(MANUAL_UNIT_COLUMN))
        quantity_text = _clean_text(row.get(MANUAL_QUANTITY_COLUMN))
        quantity = _parse_positive_decimal(quantity_text)
        if not project_name:
            row_issues.append("项目名称不能为空")
        if not feature_text:
            row_issues.append("项目特征不能为空")
        if any(marker in feature_text for marker in _PLACEHOLDER_MARKERS):
            row_issues.append("项目特征仍包含待确认/缺失提示，需要业务员补全或改写")
        if not unit:
            row_issues.append("确认单位不能为空")
        if quantity is None:
            row_issues.append("人工工程量必须填写大于 0 的数字")
        if not _clean_text(row.get(QUANTITY_SOURCE_COLUMN)):
            row_issues.append("工程量来源说明不能为空")
        if row_issues:
            issues.append({"row_number": row_index, "confirmation_row_id": row_id, "issues": row_issues})
            continue
        final_rows.append(
            {
                "项目名称": project_name,
                "项目特征": feature_text,
                "单位": unit,
                "工程量": _format_decimal(quantity),
            }
        )

    return {
        "ok": not issues and bool(final_rows),
        "summary": {
            "input_row_count": len(rows),
            "adopted_final_row_count": len(final_rows),
            "skipped_row_count": skipped_count,
            "issue_count": len(issues),
        },
        "issues": issues,
        "final_rows": final_rows,
    }


def read_confirmation_workbook(path: str | Path) -> list[dict[str, Any]]:
    workbook = load_workbook(path, data_only=True)
    if CONFIRMATION_SHEET_NAME not in workbook.sheetnames:
        raise ValueError(f"确认工作簿缺少工作表：{CONFIRMATION_SHEET_NAME}")
    sheet = workbook[CONFIRMATION_SHEET_NAME]
    headers = [_clean_text(cell.value) for cell in sheet[1]]
    rows: list[dict[str, Any]] = []
    for values in sheet.iter_rows(min_row=2, values_only=True):
        if not any(value not in (None, "") for value in values):
            continue
        rows.append({header: value for header, value in zip(headers, values) if header})
    return rows


def write_confirmation_workbook(pack: Mapping[str, Any], path: str | Path) -> None:
    workbook = Workbook()
    confirmation_sheet = workbook.active
    confirmation_sheet.title = CONFIRMATION_SHEET_NAME
    _write_rows(confirmation_sheet, CONFIRMATION_HEADERS, pack.get("confirmation_rows") or [])
    _style_confirmation_sheet(confirmation_sheet)

    feature_sheet = workbook.create_sheet(FEATURE_SHEET_NAME)
    feature_headers = [
        "确认行号",
        "候选编号",
        "标准项目编码",
        "标准项目名称",
        "项目特征字段",
        "候选填充值",
        "状态",
        "置信度",
        "证据文本",
    ]
    _write_rows(feature_sheet, feature_headers, pack.get("feature_rows") or [])
    _style_detail_sheet(feature_sheet)

    evidence_sheet = workbook.create_sheet(EVIDENCE_SHEET_NAME)
    evidence_headers = [
        "确认行号",
        "候选编号",
        "标准项目编码",
        "标准项目名称",
        "证据类型",
        "证据值",
        "证据单位",
        "是否匹配工程量规则",
        "证据置信度",
        "证据文本",
        "来源文件",
        "图层",
        "布局",
        "块名",
        "X",
        "Y",
        "源行号",
        "业务标签",
    ]
    _write_rows(evidence_sheet, evidence_headers, pack.get("evidence_rows") or [])
    _style_detail_sheet(evidence_sheet)

    guide_sheet = workbook.create_sheet(GUIDE_SHEET_NAME)
    _write_guide_sheet(guide_sheet, pack)

    workbook.save(path)


def write_final_quantity_workbook(final_rows: list[Mapping[str, Any]], path: str | Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = FINAL_SHEET_NAME
    _write_rows(sheet, FINAL_HEADERS, final_rows)
    _style_final_sheet(sheet)
    workbook.save(path)


def write_confirmation_csv(path: str | Path, rows: list[Mapping[str, Any]]) -> None:
    _write_csv(Path(path), rows, CONFIRMATION_HEADERS)


def build_confirmation_markdown(pack: Mapping[str, Any]) -> str:
    summary = pack.get("summary", {})
    lines = [
        "# BIZ-2x-6 图纸识别人工确认与补量包",
        "",
        f"- 生成时间：{pack.get('generated_at', '-')}",
        f"- 待确认候选行：{summary.get('confirmation_row_count', 0)}",
        f"- 项目特征明细：{summary.get('feature_detail_count', 0)}",
        f"- 工程量证据明细：{summary.get('evidence_detail_count', 0)}",
        "- 最终四字段清单必须在业务员确认候选、补全工程量并核验通过后才能导出。",
        "",
        "## 业务员必须填写",
        "",
        f"- `{ADOPT_COLUMN}`：采用该候选填“是”，不采用填“否”。",
        f"- `{REVIEW_COLUMN}`：确认项目、特征、单位、工程量都正确后填“通过”；否则填“有问题”。",
        f"- `{MANUAL_QUANTITY_COLUMN}`：采用行必须填写大于 0 的工程量。",
        f"- `{QUANTITY_SOURCE_COLUMN}`：填写工程量来源，例如“按平面图 F-P03 手算地面面积”。",
        f"- `{MANUAL_FEATURE_COLUMN}`：不得保留“待确认/缺失”等提示文字。",
    ]
    return "\n".join(lines) + "\n"


def write_confirmation_outputs(
    pack: Mapping[str, Any],
    output_dir: str | Path,
    *,
    stem: str | None = None,
) -> dict[str, str]:
    directory = Path(output_dir)
    directory.mkdir(parents=True, exist_ok=True)
    file_stem = stem or f"BIZ2x6_图纸识别人工确认补量包_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
    json_path = directory / f"{file_stem}.json"
    md_path = directory / f"{file_stem}.md"
    csv_path = directory / f"{file_stem}_人工确认补量.csv"
    xlsx_path = directory / f"{file_stem}.xlsx"

    json_path.write_text(json.dumps(pack, ensure_ascii=False, indent=2), encoding="utf-8")
    md_path.write_text(build_confirmation_markdown(pack), encoding="utf-8")
    write_confirmation_csv(csv_path, list(pack.get("confirmation_rows") or []))
    write_confirmation_workbook(pack, xlsx_path)
    return {
        "json": str(json_path),
        "markdown": str(md_path),
        "confirmation_csv": str(csv_path),
        "confirmation_xlsx": str(xlsx_path),
    }


def write_validation_report(
    validation: Mapping[str, Any],
    output_dir: str | Path,
    *,
    stem: str | None = None,
) -> dict[str, str]:
    directory = Path(output_dir)
    directory.mkdir(parents=True, exist_ok=True)
    file_stem = stem or f"BIZ2x6_确认表校验_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
    json_path = directory / f"{file_stem}.json"
    final_xlsx_path = directory / f"{file_stem}_最终四字段清单.xlsx"
    json_path.write_text(json.dumps(validation, ensure_ascii=False, indent=2), encoding="utf-8")
    outputs = {"json": str(json_path)}
    if validation.get("ok"):
        write_final_quantity_workbook(list(validation.get("final_rows") or []), final_xlsx_path)
        outputs["final_xlsx"] = str(final_xlsx_path)
    return outputs


def _build_feature_lookup(standard_match_report: Mapping[str, Any]) -> dict[tuple[str, str], list[dict[str, Any]]]:
    lookup: dict[tuple[str, str], list[dict[str, Any]]] = {}
    for group in standard_match_report.get("candidate_groups") or []:
        candidate_key = _clean_text(group.get("candidate_key"))
        for candidate in group.get("standard_candidates") or []:
            code = _clean_text(candidate.get("item_code"))
            lookup[(candidate_key, code)] = list(candidate.get("feature_fill_candidates") or [])
    return lookup


def _build_match_lookup(standard_match_report: Mapping[str, Any]) -> dict[tuple[str, str], dict[str, Any]]:
    lookup: dict[tuple[str, str], dict[str, Any]] = {}
    for row in standard_match_report.get("standard_item_candidates") or []:
        lookup[_candidate_lookup_key(row)] = dict(row)
    return lookup


def _candidate_lookup_key(row: Mapping[str, Any]) -> tuple[str, str]:
    return (_clean_text(row.get("candidate_key")), _clean_text(row.get("standard_item_code") or row.get("item_code")))


def _build_feature_text(features: list[Mapping[str, Any]]) -> str:
    if not features:
        return ""
    parts: list[str] = []
    for feature in features:
        name = _clean_text(feature.get("field_name"))
        value = _clean_text(feature.get("candidate_value"))
        if not name:
            continue
        if not value:
            value = "待确认"
        parts.append(f"{name}：{value}")
    return "；".join(parts)


def _build_evidence_rows(quantity_evidence_report: Mapping[str, Any]) -> list[dict[str, Any]]:
    row_id_lookup: dict[tuple[str, str], str] = {}
    for index, row in enumerate(quantity_evidence_report.get("quantity_candidates") or [], start=1):
        row_id_lookup[_candidate_lookup_key(row)] = f"BIZ2x6-{index:04d}"
    rows: list[dict[str, Any]] = []
    for evidence in quantity_evidence_report.get("quantity_evidence_rows") or []:
        lookup_key = _candidate_lookup_key(evidence)
        rows.append(
            {
                "确认行号": row_id_lookup.get(lookup_key, ""),
                "候选编号": evidence.get("candidate_key", ""),
                "标准项目编码": evidence.get("standard_item_code", ""),
                "标准项目名称": evidence.get("standard_item_name", ""),
                "证据类型": evidence.get("evidence_type", ""),
                "证据值": evidence.get("value", ""),
                "证据单位": evidence.get("unit", ""),
                "是否匹配工程量规则": "是" if evidence.get("is_direct_for_formula") else "否",
                "证据置信度": evidence.get("confidence", ""),
                "证据文本": evidence.get("text", ""),
                "来源文件": evidence.get("source_file", ""),
                "图层": evidence.get("layer", ""),
                "布局": evidence.get("layout", ""),
                "块名": evidence.get("block_name", ""),
                "X": evidence.get("x", ""),
                "Y": evidence.get("y", ""),
                "源行号": evidence.get("line_number", ""),
                "业务标签": " / ".join(evidence.get("role_tags") or []),
            }
        )
    return rows


def _write_rows(sheet: Any, headers: list[str], rows: list[Mapping[str, Any]]) -> None:
    sheet.append(headers)
    for row in rows:
        sheet.append([row.get(header, "") for header in headers])


def _style_confirmation_sheet(sheet: Any) -> None:
    _style_header(sheet)
    sheet.freeze_panes = "J2"
    sheet.auto_filter.ref = sheet.dimensions
    editable_fill = PatternFill("solid", fgColor="FFF2CC")
    required_fill = PatternFill("solid", fgColor="FCE4D6")
    for column_index, header in enumerate(CONFIRMATION_HEADERS, start=1):
        width = _column_width(header)
        sheet.column_dimensions[get_column_letter(column_index)].width = width
        for cell in sheet.iter_cols(min_col=column_index, max_col=column_index, min_row=2):
            for item in cell:
                item.alignment = Alignment(vertical="top", wrap_text=True)
                if header in {
                    ADOPT_COLUMN,
                    REVIEW_COLUMN,
                    MANUAL_QUANTITY_COLUMN,
                    MANUAL_UNIT_COLUMN,
                    QUANTITY_SOURCE_COLUMN,
                    MANUAL_NAME_COLUMN,
                    MANUAL_FEATURE_COLUMN,
                    ISSUE_COLUMN,
                }:
                    item.fill = editable_fill
                if header in {ADOPT_COLUMN, REVIEW_COLUMN, MANUAL_QUANTITY_COLUMN, QUANTITY_SOURCE_COLUMN}:
                    item.fill = required_fill
    max_row = max(sheet.max_row, 2)
    _add_list_validation(sheet, ADOPT_COLUMN, '"待确认,是,否"', max_row)
    _add_list_validation(sheet, REVIEW_COLUMN, '"通过,有问题"', max_row)


def _style_detail_sheet(sheet: Any) -> None:
    _style_header(sheet)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for column_index, cell in enumerate(sheet[1], start=1):
        sheet.column_dimensions[get_column_letter(column_index)].width = _column_width(_clean_text(cell.value))
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def _style_final_sheet(sheet: Any) -> None:
    _style_header(sheet)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    widths = {"A": 24, "B": 80, "C": 12, "D": 14}
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def _write_guide_sheet(sheet: Any, pack: Mapping[str, Any]) -> None:
    summary = pack.get("summary", {})
    rows = [
        ["用途", "这张工作簿用于确认 DWG 识图后的标准项目候选，并由业务员补工程量。"],
        ["待确认候选行", summary.get("confirmation_row_count", 0)],
        ["填写 1", f"`{ADOPT_COLUMN}`：采用填“是”，不采用填“否”。"],
        ["填写 2", f"`{REVIEW_COLUMN}`：确认无误填“通过”，否则填“有问题”。"],
        ["填写 3", f"`{MANUAL_QUANTITY_COLUMN}`：采用行必须填大于 0 的数字。"],
        ["填写 4", f"`{QUANTITY_SOURCE_COLUMN}`：说明工程量来自哪张图或哪次手算。"],
        ["导出规则", "只有“是否采用=是、核验结论=通过、工程量>0、项目特征无待确认”的行，才会导出最终四字段 Excel。"],
    ]
    for row in rows:
        sheet.append(row)
    _style_header(sheet)
    sheet.column_dimensions["A"].width = 18
    sheet.column_dimensions["B"].width = 100
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def _style_header(sheet: Any) -> None:
    fill = PatternFill("solid", fgColor="1F4E78")
    font = Font(color="FFFFFF", bold=True)
    for cell in sheet[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _add_list_validation(sheet: Any, header: str, formula: str, max_row: int) -> None:
    column = CONFIRMATION_HEADERS.index(header) + 1
    letter = get_column_letter(column)
    validation = DataValidation(type="list", formula1=formula, allow_blank=True)
    sheet.add_data_validation(validation)
    validation.add(f"{letter}2:{letter}{max_row}")


def _column_width(header: str) -> int:
    long_headers = {
        MANUAL_FEATURE_COLUMN,
        "标准工程量计算规则",
        "工程量证据摘要",
        "图纸识别规格或做法",
        "待补量原因",
        "证据文本",
    }
    if header in long_headers:
        return 60
    if header in {MANUAL_NAME_COLUMN, "标准项目名称", "图纸识别名称"}:
        return 24
    if header in {"来源文件"}:
        return 34
    return max(12, min(26, len(header) * 2))


def _write_csv(path: Path, rows: list[Mapping[str, Any]], headers: list[str]) -> None:
    with path.open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.DictWriter(handle, fieldnames=headers)
        writer.writeheader()
        for row in rows:
            writer.writerow({header: row.get(header, "") for header in headers})


def _first_unit(value: Any) -> str:
    units = list(value or [])
    return _clean_text(units[0]) if units else ""


def _parse_positive_decimal(value: str) -> Decimal | None:
    try:
        parsed = Decimal(str(value).strip())
    except (InvalidOperation, ValueError):
        return None
    if parsed <= 0:
        return None
    return parsed


def _format_decimal(value: Decimal | None) -> str:
    if value is None:
        return ""
    normalized = value.normalize()
    return format(normalized, "f")


def _clean_text(value: Any) -> str:
    return str(value or "").strip()
