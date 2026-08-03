from __future__ import annotations

import argparse
import csv
import json
import re
from collections import Counter
from dataclasses import dataclass, asdict
from datetime import datetime
from difflib import SequenceMatcher
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


FOUR_FIELD_HEADERS = ["项目名称", "项目特征", "单位", "工程量"]
ITEM_ALIASES = ("项目名称", "清单项目名称", "项目", "工程内容", "名称")
FEATURE_ALIASES = ("项目特征", "项目特征描述", "特征描述", "特征", "规格", "规格/特征")
UNIT_ALIASES = ("计量单位", "单位")
QUANTITY_ALIASES = ("工程量", "数量", "工程数量")
CODE_ALIASES = ("项目编码", "编码", "清单编码")


@dataclass
class AnswerRow:
    row_no: int
    sheet_name: str
    item_code: str
    item_name: str
    feature: str
    unit: str
    quantity: str
    raw_text: str


@dataclass
class AlignmentRow:
    status: str
    match_score: float
    manual_row_no: str
    manual_item_name: str
    manual_feature: str
    manual_unit: str
    manual_quantity: str
    system_item_name: str
    system_feature: str
    system_unit: str
    system_quantity: str
    rule_gap: str
    rule_confirmation: str


def main() -> None:
    parser = argparse.ArgumentParser(description="BIZ-2x 样例答案对齐与规则反推")
    parser.add_argument("--manual-xlsx", required=True, help="人工四字段清单 .xlsx")
    parser.add_argument("--listing-json", required=True, help="DWG 上传列项 JSON")
    parser.add_argument("--output-dir", default="../outputs/biz2x_rule_reverse", help="输出目录")
    parser.add_argument("--timestamp", default="", help="输出时间戳")
    args = parser.parse_args()

    timestamp = args.timestamp or datetime.now().strftime("%Y%m%d_%H%M%S")
    output_dir = Path(args.output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    manual_rows, sheet_summaries = extract_manual_answer_rows(Path(args.manual_xlsx))
    listing_report = json.loads(Path(args.listing_json).read_text(encoding="utf-8"))
    system_rows = extract_system_four_field_rows(listing_report)
    alignment_rows = align_rows(manual_rows, system_rows)
    rule_rows = infer_rule_rows(alignment_rows, listing_report)

    stem = f"BIZ2x_样例答案反推规则_{timestamp}"
    manual_csv = output_dir / f"{stem}_人工清单.csv"
    system_csv = output_dir / f"{stem}_系统清单.csv"
    alignment_csv = output_dir / f"{stem}_对齐明细.csv"
    rule_csv = output_dir / f"{stem}_规则反推.csv"
    markdown = output_dir / f"{stem}.md"
    summary_json = output_dir / f"{stem}.json"

    write_dataclass_csv(manual_csv, manual_rows)
    write_dict_csv(system_csv, system_rows, FOUR_FIELD_HEADERS)
    write_dataclass_csv(alignment_csv, alignment_rows)
    write_dict_csv(rule_csv, rule_rows, ["规则类型", "涉及项目", "问题表现", "反推规则", "需要业务确认"])

    summary = build_summary(manual_rows, system_rows, alignment_rows, rule_rows, sheet_summaries, listing_report)
    summary_json.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    markdown.write_text(build_markdown(summary, alignment_rows, rule_rows), encoding="utf-8")

    print(json.dumps({
        "manual_count": len(manual_rows),
        "system_count": len(system_rows),
        "alignment_counts": summary["alignment_counts"],
        "outputs": {
            "markdown": str(markdown),
            "alignment_csv": str(alignment_csv),
            "rule_csv": str(rule_csv),
            "summary_json": str(summary_json),
        },
    }, ensure_ascii=False, indent=2))


def extract_manual_answer_rows(path: Path) -> tuple[list[AnswerRow], list[dict[str, Any]]]:
    workbook = load_workbook(path, data_only=True)
    rows: list[AnswerRow] = []
    sheet_summaries: list[dict[str, Any]] = []
    for worksheet in workbook.worksheets:
        values = list(worksheet.iter_rows(values_only=True))
        header = find_header(values)
        sheet_summary = {
            "sheet_name": worksheet.title,
            "max_row": worksheet.max_row,
            "max_column": worksheet.max_column,
            "header_found": bool(header),
            "header_row": header["row_index"] if header else "",
        }
        if not header:
            sheet_summaries.append(sheet_summary)
            continue
        parsed_count = 0
        for row_index, row in enumerate(values[header["row_index"] + 1 :], start=header["row_index"] + 2):
            raw_values = [clean_text(value) for value in row]
            if not any(raw_values):
                continue
            item_name = value_at(raw_values, header.get("item"))
            feature = value_at(raw_values, header.get("feature"))
            unit = value_at(raw_values, header.get("unit"))
            quantity = value_at(raw_values, header.get("quantity"))
            item_code = value_at(raw_values, header.get("code"))
            if not looks_like_bill_row(item_name, feature, unit, quantity, raw_values):
                continue
            rows.append(
                AnswerRow(
                    row_no=row_index,
                    sheet_name=worksheet.title,
                    item_code=item_code,
                    item_name=item_name,
                    feature=feature,
                    unit=unit,
                    quantity=quantity,
                    raw_text="；".join(value for value in raw_values if value),
                )
            )
            parsed_count += 1
        sheet_summary["parsed_row_count"] = parsed_count
        sheet_summaries.append(sheet_summary)
    return rows, sheet_summaries


def find_header(values: list[tuple[Any, ...]]) -> dict[str, int] | None:
    best: dict[str, int] | None = None
    best_score = 0
    for row_index, row in enumerate(values[:80]):
        cells = [clean_text(value) for value in row]
        mapping = {
            "item": find_alias_column(cells, ITEM_ALIASES),
            "feature": find_alias_column(cells, FEATURE_ALIASES),
            "unit": find_alias_column(cells, UNIT_ALIASES),
            "quantity": find_alias_column(cells, QUANTITY_ALIASES),
            "code": find_alias_column(cells, CODE_ALIASES),
        }
        score = sum(1 for key in ("item", "feature", "unit", "quantity") if mapping[key] is not None)
        if score > best_score:
            best_score = score
            best = {"row_index": row_index, **{k: v for k, v in mapping.items() if v is not None}}
        if score >= 4:
            return best
    if best_score >= 3 and best:
        return best
    return None


def find_alias_column(cells: list[str], aliases: tuple[str, ...]) -> int | None:
    for index, cell in enumerate(cells):
        normalized = normalize_header(cell)
        if not normalized:
            continue
        for alias in aliases:
            if normalize_header(alias) == normalized or normalize_header(alias) in normalized:
                return index
    return None


def looks_like_bill_row(item_name: str, feature: str, unit: str, quantity: str, raw_values: list[str]) -> bool:
    raw_text = "".join(raw_values)
    if not item_name and not feature:
        return False
    if any(term in raw_text for term in ("合计", "小计", "总价", "综合单价", "暂列金额", "措施项目")) and not unit:
        return False
    if item_name in {"项目名称", "清单项目名称", "名称"}:
        return False
    return bool(item_name and (unit or quantity or feature))


def extract_system_four_field_rows(report: dict[str, Any]) -> list[dict[str, str]]:
    rows = report.get("quantity_list_rows") or []
    if not rows:
        traces = {
            str(row.get("识别项目编号") or ""): row
            for row in report.get("special_quantity_trace_rows") or []
            if row.get("识别项目编号")
        }
        rows = []
        for project in report.get("project_rows") or []:
            trace = traces.get(str(project.get("识别项目编号") or "")) or {}
            has_quantity = trace.get("建议工程量") not in (None, "") and str(trace.get("是否可复核") or "") == "是"
            rows.append(
                {
                    "项目名称": clean_text(project.get("项目名称")),
                    "项目特征": clean_text(project.get("项目特征")),
                    "单位": clean_text(trace.get("建议单位") if has_quantity else project.get("单位")),
                    "工程量": clean_text(trace.get("建议工程量") if has_quantity else project.get("工程量") or "待算量"),
                }
            )
    return [
        {
            "项目名称": clean_text(row.get("项目名称")),
            "项目特征": clean_text(row.get("项目特征")),
            "单位": clean_text(row.get("单位")),
            "工程量": clean_text(row.get("工程量")),
        }
        for row in rows
        if clean_text(row.get("项目名称"))
    ]


def align_rows(manual_rows: list[AnswerRow], system_rows: list[dict[str, str]]) -> list[AlignmentRow]:
    used_system: set[int] = set()
    alignments: list[AlignmentRow] = []
    for manual in manual_rows:
        best_index = -1
        best_score = 0.0
        for index, system in enumerate(system_rows):
            if index in used_system:
                continue
            score = match_score(manual, system)
            if score > best_score:
                best_score = score
                best_index = index
        if best_index >= 0 and best_score >= 0.48:
            used_system.add(best_index)
            system = system_rows[best_index]
            status = "已匹配"
        else:
            system = {}
            status = "人工有，系统未识别"
        gap, confirmation = infer_gap(
            manual_item=manual.item_name,
            manual_feature=manual.feature,
            manual_unit=manual.unit,
            manual_quantity=manual.quantity,
            system_item=system.get("项目名称", ""),
            system_feature=system.get("项目特征", ""),
            system_unit=system.get("单位", ""),
            system_quantity=system.get("工程量", ""),
            matched=status == "已匹配",
        )
        alignments.append(
            AlignmentRow(
                status=status,
                match_score=round(best_score, 3),
                manual_row_no=str(manual.row_no),
                manual_item_name=manual.item_name,
                manual_feature=manual.feature,
                manual_unit=manual.unit,
                manual_quantity=manual.quantity,
                system_item_name=system.get("项目名称", ""),
                system_feature=system.get("项目特征", ""),
                system_unit=system.get("单位", ""),
                system_quantity=system.get("工程量", ""),
                rule_gap=gap,
                rule_confirmation=confirmation,
            )
        )

    for index, system in enumerate(system_rows):
        if index in used_system:
            continue
        gap, confirmation = infer_gap(
            manual_item="",
            manual_feature="",
            manual_unit="",
            manual_quantity="",
            system_item=system.get("项目名称", ""),
            system_feature=system.get("项目特征", ""),
            system_unit=system.get("单位", ""),
            system_quantity=system.get("工程量", ""),
            matched=False,
        )
        alignments.append(
            AlignmentRow(
                status="系统有，人工未列",
                match_score=0.0,
                manual_row_no="",
                manual_item_name="",
                manual_feature="",
                manual_unit="",
                manual_quantity="",
                system_item_name=system.get("项目名称", ""),
                system_feature=system.get("项目特征", ""),
                system_unit=system.get("单位", ""),
                system_quantity=system.get("工程量", ""),
                rule_gap=gap,
                rule_confirmation=confirmation,
            )
        )
    return alignments


def match_score(manual: AnswerRow, system: dict[str, str]) -> float:
    manual_name = normalize_name(manual.item_name)
    system_name = normalize_name(system.get("项目名称", ""))
    if not manual_name or not system_name:
        return 0.0
    score = SequenceMatcher(None, manual_name, system_name).ratio()
    if manual_name in system_name or system_name in manual_name:
        score = max(score, 0.72)
    manual_feature = normalize_name(manual.feature)
    system_feature = normalize_name(system.get("项目特征", ""))
    if manual_feature and system_feature:
        score = max(score, 0.65 * score + 0.35 * SequenceMatcher(None, manual_feature, system_feature).ratio())
    if clean_text(manual.unit) and clean_text(system.get("单位")) and clean_text(manual.unit) == clean_text(system.get("单位")):
        score += 0.08
    return min(score, 1.0)


def infer_gap(
    *,
    manual_item: str,
    manual_feature: str,
    manual_unit: str,
    manual_quantity: str,
    system_item: str,
    system_feature: str,
    system_unit: str,
    system_quantity: str,
    matched: bool,
) -> tuple[str, str]:
    item = manual_item or system_item
    kind = classify_rule_kind(item + " " + manual_feature + " " + system_feature)
    if not matched and manual_item:
        return "漏项识别", f"需要确认图纸中哪些文字/材料做法应生成“{manual_item}”，并加入项目识别关键词和标准库映射。"
    if not manual_item and system_item:
        return "疑似误识别", f"需要确认“{system_item}”是否为本工程实际清单项；若不是，应加入非项目/通用做法过滤规则。"
    if system_quantity in {"", "待算量", "待 CAD 区域/边界绑定后按标准规则计算"} and manual_quantity:
        return "工程量未算出", rule_confirmation_for_kind(kind, item)
    if manual_unit and system_unit and manual_unit != system_unit:
        return "单位口径不一致", f"需要确认“{item}”按人工清单单位 {manual_unit}，还是标准库/系统单位 {system_unit}。"
    if manual_quantity and system_quantity and not quantities_close(manual_quantity, system_quantity):
        return "工程量差异", rule_confirmation_for_kind(kind, item)
    return "可沉淀规则", rule_confirmation_for_kind(kind, item)


def classify_rule_kind(text: str) -> str:
    if any(term in text for term in ("踢脚", "踢脚线")):
        return "baseboard"
    if any(term in text for term in ("防水", "涂膜")):
        return "waterproof"
    if any(term in text for term in ("吊顶", "天棚", "天花", "格栅", "软膜")):
        return "ceiling"
    if any(term in text for term in ("地面", "楼地面", "地砖", "地板")):
        return "floor"
    if any(term in text for term in ("墙面", "墙柱面", "隔墙", "隔断")):
        return "wall"
    if any(term in text for term in ("窗帘盒", "窗台板")):
        return "linear_finish"
    if any(term in text for term in ("拆除", "铲除")):
        return "demolition"
    return "other"


def rule_confirmation_for_kind(kind: str, item: str) -> str:
    rules = {
        "ceiling": "确认吊顶/天棚项目对应的 CAD 闭合区域、材料编号和是否按水平投影面积计算；不同吊顶做法需按材料/房间分区。",
        "floor": "确认地面项目对应的房间/区域边界、面层材料编号，以及是否按闭合区域面积计算并扣除洞口。",
        "waterproof": "确认防水部位、适用房间、防水高度、净周长/地面面积、门洞或开口扣减规则。",
        "baseboard": "确认踢脚线适用房间、沿墙净长度、门洞扣减、无踢脚线墙段和高度/材质口径。",
        "wall": "确认墙面/隔墙对应边界、墙高、洞口扣减、饰面材料和是否按展开面积计算。",
        "linear_finish": "确认窗帘盒/窗台板是按长度还是面积，CAD 来源取线段、块参照还是节点尺寸。",
        "demolition": "确认拆除项是否纳入本 GB/T 标准库范围，以及拆除范围对应的原有构件区域/长度/数量。",
        "other": f"确认“{item}”的适用图纸来源、标准项目编码、计量单位和 CAD 取量方式。",
    }
    return rules.get(kind, rules["other"])


def infer_rule_rows(alignment_rows: list[AlignmentRow], listing_report: dict[str, Any]) -> list[dict[str, str]]:
    grouped: dict[str, list[AlignmentRow]] = {}
    for row in alignment_rows:
        kind = classify_rule_kind(row.manual_item_name + " " + row.system_item_name + " " + row.manual_feature + " " + row.system_feature)
        grouped.setdefault(kind, []).append(row)

    rows: list[dict[str, str]] = []
    for kind, items in grouped.items():
        project_names = sorted({row.manual_item_name or row.system_item_name for row in items if row.manual_item_name or row.system_item_name})
        gap_counts = Counter(row.rule_gap for row in items)
        rows.append(
            {
                "规则类型": kind,
                "涉及项目": "；".join(project_names[:12]),
                "问题表现": "；".join(f"{key}{value}项" for key, value in gap_counts.most_common()),
                "反推规则": rule_confirmation_for_kind(kind, project_names[0] if project_names else ""),
                "需要业务确认": business_confirmation_for_group(kind, items),
            }
        )

    special_summary = (listing_report.get("special_quantity_summary") or {})
    if special_summary.get("blocked_trace_count"):
        rows.append(
            {
                "规则类型": "cad_region_binding",
                "涉及项目": "全部专项算量项目",
                "问题表现": f"专项算量阻断 {special_summary.get('blocked_trace_count')} 项",
                "反推规则": "需要用人工清单工程量反查每个项目应绑定的房间/区域/构件，并沉淀区域绑定优先级。",
                "需要业务确认": "确认每个人工清单项目对应哪张图、哪个房间/区域、哪个材料编号或做法编号。",
            }
        )
    return rows


def business_confirmation_for_group(kind: str, rows: list[AlignmentRow]) -> str:
    if kind == "ceiling":
        return "吊顶做法编号、适用房间、是否复用天棚涂料面积、异形/灯槽是否另算。"
    if kind == "floor":
        return "地面材料编号、适用房间、门洞/洞口扣减口径。"
    if kind == "waterproof":
        return "防水高度、墙面/地面防水是否拆分、门洞扣减规则。"
    if kind == "baseboard":
        return "哪些房间有踢脚线、门洞和固定家具段是否扣除。"
    if kind == "wall":
        return "墙高、洞口扣减、墙面材料编号和是否展开计算。"
    if kind == "linear_finish":
        return "按长度/面积计量的选择，以及 CAD 取线/取块/取节点尺寸的来源。"
    if kind == "demolition":
        return "拆除范围是否以原图区域为准，拆除项对应标准编码。"
    return "确认项目是否为实际施工项及其标准编码、单位、取量来源。"


def build_summary(
    manual_rows: list[AnswerRow],
    system_rows: list[dict[str, str]],
    alignment_rows: list[AlignmentRow],
    rule_rows: list[dict[str, str]],
    sheet_summaries: list[dict[str, Any]],
    listing_report: dict[str, Any],
) -> dict[str, Any]:
    alignment_counts = Counter(row.status for row in alignment_rows)
    gap_counts = Counter(row.rule_gap for row in alignment_rows)
    return {
        "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "manual_row_count": len(manual_rows),
        "system_row_count": len(system_rows),
        "alignment_counts": dict(alignment_counts.most_common()),
        "gap_counts": dict(gap_counts.most_common()),
        "sheet_summaries": sheet_summaries,
        "dwg_summary": listing_report.get("summary") or {},
        "rule_row_count": len(rule_rows),
    }


def build_markdown(summary: dict[str, Any], alignment_rows: list[AlignmentRow], rule_rows: list[dict[str, str]]) -> str:
    lines = [
        "# BIZ-2x 样例答案反推规则报告",
        "",
        f"- 生成时间：{summary['generated_at']}",
        f"- 人工清单行数：{summary['manual_row_count']}",
        f"- 系统识别四字段行数：{summary['system_row_count']}",
        f"- 对齐结果：{summary['alignment_counts']}",
        f"- 规则缺口：{summary['gap_counts']}",
        "",
        "## 关键结论",
        "",
    ]
    if summary["manual_row_count"] == 0:
        lines.append("- 未能从人工清单中识别四字段表头，需要先确认人工清单 Sheet 和表头位置。")
    else:
        lines.extend(
            [
                "- 人工答案可用于反推项目识别、项目特征填充、CAD 区域绑定和工程量计算规则。",
                "- 系统未识别的人工项应进入“疑似漏项 -> 标准库映射 -> 图纸证据定位”。",
                "- 系统多识别但人工未列的项应进入“非项目/通用做法/专业范围过滤”。",
                "- 已匹配但工程量差异或待算量的项，应优先补 CAD 区域绑定、净周长、洞口扣减和展开面积规则。",
            ]
        )
    lines.extend(["", "## 规则反推", "", "| 规则类型 | 涉及项目 | 问题表现 | 反推规则 | 需要业务确认 |", "| --- | --- | --- | --- | --- |"])
    for row in rule_rows:
        lines.append(
            "| "
            + " | ".join(
                md(row.get(key, ""))
                for key in ["规则类型", "涉及项目", "问题表现", "反推规则", "需要业务确认"]
            )
            + " |"
        )
    lines.extend(["", "## 对齐明细节选", "", "| 状态 | 分数 | 人工项目 | 人工工程量 | 系统项目 | 系统工程量 | 缺口 |", "| --- | ---: | --- | --- | --- | --- | --- |"])
    for row in alignment_rows[:80]:
        lines.append(
            "| "
            + " | ".join(
                [
                    md(row.status),
                    md(row.match_score),
                    md(row.manual_item_name),
                    md(row.manual_quantity),
                    md(row.system_item_name),
                    md(row.system_quantity),
                    md(row.rule_gap),
                ]
            )
            + " |"
        )
    return "\n".join(lines) + "\n"


def quantities_close(left: str, right: str, tolerance: float = 0.05) -> bool:
    l_val = parse_number(left)
    r_val = parse_number(right)
    if l_val is None or r_val is None:
        return False
    if abs(l_val - r_val) <= 0.01:
        return True
    denom = max(abs(l_val), 1.0)
    return abs(l_val - r_val) / denom <= tolerance


def parse_number(value: str) -> float | None:
    match = re.search(r"-?\d+(?:\.\d+)?", str(value or "").replace(",", ""))
    if not match:
        return None
    try:
        return float(match.group())
    except ValueError:
        return None


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).replace("\n", " ").replace("\r", " ").strip()
    return re.sub(r"\s+", " ", text)


def normalize_header(value: str) -> str:
    return re.sub(r"[\s:：/\\（）()【】\[\]、，,；;|]+", "", clean_text(value).lower())


def normalize_name(value: str) -> str:
    text = clean_text(value).lower()
    text = re.sub(r"\d+(?:\.\d+)?\s*(mm|cm|m|㎡|m2|平方米|厚|高|宽)?", "", text)
    text = re.sub(r"[\s:：/\\（）()【】\[\]、，,；;|·•+-]+", "", text)
    return text


def value_at(values: list[str], index: int | None) -> str:
    if index is None or index >= len(values):
        return ""
    return values[index]


def write_dataclass_csv(path: Path, rows: list[Any]) -> None:
    if not rows:
        path.write_text("", encoding="utf-8-sig")
        return
    fieldnames = list(asdict(rows[0]).keys())
    with path.open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        for row in rows:
            writer.writerow(asdict(row))


def write_dict_csv(path: Path, rows: list[dict[str, Any]], fieldnames: list[str]) -> None:
    with path.open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        for row in rows:
            writer.writerow({field: row.get(field, "") for field in fieldnames})


def md(value: Any) -> str:
    return str(value if value is not None else "").replace("|", "\\|").replace("\n", " ")


if __name__ == "__main__":
    main()
