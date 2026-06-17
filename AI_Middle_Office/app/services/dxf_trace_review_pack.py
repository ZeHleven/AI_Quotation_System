from __future__ import annotations

import csv
import json
from datetime import datetime
from pathlib import Path
from typing import Any, Mapping

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


TRACE_REVIEW_SHEET_NAME = "标准规则trace复核"
BLOCKED_SHEET_NAME = "阻断项明细"
TRACE_DETAIL_SHEET_NAME = "计算追溯明细"
GUIDE_SHEET_NAME = "填写说明"

AUTO_CONCLUSION_COLUMN = "系统初判结论"
AUTO_ACTION_COLUMN = "系统建议动作"
RISK_LEVEL_COLUMN = "风险等级"
AUTO_REASON_COLUMN = "自动初判原因"
ADOPT_COLUMN = "是否采用（业务填写：是/否）"
REVIEW_COLUMN = "核验结论（业务填写：通过/有问题）"
MANUAL_QUANTITY_COLUMN = "确认工程量（业务填写）"
MANUAL_UNIT_COLUMN = "确认单位（业务填写）"
MANUAL_NAME_COLUMN = "项目名称（业务确认/可修改）"
MANUAL_FEATURE_COLUMN = "项目特征（业务填写，按标准字段口径）"
DEDUCTION_REVIEW_COLUMN = "扣减/合并规则复核（业务填写）"
QUANTITY_SOURCE_COLUMN = "工程量来源说明（业务填写）"
ISSUE_COLUMN = "问题说明（业务填写）"

TRACE_REVIEW_HEADERS = [
    "复核行号",
    AUTO_CONCLUSION_COLUMN,
    AUTO_ACTION_COLUMN,
    RISK_LEVEL_COLUMN,
    AUTO_REASON_COLUMN,
    ADOPT_COLUMN,
    REVIEW_COLUMN,
    MANUAL_QUANTITY_COLUMN,
    MANUAL_UNIT_COLUMN,
    QUANTITY_SOURCE_COLUMN,
    MANUAL_NAME_COLUMN,
    MANUAL_FEATURE_COLUMN,
    DEDUCTION_REVIEW_COLUMN,
    ISSUE_COLUMN,
    "是否可进入人工复核",
    "trace状态",
    "建议编号",
    "标准项目编码",
    "标准项目名称",
    "标准单位",
    "标准规则类型",
    "标准工程量计算规则",
    "几何建议量",
    "几何单位",
    "标准规则建议量",
    "建议单位",
    "阻断原因",
    "未解决事项",
    "CAD几何公式",
    "CAD来源图元行号",
    "CAD图元类型统计",
]

BLOCKED_HEADERS = [
    "建议编号",
    "绑定状态",
    "文件名",
    "图层",
    "块名",
    "业务提示",
    "建议量类型",
    "建议量",
    "单位",
    "标准候选数",
    "可复核trace数",
    "说明",
    "风险提示",
]

TRACE_DETAIL_HEADERS = [
    "复核行号",
    "建议编号",
    "trace状态",
    "标准项目编码",
    "标准项目名称",
    "标准工程量计算规则",
    "几何来源key",
    "几何公式",
    "几何输入量",
    "几何输入单位",
    "标准规则应用方式",
    "结果工程量",
    "结果单位",
    "源追溯JSON",
]


def load_json_report(path: str | Path) -> dict[str, Any]:
    report_path = Path(path)
    if not report_path.exists():
        raise ValueError(f"Report not found: {report_path}")
    return json.loads(report_path.read_text(encoding="utf-8"))


def build_trace_review_pack(standard_rule_binding_report: Mapping[str, Any]) -> dict[str, Any]:
    trace_rows: list[dict[str, Any]] = []
    detail_rows: list[dict[str, Any]] = []
    traces = list(standard_rule_binding_report.get("standard_rule_traces") or [])
    auto_decisions = _build_auto_decisions(traces)
    for index, trace in enumerate(traces, start=1):
        review_row_id = f"BIZ2x9h-{index:04d}"
        row = _build_review_row(review_row_id, trace, auto_decisions.get(id(trace), _blocked_decision("缺少系统初判上下文。")))
        trace_rows.append(row)
        detail_rows.append(_build_detail_row(review_row_id, trace))

    blocked_rows = _build_blocked_rows(standard_rule_binding_report)
    ready_count = sum(1 for row in trace_rows if row["是否可进入人工复核"] == "是")
    blocked_trace_count = len(trace_rows) - ready_count
    action_counts = _count_values(trace_rows, AUTO_ACTION_COLUMN)
    risk_counts = _count_values(trace_rows, RISK_LEVEL_COLUMN)
    return {
        "ok": True,
        "phase": "BIZ-2x-9h-2-standard-rule-trace-auto-decision-review-pack",
        "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "safe_for_final_quantity_list": False,
        "summary": {
            "trace_review_row_count": len(trace_rows),
            "ready_for_manual_review_count": ready_count,
            "blocked_trace_row_count": blocked_trace_count,
            "blocked_suggestion_row_count": len(blocked_rows),
            "auto_action_counts": action_counts,
            "risk_level_counts": risk_counts,
            "final_export_requires_manual_confirmation": True,
            "final_ready_count": 0,
        },
        "trace_review_rows": trace_rows,
        "blocked_rows": blocked_rows,
        "trace_detail_rows": detail_rows,
    }


def build_trace_review_markdown(pack: Mapping[str, Any]) -> str:
    summary = pack.get("summary", {})
    lines = [
        "# BIZ-2x-9h-2 标准规则 trace 自动初判复核包",
        "",
        f"- 生成时间：{pack.get('generated_at', '-')}",
        f"- trace 复核行：{summary.get('trace_review_row_count', 0)}",
        f"- 可进入人工复核：{summary.get('ready_for_manual_review_count', 0)}",
        f"- 阻断 trace 行：{summary.get('blocked_trace_row_count', 0)}",
        f"- 阻断建议量：{summary.get('blocked_suggestion_row_count', 0)}",
        f"- 系统建议采用：{summary.get('auto_action_counts', {}).get('建议采用', 0)}",
        f"- 系统建议不采用：{summary.get('auto_action_counts', {}).get('建议不采用', 0)}",
        f"- 系统要求人工确认：{summary.get('auto_action_counts', {}).get('需人工确认', 0)}",
        "- 系统已经自动初判采用/不采用/需人工确认，但本包仍不生成最终四字段清单，必须人工确认后再进入后续导出或报价。",
        "",
        "## 业务员填写重点",
        "",
        f"- `{AUTO_CONCLUSION_COLUMN}` / `{AUTO_ACTION_COLUMN}` / `{RISK_LEVEL_COLUMN}`：系统已预先判断，业务员优先处理“需人工确认”和“高”风险行。",
        f"- `{ADOPT_COLUMN}`：系统会按初判预填“是/否/待确认”，业务员可覆盖。",
        f"- `{REVIEW_COLUMN}`：标准项目、项目特征、工程量和扣减规则都确认后填“通过”；有疑问填“有问题”。",
        f"- `{MANUAL_FEATURE_COLUMN}`：必须按标准项目特征字段口径填写，不要自由发挥字段名。",
        f"- `{DEDUCTION_REVIEW_COLUMN}`：说明是否已按标准规则处理扣减、合并、展开面积等问题。",
        f"- `{QUANTITY_SOURCE_COLUMN}`：说明工程量采用 CAD 建议量、人工复核量或其他来源。",
    ]
    return "\n".join(lines) + "\n"


def write_trace_review_outputs(
    pack: Mapping[str, Any],
    output_dir: str | Path,
    *,
    stem: str | None = None,
) -> dict[str, str]:
    target_dir = Path(output_dir)
    target_dir.mkdir(parents=True, exist_ok=True)
    file_stem = stem or f"BIZ2x9h2_标准规则trace自动初判复核包_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
    json_path = target_dir / f"{file_stem}.json"
    md_path = target_dir / f"{file_stem}.md"
    csv_path = target_dir / f"{file_stem}_trace复核.csv"
    xlsx_path = target_dir / f"{file_stem}.xlsx"

    json_path.write_text(json.dumps(pack, ensure_ascii=False, indent=2), encoding="utf-8")
    md_path.write_text(build_trace_review_markdown(pack), encoding="utf-8")
    _write_csv(csv_path, list(pack.get("trace_review_rows") or []), TRACE_REVIEW_HEADERS)
    write_trace_review_workbook(pack, xlsx_path)
    return {
        "json": str(json_path),
        "markdown": str(md_path),
        "trace_review_csv": str(csv_path),
        "trace_review_xlsx": str(xlsx_path),
    }


def write_trace_review_workbook(pack: Mapping[str, Any], path: str | Path) -> None:
    workbook = Workbook()
    review_sheet = workbook.active
    review_sheet.title = TRACE_REVIEW_SHEET_NAME
    _write_rows(review_sheet, TRACE_REVIEW_HEADERS, list(pack.get("trace_review_rows") or []))
    _style_trace_review_sheet(review_sheet)

    blocked_sheet = workbook.create_sheet(BLOCKED_SHEET_NAME)
    _write_rows(blocked_sheet, BLOCKED_HEADERS, list(pack.get("blocked_rows") or []))
    _style_detail_sheet(blocked_sheet)

    detail_sheet = workbook.create_sheet(TRACE_DETAIL_SHEET_NAME)
    _write_rows(detail_sheet, TRACE_DETAIL_HEADERS, list(pack.get("trace_detail_rows") or []))
    _style_detail_sheet(detail_sheet)

    guide_sheet = workbook.create_sheet(GUIDE_SHEET_NAME)
    _write_guide_sheet(guide_sheet, pack)
    workbook.save(path)


def _build_review_row(row_id: str, trace: Mapping[str, Any], auto_decision: Mapping[str, str]) -> dict[str, Any]:
    ready = bool(trace.get("ready_for_manual_review"))
    source_trace = _source_trace(trace)
    quantity = trace.get("standard_rule_suggested_quantity") if ready else ""
    unit = trace.get("suggested_unit") if ready else _first_unit(trace.get("unit_options"))
    source_description = "采用系统 CAD 几何建议量，需人工复核扣减/合并规则" if ready else ""
    return {
        "复核行号": row_id,
        AUTO_CONCLUSION_COLUMN: auto_decision.get(AUTO_CONCLUSION_COLUMN, ""),
        AUTO_ACTION_COLUMN: auto_decision.get(AUTO_ACTION_COLUMN, ""),
        RISK_LEVEL_COLUMN: auto_decision.get(RISK_LEVEL_COLUMN, ""),
        AUTO_REASON_COLUMN: auto_decision.get(AUTO_REASON_COLUMN, ""),
        ADOPT_COLUMN: auto_decision.get(ADOPT_COLUMN, "待确认" if ready else "否"),
        REVIEW_COLUMN: "",
        MANUAL_QUANTITY_COLUMN: quantity or "",
        MANUAL_UNIT_COLUMN: unit or "",
        QUANTITY_SOURCE_COLUMN: source_description,
        MANUAL_NAME_COLUMN: trace.get("item_name", ""),
        MANUAL_FEATURE_COLUMN: _feature_placeholder(trace),
        DEDUCTION_REVIEW_COLUMN: "",
        ISSUE_COLUMN: "",
        "是否可进入人工复核": "是" if ready else "否",
        "trace状态": trace.get("trace_status", ""),
        "建议编号": trace.get("suggestion_key", ""),
        "标准项目编码": trace.get("item_code", ""),
        "标准项目名称": trace.get("item_name", ""),
        "标准单位": "、".join(trace.get("unit_options") or []),
        "标准规则类型": trace.get("quantity_formula_type", ""),
        "标准工程量计算规则": trace.get("quantity_rule_text", ""),
        "几何建议量": trace.get("geometry_quantity", ""),
        "几何单位": trace.get("geometry_unit", ""),
        "标准规则建议量": trace.get("standard_rule_suggested_quantity") or "",
        "建议单位": trace.get("suggested_unit", ""),
        "阻断原因": trace.get("block_reason", ""),
        "未解决事项": "；".join(trace.get("unresolved_requirements") or []),
        "CAD几何公式": source_trace.get("formula", "") or trace.get("calculation_trace", {}).get("geometry_formula", ""),
        "CAD来源图元行号": "、".join(str(item) for item in source_trace.get("sample_line_numbers") or []),
        "CAD图元类型统计": json.dumps(source_trace.get("sample_entity_types", {}), ensure_ascii=False),
    }


def _build_detail_row(row_id: str, trace: Mapping[str, Any]) -> dict[str, Any]:
    calculation_trace = dict(trace.get("calculation_trace") or {})
    return {
        "复核行号": row_id,
        "建议编号": trace.get("suggestion_key", ""),
        "trace状态": trace.get("trace_status", ""),
        "标准项目编码": trace.get("item_code", ""),
        "标准项目名称": trace.get("item_name", ""),
        "标准工程量计算规则": trace.get("quantity_rule_text", ""),
        "几何来源key": calculation_trace.get("geometry_source_key", ""),
        "几何公式": calculation_trace.get("geometry_formula", ""),
        "几何输入量": calculation_trace.get("geometry_input_quantity", ""),
        "几何输入单位": calculation_trace.get("geometry_input_unit", ""),
        "标准规则应用方式": calculation_trace.get("standard_rule_application", ""),
        "结果工程量": calculation_trace.get("result_quantity", ""),
        "结果单位": calculation_trace.get("result_unit", ""),
        "源追溯JSON": json.dumps(calculation_trace.get("source_calculation_trace", {}), ensure_ascii=False),
    }


def _build_blocked_rows(report: Mapping[str, Any]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for binding in report.get("bindings") or []:
        if _as_int(binding.get("compatible_trace_count")) > 0:
            continue
        rows.append(
            {
                "建议编号": binding.get("suggestion_key", ""),
                "绑定状态": binding.get("binding_status", ""),
                "文件名": binding.get("source_file", ""),
                "图层": binding.get("layer", ""),
                "块名": binding.get("block_name", ""),
                "业务提示": binding.get("business_hint", ""),
                "建议量类型": binding.get("quantity_kind", ""),
                "建议量": binding.get("suggested_quantity", ""),
                "单位": binding.get("suggested_unit", ""),
                "标准候选数": binding.get("standard_candidate_count", ""),
                "可复核trace数": binding.get("compatible_trace_count", ""),
                "说明": "；".join(binding.get("binding_notes") or []),
                "风险提示": "；".join(binding.get("risk_flags") or []),
            }
        )
    return rows


def _as_int(value: Any) -> int:
    try:
        return int(value)
    except (TypeError, ValueError):
        return 0


def _write_guide_sheet(sheet: Any, pack: Mapping[str, Any]) -> None:
    summary = pack.get("summary", {})
    rows = [
        ["用途", "本工作簿用于复核 CAD 几何建议量套用 GB/T 标准工程量规则后的 trace，并由系统先自动初判采用/不采用/需人工确认。"],
        ["trace 复核行", summary.get("trace_review_row_count", 0)],
        ["可进入人工复核", summary.get("ready_for_manual_review_count", 0)],
        ["阻断 trace 行", summary.get("blocked_trace_row_count", 0)],
        ["系统初判", f"`{AUTO_CONCLUSION_COLUMN}`、`{AUTO_ACTION_COLUMN}`、`{RISK_LEVEL_COLUMN}` 是系统预判结果，业务员优先处理“需人工确认”和高风险行。"],
        ["填写 1", f"`{ADOPT_COLUMN}` 已按系统初判预填，业务员确认后可覆盖。"],
        ["填写 2", f"`{REVIEW_COLUMN}`：确认无误填“通过”，否则填“有问题”。"],
        ["填写 3", f"`{MANUAL_FEATURE_COLUMN}`：按标准字段口径填写项目特征。"],
        ["填写 4", f"`{DEDUCTION_REVIEW_COLUMN}`：说明扣减、合并、展开面积等标准规则是否已复核。"],
        ["边界", "本包不直接导出最终四字段清单；所有采用行仍需人工确认后进入最终导出或报价。"],
    ]
    for row in rows:
        sheet.append(row)
    _style_header(sheet)
    sheet.column_dimensions["A"].width = 20
    sheet.column_dimensions["B"].width = 110
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def _style_trace_review_sheet(sheet: Any) -> None:
    _style_header(sheet)
    sheet.freeze_panes = "O2"
    sheet.auto_filter.ref = sheet.dimensions
    editable_fill = PatternFill("solid", fgColor="FFF2CC")
    required_fill = PatternFill("solid", fgColor="FCE4D6")
    auto_fill = PatternFill("solid", fgColor="D9EAD3")
    editable_headers = {
        ADOPT_COLUMN,
        REVIEW_COLUMN,
        MANUAL_QUANTITY_COLUMN,
        MANUAL_UNIT_COLUMN,
        QUANTITY_SOURCE_COLUMN,
        MANUAL_NAME_COLUMN,
        MANUAL_FEATURE_COLUMN,
        DEDUCTION_REVIEW_COLUMN,
        ISSUE_COLUMN,
    }
    required_headers = {ADOPT_COLUMN, REVIEW_COLUMN, MANUAL_QUANTITY_COLUMN, QUANTITY_SOURCE_COLUMN, MANUAL_FEATURE_COLUMN}
    for column_index, header in enumerate(TRACE_REVIEW_HEADERS, start=1):
        sheet.column_dimensions[get_column_letter(column_index)].width = _column_width(header)
        for column in sheet.iter_cols(min_col=column_index, max_col=column_index, min_row=2):
            for cell in column:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                if header in {AUTO_CONCLUSION_COLUMN, AUTO_ACTION_COLUMN, RISK_LEVEL_COLUMN, AUTO_REASON_COLUMN}:
                    cell.fill = auto_fill
                if header in editable_headers:
                    cell.fill = editable_fill
                if header in required_headers:
                    cell.fill = required_fill
    max_row = max(sheet.max_row, 2)
    _add_list_validation(sheet, ADOPT_COLUMN, '"待确认,是,否"', max_row)
    _add_list_validation(sheet, REVIEW_COLUMN, '"通过,有问题"', max_row)


def _style_detail_sheet(sheet: Any) -> None:
    _style_header(sheet)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for column_index, cell in enumerate(sheet[1], start=1):
        sheet.column_dimensions[get_column_letter(column_index)].width = _column_width(str(cell.value or ""))
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def _style_header(sheet: Any) -> None:
    fill = PatternFill("solid", fgColor="1F4E78")
    font = Font(color="FFFFFF", bold=True)
    for cell in sheet[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _write_rows(sheet: Any, headers: list[str], rows: list[Mapping[str, Any]]) -> None:
    sheet.append(headers)
    for row in rows:
        sheet.append([row.get(header, "") for header in headers])


def _write_csv(path: Path, rows: list[Mapping[str, Any]], headers: list[str]) -> None:
    with path.open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.DictWriter(handle, fieldnames=headers)
        writer.writeheader()
        for row in rows:
            writer.writerow({header: row.get(header, "") for header in headers})


def _add_list_validation(sheet: Any, header: str, formula: str, max_row: int) -> None:
    column = TRACE_REVIEW_HEADERS.index(header) + 1
    letter = get_column_letter(column)
    validation = DataValidation(type="list", formula1=formula, allow_blank=True)
    sheet.add_data_validation(validation)
    validation.add(f"{letter}2:{letter}{max_row}")


def _column_width(header: str) -> int:
    long_headers = {
        AUTO_REASON_COLUMN,
        MANUAL_FEATURE_COLUMN,
        DEDUCTION_REVIEW_COLUMN,
        QUANTITY_SOURCE_COLUMN,
        "标准工程量计算规则",
        "未解决事项",
        "CAD来源图元行号",
        "CAD图元类型统计",
        "说明",
        "风险提示",
        "源追溯JSON",
    }
    if header in long_headers:
        return 58
    if header in {AUTO_CONCLUSION_COLUMN, AUTO_ACTION_COLUMN, RISK_LEVEL_COLUMN}:
        return 18
    if header in {MANUAL_NAME_COLUMN, "标准项目名称", "项目名称（业务确认/可修改）"}:
        return 26
    if header in {"文件名", "几何来源key"}:
        return 36
    return max(12, min(28, len(header) * 2))


def _source_trace(trace: Mapping[str, Any]) -> dict[str, Any]:
    calculation_trace = trace.get("calculation_trace") or {}
    source_trace = calculation_trace.get("source_calculation_trace") or {}
    return dict(source_trace) if isinstance(source_trace, dict) else {}


def _feature_placeholder(trace: Mapping[str, Any]) -> str:
    fields = [str(item).strip() for item in trace.get("feature_fields") or [] if str(item).strip()]
    if not fields:
        return "待确认：标准表该项目未拆出项目特征字段或需人工补充"
    return "；".join(f"{field}：待确认" for field in fields)


def _first_unit(value: Any) -> str:
    units = list(value or [])
    return str(units[0]).strip() if units else ""


def _build_auto_decisions(traces: list[Mapping[str, Any]]) -> dict[int, dict[str, str]]:
    decisions: dict[int, dict[str, str]] = {}
    traces_by_suggestion: dict[str, list[Mapping[str, Any]]] = {}
    for trace in traces:
        traces_by_suggestion.setdefault(str(trace.get("suggestion_key") or ""), []).append(trace)

    for group in traces_by_suggestion.values():
        ready_traces = [trace for trace in group if bool(trace.get("ready_for_manual_review"))]
        selected_trace = _select_preferred_trace(ready_traces) if ready_traces else None
        for trace in group:
            if not bool(trace.get("ready_for_manual_review")):
                decisions[id(trace)] = _blocked_decision(_blocked_reason(trace))
                continue
            if selected_trace is not trace:
                decisions[id(trace)] = {
                    AUTO_CONCLUSION_COLUMN: "系统建议不采用：同一建议量已有优先标准候选",
                    AUTO_ACTION_COLUMN: "建议不采用",
                    RISK_LEVEL_COLUMN: "中",
                    AUTO_REASON_COLUMN: "同一 CAD 建议量对应多个标准项目，系统已按项目名称、规则类型和场景关键词选择更优候选，本行作为备选不采用。",
                    ADOPT_COLUMN: "否",
                }
                continue
            decisions[id(trace)] = _selected_decision(trace, len(ready_traces))
    return decisions


def _select_preferred_trace(ready_traces: list[Mapping[str, Any]]) -> Mapping[str, Any] | None:
    if not ready_traces:
        return None
    return sorted(ready_traces, key=_preference_score, reverse=True)[0]


def _preference_score(trace: Mapping[str, Any]) -> tuple[int, str]:
    code = str(trace.get("item_code") or "")
    name = str(trace.get("item_name") or "")
    source_text = _trace_search_text(trace)
    score = 0
    if code == "011502001":
        score += 90
    if code == "011302003" and "造型" in source_text:
        score += 95
    if code == "011302001":
        score += 85
    if code == "011102003":
        score += 90
    if code == "010904002" and "防水" in source_text:
        score += 92
    if "喷刷" in name or "油漆" in name or "涂料" in name:
        score -= 8
    if "展开面积" in str(trace.get("block_reason") or ""):
        score -= 100
    return (score, code)


def _selected_decision(trace: Mapping[str, Any], ready_count: int) -> dict[str, str]:
    risk_level = "中"
    action = "建议采用"
    adopt = "是"
    conclusion = "系统建议采用，待业务确认"
    reasons = [
        "标准项目已绑定 active GB/T 标准库，且几何量类型与标准库 quantity_rule 类型一致。",
        "系统已在同一 CAD 建议量的多个候选中选择优先标准项目。",
    ]
    quantity = _as_float(trace.get("standard_rule_suggested_quantity"))
    if ready_count > 1:
        risk_level = "中"
        reasons.append("同一 CAD 建议量仍存在其他标准候选，需业务确认项目名称和做法。")
    if _is_node_or_detail_trace(trace):
        risk_level = "高"
        action = "需人工确认"
        adopt = "待确认"
        conclusion = "系统不自动采用：疑似节点/大样或非施工范围"
        reasons.append("来源图纸或 CAD 量值疑似节点/大样/图例，不能自动认定为施工范围。")
    elif quantity is not None and str(trace.get("quantity_formula_type") or "") == "area" and quantity < 1:
        risk_level = "高"
        action = "需人工确认"
        adopt = "待确认"
        conclusion = "系统不自动采用：面积过小需确认施工范围"
        reasons.append("面积小于 1㎡，可能是节点、局部符号或非完整施工区域。")
    return {
        AUTO_CONCLUSION_COLUMN: conclusion,
        AUTO_ACTION_COLUMN: action,
        RISK_LEVEL_COLUMN: risk_level,
        AUTO_REASON_COLUMN: "；".join(reasons),
        ADOPT_COLUMN: adopt,
    }


def _blocked_decision(reason: str) -> dict[str, str]:
    return {
        AUTO_CONCLUSION_COLUMN: "系统自动阻断",
        AUTO_ACTION_COLUMN: "建议不采用",
        RISK_LEVEL_COLUMN: "高",
        AUTO_REASON_COLUMN: reason,
        ADOPT_COLUMN: "否",
    }


def _blocked_reason(trace: Mapping[str, Any]) -> str:
    block_reason = str(trace.get("block_reason") or "").strip()
    if block_reason:
        return block_reason
    return "该 trace 未通过标准规则兼容性校验，不能进入最终清单。"


def _is_node_or_detail_trace(trace: Mapping[str, Any]) -> bool:
    text = _trace_search_text(trace)
    return any(term in text for term in ("通用节点", "节点", "详图", "大样", "图例"))


def _trace_search_text(trace: Mapping[str, Any]) -> str:
    calculation_trace = trace.get("calculation_trace") or {}
    source_trace = calculation_trace.get("source_calculation_trace") or {}
    return " ".join(
        str(value or "")
        for value in [
            calculation_trace.get("geometry_source_key"),
            source_trace.get("source_key") if isinstance(source_trace, Mapping) else "",
            source_trace.get("mapping_business_hint") if isinstance(source_trace, Mapping) else "",
            source_trace.get("matched_reason") if isinstance(source_trace, Mapping) else "",
            trace.get("item_name"),
        ]
    )


def _as_float(value: Any) -> float | None:
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _count_values(rows: list[Mapping[str, Any]], column: str) -> dict[str, int]:
    counts: dict[str, int] = {}
    for row in rows:
        key = str(row.get(column) or "")
        counts[key] = counts.get(key, 0) + 1
    return dict(sorted(counts.items()))
