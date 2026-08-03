from __future__ import annotations

import csv
import json
import re
import unicodedata
from collections import Counter, defaultdict
from dataclasses import dataclass
from datetime import datetime, timezone
from difflib import SequenceMatcher
from pathlib import Path
from typing import Any, Iterable

from openpyxl import Workbook

from app.models.cost_item import COST_STATUS_ACTIVE, CostItem


REFERENCE_PRICE_FIELDS: tuple[tuple[str, str], ...] = (
    ("price", "主参考价"),
    ("client_tax_excluded_price", "对甲税前综合单价"),
    ("subcontract_composite_price", "劳务发包综合单价"),
    ("crew_benchmark_price", "班组标底税前价"),
)

DISPLAY_PRICE_FIELDS: tuple[tuple[str, str], ...] = (
    *REFERENCE_PRICE_FIELDS,
    ("client_labor_price", "对甲人工费"),
    ("client_main_material_price", "对甲主材费"),
    ("client_auxiliary_material_price", "对甲辅材费"),
    ("subcontract_labor_price", "劳务人工费"),
    ("subcontract_main_material_price", "劳务主材费"),
    ("subcontract_auxiliary_material_price", "劳务辅材费"),
)

_TEXT_SPLITTER = re.compile(r"[\s\-_—–,，、;；:：|+&()（）\[\]{}【】<>《》\"'“”‘’/\\]+")
_GENERIC_NAME_TERMS = ("工程", "施工", "项目", "报价")

_LENGTH_UNITS = {"m", "米", "延米", "延长米", "米长", "lm", "linearm"}
_AREA_UNITS = {"m2", "㎡", "平方米", "平方", "平米"}
_VOLUME_UNITS = {"m3", "㎥", "立方米", "方"}
_WEIGHT_UNITS = {"kg", "公斤", "千克", "t", "吨"}
_COUNT_UNITS = {"项", "套", "个", "件", "台", "头", "组", "根", "块", "张", "樘", "只", "盏", "户", "副", "片"}
_KNOWN_UNIT_KEYS = _LENGTH_UNITS | _AREA_UNITS | _VOLUME_UNITS | _WEIGHT_UNITS | _COUNT_UNITS


@dataclass(frozen=True)
class CostQualityIssue:
    severity: str
    category: str
    message: str
    suggestion: str
    cost_item_id: int | None = None
    related_item_ids: tuple[int, ...] = ()
    item_name: str | None = None
    spec: str | None = None
    unit: str | None = None
    price: float | None = None
    evidence: dict[str, Any] | None = None

    def as_dict(self) -> dict[str, Any]:
        return {
            "severity": self.severity,
            "category": self.category,
            "cost_item_id": self.cost_item_id,
            "related_item_ids": list(self.related_item_ids),
            "item_name": self.item_name,
            "spec": self.spec,
            "unit": self.unit,
            "price": self.price,
            "message": self.message,
            "suggestion": self.suggestion,
            "evidence": self.evidence or {},
        }


def _clean_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _canonical_text(value: Any) -> str:
    return unicodedata.normalize("NFKC", _clean_text(value)).lower()


def normalize_quality_text(value: Any) -> str:
    return _TEXT_SPLITTER.sub("", _canonical_text(value))


def _name_key(value: Any) -> str:
    text = normalize_quality_text(value)
    for term in _GENERIC_NAME_TERMS:
        text = text.replace(term, "")
    return text


def _unit_key(value: Any) -> str:
    return normalize_quality_text(value).replace("m²", "m2").replace("m³", "m3")


def unit_family(value: Any) -> str | None:
    key = _unit_key(value)
    if not key:
        return None
    if key in _LENGTH_UNITS:
        return "length"
    if key in _AREA_UNITS:
        return "area"
    if key in _VOLUME_UNITS:
        return "volume"
    if key in _WEIGHT_UNITS:
        return "weight"
    if key in _COUNT_UNITS:
        return f"count:{key}"
    return None


def _units_compatible(left: Any, right: Any) -> bool:
    left_key = _unit_key(left)
    right_key = _unit_key(right)
    if not left_key or not right_key:
        return True
    if left_key == right_key:
        return True
    left_family = unit_family(left)
    right_family = unit_family(right)
    return bool(left_family and right_family and left_family == right_family)


def _positive_number(value: Any) -> float | None:
    if value is None:
        return None
    try:
        number = float(value)
    except (TypeError, ValueError):
        return None
    if number <= 0:
        return None
    return round(number, 6)


def _item_id(item: CostItem) -> int:
    return int(getattr(item, "id", 0) or 0)


def _item_snapshot(item: CostItem) -> dict[str, Any]:
    data = {
        "id": _item_id(item),
        "category": _clean_text(getattr(item, "category", "")),
        "subcategory": _clean_text(getattr(item, "subcategory", "")),
        "item_name": _clean_text(getattr(item, "item_name", "")),
        "spec": _clean_text(getattr(item, "spec", "")),
        "unit": _clean_text(getattr(item, "unit", "")),
        "status": _clean_text(getattr(item, "status", "")),
        "source": _clean_text(getattr(item, "source", "")),
        "notes": _clean_text(getattr(item, "notes", "")),
    }
    for field_name, _ in DISPLAY_PRICE_FIELDS:
        value = getattr(item, field_name, None)
        data[field_name] = round(float(value), 6) if isinstance(value, (int, float)) else value
    return data


def _issue_for_item(
    item: CostItem,
    *,
    severity: str,
    category: str,
    message: str,
    suggestion: str,
    evidence: dict[str, Any] | None = None,
    related_item_ids: Iterable[int] = (),
) -> CostQualityIssue:
    return CostQualityIssue(
        severity=severity,
        category=category,
        cost_item_id=_item_id(item),
        related_item_ids=tuple(sorted({item_id for item_id in related_item_ids if item_id})),
        item_name=_clean_text(item.item_name),
        spec=_clean_text(item.spec),
        unit=_clean_text(item.unit),
        price=_positive_number(getattr(item, "price", None)),
        message=message,
        suggestion=suggestion,
        evidence=evidence or {},
    )


def _group_ids(items: Iterable[CostItem]) -> tuple[int, ...]:
    return tuple(sorted(_item_id(item) for item in items if _item_id(item)))


def _same_name_groups(items: list[CostItem]) -> dict[str, list[CostItem]]:
    groups: dict[str, list[CostItem]] = defaultdict(list)
    for item in items:
        key = _name_key(item.item_name)
        if key:
            groups[key].append(item)
    return {key: group for key, group in groups.items() if len(group) > 1}


def _duplicate_key(item: CostItem) -> tuple[str, str, str, str, str]:
    return (
        _name_key(item.category),
        _name_key(item.subcategory),
        _name_key(item.item_name),
        normalize_quality_text(item.spec),
        _unit_key(item.unit),
    )


def _collect_price_issues(items: list[CostItem]) -> list[CostQualityIssue]:
    issues: list[CostQualityIssue] = []
    named_reference_fields = ("client_tax_excluded_price", "subcontract_composite_price", "crew_benchmark_price")
    for item in items:
        main_price = _positive_number(getattr(item, "price", None))
        named_prices = [_positive_number(getattr(item, field_name, None)) for field_name in named_reference_fields]
        if main_price is None:
            issues.append(
                _issue_for_item(
                    item,
                    severity="high",
                    category="invalid_main_price",
                    message="active 成本条目的主参考价为空、为 0 或无效。",
                    suggestion="请成本部核定主参考价，核定前不建议作为演示样例。",
                    evidence={field_name: getattr(item, field_name, None) for field_name, _ in REFERENCE_PRICE_FIELDS},
                )
            )
        if not any(price is not None for price in named_prices):
            issues.append(
                _issue_for_item(
                    item,
                    severity="medium",
                    category="missing_named_reference_price",
                    message="对甲税前价、劳务发包价、班组标底价均缺少有效值。",
                    suggestion="请补齐至少一个业务可解释的来源价，便于后续演示报价依据。",
                    evidence={field_name: getattr(item, field_name, None) for field_name in named_reference_fields},
                )
            )
    return issues


def _collect_unit_issues(items: list[CostItem], same_name_groups: dict[str, list[CostItem]]) -> list[CostQualityIssue]:
    issues: list[CostQualityIssue] = []
    for item in items:
        unit = _clean_text(item.unit)
        if not unit:
            issues.append(
                _issue_for_item(
                    item,
                    severity="high",
                    category="missing_unit",
                    message="active 成本条目缺少单位。",
                    suggestion="请补齐单位，否则报价匹配和参考价解释都容易失真。",
                )
            )
            continue
        if unit_family(unit) is None and _unit_key(unit) not in _KNOWN_UNIT_KEYS:
            issues.append(
                _issue_for_item(
                    item,
                    severity="low",
                    category="unit_needs_review",
                    message="单位不在常见单位族中，建议人工确认是否为录入差异或特殊单位。",
                    suggestion="如属于 m/㎡/m³/kg/项/套/个等常见口径，请统一写法；特殊单位可保留。",
                    evidence={"unit_key": _unit_key(unit)},
                )
            )

    for group in same_name_groups.values():
        unit_labels = {_unit_label(item.unit) for item in group if _clean_text(item.unit)}
        if len(unit_labels) > 1:
            first = group[0]
            issues.append(
                CostQualityIssue(
                    severity="medium",
                    category="same_name_mixed_units",
                    cost_item_id=_item_id(first),
                    related_item_ids=_group_ids(group),
                    item_name=_clean_text(first.item_name),
                    spec=None,
                    unit=", ".join(sorted(unit_labels)),
                    price=None,
                    message="同名 active 成本条目存在多个单位口径。",
                    suggestion="请确认这是合理规格差异，还是单位录入不统一导致的匹配歧义。",
                    evidence={"units": sorted({_clean_text(item.unit) for item in group})},
                )
            )
    return issues


def _unit_label(unit: Any) -> str:
    family = unit_family(unit)
    if family:
        return family
    return _unit_key(unit) or "-"


def _collect_group_issues(items: list[CostItem], same_name_groups: dict[str, list[CostItem]]) -> list[CostQualityIssue]:
    issues: list[CostQualityIssue] = []

    duplicate_groups: dict[tuple[str, str, str, str, str], list[CostItem]] = defaultdict(list)
    for item in items:
        duplicate_groups[_duplicate_key(item)].append(item)
    for group in duplicate_groups.values():
        if len(group) <= 1:
            continue
        first = group[0]
        issues.append(
            CostQualityIssue(
                severity="high",
                category="exact_active_duplicate",
                cost_item_id=_item_id(first),
                related_item_ids=_group_ids(group),
                item_name=_clean_text(first.item_name),
                spec=_clean_text(first.spec),
                unit=_clean_text(first.unit),
                price=_positive_number(first.price),
                message="active 成本库中存在完全相同的名称、规格、单位和分类键。",
                suggestion="请成本部确认是否需要保留多条；如无业务差异，建议后续治理为一条。",
                evidence={"duplicate_count": len(group)},
            )
        )

    for group in same_name_groups.values():
        specs = {normalize_quality_text(item.spec) or "(空规格)" for item in group}
        if len(specs) > 1:
            first = group[0]
            issues.append(
                CostQualityIssue(
                    severity="low",
                    category="same_name_multi_spec",
                    cost_item_id=_item_id(first),
                    related_item_ids=_group_ids(group),
                    item_name=_clean_text(first.item_name),
                    spec=None,
                    unit=None,
                    price=None,
                    message="同名 active 成本条目存在多个规格，是演示条目切换的候选，也需要确认规格表达清晰。",
                    suggestion="演示时可用于展示同名不同规格切换；治理时请确认每条规格足够可辨认。",
                    evidence={"spec_count": len(specs), "item_count": len(group)},
                )
            )
        for item in group:
            if not _clean_text(item.spec):
                issues.append(
                    _issue_for_item(
                        item,
                        severity="medium",
                        category="missing_spec_on_multi_name",
                        message="同名多条 active 成本条目中存在空规格，容易影响预审阶段人工切换判断。",
                        suggestion="请补充规格/特征，至少说明适用范围或工艺差异。",
                        related_item_ids=_group_ids(group),
                    )
                )
            if not _clean_text(item.notes):
                issues.append(
                    _issue_for_item(
                        item,
                        severity="low",
                        category="missing_notes_on_multi_name",
                        message="同名多条 active 成本条目中存在空备注，演示解释信息不足。",
                        suggestion="建议补充计算规则、工作内容或价格来源备注。",
                        related_item_ids=_group_ids(group),
                    )
                )

    for item in items:
        if not _clean_text(item.spec):
            issues.append(
                _issue_for_item(
                    item,
                    severity="low",
                    category="missing_spec",
                    message="active 成本条目缺少规格/特征。",
                    suggestion="如该条目确实不区分规格可保留；否则建议补齐，便于报价依据解释。",
                )
            )
    return issues


def _similarity(left: CostItem, right: CostItem) -> float:
    left_name = _name_key(left.item_name)
    right_name = _name_key(right.item_name)
    if not left_name or not right_name:
        return 0.0
    name_ratio = SequenceMatcher(None, left_name, right_name).ratio()
    left_spec = normalize_quality_text(left.spec)
    right_spec = normalize_quality_text(right.spec)
    spec_ratio = SequenceMatcher(None, left_spec, right_spec).ratio() if left_spec and right_spec else 0.0
    return round(max(name_ratio, (name_ratio * 0.8) + (spec_ratio * 0.2)), 4)


def _collect_similar_item_issues(items: list[CostItem], *, max_pairs: int = 200) -> list[CostQualityIssue]:
    issues: list[CostQualityIssue] = []
    seen_pairs: set[tuple[int, int]] = set()
    sorted_items = sorted(items, key=_item_id)
    for index, left in enumerate(sorted_items):
        for right in sorted_items[index + 1 :]:
            if _duplicate_key(left) == _duplicate_key(right):
                continue
            if not _units_compatible(left.unit, right.unit):
                continue
            score = _similarity(left, right)
            if score < 0.88:
                continue
            pair = tuple(sorted((_item_id(left), _item_id(right))))
            if pair in seen_pairs:
                continue
            seen_pairs.add(pair)
            issues.append(
                CostQualityIssue(
                    severity="low",
                    category="similar_active_items",
                    cost_item_id=_item_id(left),
                    related_item_ids=pair,
                    item_name=_clean_text(left.item_name),
                    spec=_clean_text(left.spec),
                    unit=_clean_text(left.unit),
                    price=_positive_number(left.price),
                    message="发现名称/规格高度相似且单位兼容的 active 成本条目。",
                    suggestion="请人工确认是否为合理拆分，或后续需要合并/补充规格说明。",
                    evidence={
                        "similarity": score,
                        "left": _item_snapshot(left),
                        "right": _item_snapshot(right),
                    },
                )
            )
            if len(issues) >= max_pairs:
                return issues
    return issues


def _serialize_sync_run(run: Any) -> dict[str, Any]:
    if not run:
        return {}
    started_at = getattr(run, "started_at", None)
    finished_at = getattr(run, "finished_at", None)
    return {
        "id": getattr(run, "id", None),
        "source": getattr(run, "source", None),
        "status": getattr(run, "status", None),
        "requested_count": getattr(run, "requested_count", None),
        "synced_count": getattr(run, "synced_count", None),
        "started_at": started_at.isoformat(timespec="seconds") if hasattr(started_at, "isoformat") else started_at,
        "finished_at": finished_at.isoformat(timespec="seconds") if hasattr(finished_at, "isoformat") else finished_at,
        "message": getattr(run, "message", None),
        "error": getattr(run, "error", None),
    }


def _collect_sync_issues(active_count: int, sync_runs: list[Any]) -> tuple[list[CostQualityIssue], dict[str, Any]]:
    if not sync_runs:
        return (
            [
                CostQualityIssue(
                    severity="low",
                    category="rag_sync_not_checked",
                    message="未读取到成本库 RAG 同步记录。",
                    suggestion="演示前可手工查看同步记录，确认 active 成本库已经同步到 RAG。",
                    evidence={},
                )
            ],
            {},
        )
    latest = sync_runs[0]
    sync_summary = _serialize_sync_run(latest)
    issues: list[CostQualityIssue] = []
    status = _clean_text(getattr(latest, "status", ""))
    synced_count = getattr(latest, "synced_count", None)
    if status and status != "success":
        issues.append(
            CostQualityIssue(
                severity="medium",
                category="rag_sync_latest_not_success",
                message="最近一次成本库 RAG 同步不是成功状态。",
                suggestion="演示前建议在成本库页面检查同步记录；如需同步，由管理员手动触发。",
                evidence=sync_summary,
            )
        )
    if status == "success" and synced_count is not None and int(synced_count) != active_count:
        issues.append(
            CostQualityIssue(
                severity="medium",
                category="rag_sync_count_mismatch",
                message="最近一次成功同步数量与当前 active 成本条目数量不一致。",
                suggestion="这可能表示 active 成本库在同步后发生过变化；演示前建议管理员确认是否需要重新同步。",
                evidence={"active_count": active_count, **sync_summary},
            )
        )
    return issues, sync_summary


def _demo_case_from_item(case_type: str, item: CostItem, note: str, related_items: Iterable[CostItem] = ()) -> dict[str, Any]:
    return {
        "case_type": case_type,
        "cost_item_id": _item_id(item),
        "related_item_ids": _group_ids(related_items),
        "item_name": _clean_text(item.item_name),
        "spec": _clean_text(item.spec),
        "unit": _clean_text(item.unit),
        "price": _positive_number(item.price),
        "sample_demand": _sample_demand_text(item),
        "note": note,
    }


def _sample_demand_text(item: CostItem) -> str:
    parts = [_clean_text(item.item_name)]
    if _clean_text(item.spec):
        parts.append(_clean_text(item.spec))
    unit = _clean_text(item.unit) or "项"
    return f"{' '.join(parts)} 1{unit}"


def _build_demo_cases(items: list[CostItem], issues: list[CostQualityIssue], same_name_groups: dict[str, list[CostItem]]) -> list[dict[str, Any]]:
    cases: list[dict[str, Any]] = []
    healthy_items = [
        item
        for item in items
        if _positive_number(item.price) is not None and _clean_text(item.spec) and _clean_text(item.unit)
    ]
    if healthy_items:
        cases.append(_demo_case_from_item("exact_cost_reference", healthy_items[0], "演示成本库参考价命中与报价依据展示。"))

    multi_spec_group = next((group for group in same_name_groups.values() if len({normalize_quality_text(item.spec) for item in group}) > 1), None)
    if multi_spec_group:
        cases.append(
            _demo_case_from_item(
                "same_name_multi_spec_switch",
                multi_spec_group[0],
                "演示预审阶段同名不同规格成本条目切换。",
                related_items=multi_spec_group,
            )
        )

    compatible_item = next((item for item in healthy_items if unit_family(item.unit) in {"length", "area", "volume"}), None)
    if compatible_item:
        cases.append(_demo_case_from_item("unit_family_match", compatible_item, "演示单位族兼容命中，如 m/米、m2/㎡。"))

    issue_item_ids = {issue.cost_item_id for issue in issues if issue.severity in {"high", "medium"} and issue.cost_item_id}
    review_item = next((item for item in items if _item_id(item) in issue_item_ids), None)
    if review_item:
        cases.append(_demo_case_from_item("manual_review_needed", review_item, "演示时不建议直接作为正向样例，可用于成本部数据治理讨论。"))
    return cases


def analyze_cost_items_quality(
    items: Iterable[CostItem],
    *,
    sync_runs: Iterable[Any] | None = None,
    generated_at: datetime | None = None,
    max_similar_pairs: int = 200,
) -> dict[str, Any]:
    generated_at = generated_at or datetime.now(timezone.utc)
    all_items = list(items)
    active_items = [item for item in all_items if _clean_text(getattr(item, "status", "")) == COST_STATUS_ACTIVE]
    active_items.sort(key=_item_id)

    same_name_groups = _same_name_groups(active_items)
    issues: list[CostQualityIssue] = []
    issues.extend(_collect_price_issues(active_items))
    issues.extend(_collect_group_issues(active_items, same_name_groups))
    issues.extend(_collect_unit_issues(active_items, same_name_groups))
    issues.extend(_collect_similar_item_issues(active_items, max_pairs=max_similar_pairs))

    sync_issue_list, sync_summary = _collect_sync_issues(len(active_items), list(sync_runs or []))
    issues.extend(sync_issue_list)

    issue_dicts = [issue.as_dict() for issue in issues]
    severity_counts = Counter(issue.severity for issue in issues)
    category_counts = Counter(issue.category for issue in issues)
    demo_cases = _build_demo_cases(active_items, issues, same_name_groups)
    active_snapshots = [_item_snapshot(item) for item in active_items]

    return {
        "generated_at": generated_at.isoformat(timespec="seconds"),
        "scope": "cost_items.active",
        "input_item_count": len(all_items),
        "active_count": len(active_items),
        "issue_count": len(issue_dicts),
        "severity_counts": dict(sorted(severity_counts.items())),
        "category_counts": dict(category_counts.most_common()),
        "sync_summary": sync_summary,
        "issues": issue_dicts,
        "demo_cases": demo_cases,
        "active_items": active_snapshots,
    }


def build_markdown_report(result: dict[str, Any]) -> str:
    severity = result.get("severity_counts", {})
    category_counts = result.get("category_counts", {})
    lines = [
        "# BIZ-2k 成本库数据质量体检报告",
        "",
        f"- 生成时间：{result.get('generated_at')}",
        f"- 数据范围：{result.get('scope')}",
        f"- active 条目数：{result.get('active_count', 0)}",
        f"- 问题总数：{result.get('issue_count', 0)}",
        f"- 严重程度：high={severity.get('high', 0)} / medium={severity.get('medium', 0)} / low={severity.get('low', 0)}",
        "",
        "## 体检边界",
        "",
        "- 本报告只读分析 `cost_items.active`，不修改报价逻辑、不写数据库、不触发 RAG 同步。",
        "- 报告中的问题均为人工复核建议，不代表系统自动判错。",
        "",
        "## 问题分类统计",
        "",
        "| 分类 | 数量 |",
        "| --- | ---: |",
    ]
    for category, count in category_counts.items():
        lines.append(f"| {category} | {count} |")
    if not category_counts:
        lines.append("| 无 | 0 |")

    sync_summary = result.get("sync_summary") or {}
    lines.extend(["", "## 最近 RAG 同步记录", ""])
    if sync_summary:
        lines.extend(
            [
                f"- run_id：{sync_summary.get('id')}",
                f"- 状态：{sync_summary.get('status')}",
                f"- 请求数量：{sync_summary.get('requested_count')}",
                f"- 同步数量：{sync_summary.get('synced_count')}",
                f"- 开始时间：{sync_summary.get('started_at')}",
            ]
        )
    else:
        lines.append("- 未读取到同步记录。")

    lines.extend(["", "## 演示回归样例", "", "| 类型 | 成本条目 | 示例需求 | 说明 |", "| --- | --- | --- | --- |"])
    for case in result.get("demo_cases", []):
        label = f"#{case.get('cost_item_id')} {case.get('item_name') or '-'}"
        lines.append(f"| {case.get('case_type')} | {label} | {case.get('sample_demand') or '-'} | {case.get('note') or '-'} |")
    if not result.get("demo_cases"):
        lines.append("| 无 | - | - | 未找到适合自动推荐的样例 |")

    lines.extend(["", "## 问题明细（前 80 条）", "", "| 严重程度 | 分类 | 条目 | 说明 | 建议 |", "| --- | --- | --- | --- | --- |"])
    for issue in result.get("issues", [])[:80]:
        item_label = f"#{issue.get('cost_item_id')}" if issue.get("cost_item_id") else "-"
        if issue.get("item_name"):
            item_label = f"{item_label} {issue.get('item_name')}"
        lines.append(
            f"| {issue.get('severity')} | {issue.get('category')} | {item_label} | "
            f"{issue.get('message')} | {issue.get('suggestion')} |"
        )
    if not result.get("issues"):
        lines.append("| - | - | - | 未发现问题 | - |")
    lines.append("")
    return "\n".join(lines)


def build_demo_regression_pack(result: dict[str, Any]) -> str:
    lines = [
        "# BIZ-2k 演示回归包",
        "",
        f"- 生成时间：{result.get('generated_at')}",
        "- 边界：只做演示前人工回归，不修改业务口径。",
        "",
        "## 回归步骤",
        "",
        "1. 登录后台，确认 `PUBLIC_ACCESS_ENABLED=false` 且当前仍为内网验证环境。",
        "2. 打开 `/admin/cost-db`，用报告中的样例名称检索成本库 active 条目。",
        "3. 查看成本库同步记录，只确认最近状态，不自动触发同步。",
        "4. 在旧业务工作台提交样例需求，进入预审后查看成本库参考价与报价依据。",
        "5. 对同名不同规格样例，手工切换成本条目并确认单价/合计展示变化。",
        "6. 对需人工复核样例，只观察风险提示，不将其作为正式报价口径。",
        "7. 打回一条预审结果，确认后台报价运营详情可看到打回原因。",
        "",
        "## 推荐样例",
        "",
    ]
    for case in result.get("demo_cases", []):
        related = ", ".join(str(item_id) for item_id in case.get("related_item_ids", []) if item_id)
        lines.extend(
            [
                f"### {case.get('case_type')}",
                "",
                f"- 成本条目：#{case.get('cost_item_id')} {case.get('item_name')}",
                f"- 规格：{case.get('spec') or '-'}",
                f"- 单位/价格：{case.get('unit') or '-'} / {case.get('price') if case.get('price') is not None else '-'}",
                f"- 关联条目：{related or '-'}",
                f"- 示例需求：{case.get('sample_demand')}",
                f"- 说明：{case.get('note')}",
                "",
            ]
        )
    if not result.get("demo_cases"):
        lines.append("暂无自动推荐样例，请先查看数据质量报告。")
    return "\n".join(lines)


def write_issues_csv(result: dict[str, Any], path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    fieldnames = [
        "severity",
        "category",
        "cost_item_id",
        "related_item_ids",
        "item_name",
        "spec",
        "unit",
        "price",
        "message",
        "suggestion",
        "evidence",
    ]
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        for issue in result.get("issues", []):
            row = {field: issue.get(field) for field in fieldnames}
            row["related_item_ids"] = ",".join(str(item_id) for item_id in issue.get("related_item_ids", []))
            row["evidence"] = json.dumps(issue.get("evidence") or {}, ensure_ascii=False, sort_keys=True)
            writer.writerow(row)


def write_xlsx_report(result: dict[str, Any], path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()

    summary = workbook.active
    summary.title = "Summary"
    summary.append(["字段", "值"])
    for key in ("generated_at", "scope", "input_item_count", "active_count", "issue_count"):
        summary.append([key, result.get(key)])
    summary.append([])
    summary.append(["severity", "count"])
    for severity, count in (result.get("severity_counts") or {}).items():
        summary.append([severity, count])
    summary.append([])
    summary.append(["category", "count"])
    for category, count in (result.get("category_counts") or {}).items():
        summary.append([category, count])

    issues_sheet = workbook.create_sheet("Issues")
    issue_headers = [
        "severity",
        "category",
        "cost_item_id",
        "related_item_ids",
        "item_name",
        "spec",
        "unit",
        "price",
        "message",
        "suggestion",
    ]
    issues_sheet.append(issue_headers)
    for issue in result.get("issues", []):
        issues_sheet.append(
            [
                issue.get("severity"),
                issue.get("category"),
                issue.get("cost_item_id"),
                ",".join(str(item_id) for item_id in issue.get("related_item_ids", [])),
                issue.get("item_name"),
                issue.get("spec"),
                issue.get("unit"),
                issue.get("price"),
                issue.get("message"),
                issue.get("suggestion"),
            ]
        )

    demo_sheet = workbook.create_sheet("Demo Cases")
    demo_headers = ["case_type", "cost_item_id", "related_item_ids", "item_name", "spec", "unit", "price", "sample_demand", "note"]
    demo_sheet.append(demo_headers)
    for case in result.get("demo_cases", []):
        demo_sheet.append(
            [
                case.get("case_type"),
                case.get("cost_item_id"),
                ",".join(str(item_id) for item_id in case.get("related_item_ids", [])),
                case.get("item_name"),
                case.get("spec"),
                case.get("unit"),
                case.get("price"),
                case.get("sample_demand"),
                case.get("note"),
            ]
        )

    active_sheet = workbook.create_sheet("Active Items")
    active_headers = [
        "id",
        "category",
        "subcategory",
        "item_name",
        "spec",
        "unit",
        "price",
        "client_tax_excluded_price",
        "subcontract_composite_price",
        "crew_benchmark_price",
        "source",
        "notes",
    ]
    active_sheet.append(active_headers)
    for item in result.get("active_items", []):
        active_sheet.append([item.get(header) for header in active_headers])

    for sheet in workbook.worksheets:
        for column_cells in sheet.columns:
            max_length = max(len(str(cell.value or "")) for cell in column_cells)
            sheet.column_dimensions[column_cells[0].column_letter].width = min(max(max_length + 2, 10), 48)
    workbook.save(path)
