from __future__ import annotations

from datetime import datetime, timezone
from types import SimpleNamespace

from openpyxl import load_workbook

from app.models.cost_item import COST_SOURCE_AI_SUGGESTED, COST_SOURCE_MANUAL, COST_STATUS_ACTIVE, COST_STATUS_DRAFT
from app.services.cost_governance import (
    build_cost_governance_pack,
    build_governance_summary_markdown,
    write_governance_actions_csv,
    write_governance_actions_xlsx,
)


def _item(
    item_id: int,
    *,
    status: str = COST_STATUS_ACTIVE,
    source: str = COST_SOURCE_MANUAL,
    item_name: str = "wall paint",
    spec: str = "standard",
    unit: str = "m2",
    price: float | None = 10.0,
):
    return SimpleNamespace(
        id=item_id,
        category="test",
        subcategory="sub",
        item_name=item_name,
        spec=spec,
        unit=unit,
        price=price,
        status=status,
        source=source,
        created_at=datetime(2026, 5, 28, tzinfo=timezone.utc),
        updated_at=datetime(2026, 5, 28, tzinfo=timezone.utc),
    )


def test_governance_pack_upgrades_quoted_medium_issue_to_high():
    items = [_item(1), _item(2, status=COST_STATUS_DRAFT, source=COST_SOURCE_AI_SUGGESTED, price=0)]
    quality_result = {
        "issues": [
            {
                "severity": "medium",
                "category": "missing_named_reference_price",
                "cost_item_id": 1,
                "related_item_ids": [],
                "item_name": "wall paint",
                "spec": "standard",
                "unit": "m2",
                "price": 10,
                "message": "missing named price",
                "suggestion": "fill named price",
                "evidence": {"field": None},
            }
        ],
        "sync_summary": {"status": "success", "synced_count": 1},
    }
    quote_usage = {1: {"count": 2, "latest_used_at": datetime(2026, 5, 28, tzinfo=timezone.utc)}}

    pack = build_cost_governance_pack(items, quality_result, quote_usage=quote_usage)

    assert pack["summary"]["action_count"] == 2
    assert pack["summary"]["risk_counts"]["high"] == 2
    assert pack["trial_readiness"]["recommendation"] == "cleanup_before_trial"
    quoted_action = next(action for action in pack["actions"] if action["cost_item_id"] == 1)
    assert quoted_action["risk_level"] == "high"
    assert quoted_action["quote_usage_count"] == 2
    draft_action = next(action for action in pack["actions"] if action["cost_item_id"] == 2)
    assert "draft_missing_price_or_unit" in draft_action["issue_type"]
    assert "draft_duplicate_with_active" in draft_action["issue_type"]
    assert draft_action["owner"] == "cost_editor/cost_approver"


def test_governance_pack_writes_summary_csv_and_xlsx(tmp_path):
    items = [_item(10, item_name="floor leveling")]
    quality_result = {
        "issues": [
            {
                "severity": "low",
                "category": "missing_spec",
                "cost_item_id": 10,
                "related_item_ids": [],
                "item_name": "floor leveling",
                "spec": "",
                "unit": "m2",
                "price": 25,
                "message": "missing spec",
                "suggestion": "observe",
                "evidence": {},
            }
        ],
        "sync_summary": {"id": 1, "status": "success", "synced_count": 1},
    }
    pack = build_cost_governance_pack(items, quality_result, quote_usage={})

    summary = build_governance_summary_markdown(pack)
    csv_path = tmp_path / "actions.csv"
    xlsx_path = tmp_path / "actions.xlsx"
    write_governance_actions_csv(pack, csv_path)
    write_governance_actions_xlsx(pack, xlsx_path)

    assert "BIZ-2t Cost Data Governance Summary" in summary
    assert csv_path.read_text(encoding="utf-8-sig").startswith("issue_id,risk_level")
    workbook = load_workbook(xlsx_path)
    assert set(workbook.sheetnames) == {"Summary", "Actions"}
    assert workbook["Actions"]["A1"].value == "issue_id"
