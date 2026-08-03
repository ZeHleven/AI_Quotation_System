from __future__ import annotations

import json

from openpyxl import load_workbook

from app.services.dxf_trace_review_pack import (
    ADOPT_COLUMN,
    AUTO_ACTION_COLUMN,
    AUTO_CONCLUSION_COLUMN,
    MANUAL_FEATURE_COLUMN,
    REVIEW_COLUMN,
    RISK_LEVEL_COLUMN,
    TRACE_REVIEW_SHEET_NAME,
    build_trace_review_pack,
    write_trace_review_outputs,
    write_trace_review_workbook,
)


def _binding_report() -> dict[str, object]:
    return {
        "ok": True,
        "phase": "BIZ-2x-9f-9g-standard-item-binding-and-rule-trace",
        "bindings": [
            {
                "suggestion_key": "S-area",
                "binding_status": "blocked_multiple_standard_candidates_need_selection",
                "source_file": "sample.dxf",
                "layer": "D-顶面造型轮廓",
                "block_name": "",
                "business_hint": "天棚/吊顶面积候选",
                "quantity_kind": "area",
                "suggested_quantity": 12.5,
                "suggested_unit": "㎡",
                "standard_candidate_count": 2,
                "compatible_trace_count": 1,
                "binding_notes": ["存在多个标准项目候选"],
                "risk_flags": ["manual_review_required"],
            },
            {
                "suggestion_key": "S-socket",
                "binding_status": "blocked_out_of_scope_or_no_active_standard_candidate",
                "source_file": "sample.dxf",
                "layer": "C-平面插座",
                "block_name": "P-普通插座",
                "business_hint": "插座数量候选",
                "quantity_kind": "count",
                "suggested_quantity": 3,
                "suggested_unit": "个",
                "standard_candidate_count": 0,
                "compatible_trace_count": 0,
                "binding_notes": ["不强行套用 GB/T 50854"],
                "risk_flags": ["out_of_scope"],
            },
        ],
        "standard_rule_traces": [
            {
                "suggestion_key": "S-area",
                "item_code": "011302001",
                "item_name": "平面吊顶 | 天棚",
                "unit_options": ["m²", "㎡"],
                "feature_fields": ["吊顶形式", "龙骨材料种类"],
                "quantity_formula_type": "area",
                "quantity_rule_text": "按设计图示尺寸以水平投影面积计算",
                "trace_status": "standard_rule_trace_ready_for_manual_review",
                "ready_for_manual_review": True,
                "geometry_quantity": 12.5,
                "geometry_unit": "㎡",
                "standard_rule_suggested_quantity": 12.5,
                "suggested_unit": "㎡",
                "block_reason": "",
                "unresolved_requirements": ["standard_item_manual_confirmation"],
                "calculation_trace": {
                    "geometry_source_key": "sample|面积候选|D-顶面造型轮廓|",
                    "geometry_formula": "sum(CAD_area_mm2) * area_to_square_meter_factor",
                    "geometry_input_quantity": 12.5,
                    "geometry_input_unit": "㎡",
                    "standard_rule_application": "use_geometry_input_as_manual_review_candidate",
                    "result_quantity": 12.5,
                    "result_unit": "㎡",
                    "source_calculation_trace": {
                        "formula": "sum(CAD_area_mm2) * area_to_square_meter_factor",
                        "sample_line_numbers": [10, 20],
                        "sample_entity_types": {"LWPOLYLINE": 2},
                    },
                },
            },
            {
                "suggestion_key": "S-door",
                "item_code": "010801001",
                "item_name": "木质门",
                "unit_options": ["m²", "㎡"],
                "feature_fields": ["门代号及洞口尺寸"],
                "quantity_formula_type": "area",
                "quantity_rule_text": "按设计图示洞口尺寸以面积计算",
                "trace_status": "blocked_standard_rule_incompatible_with_geometry_kind",
                "ready_for_manual_review": False,
                "geometry_quantity": 2,
                "geometry_unit": "个",
                "standard_rule_suggested_quantity": None,
                "suggested_unit": "",
                "block_reason": "标准规则类型为 area，与几何建议量类型 count 不一致。",
                "unresolved_requirements": ["standard_item_manual_confirmation"],
                "calculation_trace": {},
            },
        ],
    }


def test_biz2x9h_builds_trace_review_pack_and_workbook(tmp_path):
    pack = build_trace_review_pack(_binding_report())

    assert pack["phase"] == "BIZ-2x-9h-2-standard-rule-trace-auto-decision-review-pack"
    assert pack["summary"]["trace_review_row_count"] == 2
    assert pack["summary"]["ready_for_manual_review_count"] == 1
    assert pack["summary"]["blocked_suggestion_row_count"] == 1
    assert pack["summary"]["auto_action_counts"]["建议采用"] == 1
    assert pack["summary"]["auto_action_counts"]["建议不采用"] == 1
    ready_row = pack["trace_review_rows"][0]
    blocked_row = pack["trace_review_rows"][1]
    assert ready_row[AUTO_CONCLUSION_COLUMN] == "系统建议采用，待业务确认"
    assert ready_row[AUTO_ACTION_COLUMN] == "建议采用"
    assert ready_row[RISK_LEVEL_COLUMN] == "中"
    assert ready_row[ADOPT_COLUMN] == "是"
    assert ready_row["确认工程量（业务填写）"] == 12.5
    assert "吊顶形式：待确认" in ready_row[MANUAL_FEATURE_COLUMN]
    assert blocked_row[AUTO_CONCLUSION_COLUMN] == "系统自动阻断"
    assert blocked_row[ADOPT_COLUMN] == "否"
    assert blocked_row["是否可进入人工复核"] == "否"

    workbook_path = tmp_path / "trace-review.xlsx"
    write_trace_review_workbook(pack, workbook_path)
    workbook = load_workbook(workbook_path)
    assert {TRACE_REVIEW_SHEET_NAME, "阻断项明细", "计算追溯明细", "填写说明"} <= set(workbook.sheetnames)
    assert workbook[TRACE_REVIEW_SHEET_NAME]["A1"].value == "复核行号"
    assert workbook[TRACE_REVIEW_SHEET_NAME]["B1"].value == AUTO_CONCLUSION_COLUMN
    assert workbook[TRACE_REVIEW_SHEET_NAME]["F1"].value == ADOPT_COLUMN
    assert workbook[TRACE_REVIEW_SHEET_NAME]["G1"].value == REVIEW_COLUMN


def test_biz2x9h_writes_outputs(tmp_path):
    pack = build_trace_review_pack(_binding_report())
    outputs = write_trace_review_outputs(pack, tmp_path, stem="trace-review")

    assert set(outputs) == {"json", "markdown", "trace_review_csv", "trace_review_xlsx"}
    assert json.loads((tmp_path / "trace-review.json").read_text(encoding="utf-8"))["summary"]["ready_for_manual_review_count"] == 1
    assert (tmp_path / "trace-review_trace复核.csv").read_text(encoding="utf-8-sig").startswith("复核行号")
