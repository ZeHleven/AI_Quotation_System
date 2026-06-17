from __future__ import annotations

from openpyxl import load_workbook

from app.services.drawing_quantity_confirmation import (
    ADOPT_COLUMN,
    MANUAL_FEATURE_COLUMN,
    MANUAL_NAME_COLUMN,
    MANUAL_QUANTITY_COLUMN,
    MANUAL_UNIT_COLUMN,
    QUANTITY_SOURCE_COLUMN,
    REVIEW_COLUMN,
    build_drawing_confirmation_pack,
    read_confirmation_workbook,
    validate_confirmation_rows,
    write_confirmation_workbook,
    write_final_quantity_workbook,
    write_validation_report,
)


def _standard_match_report() -> dict[str, object]:
    return {
        "phase": "BIZ-2x-4-standard-match-preview",
        "candidate_groups": [
            {
                "candidate_key": "BIZ2x4-0001",
                "standard_candidates": [
                    {
                        "item_code": "011102003",
                        "item_name": "块料楼地面",
                        "feature_fill_candidates": [
                            {
                                "field_name": "结合层厚度、材料种类及强度等级",
                                "candidate_value": "20mm 1:3水泥砂浆",
                                "status": "candidate_from_drawing_text",
                                "confidence": 0.82,
                                "evidence_text": "地面做法",
                            },
                            {
                                "field_name": "面层材料品种、规格",
                                "candidate_value": "",
                                "status": "missing_needs_manual_review",
                                "confidence": 0.0,
                                "evidence_text": "",
                            },
                        ],
                    }
                ],
            }
        ],
        "standard_item_candidates": [
            {
                "candidate_key": "BIZ2x4-0001",
                "standard_item_code": "011102003",
                "standard_item_name": "块料楼地面",
                "match_confidence": 0.91,
            }
        ],
    }


def _quantity_evidence_report() -> dict[str, object]:
    return {
        "phase": "BIZ-2x-5-quantity-evidence-preview",
        "quantity_candidates": [
            {
                "candidate_key": "BIZ2x4-0001",
                "source_file": "03.dxf",
                "source_row_number": 8,
                "source_name": "地砖地面做法",
                "source_spec_or_method": "20mm 1:3水泥砂浆",
                "standard_item_code": "011102003",
                "standard_item_name": "块料楼地面",
                "unit_options": ["m2", "㎡"],
                "quantity_status": "partial_quantity_evidence_needs_manual_measurement",
                "quantity_block_reason": "缺少面积证据",
                "quantity_formula_type": "area",
                "quantity_rule_text": "按设计图示尺寸以面积计算",
                "suggested_quantity": "",
                "suggested_unit": "",
                "evidence_summary": "dimension_text:20mm",
            }
        ],
        "quantity_evidence_rows": [
            {
                "candidate_key": "BIZ2x4-0001",
                "standard_item_code": "011102003",
                "standard_item_name": "块料楼地面",
                "evidence_type": "dimension_text",
                "value": "20",
                "unit": "mm",
                "is_direct_for_formula": False,
                "confidence": 0.9,
                "text": "20mm 1:3水泥砂浆",
                "source_file": "03.dxf",
                "layer": "地面",
                "layout": "",
                "block_name": "",
                "x": 1,
                "y": 2,
                "line_number": 30,
                "role_tags": ["plan"],
            }
        ],
    }


def test_biz2x6_builds_manual_confirmation_pack_and_workbook(tmp_path):
    pack = build_drawing_confirmation_pack(_standard_match_report(), _quantity_evidence_report())

    assert pack["summary"]["confirmation_row_count"] == 1
    row = pack["confirmation_rows"][0]
    assert row["确认行号"] == "BIZ2x6-0001"
    assert row[ADOPT_COLUMN] == "待确认"
    assert "面层材料品种、规格：待确认" in row[MANUAL_FEATURE_COLUMN]
    assert row["项目特征缺失字段"] == "面层材料品种、规格"

    workbook_path = tmp_path / "confirmation.xlsx"
    write_confirmation_workbook(pack, workbook_path)
    workbook = load_workbook(workbook_path)
    assert {"人工确认补量", "项目特征明细", "工程量证据明细", "填写说明"} <= set(workbook.sheetnames)
    assert workbook["人工确认补量"]["A1"].value == "确认行号"


def test_biz2x6_validates_completed_rows_and_exports_four_field_workbook(tmp_path):
    pack = build_drawing_confirmation_pack(_standard_match_report(), _quantity_evidence_report())
    row = pack["confirmation_rows"][0]
    row[ADOPT_COLUMN] = "是"
    row[REVIEW_COLUMN] = "通过"
    row[MANUAL_QUANTITY_COLUMN] = "12.5"
    row[MANUAL_UNIT_COLUMN] = "m2"
    row[QUANTITY_SOURCE_COLUMN] = "按 F-P03 地面铺装图手算面积"
    row[MANUAL_NAME_COLUMN] = "块料楼地面"
    row[MANUAL_FEATURE_COLUMN] = "结合层厚度、材料种类及强度等级：20mm 1:3水泥砂浆；面层材料品种、规格：玻化砖 600x600"

    workbook_path = tmp_path / "completed-confirmation.xlsx"
    write_confirmation_workbook(pack, workbook_path)
    rows = read_confirmation_workbook(workbook_path)
    validation = validate_confirmation_rows(rows)

    assert validation["ok"] is True
    assert validation["summary"]["adopted_final_row_count"] == 1
    assert validation["final_rows"][0] == {
        "项目名称": "块料楼地面",
        "项目特征": "结合层厚度、材料种类及强度等级：20mm 1:3水泥砂浆；面层材料品种、规格：玻化砖 600x600",
        "单位": "m2",
        "工程量": "12.5",
    }

    final_path = tmp_path / "final.xlsx"
    write_final_quantity_workbook(validation["final_rows"], final_path)
    final_workbook = load_workbook(final_path)
    final_sheet = final_workbook["最终四字段清单"]
    assert [cell.value for cell in final_sheet[1]] == ["项目名称", "项目特征", "单位", "工程量"]
    assert final_sheet["A2"].value == "块料楼地面"

    outputs = write_validation_report(validation, tmp_path, stem="validation")
    assert "final_xlsx" in outputs


def test_biz2x6_blocks_adopted_rows_with_placeholders_or_missing_quantity():
    pack = build_drawing_confirmation_pack(_standard_match_report(), _quantity_evidence_report())
    row = pack["confirmation_rows"][0]
    row[ADOPT_COLUMN] = "是"
    row[REVIEW_COLUMN] = "通过"
    row[MANUAL_QUANTITY_COLUMN] = ""
    row[QUANTITY_SOURCE_COLUMN] = ""

    validation = validate_confirmation_rows([row])

    assert validation["ok"] is False
    issue_text = "；".join(validation["issues"][0]["issues"])
    assert "人工工程量必须填写大于 0 的数字" in issue_text
    assert "项目特征仍包含待确认" in issue_text
