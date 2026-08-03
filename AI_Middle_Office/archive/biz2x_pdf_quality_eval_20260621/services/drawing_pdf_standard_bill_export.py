from __future__ import annotations

import csv
import json
from datetime import datetime
from pathlib import Path
from typing import Any, Iterable, Mapping, Sequence

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


PHASE = "BIZ-2x-pdf-standard-bill-preview"

BILL_HEADERS = [
    "row_no",
    "export_status",
    "three_field_status",
    "division",
    "standard_code",
    "item_code",
    "item_name",
    "feature",
    "unit",
    "quantity",
    "standard_item_name",
    "standard_match_score",
    "source_files",
    "evidence_ids",
    "review_note",
]

CHINESE_BILL_HEADERS = [
    "序号",
    "导出状态",
    "三字段状态",
    "分部分项",
    "标准来源",
    "项目编码",
    "项目名称",
    "项目特征",
    "计量单位",
    "工程量",
    "国标项目名称",
    "标准匹配分",
    "图纸来源",
    "证据编号",
    "复核说明",
]


def build_standard_bill_preview_report(
    v2_report: Mapping[str, Any],
    *,
    gate_report: Mapping[str, Any] | None = None,
) -> dict[str, Any]:
    human_rows = [dict(row) for row in v2_report.get("human_style_rows") or []]
    comparison_by_candidate = _comparison_by_candidate(v2_report)
    gate_passed = bool((gate_report or {}).get("can_enable_quantity"))
    rows: list[dict[str, Any]] = []
    for index, row in enumerate(human_rows, start=1):
        comparison = comparison_by_candidate.get(index, {})
        standard_item_code = str(row.get("standard_item_code") or "").strip()
        three_field_status = str(comparison.get("status") or "not_compared")
        mapped = bool(standard_item_code)
        export_status = _export_status(
            gate_passed=gate_passed,
            mapped=mapped,
            three_field_status=three_field_status,
        )
        rows.append(
            {
                "row_no": index,
                "export_status": export_status,
                "three_field_status": three_field_status,
                "division": row.get("division", ""),
                "standard_code": row.get("standard_code", ""),
                "item_code": standard_item_code,
                "item_name": row.get("item_name", ""),
                "feature": row.get("feature", ""),
                "unit": row.get("unit", ""),
                "quantity": "",
                "standard_item_name": row.get("standard_item_name", ""),
                "standard_match_score": row.get("standard_match_score", ""),
                "source_files": row.get("source_files", ""),
                "evidence_ids": row.get("evidence_ids", ""),
                "review_note": _review_note(row, comparison, gate_passed=gate_passed, mapped=mapped),
            }
        )

    mapped_count = sum(1 for row in rows if row.get("item_code"))
    final_candidate_count = sum(1 for row in rows if row.get("export_status") == "final_candidate")
    export_mode = "final_candidate" if gate_passed and mapped_count == len(rows) else "review_only"
    return {
        "ok": True,
        "phase": PHASE,
        "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "safe_for_final_standard_bill": export_mode == "final_candidate",
        "quantity_acceptance_enabled": False,
        "summary": {
            "standard_bill_row_count": len(rows),
            "standard_mapped_count": mapped_count,
            "standard_unmapped_count": len(rows) - mapped_count,
            "final_candidate_count": final_candidate_count,
            "review_only_count": len(rows) - final_candidate_count,
            "can_enable_quantity": gate_passed,
            "export_mode": export_mode,
            "quantity_status": "deferred_until_three_fields_accepted"
            if not gate_passed
            else "quantity_still_blank_until_quantity_stage",
        },
        "bill_rows": rows,
    }


def write_standard_bill_preview_outputs(
    report: Mapping[str, Any],
    output_dir: str | Path,
    *,
    stem: str | None = None,
) -> dict[str, str]:
    target = Path(output_dir)
    target.mkdir(parents=True, exist_ok=True)
    file_stem = stem or f"BIZ2x_PDF_standard_bill_preview_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
    json_path = target / f"{file_stem}.json"
    csv_path = target / f"{file_stem}.csv"
    markdown_path = target / f"{file_stem}.md"
    xlsx_path = target / f"{file_stem}.xlsx"
    outputs = {
        "json": str(json_path),
        "csv": str(csv_path),
        "markdown": str(markdown_path),
        "xlsx": str(xlsx_path),
    }
    payload = {**dict(report), "outputs": outputs}
    json_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    _write_csv(csv_path, payload.get("bill_rows") or [], BILL_HEADERS)
    markdown_path.write_text(_build_markdown(payload), encoding="utf-8")
    _write_workbook(xlsx_path, payload)
    return outputs


def _comparison_by_candidate(v2_report: Mapping[str, Any]) -> dict[int, Mapping[str, Any]]:
    result: dict[int, Mapping[str, Any]] = {}
    comparisons = ((v2_report.get("three_field_acceptance_report") or {}).get("comparison_rows") or [])
    for row in comparisons:
        candidate_no = _int(row.get("candidate_row_no"), 0)
        if candidate_no and candidate_no not in result:
            result[candidate_no] = row
    return result


def _export_status(*, gate_passed: bool, mapped: bool, three_field_status: str) -> str:
    if not mapped:
        return "standard_unmapped_review"
    if not gate_passed:
        return "review_only_three_field_gate_failed"
    if three_field_status != "matched_three_fields":
        return "review_only_three_field_not_matched"
    return "final_candidate"


def _review_note(
    row: Mapping[str, Any],
    comparison: Mapping[str, Any],
    *,
    gate_passed: bool,
    mapped: bool,
) -> str:
    notes = []
    if not gate_passed:
        notes.append("三字段门禁未通过，暂不可作为最终清单")
    if not mapped:
        notes.append("未匹配到国标标准库项目编码")
    status = str(comparison.get("status") or "")
    if status and status != "matched_three_fields":
        notes.append(f"三字段状态：{status}")
    issue = str(comparison.get("issue") or "").strip()
    if issue:
        notes.append(issue)
    local_note = str(row.get("review_note") or "").strip()
    if local_note:
        notes.append(local_note)
    return "；".join(notes)


def _build_markdown(report: Mapping[str, Any]) -> str:
    summary = report.get("summary") or {}
    lines = [
        "# BIZ-2x PDF Standard Bill Preview",
        "",
        f"- generated_at: {report.get('generated_at', '-')}",
        f"- export_mode: {summary.get('export_mode', '-')}",
        f"- rows: {summary.get('standard_bill_row_count', 0)}",
        f"- mapped: {summary.get('standard_mapped_count', 0)}",
        f"- unmapped: {summary.get('standard_unmapped_count', 0)}",
        f"- quantity_status: {summary.get('quantity_status', '-')}",
        "",
        "| row | status | item_code | item_name | unit | quantity | review_note |",
        "| ---: | --- | --- | --- | --- | --- | --- |",
    ]
    for row in (report.get("bill_rows") or [])[:80]:
        lines.append(
            "| "
            + " | ".join(
                [
                    _md(row.get("row_no")),
                    _md(row.get("export_status")),
                    _md(row.get("item_code")),
                    _md(row.get("item_name")),
                    _md(row.get("unit")),
                    _md(row.get("quantity")),
                    _md(row.get("review_note")),
                ]
            )
            + " |"
        )
    return "\n".join(lines) + "\n"


def _write_workbook(path: Path, report: Mapping[str, Any]) -> None:
    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "summary"
    _append_rows(summary_sheet, [["metric", "value"]])
    _append_rows(summary_sheet, [[key, _cell_value(value)] for key, value in (report.get("summary") or {}).items()])
    _style_sheet(summary_sheet)

    bill_sheet = workbook.create_sheet("standard_bill_preview")
    _append_rows(bill_sheet, [CHINESE_BILL_HEADERS])
    _append_rows(
        bill_sheet,
        [[_cell_value(row.get(header)) for header in BILL_HEADERS] for row in report.get("bill_rows") or []],
    )
    _style_sheet(bill_sheet)
    workbook.save(path)


def _write_csv(path: Path, rows: Iterable[Mapping[str, Any]], fieldnames: Sequence[str]) -> None:
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        for row in rows:
            writer.writerow({field: _cell_value(row.get(field)) for field in fieldnames})


def _append_rows(sheet: Any, rows: Sequence[Sequence[Any]]) -> None:
    for row in rows:
        sheet.append(list(row))


def _style_sheet(sheet: Any) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    if sheet.max_row >= 1:
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for col_idx, column_cells in enumerate(sheet.columns, start=1):
        values = [str(cell.value or "") for cell in column_cells[:200]]
        width = min(max([len(value) for value in values] + [10]) + 2, 60)
        sheet.column_dimensions[get_column_letter(col_idx)].width = width


def _cell_value(value: Any) -> Any:
    if isinstance(value, (dict, list, tuple)):
        return json.dumps(value, ensure_ascii=False)
    return value


def _md(value: Any) -> str:
    return str(value or "").replace("|", "\\|").replace("\n", "<br>")


def _int(value: Any, default: int = 0) -> int:
    try:
        return int(float(str(value).strip()))
    except (TypeError, ValueError):
        return default
