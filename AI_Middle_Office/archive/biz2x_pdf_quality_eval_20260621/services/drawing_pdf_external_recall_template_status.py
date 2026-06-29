from __future__ import annotations

import csv
import json
from collections import Counter
from datetime import datetime
from pathlib import Path
from typing import Any, Iterable, Mapping, Sequence

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.services.drawing_pdf_gap_recall_importer import load_external_recall_results


PHASE = "BIZ-2x-pdf-external-recall-template-status"

STATUS_HEADERS = [
    "row_no",
    "fill_status",
    "task_no",
    "gap_no",
    "gap_priority",
    "gap_type",
    "recommended_pass",
    "source_file",
    "page",
    "tile_id",
    "answer_item_name",
    "answer_unit",
    "item_hint",
    "spec_or_method",
    "suggested_unit",
    "text",
    "evidence_field_count",
    "missing_recommended_fields",
    "issue",
]

AGGREGATE_HEADERS = [
    "group_key",
    "input_row_count",
    "importable_row_count",
    "strong_importable_row_count",
    "weak_importable_row_count",
    "answer_only_count",
    "blank_task_count",
    "missing_item_hint_count",
    "missing_spec_or_method_count",
    "missing_suggested_unit_count",
    "missing_text_count",
]

EVIDENCE_FIELDS = ("item_hint", "spec_or_method", "suggested_unit", "text")
REFERENCE_FIELDS = ("answer_item_name", "answer_feature", "answer_unit", "target_item_name", "target_feature", "target_unit")


def build_external_recall_template_status(
    external_results: Mapping[str, Any],
    *,
    source_path: str | Path | None = None,
) -> dict[str, Any]:
    rows = _external_rows(external_results)
    status_rows = [_build_status_row(row, row_no=index) for index, row in enumerate(rows, start=1)]
    status_counts = dict(Counter(str(row.get("fill_status") or "") for row in status_rows))
    missing_field_counts = _missing_field_counts(status_rows)
    pass_summary_rows = _aggregate_rows(status_rows, group_fields=("recommended_pass",))
    source_summary_rows = _aggregate_rows(status_rows, group_fields=("source_file", "page"))
    summary = {
        "source_path": str(source_path or ""),
        "input_row_count": len(rows),
        "importable_row_count": sum(1 for row in status_rows if str(row.get("fill_status", "")).startswith("importable")),
        "strong_importable_row_count": status_counts.get("importable", 0),
        "weak_importable_row_count": status_counts.get("importable_weak", 0),
        "blank_task_count": status_counts.get("blank_task", 0),
        "answer_only_count": status_counts.get("answer_only_reference", 0),
        "status_counts": status_counts,
        "pass_counts": dict(Counter(str(row.get("recommended_pass") or "") for row in status_rows)),
        "missing_field_counts": missing_field_counts,
        "pass_group_count": len(pass_summary_rows),
        "source_page_group_count": len(source_summary_rows),
        "answer_columns_count_as_evidence": False,
        "ready_for_external_import": any(str(row.get("fill_status", "")).startswith("importable") for row in status_rows),
    }
    return {
        "ok": True,
        "phase": PHASE,
        "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "summary": summary,
        "pass_summary_rows": pass_summary_rows,
        "source_summary_rows": source_summary_rows,
        "status_rows": status_rows,
    }


def build_external_recall_template_status_from_path(path: str | Path) -> dict[str, Any]:
    source = Path(path)
    return build_external_recall_template_status(load_external_recall_results(source), source_path=source)


def write_external_recall_template_status_outputs(
    report: Mapping[str, Any],
    output_dir: str | Path,
    *,
    stem: str | None = None,
) -> dict[str, str]:
    target = Path(output_dir)
    target.mkdir(parents=True, exist_ok=True)
    file_stem = stem or f"BIZ2x_PDF_external_recall_template_status_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
    json_path = target / f"{file_stem}.json"
    csv_path = target / f"{file_stem}.csv"
    markdown_path = target / f"{file_stem}.md"
    xlsx_path = target / f"{file_stem}.xlsx"
    outputs = {
        "json": str(json_path),
        "csv": str(csv_path),
        "markdown": str(markdown_path),
        "xlsx": str(xlsx_path),
    }
    payload = {**dict(report), "outputs": outputs}
    json_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    _write_csv(csv_path, payload.get("status_rows") or [], STATUS_HEADERS)
    markdown_path.write_text(_build_markdown(payload), encoding="utf-8")
    _write_workbook(xlsx_path, payload)
    return outputs


def _external_rows(external_results: Mapping[str, Any]) -> list[dict[str, Any]]:
    rows = external_results.get("evidence_rows") or external_results.get("rows") or external_results.get("recall_evidence")
    if isinstance(rows, list):
        return [dict(row) for row in rows if isinstance(row, Mapping)]
    return []


def _build_status_row(row: Mapping[str, Any], *, row_no: int) -> dict[str, Any]:
    item_hint = _first(row, "item_hint", "evidence_item_hint", "raw_item_name", "item_name")
    spec = _first(row, "spec_or_method", "evidence_spec_or_method", "feature", "spec")
    unit = _first(row, "suggested_unit", "evidence_suggested_unit", "unit")
    text = _first(row, "text", "evidence_text", "normalized_text")
    evidence_values = {
        "item_hint": item_hint,
        "spec_or_method": spec,
        "suggested_unit": unit,
        "text": text,
    }
    evidence_field_count = sum(1 for value in evidence_values.values() if value)
    reference_count = sum(1 for key in REFERENCE_FIELDS if _first(row, key))
    missing = [field for field, value in evidence_values.items() if not value]
    fill_status, issue = _fill_status(evidence_values, evidence_field_count, reference_count)
    return {
        "row_no": row_no,
        "fill_status": fill_status,
        "task_no": _first(row, "task_no"),
        "gap_no": _first(row, "gap_no"),
        "gap_priority": _first(row, "gap_priority"),
        "gap_type": _first(row, "gap_type"),
        "recommended_pass": _first(row, "recommended_pass", "vision_pass", "prompt_mode"),
        "source_file": _first(row, "source_file", "candidate_source_files"),
        "page": _first(row, "page", "evidence_pages"),
        "tile_id": _first(row, "tile_id", "source_tile_id", "evidence_tiles"),
        "answer_item_name": _first(row, "answer_item_name", "target_item_name"),
        "answer_unit": _first(row, "answer_unit", "target_unit"),
        "item_hint": item_hint,
        "spec_or_method": spec,
        "suggested_unit": unit,
        "text": text,
        "evidence_field_count": evidence_field_count,
        "missing_recommended_fields": ",".join(missing),
        "issue": issue,
    }


def _fill_status(evidence_values: Mapping[str, str], evidence_field_count: int, reference_count: int) -> tuple[str, str]:
    has_item_or_spec = bool(evidence_values.get("item_hint") or evidence_values.get("spec_or_method"))
    has_text = bool(evidence_values.get("text"))
    if has_item_or_spec and has_text:
        return "importable", ""
    if evidence_field_count:
        return "importable_weak", "evidence is present but item/spec or source text is incomplete"
    if reference_count:
        return "answer_only_reference", "only answer reference columns are filled; no drawing evidence"
    return "blank_task", "no evidence fields filled"


def _build_markdown(report: Mapping[str, Any]) -> str:
    summary = report.get("summary") or {}
    lines = [
        "# BIZ-2x PDF External Recall Template Status",
        "",
        f"- generated_at: {report.get('generated_at', '-')}",
        f"- input_row_count: {summary.get('input_row_count', 0)}",
        f"- importable_row_count: {summary.get('importable_row_count', 0)}",
        f"- blank_task_count: {summary.get('blank_task_count', 0)}",
        f"- answer_only_count: {summary.get('answer_only_count', 0)}",
        f"- ready_for_external_import: {summary.get('ready_for_external_import', False)}",
        "",
        "## Pass Summary",
        "",
        "| pass | rows | importable | answer-only | blank | missing text |",
        "| --- | ---: | ---: | ---: | ---: | ---: |",
    ]
    for row in report.get("pass_summary_rows") or []:
        lines.append(
            "| "
            + " | ".join(
                [
                    _md(row.get("group_key")),
                    _md(row.get("input_row_count")),
                    _md(row.get("importable_row_count")),
                    _md(row.get("answer_only_count")),
                    _md(row.get("blank_task_count")),
                    _md(row.get("missing_text_count")),
                ]
            )
            + " |"
        )
    lines.extend(
        [
            "",
            "## Row Detail",
            "",
        "| row | status | pass | source | page | tile | answer item | missing | issue |",
        "| ---: | --- | --- | --- | --- | --- | --- | --- | --- |",
        ]
    )
    for row in (report.get("status_rows") or [])[:120]:
        lines.append(
            "| "
            + " | ".join(
                [
                    _md(row.get("row_no")),
                    _md(row.get("fill_status")),
                    _md(row.get("recommended_pass")),
                    _md(row.get("source_file")),
                    _md(row.get("page")),
                    _md(row.get("tile_id")),
                    _md(row.get("answer_item_name")),
                    _md(row.get("missing_recommended_fields")),
                    _md(row.get("issue")),
                ]
            )
            + " |"
        )
    return "\n".join(lines) + "\n"


def _write_workbook(path: Path, report: Mapping[str, Any]) -> None:
    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "summary"
    _append_rows(summary_sheet, [["metric", "value"]])
    _append_rows(summary_sheet, [[key, _cell_value(value)] for key, value in (report.get("summary") or {}).items()])
    _style_sheet(summary_sheet)

    status_sheet = workbook.create_sheet("template_status")
    _append_rows(status_sheet, [STATUS_HEADERS])
    _append_rows(
        status_sheet,
        [[_cell_value(row.get(header)) for header in STATUS_HEADERS] for row in report.get("status_rows") or []],
    )
    _style_sheet(status_sheet)

    pass_sheet = workbook.create_sheet("pass_summary")
    _append_rows(pass_sheet, [AGGREGATE_HEADERS])
    _append_rows(
        pass_sheet,
        [[_cell_value(row.get(header)) for header in AGGREGATE_HEADERS] for row in report.get("pass_summary_rows") or []],
    )
    _style_sheet(pass_sheet)

    source_sheet = workbook.create_sheet("source_page_summary")
    _append_rows(source_sheet, [AGGREGATE_HEADERS])
    _append_rows(
        source_sheet,
        [[_cell_value(row.get(header)) for header in AGGREGATE_HEADERS] for row in report.get("source_summary_rows") or []],
    )
    _style_sheet(source_sheet)
    workbook.save(path)


def _write_csv(path: Path, rows: Iterable[Mapping[str, Any]], fieldnames: Sequence[str]) -> None:
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        for row in rows:
            writer.writerow({field: _cell_value(row.get(field)) for field in fieldnames})


def _append_rows(sheet: Any, rows: Sequence[Sequence[Any]]) -> None:
    for row in rows:
        sheet.append(list(row))


def _style_sheet(sheet: Any) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    if sheet.max_row >= 1:
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for col_idx, column_cells in enumerate(sheet.columns, start=1):
        values = [str(cell.value or "") for cell in column_cells[:200]]
        width = min(max([len(value) for value in values] + [10]) + 2, 70)
        sheet.column_dimensions[get_column_letter(col_idx)].width = width


def _aggregate_rows(status_rows: Sequence[Mapping[str, Any]], *, group_fields: Sequence[str]) -> list[dict[str, Any]]:
    grouped: dict[str, list[Mapping[str, Any]]] = {}
    for row in status_rows:
        key = " | ".join(str(row.get(field) or "") for field in group_fields)
        grouped.setdefault(key, []).append(row)
    return [_aggregate_group(key, rows) for key, rows in sorted(grouped.items(), key=lambda item: item[0])]


def _aggregate_group(group_key: str, rows: Sequence[Mapping[str, Any]]) -> dict[str, Any]:
    status_counts = Counter(str(row.get("fill_status") or "") for row in rows)
    missing_counts = _missing_field_counts(rows)
    return {
        "group_key": group_key,
        "input_row_count": len(rows),
        "importable_row_count": sum(1 for row in rows if str(row.get("fill_status", "")).startswith("importable")),
        "strong_importable_row_count": status_counts.get("importable", 0),
        "weak_importable_row_count": status_counts.get("importable_weak", 0),
        "answer_only_count": status_counts.get("answer_only_reference", 0),
        "blank_task_count": status_counts.get("blank_task", 0),
        "missing_item_hint_count": missing_counts.get("item_hint", 0),
        "missing_spec_or_method_count": missing_counts.get("spec_or_method", 0),
        "missing_suggested_unit_count": missing_counts.get("suggested_unit", 0),
        "missing_text_count": missing_counts.get("text", 0),
    }


def _missing_field_counts(status_rows: Sequence[Mapping[str, Any]]) -> dict[str, int]:
    counts = {field: 0 for field in EVIDENCE_FIELDS}
    for row in status_rows:
        missing_fields = [field.strip() for field in str(row.get("missing_recommended_fields") or "").split(",") if field.strip()]
        for field in missing_fields:
            if field in counts:
                counts[field] += 1
    return counts


def _first(row: Mapping[str, Any], *keys: str) -> str:
    for key in keys:
        value = row.get(key)
        if value is None:
            continue
        text = str(value).strip()
        if text:
            return text
    return ""


def _cell_value(value: Any) -> Any:
    if isinstance(value, (dict, list, tuple)):
        return json.dumps(value, ensure_ascii=False)
    return value


def _md(value: Any) -> str:
    return str(value or "").replace("|", "\\|").replace("\n", "<br>")
