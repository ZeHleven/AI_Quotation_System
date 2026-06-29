from __future__ import annotations

import csv
import json
from datetime import datetime
from pathlib import Path
from typing import Any, Iterable, Mapping, Sequence

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


PHASE = "BIZ-2x-pdf-external-recall-template"

TEMPLATE_HEADERS = [
    "task_no",
    "gap_no",
    "gap_priority",
    "gap_type",
    "recommended_pass",
    "source_file",
    "page",
    "tile_id",
    "tile_type",
    "image_path",
    "answer_item_name",
    "answer_feature",
    "answer_unit",
    "evidence_role",
    "discipline",
    "item_hint",
    "space",
    "material_codes",
    "spec_or_method",
    "suggested_unit",
    "text",
    "normalized_text",
    "confidence",
    "model",
    "needs_manual_review",
    "reason",
]

TEMPLATE_HELPER_HEADERS = [
    "image_exists",
    "image_link",
    "fill_status_formula",
    "missing_required_fields_formula",
    "fill_hint",
]

WORKBOOK_TEMPLATE_HEADERS = TEMPLATE_HEADERS + TEMPLATE_HELPER_HEADERS

README_ROWS = [
    ["field", "how_to_fill"],
    [
        "answer_*",
        "Reference-only acceptance-answer columns. They help locate the target gap, but importers must not treat them as drawing evidence.",
    ],
    ["item_hint", "Fill with the item name recognized from visible drawing evidence."],
    ["spec_or_method", "Fill with visible material, specification, model, size, method, or construction-note evidence."],
    ["suggested_unit", "Fill with the suggested unit only when supported by drawing evidence; leave blank if uncertain."],
    ["text", "Fill with the original drawing/table/legend/node text that supports the candidate item."],
    ["confidence", "Optional confidence in the 0-1 range."],
    ["image_exists", "Workbook helper column. TRUE means the referenced image file existed when the template was generated."],
    ["image_link", "Workbook helper column. Click open_image to inspect the referenced drawing tile/page."],
    ["fill_status_formula", "Workbook helper formula only. Do not overwrite unless you intentionally remove helper checks."],
    ["Required minimum", "For a strong importable row, fill text and at least one of item_hint or spec_or_method."],
]


def build_external_recall_template(recall_plan: Mapping[str, Any]) -> dict[str, Any]:
    rows = []
    for plan_row in recall_plan.get("plan_rows") or []:
        rows.append(
            {
                "task_no": plan_row.get("task_no", ""),
                "gap_no": plan_row.get("gap_no", ""),
                "gap_priority": plan_row.get("gap_priority", ""),
                "gap_type": plan_row.get("gap_type", ""),
                "recommended_pass": plan_row.get("recommended_pass", ""),
                "source_file": plan_row.get("source_file", ""),
                "page": plan_row.get("page", ""),
                "tile_id": plan_row.get("tile_id", ""),
                "tile_type": plan_row.get("tile_type", ""),
                "image_path": plan_row.get("image_path", ""),
                "answer_item_name": plan_row.get("answer_item_name", ""),
                "answer_feature": plan_row.get("answer_feature", ""),
                "answer_unit": plan_row.get("answer_unit", ""),
                "evidence_role": "",
                "discipline": "",
                "item_hint": "",
                "space": "",
                "material_codes": "",
                "spec_or_method": "",
                "suggested_unit": "",
                "text": "",
                "normalized_text": "",
                "confidence": "",
                "model": "",
                "needs_manual_review": "true",
                "reason": "",
            }
        )
    return {
        "ok": True,
        "phase": PHASE,
        "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "summary": {
            "template_row_count": len(rows),
            "unique_image_count": len({str(row.get("image_path") or "") for row in rows if row.get("image_path")}),
            "pass_counts": _counts(row.get("recommended_pass") for row in rows),
            "safe_to_import_without_filling_evidence": False,
        },
        "template_rows": rows,
    }


def write_external_recall_template_outputs(
    report: Mapping[str, Any],
    output_dir: str | Path,
    *,
    stem: str | None = None,
) -> dict[str, str]:
    target = Path(output_dir)
    target.mkdir(parents=True, exist_ok=True)
    file_stem = stem or f"BIZ2x_PDF_external_recall_template_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
    json_path = target / f"{file_stem}.json"
    csv_path = target / f"{file_stem}.csv"
    markdown_path = target / f"{file_stem}.md"
    xlsx_path = target / f"{file_stem}.xlsx"
    outputs = {
        "json": str(json_path),
        "csv": str(csv_path),
        "markdown": str(markdown_path),
        "xlsx": str(xlsx_path),
    }
    payload = {**dict(report), "outputs": outputs}
    json_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    _write_csv(csv_path, payload.get("template_rows") or [], TEMPLATE_HEADERS)
    markdown_path.write_text(_build_markdown(payload), encoding="utf-8")
    _write_workbook(xlsx_path, payload)
    return outputs


def _build_markdown(report: Mapping[str, Any]) -> str:
    summary = report.get("summary") or {}
    lines = [
        "# BIZ-2x PDF External Recall Template",
        "",
        f"- generated_at: {report.get('generated_at', '-')}",
        f"- template_row_count: {summary.get('template_row_count', 0)}",
        f"- unique_image_count: {summary.get('unique_image_count', 0)}",
        f"- safe_to_import_without_filling_evidence: {summary.get('safe_to_import_without_filling_evidence', False)}",
        "",
        "| task | pass | source | page | tile | answer item | evidence item_hint | evidence text |",
        "| ---: | --- | --- | --- | --- | --- | --- | --- |",
    ]
    for row in (report.get("template_rows") or [])[:80]:
        lines.append(
            "| "
            + " | ".join(
                [
                    _md(row.get("task_no")),
                    _md(row.get("recommended_pass")),
                    _md(row.get("source_file")),
                    _md(row.get("page")),
                    _md(row.get("tile_id")),
                    _md(row.get("answer_item_name")),
                    _md(row.get("item_hint")),
                    _md(row.get("text")),
                ]
            )
            + " |"
        )
    return "\n".join(lines) + "\n"


def _write_workbook(path: Path, report: Mapping[str, Any]) -> None:
    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "summary"
    _append_rows(summary_sheet, [["metric", "value"]])
    _append_rows(summary_sheet, [[key, _cell_value(value)] for key, value in (report.get("summary") or {}).items()])
    _style_sheet(summary_sheet)

    readme_sheet = workbook.create_sheet("readme")
    _append_rows(readme_sheet, README_ROWS)
    _style_sheet(readme_sheet)

    template_sheet = workbook.create_sheet("external_recall_template")
    _append_rows(template_sheet, [WORKBOOK_TEMPLATE_HEADERS])
    _append_rows(
        template_sheet,
        [_template_workbook_row(row, row_idx=index) for index, row in enumerate(report.get("template_rows") or [], start=2)],
    )
    _style_sheet(template_sheet)
    _enhance_template_sheet(template_sheet)
    workbook.save(path)


def _write_csv(path: Path, rows: Iterable[Mapping[str, Any]], fieldnames: Sequence[str]) -> None:
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        for row in rows:
            writer.writerow({field: _cell_value(row.get(field)) for field in fieldnames})


def _append_rows(sheet: Any, rows: Sequence[Sequence[Any]]) -> None:
    for row in rows:
        sheet.append(list(row))


def _style_sheet(sheet: Any) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    if sheet.max_row >= 1:
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for col_idx, column_cells in enumerate(sheet.columns, start=1):
        values = [str(cell.value or "") for cell in column_cells[:200]]
        width = min(max([len(value) for value in values] + [10]) + 2, 65)
        sheet.column_dimensions[get_column_letter(col_idx)].width = width


def _enhance_template_sheet(sheet: Any) -> None:
    sheet.freeze_panes = "N2"
    _color_columns(sheet, ("answer_item_name", "answer_feature", "answer_unit"), "E7E6E6")
    _color_columns(sheet, ("item_hint", "spec_or_method", "suggested_unit", "text"), "FFF2CC")
    _color_columns(sheet, TEMPLATE_HELPER_HEADERS, "D9EAD3")
    _add_data_validations(sheet)
    _add_image_links(sheet)


def _add_data_validations(sheet: Any) -> None:
    max_row = max(sheet.max_row + 200, 500)
    validations = [
        ("evidence_role", DataValidation(type="list", formula1='"construction_note,material_schedule,legend,node_detail,dimension_note,manual_review_note"', allow_blank=True)),
        ("discipline", DataValidation(type="list", formula1='"decoration,electrical,plumbing,hvac,fire,unknown"', allow_blank=True)),
        ("suggested_unit", DataValidation(type="list", formula1='"m,m2,m3,set,each,item,group,point,box,kg,t"', allow_blank=True)),
        ("needs_manual_review", DataValidation(type="list", formula1='"true,false"', allow_blank=True)),
        ("confidence", DataValidation(type="decimal", operator="between", formula1="0", formula2="1", allow_blank=True)),
    ]
    for header, validation in validations:
        col_idx = _header_index(header)
        if not col_idx:
            continue
        sheet.add_data_validation(validation)
        col = get_column_letter(col_idx)
        validation.add(f"{col}2:{col}{max_row}")


def _color_columns(sheet: Any, headers: Sequence[str], fill_color: str) -> None:
    fill = PatternFill("solid", fgColor=fill_color)
    for header in headers:
        col_idx = _header_index(header)
        if not col_idx:
            continue
        for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=col_idx, max_col=col_idx):
            for cell in row:
                cell.fill = fill


def _template_workbook_row(row: Mapping[str, Any], *, row_idx: int) -> list[Any]:
    values = [_cell_value(row.get(header)) for header in TEMPLATE_HEADERS]
    image_path = str(row.get("image_path") or "")
    image_exists = Path(image_path).exists() if image_path else False
    image_link = "open_image" if image_path else ""
    item_col = get_column_letter(_header_index("item_hint"))
    spec_col = get_column_letter(_header_index("spec_or_method"))
    unit_col = get_column_letter(_header_index("suggested_unit"))
    text_col = get_column_letter(_header_index("text"))
    answer_item_col = get_column_letter(_header_index("answer_item_name"))
    answer_feature_col = get_column_letter(_header_index("answer_feature"))
    answer_unit_col = get_column_letter(_header_index("answer_unit"))
    fill_status_formula = (
        f'=IF(AND(LEN(TRIM({text_col}{row_idx}&""))>0,'
        f'LEN(TRIM({item_col}{row_idx}&{spec_col}{row_idx}&""))>0),"importable",'
        f'IF(LEN(TRIM({item_col}{row_idx}&{spec_col}{row_idx}&{unit_col}{row_idx}&{text_col}{row_idx}&""))>0,'
        f'"importable_weak",IF(LEN(TRIM({answer_item_col}{row_idx}&{answer_feature_col}{row_idx}&{answer_unit_col}{row_idx}&""))>0,'
        f'"answer_only_reference","blank_task")))'
    )
    missing_formula = (
        f'=TEXTJOIN(",",TRUE,'
        f'IF(LEN(TRIM({item_col}{row_idx}&{spec_col}{row_idx}&""))=0,"item_hint_or_spec",""),'
        f'IF(LEN(TRIM({text_col}{row_idx}&""))=0,"text",""))'
    )
    fill_hint = "Fill drawing evidence columns only; answer_* columns are reference-only."
    return values + [image_exists, image_link, fill_status_formula, missing_formula, fill_hint]


def _add_image_links(sheet: Any) -> None:
    image_path_col = _header_index("image_path")
    image_link_col = _header_index("image_link")
    if not image_path_col or not image_link_col:
        return
    for row_idx in range(2, sheet.max_row + 1):
        image_path = str(sheet.cell(row=row_idx, column=image_path_col).value or "").strip()
        if not image_path:
            continue
        cell = sheet.cell(row=row_idx, column=image_link_col)
        cell.value = "open_image"
        cell.hyperlink = image_path
        cell.style = "Hyperlink"


def _header_index(header: str) -> int:
    try:
        return WORKBOOK_TEMPLATE_HEADERS.index(header) + 1
    except ValueError:
        return 0


def _counts(values: Iterable[Any]) -> dict[str, int]:
    result: dict[str, int] = {}
    for value in values:
        key = str(value or "")
        result[key] = result.get(key, 0) + 1
    return result


def _cell_value(value: Any) -> Any:
    if isinstance(value, (dict, list, tuple)):
        return json.dumps(value, ensure_ascii=False)
    return value


def _md(value: Any) -> str:
    return str(value or "").replace("|", "\\|").replace("\n", "<br>")
