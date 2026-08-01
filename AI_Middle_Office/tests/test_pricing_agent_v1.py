from __future__ import annotations

import asyncio
import importlib.util
import json
from decimal import Decimal
from io import BytesIO
from pathlib import Path
from uuid import uuid4

import pytest
from alembic.migration import MigrationContext
from alembic.operations import Operations
from fastapi import FastAPI
from fastapi.testclient import TestClient
from openpyxl import Workbook
from sqlalchemy import Column, Integer, MetaData, String, Table, create_engine, inspect
from sqlalchemy.orm import sessionmaker
from sqlalchemy.pool import StaticPool

import app.main  # noqa: F401 - register all model metadata
from app.core.config import settings
from app.core.database import Base, get_db
from app.dependencies import get_current_user
from app.api.v1 import pricing_agent as pricing_agent_api
from app.api.v1.quote import _unpriced_requirement_placeholders
from app.models.account import Account, AccountMembership
from app.models.enterprise_quota import EnterpriseQuotaItem, EnterpriseQuotaVersion
from app.models.pricing_agent import (
    PricingAgentRun,
    PricingAgentRunLine,
    PricingArchiveFile,
    PricingArchiveLine,
)
from app.models.quote_job import QuoteJob
from app.models.quote_preview_draft import QuotePreviewDraft
from app.models.user import User
from app.schemas.pricing_agent import (
    PricingAgentCandidateSelectIn,
    PricingAgentManualPriceIn,
    PricingAgentRunCreateIn,
)
from app.services.pricing_agent import (
    PricingAgentError,
    confirm_pricing_agent_run_to_quote_draft,
    create_pricing_agent_run,
    get_pricing_agent_run,
    select_pricing_agent_candidate,
    serialize_pricing_agent_run,
    set_pricing_agent_manual_price,
)
from app.services.pricing_agent_hybrid import (
    PricingHybridDocument,
    PricingHybridHit,
    PricingHybridResult,
    search_pricing_hybrid,
)
from app.services.pricing_archive_parser import (
    normalize_text,
    normalize_unit,
    parse_demand_workbook,
    parse_priced_workbook,
)
from app.services.pricing_archive_storage import PricingArchiveStorageError, store_archive_bytes
from mcp_servers.tender_evidence.hybrid_client import (
    HybridReindexResult,
    HybridSearchHit,
    TenderHybridIndexStale,
)


MIGRATION_PATH = (
    Path(__file__).resolve().parents[1]
    / "alembic"
    / "versions"
    / "20260730_0075_add_pricing_agent_v1.py"
)
DECISION_MIGRATION_PATH = (
    Path(__file__).resolve().parents[1]
    / "alembic"
    / "versions"
    / "20260731_0077_persist_pricing_agent_decisions.py"
)
FRONTEND_ROOT = Path(__file__).resolve().parents[2] / "ai-web" / "src"


@pytest.fixture()
def db():
    engine = create_engine(
        "sqlite://",
        connect_args={"check_same_thread": False},
        poolclass=StaticPool,
    )
    Base.metadata.create_all(engine)
    session = sessionmaker(bind=engine)()
    try:
        yield session
    finally:
        session.close()
        Base.metadata.drop_all(engine)
        engine.dispose()


def _seed_user_account(db):
    user = User(
        username=f"pricing-agent-{uuid4().hex[:8]}",
        hashed_password="x",
        role="admin",
        role_version=1,
        quota=10,
        is_active=True,
    )
    db.add(user)
    db.flush()
    account = Account(
        account_uuid=str(uuid4()),
        account_code=f"pricing-agent-{uuid4().hex[:8]}",
        account_name="报价 Agent 测试账户",
        status="active",
        created_by=user.id,
    )
    db.add(account)
    db.flush()
    db.add(
        AccountMembership(
            account_id=account.id,
            user_id=user.id,
            member_role="owner",
            status="active",
            is_default=True,
            created_by=user.id,
        )
    )
    db.commit()
    return user, account


def _workbook_bytes() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "报价清单"
    sheet.append(["序号", "项目名称", "项目特征", "单位", "工程量", "综合单价", "合价"])
    sheet.append([1, "石膏板吊顶", "轻钢龙骨双层石膏板", "㎡", 20, 128.5, 2570])
    sheet.append([2, "乳胶漆墙面", "两遍腻子一底两面", "㎡", 100, None, 6800])
    sheet.append([None, "合计", None, None, None, None, 9370])
    stream = BytesIO()
    workbook.save(stream)
    workbook.close()
    return stream.getvalue()


def _seed_archive_line(db, *, account: Account, user: User):
    archive = PricingArchiveFile(
        archive_uuid=str(uuid4()),
        account_id=account.id,
        original_filename="历史报价.xlsx",
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        file_sha256="a" * 64,
        size_bytes=100,
        storage_backend="local",
        storage_object_name="pricing-agent-archives/test.xlsx",
        parser_version="pricing-archive-parser-v1",
        status="ready",
        indexed_row_count=1,
        rejected_row_count=0,
        created_by=user.id,
    )
    db.add(archive)
    db.flush()
    line = PricingArchiveLine(
        line_uuid=str(uuid4()),
        account_id=account.id,
        archive_file_id=archive.id,
        source_sheet="报价清单",
        source_row_index=2,
        sort_order=1,
        item_code="A-001",
        item_name="石膏板吊顶",
        specification="轻钢龙骨双层石膏板",
        unit="㎡",
        quantity=Decimal("20"),
        unit_price=Decimal("128.5"),
        total_price=Decimal("2570"),
        normalized_code="a001",
        normalized_name="石膏板吊顶",
        normalized_spec="轻钢龙骨双层石膏板",
        normalized_unit="m2",
        searchable=True,
        price_derivation="source_unit_price",
        fingerprint="b" * 64,
        raw_text="石膏板吊顶 轻钢龙骨双层石膏板 ㎡ 128.5",
        raw_row_json="{}",
    )
    db.add(line)
    db.commit()
    return archive, line


def _append_archive_line(
    db,
    *,
    archive: PricingArchiveFile,
    account: Account,
    source_row_index: int,
    item_name: str,
    specification: str | None,
    unit: str | None,
    unit_price: str,
    item_code: str | None = None,
):
    line = PricingArchiveLine(
        line_uuid=str(uuid4()),
        account_id=account.id,
        archive_file_id=archive.id,
        source_sheet="报价清单",
        source_row_index=source_row_index,
        sort_order=source_row_index,
        item_code=item_code,
        item_name=item_name,
        specification=specification,
        unit=unit,
        quantity=Decimal("1"),
        unit_price=Decimal(unit_price),
        total_price=Decimal(unit_price),
        normalized_code=normalize_text(item_code) or None,
        normalized_name=normalize_text(item_name),
        normalized_spec=normalize_text(specification) or None,
        normalized_unit=normalize_unit(unit),
        searchable=True,
        price_derivation="source_unit_price",
        fingerprint=uuid4().hex * 2,
        raw_text=f"{item_name} {specification or ''} {unit or ''} {unit_price}",
        raw_row_json="{}",
    )
    db.add(line)
    db.commit()
    return line


def _seed_enterprise_line(db, *, user: User):
    version = EnterpriseQuotaVersion(
        version_code=f"pricing-agent-{uuid4().hex[:8]}",
        version_name="报价 Agent 企业定额",
        status="active",
        is_active=True,
        created_by=user.id,
    )
    db.add(version)
    db.flush()
    item = EnterpriseQuotaItem(
        version_id=version.id,
        quota_code="E-001",
        item_name="乳胶漆墙面",
        specification="两遍腻子一底两面",
        unit="㎡",
        unit_price=66.8,
        sort_order=1,
    )
    db.add(item)
    db.commit()
    return version, item


def test_archive_parser_uses_fixed_aliases_and_derives_price_from_total():
    parsed = parse_priced_workbook(_workbook_bytes(), "历史报价.xlsx")
    assert parsed.summary["mapping_mode"] == "automatic_fixed_system_fields"
    assert parsed.summary["indexed_row_count"] == 2
    assert [line.item_name for line in parsed.lines] == ["石膏板吊顶", "乳胶漆墙面"]
    assert parsed.lines[0].unit_price == Decimal("128.500000")
    assert parsed.lines[0].normalized_unit == "m2"
    assert parsed.lines[1].unit_price == Decimal("68.000000")
    assert parsed.lines[1].price_derivation == "derived_from_total"


def test_demand_parser_does_not_require_a_price_column():
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["清单编码", "项目名称", "项目特征", "单位", "工程量"])
    sheet.append(["Q-1", "轻钢龙骨隔墙", "75系列双层石膏板", "㎡", 12.5])
    stream = BytesIO()
    workbook.save(stream)
    workbook.close()
    parsed = parse_demand_workbook(stream.getvalue(), "需求清单.xlsx")
    assert parsed.summary["line_count"] == 1
    assert parsed.lines[0] == {
        "row_key": f"{sheet.title}:2",
        "source_sheet": sheet.title,
        "source_row_index": 2,
        "item_code": "Q-1",
        "item_name": "轻钢龙骨隔墙",
        "specification": "75系列双层石膏板",
        "quantity": "12.500000",
        "unit": "㎡",
    }


def test_demand_parser_skips_summary_and_section_rows_without_user_mapping():
    workbook = Workbook()
    detail = workbook.active
    detail.title = "装修工程量清单"
    detail.append(["项目名称", "项目特征", "单位", "工程量"])
    detail.append(["一、装修部分", None, None, None])
    detail.append(["地砖铺贴", "800*800mm", "㎡", 25])
    summary = workbook.create_sheet("汇总表")
    summary.append(["项目", "金额"])
    summary.append(["序号", None])
    summary.append([1, 100])
    summary.append([2, 200])
    stream = BytesIO()
    workbook.save(stream)
    workbook.close()

    parsed = parse_demand_workbook(stream.getvalue(), "需求清单.xlsx")

    assert [line["item_name"] for line in parsed.lines] == ["地砖铺贴"]
    assert parsed.summary["line_count"] == 1
    assert parsed.summary["skipped_non_item_row_count"] == 4


def test_local_storage_permission_error_is_reported_as_storage_error(monkeypatch, tmp_path):
    previous = {
        "pricing_agent_archive_storage_backend": settings.pricing_agent_archive_storage_backend,
        "pricing_agent_archive_local_root": settings.pricing_agent_archive_local_root,
    }
    object.__setattr__(settings, "pricing_agent_archive_storage_backend", "local")
    object.__setattr__(settings, "pricing_agent_archive_local_root", str(tmp_path))

    def denied(*_args, **_kwargs):
        raise PermissionError("denied")

    monkeypatch.setattr(Path, "write_bytes", denied)
    try:
        with pytest.raises(PricingArchiveStorageError, match="PRICING_ARCHIVE_LOCAL_WRITE_FAILED"):
            store_archive_bytes(
                content=b"x",
                account_uuid=str(uuid4()),
                file_sha256="f" * 64,
                suffix=".xlsx",
                content_type="application/octet-stream",
            )
    finally:
        for name, value in previous.items():
            object.__setattr__(settings, name, value)


def test_exact_mode_uses_archive_only_and_never_adds_ai_or_lexical(db, monkeypatch):
    user, account = _seed_user_account(db)
    _seed_archive_line(db, account=account, user=user)

    async def forbidden_ai(**_kwargs):
        pytest.fail("准确模式不能调用行业 AI")

    def forbidden_hybrid(**_kwargs):
        pytest.fail("准确模式不能调用关键词或向量混合检索")

    monkeypatch.setattr("app.services.pricing_agent.estimate_industry_prices", forbidden_ai)
    monkeypatch.setattr(
        "app.services.pricing_agent_retrieval.search_pricing_hybrid",
        forbidden_hybrid,
    )
    payload = PricingAgentRunCreateIn(
        mode="exact",
        sources=["archive"],
        context={"city": "杭州", "project_type": "写字楼", "decoration_level": "精装"},
        lines=[
            {
                "row_key": "row-1",
                "item_code": "A-001",
                "item_name": "石膏板吊顶",
                "specification": "轻钢龙骨双层石膏板",
                "quantity": "10",
                "unit": "㎡",
            }
        ],
    )
    run = asyncio.run(create_pricing_agent_run(db, current_user=user, payload=payload))
    result = serialize_pricing_agent_run(run)["result"]["lines"][0]
    assert result["selected_source"] == "archive"
    assert result["match_type"] == "code_exact"
    assert result["unit_price"] == "128.500000"
    assert result["total_price"] == "1285.000000"
    assert result["query_plan"]["channels"] == ["exact"]
    assert result["source_evidence"]["archive"]["channel_status"]["lexical"] == "not_used"


def test_exact_mode_rejects_same_name_when_specification_is_different(db):
    user, account = _seed_user_account(db)
    _seed_archive_line(db, account=account, user=user)
    payload = PricingAgentRunCreateIn(
        mode="exact",
        sources=["archive"],
        context={"city": "杭州", "project_type": "写字楼", "decoration_level": "精装"},
        lines=[
            {
                "row_key": "different-spec",
                "item_name": "石膏板吊顶",
                "specification": "完全不同的测试规格",
                "quantity": "10",
                "unit": "㎡",
            }
        ],
    )

    run = asyncio.run(create_pricing_agent_run(db, current_user=user, payload=payload))
    result = serialize_pricing_agent_run(run)["result"]["lines"][0]

    assert result["status"] == "unpriced"
    assert result["unit_price"] is None
    assert result["candidates"] == []
    assert result["requires_review"] is False


def test_exact_mode_uses_unique_name_spec_unit_without_false_review(db):
    user, account = _seed_user_account(db)
    archive, _line = _seed_archive_line(db, account=account, user=user)
    _append_archive_line(
        db,
        archive=archive,
        account=account,
        source_row_index=3,
        item_name="石膏板吊顶",
        specification="轻钢龙骨单层石膏板",
        unit="㎡",
        unit_price="98.50",
    )
    payload = PricingAgentRunCreateIn(
        mode="exact",
        sources=["archive"],
        context={"city": "杭州", "project_type": "写字楼", "decoration_level": "精装"},
        lines=[
            {
                "row_key": "unique-spec",
                "item_name": "石膏板吊顶",
                "specification": "轻钢龙骨双层石膏板",
                "quantity": "10",
                "unit": "㎡",
            }
        ],
    )

    run = asyncio.run(create_pricing_agent_run(db, current_user=user, payload=payload))
    result = serialize_pricing_agent_run(run)["result"]["lines"][0]

    assert result["unit_price"] == "128.500000"
    assert result["requires_review"] is False
    assert len(result["candidates"]) == 1
    assert result["candidates"][0]["specification"] == "轻钢龙骨双层石膏板"


def test_exact_mode_marks_conflicting_exact_prices_for_review(db):
    user, account = _seed_user_account(db)
    archive, _line = _seed_archive_line(db, account=account, user=user)
    _append_archive_line(
        db,
        archive=archive,
        account=account,
        source_row_index=3,
        item_name="石膏板吊顶",
        specification="轻钢龙骨双层石膏板",
        unit="㎡",
        unit_price="135.00",
    )
    payload = PricingAgentRunCreateIn(
        mode="exact",
        sources=["archive"],
        context={"city": "杭州", "project_type": "写字楼", "decoration_level": "精装"},
        lines=[
            {
                "row_key": "conflicting-prices",
                "item_name": "石膏板吊顶",
                "specification": "轻钢龙骨双层石膏板",
                "quantity": "10",
                "unit": "㎡",
            }
        ],
    )

    run = asyncio.run(create_pricing_agent_run(db, current_user=user, payload=payload))
    result = serialize_pricing_agent_run(run)["result"]["lines"][0]

    assert result["status"] == "unpriced"
    assert result["unit_price"] is None
    assert result["requires_review"] is True
    assert {candidate["unit_price"] for candidate in result["candidates"]} == {
        "128.500000",
        "135.000000",
    }


def test_manual_candidate_selection_is_persisted_and_restored(db):
    user, account = _seed_user_account(db)
    archive, _line = _seed_archive_line(db, account=account, user=user)
    selected_archive_line = _append_archive_line(
        db,
        archive=archive,
        account=account,
        source_row_index=3,
        item_name="石膏板吊顶",
        specification="轻钢龙骨双层石膏板",
        unit="㎡",
        unit_price="135.00",
    )
    payload = PricingAgentRunCreateIn(
        mode="exact",
        sources=["archive"],
        context={"city": "杭州", "project_type": "写字楼", "decoration_level": "精装"},
        lines=[
            {
                "row_key": "persisted-selection",
                "item_name": "石膏板吊顶",
                "specification": "轻钢龙骨双层石膏板",
                "quantity": "10",
                "unit": "㎡",
            }
        ],
    )
    run = asyncio.run(create_pricing_agent_run(db, current_user=user, payload=payload))
    before = serialize_pricing_agent_run(run)["result"]["lines"][0]

    updated = select_pricing_agent_candidate(
        db,
        current_user=user,
        run_uuid=run.run_uuid,
        line_uuid=before["line_uuid"],
        payload=PricingAgentCandidateSelectIn(
            source="archive",
            source_record_id=selected_archive_line.line_uuid,
        ),
    )
    db.commit()
    db.expire_all()

    restored = get_pricing_agent_run(
        db,
        current_user=user,
        run_uuid=updated.run_uuid,
    )
    data = serialize_pricing_agent_run(restored)
    line = data["result"]["lines"][0]
    stored_line = db.query(PricingAgentRunLine).filter(
        PricingAgentRunLine.line_uuid == line["line_uuid"]
    ).one()

    assert line["unit_price"] == "135.000000"
    assert line["total_price"] == "1350.000000"
    assert line["requires_review"] is False
    assert line["manual_candidate_selected"] is True
    assert line["selection_origin"] == "manual"
    assert line["decision_revision"] == 1
    assert data["summary"]["priced_count"] == 1
    assert data["summary"]["requires_review_count"] == 0
    assert stored_line.selected_candidate_json is not None
    assert stored_line.manual_selected_by == user.id
    assert stored_line.manual_selected_at is not None
    assert "selection_history" in stored_line.evidence_json


def test_manual_price_is_persisted_with_actor_time_and_revision_history(db):
    user, _account = _seed_user_account(db)
    payload = PricingAgentRunCreateIn(
        mode="exact",
        sources=["archive"],
        context={"city": "杭州市", "project_type": "写字楼", "decoration_level": "精装"},
        lines=[
            {
                "row_key": "manual-price-row",
                "item_name": "无历史价格的新项目",
                "specification": "首次出现",
                "quantity": "4",
                "unit": "项",
            }
        ],
    )
    run = asyncio.run(create_pricing_agent_run(db, current_user=user, payload=payload))
    before = serialize_pricing_agent_run(run)["result"]["lines"][0]
    assert before["unit_price"] is None

    set_pricing_agent_manual_price(
        db,
        current_user=user,
        run_uuid=run.run_uuid,
        line_uuid=before["line_uuid"],
        payload=PricingAgentManualPriceIn(
            unit_price=Decimal("12.50"),
            reason="首次人工测价",
        ),
    )
    db.commit()
    db.expire_all()

    restored = get_pricing_agent_run(
        db,
        current_user=user,
        run_uuid=run.run_uuid,
    )
    first = serialize_pricing_agent_run(restored)
    line = first["result"]["lines"][0]
    stored_line = db.query(PricingAgentRunLine).filter(
        PricingAgentRunLine.line_uuid == line["line_uuid"]
    ).one()
    first_evidence = json.loads(stored_line.evidence_json)

    assert line["selected_source"] == "manual"
    assert line["source_label"] == "人工补价"
    assert line["match_type"] == "manual_price"
    assert line["unit_price"] == "12.500000"
    assert line["total_price"] == "50.000000"
    assert line["manual_price_entered"] is True
    assert line["manual_selected_by"] == user.id
    assert line["manual_selected_at"] is not None
    assert line["decision_revision"] == 1
    assert first["summary"]["source_counts"]["manual"] == 1
    assert first_evidence["selection_history"][0]["manual_price"]["reason"] == "首次人工测价"

    set_pricing_agent_manual_price(
        db,
        current_user=user,
        run_uuid=run.run_uuid,
        line_uuid=line["line_uuid"],
        payload=PricingAgentManualPriceIn(unit_price=Decimal("15.00")),
    )
    db.commit()
    db.expire_all()

    updated = serialize_pricing_agent_run(
        get_pricing_agent_run(db, current_user=user, run_uuid=run.run_uuid)
    )
    updated_line = updated["result"]["lines"][0]
    updated_evidence = json.loads(
        db.query(PricingAgentRunLine)
        .filter(PricingAgentRunLine.line_uuid == line["line_uuid"])
        .one()
        .evidence_json
    )
    assert updated_line["unit_price"] == "15.000000"
    assert updated_line["total_price"] == "60.000000"
    assert updated_line["decision_revision"] == 2
    assert len(updated_evidence["selection_history"]) == 2


def test_confirmed_agent_run_creates_one_existing_quote_preview_draft(db):
    user, account = _seed_user_account(db)
    _seed_archive_line(db, account=account, user=user)
    payload = PricingAgentRunCreateIn(
        mode="exact",
        sources=["archive"],
        context={"city": "杭州市", "project_type": "写字楼", "decoration_level": "精装"},
        lines=[
            {
                "row_key": "confirmed-draft",
                "item_code": "A-001",
                "item_name": "石膏板吊顶",
                "specification": "轻钢龙骨双层石膏板",
                "quantity": "2",
                "unit": "㎡",
            }
        ],
    )
    run = asyncio.run(create_pricing_agent_run(db, current_user=user, payload=payload))

    first = confirm_pricing_agent_run_to_quote_draft(
        db,
        current_user=user,
        run_uuid=run.run_uuid,
    )
    db.commit()
    second = confirm_pricing_agent_run_to_quote_draft(
        db,
        current_user=user,
        run_uuid=run.run_uuid,
    )
    db.commit()

    assert second["quote_job_id"] == first["quote_job_id"]
    assert second["preview_draft_id"] == first["preview_draft_id"]
    assert db.query(QuoteJob).count() == 1
    assert db.query(QuotePreviewDraft).count() == 1
    job = db.query(QuoteJob).filter(QuoteJob.job_id == first["quote_job_id"]).one()
    draft = db.query(QuotePreviewDraft).filter(
        QuotePreviewDraft.id == first["preview_draft_id"]
    ).one()
    draft_payload = json.loads(draft.draft_json)
    row = draft_payload["project_details"][0]

    assert job.status == "succeeded"
    assert job.stage == "completed"
    assert job.result_item_count == 1
    assert job.result_total_amount == 257
    assert draft.status == "editing"
    assert draft.row_count == 1
    assert draft.priced_row_count == 1
    assert row["manual_unit_price"] == 128.5
    assert row["confirmed_total_price"] == 257
    assert row["price_confirmed_by_user"] is True
    assert row["pricing_agent"]["selection_origin"] == "automatic"
    assert first["draft_url"].startswith("/index.html?quote_job_id=")

    restored = serialize_pricing_agent_run(
        get_pricing_agent_run(db, current_user=user, run_uuid=run.run_uuid)
    )
    assert restored["confirmation"]["confirmed"] is True
    assert restored["confirmation"]["quote_job_id"] == first["quote_job_id"]
    with pytest.raises(PricingAgentError, match="PRICING_AGENT_RUN_ALREADY_CONFIRMED"):
        select_pricing_agent_candidate(
            db,
            current_user=user,
            run_uuid=run.run_uuid,
            line_uuid=restored["result"]["lines"][0]["line_uuid"],
            payload=PricingAgentCandidateSelectIn(
                source="archive",
                source_record_id=restored["result"]["lines"][0]["candidates"][0]["source_record_id"],
            ),
        )


def test_unpriced_agent_line_can_enter_quote_draft_but_remains_push_blocked(db):
    user, _account = _seed_user_account(db)
    payload = PricingAgentRunCreateIn(
        mode="exact",
        sources=["archive"],
        context={"city": "杭州市", "project_type": "餐厅", "decoration_level": "精装"},
        lines=[
            {
                "row_key": "unpriced-placeholder",
                "item_name": "完全没有历史依据的新项目",
                "specification": "定制做法",
                "quantity": "3",
                "unit": "项",
            }
        ],
    )
    run = asyncio.run(create_pricing_agent_run(db, current_user=user, payload=payload))

    confirmation = confirm_pricing_agent_run_to_quote_draft(
        db,
        current_user=user,
        run_uuid=run.run_uuid,
    )
    db.commit()

    draft = db.query(QuotePreviewDraft).filter(
        QuotePreviewDraft.id == confirmation["preview_draft_id"]
    ).one()
    row = json.loads(draft.draft_json)["project_details"][0]

    assert draft.row_count == 1
    assert draft.priced_row_count == 0
    assert draft.unpriced_row_count == 1
    assert confirmation["contains_unpriced_placeholders"] is True
    assert confirmation["unpriced_row_count"] == 1
    assert row["requirement_placeholder"] is True
    assert row["pricing_agent_unpriced_placeholder"] is True
    assert row["needs_manual_pricing"] is True
    assert row["quote_source"] == "pricing_agent_unpriced_placeholder"
    assert row["unit_price"] == 0
    assert row["total_price"] == 0
    assert len(_unpriced_requirement_placeholders([row])) == 1


def test_pricing_agent_run_timestamps_use_one_clock(db):
    user, account = _seed_user_account(db)
    _seed_archive_line(db, account=account, user=user)
    payload = PricingAgentRunCreateIn(
        mode="exact",
        sources=["archive"],
        context={"city": "杭州", "project_type": "写字楼", "decoration_level": "精装"},
        lines=[
            {
                "row_key": "timestamp-row",
                "item_name": "石膏板吊顶",
                "specification": "轻钢龙骨双层石膏板",
                "quantity": "1",
                "unit": "㎡",
            }
        ],
    )

    run = asyncio.run(create_pricing_agent_run(db, current_user=user, payload=payload))

    assert run.started_at is not None
    assert run.finished_at is not None
    assert run.finished_at >= run.started_at


def test_expanded_mode_returns_keyword_candidate_without_automatic_pricing(db):
    user, account = _seed_user_account(db)
    _seed_archive_line(db, account=account, user=user)
    payload = PricingAgentRunCreateIn(
        mode="expanded",
        sources=["archive"],
        context={"city": "杭州", "project_type": "写字楼", "decoration_level": "精装"},
        lines=[
            {
                "row_key": "row-2",
                "item_name": "双层石膏板吊顶",
                "specification": "轻钢龙骨",
                "quantity": "8",
                "unit": "㎡",
            }
        ],
    )
    run = asyncio.run(create_pricing_agent_run(db, current_user=user, payload=payload))
    result = serialize_pricing_agent_run(run)["result"]["lines"][0]
    assert result["selected_source"] is None
    assert result["match_type"] is None
    assert result["unit_price"] is None
    assert result["requires_review"] is True
    assert len(result["candidates"]) == 1
    candidate = result["candidates"][0]
    assert candidate["source"] == "archive"
    assert candidate["match_type"] == "keyword_similar"
    assert result["query_plan"]["context_policy"] == "soft_rerank_only"
    assert result["source_evidence"]["archive"]["channel_status"] == {
        "exact": "used",
        "lexical": "used",
        "vector": "disabled_fallback_keyword",
        "fusion": "keyword_only",
    }

    adopted = select_pricing_agent_candidate(
        db,
        current_user=user,
        run_uuid=run.run_uuid,
        line_uuid=result["line_uuid"],
        payload=PricingAgentCandidateSelectIn(
            source="archive",
            source_record_id=candidate["source_record_id"],
        ),
    )
    adopted_line = serialize_pricing_agent_run(adopted)["result"]["lines"][0]
    assert adopted_line["selected_source"] == "archive"
    assert adopted_line["match_type"] == "keyword_similar"
    assert adopted_line["unit_price"] == "128.500000"
    assert adopted_line["manual_candidate_selected"] is True


def test_pricing_hybrid_adapter_lazily_indexes_missing_snapshot():
    class FakeHybridClient:
        def __init__(self):
            self.search_calls = 0
            self.reindex_calls = 0

        def search(self, **kwargs):
            self.search_calls += 1
            if self.search_calls == 1:
                raise TenderHybridIndexStale("missing")
            return [
                HybridSearchHit(
                    evidence_id="record-1",
                    block_id="record-1",
                    rrf_score=0.032,
                    vector_score=0.86,
                    bm25_score=2.4,
                )
            ]

        def reindex(self, **kwargs):
            self.reindex_calls += 1
            return HybridReindexResult(
                case_id=kwargs["case_id"],
                manifest_version=kwargs["manifest_version"],
                manifest_hash=kwargs["manifest_hash"],
                indexed_block_count=len(kwargs["blocks"]),
                idempotent=False,
            )

    client = FakeHybridClient()
    result = search_pricing_hybrid(
        documents=[
            PricingHybridDocument(
                record_id="record-1",
                content="项目名称：轻钢龙骨石膏板天棚；单位：平方米。",
                document_key="archive.xlsx:Sheet1:2",
                keywords=("轻钢龙骨", "石膏板", "天棚"),
            )
        ],
        scope_key="test-account-archive",
        manifest_version=1,
        query="双层石膏板吊顶 平方米",
        client=client,
        enabled=True,
    )

    assert result.status == "used"
    assert result.indexed_shard_count == 1
    assert result.searched_shard_count == 1
    assert result.hits[0].record_id == "record-1"
    assert result.hits[0].vector_score == 0.86
    assert client.search_calls == 2
    assert client.reindex_calls == 1


def test_expanded_hybrid_candidate_keeps_vector_and_rrf_evidence(db, monkeypatch):
    user, account = _seed_user_account(db)
    _archive, archive_line = _seed_archive_line(db, account=account, user=user)

    def fake_hybrid(**_kwargs):
        return PricingHybridResult(
            hits=(
                PricingHybridHit(
                    record_id=archive_line.line_uuid,
                    rrf_score=0.031,
                    vector_score=0.86,
                    bm25_score=1.7,
                ),
            ),
            status="used",
            indexed_shard_count=1,
            searched_shard_count=1,
        )

    monkeypatch.setattr(
        "app.services.pricing_agent_retrieval.search_pricing_hybrid",
        fake_hybrid,
    )
    payload = PricingAgentRunCreateIn(
        mode="expanded",
        sources=["archive"],
        context={"city": "杭州", "project_type": "写字楼", "decoration_level": "精装"},
        lines=[
            {
                "row_key": "hybrid-vector",
                "item_name": "轻质板材天棚",
                "specification": "双层板面轻钢骨架",
                "quantity": "8",
                "unit": "㎡",
            }
        ],
    )

    run = asyncio.run(create_pricing_agent_run(db, current_user=user, payload=payload))
    result = serialize_pricing_agent_run(run)["result"]["lines"][0]
    candidate = result["candidates"][0]

    assert result["selected_source"] is None
    assert result["unit_price"] is None
    assert result["requires_review"] is True
    assert candidate["match_type"] == "hybrid_similar"
    assert candidate["retrieval_channels"] == ["keyword", "vector"]
    assert candidate["vector_score"] == "0.860000"
    assert candidate["rrf_score"] == "0.031000"
    assert result["source_evidence"]["archive"]["channel_status"]["vector"] == "used"
    assert result["source_evidence"]["archive"]["channel_status"]["fusion"] == "rrf"


def test_expanded_hybrid_rejects_conflicting_action_even_with_high_vector_score(db, monkeypatch):
    user, account = _seed_user_account(db)
    archive, _line = _seed_archive_line(db, account=account, user=user)
    removal = _append_archive_line(
        db,
        archive=archive,
        account=account,
        source_row_index=3,
        item_name="拆除木门",
        specification="单开木门保护性拆除",
        unit="樘",
        unit_price="25.00",
    )

    def fake_hybrid(**_kwargs):
        return PricingHybridResult(
            hits=(
                PricingHybridHit(
                    record_id=removal.line_uuid,
                    rrf_score=0.032,
                    vector_score=0.93,
                    bm25_score=4.5,
                ),
            ),
            status="used",
            indexed_shard_count=0,
            searched_shard_count=1,
        )

    monkeypatch.setattr(
        "app.services.pricing_agent_retrieval.search_pricing_hybrid",
        fake_hybrid,
    )
    payload = PricingAgentRunCreateIn(
        mode="expanded",
        sources=["archive"],
        context={"city": "杭州", "project_type": "住宅", "decoration_level": "精装"},
        lines=[
            {
                "row_key": "action-conflict",
                "item_name": "安装木门",
                "specification": "单开木门安装",
                "quantity": "1",
                "unit": "樘",
            }
        ],
    )

    run = asyncio.run(create_pricing_agent_run(db, current_user=user, payload=payload))
    result = serialize_pricing_agent_run(run)["result"]["lines"][0]

    assert result["selected_source"] is None
    assert result["candidates"] == []
    assert result["requires_review"] is False


def test_industry_estimate_does_not_override_available_approximate_candidate(db, monkeypatch):
    user, account = _seed_user_account(db)
    _seed_archive_line(db, account=account, user=user)

    async def forbidden_industry(**_kwargs):
        pytest.fail("存在存档或企业近似候选时不能直接调用行业估价")

    monkeypatch.setattr(
        "app.services.pricing_agent.estimate_industry_prices",
        forbidden_industry,
    )
    payload = PricingAgentRunCreateIn(
        mode="expanded",
        sources=["archive", "industry"],
        context={"city": "杭州", "project_type": "写字楼", "decoration_level": "精装"},
        lines=[
            {
                "row_key": "candidate-before-industry",
                "item_name": "双层石膏板吊顶",
                "specification": "轻钢龙骨",
                "quantity": "8",
                "unit": "㎡",
            }
        ],
    )

    run = asyncio.run(create_pricing_agent_run(db, current_user=user, payload=payload))
    result = serialize_pricing_agent_run(run)["result"]["lines"][0]

    assert result["selected_source"] is None
    assert result["unit_price"] is None
    assert result["requires_review"] is True
    assert result["candidates"][0]["source"] == "archive"
    assert "industry" not in result["source_evidence"]


def test_exact_mode_can_read_enterprise_source_without_mutating_it(db):
    user, _account = _seed_user_account(db)
    version, item = _seed_enterprise_line(db, user=user)
    payload = PricingAgentRunCreateIn(
        mode="exact",
        sources=["enterprise"],
        context={"city": "杭州市", "project_type": "写字楼", "decoration_level": "精装"},
        lines=[
            {
                "row_key": "enterprise-row",
                "item_code": "E-001",
                "item_name": "乳胶漆墙面",
                "quantity": "3",
                "unit": "㎡",
            }
        ],
    )
    run = asyncio.run(create_pricing_agent_run(db, current_user=user, payload=payload))
    result = serialize_pricing_agent_run(run)["result"]["lines"][0]
    assert result["selected_source"] == "enterprise"
    assert result["source_label"] == "企业数据"
    assert result["unit_price"] == "66.800000"
    assert db.get(EnterpriseQuotaVersion, version.id).status == "active"
    assert float(db.get(EnterpriseQuotaItem, item.id).unit_price) == 66.8


def test_expanded_enterprise_hybrid_candidate_requires_manual_adoption(db, monkeypatch):
    user, _account = _seed_user_account(db)
    version, item = _seed_enterprise_line(db, user=user)

    def fake_hybrid(**_kwargs):
        return PricingHybridResult(
            hits=(
                PricingHybridHit(
                    record_id=str(item.id),
                    rrf_score=0.031,
                    vector_score=0.82,
                    bm25_score=3.2,
                ),
            ),
            status="used",
            indexed_shard_count=0,
            searched_shard_count=1,
        )

    monkeypatch.setattr(
        "app.services.pricing_agent_retrieval.search_pricing_hybrid",
        fake_hybrid,
    )
    payload = PricingAgentRunCreateIn(
        mode="expanded",
        sources=["enterprise"],
        context={"city": "杭州市", "project_type": "写字楼", "decoration_level": "精装"},
        lines=[
            {
                "row_key": "enterprise-near",
                "item_name": "墙面乳胶漆涂刷",
                "specification": "基层处理后涂刷",
                "quantity": "3",
                "unit": "㎡",
            }
        ],
    )

    run = asyncio.run(create_pricing_agent_run(db, current_user=user, payload=payload))
    result = serialize_pricing_agent_run(run)["result"]["lines"][0]
    candidate = result["candidates"][0]

    assert result["selected_source"] is None
    assert result["unit_price"] is None
    assert result["requires_review"] is True
    assert candidate["source"] == "enterprise"
    assert candidate["source_record_id"] == str(item.id)
    assert candidate["match_type"] == "hybrid_similar"

    adopted = select_pricing_agent_candidate(
        db,
        current_user=user,
        run_uuid=run.run_uuid,
        line_uuid=result["line_uuid"],
        payload=PricingAgentCandidateSelectIn(
            source="enterprise",
            source_record_id=str(item.id),
        ),
    )
    adopted_line = serialize_pricing_agent_run(adopted)["result"]["lines"][0]
    assert adopted_line["selected_source"] == "enterprise"
    assert adopted_line["unit_price"] == "66.800000"
    assert db.get(EnterpriseQuotaVersion, version.id).status == "active"
    assert float(db.get(EnterpriseQuotaItem, item.id).unit_price) == 66.8


def test_expanded_industry_source_receives_user_context_and_is_labeled_as_ai(db, monkeypatch):
    user, _account = _seed_user_account(db)
    captured = {}

    async def fake_industry(*, rows, context, current_user):
        captured.update({"rows": rows, "context": context, "user_id": current_user.id})
        return {
            rows[0]["row_id"]: {
                "row_id": rows[0]["row_id"],
                "unit_price": "315.000000",
                "confidence": 0.62,
                "basis": "测试行业估价",
                "risks": ["需人工复核"],
                "source_label": "行业数据·AI推算",
            }
        }

    monkeypatch.setattr("app.services.pricing_agent.estimate_industry_prices", fake_industry)
    payload = PricingAgentRunCreateIn(
        mode="expanded",
        sources=["industry"],
        context={"city": "杭州市", "project_type": "餐厅", "decoration_level": "精装"},
        lines=[
            {
                "row_key": "industry-row",
                "item_name": "定制金属隔断",
                "quantity": "2",
                "unit": "㎡",
            }
        ],
    )
    run = asyncio.run(create_pricing_agent_run(db, current_user=user, payload=payload))
    result = serialize_pricing_agent_run(run)["result"]["lines"][0]
    assert captured["context"] == {
        "city": "杭州市",
        "project_type": "餐厅",
        "decoration_level": "精装",
    }
    assert result["selected_source"] == "industry"
    assert result["source_label"] == "行业数据·AI推算"
    assert result["match_type"] == "ai_estimate"
    assert result["requires_review"] is True
    assert result["total_price"] == "630.000000"


def test_exact_schema_rejects_industry_source():
    with pytest.raises(ValueError, match="准确模式不能使用行业数据"):
        PricingAgentRunCreateIn(
            mode="exact",
            sources=["industry"],
            context={"city": "杭州", "project_type": "写字楼", "decoration_level": "精装"},
            lines=[{"row_key": "row-1", "item_name": "石膏板吊顶"}],
        )


def test_pricing_agent_api_upload_deduplicate_and_exact_run(db, tmp_path):
    user, _account = _seed_user_account(db)
    test_app = FastAPI()
    test_app.include_router(pricing_agent_api.router, prefix="/api/v1")

    def override_user():
        return user

    def override_db():
        yield db

    test_app.dependency_overrides[get_current_user] = override_user
    test_app.dependency_overrides[get_db] = override_db
    previous = {
        "feature_pricing_agent": settings.feature_pricing_agent,
        "pricing_agent_archive_storage_backend": settings.pricing_agent_archive_storage_backend,
        "pricing_agent_archive_local_root": settings.pricing_agent_archive_local_root,
    }
    object.__setattr__(settings, "feature_pricing_agent", True)
    object.__setattr__(settings, "pricing_agent_archive_storage_backend", "local")
    object.__setattr__(settings, "pricing_agent_archive_local_root", str(tmp_path))
    try:
        with TestClient(test_app) as client:
            workbook_content = _workbook_bytes()
            first = client.post(
                "/api/v1/pricing-agent/archives",
                files={
                    "file": (
                        "历史报价.xlsx",
                        workbook_content,
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    )
                },
            )
            assert first.status_code == 200, first.text
            assert first.json()["data"]["indexed_row_count"] == 2
            assert first.json()["data"]["storage_backend"] == "local"
            duplicate = client.post(
                "/api/v1/pricing-agent/archives",
                files={
                    "file": (
                        "历史报价.xlsx",
                        workbook_content,
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    )
                },
            )
            assert duplicate.status_code == 200, duplicate.text
            assert duplicate.json()["duplicate"] is True
            assert db.query(PricingArchiveFile).count() == 1

            response = client.post(
                "/api/v1/pricing-agent/runs",
                json={
                    "mode": "exact",
                    "sources": ["archive"],
                    "context": {
                        "city": "杭州",
                        "project_type": "写字楼",
                        "decoration_level": "精装",
                    },
                    "lines": [
                        {
                            "row_key": "api-row-1",
                            "item_name": "石膏板吊顶",
                            "specification": "轻钢龙骨双层石膏板",
                            "quantity": "2",
                            "unit": "㎡",
                        }
                    ],
                },
            )
            assert response.status_code == 200, response.text
            line = response.json()["data"]["result"]["lines"][0]
            assert line["source_label"] == "存档数据"
            assert line["unit_price"] == "128.500000"
            assert line["total_price"] == "257.000000"
    finally:
        for name, value in previous.items():
            object.__setattr__(settings, name, value)


def test_pricing_agent_api_persists_candidate_and_confirms_existing_draft(db):
    user, account = _seed_user_account(db)
    archive, _line = _seed_archive_line(db, account=account, user=user)
    selected = _append_archive_line(
        db,
        archive=archive,
        account=account,
        source_row_index=3,
        item_name="石膏板吊顶",
        specification="轻钢龙骨双层石膏板",
        unit="㎡",
        unit_price="135.00",
    )
    test_app = FastAPI()
    test_app.include_router(pricing_agent_api.router, prefix="/api/v1")

    def override_user():
        return user

    def override_db():
        yield db

    test_app.dependency_overrides[get_current_user] = override_user
    test_app.dependency_overrides[get_db] = override_db
    previous = settings.feature_pricing_agent
    object.__setattr__(settings, "feature_pricing_agent", True)
    try:
        with TestClient(test_app) as client:
            response = client.post(
                "/api/v1/pricing-agent/runs",
                json={
                    "mode": "exact",
                    "sources": ["archive"],
                    "context": {
                        "city": "杭州市",
                        "project_type": "写字楼",
                        "decoration_level": "精装",
                    },
                    "lines": [
                        {
                            "row_key": "api-decision",
                            "item_name": "石膏板吊顶",
                            "specification": "轻钢龙骨双层石膏板",
                            "quantity": "2",
                            "unit": "㎡",
                        }
                    ],
                },
            )
            assert response.status_code == 200, response.text
            run_data = response.json()["data"]
            line = run_data["result"]["lines"][0]
            assert line["unit_price"] is None

            response = client.put(
                (
                    f"/api/v1/pricing-agent/runs/{run_data['run_uuid']}"
                    f"/lines/{line['line_uuid']}/selection"
                ),
                json={
                    "source": "archive",
                    "source_record_id": selected.line_uuid,
                },
            )
            assert response.status_code == 200, response.text
            persisted = response.json()["data"]["result"]["lines"][0]
            assert persisted["unit_price"] == "135.000000"
            assert persisted["manual_candidate_selected"] is True

            response = client.post(
                f"/api/v1/pricing-agent/runs/{run_data['run_uuid']}/confirm-to-quote-draft"
            )
            assert response.status_code == 200, response.text
            confirmation = response.json()["data"]
            assert confirmation["confirmed"] is True
            assert confirmation["priced_row_count"] == 1

            response = client.get(
                f"/api/v1/pricing-agent/runs/{run_data['run_uuid']}"
            )
            assert response.status_code == 200, response.text
            restored = response.json()["data"]
            assert restored["result"]["lines"][0]["unit_price"] == "135.000000"
            assert restored["confirmation"]["quote_job_id"] == confirmation["quote_job_id"]
    finally:
        object.__setattr__(settings, "feature_pricing_agent", previous)


def test_pricing_agent_api_saves_manual_price_for_unmatched_line(db):
    user, _account = _seed_user_account(db)
    test_app = FastAPI()
    test_app.include_router(pricing_agent_api.router, prefix="/api/v1")

    def override_user():
        return user

    def override_db():
        yield db

    test_app.dependency_overrides[get_current_user] = override_user
    test_app.dependency_overrides[get_db] = override_db
    previous = settings.feature_pricing_agent
    object.__setattr__(settings, "feature_pricing_agent", True)
    try:
        with TestClient(test_app) as client:
            response = client.post(
                "/api/v1/pricing-agent/runs",
                json={
                    "mode": "exact",
                    "sources": ["archive"],
                    "context": {
                        "city": "杭州市",
                        "project_type": "写字楼",
                        "decoration_level": "精装",
                    },
                    "lines": [
                        {
                            "row_key": "api-manual-price",
                            "item_name": "未命中新项目",
                            "quantity": "2",
                            "unit": "项",
                        }
                    ],
                },
            )
            assert response.status_code == 200, response.text
            run_data = response.json()["data"]
            line = run_data["result"]["lines"][0]
            assert line["unit_price"] is None

            response = client.put(
                (
                    f"/api/v1/pricing-agent/runs/{run_data['run_uuid']}"
                    f"/lines/{line['line_uuid']}/manual-price"
                ),
                json={"unit_price": "88.50", "reason": "测试人工补价"},
            )
            assert response.status_code == 200, response.text
            saved = response.json()["data"]["result"]["lines"][0]
            assert saved["source_label"] == "人工补价"
            assert saved["unit_price"] == "88.500000"
            assert saved["total_price"] == "177.000000"
            assert saved["manual_price_entered"] is True
            assert saved["decision_revision"] == 1
    finally:
        object.__setattr__(settings, "feature_pricing_agent", previous)


def test_migration_declares_isolated_pricing_agent_tables():
    text = MIGRATION_PATH.read_text(encoding="utf-8")
    for table_name in (
        "pricing_archive_files",
        "pricing_archive_lines",
        "pricing_agent_runs",
        "pricing_agent_run_lines",
    ):
        assert f'"{table_name}"' in text
        assert table_name in Base.metadata.tables


def test_0075_migration_up_and_down_on_sqlite():
    spec = importlib.util.spec_from_file_location("pricing_agent_0075", MIGRATION_PATH)
    assert spec and spec.loader
    migration = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(migration)
    assert migration.revision == "20260730_0075"
    assert migration.down_revision == "20260729_0074"

    metadata = MetaData()
    Table("users", metadata, Column("id", Integer, primary_key=True))
    Table("accounts", metadata, Column("id", Integer, primary_key=True))
    engine = create_engine("sqlite://")
    with engine.begin() as connection:
        metadata.create_all(connection)
        operations = Operations(MigrationContext.configure(connection))
        original_op = migration.op
        migration.op = operations
        try:
            migration.upgrade()
            created = {
                "pricing_archive_files",
                "pricing_archive_lines",
                "pricing_agent_runs",
                "pricing_agent_run_lines",
            }
            assert created.issubset(inspect(connection).get_table_names())
            migration.downgrade()
            assert not created.intersection(inspect(connection).get_table_names())
        finally:
            migration.op = original_op
    engine.dispose()


def test_0077_migration_adds_and_removes_decision_columns_on_sqlite():
    spec = importlib.util.spec_from_file_location("pricing_agent_0077", DECISION_MIGRATION_PATH)
    assert spec and spec.loader
    migration = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(migration)
    assert migration.revision == "20260731_0077"
    assert migration.down_revision == "20260731_0076"

    metadata = MetaData()
    Table("users", metadata, Column("id", Integer, primary_key=True))
    Table("accounts", metadata, Column("id", Integer, primary_key=True))
    Table(
        "quote_jobs",
        metadata,
        Column("id", Integer, primary_key=True),
        Column("job_id", String(36), unique=True, nullable=False),
    )
    Table(
        "quote_preview_drafts",
        metadata,
        Column("id", Integer, primary_key=True),
        Column("quote_job_id", String(36), nullable=False),
    )
    Table(
        "pricing_agent_runs",
        metadata,
        Column("id", Integer, primary_key=True),
        Column("run_uuid", String(36), unique=True, nullable=False),
        Column("account_id", Integer, nullable=False),
    )
    Table(
        "pricing_agent_run_lines",
        metadata,
        Column("id", Integer, primary_key=True),
        Column("line_uuid", String(36), unique=True, nullable=False),
        Column("run_id", Integer, nullable=False),
    )
    engine = create_engine("sqlite://")
    with engine.begin() as connection:
        metadata.create_all(connection)
        operations = Operations(MigrationContext.configure(connection))
        original_op = migration.op
        migration.op = operations
        try:
            migration.upgrade()
            run_columns = {
                item["name"]
                for item in inspect(connection).get_columns("pricing_agent_runs")
            }
            line_columns = {
                item["name"]
                for item in inspect(connection).get_columns("pricing_agent_run_lines")
            }
            assert {
                "confirmed_quote_job_id",
                "confirmed_preview_draft_id",
                "confirmed_by",
                "confirmed_at",
                "confirmation_hash",
                "confirmation_json",
            }.issubset(run_columns)
            assert {
                "selection_origin",
                "selected_candidate_json",
                "manual_selected_by",
                "manual_selected_at",
                "decision_revision",
            }.issubset(line_columns)

            migration.downgrade()
            run_columns = {
                item["name"]
                for item in inspect(connection).get_columns("pricing_agent_runs")
            }
            line_columns = {
                item["name"]
                for item in inspect(connection).get_columns("pricing_agent_run_lines")
            }
            assert "confirmed_quote_job_id" not in run_columns
            assert "selection_origin" not in line_columns
        finally:
            migration.op = original_op
    engine.dispose()


def test_frontend_exposes_feature_gated_pricing_agent_lab():
    app_text = (FRONTEND_ROOT / "App.vue").read_text(encoding="utf-8")
    component_text = (FRONTEND_ROOT / "PricingAgentLab.vue").read_text(encoding="utf-8")
    api_text = (FRONTEND_ROOT / "pricingAgentApi.js").read_text(encoding="utf-8")
    assert "routeName === 'pricingAgent'" in app_text
    assert "module.key === 'pricing_agent'" in app_text
    assert "准确模式只做精准匹配" in component_text
    assert "关键词与向量混合召回" in component_text
    assert "近似候选必须人工采用" in component_text
    assert "关键词＋向量＋RRF" in component_text
    assert "行业数据·AI推算" in component_text
    assert "采用此价格" in component_text
    assert "选择结果已持久化" in component_text
    assert "确认并生成报价草稿" in component_text
    assert "confirmToQuoteDraft" in api_text
    assert "selectCandidate" in api_text
    assert "setManualPrice" in api_text
    assert "/manual-price" in api_text
    assert "可带占位行生成报价草稿" in component_text
    assert "人工补价" in component_text
    assert "pricing-agent-exact-v1.1" in component_text
    assert '"rules_version": "pricing-agent-exact-v1.1"' in (
        Path(pricing_agent_api.__file__).read_text(encoding="utf-8")
    )
    assert "/pricing-agent/demand-preview" in api_text
