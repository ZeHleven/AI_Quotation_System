from __future__ import annotations

import json

from openpyxl import load_workbook

from app.services import drawing_quantity_confirmation as confirmation
from app.services import dxf_trace_review_pack as trace_pack
from app.services.dxf_trace_review_converter import (
    build_trace_review_conversion,
    read_trace_review_workbook,
    write_trace_review_conversion_outputs,
)


def _trace_row(**overrides: object) -> dict[str, object]:
    row: dict[str, object] = {header: "" for header in trace_pack.TRACE_REVIEW_HEADERS}
    row.update(
        {
            "复核行号": "BIZ2x9h-0001",
            trace_pack.AUTO_ACTION_COLUMN: "建议采用",
            trace_pack.ADOPT_COLUMN: "是",
            trace_pack.REVIEW_COLUMN: "通过",
            trace_pack.MANUAL_QUANTITY_COLUMN: "12.5",
            trace_pack.MANUAL_UNIT_COLUMN: "㎡",
            trace_pack.QUANTITY_SOURCE_COLUMN: "采用 CAD 几何建议量并复核标准扣减规则",
            trace_pack.MANUAL_NAME_COLUMN: "块料楼地面",
            trace_pack.MANUAL_FEATURE_COLUMN: "结合层厚度、材料种类及强度等级：20mm 1:3水泥砂浆；面层材料品种、规格：600x600地砖",
            trace_pack.DEDUCTION_REVIEW_COLUMN: "已按标准规则复核门洞等扣减范围，本行无需额外扣减",
            "建议编号": "S-floor",
            "标准项目编码": "011102003",
            "标准项目名称": "块料楼地面",
            "标准单位": "㎡",
            "标准规则类型": "area",
            "标准工程量计算规则": "按设计图示尺寸以面积计算",
            "标准规则建议量": "12.5",
            "建议单位": "㎡",
            "CAD几何公式": "sum(CAD_area_mm2) * 0.000001",
            "CAD来源图元行号": "10、20",
        }
    )
    row.update(overrides)
    return row


def test_biz2x9h3_converts_only_passed_suggested_adopt_rows():
    passed_row = _trace_row()
    incomplete_row = _trace_row(
        **{
            "复核行号": "BIZ2x9h-0002",
            trace_pack.REVIEW_COLUMN: "",
            trace_pack.MANUAL_FEATURE_COLUMN: "结合层厚度、材料种类及强度等级：待确认",
            trace_pack.DEDUCTION_REVIEW_COLUMN: "",
        }
    )
    skipped_row = _trace_row(
        **{
            "复核行号": "BIZ2x9h-0003",
            trace_pack.AUTO_ACTION_COLUMN: "建议不采用",
            trace_pack.ADOPT_COLUMN: "否",
        }
    )

    conversion = build_trace_review_conversion([passed_row, incomplete_row, skipped_row])

    assert conversion["ok"] is False
    assert conversion["summary"]["system_suggested_adopt_count"] == 2
    assert conversion["summary"]["converted_confirmation_row_count"] == 1
    assert conversion["summary"]["conversion_issue_count"] == 1
    assert conversion["summary"]["skipped_row_count"] == 1
    assert conversion["summary"]["biz2x6_validation_ok"] is True
    assert conversion["summary"]["final_ready_count"] == 1

    confirmation_row = conversion["confirmation_pack"]["confirmation_rows"][0]
    assert confirmation_row["确认行号"] == "BIZ2x9h3-0001"
    assert confirmation_row[confirmation.ADOPT_COLUMN] == "是"
    assert confirmation_row[confirmation.REVIEW_COLUMN] == "通过"
    assert confirmation_row[confirmation.MANUAL_QUANTITY_COLUMN] == "12.5"
    assert confirmation_row["工程量状态"] == "trace_review_confirmed"
    assert "按设计图示尺寸以面积计算" in confirmation_row["工程量证据摘要"]
    assert "核验结论必须填写“通过”" in conversion["issues"][0]["问题说明"]


def test_biz2x9h3_reads_workbook_and_writes_conversion_outputs(tmp_path):
    workbook_path = tmp_path / "trace-review.xlsx"
    trace_pack.write_trace_review_workbook(
        {
            "summary": {"trace_review_row_count": 1},
            "trace_review_rows": [_trace_row()],
            "blocked_rows": [],
            "trace_detail_rows": [],
        },
        workbook_path,
    )

    rows = read_trace_review_workbook(workbook_path)
    conversion = build_trace_review_conversion(rows)
    outputs = write_trace_review_conversion_outputs(conversion, tmp_path, stem="conversion")

    assert conversion["ok"] is True
    assert set(outputs) >= {
        "json",
        "markdown",
        "issue_csv",
        "skipped_csv",
        "converted_confirmation_csv",
        "confirmation_confirmation_xlsx",
        "validation_final_xlsx",
    }
    assert json.loads((tmp_path / "conversion.json").read_text(encoding="utf-8"))["summary"]["converted_confirmation_row_count"] == 1

    confirmation_workbook = load_workbook(outputs["confirmation_confirmation_xlsx"])
    assert confirmation.CONFIRMATION_SHEET_NAME in confirmation_workbook.sheetnames
    assert confirmation_workbook[confirmation.CONFIRMATION_SHEET_NAME]["A2"].value == "BIZ2x9h3-0001"

    final_workbook = load_workbook(outputs["validation_final_xlsx"])
    assert confirmation.FINAL_SHEET_NAME in final_workbook.sheetnames
    assert final_workbook[confirmation.FINAL_SHEET_NAME]["A2"].value == "块料楼地面"
