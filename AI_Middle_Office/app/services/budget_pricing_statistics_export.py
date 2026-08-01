"""Excel export for pricing-draft statistics and reconciled resource details."""

from __future__ import annotations

import re
from dataclasses import dataclass
from io import BytesIO
from typing import Any, Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from app.models.budget_pricing_draft import BudgetProjectPricingDraft
from app.services.budget_pricing_drafts import serialize_budget_pricing_draft
from app.services.budget_pricing_resource_details import (
    build_budget_pricing_resource_details,
)


WORKBOOK_CONTENT_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
_HEADER_FILL = PatternFill("solid", fgColor="2563EB")
_HEADER_FONT = Font(color="FFFFFF", bold=True)
_MONEY_FORMAT = "#,##0.00"
_QUANTITY_FORMAT = "#,##0.000000"

_SUMMARY_FIELDS: tuple[tuple[str, str], ...] = (
    ("labor_total", "人工费"),
    ("main_material_total", "主材费"),
    ("auxiliary_material_total", "辅材费"),
    ("subcontract_total", "分包费"),
    ("direct_subtotal", "直接费小计"),
    ("measures_fee", "措施费"),
    ("management_fee", "管理费"),
    ("other_fee", "其它费用"),
    ("suspended_amount", "暂列金额"),
    ("tax_excluded_total", "不含税合计"),
    ("tax_total", "税费"),
    ("tax_included_total", "含税合计"),
    ("cost_total", "成本合计"),
    ("quote_amount", "报价金额"),
    ("unit_cost", "单位面积造价"),
)

_MATERIAL_COLUMNS: tuple[tuple[str, str], ...] = (
    ("category", "分类"),
    ("resource_code", "材料编码"),
    ("resource_type", "类型"),
    ("resource_name", "材料名称"),
    ("specification", "规格"),
    ("brand", "品牌"),
    ("unit", "单位"),
    ("quantity", "数量"),
    ("price", "除税单价"),
    ("amount", "总价"),
)

_LABOR_COLUMNS: tuple[tuple[str, str], ...] = (
    ("resource_code", "编码"),
    ("resource_type", "类型"),
    ("resource_name", "项目名称"),
    ("work_content", "工作内容"),
    ("calculation_rule", "计算规则"),
    ("unit", "单位"),
    ("quantity", "数量"),
    ("price", "不含税人工单价"),
    ("amount", "人工总价"),
)

_EXPORT_SECTIONS: tuple[str, ...] = (
    "summary",
    "main_material",
    "auxiliary_material",
    "labor",
)
_EXPORT_SECTION_TITLES = {
    "summary": "统计汇总",
    "main_material": "主材明细",
    "auxiliary_material": "辅材明细",
    "labor": "人工明细",
}


@dataclass(frozen=True)
class BudgetPricingStatisticsExportResult:
    content: bytes
    filename: str
    content_type: str = WORKBOOK_CONTENT_TYPE


def _number(value: Any) -> float | None:
    if value in (None, ""):
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _safe_filename(value: str) -> str:
    cleaned = re.sub(r'[\\/:*?"<>|]+', "_", str(value or "").strip())
    return cleaned[:80] or "预算项目"


def _normalize_export_sections(sections: Iterable[str] | None) -> tuple[str, ...]:
    if sections is None:
        return _EXPORT_SECTIONS
    requested = {
        str(section or "").strip()
        for section in sections
        if str(section or "").strip()
    }
    invalid = sorted(requested.difference(_EXPORT_SECTIONS))
    if invalid:
        raise ValueError(f"不支持的统计导出内容：{', '.join(invalid)}")
    selected = tuple(section for section in _EXPORT_SECTIONS if section in requested)
    if not selected:
        raise ValueError("请至少选择一项统计导出内容")
    return selected


def _style_sheet(worksheet: Any, *, widths: tuple[int, ...]) -> None:
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    for cell in worksheet[1]:
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for index, width in enumerate(widths, start=1):
        worksheet.column_dimensions[get_column_letter(index)].width = width


def _append_detail_sheet(
    workbook: Workbook,
    *,
    title: str,
    columns: tuple[tuple[str, str], ...],
    rows: list[dict[str, Any]],
) -> None:
    worksheet = workbook.create_sheet(title)
    worksheet.append([label for _key, label in columns])
    for row in rows:
        values: list[Any] = []
        for key, _label in columns:
            value = row.get(key)
            values.append(_number(value) if key in {"quantity", "price", "amount"} else value)
        worksheet.append(values)
    for row in worksheet.iter_rows(min_row=2):
        for index, (key, _label) in enumerate(columns, start=1):
            if key == "quantity":
                row[index - 1].number_format = _QUANTITY_FORMAT
            elif key in {"price", "amount"}:
                row[index - 1].number_format = _MONEY_FORMAT
    widths = tuple(
        18 if key in {"resource_name", "work_content", "calculation_rule"} else 14
        for key, _label in columns
    )
    _style_sheet(worksheet, widths=widths)


def render_budget_pricing_statistics_export(
    db: Session,
    draft: BudgetProjectPricingDraft,
    *,
    project_name: str,
    sections: Iterable[str] | None = None,
) -> BudgetPricingStatisticsExportResult:
    """Render selected statistics cards and labor/material detail sheets."""

    selected_sections = _normalize_export_sections(sections)
    draft_payload = serialize_budget_pricing_draft(draft)
    summary = draft_payload.get("summary") or {}
    totals = summary.get("totals") or {}
    details = {
        bucket: build_budget_pricing_resource_details(db, draft, bucket=bucket)
        for bucket in ("main_material", "auxiliary_material", "labor")
        if bucket in selected_sections
    }

    workbook = Workbook()
    if "summary" in selected_sections:
        summary_sheet = workbook.active
        summary_sheet.title = _EXPORT_SECTION_TITLES["summary"]
        summary_sheet.append(["统计项", "金额"])
        for key, label in _SUMMARY_FIELDS:
            summary_sheet.append([label, _number(totals.get(key))])
        for row in summary_sheet.iter_rows(min_row=2, min_col=2, max_col=2):
            row[0].number_format = _MONEY_FORMAT
        _style_sheet(summary_sheet, widths=(22, 18))
    else:
        workbook.remove(workbook.active)

    if "main_material" in selected_sections:
        _append_detail_sheet(
            workbook,
            title=_EXPORT_SECTION_TITLES["main_material"],
            columns=_MATERIAL_COLUMNS,
            rows=details["main_material"]["rows"],
        )
    if "auxiliary_material" in selected_sections:
        _append_detail_sheet(
            workbook,
            title=_EXPORT_SECTION_TITLES["auxiliary_material"],
            columns=_MATERIAL_COLUMNS,
            rows=details["auxiliary_material"]["rows"],
        )
    if "labor" in selected_sections:
        _append_detail_sheet(
            workbook,
            title=_EXPORT_SECTION_TITLES["labor"],
            columns=_LABOR_COLUMNS,
            rows=details["labor"]["rows"],
        )

    output = BytesIO()
    workbook.save(output)
    export_title = (
        _EXPORT_SECTION_TITLES[selected_sections[0]]
        if len(selected_sections) == 1
        else "报价统计"
    )
    return BudgetPricingStatisticsExportResult(
        content=output.getvalue(),
        filename=f"{_safe_filename(project_name)}_{export_title}_R{draft.revision}.xlsx",
    )
