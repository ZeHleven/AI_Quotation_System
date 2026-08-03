from __future__ import annotations

import csv
import json
import re
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any, Mapping, Sequence

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


PHASE = "BIZ-2x-pdf-object-recall-capture-pack"

CAPTURE_HEADERS = [
    "capture_no",
    "status",
    "recommended_pass",
    "source_file",
    "page",
    "tile_id",
    "image_path",
    "image_exists",
    "task_nos",
    "object_classes",
    "task_count",
    "prompt_file",
    "prompt_text",
    "target_fields_in_prompt",
]

EVIDENCE_TEMPLATE_HEADERS = [
    "call_no",
    "task_no",
    "source_file",
    "page",
    "tile_id",
    "vision_pass",
    "evidence_role",
    "discipline",
    "evidence_item_hint",
    "evidence_spec_or_method",
    "evidence_suggested_unit",
    "evidence_text",
    "confidence",
    "needs_manual_review",
    "reason",
]

SUMMARY_HEADERS = ["metric", "value"]

PASS_GUIDANCE = {
    "door_window_demolition": "Focus on door/window demolition notes, door schedules, frame/leaf/hardware removal, and visible finish notes.",
    "demolition_node": "Focus on demolition notes, removed finishes, removed fixtures, haul-away notes, and visible renovation scope.",
    "finish_schedule": "Focus on floor, wall, ceiling, finish schedule, material codes, finish legends, and room finish tables.",
    "electrical_mep": "Focus on lighting, switches, sockets, distribution boxes, conduits, wiring, symbols, and electrical legends.",
    "fixture_valve_schedule": "Focus on plumbing fixtures, valves, drains, water meters, faucets, pipe notes, and plumbing legends.",
    "table_legend": "Focus on tables, legends, schedules, symbol explanations, material descriptions, and units.",
}


def build_object_recall_capture_pack(
    workbench_rows: Sequence[Mapping[str, Any]],
    *,
    include_importable: bool = False,
    max_tasks_per_call: int = 12,
) -> dict[str, Any]:
    """Build an answer-blind capture pack for the next external vision pass.

    The pack intentionally excludes target_item_name / target_feature / target_unit
    from prompts and task rows. Those fields are answers for validation, not
    evidence for recall.
    """

    source_rows = [dict(row) for row in workbench_rows if isinstance(row, Mapping)]
    filtered_rows = [
        row for row in source_rows if include_importable or not _is_importable(row)
    ]
    grouped = _group_rows(filtered_rows)

    capture_rows: list[dict[str, Any]] = []
    evidence_template_rows: list[dict[str, Any]] = []
    for group_rows in grouped:
        for chunk in _chunks(group_rows, max(1, max_tasks_per_call)):
            first = chunk[0]
            capture_no = len(capture_rows) + 1
            task_nos = _join_unique(row.get("task_no") for row in chunk)
            object_classes = _join_unique(row.get("object_class") for row in chunk)
            prompt_text = _build_prompt(
                recommended_pass=_text(first.get("recommended_pass")),
                source_file=_text(first.get("source_file")),
                page=_text(first.get("page")),
                tile_id=_text(first.get("tile_id")),
                object_classes=object_classes,
            )
            capture_row = {
                "capture_no": capture_no,
                "status": "needs_external_vision",
                "recommended_pass": first.get("recommended_pass", ""),
                "source_file": first.get("source_file", ""),
                "page": first.get("page", ""),
                "tile_id": first.get("tile_id", ""),
                "image_path": first.get("image_path", ""),
                "image_exists": _truthy(first.get("image_exists")),
                "task_nos": task_nos,
                "object_classes": object_classes,
                "task_count": len(chunk),
                "prompt_file": f"capture_{capture_no:04d}.txt",
                "prompt_text": prompt_text,
                "target_fields_in_prompt": False,
            }
            capture_rows.append(capture_row)
            evidence_template_rows.append(
                {
                    "call_no": capture_no,
                    "task_no": task_nos,
                    "source_file": first.get("source_file", ""),
                    "page": first.get("page", ""),
                    "tile_id": first.get("tile_id", ""),
                    "vision_pass": first.get("recommended_pass", ""),
                    "evidence_role": "",
                    "discipline": "",
                    "evidence_item_hint": "",
                    "evidence_spec_or_method": "",
                    "evidence_suggested_unit": "",
                    "evidence_text": "",
                    "confidence": "",
                    "needs_manual_review": "true",
                    "reason": "",
                }
            )

    summary = {
        "source_workbench_row_count": len(source_rows),
        "capture_source_row_count": len(filtered_rows),
        "capture_call_count": len(capture_rows),
        "image_exists_call_count": sum(1 for row in capture_rows if row.get("image_exists")),
        "missing_image_call_count": sum(1 for row in capture_rows if not row.get("image_exists")),
        "evidence_template_row_count": len(evidence_template_rows),
        "include_importable": include_importable,
        "target_fields_in_prompt": False,
        "answer_columns_count_as_evidence": False,
        "quantity_status": "deferred_until_three_fields_accepted",
        "recommended_pass_counts": dict(Counter(_text(row.get("recommended_pass")) for row in capture_rows)),
    }
    return {
        "ok": True,
        "phase": PHASE,
        "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "summary": summary,
        "capture_rows": capture_rows,
        "evidence_template_rows": evidence_template_rows,
    }


def write_object_recall_capture_pack_outputs(
    report: Mapping[str, Any],
    output_dir: str | Path,
    *,
    stem: str | None = None,
) -> dict[str, str]:
    target = Path(output_dir)
    target.mkdir(parents=True, exist_ok=True)
    file_stem = stem or f"BIZ2x_PDF_object_recall_capture_pack_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
    prompt_dir = target / f"{file_stem}_prompts"
    prompt_dir.mkdir(parents=True, exist_ok=True)

    capture_rows = [dict(row) for row in report.get("capture_rows") or [] if isinstance(row, Mapping)]
    for row in capture_rows:
        prompt_file = str(row.get("prompt_file") or f"capture_{row.get('capture_no', 0):04d}.txt")
        prompt_path = prompt_dir / _safe_file_name(prompt_file)
        prompt_path.parent.mkdir(parents=True, exist_ok=True)
        prompt_path.write_text(str(row.get("prompt_text") or ""), encoding="utf-8")
        row["prompt_file"] = str(prompt_path)

    evidence_rows = [dict(row) for row in report.get("evidence_template_rows") or [] if isinstance(row, Mapping)]
    json_path = target / f"{file_stem}.json"
    capture_csv = target / f"{file_stem}_capture_tasks.csv"
    evidence_template_csv = target / f"{file_stem}_evidence_template.csv"
    markdown_path = target / f"{file_stem}.md"
    xlsx_path = target / f"{file_stem}.xlsx"
    outputs = {
        "json": str(json_path),
        "capture_csv": str(capture_csv),
        "evidence_template_csv": str(evidence_template_csv),
        "markdown": str(markdown_path),
        "xlsx": str(xlsx_path),
        "prompt_dir": str(prompt_dir),
    }
    payload = {
        **dict(report),
        "capture_rows": capture_rows,
        "evidence_template_rows": evidence_rows,
        "outputs": outputs,
    }
    json_path.write_text(json.dumps(_json_safe(payload), ensure_ascii=False, indent=2), encoding="utf-8")
    _write_csv(capture_csv, capture_rows, CAPTURE_HEADERS)
    _write_csv(evidence_template_csv, evidence_rows, EVIDENCE_TEMPLATE_HEADERS)
    markdown_path.write_text(_build_markdown(payload), encoding="utf-8")
    _write_workbook(xlsx_path, payload)
    return outputs


def _group_rows(rows: Sequence[Mapping[str, Any]]) -> list[list[Mapping[str, Any]]]:
    grouped: dict[tuple[str, str, str, str, str], list[Mapping[str, Any]]] = defaultdict(list)
    for row in rows:
        key = (
            _text(row.get("recommended_pass")),
            _file_key(row.get("source_file")),
            _text(row.get("page")),
            _text(row.get("tile_id")),
            _text(row.get("image_path")),
        )
        grouped[key].append(row)
    result = list(grouped.values())
    result.sort(key=lambda group: (_text(group[0].get("recommended_pass")), _text(group[0].get("source_file")), _text(group[0].get("page")), _text(group[0].get("tile_id"))))
    return result


def _chunks(rows: Sequence[Mapping[str, Any]], size: int) -> list[list[Mapping[str, Any]]]:
    return [list(rows[index : index + size]) for index in range(0, len(rows), size)]


def _build_prompt(
    *,
    recommended_pass: str,
    source_file: str,
    page: str,
    tile_id: str,
    object_classes: str,
) -> str:
    guidance = PASS_GUIDANCE.get(recommended_pass, "Inspect all visible construction drawing evidence that may become bill-of-quantities line items.")
    return "\n".join(
        [
            "You are extracting bill-of-quantities evidence from a construction drawing image.",
            "Use only visible drawing evidence: notes, labels, schedules, legends, symbols, dimensions, and material descriptions.",
            "Do not infer quantities. Do not copy expected answers. Do not invent items that are not visible.",
            "Use Chinese for item_hint/spec_or_method/text when the drawing content is Chinese.",
            "item_hint must be a concrete visible construction object or material, not an English category such as Removed Finishes, Removed Fixtures, or Haul-Away Notes.",
            "If the unit is not clearly inferable from BOQ convention, leave suggested_unit empty. Do not write unknown, N/A, or unclear.",
            "If only a generic demolition/haul-away note is visible, keep item_hint empty and put the visible note in text with needs_manual_review=true.",
            "",
            f"source_file: {source_file}",
            f"page: {page}",
            f"tile_id: {tile_id}",
            f"vision_pass: {recommended_pass}",
            f"object_classes: {object_classes}",
            f"focus: {guidance}",
            "",
            "Return JSON only, using this schema:",
            "{",
            '  "evidence_items": [',
            "    {",
            '      "evidence_role": "construction_note|schedule|legend|symbol|dimension|unknown",',
            '      "discipline": "decoration|electrical|plumbing|unknown",',
            '      "item_hint": "concrete Chinese item/material name supported by visible drawing evidence, or empty",',
            '      "spec_or_method": "visible material/spec/method text, preferably Chinese",',
            '      "suggested_unit": "m2|m|set|item|point or empty",',
            '      "text": "exact or paraphrased visible evidence from the drawing",',
            '      "confidence": 0.0,',
            '      "needs_manual_review": true,',
            '      "reason": "why this evidence supports a BOQ line"',
            "    }",
            "  ]",
            "}",
        ]
    )


def _is_importable(row: Mapping[str, Any]) -> bool:
    if str(row.get("ready_for_import") or "").strip().lower() == "true":
        return True
    return str(row.get("fill_status") or "").strip() == "importable"


def _join_unique(values: Any) -> str:
    if values is None or isinstance(values, (str, bytes)):
        iterable = [values]
    else:
        try:
            iterable = iter(values)
        except TypeError:
            iterable = iter([values])
    result: list[str] = []
    seen: set[str] = set()
    for value in iterable:
        text = _text(value)
        if not text or text in seen:
            continue
        seen.add(text)
        result.append(text)
    return ";".join(result)


def _truthy(value: Any) -> bool:
    if isinstance(value, bool):
        return value
    return str(value or "").strip().lower() in {"1", "true", "yes", "y"}


def _file_key(value: Any) -> str:
    text = _text(value).replace("\\", "/")
    return Path(text).name.lower() if text else ""


def _text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _safe_file_name(value: str) -> str:
    name = re.sub(r"[^A-Za-z0-9_.-]+", "_", value).strip("._")
    return name or "capture.txt"


def _write_csv(path: Path, rows: Sequence[Mapping[str, Any]], headers: Sequence[str]) -> None:
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=list(headers), extrasaction="ignore")
        writer.writeheader()
        for row in rows:
            writer.writerow({header: row.get(header, "") for header in headers})


def _build_markdown(payload: Mapping[str, Any]) -> str:
    summary = payload.get("summary") or {}
    outputs = payload.get("outputs") or {}
    lines = [
        "# BIZ-2x PDF Object Recall Capture Pack",
        "",
        "This is an answer-blind evidence capture pack for the next drawing-recognition pass.",
        "Target answer fields are intentionally excluded from prompts and task rows.",
        "",
        "## Summary",
        "",
    ]
    for key in (
        "source_workbench_row_count",
        "capture_source_row_count",
        "capture_call_count",
        "image_exists_call_count",
        "missing_image_call_count",
        "target_fields_in_prompt",
        "quantity_status",
    ):
        lines.append(f"- {key}: {summary.get(key, '')}")
    lines.extend(["", "## Outputs", ""])
    for key, value in outputs.items():
        lines.append(f"- {key}: `{value}`")
    lines.extend(
        [
            "",
            "## How To Use",
            "",
            "1. Open `capture_tasks` and the prompt files.",
            "2. Run each image with its prompt in the external vision model or manual review.",
            "3. Fill `blank_evidence_template` rows, duplicating rows when one image yields multiple evidence items.",
            "4. Feed the filled workbook/CSV back into `biz2x_pdf_external_recall_acceptance_pipeline.py --external-results`.",
        ]
    )
    return "\n".join(lines) + "\n"


def _write_workbook(path: Path, payload: Mapping[str, Any]) -> None:
    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "summary"
    _write_sheet(
        summary_sheet,
        [{"metric": key, "value": value} for key, value in (payload.get("summary") or {}).items()],
        SUMMARY_HEADERS,
    )
    capture_sheet = workbook.create_sheet("capture_tasks")
    _write_sheet(capture_sheet, payload.get("capture_rows") or [], CAPTURE_HEADERS)
    template_sheet = workbook.create_sheet("blank_evidence_template")
    _write_sheet(template_sheet, payload.get("evidence_template_rows") or [], EVIDENCE_TEMPLATE_HEADERS)
    workbook.save(path)


def _write_sheet(sheet: Any, rows: Sequence[Mapping[str, Any]], headers: Sequence[str]) -> None:
    fill = PatternFill("solid", fgColor="1F4E78")
    font = Font(color="FFFFFF", bold=True)
    for col_index, header in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=col_index, value=header)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row_index, row in enumerate(rows, start=2):
        for col_index, header in enumerate(headers, start=1):
            value = row.get(header, "")
            if isinstance(value, (dict, list)):
                value = json.dumps(value, ensure_ascii=False)
            sheet.cell(row=row_index, column=col_index, value=value)
    for col_index, header in enumerate(headers, start=1):
        width = min(max(len(header) + 2, 12), 48)
        sheet.column_dimensions[get_column_letter(col_index)].width = width
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions


def _json_safe(value: Any) -> Any:
    if isinstance(value, Mapping):
        return {str(key): _json_safe(item) for key, item in value.items()}
    if isinstance(value, list):
        return [_json_safe(item) for item in value]
    if isinstance(value, Path):
        return str(value)
    return value
