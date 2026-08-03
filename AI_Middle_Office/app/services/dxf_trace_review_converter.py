from __future__ import annotations

import csv
import json
from datetime import datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any, Mapping

from openpyxl import load_workbook

from app.services import drawing_quantity_confirmation as confirmation
from app.services.dxf_trace_review_pack import (
    ADOPT_COLUMN,
    AUTO_ACTION_COLUMN,
    DEDUCTION_REVIEW_COLUMN,
    ISSUE_COLUMN,
    MANUAL_FEATURE_COLUMN,
    MANUAL_NAME_COLUMN,
    MANUAL_QUANTITY_COLUMN,
    MANUAL_UNIT_COLUMN,
    QUANTITY_SOURCE_COLUMN,
    REVIEW_COLUMN,
    TRACE_REVIEW_SHEET_NAME,
)


PHASE = "BIZ-2x-9h-3-trace-review-to-confirmation"
SUGGESTED_ADOPT_ACTION = "建议采用"
PASS_REVIEW_VALUE = "通过"
TRACE_PLACEHOLDER_MARKERS = ("待确认", "待补", "缺失", "missing_needs_manual_review")

ISSUE_HEADERS = [
    "复核行号",
    "建议编号",
    "系统建议动作",
    "是否采用",
    "核验结论",
    "问题说明",
    "处理建议",
]

SKIPPED_HEADERS = [
    "复核行号",
    "建议编号",
    "系统建议动作",
    "是否采用",
    "跳过原因",
]


def read_trace_review_workbook(path: str | Path) -> list[dict[str, Any]]:
    workbook = load_workbook(path, data_only=True)
    if TRACE_REVIEW_SHEET_NAME not in workbook.sheetnames:
        raise ValueError(f"trace 复核工作簿缺少工作表：{TRACE_REVIEW_SHEET_NAME}")
    sheet = workbook[TRACE_REVIEW_SHEET_NAME]
    headers = [_clean_text(cell.value) for cell in sheet[1]]
    rows: list[dict[str, Any]] = []
    for values in sheet.iter_rows(min_row=2, values_only=True):
        if not any(value not in (None, "") for value in values):
            continue
        rows.append({header: value for header, value in zip(headers, values) if header})
    return rows


def build_trace_review_conversion(trace_review_rows: list[Mapping[str, Any]]) -> dict[str, Any]:
    confirmation_rows: list[dict[str, Any]] = []
    feature_rows: list[dict[str, Any]] = []
    evidence_rows: list[dict[str, Any]] = []
    issues: list[dict[str, Any]] = []
    skipped_rows: list[dict[str, Any]] = []

    suggested_adopt_count = 0
    manually_not_adopted_count = 0
    for row_index, row in enumerate(trace_review_rows, start=1):
        row_id = _clean_text(row.get("复核行号")) or f"TRACE-ROW-{row_index}"
        suggestion_key = _clean_text(row.get("建议编号"))
        action = _clean_text(row.get(AUTO_ACTION_COLUMN))
        adopt = _clean_text(row.get(ADOPT_COLUMN))
        review = _clean_text(row.get(REVIEW_COLUMN))

        if action != SUGGESTED_ADOPT_ACTION:
            skipped_rows.append(_skipped_row(row, "系统建议动作不是“建议采用”"))
            continue

        suggested_adopt_count += 1
        if not _is_yes(adopt):
            manually_not_adopted_count += 1
            skipped_rows.append(_skipped_row(row, "业务未确认采用"))
            continue

        row_issues = _validate_convertible_trace_row(row)
        if row_issues:
            issues.append(
                {
                    "复核行号": row_id,
                    "建议编号": suggestion_key,
                    "系统建议动作": action,
                    "是否采用": adopt,
                    "核验结论": review,
                    "问题说明": "；".join(row_issues),
                    "处理建议": "补齐/修正后重新运行 BIZ-2x-9h-3 转换；本行当前不会进入确认清单。",
                }
            )
            continue

        confirmation_row_id = f"BIZ2x9h3-{len(confirmation_rows) + 1:04d}"
        confirmation_row = _to_confirmation_row(confirmation_row_id, row)
        confirmation_rows.append(confirmation_row)
        feature_rows.extend(_to_feature_rows(confirmation_row_id, row))
        evidence_rows.append(_to_evidence_row(confirmation_row_id, row))

    confirmation_pack = {
        "ok": True,
        "phase": "BIZ-2x-9h-3-converted-biz2x6-confirmation-pack",
        "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "safe_for_final_quantity_list": False,
        "summary": {
            "confirmation_row_count": len(confirmation_rows),
            "feature_detail_count": len(feature_rows),
            "evidence_detail_count": len(evidence_rows),
            "source_phase": PHASE,
            "final_export_requires_manual_confirmation": False,
        },
        "confirmation_rows": confirmation_rows,
        "feature_rows": feature_rows,
        "evidence_rows": evidence_rows,
    }
    validation = confirmation.validate_confirmation_rows(confirmation_rows)
    validation["phase"] = "BIZ-2x-9h-3-converted-confirmation-validation"
    validation["generated_at"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    conversion_ok = bool(confirmation_rows) and not issues and validation.get("ok") is True
    return {
        "ok": conversion_ok,
        "phase": PHASE,
        "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "safe_for_final_quantity_list": bool(validation.get("ok")),
        "summary": {
            "trace_review_row_count": len(trace_review_rows),
            "system_suggested_adopt_count": suggested_adopt_count,
            "converted_confirmation_row_count": len(confirmation_rows),
            "conversion_issue_count": len(issues),
            "skipped_row_count": len(skipped_rows),
            "manually_not_adopted_count": manually_not_adopted_count,
            "biz2x6_validation_ok": bool(validation.get("ok")),
            "final_ready_count": validation.get("summary", {}).get("adopted_final_row_count", 0),
        },
        "issues": issues,
        "skipped_rows": skipped_rows,
        "confirmation_pack": confirmation_pack,
        "confirmation_validation": validation,
    }


def build_trace_review_conversion_markdown(conversion: Mapping[str, Any]) -> str:
    summary = conversion.get("summary", {})
    lines = [
        "# BIZ-2x-9h-3 trace 复核转确认行预检",
        "",
        f"- 生成时间：{conversion.get('generated_at', '-')}",
        f"- trace 复核行：{summary.get('trace_review_row_count', 0)}",
        f"- 系统建议采用行：{summary.get('system_suggested_adopt_count', 0)}",
        f"- 已转换为 BIZ-2x-6 确认行：{summary.get('converted_confirmation_row_count', 0)}",
        f"- 转换问题行：{summary.get('conversion_issue_count', 0)}",
        f"- 跳过行：{summary.get('skipped_row_count', 0)}",
        f"- BIZ-2x-6 校验是否通过：{'是' if summary.get('biz2x6_validation_ok') else '否'}",
        f"- 最终四字段可导出行：{summary.get('final_ready_count', 0)}",
        "",
        "## 转换条件",
        "",
        "- `系统建议动作=建议采用`。",
        "- `是否采用=是`。",
        "- `核验结论=通过`。",
        "- 项目名称、项目特征、确认单位、确认工程量、工程量来源说明、扣减/合并规则复核均已填写。",
        "- 项目特征中不能保留“待确认/待补/缺失”等占位提示。",
        "",
        "## 边界",
        "",
        "- 本步骤不写数据库、不接报价，只生成可追溯的转换报告和 BIZ-2x-6 兼容确认行。",
        "- 若没有转换行或校验未通过，不会生成最终四字段清单。",
    ]
    if conversion.get("issues"):
        lines.extend(["", "## 前 10 条问题", ""])
        for issue in list(conversion.get("issues") or [])[:10]:
            lines.append(f"- {issue.get('复核行号', '')}：{issue.get('问题说明', '')}")
    return "\n".join(lines) + "\n"


def write_trace_review_conversion_outputs(
    conversion: Mapping[str, Any],
    output_dir: str | Path,
    *,
    stem: str | None = None,
) -> dict[str, str]:
    directory = Path(output_dir)
    directory.mkdir(parents=True, exist_ok=True)
    file_stem = stem or f"BIZ2x9h3_trace复核转确认行_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
    json_path = directory / f"{file_stem}.json"
    md_path = directory / f"{file_stem}.md"
    issue_csv_path = directory / f"{file_stem}_转换问题.csv"
    skipped_csv_path = directory / f"{file_stem}_跳过行.csv"
    converted_csv_path = directory / f"{file_stem}_BIZ2x6确认行.csv"

    json_path.write_text(json.dumps(conversion, ensure_ascii=False, indent=2), encoding="utf-8")
    md_path.write_text(build_trace_review_conversion_markdown(conversion), encoding="utf-8")
    _write_csv(issue_csv_path, list(conversion.get("issues") or []), ISSUE_HEADERS)
    _write_csv(skipped_csv_path, list(conversion.get("skipped_rows") or []), SKIPPED_HEADERS)
    confirmation_rows = list((conversion.get("confirmation_pack") or {}).get("confirmation_rows") or [])
    _write_csv(converted_csv_path, confirmation_rows, confirmation.CONFIRMATION_HEADERS)

    outputs = {
        "json": str(json_path),
        "markdown": str(md_path),
        "issue_csv": str(issue_csv_path),
        "skipped_csv": str(skipped_csv_path),
        "converted_confirmation_csv": str(converted_csv_path),
    }
    if confirmation_rows:
        confirmation_outputs = confirmation.write_confirmation_outputs(
            conversion.get("confirmation_pack") or {},
            directory,
            stem=f"{file_stem}_BIZ2x6确认行",
        )
        outputs.update({f"confirmation_{key}": value for key, value in confirmation_outputs.items()})
        validation = conversion.get("confirmation_validation") or {}
        validation_outputs = confirmation.write_validation_report(
            validation,
            directory,
            stem=f"{file_stem}_BIZ2x6确认行校验",
        )
        outputs.update({f"validation_{key}": value for key, value in validation_outputs.items()})
    return outputs


def _validate_convertible_trace_row(row: Mapping[str, Any]) -> list[str]:
    row_issues: list[str] = []
    if _clean_text(row.get(REVIEW_COLUMN)) != PASS_REVIEW_VALUE:
        row_issues.append("核验结论必须填写“通过”")
    project_name = _clean_text(row.get(MANUAL_NAME_COLUMN)) or _clean_text(row.get("标准项目名称"))
    feature_text = _clean_text(row.get(MANUAL_FEATURE_COLUMN))
    unit = _clean_text(row.get(MANUAL_UNIT_COLUMN)) or _clean_text(row.get("建议单位"))
    quantity = _parse_positive_decimal(row.get(MANUAL_QUANTITY_COLUMN))
    source = _clean_text(row.get(QUANTITY_SOURCE_COLUMN))
    deduction_review = _clean_text(row.get(DEDUCTION_REVIEW_COLUMN))
    if not project_name:
        row_issues.append("项目名称不能为空")
    if not feature_text:
        row_issues.append("项目特征不能为空")
    if any(marker in feature_text for marker in TRACE_PLACEHOLDER_MARKERS):
        row_issues.append("项目特征仍包含待确认/待补/缺失提示")
    if not unit:
        row_issues.append("确认单位不能为空")
    if quantity is None:
        row_issues.append("确认工程量必须填写大于 0 的数字")
    if not source:
        row_issues.append("工程量来源说明不能为空")
    if not deduction_review:
        row_issues.append("扣减/合并规则复核不能为空")
    return row_issues


def _to_confirmation_row(row_id: str, row: Mapping[str, Any]) -> dict[str, Any]:
    quantity = _parse_positive_decimal(row.get(MANUAL_QUANTITY_COLUMN))
    feature_text = _clean_text(row.get(MANUAL_FEATURE_COLUMN))
    deduction_review = _clean_text(row.get(DEDUCTION_REVIEW_COLUMN))
    source_description = _clean_text(row.get(QUANTITY_SOURCE_COLUMN))
    evidence_summary = _evidence_summary(row, deduction_review)
    project_name = _clean_text(row.get(MANUAL_NAME_COLUMN)) or _clean_text(row.get("标准项目名称"))
    unit = _clean_text(row.get(MANUAL_UNIT_COLUMN)) or _clean_text(row.get("建议单位"))
    return {
        "确认行号": row_id,
        confirmation.ADOPT_COLUMN: "是",
        confirmation.REVIEW_COLUMN: PASS_REVIEW_VALUE,
        confirmation.MANUAL_QUANTITY_COLUMN: _format_decimal(quantity),
        confirmation.MANUAL_UNIT_COLUMN: unit,
        confirmation.QUANTITY_SOURCE_COLUMN: source_description,
        confirmation.MANUAL_NAME_COLUMN: project_name,
        confirmation.MANUAL_FEATURE_COLUMN: feature_text,
        confirmation.ISSUE_COLUMN: _clean_text(row.get(ISSUE_COLUMN)),
        "候选编号": _clean_text(row.get("建议编号")) or _clean_text(row.get("复核行号")),
        "标准项目编码": _clean_text(row.get("标准项目编码")),
        "标准项目名称": _clean_text(row.get("标准项目名称")) or project_name,
        "标准单位": _clean_text(row.get("标准单位")),
        "工程量状态": "trace_review_confirmed",
        "待补量原因": "",
        "工程量规则类型": _clean_text(row.get("标准规则类型")),
        "标准工程量计算规则": _clean_text(row.get("标准工程量计算规则")),
        "建议工程量": _clean_text(row.get("标准规则建议量")) or _clean_text(row.get("几何建议量")),
        "建议单位": _clean_text(row.get("建议单位")) or unit,
        "工程量证据摘要": evidence_summary,
        "图纸识别名称": project_name,
        "图纸识别规格或做法": deduction_review,
        "来源文件": "",
        "来源行号": _clean_text(row.get("CAD来源图元行号")),
        "匹配置信度": "",
        "项目特征缺失字段": "",
    }


def _to_feature_rows(row_id: str, row: Mapping[str, Any]) -> list[dict[str, Any]]:
    feature_text = _clean_text(row.get(MANUAL_FEATURE_COLUMN))
    features = _split_feature_text(feature_text)
    if not features:
        return []
    return [
        {
            "确认行号": row_id,
            "候选编号": _clean_text(row.get("建议编号")),
            "标准项目编码": _clean_text(row.get("标准项目编码")),
            "标准项目名称": _clean_text(row.get("标准项目名称")),
            "项目特征字段": name,
            "候选填充值": value,
            "状态": "confirmed_from_trace_review",
            "置信度": "",
            "证据文本": feature_text,
        }
        for name, value in features
    ]


def _to_evidence_row(row_id: str, row: Mapping[str, Any]) -> dict[str, Any]:
    return {
        "确认行号": row_id,
        "候选编号": _clean_text(row.get("建议编号")),
        "标准项目编码": _clean_text(row.get("标准项目编码")),
        "标准项目名称": _clean_text(row.get("标准项目名称")),
        "证据类型": "standard_rule_trace",
        "证据值": _clean_text(row.get("标准规则建议量")) or _clean_text(row.get(MANUAL_QUANTITY_COLUMN)),
        "证据单位": _clean_text(row.get("建议单位")) or _clean_text(row.get(MANUAL_UNIT_COLUMN)),
        "是否匹配工程量规则": "是",
        "证据置信度": "",
        "证据文本": _evidence_summary(row, _clean_text(row.get(DEDUCTION_REVIEW_COLUMN))),
        "来源文件": "",
        "图层": "",
        "布局": "",
        "块名": "",
        "X": "",
        "Y": "",
        "源行号": _clean_text(row.get("CAD来源图元行号")),
        "业务标签": "BIZ-2x-9h-3 trace复核通过",
    }


def _evidence_summary(row: Mapping[str, Any], deduction_review: str) -> str:
    parts = [
        f"标准规则：{_clean_text(row.get('标准工程量计算规则'))}",
        f"CAD公式：{_clean_text(row.get('CAD几何公式'))}",
        f"CAD来源图元行号：{_clean_text(row.get('CAD来源图元行号'))}",
        f"扣减/合并复核：{deduction_review}",
    ]
    return "；".join(part for part in parts if part and not part.endswith("："))


def _split_feature_text(feature_text: str) -> list[tuple[str, str]]:
    features: list[tuple[str, str]] = []
    for part in feature_text.replace("\n", "；").split("；"):
        item = part.strip()
        if not item:
            continue
        if "：" in item:
            name, value = item.split("：", 1)
        elif ":" in item:
            name, value = item.split(":", 1)
        else:
            name, value = item, ""
        name = name.strip()
        value = value.strip()
        if name:
            features.append((name, value))
    return features


def _skipped_row(row: Mapping[str, Any], reason: str) -> dict[str, Any]:
    return {
        "复核行号": _clean_text(row.get("复核行号")),
        "建议编号": _clean_text(row.get("建议编号")),
        "系统建议动作": _clean_text(row.get(AUTO_ACTION_COLUMN)),
        "是否采用": _clean_text(row.get(ADOPT_COLUMN)),
        "跳过原因": reason,
    }


def _write_csv(path: Path, rows: list[Mapping[str, Any]], headers: list[str]) -> None:
    with path.open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.DictWriter(handle, fieldnames=headers)
        writer.writeheader()
        for row in rows:
            writer.writerow({header: row.get(header, "") for header in headers})


def _parse_positive_decimal(value: Any) -> Decimal | None:
    try:
        parsed = Decimal(str(value).strip())
    except (InvalidOperation, ValueError):
        return None
    if parsed <= 0:
        return None
    return parsed


def _format_decimal(value: Decimal | None) -> str:
    if value is None:
        return ""
    normalized = value.normalize()
    return format(normalized, "f")


def _is_yes(value: Any) -> bool:
    return _clean_text(value) in {"是", "Y", "y", "yes", "YES", "true", "TRUE", "1"}


def _clean_text(value: Any) -> str:
    return str(value or "").strip()
