from __future__ import annotations

import csv
import json
import re
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any, Mapping, Sequence

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.services.drawing_three_field_acceptance import _specific_feature_tokens


PHASE = "BIZ-2x-pdf-feature-precision-capture-pack"

DEFAULT_ROUTES = ("feature_enrichment", "split_variant_review")

CAPTURE_HEADERS = [
    "capture_no",
    "status",
    "recommended_pass",
    "source_file",
    "page",
    "tile_id",
    "image_path",
    "image_exists",
    "defect_nos",
    "object_classes",
    "feature_gap_families",
    "task_count",
    "prompt_file",
    "prompt_text",
    "target_fields_in_prompt",
]

EVIDENCE_TEMPLATE_HEADERS = [
    "call_no",
    "task_no",
    "source_file",
    "page",
    "tile_id",
    "vision_pass",
    "evidence_role",
    "discipline",
    "evidence_item_hint",
    "evidence_spec_or_method",
    "evidence_suggested_unit",
    "evidence_text",
    "confidence",
    "needs_manual_review",
    "reason",
]

SUMMARY_HEADERS = ["metric", "value"]

PRECISION_PASS_GUIDANCE = {
    "electrical_mep": "Inspect electrical legends, system diagrams, cable schedules, conduit/wiring notes, distribution-box notes, and labels. Extract exact visible conduit types, wire/cable models, sizes, installation method, and units.",
    "fixture_valve_schedule": "Inspect plumbing legends, riser/system diagrams, pipe/valve schedules, fixture schedules, and notes. Extract exact visible pipe material, diameter, valve type, fixture spec, connection method, and units.",
    "finish_schedule": "Inspect finish schedules, room finish tables, material legends, node notes, and material labels. Extract exact visible finish code, size, material, substrate, layer build-up, and units.",
    "demolition_node": "Inspect demolition notes and renovation scope. Extract exact visible removed object, material, size, haul-away/scope notes, and units.",
    "table_legend": "Inspect schedules, tables, legends, and symbol notes. Extract exact visible item/spec/method text and units.",
}


def build_feature_precision_capture_pack(
    defect_router_report: Mapping[str, Any],
    *,
    recall_plans: Sequence[Mapping[str, Any]] | None = None,
    image_roots: Sequence[str | Path] | None = None,
    routes: Sequence[str] | None = None,
    max_rows: int | None = None,
    max_tasks_per_call: int = 8,
) -> dict[str, Any]:
    wanted_routes = tuple(routes or DEFAULT_ROUTES)
    defect_rows = [dict(row) for row in defect_router_report.get("defect_rows") or [] if isinstance(row, Mapping)]
    selected_rows = [
        row
        for row in defect_rows
        if str(row.get("repair_route") or "") in wanted_routes and _needs_precision_capture(row)
    ]
    if max_rows is not None:
        selected_rows = selected_rows[:max_rows]

    resolver = _ImageResolver(recall_plans=recall_plans or [], image_roots=image_roots or [])
    task_rows = [_build_task_row(row, resolver=resolver, task_no=index) for index, row in enumerate(selected_rows, start=1)]
    grouped = _group_task_rows(task_rows)

    capture_rows: list[dict[str, Any]] = []
    evidence_template_rows: list[dict[str, Any]] = []
    for group_rows in grouped:
        for chunk in _chunks(group_rows, max(1, max_tasks_per_call)):
            first = chunk[0]
            capture_no = len(capture_rows) + 1
            defect_nos = _join_unique(row.get("defect_no") for row in chunk)
            object_classes = _join_unique(row.get("object_class") for row in chunk)
            gap_families = _join_unique(row.get("feature_gap_families") for row in chunk)
            prompt_text = _build_prompt(
                recommended_pass=_text(first.get("recommended_pass")),
                source_file=_text(first.get("source_file")),
                page=_text(first.get("page")),
                tile_id=_text(first.get("tile_id")),
                object_classes=object_classes,
                feature_gap_families=gap_families,
            )
            capture_rows.append(
                {
                    "capture_no": capture_no,
                    "status": "needs_external_vision",
                    "recommended_pass": first.get("recommended_pass", ""),
                    "source_file": first.get("source_file", ""),
                    "page": first.get("page", ""),
                    "tile_id": first.get("tile_id", ""),
                    "image_path": first.get("image_path", ""),
                    "image_exists": bool(first.get("image_exists")),
                    "defect_nos": defect_nos,
                    "object_classes": object_classes,
                    "feature_gap_families": gap_families,
                    "task_count": len(chunk),
                    "prompt_file": f"feature_precision_{capture_no:04d}.txt",
                    "prompt_text": prompt_text,
                    "target_fields_in_prompt": False,
                }
            )
            evidence_template_rows.append(
                {
                    "call_no": capture_no,
                    "task_no": defect_nos,
                    "source_file": first.get("source_file", ""),
                    "page": first.get("page", ""),
                    "tile_id": first.get("tile_id", ""),
                    "vision_pass": first.get("recommended_pass", ""),
                    "evidence_role": "",
                    "discipline": "",
                    "evidence_item_hint": "",
                    "evidence_spec_or_method": "",
                    "evidence_suggested_unit": "",
                    "evidence_text": "",
                    "confidence": "",
                    "needs_manual_review": "true",
                    "reason": "",
                }
            )

    summary = {
        "source_defect_count": len(defect_rows),
        "selected_defect_count": len(selected_rows),
        "capture_call_count": len(capture_rows),
        "image_exists_call_count": sum(1 for row in capture_rows if row.get("image_exists")),
        "missing_image_call_count": sum(1 for row in capture_rows if not row.get("image_exists")),
        "evidence_template_row_count": len(evidence_template_rows),
        "routes": list(wanted_routes),
        "route_counts": dict(Counter(str(row.get("repair_route") or "") for row in task_rows)),
        "object_class_counts": dict(Counter(str(row.get("object_class") or "") for row in task_rows)),
        "recommended_pass_counts": dict(Counter(str(row.get("recommended_pass") or "") for row in capture_rows)),
        "feature_gap_family_counts": dict(Counter(family for row in task_rows for family in _split_values(row.get("feature_gap_families")))),
        "target_fields_in_prompt": False,
        "answer_columns_count_as_evidence": False,
        "quantity_status": "deferred_until_three_fields_accepted",
    }
    return {
        "ok": True,
        "phase": PHASE,
        "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "safe_to_import_without_evidence": False,
        "answer_columns_count_as_evidence": False,
        "summary": summary,
        "capture_rows": capture_rows,
        "evidence_template_rows": evidence_template_rows,
    }


def write_feature_precision_capture_pack_outputs(
    report: Mapping[str, Any],
    output_dir: str | Path,
    *,
    stem: str | None = None,
) -> dict[str, str]:
    target = Path(output_dir)
    target.mkdir(parents=True, exist_ok=True)
    file_stem = stem or f"BIZ2x_PDF_feature_precision_capture_pack_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
    prompt_dir = target / f"{file_stem}_prompts"
    prompt_dir.mkdir(parents=True, exist_ok=True)

    capture_rows = [dict(row) for row in report.get("capture_rows") or [] if isinstance(row, Mapping)]
    for row in capture_rows:
        prompt_file = str(row.get("prompt_file") or f"feature_precision_{row.get('capture_no', 0):04d}.txt")
        prompt_path = prompt_dir / _safe_file_name(prompt_file)
        prompt_path.parent.mkdir(parents=True, exist_ok=True)
        prompt_path.write_text(str(row.get("prompt_text") or ""), encoding="utf-8")
        row["prompt_file"] = str(prompt_path)

    evidence_rows = [dict(row) for row in report.get("evidence_template_rows") or [] if isinstance(row, Mapping)]
    json_path = target / f"{file_stem}.json"
    capture_csv = target / f"{file_stem}_capture_tasks.csv"
    evidence_template_csv = target / f"{file_stem}_evidence_template.csv"
    markdown_path = target / f"{file_stem}.md"
    xlsx_path = target / f"{file_stem}.xlsx"
    outputs = {
        "json": str(json_path),
        "capture_csv": str(capture_csv),
        "evidence_template_csv": str(evidence_template_csv),
        "markdown": str(markdown_path),
        "xlsx": str(xlsx_path),
        "prompt_dir": str(prompt_dir),
    }
    payload = {
        **dict(report),
        "capture_rows": capture_rows,
        "evidence_template_rows": evidence_rows,
        "outputs": outputs,
    }
    json_path.write_text(json.dumps(_json_safe(payload), ensure_ascii=False, indent=2), encoding="utf-8")
    _write_csv(capture_csv, capture_rows, CAPTURE_HEADERS)
    _write_csv(evidence_template_csv, evidence_rows, EVIDENCE_TEMPLATE_HEADERS)
    markdown_path.write_text(_build_markdown(payload), encoding="utf-8")
    _write_workbook(xlsx_path, payload)
    return outputs


def _needs_precision_capture(row: Mapping[str, Any]) -> bool:
    answer_text = " ".join([_text(row.get("answer_item_name")), _text(row.get("answer_feature"))])
    candidate_text = " ".join([_text(row.get("candidate_item_name")), _text(row.get("candidate_feature"))])
    answer_tokens = _specific_feature_tokens(answer_text)
    candidate_tokens = _specific_feature_tokens(candidate_text)
    if answer_tokens - candidate_tokens:
        return True
    candidate_compact = _compact(candidate_text)
    if any(term in candidate_compact for term in ("具体型号和规格", "按图纸型号", "可见文本", "图例说明")):
        return True
    object_class = _text(row.get("object_class"))
    return object_class in {"electrical_mep", "fixture_valve_schedule"} and str(row.get("repair_route") or "") == "feature_enrichment"


def _build_task_row(row: Mapping[str, Any], *, resolver: "_ImageResolver", task_no: int) -> dict[str, Any]:
    answer_tokens = _specific_feature_tokens(" ".join([_text(row.get("answer_item_name")), _text(row.get("answer_feature"))]))
    candidate_tokens = _specific_feature_tokens(" ".join([_text(row.get("candidate_item_name")), _text(row.get("candidate_feature"))]))
    image = resolver.resolve(row)
    object_class = _text(row.get("object_class")) or _object_class_from_text(row)
    return {
        "task_no": task_no,
        "defect_no": row.get("defect_no", ""),
        "repair_route": row.get("repair_route", ""),
        "object_class": object_class,
        "recommended_pass": _recommended_pass(object_class),
        "source_file": row.get("candidate_source_files", ""),
        "page": row.get("evidence_pages", ""),
        "tile_id": row.get("evidence_tiles", ""),
        "image_path": image.get("image_path", ""),
        "image_exists": image.get("image_exists", False),
        "image_source": image.get("image_source", ""),
        "current_candidate_evidence_ids": row.get("evidence_ids", ""),
        "feature_gap_families": ";".join(_specific_token_family(token) for token in sorted(answer_tokens - candidate_tokens)),
    }


def _recommended_pass(object_class: str) -> str:
    if object_class == "electrical_mep":
        return "electrical_mep"
    if object_class == "fixture_valve_schedule":
        return "fixture_valve_schedule"
    if object_class in {"door_window_demolition", "fixture_demolition", "demolition_node"}:
        return "demolition_node"
    if object_class.startswith("finish_"):
        return "finish_schedule"
    return "table_legend"


def _object_class_from_text(row: Mapping[str, Any]) -> str:
    text = _compact(" ".join([row.get("answer_section", ""), row.get("answer_item_name", ""), row.get("answer_feature", "")]))
    if any(term in text for term in ("电缆", "配线", "配管", "桥架", "灯具", "开关", "插座", "配电箱")):
        return "electrical_mep"
    if any(term in text for term in ("给水", "排水", "阀", "水表", "地漏", "马桶", "台盆", "龙头")):
        return "fixture_valve_schedule"
    if "拆除" in text:
        return "demolition_node"
    if any(term in text for term in ("地面", "地砖", "石材", "防水")):
        return "finish_floor"
    if any(term in text for term in ("吊顶", "天花", "灯槽", "窗帘盒")):
        return "finish_ceiling"
    if any(term in text for term in ("墙面", "隔墙", "抹灰")):
        return "finish_wall"
    return "general_object"


def _specific_token_family(token: str) -> str:
    lowered = token.lower()
    for prefix in ("sc", "mt", "jdg", "dn", "de"):
        if lowered.startswith(prefix):
            return prefix
    for family in ("wdzcbyj", "wdzcyjy", "wdzcyjv", "wdznyjy", "wdznyjv", "nhyjy", "nhyjv", "byj", "bv", "yjy", "yjv"):
        if lowered.startswith(family):
            return family
    if lowered in {"sus304", "ppr", "pvc", "pvc-u"}:
        return lowered
    if lowered in {"不锈钢", "铜质", "柔性铸铁", "铸铁", "镀锌钢管", "金属波纹软管"}:
        return "material"
    return lowered


def _group_task_rows(rows: Sequence[Mapping[str, Any]]) -> list[list[Mapping[str, Any]]]:
    grouped: dict[tuple[str, str, str, str, str], list[Mapping[str, Any]]] = defaultdict(list)
    for row in rows:
        key = (
            _text(row.get("recommended_pass")),
            _file_key(row.get("source_file")),
            _text(row.get("page")),
            _text(row.get("tile_id")),
            _text(row.get("image_path")),
        )
        grouped[key].append(row)
    result = list(grouped.values())
    result.sort(key=lambda group: (_text(group[0].get("recommended_pass")), _text(group[0].get("source_file")), _text(group[0].get("page")), _text(group[0].get("tile_id"))))
    return result


def _chunks(rows: Sequence[Mapping[str, Any]], size: int) -> list[list[Mapping[str, Any]]]:
    return [list(rows[index : index + size]) for index in range(0, len(rows), size)]


def _build_prompt(
    *,
    recommended_pass: str,
    source_file: str,
    page: str,
    tile_id: str,
    object_classes: str,
    feature_gap_families: str,
) -> str:
    guidance = PRECISION_PASS_GUIDANCE.get(recommended_pass, PRECISION_PASS_GUIDANCE["table_legend"])
    family_guidance = _family_guidance(feature_gap_families)
    return "\n".join(
        [
            "You are extracting precise bill-of-quantities feature evidence from a construction drawing image.",
            "Use only visible drawing evidence: notes, labels, schedules, legends, symbols, system diagrams, and material descriptions.",
            "Do not infer quantities. Do not copy expected answers. Do not invent specs that are not visible.",
            "Do not use placeholders such as 具体型号和规格, 按图纸型号, 可见文本, unclear, N/A, or unknown.",
            "Use Chinese for item_hint/spec_or_method/text when the drawing content is Chinese.",
            "Extract exact visible specification text. If a full model/diameter/material is not visible, leave spec_or_method empty and explain the visible evidence in text.",
            "",
            f"source_file: {source_file}",
            f"page: {page}",
            f"tile_id: {tile_id}",
            f"vision_pass: {recommended_pass}",
            f"object_classes: {object_classes}",
            f"feature_gap_families: {feature_gap_families}",
            f"focus: {guidance}",
            f"exact-spec focus: {family_guidance}",
            "",
            "Return JSON only, using this schema:",
            "{",
            '  "evidence_items": [',
            "    {",
            '      "evidence_role": "schedule|legend|system_diagram|note|symbol|unknown",',
            '      "discipline": "decoration|electrical|plumbing|unknown",',
            '      "item_hint": "concrete Chinese item/material name supported by visible evidence, or empty",',
            '      "spec_or_method": "exact visible material/spec/model/method text; no placeholders",',
            '      "suggested_unit": "m2|m|set|item|point or empty",',
            '      "text": "short quote or paraphrase of the visible drawing evidence",',
            '      "confidence": 0.0,',
            '      "needs_manual_review": true,',
            '      "reason": "why this evidence is relevant or why it is incomplete"',
            "    }",
            "  ]",
            "}",
        ]
    )


def _family_guidance(feature_gap_families: str) -> str:
    families = set(_split_values(feature_gap_families))
    hints: list[str] = []
    if {"sc", "mt", "jdg"} & families:
        hints.append("conduit type plus size, for example SC/MT/JDG followed by visible diameter or size")
    if {"dn", "de"} & families:
        hints.append("pipe/valve diameter with DN or De exactly as visible")
    if any(family in families for family in ("wdzcbyj", "wdzcyjy", "wdzcyjv", "yjy", "yjv", "byj", "bv")):
        hints.append("full wire/cable model, core count, section size, and installation method")
    if "material" in families or any(family in families for family in ("sus304", "ppr", "pvc", "pvc-u")):
        hints.append("material text such as stainless steel, copper, PPR/PVC, cast iron, galvanized steel, or equivalent visible wording")
    if not hints:
        hints.append("exact material code, size, substrate, finish code, model, or construction method visible in the image")
    return "; ".join(hints)


class _ImageResolver:
    def __init__(self, *, recall_plans: Sequence[Mapping[str, Any]], image_roots: Sequence[str | Path]) -> None:
        self.by_evidence_id: dict[str, str] = {}
        self.by_source_page_tile: dict[tuple[str, str, str], str] = {}
        self.image_by_evidence_id: dict[str, str] = {}
        self.root_images: list[tuple[str, str]] = []
        for plan in recall_plans:
            self._load_plan(plan)
        for root in image_roots:
            self._load_image_root(Path(root))

    def resolve(self, row: Mapping[str, Any]) -> dict[str, Any]:
        candidates = [
            self._by_evidence_id(row),
            self._by_source_page_tile(row),
            self._image_by_evidence_id(row),
            self._image_root_tile(row),
            self._image_root_source_page(row),
        ]
        for image_source, path in candidates:
            if path:
                image_path = str(path)
                return {
                    "image_path": image_path,
                    "image_exists": Path(image_path).exists(),
                    "image_source": image_source,
                }
        return {"image_path": "", "image_exists": False, "image_source": ""}

    def _load_plan(self, plan: Mapping[str, Any]) -> None:
        for row in plan.get("plan_rows") or []:
            if not isinstance(row, Mapping):
                continue
            image_path = _first(row, "image_path")
            if not image_path:
                continue
            evidence_id = _first(row, "evidence_id", "current_candidate_evidence_ids", "evidence_ids")
            source_key = _source_page_tile_key(_first(row, "source_file"), _first(row, "page"), _first(row, "tile_id"))
            if evidence_id:
                for value in _split_values(evidence_id):
                    self.by_evidence_id.setdefault(value, image_path)
            if source_key:
                self.by_source_page_tile.setdefault(source_key, image_path)

    def _load_image_root(self, root: Path) -> None:
        if not root.exists():
            return
        for path in sorted(root.rglob("*")):
            if path.suffix.lower() not in {".png", ".jpg", ".jpeg"}:
                continue
            compact_name = _compact(path.stem)
            self.root_images.append((compact_name, str(path)))
            for evidence_id in _evidence_ids_from_text(path.name):
                self.image_by_evidence_id.setdefault(evidence_id, str(path))

    def _by_evidence_id(self, row: Mapping[str, Any]) -> tuple[str, str]:
        for evidence_id in _split_values(row.get("evidence_ids")):
            if evidence_id in self.by_evidence_id:
                return "recall_plan_evidence_id", self.by_evidence_id[evidence_id]
        return "", ""

    def _by_source_page_tile(self, row: Mapping[str, Any]) -> tuple[str, str]:
        for source_file in _split_values(row.get("candidate_source_files")) or [_first(row, "candidate_source_files")]:
            for page in _split_values(row.get("evidence_pages")) or [_first(row, "evidence_pages")]:
                for tile in _split_values(row.get("evidence_tiles")) or [_first(row, "evidence_tiles")]:
                    key = _source_page_tile_key(source_file, page, tile)
                    if key in self.by_source_page_tile:
                        return "recall_plan_source_page_tile", self.by_source_page_tile[key]
        return "", ""

    def _image_by_evidence_id(self, row: Mapping[str, Any]) -> tuple[str, str]:
        for evidence_id in _split_values(row.get("evidence_ids")):
            if evidence_id in self.image_by_evidence_id:
                return "image_root_evidence_id", self.image_by_evidence_id[evidence_id]
        return "", ""

    def _image_root_tile(self, row: Mapping[str, Any]) -> tuple[str, str]:
        source_prefixes = _source_file_prefixes(row.get("candidate_source_files"))
        for tile_id in _split_values(row.get("evidence_tiles")):
            compact_tile = _compact(tile_id)
            if not compact_tile:
                continue
            for image_name, image_path in self.root_images:
                if compact_tile not in image_name:
                    continue
                if source_prefixes and not _image_name_matches_source_prefix(image_name, source_prefixes):
                    continue
                return "image_root_tile", image_path
        return "", ""

    def _image_root_source_page(self, row: Mapping[str, Any]) -> tuple[str, str]:
        pages = _split_values(row.get("evidence_pages")) or [""]
        for source_file in _split_values(row.get("candidate_source_files")):
            source_key = _compact_file_stem(source_file)
            if not source_key:
                continue
            for page in pages:
                for image_name, image_path in self.root_images:
                    if source_key in image_name and _image_name_matches_page(image_name, page):
                        return "image_root_source_page", image_path
        return "", ""


def _source_page_tile_key(source_file: Any, page: Any, tile_id: Any) -> tuple[str, str, str]:
    return (_compact_file(source_file), _text(page), _text(tile_id))


def _file_key(value: Any) -> str:
    return _compact_file(value)


def _compact_file(value: Any) -> str:
    return _compact(Path(str(value or "")).name)


def _compact_file_stem(value: Any) -> str:
    return _compact(Path(str(value or "")).stem)


def _image_name_matches_page(image_name: str, page: Any) -> bool:
    page_text = _text(page)
    if not page_text:
        return True
    try:
        page_no = int(float(page_text))
    except ValueError:
        return True
    return f"p{page_no:03d}" in image_name or image_name.endswith(str(page_no)) or f"-{page_no}" in image_name


def _source_file_prefixes(value: Any) -> list[str]:
    prefixes: list[str] = []
    for source_file in _split_values(value):
        match = re.match(r"\s*(\d{1,2})[.、_-]", Path(source_file).name)
        if match:
            prefixes.append(match.group(1).zfill(2))
    return _unique(prefixes)


def _image_name_matches_source_prefix(image_name: str, prefixes: Sequence[str]) -> bool:
    if not prefixes:
        return True
    return any(image_name.startswith(prefix) or image_name.startswith(f"{prefix}.") or image_name.startswith(f"{prefix}_") for prefix in prefixes)


def _evidence_ids_from_text(value: str) -> list[str]:
    return re.findall(r"(?:PDFEV|PDFGAP|PDFCAP)-[A-Z0-9-]+", value)


def _split_values(value: Any) -> list[str]:
    if value is None:
        return []
    if isinstance(value, (list, tuple, set)):
        return [str(item).strip() for item in value if str(item).strip()]
    text = str(value or "").strip()
    if not text:
        return []
    return [part.strip() for part in re.split(r"[;；,，]+", text) if part.strip()]


def _join_unique(values: Sequence[Any]) -> str:
    return ";".join(_unique(_split_values(";".join(_text(value) for value in values))))


def _unique(values: Sequence[str]) -> list[str]:
    seen: set[str] = set()
    result: list[str] = []
    for value in values:
        text = str(value or "").strip()
        if not text or text in seen:
            continue
        seen.add(text)
        result.append(text)
    return result


def _first(row: Mapping[str, Any], *keys: str) -> str:
    for key in keys:
        value = row.get(key)
        if value not in (None, ""):
            return str(value).strip()
    return ""


def _text(value: Any) -> str:
    return str(value or "").strip()


def _compact(value: Any) -> str:
    return "".join(ch.lower() for ch in str(value or "") if ch.isalnum() or ch in {"-", "+"})


def _safe_file_name(value: str) -> str:
    return re.sub(r"[^A-Za-z0-9._-]+", "_", value).strip("_") or "prompt.txt"


def _write_csv(path: Path, rows: Sequence[Mapping[str, Any]], headers: Sequence[str]) -> None:
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=list(headers), extrasaction="ignore")
        writer.writeheader()
        for row in rows:
            writer.writerow({key: _csv_value(row.get(key, "")) for key in headers})


def _csv_value(value: Any) -> str:
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, (dict, list)):
        return json.dumps(value, ensure_ascii=False)
    return str(value or "")


def _build_markdown(report: Mapping[str, Any]) -> str:
    summary = report.get("summary") or {}
    lines = [
        "# BIZ-2x PDF 精确规格补召回包",
        "",
        f"- 生成时间：{report.get('generated_at', '-')}",
        f"- 选中缺陷：{summary.get('selected_defect_count', 0)}",
        f"- 视觉调用：{summary.get('capture_call_count', 0)}",
        f"- 有图调用：{summary.get('image_exists_call_count', 0)}",
        f"- 缺图调用：{summary.get('missing_image_call_count', 0)}",
        f"- target_fields_in_prompt：{summary.get('target_fields_in_prompt')}",
        f"- 工程量状态：{summary.get('quantity_status')}",
        "",
        "## 说明",
        "",
        "- 本包用于补项目特征中的精确规格/材质证据，不用于工程量。",
        "- Prompt 不包含人工答案、目标规格或目标单位；只要求模型提取图纸中可见的精确规格。",
        "- 回灌证据必须填入 `blank_evidence_template` 的 evidence_* 字段，不能把答案字段当证据。",
        "",
        "## 输出",
        "",
    ]
    for key, value in (report.get("outputs") or {}).items():
        lines.append(f"- {key}: `{value}`")
    return "\n".join(lines) + "\n"


def _write_workbook(path: Path, report: Mapping[str, Any]) -> None:
    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "summary"
    _write_summary_sheet(summary_sheet, report.get("summary") or {})
    _write_rows_sheet(workbook.create_sheet("capture_tasks"), report.get("capture_rows") or [], CAPTURE_HEADERS)
    _write_rows_sheet(workbook.create_sheet("blank_evidence_template"), report.get("evidence_template_rows") or [], EVIDENCE_TEMPLATE_HEADERS)
    workbook.save(path)


def _write_summary_sheet(sheet: Any, summary: Mapping[str, Any]) -> None:
    sheet.append(SUMMARY_HEADERS)
    for key, value in summary.items():
        sheet.append([key, _csv_value(value)])
    _style_sheet(sheet)


def _write_rows_sheet(sheet: Any, rows: Sequence[Mapping[str, Any]], headers: Sequence[str]) -> None:
    sheet.append(list(headers))
    for row in rows:
        sheet.append([_csv_value(row.get(header, "")) for header in headers])
    _style_sheet(sheet)


def _style_sheet(sheet: Any) -> None:
    fill = PatternFill("solid", fgColor="1F4E78")
    font = Font(color="FFFFFF", bold=True)
    for cell in sheet[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for column in sheet.columns:
        column_letter = get_column_letter(column[0].column)
        max_length = min(max(len(str(cell.value or "")) for cell in column) + 2, 60)
        sheet.column_dimensions[column_letter].width = max(12, max_length)


def _json_safe(value: Any) -> Any:
    if isinstance(value, Mapping):
        return {str(key): _json_safe(item) for key, item in value.items()}
    if isinstance(value, list):
        return [_json_safe(item) for item in value]
    if isinstance(value, Path):
        return str(value)
    return value
