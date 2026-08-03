from __future__ import annotations

import json
import sys
from pathlib import Path

from openpyxl import load_workbook

from app.services.drawing_pdf_feature_precision_capture_pack import (
    build_feature_precision_capture_pack,
    write_feature_precision_capture_pack_outputs,
)
from app.services.drawing_pdf_feature_precision_capture_runner import (
    run_feature_precision_capture_pack,
    write_feature_precision_capture_run_outputs,
)
from scripts import biz2x_pdf_feature_precision_capture_pack
from scripts import biz2x_pdf_feature_precision_capture_run


def test_feature_precision_capture_pack_is_answer_blind(tmp_path: Path):
    image_path = tmp_path / "R03-PDFEV-000021.png"
    image_path.write_bytes(b"image")
    report = build_feature_precision_capture_pack(
        {"defect_rows": _defect_rows()},
        recall_plans=[_recall_plan(image_path)],
    )

    assert report["summary"]["source_defect_count"] == 3
    assert report["summary"]["selected_defect_count"] == 2
    assert report["summary"]["capture_call_count"] == 1
    assert report["summary"]["target_fields_in_prompt"] is False
    row = report["capture_rows"][0]
    assert row["defect_nos"] == "60;61"
    assert row["target_fields_in_prompt"] is False
    assert row["image_exists"] is True
    assert "Do not use placeholders" in row["prompt_text"]
    assert "conduit type plus size" in row["prompt_text"]

    serialized = json.dumps(report, ensure_ascii=False)
    assert "SC40" not in serialized
    assert "SC50" not in serialized
    assert "SECRET_TARGET" not in serialized
    assert "镀锌钢管" not in serialized


def test_feature_precision_capture_pack_resolves_image_root_by_tile_and_source_prefix(tmp_path: Path):
    image_root = tmp_path / "images"
    image_root.mkdir()
    wrong_pdf_image = image_root / "04.some_c_c7404304_p001_g03_r02_c03.png"
    expected_image = image_root / "03.some_09fa8fe4_p001_g03_r02_c03.png"
    wrong_pdf_image.write_bytes(b"wrong")
    expected_image.write_bytes(b"right")
    report = build_feature_precision_capture_pack(
        {
            "defect_rows": [
                {
                    "defect_no": 70,
                    "repair_route": "feature_enrichment",
                    "object_class": "electrical_mep",
                    "candidate_source_files": "03.整理完毕1017信达资产职工餐厅施工图(cad2pdf)(1).pdf",
                    "evidence_pages": "1",
                    "evidence_tiles": "p001_g03_r02_c03",
                    "answer_item_name": "SECRET_TARGET 电气配管",
                    "answer_feature": "镀锌钢管 SC40",
                    "candidate_item_name": "电气配管",
                    "candidate_feature": "具体型号和规格",
                }
            ]
        },
        image_roots=[image_root],
    )

    assert report["summary"]["selected_defect_count"] == 1
    assert report["summary"]["image_exists_call_count"] == 1
    row = report["capture_rows"][0]
    assert row["image_exists"] is True
    assert row["image_path"] == str(expected_image)
    assert "SC40" not in row["prompt_text"]


def test_feature_precision_capture_pack_outputs_prompt_and_template(tmp_path: Path):
    image_path = tmp_path / "R03-PDFEV-000021.png"
    image_path.write_bytes(b"image")
    report = build_feature_precision_capture_pack(
        {"defect_rows": _defect_rows()},
        recall_plans=[_recall_plan(image_path)],
    )

    outputs = write_feature_precision_capture_pack_outputs(report, tmp_path / "out", stem="precision")

    assert Path(outputs["json"]).exists()
    assert Path(outputs["capture_csv"]).exists()
    assert Path(outputs["evidence_template_csv"]).exists()
    assert Path(outputs["xlsx"]).exists()
    prompt_files = sorted(Path(outputs["prompt_dir"]).glob("*.txt"))
    assert len(prompt_files) == 1
    prompt_text = prompt_files[0].read_text(encoding="utf-8")
    assert "SC40" not in prompt_text
    assert "SECRET_TARGET" not in prompt_text

    workbook = load_workbook(outputs["xlsx"])
    assert workbook.sheetnames == ["summary", "capture_tasks", "blank_evidence_template"]
    capture_sheet = workbook["capture_tasks"]
    capture_headers = {cell.value: cell.column for cell in capture_sheet[1]}
    assert capture_sheet.cell(row=2, column=capture_headers["defect_nos"]).value == "60;61"
    assert capture_sheet.cell(row=2, column=capture_headers["target_fields_in_prompt"]).value == "false"

    template_sheet = workbook["blank_evidence_template"]
    template_headers = {cell.value: cell.column for cell in template_sheet[1]}
    assert template_sheet.cell(row=2, column=template_headers["call_no"]).value == "1"
    assert template_sheet.cell(row=2, column=template_headers["evidence_text"]).value is None


def test_feature_precision_capture_pack_cli(tmp_path: Path, monkeypatch, capsys):
    image_path = tmp_path / "R03-PDFEV-000021.png"
    image_path.write_bytes(b"image")
    defect_router_json = tmp_path / "defect_router.json"
    recall_plan_json = tmp_path / "recall_plan.json"
    defect_router_json.write_text(json.dumps({"defect_rows": _defect_rows()}, ensure_ascii=False), encoding="utf-8")
    recall_plan_json.write_text(json.dumps(_recall_plan(image_path), ensure_ascii=False), encoding="utf-8")
    output_dir = tmp_path / "cli_out"
    monkeypatch.setattr(
        sys,
        "argv",
        [
            "biz2x_pdf_feature_precision_capture_pack.py",
            "--defect-router",
            str(defect_router_json),
            "--recall-plan-json",
            str(recall_plan_json),
            "--output-dir",
            str(output_dir),
            "--timestamp",
            "fixed",
        ],
    )

    assert biz2x_pdf_feature_precision_capture_pack.main() == 0
    payload = json.loads(capsys.readouterr().out)

    assert payload["ok"] is True
    assert payload["summary"]["selected_defect_count"] == 2
    assert payload["summary"]["capture_call_count"] == 1
    assert payload["summary"]["target_fields_in_prompt"] is False
    assert Path(payload["outputs"]["xlsx"]).exists()


def test_feature_precision_runner_execute_is_answer_blind(tmp_path: Path):
    pack = _capture_pack(tmp_path)
    calls = []

    async def fake_vision_client(base64_image, mime_type, **kwargs):
        calls.append({"base64_image": base64_image, "mime_type": mime_type, "kwargs": kwargs})
        return {
            "raw_content": '{"evidence_items":[]}',
            "evidence_items": [
                {
                    "evidence_role": "system_diagram",
                    "discipline": "electrical",
                    "item_hint": "电气配管",
                    "spec_or_method": "SC40",
                    "suggested_unit": "m",
                    "text": "系统图可见电气配管 SC40",
                    "confidence": 0.86,
                    "needs_manual_review": True,
                    "reason": "visible label",
                }
            ],
        }

    report = run_feature_precision_capture_pack(pack, execute=True, vision_client=fake_vision_client)

    assert report["phase"] == "BIZ-2x-pdf-feature-precision-capture-runner"
    assert report["summary"]["execute"] is True
    assert report["summary"]["evidence_count"] == 1
    assert report["summary"]["target_fields_sent_to_model"] is False
    assert report["call_rows"][0]["defect_nos"] == "60;61"
    assert report["call_rows"][0]["feature_gap_families"] == "sc;material"
    evidence = report["evidence_rows"][0]
    assert evidence["source_kind"] == "pdf_feature_precision_capture_llm"
    assert evidence["task_no"] == "60"
    assert evidence["task_nos"] == "60;61"
    assert evidence["defect_nos"] == "60;61"
    assert evidence["feature_gap_families"] == "sc;material"
    assert evidence["spec_or_method"] == "SC40"
    kwargs = calls[0]["kwargs"]
    assert kwargs["prompt_override"] == "answer blind feature prompt"
    assert kwargs["tile_context"]["capture_pack_answer_blind"] is True
    assert "target_item_name" not in kwargs["tile_context"]
    assert "SECRET_TARGET" not in json.dumps(kwargs, ensure_ascii=False)


def test_feature_precision_runner_outputs_are_external_import_compatible(tmp_path: Path):
    report = run_feature_precision_capture_pack(_capture_pack(tmp_path), execute=True, vision_client=_fake_precision_client)

    outputs = write_feature_precision_capture_run_outputs(report, tmp_path / "out", stem="feature_run")

    assert Path(outputs["json"]).exists()
    assert Path(outputs["evidence_csv"]).exists()
    assert Path(outputs["xlsx"]).exists()
    payload = json.loads(Path(outputs["json"]).read_text(encoding="utf-8"))
    assert payload["phase"] == "BIZ-2x-pdf-feature-precision-capture-runner"
    workbook = load_workbook(outputs["xlsx"])
    assert workbook.sheetnames == ["run_summary", "capture_calls", "recall_evidence"]
    evidence_sheet = workbook["recall_evidence"]
    headers = {cell.value: cell.column for cell in evidence_sheet[1]}
    assert evidence_sheet.cell(row=2, column=headers["source_kind"]).value == "pdf_feature_precision_capture_llm"
    assert evidence_sheet.cell(row=2, column=headers["spec_or_method"]).value == "SC40"


def test_feature_precision_runner_cli_dry_run(tmp_path: Path, monkeypatch, capsys):
    pack_json = tmp_path / "capture_pack.json"
    pack_json.write_text(json.dumps(_capture_pack(tmp_path), ensure_ascii=False), encoding="utf-8")
    output_dir = tmp_path / "cli_run"
    monkeypatch.setattr(
        sys,
        "argv",
        [
            "biz2x_pdf_feature_precision_capture_run.py",
            "--capture-pack-json",
            str(pack_json),
            "--output-dir",
            str(output_dir),
            "--timestamp",
            "fixed",
        ],
    )

    assert biz2x_pdf_feature_precision_capture_run.main() == 0
    payload = json.loads(capsys.readouterr().out)

    assert payload["ok"] is True
    assert payload["phase"] == "BIZ-2x-pdf-feature-precision-capture-runner"
    assert payload["summary"]["execute"] is False
    assert payload["summary"]["capture_call_count"] == 1
    assert Path(payload["outputs"]["xlsx"]).exists()


def _defect_rows() -> list[dict[str, object]]:
    return [
        {
            "defect_no": 60,
            "repair_route": "feature_enrichment",
            "object_class": "electrical_mep",
            "candidate_source_files": "drawing.pdf",
            "evidence_pages": "1",
            "evidence_tiles": "p001_g03_r02_c03",
            "evidence_ids": "R03-PDFEV-000021",
            "answer_item_name": "SECRET_TARGET 电气配管",
            "answer_feature": "镀锌钢管 SC40",
            "candidate_item_name": "电气配管",
            "candidate_feature": "具体型号和规格",
        },
        {
            "defect_no": 61,
            "repair_route": "feature_enrichment",
            "object_class": "electrical_mep",
            "candidate_source_files": "drawing.pdf",
            "evidence_pages": "1",
            "evidence_tiles": "p001_g03_r02_c03",
            "evidence_ids": "R03-PDFEV-000021",
            "answer_item_name": "SECRET_TARGET 电气配管",
            "answer_feature": "镀锌钢管 SC50",
            "candidate_item_name": "电气配管",
            "candidate_feature": "按图纸型号",
        },
        {
            "defect_no": 62,
            "repair_route": "object_evidence_recall",
            "object_class": "finish_floor",
            "candidate_source_files": "drawing.pdf",
            "evidence_pages": "1",
            "evidence_tiles": "p001_g03_r02_c03",
            "evidence_ids": "R03-PDFEV-000021",
            "answer_item_name": "SECRET_TARGET 地砖拆除",
            "answer_feature": "地砖",
            "candidate_item_name": "",
            "candidate_feature": "",
        },
    ]


def _recall_plan(image_path: Path) -> dict[str, object]:
    return {
        "plan_rows": [
            {
                "evidence_id": "R03-PDFEV-000021",
                "source_file": "drawing.pdf",
                "page": "1",
                "tile_id": "p001_g03_r02_c03",
                "image_path": str(image_path),
            }
        ]
    }


async def _fake_precision_client(base64_image, mime_type, **kwargs):
    return {
        "raw_content": '{"evidence_items":[]}',
        "evidence_items": [
            {
                "evidence_role": "system_diagram",
                "discipline": "electrical",
                "item_hint": "电气配管",
                "spec_or_method": "SC40",
                "suggested_unit": "m",
                "text": "系统图可见电气配管 SC40",
                "confidence": 0.86,
                "needs_manual_review": True,
                "reason": "visible label",
            }
        ],
    }


def _capture_pack(tmp_path: Path) -> dict[str, object]:
    image = tmp_path / "tile.png"
    image.write_bytes(b"fake image")
    return {
        "capture_rows": [
            {
                "capture_no": 1,
                "recommended_pass": "electrical_mep",
                "source_file": "drawing.pdf",
                "page": 1,
                "tile_id": "p001_g03_r02_c03",
                "image_path": str(image),
                "defect_nos": "60;61",
                "object_classes": "electrical_mep",
                "feature_gap_families": "sc;material",
                "task_count": 2,
                "prompt_text": "answer blind feature prompt",
                "target_item_name": "SECRET_TARGET",
                "target_feature": "镀锌钢管 SC40",
            }
        ]
    }
