from __future__ import annotations

import json
import sys
from pathlib import Path

from openpyxl import load_workbook

from app.services.drawing_pdf_gap_recall_eval import (
    build_gap_recall_v2_evaluation,
    write_gap_recall_v2_evaluation_outputs,
)
from app.services.drawing_pdf_external_recall_template import (
    build_external_recall_template,
    write_external_recall_template_outputs,
)
from app.services.drawing_pdf_object_recall_pack import (
    build_object_recall_pack,
    write_object_recall_pack_outputs,
)
from app.services.drawing_pdf_v2_takeoff import build_pdf_v2_takeoff_report
from app.services.drawing_three_field_acceptance import ThreeFieldAnswerRow
from scripts import (
    biz2x_pdf_external_recall_acceptance_pipeline,
    biz2x_pdf_gap_recall_acceptance_pipeline,
    biz2x_pdf_gap_recall_eval,
)


def _base_v2_report() -> dict[str, object]:
    answer_rows = [
        ThreeFieldAnswerRow(
            sheet_name="manual",
            row_no=7,
            section="demolition",
            seq="1",
            item_code="",
            item_name="stainless glass door demolition",
            feature="remove stainless glass door, door frame, leaf, and hardware",
            unit="set",
        )
    ]
    return build_pdf_v2_takeoff_report(
        {
            "summary": {"pdf_file_count": 1, "pdf_page_count": 1, "pdf_render_status": "rendered"},
            "evidence_rows": [
                {
                    "evidence_id": "BASE-000001",
                    "source_file": "drawing.pdf",
                    "page": 1,
                    "tile_id": "p001_g03_r02_c02",
                    "evidence_type": "demolition",
                    "discipline": "decoration",
                    "raw_item_name": "floor tile demolition",
                    "spec_or_method": "remove floor tile",
                    "suggested_unit": "m2",
                    "evidence_text": "remove floor tile",
                    "confidence": 0.7,
                    "needs_review": True,
                }
            ],
        },
        answer_rows=answer_rows,
        style_prompt_text="test",
    )


def _recall_run_report() -> dict[str, object]:
    return {
        "phase": "BIZ-2x-pdf-gap-recall-visual-runner",
        "summary": {"execute": True, "evidence_count": 1},
        "evidence_rows": [
            {
                "evidence_id": "PDFGAP-000001",
                "source_kind": "pdf_gap_recall_llm",
                "source_file": "drawing.pdf",
                "page": 1,
                "tile_id": "p001_whole",
                "vision_pass": "door_window_demolition",
                "evidence_role": "construction_note",
                "discipline": "decoration",
                "item_hint": "stainless glass door demolition",
                "spec_or_method": "remove stainless glass door, door frame, leaf, and hardware",
                "suggested_unit": "set",
                "text": "stainless glass door demolition; remove frame, leaf, and hardware",
                "normalized_text": "stainless glass door demolition",
                "confidence": 0.86,
                "model": "test",
                "needs_manual_review": True,
                "reason": "door/window demolition note is visible",
            }
        ],
    }


def _recall_plan(tmp_path: Path) -> dict[str, object]:
    image = tmp_path / "whole.png"
    image.write_bytes(b"fake image bytes")
    return {
        "plan_rows": [
            {
                "task_no": 1,
                "gap_no": 1,
                "gap_priority": "P1_missing_core",
                "gap_type": "missing_candidate",
                "answer_item_name": "stainless glass door demolition",
                "answer_feature": "remove stainless glass door, door frame, leaf, and hardware",
                "answer_unit": "set",
                "recommended_pass": "door_window_demolition",
                "source_file": "drawing.pdf",
                "page": 1,
                "tile_id": "p001_whole",
                "tile_type": "whole_page_preview",
                "image_path": str(image),
            }
        ]
    }


def _object_recall_review_report() -> dict[str, object]:
    return {
        "ok": True,
        "phase": "BIZ-2x-pdf-three-field-human-review",
        "summary": {"answer_count": 1, "matched_three_fields_count": 0},
        "review_rows": [
            {
                "review_no": 1,
                "status": "missing_candidate",
                "answer_sheet": "manual",
                "answer_row_no": 7,
                "answer_section": "demolition",
                "answer_item_name": "stainless glass door demolition",
                "answer_feature": "remove stainless glass door, door frame, leaf, and hardware",
                "answer_unit": "set",
                "candidate_item_name": "",
                "candidate_unit": "",
                "candidate_source_files": "drawing.pdf",
                "evidence_pages": "1",
                "evidence_tiles": "p001_whole",
            }
        ],
    }


def test_gap_recall_eval_rebuilds_v2_and_improves_three_field_match():
    base = _base_v2_report()
    evaluation = build_gap_recall_v2_evaluation(base, _recall_run_report(), style_prompt_text="test")

    assert evaluation["summary"]["recall_evidence_count"] == 1
    assert evaluation["summary"]["matched_delta"] >= 1
    assert evaluation["augmented_summary"]["three_field_matched_count"] > base["summary"]["three_field_matched_count"]
    assert evaluation["augmented_summary"]["quantity_status"] == "deferred_until_three_fields_accepted"
    assert any(row["metric"] == "three_field_matched_count" for row in evaluation["metric_rows"])


def test_gap_recall_eval_writes_eval_and_augmented_v2_outputs(tmp_path: Path):
    evaluation = build_gap_recall_v2_evaluation(_base_v2_report(), _recall_run_report(), style_prompt_text="test")

    outputs = write_gap_recall_v2_evaluation_outputs(evaluation, tmp_path / "out", stem="eval")

    assert Path(outputs["json"]).exists()
    assert Path(outputs["metrics_csv"]).exists()
    assert Path(outputs["markdown"]).exists()
    assert Path(outputs["xlsx"]).exists()
    assert Path(outputs["augmented_v2_json"]).exists()
    assert Path(outputs["augmented_v2_xlsx"]).exists()
    workbook = load_workbook(outputs["xlsx"])
    assert workbook.sheetnames == ["eval_summary", "metric_delta"]
    assert workbook["metric_delta"]["A2"].value == "evidence_count"


def test_gap_recall_eval_cli(tmp_path: Path, monkeypatch, capsys):
    base_json = tmp_path / "base.json"
    run_json = tmp_path / "run.json"
    base_json.write_text(json.dumps(_base_v2_report(), ensure_ascii=False), encoding="utf-8")
    run_json.write_text(json.dumps(_recall_run_report(), ensure_ascii=False), encoding="utf-8")
    output_dir = tmp_path / "cli_out"
    monkeypatch.setattr(
        sys,
        "argv",
        [
            "biz2x_pdf_gap_recall_eval.py",
            "--base-v2-json",
            str(base_json),
            "--recall-run-json",
            str(run_json),
            "--output-dir",
            str(output_dir),
            "--timestamp",
            "fixed",
        ],
    )

    assert biz2x_pdf_gap_recall_eval.main() == 0
    payload = json.loads(capsys.readouterr().out)

    assert payload["ok"] is True
    assert payload["summary"]["matched_delta"] >= 1
    assert Path(payload["outputs"]["augmented_v2_xlsx"]).exists()


def test_gap_recall_acceptance_pipeline_cli(tmp_path: Path, monkeypatch, capsys):
    base_json = tmp_path / "base.json"
    run_json = tmp_path / "run.json"
    base_json.write_text(json.dumps(_base_v2_report(), ensure_ascii=False), encoding="utf-8")
    run_json.write_text(json.dumps(_recall_run_report(), ensure_ascii=False), encoding="utf-8")
    output_dir = tmp_path / "pipeline_out"
    monkeypatch.setattr(
        sys,
        "argv",
        [
            "biz2x_pdf_gap_recall_acceptance_pipeline.py",
            "--base-v2-json",
            str(base_json),
            "--recall-run-json",
            str(run_json),
            "--output-dir",
            str(output_dir),
            "--timestamp",
            "fixed",
        ],
    )

    assert biz2x_pdf_gap_recall_acceptance_pipeline.main() == 0
    payload = json.loads(capsys.readouterr().out)

    assert payload["ok"] is True
    assert payload["evaluation_summary"]["matched_delta"] >= 1
    assert payload["review_summary"]["review_row_count"] == 1
    assert payload["can_enable_quantity"] is True
    assert Path(payload["review_outputs"]["xlsx"]).exists()
    assert Path(payload["object_recall_outputs"]["xlsx"]).exists()
    assert Path(payload["object_workbench_outputs"]["xlsx"]).exists()
    assert Path(payload["object_capture_outputs"]["xlsx"]).exists()
    assert payload["object_capture_summary"]["target_fields_in_prompt"] is False
    assert Path(payload["gate_outputs"]["xlsx"]).exists()
    assert Path(payload["standard_bill_outputs"]["xlsx"]).exists()
    assert payload["standard_bill_summary"]["standard_bill_row_count"] >= 1
    assert Path(payload["quantity_placeholder_outputs"]["xlsx"]).exists()
    assert payload["quantity_placeholder_summary"]["quantity_filled_count"] == 0
    assert payload["closed_loop_summary"]["can_enable_quantity"] is True
    assert Path(payload["closed_loop_outputs"]["xlsx"]).exists()


def test_external_recall_acceptance_pipeline_cli(tmp_path: Path, monkeypatch, capsys):
    base_json = tmp_path / "base.json"
    external_json = tmp_path / "external.json"
    base_json.write_text(json.dumps(_base_v2_report(), ensure_ascii=False), encoding="utf-8")
    external_json.write_text(
        json.dumps({"evidence_rows": _recall_run_report()["evidence_rows"]}, ensure_ascii=False),
        encoding="utf-8",
    )
    output_dir = tmp_path / "external_pipeline_out"
    monkeypatch.setattr(
        sys,
        "argv",
        [
            "biz2x_pdf_external_recall_acceptance_pipeline.py",
            "--base-v2-json",
            str(base_json),
            "--external-results",
            str(external_json),
            "--output-dir",
            str(output_dir),
            "--source-name",
            "offline-glm-export",
            "--timestamp",
            "fixed",
            "--require-importable",
        ],
    )

    assert biz2x_pdf_external_recall_acceptance_pipeline.main() == 0
    payload = json.loads(capsys.readouterr().out)

    assert payload["ok"] is True
    assert payload["template_status_summary"]["importable_row_count"] == 1
    assert payload["template_status_summary"]["ready_for_external_import"] is True
    assert payload["import_summary"]["evidence_count"] == 1
    assert payload["evaluation_summary"]["matched_delta"] >= 1
    assert payload["can_enable_quantity"] is True
    assert "object_recall_summary" in payload
    assert payload["defect_router_summary"]["defect_count"] == 0
    assert Path(payload["template_status_outputs"]["xlsx"]).exists()
    assert Path(payload["import_outputs"]["xlsx"]).exists()
    assert Path(payload["review_outputs"]["xlsx"]).exists()
    assert Path(payload["defect_router_outputs"]["xlsx"]).exists()
    assert Path(payload["object_recall_outputs"]["xlsx"]).exists()
    assert Path(payload["object_workbench_outputs"]["xlsx"]).exists()
    assert Path(payload["object_capture_outputs"]["xlsx"]).exists()
    assert payload["object_capture_summary"]["target_fields_in_prompt"] is False
    assert Path(payload["gate_outputs"]["xlsx"]).exists()
    assert Path(payload["standard_bill_outputs"]["xlsx"]).exists()
    assert payload["standard_bill_summary"]["standard_bill_row_count"] >= 1
    assert Path(payload["quantity_placeholder_outputs"]["xlsx"]).exists()
    assert payload["quantity_placeholder_summary"]["quantity_filled_count"] == 0
    assert payload["closed_loop_summary"]["can_enable_quantity"] is True
    assert Path(payload["closed_loop_outputs"]["xlsx"]).exists()


def test_external_recall_acceptance_pipeline_quality_filter_drops_generic_rows(tmp_path: Path, monkeypatch, capsys):
    base_json = tmp_path / "base.json"
    external_json = tmp_path / "external.json"
    base_json.write_text(json.dumps(_base_v2_report(), ensure_ascii=False), encoding="utf-8")
    concrete = dict(_recall_run_report()["evidence_rows"][0])
    generic = {
        **concrete,
        "evidence_id": "PDFGAP-GENERIC",
        "item_hint": "Removed Finishes",
        "spec_or_method": "",
        "suggested_unit": "unknown",
        "text": "All existing finishes removed.",
        "normalized_text": "All existing finishes removed.",
    }
    external_json.write_text(
        json.dumps({"evidence_rows": [concrete, generic]}, ensure_ascii=False),
        encoding="utf-8",
    )
    output_dir = tmp_path / "external_pipeline_quality_out"
    monkeypatch.setattr(
        sys,
        "argv",
        [
            "biz2x_pdf_external_recall_acceptance_pipeline.py",
            "--base-v2-json",
            str(base_json),
            "--external-results",
            str(external_json),
            "--output-dir",
            str(output_dir),
            "--source-name",
            "offline-glm-export",
            "--timestamp",
            "quality",
            "--quality-filter",
            "--quality-include-review",
            "--require-importable",
        ],
    )

    assert biz2x_pdf_external_recall_acceptance_pipeline.main() == 0
    payload = json.loads(capsys.readouterr().out)

    assert payload["template_status_summary"]["importable_row_count"] == 2
    assert payload["quality_summary"]["accepted_row_count"] == 0
    assert payload["quality_summary"]["review_row_count"] == 1
    assert payload["quality_summary"]["rejected_row_count"] == 1
    assert payload["quality_summary"]["filtered_importable_row_count"] == 1
    assert payload["import_summary"]["input_row_count"] == 1
    assert payload["import_summary"]["evidence_count"] == 1
    assert Path(payload["quality_outputs"]["xlsx"]).exists()
    assert Path(payload["import_outputs"]["xlsx"]).exists()
    assert Path(payload["defect_router_outputs"]["xlsx"]).exists()


def test_external_recall_acceptance_pipeline_merges_multiple_external_results(tmp_path: Path, monkeypatch, capsys):
    base_json = tmp_path / "base.json"
    external_json_1 = tmp_path / "external_1.json"
    external_json_2 = tmp_path / "external_2.json"
    base_json.write_text(json.dumps(_base_v2_report(), ensure_ascii=False), encoding="utf-8")
    first = dict(_recall_run_report()["evidence_rows"][0])
    second = {**first, "item_hint": "second evidence", "text": "second evidence"}
    external_json_1.write_text(json.dumps({"evidence_rows": [first]}, ensure_ascii=False), encoding="utf-8")
    external_json_2.write_text(json.dumps({"evidence_rows": [second]}, ensure_ascii=False), encoding="utf-8")
    output_dir = tmp_path / "external_pipeline_multi_out"
    monkeypatch.setattr(
        sys,
        "argv",
        [
            "biz2x_pdf_external_recall_acceptance_pipeline.py",
            "--base-v2-json",
            str(base_json),
            "--external-results",
            str(external_json_1),
            "--external-results",
            str(external_json_2),
            "--output-dir",
            str(output_dir),
            "--source-name",
            "merged-offline-results",
            "--timestamp",
            "multi",
            "--require-importable",
        ],
    )

    assert biz2x_pdf_external_recall_acceptance_pipeline.main() == 0
    payload = json.loads(capsys.readouterr().out)

    assert payload["template_status_summary"]["input_row_count"] == 2
    assert payload["template_status_summary"]["importable_row_count"] == 2
    assert payload["import_summary"]["evidence_count"] == 2
    import_payload = json.loads(Path(payload["import_outputs"]["json"]).read_text(encoding="utf-8"))
    evidence_ids = [row["evidence_id"] for row in import_payload["evidence_rows"]]
    assert len(set(evidence_ids)) == 2
    assert evidence_ids[0].startswith("external_1__")
    assert evidence_ids[1].startswith("external_2__")
    assert Path(payload["import_outputs"]["xlsx"]).exists()


def test_external_recall_acceptance_pipeline_accepts_filled_object_pack(tmp_path: Path, monkeypatch, capsys):
    base_json = tmp_path / "base.json"
    base_json.write_text(json.dumps(_base_v2_report(), ensure_ascii=False), encoding="utf-8")
    object_pack = build_object_recall_pack(_object_recall_review_report())
    object_outputs = write_object_recall_pack_outputs(object_pack, tmp_path / "object_pack", stem="object_pack")
    workbook = load_workbook(object_outputs["xlsx"])
    sheet = workbook["object_recall_tasks"]
    headers = {cell.value: cell.column for cell in sheet[1]}
    sheet.cell(row=2, column=headers["evidence_item_hint"]).value = "stainless glass door demolition"
    sheet.cell(row=2, column=headers["evidence_spec_or_method"]).value = "remove stainless glass door, door frame, leaf, and hardware"
    sheet.cell(row=2, column=headers["evidence_suggested_unit"]).value = "set"
    sheet.cell(row=2, column=headers["evidence_text"]).value = "door schedule note: stainless glass door demolition; remove frame, leaf and hardware"
    workbook.save(object_outputs["xlsx"])
    workbook.close()
    output_dir = tmp_path / "object_pipeline_out"
    monkeypatch.setattr(
        sys,
        "argv",
        [
            "biz2x_pdf_external_recall_acceptance_pipeline.py",
            "--base-v2-json",
            str(base_json),
            "--external-results",
            object_outputs["xlsx"],
            "--output-dir",
            str(output_dir),
            "--source-name",
            "filled-object-pack",
            "--timestamp",
            "object",
            "--require-importable",
        ],
    )

    assert biz2x_pdf_external_recall_acceptance_pipeline.main() == 0
    payload = json.loads(capsys.readouterr().out)

    assert payload["template_status_summary"]["importable_row_count"] == 1
    assert payload["import_summary"]["evidence_count"] == 1
    assert payload["review_summary"]["status_counts"]["matched_three_fields"] == 1
    assert payload["defect_router_summary"]["defect_count"] == 0
    assert Path(payload["review_outputs"]["xlsx"]).exists()
    assert Path(payload["defect_router_outputs"]["xlsx"]).exists()
    assert Path(payload["object_workbench_outputs"]["xlsx"]).exists()
    assert Path(payload["object_capture_outputs"]["xlsx"]).exists()
    assert payload["object_capture_summary"]["target_fields_in_prompt"] is False
    assert payload["closed_loop_summary"]["current_stage_no"] in {2, 4, 6, 7}
    assert Path(payload["closed_loop_outputs"]["xlsx"]).exists()


def test_external_recall_acceptance_pipeline_passes_fallback_images_to_object_workbench(
    tmp_path: Path, monkeypatch, capsys
):
    base_json = tmp_path / "base.json"
    external_json = tmp_path / "external.json"
    fallback_image = tmp_path / "door_window.png"
    base_json.write_text(json.dumps(_base_v2_report(), ensure_ascii=False), encoding="utf-8")
    fallback_image.write_bytes(b"fake image bytes")
    external_json.write_text(
        json.dumps(
            {
                "evidence_rows": [
                    {
                        "evidence_id": "PDFGAP-UNRELATED",
                        "source_file": "drawing.pdf",
                        "page": 1,
                        "tile_id": "p001_whole",
                        "vision_pass": "door_window_demolition",
                        "item_hint": "unrelated loose furniture",
                        "spec_or_method": "movable chair",
                        "suggested_unit": "set",
                        "text": "movable chair note",
                        "confidence": 0.8,
                    }
                ]
            },
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )
    output_dir = tmp_path / "external_pipeline_fallback_out"
    monkeypatch.setattr(
        sys,
        "argv",
        [
            "biz2x_pdf_external_recall_acceptance_pipeline.py",
            "--base-v2-json",
            str(base_json),
            "--external-results",
            str(external_json),
            "--output-dir",
            str(output_dir),
            "--source-name",
            "fallback-image-check",
            "--timestamp",
            "fallback",
            "--require-importable",
            "--fallback-image",
            f"table_legend={fallback_image}",
        ],
    )

    assert biz2x_pdf_external_recall_acceptance_pipeline.main() == 0
    payload = json.loads(capsys.readouterr().out)

    assert payload["can_enable_quantity"] is False
    assert payload["object_workbench_summary"]["image_link_count"] == 1
    assert payload["object_capture_summary"]["image_exists_call_count"] == 1
    assert Path(payload["object_workbench_outputs"]["xlsx"]).exists()
    assert Path(payload["object_capture_outputs"]["xlsx"]).exists()


def test_external_recall_acceptance_pipeline_passes_task_images_to_object_workbench(
    tmp_path: Path, monkeypatch, capsys
):
    base_json = tmp_path / "base.json"
    external_json = tmp_path / "external.json"
    task_image = tmp_path / "task_1_zoom.png"
    base_json.write_text(json.dumps(_base_v2_report(), ensure_ascii=False), encoding="utf-8")
    task_image.write_bytes(b"fake task image")
    external_json.write_text(
        json.dumps(
            {
                "evidence_rows": [
                    {
                        "evidence_id": "PDFGAP-UNRELATED",
                        "source_file": "drawing.pdf",
                        "page": 1,
                        "tile_id": "p001_whole",
                        "vision_pass": "door_window_demolition",
                        "item_hint": "unrelated loose furniture",
                        "spec_or_method": "movable chair",
                        "suggested_unit": "set",
                        "text": "movable chair note",
                        "confidence": 0.8,
                    }
                ]
            },
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )
    output_dir = tmp_path / "external_pipeline_task_image_out"
    monkeypatch.setattr(
        sys,
        "argv",
        [
            "biz2x_pdf_external_recall_acceptance_pipeline.py",
            "--base-v2-json",
            str(base_json),
            "--external-results",
            str(external_json),
            "--output-dir",
            str(output_dir),
            "--source-name",
            "task-image-check",
            "--timestamp",
            "taskimage",
            "--require-importable",
            "--task-image",
            f"1={task_image}",
        ],
    )

    assert biz2x_pdf_external_recall_acceptance_pipeline.main() == 0
    payload = json.loads(capsys.readouterr().out)
    capture_payload = json.loads(Path(payload["object_capture_outputs"]["json"]).read_text(encoding="utf-8"))

    assert payload["object_workbench_summary"]["image_source_counts"] == {"task_image:1": 1}
    assert payload["object_capture_summary"]["image_exists_call_count"] == 1
    assert capture_payload["capture_rows"][0]["source_file"] == str(task_image)
    assert capture_payload["capture_rows"][0]["page"] == "task_image"
    assert capture_payload["capture_rows"][0]["tile_id"] == "task_image:1"
    assert capture_payload["capture_rows"][0]["target_fields_in_prompt"] is False
    assert "stainless glass door demolition" not in capture_payload["capture_rows"][0]["prompt_text"]


def test_external_recall_acceptance_pipeline_safe_stem_truncates_long_names():
    stem = biz2x_pdf_external_recall_acceptance_pipeline._safe_stem(
        "BIZ2x_PDF_external_recall_acceptance_" + "source_page_filtered_pipeline_" * 8,
        max_length=80,
    )

    assert len(stem) <= 80
    assert stem.startswith("BIZ2x_PDF_external_recall_acceptance")
    assert len(stem.rsplit("_", 1)[-1]) == 8


def test_external_recall_acceptance_pipeline_reports_blank_template_status(tmp_path: Path, monkeypatch, capsys):
    base_json = tmp_path / "base.json"
    plan_json = tmp_path / "plan.json"
    base_json.write_text(json.dumps(_base_v2_report(), ensure_ascii=False), encoding="utf-8")
    plan = _recall_plan(tmp_path)
    plan_json.write_text(json.dumps(plan, ensure_ascii=False), encoding="utf-8")
    template = build_external_recall_template(plan)
    template_outputs = write_external_recall_template_outputs(template, tmp_path / "template", stem="blank")
    output_dir = tmp_path / "blank_external_pipeline_out"
    monkeypatch.setattr(
        sys,
        "argv",
        [
            "biz2x_pdf_external_recall_acceptance_pipeline.py",
            "--base-v2-json",
            str(base_json),
            "--external-results",
            template_outputs["xlsx"],
            "--recall-plan-json",
            str(plan_json),
            "--output-dir",
            str(output_dir),
            "--source-name",
            "blank-template",
            "--timestamp",
            "fixed",
        ],
    )

    assert biz2x_pdf_external_recall_acceptance_pipeline.main() == 0
    payload = json.loads(capsys.readouterr().out)

    assert payload["ok"] is True
    assert payload["template_status_summary"]["input_row_count"] == 1
    assert payload["template_status_summary"]["importable_row_count"] == 0
    assert payload["template_status_summary"]["answer_only_count"] == 1
    assert payload["template_status_summary"]["ready_for_external_import"] is False
    assert payload["import_summary"]["evidence_count"] == 0
    assert payload["can_enable_quantity"] is False
    assert payload["object_recall_summary"]["object_recall_task_count"] >= 1
    assert payload["defect_router_summary"]["defect_count"] >= 1
    assert payload["closed_loop_summary"]["can_enable_quantity"] is False
    assert Path(payload["template_status_outputs"]["xlsx"]).exists()
    assert Path(payload["import_outputs"]["xlsx"]).exists()
    assert Path(payload["defect_router_outputs"]["xlsx"]).exists()
    assert Path(payload["object_recall_outputs"]["xlsx"]).exists()
    assert Path(payload["object_workbench_outputs"]["xlsx"]).exists()
    assert Path(payload["object_capture_outputs"]["xlsx"]).exists()
    assert payload["object_capture_summary"]["capture_call_count"] >= 1
    assert Path(payload["closed_loop_outputs"]["xlsx"]).exists()


def test_external_recall_acceptance_pipeline_require_importable_stops_blank_template(
    tmp_path: Path, monkeypatch, capsys
):
    base_json = tmp_path / "base.json"
    plan_json = tmp_path / "plan.json"
    base_json.write_text(json.dumps(_base_v2_report(), ensure_ascii=False), encoding="utf-8")
    plan = _recall_plan(tmp_path)
    plan_json.write_text(json.dumps(plan, ensure_ascii=False), encoding="utf-8")
    template = build_external_recall_template(plan)
    template_outputs = write_external_recall_template_outputs(template, tmp_path / "template", stem="blank")
    output_dir = tmp_path / "protected_external_pipeline_out"
    monkeypatch.setattr(
        sys,
        "argv",
        [
            "biz2x_pdf_external_recall_acceptance_pipeline.py",
            "--base-v2-json",
            str(base_json),
            "--external-results",
            template_outputs["xlsx"],
            "--recall-plan-json",
            str(plan_json),
            "--output-dir",
            str(output_dir),
            "--source-name",
            "blank-template",
            "--timestamp",
            "fixed",
            "--require-importable",
        ],
    )

    assert biz2x_pdf_external_recall_acceptance_pipeline.main() == 2
    payload = json.loads(capsys.readouterr().out)

    assert payload["ok"] is False
    assert payload["stopped_at"] == "template_status"
    assert payload["reason"] == "no_importable_external_recall_rows"
    assert payload["template_status_summary"]["ready_for_external_import"] is False
    assert payload["template_status_summary"]["importable_row_count"] == 0
    assert Path(payload["template_status_outputs"]["xlsx"]).exists()
    assert "import_outputs" not in payload
