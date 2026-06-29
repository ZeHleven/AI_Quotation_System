from __future__ import annotations

import json
import sys
from pathlib import Path

from openpyxl import load_workbook

from app.services.drawing_pdf_external_recall_template import (
    build_external_recall_template,
    write_external_recall_template_outputs,
)
from app.services.drawing_pdf_gap_recall_importer import (
    build_gap_recall_external_import_report,
    load_external_recall_results,
)
from scripts import biz2x_pdf_external_recall_template


def _recall_plan(tmp_path: Path) -> dict[str, object]:
    image = tmp_path / "page.png"
    image.write_bytes(b"fake image bytes")
    return {
        "plan_rows": [
            {
                "task_no": 1,
                "gap_no": 11,
                "gap_priority": "P1_missing_core",
                "gap_type": "missing_candidate",
                "recommended_pass": "door_window_demolition",
                "source_file": "drawing.pdf",
                "page": 1,
                "tile_id": "p001_whole",
                "tile_type": "whole_page_preview",
                "image_path": str(image),
                "answer_item_name": "door demolition",
                "answer_feature": "remove stainless glass door",
                "answer_unit": "set",
            },
            {
                "task_no": 2,
                "gap_no": 12,
                "gap_priority": "P2_missing_mep",
                "gap_type": "missing_candidate",
                "recommended_pass": "fixture_valve_schedule",
                "source_file": "drawing.pdf",
                "page": 1,
                "tile_id": "p001_whole",
                "tile_type": "whole_page_preview",
                "image_path": str(image),
                "answer_item_name": "water meter installation",
                "answer_feature": "DN25",
                "answer_unit": "group",
            },
        ]
    }


def test_external_recall_template_keeps_answers_reference_only(tmp_path: Path):
    report = build_external_recall_template(_recall_plan(tmp_path))

    assert report["summary"]["template_row_count"] == 2
    assert report["summary"]["unique_image_count"] == 1
    assert report["summary"]["safe_to_import_without_filling_evidence"] is False
    row = report["template_rows"][0]
    assert row["answer_item_name"] == "door demolition"
    assert row["item_hint"] == ""
    assert row["spec_or_method"] == ""
    assert row["suggested_unit"] == ""
    assert row["text"] == ""
    assert row["needs_manual_review"] == "true"


def test_external_recall_template_writes_outputs(tmp_path: Path):
    report = build_external_recall_template(_recall_plan(tmp_path))

    outputs = write_external_recall_template_outputs(report, tmp_path / "out", stem="template")

    assert Path(outputs["json"]).exists()
    assert Path(outputs["csv"]).exists()
    assert Path(outputs["markdown"]).exists()
    assert Path(outputs["xlsx"]).exists()
    workbook = load_workbook(outputs["xlsx"])
    assert workbook.sheetnames == ["summary", "readme", "external_recall_template"]
    assert workbook["external_recall_template"]["K2"].value == "door demolition"
    assert workbook["external_recall_template"]["P2"].value is None
    assert workbook["external_recall_template"]["AA1"].value == "image_exists"
    assert workbook["external_recall_template"]["AA2"].value is True
    assert workbook["external_recall_template"]["AB1"].value == "image_link"
    assert workbook["external_recall_template"]["AB2"].value == "open_image"
    assert workbook["external_recall_template"]["AB2"].hyperlink.target == str(Path(_recall_plan(tmp_path)["plan_rows"][0]["image_path"]))
    assert workbook["external_recall_template"]["AC1"].value == "fill_status_formula"
    assert str(workbook["external_recall_template"]["AC2"].value).startswith("=IF(")
    assert workbook["external_recall_template"].freeze_panes == "N2"
    assert len(workbook["external_recall_template"].data_validations.dataValidation) >= 5


def test_blank_external_recall_template_is_not_imported_as_evidence(tmp_path: Path):
    report = build_external_recall_template(_recall_plan(tmp_path))
    outputs = write_external_recall_template_outputs(report, tmp_path / "out", stem="template")
    external_results = load_external_recall_results(outputs["xlsx"])

    imported = build_gap_recall_external_import_report(
        external_results,
        recall_plan=_recall_plan(tmp_path),
        source_name="blank-template",
    )

    assert imported["summary"]["input_row_count"] == 2
    assert imported["summary"]["evidence_count"] == 0
    assert imported["summary"]["skipped_input_row_count"] == 2
    assert imported["summary"]["validation_status_counts"] == {"skipped": 2}


def test_external_recall_template_cli(tmp_path: Path, monkeypatch, capsys):
    plan_json = tmp_path / "plan.json"
    plan_json.write_text(json.dumps(_recall_plan(tmp_path), ensure_ascii=False), encoding="utf-8")
    output_dir = tmp_path / "cli_out"
    monkeypatch.setattr(
        sys,
        "argv",
        [
            "biz2x_pdf_external_recall_template.py",
            "--recall-plan-json",
            str(plan_json),
            "--output-dir",
            str(output_dir),
            "--timestamp",
            "fixed",
        ],
    )

    assert biz2x_pdf_external_recall_template.main() == 0
    payload = json.loads(capsys.readouterr().out)

    assert payload["ok"] is True
    assert payload["summary"]["template_row_count"] == 2
    assert payload["summary"]["safe_to_import_without_filling_evidence"] is False
    assert Path(payload["outputs"]["xlsx"]).exists()
