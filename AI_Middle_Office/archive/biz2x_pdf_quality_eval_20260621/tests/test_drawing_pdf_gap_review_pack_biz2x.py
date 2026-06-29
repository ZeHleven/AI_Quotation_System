from __future__ import annotations

import json
import sys
from pathlib import Path

from openpyxl import load_workbook

from app.services.drawing_pdf_gap_review_pack import (
    build_gap_review_evidence_pack,
    write_gap_review_evidence_pack,
)
from scripts import biz2x_pdf_gap_review_pack


def _source_report(tmp_path: Path, *, image_path: Path) -> Path:
    path = tmp_path / "source_evidence.json"
    path.write_text(
        json.dumps(
            {
                "tile_report": {
                    "tile_rows": [
                        {
                            "source_file": "drawing.pdf",
                            "tile_id": "p001_whole",
                            "image_path": str(image_path),
                        }
                    ]
                }
            },
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )
    return path


def _v2_report(source_report_name: str) -> dict[str, object]:
    return {
        "evidence_rows": [
            {
                "evidence_id": "PDFEV-000001",
                "evidence_type": "finish_schedule",
                "raw_item_name": "floor tile",
                "spec_or_method": "600x600",
                "evidence_text": "finish schedule floor tile 600x600",
                "source_file": "drawing.pdf",
                "page": 2,
                "tile_id": "p001_whole",
                "raw": {
                    "source_report_file": source_report_name,
                    "source_file": "drawing.pdf",
                    "tile_id": "p001_whole",
                    "page": 2,
                    "bbox_pdf": [1, 2, 3, 4],
                    "bbox_pixel": [10, 20, 30, 40],
                },
            }
        ],
        "three_field_gap_rows": [
            {
                "gap_priority": "P1_missing_core",
                "gap_type": "missing_candidate",
                "section": "floor",
                "answer_sheet": "answer",
                "answer_row_no": 8,
                "answer_item_name": "floor tile",
                "answer_feature": "600x600",
                "answer_unit": "m2",
                "candidate_row_no": "",
                "candidate_item_name": "",
                "candidate_unit": "",
                "candidate_standard_item_code": "",
                "candidate_standard_item_name": "",
                "candidate_evidence_ids": "PDFEV-000001",
                "suggested_next_action": "review source drawing",
            },
            {
                "gap_priority": "P2_missing",
                "gap_type": "missing_candidate",
                "answer_item_name": "wall paint",
                "candidate_evidence_ids": "",
            },
        ],
    }


def test_gap_review_pack_links_gap_rows_to_tile_images(tmp_path: Path):
    image_path = tmp_path / "tile.png"
    image_path.write_bytes(b"not-a-real-image-but-copyable")
    source_report = _source_report(tmp_path, image_path=image_path)
    report = _v2_report(source_report.name)

    pack = build_gap_review_evidence_pack(report, source_report_dir=tmp_path)

    assert pack["summary"]["gap_count"] == 1
    assert pack["summary"]["manifest_row_count"] == 1
    row = pack["manifest_rows"][0]
    assert row["gap_priority"] == "P1_missing_core"
    assert row["evidence_id"] == "PDFEV-000001"
    assert row["tile_image_path"] == str(image_path)
    assert row["bbox_pdf"] == "[1, 2, 3, 4]"


def test_gap_review_pack_writes_review_artifacts_and_copied_images(tmp_path: Path):
    image_path = tmp_path / "tile.png"
    image_path.write_bytes(b"copyable")
    source_report = _source_report(tmp_path, image_path=image_path)
    pack = build_gap_review_evidence_pack(_v2_report(source_report.name), source_report_dir=tmp_path)

    outputs = write_gap_review_evidence_pack(pack, tmp_path / "out", stem="pack")

    assert Path(outputs["json"]).exists()
    assert Path(outputs["manifest_csv"]).exists()
    assert Path(outputs["markdown"]).exists()
    assert Path(outputs["xlsx"]).exists()
    copied = list(Path(outputs["review_image_dir"]).glob("*.png"))
    assert len(copied) == 1
    workbook = load_workbook(outputs["xlsx"])
    assert workbook.sheetnames == ["gap_summary", "evidence_manifest"]
    assert workbook["gap_summary"]["A1"].value == "metric"
    assert workbook["evidence_manifest"]["A2"].value == 1


def test_gap_review_pack_cli_writes_from_v2_json(tmp_path: Path, monkeypatch, capsys):
    image_path = tmp_path / "tile.png"
    image_path.write_bytes(b"copyable")
    source_report = _source_report(tmp_path, image_path=image_path)
    v2_json = tmp_path / "v2.json"
    v2_json.write_text(json.dumps(_v2_report(source_report.name), ensure_ascii=False), encoding="utf-8")
    output_dir = tmp_path / "cli_out"
    monkeypatch.setattr(
        sys,
        "argv",
        [
            "biz2x_pdf_gap_review_pack.py",
            "--v2-json",
            str(v2_json),
            "--source-report-dir",
            str(tmp_path),
            "--output-dir",
            str(output_dir),
            "--priority-prefixes",
            "all",
            "--max-gaps",
            "1",
            "--timestamp",
            "fixed",
        ],
    )

    assert biz2x_pdf_gap_review_pack.main() == 0
    payload = json.loads(capsys.readouterr().out)

    assert payload["ok"] is True
    assert payload["summary"]["gap_count"] == 1
    assert Path(payload["outputs"]["xlsx"]).exists()


def test_gap_review_pack_cli_parses_priority_prefixes():
    assert biz2x_pdf_gap_review_pack._parse_priority_prefixes("P1,P2;P3") == ["P1", "P2", "P3"]
    assert biz2x_pdf_gap_review_pack._parse_priority_prefixes("all") == [""]
