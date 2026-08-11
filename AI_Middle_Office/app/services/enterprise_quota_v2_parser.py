from __future__ import annotations

import hashlib
import json
import re
from collections import Counter, defaultdict
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from io import BytesIO
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from app.services.enterprise_quota_units import normalize_enterprise_quota_unit


ENTERPRISE_QUOTA_V2_SCHEMA = "enterprise-quota-v2"
ENTERPRISE_SHEET = "企业定额库"
LABOR_SHEET = "人工价格库"
MATERIAL_SHEET = "材料价格库"
VALIDATION_SHEET = "消耗量校验报告"

SHEET_ALIASES = {
    ENTERPRISE_SHEET: (ENTERPRISE_SHEET, "企业定额"),
    LABOR_SHEET: (LABOR_SHEET, "劳务指导价"),
    MATERIAL_SHEET: (MATERIAL_SHEET,),
    VALIDATION_SHEET: (VALIDATION_SHEET,),
}

ENTERPRISE_HEADERS = (
    "定额编码",
    "类型",
    "项目名称",
    "项目特征及工作内容",
    "规格",
    "品牌",
    "单位",
    "含量",
    "单价",
    "人工费",
    "主材费",
    "辅材费",
    "机械费",
)
LABOR_HEADERS = (
    "编码",
    "类型",
    "项目名称",
    "工作内容",
    "计算规则",
    "单位",
    "含量",
    "不含税人工费",
)
MATERIAL_HEADERS = (
    "材料类别",
    "材料编码",
    "类型",
    "材料名称",
    "规格/型号",
    "厂家/品牌",
    "单位",
    "除税单价",
)
VALIDATION_HEADERS = (
    "定额编码",
    "专业",
    "章节",
    "项目名称",
    "单位",
    "人工费",
    "主材费",
    "辅材费",
    "机械费",
    "合计",
)

_MAJOR_SECTION_RE = re.compile(r"^[一二三四五六七八九十]+、")
_FORMULA_ERROR_PREFIX = "#"
_SHEET_NAME_RE = re.compile(r"(?:'([^']+)'|([\w\u4e00-\u9fff]+))!")
_COMPONENT_ROW_TYPES = frozenset({"人工", "主材", "辅材", "机械"})


class EnterpriseQuotaV2ParseError(ValueError):
    pass


def parse_enterprise_quota_v2_file(path: str | Path) -> dict[str, Any]:
    source = Path(path)
    if not source.is_file():
        raise EnterpriseQuotaV2ParseError(f"企业定额文件不存在: {source}")
    return parse_enterprise_quota_v2_bytes(source.read_bytes(), filename=source.name)


def parse_enterprise_quota_v2_bytes(content: bytes, *, filename: str = "enterprise_quota_v2.xlsx") -> dict[str, Any]:
    if not content:
        raise EnterpriseQuotaV2ParseError("企业定额文件为空")
    suffix = Path(filename).suffix.lower()
    if suffix not in {".xlsx", ".xlsm"}:
        raise EnterpriseQuotaV2ParseError("企业定额 2.0 仅支持 .xlsx/.xlsm")

    try:
        formula_book = load_workbook(BytesIO(content), data_only=False, read_only=False)
        value_book = load_workbook(BytesIO(content), data_only=True, read_only=False)
    except Exception as exc:  # pragma: no cover - openpyxl error types vary.
        raise EnterpriseQuotaV2ParseError(f"企业定额 2.0 Excel 解析失败: {exc}") from exc

    resolved_names = _resolve_sheet_names(formula_book.sheetnames)
    missing = [canonical for canonical in SHEET_ALIASES if canonical not in resolved_names]
    if missing:
        raise EnterpriseQuotaV2ParseError(f"缺少企业定额 2.0 Sheet: {', '.join(missing)}")

    formula_sheets = {
        canonical: formula_book[actual_name] for canonical, actual_name in resolved_names.items()
    }
    value_sheets = {
        canonical: value_book[actual_name] for canonical, actual_name in resolved_names.items()
    }
    _validate_headers(formula_sheets)

    workbook_rows: list[dict[str, Any]] = []
    formula_count = 0
    formula_error_cells: list[dict[str, Any]] = []
    workbook_metadata = {
        "schema": ENTERPRISE_QUOTA_V2_SCHEMA,
        "sheet_order": list(SHEET_ALIASES),
        "sheets": {},
        "defined_names": [
            {
                "name": item.name,
                "attr_text": item.attr_text,
                "local_sheet_id": item.localSheetId,
                "hidden": bool(item.hidden),
            }
            for item in formula_book.defined_names.values()
        ],
        "calculation": {
            "calc_mode": getattr(formula_book.calculation, "calcMode", None),
            "full_calc_on_load": getattr(formula_book.calculation, "fullCalcOnLoad", None),
            "force_full_calc": getattr(formula_book.calculation, "forceFullCalc", None),
        },
    }

    for sheet_order, canonical in enumerate(SHEET_ALIASES):
        formula_sheet = formula_sheets[canonical]
        value_sheet = value_sheets[canonical]
        sheet_rows, sheet_formula_count, sheet_errors, metadata = _extract_sheet_rows(
            formula_sheet,
            value_sheet,
            canonical_name=canonical,
            sheet_order=sheet_order,
        )
        workbook_rows.extend(sheet_rows)
        formula_count += sheet_formula_count
        formula_error_cells.extend(sheet_errors)
        workbook_metadata["sheets"][canonical] = metadata

    labor_resources = _parse_labor_resources(workbook_rows)
    material_resources = _parse_material_resources(workbook_rows)
    resources = labor_resources + material_resources
    resource_lookup = _resource_lookup(resources)

    sections, items, components = _parse_enterprise_rows(workbook_rows, resource_lookup)
    validation_rows = _parse_validation_rows(workbook_rows)
    _assign_row_entity_keys(workbook_rows, sections, items, components, resources)

    issues = _build_quality_issues(
        items=items,
        components=components,
        resources=resources,
        validation_rows=validation_rows,
        formula_error_cells=formula_error_cells,
        resource_lookup=resource_lookup,
    )
    severity_counts = Counter(issue["severity"] for issue in issues)
    summary = {
        "schema_version": ENTERPRISE_QUOTA_V2_SCHEMA,
        "sheet_count": len(SHEET_ALIASES),
        "major_section_count": sum(section["level"] == 1 for section in sections),
        "chapter_count": sum(section["level"] == 2 for section in sections),
        "section_count": len(sections),
        "quota_item_count": len(items),
        "component_count": len(components),
        "labor_resource_count": len(labor_resources),
        "material_resource_count": len(material_resources),
        "resource_count": len(resources),
        "validation_row_count": len(validation_rows),
        "formula_count": formula_count,
        "formula_error_cell_count": len(formula_error_cells),
        "unresolved_component_count": sum(
            component["formula_link_status"] == "unresolved" for component in components
        ),
        "ambiguous_component_count": sum(
            component["formula_link_status"] == "ambiguous" for component in components
        ),
        "error_count": severity_counts.get("error", 0),
        "warning_count": severity_counts.get("warning", 0),
    }
    quality_status = "blocked" if summary["error_count"] else ("warning" if summary["warning_count"] else "ready")
    workbook_title = _clean(workbook_rows[0]["values"].get("A")) if workbook_rows else None

    return {
        "ok": True,
        "schema_version": ENTERPRISE_QUOTA_V2_SCHEMA,
        "source": {
            "file_name": filename,
            "file_type": suffix.lstrip("."),
            "file_size": len(content),
            "sha256": hashlib.sha256(content).hexdigest(),
        },
        "workbook_title": workbook_title,
        "workbook_metadata": workbook_metadata,
        "summary": summary,
        "quality": {
            "status": quality_status,
            "blocker_count": summary["error_count"],
            "warning_count": summary["warning_count"],
            "issues": issues,
        },
        "sections": sections,
        "items": items,
        "components": components,
        "resources": resources,
        "validation_rows": validation_rows,
        "workbook_rows": workbook_rows,
    }


def _resolve_sheet_names(sheet_names: Iterable[str]) -> dict[str, str]:
    available = {str(name).strip(): str(name) for name in sheet_names}
    resolved: dict[str, str] = {}
    for canonical, aliases in SHEET_ALIASES.items():
        for alias in aliases:
            if alias in available:
                resolved[canonical] = available[alias]
                break
    return resolved


def _validate_headers(sheets: dict[str, Any]) -> None:
    expected = {
        ENTERPRISE_SHEET: ENTERPRISE_HEADERS,
        LABOR_SHEET: LABOR_HEADERS,
        MATERIAL_SHEET: MATERIAL_HEADERS,
        VALIDATION_SHEET: VALIDATION_HEADERS,
    }
    for sheet_name, headers in expected.items():
        actual = tuple(_clean(sheets[sheet_name].cell(2, index).value) for index in range(1, len(headers) + 1))
        if actual != headers:
            raise EnterpriseQuotaV2ParseError(
                f"{sheet_name} 表头不符合 2.0 合同，期望 {headers}，实际 {actual}"
            )


def _extract_sheet_rows(
    formula_sheet,
    value_sheet,
    *,
    canonical_name: str,
    sheet_order: int,
) -> tuple[list[dict[str, Any]], int, list[dict[str, Any]], dict[str, Any]]:
    max_column = {
        ENTERPRISE_SHEET: len(ENTERPRISE_HEADERS),
        LABOR_SHEET: len(LABOR_HEADERS),
        MATERIAL_SHEET: len(MATERIAL_HEADERS),
        VALIDATION_SHEET: len(VALIDATION_HEADERS),
    }[canonical_name]
    merges_by_row: defaultdict[int, list[str]] = defaultdict(list)
    merged_ranges = [str(merged) for merged in formula_sheet.merged_cells.ranges]
    for merged in formula_sheet.merged_cells.ranges:
        merges_by_row[int(merged.min_row)].append(str(merged))

    rows: list[dict[str, Any]] = []
    formula_count = 0
    formula_errors: list[dict[str, Any]] = []
    parent_context: dict[str, int | None] = {
        "major_section": None,
        "chapter": None,
        "quota_item": None,
    }
    for row_number in range(1, formula_sheet.max_row + 1):
        values: dict[str, Any] = {}
        formulas: dict[str, str] = {}
        styles: dict[str, Any] = {}
        for column_index in range(1, max_column + 1):
            column = get_column_letter(column_index)
            formula_cell = formula_sheet.cell(row_number, column_index)
            value_cell = value_sheet.cell(row_number, column_index)
            formula = _formula_text(formula_cell)
            value = _json_value(value_cell.value if formula else formula_cell.value)
            values[column] = value
            if formula:
                formulas[column] = formula
                formula_count += 1
                if isinstance(value, str) and value.startswith(_FORMULA_ERROR_PREFIX):
                    formula_errors.append(
                        {
                            "sheet": canonical_name,
                            "cell": f"{column}{row_number}",
                            "formula": formula,
                            "cached_value": value,
                        }
                    )
            styles[column] = _cell_style(formula_cell)

        row_kind, parent_row_number = _classify_row(
            canonical_name,
            row_number,
            values,
            formula_sheet.row_dimensions[row_number].outlineLevel or 0,
            parent_context,
        )
        if row_kind == "empty" and row_number > 2:
            continue
        row_dimension = formula_sheet.row_dimensions[row_number]
        rows.append(
            {
                "sheet_name": canonical_name,
                "sheet_order": sheet_order,
                "row_number": row_number,
                "row_kind": row_kind,
                "outline_level": int(row_dimension.outlineLevel or 0),
                "parent_row_number": parent_row_number,
                "values": values,
                "formulas": formulas,
                "styles": styles,
                "merge_ranges": merges_by_row.get(row_number, []),
                "row_height": _number_or_none(row_dimension.height),
                "hidden": bool(row_dimension.hidden),
                "collapsed": bool(row_dimension.collapsed),
                "entity_type": None,
                "entity_key": None,
            }
        )

    metadata = {
        "title": _clean(formula_sheet.cell(1, 1).value),
        "max_row": formula_sheet.max_row,
        "max_column": max_column,
        "freeze_panes": str(formula_sheet.freeze_panes) if formula_sheet.freeze_panes else None,
        "merged_ranges": merged_ranges,
        "auto_filter": str(formula_sheet.auto_filter.ref) if formula_sheet.auto_filter.ref else None,
        "show_grid_lines": bool(formula_sheet.sheet_view.showGridLines),
        "column_widths": {
            get_column_letter(index): _number_or_none(formula_sheet.column_dimensions[get_column_letter(index)].width)
            for index in range(1, max_column + 1)
        },
    }
    return rows, formula_count, formula_errors, metadata


def _classify_row(
    sheet_name: str,
    row_number: int,
    values: dict[str, Any],
    outline_level: int,
    parent_context: dict[str, int | None],
) -> tuple[str, int | None]:
    if row_number == 1:
        return "title", None
    if row_number == 2:
        return "header", None
    if not any(value not in (None, "") for value in values.values()):
        return "empty", None
    if sheet_name != ENTERPRISE_SHEET:
        return "data", None

    first_value = _clean(values.get("A"))
    row_type = _clean(values.get("B"))
    if row_type == "定额":
        parent = parent_context.get("chapter") or parent_context.get("major_section")
        parent_context["quota_item"] = row_number
        return "quota_item", parent
    if parent_context.get("quota_item") and (
        outline_level > 0 or row_type in _COMPONENT_ROW_TYPES
    ):
        return "component", parent_context.get("quota_item")
    if outline_level > 0:
        return "component", parent_context.get("quota_item")
    parent_context["quota_item"] = None
    if first_value and _MAJOR_SECTION_RE.match(first_value):
        parent_context["major_section"] = row_number
        parent_context["chapter"] = None
        return "major_section", None
    if first_value:
        parent_context["chapter"] = row_number
        return "chapter", parent_context.get("major_section")
    return "data", None


def _parse_labor_resources(workbook_rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    resources = []
    for row in _rows_for_sheet(workbook_rows, LABOR_SHEET, row_kind="data"):
        values = row["values"]
        if not any(values.get(column) not in (None, "") for column in "ABCDEFGH"):
            continue
        row_number = row["row_number"]
        resource_type = _resource_type(_clean(values.get("B")))
        resources.append(
            {
                "key": f"labor:{row_number}",
                "library_kind": "labor",
                "category": None,
                "resource_code": _clean(values.get("A")),
                "resource_type": resource_type,
                "resource_name": _clean(values.get("C")),
                "work_content": _clean(values.get("D")),
                "calculation_rule": _clean(values.get("E")),
                "specification": None,
                "brand": None,
                "unit": normalize_enterprise_quota_unit(values.get("F")),
                "default_quantity": _number_or_none(values.get("G")),
                "price": _number_or_none(values.get("H")),
                "source_sheet": LABOR_SHEET,
                "source_row_index": row_number,
                "formulas": row["formulas"],
                "raw_values": values,
            }
        )
    return resources


def _parse_material_resources(workbook_rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    resources = []
    for row in _rows_for_sheet(workbook_rows, MATERIAL_SHEET, row_kind="data"):
        values = row["values"]
        if not any(values.get(column) not in (None, "") for column in "ABCDEFGH"):
            continue
        row_number = row["row_number"]
        resources.append(
            {
                "key": f"material:{row_number}",
                "library_kind": "material",
                "category": _clean(values.get("A")),
                "resource_code": _clean(values.get("B")),
                "resource_type": _resource_type(_clean(values.get("C"))),
                "resource_name": _clean(values.get("D")),
                "work_content": None,
                "calculation_rule": None,
                "specification": _clean(values.get("E")),
                "brand": _clean(values.get("F")),
                "unit": normalize_enterprise_quota_unit(values.get("G")),
                "default_quantity": None,
                "price": _number_or_none(values.get("H")),
                "source_sheet": MATERIAL_SHEET,
                "source_row_index": row_number,
                "formulas": row["formulas"],
                "raw_values": values,
            }
        )
    return resources


def _resource_lookup(resources: list[dict[str, Any]]) -> dict[str, Any]:
    by_key = {resource["key"]: resource for resource in resources}
    by_library_name: defaultdict[tuple[str, str], list[dict[str, Any]]] = defaultdict(list)
    by_library_code: defaultdict[tuple[str, str], list[dict[str, Any]]] = defaultdict(list)
    for resource in resources:
        name = _clean(resource.get("resource_name"))
        code = _clean(resource.get("resource_code"))
        if name:
            by_library_name[(resource["library_kind"], name)].append(resource)
        if code:
            by_library_code[(resource["library_kind"], code)].append(resource)
    return {
        "by_key": by_key,
        "by_library_name": by_library_name,
        "by_library_code": by_library_code,
    }


def _parse_enterprise_rows(
    workbook_rows: list[dict[str, Any]],
    resource_lookup: dict[str, Any],
) -> tuple[list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]]]:
    sections: list[dict[str, Any]] = []
    items: list[dict[str, Any]] = []
    components: list[dict[str, Any]] = []
    section_by_row: dict[int, dict[str, Any]] = {}
    item_by_row: dict[int, dict[str, Any]] = {}
    current_major: dict[str, Any] | None = None
    current_chapter: dict[str, Any] | None = None
    current_item: dict[str, Any] | None = None

    major_count = 0
    chapter_counts: Counter[int] = Counter()
    for row in _rows_for_sheet(workbook_rows, ENTERPRISE_SHEET):
        values = row["values"]
        if row["row_kind"] == "major_section":
            major_count += 1
            current_major = {
                "key": f"section:{row['row_number']}",
                "section_code": f"S{major_count}",
                "section_name": _clean(values.get("A")),
                "level": 1,
                "outline_level": row["outline_level"],
                "parent_key": None,
                "source_sheet": ENTERPRISE_SHEET,
                "source_row_index": row["row_number"],
                "sort_order": len(sections) + 1,
                "raw_values": values,
            }
            current_chapter = None
            current_item = None
            sections.append(current_major)
            section_by_row[row["row_number"]] = current_major
            continue
        if row["row_kind"] == "chapter":
            major_index = major_count or 1
            chapter_counts[major_index] += 1
            current_chapter = {
                "key": f"section:{row['row_number']}",
                "section_code": f"S{major_index}.{chapter_counts[major_index]}",
                "section_name": _clean(values.get("A")),
                "level": 2,
                "outline_level": row["outline_level"],
                "parent_key": current_major["key"] if current_major else None,
                "source_sheet": ENTERPRISE_SHEET,
                "source_row_index": row["row_number"],
                "sort_order": len(sections) + 1,
                "raw_values": values,
            }
            current_item = None
            sections.append(current_chapter)
            section_by_row[row["row_number"]] = current_chapter
            continue
        if row["row_kind"] == "quota_item":
            current_item = {
                "key": f"item:{row['row_number']}",
                "section_key": (current_chapter or current_major or {}).get("key"),
                "quota_code": _clean(values.get("A")),
                "row_type": _clean(values.get("B")),
                "item_name": _clean(values.get("C")),
                "work_content": _clean(values.get("D")),
                "specification": _clean(values.get("E")),
                "brand": _clean(values.get("F")),
                "unit": normalize_enterprise_quota_unit(values.get("G")),
                "quantity": _number_or_none(values.get("H")),
                "unit_price": _number_or_none(values.get("I")),
                "labor_fee": _number_or_none(values.get("J")),
                "main_material_fee": _number_or_none(values.get("K")),
                "auxiliary_material_fee": _number_or_none(values.get("L")),
                "machinery_fee": _number_or_none(values.get("M")),
                "outline_level": row["outline_level"],
                "source_sheet": ENTERPRISE_SHEET,
                "source_row_index": row["row_number"],
                "sort_order": len(items) + 1,
                "formulas": row["formulas"],
                "raw_values": values,
            }
            items.append(current_item)
            item_by_row[row["row_number"]] = current_item
            continue
        if row["row_kind"] != "component":
            continue

        formulas_text = " ".join(row["formulas"].values())
        library_kind = _formula_library_kind(formulas_text, values)
        resource_name = _clean(values.get("C"))
        matches = list(
            resource_lookup["by_library_name"].get((library_kind, resource_name), [])
            if library_kind and resource_name
            else []
        )
        linked_resource = matches[0] if matches else None
        link_status = "linked"
        if not linked_resource:
            link_status = "unresolved"
        elif len(matches) > 1:
            link_status = "ambiguous"
        component_type = (
            _resource_type_label(linked_resource["resource_type"])
            if linked_resource
            else _clean(values.get("B"))
        )
        unit_price = linked_resource.get("price") if linked_resource else _number_or_none(values.get("I"))
        quantity = _number_or_none(values.get("H"))
        amount = _multiply(quantity, unit_price)
        component = {
            "key": f"component:{row['row_number']}",
            "item_key": current_item["key"] if current_item else None,
            "parent_quota_code": current_item.get("quota_code") if current_item else None,
            "resource_key": linked_resource["key"] if linked_resource else None,
            "resource_code": linked_resource.get("resource_code") if linked_resource else _clean(values.get("A")),
            "component_type": component_type,
            "resource_name": resource_name,
            "work_content": linked_resource.get("work_content") if linked_resource else _clean(values.get("D")),
            "specification": linked_resource.get("specification") if linked_resource else _clean(values.get("E")),
            "brand": linked_resource.get("brand") if linked_resource else _clean(values.get("F")),
            "unit": linked_resource.get("unit") if linked_resource else normalize_enterprise_quota_unit(values.get("G")),
            "quantity": quantity,
            "unit_price": unit_price,
            "amount": amount,
            "fee_bucket": _fee_bucket(component_type),
            "outline_level": row["outline_level"],
            "formula_library_kind": library_kind,
            "formula_link_status": link_status,
            "source_sheet": ENTERPRISE_SHEET,
            "source_row_index": row["row_number"],
            "sort_order": len(components) + 1,
            "formulas": row["formulas"],
            "raw_values": values,
        }
        components.append(component)
        if current_item:
            current_item.setdefault("component_keys", []).append(component["key"])

    _recalculate_parsed_items(items, components)
    return sections, items, components


def _parse_validation_rows(workbook_rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    result = []
    for row in _rows_for_sheet(workbook_rows, VALIDATION_SHEET, row_kind="data"):
        values = row["values"]
        code = _clean(values.get("A"))
        if not code:
            continue
        result.append(
            {
                "row_number": row["row_number"],
                "quota_code": code,
                "professional": _clean(values.get("B")),
                "chapter": _clean(values.get("C")),
                "item_name": _clean(values.get("D")),
                "unit": normalize_enterprise_quota_unit(values.get("E")),
                "labor_fee": _number_or_none(values.get("F")),
                "main_material_fee": _number_or_none(values.get("G")),
                "auxiliary_material_fee": _number_or_none(values.get("H")),
                "machinery_fee": _number_or_none(values.get("I")),
                "total": _number_or_none(values.get("J")),
                "raw_values": values,
            }
        )
    return result


def _recalculate_parsed_items(items: list[dict[str, Any]], components: list[dict[str, Any]]) -> None:
    component_by_item: defaultdict[str, list[dict[str, Any]]] = defaultdict(list)
    for component in components:
        if component["item_key"]:
            component_by_item[component["item_key"]].append(component)
    field_by_bucket = {
        "labor": "labor_fee",
        "main_material": "main_material_fee",
        "auxiliary_material": "auxiliary_material_fee",
        "machinery": "machinery_fee",
    }
    for item in items:
        sums = Counter()
        for component in component_by_item.get(item["key"], []):
            amount = _decimal_or_none(component.get("amount"))
            if amount is not None and component.get("fee_bucket"):
                sums[component["fee_bucket"]] += amount
        for bucket, field in field_by_bucket.items():
            item[field] = _float_or_none(sums[bucket])
        item["unit_price"] = _float_or_none(sum(sums.values(), Decimal("0")))


def _assign_row_entity_keys(
    workbook_rows: list[dict[str, Any]],
    sections: list[dict[str, Any]],
    items: list[dict[str, Any]],
    components: list[dict[str, Any]],
    resources: list[dict[str, Any]],
) -> None:
    by_sheet_row = {}
    for entity_type, rows in (
        ("section", sections),
        ("quota_item", items),
        ("component", components),
        ("resource", resources),
    ):
        for row in rows:
            by_sheet_row[(row["source_sheet"], row["source_row_index"])] = (
                entity_type,
                row["key"],
            )
    for row in workbook_rows:
        entity = by_sheet_row.get((row["sheet_name"], row["row_number"]))
        if entity:
            row["entity_type"], row["entity_key"] = entity


def _build_quality_issues(
    *,
    items: list[dict[str, Any]],
    components: list[dict[str, Any]],
    resources: list[dict[str, Any]],
    validation_rows: list[dict[str, Any]],
    formula_error_cells: list[dict[str, Any]],
    resource_lookup: dict[str, Any],
) -> list[dict[str, Any]]:
    issues: list[dict[str, Any]] = []
    item_codes: defaultdict[str, list[int]] = defaultdict(list)
    for item in items:
        if item["quota_code"]:
            item_codes[item["quota_code"]].append(item["source_row_index"])
        missing = [
            label
            for field, label in (
                ("quota_code", "定额编码"),
                ("item_name", "项目名称"),
                ("unit", "单位"),
            )
            if not item.get(field)
        ]
        if missing:
            issues.append(
                _issue(
                    "error",
                    "ITEM_REQUIRED_FIELD_MISSING",
                    f"定额主项缺少字段：{', '.join(missing)}",
                    sheet=ENTERPRISE_SHEET,
                    row=item["source_row_index"],
                    evidence={"quota_code": item.get("quota_code"), "item_name": item.get("item_name")},
                )
            )
    for code, rows in item_codes.items():
        if len(rows) > 1:
            issues.append(
                _issue(
                    "error",
                    "ITEM_CODE_DUPLICATED",
                    f"定额编码重复：{code}",
                    sheet=ENTERPRISE_SHEET,
                    row=rows[0],
                    evidence={"rows": rows},
                )
            )

    unresolved = [component for component in components if component["formula_link_status"] == "unresolved"]
    for component in unresolved:
        issues.append(
            _issue(
                "error",
                "FORMULA_RESOURCE_UNRESOLVED",
                f"组成行无法在{_library_label(component.get('formula_library_kind'))}找到资源：{component.get('resource_name') or '-'}",
                sheet=ENTERPRISE_SHEET,
                row=component["source_row_index"],
                evidence={
                    "quota_code": component.get("parent_quota_code"),
                    "resource_name": component.get("resource_name"),
                    "library_kind": component.get("formula_library_kind"),
                    "formulas": component.get("formulas"),
                },
            )
        )

    ambiguous_groups = []
    for (library_kind, name), matches in resource_lookup["by_library_name"].items():
        if len(matches) > 1:
            ambiguous_groups.append(
                {
                    "library_kind": library_kind,
                    "resource_name": name,
                    "rows": [item["source_row_index"] for item in matches],
                    "prices": [item["price"] for item in matches],
                }
            )
    if ambiguous_groups:
        issues.append(
            _issue(
                "warning",
                "FORMULA_LOOKUP_NAME_AMBIGUOUS",
                f"存在 {len(ambiguous_groups)} 组同名资源；按 Excel MATCH 规则使用首条记录",
                evidence={"groups": ambiguous_groups[:50]},
            )
        )

    duplicate_codes = []
    by_library_code: defaultdict[tuple[str, str], list[dict[str, Any]]] = defaultdict(list)
    for resource in resources:
        if resource.get("resource_code"):
            by_library_code[(resource["library_kind"], resource["resource_code"])].append(resource)
    for (library_kind, code), matches in by_library_code.items():
        if len(matches) > 1:
            duplicate_codes.append(
                {
                    "library_kind": library_kind,
                    "resource_code": code,
                    "rows": [item["source_row_index"] for item in matches],
                }
            )
    if duplicate_codes:
        issues.append(
            _issue(
                "warning",
                "RESOURCE_CODE_DUPLICATED",
                f"人工/材料价格库存在 {len(duplicate_codes)} 组重复编码",
                evidence={"groups": duplicate_codes[:50]},
            )
        )

    item_by_code = {item["quota_code"]: item for item in items if item.get("quota_code")}
    validation_mismatches = []
    for report in validation_rows:
        item = item_by_code.get(report["quota_code"])
        if not item:
            validation_mismatches.append(
                {"quota_code": report["quota_code"], "reason": "quota_item_missing"}
            )
            continue
        differences = {}
        for item_field, report_field in (
            ("labor_fee", "labor_fee"),
            ("main_material_fee", "main_material_fee"),
            ("auxiliary_material_fee", "auxiliary_material_fee"),
            ("machinery_fee", "machinery_fee"),
            ("unit_price", "total"),
        ):
            if not _numbers_equal(item.get(item_field), report.get(report_field), tolerance=Decimal("0.01")):
                differences[item_field] = {
                    "calculated": item.get(item_field),
                    "report": report.get(report_field),
                }
        if differences:
            validation_mismatches.append(
                {"quota_code": report["quota_code"], "differences": differences}
            )
    if validation_mismatches:
        issues.append(
            _issue(
                "warning",
                "VALIDATION_REPORT_MISMATCH",
                f"当前价格联动结果与静态校验报告有 {len(validation_mismatches)} 条差异",
                sheet=VALIDATION_SHEET,
                evidence={"count": len(validation_mismatches), "examples": validation_mismatches[:50]},
            )
        )

    if formula_error_cells:
        issues.append(
            _issue(
                "warning",
                "SOURCE_FORMULA_CACHED_ERROR",
                f"源工作簿缓存结果包含 {len(formula_error_cells)} 个公式错误；系统按结构化公式重新计算",
                sheet=ENTERPRISE_SHEET,
                evidence={"examples": formula_error_cells[:50]},
            )
        )
    return issues


def _rows_for_sheet(
    workbook_rows: list[dict[str, Any]],
    sheet_name: str,
    *,
    row_kind: str | None = None,
) -> list[dict[str, Any]]:
    return [
        row
        for row in workbook_rows
        if row["sheet_name"] == sheet_name and (row_kind is None or row["row_kind"] == row_kind)
    ]


def _formula_text(cell) -> str | None:
    if cell.data_type == "f":
        value = str(cell.value or "")
        return value if value.startswith("=") else f"={value}"
    value = cell.value
    if isinstance(value, str) and value.startswith("="):
        return value
    return None


def _formula_library_kind(formulas_text: str, values: dict[str, Any]) -> str | None:
    if LABOR_SHEET in formulas_text or "劳务指导价" in formulas_text:
        return "labor"
    if MATERIAL_SHEET in formulas_text:
        return "material"
    row_type = _clean(values.get("B"))
    if row_type in {"人工", "机械"}:
        return "labor"
    if row_type in {"主材", "辅材"}:
        return "material"
    return None


def _resource_type(value: str | None) -> str:
    mapping = {
        "人工": "labor",
        "主材": "main_material",
        "辅材": "auxiliary_material",
        "机械": "machinery",
    }
    return mapping.get(value or "", "unknown")


def _resource_type_label(value: str | None) -> str:
    mapping = {
        "labor": "人工",
        "main_material": "主材",
        "auxiliary_material": "辅材",
        "machinery": "机械",
    }
    return mapping.get(value or "", "未知")


def _fee_bucket(component_type: str | None) -> str | None:
    mapping = {
        "人工": "labor",
        "主材": "main_material",
        "辅材": "auxiliary_material",
        "机械": "machinery",
    }
    return mapping.get(component_type or "")


def _library_label(library_kind: str | None) -> str:
    return {"labor": LABOR_SHEET, "material": MATERIAL_SHEET}.get(library_kind or "", "价格库")


def _cell_style(cell) -> dict[str, Any]:
    font_color = _color_value(getattr(cell.font, "color", None))
    fill_color = _color_value(getattr(cell.fill, "fgColor", None))
    return {
        "style_id": int(cell.style_id or 0),
        "number_format": cell.number_format,
        "font": {
            "bold": bool(cell.font.bold),
            "italic": bool(cell.font.italic),
            "size": _number_or_none(cell.font.sz),
            "color": font_color,
        },
        "fill": fill_color,
        "alignment": {
            "horizontal": cell.alignment.horizontal,
            "vertical": cell.alignment.vertical,
            "wrap_text": bool(cell.alignment.wrap_text),
        },
    }


def _color_value(color) -> str | None:
    if color is None:
        return None
    color_type = getattr(color, "type", None)
    if color_type == "rgb":
        return getattr(color, "rgb", None)
    if color_type == "theme":
        theme = getattr(color, "theme", None)
        tint = getattr(color, "tint", None)
        return f"theme:{theme}:{tint}"
    if color_type == "indexed":
        return f"indexed:{getattr(color, 'indexed', None)}"
    return None


def _issue(
    severity: str,
    code: str,
    message: str,
    *,
    sheet: str | None = None,
    row: int | None = None,
    evidence: dict[str, Any] | None = None,
) -> dict[str, Any]:
    return {
        "severity": severity,
        "code": code,
        "message": message,
        "sheet": sheet,
        "row": row,
        "evidence": evidence or {},
    }


def _clean(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _json_value(value: Any) -> Any:
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    if isinstance(value, Decimal):
        return float(value)
    return value


def _decimal_or_none(value: Any) -> Decimal | None:
    if value in (None, "") or isinstance(value, bool):
        return None
    try:
        decimal_value = Decimal(str(value))
        return decimal_value if decimal_value.is_finite() else None
    except (InvalidOperation, TypeError, ValueError):
        return None


def _number_or_none(value: Any) -> float | None:
    decimal_value = _decimal_or_none(value)
    return float(decimal_value) if decimal_value is not None else None


def _float_or_none(value: Decimal | Any) -> float | None:
    decimal_value = _decimal_or_none(value)
    return float(decimal_value) if decimal_value is not None else None


def _multiply(left: Any, right: Any) -> float | None:
    left_value = _decimal_or_none(left)
    right_value = _decimal_or_none(right)
    if left_value is None or right_value is None:
        return None
    return float(left_value * right_value)


def _numbers_equal(left: Any, right: Any, *, tolerance: Decimal) -> bool:
    left_value = _decimal_or_none(left)
    right_value = _decimal_or_none(right)
    if left_value is None or right_value is None:
        return left_value is right_value
    return abs(left_value - right_value) <= tolerance


def compact_json(value: Any) -> str:
    return json.dumps(value, ensure_ascii=False, separators=(",", ":"), default=str)
