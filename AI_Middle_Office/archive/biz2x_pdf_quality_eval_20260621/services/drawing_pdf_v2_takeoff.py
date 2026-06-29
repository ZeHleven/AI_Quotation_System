from __future__ import annotations

import csv
import json
import re
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any, Iterable, Mapping, Sequence

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.services.drawing_three_field_acceptance import (
    ThreeFieldAnswerRow,
    ThreeFieldCandidateRow,
    build_three_field_acceptance_report,
)
from app.services.quantity_standard_index import (
    find_standard_item,
    LoadedStandardLibraryIndex,
    load_standard_library_index,
    search_standard_index,
)


PHASE = "BIZ-2x-pdf-v2-evidence-driven-takeoff"

EVIDENCE_HEADERS = [
    "evidence_id",
    "source_file",
    "page",
    "tile_id",
    "source_kind",
    "vision_pass",
    "evidence_role",
    "evidence_type",
    "discipline",
    "raw_item_name",
    "space",
    "material_codes",
    "spec_or_method",
    "suggested_unit",
    "evidence_text",
    "confidence",
    "needs_review",
]
NORMALIZED_HEADERS = [
    "evidence_group_id",
    "evidence_type",
    "discipline",
    "source_kind",
    "vision_pass",
    "evidence_role",
    "source_files",
    "raw_names",
    "space",
    "material_codes",
    "spec_or_method",
    "suggested_unit",
    "evidence_ids",
    "evidence_text",
]
HUMAN_ROW_HEADERS = [
    "row_id",
    "division",
    "item_name",
    "feature",
    "unit",
    "quantity",
    "standard_code",
    "standard_item_code",
    "standard_item_name",
    "standard_match_score",
    "evidence_ids",
    "source_files",
    "confidence",
    "review_note",
]
GAP_ROW_HEADERS = [
    "gap_type",
    "gap_priority",
    "answer_sheet",
    "answer_row_no",
    "section",
    "answer_item_name",
    "answer_feature",
    "answer_unit",
    "candidate_row_no",
    "candidate_source",
    "candidate_source_files",
    "candidate_evidence_ids",
    "candidate_standard_item_code",
    "candidate_standard_item_name",
    "candidate_item_name",
    "candidate_feature",
    "candidate_unit",
    "name_score",
    "feature_score",
    "unit_score",
    "issue",
    "suggested_next_action",
]


def build_pdf_v2_takeoff_report(
    pdf_direct_report: Mapping[str, Any],
    *,
    answer_rows: Sequence[ThreeFieldAnswerRow] | None = None,
    style_prompt_text: str = "",
    standard_index: LoadedStandardLibraryIndex | None = None,
) -> dict[str, Any]:
    index = standard_index or load_standard_library_index()
    pdf_summary = dict(pdf_direct_report.get("summary") or {})
    evidence_rows = extract_evidence_rows(pdf_direct_report)
    normalized_rows = normalize_evidence_rows(evidence_rows)
    human_rows = build_human_style_candidate_rows(normalized_rows, standard_index=index)
    candidate_rows = [
        ThreeFieldCandidateRow(
            source=row.get("source_files", ""),
            row_no=index_no,
            item_name=row.get("item_name", ""),
            feature=row.get("feature", ""),
            unit=row.get("unit", ""),
            quantity="",
            raw=json.dumps(_candidate_raw_payload(row), ensure_ascii=False, default=str),
        )
        for index_no, row in enumerate(human_rows, start=1)
    ]
    acceptance_report = None
    if answer_rows is not None:
        acceptance_report = build_three_field_acceptance_report(
            answer_rows=answer_rows,
            candidate_rows=candidate_rows,
            source_name="pdf_v2_evidence_driven_takeoff",
        )
    gap_rows = build_three_field_gap_rows(acceptance_report)

    summary = _build_summary(pdf_summary, evidence_rows, normalized_rows, human_rows, acceptance_report, gap_rows)
    return {
        "ok": True,
        "phase": PHASE,
        "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "safe_for_final_quantity_list": False,
        "quantity_acceptance_enabled": False,
        "style_prompt_summary": _style_prompt_summary(style_prompt_text),
        "summary": summary,
        "stage_results": _build_stage_results(summary, pdf_summary, acceptance_report),
        "pdf_direct_summary": pdf_summary,
        "evidence_rows": evidence_rows,
        "normalized_evidence_rows": normalized_rows,
        "human_style_rows": human_rows,
        "three_field_acceptance_report": acceptance_report,
        "three_field_gap_rows": gap_rows,
        "quantity_stage": {
            "status": "deferred",
            "reason": "当前升级只验收项目名称、项目特征、单位；工程量待三字段稳定后再启用。",
        },
    }


def extract_evidence_rows(pdf_direct_report: Mapping[str, Any]) -> list[dict[str, Any]]:
    raw_rows = _source_evidence_rows(pdf_direct_report)
    rows: list[dict[str, Any]] = []
    for index, raw in enumerate(raw_rows, start=1):
        existing_id = _first(raw, "evidence_id", "证据编号")
        role = _first(raw, "evidence_role", "source_kind", "角色")
        item_name = _first(
            raw,
            "item_hint",
            "raw_item_name",
            "清单项目提示",
            "图纸项目名称",
            "item_name",
            "项目名称",
            "project_name",
        )
        spec = _first(raw, "规格/做法", "spec_or_method", "项目特征", "feature", "spec")
        evidence_text = _first(raw, "证据文本", "evidence_text", "text", "normalized_text", "识别依据", "reason")
        space = _first(raw, "空间/部位", "space", "部位")
        discipline = _normalize_discipline(_first(raw, "discipline", "专业"))
        material_codes = _material_codes(raw)
        unit = _normalize_unit(_first(raw, "建议单位", "suggested_unit", "单位", "unit"))
        source_file = _first(raw, "PDF文件", "source_file", "file_name")
        source_kind = _first(raw, "source_kind")
        vision_pass = _first(raw, "vision_pass", "recommended_pass", "prompt_mode")
        evidence_role = _first(raw, "evidence_role", "role")
        evidence_type = _first(raw, "evidence_type") or classify_evidence_type(
            " ".join([role, discipline, item_name, spec, evidence_text, " ".join(material_codes)])
        )
        if not spec and item_name and evidence_type not in {"context", "drawing_meta", "relation"}:
            spec = evidence_text
        if not item_name and evidence_type not in {"context", "drawing_meta", "relation"}:
            item_name = _first(raw, "text", "normalized_text")
        if not item_name and not spec and not evidence_text:
            continue
        if _is_low_signal_visual_evidence(
            item_name=item_name,
            spec=spec,
            evidence_text=evidence_text,
            material_codes=material_codes,
        ):
            continue
        rows.append(
            {
                "evidence_id": existing_id or f"PDFEVD-{index:06d}",
                "source_file": source_file,
                "page": _first(raw, "页码", "page"),
                "tile_id": _first(raw, "tile_id", "source_tile_id"),
                "source_kind": source_kind,
                "vision_pass": vision_pass,
                "evidence_role": evidence_role,
                "evidence_type": evidence_type,
                "discipline": discipline,
                "raw_item_name": item_name,
                "space": space,
                "material_codes": material_codes,
                "spec_or_method": spec,
                "suggested_unit": unit,
                "evidence_text": evidence_text,
                "confidence": _float(_first(raw, "置信度", "confidence"), 0.0),
                "needs_review": _boolish(_first(raw, "需人工复核", "needs_manual_review"), default=True),
                "raw": dict(raw),
            }
        )
    return rows


def _is_low_signal_visual_evidence(
    *,
    item_name: str,
    spec: str,
    evidence_text: str,
    material_codes: Sequence[str],
) -> bool:
    if _is_prompt_template_echo(item_name, spec, evidence_text):
        return True
    if _is_code_only_evidence(item_name, spec, evidence_text):
        return True
    if _is_terminal_symbol_evidence(item_name, spec, evidence_text):
        return True
    if material_codes:
        return False
    parts = [_normalize_text(value) for value in (item_name, spec, evidence_text) if _normalize_text(value)]
    if not parts:
        return True
    low_signal_tokens = {"t", "n", "e", "w", "s", "l", "x", "y"}
    return all(part in low_signal_tokens or re.fullmatch(r"[a-z]", part or "") for part in parts)


def _is_prompt_template_echo(item_name: str, spec: str, evidence_text: str) -> bool:
    text = _normalize_text(" ".join([item_name, spec, evidence_text]))
    return any(
        token in text
        for token in (
            "可能形成的清单项目提示不确定则为空",
            "空间或部位不确定则为空",
            "规格材料做法安装方式或构造说明",
            "识别到的原文",
        )
    )


def _is_code_only_evidence(item_name: str, spec: str, evidence_text: str) -> bool:
    parts = [_clean(value).upper() for value in (item_name, spec, evidence_text) if _clean(value)]
    if not parts:
        return False
    compact_parts = [re.sub(r"[\s,，。；;:：/\\()（）\[\]【】<>《》]+", "", part) for part in parts]
    return all(re.fullmatch(r"[A-Z]{1,5}-?\d{1,4}[A-Z]?", part or "") for part in compact_parts)


def _is_terminal_symbol_evidence(item_name: str, spec: str, evidence_text: str) -> bool:
    item = _normalize_text(item_name)
    if item not in {"接地端子", "中性线端子", "相线端子", "工作零线端子"}:
        return False
    detail_parts = [_normalize_text(value) for value in (spec, evidence_text) if _normalize_text(value)]
    if not detail_parts:
        return True
    low_signal_tokens = {"t", "n", "e", "w"}
    return all(part in low_signal_tokens for part in detail_parts)


def _source_evidence_rows(report: Mapping[str, Any]) -> list[Mapping[str, Any]]:
    direct_rows = (
        list(report.get("pdf_direct_item_rows") or [])
        or list(report.get("item_rows") or [])
        or list(report.get("drawing_items") or [])
    )
    if direct_rows:
        return direct_rows
    visual_report = report.get("visual_evidence_report") if isinstance(report.get("visual_evidence_report"), Mapping) else {}
    evidence_rows = (
        list(report.get("evidence_rows") or [])
        or list(visual_report.get("evidence_rows") or [])
        or list(report.get("pdf_evidence_rows") or [])
    )
    return evidence_rows


def normalize_evidence_rows(evidence_rows: Sequence[Mapping[str, Any]]) -> list[dict[str, Any]]:
    grouped: dict[str, list[Mapping[str, Any]]] = defaultdict(list)
    for row in evidence_rows:
        key = "|".join(
            [
                _clean(row.get("evidence_type")),
                _normalize_discipline(row.get("discipline")),
                _normalize_text(row.get("raw_item_name")),
                _normalize_text(row.get("spec_or_method")),
                _normalize_text(",".join(row.get("material_codes") or [])),
                _normalize_unit(row.get("suggested_unit")),
            ]
        )
        grouped[key].append(row)

    normalized: list[dict[str, Any]] = []
    for index, rows in enumerate(grouped.values(), start=1):
        material_codes = _unique(value for row in rows for value in (row.get("material_codes") or []))
        raw_names = _unique(row.get("raw_item_name") for row in rows)
        specs = _unique(row.get("spec_or_method") for row in rows)
        spaces = _unique(row.get("space") for row in rows)
        units = _unique(row.get("suggested_unit") for row in rows if row.get("suggested_unit"))
        evidence_ids = [str(row.get("evidence_id") or "") for row in rows if row.get("evidence_id")]
        normalized.append(
            {
                "evidence_group_id": f"PDFGRP-{index:06d}",
                "evidence_type": _dominant(row.get("evidence_type") for row in rows),
                "discipline": _dominant(row.get("discipline") for row in rows),
                "source_kind": _dominant(row.get("source_kind") for row in rows),
                "vision_pass": _dominant(row.get("vision_pass") for row in rows),
                "evidence_role": _dominant(row.get("evidence_role") for row in rows),
                "source_files": "；".join(_unique(row.get("source_file") for row in rows)),
                "raw_names": "；".join(raw_names),
                "space": "；".join(spaces),
                "material_codes": material_codes,
                "spec_or_method": "；".join(specs),
                "suggested_unit": units[0] if units else "",
                "evidence_ids": evidence_ids,
                "evidence_text": "；".join(_unique(row.get("evidence_text") for row in rows))[:1200],
                "confidence": max((_float(row.get("confidence"), 0.0) for row in rows), default=0.0),
                "needs_review": any(bool(row.get("needs_review")) for row in rows),
            }
        )
    return normalized


def build_human_style_candidate_rows(
    normalized_rows: Sequence[Mapping[str, Any]],
    *,
    standard_index: LoadedStandardLibraryIndex | None = None,
) -> list[dict[str, Any]]:
    index = standard_index or load_standard_library_index()
    rows: list[dict[str, Any]] = []
    for row in normalized_rows:
        if row.get("evidence_type") in {"context", "drawing_meta", "relation"}:
            continue
        item_name = _human_item_name(row)
        if not item_name:
            continue
        unit = _human_unit(row, item_name)
        feature = _human_feature(row)
        if _is_business_measure_rule(row):
            standard_candidates: list[dict[str, Any]] = []
            direct_candidate = {}
        else:
            query = " ".join([item_name, feature, row.get("evidence_text", ""), row.get("raw_names", "")])
            standard_candidates = search_standard_index(query, index=index, limit=3)
            direct_candidate = _direct_standard_candidate(item_name, feature, row, index=index)
            if direct_candidate:
                standard_candidates = [direct_candidate, *[candidate for candidate in standard_candidates if candidate.get("item_code") != direct_candidate.get("item_code")]]
            if not direct_candidate:
                standard_candidates = [
                    candidate
                    for candidate in standard_candidates
                    if _standard_candidate_compatible(item_name, feature, row, candidate)
                ]
        selected = direct_candidate or (standard_candidates[0] if standard_candidates else {})
        if not unit:
            unit = _normalize_unit(_first(selected, "unit", "selected_unit") or _first_option(selected.get("unit_options")))
        rows.append(
            {
                "row_id": f"PDFV2-{len(rows) + 1:06d}",
                "division": _division_for_type(str(row.get("evidence_type") or "")),
                "item_name": item_name,
                "feature": feature,
                "unit": unit,
                "quantity": "",
                "standard_code": selected.get("standard_code", ""),
                "standard_item_code": selected.get("item_code", ""),
                "standard_item_name": selected.get("item_name", ""),
                "standard_match_score": selected.get("score", ""),
                "standard_candidates": standard_candidates,
                "evidence_ids": "；".join(row.get("evidence_ids") or []),
                "source_files": row.get("source_files", ""),
                "confidence": row.get("confidence", 0),
                "review_note": "工程量未启用；需人工复核图纸证据和列项拆分。",
                "normalized_evidence": dict(row),
            }
        )
    return rows


def build_three_field_gap_rows(acceptance_report: Mapping[str, Any] | None) -> list[dict[str, Any]]:
    if not acceptance_report:
        return []
    answer_by_key = {
        (
            row.get("sheet_name", ""),
            row.get("row_no", 0),
            row.get("item_name", ""),
            row.get("feature", ""),
            row.get("unit", ""),
        ): row
        for row in acceptance_report.get("answer_rows") or []
    }
    candidate_by_row_no = {int(row.get("row_no") or 0): row for row in acceptance_report.get("candidate_rows") or []}
    rows: list[dict[str, Any]] = []
    for comparison in acceptance_report.get("comparison_rows") or []:
        status = str(comparison.get("status") or "")
        if status == "matched_three_fields":
            continue
        answer = answer_by_key.get(
            (
                comparison.get("answer_sheet", ""),
                comparison.get("answer_row_no", 0),
                comparison.get("answer_item_name", ""),
                comparison.get("answer_feature", ""),
                comparison.get("answer_unit", ""),
            ),
            {},
        )
        candidate_row = candidate_by_row_no.get(int(_float(comparison.get("candidate_row_no"), 0)))
        candidate_trace = _candidate_trace_payload(candidate_row)
        rows.append(
            {
                "gap_type": status,
                "gap_priority": _gap_priority(comparison, answer),
                "answer_sheet": comparison.get("answer_sheet", ""),
                "answer_row_no": comparison.get("answer_row_no", ""),
                "section": answer.get("section", ""),
                "answer_item_name": comparison.get("answer_item_name", ""),
                "answer_feature": comparison.get("answer_feature", ""),
                "answer_unit": comparison.get("answer_unit", ""),
                "candidate_row_no": comparison.get("candidate_row_no", ""),
                "candidate_source": comparison.get("candidate_source", ""),
                "candidate_source_files": candidate_trace.get("source_files", ""),
                "candidate_evidence_ids": candidate_trace.get("evidence_ids", ""),
                "candidate_standard_item_code": candidate_trace.get("standard_item_code", ""),
                "candidate_standard_item_name": candidate_trace.get("standard_item_name", ""),
                "candidate_item_name": comparison.get("candidate_item_name", ""),
                "candidate_feature": comparison.get("candidate_feature", ""),
                "candidate_unit": comparison.get("candidate_unit", ""),
                "name_score": comparison.get("name_score", ""),
                "feature_score": comparison.get("feature_score", ""),
                "unit_score": comparison.get("unit_score", ""),
                "issue": comparison.get("issue", ""),
                "suggested_next_action": _gap_next_action(comparison, answer),
            }
        )
    return rows


def _candidate_raw_payload(row: Mapping[str, Any]) -> dict[str, Any]:
    return {
        "row_id": row.get("row_id", ""),
        "division": row.get("division", ""),
        "item_name": row.get("item_name", ""),
        "unit": row.get("unit", ""),
        "standard_code": row.get("standard_code", ""),
        "standard_item_code": row.get("standard_item_code", ""),
        "standard_item_name": row.get("standard_item_name", ""),
        "evidence_ids": row.get("evidence_ids", ""),
        "source_files": row.get("source_files", ""),
        "confidence": row.get("confidence", 0),
    }


def _candidate_trace_payload(candidate_row: Mapping[str, Any] | None) -> dict[str, Any]:
    if not candidate_row:
        return {}
    try:
        payload = json.loads(str(candidate_row.get("raw") or "{}"))
        if isinstance(payload, dict):
            return payload
    except (json.JSONDecodeError, TypeError, ValueError):
        return {}
    return {}


def _gap_priority(comparison: Mapping[str, Any], answer: Mapping[str, Any]) -> str:
    status = str(comparison.get("status") or "")
    section = _clean(answer.get("section"))
    item_text = _normalize_text(" ".join([str(comparison.get("answer_item_name") or ""), str(comparison.get("answer_feature") or "")]))
    if status == "unit_conflict":
        return "P1_unit"
    if "关键细分差异" in _clean(comparison.get("issue")):
        return "P1_specificity"
    if status == "missing_candidate" and any(term in item_text for term in ("门", "窗", "洁具", "马桶", "台盆", "花洒", "龙头", "阀门", "水表", "地漏")):
        return "P1_missing_core"
    if status == "missing_candidate" and ("电" in section or "水" in section):
        return "P2_missing_mep"
    if status == "missing_candidate":
        return "P2_missing"
    if status == "matched_name_unit_feature_review":
        return "P2_feature"
    return "P3_review"


def _gap_next_action(comparison: Mapping[str, Any], answer: Mapping[str, Any]) -> str:
    status = str(comparison.get("status") or "")
    item_name = _normalize_text(" ".join([str(comparison.get("answer_item_name") or ""), str(comparison.get("answer_feature") or "")]))
    section = _clean(answer.get("section"))
    issue = _clean(comparison.get("issue"))
    if status == "missing_candidate":
        if any(term in item_name for term in ("门", "窗", "隔断", "隔墙", "洗手台", "洁具", "马桶", "花洒", "龙头", "阀门", "水表", "地漏")):
            return "补视觉证据：优先识别门窗表、洁具五金表、给排水图例或节点详图。"
        if any(term in item_name for term in ("ct", "st", "pm", "mr", "墙砖", "地砖", "瓷砖", "石材", "涂料", "美缝")):
            return "补材料表证据：需要识别材料编号、规格、墙地天部位及做法。"
        if "拆除" in item_name:
            return "补拆除对象证据：需要区分地面、墙面、门窗、洁具、天花和管线拆除。"
        if "电" in section or any(term in item_name for term in ("灯", "配线", "配管", "电缆", "开关", "插座")):
            return "补电气图例/设备表证据：需要具体型号、规格、功率或回路信息。"
        return "补图纸证据：当前候选不足以生成该人工清单项。"
    if status == "unit_conflict":
        return "复核单位口径：确认系统单位与人工清单单位/国标单位是否可等价或需保留差异。"
    if "关键细分差异" in issue:
        return "补细分特征证据：需要具体灯型、线缆规格、拆除对象或节点做法后才能通过。"
    if status == "matched_name_unit_feature_review":
        return "补项目特征：名称和单位基本接近，但做法、规格、材质或部位不足。"
    return "人工复核：判断是否误识别、漏项或候选拆分粒度不一致。"


def write_pdf_v2_takeoff_outputs(
    report: Mapping[str, Any],
    output_dir: str | Path,
    *,
    stem: str | None = None,
) -> dict[str, str]:
    target = Path(output_dir)
    target.mkdir(parents=True, exist_ok=True)
    file_stem = stem or f"BIZ2x_PDF_V2证据驱动列项_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
    json_path = target / f"{file_stem}.json"
    markdown_path = target / f"{file_stem}.md"
    evidence_csv = target / f"{file_stem}_图纸证据.csv"
    normalized_csv = target / f"{file_stem}_归一化证据.csv"
    human_csv = target / f"{file_stem}_人工清单风格候选.csv"
    gap_csv = target / f"{file_stem}_三字段缺口复核.csv"
    xlsx_path = target / f"{file_stem}.xlsx"
    outputs = {
        "json": str(json_path),
        "markdown": str(markdown_path),
        "evidence_csv": str(evidence_csv),
        "normalized_evidence_csv": str(normalized_csv),
        "human_style_csv": str(human_csv),
        "three_field_gap_csv": str(gap_csv),
        "xlsx": str(xlsx_path),
    }

    report_with_outputs = {**dict(report), "outputs": outputs}
    json_path.write_text(json.dumps(_json_safe(report_with_outputs), ensure_ascii=False, indent=2), encoding="utf-8")
    markdown_path.write_text(build_pdf_v2_takeoff_markdown(report_with_outputs), encoding="utf-8")
    _write_csv(evidence_csv, report.get("evidence_rows") or [], EVIDENCE_HEADERS)
    _write_csv(normalized_csv, report.get("normalized_evidence_rows") or [], NORMALIZED_HEADERS)
    _write_csv(human_csv, report.get("human_style_rows") or [], HUMAN_ROW_HEADERS)
    _write_csv(gap_csv, report.get("three_field_gap_rows") or [], GAP_ROW_HEADERS)
    _write_workbook(xlsx_path, report)
    return outputs


def build_pdf_v2_takeoff_markdown(report: Mapping[str, Any]) -> str:
    summary = report.get("summary") or {}
    lines = [
        "# BIZ-2x PDF V2 证据驱动列项报告",
        "",
        f"- 生成时间：{report.get('generated_at', '-')}",
        f"- 图纸证据：{summary.get('evidence_count', 0)}",
        f"- 归一化证据组：{summary.get('normalized_evidence_count', 0)}",
        f"- 人工清单风格候选：{summary.get('human_style_row_count', 0)}",
        f"- 国标已映射：{summary.get('standard_mapped_count', 0)}",
        f"- 三字段全匹配：{summary.get('three_field_matched_count', 0)}",
        f"- 三字段缺口复核：{summary.get('three_field_gap_count', 0)}",
        f"- 工程量状态：{summary.get('quantity_status', '-')}",
        "",
        "## 阶段状态",
        "",
    ]
    for stage in report.get("stage_results") or []:
        lines.append(f"- {stage.get('stage')}: {stage.get('status')} - {stage.get('summary')}")
    lines.extend(["", "## 候选预览", "", "| 项目名称 | 单位 | 国标项 | 证据 |", "| --- | --- | --- | --- |"])
    for row in (report.get("human_style_rows") or [])[:60]:
        lines.append(
            "| "
            + " | ".join(
                [
                    _md(row.get("item_name")),
                    _md(row.get("unit")),
                    _md(row.get("standard_item_name")),
                    _md(row.get("evidence_ids")),
                ]
            )
            + " |"
        )
    gap_rows = report.get("three_field_gap_rows") or []
    if gap_rows:
        lines.extend(["", "## 三字段缺口复核预览", "", "| 类型 | 人工项目 | 系统候选 | 下一步 |", "| --- | --- | --- | --- |"])
        for row in gap_rows[:40]:
            lines.append(
                "| "
                + " | ".join(
                    [
                        _md(row.get("gap_type")),
                        _md(row.get("answer_item_name")),
                        _md(row.get("candidate_item_name")),
                        _md(row.get("suggested_next_action")),
                    ]
                )
                + " |"
            )
    return "\n".join(lines) + "\n"


def classify_evidence_type(text: str) -> str:
    normalized = _normalize_text(text)
    raw = _clean(text).lower()
    if any(token in raw for token in ("room_name", "drawing_title", "drawing_code", "arrow_relation")):
        if "room_name" in raw:
            return "context"
        if "arrow_relation" in raw:
            return "relation"
        return "drawing_meta"
    if any(token in normalized for token in ("拆除", "拆", "清运")):
        return "demolition"
    if any(token in normalized for token in ("灯槽", "窗帘盒")):
        return "ceiling"
    if "门槛" in normalized:
        return "floor"
    if (
        any(token in raw for token in ("plumbing_spec", "plumbing"))
        or any(token in normalized for token in ("给水", "排水", "地漏", "阀", "水表", "洁具", "马桶", "台盆", "龙头", "管道"))
        or re.search(r"\b(?:dn|de)\s*\d+\b", raw, flags=re.IGNORECASE)
    ):
        return "plumbing"
    if any(token in normalized for token in ("淋浴隔断", "隔断", "隔墙")):
        return "wall"
    if any(token in raw for token in ("door_window_mark",)):
        return "door_window"
    if any(token in normalized for token in ("门", "窗")):
        return "door_window"
    if any(token in normalized for token in ("金属线条", "不锈钢线条", "黑色不锈钢线条", "收口线条")):
        return "wall"
    if (
        any(token in raw for token in ("electrical_spec", "device_symbol", "electrical"))
        or any(
            token in normalized
            for token in (
                "配电",
                "配管",
                "配线",
                "电缆",
                "电线",
                "灯",
                "插座",
                "开关",
                "桥架",
                "弱电",
                "wdzc",
                "byj",
                "yjv",
                "排气扇",
                "换气扇",
                "电热水器",
                "热水器",
            )
        )
        or re.search(r"\b(?:sc|mt|jdg)\s*\d+\b", raw, flags=re.IGNORECASE)
    ):
        return "electrical"
    if any(token in normalized for token in ("吊顶", "天棚", "天花", "灯槽", "窗帘盒")):
        return "ceiling"
    if any(token in normalized for token in ("铝扣板", "石膏板", "轻钢龙骨")) and not any(
        token in normalized for token in ("隔墙", "墙体")
    ):
        return "ceiling"
    if any(token in normalized for token in ("墙面", "墙砖", "包柱", "柱面", "墙柱", "装饰板", "墙布", "硬包")):
        return "wall"
    if any(token in normalized for token in ("地面", "地砖", "瓷砖", "石材", "踢脚", "门槛")):
        return "floor"
    if any(token in normalized for token in ("墙", "柱", "隔断", "隔墙", "墙布", "硬包", "石材", "玻璃")):
        return "wall"
    return "general"


def _human_item_name(row: Mapping[str, Any]) -> str:
    evidence_type = str(row.get("evidence_type") or "")
    raw = _clean(row.get("raw_names"))
    spec = _clean(row.get("spec_or_method"))
    text = " ".join([raw, spec, _clean(row.get("evidence_text"))])
    codes = list(row.get("material_codes") or [])
    code_suffix = f" {codes[0]}" if codes else ""
    lower = _normalize_text(text)

    if evidence_type == "demolition":
        target = _remove_words(raw or spec, ["拆除", "拆"])
        return f"拆除{target or '原有构件'}".strip()
    if evidence_type == "floor":
        if "挡水条" in lower:
            return f"人造石挡水条{code_suffix}".strip() if "人造石" in lower else f"挡水条{code_suffix}".strip()
        if "防水保护层" in lower or ("保护层" in lower and "防水" in lower):
            return f"防水保护层{code_suffix}".strip()
        if "零星砌筑" in lower or "砖砌地台" in lower or "地台" in lower:
            return f"零星砌筑{code_suffix}".strip()
        if "门槛" in lower:
            return f"门槛石{code_suffix}".strip()
        if any(term in lower for term in ("石材", "st")):
            return f"石材地面{code_suffix}".strip()
        if any(term in lower for term in ("瓷砖", "地砖", "ct", "块料")):
            return f"瓷砖地面{code_suffix}".strip()
        if "踢脚" in lower:
            return f"踢脚线{code_suffix}".strip()
        return f"{raw or '地面工程'}{code_suffix}".strip()
    if evidence_type == "ceiling":
        if "铝扣" in lower:
            return f"铝扣板吊顶{code_suffix}".strip()
        if "防潮无机涂料" in lower:
            return "黑色防潮无机涂料" if "黑色" in lower else "防潮无机涂料"
        if "白色无机涂料" in lower:
            return "白色无机涂料"
        if "木饰面" in lower and "天花" in lower:
            return f"木饰面天花吊顶{code_suffix}".strip()
        if "圆形" in lower and "灯槽" in lower:
            return f"圆形灯槽{code_suffix}".strip()
        if "灯槽" in lower:
            return f"灯槽{code_suffix}".strip()
        if "窗帘盒" in lower:
            return f"窗帘盒{code_suffix}".strip()
        if "石膏" in lower or "轻钢" in lower:
            prefix = "轻钢龙骨防水石膏板" if "防水" in lower else "轻钢龙骨石膏板"
            suffix = "造型吊顶" if any(term in lower for term in ("造型", "跌级", "二级")) else "平级吊顶"
            return f"{prefix}{suffix}{code_suffix}".strip()
        return f"{raw or '吊顶'}{code_suffix}".strip()
    if evidence_type == "wall":
        if "钢化玻璃造型柱" in lower or ("玻璃造型柱" in lower):
            return f"钢化玻璃造型柱{code_suffix}".strip()
        if "隔断底座" in lower or ("卡座" in lower and "底座" in lower):
            return f"隔断底座{code_suffix}".strip()
        if "陶粒回填" in lower:
            return f"陶粒回填{code_suffix}".strip()
        if "砖砌隔墙" in lower or "蒸压加气砼砌块墙" in lower or "新增隔墙" in lower:
            return f"砖砌隔墙{code_suffix}".strip()
        if "窗台石" in lower and "人造石" in lower:
            return f"人造石窗台石{code_suffix}".strip()
        if "淋浴隔断" in lower:
            return f"淋浴隔断{code_suffix}".strip()
        if "包柱" in lower and any(term in lower for term in ("瓷砖", "墙砖", "ct")):
            return f"瓷砖包柱{code_suffix}".strip()
        if "玻璃" in lower and any(term in lower for term in ("隔", "隔断", "隔墙")):
            return f"玻璃隔墙{code_suffix}".strip()
        if "墙布" in lower:
            return f"墙布墙面{code_suffix}".strip()
        if "硬包" in lower:
            return f"硬包墙面{code_suffix}".strip()
        if "瓷砖" in lower or "墙砖" in lower:
            return f"墙面瓷砖湿贴{code_suffix}".strip()
        if "石材" in lower:
            return f"墙面石材湿贴{code_suffix}".strip()
        if "抹灰" in lower:
            return f"墙面抹灰{code_suffix}".strip()
        return f"{raw or '墙面工程'}{code_suffix}".strip()
    if evidence_type == "door_window":
        if "玻璃" in lower and "门" in lower:
            prefix = "成品不锈钢玻璃门" if ("不锈钢" in lower or "mt01" in lower) else "成品玻璃门"
            if "双开" in lower and "双开" not in prefix:
                prefix = prefix.replace("玻璃门", "双开玻璃门")
            return f"{prefix}{code_suffix}".strip()
        if "铝合金" in lower and "窗" in lower:
            return f"铝合金窗{code_suffix}".strip()
        if "门" in lower:
            return f"{raw or '成品门'}{code_suffix}".strip()
        return f"{raw or '门窗工程'}{code_suffix}".strip()
    if evidence_type == "electrical":
        return _electrical_item_name(raw, " ".join([spec, _clean(row.get("evidence_text"))]), codes)
    if evidence_type == "plumbing":
        return _plumbing_item_name(raw, " ".join([spec, _clean(row.get("evidence_text"))]), codes)
    return f"{raw or spec}{code_suffix}".strip()


def _electrical_item_name(raw: str, spec: str, codes: Sequence[str]) -> str:
    text = " ".join([raw, spec, " ".join(codes)])
    lower = _normalize_text(text)
    spec_lower = _normalize_text(spec)
    cable_wire_code = _first_regex(text, r"(?:WDZC-|WDZN-|NH-|BV|BYJ|YJV|YJY)[A-Z0-9\-*xX\.]+")
    conduit_code = _first_regex(text, r"(?:SC|MT|JDG)\s*\d+")
    spec_code = cable_wire_code or conduit_code
    if "配电箱" in lower:
        return f"配电箱 {codes[0]}" if codes else "配电箱"
    if any(term in lower for term in ("排气扇", "换气扇", "通风扇")):
        return "排气扇安装"
    if any(term in lower for term in ("电热水器", "热水器")):
        return "电热水器供货及安装"
    if conduit_code and not cable_wire_code:
        return f"电气配管 {conduit_code}".strip()
    if "电缆" in lower:
        return f"电缆敷设 {spec_code}".strip()
    if "配线" in lower or "电线" in lower:
        return f"电气配线 {spec_code}".strip()
    if "配管" in lower or re.search(r"\b(?:sc|mt|jdg)\s*\d+", lower):
        return f"电气配管 {conduit_code or spec_code}".strip()
    if "开关" in spec_lower and "插座" not in spec_lower:
        return f"开关安装 {codes[0]}".strip() if codes else "开关安装"
    if "插座" in lower:
        return f"插座安装 {codes[0]}".strip() if codes else "插座安装"
    if "开关" in lower:
        return f"开关安装 {codes[0]}".strip() if codes else "开关安装"
    if "灯带" in lower:
        return f"LED灯带 {codes[0]}".strip() if codes else "LED灯带"
    if "筒灯" in lower:
        return f"LED筒灯 {codes[0]}".strip() if codes else "LED筒灯"
    if "灯" in lower:
        return f"灯具安装 {codes[0]}".strip() if codes else "灯具安装"
    return raw or "电气安装"


def _plumbing_item_name(raw: str, spec: str, codes: Sequence[str]) -> str:
    text = " ".join([raw, spec, " ".join(codes)])
    lower = _normalize_text(text)
    pipe_size = _first_regex(text, r"(?:DN|De)\s*\d+")
    if "给水" in lower and ("管" in lower or pipe_size):
        return f"给水管 {pipe_size}".strip()
    if "排水" in lower and ("管" in lower or pipe_size):
        return f"排水管 {pipe_size}".strip()
    if "地漏" in lower:
        return "地漏供货及安装"
    if "水表" in lower:
        return f"水表供货及安装 {pipe_size}".strip()
    if "阀" in lower:
        if any(term in lower for term in ("塑料", "ppr", "pvc", "pe")):
            return f"塑料阀门供货及安装 {pipe_size}".strip()
        if any(term in lower for term in ("金属", "不锈钢", "铜")):
            return f"金属阀门供货及安装 {pipe_size}".strip()
        return f"阀门供货及安装 {pipe_size}".strip()
    if any(term in lower for term in ("洗脸盆", "台盆", "面盆")):
        return "台盆供货及安装"
    if any(term in lower for term in ("梳妆镜", "镜面", "镜子")):
        return "梳妆镜供货及安装"
    if any(term in lower for term in ("厕纸架", "纸巾架", "卫生纸架")):
        return "厕纸架供货及安装"
    if any(term in lower for term in ("马桶", "坐便", "大便器")):
        return "马桶供货及安装"
    if "小便器" in lower:
        return "小便器供货及安装"
    if "浴缸" in lower:
        return "浴缸供货及安装"
    if "洗涤盆" in lower:
        return "洗涤盆供货及安装"
    if any(term in lower for term in ("花洒", "淋浴喷头")):
        return "淋浴花洒供货及安装"
    if "龙头" in lower:
        hot_cold = "冷热" in lower or ("冷水" in lower and "热水" in lower)
        return "冷热水龙头供货及安装" if hot_cold else "龙头供货及安装"
    if any(term in lower for term in ("洁具", "五金")):
        return raw or "洁具五金供货及安装"
    return raw or "给排水安装"


def _human_unit(row: Mapping[str, Any], item_name: str) -> str:
    suggested = _normalize_unit(row.get("suggested_unit"))
    item_text = _normalize_text(" ".join([item_name, _clean(row.get("spec_or_method")), _clean(row.get("evidence_text"))]))
    if "淋浴隔断" in item_text:
        return "㎡"
    if "窗台石" in item_text and suggested:
        return suggested
    if _door_window_unit(item_text):
        return _door_window_unit(item_text)
    if "配电箱" in item_text:
        return "套"
    if "排气扇安装" in item_text:
        return "个"
    if "电热水器供货及安装" in item_text:
        return "套"
    if any(
        term in item_text
        for term in (
            "马桶供货及安装",
            "小便器供货及安装",
            "台盆供货及安装",
            "淋浴花洒供货及安装",
            "花洒供货及安装",
            "冷热水龙头供货及安装",
            "龙头供货及安装",
            "浴缸供货及安装",
            "洗涤盆供货及安装",
        )
    ):
        return "套"
    if any(term in item_text for term in ("阀门供货及安装", "地漏供货及安装", "水表供货及安装", "厕纸架供货及安装", "梳妆镜供货及安装", "洁具五金供货及安装")):
        return "个"
    if any(term in item_text for term in ("灯槽", "窗帘盒", "踢脚")):
        return "m"
    if any(term in item_text for term in ("挡水条", "隔断底座")):
        return "m"
    if any(term in item_text for term in ("零星砌筑", "砖砌隔墙", "陶粒回填")):
        return "m³"
    if "钢化玻璃造型柱" in item_text:
        return "套"
    if any(term in item_text for term in ("开荒精保洁", "材料二次运输", "防水保护层", "无机涂料")):
        return "㎡"
    if any(term in item_text for term in ("灯具", "筒灯", "格栅灯", "射灯")) and "灯带" not in item_text:
        return "套"
    if suggested in {"㎡，m", "㎡,m", "㎡/m", "m/㎡"}:
        if any(term in item_text for term in ("地面", "墙面", "吊顶", "天棚", "隔墙", "隔断", "石材", "瓷砖", "墙布", "硬包")):
            return "㎡"
        if any(term in item_text for term in ("管", "线", "电缆", "配管", "配线", "灯带")):
            return "m"
    if suggested:
        return suggested
    text = item_text
    if any(term in text for term in ("地面", "墙面", "吊顶", "隔墙", "隔断", "石材", "瓷砖")):
        return "㎡"
    if any(term in text for term in ("防水保护层", "无机涂料", "开荒精保洁", "材料二次运输")):
        return "㎡"
    if any(term in text for term in ("零星砌筑", "砖砌隔墙", "陶粒回填")):
        return "m³"
    if any(term in text for term in ("挡水条", "隔断底座")):
        return "m"
    if any(term in text for term in ("管", "线", "电缆", "配管", "配线", "灯带", "踢脚", "灯槽", "窗帘盒")):
        return "m"
    if "水表" in text:
        return "个"
    if any(term in text for term in ("灯具", "筒灯", "格栅灯", "射灯")):
        return "套"
    if any(term in text for term in ("排气扇", "换气扇", "通风扇")):
        return "个"
    if any(term in text for term in ("电热水器", "热水器")):
        return "套"
    if any(term in text for term in ("洗脸盆", "大便器", "小便器", "浴缸", "洗涤盆", "洁具", "马桶", "台盆")):
        return "套"
    if any(term in text for term in ("插座", "开关", "地漏", "阀", "灯", "花洒", "龙头")):
        return "套"
    return ""


def _door_window_unit(text: str) -> str:
    if "淋浴隔断" in text:
        return "㎡"
    if "拆除" in text and any(term in text for term in ("门", "窗", "玻璃门", "实木门", "铝合金门")):
        return "套"
    if any(term in text for term in ("成品门", "玻璃门", "实木门", "铝合金门", "不锈钢门")):
        return "樘"
    if "门套" in text:
        return "m"
    if "窗台石" in text:
        return "m"
    if "售卖窗口" in text:
        return "个"
    return ""


def _human_feature(row: Mapping[str, Any]) -> str:
    parts: list[str] = []
    item_name = _human_item_name(row)
    if row.get("space"):
        parts.append(f"1. 部位：{row.get('space')}")
    codes = "、".join(row.get("material_codes") or [])
    spec = _clean(row.get("spec_or_method"))
    if codes or spec:
        parts.append(f"{len(parts)+1}. 材料/规格/做法：{'; '.join([value for value in [codes, spec] if value])}")
    for template in _human_feature_templates(item_name=item_name, spec=spec, evidence_text=_clean(row.get("evidence_text"))):
        parts.append(f"{len(parts)+1}. {template}")
    evidence = _clean(row.get("evidence_text"))
    if evidence:
        parts.append(f"{len(parts)+1}. {_feature_evidence_label(row)}：{evidence[:300]}")
    parts.append(f"{len(parts)+1}. {_feature_scope_text(row)}")
    return "\n".join(parts)


def _feature_evidence_label(row: Mapping[str, Any]) -> str:
    return "规则依据" if _is_business_measure_rule(row) else "图纸证据"


def _feature_scope_text(row: Mapping[str, Any]) -> str:
    if _is_business_measure_rule(row):
        return "措施范围：按项目整体交付与施工组织范围计入，需人工复核。"
    return "报价范围：按图纸及材料表完成供货、安装、辅材、收口及成品保护，需人工复核。"


def _is_business_measure_rule(row: Mapping[str, Any]) -> bool:
    return any(
        str(row.get(key) or "").strip() == "business_measure_rule"
        for key in ("source_kind", "vision_pass")
    ) or str(row.get("evidence_role") or "").strip() == "measure_item"


def _human_feature_templates(*, item_name: str, spec: str, evidence_text: str) -> list[str]:
    text = _normalize_text(" ".join([item_name, spec, evidence_text]))
    templates: list[str] = []
    if "门槛石" in text:
        templates.append("界面剂一道，50厚CL7.5轻集料混凝土垫层；20厚1:3细石水泥砂浆找平层；20厚水泥砂浆结合层；20厚大理石门槛石")
    elif "瓷砖地面" in text:
        templates.append("界面剂一道，50厚CL7.5轻集料混凝土垫层；20厚1:3细石水泥砂浆找平层；20厚水泥砂浆结合层；地砖铺贴")
    elif "墙面瓷砖" in text:
        templates.append("墙面基层处理；水泥砂浆找平层；水泥砂浆结合层；墙砖湿贴")
    elif "墙面石材" in text:
        templates.append("墙面基层处理；水泥砂浆找平层；水泥砂浆结合层；石材湿贴")

    if "铝扣板吊顶" in text:
        templates.append("专用镀锌轻钢龙骨，主龙骨间距800以内；600*600铝扣板安装")
    elif "造型吊顶" in text:
        templates.append("U型50系列轻钢天棚龙骨跌级；造型处为15厚阻燃板基层；双层9.5mm防水石膏板，自攻螺丝固定")
    elif "防水石膏板" in text or "轻钢龙骨石膏板" in text:
        templates.append("U型50系列轻钢天棚龙骨；双层9.5mm防水石膏板，自攻螺丝固定，满足规范要求")
    elif "灯槽" in text:
        templates.append("尺寸约100*50mm；15厚阻燃板基层，防腐、防蛀处理，面贴单层9.5mm石膏板")
    elif "窗帘盒" in text:
        templates.append("尺寸约200mm宽；15厚阻燃板基层，防腐、防蛀处理，面贴单层9.5mm石膏板")

    if "防水" in text and "石膏板" not in text and "吊顶" not in text:
        templates.append("1.5厚聚氨酯涂膜防水三遍")
    if "零星砌筑" in text or "砖砌地台" in text:
        templates.append("过厅砖砌地台，抬高240mm")
    if "防水保护层" in text:
        templates.append("10厚水泥砂浆防水保护层")
    if "挡水条" in text:
        templates.append("60宽人造石挡水条")
    if "砖砌隔墙" in text:
        templates.append("100mm宽；新建蒸压加气砼砌块墙")
    if "陶粒回填" in text:
        templates.append("隔墙陶粒回填")
    if "墙布墙面" in text:
        templates.append("基层处理，批腻子；墙布专用粘结胶；素色墙布")
    elif "硬包墙面" in text:
        templates.append("木方+15厚阻燃板基层；布艺硬包")
    elif "墙面抹灰" in text:
        templates.append("20厚1:2.5水泥砂浆抹灰找平；清理、修补、湿润基层表面")
    if "防潮无机涂料" in text:
        color = "黑色" if "黑色" in text else ""
        templates.append(f"基层处理、贴绷带及点防锈漆；防水腻子，面油{color}防潮无机涂料三遍")
    elif "白色无机涂料" in text:
        templates.append("基层处理、贴绷带及点防锈漆；防水腻子，面油白色无机涂料三遍")

    if "踢脚线" in text:
        templates.append("不锈钢踢脚线，基层处理，成品安装")
    if "成品保护" in text:
        templates.append("墙地面成品保护")

    if "拆除" in text and any(term in text for term in ("玻璃门", "实木门", "铝合金门", "门套", "门扇", "五金")):
        templates.append("拆除门套、门扇及五金并清运，保护周边完成面")
    if "售卖窗口" in text:
        templates.append("售卖窗口基层、面层及收口按图纸节点施工")
    if "窗台石" in text:
        templates.append("人造石窗台石，基层处理、粘结层、磨边及收口按图纸完成")
    if "隔断底座" in text:
        templates.append("钢通结构+15厚阻燃夹板基层；古堡灰大理石ST-1底座，1470*240*200mm")
    if "钢化玻璃造型柱" in text:
        templates.append("10mm钢化玻璃造型柱，200*200*2366mm，4套")
    if "门套" in text and "拆除" not in text:
        templates.append("不锈钢门套，基层处理、固定、收口及成品保护")
    if any(term in text for term in ("成品玻璃门", "成品门", "实木门", "铝合金门")) and "拆除" not in text:
        templates.append("成品门供货及安装，含门套、门扇、五金、固定及调试")
    if "淋浴隔断" in text:
        templates.append("淋浴隔断供货及安装，含玻璃、五金、固定件及收口")

    if "灯带" in text:
        templates.append("LED灯带，7.2W/m，DC24V，3000K色温，电源驱动器按图纸配置")
    elif "灯具安装" in text or "筒灯" in text or "射灯" in text or "格栅灯" in text:
        templates.append("LED灯具，功率按图纸型号，3000K色温")
    if "开关安装" in text:
        templates.append("86型10A开关，单联/双联/三联按图纸型号")
    if "插座安装" in text:
        templates.append("86型插座，普通插座或16A专用插座按图纸型号")
    if "电气配线" in text:
        templates.append("管内穿线，WDZC-BYJ导线，规格按图纸型号")
    if "电缆敷设" in text:
        templates.append("WDZC-YJY电缆敷设，规格按图纸型号")
    if "配电箱" in text:
        templates.append("成套配电箱，暗装，包括箱体及元器件")
    if "排气扇安装" in text:
        templates.append("名称、型号：排气扇")
        templates.append("排气扇安装，型号按图纸，含本体、固定、接线及调试")
    if "电热水器供货及安装" in text:
        templates.append("名称、型号：电热水器")
        templates.append("电热水器供货及安装，型号按图纸，含本体、固定、接线及调试")
    if "阀门供货及安装" in text:
        templates.append("阀门供货及安装，规格、材质、连接方式按图纸及材料表")
    if "水表供货及安装" in text:
        templates.append("水表供货及安装，规格、连接方式及配件按图纸及材料表")
    if "地漏供货及安装" in text:
        templates.append("地漏供货及安装，含本体、接口连接及收口")
    if "厕纸架供货及安装" in text:
        templates.append("厕纸架供货及安装，含本体、固定件及安装")
    if "梳妆镜供货及安装" in text:
        templates.append("梳妆镜供货及安装，含镜面、固定件及收口")
    if "马桶供货及安装" in text:
        templates.append("马桶供货及安装，含洁具本体、角阀、软管、密封及调试")
    if "台盆供货及安装" in text:
        templates.append("台盆供货及安装，含本体、支架或台面固定、下水及五金")
    if "花洒供货及安装" in text:
        templates.append("淋浴花洒供货及安装，含本体、混水阀、连接件及调试")
    if "龙头供货及安装" in text:
        if "冷热水龙头供货及安装" in text:
            templates.append("冷热水龙头供货及安装，含本体、软管、角阀及调试")
        else:
            templates.append("龙头供货及安装，含本体、软管、角阀及调试")
    return _unique(templates)


def _direct_standard_candidate(
    item_name: str,
    feature: str,
    row: Mapping[str, Any],
    *,
    index: LoadedStandardLibraryIndex,
) -> dict[str, Any]:
    text = _normalize_text(" ".join([item_name, feature, _clean(row.get("evidence_text")), _clean(row.get("raw_names"))]))
    target: tuple[str, str, str] | None = None
    if "配电箱" in text:
        target = ("GBT50856-2024", "030402011", "PDF识图配电箱直连成套配电箱标准项")
    elif "灯带" in text:
        target = ("GBT50856-2024", "030413002", "PDF识图灯带直连装饰灯标准项")
    elif any(term in text for term in ("灯具安装", "普通灯具", "筒灯", "格栅灯", "射灯")) and "灯带" not in text:
        target = ("GBT50856-2024", "030413001", "PDF识图灯具安装直连普通灯具标准项")
    elif "开关安装" in text or "照明开关" in text:
        target = ("GBT50856-2024", "030413013", "PDF识图开关安装直连照明开关/按钮标准项")
    elif "插座安装" in text or "插座" in text:
        target = ("GBT50856-2024", "030413014", "PDF识图插座安装直连插座标准项")
    elif "电气配管" in text or ("配管" in text and "给排水" not in text):
        target = ("GBT50856-2024", "030412001", "PDF识图电气配管直连配管标准项")
    elif "电气配线" in text or ("配线" in text and "配线架" not in text):
        target = ("GBT50856-2024", "030412004", "PDF识图电气配线直连配线标准项")
    elif "电缆敷设" in text and any(term in text for term in ("yjv", "yjy", "wdzc", "wdzn", "nh")):
        target = ("GBT50856-2024", "030409001", "PDF识图电缆敷设直连电力电缆标准项")
    elif "桥架" in text:
        target = ("GBT50856-2024", "030412003", "PDF识图桥架直连桥架标准项")
    elif "防潮无机涂料" in text or (
        any(term in text for term in ("无机涂料", "乳胶漆", "喷刷涂料"))
        and any(term in text for term in ("天棚", "天花", "顶面", "吊顶", "石膏板"))
    ):
        target = ("GBT50854-2024", "011404002", "PDF识图天棚/顶面涂料直连天棚喷刷涂料标准项")
    elif any(term in text for term in ("墙面无机涂料", "墙面乳胶漆", "墙面喷刷涂料")):
        target = ("GBT50854-2024", "011404001", "PDF识图墙面涂料直连墙面喷刷涂料标准项")
    elif any(term in text for term in ("平级吊顶", "铝扣板吊顶", "石膏板平级吊顶", "吊顶")) and not any(term in text for term in ("造型吊顶", "跌级吊顶")):
        target = ("GBT50854-2024", "011302001", "PDF识图平面/铝扣板/石膏板吊顶直连平面吊顶天棚标准项")
    elif any(term in text for term in ("造型吊顶", "跌级吊顶")):
        target = ("GBT50854-2024", "011302002", "PDF识图造型/跌级吊顶直连跌级吊顶天棚标准项")
    elif "窗帘盒" in text:
        target = ("GBT50854-2024", "010810002", "PDF识图窗帘盒直连窗帘盒标准项")
    elif "门槛石" in text:
        target = ("GBT50854-2024", "011102001", "PDF识图门槛石直连石材楼地面标准项")
    elif "石材踢脚" in text:
        target = ("GBT50854-2024", "011105002", "PDF识图石材踢脚线直连石材踢脚线标准项")
    elif "踢脚线" in text or "踢脚" in text:
        target = ("GBT50854-2024", "011105001", "PDF识图踢脚线直连踢脚线标准项")
    elif "墙布墙面" in text or "墙纸" in text or "墙布" in text:
        target = ("GBT50854-2024", "011405001", "PDF识图墙布/墙纸直连墙纸裱糊标准项")
    elif "墙面石材" in text or ("石材" in text and "墙面" in text):
        target = ("GBT50854-2024", "011203001", "PDF识图墙面石材直连石材墙柱面标准项")
    elif "硬包墙面" in text or "墙面装饰板" in text or "生态板墙面" in text:
        target = ("GBT50854-2024", "011205001", "PDF识图墙面装饰板/硬包直连墙柱面装饰板标准项")
    elif "给水管" in text and any(term in text for term in ("sus304", "不锈钢")):
        target = ("GBT50856-2024", "031001006", "PDF识图不锈钢给水管直连给排水不锈钢管标准项")
    elif any(term in text for term in ("排水管", "管道安装")) and any(term in text for term in ("柔性铸铁", "铸铁")):
        target = ("GBT50856-2024", "031001001", "PDF识图铸铁排水管直连给排水铸铁管标准项")
    elif any(term in text for term in ("给水管", "排水管", "管道安装")) and any(
        term in text for term in ("ppr", "pvc", "pe", "塑料", "de")
    ):
        target = ("GBT50856-2024", "031001008", "PDF识图塑料给排水管直连给排水塑料管标准项")
    elif "地漏" in text:
        target = ("GBT50856-2024", "031003014", "PDF识图地漏直连给、排水附件标准项")
    elif "水表" in text:
        target = ("GBT50856-2024", "031002011", "PDF识图水表直连水表标准项")
    elif "洗脸盆" in text or "台盆" in text or "面盆" in text:
        target = ("GBT50856-2024", "031003003", "PDF识图洗脸盆/台盆直连洗脸盆标准项")
    elif any(term in text for term in ("厕纸架", "纸巾架", "卫生纸架", "梳妆镜", "镜面", "镜子", "花洒", "淋浴喷头", "龙头", "给、排水附件")):
        target = ("GBT50856-2024", "031003014", "PDF识图洁具五金直连给、排水附件标准项")
    elif "大便器" in text or "坐便" in text or "马桶" in text:
        target = ("GBT50856-2024", "031003006", "PDF识图马桶/坐便器直连大便器标准项")
    elif "小便器" in text:
        target = ("GBT50856-2024", "031003007", "PDF识图小便器直连小便器标准项")
    elif "浴缸" in text:
        target = ("GBT50856-2024", "031003001", "PDF识图浴缸直连浴缸标准项")
    elif "洗涤盆" in text:
        target = ("GBT50856-2024", "031003004", "PDF识图洗涤盆直连洗涤盆标准项")
    elif "塑料阀门" in text:
        target = ("GBT50856-2024", "031002003", "PDF识图塑料阀门直连塑料阀门标准项")
    elif "金属阀门" in text or "不锈钢阀门" in text or "铜阀" in text:
        target = ("GBT50856-2024", "031002001", "PDF识图金属阀门直连金属阀门标准项")

    if not target:
        return {}
    standard_code, item_code, reason = target
    item = find_standard_item(index, standard_code, item_code)
    if not item:
        return {}
    return {
        **_standard_item_payload(index, standard_code=standard_code, item=item.as_dict()),
        "score": 80.0,
        "matched_fields": ["pdf_v2_direct_standard_hint"],
        "match_reason": reason,
    }


def _standard_item_payload(
    index: LoadedStandardLibraryIndex,
    *,
    standard_code: str,
    item: Mapping[str, Any],
) -> dict[str, Any]:
    library = index.quantity_libraries.get(standard_code)
    return {
        "standard_code": standard_code,
        "standard_name": (library.standard.get("name", "") if library else ""),
        "item_code": item.get("item_code", ""),
        "official_item_code": item.get("official_item_code") or item.get("item_code", ""),
        "item_name": item.get("item_name", ""),
        "chapter_name": item.get("chapter_name", ""),
        "unit_options": list(item.get("unit_options") or []),
        "feature_fields": list(item.get("feature_fields") or []),
        "quantity_rule": dict(item.get("quantity_rule") or {}),
        "drawing_evidence_requirements": list(item.get("drawing_evidence_requirements") or []),
        "source_note": item.get("source_note", ""),
    }


def _standard_candidate_compatible(
    item_name: str,
    feature: str,
    row: Mapping[str, Any],
    candidate: Mapping[str, Any],
) -> bool:
    text = _normalize_text(" ".join([item_name, feature, _clean(row.get("evidence_text")), _clean(row.get("raw_names"))]))
    candidate_text = _normalize_text(
        " ".join([_clean(candidate.get("item_name")), _clean(candidate.get("chapter_name")), _clean(candidate.get("standard_name"))])
    )
    if any(term in text for term in ("拆除", "铲除", "清运", "基层处理")):
        return any(term in candidate_text for term in ("拆除", "铲除", "修缮", "建筑垃圾", "清运", "基层"))
    if "灯槽" in text:
        return any(term in candidate_text for term in ("吊顶", "天棚")) and "干燥机" not in candidate_text
    if any(term in text for term in ("给水管", "排水管", "管道", "dn", "de")):
        source_family = _pipe_material_family(text)
        candidate_family = _pipe_material_family(candidate_text)
        if source_family == "unknown":
            return False
        if source_family != candidate_family:
            return False
        return any(term in candidate_text for term in ("管", "给排水", "排水", "给水"))
    if any(term in text for term in ("水表", "地漏", "洗脸盆", "大便器", "小便器", "浴缸", "洗涤盆")):
        return any(term in candidate_text for term in ("水表", "给排水附件", "洗脸盆", "大便器", "小便器", "浴缸", "洗涤盆"))
    if "阀门" in text:
        return "阀门" in candidate_text
    if any(term in text for term in ("电气配管", "jdg", "sc", "mt")):
        return "配管" in candidate_text
    if any(term in text for term in ("电气配线", "byj", "bv")):
        return "配线" in candidate_text and "配线架" not in candidate_text
    if any(term in text for term in ("电缆敷设", "yjv", "yjy")):
        return "电缆" in candidate_text and "桥架" not in candidate_text
    if any(term in text for term in ("灯具", "筒灯", "格栅灯", "射灯")) and "灯带" not in text:
        return any(term in candidate_text for term in ("灯具", "照明", "装饰灯"))
    if any(term in text for term in ("排气扇", "换气扇", "通风扇")):
        return any(term in candidate_text for term in ("风机", "排气", "通风", "换气"))
    if any(term in text for term in ("电热水器", "热水器")):
        return any(term in candidate_text for term in ("热水器", "电热", "设备"))
    if any(term in text for term in ("吊顶", "天棚", "铝扣板", "石膏板")):
        return any(term in candidate_text for term in ("吊顶", "天棚"))
    if any(term in text for term in ("墙布", "墙纸")):
        return any(term in candidate_text for term in ("裱糊", "墙纸", "涂料", "油漆"))
    if "石材" in text and "墙面" in text:
        return "石材" in candidate_text and any(term in candidate_text for term in ("墙", "柱"))
    return True


def _pipe_material_family(text: str) -> str:
    if any(term in text for term in ("柔性铸铁", "铸铁")):
        return "cast_iron"
    if any(term in text for term in ("ppr", "pvc", "pe", "塑料", "de")):
        return "plastic"
    if any(term in text for term in ("sus304", "不锈钢")):
        return "stainless"
    if any(term in text for term in ("钢管", "焊接钢管", "镀锌钢管", "碳钢", "金属管")):
        return "steel"
    return "unknown"


def _build_summary(
    pdf_summary: Mapping[str, Any],
    evidence_rows: Sequence[Mapping[str, Any]],
    normalized_rows: Sequence[Mapping[str, Any]],
    human_rows: Sequence[Mapping[str, Any]],
    acceptance_report: Mapping[str, Any] | None,
    gap_rows: Sequence[Mapping[str, Any]] | None = None,
) -> dict[str, Any]:
    standard_mapped = sum(1 for row in human_rows if row.get("standard_item_code"))
    acceptance_summary = (acceptance_report or {}).get("summary") or {}
    gap_status_counts = dict(Counter(row.get("gap_type", "") for row in (gap_rows or [])).most_common())
    return {
        "pdf_file_count": pdf_summary.get("pdf_file_count", 0),
        "pdf_page_count": pdf_summary.get("pdf_page_count", 0),
        "pdf_render_status": pdf_summary.get("pdf_render_status", ""),
        "evidence_count": len(evidence_rows),
        "normalized_evidence_count": len(normalized_rows),
        "human_style_row_count": len(human_rows),
        "standard_mapped_count": standard_mapped,
        "standard_unmapped_count": len(human_rows) - standard_mapped,
        "evidence_type_counts": dict(Counter(row.get("evidence_type", "") for row in evidence_rows).most_common()),
        "division_counts": dict(Counter(row.get("division", "") for row in human_rows).most_common()),
        "three_field_answer_count": acceptance_summary.get("answer_count", 0),
        "three_field_candidate_count": acceptance_summary.get("candidate_count", len(human_rows)),
        "three_field_matched_count": acceptance_summary.get("matched_three_fields_count", 0),
        "three_field_pass_rate": acceptance_summary.get("three_field_pass_rate", 0),
        "three_field_gap_count": len(gap_rows or []),
        "three_field_gap_status_counts": gap_status_counts,
        "quantity_status": "deferred_until_three_fields_accepted",
    }


def _build_stage_results(
    summary: Mapping[str, Any],
    pdf_summary: Mapping[str, Any],
    acceptance_report: Mapping[str, Any] | None,
) -> list[dict[str, Any]]:
    return [
        {
            "stage": "S1_pdf_render_and_tile",
            "status": "completed" if pdf_summary else "provided_report",
            "summary": f"PDF {summary.get('pdf_file_count', 0)} 份，页面 {summary.get('pdf_page_count', 0)}，渲染状态 {summary.get('pdf_render_status') or '-'}",
        },
        {
            "stage": "S2_visual_evidence_extraction",
            "status": "completed" if summary.get("evidence_count", 0) else "needs_glm_or_ocr",
            "summary": f"抽取图纸证据 {summary.get('evidence_count', 0)} 条",
        },
        {
            "stage": "S3_evidence_normalization",
            "status": "completed" if summary.get("normalized_evidence_count", 0) else "pending",
            "summary": f"归一化证据组 {summary.get('normalized_evidence_count', 0)} 组",
        },
        {
            "stage": "S4_human_style_itemization",
            "status": "completed" if summary.get("human_style_row_count", 0) else "pending",
            "summary": f"生成人工清单风格候选 {summary.get('human_style_row_count', 0)} 行",
        },
        {
            "stage": "S5_standard_mapping",
            "status": "completed" if summary.get("standard_mapped_count", 0) else "pending",
            "summary": f"国标映射 {summary.get('standard_mapped_count', 0)} 行，未映射 {summary.get('standard_unmapped_count', 0)} 行",
        },
        {
            "stage": "S6_excel_and_three_field_acceptance",
            "status": "completed" if acceptance_report else "excel_only_without_answer",
            "summary": f"三字段候选 {summary.get('three_field_candidate_count', 0)} 行，全匹配 {summary.get('three_field_matched_count', 0)} 行",
        },
        {
            "stage": "S7_quantity_takeoff",
            "status": "deferred",
            "summary": "工程量列保留空值，等待三字段验收稳定后再做算量。",
        },
    ]


def _write_workbook(path: Path, report: Mapping[str, Any]) -> None:
    workbook = Workbook()
    overview = workbook.active
    overview.title = "阶段概览"
    summary = report.get("summary") or {}
    _append_rows(
        overview,
        [
            ["指标", "值"],
            ["图纸证据", summary.get("evidence_count", 0)],
            ["归一化证据组", summary.get("normalized_evidence_count", 0)],
            ["人工清单风格候选", summary.get("human_style_row_count", 0)],
            ["国标已映射", summary.get("standard_mapped_count", 0)],
            ["三字段候选", summary.get("three_field_candidate_count", 0)],
            ["三字段全匹配", summary.get("three_field_matched_count", 0)],
            ["三字段缺口复核", summary.get("three_field_gap_count", 0)],
            ["工程量状态", summary.get("quantity_status", "")],
        ],
    )
    _style_sheet(overview)
    _sheet_from_rows(workbook, "图纸证据", EVIDENCE_HEADERS, report.get("evidence_rows") or [])
    _sheet_from_rows(workbook, "归一化证据", NORMALIZED_HEADERS, report.get("normalized_evidence_rows") or [])
    _sheet_from_rows(workbook, "人工清单风格候选", HUMAN_ROW_HEADERS, report.get("human_style_rows") or [])
    comparison = ((report.get("three_field_acceptance_report") or {}).get("comparison_rows") or [])
    if comparison:
        headers = list(comparison[0].keys())
        _sheet_from_rows(workbook, "三字段验收", headers, comparison)
    _sheet_from_rows(workbook, "三字段缺口复核", GAP_ROW_HEADERS, report.get("three_field_gap_rows") or [])
    workbook.save(path)


def _sheet_from_rows(workbook: Workbook, title: str, headers: Sequence[str], rows: Sequence[Mapping[str, Any]]) -> None:
    sheet = workbook.create_sheet(title[:31])
    _append_rows(sheet, [headers, *[[_cell(row.get(header)) for header in headers] for row in rows]])
    _style_sheet(sheet)


def _append_rows(sheet: Any, rows: Sequence[Sequence[Any]]) -> None:
    for row in rows:
        sheet.append(list(row))


def _style_sheet(sheet: Any) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    if sheet.max_row >= 1:
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for col_idx, column_cells in enumerate(sheet.columns, start=1):
        max_len = max(len(str(cell.value or "")) for cell in column_cells[:200])
        sheet.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 10), 48)


def _write_csv(path: Path, rows: Iterable[Mapping[str, Any]], fieldnames: Sequence[str]) -> None:
    with path.open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.DictWriter(handle, fieldnames=list(fieldnames))
        writer.writeheader()
        for row in rows:
            writer.writerow({field: _cell(row.get(field)) for field in fieldnames})


def _style_prompt_summary(text: str) -> dict[str, Any]:
    cleaned = _clean(text)
    return {
        "provided": bool(cleaned),
        "length": len(cleaned),
        "uses_human_listing_rules": any(term in cleaned for term in ("人工清单", "材料编号", "拆分", "项目名称")),
    }


def _division_for_type(evidence_type: str) -> str:
    return {
        "demolition": "拆除工程",
        "floor": "楼地面工程",
        "ceiling": "天棚工程",
        "wall": "墙柱面工程",
        "door_window": "门窗工程",
        "electrical": "电气工程",
        "plumbing": "给排水工程",
    }.get(evidence_type, "其他工程")


def _material_codes(row: Mapping[str, Any]) -> list[str]:
    values: list[str] = []
    raw_codes = row.get("材料编号") or row.get("material_codes") or row.get("material_code") or row.get("material_code_list") or []
    if isinstance(raw_codes, str):
        values.extend(re.split(r"[,，、;\s]+", raw_codes))
    elif isinstance(raw_codes, Sequence):
        values.extend(str(item) for item in raw_codes)
    searchable_text = " ".join(
        _clean(row.get(key))
        for key in (
            "item_hint",
            "清单项目提示",
            "图纸项目名称",
            "item_name",
            "项目名称",
            "project_name",
            "text",
            "normalized_text",
            "证据文本",
            "evidence_text",
            "识别依据",
            "reason",
            "规格/做法",
            "spec_or_method",
            "项目特征",
            "feature",
            "spec",
        )
    )
    values.extend(re.findall(r"\b[A-Z]{1,5}[-]?\d{1,4}[A-Z]?\b", searchable_text))
    return _unique(value.strip().upper() for value in values if value and len(value.strip()) >= 2)


def _normalize_unit(value: Any) -> str:
    cleaned = re.sub(r"[，,。；;：:]+$", "", _clean(value))
    text = cleaned.lower()
    mapping = {
        "m2": "㎡",
        "m²": "㎡",
        "㎡": "㎡",
        "平方米": "㎡",
        "m3": "m³",
        "m³": "m³",
        "立方米": "m³",
        "米": "m",
    }
    return mapping.get(text, cleaned)


def _normalize_discipline(value: Any) -> str:
    text = _clean(value).lower()
    mapping = {
        "装饰": "decoration",
        "装修": "decoration",
        "建筑装饰": "decoration",
        "电气": "electrical",
        "强电": "electrical",
        "弱电": "electrical",
        "管道": "plumbing",
        "管道工程": "plumbing",
        "给排水": "plumbing",
        "给排水工程": "plumbing",
        "给水排水": "plumbing",
        "给水排水工程": "plumbing",
    }
    text = mapping.get(text, text)
    return text if text in {"decoration", "electrical", "plumbing", "unknown"} else ""


def _clean(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).replace("\r", " ").replace("\n", " ").strip()
    return re.sub(r"\s+", " ", text)


def _normalize_text(value: Any) -> str:
    return re.sub(r"[\s,，。；;:：、\-_/\\()（）\[\]【】<>《》|]+", "", _clean(value).lower())


def _unique(values: Iterable[Any]) -> list[str]:
    seen: set[str] = set()
    result: list[str] = []
    for value in values:
        cleaned = _clean(value)
        if not cleaned:
            continue
        key = _normalize_text(cleaned)
        if key in seen:
            continue
        seen.add(key)
        result.append(cleaned)
    return result


def _dominant(values: Iterable[Any]) -> str:
    counter = Counter(_clean(value) for value in values if _clean(value))
    return counter.most_common(1)[0][0] if counter else ""


def _first(row: Mapping[str, Any], *keys: str) -> str:
    for key in keys:
        if key in row and row.get(key) not in (None, ""):
            return _clean(row.get(key))
    return ""


def _first_option(values: Any) -> str:
    if isinstance(values, Sequence) and not isinstance(values, str) and values:
        return _clean(values[0])
    return ""


def _first_regex(text: str, pattern: str) -> str:
    match = re.search(pattern, text, flags=re.IGNORECASE)
    return match.group(0).replace(" ", "") if match else ""


def _float(value: Any, default: float = 0.0) -> float:
    try:
        if value in (None, ""):
            return default
        return float(str(value))
    except (TypeError, ValueError):
        return default


def _boolish(value: Any, *, default: bool = False) -> bool:
    if value in (None, ""):
        return default
    text = _clean(value).lower()
    if text in {"true", "1", "yes", "y", "是", "需", "需要"}:
        return True
    if text in {"false", "0", "no", "n", "否", "不需要"}:
        return False
    return bool(value)


def _remove_words(text: str, words: Sequence[str]) -> str:
    result = _clean(text)
    for word in words:
        result = result.replace(word, "")
    return result.strip()


def _cell(value: Any) -> Any:
    if isinstance(value, (list, tuple)):
        return "；".join(_clean(item) for item in value if _clean(item))
    if isinstance(value, dict):
        return json.dumps(value, ensure_ascii=False, default=str)
    return value if value is not None else ""


def _json_safe(value: Any) -> Any:
    if isinstance(value, Mapping):
        return {str(key): _json_safe(item) for key, item in value.items()}
    if isinstance(value, list):
        return [_json_safe(item) for item in value]
    if isinstance(value, tuple):
        return [_json_safe(item) for item in value]
    return value


def _md(value: Any) -> str:
    return _clean(value).replace("|", "\\|")
