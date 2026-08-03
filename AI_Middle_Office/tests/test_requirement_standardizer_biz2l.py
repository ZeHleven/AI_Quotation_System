from __future__ import annotations

import json
from io import BytesIO

import pytest
from openpyxl import Workbook

from app.services.requirement_standardizer import (
    RequirementStandardizationError,
    build_standardization_csv,
    build_standardization_markdown,
    standardize_requirement_excel_bytes,
    write_standardization_outputs,
)


def _workbook_bytes(rows, *, sheet_title="需求单", merges=None, extra_sheets=None) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = sheet_title
    for row in rows:
        sheet.append(row)
    for merge_range in merges or []:
        sheet.merge_cells(merge_range)
    for title, sheet_rows in extra_sheets or []:
        extra = workbook.create_sheet(title)
        for row in sheet_rows:
            extra.append(row)
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _data_rows(result):
    return [row for row in result["rows"] if row["row_type"] == "data_row"]


def _reference_rows(result):
    return [row for row in result["rows"] if row["row_type"] == "reference_row"]


def test_standardizer_reads_standard_headers_and_keeps_price_read_only():
    content = _workbook_bytes(
        [
            ["客户需求单"],
            ["区域", "项目名称", "规格", "数量", "单位", "综合单价", "备注"],
            ["客厅", "拆除复合木地板", "", 20, "㎡", 99, "不含清运"],
            ["客厅", "窗帘盒/灯槽拆除", "", 18, "m", None, ""],
            ["合计", None, None, None, None, 999, None],
        ]
    )

    result = standardize_requirement_excel_bytes(content, filename="quote.xlsx")
    rows = _data_rows(result)

    assert result["version"] == "biz2l-standard-v0"
    assert result["summary"]["standard_row_count"] == 2
    assert result["field_mapping"]["B"] == "item_name"
    assert result["field_mapping"]["F"] == "price_ignored"
    assert rows[0]["item_name"] == "拆除复合木地板"
    assert rows[0]["quantity"] == 20
    assert rows[0]["unit"] == "㎡"
    assert rows[0]["location"] == "客厅"
    assert "PRICE_COLUMN_PRESENT" in rows[0]["warnings"]
    assert rows[0]["requires_confirmation"] is True
    assert any(issue["code"] == "PRICE_COLUMN_PRESENT" for issue in result["issues"])
    assert any(row["row_type"] == "summary_row" for row in result["rows"])


def test_standardizer_keeps_instruction_and_calculation_rule_sheets_as_reference_rows():
    content = _workbook_bytes(
        [["项目名称", "工程量", "单位"], ["墙面抹灰", 12, "㎡"]],
        sheet_title="装饰部分清单",
        extra_sheets=[
            (
                "编制说明",
                [
                    ["编制说明"],
                    [""],
                    ["1 投标人应充分考虑材料价格波动"],
                    ["2 成品保护费用综合考虑"],
                ],
            ),
            (
                "计算规则",
                [
                    ["精装修计算规则"],
                    ["一、楼地面"],
                    ["1 按设计图示尺寸以面积计算"],
                    ["2 扣除门窗洞口超过0.3m2的面积"],
                ],
            ),
        ],
    )

    result = standardize_requirement_excel_bytes(content, filename="rules.xlsx")
    refs = _reference_rows(result)

    assert result["summary"]["standard_row_count"] == 1
    assert [row["item_name"] for row in _data_rows(result)] == ["墙面抹灰"]
    assert {item["sheet_name"]: item["sheet_role"] for item in result["sheet_mappings"]} == {
        "装饰部分清单": "bill",
        "编制说明": "metadata",
        "计算规则": "calculation_rule",
    }
    assert [(row["source_sheet"], row["raw_row_index"], row["item_name"]) for row in refs] == [
        ("编制说明", 1, "编制说明"),
        ("编制说明", 3, "1 投标人应充分考虑材料价格波动"),
        ("编制说明", 4, "2 成品保护费用综合考虑"),
        ("计算规则", 1, "精装修计算规则"),
        ("计算规则", 2, "一、楼地面"),
        ("计算规则", 3, "1 按设计图示尺寸以面积计算"),
        ("计算规则", 4, "2 扣除门窗洞口超过0.3m2的面积"),
    ]
    assert all(row["quantity"] is None and row["requires_confirmation"] is False for row in refs)


def test_standardizer_keeps_cover_summary_and_formula_errors_out_of_data_rows():
    content = _workbook_bytes(
        [
            ["标前成本分析表"],
            ["H", 0, 0, "I", "#REF!", "#REF!"],
        ],
        sheet_title="封面",
        extra_sheets=[
            ("报价说明", [["说明：本表仅供投标前分析使用"]]),
            ("报价汇总表", [["合计", None, 123]]),
            (
                "分部分项清单",
                [
                    ["项目名称", "项目特征", "数量", "单位", "", "", "", "", "I", "J", "K", "L"],
                    ["拆除木地板", "含清运", 12, "㎡", "", "", "", "", "-", "-", "-", "-"],
                ],
            ),
        ],
    )

    result = standardize_requirement_excel_bytes(content, filename="bid-analysis.xlsx")
    rows = _data_rows(result)

    assert result["summary"]["standard_row_count"] == 1
    assert rows[0]["source_sheet"] == "分部分项清单"
    assert rows[0]["item_name"] == "拆除木地板"
    assert rows[0]["quantity"] == 12
    assert all(row["source_sheet"] not in {"封面", "报价说明", "报价汇总表"} for row in rows)
    assert any(row["source_sheet"] == "封面" and row["row_type"] == "empty_row" for row in result["rows"])
    list_sheet_mapping = next(item for item in result["sheet_mappings"] if item["sheet_name"] == "分部分项清单")
    assert [column["column"] for column in list_sheet_mapping["columns"]] == ["A", "B", "C", "D"]


def test_standardizer_mapping_columns_do_not_promote_samples_fee_or_group_headers():
    content = _workbook_bytes(
        [
            ["", "", "", "", "工程量", "工程量", "工程量", "工程量"],
            ["项目名称", "项目特征", "单位", "数量", "人工费", "管理费", "", ""],
            ["拆除木地板", "含清运", "㎡", 12, 100, 20, "一--五楼", "人工费"],
        ],
        sheet_title="装饰清单",
    )

    result = standardize_requirement_excel_bytes(content, filename="decor.xlsx")
    mapping = result["sheet_mappings"][0]

    assert result["summary"]["standard_row_count"] == 1
    assert mapping["field_mapping"]["D"] == "quantity"
    assert [column["column"] for column in mapping["columns"]] == ["A", "B", "C", "D"]
    assert all(column["label"] not in {"人工费", "管理费", "一--五楼"} for column in mapping["columns"])


def test_standardizer_exposes_quantity_source_and_multiple_candidates():
    content = _workbook_bytes(
        [
            ["项目名称", "项目特征", "单位", "工程量", "工程量", "工程量"],
            ["拆除木地板", "含清运", "㎡", 12, 3, 5],
        ],
        sheet_title="装饰清单",
    )

    result = standardize_requirement_excel_bytes(content, filename="decor-quantities.xlsx")
    row = _data_rows(result)[0]
    mapping = result["sheet_mappings"][0]

    assert row["quantity"] == 12
    assert row["quantity_source"]["column"] == "D"
    assert [candidate["column"] for candidate in row["quantity_candidates"]] == ["D", "E", "F"]
    assert [candidate["quantity"] for candidate in row["quantity_candidates"]] == [12, 3, 5]
    assert "MULTIPLE_QUANTITY_CANDIDATES" in row["warnings"]
    assert row["requires_confirmation"] is True
    assert [column["column"] for column in mapping["columns"]] == ["A", "B", "C", "D"]


def test_standardizer_prefers_project_feature_as_spec_not_item_name():
    content = _workbook_bytes(
        [
            ["序号", "分部分项工程", "项目特征", "单位", "主材", "工程量", "不含税单价"],
            ["1", "拆除铝合金玻璃门", "人工及机械拆除，垃圾外运", "㎡", "", 6.6, 0],
        ]
    )

    result = standardize_requirement_excel_bytes(content, filename="engineering-list.xlsx")
    row = _data_rows(result)[0]

    assert result["field_mapping"]["B"] == "item_name"
    assert result["field_mapping"]["C"] == "spec"
    assert row["item_name"] == "拆除铝合金玻璃门"
    assert row["spec"] == "人工及机械拆除，垃圾外运"
    assert row["quantity"] == 6.6
    assert row["unit"] == "㎡"
    assert "PRICE_COLUMN_PRESENT" in row["warnings"]


def test_standardizer_does_not_flag_spec_price_or_number_as_quantity_ambiguity():
    content = _workbook_bytes(
        [
            ["序号", "项目名称", "项目特征", "数量", "单位", "综合单价"],
            [9, "成品木门安装", "900*2100mm 成品复合木门，含门套、五金", 2, "樘", 880],
        ]
    )

    result = standardize_requirement_excel_bytes(content, filename="mapped-dimensions.xlsx")
    row = _data_rows(result)[0]

    assert row["quantity"] == 2
    assert row["unit"] == "樘"
    assert row["spec"] == "900*2100mm 成品复合木门，含门套、五金"
    assert "PRICE_COLUMN_PRESENT" in row["warnings"]
    assert "MULTIPLE_NUMBERS" not in row["warnings"]


def test_standardizer_treats_estimated_quantity_column_as_quantity():
    content = _workbook_bytes(
        [
            ["位置", "工作内容", "现场描述", "预估数量", "单位", "图片编号"],
            ["茶水间", "墙砖空鼓修补", "局部敲除重贴，甲方描述约 12~15 片", 15, "片", "IMG_002"],
        ]
    )

    result = standardize_requirement_excel_bytes(content, filename="estimated-quantity.xlsx")
    row = _data_rows(result)[0]

    assert result["field_mapping"]["D"] == "quantity"
    assert row["quantity"] == 15
    assert row["unit"] == "片"
    assert row["quantity_source"]["column"] == "D"
    assert "MULTIPLE_NUMBERS" not in row["warnings"]


def test_standardizer_does_not_guess_quantity_from_spec_when_mapped_quantity_is_blank():
    content = _workbook_bytes(
        [
            ["位置", "工作内容", "现场描述", "预估数量", "单位", "图片编号"],
            ["茶水间", "墙砖空鼓修补", "局部敲除重贴，甲方描述约 12~15 片", "", "片", "IMG_002"],
        ]
    )

    result = standardize_requirement_excel_bytes(content, filename="blank-estimated-quantity.xlsx")
    row = _data_rows(result)[0]

    assert result["field_mapping"]["D"] == "quantity"
    assert row["quantity"] is None
    assert row["unit"] == "片"
    assert "MISSING_QUANTITY" in row["warnings"]
    assert "MULTIPLE_NUMBERS" not in row["warnings"]


def test_standardizer_infers_no_header_mixed_name_quantity_and_notes():
    content = _workbook_bytes(
        [
            ["一、拆除工程"],
            ["拆除木地板20㎡"],
            ["墙面乳胶漆 两遍 120㎡"],
            ["备注：现场保护由施工方负责"],
        ]
    )

    result = standardize_requirement_excel_bytes(content, filename="mixed.xlsx")
    rows = _data_rows(result)

    assert result["summary"]["standard_row_count"] == 2
    assert rows[0]["item_name"] == "拆除木地板"
    assert rows[0]["quantity"] == 20
    assert rows[0]["unit"] == "㎡"
    assert rows[0]["requires_confirmation"] is True
    assert "AMBIGUOUS_HEADER" in rows[0]["warnings"]
    assert rows[1]["item_name"] == "墙面乳胶漆"
    assert rows[1]["spec"] == "两遍"
    assert rows[1]["quantity"] == 120
    assert rows[1]["unit"] == "㎡"
    assert any(row["row_type"] == "note_row" for row in result["rows"])
    assert any(row["row_type"] == "section_row" for row in result["rows"])


def test_standardizer_flags_missing_unit_range_quantity_and_multiple_numbers():
    content = _workbook_bytes(
        [
            ["工作内容", "工程量", "计量单位"],
            ["安装柜体", "20-30", "m"],
            ["拆除踢脚线", 30, ""],
            ["墙面乳胶漆 2遍", 120, "㎡"],
        ]
    )

    result = standardize_requirement_excel_bytes(content, filename="warnings.xlsx")
    rows = _data_rows(result)

    assert "RANGE_QUANTITY" in rows[0]["warnings"]
    assert rows[0]["requires_confirmation"] is True
    assert rows[1]["item_name"] == "拆除踢脚线"
    assert "MISSING_UNIT" in rows[1]["warnings"]
    assert rows[1]["requires_confirmation"] is True
    assert rows[2]["item_name"] == "墙面乳胶漆"
    assert rows[2]["spec"] == "2遍"
    assert "MULTIPLE_NUMBERS" not in rows[2]["warnings"]


def test_standardizer_marks_merged_cells_and_multi_sheet_context():
    content = _workbook_bytes(
        [
            ["区域", "项目名称", "数量", "单位"],
            ["客厅", "拆除木地板", 20, "㎡"],
            [None, "拆除木脚线", 30, "m"],
        ],
        merges=["A2:A3"],
        extra_sheets=[
            (
                "安装工程",
                [
                    ["项目名称", "数量", "单位"],
                    ["安装门套", 2, "套"],
                ],
            )
        ],
    )

    result = standardize_requirement_excel_bytes(content, filename="multi.xlsx")
    rows = _data_rows(result)

    assert result["summary"]["sheet_count"] == 2
    assert result["summary"]["standard_row_count"] == 3
    assert rows[1]["location"] == "客厅"
    assert "MERGED_CELL_EXPANDED" in rows[1]["warnings"]
    assert all("MULTI_SHEET_DETECTED" in row["warnings"] for row in result["rows"])


def test_standardizer_accepts_door_and_electrical_point_units():
    content = _workbook_bytes(
        [
            ["项目名称", "数量", "单位"],
            ["成品木门安装", 2, "樘"],
            ["强弱电点位安装", 16, "点"],
            ["门禁点位调试", 3, "点位"],
        ],
    )

    result = standardize_requirement_excel_bytes(content, filename="units.xlsx")
    rows = _data_rows(result)

    assert [row["unit"] for row in rows] == ["樘", "点", "点"]
    assert [row["unit_family"] for row in rows] == ["count", "count", "count"]
    assert all("MISSING_UNIT" not in row["warnings"] for row in rows)


def test_standardizer_outputs_markdown_csv_and_json_files(tmp_path):
    content = _workbook_bytes(
        [
            ["项目名称", "数量", "单位"],
            ["窗帘盒/灯槽拆除", 18, "m"],
        ]
    )
    result = standardize_requirement_excel_bytes(content, filename="quote.xlsx")

    markdown = build_standardization_markdown(result)
    csv_text = build_standardization_csv(result)
    paths = write_standardization_outputs(result, tmp_path, stem="preview")

    assert "BIZ-2l-1 需求单标准化预览" in markdown
    assert "窗帘盒/灯槽拆除" in csv_text
    assert set(paths) == {"json", "csv", "markdown"}
    loaded = json.loads((tmp_path / "preview.json").read_text(encoding="utf-8"))
    assert loaded["summary"]["standard_row_count"] == 1


def test_standardizer_rejects_legacy_xls_suffix():
    with pytest.raises(RequirementStandardizationError, match="只支持 .xlsx/.xlsm"):
        standardize_requirement_excel_bytes(b"not-really-excel", filename="legacy.xls")
