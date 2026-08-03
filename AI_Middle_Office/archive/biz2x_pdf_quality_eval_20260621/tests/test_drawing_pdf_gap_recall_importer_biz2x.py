from __future__ import annotations

import json
import sys
from pathlib import Path

from openpyxl import load_workbook

from app.services.drawing_pdf_gap_recall_importer import (
    build_gap_recall_external_import_report,
    load_external_recall_results,
    write_gap_recall_external_import_outputs,
)
from app.services.drawing_pdf_external_recall_template import (
    build_external_recall_template,
    write_external_recall_template_outputs,
)
from scripts import biz2x_pdf_gap_recall_import


def _recall_plan(tmp_path: Path) -> dict[str, object]:
    image = tmp_path / "whole.png"
    image.write_bytes(b"fake image bytes")
    return {
        "plan_rows": [
            {
                "task_no": 1,
                "gap_no": 1,
                "gap_priority": "P1_missing_core",
                "answer_item_name": "door demolition",
                "recommended_pass": "door_window_demolition",
                "source_file": "drawing.pdf",
                "page": 1,
                "tile_id": "p001_whole",
                "tile_type": "whole_page_preview",
                "image_path": str(image),
            }
        ]
    }


def _external_results() -> dict[str, object]:
    return {
        "call_results": [
            {
                "call_no": 1,
                "model": "offline-glm-export",
                "evidence_items": [
                    {
                        "evidence_role": "construction_note",
                        "discipline": "decoration",
                        "item_hint": "door demolition",
                        "spec_or_method": "remove stainless glass door",
                        "suggested_unit": "set",
                        "text": "remove stainless glass door and hardware",
                        "confidence": 0.88,
                    }
                ],
            }
        ]
    }


def test_gap_recall_importer_matches_external_call_results(tmp_path: Path):
    report = build_gap_recall_external_import_report(
        _external_results(),
        recall_plan=_recall_plan(tmp_path),
        source_name="offline-glm-export",
    )

    assert report["summary"]["input_row_count"] == 1
    assert report["summary"]["evidence_count"] == 1
    assert report["summary"]["status_counts"] == {"external_imported": 1}
    assert report["summary"]["validation_status_counts"] == {"imported": 1}
    row = report["evidence_rows"][0]
    assert row["source_kind"] == "pdf_gap_recall_external_import"
    assert row["source_file"] == "drawing.pdf"
    assert row["page"] == 1
    assert row["tile_id"] == "p001_whole"
    assert row["vision_pass"] == "door_window_demolition"
    assert row["item_hint"] == "door demolition"


def test_gap_recall_importer_keeps_unmatched_evidence_for_review():
    report = build_gap_recall_external_import_report(
        {
            "evidence_rows": [
                {
                    "source_file": "unknown.pdf",
                    "page": 9,
                    "tile_id": "p009_whole",
                    "vision_pass": "table_legend",
                    "item_hint": "fixture",
                    "text": "fixture note",
                }
            ]
        },
        recall_plan={"plan_rows": []},
        source_name="manual-json",
    )

    assert report["summary"]["evidence_count"] == 1
    assert report["summary"]["unassigned_evidence_count"] == 1
    assert report["validation_rows"][0]["status"] == "imported_unmatched_call"


def test_gap_recall_importer_matches_task_no_through_covered_task_numbers(tmp_path: Path):
    image = tmp_path / "whole.png"
    image.write_bytes(b"fake image bytes")
    recall_plan = {
        "plan_rows": [
            {
                "task_no": 7,
                "gap_no": 7,
                "answer_item_name": "door demolition",
                "recommended_pass": "door_window_demolition",
                "source_file": "drawing.pdf",
                "page": 1,
                "tile_id": "p001_whole",
                "tile_type": "whole_page_preview",
                "image_path": str(image),
            }
        ]
    }

    report = build_gap_recall_external_import_report(
        {
            "evidence_rows": [
                {
                    "task_no": 7,
                    "item_hint": "door demolition",
                    "text": "door schedule note",
                }
            ]
        },
        recall_plan=recall_plan,
        source_name="filled-workbench",
    )

    assert report["summary"]["validation_status_counts"] == {"imported": 1}
    assert report["summary"]["status_counts"] == {"external_imported": 1}
    assert report["evidence_rows"][0]["source_file"] == "drawing.pdf"
    assert report["evidence_rows"][0]["vision_pass"] == "door_window_demolition"


def test_gap_recall_importer_writes_outputs(tmp_path: Path):
    report = build_gap_recall_external_import_report(
        _external_results(),
        recall_plan=_recall_plan(tmp_path),
        source_name="offline-glm-export",
    )

    outputs = write_gap_recall_external_import_outputs(report, tmp_path / "out", stem="imported")

    assert Path(outputs["json"]).exists()
    assert Path(outputs["call_csv"]).exists()
    assert Path(outputs["evidence_csv"]).exists()
    assert Path(outputs["validation_csv"]).exists()
    assert Path(outputs["markdown"]).exists()
    assert Path(outputs["xlsx"]).exists()
    workbook = load_workbook(outputs["xlsx"])
    assert workbook.sheetnames == ["import_summary", "visual_calls", "recall_evidence", "validation"]
    assert workbook["recall_evidence"]["A2"].value == "PDFGAP-EXT-000001"


def test_gap_recall_importer_loads_csv(tmp_path: Path):
    csv_path = tmp_path / "external.csv"
    csv_path.write_text(
        "source_file,page,tile_id,vision_pass,item_hint,text\n"
        "drawing.pdf,1,p001_whole,door_window_demolition,door demolition,remove door\n",
        encoding="utf-8-sig",
    )

    payload = load_external_recall_results(csv_path)

    assert payload["evidence_rows"][0]["item_hint"] == "door demolition"


def test_gap_recall_importer_loads_filled_external_template_xlsx(tmp_path: Path):
    template = build_external_recall_template(_recall_plan(tmp_path))
    outputs = write_external_recall_template_outputs(template, tmp_path / "template", stem="filled")
    workbook = load_workbook(outputs["xlsx"])
    sheet = workbook["external_recall_template"]
    headers = {cell.value: cell.column for cell in sheet[1]}
    sheet.cell(row=2, column=headers["item_hint"]).value = "door demolition"
    sheet.cell(row=2, column=headers["spec_or_method"]).value = "remove stainless glass door and hardware"
    sheet.cell(row=2, column=headers["suggested_unit"]).value = "set"
    sheet.cell(row=2, column=headers["text"]).value = "door schedule note: remove stainless glass door and hardware"
    sheet.cell(row=2, column=headers["confidence"]).value = 0.9
    workbook.save(outputs["xlsx"])
    workbook.close()

    external_results = load_external_recall_results(outputs["xlsx"])
    report = build_gap_recall_external_import_report(
        external_results,
        recall_plan=_recall_plan(tmp_path),
        source_name="filled-template",
    )

    assert report["summary"]["input_row_count"] == 1
    assert report["summary"]["evidence_count"] == 1
    assert report["summary"]["validation_status_counts"] == {"imported": 1}
    row = report["evidence_rows"][0]
    assert row["source_file"] == "drawing.pdf"
    assert row["vision_pass"] == "door_window_demolition"
    assert row["item_hint"] == "door demolition"
    assert row["spec_or_method"] == "remove stainless glass door and hardware"
    assert row["suggested_unit"] == "set"


def test_gap_recall_importer_cli(tmp_path: Path, monkeypatch, capsys):
    plan_json = tmp_path / "plan.json"
    external_json = tmp_path / "external.json"
    plan_json.write_text(json.dumps(_recall_plan(tmp_path), ensure_ascii=False), encoding="utf-8")
    external_json.write_text(json.dumps(_external_results(), ensure_ascii=False), encoding="utf-8")
    output_dir = tmp_path / "cli_out"
    monkeypatch.setattr(
        sys,
        "argv",
        [
            "biz2x_pdf_gap_recall_import.py",
            "--external-results",
            str(external_json),
            "--recall-plan-json",
            str(plan_json),
            "--output-dir",
            str(output_dir),
            "--source-name",
            "offline-glm-export",
            "--timestamp",
            "fixed",
        ],
    )

    assert biz2x_pdf_gap_recall_import.main() == 0
    payload = json.loads(capsys.readouterr().out)

    assert payload["ok"] is True
    assert payload["summary"]["evidence_count"] == 1
    assert Path(payload["outputs"]["json"]).exists()
