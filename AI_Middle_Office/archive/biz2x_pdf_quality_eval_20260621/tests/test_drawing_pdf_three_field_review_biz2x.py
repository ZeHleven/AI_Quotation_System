from __future__ import annotations

import json
import sys
from pathlib import Path

from openpyxl import load_workbook

from app.services.drawing_pdf_three_field_review import (
    build_three_field_human_review_report,
    write_three_field_human_review_outputs,
)
from scripts import biz2x_pdf_three_field_review


def _v2_report() -> dict[str, object]:
    return {
        "summary": {
            "quantity_status": "deferred_until_three_fields_accepted",
        },
        "evidence_rows": [
            {
                "evidence_id": "EVD-001",
                "source_file": "03.pdf",
                "page": 2,
                "tile_id": "p002_whole",
            }
        ],
        "human_style_rows": [
            {
                "source_files": "03.pdf",
                "evidence_ids": "EVD-001",
                "standard_code": "GBT50854-2024",
                "standard_item_code": "011102003",
                "standard_item_name": "块料楼地面",
            }
        ],
        "three_field_acceptance_report": {
            "summary": {
                "answer_count": 1,
                "candidate_count": 1,
                "matched_three_fields_count": 1,
                "three_field_pass_rate": 1.0,
            },
            "answer_rows": [
                {
                    "sheet_name": "装修工程量清单",
                    "row_no": 5,
                    "section": "楼地面工程",
                    "item_code": "011102003",
                    "item_name": "块料楼地面",
                    "feature": "600x600地砖",
                    "unit": "m2",
                }
            ],
            "candidate_rows": [
                {
                    "source": "03.pdf",
                    "row_no": 1,
                    "item_name": "块料楼地面",
                    "feature": "600x600地砖",
                    "unit": "m2",
                    "raw": "{}",
                }
            ],
            "comparison_rows": [
                {
                    "status": "matched_three_fields",
                    "answer_sheet": "装修工程量清单",
                    "answer_row_no": 5,
                    "answer_item_name": "块料楼地面",
                    "answer_feature": "600x600地砖",
                    "answer_unit": "m2",
                    "candidate_row_no": 1,
                    "candidate_source": "03.pdf",
                    "candidate_item_name": "块料楼地面",
                    "candidate_feature": "600x600地砖",
                    "candidate_unit": "m2",
                    "name_score": 1.0,
                    "feature_score": 1.0,
                    "unit_score": 1.0,
                    "overall_score": 1.0,
                    "issue": "三字段通过",
                }
            ],
        },
    }


def test_three_field_human_review_adds_field_flags_and_evidence_page():
    report = build_three_field_human_review_report(_v2_report())

    assert report["summary"]["review_row_count"] == 1
    assert report["summary"]["quantity_status"] == "deferred_until_three_fields_accepted"
    row = report["review_rows"][0]
    assert row["review_result"] == "pass"
    assert row["name_match"] == "yes"
    assert row["feature_match"] == "yes"
    assert row["unit_match"] == "yes"
    assert row["answer_section"] == "楼地面工程"
    assert row["evidence_pages"] == "2"
    assert row["evidence_tiles"] == "p002_whole"
    assert row["standard_item_code"] == "011102003"


def test_three_field_human_review_writes_outputs(tmp_path: Path):
    report = build_three_field_human_review_report(_v2_report())

    outputs = write_three_field_human_review_outputs(report, tmp_path, stem="review")

    assert Path(outputs["json"]).exists()
    assert Path(outputs["csv"]).exists()
    assert Path(outputs["markdown"]).exists()
    assert Path(outputs["xlsx"]).exists()
    workbook = load_workbook(outputs["xlsx"])
    assert workbook.sheetnames == ["summary", "three_field_review"]
    assert workbook["three_field_review"]["A1"].value == "review_no"
    assert workbook["three_field_review"]["B2"].value == "pass"


def test_three_field_human_review_cli(tmp_path: Path, monkeypatch, capsys):
    v2_json = tmp_path / "v2.json"
    v2_json.write_text(json.dumps(_v2_report(), ensure_ascii=False), encoding="utf-8")
    output_dir = tmp_path / "cli_out"
    monkeypatch.setattr(
        sys,
        "argv",
        [
            "biz2x_pdf_three_field_review.py",
            "--v2-json",
            str(v2_json),
            "--output-dir",
            str(output_dir),
            "--timestamp",
            "fixed",
        ],
    )

    assert biz2x_pdf_three_field_review.main() == 0
    payload = json.loads(capsys.readouterr().out)

    assert payload["ok"] is True
    assert payload["summary"]["review_row_count"] == 1
    assert Path(payload["outputs"]["xlsx"]).exists()
