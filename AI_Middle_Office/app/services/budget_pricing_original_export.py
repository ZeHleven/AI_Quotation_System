from __future__ import annotations

import json
import re
from copy import copy
from dataclasses import dataclass
from datetime import datetime
from decimal import Decimal, InvalidOperation
from io import BytesIO
from typing import Any, Iterable

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from app.models.budget_project import BudgetProjectImportBatch, BudgetProjectImportSheetMapping, BudgetProjectProfile
from app.models.budget_pricing_draft import BudgetProjectPricingDraft, BudgetProjectPricingDraftLine
from app.models.file_object import FileObject
from app.models.user import User
from app.services.budget_pricing import BudgetPricingError
from app.services.budget_pricing_drafts import (
    ensure_budget_pricing_draft_uses_active_import,
    get_current_budget_pricing_draft,
)


MONEY_NUMBER_FORMAT = "#,##0.00"
RATE_NUMBER_FORMAT = "0.00%"
WORKBOOK_CONTENT_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
WORKBOOK_MACRO_CONTENT_TYPE = "application/vnd.ms-excel.sheet.macroEnabled.12"


@dataclass(frozen=True)
class ExportField:
    key: str
    label: str
    aliases: tuple[str, ...]
    number_format: str | None = None
    preserve_existing_value: bool = False


@dataclass(frozen=True)
class HeaderLayout:
    rows: tuple[int, ...]
    table_right_column: int


EXPORT_FIELDS: tuple[ExportField, ...] = (
    ExportField("item_name", "项目名称", ("项目名称", "分部分项工程", "清单项目", "项目", "名称", "子目名称"), preserve_existing_value=True),
    ExportField("spec", "特征描述", ("项目特征描述", "特征描述", "项目特征", "清单特征", "特征", "规格描述", "规格"), preserve_existing_value=True),
    ExportField("region", "区域", ("区域", "分区", "部位", "位置", "施工区域"), preserve_existing_value=True),
    ExportField("material_supply_mode", "主材采购方式", ("主材采购方式", "主材供应方式", "主材采购", "采购方式", "甲乙供"), preserve_existing_value=True),
    ExportField("quantity", "工程量", ("工程量", "工程量小计", "数量", "计量数量", "工程数量"), preserve_existing_value=True),
    ExportField("effective_unit_price", "不含税综合单价", ("不含税综合单价", "税前综合单价", "除税综合单价", "综合单价", "单价"), MONEY_NUMBER_FORMAT),
    ExportField("line_total", "不含税综合合价", ("不含税综合合价", "税前综合合价", "除税综合合价", "综合合价", "合价", "金额", "总价"), MONEY_NUMBER_FORMAT),
    ExportField("labor_unit_cost", "人工费", ("人工费", "人工单价", "人工成本"), MONEY_NUMBER_FORMAT),
    ExportField("main_material_unit_cost", "主材费", ("主材费", "主材单价", "主材成本"), MONEY_NUMBER_FORMAT),
    ExportField("auxiliary_material_unit_cost", "辅材费", ("辅材费", "辅助材料费", "辅材单价", "辅材成本"), MONEY_NUMBER_FORMAT),
    ExportField("tax_amount", "税金", ("税金", "税费", "税额"), MONEY_NUMBER_FORMAT),
    ExportField("main_material_without_loss", "主材费不含损耗", ("主材费不含损耗", "主材不含损耗", "主材净价"), MONEY_NUMBER_FORMAT),
    ExportField("loss_rate", "损耗率", ("损耗率", "损耗"), RATE_NUMBER_FORMAT),
    ExportField("machinery_unit_cost", "机械费", ("机械费", "机械单价", "机械成本"), MONEY_NUMBER_FORMAT),
    ExportField("comprehensive_unit_cost", "综合费", ("综合费", "综合费用", "综合单价费用"), MONEY_NUMBER_FORMAT),
    ExportField("management_unit_cost", "管理费", ("管理费", "管理单价", "管理成本"), MONEY_NUMBER_FORMAT),
    ExportField("profit_unit_cost", "利润费", ("利润费", "利润", "利润单价"), MONEY_NUMBER_FORMAT),
    ExportField("measure_unit_cost", "措施费", ("措施费", "措施项目费", "措施单价"), MONEY_NUMBER_FORMAT),
    ExportField("owner_material_unit_price", "甲供材单价", ("甲供材单价", "甲供材料单价", "甲供材价"), MONEY_NUMBER_FORMAT),
    ExportField("owner_material_loss_amount", "甲供材损耗金", ("甲供材损耗金", "甲供材损耗", "甲供材料损耗金"), MONEY_NUMBER_FORMAT),
)


@dataclass(frozen=True)
class BudgetPricingOriginalExportSource:
    draft: BudgetProjectPricingDraft
    batch: BudgetProjectImportBatch
    source_file: FileObject
    sheet_mappings: tuple[BudgetProjectImportSheetMapping, ...]
    lines: tuple[BudgetProjectPricingDraftLine, ...]


@dataclass(frozen=True)
class BudgetPricingOriginalExportResult:
    content: bytes
    filename: str
    content_type: str
    summary: dict[str, Any]


def prepare_budget_pricing_original_export(
    db: Session,
    profile: BudgetProjectProfile,
    current_user: User,
    *,
    pricing_mode: str | None = None,
) -> BudgetPricingOriginalExportSource:
    draft = get_current_budget_pricing_draft(
        db,
        profile,
        current_user,
        pricing_mode=pricing_mode,
    )
    if draft is None:
        raise BudgetPricingError("BUDGET_PRICING_DRAFT_NOT_FOUND", status_code=404)
    ensure_budget_pricing_draft_uses_active_import(profile, draft)

    batch = db.query(BudgetProjectImportBatch).filter(BudgetProjectImportBatch.id == draft.source_import_batch_id).one_or_none()
    if batch is None:
        raise BudgetPricingError("BUDGET_PRICING_EXPORT_IMPORT_BATCH_NOT_FOUND", status_code=409)
    source_file = batch.source_file_object
    if source_file is None:
        raise BudgetPricingError("BUDGET_PRICING_EXPORT_SOURCE_FILE_NOT_RETAINED", status_code=409)

    sheet_mappings = (
        db.query(BudgetProjectImportSheetMapping)
        .filter(BudgetProjectImportSheetMapping.batch_id == batch.id)
        .order_by(BudgetProjectImportSheetMapping.id.asc())
        .all()
    )
    lines = (
        db.query(BudgetProjectPricingDraftLine)
        .filter(BudgetProjectPricingDraftLine.draft_id == draft.id)
        .order_by(BudgetProjectPricingDraftLine.source_sort_order.asc(), BudgetProjectPricingDraftLine.id.asc())
        .all()
    )
    return BudgetPricingOriginalExportSource(
        draft=draft,
        batch=batch,
        source_file=source_file,
        sheet_mappings=tuple(sheet_mappings),
        lines=tuple(lines),
    )


def render_budget_pricing_original_export(
    source_content: bytes,
    source: BudgetPricingOriginalExportSource,
    *,
    project_name: str,
) -> BudgetPricingOriginalExportResult:
    keep_vba = source.batch.source_filename.lower().endswith(".xlsm")
    try:
        workbook = load_workbook(BytesIO(source_content), data_only=False, keep_vba=keep_vba)
    except Exception as exc:  # pragma: no cover - openpyxl has implementation-specific errors.
        raise BudgetPricingError("BUDGET_PRICING_EXPORT_WORKBOOK_INVALID", status_code=422) from exc

    mappings_by_sheet = {mapping.sheet_name: mapping for mapping in source.sheet_mappings}
    lines_by_sheet: dict[str, list[BudgetProjectPricingDraftLine]] = {}
    for line in source.lines:
        lines_by_sheet.setdefault(line.source_sheet, []).append(line)

    summary: dict[str, Any] = {
        "total_lines": len(source.lines),
        "priced_lines": sum(1 for line in source.lines if line.effective_unit_price is not None),
        "pending_lines": sum(1 for line in source.lines if line.effective_unit_price is None),
        "sheets": [],
        "unresolved": [],
    }
    for sheet_name, lines in lines_by_sheet.items():
        worksheet = workbook[sheet_name] if sheet_name in workbook.sheetnames else None
        mapping = mappings_by_sheet.get(sheet_name)
        if worksheet is None:
            summary["unresolved"].append({"sheet": sheet_name, "reason": "WORKSHEET_NOT_FOUND", "line_count": len(lines)})
            continue
        if mapping is None or not mapping.header_row_index:
            summary["unresolved"].append({"sheet": sheet_name, "reason": "HEADER_ROW_NOT_FOUND", "line_count": len(lines)})
            continue
        sheet_summary = _write_sheet_pricing_fields(
            worksheet,
            header_row=int(mapping.header_row_index),
            lines=lines,
        )
        summary["sheets"].append(sheet_summary)
        for unresolved in sheet_summary["unresolved"]:
            summary["unresolved"].append({"sheet": sheet_name, **unresolved})

    _append_export_note_sheet(workbook, source=source, project_name=project_name, summary=summary)
    output = BytesIO()
    workbook.save(output)
    extension = ".xlsm" if keep_vba else ".xlsx"
    filename = _export_filename(project_name, source.draft.pricing_mode, source.draft.revision, extension)
    return BudgetPricingOriginalExportResult(
        content=output.getvalue(),
        filename=filename,
        content_type=WORKBOOK_MACRO_CONTENT_TYPE if keep_vba else WORKBOOK_CONTENT_TYPE,
        summary=summary,
    )


def _write_sheet_pricing_fields(worksheet: Any, *, header_row: int, lines: Iterable[BudgetProjectPricingDraftLine]) -> dict[str, Any]:
    layout = _detect_header_layout(worksheet, header_row)
    header_columns, unresolved = _match_existing_header_columns(worksheet, layout)
    missing_fields = [
        field
        for field in EXPORT_FIELDS
        if field.key not in header_columns and not any(item["field"] == field.label for item in unresolved)
    ]
    appended_columns = _append_missing_fields(
        worksheet,
        layout=layout,
        header_columns=header_columns,
        fields=missing_fields,
    )
    header_columns.update(appended_columns)
    appended_fields = [field.label for field in missing_fields]

    written_count = 0
    pending_count = 0
    for line in lines:
        if line.source_raw_row_index <= layout.rows[-1] or line.source_raw_row_index > worksheet.max_row:
            unresolved.append({"field": "整行", "reason": "SOURCE_ROW_OUT_OF_RANGE", "row": line.source_raw_row_index})
            continue
        values = _line_field_values(line)
        if line.effective_unit_price is None:
            pending_count += 1
        for field in EXPORT_FIELDS:
            column_index = header_columns.get(field.key)
            if column_index is None:
                continue
            value = values.get(field.key)
            if value is None:
                continue
            cell = worksheet.cell(row=line.source_raw_row_index, column=column_index)
            if field.preserve_existing_value and cell.value not in (None, ""):
                continue
            cell.value = value
            if field.number_format and (
                column_index in appended_columns.values() or _is_generic_number_format(cell.number_format)
            ):
                cell.number_format = field.number_format
        written_count += 1
    return {
        "sheet": worksheet.title,
        "written_line_count": written_count,
        "pending_line_count": pending_count,
        "appended_fields": appended_fields,
        "matched_fields": [field.label for field in EXPORT_FIELDS if field.key in header_columns],
        "unresolved": unresolved,
    }


def _detect_header_layout(worksheet: Any, header_row: int) -> HeaderLayout:
    max_header_row = header_row
    for merged_range in worksheet.merged_cells.ranges:
        if merged_range.min_row == header_row and merged_range.max_row > header_row:
            max_header_row = max(max_header_row, merged_range.max_row)
        if (
            merged_range.min_row == header_row
            and merged_range.max_row == header_row
            and merged_range.max_col > merged_range.min_col
            and any(
                worksheet.cell(row=header_row + 1, column=column_index).value not in (None, "")
                for column_index in range(merged_range.min_col, merged_range.max_col + 1)
            )
        ):
            max_header_row = max(max_header_row, header_row + 1)

    if max_header_row == header_row:
        child_header_count = sum(
            1
            for column_index in range(1, worksheet.max_column + 1)
            if _normalize_header(worksheet.cell(row=header_row + 1, column=column_index).value)
            in {_normalize_header(alias) for field in EXPORT_FIELDS for alias in field.aliases}
        )
        if child_header_count >= 2:
            max_header_row = header_row + 1

    header_rows = tuple(range(header_row, min(max_header_row, header_row + 2) + 1))
    table_right_column = _find_table_right_column(worksheet, header_rows)
    return HeaderLayout(rows=header_rows, table_right_column=table_right_column)


def _find_table_right_column(worksheet: Any, header_rows: tuple[int, ...]) -> int:
    right_column = 1
    for row_index in range(header_rows[0], worksheet.max_row + 1):
        for column_index in range(1, worksheet.max_column + 1):
            if worksheet.cell(row=row_index, column=column_index).value not in (None, ""):
                right_column = max(right_column, column_index)
    for merged_range in worksheet.merged_cells.ranges:
        if merged_range.max_row >= header_rows[0] and merged_range.min_row <= header_rows[-1]:
            right_column = max(right_column, merged_range.max_col)
    return right_column


def _match_existing_header_columns(worksheet: Any, layout: HeaderLayout) -> tuple[dict[str, int], list[dict[str, Any]]]:
    normalized_headers: dict[str, list[int]] = {}
    for column_index in range(1, layout.table_right_column + 1):
        for normalized in _column_header_labels(worksheet, layout.rows, column_index):
            normalized_headers.setdefault(normalized, []).append(column_index)

    resolved: dict[str, int] = {}
    unresolved: list[dict[str, Any]] = []
    occupied: set[int] = set()
    for field in EXPORT_FIELDS:
        candidate_scores: dict[int, int] = {}
        for alias in field.aliases:
            normalized_alias = _normalize_header(alias)
            for normalized_header, column_indexes in normalized_headers.items():
                match_score = _header_alias_match_score(normalized_header, normalized_alias)
                if match_score is None:
                    continue
                for column_index in column_indexes:
                    candidate_scores[column_index] = max(candidate_scores.get(column_index, 0), match_score)
        candidate_scores = {
            column_index: score
            for column_index, score in candidate_scores.items()
            if column_index not in occupied
        }
        highest_score = max(candidate_scores.values(), default=None)
        candidates = [
            column_index
            for column_index, score in candidate_scores.items()
            if score == highest_score
        ]
        if len(candidates) == 1:
            resolved[field.key] = candidates[0]
            occupied.add(candidates[0])
        elif field.key == "quantity" and candidates:
            # Multi-region bills repeat the same quantity header. The import mapper uses
            # the first source column for its calculation quantity, so retain that column.
            selected_column = min(candidates)
            resolved[field.key] = selected_column
            occupied.add(selected_column)
        elif len(candidates) > 1:
            unresolved.append({
                "field": field.label,
                "reason": "MULTIPLE_HEADER_CANDIDATES",
                "columns": [get_column_letter(column_index) for column_index in candidates],
            })
    return resolved, unresolved


def _column_header_labels(worksheet: Any, header_rows: tuple[int, ...], column_index: int) -> set[str]:
    top_label = _normalize_header(_merged_cell_value(worksheet, header_rows[0], column_index))
    if not top_label:
        return set()
    labels = {top_label}
    for row_index in header_rows[1:]:
        child_label = _normalize_header(_merged_cell_value(worksheet, row_index, column_index))
        if child_label:
            labels.add(child_label)
    return labels


def _header_alias_match_score(normalized_header: str, normalized_alias: str) -> int | None:
    if normalized_header == normalized_alias:
        return 1000 + len(normalized_alias)
    if (
        len(normalized_alias) >= 4
        and normalized_header.startswith(normalized_alias)
        and "分析" not in normalized_header
    ):
        return 500 + len(normalized_alias)
    return None


def _merged_cell_value(worksheet: Any, row_index: int, column_index: int) -> Any:
    for merged_range in worksheet.merged_cells.ranges:
        if (
            merged_range.min_row <= row_index <= merged_range.max_row
            and merged_range.min_col <= column_index <= merged_range.max_col
        ):
            return worksheet.cell(row=merged_range.min_row, column=merged_range.min_col).value
    return worksheet.cell(row=row_index, column=column_index).value


def _append_missing_fields(
    worksheet: Any,
    *,
    layout: HeaderLayout,
    header_columns: dict[str, int],
    fields: list[ExportField],
) -> dict[str, int]:
    if not fields:
        return {}

    start_column = layout.table_right_column + 1
    appended_columns = {field.key: start_column + index for index, field in enumerate(fields)}
    for field in fields:
        target_column = appended_columns[field.key]
        anchor_column = _style_anchor_column(field, header_columns, layout.table_right_column)
        _copy_appended_column_styles(
            worksheet,
            layout=layout,
            target_column=target_column,
            anchor_column=anchor_column,
            field=field,
        )

    if len(layout.rows) > 1:
        _write_grouped_headers(
            worksheet,
            layout=layout,
            header_columns=header_columns,
            fields=fields,
            appended_columns=appended_columns,
        )
    else:
        for field in fields:
            target_column = appended_columns[field.key]
            anchor_column = _style_anchor_column(field, header_columns, layout.table_right_column)
            source_header = _style_source_cell(worksheet, layout.rows[0], anchor_column)
            target_header = worksheet.cell(row=layout.rows[0], column=target_column)
            _copy_cell_style(source_header, target_header)
            target_header.value = field.label
    return appended_columns


def _write_grouped_headers(
    worksheet: Any,
    *,
    layout: HeaderLayout,
    header_columns: dict[str, int],
    fields: list[ExportField],
    appended_columns: dict[str, int],
) -> None:
    first_column = appended_columns[fields[0].key]
    last_column = appended_columns[fields[-1].key]
    group_source = _group_header_style_source(worksheet, layout)
    group_cell = worksheet.cell(row=layout.rows[0], column=first_column)
    _copy_cell_style(group_source, group_cell)
    group_cell.value = "系统报价补充"
    if first_column != last_column:
        worksheet.merge_cells(
            start_row=layout.rows[0],
            start_column=first_column,
            end_row=layout.rows[0],
            end_column=last_column,
        )

    child_row = layout.rows[-1]
    for field in fields:
        target_column = appended_columns[field.key]
        anchor_column = _style_anchor_column(field, header_columns, layout.table_right_column)
        target_header = worksheet.cell(row=child_row, column=target_column)
        _copy_cell_style(_style_source_cell(worksheet, child_row, anchor_column), target_header)
        target_header.value = field.label


def _copy_appended_column_styles(
    worksheet: Any,
    *,
    layout: HeaderLayout,
    target_column: int,
    anchor_column: int,
    field: ExportField,
) -> None:
    source_letter = get_column_letter(anchor_column)
    target_letter = get_column_letter(target_column)
    source_width = worksheet.column_dimensions[source_letter].width
    worksheet.column_dimensions[target_letter].width = _appended_column_width(field, source_width)
    for row_index in range(layout.rows[-1] + 1, worksheet.max_row + 1):
        _copy_cell_style(
            _style_source_cell(worksheet, row_index, anchor_column),
            worksheet.cell(row=row_index, column=target_column),
        )


def _style_anchor_column(field: ExportField, header_columns: dict[str, int], table_right_column: int) -> int:
    if field.key in {"spec", "region", "material_supply_mode"}:
        candidates = ("spec", "quantity", "item_name")
    elif field.key == "quantity":
        candidates = ("quantity", "effective_unit_price", "line_total")
    elif field.key == "loss_rate":
        candidates = ("loss_rate", "effective_unit_price", "line_total", "quantity")
    else:
        candidates = (
            "effective_unit_price",
            "line_total",
            "labor_unit_cost",
            "main_material_unit_cost",
            "comprehensive_unit_cost",
            "quantity",
        )
    return next((header_columns[key] for key in candidates if key in header_columns), table_right_column)


def _group_header_style_source(worksheet: Any, layout: HeaderLayout) -> Any:
    for merged_range in worksheet.merged_cells.ranges:
        if (
            merged_range.min_row == layout.rows[0]
            and merged_range.max_row == layout.rows[0]
            and merged_range.max_col > merged_range.min_col
        ):
            return worksheet.cell(row=merged_range.min_row, column=merged_range.min_col)
    return _style_source_cell(worksheet, layout.rows[0], layout.table_right_column)


def _style_source_cell(worksheet: Any, row_index: int, column_index: int) -> Any:
    for merged_range in worksheet.merged_cells.ranges:
        if (
            merged_range.min_row <= row_index <= merged_range.max_row
            and merged_range.min_col <= column_index <= merged_range.max_col
        ):
            return worksheet.cell(row=merged_range.min_row, column=merged_range.min_col)
    return worksheet.cell(row=row_index, column=column_index)


def _appended_column_width(field: ExportField, source_width: float | None) -> float:
    preferred_widths = {
        "spec": 26,
        "region": 12,
        "material_supply_mode": 14,
        "quantity": 12,
        "effective_unit_price": 15,
        "line_total": 15,
        "main_material_without_loss": 16,
        "owner_material_unit_price": 15,
        "owner_material_loss_amount": 16,
    }
    preferred = preferred_widths.get(field.key, 12)
    if not source_width:
        return float(preferred)
    return float(min(max(source_width, 10), preferred))


def _is_generic_number_format(number_format: str | None) -> bool:
    return not number_format or number_format.lower() == "general"


def _copy_cell_style(source: Any, target: Any) -> None:
    if source.has_style:
        target._style = copy(source._style)
    if source.number_format:
        target.number_format = source.number_format
    if source.alignment:
        target.alignment = copy(source.alignment)
    if source.protection:
        target.protection = copy(source.protection)


def _line_field_values(line: BudgetProjectPricingDraftLine) -> dict[str, Any]:
    source_snapshot = _json_dict(line.source_row_snapshot_json)
    raw_fields = source_snapshot.get("raw_fields") if isinstance(source_snapshot.get("raw_fields"), dict) else {}
    breakdown = _json_dict(line.pricing_breakdown_json)
    return {
        "item_name": line.item_name,
        "spec": line.spec,
        "region": _first_value(source_snapshot.get("region"), source_snapshot.get("area"), source_snapshot.get("work_area"), raw_fields.get("区域"), raw_fields.get("部位")),
        "material_supply_mode": breakdown.get("material_supply_mode"),
        "quantity": _number_value(line.calculation_quantity),
        "effective_unit_price": _number_value(line.effective_unit_price),
        "line_total": _number_value(line.line_total),
        "labor_unit_cost": _number_value(breakdown.get("labor_unit_cost")),
        "main_material_unit_cost": _number_value(breakdown.get("main_material_unit_cost")),
        "auxiliary_material_unit_cost": _number_value(breakdown.get("auxiliary_material_unit_cost")),
        "tax_amount": _number_value(breakdown.get("tax_amount")),
        "main_material_without_loss": _number_value(breakdown.get("main_material_without_loss")),
        "loss_rate": _number_value(breakdown.get("loss_rate")),
        "machinery_unit_cost": _number_value(breakdown.get("machinery_unit_cost")),
        "comprehensive_unit_cost": _number_value(breakdown.get("comprehensive_unit_cost")),
        "management_unit_cost": _number_value(breakdown.get("management_unit_cost")),
        "profit_unit_cost": _number_value(breakdown.get("profit_unit_cost")),
        "measure_unit_cost": _number_value(breakdown.get("measure_unit_cost")),
        "owner_material_unit_price": _number_value(breakdown.get("owner_material_unit_price")),
        "owner_material_loss_amount": _number_value(breakdown.get("owner_material_loss_amount")),
    }


def _append_export_note_sheet(workbook: Any, *, source: BudgetPricingOriginalExportSource, project_name: str, summary: dict[str, Any]) -> None:
    title = "系统估算说明"
    suffix = 2
    while title in workbook.sheetnames:
        title = f"系统估算说明{suffix}"
        suffix += 1
    worksheet = workbook.create_sheet(title=title)
    worksheet.append(["项目名称", project_name])
    worksheet.append(["原导入文件", source.batch.source_filename])
    worksheet.append(["导入批次", source.batch.batch_uuid])
    worksheet.append(["计价模式", source.draft.pricing_mode])
    worksheet.append(["草稿修订", source.draft.revision])
    worksheet.append(["导出时间", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
    worksheet.append(["计价行数", summary["total_lines"]])
    worksheet.append(["已计价行", summary["priced_lines"]])
    worksheet.append(["待人工补价行", summary["pending_lines"]])
    worksheet.append([])
    worksheet.append(["Sheet", "自动追加字段", "未写入原因"])
    for item in summary["sheets"]:
        reasons = "; ".join(
            f"{issue.get('field')}: {issue.get('reason')}"
            for issue in item["unresolved"]
        )
        worksheet.append([item["sheet"], "、".join(item["appended_fields"]), reasons])
    for item in summary["unresolved"]:
        if item.get("sheet") not in {entry["sheet"] for entry in summary["sheets"]}:
            worksheet.append([item.get("sheet"), "", item.get("reason")])
    worksheet.column_dimensions["A"].width = 20
    worksheet.column_dimensions["B"].width = 48
    worksheet.column_dimensions["C"].width = 42


def _json_dict(value: Any) -> dict[str, Any]:
    if isinstance(value, dict):
        return value
    if not value:
        return {}
    try:
        parsed = json.loads(value)
    except (TypeError, ValueError):
        return {}
    return parsed if isinstance(parsed, dict) else {}


def _first_value(*values: Any) -> Any:
    for value in values:
        if value not in (None, ""):
            return value
    return None


def _number_value(value: Any) -> float | None:
    if value in (None, ""):
        return None
    try:
        return float(Decimal(str(value)))
    except (InvalidOperation, TypeError, ValueError):
        return None


def _normalize_header(value: Any) -> str:
    text = str(value or "").strip().lower()
    return re.sub(r"[\s\-_（）()【】\[\]：:，,。./\\]+", "", text)


def _export_filename(project_name: str, pricing_mode: str, revision: int, extension: str) -> str:
    safe_project_name = re.sub(r'[\\/:*?"<>|]+', "_", project_name or "项目")[:80].strip(" .") or "项目"
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return f"{safe_project_name}_原格式报价_{pricing_mode}_Rev{revision}_{timestamp}{extension}"
