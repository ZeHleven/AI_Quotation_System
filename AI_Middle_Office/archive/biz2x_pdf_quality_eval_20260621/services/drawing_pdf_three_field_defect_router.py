from __future__ import annotations

import csv
import json
import re
from collections import Counter
from datetime import datetime
from pathlib import Path
from typing import Any, Iterable, Mapping, Sequence

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


PHASE = "BIZ-2x-pdf-three-field-defect-router"

DEFECT_HEADERS = [
    "defect_no",
    "review_no",
    "status",
    "repair_route",
    "repair_priority",
    "repair_action",
    "object_class",
    "answer_sheet",
    "answer_row_no",
    "answer_section",
    "answer_item_name",
    "answer_feature",
    "answer_unit",
    "candidate_row_no",
    "candidate_item_name",
    "candidate_feature",
    "candidate_unit",
    "name_score",
    "feature_score",
    "unit_score",
    "evidence_ids",
    "candidate_source_files",
    "evidence_pages",
    "evidence_tiles",
    "issue",
    "suggested_review_focus",
    "suggested_unit_resolution",
    "suggested_feature_focus",
    "ready_for_auto_accept",
    "quantity_status",
]

ROUTE_HEADERS = [
    "repair_route",
    "task_count",
    "p1_count",
    "p2_count",
    "p3_count",
    "object_classes",
    "next_action",
]

SUMMARY_HEADERS = ["metric", "value"]


def build_three_field_defect_router_report(review_report: Mapping[str, Any]) -> dict[str, Any]:
    review_rows = [dict(row) for row in review_report.get("review_rows") or [] if isinstance(row, Mapping)]
    defect_rows: list[dict[str, Any]] = []
    for row in review_rows:
        if str(row.get("status") or "") == "matched_three_fields":
            continue
        defect_rows.append(_build_defect_row(row, defect_no=len(defect_rows) + 1))

    route_summary_rows = _route_summary(defect_rows)
    summary = {
        "answer_count": (review_report.get("summary") or {}).get("answer_count", 0),
        "candidate_count": (review_report.get("summary") or {}).get("candidate_count", 0),
        "matched_three_fields_count": (review_report.get("summary") or {}).get("matched_three_fields_count", 0),
        "defect_count": len(defect_rows),
        "route_counts": dict(Counter(str(row.get("repair_route") or "") for row in defect_rows)),
        "status_counts": dict(Counter(str(row.get("status") or "") for row in defect_rows)),
        "priority_counts": dict(Counter(str(row.get("repair_priority") or "") for row in defect_rows)),
        "object_class_counts": dict(Counter(str(row.get("object_class") or "") for row in defect_rows)),
        "auto_acceptable_count": 0,
        "quantity_status": "deferred_until_three_fields_accepted",
        "can_enable_quantity": False,
    }
    return {
        "ok": True,
        "phase": PHASE,
        "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "summary": summary,
        "route_summary_rows": route_summary_rows,
        "defect_rows": defect_rows,
    }


def write_three_field_defect_router_outputs(
    report: Mapping[str, Any],
    output_dir: str | Path,
    *,
    stem: str | None = None,
) -> dict[str, str]:
    target = Path(output_dir)
    target.mkdir(parents=True, exist_ok=True)
    file_stem = stem or f"BIZ2x_PDF_three_field_defect_router_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
    json_path = target / f"{file_stem}.json"
    csv_path = target / f"{file_stem}.csv"
    markdown_path = target / f"{file_stem}.md"
    xlsx_path = target / f"{file_stem}.xlsx"
    outputs = {
        "json": str(json_path),
        "csv": str(csv_path),
        "markdown": str(markdown_path),
        "xlsx": str(xlsx_path),
    }
    payload = {**dict(report), "outputs": outputs}
    json_path.write_text(json.dumps(_json_safe(payload), ensure_ascii=False, indent=2), encoding="utf-8")
    _write_csv(csv_path, payload.get("defect_rows") or [], DEFECT_HEADERS)
    markdown_path.write_text(_build_markdown(payload), encoding="utf-8")
    _write_workbook(xlsx_path, payload)
    return outputs


def _build_defect_row(row: Mapping[str, Any], *, defect_no: int) -> dict[str, Any]:
    status = str(row.get("status") or "")
    object_class = _object_class(row)
    route = _repair_route(row, object_class)
    priority = _repair_priority(row, object_class, route)
    return {
        "defect_no": defect_no,
        "review_no": row.get("review_no", ""),
        "status": status,
        "repair_route": route,
        "repair_priority": priority,
        "repair_action": _repair_action(route, object_class, row),
        "object_class": object_class,
        "answer_sheet": row.get("answer_sheet", ""),
        "answer_row_no": row.get("answer_row_no", ""),
        "answer_section": row.get("answer_section", ""),
        "answer_item_name": row.get("answer_item_name", ""),
        "answer_feature": row.get("answer_feature", ""),
        "answer_unit": row.get("answer_unit", ""),
        "candidate_row_no": row.get("candidate_row_no", ""),
        "candidate_item_name": row.get("candidate_item_name", ""),
        "candidate_feature": row.get("candidate_feature", ""),
        "candidate_unit": row.get("candidate_unit", ""),
        "name_score": row.get("name_score", ""),
        "feature_score": row.get("feature_score", ""),
        "unit_score": row.get("unit_score", ""),
        "evidence_ids": row.get("evidence_ids", ""),
        "candidate_source_files": row.get("candidate_source_files", ""),
        "evidence_pages": row.get("evidence_pages", ""),
        "evidence_tiles": row.get("evidence_tiles", ""),
        "issue": row.get("issue", ""),
        "suggested_review_focus": _review_focus(route, object_class, row),
        "suggested_unit_resolution": _unit_resolution(row) if route == "unit_rule_review" else "",
        "suggested_feature_focus": _feature_focus(row) if route == "feature_enrichment" else "",
        "ready_for_auto_accept": "false",
        "quantity_status": "deferred_until_three_fields_accepted",
    }


def _repair_route(row: Mapping[str, Any], object_class: str) -> str:
    status = str(row.get("status") or "")
    if status == "missing_candidate":
        return "object_evidence_recall"
    if status == "unit_conflict":
        return "unit_rule_review"
    if status == "matched_name_unit_feature_review":
        if _has_variant_conflict(row):
            return "split_variant_review"
        return "feature_enrichment"
    if status == "weak_match_review":
        return "object_match_review"
    return "manual_review"


def _repair_priority(row: Mapping[str, Any], object_class: str, route: str) -> str:
    if route == "object_evidence_recall" and object_class in {"electrical_mep", "water_heater", "fixture_valve_schedule", "door_window_demolition"}:
        return "P1"
    if route in {"unit_rule_review", "object_match_review", "split_variant_review"}:
        return "P1"
    if route == "object_evidence_recall":
        return "P2"
    return "P3"


def _repair_action(route: str, object_class: str, row: Mapping[str, Any]) -> str:
    if route == "object_evidence_recall":
        return "补真实图纸证据后再回灌；只允许填写 item_hint/spec_or_method/suggested_unit/text。"
    if route == "unit_rule_review":
        return "核对对象类别和国标/人工清单单位口径；必要时补单位规则，不直接解锁。"
    if route == "split_variant_review":
        return "拆分或补充材料编号、部位、规格、平级/造型等关键差异证据。"
    if route == "feature_enrichment":
        return "补项目特征证据，重点补材料编号、规格、部位、做法和报价边界。"
    if route == "object_match_review":
        return "核对候选是否为同一清单对象；不同对象需退回缺候选。"
    return "人工复核该行三字段。"


def _review_focus(route: str, object_class: str, row: Mapping[str, Any]) -> str:
    target = str(row.get("answer_item_name") or "").strip()
    if route == "object_evidence_recall":
        return f"找“{target}”本体证据，避免只写泛化说明。"
    if route == "unit_rule_review":
        return f"核对“{target}”与候选“{row.get('candidate_item_name', '')}”是否同一对象，再核单位。"
    if route == "split_variant_review":
        return f"核对“{target}”的材料编号、规格、部位或做法是否被候选合并。"
    if route == "feature_enrichment":
        return f"补足“{target}”项目特征中缺失的施工做法或规格。"
    return f"判断“{target}”与候选是否同一项目。"


def _unit_resolution(row: Mapping[str, Any]) -> str:
    answer_unit = str(row.get("answer_unit") or "").strip() or "-"
    candidate_unit = str(row.get("candidate_unit") or "").strip() or "-"
    answer_name = str(row.get("answer_item_name") or "")
    candidate_name = str(row.get("candidate_item_name") or "")
    if "配电箱" in answer_name and "配电箱" in candidate_name and {answer_unit, candidate_unit} <= {"套", "台"}:
        return "配电箱出现“套/台”口径差异：先按人工验收口径复核，后续可配置同类单位映射。"
    return f"人工单位={answer_unit}；候选单位={candidate_unit}；需按对象类别和标准库单位复核。"


def _feature_focus(row: Mapping[str, Any]) -> str:
    issue = str(row.get("issue") or "")
    answer_text = _compact(" ".join([row.get("answer_item_name", ""), row.get("answer_feature", "")]))
    if "平级/造型" in issue or any(term in answer_text for term in ("平级", "造型", "跌级")):
        return "补平级/造型/跌级、基层、龙骨和面层做法。"
    if "拆除" in answer_text:
        return "补拆除对象、规格尺寸、清运范围和保护要求。"
    if any(term in answer_text for term in ("灯", "配管", "配线", "电缆", "插座", "开关")):
        return "补型号、规格、回路/点位或安装方式。"
    if any(term in answer_text for term in ("水表", "阀", "地漏", "马桶", "台盆", "龙头")):
        return "补规格、材质、连接方式、附件和供货安装范围。"
    return "补材料编号、部位、规格、基层做法和报价范围。"


def _has_variant_conflict(row: Mapping[str, Any]) -> bool:
    issue = str(row.get("issue") or "")
    text = _compact(" ".join([row.get("answer_item_name", ""), row.get("answer_feature", ""), row.get("candidate_item_name", ""), row.get("candidate_feature", "")]))
    return bool("平级/造型" in issue or any(term in text for term in ("平级", "造型", "跌级", "单开", "双开", "dn40", "dn20", "dn15", "de110", "de63")))


def _object_class(row: Mapping[str, Any]) -> str:
    text = _compact(" ".join([row.get("answer_section", ""), row.get("answer_item_name", ""), row.get("answer_feature", "")]))
    if "拆除" in text and any(term in text for term in ("门", "窗", "售卖窗口", "隔断")):
        return "door_window_demolition"
    if "拆除" in text and any(term in text for term in ("马桶", "台盆", "洗手台", "洁具", "地漏", "阀", "水表")):
        return "fixture_demolition"
    if any(term in text for term in ("阀", "水表", "地漏", "马桶", "台盆", "洗脸盆", "花洒", "龙头", "给水", "排水")):
        return "fixture_valve_schedule"
    if any(term in text for term in ("电热水器", "热水器")):
        return "water_heater"
    if any(term in text for term in ("电缆", "配线", "配管", "桥架", "灯具", "筒灯", "射灯", "格栅灯", "灯带", "开关", "插座", "配电箱", "电热水器")):
        return "electrical_mep"
    if "拆除" in text:
        return "demolition_node"
    if any(term in text for term in ("地面", "地砖", "瓷砖", "石材", "防水", "保护层", "挡水条", "美缝")):
        return "finish_floor"
    if any(term in text for term in ("天花", "天棚", "吊顶", "灯槽", "窗帘盒", "矿棉板", "扣板", "涂料")):
        return "finish_ceiling"
    if any(term in text for term in ("墙面", "隔墙", "抹灰", "墙布", "硬包", "回填", "砌筑")):
        return "finish_wall"
    return "general_object"


def _route_summary(rows: Sequence[Mapping[str, Any]]) -> list[dict[str, Any]]:
    grouped: dict[str, list[Mapping[str, Any]]] = {}
    for row in rows:
        grouped.setdefault(str(row.get("repair_route") or ""), []).append(row)
    result: list[dict[str, Any]] = []
    for route, items in sorted(grouped.items()):
        result.append(
            {
                "repair_route": route,
                "task_count": len(items),
                "p1_count": sum(1 for item in items if item.get("repair_priority") == "P1"),
                "p2_count": sum(1 for item in items if item.get("repair_priority") == "P2"),
                "p3_count": sum(1 for item in items if item.get("repair_priority") == "P3"),
                "object_classes": "；".join(sorted(set(str(item.get("object_class") or "") for item in items))),
                "next_action": _route_next_action(route),
            }
        )
    return result


def _route_next_action(route: str) -> str:
    if route == "object_evidence_recall":
        return "进入对象召回工作台或答案盲审 GLM，补证据字段。"
    if route == "unit_rule_review":
        return "梳理单位规则和对象类别，不同对象不得互相匹配。"
    if route == "split_variant_review":
        return "按材料编号、规格、部位、平级/造型拆分候选。"
    if route == "feature_enrichment":
        return "用材料表、做法表、图例和节点补项目特征。"
    if route == "object_match_review":
        return "人工判定同一对象与否，不同对象退回缺候选。"
    return "人工复核。"


def _build_markdown(report: Mapping[str, Any]) -> str:
    summary = report.get("summary") or {}
    lines = [
        "# BIZ-2x PDF Three Field Defect Router",
        "",
        f"- generated_at: {report.get('generated_at', '-')}",
        f"- defect_count: {summary.get('defect_count', 0)}",
        f"- quantity_status: {summary.get('quantity_status', '-')}",
        f"- can_enable_quantity: {summary.get('can_enable_quantity', False)}",
        "",
        "## Routes",
        "",
        "| route | tasks | P1 | P2 | P3 | next action |",
        "| --- | ---: | ---: | ---: | ---: | --- |",
    ]
    for row in report.get("route_summary_rows") or []:
        lines.append(
            "| "
            + " | ".join(
                [
                    _md(row.get("repair_route")),
                    _md(row.get("task_count")),
                    _md(row.get("p1_count")),
                    _md(row.get("p2_count")),
                    _md(row.get("p3_count")),
                    _md(row.get("next_action")),
                ]
            )
            + " |"
        )
    lines.extend(
        [
            "",
            "## Top Defects",
            "",
            "| priority | route | status | target | candidate | focus |",
            "| --- | --- | --- | --- | --- | --- |",
        ]
    )
    for row in (report.get("defect_rows") or [])[:120]:
        lines.append(
            "| "
            + " | ".join(
                [
                    _md(row.get("repair_priority")),
                    _md(row.get("repair_route")),
                    _md(row.get("status")),
                    _md(row.get("answer_item_name")),
                    _md(row.get("candidate_item_name")),
                    _md(row.get("suggested_review_focus")),
                ]
            )
            + " |"
        )
    return "\n".join(lines) + "\n"


def _write_workbook(path: Path, report: Mapping[str, Any]) -> None:
    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "defect_summary"
    _append_rows(summary_sheet, [SUMMARY_HEADERS])
    _append_rows(summary_sheet, [[key, _cell_value(value)] for key, value in (report.get("summary") or {}).items()])
    _style_sheet(summary_sheet)

    route_sheet = workbook.create_sheet("route_summary")
    _append_rows(route_sheet, [ROUTE_HEADERS])
    _append_rows(route_sheet, [[_cell_value(row.get(header)) for header in ROUTE_HEADERS] for row in report.get("route_summary_rows") or []])
    _style_sheet(route_sheet)

    defect_sheet = workbook.create_sheet("defect_tasks")
    _append_rows(defect_sheet, [DEFECT_HEADERS])
    _append_rows(defect_sheet, [[_cell_value(row.get(header)) for header in DEFECT_HEADERS] for row in report.get("defect_rows") or []])
    _style_sheet(defect_sheet)
    workbook.save(path)


def _write_csv(path: Path, rows: Iterable[Mapping[str, Any]], fieldnames: Sequence[str]) -> None:
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        for row in rows:
            writer.writerow({field: _cell_value(row.get(field)) for field in fieldnames})


def _append_rows(sheet: Any, rows: Sequence[Sequence[Any]]) -> None:
    for row in rows:
        sheet.append(list(row))


def _style_sheet(sheet: Any) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    if sheet.max_row >= 1:
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for col_idx, column_cells in enumerate(sheet.columns, start=1):
        values = [str(cell.value or "") for cell in column_cells[:200]]
        width = min(max([len(value) for value in values] + [10]) + 2, 72)
        sheet.column_dimensions[get_column_letter(col_idx)].width = width


def _compact(value: Any) -> str:
    return re.sub(r"[\s,，。；;:：、\-_/\\()（）\[\]【】<>《》|]+", "", str(value or "").lower())


def _cell_value(value: Any) -> Any:
    if isinstance(value, (dict, list, tuple)):
        return json.dumps(value, ensure_ascii=False)
    return value


def _json_safe(value: Any) -> Any:
    if isinstance(value, Mapping):
        return {str(key): _json_safe(item) for key, item in value.items()}
    if isinstance(value, (list, tuple)):
        return [_json_safe(item) for item in value]
    if isinstance(value, Path):
        return str(value)
    return value


def _md(value: Any) -> str:
    return str(value or "").replace("|", "\\|").replace("\n", "<br>")
