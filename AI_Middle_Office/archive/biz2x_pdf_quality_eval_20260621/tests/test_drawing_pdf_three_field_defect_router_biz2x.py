from __future__ import annotations

import json
import sys
from pathlib import Path

from openpyxl import load_workbook

from app.services.drawing_pdf_three_field_defect_router import (
    build_three_field_defect_router_report,
    write_three_field_defect_router_outputs,
)
from scripts import biz2x_pdf_three_field_defect_router


def _review_report() -> dict[str, object]:
    return {
        "summary": {
            "answer_count": 4,
            "candidate_count": 3,
            "matched_three_fields_count": 1,
        },
        "review_rows": [
            {
                "review_no": 1,
                "status": "matched_three_fields",
                "answer_item_name": "石材地面 ST-01",
                "answer_unit": "㎡",
                "candidate_item_name": "石材地面 ST-01",
                "candidate_unit": "㎡",
            },
            {
                "review_no": 2,
                "status": "missing_candidate",
                "answer_sheet": "给排水",
                "answer_row_no": 10,
                "answer_section": "给排水",
                "answer_item_name": "水表供货及安装",
                "answer_feature": "水表 DN40，铜质",
                "answer_unit": "个",
            },
            {
                "review_no": 3,
                "status": "unit_conflict",
                "answer_sheet": "电气",
                "answer_row_no": 11,
                "answer_section": "配电",
                "answer_item_name": "配电箱AL",
                "answer_feature": "成套配电箱20KW，暗装",
                "answer_unit": "套",
                "candidate_row_no": 7,
                "candidate_item_name": "配电箱",
                "candidate_feature": "成套配电箱，暗装",
                "candidate_unit": "台",
                "name_score": 0.82,
                "feature_score": 0.7,
                "unit_score": 0,
            },
            {
                "review_no": 4,
                "status": "matched_name_unit_feature_review",
                "answer_sheet": "装修",
                "answer_row_no": 12,
                "answer_section": "天棚",
                "answer_item_name": "轻钢龙骨防水石膏板造型吊顶",
                "answer_feature": "U型50系列轻钢天棚龙骨 跌级；双层9.5mm防水石膏板",
                "answer_unit": "㎡",
                "candidate_row_no": 8,
                "candidate_item_name": "轻钢龙骨防水石膏板平级吊顶",
                "candidate_feature": "U型50系列轻钢天棚龙骨；双层9.5mm防水石膏板",
                "candidate_unit": "㎡",
                "issue": "项目名称存在平级/造型等关键列项冲突，需人工复核。",
            },
        ],
    }


def test_three_field_defect_router_classifies_core_routes():
    report = build_three_field_defect_router_report(_review_report())

    assert report["summary"]["defect_count"] == 3
    assert report["summary"]["matched_three_fields_count"] == 1
    assert report["summary"]["can_enable_quantity"] is False
    assert report["summary"]["route_counts"] == {
        "object_evidence_recall": 1,
        "unit_rule_review": 1,
        "split_variant_review": 1,
    }
    routes = {row["review_no"]: row for row in report["defect_rows"]}
    assert routes[2]["repair_priority"] == "P1"
    assert routes[2]["object_class"] == "fixture_valve_schedule"
    assert "补真实图纸证据" in routes[2]["repair_action"]
    assert routes[3]["repair_route"] == "unit_rule_review"
    assert "套/台" in routes[3]["suggested_unit_resolution"]
    assert routes[4]["repair_route"] == "split_variant_review"
    assert "平级/造型" in routes[4]["repair_action"]


def test_three_field_defect_router_writes_workbook(tmp_path: Path):
    report = build_three_field_defect_router_report(_review_report())

    outputs = write_three_field_defect_router_outputs(report, tmp_path / "out", stem="router")

    assert Path(outputs["json"]).exists()
    assert Path(outputs["csv"]).exists()
    assert Path(outputs["markdown"]).exists()
    assert Path(outputs["xlsx"]).exists()
    workbook = load_workbook(outputs["xlsx"])
    assert workbook.sheetnames == ["defect_summary", "route_summary", "defect_tasks"]
    assert workbook["defect_tasks"]["D2"].value == "object_evidence_recall"
    workbook.close()


def test_three_field_defect_router_cli(tmp_path: Path, monkeypatch, capsys):
    review_json = tmp_path / "review.json"
    review_json.write_text(json.dumps(_review_report(), ensure_ascii=False), encoding="utf-8")
    output_dir = tmp_path / "cli_out"
    monkeypatch.setattr(
        sys,
        "argv",
        [
            "biz2x_pdf_three_field_defect_router.py",
            "--three-field-review-json",
            str(review_json),
            "--output-dir",
            str(output_dir),
            "--timestamp",
            "fixed",
        ],
    )

    assert biz2x_pdf_three_field_defect_router.main() == 0
    payload = json.loads(capsys.readouterr().out)

    assert payload["ok"] is True
    assert payload["summary"]["defect_count"] == 3
    assert Path(payload["outputs"]["xlsx"]).exists()
