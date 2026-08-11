import hashlib
import json
from io import BytesIO
from types import SimpleNamespace

import pytest
import xlwt
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, PatternFill, Side

from app.services import budget_pricing_original_export as export_service
from app.services.budget_pricing import BudgetPricingError
from app.services.budget_pricing_original_export import (
    BudgetPricingOriginalExportSource,
    load_budget_pricing_original_source_content,
    render_budget_pricing_original_export,
)


def _source(*, line, header_row_index=1, source_filename="甲方清单.xlsx", source_file=None, source_sha256=""):
    return BudgetPricingOriginalExportSource(
        draft=SimpleNamespace(pricing_mode="enterprise_ai", revision=3),
        batch=SimpleNamespace(
            source_filename=source_filename,
            source_file_sha256=source_sha256,
            batch_uuid="batch-1",
        ),
        source_file=source_file,
        sheet_mappings=(SimpleNamespace(sheet_name="清单", header_row_index=header_row_index),),
        lines=(line,),
    )


def _line():
    return SimpleNamespace(
        source_sheet="清单",
        source_raw_row_index=2,
        item_name="轻钢龙骨吊顶",
        spec="600*600 矿棉板",
        calculation_quantity="12.5",
        effective_unit_price="88.8",
        line_total="1110",
        pricing_breakdown_json=json.dumps(
            {
                "material_supply_mode": "乙供",
                "labor_unit_cost": "20",
                "main_material_unit_cost": "40",
                "auxiliary_material_unit_cost": "8",
                "tax_amount": "6.12",
                "main_material_without_loss": "38",
                "loss_rate": "0.05",
                "machinery_unit_cost": "2",
                "comprehensive_unit_cost": "1",
                "management_unit_cost": "5",
                "profit_unit_cost": "4",
                "measure_unit_cost": "2",
                "owner_material_unit_price": "0",
                "owner_material_loss_amount": "0",
            }
        ),
        source_row_snapshot_json=json.dumps({"region": "一层", "raw_fields": {}}),
    )


def _workbook_bytes(headers):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "清单"
    for index, value in enumerate(headers, start=1):
        sheet.cell(row=1, column=index, value=value)
    sheet.cell(row=2, column=1, value="客户原项目")
    sheet.cell(row=2, column=2, value=12.5)
    sheet.cell(row=2, column=2).fill = PatternFill("solid", fgColor="FFF2CC")
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def _legacy_workbook_bytes(headers):
    workbook = xlwt.Workbook()
    sheet = workbook.add_sheet("清单")
    for column_index, value in enumerate(headers):
        sheet.write(0, column_index, value)
    sheet.write(1, 0, "客户原项目")
    sheet.write(1, 1, 12.5)
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def _headers(sheet):
    return [sheet.cell(row=1, column=index).value for index in range(1, sheet.max_column + 1)]


def test_local_source_fallback_reads_hash_matched_workbook(tmp_path, monkeypatch):
    content = _workbook_bytes(["项目名称", "工程量", "综合单价"])
    source_path = tmp_path / "甲方清单.xlsx"
    source_path.write_bytes(content)
    monkeypatch.setattr(
        export_service,
        "settings",
        SimpleNamespace(app_env="development", budget_pricing_local_source_root=str(tmp_path)),
    )
    source = _source(
        line=_line(),
        source_filename=source_path.name,
        source_sha256=hashlib.sha256(content).hexdigest(),
    )

    assert load_budget_pricing_original_source_content(source) == content


def test_local_source_fallback_rejects_hash_mismatch(tmp_path, monkeypatch):
    source_path = tmp_path / "甲方清单.xlsx"
    source_path.write_bytes(_workbook_bytes(["项目名称", "工程量"]))
    monkeypatch.setattr(
        export_service,
        "settings",
        SimpleNamespace(app_env="development", budget_pricing_local_source_root=str(tmp_path)),
    )
    source = _source(line=_line(), source_filename=source_path.name, source_sha256="0" * 64)

    with pytest.raises(BudgetPricingError) as exc_info:
        load_budget_pricing_original_source_content(source)

    assert exc_info.value.code == "BUDGET_PRICING_EXPORT_LOCAL_SOURCE_HASH_MISMATCH"


def test_local_source_fallback_is_disabled_in_production(tmp_path, monkeypatch):
    monkeypatch.setattr(
        export_service,
        "settings",
        SimpleNamespace(app_env="production", budget_pricing_local_source_root=str(tmp_path)),
    )
    source = _source(line=_line(), source_filename="甲方清单.xlsx", source_sha256="0" * 64)

    with pytest.raises(BudgetPricingError) as exc_info:
        load_budget_pricing_original_source_content(source)

    assert exc_info.value.code == "BUDGET_PRICING_EXPORT_SOURCE_FILE_NOT_RETAINED"


def test_export_appends_only_missing_primary_price_fields_without_moving_original_columns():
    result = render_budget_pricing_original_export(
        _workbook_bytes(["项目名称", "工程量"]),
        _source(line=_line()),
        project_name="测试项目",
    )

    workbook = load_workbook(BytesIO(result.content), data_only=False)
    sheet = workbook["清单"]
    headers = _headers(sheet)

    assert headers[:2] == ["项目名称", "工程量"]
    assert headers == ["项目名称", "工程量", "不含税综合单价", "不含税综合合价"]
    assert sheet.cell(row=2, column=1).value == "客户原项目"
    assert sheet.cell(row=2, column=2).value == 12.5
    assert sheet.cell(row=2, column=3).value == 88.8
    assert sheet.cell(row=2, column=4).value == 1110
    assert sheet.cell(row=2, column=3).fill.fgColor.rgb == sheet.cell(row=2, column=2).fill.fgColor.rgb
    assert "系统估算说明" in workbook.sheetnames


def test_export_fills_existing_unit_price_alias_instead_of_adding_a_duplicate_column():
    result = render_budget_pricing_original_export(
        _workbook_bytes(["项目名称", "工程量", "综合单价"]),
        _source(line=_line()),
        project_name="测试项目",
    )

    workbook = load_workbook(BytesIO(result.content), data_only=False)
    sheet = workbook["清单"]
    headers = _headers(sheet)

    assert headers.count("不含税综合单价") == 0
    assert sheet.cell(row=2, column=3).value == 88.8


def test_export_converts_legacy_xls_source_to_openxml_before_writing_prices():
    result = render_budget_pricing_original_export(
        _legacy_workbook_bytes(["项目名称", "工程量", "综合单价"]),
        _source(line=_line(), source_filename="甲方旧版清单.xls"),
        project_name="测试项目",
    )

    workbook = load_workbook(BytesIO(result.content), data_only=False)
    sheet = workbook["清单"]
    assert result.filename.endswith(".xlsx")
    assert sheet.cell(row=2, column=3).value == 88.8


def test_export_matches_source_sheet_when_original_title_has_trailing_whitespace():
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "清单 "
    sheet.append(["项目名称", "工程量", "综合单价"])
    sheet.append(["客户原项目", 12.5, None])
    output = BytesIO()
    workbook.save(output)

    result = render_budget_pricing_original_export(
        output.getvalue(),
        _source(line=_line()),
        project_name="测试项目",
    )

    rendered = load_workbook(BytesIO(result.content), data_only=False)
    assert rendered["清单 "]["C2"].value == 88.8
    assert not any(item["reason"] == "WORKSHEET_NOT_FOUND" for item in result.summary["unresolved"])


def test_export_preserves_existing_formats_and_fills_combined_cost_columns_without_extending_layout():
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "清单"
    sheet.append(["项目名称", "项目特征", "单位", "工程量", "人工费", "主材费", "辅材及机械费", "综合取费", "综合单价", "金额"])
    sheet.append(["客户原项目", "矿棉板吊顶", "㎡", 12.5, None, None, None, None, "=SUM(E2:H2)", "=I2*D2"])
    sheet["E2"].number_format = "General"
    sheet["I2"].number_format = "¥#,##0.000"
    sheet["J2"].number_format = "¥#,##0.000"
    output = BytesIO()
    workbook.save(output)
    line = _line()
    line.effective_unit_price = "82"
    line.line_total = "1025"

    result = render_budget_pricing_original_export(
        output.getvalue(),
        _source(line=line),
        project_name="测试项目",
    )

    rendered = load_workbook(BytesIO(result.content), data_only=False)
    sheet = rendered["清单"]
    assert sheet["E2"].value == 20
    assert sheet["F2"].value == 40
    assert sheet["G2"].value == 10
    assert sheet["H2"].value == 12
    assert sheet["I2"].value == 82
    assert sheet["J2"].value == 1025
    assert sheet["E2"].number_format == "General"
    assert sheet["I2"].number_format == "¥#,##0.000"
    assert sheet["J2"].number_format == "¥#,##0.000"
    assert sheet.max_column == 10
    assert result.summary["sheets"][0]["appended_fields"] == []


def test_export_preserves_two_level_header_styles_without_appending_optional_audit_columns():
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "清单"
    thin_black = Side(style="thin", color="000000")
    formal_border = Border(left=thin_black, right=thin_black, top=thin_black, bottom=thin_black)
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for row_index in range(1, 4):
        for column_index in range(1, 9):
            cell = sheet.cell(row=row_index, column=column_index)
            cell.border = formal_border
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            if row_index <= 2:
                cell.fill = header_fill

    sheet.merge_cells("A1:A2")
    sheet.merge_cells("B1:B2")
    sheet.merge_cells("C1:C2")
    sheet.merge_cells("D1:E1")
    sheet.merge_cells("F1:F2")
    sheet["A1"] = "项目名称"
    sheet["B1"] = "项目特征描述"
    sheet["C1"] = "工程量"
    sheet["D1"] = "不含税金额（元）"
    sheet["D2"] = "综合单价"
    sheet["E2"] = "综合合价"
    sheet["F1"] = "备注"
    sheet["A3"] = "客户原项目"
    sheet["B3"] = "矿棉板吊顶"
    sheet["C3"] = 12.5
    sheet["D3"] = 0
    sheet["D3"].number_format = "¥#,##0.000"
    sheet["E3"] = 0

    output = BytesIO()
    workbook.save(output)
    line = _line()
    line.source_raw_row_index = 3
    result = render_budget_pricing_original_export(
        output.getvalue(),
        _source(line=line, header_row_index=1),
        project_name="测试项目",
    )

    rendered = load_workbook(BytesIO(result.content), data_only=False)
    sheet = rendered["清单"]
    assert sheet.max_column == 8
    assert {str(item) for item in sheet.merged_cells.ranges} == {"A1:A2", "B1:B2", "C1:C2", "D1:E1", "F1:F2"}
    assert sheet["D1"].fill.fgColor.rgb == header_fill.fgColor.rgb
    assert sheet["D2"].border.left.style == formal_border.left.style
    assert sheet["D3"].value == 88.8
    assert sheet["E3"].value == 1110
    assert sheet["D3"].number_format == "¥#,##0.000"
    assert result.summary["sheets"][0]["appended_fields"] == []
