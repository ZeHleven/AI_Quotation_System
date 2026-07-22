from __future__ import annotations

import csv
import json
import re
from datetime import date, datetime
from io import BytesIO, StringIO
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string, get_column_letter


STANDARDIZATION_VERSION = "biz2l-standard-v0"
SUPPORTED_REQUIREMENT_EXTENSIONS = {".xlsx", ".xlsm"}
MAX_SCAN_ROWS = 30
MAX_ROWS_PER_SHEET = 800
MAX_COLUMNS_PER_SHEET = 300
EXCEL_ERROR_VALUES = {"#REF!", "#DIV/0!", "#VALUE!", "#NAME?", "#N/A", "#NUM!", "#NULL!"}


FIELD_ALIASES: dict[str, dict[str, tuple[str, ...]]] = {
    "item_name": {
        "strong": ("项目名称", "施工项目", "工作内容", "清单名称", "材料名称", "分部分项工程", "工程名称", "名称"),
        "weak": ("描述", "内容", "施工内容"),
    },
    "spec": {
        "strong": ("规格", "型号", "项目特征", "特征描述", "规格型号", "做法", "工艺"),
        "weak": ("描述", "备注", "说明"),
    },
    "quantity": {
        "strong": ("数量", "工程量", "数量/工程量", "工程数量", "预估数量", "暂定数量"),
        "weak": ("计量",),
    },
    "unit": {
        "strong": ("单位", "计量单位", "报价单位"),
        "weak": ("计价单位",),
    },
    "remark": {
        "strong": ("备注", "说明", "附注", "施工说明"),
        "weak": ("其他说明",),
    },
    "location": {
        "strong": ("区域", "房间", "楼层", "部位", "位置"),
        "weak": ("空间", "施工区域"),
    },
}

PRICE_ALIASES = (
    "单价",
    "综合单价",
    "合价",
    "金额",
    "总价",
    "税前价",
    "含税价",
    "备注价",
    "人工费",
    "管理费",
    "主材费",
    "辅材费",
    "材料费",
    "机械费",
    "利润",
    "税金",
    "规费",
)
HEADER_MATCH_PRIORITY = ("spec", "unit", "quantity", "remark", "location", "item_name")
CONTAINS_EXACT_ONLY_ALIASES = {"名称", "描述", "内容", "计量"}
SUMMARY_TERMS = ("合计", "小计", "总计", "汇总", "金额合计", "总金额")
NOTE_PREFIXES = ("备注", "说明", "注：", "注:", "施工说明", "要求")

UNIT_NORMALIZATION = {
    "M": "m",
    "米": "m",
    "延米": "m",
    "m2": "㎡",
    "M2": "㎡",
    "平方米": "㎡",
    "平米": "㎡",
    "m3": "m3",
    "M3": "m3",
    "立方米": "m3",
    "方": "m3",
    "点位": "点",
}

UNIT_FAMILIES = {
    "m": "length",
    "㎡": "area",
    "m3": "volume",
    "个": "count",
    "只": "count",
    "块": "count",
    "片": "count",
    "根": "count",
    "张": "count",
    "樘": "count",
    "点": "count",
    "套": "set",
    "组": "set",
    "项": "set",
}

UNIT_PATTERN = (
    r"㎡|m2|M2|平方米|平米|m3|M3|立方米|方|延米|米|m|M|点位|套|组|项|个|只|块|片|根|张|樘|点"
)
QUANTITY_UNIT_RE = re.compile(r"(?P<quantity>约?\d+(?:\.\d+)?)\s*(?P<unit>" + UNIT_PATTERN + r")")
RANGE_RE = re.compile(r"\d+(?:\.\d+)?\s*(?:-|~|至|到)\s*\d+(?:\.\d+)?")
NUMBER_RE = re.compile(r"\d+(?:\.\d+)?")
SPEC_TAIL_RE = re.compile(
    r"(?P<name>.+?)\s*(?P<spec>(?:\d+|一|二|两|三|四|五|六|七|八|九|十)遍|"
    r"\d+(?:\.\d+)?\s*(?:mm|MM|cm|CM|公分|厚|宽|高).*)$"
)


WARNING_MESSAGES = {
    "MISSING_ITEM_NAME": "缺少项目名称，不能自动进入报价。",
    "MISSING_QUANTITY": "缺少数量，不能自动进入报价。",
    "MISSING_UNIT": "缺少单位，需要人工确认。",
    "INVALID_QUANTITY": "数量为 0、负数或非法文本，不能自动进入报价。",
    "RANGE_QUANTITY": "数量为范围值，需要人工确认。",
    "APPROXIMATE_QUANTITY": "数量为约数，需要人工确认。",
    "MULTIPLE_NUMBERS": "同一行存在多个数字，需要确认哪个是工程量。",
    "POSSIBLE_SUMMARY_ROW": "疑似合计/小计行，默认不进入报价。",
    "POSSIBLE_NOTE_ROW": "疑似说明行，默认不进入报价。",
    "POSSIBLE_SECTION_ROW": "疑似区域/章节行，默认不进入报价。",
    "AMBIGUOUS_HEADER": "表头不清晰，需要人工确认列映射。",
    "MERGED_CELL_EXPANDED": "合并单元格已展开，需要人工核对。",
    "MULTI_SHEET_DETECTED": "检测到多 Sheet 文件，每个 Sheet 单独输出。",
    "PRICE_COLUMN_PRESENT": "原表包含价格列，系统不会采用该价格。",
    "LOW_CONFIDENCE": "低置信度，必须人工确认。",
    "MULTIPLE_QUANTITY_CANDIDATES": "同一行存在多个工程量候选，需人工确认采用哪一个数量。",
}

VALIDATION_ERROR_MESSAGES = {
    **WARNING_MESSAGES,
    "CONFIRMATION_REQUIRED": "该行存在需人工确认的风险，请先勾选人工确认。",
    "NOT_DATA_ROW": "该行不是有效清单数据行，可能是说明、汇总或空白行。",
}

FORCE_CONFIRMATION_WARNINGS = {
    "MISSING_ITEM_NAME",
    "MISSING_QUANTITY",
    "MISSING_UNIT",
    "INVALID_QUANTITY",
    "RANGE_QUANTITY",
    "APPROXIMATE_QUANTITY",
    "MULTIPLE_NUMBERS",
    "AMBIGUOUS_HEADER",
    "AMBIGUOUS_ROW_TYPE",
    "PRICE_COLUMN_PRESENT",
    "LOW_CONFIDENCE",
    "MULTIPLE_QUANTITY_CANDIDATES",
}

MANUAL_FIELD_VALUES = {
    "item_name",
    "spec",
    "quantity",
    "unit",
    "remark",
    "location",
    "ignore",
    "price_ignored",
}

CONFIRMED_EXPORT_FIELDS = [
    "source_sheet",
    "raw_row_index",
    "item_name",
    "spec",
    "quantity",
    "unit",
    "remark",
    "location",
    "confidence",
    "warnings",
    "raw_text",
]


class RequirementStandardizationError(ValueError):
    pass


def standardize_requirement_excel_path(path: str | Path) -> dict[str, Any]:
    excel_path = Path(path)
    return standardize_requirement_excel_bytes(excel_path.read_bytes(), filename=excel_path.name)


def standardize_requirement_excel_bytes(file_content: bytes, *, filename: str | None = None) -> dict[str, Any]:
    if not file_content:
        raise RequirementStandardizationError("Excel 需求单为空，请重新上传")

    suffix = Path(filename or "").suffix.lower()
    if suffix and suffix not in SUPPORTED_REQUIREMENT_EXTENSIONS:
        raise RequirementStandardizationError("BIZ-2l-1 只支持 .xlsx/.xlsm，旧 .xls 请另存后再解析")

    try:
        workbook = load_workbook(BytesIO(file_content), data_only=True, read_only=False)
    except Exception as exc:  # pragma: no cover - openpyxl raises several implementation-specific errors.
        raise RequirementStandardizationError(f"Excel 需求单读取失败，请确认文件为 .xlsx/.xlsm 格式: {exc}") from exc

    rows: list[dict[str, Any]] = []
    sheet_mappings: list[dict[str, Any]] = []
    issues: list[dict[str, Any]] = []
    standard_row_count = 0
    requires_confirmation_count = 0
    ignored_row_count = 0

    multi_sheet = len(workbook.worksheets) > 1
    for sheet in workbook.worksheets:
        parsed_sheet = _standardize_sheet(sheet, multi_sheet=multi_sheet)
        sheet_mappings.append(parsed_sheet["sheet_mapping"])
        rows.extend(parsed_sheet["rows"])
        issues.extend(parsed_sheet["issues"])
        standard_row_count += parsed_sheet["standard_row_count"]
        requires_confirmation_count += parsed_sheet["requires_confirmation_count"]
        ignored_row_count += parsed_sheet["ignored_row_count"]

    return {
        "version": STANDARDIZATION_VERSION,
        "source": {
            "file_name": filename or "uploaded.xlsx",
            "file_type": suffix.lstrip(".") if suffix else "",
        },
        "summary": {
            "sheet_count": len(workbook.worksheets),
            "standard_row_count": standard_row_count,
            "requires_confirmation_count": requires_confirmation_count,
            "ignored_row_count": ignored_row_count,
            "total_output_row_count": len(rows),
        },
        "field_mapping": sheet_mappings[0]["field_mapping"] if sheet_mappings else {},
        "sheet_mappings": sheet_mappings,
        "rows": rows,
        "issues": issues,
    }


def build_standardization_markdown(result: dict[str, Any]) -> str:
    summary = result.get("summary", {})
    source = result.get("source", {})
    lines = [
        "# BIZ-2l-1 需求单标准化预览",
        "",
        f"- 文件：{source.get('file_name', '')}",
        f"- Sheet 数：{summary.get('sheet_count', 0)}",
        f"- 标准行：{summary.get('standard_row_count', 0)}",
        f"- 需确认：{summary.get('requires_confirmation_count', 0)}",
        f"- 忽略行：{summary.get('ignored_row_count', 0)}",
        "",
        "| Sheet | 行号 | 行类型 | 项目名称 | 规格 | 数量 | 单位 | 置信度 | 需确认 | 警告 |",
        "|---|---:|---|---|---|---:|---|---|---|---|",
    ]
    for row in result.get("rows", []):
        lines.append(
            "| {sheet} | {row_index} | {row_type} | {item_name} | {spec} | {quantity} | {unit} | {confidence} | {confirm} | {warnings} |".format(
                sheet=_md_cell(row.get("source_sheet")),
                row_index=row.get("raw_row_index", ""),
                row_type=_md_cell(row.get("row_type")),
                item_name=_md_cell(row.get("item_name")),
                spec=_md_cell(row.get("spec")),
                quantity=_md_cell(row.get("quantity")),
                unit=_md_cell(row.get("unit")),
                confidence=_md_cell(row.get("confidence")),
                confirm="是" if row.get("requires_confirmation") else "否",
                warnings=_md_cell(",".join(row.get("warnings", []))),
            )
        )
    if result.get("issues"):
        lines.extend(["", "## 问题提示", ""])
        for issue in result["issues"]:
            lines.append(
                f"- {issue.get('source_sheet')} 第 {issue.get('raw_row_index')} 行：{issue.get('code')} - {issue.get('message')}"
            )
    return "\n".join(lines) + "\n"


def build_standardization_csv(result: dict[str, Any]) -> str:
    output = StringIO()
    fieldnames = [
        "source_sheet",
        "raw_row_index",
        "row_type",
        "item_name",
        "spec",
        "quantity",
        "unit",
        "remark",
        "location",
        "confidence",
        "requires_confirmation",
        "warnings",
        "raw_text",
    ]
    writer = csv.DictWriter(output, fieldnames=fieldnames, extrasaction="ignore")
    writer.writeheader()
    for row in result.get("rows", []):
        copy = dict(row)
        copy["warnings"] = ",".join(copy.get("warnings") or [])
        writer.writerow(copy)
    return output.getvalue()


def write_standardization_outputs(result: dict[str, Any], output_dir: str | Path, *, stem: str) -> dict[str, str]:
    output_path = Path(output_dir)
    output_path.mkdir(parents=True, exist_ok=True)
    safe_stem = re.sub(r"[^0-9A-Za-z_.-]+", "_", stem).strip("_") or "requirement_standardization"
    paths = {
        "json": output_path / f"{safe_stem}.json",
        "csv": output_path / f"{safe_stem}.csv",
        "markdown": output_path / f"{safe_stem}.md",
    }
    paths["json"].write_text(json.dumps(result, ensure_ascii=False, indent=2), encoding="utf-8")
    paths["csv"].write_text(build_standardization_csv(result), encoding="utf-8-sig")
    paths["markdown"].write_text(build_standardization_markdown(result), encoding="utf-8")
    return {key: str(path) for key, path in paths.items()}


def apply_manual_field_mappings(result: dict[str, Any], mappings: Any) -> dict[str, Any]:
    """Rebuild preview rows from in-memory raw cells with user-selected columns."""
    mapping_by_sheet = _normalize_manual_mappings(mappings)
    if not mapping_by_sheet:
        return result

    sheet_rows: dict[str, list[dict[str, Any]]] = {}
    for row in result.get("rows", []):
        sheet_rows.setdefault(str(row.get("source_sheet") or ""), []).append(row)

    original_sheet_mappings = result.get("sheet_mappings") or []
    original_by_sheet = {str(item.get("sheet_name") or ""): item for item in original_sheet_mappings}
    all_sheet_names = list(original_by_sheet) or list(sheet_rows)
    multi_sheet = (result.get("summary", {}).get("sheet_count") or len(all_sheet_names)) > 1

    rebuilt_rows: list[dict[str, Any]] = []
    rebuilt_sheet_mappings: list[dict[str, Any]] = []
    issues: list[dict[str, Any]] = []

    for sheet_name in all_sheet_names:
        source_mapping = original_by_sheet.get(sheet_name, {})
        field_mapping = mapping_by_sheet.get(sheet_name) or source_mapping.get("field_mapping", {})
        field_mapping = _clean_manual_field_mapping(field_mapping)
        header_context = _header_context_from_field_mapping(
            field_mapping,
            header_row_index=source_mapping.get("header_row_index"),
        )
        sheet_output_rows: list[dict[str, Any]] = []
        current_section = ""
        for original_row in sheet_rows.get(sheet_name, []):
            row = _row_from_preview_row(original_row, field_mapping)
            standardized = _standardize_row(
                sheet_name,
                row,
                header_context=header_context,
                current_section=current_section,
                multi_sheet=multi_sheet,
            )
            if standardized["row_type"] == "section_row":
                current_section = standardized.get("raw_text") or current_section
            sheet_output_rows.append(standardized)
            issues.extend(_issues_for_row(standardized))
        rebuilt_rows.extend(sheet_output_rows)
        rebuilt_sheet_mappings.append(
            {
                "sheet_name": sheet_name,
                "header_row_index": source_mapping.get("header_row_index"),
                "field_mapping": field_mapping,
                "columns": _columns_from_preview_rows(sheet_output_rows, field_mapping),
            }
        )

    summary = _summary_for_rows(rebuilt_rows, result.get("summary", {}))
    return {
        **result,
        "summary": summary,
        "field_mapping": rebuilt_sheet_mappings[0]["field_mapping"] if rebuilt_sheet_mappings else {},
        "sheet_mappings": rebuilt_sheet_mappings,
        "rows": rebuilt_rows,
        "issues": issues,
    }


def confirm_standardized_rows(rows: list[dict[str, Any]]) -> dict[str, Any]:
    confirmed_rows: list[dict[str, Any]] = []
    blocked_rows: list[dict[str, Any]] = []

    for row in rows:
        if not (row.get("include") or row.get("selected")):
            continue
        cleaned, errors = _clean_confirmed_row(row)
        if errors:
            error_messages = _validation_error_messages(errors, row)
            blocked_rows.append(
                {
                    "requirement_row_key": row.get("requirement_row_key"),
                    "source_sheet": row.get("source_sheet"),
                    "raw_row_index": row.get("raw_row_index"),
                    "item_name": row.get("item_name"),
                    "spec": row.get("spec"),
                    "quantity": row.get("quantity"),
                    "unit": row.get("unit"),
                    "remark": row.get("remark"),
                    "raw_text": row.get("raw_text"),
                    "raw_cells": row.get("raw_cells") or [],
                    "warnings": row.get("warnings") or [],
                    "errors": errors,
                    "error_messages": error_messages,
                    "error_summary": "；".join(error_messages),
                }
            )
            continue
        confirmed_rows.append(cleaned)

    return {
        "summary": {
            "selected_row_count": len(confirmed_rows) + len(blocked_rows),
            "confirmed_row_count": len(confirmed_rows),
            "blocked_row_count": len(blocked_rows),
        },
        "rows": confirmed_rows,
        "blocked_rows": blocked_rows,
        "csv": build_confirmed_standardization_csv(confirmed_rows),
        "quote_text": build_confirmed_standardization_text(confirmed_rows),
    }


def build_confirmed_standardization_csv(rows: list[dict[str, Any]]) -> str:
    output = StringIO()
    writer = csv.DictWriter(output, fieldnames=CONFIRMED_EXPORT_FIELDS, extrasaction="ignore")
    writer.writeheader()
    for row in rows:
        copy = dict(row)
        copy["warnings"] = ",".join(copy.get("warnings") or [])
        writer.writerow(copy)
    return output.getvalue()


def build_confirmed_standardization_text(rows: list[dict[str, Any]]) -> str:
    lines: list[str] = []
    for index, row in enumerate(rows, start=1):
        parts = [f"{index}. {row.get('item_name', '')}"]
        if row.get("spec"):
            parts.append(f"规格/特征: {row['spec']}")
        quantity = row.get("quantity", "")
        unit = row.get("unit", "")
        parts.append(f"数量: {quantity}{unit}")
        if row.get("remark"):
            parts.append(f"备注: {row['remark']}")
        if row.get("source_sheet") or row.get("raw_row_index"):
            parts.append(f"来源: {row.get('source_sheet', '')} 第{row.get('raw_row_index', '')}行")
        lines.append("; ".join(parts))
    return "\n".join(lines)


def _validation_error_message(code: str) -> str:
    return VALIDATION_ERROR_MESSAGES.get(code) or code or "未通过确认校验。"


def _validation_error_messages(errors: list[str], row: dict[str, Any]) -> list[str]:
    messages: list[str] = []
    for error in errors:
        if error == "CONFIRMATION_REQUIRED":
            messages.extend(_confirmation_required_messages(row, errors))
        else:
            messages.append(_validation_error_message(error))
    return _unique(messages) or [_validation_error_message("CONFIRMATION_REQUIRED")]


def _confirmation_required_messages(row: dict[str, Any], errors: list[str]) -> list[str]:
    warnings = [str(code) for code in (row.get("warnings") or []) if code]
    details: list[str] = []
    for code in warnings:
        if code in {"MULTI_SHEET_DETECTED"}:
            continue
        if code not in FORCE_CONFIRMATION_WARNINGS:
            continue
        if code in errors:
            continue
        details.append(WARNING_MESSAGES.get(code, code))
    if row.get("confidence") == "low" and "LOW_CONFIDENCE" not in warnings:
        details.append(WARNING_MESSAGES["LOW_CONFIDENCE"])
    if bool(row.get("requires_confirmation")) and not details:
        details.append("该行被标记为需人工确认，请核对项目名称、数量、单位、风险提示和原始行内容。")
    return details


def _reference_sheet_role(sheet_name: str) -> str:
    normalized = _normalize_compact(sheet_name)
    if any(marker in normalized for marker in ("\u7f16\u5236\u8bf4\u660e", "\u62a5\u4ef7\u8bf4\u660e", "\u8bf4\u660e")):
        return "metadata"
    if any(marker in normalized for marker in ("\u8ba1\u7b97\u89c4\u5219", "\u8ba1\u91cf\u89c4\u5219", "\u5de5\u7a0b\u91cf\u8ba1\u7b97\u89c4\u5219")):
        return "calculation_rule"
    if any(marker in normalized for marker in ("\u4e3b\u6750\u54c1\u724c", "\u54c1\u724c\u8868", "\u6750\u6599\u54c1\u724c")):
        return "material_reference"
    if "\u635f\u8017" in normalized:
        return "loss_reference"
    if any(marker in normalized for marker in ("\u5907\u7528\u6e05\u5355", "\u6682\u5217\u6e05\u5355")):
        return "optional_backup"
    if any(marker in normalized for marker in ("\u6c47\u603b", "\u6c47\u603b\u8868", "\u62a5\u4ef7\u6c47\u603b")):
        return "summary_analysis"
    return ""


def _as_reference_context_row(row: dict[str, Any], sheet_role: str) -> dict[str, Any]:
    converted = {**row, "sheet_role": sheet_role}
    raw_text = _clean_text(converted.get("raw_text"))
    if not raw_text:
        converted.update(
            {
                "row_type": "empty_row",
                "item_name": "",
                "spec": "",
                "quantity": None,
                "unit": "",
                "remark": "",
                "requires_confirmation": False,
            }
        )
        return converted
    converted.update(
        {
            "row_type": "reference_row",
            "item_name": raw_text[:255],
            "spec": "",
            "quantity": None,
            "quantity_source": {},
            "quantity_candidates": [],
            "unit": "",
            "unit_raw": "",
            "unit_family": "",
            "remark": raw_text,
            "location": "",
            "work_area": "",
            "field_mapping": {},
            "normalized_name": "",
            "requires_confirmation": False,
            "warnings": _unique([*(converted.get("warnings") or []), "REFERENCE_CONTEXT_ROW"]),
        }
    )
    return converted


def _standardize_sheet(sheet, *, multi_sheet: bool) -> dict[str, Any]:
    row_data = _read_sheet_rows(sheet)
    header_context = _detect_header(row_data)
    field_mapping = _field_mapping_for_context(header_context)
    sheet_role = _reference_sheet_role(sheet.title)
    sheet_issues: list[dict[str, Any]] = []
    output_rows: list[dict[str, Any]] = []
    current_section = ""
    standard_row_count = 0
    requires_confirmation_count = 0
    ignored_row_count = 0

    for row in row_data:
        standardized = _standardize_row(
            sheet.title,
            row,
            header_context=header_context,
            current_section=current_section,
            multi_sheet=multi_sheet,
        )
        if sheet_role:
            standardized = _as_reference_context_row(standardized, sheet_role)
        elif sheet_role == "":
            standardized["sheet_role"] = "bill"
        if standardized["row_type"] == "section_row":
            current_section = standardized.get("raw_text") or current_section
        if standardized["row_type"] == "data_row":
            standard_row_count += 1
        else:
            ignored_row_count += 1
        if standardized["requires_confirmation"]:
            requires_confirmation_count += 1
        output_rows.append(standardized)
        sheet_issues.extend(_issues_for_row(standardized))

    return {
        "sheet_mapping": {
            "sheet_name": sheet.title,
            "sheet_role": sheet_role or "bill",
            "header_row_index": header_context["header_row_index"],
            "field_mapping": field_mapping,
            "columns": _sheet_columns(row_data, header_context),
        },
        "rows": output_rows,
        "issues": sheet_issues,
        "standard_row_count": standard_row_count,
        "requires_confirmation_count": requires_confirmation_count,
        "ignored_row_count": ignored_row_count,
    }


def _read_sheet_rows(sheet) -> list[dict[str, Any]]:
    max_row = min(sheet.max_row or 0, MAX_ROWS_PER_SHEET)
    max_col = min(sheet.max_column or 0, MAX_COLUMNS_PER_SHEET)
    merged_values: dict[tuple[int, int], str] = {}
    merged_sources: set[tuple[int, int]] = set()
    for merged_range in sheet.merged_cells.ranges:
        top_left = sheet.cell(merged_range.min_row, merged_range.min_col)
        top_left_text = _clean_text(top_left.value)
        if not top_left_text:
            continue
        for row_index in range(merged_range.min_row, merged_range.max_row + 1):
            for col_index in range(merged_range.min_col, merged_range.max_col + 1):
                if row_index == merged_range.min_row and col_index == merged_range.min_col:
                    continue
                merged_values[(row_index, col_index)] = top_left_text
                merged_sources.add((row_index, col_index))

    rows: list[dict[str, Any]] = []
    for row_index in range(1, max_row + 1):
        values: list[str] = []
        merged_expanded = False
        for col_index in range(1, max_col + 1):
            value = _clean_text(sheet.cell(row_index, col_index).value)
            if not value and (row_index, col_index) in merged_values:
                value = merged_values[(row_index, col_index)]
                merged_expanded = True
            if (row_index, col_index) in merged_sources:
                merged_expanded = True
            values.append(value)
        rows.append({"index": row_index, "values": values, "merged_expanded": merged_expanded})
    return rows


def _detect_header(rows: list[dict[str, Any]]) -> dict[str, Any]:
    best: dict[str, Any] = {
        "header_row_index": None,
        "column_map": {},
        "headers": {},
        "all_headers": {},
        "header_groups": {},
        "strengths": {},
        "score": 0,
    }
    for pos, row in enumerate(rows[:MAX_SCAN_ROWS]):
        previous_values = rows[pos - 1]["values"] if pos > 0 else []
        labels = _header_labels(row["values"], previous_values)
        column_map: dict[str, int] = {}
        headers: dict[int, str] = {}
        strengths: dict[str, str] = {}
        score = 0
        for index, label in enumerate(labels):
            kind, strength = _classify_header(label)
            if not kind:
                continue
            if kind == "price":
                headers[index] = label
                column_map.setdefault("price", index)
                continue
            if kind in column_map and strengths.get(kind) == "strong":
                continue
            column_map[kind] = index
            headers[index] = label
            strengths[kind] = strength
            score += {"item_name": 5, "quantity": 4, "unit": 4, "spec": 2, "remark": 1, "location": 1}.get(kind, 1)
        if "item_name" in column_map and ("quantity" in column_map or "unit" in column_map) and score > best["score"]:
            best = {
                "header_row_index": row["index"],
                "column_map": column_map,
                "headers": headers,
                "all_headers": {index: label for index, label in enumerate(labels) if label},
                "header_groups": {index: value for index, value in enumerate(previous_values) if value},
                "strengths": strengths,
                "score": score,
            }
    return best


def _header_labels(values: list[str], previous_values: list[str]) -> list[str]:
    labels: list[str] = []
    for index, value in enumerate(values):
        previous = previous_values[index] if index < len(previous_values) else ""
        if value and previous and not _classify_header(value)[0] and not _classify_header(previous)[0]:
            labels.append(f"{previous}{value}")
        else:
            labels.append(value)
    return labels


def _field_mapping_for_context(header_context: dict[str, Any]) -> dict[str, str]:
    mapping: dict[str, str] = {}
    for field, index in header_context.get("column_map", {}).items():
        if field == "price":
            mapping[get_column_letter(index + 1)] = "price_ignored"
        else:
            mapping[get_column_letter(index + 1)] = field
    return mapping


def _sheet_columns(rows: list[dict[str, Any]], header_context: dict[str, Any]) -> list[dict[str, Any]]:
    field_mapping = _field_mapping_for_context(header_context)
    header_labels = header_context.get("all_headers") or header_context.get("headers", {})
    max_col = max((len(row.get("values", [])) for row in rows), default=0)
    header_row_index = header_context.get("header_row_index")
    columns: list[dict[str, Any]] = []
    for index in range(max_col):
        column = get_column_letter(index + 1)
        samples: list[str] = []
        for row in rows:
            if header_row_index and row.get("index") <= header_row_index:
                continue
            value = _value_at(row.get("values", []), index)
            if value and value not in samples:
                samples.append(value)
            if len(samples) >= 5:
                break
        detected_field = field_mapping.get(column, "ignore")
        label = header_labels.get(index) or ""
        if not _should_show_mapping_column(column, label, detected_field, samples, has_header=bool(header_row_index)):
            continue
        columns.append(
            {
                "column": column,
                "index": index,
                "label": label,
                "detected_field": detected_field,
                "is_price": detected_field == "price_ignored",
                "sample_values": samples,
            }
        )
    return columns


def _columns_from_preview_rows(rows: list[dict[str, Any]], field_mapping: dict[str, str]) -> list[dict[str, Any]]:
    labels: dict[str, str] = {}
    samples: dict[str, list[str]] = {}
    max_index = 0
    for row in rows:
        for cell in row.get("raw_cells") or []:
            column = str(cell.get("column") or "")
            if not column:
                continue
            value = _clean_text(cell.get("value"))
            try:
                max_index = max(max_index, column_index_from_string(column))
            except ValueError:
                continue
            if row.get("row_type") == "header_row" and value and column not in labels:
                labels[column] = value
            if row.get("row_type") == "header_row":
                continue
            if value:
                bucket = samples.setdefault(column, [])
                if value not in bucket and len(bucket) < 5:
                    bucket.append(value)
    columns: list[dict[str, Any]] = []
    for index in range(1, max_index + 1):
        column = get_column_letter(index)
        detected_field = field_mapping.get(column, "ignore")
        label = labels.get(column, "")
        sample_values = samples.get(column, [])
        if not _should_show_mapping_column(column, label, detected_field, sample_values, has_header=bool(labels)):
            continue
        columns.append(
            {
                "column": column,
                "index": index - 1,
                "label": label,
                "detected_field": detected_field,
                "is_price": detected_field == "price_ignored",
                "sample_values": sample_values,
            }
        )
    return columns


def _first_non_empty_column_value(rows: list[dict[str, Any]], index: int) -> str:
    for row in rows[:MAX_SCAN_ROWS]:
        value = _value_at(row.get("values", []), index)
        if value:
            return value
    return ""


def _standardize_row(
    sheet_name: str,
    row: dict[str, Any],
    *,
    header_context: dict[str, Any],
    current_section: str,
    multi_sheet: bool,
) -> dict[str, Any]:
    values = row["values"]
    raw_text = _join_values(values)
    header_row_index = header_context.get("header_row_index")
    column_map = header_context.get("column_map", {})
    header_labels = header_context.get("headers", {})
    warnings: list[str] = []

    if row["merged_expanded"]:
        warnings.append("MERGED_CELL_EXPANDED")
    if multi_sheet:
        warnings.append("MULTI_SHEET_DETECTED")

    if header_row_index == row["index"]:
        return _base_row(
            sheet_name,
            row,
            row_type="header_row",
            raw_text=raw_text,
            raw_fields=_raw_fields(values, header_labels),
            warnings=warnings,
            confidence="high",
            requires_confirmation=False,
        )

    row_type = _classify_non_data_row(values, raw_text)
    if row_type in {"section_row", "ambiguous_row"} and _has_mapped_data_signal(values, column_map, header_row_index, row["index"]):
        row_type = "data_row"
    if row_type != "data_row":
        row_warnings = list(warnings)
        if row_type == "summary_row":
            row_warnings.append("POSSIBLE_SUMMARY_ROW")
        elif row_type == "note_row":
            row_warnings.append("POSSIBLE_NOTE_ROW")
        elif row_type == "section_row":
            row_warnings.append("POSSIBLE_SECTION_ROW")
        elif row_type == "ambiguous_row":
            row_warnings.extend(["AMBIGUOUS_ROW_TYPE", "LOW_CONFIDENCE"])
        return _base_row(
            sheet_name,
            row,
            row_type=row_type,
            raw_text=raw_text,
            raw_fields=_raw_fields(values, header_labels),
            warnings=_unique(row_warnings),
            confidence="low" if row_type == "ambiguous_row" else "medium",
            requires_confirmation=row_type == "ambiguous_row",
        )

    extraction = _extract_data_fields(values, raw_text, header_context, current_section)
    warnings.extend(extraction["warnings"])
    confidence = _confidence_for(extraction, header_context, warnings)
    if confidence == "low":
        warnings.append("LOW_CONFIDENCE")
    warnings = _unique(warnings)
    requires_confirmation = confidence == "low" or any(code in FORCE_CONFIRMATION_WARNINGS for code in warnings)

    return {
        **_base_row(
            sheet_name,
            row,
            row_type="data_row",
            raw_text=raw_text,
            raw_fields=_raw_fields(values, header_labels),
            warnings=warnings,
            confidence=confidence,
            requires_confirmation=requires_confirmation,
        ),
        "item_name": extraction["item_name"],
        "spec": extraction["spec"],
        "quantity": extraction["quantity"],
        "quantity_source": extraction["quantity_source"],
        "quantity_candidates": extraction["quantity_candidates"],
        "unit": extraction["unit"],
        "unit_raw": extraction["unit_raw"],
        "unit_family": extraction["unit_family"],
        "remark": extraction["remark"],
        "location": extraction["location"],
        "work_area": extraction["location"],
        "field_mapping": _field_mapping_for_context(header_context),
        "normalized_name": _normalize_name(extraction["item_name"]),
        "cost_candidates": [],
    }


def _base_row(
    sheet_name: str,
    row: dict[str, Any],
    *,
    row_type: str,
    raw_text: str,
    raw_fields: dict[str, str],
    warnings: list[str],
    confidence: str,
    requires_confirmation: bool,
) -> dict[str, Any]:
    return {
        "source_sheet": sheet_name,
        "raw_row_index": row["index"],
        "row_type": row_type,
        "item_name": "",
        "spec": "",
        "quantity": None,
        "quantity_source": {},
        "quantity_candidates": [],
        "unit": "",
        "unit_raw": "",
        "unit_family": "",
        "remark": "",
        "location": "",
        "work_area": "",
        "raw_fields": raw_fields,
        "raw_cells": _raw_cells(row.get("values", [])),
        "raw_text": raw_text,
        "field_mapping": {},
        "normalized_name": "",
        "confidence": confidence,
        "warnings": _unique(warnings),
        "requires_confirmation": requires_confirmation,
        "cost_candidates": [],
    }


def _classify_non_data_row(values: list[str], raw_text: str) -> str:
    non_empty = [value for value in values if value]
    if not non_empty:
        return "empty_row"
    if _looks_like_noise_row(non_empty):
        return "empty_row"
    compact = _normalize_compact(raw_text)
    if any(_normalize_compact(term) in compact for term in SUMMARY_TERMS):
        return "summary_row"
    if len(non_empty) == 1:
        only = non_empty[0]
        if _looks_like_note(only):
            return "note_row"
        if _looks_like_section(only):
            return "section_row"
        if QUANTITY_UNIT_RE.search(only):
            return "data_row"
        return "section_row" if len(only) <= 16 else "ambiguous_row"
    if _looks_like_note(non_empty[0]) and not any(QUANTITY_UNIT_RE.search(value) for value in non_empty):
        return "note_row"
    if _row_has_data_signal(non_empty):
        return "data_row"
    return "ambiguous_row"


def _row_has_data_signal(values: list[str]) -> bool:
    joined = " ".join(values)
    if QUANTITY_UNIT_RE.search(joined):
        return True
    has_number = bool(NUMBER_RE.search(joined))
    has_unit = any(_normalize_unit(value) for value in values)
    return has_number and has_unit


def _has_mapped_data_signal(
    values: list[str],
    column_map: dict[str, int],
    header_row_index: int | None,
    row_index: int,
) -> bool:
    if header_row_index is None or row_index <= header_row_index:
        return False
    item_name = _value_at(values, column_map.get("item_name"))
    quantity = _value_at(values, column_map.get("quantity"))
    unit = _value_at(values, column_map.get("unit"))
    spec = _value_at(values, column_map.get("spec"))
    if not _looks_like_item_text(item_name):
        return False
    quantity_value, _ = _parse_quantity(quantity)
    return bool(quantity_value is not None or _normalize_unit(unit) or _looks_like_item_text(spec))


def _extract_data_fields(
    values: list[str],
    raw_text: str,
    header_context: dict[str, Any],
    current_section: str,
) -> dict[str, Any]:
    column_map = header_context.get("column_map", {})
    header_labels = header_context.get("headers", {})
    warnings: list[str] = []
    item_name = _value_at(values, column_map.get("item_name"))
    spec = _value_at(values, column_map.get("spec"))
    selected_quantity_index = column_map.get("quantity")
    quantity_raw = _value_at(values, column_map.get("quantity"))
    unit_raw = _value_at(values, column_map.get("unit"))
    remark = _value_at(values, column_map.get("remark"))
    location = _value_at(values, column_map.get("location")) or current_section
    price_value = _value_at(values, column_map.get("price"))
    if price_value:
        warnings.append("PRICE_COLUMN_PRESENT")

    if "item_name" not in column_map and "quantity" not in column_map and "unit" not in column_map:
        warnings.append("AMBIGUOUS_HEADER")
        inferred = _infer_from_text(raw_text)
        item_name = inferred["item_name"]
        if not spec and "spec" not in column_map:
            spec = inferred["spec"]
        quantity_raw = quantity_raw or inferred["quantity_raw"]
        unit_raw = unit_raw or inferred["unit_raw"]
        warnings.extend(inferred["warnings"])
    else:
        if not quantity_raw or not unit_raw:
            inferred = _infer_from_text(" ".join(value for value in (item_name, spec, remark, raw_text) if value))
            if "quantity" not in column_map and not quantity_raw:
                quantity_raw = inferred["quantity_raw"]
            if "unit" not in column_map and not unit_raw:
                unit_raw = inferred["unit_raw"]
            if item_name and inferred["item_name"] and item_name == raw_text:
                item_name = inferred["item_name"]
            if not spec and "spec" not in column_map:
                spec = inferred["spec"]
            if "quantity" not in column_map:
                warnings.extend(inferred["warnings"])

    if not item_name and raw_text:
        item_name = _first_text_cell(values, skip_indexes={column_map.get("quantity"), column_map.get("unit"), column_map.get("price")})

    quantity, quantity_warnings = _parse_quantity(quantity_raw)
    warnings.extend(quantity_warnings)
    quantity_candidates = _quantity_candidates(values, header_context, selected_quantity_index=selected_quantity_index)
    quantity_source = _quantity_source_for(
        header_context,
        selected_quantity_index=selected_quantity_index,
        quantity_raw=quantity_raw,
        quantity=quantity,
        quantity_candidates=quantity_candidates,
    )
    if len(quantity_candidates) > 1:
        warnings.append("MULTIPLE_QUANTITY_CANDIDATES")
    unit = _normalize_unit(unit_raw)
    if not item_name:
        warnings.append("MISSING_ITEM_NAME")
    if quantity is None:
        warnings.append("MISSING_QUANTITY")
    if not unit:
        warnings.append("MISSING_UNIT")

    if _needs_multiple_number_review(raw_text, selected_quantity_index, quantity_candidates):
        warnings.append("MULTIPLE_NUMBERS")

    # An independently mapped specification column is authoritative.  Splitting
    # digits out of the project name as an inferred spec in that case corrupts
    # legitimate names such as ``新建120-200mm轻质砖墙``.
    if item_name and not spec and "spec" not in column_map:
        split = _split_name_spec(item_name)
        item_name = split["item_name"]
        if split["spec"]:
            spec = split["spec"]

    return {
        "item_name": item_name,
        "spec": spec,
        "quantity": quantity,
        "quantity_source": quantity_source,
        "quantity_candidates": quantity_candidates,
        "unit": unit,
        "unit_raw": unit_raw,
        "unit_family": UNIT_FAMILIES.get(unit, ""),
        "remark": remark,
        "location": location,
        "warnings": _unique(warnings),
        "has_header": bool(header_labels),
    }


def _infer_from_text(text: str) -> dict[str, Any]:
    warnings: list[str] = []
    match = None
    for candidate in QUANTITY_UNIT_RE.finditer(text):
        match = candidate
    if not match:
        return {"item_name": text.strip(), "spec": "", "quantity_raw": "", "unit_raw": "", "warnings": warnings}

    before = text[: match.start()].strip(" ,，;；-")
    after = text[match.end() :].strip(" ,，;；-")
    split = _split_name_spec(before)
    if after:
        warnings.append("MULTIPLE_NUMBERS") if NUMBER_RE.search(after) else None
    quantity_raw = match.group("quantity")
    unit_raw = match.group("unit")
    return {
        "item_name": split["item_name"],
        "spec": split["spec"] or after,
        "quantity_raw": quantity_raw,
        "unit_raw": unit_raw,
        "warnings": warnings,
    }


def _quantity_candidates(
    values: list[str],
    header_context: dict[str, Any],
    *,
    selected_quantity_index: int | None,
) -> list[dict[str, Any]]:
    headers = header_context.get("all_headers") or header_context.get("headers") or {}
    groups = header_context.get("header_groups") or {}
    column_map = header_context.get("column_map") or {}
    price_index = column_map.get("price")
    candidates: list[dict[str, Any]] = []

    for index, raw_value in enumerate(values):
        raw_text = _clean_text(raw_value)
        if not raw_text:
            continue
        quantity, quantity_warnings = _parse_quantity(raw_text)
        if quantity is None:
            continue

        label = _clean_text(headers.get(index))
        group_label = _clean_text(groups.get(index))
        label_kind = _classify_header(label)[0]
        group_kind = _classify_header(group_label)[0]
        is_selected = index == selected_quantity_index
        is_quantity_column = label_kind == "quantity" or group_kind == "quantity"
        is_price_column = index == price_index or label_kind == "price" or group_kind == "price"
        if is_price_column and not is_selected:
            continue
        if not is_selected and not is_quantity_column:
            continue

        column = get_column_letter(index + 1)
        candidates.append(
            {
                "key": column,
                "column": column,
                "label": label or column,
                "group_label": group_label,
                "raw_value": raw_text,
                "quantity": quantity,
                "selected": is_selected,
                "warnings": quantity_warnings,
            }
        )
    return candidates


def _needs_multiple_number_review(
    raw_text: str,
    selected_quantity_index: int | None,
    quantity_candidates: list[dict[str, Any]],
) -> bool:
    if len(NUMBER_RE.findall(raw_text)) <= 1:
        return False
    if selected_quantity_index is not None:
        return False
    if len(quantity_candidates) > 1:
        return False
    return True


def _quantity_source_for(
    header_context: dict[str, Any],
    *,
    selected_quantity_index: int | None,
    quantity_raw: str,
    quantity: float | int | None,
    quantity_candidates: list[dict[str, Any]],
) -> dict[str, Any]:
    if selected_quantity_index is not None and quantity is not None:
        column = get_column_letter(selected_quantity_index + 1)
        for candidate in quantity_candidates:
            if candidate.get("column") == column:
                return {**candidate, "method": "mapped_column"}
        headers = header_context.get("all_headers") or header_context.get("headers") or {}
        groups = header_context.get("header_groups") or {}
        return {
            "key": column,
            "column": column,
            "label": _clean_text(headers.get(selected_quantity_index)) or column,
            "group_label": _clean_text(groups.get(selected_quantity_index)),
            "raw_value": _clean_text(quantity_raw),
            "quantity": quantity,
            "selected": True,
            "method": "mapped_column",
            "warnings": [],
        }
    if quantity is not None:
        return {
            "key": "inline_text",
            "column": "",
            "label": "行内识别",
            "group_label": "",
            "raw_value": _clean_text(quantity_raw),
            "quantity": quantity,
            "selected": True,
            "method": "inline_text",
            "warnings": [],
        }
    return {}


def _parse_quantity(value: str) -> tuple[float | int | None, list[str]]:
    text = _clean_text(value)
    warnings: list[str] = []
    if not text:
        return None, warnings
    if RANGE_RE.search(text):
        return None, ["RANGE_QUANTITY"]
    if "约" in text or "左右" in text or "大约" in text:
        warnings.append("APPROXIMATE_QUANTITY")
    match = NUMBER_RE.search(text)
    if not match:
        return None, ["INVALID_QUANTITY"]
    number = float(match.group(0))
    if number <= 0:
        return None, ["INVALID_QUANTITY"]
    if number.is_integer():
        return int(number), warnings
    return number, warnings


def _confidence_for(extraction: dict[str, Any], header_context: dict[str, Any], warnings: list[str]) -> str:
    core_ok = bool(extraction["item_name"] and extraction["quantity"] is not None and extraction["unit"])
    if not core_ok:
        return "low"
    if "AMBIGUOUS_HEADER" in warnings or "RANGE_QUANTITY" in warnings or "MULTIPLE_NUMBERS" in warnings:
        return "medium"
    strengths = header_context.get("strengths", {})
    if strengths.get("item_name") == "strong" and strengths.get("quantity") == "strong" and strengths.get("unit") == "strong":
        return "high"
    return "medium"


def _issues_for_row(row: dict[str, Any]) -> list[dict[str, Any]]:
    issues: list[dict[str, Any]] = []
    for code in row.get("warnings", []):
        if code in {"MULTI_SHEET_DETECTED"}:
            continue
        issues.append(
            {
                "source_sheet": row.get("source_sheet"),
                "raw_row_index": row.get("raw_row_index"),
                "code": code,
                "message": WARNING_MESSAGES.get(code, code),
            }
        )
    return issues


def _normalize_manual_mappings(mappings: Any) -> dict[str, dict[str, str]]:
    if not mappings:
        return {}
    source = mappings
    if isinstance(source, dict) and "sheet_mappings" in source:
        source = source["sheet_mappings"]
    normalized: dict[str, dict[str, str]] = {}
    if isinstance(source, list):
        for item in source:
            if not isinstance(item, dict):
                continue
            sheet_name = str(item.get("sheet_name") or item.get("source_sheet") or "")
            if not sheet_name:
                continue
            normalized[sheet_name] = _clean_manual_field_mapping(item.get("field_mapping") or {})
        return normalized
    if isinstance(source, dict):
        for sheet_name, field_mapping in source.items():
            if isinstance(field_mapping, dict) and "field_mapping" in field_mapping:
                field_mapping = field_mapping["field_mapping"]
            normalized[str(sheet_name)] = _clean_manual_field_mapping(field_mapping or {})
    return normalized


def _clean_manual_field_mapping(field_mapping: dict[str, Any]) -> dict[str, str]:
    cleaned: dict[str, str] = {}
    for column, field in (field_mapping or {}).items():
        column_text = str(column or "").strip().upper()
        if not column_text:
            continue
        try:
            column_text = get_column_letter(column_index_from_string(column_text))
        except ValueError:
            continue
        field_text = str(field or "ignore").strip()
        if field_text not in MANUAL_FIELD_VALUES:
            field_text = "ignore"
        cleaned[column_text] = field_text
    return cleaned


def _header_context_from_field_mapping(field_mapping: dict[str, str], *, header_row_index: int | None) -> dict[str, Any]:
    column_map: dict[str, int] = {}
    headers: dict[int, str] = {}
    strengths: dict[str, str] = {}
    for column, field in field_mapping.items():
        if field == "ignore":
            continue
        try:
            index = column_index_from_string(column) - 1
        except ValueError:
            continue
        kind = "price" if field == "price_ignored" else field
        column_map[kind] = index
        headers[index] = column
        if kind != "price":
            strengths[kind] = "strong"
    return {
        "header_row_index": header_row_index,
        "column_map": column_map,
        "headers": headers,
        "strengths": strengths,
        "score": len(column_map),
    }


def _row_from_preview_row(row: dict[str, Any], field_mapping: dict[str, str]) -> dict[str, Any]:
    max_index = 0
    cells = row.get("raw_cells") or []
    for cell in cells:
        column = str(cell.get("column") or "")
        if not column:
            continue
        try:
            max_index = max(max_index, column_index_from_string(column))
        except ValueError:
            continue
    for column in field_mapping:
        try:
            max_index = max(max_index, column_index_from_string(column))
        except ValueError:
            continue
    values = [""] * max(max_index, 0)
    for cell in cells:
        column = str(cell.get("column") or "")
        if not column:
            continue
        try:
            index = column_index_from_string(column) - 1
        except ValueError:
            continue
        if 0 <= index < len(values):
            values[index] = _clean_text(cell.get("value"))
    return {
        "index": int(row.get("raw_row_index") or 0),
        "values": values,
        "merged_expanded": "MERGED_CELL_EXPANDED" in (row.get("warnings") or []),
    }


def _summary_for_rows(rows: list[dict[str, Any]], original_summary: dict[str, Any]) -> dict[str, Any]:
    standard_row_count = sum(1 for row in rows if row.get("row_type") == "data_row")
    requires_confirmation_count = sum(1 for row in rows if row.get("requires_confirmation"))
    return {
        **(original_summary or {}),
        "standard_row_count": standard_row_count,
        "requires_confirmation_count": requires_confirmation_count,
        "ignored_row_count": len(rows) - standard_row_count,
        "total_output_row_count": len(rows),
    }


def _clean_confirmed_row(row: dict[str, Any]) -> tuple[dict[str, Any], list[str]]:
    errors: list[str] = []
    warnings = list(row.get("warnings") or [])
    manually_confirmed = bool(row.get("confirmed") or row.get("manual_confirmed"))
    requires_manual = (
        bool(row.get("requires_confirmation"))
        or row.get("confidence") == "low"
        or any(code in FORCE_CONFIRMATION_WARNINGS for code in warnings)
    )
    item_name = _clean_text(row.get("item_name"))
    quantity, quantity_warnings = _parse_quantity(_clean_text(row.get("quantity")))
    unit = _clean_text(row.get("unit"))

    if row.get("row_type") != "data_row" and not manually_confirmed:
        errors.append("NOT_DATA_ROW")
    if not item_name:
        errors.append("MISSING_ITEM_NAME")
    if quantity is None:
        errors.append(quantity_warnings[0] if quantity_warnings else "MISSING_QUANTITY")
    if not unit:
        errors.append("MISSING_UNIT")
    if requires_manual and not manually_confirmed:
        errors.append("CONFIRMATION_REQUIRED")

    if errors:
        return {}, _unique(errors)

    unit_normalized = _normalize_unit(unit) or unit
    return (
        {
            "requirement_row_key": row.get("requirement_row_key"),
            "source_sheet": row.get("source_sheet"),
            "raw_row_index": row.get("raw_row_index"),
            "row_type": "data_row",
            "item_name": item_name,
            "spec": _clean_text(row.get("spec")),
            "quantity": quantity,
            "quantity_source": row.get("quantity_source") or {},
            "quantity_candidates": row.get("quantity_candidates") or [],
            "unit": unit_normalized,
            "unit_raw": _clean_text(row.get("unit_raw")) or unit,
            "unit_family": UNIT_FAMILIES.get(unit_normalized, row.get("unit_family", "")),
            "remark": _clean_text(row.get("remark")),
            "location": _clean_text(row.get("location")),
            "work_area": _clean_text(row.get("work_area")) or _clean_text(row.get("location")),
            "confidence": row.get("confidence", ""),
            "warnings": _unique(warnings),
            "requires_confirmation": False,
            "manual_confirmed": manually_confirmed,
            "raw_fields": row.get("raw_fields") or {},
            "raw_cells": row.get("raw_cells") or [],
            "raw_text": row.get("raw_text") or "",
            "field_mapping": row.get("field_mapping") or {},
            "normalized_name": _normalize_name(item_name),
            "cost_candidates": row.get("cost_candidates") or [],
        },
        [],
    )


def _classify_header(label: str) -> tuple[str | None, str]:
    key = _normalize_compact(label)
    if not key:
        return None, ""
    if any(_normalize_compact(alias) in key for alias in PRICE_ALIASES):
        return "price", "strong"
    for field in HEADER_MATCH_PRIORITY:
        aliases = FIELD_ALIASES[field]
        for strength in ("strong", "weak"):
            for alias in aliases[strength]:
                alias_key = _normalize_compact(alias)
                if key == alias_key:
                    return field, strength
    for field in HEADER_MATCH_PRIORITY:
        aliases = FIELD_ALIASES[field]
        for strength in ("strong", "weak"):
            for alias in sorted(aliases[strength], key=lambda value: len(_normalize_compact(value)), reverse=True):
                alias_key = _normalize_compact(alias)
                if len(alias_key) < 2 or alias_key in CONTAINS_EXACT_ONLY_ALIASES:
                    continue
                if alias_key in key:
                    return field, strength
    return None, ""


def _raw_cells(values: list[str]) -> list[dict[str, str]]:
    cells: list[dict[str, str]] = []
    for index, value in enumerate(values):
        if not value:
            continue
        cells.append({"column": get_column_letter(index + 1), "value": value})
    return cells


def _raw_fields(values: list[str], header_labels: dict[int, str]) -> dict[str, str]:
    fields: dict[str, str] = {}
    for index, value in enumerate(values):
        if not value:
            continue
        label = header_labels.get(index) or get_column_letter(index + 1)
        fields[label] = value
    return fields


def _value_at(values: list[str], index: int | None) -> str:
    if index is None or index >= len(values):
        return ""
    return values[index].strip()


def _first_text_cell(values: list[str], *, skip_indexes: set[int | None]) -> str:
    for index, value in enumerate(values):
        if index in skip_indexes or not value:
            continue
        if _looks_like_item_text(value):
            return value
    return ""


def _split_name_spec(value: str) -> dict[str, str]:
    text = value.strip()
    match = SPEC_TAIL_RE.match(text)
    if not match:
        return {"item_name": text, "spec": ""}
    return {"item_name": match.group("name").strip(), "spec": match.group("spec").strip()}


def _looks_like_note(value: str) -> bool:
    text = value.strip()
    return any(text.startswith(prefix) for prefix in NOTE_PREFIXES)


def _looks_like_section(value: str) -> bool:
    text = value.strip()
    if re.match(r"^[一二三四五六七八九十]+[、.．]", text):
        return True
    if re.match(r"^\d+[、.．]\s*", text):
        return True
    return len(text) <= 12 and not NUMBER_RE.search(text)


def _looks_like_noise_row(values: list[str]) -> bool:
    return bool(values) and all(_looks_like_noise_cell(value) for value in values)


def _looks_like_noise_cell(value: str) -> bool:
    text = _clean_text(value).upper()
    if not text:
        return True
    if text in {"-", "--", "—", "–", "/"}:
        return True
    if text in EXCEL_ERROR_VALUES:
        return True
    if NUMBER_RE.fullmatch(text):
        return True
    return bool(re.fullmatch(r"[A-Z]{1,3}", text))


def _looks_like_item_text(value: str) -> bool:
    text = _clean_text(value)
    if not text:
        return False
    if _looks_like_noise_cell(text):
        return False
    return True


def _should_show_mapping_column(
    column: str,
    label: str,
    detected_field: str,
    samples: list[str],
    *,
    has_header: bool,
) -> bool:
    if detected_field == "price_ignored":
        return False
    if label and _classify_header(label)[0] == "price":
        return False
    if detected_field == "ignore" and label and _classify_header(label)[0] == "quantity":
        return False
    if detected_field != "ignore":
        return True
    if label and label != column and _looks_like_item_text(label):
        return True
    return not has_header and any(_looks_like_item_text(sample) for sample in samples)


def _normalize_unit(value: str) -> str:
    text = _clean_text(value)
    if not text:
        return ""
    direct = UNIT_NORMALIZATION.get(text, text)
    if direct in UNIT_FAMILIES:
        return direct
    match = re.fullmatch(UNIT_PATTERN, text)
    if match:
        return UNIT_NORMALIZATION.get(match.group(0), match.group(0))
    return ""


def _normalize_name(value: str) -> str:
    return _normalize_compact(value)


def _normalize_compact(value: str) -> str:
    return re.sub(r"[\s:：|,，、。；;（）()\[\]【】<>《》/\\_-]+", "", _clean_text(value).lower())


def _join_values(values: list[str]) -> str:
    return " ".join(value for value in values if value).strip()


def _clean_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M")
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    text = str(value).strip()
    if text.upper() in EXCEL_ERROR_VALUES:
        return ""
    return text


def _unique(values: list[str]) -> list[str]:
    seen: set[str] = set()
    result: list[str] = []
    for value in values:
        if not value or value in seen:
            continue
        seen.add(value)
        result.append(value)
    return result


def _md_cell(value: Any) -> str:
    text = "" if value is None else str(value)
    return text.replace("|", "\\|").replace("\n", " ")
