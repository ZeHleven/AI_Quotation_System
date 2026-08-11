from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime
from io import BytesIO
from pathlib import Path
import re
from typing import Any

from openpyxl import load_workbook

from app.services.legacy_excel import LegacyExcelConversionError, normalize_excel_workbook_bytes


SUPPORTED_EXCEL_EXTENSIONS = {".xlsx", ".xlsm", ".xls"}
LEGACY_EXCEL_EXTENSIONS = {".xls"}
SUPPORTED_EXCEL_MIME_TYPES = {
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    "application/vnd.ms-excel.sheet.macroenabled.12",
    "application/vnd.ms-excel",
}

HEADER_ALIASES = {
    "name": ("施工项目", "项目名称", "工作内容", "材料名称", "清单名称", "名称"),
    "quantity": ("数量", "工程量", "计量数量"),
    "unit": ("计量单位", "单位"),
    "spec": ("规格型号", "规格", "特征描述", "项目特征", "特征"),
    "notes": ("工艺说明", "备注说明", "备注", "说明"),
}

TOTAL_TERMS = ("合计", "小计", "总计", "汇总", "总价")
HEADER_SCAN_ROWS = 30
MAX_ITEMS = 300


class QuoteExcelParseError(ValueError):
    pass


@dataclass(frozen=True)
class QuoteExcelParseResult:
    sheet_name: str
    item_count: int
    text: str
    items: tuple[dict[str, str], ...]


def _suffix(filename: str | None) -> str:
    if not filename:
        return ""
    return Path(filename).suffix.lower()


def is_quote_excel_file(filename: str | None, mime_type: str | None) -> bool:
    mime = (mime_type or "").lower()
    return _suffix(filename) in SUPPORTED_EXCEL_EXTENSIONS or mime in SUPPORTED_EXCEL_MIME_TYPES


def is_legacy_excel_file(filename: str | None, mime_type: str | None) -> bool:
    mime = (mime_type or "").lower()
    return _suffix(filename) in LEGACY_EXCEL_EXTENSIONS or mime == "application/vnd.ms-excel"


def _clean_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M")
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _normalize_header(value: Any) -> str:
    return re.sub(r"[\s:：/\\|,，、。;；（）()\[\]【】<>《》]+", "", _clean_text(value).lower())


def _header_kind(value: Any) -> str | None:
    key = _normalize_header(value)
    if not key:
        return None
    for kind, aliases in HEADER_ALIASES.items():
        for alias in aliases:
            alias_key = _normalize_header(alias)
            if key == alias_key or alias_key in key:
                return kind
    return None


def _row_values(row: tuple[Any, ...]) -> list[str]:
    return [_clean_text(cell) for cell in row]


def _is_total_or_empty_row(values: list[str]) -> bool:
    joined = "".join(values).strip()
    if not joined:
        return True
    compact = _normalize_header(joined)
    return any(term in compact for term in TOTAL_TERMS)


def _is_repeated_header(values: list[str]) -> bool:
    kinds = {_header_kind(value) for value in values}
    kinds.discard(None)
    return len(kinds) >= 2 and "name" in kinds


def _detect_header(rows: list[tuple[Any, ...]]) -> tuple[int, dict[str, int]]:
    best_index = -1
    best_score = 0
    best_map: dict[str, int] = {}

    for index, row in enumerate(rows[:HEADER_SCAN_ROWS]):
        column_map: dict[str, int] = {}
        score = 0
        for column_index, value in enumerate(row):
            kind = _header_kind(value)
            if not kind or kind in column_map:
                continue
            column_map[kind] = column_index
            score += {"name": 4, "quantity": 3, "unit": 3, "spec": 1, "notes": 1}.get(kind, 1)
        if "name" in column_map and ("quantity" in column_map or "unit" in column_map) and score > best_score:
            best_index = index
            best_score = score
            best_map = column_map

    if best_index < 0:
        raise QuoteExcelParseError("Excel 需求单未识别到可报价项目，请检查是否包含项目名称、数量、单位列")
    return best_index, best_map


def _cell(values: list[str], column_map: dict[str, int], key: str) -> str:
    index = column_map.get(key)
    if index is None or index >= len(values):
        return ""
    return values[index].strip()


def _format_item(values: list[str], column_map: dict[str, int]) -> tuple[str, dict[str, str]] | None:
    if _is_total_or_empty_row(values) or _is_repeated_header(values):
        return None

    name = _cell(values, column_map, "name")
    if not name:
        return None

    compact_name = _normalize_header(name)
    if any(term in compact_name for term in TOTAL_TERMS):
        return None

    spec = _cell(values, column_map, "spec")
    quantity = _cell(values, column_map, "quantity")
    unit = _cell(values, column_map, "unit")
    notes = _cell(values, column_map, "notes")

    parts = [name]
    if spec:
        parts.append(f"规格/特征：{spec}")
    if quantity:
        parts.append(f"数量：{quantity}")
    if unit:
        parts.append(f"单位：{unit}")
    if notes:
        parts.append(f"备注：{notes}")
    item = {
        "project_name": name,
        "quantity": quantity,
        "unit": unit,
        "spec": spec,
        "notes": notes,
    }
    return "，".join(parts), {key: value for key, value in item.items() if value}


def _parse_sheet(sheet) -> QuoteExcelParseResult | None:
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return None

    try:
        header_index, column_map = _detect_header(rows)
    except QuoteExcelParseError:
        return None
    item_texts: list[str] = []
    items: list[dict[str, str]] = []
    for row in rows[header_index + 1 :]:
        values = _row_values(row)
        item = _format_item(values, column_map)
        if item:
            item_text, item_data = item
            item_texts.append(item_text)
            items.append(item_data)
        if len(item_texts) >= MAX_ITEMS:
            break

    if not item_texts:
        return None

    return QuoteExcelParseResult(
        sheet_name=sheet.title,
        item_count=len(item_texts),
        text="；".join(item_texts),
        items=tuple(items),
    )


def parse_quote_excel_bytes(file_content: bytes, *, filename: str | None = None) -> QuoteExcelParseResult:
    if not file_content:
        raise QuoteExcelParseError("Excel 需求单为空，请重新上传")

    try:
        normalized_content = normalize_excel_workbook_bytes(file_content, filename=filename)
        workbook = load_workbook(BytesIO(normalized_content), data_only=True, read_only=True)
    except LegacyExcelConversionError as exc:
        raise QuoteExcelParseError(str(exc)) from exc
    except Exception as exc:
        raise QuoteExcelParseError(f"Excel 需求单读取失败，请确认文件为 .xls/.xlsx/.xlsm 格式: {exc}") from exc

    for sheet in workbook.worksheets:
        result = _parse_sheet(sheet)
        if result:
            return result

    name = filename or "上传文件"
    raise QuoteExcelParseError(f"{name} 未识别到可报价项目，请检查是否包含项目名称、数量、单位列和明细行")
