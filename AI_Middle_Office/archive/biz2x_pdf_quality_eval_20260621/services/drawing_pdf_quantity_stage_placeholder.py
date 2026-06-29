from __future__ import annotations

import csv
import json
from datetime import datetime
from pathlib import Path
from typing import Any, Iterable, Mapping, Sequence

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


PHASE = "BIZ-2x-pdf-quantity-stage-placeholder"

QUANTITY_HEADERS = [
    "row_no",
    "quantity_status",
    "block_reason",
    "division",
    "standard_code",
    "item_code",
    "item_name",
    "feature",
    "unit",
    "quantity",
    "formula_type",
    "quantity_rule_text",
    "required_evidence",
    "source_files",
    "evidence_ids",
    "manual_action",
]

CHINESE_QUANTITY_HEADERS = [
    "序号",
    "算量状态",
    "阻断原因",
    "分部分项",
    "标准来源",
    "项目编码",
    "项目名称",
    "项目特征",
    "计量单位",
    "工程量",
    "计算规则类型",
    "工程量计算规则",
    "所需图纸证据",
    "图纸来源",
    "证据编号",
    "下一步动作",
]


def build_quantity_stage_placeholder_report(
    standard_bill_report: Mapping[str, Any],
    *,
    v2_report: Mapping[str, Any] | None = None,
    gate_report: Mapping[str, Any] | None = None,
) -> dict[str, Any]:
    bill_rows = [dict(row) for row in standard_bill_report.get("bill_rows") or []]
    human_rows = [dict(row) for row in (v2_report or {}).get("human_style_rows") or []]
    gate_passed = bool(
        (gate_report or {}).get("can_enable_quantity")
        or (standard_bill_report.get("summary") or {}).get("can_enable_quantity")
    )

    quantity_rows: list[dict[str, Any]] = []
    for index, row in enumerate(bill_rows, start=1):
        human_row = human_rows[index - 1] if index <= len(human_rows) else {}
        rule = _quantity_rule_for_row(row, human_row)
        status, reason, action = _quantity_status(row, gate_passed=gate_passed)
        quantity_rows.append(
            {
                "row_no": row.get("row_no") or index,
                "quantity_status": status,
                "block_reason": reason,
                "division": row.get("division", ""),
                "standard_code": row.get("standard_code", ""),
                "item_code": row.get("item_code", ""),
                "item_name": row.get("item_name", ""),
                "feature": row.get("feature", ""),
                "unit": row.get("unit", ""),
                "quantity": "",
                "formula_type": rule.get("formula_type", ""),
                "quantity_rule_text": rule.get("rule_text", ""),
                "required_evidence": _join(rule.get("required_evidence") or rule.get("drawing_evidence_requirements") or []),
                "source_files": row.get("source_files", ""),
                "evidence_ids": row.get("evidence_ids", ""),
                "manual_action": action,
            }
        )

    blocked_count = sum(1 for row in quantity_rows if str(row.get("quantity_status") or "").startswith("blocked"))
    ready_placeholder_count = sum(1 for row in quantity_rows if row.get("quantity_status") == "ready_for_quantity_engine")
    gate_blocked_count = sum(1 for row in quantity_rows if row.get("quantity_status") == "blocked_three_field_gate_failed")
    standard_unmapped_count = sum(1 for row in quantity_rows if row.get("quantity_status") == "blocked_standard_unmapped")
    return {
        "ok": True,
        "phase": PHASE,
        "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "safe_for_final_quantity_list": False,
        "quantity_acceptance_enabled": False,
        "summary": {
            "quantity_row_count": len(quantity_rows),
            "ready_placeholder_count": ready_placeholder_count,
            "blocked_count": blocked_count,
            "gate_blocked_count": gate_blocked_count,
            "standard_unmapped_count": standard_unmapped_count,
            "quantity_filled_count": 0,
            "can_enable_quantity": gate_passed,
            "quantity_status": "placeholder_only_ready_for_quantity_engine"
            if gate_passed
            else "blocked_until_three_field_gate_passed",
        },
        "quantity_rows": quantity_rows,
    }


def write_quantity_stage_placeholder_outputs(
    report: Mapping[str, Any],
    output_dir: str | Path,
    *,
    stem: str | None = None,
) -> dict[str, str]:
    target = Path(output_dir)
    target.mkdir(parents=True, exist_ok=True)
    file_stem = stem or f"BIZ2x_PDF_quantity_stage_placeholder_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
    json_path = target / f"{file_stem}.json"
    csv_path = target / f"{file_stem}.csv"
    markdown_path = target / f"{file_stem}.md"
    xlsx_path = target / f"{file_stem}.xlsx"
    outputs = {
        "json": str(json_path),
        "csv": str(csv_path),
        "markdown": str(markdown_path),
        "xlsx": str(xlsx_path),
    }
    payload = {**dict(report), "outputs": outputs}
    json_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    _write_csv(csv_path, payload.get("quantity_rows") or [], QUANTITY_HEADERS)
    markdown_path.write_text(_build_markdown(payload), encoding="utf-8")
    _write_workbook(xlsx_path, payload)
    return outputs


def _quantity_rule_for_row(row: Mapping[str, Any], human_row: Mapping[str, Any]) -> dict[str, Any]:
    item_code = str(row.get("item_code") or "").strip()
    candidates = human_row.get("standard_candidates") or []
    if isinstance(candidates, list):
        for candidate in candidates:
            if not isinstance(candidate, Mapping):
                continue
            candidate_code = str(candidate.get("official_item_code") or candidate.get("item_code") or "").strip()
            if item_code and candidate_code == item_code:
                rule = dict(candidate.get("quantity_rule") or {})
                if candidate.get("drawing_evidence_requirements") and not rule.get("required_evidence"):
                    rule["required_evidence"] = candidate.get("drawing_evidence_requirements")
                return rule
        for candidate in candidates:
            if isinstance(candidate, Mapping) and candidate.get("quantity_rule"):
                rule = dict(candidate.get("quantity_rule") or {})
                if candidate.get("drawing_evidence_requirements") and not rule.get("required_evidence"):
                    rule["required_evidence"] = candidate.get("drawing_evidence_requirements")
                return rule
    return {}


def _quantity_status(row: Mapping[str, Any], *, gate_passed: bool) -> tuple[str, str, str]:
    if not gate_passed:
        return (
            "blocked_three_field_gate_failed",
            "三字段验收门禁未通过，禁止进入工程量计算",
            "先补识图证据并通过项目名称、项目特征、单位验收",
        )
    if not str(row.get("item_code") or "").strip():
        return (
            "blocked_standard_unmapped",
            "未匹配国标项目编码，无法选择可靠计算规则",
            "先完成人工复核或标准库映射",
        )
    if row.get("export_status") != "final_candidate":
        return (
            "blocked_standard_bill_review",
            "标准清单行仍处于复核状态，暂不可计算工程量",
            "先处理标准清单预览中的复核说明",
        )
    return (
        "ready_for_quantity_engine",
        "",
        "读取图纸尺寸、范围、表格或点位后再填充工程量",
    )


def _build_markdown(report: Mapping[str, Any]) -> str:
    summary = report.get("summary") or {}
    lines = [
        "# BIZ-2x PDF Quantity Stage Placeholder",
        "",
        f"- generated_at: {report.get('generated_at', '-')}",
        f"- quantity_status: {summary.get('quantity_status', '-')}",
        f"- rows: {summary.get('quantity_row_count', 0)}",
        f"- blocked: {summary.get('blocked_count', 0)}",
        f"- ready_placeholder: {summary.get('ready_placeholder_count', 0)}",
        "",
        "| row | status | item_code | item_name | unit | rule | required_evidence | action |",
        "| ---: | --- | --- | --- | --- | --- | --- | --- |",
    ]
    for row in (report.get("quantity_rows") or [])[:80]:
        lines.append(
            "| "
            + " | ".join(
                [
                    _md(row.get("row_no")),
                    _md(row.get("quantity_status")),
                    _md(row.get("item_code")),
                    _md(row.get("item_name")),
                    _md(row.get("unit")),
                    _md(row.get("formula_type")),
                    _md(row.get("required_evidence")),
                    _md(row.get("manual_action")),
                ]
            )
            + " |"
        )
    return "\n".join(lines) + "\n"


def _write_workbook(path: Path, report: Mapping[str, Any]) -> None:
    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "summary"
    _append_rows(summary_sheet, [["metric", "value"]])
    _append_rows(summary_sheet, [[key, _cell_value(value)] for key, value in (report.get("summary") or {}).items()])
    _style_sheet(summary_sheet)

    quantity_sheet = workbook.create_sheet("quantity_placeholder")
    _append_rows(quantity_sheet, [CHINESE_QUANTITY_HEADERS])
    _append_rows(
        quantity_sheet,
        [[_cell_value(row.get(header)) for header in QUANTITY_HEADERS] for row in report.get("quantity_rows") or []],
    )
    _style_sheet(quantity_sheet)
    workbook.save(path)


def _write_csv(path: Path, rows: Iterable[Mapping[str, Any]], fieldnames: Sequence[str]) -> None:
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        for row in rows:
            writer.writerow({field: _cell_value(row.get(field)) for field in fieldnames})


def _append_rows(sheet: Any, rows: Sequence[Sequence[Any]]) -> None:
    for row in rows:
        sheet.append(list(row))


def _style_sheet(sheet: Any) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    if sheet.max_row >= 1:
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for col_idx, column_cells in enumerate(sheet.columns, start=1):
        values = [str(cell.value or "") for cell in column_cells[:200]]
        width = min(max([len(value) for value in values] + [10]) + 2, 65)
        sheet.column_dimensions[get_column_letter(col_idx)].width = width


def _join(value: Any) -> str:
    if isinstance(value, str):
        return value
    if isinstance(value, (list, tuple, set)):
        return "；".join(str(item) for item in value if str(item).strip())
    return str(value or "")


def _cell_value(value: Any) -> Any:
    if isinstance(value, (dict, list, tuple)):
        return json.dumps(value, ensure_ascii=False)
    return value


def _md(value: Any) -> str:
    return str(value or "").replace("|", "\\|").replace("\n", "<br>")
