from __future__ import annotations

import csv
import json
from datetime import datetime
from pathlib import Path
from typing import Any, Iterable, Mapping, Sequence

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.services.drawing_pdf_v2_takeoff import (
    build_pdf_v2_takeoff_report,
    write_pdf_v2_takeoff_outputs,
)
from app.services.drawing_three_field_acceptance import ThreeFieldAnswerRow


PHASE = "BIZ-2x-pdf-gap-recall-v2-evaluation"

METRIC_KEYS = [
    "evidence_count",
    "normalized_evidence_count",
    "human_style_row_count",
    "standard_mapped_count",
    "standard_unmapped_count",
    "three_field_candidate_count",
    "three_field_matched_count",
    "three_field_pass_rate",
    "three_field_gap_count",
]

METRIC_HEADERS = ["metric", "base", "augmented", "delta"]


def build_gap_recall_v2_evaluation(
    base_v2_report: Mapping[str, Any],
    recall_run_report: Mapping[str, Any],
    *,
    style_prompt_text: str = "",
) -> dict[str, Any]:
    base_summary = dict(base_v2_report.get("summary") or {})
    recall_rows = _recall_evidence_rows(recall_run_report)
    merged_evidence_rows = [dict(row) for row in base_v2_report.get("evidence_rows") or []]
    merged_evidence_rows.extend(recall_rows)
    answer_rows = _answer_rows_from_base_report(base_v2_report)
    pdf_summary = dict(base_v2_report.get("pdf_direct_summary") or {})
    if not pdf_summary:
        pdf_summary = {
            "pdf_file_count": base_summary.get("pdf_file_count", 0),
            "pdf_page_count": base_summary.get("pdf_page_count", 0),
            "pdf_render_status": base_summary.get("pdf_render_status", ""),
        }

    augmented_report = build_pdf_v2_takeoff_report(
        {
            "summary": pdf_summary,
            "evidence_rows": merged_evidence_rows,
        },
        answer_rows=answer_rows,
        style_prompt_text=style_prompt_text,
    )
    augmented_report["inputs"] = {
        "base_phase": base_v2_report.get("phase", ""),
        "recall_phase": recall_run_report.get("phase", ""),
        "recall_execute": (recall_run_report.get("summary") or {}).get("execute", False),
        "recall_evidence_count": len(recall_rows),
        "quantity_acceptance_enabled": False,
    }
    augmented_summary = dict(augmented_report.get("summary") or {})
    metric_rows = _metric_rows(base_summary, augmented_summary)
    return {
        "ok": True,
        "phase": PHASE,
        "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "summary": {
            "base_three_field_matched_count": base_summary.get("three_field_matched_count", 0),
            "augmented_three_field_matched_count": augmented_summary.get("three_field_matched_count", 0),
            "matched_delta": _delta(
                base_summary.get("three_field_matched_count", 0),
                augmented_summary.get("three_field_matched_count", 0),
            ),
            "base_three_field_gap_count": base_summary.get("three_field_gap_count", 0),
            "augmented_three_field_gap_count": augmented_summary.get("three_field_gap_count", 0),
            "gap_delta": _delta(
                base_summary.get("three_field_gap_count", 0),
                augmented_summary.get("three_field_gap_count", 0),
            ),
            "recall_evidence_count": len(recall_rows),
            "quantity_status": augmented_summary.get("quantity_status", "deferred_until_three_fields_accepted"),
        },
        "base_summary": base_summary,
        "recall_run_summary": dict(recall_run_report.get("summary") or {}),
        "augmented_summary": augmented_summary,
        "metric_rows": metric_rows,
        "augmented_v2_report": augmented_report,
    }


def write_gap_recall_v2_evaluation_outputs(
    evaluation: Mapping[str, Any],
    output_dir: str | Path,
    *,
    stem: str | None = None,
) -> dict[str, str]:
    target = Path(output_dir)
    target.mkdir(parents=True, exist_ok=True)
    file_stem = stem or f"BIZ2x_PDF_gap_recall_v2_eval_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
    json_path = target / f"{file_stem}.json"
    csv_path = target / f"{file_stem}_metrics.csv"
    markdown_path = target / f"{file_stem}.md"
    xlsx_path = target / f"{file_stem}.xlsx"
    augmented_outputs = write_pdf_v2_takeoff_outputs(
        evaluation.get("augmented_v2_report") or {},
        target,
        stem=f"{file_stem}_augmented_v2",
    )
    outputs = {
        "json": str(json_path),
        "metrics_csv": str(csv_path),
        "markdown": str(markdown_path),
        "xlsx": str(xlsx_path),
        "augmented_v2_json": augmented_outputs["json"],
        "augmented_v2_xlsx": augmented_outputs["xlsx"],
    }
    report = {**dict(evaluation), "outputs": outputs}
    json_path.write_text(json.dumps(_json_safe(report), ensure_ascii=False, indent=2), encoding="utf-8")
    _write_csv(csv_path, report.get("metric_rows") or [], METRIC_HEADERS)
    markdown_path.write_text(_build_markdown(report), encoding="utf-8")
    _write_workbook(xlsx_path, report)
    return outputs


def _answer_rows_from_base_report(base_v2_report: Mapping[str, Any]) -> list[ThreeFieldAnswerRow]:
    raw_rows = ((base_v2_report.get("three_field_acceptance_report") or {}).get("answer_rows") or [])
    rows: list[ThreeFieldAnswerRow] = []
    for row in raw_rows:
        rows.append(
            ThreeFieldAnswerRow(
                sheet_name=str(row.get("sheet_name") or ""),
                row_no=int(_float(row.get("row_no"), 0)),
                section=str(row.get("section") or ""),
                seq=str(row.get("seq") or ""),
                item_code=str(row.get("item_code") or ""),
                item_name=str(row.get("item_name") or ""),
                feature=str(row.get("feature") or ""),
                unit=str(row.get("unit") or ""),
                quantity=str(row.get("quantity") or ""),
            )
        )
    return rows


def _recall_evidence_rows(recall_run_report: Mapping[str, Any]) -> list[dict[str, Any]]:
    rows = list(recall_run_report.get("evidence_rows") or [])
    if not rows and isinstance(recall_run_report.get("visual_evidence_report"), Mapping):
        rows = list((recall_run_report.get("visual_evidence_report") or {}).get("evidence_rows") or [])
    result: list[dict[str, Any]] = []
    for index, row in enumerate(rows, start=1):
        item = dict(row)
        existing_id = str(item.get("evidence_id") or "").strip()
        item["evidence_id"] = existing_id or f"PDFGAP-{index:06d}"
        item.setdefault("source_kind", "pdf_gap_recall_llm")
        result.append(item)
    return result


def _metric_rows(base_summary: Mapping[str, Any], augmented_summary: Mapping[str, Any]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for key in METRIC_KEYS:
        base = base_summary.get(key, 0)
        augmented = augmented_summary.get(key, 0)
        rows.append({"metric": key, "base": base, "augmented": augmented, "delta": _delta(base, augmented)})
    return rows


def _delta(base: Any, augmented: Any) -> Any:
    try:
        return round(float(augmented or 0) - float(base or 0), 4)
    except (TypeError, ValueError):
        return ""


def _build_markdown(report: Mapping[str, Any]) -> str:
    summary = report.get("summary") or {}
    lines = [
        "# BIZ-2x PDF Gap Recall V2 Evaluation",
        "",
        f"- generated_at: {report.get('generated_at', '-')}",
        f"- recall evidence: {summary.get('recall_evidence_count', 0)}",
        f"- matched delta: {summary.get('matched_delta', 0)}",
        f"- gap delta: {summary.get('gap_delta', 0)}",
        f"- quantity status: {summary.get('quantity_status', '-')}",
        "",
        "## Metrics",
        "",
        "| metric | base | augmented | delta |",
        "| --- | ---: | ---: | ---: |",
    ]
    for row in report.get("metric_rows") or []:
        lines.append(
            f"| {_md(row.get('metric'))} | {_md(row.get('base'))} | {_md(row.get('augmented'))} | {_md(row.get('delta'))} |"
        )
    return "\n".join(lines) + "\n"


def _write_workbook(path: Path, report: Mapping[str, Any]) -> None:
    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "eval_summary"
    _append_rows(summary_sheet, [["metric", "value"]])
    for key, value in (report.get("summary") or {}).items():
        _append_rows(summary_sheet, [[key, value]])
    _style_sheet(summary_sheet)

    metrics_sheet = workbook.create_sheet("metric_delta")
    _append_rows(metrics_sheet, [METRIC_HEADERS])
    _append_rows(metrics_sheet, [[row.get(header, "") for header in METRIC_HEADERS] for row in report.get("metric_rows") or []])
    _style_sheet(metrics_sheet)
    workbook.save(path)


def _append_rows(sheet: Any, rows: Iterable[Sequence[Any]]) -> None:
    for row in rows:
        sheet.append(list(row))


def _style_sheet(sheet: Any) -> None:
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for col_idx, column_cells in enumerate(sheet.columns, start=1):
        values = [str(cell.value or "") for cell in column_cells[:80]]
        width = min(max([len(value) for value in values] + [10]) + 2, 80)
        sheet.column_dimensions[get_column_letter(col_idx)].width = width
    sheet.freeze_panes = "A2"


def _write_csv(path: Path, rows: Iterable[Mapping[str, Any]], fieldnames: Sequence[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        for row in rows:
            writer.writerow({key: _cell_value(row.get(key)) for key in fieldnames})


def _cell_value(value: Any) -> Any:
    if isinstance(value, (dict, list, tuple)):
        return json.dumps(value, ensure_ascii=False)
    return value


def _json_safe(value: Any) -> Any:
    if isinstance(value, Mapping):
        return {str(key): _json_safe(item) for key, item in value.items()}
    if isinstance(value, (list, tuple)):
        return [_json_safe(item) for item in value]
    if isinstance(value, Path):
        return str(value)
    return value


def _md(value: Any) -> str:
    return str(value or "").replace("|", "\\|").replace("\n", "<br>")


def _float(value: Any, default: float = 0.0) -> float:
    try:
        return float(str(value).strip())
    except (TypeError, ValueError):
        return default
