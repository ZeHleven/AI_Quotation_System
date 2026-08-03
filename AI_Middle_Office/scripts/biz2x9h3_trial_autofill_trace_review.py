from __future__ import annotations

import argparse
import json
import sys
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8")

BACKEND_ROOT = Path(__file__).resolve().parents[1]
WORKSPACE_ROOT = BACKEND_ROOT.parent
if str(BACKEND_ROOT) not in sys.path:
    sys.path.insert(0, str(BACKEND_ROOT))

from app.services.dxf_trace_review_pack import (  # noqa: E402
    ADOPT_COLUMN,
    DEDUCTION_REVIEW_COLUMN,
    ISSUE_COLUMN,
    MANUAL_FEATURE_COLUMN,
    QUANTITY_SOURCE_COLUMN,
    REVIEW_COLUMN,
    TRACE_REVIEW_SHEET_NAME,
)


TRIAL_AUTOFILL_ROWS = {
    "BIZ2x9h-0007": {
        "feature": (
            "吊顶部位：职工餐厅顶面造型轮廓区域；"
            "吊顶形式、吊杆规格、高度：不上人艺术造型吊顶，∅8吊筋，吊顶高度按设计图示，完成面距楼板超过1.5m增加反支撑、超过2.5m增加转换层；"
            "龙骨材料种类、规格、中距：50或75龙骨，50副龙骨，中距900mm；"
            "基层材料种类、规格：木方、木阻燃板基层，灯箱处木基层箱体，内部刷白；"
            "面板材料品种、规格：9.5或12厚纸面石膏板、透光软膜、无机涂料饰面；"
            "压条材料种类、规格：L型收边条、软膜卡件，按图纸做法；"
            "嵌缝材料种类：按图纸做法，石膏板沿墙边缘留10mm槽；"
            "防护材料种类：木方、木阻燃板双面刷石油沥青，玛蹄脂三遍，防火涂料三遍，满刷氯偏乳液或乳化光油防潮涂料2道"
        ),
        "deduction_review": (
            "按标准库 011302003 工程量计算规则采用顶面造型轮廓 CAD 水平投影面积；"
            "天棚面中的灯槽及跌级天棚面积不展开计算；扣除与天棚相连的窗帘盒所占面积；"
            "不扣除检查口、附墙烟囱、柱垛、管道以及单个面积≤0.3㎡的独立柱、孔洞所占面积；"
            "当前 trace 未识别需另行扣减项，按标准规则建议量采用。"
        ),
        "quantity_source": (
            "系统代填试运行闭环：来源 04.信达资产职工餐厅水电施工图.dxf，图层 D-顶面造型轮廓；"
            "CAD 公式 sum(CAD_area_mm2) * area_to_square_meter_factor，按 GB/T 011302003 面积规则计算。"
        ),
    },
    "BIZ2x9h-0010": {
        "feature": (
            "吊顶部位：职工餐厅顶面造型轮廓区域；"
            "吊顶形式、吊杆规格、高度：不上人艺术造型吊顶，∅8吊筋，吊顶高度按设计图示，完成面距楼板超过1.5m增加反支撑、超过2.5m增加转换层；"
            "龙骨材料种类、规格、中距：50或75龙骨，50副龙骨，中距900mm；"
            "基层材料种类、规格：木方、木阻燃板基层，灯箱处木基层箱体，内部刷白；"
            "面板材料品种、规格：9.5或12厚纸面石膏板、透光软膜、无机涂料饰面；"
            "压条材料种类、规格：L型收边条、软膜卡件，按图纸做法；"
            "嵌缝材料种类：按图纸做法，石膏板沿墙边缘留10mm槽；"
            "防护材料种类：木方、木阻燃板双面刷石油沥青，玛蹄脂三遍，防火涂料三遍，满刷氯偏乳液或乳化光油防潮涂料2道"
        ),
        "deduction_review": (
            "按标准库 011302003 工程量计算规则采用顶面造型轮廓 CAD 水平投影面积；"
            "天棚面中的灯槽及跌级天棚面积不展开计算；扣除与天棚相连的窗帘盒所占面积；"
            "不扣除检查口、附墙烟囱、柱垛、管道以及单个面积≤0.3㎡的独立柱、孔洞所占面积；"
            "当前 trace 未识别需另行扣减项，按标准规则建议量采用。"
        ),
        "quantity_source": (
            "系统代填试运行闭环：来源 01.前言文件.dxf，图层 D-顶面造型轮廓；"
            "CAD 公式 sum(CAD_area_mm2) * area_to_square_meter_factor，按 GB/T 011302003 面积规则计算。"
        ),
    },
}


def main() -> int:
    parser = argparse.ArgumentParser(description="BIZ-2x-9h-3 试运行闭环：系统代填当前 2 条建议采用 trace")
    parser.add_argument("--trace-review-workbook", default="", help="BIZ-2x-9h-2 trace 复核工作簿；不填则取最新")
    parser.add_argument("--output-dir", default=str(WORKSPACE_ROOT / "outputs" / "biz2x9h3"))
    parser.add_argument("--print-summary-only", action="store_true")
    args = parser.parse_args()

    try:
        source_path = Path(args.trace_review_workbook) if args.trace_review_workbook else _latest(
            WORKSPACE_ROOT / "outputs" / "biz2x9h",
            "BIZ2x9h2_标准规则trace自动初判复核包_*.xlsx",
        )
        result = _autofill(source_path, Path(args.output_dir))
    except (OSError, ValueError) as exc:
        print(json.dumps({"ok": False, "error": "BIZ2X9H3_TRIAL_AUTOFILL_FAILED", "message": str(exc)}, ensure_ascii=False, indent=2))
        return 2

    if args.print_summary_only:
        print(json.dumps({"ok": result["ok"], "summary": result["summary"], "outputs": result["outputs"]}, ensure_ascii=False, indent=2))
    else:
        print(json.dumps(result, ensure_ascii=False, indent=2))
    return 0


def _autofill(source_path: Path, output_dir: Path) -> dict[str, Any]:
    workbook = load_workbook(source_path)
    if TRACE_REVIEW_SHEET_NAME not in workbook.sheetnames:
        raise ValueError(f"trace 复核工作簿缺少工作表：{TRACE_REVIEW_SHEET_NAME}")
    sheet = workbook[TRACE_REVIEW_SHEET_NAME]
    headers = [str(cell.value or "").strip() for cell in sheet[1]]
    header_to_column = {header: index + 1 for index, header in enumerate(headers) if header}
    required_headers = [
        "复核行号",
        ADOPT_COLUMN,
        REVIEW_COLUMN,
        MANUAL_FEATURE_COLUMN,
        DEDUCTION_REVIEW_COLUMN,
        QUANTITY_SOURCE_COLUMN,
        ISSUE_COLUMN,
    ]
    missing_headers = [header for header in required_headers if header not in header_to_column]
    if missing_headers:
        raise ValueError(f"trace 复核工作簿缺少字段：{', '.join(missing_headers)}")

    filled_rows: list[str] = []
    for row_index in range(2, sheet.max_row + 1):
        row_id = str(sheet.cell(row=row_index, column=header_to_column["复核行号"]).value or "").strip()
        fill = TRIAL_AUTOFILL_ROWS.get(row_id)
        if not fill:
            continue
        sheet.cell(row=row_index, column=header_to_column[ADOPT_COLUMN]).value = "是"
        sheet.cell(row=row_index, column=header_to_column[REVIEW_COLUMN]).value = "通过"
        sheet.cell(row=row_index, column=header_to_column[MANUAL_FEATURE_COLUMN]).value = fill["feature"]
        sheet.cell(row=row_index, column=header_to_column[DEDUCTION_REVIEW_COLUMN]).value = fill["deduction_review"]
        sheet.cell(row=row_index, column=header_to_column[QUANTITY_SOURCE_COLUMN]).value = fill["quantity_source"]
        sheet.cell(row=row_index, column=header_to_column[ISSUE_COLUMN]).value = "系统代填用于试运行闭环验证；正式试运行前仍建议业务复核。"
        filled_rows.append(row_id)

    if not filled_rows:
        raise ValueError("未找到可代填的目标行。")

    output_dir.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    xlsx_path = output_dir / f"BIZ2x9h2_系统代填通过版_{timestamp}.xlsx"
    json_path = output_dir / f"BIZ2x9h2_系统代填通过版_{timestamp}.json"
    workbook.save(xlsx_path)
    result = {
        "ok": True,
        "phase": "BIZ-2x-9h-3-trial-autofill-trace-review",
        "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "inputs": {"trace_review_workbook": str(source_path)},
        "summary": {
            "target_row_count": len(TRIAL_AUTOFILL_ROWS),
            "filled_row_count": len(filled_rows),
            "filled_rows": filled_rows,
            "final_export_requires_biz2x9h3_validation": True,
        },
        "outputs": {
            "trial_autofilled_xlsx": str(xlsx_path),
            "json": str(json_path),
        },
    }
    json_path.write_text(json.dumps(result, ensure_ascii=False, indent=2), encoding="utf-8")
    return result


def _latest(directory: Path, pattern: str) -> Path:
    matches = sorted(directory.glob(pattern), key=lambda item: item.stat().st_mtime, reverse=True)
    if not matches:
        raise ValueError(f"No file matched {pattern} in {directory}")
    return matches[0]


if __name__ == "__main__":
    raise SystemExit(main())
