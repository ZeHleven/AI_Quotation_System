from __future__ import annotations

import csv
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


QUANTITY_LIST_HEADERS = ["项目名称", "项目特征", "单位", "工程量"]


def write_quantity_list_outputs(rows: list[dict[str, Any]], output_dir: str | Path, *, stem: str) -> dict[str, str]:
    directory = Path(output_dir)
    directory.mkdir(parents=True, exist_ok=True)
    csv_path = directory / f"{stem}.csv"
    xlsx_path = directory / f"{stem}.xlsx"
    _write_four_field_csv(csv_path, rows)
    _write_four_field_workbook(xlsx_path, rows)
    return {"csv": str(csv_path), "xlsx": str(xlsx_path)}


def _write_four_field_csv(path: Path, rows: list[dict[str, Any]]) -> None:
    with path.open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.DictWriter(handle, fieldnames=QUANTITY_LIST_HEADERS)
        writer.writeheader()
        for row in rows:
            writer.writerow({header: row.get(header, "") for header in QUANTITY_LIST_HEADERS})


def _write_four_field_workbook(path: Path, rows: list[dict[str, Any]]) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "识图四字段清单"
    sheet.append(QUANTITY_LIST_HEADERS)
    for row in rows:
        sheet.append([row.get(header, "") for header in QUANTITY_LIST_HEADERS])
    _style_four_field_sheet(sheet)
    workbook.save(path)


def _style_four_field_sheet(sheet) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    widths = {
        "A": 28,
        "B": 64,
        "C": 12,
        "D": 18,
    }
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:{get_column_letter(sheet.max_column)}{sheet.max_row}"
