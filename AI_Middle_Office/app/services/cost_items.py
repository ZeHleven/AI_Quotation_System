from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime, timedelta, timezone
from io import BytesIO
from typing import Any, Mapping
from uuid import uuid4

from fastapi import HTTPException, status
from openpyxl import load_workbook
from sqlalchemy import func, or_
from sqlalchemy.orm import Session

from app.models.cost_item import (
    CHANGE_TYPE_PRICE,
    CHANGE_TYPE_STATUS,
    COST_SOURCE_IMPORTED,
    COST_SOURCE_VALUES,
    COST_STATUS_ACTIVE,
    COST_STATUS_ARCHIVED,
    COST_STATUS_DRAFT,
    COST_STATUS_VALUES,
    CostItem,
    CostItemHistory,
    PRICE_TYPE_COMBINED,
    PRICE_TYPE_VALUES,
    CostRagSyncRun,
)
from app.models.quote_cost_evidence import QuoteCostEvidence
from app.models.user import User
from app.services.cost_duplicate_guard import (
    DUPLICATE_CONFLICT_CODE,
    active_duplicate_conflicts,
    active_duplicate_conflicts_for_item,
)
from app.services.rbac import has_any_role


IMPORT_BATCH_TTL = timedelta(minutes=30)
IMPORT_BATCHES: dict[str, "ImportBatch"] = {}

PRICE_FIELDS = (
    "price",
    "client_tax_excluded_price",
    "client_labor_price",
    "client_main_material_price",
    "client_auxiliary_material_price",
    "client_direct_fee",
    "client_management_profit",
    "subcontract_composite_price",
    "subcontract_labor_price",
    "subcontract_main_material_price",
    "subcontract_auxiliary_material_price",
    "crew_benchmark_price",
)
UPDATE_FIELDS = {
    "category",
    "subcategory",
    "item_name",
    "spec",
    "unit",
    "price",
    "client_tax_excluded_price",
    "client_labor_price",
    "client_main_material_price",
    "client_auxiliary_material_price",
    "client_direct_fee",
    "client_management_profit",
    "subcontract_composite_price",
    "subcontract_labor_price",
    "subcontract_main_material_price",
    "subcontract_auxiliary_material_price",
    "crew_benchmark_price",
    "price_type",
    "source",
    "effective_date",
    "notes",
}


@dataclass
class ImportBatch:
    batch_id: str
    created_at: datetime
    items: list[dict[str, Any]]
    duplicate_warnings: list[dict[str, Any]]
    skipped_rows: list[dict[str, Any]]
    confirmed: bool = False
    result: dict[str, Any] | None = None


FULL_COST_DB_VIEW_ROLES = {"system_admin", "admin", "cost_viewer", "cost_editor", "cost_approver", "cost_exporter"}
FULL_COST_DB_EDIT_ROLES = {"system_admin", "admin", "cost_editor", "cost_approver"}
FULL_COST_DB_APPROVE_ROLES = {"system_admin", "admin", "cost_approver"}
FULL_COST_DB_EXPORT_ROLES = {"system_admin", "admin", "cost_exporter"}
QUOTE_COST_REFERENCE_ROLES = FULL_COST_DB_VIEW_ROLES | {"staff"}


def can_access_cost_db(user: User) -> bool:
    return has_any_role(user, FULL_COST_DB_VIEW_ROLES)


def can_edit_cost_db(user: User) -> bool:
    return has_any_role(user, FULL_COST_DB_EDIT_ROLES)


def can_approve_cost_db(user: User) -> bool:
    return has_any_role(user, FULL_COST_DB_APPROVE_ROLES)


def can_export_cost_db(user: User) -> bool:
    return has_any_role(user, FULL_COST_DB_EXPORT_ROLES)


def can_access_quote_cost_reference(user: User) -> bool:
    return has_any_role(user, QUOTE_COST_REFERENCE_ROLES)


def can_manage_cost_db(user: User) -> bool:
    return can_edit_cost_db(user)


def require_cost_db_access(user: User) -> None:
    if not can_access_cost_db(user):
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="PERMISSION_DENIED")


def require_cost_db_manager(user: User) -> None:
    require_cost_db_editor(user)


def require_cost_db_editor(user: User) -> None:
    if not can_edit_cost_db(user):
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="PERMISSION_DENIED")


def require_cost_db_approver(user: User) -> None:
    if not can_approve_cost_db(user):
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="PERMISSION_DENIED")


def require_cost_db_exporter(user: User) -> None:
    if not can_export_cost_db(user):
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="PERMISSION_DENIED")


def require_quote_cost_reference_access(user: User) -> None:
    if not can_access_quote_cost_reference(user):
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="PERMISSION_DENIED")


def _payload_dict(payload: Mapping[str, Any] | Any) -> dict[str, Any]:
    if hasattr(payload, "model_dump"):
        return payload.model_dump(exclude_unset=True)
    if hasattr(payload, "dict"):
        return payload.dict(exclude_unset=True)
    return dict(payload or {})


def clean_text(value: Any, max_length: int | None = None) -> str | None:
    if value is None:
        return None
    cleaned = str(value).strip()
    if not cleaned:
        return None
    return cleaned[:max_length] if max_length else cleaned


def parse_date(value: Any) -> date | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    raw = str(value).strip()
    if not raw:
        return None
    for parser in (
        date.fromisoformat,
        lambda text: datetime.strptime(text, "%Y-%m-%d").date(),
        lambda text: datetime.strptime(text, "%Y/%m/%d").date(),
    ):
        try:
            return parser(raw)
        except ValueError:
            continue
    raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_ENTITY, detail="INVALID_EFFECTIVE_DATE")


def parse_float(value: Any) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return float(value)
    raw = str(value).strip().replace(",", "")
    if not raw or raw in {"-", "/", "—"}:
        return None
    try:
        return float(raw)
    except ValueError:
        return None


def _positive_price(value: float | None, *, field_name: str) -> float | None:
    if value is None:
        return None
    if value <= 0:
        raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_ENTITY, detail=f"INVALID_{field_name.upper()}")
    return round(float(value), 6)


def _non_negative_price(value: float | None, *, field_name: str) -> float | None:
    if value is None:
        return None
    if value < 0:
        raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_ENTITY, detail=f"INVALID_{field_name.upper()}")
    return round(float(value), 6)


def derive_main_price(data: Mapping[str, Any]) -> float:
    explicit = parse_float(data.get("price"))
    if explicit is not None:
        return _positive_price(explicit, field_name="price") or 0.0
    for field_name in ("subcontract_composite_price", "crew_benchmark_price", "client_tax_excluded_price"):
        candidate = parse_float(data.get(field_name))
        if candidate is not None:
            return _positive_price(candidate, field_name=field_name) or 0.0
    raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_ENTITY, detail="PRICE_REQUIRED")


def normalize_price_type(value: Any) -> str:
    price_type = clean_text(value, 24) or PRICE_TYPE_COMBINED
    if price_type not in PRICE_TYPE_VALUES:
        raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_ENTITY, detail="INVALID_PRICE_TYPE")
    return price_type


def normalize_source(value: Any) -> str:
    source = clean_text(value, 32) or "manual"
    if source not in COST_SOURCE_VALUES:
        raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_ENTITY, detail="INVALID_SOURCE")
    return source


def normalize_status(value: Any) -> str:
    status_value = clean_text(value, 24) or COST_STATUS_DRAFT
    if status_value not in COST_STATUS_VALUES:
        raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_ENTITY, detail="INVALID_STATUS")
    return status_value


def _normalize_item_payload(data: Mapping[str, Any], *, for_import: bool = False) -> dict[str, Any]:
    category = clean_text(data.get("category"), 128)
    item_name = clean_text(data.get("item_name"), 255)
    unit = clean_text(data.get("unit"), 64)
    if not category:
        raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_ENTITY, detail="CATEGORY_REQUIRED")
    if not item_name:
        raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_ENTITY, detail="ITEM_NAME_REQUIRED")
    if not unit:
        raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_ENTITY, detail="UNIT_REQUIRED")

    normalized = {
        "category": category,
        "subcategory": clean_text(data.get("subcategory"), 128),
        "item_name": item_name,
        "spec": clean_text(data.get("spec")),
        "unit": unit,
        "price": derive_main_price(data),
        "client_tax_excluded_price": _positive_price(parse_float(data.get("client_tax_excluded_price")), field_name="client_tax_excluded_price"),
        "client_labor_price": _non_negative_price(parse_float(data.get("client_labor_price")), field_name="client_labor_price"),
        "client_main_material_price": _non_negative_price(parse_float(data.get("client_main_material_price")), field_name="client_main_material_price"),
        "client_auxiliary_material_price": _non_negative_price(parse_float(data.get("client_auxiliary_material_price")), field_name="client_auxiliary_material_price"),
        "client_direct_fee": _non_negative_price(parse_float(data.get("client_direct_fee")), field_name="client_direct_fee"),
        "client_management_profit": _non_negative_price(parse_float(data.get("client_management_profit")), field_name="client_management_profit"),
        "subcontract_composite_price": _positive_price(parse_float(data.get("subcontract_composite_price")), field_name="subcontract_composite_price"),
        "subcontract_labor_price": _non_negative_price(parse_float(data.get("subcontract_labor_price")), field_name="subcontract_labor_price"),
        "subcontract_main_material_price": _non_negative_price(parse_float(data.get("subcontract_main_material_price")), field_name="subcontract_main_material_price"),
        "subcontract_auxiliary_material_price": _non_negative_price(parse_float(data.get("subcontract_auxiliary_material_price")), field_name="subcontract_auxiliary_material_price"),
        "crew_benchmark_price": _positive_price(parse_float(data.get("crew_benchmark_price")), field_name="crew_benchmark_price"),
        "price_type": normalize_price_type(data.get("price_type")),
        "source": COST_SOURCE_IMPORTED if for_import else normalize_source(data.get("source")),
        "effective_date": parse_date(data.get("effective_date")),
        "notes": clean_text(data.get("notes")),
    }
    return normalized


def _format_dt(value: datetime | None) -> str | None:
    if not value:
        return None
    return value.isoformat(timespec="seconds")


def _format_date(value: date | None) -> str | None:
    if not value:
        return None
    return value.isoformat()


def _price_values_equal(left: Any, right: Any) -> bool:
    if left is None and right is None:
        return True
    if left is None or right is None:
        return False
    try:
        return round(float(left), 6) == round(float(right), 6)
    except (TypeError, ValueError):
        return left == right


def _changed_price_fields(history: CostItemHistory) -> list[str]:
    changed: list[str] = []
    for field in PRICE_FIELDS:
        if not _price_values_equal(getattr(history, f"old_{field}"), getattr(history, f"new_{field}")):
            changed.append(field)
    return changed


def _history_snapshot(history: CostItemHistory) -> dict[str, Any]:
    return {
        "id": history.id,
        "cost_item_id": history.cost_item_id,
        "old_price": history.old_price,
        "new_price": history.new_price,
        "old_client_tax_excluded_price": history.old_client_tax_excluded_price,
        "new_client_tax_excluded_price": history.new_client_tax_excluded_price,
        "old_client_labor_price": history.old_client_labor_price,
        "new_client_labor_price": history.new_client_labor_price,
        "old_client_main_material_price": history.old_client_main_material_price,
        "new_client_main_material_price": history.new_client_main_material_price,
        "old_client_auxiliary_material_price": history.old_client_auxiliary_material_price,
        "new_client_auxiliary_material_price": history.new_client_auxiliary_material_price,
        "old_client_direct_fee": history.old_client_direct_fee,
        "new_client_direct_fee": history.new_client_direct_fee,
        "old_client_management_profit": history.old_client_management_profit,
        "new_client_management_profit": history.new_client_management_profit,
        "old_subcontract_composite_price": history.old_subcontract_composite_price,
        "new_subcontract_composite_price": history.new_subcontract_composite_price,
        "old_subcontract_labor_price": history.old_subcontract_labor_price,
        "new_subcontract_labor_price": history.new_subcontract_labor_price,
        "old_subcontract_main_material_price": history.old_subcontract_main_material_price,
        "new_subcontract_main_material_price": history.new_subcontract_main_material_price,
        "old_subcontract_auxiliary_material_price": history.old_subcontract_auxiliary_material_price,
        "new_subcontract_auxiliary_material_price": history.new_subcontract_auxiliary_material_price,
        "old_crew_benchmark_price": history.old_crew_benchmark_price,
        "new_crew_benchmark_price": history.new_crew_benchmark_price,
        "old_status": history.old_status,
        "new_status": history.new_status,
        "change_type": history.change_type,
        "changed_by": history.changed_by,
        "change_reason": history.change_reason,
        "changed_at": _format_dt(history.changed_at),
        "changed_fields": _changed_price_fields(history) if history.change_type == CHANGE_TYPE_PRICE else [],
    }


def _usernames_for_ids(db: Session, user_ids: set[int | None]) -> dict[int, str]:
    ids = {int(user_id) for user_id in user_ids if user_id}
    if not ids:
        return {}
    rows = db.query(User.id, User.username).filter(User.id.in_(ids)).all()
    return {int(user_id): username for user_id, username in rows}


def _history_snapshot_with_user(history: CostItemHistory, usernames: dict[int, str]) -> dict[str, Any]:
    data = _history_snapshot(history)
    data["changed_by_username"] = usernames.get(history.changed_by or 0)
    return data


def serialize_cost_item(item: CostItem, *, include_history: bool = False) -> dict[str, Any]:
    data = {
        "id": item.id,
        "category": item.category,
        "subcategory": item.subcategory,
        "item_name": item.item_name,
        "spec": item.spec,
        "unit": item.unit,
        "price": item.price,
        "client_tax_excluded_price": item.client_tax_excluded_price,
        "client_labor_price": item.client_labor_price,
        "client_main_material_price": item.client_main_material_price,
        "client_auxiliary_material_price": item.client_auxiliary_material_price,
        "client_direct_fee": item.client_direct_fee,
        "client_management_profit": item.client_management_profit,
        "subcontract_composite_price": item.subcontract_composite_price,
        "subcontract_labor_price": item.subcontract_labor_price,
        "subcontract_main_material_price": item.subcontract_main_material_price,
        "subcontract_auxiliary_material_price": item.subcontract_auxiliary_material_price,
        "crew_benchmark_price": item.crew_benchmark_price,
        "price_type": item.price_type,
        "status": item.status,
        "source": item.source,
        "effective_date": _format_date(item.effective_date),
        "notes": item.notes,
        "created_by": item.created_by,
        "created_at": _format_dt(item.created_at),
        "updated_at": _format_dt(item.updated_at),
    }
    if include_history:
        data["history"] = [_history_snapshot(history) for history in sorted(item.history, key=lambda row: row.id)]
    return data


def serialize_quote_cost_candidate(item: CostItem, *, include_full_cost: bool = False) -> dict[str, Any]:
    if include_full_cost:
        return serialize_cost_item(item)
    return {
        "id": item.id,
        "category": item.category,
        "subcategory": item.subcategory,
        "item_name": item.item_name,
        "spec": item.spec,
        "unit": item.unit,
        "price": item.price,
        "price_type": item.price_type,
        "status": item.status,
        "updated_at": _format_dt(item.updated_at),
        "restricted": True,
    }


def _duplicate_conflict_detail(item: CostItem, conflicts: list[dict[str, Any]]) -> dict[str, Any]:
    message = conflicts[0]["message"] if conflicts else "已存在冲突的 active 成本项"
    return {
        "code": DUPLICATE_CONFLICT_CODE,
        "message": message,
        "item": serialize_cost_item(item),
        "matches": [conflict["existing_item"] for conflict in conflicts],
        "conflicts": conflicts,
    }


def _latest_successful_rag_sync_run(db: Session) -> CostRagSyncRun | None:
    return (
        db.query(CostRagSyncRun)
        .filter(CostRagSyncRun.status == "success")
        .order_by(CostRagSyncRun.finished_at.desc(), CostRagSyncRun.started_at.desc(), CostRagSyncRun.id.desc())
        .first()
    )


def _serialize_lineage_rag_run(run: CostRagSyncRun | None) -> dict[str, Any] | None:
    if not run:
        return None
    return {
        "id": run.id,
        "status": run.status,
        "requested_count": int(run.requested_count or 0),
        "synced_count": int(run.synced_count or 0),
        "started_at": _format_dt(run.started_at),
        "finished_at": _format_dt(run.finished_at),
        "triggered_by_username": run.triggered_by_username,
    }


def _parse_cost_notes_metadata(notes: str | None) -> dict[str, str]:
    metadata: dict[str, str] = {}
    for raw_line in (notes or "").splitlines():
        line = raw_line.strip()
        if not line or ":" not in line:
            continue
        key, value = line.split(":", 1)
        key = key.strip()
        value = value.strip()
        if key and value and value != "-":
            metadata[key] = value
    return metadata


def _source_origin(item: CostItem, usernames: dict[int, str]) -> dict[str, Any]:
    metadata = _parse_cost_notes_metadata(item.notes)
    return {
        "source": item.source,
        "created_by": item.created_by,
        "created_by_username": usernames.get(item.created_by or 0),
        "created_at": _format_dt(item.created_at),
        "quote_job_id": metadata.get("quote_job_id"),
        "quote_history_id": _parse_int(metadata.get("quote_history_id")),
        "line_no": _parse_int(metadata.get("line_no")),
        "requirement_row_key": metadata.get("requirement_row_key"),
        "source_sheet": metadata.get("source_sheet"),
        "raw_row_index": _parse_int(metadata.get("raw_row_index")),
        "confirmed_unit_price": _parse_float(metadata.get("confirmed_unit_price")),
        "confirmed_total_price": _parse_float(metadata.get("confirmed_total_price")),
        "draft_price_source": metadata.get("draft_price_source"),
        "manual_price_action": metadata.get("manual_price_action"),
        "final_price_source": metadata.get("final_price_source"),
        "price_confirmation_label": metadata.get("price_confirmation_label"),
        "cost_item_source": metadata.get("cost_item_source"),
        "notice": metadata.get("notice"),
    }


def _parse_int(value: Any) -> int | None:
    if value in (None, ""):
        return None
    try:
        return int(value)
    except (TypeError, ValueError):
        return None


def _parse_float(value: Any) -> float | None:
    if value in (None, ""):
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _lineage_destination(item: CostItem, quote_usage_count: int, latest_rag_run: CostRagSyncRun | None) -> dict[str, Any]:
    participates_in_quote = item.status == COST_STATUS_ACTIVE
    in_rag_scope = item.status == COST_STATUS_ACTIVE
    if item.status == COST_STATUS_DRAFT:
        status_text = "待成本部审核，不参与报价匹配、底价兜底或 RAG 同步"
    elif item.status == COST_STATUS_ACTIVE:
        status_text = "已启用，参与后续报价匹配、底价兜底，并可同步到 RAG"
    else:
        status_text = "已归档冻结，不参与后续报价匹配、底价兜底或 RAG 同步"
    return {
        "status": item.status,
        "status_text": status_text,
        "participates_in_quote": participates_in_quote,
        "in_rag_sync_scope": in_rag_scope,
        "quote_usage_count": quote_usage_count,
        "latest_successful_rag_sync": _serialize_lineage_rag_run(latest_rag_run) if in_rag_scope else None,
        "rag_sync_note": "当前仅按 active 范围推断是否进入同步；尚未记录单条同步明细",
    }


def _quote_usage_stats(db: Session, item_ids: list[int]) -> dict[int, dict[str, Any]]:
    if not item_ids:
        return {}
    counts = {
        int(cost_item_id): int(count or 0)
        for cost_item_id, count in (
            db.query(QuoteCostEvidence.cost_item_id, func.count(QuoteCostEvidence.id))
            .filter(QuoteCostEvidence.cost_item_id.in_(item_ids))
            .group_by(QuoteCostEvidence.cost_item_id)
            .all()
        )
        if cost_item_id
    }
    latest_rows = (
        db.query(QuoteCostEvidence)
        .filter(QuoteCostEvidence.cost_item_id.in_(item_ids))
        .order_by(QuoteCostEvidence.cost_item_id.asc(), QuoteCostEvidence.created_at.desc(), QuoteCostEvidence.id.desc())
        .all()
    )
    latest_by_item: dict[int, QuoteCostEvidence] = {}
    for row in latest_rows:
        if row.cost_item_id and row.cost_item_id not in latest_by_item:
            latest_by_item[int(row.cost_item_id)] = row
    return {
        item_id: {
            "quote_usage_count": counts.get(item_id, 0),
            "latest_quote_used_at": _format_dt(latest_by_item[item_id].created_at) if item_id in latest_by_item else None,
            "latest_quote_job_id": latest_by_item[item_id].quote_job_id if item_id in latest_by_item else None,
            "latest_quote_history_id": latest_by_item[item_id].quote_history_id if item_id in latest_by_item else None,
        }
        for item_id in item_ids
    }


def _lineage_item(
    item: CostItem,
    *,
    usernames: dict[int, str],
    usage: dict[str, Any] | None = None,
    latest_rag_run: CostRagSyncRun | None = None,
) -> dict[str, Any]:
    usage = usage or {}
    quote_usage_count = int(usage.get("quote_usage_count") or 0)
    data = serialize_cost_item(item)
    data["origin"] = _source_origin(item, usernames)
    data["destination"] = _lineage_destination(item, quote_usage_count, latest_rag_run)
    data["quote_usage"] = {
        "count": quote_usage_count,
        "latest_used_at": usage.get("latest_quote_used_at"),
        "latest_quote_job_id": usage.get("latest_quote_job_id"),
        "latest_quote_history_id": usage.get("latest_quote_history_id"),
    }
    return data


def cost_item_lineage_summary(db: Session, user: User) -> dict[str, Any]:
    require_cost_db_access(user)
    status_counts = {status_value: 0 for status_value in COST_STATUS_VALUES}
    for status_value, count in db.query(CostItem.status, func.count(CostItem.id)).group_by(CostItem.status).all():
        status_counts[status_value] = int(count or 0)

    source_counts = {source_value: 0 for source_value in COST_SOURCE_VALUES}
    for source_value, count in db.query(CostItem.source, func.count(CostItem.id)).group_by(CostItem.source).all():
        source_counts[source_value] = int(count or 0)

    used_item_ids_query = db.query(QuoteCostEvidence.cost_item_id).filter(QuoteCostEvidence.cost_item_id.isnot(None)).distinct()
    quote_used_count = used_item_ids_query.count()
    active_quote_used_count = (
        db.query(func.count(func.distinct(CostItem.id)))
        .join(QuoteCostEvidence, QuoteCostEvidence.cost_item_id == CostItem.id)
        .filter(CostItem.status == COST_STATUS_ACTIVE)
        .scalar()
        or 0
    )
    latest_rag_run = _latest_successful_rag_sync_run(db)
    return {
        "total": sum(status_counts.values()),
        "by_status": status_counts,
        "by_source": source_counts,
        "ai_suggested_draft_count": (
            db.query(func.count(CostItem.id))
            .filter(CostItem.source == "ai_suggested", CostItem.status == COST_STATUS_DRAFT)
            .scalar()
            or 0
        ),
        "quote_used_count": int(quote_used_count or 0),
        "active_quote_used_count": int(active_quote_used_count or 0),
        "active_rag_scope_count": status_counts.get(COST_STATUS_ACTIVE, 0),
        "latest_successful_rag_sync": _serialize_lineage_rag_run(latest_rag_run),
    }


def list_cost_item_lineage(
    db: Session,
    user: User,
    *,
    statuses: list[str] | None = None,
    source: str | None = None,
    keyword: str | None = None,
    has_quote_usage: bool | None = None,
    page: int = 1,
    page_size: int = 20,
) -> tuple[list[dict[str, Any]], int]:
    require_cost_db_access(user)
    query = db.query(CostItem)
    if statuses:
        invalid = [status_value for status_value in statuses if status_value not in COST_STATUS_VALUES]
        if invalid:
            raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_ENTITY, detail="INVALID_STATUS")
        query = query.filter(CostItem.status.in_(statuses))
    if source:
        query = query.filter(CostItem.source == normalize_source(source))
    keyword_text = clean_text(keyword, 128)
    if keyword_text:
        pattern = f"%{keyword_text}%"
        query = query.filter(
            or_(
                CostItem.item_name.like(pattern),
                CostItem.spec.like(pattern),
                CostItem.category.like(pattern),
                CostItem.subcategory.like(pattern),
                CostItem.notes.like(pattern),
            )
        )
    used_ids = db.query(QuoteCostEvidence.cost_item_id).filter(QuoteCostEvidence.cost_item_id.isnot(None)).distinct()
    if has_quote_usage is True:
        query = query.filter(CostItem.id.in_(used_ids))
    elif has_quote_usage is False:
        query = query.filter(~CostItem.id.in_(used_ids))

    total = query.count()
    items = (
        query.order_by(CostItem.updated_at.desc(), CostItem.id.desc())
        .offset((page - 1) * page_size)
        .limit(page_size)
        .all()
    )
    item_ids = [int(item.id) for item in items]
    usage_by_item = _quote_usage_stats(db, item_ids)
    usernames = _usernames_for_ids(db, {item.created_by for item in items})
    latest_rag_run = _latest_successful_rag_sync_run(db)
    return [
        _lineage_item(
            item,
            usernames=usernames,
            usage=usage_by_item.get(int(item.id)),
            latest_rag_run=latest_rag_run,
        )
        for item in items
    ], total


def get_cost_item_lineage(db: Session, user: User, item_id: int) -> dict[str, Any]:
    require_cost_db_access(user)
    item = _get_item(db, item_id)
    evidence_rows = (
        db.query(QuoteCostEvidence)
        .filter(QuoteCostEvidence.cost_item_id == item.id)
        .order_by(QuoteCostEvidence.created_at.desc(), QuoteCostEvidence.id.desc())
        .limit(50)
        .all()
    )
    user_ids = {item.created_by}
    user_ids.update(history.changed_by for history in item.history)
    usernames = _usernames_for_ids(db, user_ids)
    usage_by_item = _quote_usage_stats(db, [int(item.id)])
    data = _lineage_item(
        item,
        usernames=usernames,
        usage=usage_by_item.get(int(item.id)),
        latest_rag_run=_latest_successful_rag_sync_run(db),
    )
    data["history"] = [
        _history_snapshot_with_user(history, usernames)
        for history in sorted(item.history, key=lambda row: row.id)
    ]
    data["status_history"] = [event for event in data["history"] if event.get("change_type") == CHANGE_TYPE_STATUS]
    data["price_history"] = [event for event in data["history"] if event.get("change_type") == CHANGE_TYPE_PRICE]
    data["quote_usages"] = [_serialize_lineage_quote_usage(row) for row in evidence_rows]
    return data


def _serialize_lineage_quote_usage(row: QuoteCostEvidence) -> dict[str, Any]:
    return {
        "id": row.id,
        "quote_id": row.quote_id,
        "quote_job_id": row.quote_job_id,
        "quote_history_id": row.quote_history_id,
        "username": row.username,
        "status": row.status,
        "item_index": row.item_index,
        "project_name": row.project_name,
        "quantity": row.quantity,
        "unit": row.unit,
        "reference_price": row.reference_price,
        "reference_total": row.reference_total,
        "final_unit_price": row.final_unit_price,
        "final_total_price": row.final_total_price,
        "price_delta": row.price_delta,
        "price_delta_rate": row.price_delta_rate,
        "fallback_applied": row.fallback_applied,
        "created_at": _format_dt(row.created_at),
        "confirmed_at": _format_dt(row.confirmed_at),
    }


def _write_price_history(db: Session, item: CostItem, user: User, old_values: dict[str, Any], reason: str | None) -> None:
    changed = any(not _price_values_equal(old_values[field], getattr(item, field)) for field in PRICE_FIELDS)
    if not changed:
        return
    db.add(
        CostItemHistory(
            cost_item_id=item.id,
            old_price=old_values["price"],
            new_price=item.price,
            old_client_tax_excluded_price=old_values["client_tax_excluded_price"],
            new_client_tax_excluded_price=item.client_tax_excluded_price,
            old_client_labor_price=old_values["client_labor_price"],
            new_client_labor_price=item.client_labor_price,
            old_client_main_material_price=old_values["client_main_material_price"],
            new_client_main_material_price=item.client_main_material_price,
            old_client_auxiliary_material_price=old_values["client_auxiliary_material_price"],
            new_client_auxiliary_material_price=item.client_auxiliary_material_price,
            old_client_direct_fee=old_values["client_direct_fee"],
            new_client_direct_fee=item.client_direct_fee,
            old_client_management_profit=old_values["client_management_profit"],
            new_client_management_profit=item.client_management_profit,
            old_subcontract_composite_price=old_values["subcontract_composite_price"],
            new_subcontract_composite_price=item.subcontract_composite_price,
            old_subcontract_labor_price=old_values["subcontract_labor_price"],
            new_subcontract_labor_price=item.subcontract_labor_price,
            old_subcontract_main_material_price=old_values["subcontract_main_material_price"],
            new_subcontract_main_material_price=item.subcontract_main_material_price,
            old_subcontract_auxiliary_material_price=old_values["subcontract_auxiliary_material_price"],
            new_subcontract_auxiliary_material_price=item.subcontract_auxiliary_material_price,
            old_crew_benchmark_price=old_values["crew_benchmark_price"],
            new_crew_benchmark_price=item.crew_benchmark_price,
            change_type=CHANGE_TYPE_PRICE,
            changed_by=user.id,
            change_reason=clean_text(reason, 2000) or "price updated",
        )
    )


def _write_status_history(db: Session, item: CostItem, user: User, old_status: str | None, reason: str | None) -> None:
    if old_status == item.status:
        return
    db.add(
        CostItemHistory(
            cost_item_id=item.id,
            old_status=old_status,
            new_status=item.status,
            change_type=CHANGE_TYPE_STATUS,
            changed_by=user.id,
            change_reason=clean_text(reason, 2000),
        )
    )


def _get_item(db: Session, item_id: int) -> CostItem:
    item = db.query(CostItem).filter(CostItem.id == item_id).first()
    if not item:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="RESOURCE_NOT_FOUND")
    return item


def create_cost_item(db: Session, user: User, payload: Mapping[str, Any] | Any) -> CostItem:
    require_cost_db_editor(user)
    data = _normalize_item_payload(_payload_dict(payload))
    item = CostItem(**data, status=COST_STATUS_DRAFT, created_by=user.id)
    db.add(item)
    db.flush()
    return item


def _cost_items_query(
    db: Session,
    *,
    category: str | None = None,
    subcategory: str | None = None,
    statuses: list[str] | None = None,
    price_type: str | None = None,
    source: str | None = None,
    keyword: str | None = None,
):
    query = db.query(CostItem)
    category_text = clean_text(category, 128)
    if category_text:
        pattern = f"%{category_text}%"
        query = query.filter(or_(CostItem.category.like(pattern), CostItem.subcategory.like(pattern)))
    subcategory_text = clean_text(subcategory, 128)
    if subcategory_text:
        query = query.filter(CostItem.subcategory.like(f"%{subcategory_text}%"))
    if statuses:
        invalid = [status_value for status_value in statuses if status_value not in COST_STATUS_VALUES]
        if invalid:
            raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_ENTITY, detail="INVALID_STATUS")
        query = query.filter(CostItem.status.in_(statuses))
    if price_type:
        query = query.filter(CostItem.price_type == normalize_price_type(price_type))
    if source:
        query = query.filter(CostItem.source == normalize_source(source))
    keyword_text = clean_text(keyword, 128)
    if keyword_text:
        pattern = f"%{keyword_text}%"
        query = query.filter(
            or_(
                CostItem.item_name.like(pattern),
                CostItem.spec.like(pattern),
                CostItem.category.like(pattern),
                CostItem.subcategory.like(pattern),
                CostItem.notes.like(pattern),
            )
        )
    return query


def list_cost_items(
    db: Session,
    user: User,
    *,
    category: str | None = None,
    subcategory: str | None = None,
    statuses: list[str] | None = None,
    price_type: str | None = None,
    source: str | None = None,
    keyword: str | None = None,
    page: int = 1,
    page_size: int = 20,
) -> tuple[list[CostItem], int]:
    require_cost_db_access(user)
    query = _cost_items_query(
        db,
        category=category,
        subcategory=subcategory,
        statuses=statuses,
        price_type=price_type,
        source=source,
        keyword=keyword,
    )
    total = query.count()
    items = query.order_by(CostItem.updated_at.desc(), CostItem.id.desc()).offset((page - 1) * page_size).limit(page_size).all()
    return items, total


def export_cost_items(
    db: Session,
    user: User,
    *,
    category: str | None = None,
    subcategory: str | None = None,
    statuses: list[str] | None = None,
    price_type: str | None = None,
    source: str | None = None,
    keyword: str | None = None,
    limit: int = 5000,
) -> tuple[list[CostItem], int]:
    require_cost_db_exporter(user)
    query = _cost_items_query(
        db,
        category=category,
        subcategory=subcategory,
        statuses=statuses,
        price_type=price_type,
        source=source,
        keyword=keyword,
    )
    total = query.count()
    items = query.order_by(CostItem.updated_at.desc(), CostItem.id.desc()).limit(limit).all()
    return items, total


def list_quote_cost_candidates(
    db: Session,
    user: User,
    *,
    keyword: str | None,
    page: int = 1,
    page_size: int = 20,
) -> tuple[list[CostItem], int]:
    require_quote_cost_reference_access(user)
    keyword_text = clean_text(keyword, 128)
    if not keyword_text:
        raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_ENTITY, detail="KEYWORD_REQUIRED")
    if len(keyword_text) < 2:
        raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_ENTITY, detail="KEYWORD_TOO_SHORT")
    query = db.query(CostItem).filter(CostItem.status == COST_STATUS_ACTIVE)
    pattern = f"%{keyword_text}%"
    query = query.filter(
        or_(
            CostItem.item_name.like(pattern),
            CostItem.spec.like(pattern),
            CostItem.category.like(pattern),
            CostItem.subcategory.like(pattern),
        )
    )
    total = query.count()
    items = query.order_by(CostItem.updated_at.desc(), CostItem.id.desc()).offset((page - 1) * page_size).limit(page_size).all()
    return items, total


def get_cost_item(db: Session, user: User, item_id: int) -> CostItem:
    require_cost_db_access(user)
    return _get_item(db, item_id)


def update_cost_item(db: Session, user: User, item_id: int, payload: Mapping[str, Any] | Any) -> CostItem:
    item = _get_item(db, item_id)
    if item.status == COST_STATUS_ACTIVE:
        require_cost_db_approver(user)
    else:
        require_cost_db_editor(user)
    if item.status == COST_STATUS_ARCHIVED:
        raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="STATE_CONFLICT")

    data = _payload_dict(payload)
    reason = data.pop("change_reason", None)
    unexpected = set(data) - UPDATE_FIELDS
    if unexpected:
        raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_ENTITY, detail="UNSUPPORTED_FIELDS")

    old_prices = {field: getattr(item, field) for field in PRICE_FIELDS}
    candidate = serialize_cost_item(item)
    candidate.update(data)
    if "price" not in data and any(field in data for field in PRICE_FIELDS if field != "price"):
        candidate["price"] = None
    normalized = _normalize_item_payload(candidate)
    if item.status == COST_STATUS_ACTIVE:
        duplicate_candidate = dict(candidate)
        duplicate_candidate.update(normalized)
        duplicate_candidate["id"] = item.id
        conflicts = active_duplicate_conflicts(db, duplicate_candidate, exclude_item_id=item.id)
        if conflicts:
            raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail=_duplicate_conflict_detail(item, conflicts))
    for field, value in normalized.items():
        setattr(item, field, value)
    db.flush()
    _write_price_history(db, item, user, old_prices, reason)
    return item


def activate_cost_item(db: Session, user: User, item_id: int, *, reason: str | None = None) -> CostItem:
    require_cost_db_approver(user)
    item = _get_item(db, item_id)
    if item.status == COST_STATUS_ACTIVE:
        return item
    if item.status == COST_STATUS_ARCHIVED:
        raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="STATE_CONFLICT")
    cleaned_reason = clean_text(reason, 2000)
    if not cleaned_reason:
        raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_ENTITY, detail="REASON_REQUIRED")
    conflicts = active_duplicate_conflicts_for_item(db, item)
    if conflicts:
        raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail=_duplicate_conflict_detail(item, conflicts))
    old_status = item.status
    item.status = COST_STATUS_ACTIVE
    db.flush()
    _write_status_history(db, item, user, old_status, cleaned_reason)
    return item


def bulk_update_cost_item_status(
    db: Session,
    user: User,
    item_ids: list[int],
    target_status: str,
    reason: str | None = None,
) -> dict[str, Any]:
    require_cost_db_approver(user)
    seen: set[int] = set()
    normalized_ids: list[int] = []
    for item_id in item_ids:
        if item_id <= 0:
            raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_ENTITY, detail="INVALID_ITEM_ID")
        if item_id in seen:
            continue
        seen.add(item_id)
        normalized_ids.append(item_id)
    if not normalized_ids:
        raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_ENTITY, detail="ITEM_IDS_REQUIRED")
    if target_status not in {COST_STATUS_ACTIVE, COST_STATUS_DRAFT, COST_STATUS_ARCHIVED}:
        raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_ENTITY, detail="INVALID_TARGET_STATUS")

    cleaned_reason = clean_text(reason, 2000)
    if target_status in {COST_STATUS_ACTIVE, COST_STATUS_DRAFT, COST_STATUS_ARCHIVED} and not cleaned_reason:
        raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_ENTITY, detail="REASON_REQUIRED")

    items = db.query(CostItem).filter(CostItem.id.in_(normalized_ids)).all()
    item_by_id = {item.id: item for item in items}
    changed_ids: list[int] = []
    skipped: list[dict[str, Any]] = []
    conflicts: list[dict[str, Any]] = []
    not_found: list[int] = []

    for item_id in normalized_ids:
        item = item_by_id.get(item_id)
        if item is None:
            not_found.append(item_id)
            continue
        if item.status == target_status:
            skipped.append({"id": item.id, "status": item.status, "reason": f"already_{target_status}"})
            continue
        if item.status == COST_STATUS_ARCHIVED:
            conflicts.append({"id": item.id, "status": item.status, "reason": "archived_locked"})
            continue
        if target_status == COST_STATUS_ACTIVE:
            duplicate_conflicts = active_duplicate_conflicts_for_item(db, item)
            if duplicate_conflicts:
                conflicts.append(
                    {
                        "id": item.id,
                        "status": item.status,
                        "reason": "duplicate_active_conflict",
                        "duplicate_conflict": _duplicate_conflict_detail(item, duplicate_conflicts),
                    }
                )
                continue

        old_status = item.status
        item.status = target_status
        _write_status_history(db, item, user, old_status, cleaned_reason)
        changed_ids.append(item.id)

    db.flush()
    return {
        "target_status": target_status,
        "requested_count": len(normalized_ids),
        "changed_count": len(changed_ids),
        "skipped_count": len(skipped),
        "conflict_count": len(conflicts),
        "not_found_count": len(not_found),
        "changed_ids": changed_ids,
        "skipped": skipped,
        "conflicts": conflicts,
        "not_found": not_found,
    }


def withdraw_cost_item_activation(db: Session, user: User, item_id: int, reason: str | None = None) -> CostItem:
    require_cost_db_approver(user)
    item = _get_item(db, item_id)
    if item.status == COST_STATUS_DRAFT:
        return item
    if item.status == COST_STATUS_ARCHIVED:
        raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="STATE_CONFLICT")
    cleaned_reason = clean_text(reason, 2000)
    if not cleaned_reason:
        raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_ENTITY, detail="REASON_REQUIRED")
    old_status = item.status
    item.status = COST_STATUS_DRAFT
    db.flush()
    _write_status_history(db, item, user, old_status, cleaned_reason)
    return item


def archive_cost_item(db: Session, user: User, item_id: int, reason: str | None = None) -> CostItem:
    require_cost_db_approver(user)
    item = _get_item(db, item_id)
    if item.status == COST_STATUS_ARCHIVED:
        return item
    cleaned_reason = clean_text(reason, 2000)
    if item.status == COST_STATUS_ACTIVE and not cleaned_reason:
        raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_ENTITY, detail="REASON_REQUIRED")
    old_status = item.status
    item.status = COST_STATUS_ARCHIVED
    db.flush()
    _write_status_history(db, item, user, old_status, cleaned_reason or "archived draft")
    return item


def _duplicate_key(data: Mapping[str, Any]) -> tuple[str, str, str, str, str]:
    return (
        str(data.get("category") or ""),
        str(data.get("subcategory") or ""),
        str(data.get("item_name") or ""),
        str(data.get("spec") or ""),
        str(data.get("unit") or ""),
    )


def _existing_by_key(db: Session, data: Mapping[str, Any]) -> CostItem | None:
    return (
        db.query(CostItem)
        .filter(
            CostItem.category == data.get("category"),
            CostItem.subcategory == data.get("subcategory"),
            CostItem.item_name == data.get("item_name"),
            CostItem.spec == data.get("spec"),
            CostItem.unit == data.get("unit"),
        )
        .order_by(CostItem.id.desc())
        .first()
    )


def _make_notes(code: str | None, calc_rule: Any, work_content: Any, client_note: Any, labor_note: Any) -> str | None:
    parts = []
    if code:
        parts.append(f"编号: {code}")
    for label, value in (
        ("计算规则", calc_rule),
        ("工作内容", work_content),
        ("对甲备注", client_note),
        ("劳务备注", labor_note),
    ):
        text = clean_text(value)
        if text:
            parts.append(f"{label}: {text}")
    return "\n".join(parts) if parts else None


def parse_cost_workbook(content: bytes) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    try:
        workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:
        raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_ENTITY, detail="INVALID_EXCEL") from exc

    sheet = workbook.worksheets[0]
    current_category: str | None = None
    items: list[dict[str, Any]] = []
    skipped_rows: list[dict[str, Any]] = []

    for row_number, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        values = list(row) + [None] * 19
        code = clean_text(values[0], 64)
        item_name = clean_text(values[1], 255)
        if code and not item_name and "章" in code:
            current_category = code
            continue
        if not code or code in {"编号", "序号"} or not item_name:
            continue

        unit = clean_text(values[5], 64)
        client_labor_price = parse_float(values[6])
        client_main_material_price = parse_float(values[7])
        client_auxiliary_material_price = parse_float(values[8])
        client_direct_fee = parse_float(values[9])
        client_management_profit = parse_float(values[10])
        client_tax_excluded_price = parse_float(values[11])
        subcontract_labor_price = parse_float(values[13])
        subcontract_main_material_price = parse_float(values[14])
        subcontract_auxiliary_material_price = parse_float(values[15])
        subcontract_composite_price = parse_float(values[16])
        crew_benchmark_price = parse_float(values[17])
        if not unit or not any(value is not None for value in (client_tax_excluded_price, subcontract_composite_price, crew_benchmark_price)):
            skipped_rows.append({"row": row_number, "reason": "missing_unit_or_price", "code": code, "item_name": item_name})
            continue

        raw_item = {
            "category": current_category or sheet.title,
            "subcategory": None,
            "item_name": item_name,
            "spec": clean_text(values[2]),
            "unit": unit,
            "client_tax_excluded_price": client_tax_excluded_price,
            "client_labor_price": client_labor_price,
            "client_main_material_price": client_main_material_price,
            "client_auxiliary_material_price": client_auxiliary_material_price,
            "client_direct_fee": client_direct_fee,
            "client_management_profit": client_management_profit,
            "subcontract_composite_price": subcontract_composite_price,
            "subcontract_labor_price": subcontract_labor_price,
            "subcontract_main_material_price": subcontract_main_material_price,
            "subcontract_auxiliary_material_price": subcontract_auxiliary_material_price,
            "crew_benchmark_price": crew_benchmark_price,
            "price_type": PRICE_TYPE_COMBINED,
            "source": COST_SOURCE_IMPORTED,
            "effective_date": None,
            "notes": _make_notes(code, values[3], values[4], values[12], values[18]),
            "source_row": row_number,
            "source_sheet": sheet.title,
        }
        try:
            normalized = _normalize_item_payload(raw_item, for_import=True)
        except HTTPException as exc:
            skipped_rows.append({"row": row_number, "reason": str(exc.detail), "code": code, "item_name": item_name})
            continue
        normalized["source_row"] = row_number
        normalized["source_sheet"] = sheet.title
        items.append(normalized)

    return items, skipped_rows


def build_import_preview(db: Session, user: User, content: bytes) -> ImportBatch:
    require_cost_db_editor(user)
    _cleanup_import_batches()
    items, skipped_rows = parse_cost_workbook(content)
    duplicate_warnings: list[dict[str, Any]] = []
    seen: set[tuple[str, str, str, str, str]] = set()
    for index, item in enumerate(items):
        key = _duplicate_key(item)
        if key in seen:
            duplicate_warnings.append({"index": index, "type": "within_file", "item_name": item["item_name"], "spec": item.get("spec")})
            continue
        seen.add(key)
        existing = _existing_by_key(db, item)
        if existing:
            duplicate_warnings.append(
                {
                    "index": index,
                    "type": f"existing_{existing.status}",
                    "cost_item_id": existing.id,
                    "item_name": item["item_name"],
                    "spec": item.get("spec"),
                }
            )
    batch_id = str(uuid4())
    batch = ImportBatch(
        batch_id=batch_id,
        created_at=datetime.now(timezone.utc),
        items=items,
        duplicate_warnings=duplicate_warnings,
        skipped_rows=skipped_rows,
    )
    IMPORT_BATCHES[batch_id] = batch
    return batch


def _cleanup_import_batches() -> None:
    now = datetime.now(timezone.utc)
    expired = [batch_id for batch_id, batch in IMPORT_BATCHES.items() if now - batch.created_at > IMPORT_BATCH_TTL]
    for batch_id in expired:
        IMPORT_BATCHES.pop(batch_id, None)


def _get_import_batch(batch_id: str) -> ImportBatch:
    _cleanup_import_batches()
    batch = IMPORT_BATCHES.get(batch_id)
    if not batch:
        raise HTTPException(status_code=status.HTTP_410_GONE, detail="BATCH_EXPIRED")
    return batch


def import_preview_response(batch: ImportBatch) -> dict[str, Any]:
    return {
        "batch_id": batch.batch_id,
        "expires_in_seconds": int(IMPORT_BATCH_TTL.total_seconds()),
        "items": batch.items,
        "duplicate_warnings": batch.duplicate_warnings,
        "skipped_rows": batch.skipped_rows,
        "item_count": len(batch.items),
    }


def confirm_import_batch(db: Session, user: User, batch_id: str) -> dict[str, Any]:
    require_cost_db_editor(user)
    batch = _get_import_batch(batch_id)
    if batch.confirmed and batch.result is not None:
        return batch.result

    created: list[int] = []
    updated: list[int] = []
    skipped: list[dict[str, Any]] = []
    seen: set[tuple[str, str, str, str, str]] = set()
    for item_data in batch.items:
        key = _duplicate_key(item_data)
        if key in seen:
            skipped.append({"item_name": item_data["item_name"], "reason": "duplicate_within_file"})
            continue
        seen.add(key)
        existing = _existing_by_key(db, item_data)
        clean_data = {key_name: value for key_name, value in item_data.items() if key_name in UPDATE_FIELDS}
        if existing and existing.status == COST_STATUS_ACTIVE:
            skipped.append({"cost_item_id": existing.id, "item_name": item_data["item_name"], "reason": "active_duplicate"})
            continue
        if existing and existing.status == COST_STATUS_DRAFT:
            old_prices = {field: getattr(existing, field) for field in PRICE_FIELDS}
            for field, value in clean_data.items():
                setattr(existing, field, value)
            existing.source = COST_SOURCE_IMPORTED
            db.flush()
            _write_price_history(db, existing, user, old_prices, "import draft overwrite")
            updated.append(existing.id)
            continue
        new_item = CostItem(**clean_data, status=COST_STATUS_DRAFT, created_by=user.id)
        db.add(new_item)
        db.flush()
        created.append(new_item.id)

    batch.confirmed = True
    batch.result = {
        "batch_id": batch.batch_id,
        "created_count": len(created),
        "updated_count": len(updated),
        "skipped_count": len(skipped),
        "created_ids": created,
        "updated_ids": updated,
        "skipped": skipped,
    }
    return batch.result
