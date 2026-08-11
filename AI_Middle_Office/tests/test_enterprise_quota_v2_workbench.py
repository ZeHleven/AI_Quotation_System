from io import BytesIO

import pytest
from openpyxl import Workbook, load_workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker

from app.core.database import Base
from app.models.enterprise_quota import (
    EnterpriseCostResource,
    EnterpriseQuotaComponent,
    EnterpriseQuotaItem,
    EnterpriseQuotaSection,
    EnterpriseQuotaSheetRow,
)
from app.services.enterprise_quota_v2_parser import (
    ENTERPRISE_HEADERS,
    ENTERPRISE_SHEET,
    LABOR_HEADERS,
    LABOR_SHEET,
    MATERIAL_HEADERS,
    MATERIAL_SHEET,
    VALIDATION_HEADERS,
    VALIDATION_SHEET,
    _classify_row,
    parse_enterprise_quota_v2_bytes,
)
from app.services.enterprise_quota_v2_workbench import (
    EnterpriseQuotaV2WorkbenchError,
    activate_version,
    clone_version_to_draft,
    import_v2_workbook_as_draft,
    list_sheet_rows,
    reparse_draft_version_from_stored_rows,
    update_resource,
)


def _set_headers(sheet, headers):
    for index, header in enumerate(headers, start=1):
        sheet.cell(2, index, header)


def _build_v2_workbook(*, with_component_outlines: bool = True) -> bytes:
    workbook = Workbook()
    enterprise = workbook.active
    enterprise.title = ENTERPRISE_SHEET
    labor = workbook.create_sheet(LABOR_SHEET)
    material = workbook.create_sheet(MATERIAL_SHEET)
    validation = workbook.create_sheet(VALIDATION_SHEET)

    enterprise["A1"] = "旗胜企业定额2.0测试"
    labor["A1"] = "人工价格库"
    material["A1"] = "材料价格库"
    validation["A1"] = "消耗量校验报告"
    _set_headers(enterprise, ENTERPRISE_HEADERS)
    _set_headers(labor, LABOR_HEADERS)
    _set_headers(material, MATERIAL_HEADERS)
    _set_headers(validation, VALIDATION_HEADERS)

    enterprise["A3"] = "一、土建类"
    enterprise["A4"] = "第一章 测试工程"
    enterprise.append(
        [
            "QS-001",
            "定额",
            "测试定额",
            "工作内容",
            "规格A",
            "旗胜",
            "项",
            1,
            "=SUM(J6:M7)",
            "=SUM(J6:J7)",
            "=SUM(K6:K7)",
            "=SUM(L6:L7)",
            "=SUM(M6:M7)",
        ]
    )
    enterprise.append(
        [
            '=IFERROR(INDEX(人工价格库!$A:$A,MATCH(C6,人工价格库!$C:$C,0)),"")',
            '=IFERROR(INDEX(人工价格库!$B:$B,MATCH(C6,人工价格库!$C:$C,0)),"")',
            "测试人工",
            '=IFERROR(INDEX(人工价格库!$D:$D,MATCH(C6,人工价格库!$C:$C,0)),"")',
            '=IFERROR(INDEX(人工价格库!$E:$E,MATCH(C6,人工价格库!$C:$C,0)),"")',
            "",
            '=IFERROR(INDEX(人工价格库!$F:$F,MATCH(C6,人工价格库!$C:$C,0)),"")',
            2,
            '=IFERROR(INDEX(人工价格库!$H:$H,MATCH(C6,人工价格库!$C:$C,0)),0)',
            "=H6*I6",
            0,
            0,
            0,
        ]
    )
    enterprise.append(
        [
            '=IFERROR(INDEX(材料价格库!$B:$B,MATCH(C7,材料价格库!$D:$D,0)),"")',
            '=IFERROR(INDEX(材料价格库!$C:$C,MATCH(C7,材料价格库!$D:$D,0)),"")',
            "测试辅材",
            "",
            '=IFERROR(INDEX(材料价格库!$E:$E,MATCH(C7,材料价格库!$D:$D,0)),"")',
            '=IFERROR(INDEX(材料价格库!$F:$F,MATCH(C7,材料价格库!$D:$D,0)),"")',
            '=IFERROR(INDEX(材料价格库!$G:$G,MATCH(C7,材料价格库!$D:$D,0)),"")',
            3,
            '=IFERROR(INDEX(材料价格库!$H:$H,MATCH(C7,材料价格库!$D:$D,0)),0)',
            0,
            0,
            "=H7*I7",
            0,
        ]
    )
    if with_component_outlines:
        enterprise.row_dimensions[6].outlineLevel = 1
        enterprise.row_dimensions[7].outlineLevel = 1
    else:
        # Real-world workbooks can omit grouping while retaining the visible
        # component type values cached by Excel.
        enterprise["B6"] = "人工"
        enterprise["B7"] = "辅材"

    labor.append(["RG-001", "人工", "测试人工", "测试工作", "按含量计算", "工日", 1, 10])
    material.append(["辅材", "CL-001", "辅材", "测试辅材", "型号A", "测试品牌", "kg", 4])
    validation.append(["QS-001", "土建", "第一章", "测试定额", "项", 20, 0, 12, 0, 32])

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def _build_classified_v2_workbook() -> bytes:
    workbook = load_workbook(BytesIO(_build_v2_workbook()))
    enterprise = workbook[ENTERPRISE_SHEET]
    enterprise["A4"] = "1.2 砌筑抹灰工程"
    enterprise.append(["二、装饰工程"])
    enterprise.append(["2.1 拆除工程"])
    enterprise.append(
        [
            "ZS-001",
            "定额",
            "装饰拆除测试",
            "拆除并清运",
            "",
            "",
            "㎡",
            1,
            20,
            20,
            0,
            0,
            0,
        ]
    )
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


@pytest.fixture()
def db_session(tmp_path):
    engine = create_engine(f"sqlite:///{(tmp_path / 'enterprise_quota_v2.db').as_posix()}")
    Base.metadata.create_all(bind=engine)
    session = sessionmaker(bind=engine)()
    try:
        yield session
    finally:
        session.close()
        engine.dispose()


def test_parser_preserves_headers_hierarchy_formulas_and_sheet_rows():
    parsed = parse_enterprise_quota_v2_bytes(_build_v2_workbook(), filename="企业定额2.0.xlsx")

    assert parsed["summary"]["sheet_count"] == 4
    assert parsed["summary"]["major_section_count"] == 1
    assert parsed["summary"]["chapter_count"] == 1
    assert parsed["summary"]["quota_item_count"] == 1
    assert parsed["summary"]["component_count"] == 2
    assert parsed["summary"]["resource_count"] == 2
    assert parsed["summary"]["formula_count"] == 19
    assert parsed["quality"]["status"] == "ready"
    assert parsed["items"][0]["specification"] == "规格A"
    assert parsed["items"][0]["brand"] == "旗胜"
    assert parsed["items"][0]["unit_price"] == 32.0
    assert parsed["components"][0]["formula_link_status"] == "linked"

    formula_row = next(
        row
        for row in parsed["workbook_rows"]
        if row["sheet_name"] == ENTERPRISE_SHEET and row["row_number"] == 6
    )
    assert formula_row["outline_level"] == 1
    assert formula_row["parent_row_number"] == 5
    assert formula_row["formulas"]["I"].startswith("=IFERROR(INDEX(人工价格库!")


@pytest.mark.parametrize("component_type", ["人工", "主材", "辅材", "机械"])
def test_row_classifier_recognizes_component_types_without_outline(component_type):
    parent_context = {
        "major_section": 3,
        "chapter": 4,
        "quota_item": 5,
    }

    row_kind, parent_row_number = _classify_row(
        ENTERPRISE_SHEET,
        6,
        {"A": "RESOURCE-001", "B": component_type, "C": "测试工料机"},
        0,
        parent_context,
    )

    assert row_kind == "component"
    assert parent_row_number == 5
    assert parent_context["quota_item"] == 5


def test_parser_recognizes_components_when_excel_outline_is_missing():
    parsed = parse_enterprise_quota_v2_bytes(
        _build_v2_workbook(with_component_outlines=False),
        filename="企业定额2.0-无大纲层级.xlsx",
    )

    assert parsed["summary"]["chapter_count"] == 1
    assert parsed["summary"]["quota_item_count"] == 1
    assert parsed["summary"]["component_count"] == 2
    assert parsed["items"][0]["unit_price"] == 32.0
    component_rows = [
        row
        for row in parsed["workbook_rows"]
        if row["sheet_name"] == ENTERPRISE_SHEET and row["row_number"] in {6, 7}
    ]
    assert [row["row_kind"] for row in component_rows] == ["component", "component"]
    assert [row["parent_row_number"] for row in component_rows] == [5, 5]


def test_reparse_draft_restores_components_from_preserved_rows(db_session):
    version = import_v2_workbook_as_draft(
        db_session,
        _build_v2_workbook(with_component_outlines=False),
        filename="企业定额2.0-待修复.xlsx",
        actor_id=None,
        version_code="enterprise-quota-v2-reparse",
    )
    db_session.flush()

    component_rows = (
        db_session.query(EnterpriseQuotaSheetRow)
        .filter(
            EnterpriseQuotaSheetRow.version_id == version.id,
            EnterpriseQuotaSheetRow.sheet_name == ENTERPRISE_SHEET,
            EnterpriseQuotaSheetRow.row_number.in_([6, 7]),
        )
        .all()
    )
    for row in component_rows:
        row.row_kind = "chapter"
        row.parent_row_number = None
        row.entity_type = None
        row.entity_id = None
    db_session.query(EnterpriseQuotaComponent).filter_by(version_id=version.id).delete(
        synchronize_session=False
    )
    item = db_session.query(EnterpriseQuotaItem).filter_by(version_id=version.id).one()
    item.unit_price = 0
    item.labor_fee = 0
    item.auxiliary_material_fee = 0
    db_session.flush()

    result = reparse_draft_version_from_stored_rows(
        db_session,
        version.id,
        actor_id=None,
        expected_revision=version.revision,
        reason="测试按新规则重新解析",
    )

    restored_item = db_session.query(EnterpriseQuotaItem).filter_by(version_id=version.id).one()
    restored_components = (
        db_session.query(EnterpriseQuotaComponent)
        .filter_by(version_id=version.id)
        .order_by(EnterpriseQuotaComponent.sort_order)
        .all()
    )
    restored_rows = (
        db_session.query(EnterpriseQuotaSheetRow)
        .filter(
            EnterpriseQuotaSheetRow.version_id == version.id,
            EnterpriseQuotaSheetRow.sheet_name == ENTERPRISE_SHEET,
            EnterpriseQuotaSheetRow.row_number.in_([6, 7]),
        )
        .order_by(EnterpriseQuotaSheetRow.row_number)
        .all()
    )
    assert result["summary"]["component_count"] == 2
    assert result["recalculation"]["linked_component_count"] == 2
    assert restored_item.unit_price == 32
    assert len(restored_components) == 2
    assert [row.row_kind for row in restored_rows] == ["component", "component"]
    assert [row.parent_row_number for row in restored_rows] == [5, 5]
    assert all(row.entity_type == "component" and row.entity_id for row in restored_rows)


def test_enterprise_rows_support_linked_major_and_chapter_filters(db_session):
    version = import_v2_workbook_as_draft(
        db_session,
        _build_classified_v2_workbook(),
        filename="企业定额2.0-分类筛选.xlsx",
        actor_id=None,
        version_code="enterprise-quota-v2-classified",
    )
    db_session.flush()
    sections = (
        db_session.query(EnterpriseQuotaSection)
        .filter_by(version_id=version.id)
        .order_by(EnterpriseQuotaSection.sort_order.asc())
        .all()
    )
    civil_major = next(section for section in sections if section.section_name == "一、土建类")
    civil_chapter = next(section for section in sections if section.section_name == "1.2 砌筑抹灰工程")
    decoration_major = next(section for section in sections if section.section_name == "二、装饰工程")
    demolition_chapter = next(section for section in sections if section.section_name == "2.1 拆除工程")

    all_rows = list_sheet_rows(
        db_session,
        version.id,
        sheet_key="enterprise",
        page=1,
        page_size=500,
    )
    assert [option["label"] for option in all_rows["classification"]["major_sections"]] == [
        "一、土建类",
        "二、装饰工程",
    ]
    assert [option["label"] for option in all_rows["classification"]["chapters"]] == [
        "1.2 砌筑抹灰工程",
        "2.1 拆除工程",
    ]

    civil_rows = list_sheet_rows(
        db_session,
        version.id,
        sheet_key="enterprise",
        major_section_id=civil_major.id,
        page=1,
        page_size=500,
    )
    civil_text = " ".join(
        str(value or "")
        for row in civil_rows["rows"]
        for value in row["values"].values()
    )
    assert civil_rows["rows"][0]["row_kind"] == "major_section"
    assert "测试定额" in civil_text
    assert "装饰拆除测试" not in civil_text

    demolition_rows = list_sheet_rows(
        db_session,
        version.id,
        sheet_key="enterprise",
        major_section_id=decoration_major.id,
        chapter_id=demolition_chapter.id,
        page=1,
        page_size=500,
    )
    demolition_text = " ".join(
        str(value or "")
        for row in demolition_rows["rows"]
        for value in row["values"].values()
    )
    assert demolition_rows["classification"]["selected_major_section_id"] == decoration_major.id
    assert demolition_rows["classification"]["selected_chapter_id"] == demolition_chapter.id
    assert demolition_rows["rows"][0]["row_kind"] == "chapter"
    assert "装饰拆除测试" in demolition_text
    assert "测试定额" not in demolition_text

    with pytest.raises(EnterpriseQuotaV2WorkbenchError) as exc_info:
        list_sheet_rows(
            db_session,
            version.id,
            sheet_key="enterprise",
            major_section_id=civil_major.id,
            chapter_id=demolition_chapter.id,
        )
    assert exc_info.value.code == "SECTION_FILTER_MISMATCH"
    assert civil_chapter.parent_section_id == civil_major.id


def test_price_change_recalculates_linked_component_item_and_excel_mirror(db_session):
    version = import_v2_workbook_as_draft(
        db_session,
        _build_v2_workbook(),
        filename="企业定额2.0.xlsx",
        actor_id=None,
        version_code="enterprise-quota-v2-test",
    )
    db_session.flush()

    item = db_session.query(EnterpriseQuotaItem).filter_by(version_id=version.id).one()
    labor = (
        db_session.query(EnterpriseCostResource)
        .filter_by(version_id=version.id, library_kind="labor")
        .one()
    )
    assert item.unit_price == 32.0

    result = update_resource(
        db_session,
        version.id,
        labor.id,
        {"price": 15},
        actor_id=None,
        expected_revision=version.revision,
        reason="专项测试人工调价",
    )
    db_session.flush()

    db_session.refresh(item)
    labor_component = (
        db_session.query(EnterpriseQuotaComponent)
        .filter_by(version_id=version.id, formula_library_kind="labor")
        .one()
    )
    item_row = (
        db_session.query(EnterpriseQuotaSheetRow)
        .filter_by(version_id=version.id, entity_type="quota_item", entity_id=item.id)
        .one()
    )
    component_row = (
        db_session.query(EnterpriseQuotaSheetRow)
        .filter_by(
            version_id=version.id,
            entity_type="component",
            entity_id=labor_component.id,
        )
        .one()
    )

    assert result["recalculation"]["linked_component_count"] == 2
    assert labor_component.unit_price == 15.0
    assert labor_component.amount == 30.0
    assert item.labor_fee == 30.0
    assert item.auxiliary_material_fee == 12.0
    assert item.unit_price == 42.0
    assert '"I":42.0' in item_row.values_json
    assert '"J":30.0' in component_row.values_json
    assert '"K":null' in component_row.values_json
    assert '"L":null' in component_row.values_json
    assert "人工价格库" in component_row.formulas_json


def test_activation_gate_and_clone_keep_active_version_immutable(db_session):
    version = import_v2_workbook_as_draft(
        db_session,
        _build_v2_workbook(),
        filename="企业定额2.0.xlsx",
        actor_id=None,
        version_code="enterprise-quota-v2-activate",
    )
    active = activate_version(
        db_session,
        version.id,
        actor_id=None,
        expected_revision=version.revision,
        reason="专项测试启用",
        acknowledge_warnings=False,
    )
    assert active.status == "active"
    assert active.is_active is True

    labor = (
        db_session.query(EnterpriseCostResource)
        .filter_by(version_id=active.id, library_kind="labor")
        .one()
    )
    with pytest.raises(EnterpriseQuotaV2WorkbenchError) as exc_info:
        update_resource(
            db_session,
            active.id,
            labor.id,
            {"price": 16},
            actor_id=None,
            expected_revision=active.revision,
            reason="不允许直接修改生效版本",
        )
    assert exc_info.value.code == "DRAFT_REQUIRED"

    clone = clone_version_to_draft(
        db_session,
        active.id,
        actor_id=None,
        version_code="enterprise-quota-v2-clone",
        version_name="企业定额2.0编辑草稿",
        reason="从生效版本创建草稿",
    )
    assert clone.status == "draft"
    assert clone.is_active is False
    assert (
        db_session.query(EnterpriseQuotaSheetRow)
        .filter_by(version_id=clone.id)
        .count()
        == db_session.query(EnterpriseQuotaSheetRow).filter_by(version_id=active.id).count()
    )
