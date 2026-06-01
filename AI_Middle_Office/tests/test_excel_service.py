import base64
from io import BytesIO

from openpyxl import load_workbook

from app.services.excel_service import build_excel_base64


def test_build_excel_base64_appends_total_quote_amount():
    excel_b64 = build_excel_base64(
        [
            {"project_name": "墙面刷新", "unit_price": 20, "total_price": 200, "notes": "标准工艺"},
            {"project_name": "灯具安装", "unit_price": 35.5, "total_price": "1,065.00", "notes": "含辅材"},
        ]
    )

    workbook = load_workbook(BytesIO(base64.b64decode(excel_b64)))
    sheet = workbook.active

    assert sheet.cell(row=4, column=1).value == "总合计报价金额"
    assert sheet.cell(row=4, column=3).value == 1265
    assert sheet.cell(row=4, column=3).number_format == "#,##0.00"
