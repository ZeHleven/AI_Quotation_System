from __future__ import annotations

import asyncio
import base64
import csv
import json
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any, Awaitable, Callable, Iterable, Mapping, Sequence

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.core.config import settings
from app.services.model_gateway import call_glm_drawing_tile_extract


PHASE = "BIZ-2x-pdf-gap-recall-visual-runner"

CALL_HEADERS = [
    "call_no",
    "status",
    "recommended_pass",
    "source_file",
    "page",
    "tile_id",
    "tile_type",
    "image_path",
    "covered_task_nos",
    "covered_gap_nos",
    "covered_answer_items",
    "task_count",
    "reason",
    "raw_content",
]

EVIDENCE_HEADERS = [
    "evidence_id",
    "source_kind",
    "source_file",
    "page",
    "tile_id",
    "vision_pass",
    "evidence_role",
    "discipline",
    "item_hint",
    "space",
    "material_codes",
    "spec_or_method",
    "suggested_unit",
    "text",
    "normalized_text",
    "confidence",
    "model",
    "needs_manual_review",
    "reason",
]

VisionClient = Callable[..., Awaitable[Mapping[str, Any]]]


def run_gap_recall_plan(
    recall_plan: Mapping[str, Any],
    *,
    execute: bool = False,
    max_calls: int | None = None,
    vision_client: VisionClient | None = None,
    trace_id: str | None = None,
) -> dict[str, Any]:
    unique_calls = _dedupe_plan_rows(recall_plan.get("plan_rows") or [])
    if max_calls is not None:
        unique_calls = unique_calls[: max(max_calls, 0)]

    call_rows: list[dict[str, Any]] = []
    evidence_rows: list[dict[str, Any]] = []
    client = vision_client or call_glm_drawing_tile_extract
    for call_index, call in enumerate(unique_calls, start=1):
        call_row = {
            "call_no": call_index,
            "status": "planned_dry_run",
            "recommended_pass": call["recommended_pass"],
            "source_file": call["source_file"],
            "page": call["page"],
            "tile_id": call["tile_id"],
            "tile_type": call["tile_type"],
            "image_path": call["image_path"],
            "covered_task_nos": "；".join(str(item) for item in call["task_nos"]),
            "covered_gap_nos": "；".join(str(item) for item in call["gap_nos"]),
            "covered_answer_items": "；".join(call["answer_items"]),
            "task_count": len(call["task_rows"]),
            "reason": "dry_run_no_external_model_call",
            "raw_content": "",
        }
        if execute:
            _execute_call(
                call,
                call_row=call_row,
                evidence_rows=evidence_rows,
                vision_client=client,
                trace_id=trace_id,
            )
        call_rows.append(call_row)

    summary = {
        "execute": execute,
        "plan_task_count": len(recall_plan.get("plan_rows") or []),
        "unique_visual_call_count": len(unique_calls),
        "call_count": len(call_rows),
        "covered_gap_count": len({gap for call in unique_calls for gap in call["gap_nos"]}),
        "evidence_count": len(evidence_rows),
        "status_counts": dict(Counter(str(row.get("status") or "") for row in call_rows)),
        "pass_counts": dict(Counter(str(row.get("recommended_pass") or "") for row in call_rows)),
        "missing_image_call_count": sum(1 for row in call_rows if row.get("status") == "skipped_missing_image"),
    }
    return {
        "ok": True,
        "phase": PHASE,
        "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "summary": summary,
        "call_rows": call_rows,
        "evidence_rows": evidence_rows,
        "visual_evidence_report": {"evidence_rows": evidence_rows},
    }


def write_gap_recall_run_outputs(
    report: Mapping[str, Any],
    output_dir: str | Path,
    *,
    stem: str | None = None,
) -> dict[str, str]:
    target = Path(output_dir)
    target.mkdir(parents=True, exist_ok=True)
    file_stem = stem or f"BIZ2x_PDF_gap_recall_run_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
    json_path = target / f"{file_stem}.json"
    call_csv = target / f"{file_stem}_calls.csv"
    evidence_csv = target / f"{file_stem}_evidence.csv"
    markdown_path = target / f"{file_stem}.md"
    xlsx_path = target / f"{file_stem}.xlsx"
    outputs = {
        "json": str(json_path),
        "call_csv": str(call_csv),
        "evidence_csv": str(evidence_csv),
        "markdown": str(markdown_path),
        "xlsx": str(xlsx_path),
    }
    report_with_outputs = {**dict(report), "outputs": outputs}
    json_path.write_text(json.dumps(_json_safe(report_with_outputs), ensure_ascii=False, indent=2), encoding="utf-8")
    _write_csv(call_csv, report.get("call_rows") or [], CALL_HEADERS)
    _write_csv(evidence_csv, report.get("evidence_rows") or [], EVIDENCE_HEADERS)
    markdown_path.write_text(_build_markdown(report_with_outputs), encoding="utf-8")
    _write_workbook(xlsx_path, report_with_outputs)
    return outputs


def _dedupe_plan_rows(plan_rows: Sequence[Mapping[str, Any]]) -> list[dict[str, Any]]:
    grouped: dict[tuple[str, str, str, str], list[Mapping[str, Any]]] = defaultdict(list)
    for row in plan_rows:
        key = (
            str(row.get("recommended_pass") or ""),
            str(row.get("image_path") or ""),
            str(row.get("source_file") or ""),
            str(row.get("tile_id") or ""),
        )
        grouped[key].append(row)
    calls: list[dict[str, Any]] = []
    for rows in grouped.values():
        first = rows[0]
        calls.append(
            {
                "recommended_pass": str(first.get("recommended_pass") or ""),
                "source_file": str(first.get("source_file") or ""),
                "page": first.get("page", ""),
                "tile_id": str(first.get("tile_id") or ""),
                "tile_type": str(first.get("tile_type") or ""),
                "image_path": str(first.get("image_path") or ""),
                "task_nos": _unique(row.get("task_no") for row in rows),
                "gap_nos": _unique(row.get("gap_no") for row in rows),
                "answer_items": _unique(row.get("answer_item_name") for row in rows),
                "task_rows": [dict(row) for row in rows],
            }
        )
    calls.sort(key=lambda row: (str(row["recommended_pass"]), str(row["source_file"]), str(row["tile_id"])))
    return calls


def _execute_call(
    call: Mapping[str, Any],
    *,
    call_row: dict[str, Any],
    evidence_rows: list[dict[str, Any]],
    vision_client: VisionClient,
    trace_id: str | None,
) -> None:
    image_path = Path(str(call.get("image_path") or ""))
    if not image_path.exists() or not image_path.is_file():
        call_row["status"] = "skipped_missing_image"
        call_row["reason"] = "image_path_missing"
        return
    try:
        model_result = asyncio.run(
            vision_client(
                base64.b64encode(image_path.read_bytes()).decode("ascii"),
                _image_mime_type(image_path),
                tile_context={
                    "source_file": call.get("source_file", ""),
                    "page": call.get("page", ""),
                    "tile_id": call.get("tile_id", ""),
                    "tile_type": call.get("tile_type", ""),
                    "gap_nos": list(call.get("gap_nos") or []),
                    "answer_items": list(call.get("answer_items") or []),
                },
                prompt_mode=str(call.get("recommended_pass") or "general"),
                trace_id=trace_id,
            )
        )
    except Exception as exc:  # pragma: no cover - concrete exceptions vary by gateway/network
        call_row["status"] = "error"
        call_row["reason"] = str(exc)[:300]
        return

    items = model_result.get("evidence_items") or []
    before_count = len(evidence_rows)
    for item in items:
        text = str(item.get("text") or "").strip()
        if not text:
            continue
        evidence_rows.append(
            {
                "evidence_id": f"PDFGAP-{len(evidence_rows) + 1:06d}",
                "source_kind": "pdf_gap_recall_llm",
                "source_file": call.get("source_file", ""),
                "page": call.get("page", ""),
                "tile_id": call.get("tile_id", ""),
                "vision_pass": call.get("recommended_pass", ""),
                "evidence_role": str(item.get("evidence_role") or "").strip() or "unknown_note",
                "discipline": str(item.get("discipline") or "").strip() or "unknown",
                "item_hint": str(item.get("item_hint") or "").strip(),
                "space": str(item.get("space") or "").strip(),
                "material_codes": _normalize_string_list(item.get("material_codes")),
                "spec_or_method": str(item.get("spec_or_method") or "").strip(),
                "suggested_unit": str(item.get("suggested_unit") or "").strip(),
                "text": text,
                "normalized_text": str(item.get("normalized_text") or "").strip() or text,
                "confidence": _bounded_float(item.get("confidence")),
                "model": settings.glm_vision_model,
                "needs_manual_review": bool(item.get("needs_manual_review", True)),
                "reason": str(item.get("reason") or "").strip(),
            }
        )
    call_row["status"] = "success"
    call_row["reason"] = f"extracted_{len(evidence_rows) - before_count}_items"
    call_row["raw_content"] = str(model_result.get("raw_content") or "")[:500]


def _image_mime_type(path: Path) -> str:
    suffix = path.suffix.lower()
    if suffix in {".jpg", ".jpeg"}:
        return "image/jpeg"
    if suffix == ".webp":
        return "image/webp"
    return "image/png"


def _normalize_string_list(value: Any) -> list[str]:
    if isinstance(value, str):
        return [item.strip() for item in value.replace("；", ";").replace("，", ",").split(",") if item.strip()]
    if isinstance(value, Sequence):
        return [str(item).strip() for item in value if str(item).strip()]
    return []


def _unique(values: Iterable[Any]) -> list[Any]:
    result: list[Any] = []
    seen: set[str] = set()
    for value in values:
        key = str(value or "").strip()
        if not key or key in seen:
            continue
        seen.add(key)
        result.append(value)
    return result


def _bounded_float(value: Any) -> float:
    try:
        parsed = float(value)
    except (TypeError, ValueError):
        return 0.0
    return max(0.0, min(1.0, parsed))


def _build_markdown(report: Mapping[str, Any]) -> str:
    summary = report.get("summary") or {}
    lines = [
        "# BIZ-2x PDF Gap Recall Visual Run",
        "",
        f"- generated_at: {report.get('generated_at', '-')}",
        f"- execute: {summary.get('execute', False)}",
        f"- plan tasks: {summary.get('plan_task_count', 0)}",
        f"- unique visual calls: {summary.get('unique_visual_call_count', 0)}",
        f"- evidence rows: {summary.get('evidence_count', 0)}",
        "",
        "## Status Counts",
        "",
    ]
    for status, count in (summary.get("status_counts") or {}).items():
        lines.append(f"- {status}: {count}")
    lines.extend(["", "## Calls", "", "| status | pass | source | tile | gaps | items |", "| --- | --- | --- | --- | --- | --- |"])
    for row in (report.get("call_rows") or [])[:80]:
        lines.append(
            "| "
            + " | ".join(
                [
                    _md(row.get("status")),
                    _md(row.get("recommended_pass")),
                    _md(row.get("source_file")),
                    _md(row.get("tile_id")),
                    _md(row.get("covered_gap_nos")),
                    _md(row.get("covered_answer_items")),
                ]
            )
            + " |"
        )
    return "\n".join(lines) + "\n"


def _write_workbook(path: Path, report: Mapping[str, Any]) -> None:
    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "run_summary"
    _append_rows(summary_sheet, [["metric", "value"]])
    for key, value in (report.get("summary") or {}).items():
        _append_rows(summary_sheet, [[key, json.dumps(value, ensure_ascii=False) if isinstance(value, (dict, list)) else value]])
    _style_sheet(summary_sheet)

    call_sheet = workbook.create_sheet("visual_calls")
    _append_rows(call_sheet, [CALL_HEADERS])
    _append_rows(call_sheet, [[_cell_value(row.get(header)) for header in CALL_HEADERS] for row in report.get("call_rows") or []])
    _style_sheet(call_sheet)

    evidence_sheet = workbook.create_sheet("recall_evidence")
    _append_rows(evidence_sheet, [EVIDENCE_HEADERS])
    _append_rows(
        evidence_sheet,
        [[_cell_value(row.get(header)) for header in EVIDENCE_HEADERS] for row in report.get("evidence_rows") or []],
    )
    _style_sheet(evidence_sheet)
    workbook.save(path)


def _append_rows(sheet: Any, rows: Iterable[Sequence[Any]]) -> None:
    for row in rows:
        sheet.append(list(row))


def _style_sheet(sheet: Any) -> None:
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for col_idx, column_cells in enumerate(sheet.columns, start=1):
        values = [str(cell.value or "") for cell in column_cells[:80]]
        width = min(max([len(value) for value in values] + [10]) + 2, 80)
        sheet.column_dimensions[get_column_letter(col_idx)].width = width
    sheet.freeze_panes = "A2"


def _write_csv(path: Path, rows: Iterable[Mapping[str, Any]], fieldnames: Sequence[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        for row in rows:
            writer.writerow({key: _cell_value(row.get(key)) for key in fieldnames})


def _cell_value(value: Any) -> Any:
    if isinstance(value, (dict, list, tuple)):
        return json.dumps(value, ensure_ascii=False)
    return value


def _json_safe(value: Any) -> Any:
    if isinstance(value, Mapping):
        return {str(key): _json_safe(item) for key, item in value.items()}
    if isinstance(value, (list, tuple)):
        return [_json_safe(item) for item in value]
    if isinstance(value, Path):
        return str(value)
    return value


def _md(value: Any) -> str:
    return str(value or "").replace("|", "\\|").replace("\n", "<br>")
