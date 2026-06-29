from __future__ import annotations

import json
import sys
from pathlib import Path

from openpyxl import load_workbook

from app.services.drawing_pdf_external_recall_template_status import build_external_recall_template_status
from app.services.drawing_pdf_gap_recall_importer import (
    build_gap_recall_external_import_report,
    load_external_recall_results,
)
from app.services.drawing_pdf_object_recall_pack import (
    build_object_recall_pack,
    write_object_recall_pack_outputs,
)
from scripts import biz2x_pdf_object_recall_pack


def _review_report() -> dict[str, object]:
    return {
        "ok": True,
        "phase": "BIZ-2x-pdf-three-field-human-review",
        "summary": {"answer_count": 3, "matched_three_fields_count": 1},
        "review_rows": [
            {
                "review_no": 1,
                "status": "missing_candidate",
                "answer_sheet": "装修工程量清单",
                "answer_row_no": 7,
                "answer_section": "拆除工程",
                "answer_item_name": "拆除不锈钢玻璃门",
                "answer_feature": "拆除不锈钢地弹门，包括门套、门扇及五金拆除清运",
                "answer_unit": "套",
                "candidate_item_name": "拆除地面",
                "candidate_unit": "㎡",
                "candidate_source_files": "03.pdf",
                "evidence_pages": "1",
                "evidence_tiles": "p001_whole",
                "evidence_ids": "EV-001",
            },
            {
                "review_no": 2,
                "status": "matched_name_unit_feature_review",
                "answer_item_name": "灯具安装",
                "answer_feature": "LED 筒灯",
                "answer_unit": "套",
                "candidate_item_name": "灯具安装",
                "candidate_unit": "套",
            },
            {
                "review_no": 3,
                "status": "matched_three_fields",
                "answer_item_name": "墙布墙面",
                "answer_feature": "墙布",
                "answer_unit": "㎡",
                "candidate_item_name": "墙布墙面",
                "candidate_unit": "㎡",
            },
        ],
    }


def test_object_recall_pack_builds_object_tasks_without_evidence_backfill():
    pack = build_object_recall_pack(_review_report())

    assert pack["summary"]["object_recall_task_count"] == 1
    assert pack["summary"]["object_class_counts"] == {"door_window_demolition": 1}
    row = pack["recall_rows"][0]
    assert row["target_item_name"] == "拆除不锈钢玻璃门"
    assert row["object_class"] == "door_window_demolition"
    assert row["recommended_pass"] == "door_window_demolition"
    assert "不锈钢玻璃门" in row["target_object_terms"]
    assert row["evidence_item_hint"] == ""
    assert row["evidence_text"] == ""
    assert row["ready_for_import"] == "false"
    assert pack["answer_columns_count_as_evidence"] is False


def test_object_recall_pack_can_include_selected_review_statuses():
    pack = build_object_recall_pack(_review_report(), statuses=["missing_candidate", "matched_name_unit_feature_review"])

    assert pack["summary"]["object_recall_task_count"] == 2
    assert pack["summary"]["status_counts"]["matched_name_unit_feature_review"] == 1
    assert any(row["object_class"] == "electrical_mep" for row in pack["recall_rows"])


def test_object_recall_pack_writes_outputs(tmp_path: Path):
    pack = build_object_recall_pack(_review_report())
    outputs = write_object_recall_pack_outputs(pack, tmp_path / "out", stem="object_pack")

    assert Path(outputs["json"]).exists()
    assert Path(outputs["csv"]).exists()
    assert Path(outputs["markdown"]).exists()
    workbook = load_workbook(outputs["xlsx"])
    assert workbook.sheetnames == ["object_recall_summary", "object_recall_tasks"]
    assert workbook["object_recall_tasks"]["H2"].value == "拆除不锈钢玻璃门"
    assert workbook["object_recall_tasks"]["V2"].value is None


def test_blank_object_recall_pack_is_answer_reference_only(tmp_path: Path):
    pack = build_object_recall_pack(_review_report())
    outputs = write_object_recall_pack_outputs(pack, tmp_path / "out", stem="object_pack")

    external_results = load_external_recall_results(outputs["xlsx"])
    status = build_external_recall_template_status(external_results)
    import_report = build_gap_recall_external_import_report(
        external_results,
        recall_plan=_object_recall_plan(),
        source_name="blank-object-pack",
    )

    assert status["summary"]["input_row_count"] == 1
    assert status["summary"]["importable_row_count"] == 0
    assert status["summary"]["answer_only_count"] == 1
    assert status["summary"]["ready_for_external_import"] is False
    assert import_report["summary"]["evidence_count"] == 0
    assert import_report["summary"]["skipped_input_row_count"] == 1


def test_filled_object_recall_pack_imports_real_evidence(tmp_path: Path):
    pack = build_object_recall_pack(_review_report())
    outputs = write_object_recall_pack_outputs(pack, tmp_path / "out", stem="object_pack")
    workbook = load_workbook(outputs["xlsx"])
    sheet = workbook["object_recall_tasks"]
    headers = {cell.value: cell.column for cell in sheet[1]}
    sheet.cell(row=2, column=headers["evidence_item_hint"]).value = "stainless glass door demolition"
    sheet.cell(row=2, column=headers["evidence_spec_or_method"]).value = "remove stainless glass door, frame, leaf and hardware"
    sheet.cell(row=2, column=headers["evidence_suggested_unit"]).value = "set"
    sheet.cell(row=2, column=headers["evidence_text"]).value = "door schedule: stainless glass door demolition includes frame, leaf and hardware"
    workbook.save(outputs["xlsx"])
    workbook.close()

    external_results = load_external_recall_results(outputs["xlsx"])
    status = build_external_recall_template_status(external_results)
    import_report = build_gap_recall_external_import_report(
        external_results,
        recall_plan=_object_recall_plan(),
        source_name="filled-object-pack",
    )

    assert status["summary"]["importable_row_count"] == 1
    assert status["summary"]["ready_for_external_import"] is True
    assert import_report["summary"]["evidence_count"] == 1
    assert import_report["summary"]["validation_status_counts"] == {"imported": 1}
    row = import_report["evidence_rows"][0]
    assert row["source_file"] == "03.pdf"
    assert row["page"] == "1"
    assert row["tile_id"] == "p001_whole"
    assert row["vision_pass"] == "door_window_demolition"
    assert row["item_hint"] == "stainless glass door demolition"
    assert row["suggested_unit"] == "set"


def test_object_recall_pack_cli(tmp_path: Path, monkeypatch, capsys):
    review_json = tmp_path / "review.json"
    review_json.write_text(json.dumps(_review_report(), ensure_ascii=False), encoding="utf-8")
    output_dir = tmp_path / "cli_out"
    monkeypatch.setattr(
        sys,
        "argv",
        [
            "biz2x_pdf_object_recall_pack.py",
            "--review-json",
            str(review_json),
            "--output-dir",
            str(output_dir),
            "--timestamp",
            "fixed",
        ],
    )

    assert biz2x_pdf_object_recall_pack.main() == 0
    payload = json.loads(capsys.readouterr().out)

    assert payload["ok"] is True
    assert payload["summary"]["object_recall_task_count"] == 1
    assert payload["safe_to_import_without_evidence"] is False
    assert Path(payload["outputs"]["xlsx"]).exists()


def _object_recall_plan() -> dict[str, object]:
    return {
        "plan_rows": [
            {
                "task_no": 1,
                "gap_no": 1,
                "gap_priority": "P1_object_missing",
                "answer_item_name": "stainless glass door demolition",
                "recommended_pass": "door_window_demolition",
                "source_file": "03.pdf",
                "page": 1,
                "tile_id": "p001_whole",
                "tile_type": "whole_page_preview",
            }
        ]
    }
