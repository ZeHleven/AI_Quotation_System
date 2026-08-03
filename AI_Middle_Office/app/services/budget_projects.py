from __future__ import annotations

import copy
import hashlib
import json
import re
import uuid
from decimal import Decimal, InvalidOperation
from io import BytesIO
from pathlib import Path
from typing import Any

from fastapi import HTTPException, status
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string, get_column_letter
from sqlalchemy import or_
from sqlalchemy.orm import Session

from app.core.config import settings
from app.core.time_utils import app_local_naive
from app.models.account import AccountBudgetProject
from app.models.budget_project import (
    BUDGET_IMPORT_STATUS_ACTIVE,
    BUDGET_IMPORT_STATUS_CONFIRMED,
    BUDGET_IMPORT_STATUS_PARSED,
    BUDGET_IMPORT_STATUS_SUPERSEDED,
    BudgetProjectImportBatch,
    BudgetProjectImportLifecycleEvent,
    BudgetProjectImportRevision,
    BudgetProjectImportSheetMapping,
    BudgetProjectProfile,
    BudgetProjectStandardRow,
)
from app.models.file_object import FileObject
from app.models.project_progress import Project, ProjectTask
from app.models.user import User
from app.services.project_progress import can_manage_project, next_project_code
from app.services.account_tenancy import (
    AccountTenancyError,
    bind_budget_project_to_current_account,
    resolve_current_account,
)
from app.services.rbac import has_any_role
from app.services.requirement_standardizer import (
    MAX_COLUMNS_PER_SHEET,
    MAX_ROWS_PER_SHEET,
    RequirementStandardizationError,
    apply_manual_field_mappings,
    standardize_requirement_excel_bytes,
)


BUDGET_PROJECT_READ_ROLES = {
    "system_admin",
    "admin",
    "staff",
    "manager",
    "viewer",
    "project_viewer",
    "project_member",
    "project_manager",
    "quote_user",
    "quote_operator",
}
BUDGET_PROJECT_CREATE_ROLES = {
    "system_admin",
    "admin",
    "staff",
    "manager",
    "project_manager",
    "quote_user",
    "quote_operator",
}
BUDGET_PROJECT_WRITE_ROLES = BUDGET_PROJECT_CREATE_ROLES
BUDGET_PROJECT_VIEW_ALL_ROLES = {"system_admin", "admin", "manager", "viewer", "project_manager"}
_PROJECT_SCOPED_READ_ROLES = {
    "system_admin",
    "admin",
    "staff",
    "manager",
    "viewer",
    "project_viewer",
    "project_member",
    "project_manager",
}
BUDGET_WORKSPACE_ACTIVE = "active"
BUDGET_WORKSPACE_ARCHIVED = "archived"
BUDGET_SHEET_ROLE_BILL = "bill"
BUDGET_SHEET_ROLE_CALCULATION_RULE = "calculation_rule"
BUDGET_SHEET_ROLE_LOSS_REFERENCE = "loss_reference"
BUDGET_SHEET_ROLE_MATERIAL_REFERENCE = "material_reference"
BUDGET_SHEET_ROLE_METADATA = "metadata"
BUDGET_SHEET_ROLE_OPTIONAL_BACKUP = "optional_backup"
BUDGET_SHEET_ROLE_SUMMARY_ANALYSIS = "summary_analysis"
MAX_IMPORT_BYTES = 30 * 1024 * 1024
BUDGET_MAX_SCAN_ROWS = 30
SUPPORTED_EXTENSIONS = {".xlsx", ".xlsm"}

_SEQUENCE_LABELS = {
    "\u5e8f\u53f7",  # sequence number
    "\u7f16\u53f7",  # number/code
    "\u884c\u53f7",  # row number
    "\u5e8f\u5217",
    "\u5e8f\u5217\u53f7",
    "\u6e05\u5355\u7f16\u7801",
    "\u9879\u76ee\u7f16\u7801",
    "no",
    "number",
    "id",
}
_ABNORMAL_MARKERS = (
    "\u7ea6",  # approximately
    "\u5927\u7ea6",
    "\u5de6\u53f3",
    "\u4ee5\u4e0a",
    "\u4ee5\u4e0b",
    "\u4e0d\u5c11\u4e8e",
    "\u4e0d\u8d85\u8fc7",
    "~",
    "\uff5e",
)
_RANGE_RE = re.compile(r"\d\s*(?:-|\u2013|\u2014|~|\uff5e|\u81f3|\u5230)\s*\d")
_NUMBER_RE = re.compile(r"[+-]?(?:\d+(?:\.\d+)?|\.\d+)")
_STRICT_QUANTITY_RE = re.compile(
    r"^\s*([+]?(?:\d{1,3}(?:,\d{3})+|\d+)(?:\.\d+)?|[+]?\.\d+)\s*"
    r"(?:m|m2|m3|mm|cm|km|\u33a1|\u33a5|\u5e73\u65b9\u7c73|\u5e73\u7c73|"
    r"\u7acb\u65b9\u7c73|\u7acb\u65b9|\u7c73|\u4e2a|\u53ea|\u4ef6|\u9879|\u5957|"
    r"\u6a18|\u6839|\u5f20|\u5757|\u7247|\u7ec4|\u70b9|kg|g|t|\u5428|l|\u5347|\u53f0)?\s*$",
    re.IGNORECASE,
)
_MAX_QUANTITY = Decimal("99999999999999.999999")
_QUANTITY_QUANTUM = Decimal("0.000001")
_UNIQUE_MAPPING_FIELDS = {"item_name", "spec", "quantity", "unit", "remark", "location"}
_ALLOWED_MAPPING_FIELDS = _UNIQUE_MAPPING_FIELDS | {"ignore"}
_BUDGET_HEADER_ALIASES = {
    "item_name": {
        "\u9879\u76ee\u540d\u79f0",
        "\u65bd\u5de5\u9879\u76ee",
        "\u5de5\u4f5c\u5185\u5bb9",
        "\u6e05\u5355\u540d\u79f0",
        "\u6750\u6599\u540d\u79f0",
        "\u5206\u90e8\u5206\u9879\u540d\u79f0",
        "\u5de5\u7a0b\u540d\u79f0",
        "\u540d\u79f0",
    },
    "spec": {
        "\u89c4\u683c",
        "\u578b\u53f7",
        "\u9879\u76ee\u7279\u5f81",
        "\u9879\u76ee\u7279\u5f81/\u505a\u6cd5\u8981\u6c42",
        "\u505a\u6cd5\u8981\u6c42",
        "\u7279\u5f81\u63cf\u8ff0",
        "\u89c4\u683c\u578b\u53f7",
        "\u505a\u6cd5",
        "\u5de5\u827a",
    },
    "quantity": {
        "\u6570\u91cf",
        "\u5de5\u7a0b\u91cf",
        "\u6570\u91cf/\u5de5\u7a0b\u91cf",
        "\u5de5\u7a0b\u6570\u91cf",
        "\u9884\u4f30\u6570\u91cf",
        "\u6682\u5b9a\u6570\u91cf",
    },
    "unit": {"\u5355\u4f4d", "\u8ba1\u91cf\u5355\u4f4d", "\u62a5\u4ef7\u5355\u4f4d"},
    "remark": {"\u5907\u6ce8", "\u8bf4\u660e", "\u9644\u6ce8", "\u65bd\u5de5\u8bf4\u660e"},
}
_PRICE_AMOUNT_MARKERS = {
    "\u5355\u4ef7",
    "\u7efc\u5408\u5355\u4ef7",
    "\u5355\u4ef7\u5206\u6790",
    "\u5408\u4ef7",
    "\u7efc\u5408\u5408\u4ef7",
    "\u91d1\u989d",
    "\u603b\u4ef7",
    "\u9020\u4ef7",
    "\u62a5\u4ef7",
    "\u4eba\u5de5\u8d39",
    "\u4e3b\u6750\u8d39",
    "\u8f85\u6750\u8d39",
    "\u6750\u6599\u8d39",
    "\u673a\u68b0\u8d39",
    "\u7ba1\u7406\u8d39",
    "\u5229\u6da6",
    "\u7a0e\u91d1",
    "\u89c4\u8d39",
    "\u6210\u672c",
    "\u542b\u7a0e\u4ef7",
    "\u4e0d\u542b\u7a0e\u4ef7",
    "单价",
    "综合单价",
    "单价分析",
    "合价",
    "综合合价",
    "金额",
    "总价",
    "造价",
    "报价",
    "人工费",
    "主材费",
    "辅材费",
    "材料费",
    "机械费",
    "管理费",
    "利润",
    "税金",
    "规费",
    "成本",
    "单方造价",
    "供应商含税报价",
    "含税价",
    "不含税价",
}
_QUANTITY_AGGREGATE_MARKERS = (
    "工程量小计",
    "工程量合计",
    "总工程量",
    "数量小计",
    "数量合计",
    "总数量",
)
_SUMMARY_FORMULA_REF_RE = re.compile(
    r"(?:'(?P<quoted>[^']+)'|(?P<plain>[^'!=+\-*/(),\s]+))!\$?(?P<column>[A-Z]{1,3})\$?(?P<row>\d+)",
    re.IGNORECASE,
)
_SIMPLE_SUM_RANGE_RE = re.compile(
    r"SUM\(\$?(?P<start_col>[A-Z]{1,3})\$?(?P<start_row>\d+)\s*:\s*"
    r"\$?(?P<end_col>[A-Z]{1,3})\$?(?P<end_row>\d+)\)",
    re.IGNORECASE,
)
_SAME_SHEET_CELL_REF_RE = re.compile(r"\$?(?P<column>[A-Z]{1,3})\$?(?P<row>\d+)", re.IGNORECASE)
_SUMMARY_SCOPE_MARKERS = (
    "\u680b",
    "\u697c",
    "\u697c\u5c42",
    "\u5c42\u6570",
    "\u6237\u578b",
    "\u533a\u57df",
    "\u6237\u6570",
    "\u5957\u6570",
    "\u6570\u91cf",
    "\u9762\u79ef",
)
_BUDGET_ROW_ANNOTATION_FIELDS = (
    "budget_summary_multiplier",
    "budget_summary_multiplier_sources",
)


def _json_dump(value: Any) -> str:
    return json.dumps(value, ensure_ascii=False, separators=(",", ":"), default=str)


def _json_load(value: str | None, fallback: Any) -> Any:
    if not value:
        return fallback
    try:
        return json.loads(value)
    except (TypeError, ValueError):
        return fallback


def _clean_text(value: Any, limit: int | None = None) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    if not text:
        return None
    return text[:limit] if limit else text


def _format_dt(value: Any) -> str | None:
    return value.isoformat() if value else None


def require_budget_project_access(current_user: User) -> None:
    if not has_any_role(current_user, BUDGET_PROJECT_READ_ROLES):
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="PERMISSION_DENIED")


def require_budget_project_create(current_user: User) -> None:
    if not has_any_role(current_user, BUDGET_PROJECT_CREATE_ROLES):
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="PERMISSION_DENIED")


def can_view_all_budget_projects(current_user: User) -> bool:
    return has_any_role(current_user, BUDGET_PROJECT_VIEW_ALL_ROLES)


def accessible_budget_profile_query(db: Session, current_user: User):
    require_budget_project_access(current_user)
    query = db.query(BudgetProjectProfile).join(Project, Project.id == BudgetProjectProfile.project_id)
    if settings.feature_budget_pricing_drafts:
        try:
            account = resolve_current_account(db, current_user)
        except AccountTenancyError as exc:
            raise HTTPException(status_code=exc.status_code, detail=exc.detail) from exc
        # Account filtering is intentionally applied before legacy role/owner
        # visibility.  An admin may view all budget projects in their current
        # account, never projects belonging to another account.
        query = query.join(
            AccountBudgetProject,
            AccountBudgetProject.project_id == BudgetProjectProfile.project_id,
        ).filter(AccountBudgetProject.account_id == account.id)
    if can_view_all_budget_projects(current_user):
        return query
    creator_filters = (
        Project.created_by == current_user.id,
        BudgetProjectProfile.created_by == current_user.id,
    )
    if has_any_role(current_user, _PROJECT_SCOPED_READ_ROLES):
        task_project_ids = db.query(ProjectTask.project_id).filter(ProjectTask.owner_user_id == current_user.id)
        return query.filter(
            or_(
                *creator_filters,
                Project.project_manager_id == current_user.id,
                Project.id.in_(task_project_ids),
            )
        )
    return query.filter(or_(*creator_filters))


def get_budget_profile(db: Session, project_id: int, current_user: User) -> BudgetProjectProfile:
    profile = (
        accessible_budget_profile_query(db, current_user)
        .filter(BudgetProjectProfile.project_id == project_id)
        .first()
    )
    if not profile:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="BUDGET_PROJECT_NOT_FOUND")
    return profile


def require_profile_mutation(profile: BudgetProjectProfile, current_user: User) -> None:
    if not has_any_role(current_user, BUDGET_PROJECT_WRITE_ROLES):
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="PERMISSION_DENIED")
    if not (can_manage_project(current_user, profile.project) or profile.created_by == current_user.id):
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="PERMISSION_DENIED")


def require_active_profile(profile: BudgetProjectProfile) -> None:
    if profile.workspace_status != BUDGET_WORKSPACE_ACTIVE:
        raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="BUDGET_PROJECT_ARCHIVED")


def create_budget_project(db: Session, payload: dict[str, Any], current_user: User) -> BudgetProjectProfile:
    require_budget_project_create(current_user)
    name = _clean_text(payload.get("name"), 255)
    if not name:
        raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_ENTITY, detail="PROJECT_NAME_REQUIRED")
    project = Project(
        project_code=next_project_code(db),
        name=name,
        client_name=_clean_text(payload.get("client_name"), 128),
        address=_clean_text(payload.get("address"), 255),
        description=_clean_text(payload.get("description"), 4000),
        status="planning",
        risk_level="normal",
        progress_percent=0,
        project_manager_id=current_user.id,
        owner_department=_clean_text(payload.get("owner_department"), 128),
        created_by=current_user.id,
    )
    db.add(project)
    db.flush()
    profile = BudgetProjectProfile(
        project_id=project.id,
        workspace_status=BUDGET_WORKSPACE_ACTIVE,
        created_by=current_user.id,
        updated_by=current_user.id,
    )
    db.add(profile)
    db.flush()
    if settings.feature_budget_pricing_drafts:
        try:
            bind_budget_project_to_current_account(
                db,
                project_id=project.id,
                current_user=current_user,
            )
        except AccountTenancyError as exc:
            raise HTTPException(status_code=exc.status_code, detail=exc.detail) from exc
    return profile


def update_budget_project(
    db: Session,
    profile: BudgetProjectProfile,
    payload: dict[str, Any],
    current_user: User,
) -> BudgetProjectProfile:
    require_profile_mutation(profile, current_user)
    require_active_profile(profile)
    project = profile.project
    for field, limit in (
        ("name", 255),
        ("client_name", 128),
        ("address", 255),
        ("description", 4000),
        ("owner_department", 128),
    ):
        if field not in payload:
            continue
        value = _clean_text(payload.get(field), limit)
        if field == "name" and not value:
            raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_ENTITY, detail="PROJECT_NAME_REQUIRED")
        setattr(project, field, value)
    profile.updated_by = current_user.id
    db.flush()
    return profile


def archive_budget_project(
    db: Session,
    profile: BudgetProjectProfile,
    reason: str | None,
    current_user: User,
) -> BudgetProjectProfile:
    require_profile_mutation(profile, current_user)
    if profile.workspace_status == BUDGET_WORKSPACE_ARCHIVED:
        return profile
    profile.workspace_status = BUDGET_WORKSPACE_ARCHIVED
    profile.archived_at = app_local_naive()
    profile.archived_by = current_user.id
    profile.archive_reason = _clean_text(reason, 2000)
    profile.updated_by = current_user.id
    db.flush()
    return profile


def _latest_import(db: Session, project_id: int) -> BudgetProjectImportBatch | None:
    return (
        db.query(BudgetProjectImportBatch)
        .filter(BudgetProjectImportBatch.project_id == project_id)
        .order_by(BudgetProjectImportBatch.id.desc())
        .first()
    )


def _can_mutate_budget_profile(
    profile: BudgetProjectProfile,
    current_user: User | None,
) -> bool:
    if profile.workspace_status != BUDGET_WORKSPACE_ACTIVE:
        return False
    if current_user is None or not has_any_role(current_user, BUDGET_PROJECT_WRITE_ROLES):
        return False
    return bool(
        can_manage_project(current_user, profile.project)
        or profile.created_by == current_user.id
    )


def _snapshot_json(
    value: str | None,
    expected_type: type,
    *,
    field: str,
    batch_id: int,
) -> Any:
    try:
        parsed = json.loads(value or "")
    except (TypeError, ValueError, json.JSONDecodeError) as exc:
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail={
                "code": "BUDGET_IMPORT_SNAPSHOT_SOURCE_INVALID",
                "batch_id": batch_id,
                "field": field,
            },
        ) from exc
    if not isinstance(parsed, expected_type):
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail={
                "code": "BUDGET_IMPORT_SNAPSHOT_SOURCE_INVALID",
                "batch_id": batch_id,
                "field": field,
                "expected_type": expected_type.__name__,
            },
        )
    return parsed


def _revision_snapshot(
    db: Session,
    batch: BudgetProjectImportBatch,
) -> dict[str, Any]:
    mappings = (
        db.query(BudgetProjectImportSheetMapping)
        .filter(BudgetProjectImportSheetMapping.batch_id == batch.id)
        .order_by(BudgetProjectImportSheetMapping.id.asc())
        .all()
    )
    rows = (
        db.query(BudgetProjectStandardRow)
        .filter(BudgetProjectStandardRow.batch_id == batch.id)
        .order_by(BudgetProjectStandardRow.sort_order.asc(), BudgetProjectStandardRow.id.asc())
        .all()
    )
    mapping_snapshot = [
        {
            "sheet_name": item.sheet_name,
            "sheet_role": item.sheet_role,
            "header_row_index": item.header_row_index,
            "detected_field_mapping": _snapshot_json(
                item.detected_field_mapping_json,
                dict,
                field=f"sheet[{item.sheet_name}].detected_field_mapping_json",
                batch_id=batch.id,
            ),
            "applied_field_mapping": _snapshot_json(
                item.applied_field_mapping_json,
                dict,
                field=f"sheet[{item.sheet_name}].applied_field_mapping_json",
                batch_id=batch.id,
            ),
            "detected_columns": _snapshot_json(
                item.detected_columns_json,
                list,
                field=f"sheet[{item.sheet_name}].detected_columns_json",
                batch_id=batch.id,
            ),
            "current_columns": _snapshot_json(
                item.current_columns_json,
                list,
                field=f"sheet[{item.sheet_name}].current_columns_json",
                batch_id=batch.id,
            ),
            "mapping_revision": item.mapping_revision,
        }
        for item in mappings
    ]
    row_snapshot = [
        {
            "row_key": item.row_key,
            "source_sheet": item.source_sheet,
            "sheet_role": item.sheet_role,
            "raw_row_index": item.raw_row_index,
            "sort_order": item.sort_order,
            "mapping_revision": item.mapping_revision,
            "row_type": item.row_type,
            "is_standard_item": bool(item.is_standard_item),
            "quantity_status": item.quantity_status,
            "standard_row": _snapshot_json(
                item.standard_row_json,
                dict,
                field=f"row[{item.row_key}].standard_row_json",
                batch_id=batch.id,
            ),
        }
        for item in rows
    ]
    summary = {
        "sheet_count": batch.sheet_count,
        "total_output_row_count": batch.total_output_row_count,
        "standard_item_count": batch.standard_item_count,
        "valid_quantity_count": batch.valid_quantity_count,
        "invalid_quantity_count": batch.invalid_quantity_count,
    }
    preview = _snapshot_json(
        batch.current_preview_json,
        dict,
        field="current_preview_json",
        batch_id=batch.id,
    )
    return {
        "preview": preview,
        "sheet_mappings": mapping_snapshot,
        "standard_rows": row_snapshot,
        "summary": summary,
    }


def _append_import_revision(
    db: Session,
    batch: BudgetProjectImportBatch,
    current_user: User,
    *,
    revision_kind: str,
) -> BudgetProjectImportRevision:
    db.flush()
    snapshot = _revision_snapshot(db, batch)
    snapshot_sha256 = hashlib.sha256(_json_dump(snapshot).encode("utf-8")).hexdigest()
    revision = BudgetProjectImportRevision(
        revision_uuid=str(uuid.uuid4()),
        batch_id=batch.id,
        revision_number=int(batch.remap_revision or 0),
        revision_kind=revision_kind,
        snapshot_sha256=snapshot_sha256,
        preview_json=_json_dump(snapshot["preview"]),
        sheet_mappings_json=_json_dump(snapshot["sheet_mappings"]),
        standard_rows_json=_json_dump(snapshot["standard_rows"]),
        summary_json=_json_dump(snapshot["summary"]),
        created_by=current_user.id,
    )
    db.add(revision)
    db.flush()
    batch.current_revision_id = revision.id
    db.flush()
    return revision


def _append_import_lifecycle_event(
    db: Session,
    *,
    batch: BudgetProjectImportBatch,
    revision: BudgetProjectImportRevision,
    event_type: str,
    from_status: str | None,
    to_status: str,
    current_user: User,
    event_data: dict[str, Any] | None = None,
) -> BudgetProjectImportLifecycleEvent:
    event = BudgetProjectImportLifecycleEvent(
        project_id=batch.project_id,
        batch_id=batch.id,
        revision_id=revision.id,
        event_type=event_type,
        from_status=from_status,
        to_status=to_status,
        actor_id=current_user.id,
        event_json=_json_dump(event_data) if event_data else None,
    )
    db.add(event)
    return event


def serialize_budget_project(
    db: Session,
    profile: BudgetProjectProfile,
    current_user: User | None = None,
) -> dict[str, Any]:
    project = profile.project
    import_count = (
        db.query(BudgetProjectImportBatch)
        .filter(BudgetProjectImportBatch.project_id == project.id)
        .count()
    )
    latest = _latest_import(db, project.id)
    latest_data = (
        serialize_import_batch(
            latest,
            profile=profile,
            current_user=current_user,
        )
        if latest
        else None
    )
    can_mutate = _can_mutate_budget_profile(profile, current_user)
    can_remap = bool(
        can_mutate
        and db.query(BudgetProjectImportBatch.id)
        .filter(
            BudgetProjectImportBatch.project_id == project.id,
            BudgetProjectImportBatch.status == BUDGET_IMPORT_STATUS_PARSED,
        )
        .first()
    )
    can_activate = bool(
        can_mutate
        and db.query(BudgetProjectImportBatch.id)
        .filter(
            BudgetProjectImportBatch.project_id == project.id,
            BudgetProjectImportBatch.status.in_(
                [BUDGET_IMPORT_STATUS_CONFIRMED, BUDGET_IMPORT_STATUS_SUPERSEDED]
            ),
        )
        .first()
    )
    can_view_pricing = bool(
        settings.feature_budget_projects
        and settings.feature_budget_pricing
        and current_user
        and has_any_role(
            current_user,
            {
                "system_admin",
                "admin",
                "cost_viewer",
                "cost_editor",
                "cost_approver",
                "cost_exporter",
            },
        )
    )
    can_create_pricing_run = bool(
        can_view_pricing
        and profile.workspace_status == BUDGET_WORKSPACE_ACTIVE
        and profile.active_import_batch_id
        and profile.active_import_revision_id
        and current_user
        and has_any_role(
            current_user,
            {"system_admin", "admin", "cost_editor", "cost_approver"},
        )
    )
    return {
        "id": project.id,
        "project_id": project.id,
        "project_code": project.project_code,
        "name": project.name,
        "client_name": project.client_name,
        "address": project.address,
        "description": project.description,
        "project_manager_id": project.project_manager_id,
        "owner_department": project.owner_department,
        "project_status": project.status,
        "status": profile.workspace_status,
        "workspace_status": profile.workspace_status,
        "archive_reason": profile.archive_reason,
        "archived_at": _format_dt(profile.archived_at),
        "archived_by": profile.archived_by,
        "created_by": profile.created_by,
        "created_at": _format_dt(profile.created_at),
        "updated_at": _format_dt(profile.updated_at),
        "import_count": import_count,
        "latest_import": latest_data,
        "latest_import_batch_uuid": latest.batch_uuid if latest else None,
        "active_import_batch_id": profile.active_import_batch_id,
        "active_import_revision_id": profile.active_import_revision_id,
        "standard_item_count": latest.standard_item_count if latest else 0,
        "valid_quantity_count": latest.valid_quantity_count if latest else 0,
        "invalid_quantity_count": latest.invalid_quantity_count if latest else 0,
        "capabilities": {
            "can_edit": can_mutate,
            "can_archive": can_mutate,
            "can_upload": can_mutate,
            "can_remap": can_remap,
            "can_activate_import": can_activate,
            "can_view_pricing": can_view_pricing,
            "can_create_pricing_run": can_create_pricing_run,
        },
    }


def _validate_workbook_limits(content: bytes, filename: str) -> None:
    suffix = Path(filename or "").suffix.lower()
    if suffix not in SUPPORTED_EXTENSIONS:
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail={"code": "UNSUPPORTED_BUDGET_IMPORT_FILE", "supported": sorted(SUPPORTED_EXTENSIONS)},
        )
    try:
        workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:
        raise RequirementStandardizationError(
            f"Excel requirement file cannot be read as .xlsx/.xlsm: {exc}"
        ) from exc
    over_limit: list[dict[str, Any]] = []
    try:
        for sheet in workbook.worksheets:
            row_count, column_count = _real_non_empty_sheet_dimensions(sheet)
            if row_count > MAX_ROWS_PER_SHEET or column_count > MAX_COLUMNS_PER_SHEET:
                over_limit.append(
                    {
                        "sheet_name": sheet.title,
                        "row_count": row_count,
                        "column_count": column_count,
                        "worksheet_row_count": int(sheet.max_row or 0),
                        "worksheet_column_count": int(sheet.max_column or 0),
                    }
                )
    finally:
        workbook.close()
    if over_limit:
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail={
                "code": "BUDGET_IMPORT_WORKBOOK_LIMIT_EXCEEDED",
                "message": "导入的 Excel 工作簿存在超过系统行列上限的 Sheet",
                "max_rows_per_sheet": MAX_ROWS_PER_SHEET,
                "max_columns_per_sheet": MAX_COLUMNS_PER_SHEET,
                "sheets": over_limit,
            },
        )


def _real_non_empty_sheet_dimensions(sheet: Any) -> tuple[int, int]:
    max_row = 0
    max_column = 0
    for row_index, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        row_has_value = False
        for column_index, value in enumerate(row, start=1):
            if _clean_text(value) is None:
                continue
            row_has_value = True
            if column_index > max_column:
                max_column = column_index
        if row_has_value:
            max_row = row_index
        if max_row > MAX_ROWS_PER_SHEET and max_column > MAX_COLUMNS_PER_SHEET:
            break
    return max_row, max_column


def _mapping_by_sheet(preview: dict[str, Any]) -> dict[str, dict[str, Any]]:
    return {
        str(item.get("sheet_name") or ""): item
        for item in preview.get("sheet_mappings") or []
        if item.get("sheet_name")
    }


def _normalize_label(value: Any) -> str:
    text = str(value or "").strip().lower()
    return re.sub(r"[\s._:\uff1a()\uff08\uff09\-]+", "", text)


def _is_sequence_label(label: Any) -> bool:
    normalized = _normalize_label(label)
    known = {_normalize_label(value) for value in _SEQUENCE_LABELS}
    return normalized in known or normalized.endswith("\u5e8f\u53f7") or normalized.endswith("\u884c\u53f7")


def _cell_value(row: dict[str, Any], column: str) -> str:
    for cell in row.get("raw_cells") or []:
        if str(cell.get("column") or "").upper() == column.upper():
            return str(cell.get("value") if cell.get("value") is not None else "").strip()
    return ""


def _resolved_header_value(sheet: Any, row: int, column: int) -> str:
    value = sheet.cell(row=row, column=column).value
    if value not in (None, ""):
        return str(value).strip()
    for merged in sheet.merged_cells.ranges:
        if merged.min_row <= row <= merged.max_row and merged.min_col <= column <= merged.max_col:
            value = sheet.cell(row=merged.min_row, column=merged.min_col).value
            return str(value).strip() if value not in (None, "") else ""
    return ""


def _header_kind(values: list[str]) -> str | None:
    for value in values:
        normalized = _normalize_label(value)
        if not normalized:
            continue
        for field, aliases in _BUDGET_HEADER_ALIASES.items():
            normalized_aliases = {_normalize_label(alias) for alias in aliases}
            if normalized in normalized_aliases or any(
                len(alias) >= 2 and alias in normalized for alias in normalized_aliases
            ):
                return field
    return None


def _detect_budget_header_row(sheet: Any, max_column: int) -> int | None:
    best_row: int | None = None
    best_score = 0
    for row_index in range(1, min(int(sheet.max_row or 0), BUDGET_MAX_SCAN_ROWS) + 1):
        kinds = {
            _header_kind([_resolved_header_value(sheet, row_index, column)])
            for column in range(1, max_column + 1)
        }
        kinds.discard(None)
        score = len(kinds & {"item_name", "spec", "quantity", "unit"})
        if score > best_score:
            best_score = score
            best_row = row_index
        if score >= 4:
            return row_index
    return best_row if best_score >= 3 else None


def _is_quantity_header(values: list[str]) -> bool:
    return _header_kind(values) == "quantity"


def _is_aggregate_quantity_header(values: list[str]) -> bool:
    compact = "|".join(_normalize_label(value) for value in values if value)
    return any(_normalize_label(marker) in compact for marker in _QUANTITY_AGGREGATE_MARKERS)


def _is_price_amount_header(values: list[str]) -> bool:
    if _is_aggregate_quantity_header(values):
        return False
    compact_values = [_normalize_label(value) for value in values if value]
    return any(
        _normalize_label(marker) in compact
        for compact in compact_values
        for marker in _PRICE_AMOUNT_MARKERS
    )


def _formula_sums_quantity_block(formulas: list[str], detail_columns: list[int]) -> bool:
    if not formulas or not detail_columns:
        return False
    first = get_column_letter(min(detail_columns))
    last = get_column_letter(max(detail_columns))
    pattern = re.compile(
        rf"SUM\(\$?{re.escape(first)}\$?\d+\s*:\s*\$?{re.escape(last)}\$?\d+\)",
        re.IGNORECASE,
    )
    matched = sum(bool(pattern.search(formula.replace(" ", ""))) for formula in formulas)
    return matched >= max(1, min(2, len(formulas)))


def _infer_sheet_role(sheet_name: str, mapping: dict[str, Any], preview: dict[str, Any]) -> str:
    normalized_name = _normalize_label(sheet_name)
    sheet_text = " ".join(
        str(row.get("raw_text") or "")
        for row in preview.get("rows") or []
        if str(row.get("source_sheet") or "") == sheet_name
    )[:2000]
    normalized_text = _normalize_label(sheet_text)
    if any(marker in normalized_name for marker in ("\u8ba1\u7b97\u89c4\u5219", "\u8ba1\u91cf\u89c4\u5219", "\u5de5\u7a0b\u91cf\u8ba1\u7b97\u89c4\u5219")):
        return BUDGET_SHEET_ROLE_CALCULATION_RULE
    if any(marker in normalized_name for marker in ("\u5907\u7528\u6e05\u5355", "\u6682\u5217\u6e05\u5355")):
        return BUDGET_SHEET_ROLE_OPTIONAL_BACKUP
    if any(marker in normalized_name for marker in ("\u635f\u8017", "\u635f\u8017\u8868")):
        return BUDGET_SHEET_ROLE_LOSS_REFERENCE
    if any(marker in normalized_name for marker in ("\u4e3b\u6750\u54c1\u724c", "\u54c1\u724c\u8868", "\u6750\u6599\u54c1\u724c")):
        return BUDGET_SHEET_ROLE_MATERIAL_REFERENCE
    if any(marker in normalized_name for marker in ("\u6c47\u603b", "\u6c47\u603b\u8868", "\u62a5\u4ef7\u6c47\u603b")):
        return BUDGET_SHEET_ROLE_SUMMARY_ANALYSIS
    if any(marker in normalized_name for marker in ("封面", "说明")) or any(
        marker in normalized_text for marker in ("编制说明", "报价说明")
    ):
        return BUDGET_SHEET_ROLE_METADATA
    if any(marker in normalized_name for marker in ("汇总", "成本分析", "报价组成", "分析表")):
        return BUDGET_SHEET_ROLE_SUMMARY_ANALYSIS
    field_mapping = mapping.get("field_mapping") or {}
    column_labels = "|".join(
        _normalize_label(column.get("label")) for column in mapping.get("columns") or []
    )
    if (
        any(marker in normalized_name for marker in ("主材表", "材料表", "主材清单"))
        or (
            "材料名称" in column_labels
            and any(marker in column_labels for marker in ("品牌厂家", "参考样板", "主材单价"))
        )
    ):
        return BUDGET_SHEET_ROLE_MATERIAL_REFERENCE
    if "item_name" in field_mapping.values() or "quantity" in field_mapping.values():
        return BUDGET_SHEET_ROLE_BILL
    return BUDGET_SHEET_ROLE_BILL


def _decimal_from_cell(value: Any) -> Decimal | None:
    if value is None or value == "":
        return None
    try:
        parsed = Decimal(str(value).replace(",", "").strip())
    except (InvalidOperation, ValueError):
        return None
    return parsed if parsed.is_finite() and parsed > 0 else None


def _summary_multiplier_with_presence(
    sheet: Any, row_index: int, formula_column: int
) -> tuple[Decimal, bool]:
    best: Decimal | None = None
    for column in range(1, formula_column):
        value = _decimal_from_cell(sheet.cell(row=row_index, column=column).value)
        if value is None:
            continue
        header_text = "".join(
            _normalize_label(_resolved_header_value(sheet, header_row, column))
            for header_row in range(1, min(row_index, 5))
        )
        if any(marker in header_text for marker in ("\u6237\u6570", "\u5957\u6570", "\u6570\u91cf", "\u5c42\u6570")):
            return value, True
        if best is None and value >= 1:
            best = value
    # A generic number may be a sequence, area or subtotal. It remains a
    # fallback value for legacy callers, but is not sufficient evidence that
    # this summary row represents a quantity scope.
    return best or Decimal("1"), False


def _summary_multiplier_for_row(sheet: Any, row_index: int, formula_column: int) -> Decimal:
    return _summary_multiplier_with_presence(sheet, row_index, formula_column)[0]


def _summary_scope_for_row(sheet: Any, row_index: int) -> list[dict[str, str]]:
    """Keep the human-readable building/floor scope alongside the multiplier."""

    scope: list[dict[str, str]] = []
    for column in range(1, sheet.max_column + 1):
        value = sheet.cell(row=row_index, column=column).value
        if value is None or str(value).strip() == "" or str(value).startswith("="):
            continue
        header_labels = [
            _normalize_label(_resolved_header_value(sheet, header_row, column))
            for header_row in range(1, min(row_index, 5))
        ]
        label = next(
            (
                candidate
                for candidate in header_labels
                if candidate and any(marker in candidate for marker in _SUMMARY_SCOPE_MARKERS)
            ),
            "",
        )
        if not label:
            continue
        scope.append({"label": str(label)[:64], "value": str(value).strip()[:128]})
    return scope[:8]


def _summary_detail_references(
    workbook: Any,
    sheet_name: str,
    coordinate: str,
    *,
    seen: set[tuple[str, str]] | None = None,
) -> set[tuple[str, str, int]]:
    """Follow same-sheet summary aliases until they reach bill subtotal cells."""

    seen = seen or set()
    key = (sheet_name, coordinate.upper())
    if key in seen or sheet_name not in workbook.sheetnames:
        return set()
    seen.add(key)
    sheet = workbook[sheet_name]
    formula = str(sheet[coordinate].value or "")
    if not formula.startswith("="):
        return set()

    references: set[tuple[str, str, int]] = set()
    for match in _SUMMARY_FORMULA_REF_RE.finditer(formula):
        ref_sheet = match.group("quoted") or match.group("plain") or ""
        ref_column = match.group("column").upper()
        ref_row = int(match.group("row"))
        if ref_sheet == sheet_name or (
            ref_sheet in workbook.sheetnames
            and _infer_sheet_role(ref_sheet, {}, {"rows": []}) == BUDGET_SHEET_ROLE_SUMMARY_ANALYSIS
        ):
            references.update(
                _summary_detail_references(
                    workbook,
                    ref_sheet or sheet_name,
                    f"{ref_column}{ref_row}",
                    seen=seen,
                )
            )
        elif ref_sheet and _infer_sheet_role(ref_sheet, {}, {"rows": []}) == BUDGET_SHEET_ROLE_BILL:
            references.add((ref_sheet, ref_column, ref_row))

    # A qualified bill reference such as '明细'!H60 is not a same-sheet H60.
    unqualified_formula = _SUMMARY_FORMULA_REF_RE.sub("", formula)
    for match in _SAME_SHEET_CELL_REF_RE.finditer(unqualified_formula):
        ref_column = match.group("column").upper()
        ref_row = int(match.group("row"))
        if f"{ref_column}{ref_row}".upper() == coordinate.upper():
            continue
        references.update(
            _summary_detail_references(
                workbook,
                sheet_name,
                f"{ref_column}{ref_row}",
                seen=seen,
            )
        )
    return references


def _detail_formula_rows(
    workbook: Any,
    sheet_name: str,
    column: str,
    row_index: int,
    *,
    seen: set[tuple[str, str, int]] | None = None,
) -> list[int]:
    seen = seen or set()
    trace_key = (sheet_name, column.upper(), row_index)
    if trace_key in seen:
        return []
    seen.add(trace_key)
    if sheet_name not in workbook.sheetnames:
        return [row_index]
    sheet = workbook[sheet_name]
    value = sheet[f"{column}{row_index}"].value
    formula = str(value or "")
    if not formula.startswith("="):
        return [row_index]
    compact = formula.replace(" ", "").lstrip("=")
    match = _SIMPLE_SUM_RANGE_RE.search(compact)
    if not match:
        nested_rows: list[int] = []
        for ref in _SAME_SHEET_CELL_REF_RE.finditer(compact):
            ref_column = ref.group("column").upper()
            ref_row = int(ref.group("row"))
            if ref_column != column.upper() or ref_row == row_index:
                continue
            nested_rows.extend(
                _detail_formula_rows(workbook, sheet_name, column, ref_row, seen=seen)
            )
        return sorted(set(nested_rows)) if nested_rows else [row_index]
    if match.group("start_col").upper() != column.upper() or match.group("end_col").upper() != column.upper():
        return [row_index]
    start = int(match.group("start_row"))
    end = int(match.group("end_row"))
    if start <= 0 or end < start:
        return [row_index]
    sum_rows = list(range(start, end + 1))
    nested_rows: list[int] = []
    for nested_row in sum_rows:
        nested_rows.extend(
            _detail_formula_rows(workbook, sheet_name, column, nested_row, seen=seen)
        )
    return sorted(set(nested_rows)) if nested_rows else sum_rows


def _workbook_summary_multipliers(content: bytes) -> dict[tuple[str, int], dict[str, Any]]:
    workbook = load_workbook(BytesIO(content), read_only=False, data_only=False)
    multipliers: dict[tuple[str, int], dict[str, Any]] = {}
    try:
        for sheet in workbook.worksheets:
            if _infer_sheet_role(sheet.title, {}, {"rows": []}) != BUDGET_SHEET_ROLE_SUMMARY_ANALYSIS:
                continue
            for row in sheet.iter_rows():
                formula_cells = [cell for cell in row if str(cell.value or "").startswith("=")]
                if not formula_cells:
                    continue
                row_index = formula_cells[0].row
                row_multiplier, has_explicit_scope = _summary_multiplier_with_presence(
                    sheet, row_index, sheet.max_column + 1
                )
                # Formula-only subtotal rows would otherwise add a second copy
                # of their children. A real scope row carries an explicit count,
                # including the valid value 1 for a public area.
                if not has_explicit_scope:
                    continue
                detail_references: set[tuple[str, str, int]] = set()
                for cell in formula_cells:
                    detail_references.update(
                        _summary_detail_references(workbook, sheet.title, cell.coordinate)
                    )
                if not detail_references:
                    continue
                scope = _summary_scope_for_row(sheet, row_index)
                for ref_sheet, ref_column, ref_row in sorted(detail_references):
                    for detail_row in _detail_formula_rows(workbook, ref_sheet, ref_column, ref_row):
                        key = (ref_sheet, detail_row)
                        current = multipliers.setdefault(
                            key,
                            {
                                "multiplier": Decimal("0"),
                                "sources": [],
                            },
                        )
                        current["multiplier"] += row_multiplier
                        current["sources"].append(
                            {
                                "summary_sheet": sheet.title,
                                "summary_row": row_index,
                                "summary_multiplier": str(row_multiplier),
                                "summary_scope": scope,
                                "detail_subtotal_cell": f"{ref_sheet}!{ref_column}{ref_row}",
                            }
                        )
    finally:
        workbook.close()
    return multipliers


def _annotate_summary_multipliers(preview: dict[str, Any], content: bytes) -> dict[str, Any]:
    multipliers = _workbook_summary_multipliers(content)
    if not multipliers:
        return preview
    annotated = copy.deepcopy(preview)
    for row in annotated.get("rows") or []:
        key = (str(row.get("source_sheet") or ""), int(row.get("raw_row_index") or 0))
        payload = multipliers.get(key)
        if not payload:
            continue
        multiplier = payload.get("multiplier") or Decimal("1")
        row["budget_summary_multiplier"] = str(multiplier)
        row["budget_summary_multiplier_sources"] = payload.get("sources") or []
        if multiplier != Decimal("1"):
            row["warnings"] = list(
                dict.fromkeys([*(row.get("warnings") or []), "BUDGET_SUMMARY_MULTIPLIER_APPLIED"])
            )
    return annotated


_REFERENCE_SHEET_ROLES = {
    BUDGET_SHEET_ROLE_CALCULATION_RULE,
    BUDGET_SHEET_ROLE_LOSS_REFERENCE,
    BUDGET_SHEET_ROLE_MATERIAL_REFERENCE,
    BUDGET_SHEET_ROLE_METADATA,
    BUDGET_SHEET_ROLE_OPTIONAL_BACKUP,
    BUDGET_SHEET_ROLE_SUMMARY_ANALYSIS,
}


def _annotate_reference_sheet_rows(preview: dict[str, Any]) -> dict[str, Any]:
    mapping_by_sheet = _mapping_by_sheet(preview)
    annotated = copy.deepcopy(preview)
    for row in annotated.get("rows") or []:
        sheet_name = str(row.get("source_sheet") or "")
        sheet_role = str(
            row.get("sheet_role")
            or (mapping_by_sheet.get(sheet_name, {}) or {}).get("sheet_role")
            or BUDGET_SHEET_ROLE_BILL
        )
        if sheet_role not in _REFERENCE_SHEET_ROLES:
            continue
        raw_text = _clean_text(row.get("raw_text"))
        if not raw_text:
            row["row_type"] = "empty_row"
            row["item_name"] = ""
            row["requires_confirmation"] = False
            row["warnings"] = list(dict.fromkeys([*(row.get("warnings") or []), "BUDGET_REFERENCE_EMPTY_ROW"]))
            continue
        row["row_type"] = "reference_row"
        row["item_name"] = raw_text[:255]
        row["spec"] = ""
        row["quantity"] = None
        row["unit"] = ""
        row["remark"] = raw_text
        row["requires_confirmation"] = False
        row["warnings"] = list(dict.fromkeys([*(row.get("warnings") or []), "BUDGET_REFERENCE_CONTEXT_ROW"]))
    return annotated


def _workbook_column_semantics(
    content: bytes, preview: dict[str, Any]
) -> tuple[dict[str, dict[str, Any]], dict[tuple[str, int, str], str]]:
    workbook = load_workbook(BytesIO(content), read_only=False, data_only=False)
    preview_mappings = _mapping_by_sheet(preview)
    semantics: dict[str, dict[str, Any]] = {}
    formula_cells: dict[tuple[str, int, str], str] = {}
    try:
        for sheet in workbook.worksheets:
            mapping = preview_mappings.get(sheet.title, {})
            header_row_index = mapping.get("header_row_index")
            max_column = min(int(sheet.max_column or 0), MAX_COLUMNS_PER_SHEET)
            header_rows: list[int] = []
            if header_row_index:
                header_rows = list(
                    range(int(header_row_index), min(sheet.max_row, int(header_row_index) + 1) + 1)
                )
            else:
                detected_header_row = _detect_budget_header_row(sheet, max_column)
                if detected_header_row:
                    header_rows = [detected_header_row]
            headers: dict[int, list[str]] = {}
            meaningful_columns: set[int] = set()
            formulas_by_column: dict[int, list[str]] = {}
            for column in range(1, max_column + 1):
                values = [
                    _resolved_header_value(sheet, row, column)
                    for row in header_rows
                ]
                headers[column] = list(dict.fromkeys(value for value in values if value))
                if headers[column]:
                    meaningful_columns.add(column)
            for row in sheet.iter_rows():
                for cell in row[:max_column]:
                    value = cell.value
                    if value not in (None, ""):
                        meaningful_columns.add(cell.column)
                    if cell.data_type == "f" or (isinstance(value, str) and value.startswith("=")):
                        formula = str(value)
                        column = get_column_letter(cell.column)
                        formula_cells[(sheet.title, cell.row, column)] = formula
                        formulas_by_column.setdefault(cell.column, []).append(formula)

            quantity_columns = [column for column, values in headers.items() if _is_quantity_header(values)]
            explicit_aggregates = [
                column for column, values in headers.items() if _is_aggregate_quantity_header(values)
            ]
            price_seed_columns = [
                column for column, values in headers.items() if _is_price_amount_header(values)
            ]
            preferred_quantity: int | None = min(explicit_aggregates) if explicit_aggregates else None
            first_price = min(price_seed_columns) if price_seed_columns else None
            if preferred_quantity is None and first_price and quantity_columns:
                candidate = first_price - 1
                details = [column for column in quantity_columns if column < candidate]
                if (
                    candidate > 0
                    and details
                    and max(details) == candidate - 1
                    and _formula_sums_quantity_block(formulas_by_column.get(candidate, []), details)
                ):
                    preferred_quantity = candidate
                    quantity_columns.append(candidate)
            if preferred_quantity is None:
                current_quantity = next(
                    (
                        column_index_from_string(column)
                        for column, field in (mapping.get("field_mapping") or {}).items()
                        if field == "quantity"
                    ),
                    None,
                )
                preferred_quantity = current_quantity

            locked_reasons: dict[str, str] = {}
            if preferred_quantity is not None:
                for column in quantity_columns:
                    if column != preferred_quantity:
                        locked_reasons[get_column_letter(column)] = "LAYER_QUANTITY_COLUMN"
            if first_price:
                is_material_reference_sheet = any(
                    marker in _normalize_label(sheet.title)
                    for marker in ("主材表", "材料表", "主材清单")
                )
                for column in sorted(meaningful_columns):
                    if column < first_price or column == preferred_quantity:
                        continue
                    kind = _header_kind(headers.get(column, []))
                    if kind in {"item_name", "spec", "quantity", "unit", "remark", "location"}:
                        continue
                    if (
                        is_material_reference_sheet
                        and column not in price_seed_columns
                        and column not in formulas_by_column
                    ):
                        continue
                    locked_reasons[get_column_letter(column)] = "PRICE_AMOUNT_COLUMN"
            for column in price_seed_columns:
                if column != preferred_quantity:
                    locked_reasons[get_column_letter(column)] = "PRICE_AMOUNT_COLUMN"
            quantity_semantic_columns = set(quantity_columns)
            if preferred_quantity:
                quantity_semantic_columns.add(preferred_quantity)
            for column in formulas_by_column:
                key = get_column_letter(column)
                if column in quantity_semantic_columns:
                    continue
                if key not in locked_reasons:
                    locked_reasons[key] = "FORMULA_COLUMN"

            detected_field_mapping: dict[str, str] = {}
            used_fields: set[str] = set()
            for column in sorted(headers):
                key = get_column_letter(column)
                kind = _header_kind(headers.get(column, []))
                if kind is None:
                    continue
                preferred_key = get_column_letter(preferred_quantity) if preferred_quantity else None
                if key in locked_reasons and not (kind == "quantity" and key == preferred_key):
                    continue
                if kind in used_fields and kind in _UNIQUE_MAPPING_FIELDS:
                    continue
                detected_field_mapping[key] = kind
                used_fields.add(kind)

            semantics[sheet.title] = {
                "locked_ignore_reasons": locked_reasons,
                "detected_budget_field_mapping": detected_field_mapping,
                "detected_budget_header_row_index": header_rows[0] if header_rows else None,
                "preferred_quantity_column": (
                    get_column_letter(preferred_quantity) if preferred_quantity else None
                ),
                "layer_quantity_columns": sorted(
                    [key for key, reason in locked_reasons.items() if reason == "LAYER_QUANTITY_COLUMN"],
                    key=column_index_from_string,
                ),
            }
    finally:
        workbook.close()
    return semantics, formula_cells


def _annotate_formula_provenance(
    preview: dict[str, Any], formula_cells: dict[tuple[str, int, str], str]
) -> dict[str, Any]:
    annotated = copy.deepcopy(preview)
    by_row: dict[tuple[str, int], dict[str, str]] = {}
    for (sheet_name, row_index, column), formula in formula_cells.items():
        by_row.setdefault((sheet_name, row_index), {})[column] = formula
    for row in annotated.get("rows") or []:
        key = (str(row.get("source_sheet") or ""), int(row.get("raw_row_index") or 0))
        formulas = by_row.get(key, {})
        if not formulas:
            continue
        cells = {
            str(cell.get("column") or "").upper(): cell
            for cell in row.get("raw_cells") or []
            if cell.get("column")
        }
        broken = False
        for column, formula in formulas.items():
            cell = cells.get(column)
            if cell is None:
                cell = {"column": column, "value": ""}
                cells[column] = cell
            cached_value = cell.get("value")
            formula_error = None
            if "#REF!" in formula.upper() or str(cached_value or "").upper() == "#REF!":
                formula_error = "BROKEN_FORMULA_REF"
                broken = True
            cell["raw_formula"] = formula
            cell["cached_value"] = cached_value
            cell["formula_error"] = formula_error
        row["raw_cells"] = sorted(
            cells.values(), key=lambda item: column_index_from_string(str(item.get("column") or "A"))
        )
        if broken:
            row["warnings"] = list(
                dict.fromkeys([*(row.get("warnings") or []), "BROKEN_FORMULA_REF"])
            )
            row["requires_confirmation"] = True
    return annotated


def _formula_provenance_from_preview(preview: dict[str, Any]) -> dict[tuple[str, int, str], str]:
    formulas: dict[tuple[str, int, str], str] = {}
    for row in preview.get("rows") or []:
        sheet_name = str(row.get("source_sheet") or "")
        row_index = int(row.get("raw_row_index") or 0)
        for cell in row.get("raw_cells") or []:
            formula = cell.get("raw_formula")
            column = str(cell.get("column") or "").upper()
            if formula and column:
                formulas[(sheet_name, row_index, column)] = str(formula)
    return formulas


def _apply_workbook_semantics(preview: dict[str, Any], content: bytes) -> dict[str, Any]:
    semantics, formula_cells = _workbook_column_semantics(content, preview)
    requested: list[dict[str, Any]] = []
    for mapping in preview.get("sheet_mappings") or []:
        sheet_name = str(mapping.get("sheet_name") or "")
        semantic = semantics.get(sheet_name, {})
        field_mapping = {
            str(column).upper(): "ignore" if field == "price_ignored" else str(field)
            for column, field in (mapping.get("field_mapping") or {}).items()
        }
        preferred = semantic.get("preferred_quantity_column")
        if preferred:
            for column, field in list(field_mapping.items()):
                if field == "quantity":
                    field_mapping[column] = "ignore"
            field_mapping[str(preferred)] = "quantity"
        for column, field in (semantic.get("detected_budget_field_mapping") or {}).items():
            if column in (semantic.get("locked_ignore_reasons") or {}) and field != "quantity":
                continue
            if field in _UNIQUE_MAPPING_FIELDS:
                for existing_column, existing_field in list(field_mapping.items()):
                    if existing_column != column and existing_field == field:
                        field_mapping[existing_column] = "ignore"
            field_mapping[str(column)] = str(field)
        for column in (semantic.get("locked_ignore_reasons") or {}):
            field_mapping[str(column)] = "ignore"
        requested.append({"sheet_name": sheet_name, "field_mapping": field_mapping})
    rebuilt = apply_manual_field_mappings(preview, requested)
    for mapping in rebuilt.get("sheet_mappings") or []:
        sheet_name = str(mapping.get("sheet_name") or "")
        semantic = semantics.get(sheet_name, {})
        mapping["budget_locked_ignore_reasons"] = semantic.get("locked_ignore_reasons") or {}
        mapping["budget_locked_ignore_columns"] = sorted(
            (semantic.get("locked_ignore_reasons") or {}).keys(), key=column_index_from_string
        )
        mapping["budget_detected_header_row_index"] = semantic.get("detected_budget_header_row_index")
        mapping["preferred_quantity_column"] = semantic.get("preferred_quantity_column")
        mapping["layer_quantity_columns"] = semantic.get("layer_quantity_columns") or []
        mapping["sheet_role"] = _infer_sheet_role(sheet_name, mapping, rebuilt)
    mapping_by_sheet = _mapping_by_sheet(rebuilt)
    for row in rebuilt.get("rows") or []:
        mapping = mapping_by_sheet.get(str(row.get("source_sheet") or ""), {})
        row["sheet_role"] = mapping.get("sheet_role") or BUDGET_SHEET_ROLE_BILL
    rebuilt = _annotate_summary_multipliers(rebuilt, content)
    rebuilt = _annotate_reference_sheet_rows(rebuilt)
    return _annotate_formula_provenance(rebuilt, formula_cells)


def _column_metadata(sheet_mapping: dict[str, Any], column: str) -> dict[str, Any]:
    for item in sheet_mapping.get("columns") or []:
        if str(item.get("column") or "").upper() == column.upper():
            return item
    return {}


def _mapped_columns(row: dict[str, Any], sheet_mapping: dict[str, Any], field: str) -> list[str]:
    field_mapping = row.get("field_mapping") or sheet_mapping.get("field_mapping") or {}
    return [str(column).upper() for column, mapped_field in field_mapping.items() if str(mapped_field) == field]


def _mapped_raw_value(row: dict[str, Any], sheet_mapping: dict[str, Any], field: str) -> str:
    columns = _mapped_columns(row, sheet_mapping, field)
    if len(columns) != 1:
        return ""
    return _cell_value(row, columns[0])


def _is_repeated_header_row(row: dict[str, Any], sheet_mapping: dict[str, Any]) -> bool:
    if row.get("row_type") != "data_row":
        return False
    matched_fields: set[str] = set()
    for field, aliases in _BUDGET_HEADER_ALIASES.items():
        raw_value = _mapped_raw_value(row, sheet_mapping, field)
        normalized = _normalize_label(raw_value)
        if not normalized:
            continue
        normalized_aliases = {_normalize_label(alias) for alias in aliases}
        if normalized in normalized_aliases:
            matched_fields.add(field)
            continue
        columns = _mapped_columns(row, sheet_mapping, field)
        if len(columns) == 1:
            label = _column_metadata(sheet_mapping, columns[0]).get("label")
            if label and normalized == _normalize_label(label):
                matched_fields.add(field)
    core_matches = matched_fields & {"item_name", "spec", "quantity", "unit"}
    return len(core_matches) >= 3 or (
        len(core_matches) >= 2 and bool(core_matches & {"quantity", "unit"})
    )


def _continuous_sequence_columns(preview: dict[str, Any], sheet_name: str) -> set[str]:
    sheet_mapping = _mapping_by_sheet(preview).get(sheet_name, {})
    if sheet_mapping.get("header_row_index") is not None:
        return set()
    rows = [
        row
        for row in preview.get("rows") or []
        if str(row.get("source_sheet") or "") == sheet_name
        and row.get("row_type") in {"data_row", "ambiguous_row"}
    ]
    all_columns = {
        str(cell.get("column") or "").upper()
        for row in rows
        for cell in row.get("raw_cells") or []
        if cell.get("column")
    }
    guarded: set[str] = set()
    for column in all_columns:
        values: list[int] = []
        invalid = False
        for row in rows:
            raw = _cell_value(row, column)
            if not raw:
                continue
            match = re.fullmatch(r"[+]?([0-9]+)[.\u3001)\uff09]?", raw)
            if not match:
                invalid = True
                break
            values.append(int(match.group(1)))
        if invalid or len(values) < 3 or values[0] > 3:
            continue
        if all(current == previous + 1 for previous, current in zip(values, values[1:])):
            guarded.add(column)
    if not guarded:
        return set()
    # In a headerless sheet, several real numeric columns can coincidentally be
    # 1/2/3. Only the left-most consecutive integer column is treated as the
    # likely row sequence; later columns remain available for manual quantity
    # mapping.
    return {min(guarded, key=column_index_from_string)}


def _guarded_sequence_columns(preview: dict[str, Any], sheet_name: str) -> set[str]:
    sheet_mapping = _mapping_by_sheet(preview).get(sheet_name, {})
    guarded = _continuous_sequence_columns(preview, sheet_name)
    for item in sheet_mapping.get("columns") or []:
        if _is_sequence_label(item.get("label")):
            guarded.add(str(item.get("column") or "").upper())
    return guarded


def _annotate_sequence_guards(preview: dict[str, Any]) -> dict[str, Any]:
    annotated = copy.deepcopy(preview)
    for mapping in annotated.get("sheet_mappings") or []:
        sheet_name = str(mapping.get("sheet_name") or "")
        guarded = _guarded_sequence_columns(annotated, sheet_name)
        for column in mapping.get("columns") or []:
            key = str(column.get("column") or "").upper()
            column["sequence_guarded"] = key in guarded
            if key in guarded:
                column["sequence_guard_reason"] = "SEQUENCE_COLUMN_NOT_QUANTITY"
    return annotated


def _ensure_budget_mapping_columns(preview: dict[str, Any]) -> dict[str, Any]:
    """Expose non-price raw columns, including numeric-only headerless columns."""
    enriched = copy.deepcopy(preview)
    rows_by_sheet: dict[str, list[dict[str, Any]]] = {}
    for row in enriched.get("rows") or []:
        rows_by_sheet.setdefault(str(row.get("source_sheet") or ""), []).append(row)
    for mapping in enriched.get("sheet_mappings") or []:
        sheet_name = str(mapping.get("sheet_name") or "")
        sheet_rows = rows_by_sheet.get(sheet_name, [])
        field_mapping = mapping.get("field_mapping") or {}
        locked_ignore_columns = {
            str(column).upper()
            for column, field in field_mapping.items()
            if field == "price_ignored"
        }
        locked_ignore_columns.update(
            str(column).upper() for column in mapping.get("budget_locked_ignore_columns") or []
        )
        locked_ignore_reasons = {
            str(column).upper(): str(reason)
            for column, reason in (mapping.get("budget_locked_ignore_reasons") or {}).items()
        }
        existing = {
            str(item.get("column") or "").upper(): item
            for item in mapping.get("columns") or []
            if item.get("column")
        }
        raw_columns = {
            str(cell.get("column") or "").upper()
            for row in sheet_rows
            for cell in row.get("raw_cells") or []
            if cell.get("column")
        }
        raw_columns.update(locked_ignore_columns)
        header_row_index = mapping.get("header_row_index")
        for column in sorted(raw_columns, key=column_index_from_string):
            if column in existing:
                if column in locked_ignore_columns:
                    existing[column]["detected_field"] = "ignore"
                    reason = locked_ignore_reasons.get(column) or "PRICE_AMOUNT_COLUMN"
                    existing[column]["is_price"] = reason == "PRICE_AMOUNT_COLUMN"
                    existing[column]["locked_ignore"] = True
                    existing[column]["lock_reason"] = reason
                continue
            label = ""
            samples: list[str] = []
            for row in sheet_rows:
                value = _cell_value(row, column)
                if not value:
                    continue
                if header_row_index is not None and row.get("raw_row_index") == header_row_index:
                    label = value
                    continue
                if value not in samples and len(samples) < 5:
                    samples.append(value)
            existing[column] = {
                "column": column,
                "index": column_index_from_string(column) - 1,
                "label": label,
                "detected_field": "ignore" if column in locked_ignore_columns else field_mapping.get(column, "ignore"),
                "is_price": locked_ignore_reasons.get(column) == "PRICE_AMOUNT_COLUMN",
                "locked_ignore": column in locked_ignore_columns,
                "lock_reason": locked_ignore_reasons.get(column),
                "sample_values": samples,
            }
        mapping["columns"] = sorted(existing.values(), key=lambda item: column_index_from_string(item["column"]))
        mapping["budget_locked_ignore_columns"] = sorted(
            locked_ignore_columns, key=column_index_from_string
        )
        mapping["budget_locked_ignore_reasons"] = {
            column: locked_ignore_reasons.get(column) or "PRICE_AMOUNT_COLUMN"
            for column in mapping["budget_locked_ignore_columns"]
        }
    return enriched


def _carry_budget_mapping_metadata(source: dict[str, Any], target: dict[str, Any]) -> dict[str, Any]:
    carried = copy.deepcopy(target)
    source_by_sheet = _mapping_by_sheet(source)
    for mapping in carried.get("sheet_mappings") or []:
        sheet_name = str(mapping.get("sheet_name") or "")
        source_mapping = source_by_sheet.get(sheet_name, {})
        mapping["sheet_role"] = source_mapping.get("sheet_role") or BUDGET_SHEET_ROLE_BILL
        mapping["preferred_quantity_column"] = source_mapping.get("preferred_quantity_column")
        mapping["layer_quantity_columns"] = source_mapping.get("layer_quantity_columns") or []
        mapping["budget_locked_ignore_reasons"] = copy.deepcopy(
            source_mapping.get("budget_locked_ignore_reasons") or {}
        )
        locked = {
            str(column).upper()
            for column in source_mapping.get("budget_locked_ignore_columns") or []
        }
        locked.update(
            str(column).upper()
            for column, field in (source_mapping.get("field_mapping") or {}).items()
            if field == "price_ignored"
        )
        locked.update(
            str(column.get("column") or "").upper()
            for column in source_mapping.get("columns") or []
            if column.get("locked_ignore") or column.get("is_price")
        )
        field_mapping = dict(mapping.get("field_mapping") or {})
        for column in locked:
            field_mapping[column] = "ignore"
        mapping["field_mapping"] = field_mapping
        mapping["budget_locked_ignore_columns"] = sorted(locked, key=column_index_from_string)
    return carried


def _carry_budget_row_annotations(source: dict[str, Any], target: dict[str, Any]) -> dict[str, Any]:
    """Keep workbook-derived multiplier annotations across manual remapping."""

    carried = copy.deepcopy(target)
    annotations: dict[tuple[str, int], dict[str, Any]] = {}
    for row in source.get("rows") or []:
        key = (str(row.get("source_sheet") or ""), int(row.get("raw_row_index") or 0))
        if key[0] and key[1] and any(field in row for field in _BUDGET_ROW_ANNOTATION_FIELDS):
            annotations[key] = row
    for row in carried.get("rows") or []:
        key = (str(row.get("source_sheet") or ""), int(row.get("raw_row_index") or 0))
        source_row = annotations.get(key)
        if not source_row:
            continue
        for field in _BUDGET_ROW_ANNOTATION_FIELDS:
            if field in source_row:
                row[field] = copy.deepcopy(source_row[field])
    return carried


def _sanitize_automatic_quantity_mappings(preview: dict[str, Any]) -> dict[str, Any]:
    requested: list[dict[str, Any]] = []
    changed = False
    for mapping in preview.get("sheet_mappings") or []:
        sheet_name = str(mapping.get("sheet_name") or "")
        field_mapping = dict(mapping.get("field_mapping") or {})
        guarded = _guarded_sequence_columns(preview, sheet_name)
        for column, field in list(field_mapping.items()):
            if field == "quantity" and str(column).upper() in guarded:
                field_mapping[column] = "ignore"
                changed = True
        requested.append({"sheet_name": sheet_name, "field_mapping": field_mapping})
    sanitized = apply_manual_field_mappings(preview, requested) if changed else preview
    if changed:
        sanitized = _carry_budget_mapping_metadata(preview, sanitized)
        sanitized = _ensure_budget_mapping_columns(sanitized)
    return _annotate_sequence_guards(sanitized)


def _validate_sheet_mapping_request(preview: dict[str, Any], sheet_mappings: list[dict[str, Any]]) -> None:
    names = [str(item.get("sheet_name") or "") for item in sheet_mappings]
    duplicates = sorted({name for name in names if name and names.count(name) > 1})
    if duplicates:
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail={"code": "DUPLICATE_BUDGET_IMPORT_SHEET_MAPPING", "sheets": duplicates},
        )
    known_sheets = set(_mapping_by_sheet(preview))
    unknown = sorted(set(names) - known_sheets)
    if unknown:
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail={"code": "UNKNOWN_BUDGET_IMPORT_SHEET", "sheets": unknown},
        )
    for item in sheet_mappings:
        sheet_name = str(item.get("sheet_name") or "")
        field_mapping = item.get("field_mapping") or {}
        sheet = _mapping_by_sheet(preview).get(sheet_name, {})
        locked_ignore_columns = {
            str(column).upper()
            for column in sheet.get("budget_locked_ignore_columns") or []
        }
        locked_ignore_columns.update(
            str(column).upper()
            for column, field in (sheet.get("field_mapping") or {}).items()
            if field == "price_ignored"
        )
        locked_ignore_columns.update(
            str(column.get("column") or "").upper()
            for column in sheet.get("columns") or []
            if column.get("locked_ignore") or column.get("is_price")
        )
        known_columns = {
            str(column.get("column") or "").upper()
            for column in sheet.get("columns") or []
            if column.get("column")
        }
        known_columns.update(str(column).upper() for column in (sheet.get("field_mapping") or {}))
        requested_columns = {str(column or "").strip().upper() for column in field_mapping}
        unknown_columns = sorted(requested_columns - known_columns)
        if unknown_columns:
            raise HTTPException(
                status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                detail={
                    "code": "UNKNOWN_BUDGET_MAPPING_COLUMN",
                    "sheet_name": sheet_name,
                    "columns": unknown_columns,
                },
            )
        unknown_fields = sorted(
            {
                str(value or "").strip()
                for column, value in field_mapping.items()
                if not (
                    str(column or "").strip().upper() in locked_ignore_columns
                    and str(value or "").strip() == "price_ignored"
                )
            }
            - _ALLOWED_MAPPING_FIELDS
        )
        if unknown_fields:
            raise HTTPException(
                status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                detail={
                    "code": "UNKNOWN_BUDGET_MAPPING_FIELD",
                    "sheet_name": sheet_name,
                    "fields": unknown_fields,
                },
            )
        invalid_locked_mappings = sorted(
            str(column or "").strip().upper()
            for column, value in field_mapping.items()
            if str(column or "").strip().upper() in locked_ignore_columns
            and str(value or "").strip() not in {"ignore", "price_ignored"}
        )
        if invalid_locked_mappings:
            raise HTTPException(
                status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                detail={
                    "code": "BUDGET_LOCKED_IGNORE_COLUMN",
                    "sheet_name": sheet_name,
                    "columns": invalid_locked_mappings,
                },
            )
        targets = [str(value) for value in field_mapping.values() if str(value) in _UNIQUE_MAPPING_FIELDS]
        duplicate_fields = sorted({value for value in targets if targets.count(value) > 1})
        if duplicate_fields:
            raise HTTPException(
                status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                detail={
                    "code": "DUPLICATE_BUDGET_FIELD_MAPPING",
                    "sheet_name": sheet_name,
                    "fields": duplicate_fields,
                },
            )
        guarded = _guarded_sequence_columns(preview, sheet_name)
        quantity_columns = [str(column).upper() for column, field in field_mapping.items() if field == "quantity"]
        blocked = sorted(set(quantity_columns) & guarded)
        if blocked:
            raise HTTPException(
                status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                detail={
                    "code": "SEQUENCE_COLUMN_CANNOT_BE_QUANTITY",
                    "sheet_name": sheet_name,
                    "columns": blocked,
                },
            )


def _quantity_column_metadata(
    row: dict[str, Any], sheet_mapping: dict[str, Any]
) -> tuple[list[str], str | None]:
    columns = _mapped_columns(row, sheet_mapping, "quantity")
    label = None
    if len(columns) == 1:
        label = _clean_text(_column_metadata(sheet_mapping, columns[0]).get("label"), 255)
    return columns, label


def _parser_quantity(value: Any) -> Decimal | None:
    if value is None or isinstance(value, bool):
        return None
    try:
        parsed = Decimal(str(value))
    except (InvalidOperation, TypeError, ValueError):
        return None
    if not parsed.is_finite():
        return None
    if abs(parsed) > _MAX_QUANTITY:
        return None
    try:
        return parsed.quantize(_QUANTITY_QUANTUM)
    except InvalidOperation:
        return None


def _budget_quantity(
    row: dict[str, Any], sheet_mapping: dict[str, Any]
) -> tuple[str, Decimal | None, Decimal, str, dict[str, Any]]:
    parser_quantity = _parser_quantity(row.get("quantity"))
    if row.get("row_type") != "data_row":
        return "", parser_quantity, Decimal("0"), "not_applicable", {}

    columns, label = _quantity_column_metadata(row, sheet_mapping)
    if not columns:
        return "", parser_quantity, Decimal("0"), "missing", {"method": "safe_zero", "reason": "NO_QUANTITY_COLUMN"}
    if len(columns) != 1:
        return "", parser_quantity, Decimal("0"), "abnormal", {"method": "safe_zero", "reason": "MULTIPLE_QUANTITY_COLUMNS"}
    column = columns[0]
    raw_quantity = _cell_value(row, column)
    source = {"column": column, "label": label or "", "method": "mapped_column"}
    if _is_sequence_label(label):
        return raw_quantity, parser_quantity, Decimal("0"), "sequence_column", {**source, "reason": "SEQUENCE_COLUMN"}
    if not raw_quantity:
        return raw_quantity, parser_quantity, Decimal("0"), "missing", {**source, "reason": "EMPTY_QUANTITY"}

    compact = raw_quantity.replace(",", "").strip()
    if _RANGE_RE.search(compact) or any(marker in compact for marker in _ABNORMAL_MARKERS):
        return raw_quantity, parser_quantity, Decimal("0"), "abnormal", {**source, "reason": "RANGE_OR_APPROXIMATE"}
    matches = _NUMBER_RE.findall(compact)
    if len(matches) > 1:
        return raw_quantity, parser_quantity, Decimal("0"), "abnormal", {**source, "reason": "MULTIPLE_NUMBERS"}
    match = _STRICT_QUANTITY_RE.fullmatch(raw_quantity)
    if not match:
        quantity_status = "non_numeric" if not matches else "abnormal"
        return raw_quantity, parser_quantity, Decimal("0"), quantity_status, {**source, "reason": "INVALID_QUANTITY"}
    try:
        value = Decimal(match.group(1).replace(",", ""))
    except InvalidOperation:
        return raw_quantity, parser_quantity, Decimal("0"), "non_numeric", {**source, "reason": "INVALID_QUANTITY"}
    if not value.is_finite() or value < 0 or value > _MAX_QUANTITY:
        return raw_quantity, parser_quantity, Decimal("0"), "abnormal", {**source, "reason": "OUT_OF_RANGE"}
    if value == 0:
        return raw_quantity, parser_quantity, Decimal("0"), "zero", {**source, "reason": "EXPLICIT_ZERO"}
    try:
        calculation_quantity = value.quantize(_QUANTITY_QUANTUM)
    except InvalidOperation:
        return raw_quantity, parser_quantity, Decimal("0"), "abnormal", {**source, "reason": "OUT_OF_RANGE"}
    if calculation_quantity <= 0:
        return (
            raw_quantity,
            parser_quantity,
            Decimal("0"),
            "precision_underflow",
            {**source, "reason": "BELOW_SUPPORTED_PRECISION"},
        )
    if calculation_quantity != value:
        return (
            raw_quantity,
            parser_quantity,
            Decimal("0"),
            "unsupported_precision",
            {**source, "reason": "UNSUPPORTED_QUANTITY_PRECISION"},
        )
    return raw_quantity, parser_quantity, calculation_quantity, "valid", source


def _row_key(row: dict[str, Any], sort_order: int) -> str:
    source_sheet = str(row.get("source_sheet") or "Sheet")
    raw_row_index = row.get("raw_row_index") or sort_order + 1
    return f"{source_sheet}:{raw_row_index}"[:255]


def _standard_row_model(
    *,
    batch_id: int,
    row: dict[str, Any],
    sort_order: int,
    mapping_revision: int,
    sheet_mapping: dict[str, Any],
) -> BudgetProjectStandardRow:
    effective_row = copy.deepcopy(row)
    if _is_repeated_header_row(effective_row, sheet_mapping):
        effective_row["row_type"] = "repeated_header"
        effective_row["requires_confirmation"] = False
        effective_row["warnings"] = list(
            dict.fromkeys([*(effective_row.get("warnings") or []), "BUDGET_REPEATED_HEADER_ROW"])
        )
    raw_quantity, parser_quantity, calculation_quantity, quantity_status, budget_source = _budget_quantity(
        effective_row, sheet_mapping
    )
    warnings = list(effective_row.get("warnings") or [])
    unit = _clean_text(effective_row.get("unit"), 64)
    unit_source = "parser" if unit else ""
    if effective_row.get("row_type") == "data_row" and not unit:
        unit = _clean_text(_mapped_raw_value(effective_row, sheet_mapping, "unit"), 64)
        if unit:
            unit_source = "mapped_raw_cell"
            warnings.append("BUDGET_UNIT_RECOVERED_FROM_MAPPED_CELL")
    if effective_row.get("row_type") == "data_row" and quantity_status != "valid":
        warnings.append("BUDGET_QUANTITY_DEFAULTED_TO_ZERO")
    warnings = list(dict.fromkeys(warnings))
    enriched_row = {
        **effective_row,
        "unit": unit or "",
        "budget_unit_source": unit_source,
        "raw_quantity": raw_quantity,
        "calculation_quantity": float(calculation_quantity),
        "quantity_status": quantity_status,
        "budget_quantity_source": budget_source,
        "warnings": warnings,
    }
    quantity_source = {**(effective_row.get("quantity_source") or {}), "budget": budget_source}
    return BudgetProjectStandardRow(
        batch_id=batch_id,
        row_key=_row_key(effective_row, sort_order),
        source_sheet=str(effective_row.get("source_sheet") or ""),
        sheet_role=str(
            effective_row.get("sheet_role")
            or sheet_mapping.get("sheet_role")
            or BUDGET_SHEET_ROLE_BILL
        ),
        raw_row_index=int(effective_row.get("raw_row_index") or sort_order + 1),
        sort_order=sort_order,
        mapping_revision=mapping_revision,
        row_type=str(effective_row.get("row_type") or "ambiguous_row"),
        is_standard_item=(
            effective_row.get("row_type") == "data_row"
            and (
                effective_row.get("sheet_role")
                or sheet_mapping.get("sheet_role")
                or BUDGET_SHEET_ROLE_BILL
            )
            == BUDGET_SHEET_ROLE_BILL
        ),
        item_name=_clean_text(effective_row.get("item_name"), 255),
        spec=_clean_text(effective_row.get("spec")),
        unit=unit,
        remark=_clean_text(effective_row.get("remark")),
        raw_quantity=raw_quantity or None,
        parser_quantity=parser_quantity,
        calculation_quantity=calculation_quantity,
        quantity_status=quantity_status,
        quantity_source_json=_json_dump(quantity_source),
        quantity_candidates_json=_json_dump(effective_row.get("quantity_candidates") or []),
        field_mapping_json=_json_dump(effective_row.get("field_mapping") or {}),
        raw_text=_clean_text(effective_row.get("raw_text")),
        raw_fields_json=_json_dump(effective_row.get("raw_fields") or {}),
        raw_cells_json=_json_dump(effective_row.get("raw_cells") or []),
        warnings_json=_json_dump(warnings),
        confidence=_clean_text(effective_row.get("confidence"), 24),
        requires_confirmation=bool(effective_row.get("requires_confirmation")),
        standard_row_json=_json_dump(enriched_row),
    )


def _batch_counts(rows: list[BudgetProjectStandardRow]) -> dict[str, int]:
    standard_rows = [row for row in rows if row.is_standard_item]
    valid_count = sum(row.quantity_status == "valid" for row in standard_rows)
    return {
        "total_output_row_count": len(rows),
        "standard_item_count": len(standard_rows),
        "valid_quantity_count": valid_count,
        "invalid_quantity_count": len(standard_rows) - valid_count,
    }


def _persist_sheet_mappings(
    db: Session,
    batch: BudgetProjectImportBatch,
    preview: dict[str, Any],
    *,
    initial: bool,
) -> None:
    existing_by_sheet = {item.sheet_name: item for item in batch.sheet_mappings}
    for mapping in preview.get("sheet_mappings") or []:
        sheet_name = str(mapping.get("sheet_name") or "")
        current_mapping_json = _json_dump(mapping.get("field_mapping") or {})
        current_columns_json = _json_dump(mapping.get("columns") or [])
        record = existing_by_sheet.get(sheet_name)
        if not record:
            record = BudgetProjectImportSheetMapping(
                batch_id=batch.id,
                sheet_name=sheet_name,
                sheet_role=str(mapping.get("sheet_role") or BUDGET_SHEET_ROLE_BILL),
                header_row_index=mapping.get("header_row_index"),
                detected_field_mapping_json=current_mapping_json,
                applied_field_mapping_json=current_mapping_json,
                detected_columns_json=current_columns_json,
                current_columns_json=current_columns_json,
                mapping_revision=batch.remap_revision,
            )
            db.add(record)
            continue
        record.header_row_index = mapping.get("header_row_index")
        record.sheet_role = str(mapping.get("sheet_role") or BUDGET_SHEET_ROLE_BILL)
        record.applied_field_mapping_json = current_mapping_json
        record.current_columns_json = current_columns_json
        record.mapping_revision = batch.remap_revision
        if initial:
            record.detected_field_mapping_json = current_mapping_json
            record.detected_columns_json = current_columns_json


def _replace_standard_rows(
    db: Session, batch: BudgetProjectImportBatch, preview: dict[str, Any]
) -> list[BudgetProjectStandardRow]:
    (
        db.query(BudgetProjectStandardRow)
        .filter(BudgetProjectStandardRow.batch_id == batch.id)
        .delete(synchronize_session=False)
    )
    mapping_by_sheet = _mapping_by_sheet(preview)
    rows = [
        _standard_row_model(
            batch_id=batch.id,
            row=row,
            sort_order=index,
            mapping_revision=batch.remap_revision,
            sheet_mapping=mapping_by_sheet.get(str(row.get("source_sheet") or ""), {}),
        )
        for index, row in enumerate(preview.get("rows") or [])
    ]
    db.add_all(rows)
    counts = _batch_counts(rows)
    for field, value in counts.items():
        setattr(batch, field, value)
    batch.sheet_count = int((preview.get("summary") or {}).get("sheet_count") or len(mapping_by_sheet))
    batch.current_preview_json = _json_dump(preview)
    batch.issues_json = _json_dump(preview.get("issues") or [])
    db.flush()
    return rows


def create_import_batch(
    db: Session,
    profile: BudgetProjectProfile,
    *,
    filename: str,
    content: bytes,
    current_user: User,
    source_file_object: FileObject | None = None,
) -> BudgetProjectImportBatch:
    require_profile_mutation(profile, current_user)
    require_active_profile(profile)
    if not content:
        raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_ENTITY, detail="BUDGET_IMPORT_FILE_EMPTY")
    if len(content) > MAX_IMPORT_BYTES:
        raise HTTPException(status_code=status.HTTP_413_REQUEST_ENTITY_TOO_LARGE, detail="BUDGET_IMPORT_FILE_TOO_LARGE")
    safe_filename = _clean_text(filename, 255) or "requirements.xlsx"
    _validate_workbook_limits(content, safe_filename)
    preview = standardize_requirement_excel_bytes(content, filename=safe_filename)
    preview = _apply_workbook_semantics(preview, content)
    preview = _ensure_budget_mapping_columns(preview)
    preview = _sanitize_automatic_quantity_mappings(preview)
    preview_json = _json_dump(preview)
    batch = BudgetProjectImportBatch(
        batch_uuid=str(uuid.uuid4()),
        project_id=profile.project_id,
        source_file_object_id=source_file_object.id if source_file_object else None,
        source_filename=safe_filename,
        source_file_sha256=hashlib.sha256(content).hexdigest(),
        source_file_size=len(content),
        source_storage_mode="file_object" if source_file_object else "metadata_only",
        parser_version=_clean_text(preview.get("version"), 64),
        status=BUDGET_IMPORT_STATUS_PARSED,
        remap_revision=0,
        original_preview_json=preview_json,
        current_preview_json=preview_json,
        issues_json=_json_dump(preview.get("issues") or []),
        created_by=current_user.id,
        updated_by=current_user.id,
    )
    db.add(batch)
    db.flush()
    _persist_sheet_mappings(db, batch, preview, initial=True)
    _replace_standard_rows(db, batch, preview)
    _append_import_revision(db, batch, current_user, revision_kind="initial")
    profile.updated_by = current_user.id
    profile.updated_at = app_local_naive()
    db.flush()
    return batch


def get_import_batch(
    db: Session, profile: BudgetProjectProfile, batch_identifier: str
) -> BudgetProjectImportBatch:
    query = db.query(BudgetProjectImportBatch).filter(
        BudgetProjectImportBatch.project_id == profile.project_id
    )
    if str(batch_identifier).isdigit():
        query = query.filter(BudgetProjectImportBatch.id == int(batch_identifier))
    else:
        query = query.filter(BudgetProjectImportBatch.batch_uuid == str(batch_identifier))
    batch = query.first()
    if not batch:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="BUDGET_IMPORT_BATCH_NOT_FOUND")
    return batch


def get_accessible_import_batch(
    db: Session, batch_identifier: str, current_user: User
) -> tuple[BudgetProjectProfile, BudgetProjectImportBatch]:
    query = db.query(BudgetProjectImportBatch)
    if str(batch_identifier).isdigit():
        query = query.filter(BudgetProjectImportBatch.id == int(batch_identifier))
    else:
        query = query.filter(BudgetProjectImportBatch.batch_uuid == str(batch_identifier))
    batch = query.first()
    if not batch:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="BUDGET_IMPORT_BATCH_NOT_FOUND")
    profile = get_budget_profile(db, batch.project_id, current_user)
    return profile, batch


def apply_import_sheet_mappings(
    db: Session,
    profile: BudgetProjectProfile,
    batch: BudgetProjectImportBatch,
    sheet_mappings: list[dict[str, Any]],
    current_user: User,
    *,
    expected_remap_revision: int,
) -> BudgetProjectImportBatch:
    require_profile_mutation(profile, current_user)
    require_active_profile(profile)
    batch = (
        db.query(BudgetProjectImportBatch)
        .filter(BudgetProjectImportBatch.id == batch.id)
        .populate_existing()
        .with_for_update()
        .one()
    )
    if batch.project_id != profile.project_id:
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail="BUDGET_IMPORT_PROJECT_MISMATCH",
        )
    if batch.status != BUDGET_IMPORT_STATUS_PARSED:
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail={
                "code": "BUDGET_IMPORT_REMAP_FROZEN",
                "status": batch.status,
            },
        )
    current_remap_revision = int(batch.remap_revision or 0)
    if int(expected_remap_revision) != current_remap_revision:
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail={
                "code": "BUDGET_IMPORT_REVISION_CONFLICT",
                "expected_remap_revision": int(expected_remap_revision),
                "current_remap_revision": current_remap_revision,
            },
        )
    current_preview = _json_load(batch.current_preview_json, {})
    original_preview = _json_load(batch.original_preview_json, {})
    _validate_sheet_mapping_request(current_preview, sheet_mappings)
    normalized_mappings = copy.deepcopy(sheet_mappings)
    for item in normalized_mappings:
        item["field_mapping"] = {
            column: "ignore" if field == "price_ignored" else field
            for column, field in (item.get("field_mapping") or {}).items()
        }
    remapped = apply_manual_field_mappings(current_preview, normalized_mappings)
    remapped = _carry_budget_mapping_metadata(current_preview, remapped)
    remapped = _carry_budget_row_annotations(original_preview, remapped)
    remapped = _annotate_formula_provenance(
        remapped, _formula_provenance_from_preview(current_preview)
    )
    remapped = _annotate_reference_sheet_rows(remapped)
    remapped = _ensure_budget_mapping_columns(remapped)
    remapped = _annotate_sequence_guards(remapped)
    batch.remap_revision = int(batch.remap_revision or 0) + 1
    batch.updated_by = current_user.id
    _persist_sheet_mappings(db, batch, remapped, initial=False)
    _replace_standard_rows(db, batch, remapped)
    _append_import_revision(db, batch, current_user, revision_kind="remap")
    profile.updated_by = current_user.id
    profile.updated_at = app_local_naive()
    db.flush()
    return batch


def _owned_revision(
    db: Session,
    batch: BudgetProjectImportBatch,
    revision_id: int | None,
) -> BudgetProjectImportRevision:
    revision = (
        db.query(BudgetProjectImportRevision)
        .filter(BudgetProjectImportRevision.id == revision_id)
        .first()
        if revision_id
        else None
    )
    if not revision or revision.batch_id != batch.id:
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail={
                "code": "BUDGET_IMPORT_REVISION_OWNERSHIP_INVALID",
                "batch_id": batch.id,
                "revision_id": revision_id,
            },
        )
    return revision


def confirm_import_batch(
    db: Session,
    profile: BudgetProjectProfile,
    batch: BudgetProjectImportBatch,
    current_user: User,
) -> BudgetProjectImportBatch:
    require_profile_mutation(profile, current_user)
    require_active_profile(profile)
    batch = (
        db.query(BudgetProjectImportBatch)
        .filter(BudgetProjectImportBatch.id == batch.id)
        .populate_existing()
        .with_for_update()
        .one()
    )
    if batch.project_id != profile.project_id:
        raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="BUDGET_IMPORT_PROJECT_MISMATCH")
    revision = _owned_revision(db, batch, batch.current_revision_id)
    if batch.status == BUDGET_IMPORT_STATUS_CONFIRMED:
        if batch.confirmed_revision_id == revision.id:
            return batch
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail="BUDGET_IMPORT_CONFIRMED_REVISION_MISMATCH",
        )
    if batch.status != BUDGET_IMPORT_STATUS_PARSED:
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail={"code": "BUDGET_IMPORT_CONFIRM_INVALID_STATUS", "status": batch.status},
        )
    previous_status = batch.status
    batch.status = BUDGET_IMPORT_STATUS_CONFIRMED
    batch.confirmed_revision_id = revision.id
    batch.confirmed_by = current_user.id
    batch.confirmed_at = app_local_naive()
    batch.updated_by = current_user.id
    _append_import_lifecycle_event(
        db,
        batch=batch,
        revision=revision,
        event_type="confirmed",
        from_status=previous_status,
        to_status=BUDGET_IMPORT_STATUS_CONFIRMED,
        current_user=current_user,
    )
    profile.updated_by = current_user.id
    profile.updated_at = app_local_naive()
    db.flush()
    return batch


def activate_import_batch(
    db: Session,
    profile: BudgetProjectProfile,
    batch: BudgetProjectImportBatch,
    current_user: User,
) -> BudgetProjectImportBatch:
    require_profile_mutation(profile, current_user)
    profile = (
        db.query(BudgetProjectProfile)
        .filter(BudgetProjectProfile.id == profile.id)
        .populate_existing()
        .with_for_update()
        .one()
    )
    require_active_profile(profile)
    batch = (
        db.query(BudgetProjectImportBatch)
        .filter(BudgetProjectImportBatch.id == batch.id)
        .populate_existing()
        .with_for_update()
        .one()
    )
    if batch.project_id != profile.project_id:
        raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="BUDGET_IMPORT_PROJECT_MISMATCH")
    if batch.status == BUDGET_IMPORT_STATUS_ACTIVE:
        if (
            profile.active_import_batch_id == batch.id
            and profile.active_import_revision_id == batch.confirmed_revision_id
        ):
            return batch
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail="BUDGET_ACTIVE_IMPORT_POINTER_INVALID",
        )
    if batch.status not in {
        BUDGET_IMPORT_STATUS_CONFIRMED,
        BUDGET_IMPORT_STATUS_SUPERSEDED,
    }:
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail={"code": "BUDGET_IMPORT_ACTIVATE_INVALID_STATUS", "status": batch.status},
        )
    revision = _owned_revision(db, batch, batch.confirmed_revision_id)
    if batch.current_revision_id != revision.id:
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail="BUDGET_IMPORT_CONFIRMED_REVISION_NOT_CURRENT",
        )

    active_batches = (
        db.query(BudgetProjectImportBatch)
        .filter(
            BudgetProjectImportBatch.project_id == profile.project_id,
            BudgetProjectImportBatch.status == BUDGET_IMPORT_STATUS_ACTIVE,
            BudgetProjectImportBatch.id != batch.id,
        )
        .with_for_update()
        .all()
    )
    now = app_local_naive()
    for previous in active_batches:
        previous_revision = _owned_revision(db, previous, previous.confirmed_revision_id)
        previous.status = BUDGET_IMPORT_STATUS_SUPERSEDED
        previous.superseded_by = current_user.id
        previous.superseded_at = now
        previous.updated_by = current_user.id
        _append_import_lifecycle_event(
            db,
            batch=previous,
            revision=previous_revision,
            event_type="superseded",
            from_status=BUDGET_IMPORT_STATUS_ACTIVE,
            to_status=BUDGET_IMPORT_STATUS_SUPERSEDED,
            current_user=current_user,
            event_data={"superseded_by_batch_id": batch.id},
        )

    previous_status = batch.status
    batch.status = BUDGET_IMPORT_STATUS_ACTIVE
    batch.activated_by = current_user.id
    batch.activated_at = now
    batch.superseded_by = None
    batch.superseded_at = None
    batch.updated_by = current_user.id
    _append_import_lifecycle_event(
        db,
        batch=batch,
        revision=revision,
        event_type="reactivated" if previous_status == BUDGET_IMPORT_STATUS_SUPERSEDED else "activated",
        from_status=previous_status,
        to_status=BUDGET_IMPORT_STATUS_ACTIVE,
        current_user=current_user,
        event_data={"previous_active_batch_id": profile.active_import_batch_id},
    )
    profile.active_import_batch_id = batch.id
    profile.active_import_revision_id = revision.id
    profile.updated_by = current_user.id
    profile.updated_at = now
    db.flush()
    return batch


def get_import_revision(
    db: Session,
    batch: BudgetProjectImportBatch,
    revision_identifier: str,
) -> BudgetProjectImportRevision:
    query = db.query(BudgetProjectImportRevision).filter(
        BudgetProjectImportRevision.batch_id == batch.id
    )
    if str(revision_identifier).isdigit():
        numeric = int(revision_identifier)
        revision = query.filter(BudgetProjectImportRevision.id == numeric).first()
        if revision is None:
            revision = query.filter(BudgetProjectImportRevision.revision_number == numeric).first()
    else:
        revision = query.filter(
            BudgetProjectImportRevision.revision_uuid == str(revision_identifier)
        ).first()
    if not revision:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="BUDGET_IMPORT_REVISION_NOT_FOUND",
        )
    return revision


def _external_field_mapping(value: str | None) -> str:
    return "ignore" if value == "price_ignored" else str(value or "ignore")


def _external_mapping_columns(
    columns: list[dict[str, Any]], field_mapping: dict[str, Any]
) -> list[dict[str, Any]]:
    serialized = copy.deepcopy(columns)
    by_column = {
        str(item.get("column") or "").upper(): item
        for item in serialized
        if item.get("column")
    }
    locked_reasons = {
        str(column).upper(): "PRICE_AMOUNT_COLUMN"
        for column, field in field_mapping.items()
        if field == "price_ignored"
    }
    for source in serialized:
        if not (source.get("locked_ignore") or source.get("is_price")):
            continue
        column = str(source.get("column") or "").upper()
        if not column:
            continue
        locked_reasons[column] = str(
            source.get("lock_reason")
            or ("PRICE_AMOUNT_COLUMN" if source.get("is_price") else "LOCKED_IGNORE_COLUMN")
        )
    for column, reason in locked_reasons.items():
        item = by_column.get(column)
        if item is None:
            item = {
                "column": column,
                "index": column_index_from_string(column) - 1,
                "label": "",
                "sample_values": [],
            }
            serialized.append(item)
            by_column[column] = item
        item["detected_field"] = "ignore"
        item["is_price"] = reason == "PRICE_AMOUNT_COLUMN"
        item["locked_ignore"] = True
        item["lock_reason"] = reason
    return sorted(serialized, key=lambda item: column_index_from_string(item["column"]))


def serialize_sheet_mapping(mapping: BudgetProjectImportSheetMapping) -> dict[str, Any]:
    detected_mapping = _json_load(mapping.detected_field_mapping_json, {})
    applied_mapping = _json_load(mapping.applied_field_mapping_json, {})
    detected_columns = _external_mapping_columns(
        _json_load(mapping.detected_columns_json, []), detected_mapping
    )
    current_columns = _external_mapping_columns(
        _json_load(mapping.current_columns_json, []), applied_mapping
    )
    return {
        "sheet_name": mapping.sheet_name,
        "sheet_role": mapping.sheet_role,
        "header_row_index": mapping.header_row_index,
        "detected_field_mapping": {
            column: _external_field_mapping(field) for column, field in detected_mapping.items()
        },
        "applied_field_mapping": {
            column: _external_field_mapping(field) for column, field in applied_mapping.items()
        },
        "detected_columns": detected_columns,
        "current_columns": current_columns,
        "mapping_revision": mapping.mapping_revision,
    }


def serialize_import_revision(
    revision: BudgetProjectImportRevision | None,
    *,
    include_snapshot: bool = False,
) -> dict[str, Any] | None:
    if revision is None:
        return None
    data: dict[str, Any] = {
        "id": revision.id,
        "revision_uuid": revision.revision_uuid,
        "batch_id": revision.batch_id,
        "revision_number": revision.revision_number,
        "revision_kind": revision.revision_kind,
        "snapshot_sha256": revision.snapshot_sha256,
        "summary": _json_load(revision.summary_json, {}),
        "created_by": revision.created_by,
        "created_at": _format_dt(revision.created_at),
    }
    if include_snapshot:
        data.update(
            {
                "preview": _json_load(revision.preview_json, {}),
                "sheet_mappings": _json_load(revision.sheet_mappings_json, []),
                "standard_rows": _json_load(revision.standard_rows_json, []),
            }
        )
    return data


def serialize_import_batch(
    batch: BudgetProjectImportBatch | None,
    *,
    include_preview: bool = False,
    profile: BudgetProjectProfile | None = None,
    current_user: User | None = None,
) -> dict[str, Any] | None:
    if batch is None:
        return None
    current_preview = _json_load(batch.current_preview_json, {})
    reference_row_count = sum(
        1
        for row in current_preview.get("rows", [])
        if isinstance(row, dict)
        and row.get("sheet_role") in _REFERENCE_SHEET_ROLES
        and row.get("row_type") == "reference_row"
    ) if isinstance(current_preview, dict) else 0
    file_object = batch.source_file_object
    source_file = {
        "filename": batch.source_filename,
        "sha256": batch.source_file_sha256,
        "size": batch.source_file_size,
        "storage_mode": batch.source_storage_mode,
        "file_object_id": file_object.file_id if file_object else None,
        "download_available": bool(file_object),
        "download_limitation": None
        if file_object
        else "Phase 1 retains filename, SHA256, mappings and parsed row snapshots; the original binary is metadata-only.",
    }
    can_mutate = bool(
        profile is not None
        and current_user is not None
        and _can_mutate_budget_profile(profile, current_user)
    )
    pointer_matches = bool(
        profile is None
        or (
            profile.active_import_batch_id == batch.id
            and profile.active_import_revision_id == batch.confirmed_revision_id
        )
    )
    data: dict[str, Any] = {
        "id": batch.id,
        "batch_uuid": batch.batch_uuid,
        "project_id": batch.project_id,
        "status": batch.status,
        "is_active": batch.status == BUDGET_IMPORT_STATUS_ACTIVE and pointer_matches,
        "is_confirmed": batch.confirmed_revision_id is not None,
        "parser_version": batch.parser_version,
        "remap_revision": batch.remap_revision,
        "current_revision_id": batch.current_revision_id,
        "confirmed_revision_id": batch.confirmed_revision_id,
        "current_revision": serialize_import_revision(batch.current_revision),
        "confirmed_revision": serialize_import_revision(batch.confirmed_revision),
        "revision_count": len(batch.revisions),
        "lifecycle_event_count": len(batch.lifecycle_events),
        "source_file": source_file,
        "summary": {
            "sheet_count": batch.sheet_count,
            "total_output_row_count": batch.total_output_row_count,
            "standard_item_count": batch.standard_item_count,
            "valid_quantity_count": batch.valid_quantity_count,
            "invalid_quantity_count": batch.invalid_quantity_count,
            "reference_row_count": reference_row_count,
        },
        "sheet_mappings": [serialize_sheet_mapping(item) for item in batch.sheet_mappings],
        "issues": _json_load(batch.issues_json, []),
        "created_by": batch.created_by,
        "confirmed_by": batch.confirmed_by,
        "confirmed_at": _format_dt(batch.confirmed_at),
        "activated_by": batch.activated_by,
        "activated_at": _format_dt(batch.activated_at),
        "superseded_by": batch.superseded_by,
        "superseded_at": _format_dt(batch.superseded_at),
        "created_at": _format_dt(batch.created_at),
        "updated_at": _format_dt(batch.updated_at),
        "capabilities": {
            "can_remap": can_mutate and batch.status == BUDGET_IMPORT_STATUS_PARSED,
            "can_confirm": bool(
                can_mutate
                and batch.status == BUDGET_IMPORT_STATUS_PARSED
                and batch.current_revision_id
            ),
            "can_activate": bool(
                can_mutate
                and batch.status in {
                    BUDGET_IMPORT_STATUS_CONFIRMED,
                    BUDGET_IMPORT_STATUS_SUPERSEDED,
                }
                and batch.confirmed_revision_id
            ),
        },
    }
    if include_preview:
        data["current_preview"] = current_preview
    return data


def serialize_standard_row(row: BudgetProjectStandardRow) -> dict[str, Any]:
    parser_quantity = float(row.parser_quantity) if row.parser_quantity is not None else None
    quantity_source = _json_load(row.quantity_source_json, {})
    budget_quantity_source = quantity_source.get("budget") or {}
    quantity_reason = budget_quantity_source.get("reason") or (
        "VALID_SOURCE_QUANTITY" if row.quantity_status == "valid" else str(row.quantity_status or "").upper()
    )
    return {
        "id": row.id,
        "project_id": row.batch.project_id,
        "batch_id": row.batch_id,
        "row_key": row.row_key,
        "source_sheet": row.source_sheet,
        "sheet_role": row.sheet_role,
        "is_reference_item": row.sheet_role == BUDGET_SHEET_ROLE_MATERIAL_REFERENCE,
        "raw_row_index": row.raw_row_index,
        "sort_order": row.sort_order,
        "mapping_revision": row.mapping_revision,
        "row_type": row.row_type,
        "is_standard_item": row.is_standard_item,
        "item_name": row.item_name,
        "spec": row.spec,
        "unit": row.unit,
        "remark": row.remark,
        "raw_quantity": row.raw_quantity or "",
        "parser_quantity": parser_quantity,
        "calculation_quantity": float(row.calculation_quantity or 0),
        "quantity_status": row.quantity_status,
        "quantity_reason": quantity_reason,
        "status_reason": quantity_reason,
        "quantity_source": quantity_source,
        "quantity_candidates": _json_load(row.quantity_candidates_json, []),
        "field_mapping": _json_load(row.field_mapping_json, {}),
        "raw_text": row.raw_text or "",
        "raw_fields": _json_load(row.raw_fields_json, {}),
        "raw_cells": _json_load(row.raw_cells_json, []),
        "warnings": _json_load(row.warnings_json, []),
        "confidence": row.confidence,
        "requires_confirmation": row.requires_confirmation,
        "standard_row": _json_load(row.standard_row_json, {}),
    }
