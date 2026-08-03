from __future__ import annotations

import csv
import json
import re
import shutil
from collections import defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any, Iterable, Mapping, Sequence

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


PHASE = "BIZ-2x-pdf-gap-review-evidence-pack"

MANIFEST_HEADERS = [
    "gap_no",
    "gap_priority",
    "gap_type",
    "section",
    "answer_sheet",
    "answer_row_no",
    "answer_item_name",
    "answer_feature",
    "answer_unit",
    "candidate_row_no",
    "candidate_item_name",
    "candidate_unit",
    "candidate_standard_item_code",
    "candidate_standard_item_name",
    "candidate_evidence_ids",
    "suggested_next_action",
    "evidence_id",
    "evidence_type",
    "raw_item_name",
    "spec_or_method",
    "evidence_text",
    "source_file",
    "page",
    "tile_id",
    "source_report_file",
    "bbox_pdf",
    "bbox_pixel",
    "tile_image_path",
    "review_image_path",
]


def build_gap_review_evidence_pack(
    v2_report: Mapping[str, Any],
    *,
    source_report_dir: str | Path | None = None,
    priority_prefixes: Sequence[str] | None = None,
    max_gaps: int | None = None,
) -> dict[str, Any]:
    prefixes = tuple(priority_prefixes or ("P1",))
    source_dir = Path(source_report_dir) if source_report_dir else None
    evidence_by_id = {str(row.get("evidence_id") or ""): row for row in v2_report.get("evidence_rows") or []}
    tile_map_by_report = _load_tile_maps_by_source_report(v2_report, source_report_dir=source_dir)

    selected_gaps: list[Mapping[str, Any]] = []
    for gap in v2_report.get("three_field_gap_rows") or []:
        priority = str(gap.get("gap_priority") or "")
        if prefixes and not priority.startswith(prefixes):
            continue
        selected_gaps.append(gap)
        if max_gaps is not None and len(selected_gaps) >= max_gaps:
            break

    manifest_rows: list[dict[str, Any]] = []
    for gap_index, gap in enumerate(selected_gaps, start=1):
        evidence_ids = _split_evidence_ids(gap.get("candidate_evidence_ids"))
        if not evidence_ids:
            manifest_rows.append(_manifest_row(gap_index, gap, None, tile_map_by_report))
            continue
        for evidence_id in evidence_ids:
            manifest_rows.append(_manifest_row(gap_index, gap, evidence_by_id.get(evidence_id), tile_map_by_report))

    return {
        "ok": True,
        "phase": PHASE,
        "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "summary": {
            "gap_count": len(selected_gaps),
            "manifest_row_count": len(manifest_rows),
            "priority_prefixes": list(prefixes),
        },
        "manifest_rows": manifest_rows,
    }


def write_gap_review_evidence_pack(
    pack: Mapping[str, Any],
    output_dir: str | Path,
    *,
    stem: str | None = None,
    copy_images: bool = True,
) -> dict[str, str]:
    target = Path(output_dir)
    target.mkdir(parents=True, exist_ok=True)
    file_stem = stem or f"BIZ2x_PDF_三字段缺口证据包_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
    image_dir = target / f"{file_stem}_review_images"
    manifest_rows = [dict(row) for row in pack.get("manifest_rows") or []]
    copied_count = 0
    if copy_images:
        image_dir.mkdir(parents=True, exist_ok=True)
        copied_count = _copy_review_images(manifest_rows, image_dir=image_dir)

    report = {**dict(pack), "summary": {**dict(pack.get("summary") or {}), "copied_image_count": copied_count}}
    json_path = target / f"{file_stem}.json"
    csv_path = target / f"{file_stem}_manifest.csv"
    markdown_path = target / f"{file_stem}.md"
    xlsx_path = target / f"{file_stem}.xlsx"
    outputs = {
        "json": str(json_path),
        "manifest_csv": str(csv_path),
        "markdown": str(markdown_path),
        "xlsx": str(xlsx_path),
        "review_image_dir": str(image_dir),
    }
    report["outputs"] = outputs

    json_path.write_text(json.dumps(_json_safe(report), ensure_ascii=False, indent=2), encoding="utf-8")
    _write_csv(csv_path, manifest_rows, MANIFEST_HEADERS)
    markdown_path.write_text(_build_markdown(report), encoding="utf-8")
    _write_workbook(xlsx_path, report)
    return outputs


def _manifest_row(
    gap_no: int,
    gap: Mapping[str, Any],
    evidence: Mapping[str, Any] | None,
    tile_map_by_report: Mapping[str, Mapping[tuple[str, str], Mapping[str, Any]]],
) -> dict[str, Any]:
    raw = evidence.get("raw") if evidence else {}
    raw = raw if isinstance(raw, Mapping) else {}
    source_report_file = str(raw.get("source_report_file") or "")
    source_file = str((evidence or {}).get("source_file") or raw.get("source_file") or "")
    tile_id = str((evidence or {}).get("tile_id") or raw.get("tile_id") or "")
    tile = (tile_map_by_report.get(source_report_file) or {}).get((source_file, tile_id), {})
    tile_image_path = str(tile.get("image_path") or "")
    return {
        "gap_no": gap_no,
        "gap_priority": gap.get("gap_priority", ""),
        "gap_type": gap.get("gap_type", ""),
        "section": gap.get("section", ""),
        "answer_sheet": gap.get("answer_sheet", ""),
        "answer_row_no": gap.get("answer_row_no", ""),
        "answer_item_name": gap.get("answer_item_name", ""),
        "answer_feature": gap.get("answer_feature", ""),
        "answer_unit": gap.get("answer_unit", ""),
        "candidate_row_no": gap.get("candidate_row_no", ""),
        "candidate_item_name": gap.get("candidate_item_name", ""),
        "candidate_unit": gap.get("candidate_unit", ""),
        "candidate_standard_item_code": gap.get("candidate_standard_item_code", ""),
        "candidate_standard_item_name": gap.get("candidate_standard_item_name", ""),
        "candidate_evidence_ids": gap.get("candidate_evidence_ids", ""),
        "suggested_next_action": gap.get("suggested_next_action", ""),
        "evidence_id": (evidence or {}).get("evidence_id", ""),
        "evidence_type": (evidence or {}).get("evidence_type", ""),
        "raw_item_name": (evidence or {}).get("raw_item_name", ""),
        "spec_or_method": (evidence or {}).get("spec_or_method", ""),
        "evidence_text": (evidence or {}).get("evidence_text", ""),
        "source_file": source_file,
        "page": (evidence or {}).get("page") or raw.get("page", ""),
        "tile_id": tile_id,
        "source_report_file": source_report_file,
        "bbox_pdf": json.dumps(raw.get("bbox_pdf") or [], ensure_ascii=False),
        "bbox_pixel": json.dumps(raw.get("bbox_pixel") or [], ensure_ascii=False),
        "tile_image_path": tile_image_path,
        "review_image_path": "",
    }


def _load_tile_maps_by_source_report(
    v2_report: Mapping[str, Any],
    *,
    source_report_dir: Path | None,
) -> dict[str, dict[tuple[str, str], Mapping[str, Any]]]:
    source_files = sorted(
        {
            str(((row.get("raw") if isinstance(row.get("raw"), Mapping) else {}) or {}).get("source_report_file") or "")
            for row in v2_report.get("evidence_rows") or []
            if isinstance(row, Mapping)
        }
        - {""}
    )
    result: dict[str, dict[tuple[str, str], Mapping[str, Any]]] = {}
    for file_name in source_files:
        report_path = _find_source_report_path(file_name, source_report_dir=source_report_dir)
        if not report_path:
            result[file_name] = {}
            continue
        try:
            source_report = json.loads(report_path.read_text(encoding="utf-8"))
        except (OSError, json.JSONDecodeError):
            result[file_name] = {}
            continue
        tile_rows = ((source_report.get("tile_report") or {}).get("tile_rows") or [])
        result[file_name] = {
            (str(row.get("source_file") or ""), str(row.get("tile_id") or "")): row
            for row in tile_rows
            if isinstance(row, Mapping)
        }
    return result


def _find_source_report_path(file_name: str, *, source_report_dir: Path | None) -> Path | None:
    if not file_name:
        return None
    path = Path(file_name)
    if path.exists():
        return path
    if source_report_dir:
        candidate = source_report_dir / file_name
        if candidate.exists():
            return candidate
        matches = list(source_report_dir.glob(file_name))
        if matches:
            return matches[0]
    return None


def _copy_review_images(rows: list[dict[str, Any]], *, image_dir: Path) -> int:
    copied = 0
    seen: dict[str, str] = {}
    for index, row in enumerate(rows, start=1):
        image_path = str(row.get("tile_image_path") or "")
        if not image_path:
            continue
        source = Path(image_path)
        if not source.exists() or not source.is_file():
            continue
        if image_path in seen:
            row["review_image_path"] = seen[image_path]
            continue
        target_name = f"{index:04d}_{_safe_name(row.get('gap_priority'))}_{_safe_name(row.get('answer_item_name'))}_{_safe_name(row.get('evidence_id'))}{source.suffix.lower() or '.png'}"
        target = image_dir / target_name[:180]
        shutil.copy2(source, target)
        row["review_image_path"] = str(target)
        seen[image_path] = str(target)
        copied += 1
    return copied


def _split_evidence_ids(value: Any) -> list[str]:
    return [item.strip() for item in re.split(r"[;,；，\s]+", str(value or "")) if item.strip()]


def _safe_name(value: Any) -> str:
    text = re.sub(r"[^\w\u4e00-\u9fff.-]+", "_", str(value or "").strip(), flags=re.UNICODE).strip("_")
    return text or "na"


def _build_markdown(report: Mapping[str, Any]) -> str:
    summary = report.get("summary") or {}
    lines = [
        "# BIZ-2x PDF 三字段缺口证据包",
        "",
        f"- 生成时间：{report.get('generated_at', '-')}",
        f"- 缺口数：{summary.get('gap_count', 0)}",
        f"- manifest 行数：{summary.get('manifest_row_count', 0)}",
        f"- 已复制本地图片：{summary.get('copied_image_count', 0)}",
        "",
        "| 优先级 | 缺口类型 | 人工项目 | 候选 | 证据 | 图片 |",
        "| --- | --- | --- | --- | --- | --- |",
    ]
    for row in (report.get("manifest_rows") or [])[:80]:
        lines.append(
            "| "
            + " | ".join(
                [
                    _md(row.get("gap_priority")),
                    _md(row.get("gap_type")),
                    _md(row.get("answer_item_name")),
                    _md(row.get("candidate_item_name")),
                    _md(row.get("evidence_id")),
                    _md(row.get("review_image_path")),
                ]
            )
            + " |"
        )
    return "\n".join(lines) + "\n"


def _write_workbook(path: Path, report: Mapping[str, Any]) -> None:
    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "gap_summary"
    _append_summary_rows(summary_sheet, report)
    _style_sheet(summary_sheet)

    manifest_sheet = workbook.create_sheet("evidence_manifest")
    _append_rows(manifest_sheet, [MANIFEST_HEADERS])
    _append_rows(
        manifest_sheet,
        [[_cell_value(row.get(header)) for header in MANIFEST_HEADERS] for row in report.get("manifest_rows") or []],
    )
    _style_sheet(manifest_sheet)
    workbook.save(path)


def _append_summary_rows(sheet: Any, report: Mapping[str, Any]) -> None:
    summary = report.get("summary") or {}
    manifest_rows = [row for row in report.get("manifest_rows") or [] if isinstance(row, Mapping)]
    rows: list[list[Any]] = [
        ["metric", "value"],
        ["gap_count", summary.get("gap_count", 0)],
        ["manifest_row_count", summary.get("manifest_row_count", 0)],
        ["copied_image_count", summary.get("copied_image_count", 0)],
        ["priority_prefixes", ", ".join(str(item) for item in summary.get("priority_prefixes") or [])],
        [],
        ["gap_priority", "gap_count"],
    ]
    rows.extend(_count_unique_gaps(manifest_rows, "gap_priority"))
    rows.extend([[], ["gap_type", "gap_count"]])
    rows.extend(_count_unique_gaps(manifest_rows, "gap_type"))
    rows.extend([[], ["section", "gap_count"]])
    rows.extend(_count_unique_gaps(manifest_rows, "section"))
    _append_rows(sheet, rows)


def _count_unique_gaps(rows: Sequence[Mapping[str, Any]], key: str) -> list[list[Any]]:
    gap_numbers: dict[str, set[Any]] = defaultdict(set)
    for row in rows:
        value = str(row.get(key) or "(blank)")
        gap_numbers[value].add(row.get("gap_no"))
    return [[value, len(gaps)] for value, gaps in sorted(gap_numbers.items(), key=lambda item: (-len(item[1]), item[0]))]


def _append_rows(sheet: Any, rows: Iterable[Sequence[Any]]) -> None:
    for row in rows:
        sheet.append(list(row))


def _style_sheet(sheet: Any) -> None:
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for col_idx, column_cells in enumerate(sheet.columns, start=1):
        values = [str(cell.value or "") for cell in column_cells[:80]]
        width = min(max([len(value) for value in values] + [10]) + 2, 80)
        sheet.column_dimensions[get_column_letter(col_idx)].width = width
    sheet.freeze_panes = "A2"


def _write_csv(path: Path, rows: Iterable[Mapping[str, Any]], fieldnames: Sequence[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        for row in rows:
            writer.writerow({key: _cell_value(row.get(key)) for key in fieldnames})


def _cell_value(value: Any) -> Any:
    if isinstance(value, (dict, list, tuple)):
        return json.dumps(value, ensure_ascii=False)
    return value


def _json_safe(value: Any) -> Any:
    if isinstance(value, Mapping):
        return {str(key): _json_safe(item) for key, item in value.items()}
    if isinstance(value, (list, tuple)):
        return [_json_safe(item) for item in value]
    if isinstance(value, Path):
        return str(value)
    return value


def _md(value: Any) -> str:
    return str(value or "").replace("|", "\\|").replace("\n", "<br>")
