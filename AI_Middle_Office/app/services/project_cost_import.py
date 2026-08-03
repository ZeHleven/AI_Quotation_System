from __future__ import annotations

import hashlib
import json
import re
import statistics
import uuid
import zipfile
from collections import defaultdict
from datetime import date, datetime
from difflib import SequenceMatcher
from io import BytesIO
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook
from sqlalchemy import func, or_
from sqlalchemy.orm import Session

from app.models.enterprise_quota import (
    IMPORT_BATCH_STATUS_IMPORTED,
    QUOTA_VERSION_STATUS_DRAFT,
    CostImportBatch,
    EnterpriseCostResource,
    EnterpriseQuotaComponent,
    EnterpriseQuotaItem,
    EnterpriseQuotaSection,
    EnterpriseQuotaVersion,
)
from app.models.project_cost_import import (
    CANDIDATE_STATUS_APPROVED,
    CANDIDATE_STATUS_PENDING,
    CANDIDATE_STATUS_REJECTED,
    IMPORT_STATUS_DRAFT_CREATED,
    IMPORT_STATUS_REVIEWING,
    EnterpriseResourcePriceObservation,
    ProjectCostImportBatch,
    ProjectCostPriceCandidate,
)
from app.services.enterprise_quota_cost_reference import active_enterprise_quota_version
from app.services.enterprise_quota_units import normalize_enterprise_quota_unit


PARSER_VERSION = "project-purchase-mvp-v1"
SUPPORTED_EXTENSIONS = {".xlsx", ".xlsm"}
MAX_ARCHIVE_FILES = 800
MAX_ARCHIVE_UNCOMPRESSED_BYTES = 512 * 1024 * 1024
MAX_TOTAL_UPLOAD_BYTES = 256 * 1024 * 1024

HEADER_ALIASES = {
    "name": ("材料名称", "物料名称", "产品名称", "品名", "货物名称", "商品名称", "项目名称"),
    "brand": ("品牌", "厂家", "生产厂家"),
    "spec": ("规格型号", "型号规格", "规格/型号", "型号/规格", "规格", "型号", "材质"),
    "brand_spec": ("型号/品牌", "品牌/型号"),
    "unit": ("单位", "计量单位"),
    "quantity": ("数量", "采购数量", "订购数量", "工程量"),
    "unit_price": ("单价", "采购单价", "含税单价", "综合单价", "报价"),
    "amount": ("金额", "合价", "总价", "小计"),
    "remark": ("备注", "说明"),
}
SERVICE_FEE_WORDS = ("运费", "税金", "车费", "加工费", "搬运费", "装卸费", "管理费", "服务费")


class ProjectCostImportError(RuntimeError):
    pass


def parse_project_purchase_files(
    files: Iterable[tuple[str, bytes]], *, max_total_bytes: int | None = MAX_TOTAL_UPLOAD_BYTES
) -> dict[str, Any]:
    observations: list[dict[str, Any]] = []
    manifest: list[dict[str, Any]] = []
    for filename, content in _expand_files(files, max_total_bytes=max_total_bytes):
        sha256 = hashlib.sha256(content).hexdigest()
        entry = {
            "file_name": filename,
            "sha256": sha256,
            "size": len(content),
            "status": "skipped",
            "sheet_count": 0,
            "observation_count": 0,
            "issues": [],
        }
        try:
            parsed = _parse_workbook(filename, content, sha256)
            entry.update(parsed["manifest"])
            observations.extend(parsed["observations"])
        except Exception as exc:
            entry["status"] = "failed"
            entry["issues"] = [str(exc)[:1000]]
        manifest.append(entry)

    usable = [row for row in observations if float(row.get("unit_price") or 0) > 0 and not row.get("excluded_reason")]
    return {
        "parser_version": PARSER_VERSION,
        "manifest": manifest,
        "observations": observations,
        "summary": {
            "file_count": len(manifest),
            "parsed_file_count": sum(1 for row in manifest if row["status"] == "parsed"),
            "skipped_file_count": sum(1 for row in manifest if row["status"] != "parsed"),
            "observation_count": len(observations),
            "usable_price_observation_count": len(usable),
            "failed_file_count": sum(1 for row in manifest if row["status"] == "failed"),
        },
    }


def parse_project_purchase_directory(source_dir: str | Path) -> dict[str, Any]:
    root = Path(source_dir).expanduser().resolve()
    if not root.is_dir() or root.parent == root:
        raise ProjectCostImportError("INVALID_SOURCE_DIRECTORY")
    paths = sorted(path for path in root.rglob("*") if path.is_file() and path.suffix.lower() in SUPPORTED_EXTENSIONS)
    if not paths:
        raise ProjectCostImportError("NO_SUPPORTED_EXCEL_FILES")
    return parse_project_purchase_files(
        ((str(path.relative_to(root)), path.read_bytes()) for path in paths),
        max_total_bytes=None,
    )


def create_project_cost_import_batch(
    db: Session,
    *,
    project_name: str,
    source_name: str | None,
    files: Iterable[tuple[str, bytes]],
    actor_user_id: int,
    max_total_bytes: int | None = MAX_TOTAL_UPLOAD_BYTES,
) -> ProjectCostImportBatch:
    project = _clean(project_name, 255)
    if not project:
        raise ProjectCostImportError("PROJECT_NAME_REQUIRED")
    parsed = parse_project_purchase_files(files, max_total_bytes=max_total_bytes)
    if not parsed["summary"]["observation_count"]:
        raise ProjectCostImportError("NO_PURCHASE_PRICE_ROWS_RECOGNIZED")

    summary = parsed["summary"]
    batch = ProjectCostImportBatch(
        batch_uuid=str(uuid.uuid4()),
        project_name=project,
        source_name=_clean(source_name, 255),
        status=IMPORT_STATUS_REVIEWING,
        parser_version=PARSER_VERSION,
        file_count=summary["file_count"],
        parsed_file_count=summary["parsed_file_count"],
        skipped_file_count=summary["skipped_file_count"],
        observation_count=summary["observation_count"],
        source_manifest_json=_json(parsed["manifest"]),
        created_by=actor_user_id,
    )
    db.add(batch)
    db.flush()

    for row in parsed["observations"]:
        batch.observations.append(EnterpriseResourcePriceObservation(**row))
    db.flush()

    candidate_rows = _build_candidates(db, batch)
    for row in candidate_rows:
        batch.candidates.append(ProjectCostPriceCandidate(**row))
    db.flush()
    batch.candidate_count = len(candidate_rows)
    batch.high_confidence_count = sum(1 for row in candidate_rows if row["confidence_score"] >= 0.75)
    batch.summary_json = _json({**summary, "candidate_count": batch.candidate_count, "high_confidence_count": batch.high_confidence_count})
    db.flush()
    return batch


def list_project_cost_import_batches(
    db: Session, *, page: int, page_size: int, status: str | None = None, keyword: str | None = None
) -> tuple[list[ProjectCostImportBatch], int]:
    query = db.query(ProjectCostImportBatch)
    if status:
        query = query.filter(ProjectCostImportBatch.status == status)
    if keyword and keyword.strip():
        pattern = f"%{keyword.strip()}%"
        query = query.filter(or_(ProjectCostImportBatch.project_name.like(pattern), ProjectCostImportBatch.source_name.like(pattern)))
    total = query.count()
    rows = query.order_by(ProjectCostImportBatch.created_at.desc(), ProjectCostImportBatch.id.desc()).offset((page - 1) * page_size).limit(page_size).all()
    return rows, total


def list_project_cost_candidates(
    db: Session,
    *,
    batch_id: int,
    page: int,
    page_size: int,
    status: str | None = None,
    risk_level: str | None = None,
    keyword: str | None = None,
) -> tuple[list[ProjectCostPriceCandidate], int]:
    query = db.query(ProjectCostPriceCandidate).filter(ProjectCostPriceCandidate.batch_id == batch_id)
    if status:
        query = query.filter(ProjectCostPriceCandidate.status == status)
    if risk_level:
        query = query.filter(ProjectCostPriceCandidate.risk_level == risk_level)
    if keyword and keyword.strip():
        pattern = f"%{keyword.strip()}%"
        query = query.filter(or_(
            ProjectCostPriceCandidate.normalized_item_name.like(pattern),
            ProjectCostPriceCandidate.brand.like(pattern),
            ProjectCostPriceCandidate.spec.like(pattern),
        ))
    total = query.count()
    rows = query.order_by(ProjectCostPriceCandidate.confidence_score.desc(), ProjectCostPriceCandidate.id.asc()).offset((page - 1) * page_size).limit(page_size).all()
    return rows, total


def list_price_observations(
    db: Session, *, batch_id: int, candidate_key: str | None, page: int, page_size: int
) -> tuple[list[EnterpriseResourcePriceObservation], int]:
    query = db.query(EnterpriseResourcePriceObservation).filter(EnterpriseResourcePriceObservation.batch_id == batch_id)
    if candidate_key:
        query = query.filter(EnterpriseResourcePriceObservation.candidate_key == candidate_key)
    total = query.count()
    rows = query.order_by(EnterpriseResourcePriceObservation.observed_at.desc(), EnterpriseResourcePriceObservation.id.desc()).offset((page - 1) * page_size).limit(page_size).all()
    return rows, total


def review_project_cost_candidates(
    db: Session,
    *,
    batch: ProjectCostImportBatch,
    candidate_ids: list[int],
    action: str,
    note: str | None,
    actor_user_id: int,
) -> list[ProjectCostPriceCandidate]:
    status_map = {"approve": CANDIDATE_STATUS_APPROVED, "reject": CANDIDATE_STATUS_REJECTED, "reopen": CANDIDATE_STATUS_PENDING}
    if action not in status_map:
        raise ProjectCostImportError("INVALID_REVIEW_ACTION")
    if batch.status == IMPORT_STATUS_DRAFT_CREATED:
        raise ProjectCostImportError("IMPORT_BATCH_ALREADY_CREATED_DRAFT")
    rows = db.query(ProjectCostPriceCandidate).filter(
        ProjectCostPriceCandidate.batch_id == batch.id,
        ProjectCostPriceCandidate.id.in_(candidate_ids),
    ).all()
    if len(rows) != len(set(candidate_ids)):
        raise ProjectCostImportError("CANDIDATE_NOT_FOUND")
    now = datetime.now()
    for row in rows:
        row.status = status_map[action]
        row.review_note = _clean(note, 2000)
        row.reviewed_by = actor_user_id if action != "reopen" else None
        row.reviewed_at = now if action != "reopen" else None
    db.flush()
    _refresh_batch_review_counts(db, batch)
    return rows


def update_project_cost_candidate(db: Session, candidate: ProjectCostPriceCandidate, changes: dict[str, Any]) -> ProjectCostPriceCandidate:
    if candidate.batch.status == IMPORT_STATUS_DRAFT_CREATED:
        raise ProjectCostImportError("IMPORT_BATCH_ALREADY_CREATED_DRAFT")
    for field in ("normalized_item_name", "brand", "spec", "unit", "resource_type", "review_note"):
        if field in changes:
            limit = 255 if field in {"normalized_item_name", "brand"} else 500 if field == "spec" else 2000 if field == "review_note" else 64
            setattr(candidate, field, _clean(changes[field], limit))
    if "recommended_price" in changes:
        price = _number(changes["recommended_price"])
        if price is None or price <= 0:
            raise ProjectCostImportError("INVALID_RECOMMENDED_PRICE")
        candidate.recommended_price = price
    if "matched_resource_id" in changes:
        resource_id = changes["matched_resource_id"]
        if resource_id in (None, 0, ""):
            candidate.matched_resource_id = None
            candidate.match_type = "new_resource"
            candidate.match_confidence = 0
        else:
            resource = db.get(EnterpriseCostResource, int(resource_id))
            active = active_enterprise_quota_version(db)
            if not resource or not active or resource.version_id != active.id:
                raise ProjectCostImportError("ACTIVE_RESOURCE_NOT_FOUND")
            candidate.matched_resource_id = resource.id
            candidate.match_type = "manual"
            candidate.match_confidence = 1.0
    db.flush()
    return candidate


def create_draft_quota_version_from_batch(
    db: Session,
    *,
    batch: ProjectCostImportBatch,
    actor_user_id: int,
    version_name: str | None = None,
) -> EnterpriseQuotaVersion:
    if batch.target_quota_version_id:
        existing = db.get(EnterpriseQuotaVersion, batch.target_quota_version_id)
        if existing:
            return existing
    approved = db.query(ProjectCostPriceCandidate).filter(
        ProjectCostPriceCandidate.batch_id == batch.id,
        ProjectCostPriceCandidate.status == CANDIDATE_STATUS_APPROVED,
    ).order_by(ProjectCostPriceCandidate.id).all()
    if not approved:
        raise ProjectCostImportError("NO_APPROVED_CANDIDATES")
    source = active_enterprise_quota_version(db)
    if source is None:
        raise ProjectCostImportError("ACTIVE_ENTERPRISE_QUOTA_VERSION_NOT_FOUND")

    source_tag = f"project-purchase:{batch.batch_uuid}"
    import_batch = CostImportBatch(
        batch_uuid=str(uuid.uuid4()),
        source_filename=f"{batch.project_name}-采购资料",
        source_file_sha256=hashlib.sha256(source_tag.encode("utf-8")).hexdigest(),
        source_file_size=None,
        parser_version=PARSER_VERSION,
        status=IMPORT_BATCH_STATUS_IMPORTED,
        error_count=0,
        warning_count=sum(1 for row in approved if row.risk_level == "high"),
        created_by=actor_user_id,
    )
    db.add(import_batch)
    db.flush()
    version = EnterpriseQuotaVersion(
        version_code=_next_version_code(db, batch),
        version_name=_clean(version_name, 255) or f"{source.version_name} + {batch.project_name}采购价",
        import_batch_id=import_batch.id,
        source_filename=import_batch.source_filename,
        source_file_sha256=import_batch.source_file_sha256,
        status=QUOTA_VERSION_STATUS_DRAFT,
        is_active=False,
        created_by=actor_user_id,
        notes=f"由项目采购资料快速入库批次 {batch.batch_uuid} 生成；需人工复核后按企业定额版本流程启用。",
    )
    db.add(version)
    db.flush()
    section_map, item_map, resource_map = _clone_version(db, source, version)
    del section_map

    affected_item_ids: set[int] = set()
    for candidate in approved:
        target_resource = resource_map.get(candidate.matched_resource_id or -1)
        if target_resource is None:
            target_resource = EnterpriseCostResource(
                version_id=version.id,
                resource_code=f"PUR-{batch.id}-{candidate.id}",
                resource_name=candidate.normalized_item_name,
                resource_type=candidate.resource_type or "main_material",
                unit=normalize_enterprise_quota_unit(candidate.unit),
                price=candidate.recommended_price,
                computed_price=candidate.recommended_price,
                price_block_label="项目采购审核价",
                source_sheet=batch.project_name,
                source_row_index=candidate.id,
                sort_order=100000 + candidate.id,
                raw_row_json=_candidate_evidence(candidate, batch),
            )
            db.add(target_resource)
            db.flush()
        else:
            target_resource.price = candidate.recommended_price
            target_resource.computed_price = candidate.recommended_price
            target_resource.price_block_label = "项目采购审核价"
            target_resource.raw_row_json = _merge_json(target_resource.raw_row_json, {
                "project_purchase_import": json.loads(_candidate_evidence(candidate, batch))
            })
            components = db.query(EnterpriseQuotaComponent).filter(
                EnterpriseQuotaComponent.version_id == version.id,
                EnterpriseQuotaComponent.resource_id == target_resource.id,
            ).all()
            for component in components:
                component.unit_price = candidate.recommended_price
                component.amount = _multiply(component.quantity, component.unit_price)
                if component.quota_item_id:
                    affected_item_ids.add(component.quota_item_id)
        candidate.draft_resource_id = target_resource.id

    for item_id in affected_item_ids:
        item = db.get(EnterpriseQuotaItem, item_id)
        if item:
            _recalculate_quota_item(item, list(item.components))

    row_counts = {
        "section_count": db.query(func.count(EnterpriseQuotaSection.id)).filter(EnterpriseQuotaSection.version_id == version.id).scalar() or 0,
        "item_count": db.query(func.count(EnterpriseQuotaItem.id)).filter(EnterpriseQuotaItem.version_id == version.id).scalar() or 0,
        "component_count": db.query(func.count(EnterpriseQuotaComponent.id)).filter(EnterpriseQuotaComponent.version_id == version.id).scalar() or 0,
        "resource_count": db.query(func.count(EnterpriseCostResource.id)).filter(EnterpriseCostResource.version_id == version.id).scalar() or 0,
    }
    summary = {
        "error_count": 0,
        "warning_count": sum(1 for row in approved if row.risk_level == "high"),
        "project_cost_import_batch_id": batch.id,
        "project_name": batch.project_name,
        "approved_candidate_count": len(approved),
        "updated_existing_resource_count": sum(1 for row in approved if row.matched_resource_id),
        "new_resource_count": sum(1 for row in approved if not row.matched_resource_id),
        "recalculated_quota_item_count": len(affected_item_ids),
        **row_counts,
    }
    version.summary_json = _json(summary)
    import_batch.summary_json = version.summary_json
    batch.target_quota_version_id = version.id
    batch.status = IMPORT_STATUS_DRAFT_CREATED
    batch.summary_json = _merge_json(batch.summary_json, {"draft_version": {"id": version.id, "version_code": version.version_code, **summary}})
    db.flush()
    return version


def serialize_project_cost_import_batch(batch: ProjectCostImportBatch) -> dict[str, Any]:
    return {
        "id": batch.id,
        "batch_uuid": batch.batch_uuid,
        "project_name": batch.project_name,
        "source_name": batch.source_name,
        "status": batch.status,
        "parser_version": batch.parser_version,
        "file_count": batch.file_count,
        "parsed_file_count": batch.parsed_file_count,
        "skipped_file_count": batch.skipped_file_count,
        "observation_count": batch.observation_count,
        "candidate_count": batch.candidate_count,
        "high_confidence_count": batch.high_confidence_count,
        "approved_count": batch.approved_count,
        "rejected_count": batch.rejected_count,
        "target_quota_version_id": batch.target_quota_version_id,
        "summary": _json_load(batch.summary_json),
        "source_manifest": _json_load(batch.source_manifest_json),
        "created_by": batch.created_by,
        "created_at": _format(batch.created_at),
        "updated_at": _format(batch.updated_at),
    }


def serialize_project_cost_candidate(row: ProjectCostPriceCandidate) -> dict[str, Any]:
    return {
        "id": row.id,
        "batch_id": row.batch_id,
        "candidate_key": row.candidate_key,
        "normalized_item_name": row.normalized_item_name,
        "brand": row.brand,
        "spec": row.spec,
        "unit": normalize_enterprise_quota_unit(row.unit),
        "resource_type": row.resource_type,
        "observation_count": row.observation_count,
        "supplier_count": row.supplier_count,
        "min_price": row.min_price,
        "median_price": row.median_price,
        "max_price": row.max_price,
        "recommended_price": row.recommended_price,
        "volatility_rate": row.volatility_rate,
        "confidence_score": row.confidence_score,
        "risk_level": row.risk_level,
        "status": row.status,
        "matched_resource_id": row.matched_resource_id,
        "matched_resource_name": row.matched_resource.resource_name if row.matched_resource else None,
        "match_type": row.match_type,
        "match_confidence": row.match_confidence,
        "review_note": row.review_note,
        "reviewed_by": row.reviewed_by,
        "reviewed_at": _format(row.reviewed_at),
        "draft_resource_id": row.draft_resource_id,
        "evidence": _json_load(row.evidence_json),
    }


def serialize_price_observation(row: EnterpriseResourcePriceObservation) -> dict[str, Any]:
    return {
        "id": row.id,
        "batch_id": row.batch_id,
        "observation_type": row.observation_type,
        "source_file_name": row.source_file_name,
        "source_file_sha256": row.source_file_sha256,
        "source_sheet": row.source_sheet,
        "source_row_index": row.source_row_index,
        "supplier_name": row.supplier_name,
        "observed_at": _format(row.observed_at),
        "raw_item_name": row.raw_item_name,
        "normalized_item_name": row.normalized_item_name,
        "brand": row.brand,
        "spec": row.spec,
        "unit": normalize_enterprise_quota_unit(row.unit),
        "quantity": row.quantity,
        "unit_price": row.unit_price,
        "amount": row.amount,
        "tax_included": row.tax_included,
        "tax_rate": row.tax_rate,
        "freight_included": row.freight_included,
        "is_return": row.is_return,
        "excluded_reason": row.excluded_reason,
        "candidate_key": row.candidate_key,
    }


def _expand_files(
    files: Iterable[tuple[str, bytes]], *, max_total_bytes: int | None
):
    total = 0
    emitted = 0
    for raw_name, content in files:
        filename = _clean(raw_name, 500) or "purchase.xlsx"
        total += len(content)
        if max_total_bytes is not None and total > max_total_bytes:
            raise ProjectCostImportError("UPLOAD_TOO_LARGE")
        suffix = Path(filename).suffix.lower()
        if suffix == ".zip":
            with zipfile.ZipFile(BytesIO(content)) as archive:
                members = [item for item in archive.infolist() if not item.is_dir() and Path(item.filename).suffix.lower() in SUPPORTED_EXTENSIONS]
                if len(members) > MAX_ARCHIVE_FILES or sum(item.file_size for item in members) > MAX_ARCHIVE_UNCOMPRESSED_BYTES:
                    raise ProjectCostImportError("ARCHIVE_TOO_LARGE")
                for item in members:
                    safe_name = str(Path(item.filename.replace("\\", "/")))
                    if safe_name.startswith(".."):
                        continue
                    emitted += 1
                    yield safe_name, archive.read(item)
        elif suffix in SUPPORTED_EXTENSIONS:
            emitted += 1
            yield filename, content
    if not emitted:
        raise ProjectCostImportError("NO_SUPPORTED_EXCEL_FILES")


def _parse_workbook(filename: str, content: bytes, sha256: str) -> dict[str, Any]:
    workbook = load_workbook(BytesIO(content), data_only=True, read_only=False)
    observations: list[dict[str, Any]] = []
    parsed_sheets = 0
    issues: list[str] = []
    observation_type = _observation_type(filename)
    for sheet in workbook.worksheets:
        values = list(sheet.iter_rows(values_only=True))
        detected = _detect_header(values)
        if not detected:
            continue
        parsed_sheets += 1
        header_index, columns = detected
        supplier = _find_label_value(values[:header_index], ("供应商", "供货单位", "报价单位"))
        order_no = _find_label_value(values[:header_index], ("订单号", "采购单号", "编号"))
        observed_at = _find_observed_date(values[:header_index], filename)
        sheet_text = " ".join(_clean(cell) for row in values[:100] for cell in row[:20] if cell not in (None, ""))
        tax_included, tax_rate = _tax_context(sheet_text)
        freight_included = _freight_context(sheet_text)
        blank_run = 0
        for row_index, row in enumerate(values[header_index + 1 :], start=header_index + 2):
            raw_name = _cell(row, columns.get("name"))
            if not _clean(raw_name):
                blank_run += 1
                if blank_run >= 8:
                    break
                continue
            blank_run = 0
            name = _clean(raw_name, 255)
            if _is_summary_row(name):
                continue
            price = _number(_cell(row, columns.get("unit_price")))
            quantity = _number(_cell(row, columns.get("quantity")))
            amount = _number(_cell(row, columns.get("amount")))
            brand = _clean(_cell(row, columns.get("brand")), 255)
            spec = _clean(_cell(row, columns.get("spec")), 500)
            brand_spec = _clean(_cell(row, columns.get("brand_spec")), 500)
            if brand_spec and not brand and not spec:
                spec = brand_spec
            unit = normalize_enterprise_quota_unit(_cell(row, columns.get("unit")))
            normalized_name = normalize_material_name(name)
            excluded_reason = _excluded_reason(name, price, quantity)
            candidate_key = _candidate_key(normalized_name, brand, spec, unit) if not excluded_reason else None
            validation = None
            if quantity is not None and price is not None and amount is not None:
                expected = quantity * price
                if abs(expected - amount) > max(0.02, abs(amount) * 0.002):
                    validation = {"code": "AMOUNT_MISMATCH", "expected": round(expected, 4), "actual": amount}
            observations.append({
                "observation_type": observation_type,
                "source_file_name": filename,
                "source_file_sha256": sha256,
                "source_sheet": _clean(sheet.title, 128),
                "source_row_index": row_index,
                "supplier_name": _clean(supplier, 255),
                "order_no": _clean(order_no, 128),
                "observed_at": observed_at,
                "raw_item_name": name,
                "normalized_item_name": normalized_name,
                "brand": brand,
                "spec": spec,
                "unit": _clean(unit, 64),
                "quantity": quantity,
                "unit_price": price,
                "amount": amount,
                "tax_included": tax_included,
                "tax_rate": tax_rate,
                "freight_included": freight_included,
                "is_return": bool((quantity is not None and quantity < 0) or (amount is not None and amount < 0)),
                "excluded_reason": excluded_reason,
                "candidate_key": candidate_key,
                "raw_json": _json({"row": list(row), "validation": validation}),
            })
    workbook.close()
    return {
        "observations": observations,
        "manifest": {
            "status": "parsed" if observations else "skipped",
            "sheet_count": len(workbook.sheetnames),
            "parsed_sheet_count": parsed_sheets,
            "observation_count": len(observations),
            "issues": issues,
        },
    }


def _detect_header(rows: list[tuple[Any, ...]]) -> tuple[int, dict[str, int]] | None:
    best: tuple[int, dict[str, int], int] | None = None
    for row_index, row in enumerate(rows[:60]):
        columns: dict[str, int] = {}
        for column_index, value in enumerate(row):
            text = _header_text(value)
            if not text:
                continue
            for key, aliases in HEADER_ALIASES.items():
                if key in columns:
                    continue
                if any(_header_text(alias) == text or _header_text(alias) in text for alias in aliases):
                    columns[key] = column_index
                    break
        score = len(columns) + (2 if "name" in columns else 0) + (2 if "unit_price" in columns else 0)
        if "name" in columns and "unit_price" in columns and (best is None or score > best[2]):
            best = (row_index, columns, score)
    return (best[0], best[1]) if best else None


def _build_candidates(db: Session, batch: ProjectCostImportBatch) -> list[dict[str, Any]]:
    grouped: dict[str, list[EnterpriseResourcePriceObservation]] = defaultdict(list)
    for row in batch.observations:
        if row.candidate_key and row.unit_price and row.unit_price > 0 and not row.is_return and not row.excluded_reason:
            grouped[row.candidate_key].append(row)
    active = active_enterprise_quota_version(db)
    resources = list(active.resources) if active else []
    result = []
    for key, rows in grouped.items():
        prices = sorted(float(row.unit_price) for row in rows if row.unit_price and row.unit_price > 0)
        median = float(statistics.median(prices))
        volatility = (max(prices) - min(prices)) / median if median else None
        sample = rows[0]
        score = 0.45
        score += 0.12 if sample.spec else 0
        score += 0.08 if sample.brand else 0
        score += 0.1 if len(rows) >= 2 else 0
        score += 0.08 if len({row.supplier_name for row in rows if row.supplier_name}) >= 2 else 0
        score += 0.12 if volatility is not None and volatility <= 0.1 else 0
        score -= 0.2 if volatility is not None and volatility >= 0.2 else 0
        score = max(0, min(round(score, 4), 1))
        risk = "high" if volatility is not None and volatility >= 0.2 else "low" if score >= 0.75 else "medium"
        match = _match_resource(sample, resources)
        result.append({
            "candidate_key": key,
            "normalized_item_name": sample.normalized_item_name,
            "brand": sample.brand,
            "spec": sample.spec,
            "unit": sample.unit,
            "resource_type": "main_material",
            "observation_count": len(rows),
            "supplier_count": len({row.supplier_name for row in rows if row.supplier_name}),
            "min_price": min(prices),
            "median_price": median,
            "max_price": max(prices),
            "recommended_price": median,
            "volatility_rate": volatility,
            "confidence_score": score,
            "risk_level": risk,
            "status": CANDIDATE_STATUS_PENDING,
            "matched_resource_id": match[0].id if match[0] else None,
            "match_type": match[1],
            "match_confidence": match[2],
            "evidence_json": _json([{
                "observation_id": row.id,
                "source_file_name": row.source_file_name,
                "source_sheet": row.source_sheet,
                "source_row_index": row.source_row_index,
                "supplier_name": row.supplier_name,
                "unit_price": row.unit_price,
                "observed_at": _format(row.observed_at),
            } for row in rows[:20]]),
        })
    return result


def _match_resource(observation: EnterpriseResourcePriceObservation, resources: list[EnterpriseCostResource]) -> tuple[EnterpriseCostResource | None, str, float]:
    name = _match_text(observation.normalized_item_name)
    unit = normalize_enterprise_quota_unit(observation.unit)
    best: tuple[EnterpriseCostResource | None, str, float] = (None, "new_resource", 0)
    for resource in resources:
        resource_name = _match_text(resource.resource_name)
        if not resource_name:
            continue
        if unit and normalize_enterprise_quota_unit(resource.unit) not in (None, "", unit):
            continue
        if name == resource_name:
            score, match_type = 0.98, "exact_name_unit"
        elif name in resource_name or resource_name in name:
            score, match_type = 0.82, "contains_name_unit"
        else:
            score = SequenceMatcher(None, name, resource_name).ratio()
            match_type = "similar_name_unit"
        if score > best[2] and score >= 0.72:
            best = (resource, match_type, round(score, 4))
    return best


def _clone_version(db: Session, source: EnterpriseQuotaVersion, target: EnterpriseQuotaVersion):
    section_map: dict[int, EnterpriseQuotaSection] = {}
    for row in source.sections:
        clone = EnterpriseQuotaSection(version_id=target.id, section_code=row.section_code, section_name=row.section_name, source_sheet=row.source_sheet, source_row_index=row.source_row_index, sort_order=row.sort_order, raw_row_json=row.raw_row_json)
        db.add(clone)
        db.flush()
        section_map[row.id] = clone
    resource_map: dict[int, EnterpriseCostResource] = {}
    for row in source.resources:
        clone = EnterpriseCostResource(version_id=target.id, resource_code=row.resource_code, resource_name=row.resource_name, resource_type=row.resource_type, unit=row.unit, price=row.price, tax_rate=row.tax_rate, computed_price=row.computed_price, price_block_label=row.price_block_label, source_sheet=row.source_sheet, source_row_index=row.source_row_index, sort_order=row.sort_order, raw_row_json=row.raw_row_json)
        db.add(clone)
        db.flush()
        resource_map[row.id] = clone
    item_map: dict[int, EnterpriseQuotaItem] = {}
    for row in source.items:
        clone = EnterpriseQuotaItem(version_id=target.id, section_id=section_map.get(row.section_id).id if row.section_id in section_map else None, quota_code=row.quota_code, item_name=row.item_name, work_content=row.work_content, worker_or_subtype=row.worker_or_subtype, unit=row.unit, quantity=row.quantity, unit_price=row.unit_price, labor_fee=row.labor_fee, main_material_fee=row.main_material_fee, auxiliary_material_fee=row.auxiliary_material_fee, machinery_fee=row.machinery_fee, source_sheet=row.source_sheet, source_row_index=row.source_row_index, sort_order=row.sort_order, raw_row_json=row.raw_row_json)
        db.add(clone)
        db.flush()
        item_map[row.id] = clone
    for row in source.components:
        clone = EnterpriseQuotaComponent(version_id=target.id, quota_item_id=item_map.get(row.quota_item_id).id if row.quota_item_id in item_map else None, resource_id=resource_map.get(row.resource_id).id if row.resource_id in resource_map else None, parent_quota_code=row.parent_quota_code, component_type=row.component_type, resource_code=row.resource_code, resource_name=row.resource_name, worker_or_subtype=row.worker_or_subtype, unit=row.unit, quantity=row.quantity, unit_price=row.unit_price, amount=row.amount, fee_bucket=row.fee_bucket, source_sheet=row.source_sheet, source_row_index=row.source_row_index, sort_order=row.sort_order, raw_row_json=row.raw_row_json)
        db.add(clone)
    db.flush()
    return section_map, item_map, resource_map


def _recalculate_quota_item(item: EnterpriseQuotaItem, components: list[EnterpriseQuotaComponent]) -> None:
    buckets = {"labor": 0.0, "main_material": 0.0, "auxiliary_material": 0.0, "machinery": 0.0}
    for row in components:
        bucket = _fee_bucket(row.fee_bucket, row.component_type)
        if bucket:
            buckets[bucket] += float(row.amount or 0)
    item.labor_fee = round(buckets["labor"], 6)
    item.main_material_fee = round(buckets["main_material"], 6)
    item.auxiliary_material_fee = round(buckets["auxiliary_material"], 6)
    item.machinery_fee = round(buckets["machinery"], 6)
    item.unit_price = round(sum(buckets.values()), 6)


def _fee_bucket(value: Any, component_type: Any) -> str | None:
    text = f"{_clean(value)} {_clean(component_type)}".lower()
    if "人工" in text or "labor" in text or "rg" in text:
        return "labor"
    if "主材" in text or "main" in text or "ca" in text:
        return "main_material"
    if "辅材" in text or "aux" in text or "cb" in text:
        return "auxiliary_material"
    if "机械" in text or "machine" in text or "machinery" in text or "jx" in text:
        return "machinery"
    return None


def normalize_material_name(value: Any) -> str:
    text = _clean(value, 255).lower()
    text = re.sub(r"[\s\u3000,，。;；:：/\\|·•\-—_]+", "", text)
    return text


def _candidate_key(name: str, brand: str | None, spec: str | None, unit: str | None) -> str:
    raw = "|".join((_match_text(name), _match_text(brand), _match_text(spec), _match_text(unit)))
    return hashlib.sha256(raw.encode("utf-8")).hexdigest()


def _observation_type(filename: str) -> str:
    text = filename.lower()
    if "比价" in text or "对比" in text:
        return "comparison"
    if "询价" in text:
        return "inquiry"
    if "报价" in text:
        return "quotation"
    if "计划" in text or "请购" in text:
        return "requisition"
    return "order"


def _excluded_reason(name: str, price: float | None, quantity: float | None) -> str | None:
    if price is None or price <= 0:
        return "missing_or_zero_price"
    normalized = normalize_material_name(name)
    if any(word in normalized for word in SERVICE_FEE_WORDS) and len(normalized) <= 16:
        return "service_fee"
    if quantity is not None and quantity < 0:
        return "return"
    return None


def _is_summary_row(name: str) -> bool:
    text = normalize_material_name(name)
    return text in {"合计", "总计", "小计", "价税合计", "金额合计"} or text.startswith("备注")


def _find_label_value(rows: list[tuple[Any, ...]], labels: tuple[str, ...]) -> str | None:
    for row in rows:
        for index, value in enumerate(row):
            text = _clean(value)
            for label in labels:
                match = re.search(rf"{re.escape(label)}\s*[：:]\s*(.+)", text)
                if match:
                    return match.group(1).strip()
                if text == label and index + 1 < len(row):
                    return _clean(row[index + 1])
    return None


def _find_observed_date(rows: list[tuple[Any, ...]], filename: str) -> datetime | None:
    for row in rows:
        for index, value in enumerate(row):
            text = _clean(value)
            if any(label in text for label in ("日期", "时间")):
                parsed = _parse_date(value)
                if parsed:
                    return parsed
                if index + 1 < len(row):
                    parsed = _parse_date(row[index + 1])
                    if parsed:
                        return parsed
    return _parse_date(filename)


def _parse_date(value: Any) -> datetime | None:
    if isinstance(value, datetime):
        return value.replace(tzinfo=None)
    if isinstance(value, date):
        return datetime(value.year, value.month, value.day)
    text = _clean(value)
    match = re.search(r"(20\d{2})[^0-9]?(0?[1-9]|1[0-2])[^0-9]?(0?[1-9]|[12]\d|3[01])", text)
    if not match:
        return None
    try:
        return datetime(int(match.group(1)), int(match.group(2)), int(match.group(3)))
    except ValueError:
        return None


def _tax_context(text: str) -> tuple[bool | None, float | None]:
    rate_match = re.search(r"(1|3|6|9|13)\s*%", text)
    rate = float(rate_match.group(1)) / 100 if rate_match else None
    if "不含税" in text:
        return False, rate
    if "含税" in text or "专票" in text or "发票" in text:
        return True, rate
    return None, rate


def _freight_context(text: str) -> bool | None:
    if "不含运" in text or "运费另计" in text:
        return False
    if "含运" in text or "包运" in text or "运费已含" in text:
        return True
    return None


def _refresh_batch_review_counts(db: Session, batch: ProjectCostImportBatch) -> None:
    counts = dict(db.query(ProjectCostPriceCandidate.status, func.count(ProjectCostPriceCandidate.id)).filter(ProjectCostPriceCandidate.batch_id == batch.id).group_by(ProjectCostPriceCandidate.status).all())
    batch.approved_count = int(counts.get(CANDIDATE_STATUS_APPROVED, 0))
    batch.rejected_count = int(counts.get(CANDIDATE_STATUS_REJECTED, 0))


def _next_version_code(db: Session, batch: ProjectCostImportBatch) -> str:
    base = f"purchase-{datetime.now().strftime('%Y%m%d')}-{batch.batch_uuid[:8]}"
    code = base
    suffix = 1
    while db.query(EnterpriseQuotaVersion.id).filter(EnterpriseQuotaVersion.version_code == code).first():
        suffix += 1
        code = f"{base}-{suffix}"
    return code[:64]


def _candidate_evidence(candidate: ProjectCostPriceCandidate, batch: ProjectCostImportBatch) -> str:
    return _json({
        "project_cost_import_batch_id": batch.id,
        "project_name": batch.project_name,
        "candidate_id": candidate.id,
        "candidate_key": candidate.candidate_key,
        "recommended_price": candidate.recommended_price,
        "review_note": candidate.review_note,
        "evidence": _json_load(candidate.evidence_json),
    })


def _merge_json(raw: str | None, update: dict[str, Any]) -> str:
    value = _json_load(raw)
    if not isinstance(value, dict):
        value = {"previous_raw": value}
    value.update(update)
    return _json(value)


def _multiply(left: float | None, right: float | None) -> float | None:
    if left is None or right is None:
        return None
    return round(float(left) * float(right), 6)


def _header_text(value: Any) -> str:
    return re.sub(r"[\s\n\r\t/\\（）()\-—_]+", "", _clean(value)).lower()


def _match_text(value: Any) -> str:
    return re.sub(r"[^0-9a-z\u4e00-\u9fff]+", "", _clean(value).lower())


def _cell(row: tuple[Any, ...], index: int | None) -> Any:
    return row[index] if index is not None and index < len(row) else None


def _number(value: Any) -> float | None:
    if value in (None, ""):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    match = re.search(r"-?\d+(?:\.\d+)?", _clean(value).replace(",", ""))
    return float(match.group(0)) if match else None


def _clean(value: Any, limit: int | None = None) -> str:
    if value is None:
        return ""
    text = re.sub(r"[\x00-\x08\x0b\x0c\x0e-\x1f]", "", str(value)).replace("\r", " ").replace("\n", " ").strip()
    return text[:limit] if limit else text


def _json(value: Any) -> str:
    return json.dumps(value, ensure_ascii=False, sort_keys=True, default=str)


def _json_load(value: str | None) -> Any:
    if not value:
        return None
    try:
        return json.loads(value)
    except (TypeError, ValueError):
        return value


def _format(value: Any) -> str | None:
    return value.isoformat() if value else None
