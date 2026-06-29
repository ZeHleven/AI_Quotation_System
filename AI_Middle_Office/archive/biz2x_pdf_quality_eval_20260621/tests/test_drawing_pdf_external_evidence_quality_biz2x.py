from __future__ import annotations

import json
import sys
from pathlib import Path

from openpyxl import load_workbook

from app.services.drawing_pdf_external_evidence_quality import (
    build_external_evidence_quality_report,
    filter_external_results_by_quality,
    write_external_evidence_quality_outputs,
)
from scripts import biz2x_pdf_external_evidence_quality


def _external_results() -> dict[str, object]:
    return {
        "evidence_rows": [
            {
                "source_file": "drawing.pdf",
                "page": 1,
                "tile_id": "p001_whole",
                "vision_pass": "door_window_demolition",
                "task_no": 7,
                "item_hint": "拆除不锈钢玻璃门",
                "spec_or_method": "含门套、门扇及五金拆除并清运",
                "suggested_unit": "套",
                "text": "图纸可见拆除不锈钢玻璃门说明",
                "confidence": 0.91,
            },
            {
                "source_file": "drawing.pdf",
                "page": 2,
                "tile_id": "p002_whole",
                "vision_pass": "demolition_notes",
                "task_no": 8,
                "item_hint": "Removed Finishes",
                "suggested_unit": "unknown",
                "text": "All existing finishes removed.",
                "confidence": 0.95,
            },
            {
                "source_file": "drawing.pdf",
                "page": 3,
                "tile_id": "p003_whole",
                "vision_pass": "water_supply",
                "task_no": 9,
                "item_hint": "给水管 SUS304 薄壁不锈钢管 DN40",
                "spec_or_method": "SUS304 薄壁不锈钢管 DN40，按图示安装",
                "text": "材料表可见 SUS304 薄壁不锈钢管 DN40",
                "confidence": 0.82,
            },
        ]
    }


def test_external_evidence_quality_scores_and_filters_conservative_rows():
    report = build_external_evidence_quality_report(_external_results())

    assert report["summary"]["input_row_count"] == 3
    assert report["summary"]["accepted_row_count"] == 1
    assert report["summary"]["review_row_count"] == 1
    assert report["summary"]["rejected_row_count"] == 1
    assert report["summary"]["filtered_importable_row_count"] == 1
    rows = report["quality_rows"]
    assert rows[0]["quality_status"] == "accepted"
    assert rows[1]["quality_status"] == "rejected"
    assert "unknown_or_invalid_unit" in rows[1]["issue_codes"]
    assert rows[2]["quality_status"] == "review"
    assert "missing_unit" in rows[2]["issue_codes"]


def test_external_evidence_quality_can_include_review_rows():
    filtered = filter_external_results_by_quality(_external_results(), include_review=True)

    assert len(filtered["evidence_rows"]) == 2
    assert filtered["evidence_rows"][0]["item_hint"] == "拆除不锈钢玻璃门"
    assert filtered["evidence_rows"][1]["item_hint"].startswith("给水管")


def test_external_evidence_quality_accepts_short_domain_item_terms():
    report = build_external_evidence_quality_report(
        {
            "evidence_rows": [
                {
                    "source_file": "drawing.pdf",
                    "page": 1,
                    "tile_id": "p001_g02",
                    "vision_pass": "finish_schedule",
                    "item_hint": "灯槽",
                    "spec_or_method": "灯槽",
                    "suggested_unit": "㎡",
                    "text": "灯槽",
                    "confidence": 0.7,
                }
            ]
        }
    )

    assert report["quality_rows"][0]["quality_status"] == "accepted"
    assert report["summary"]["filtered_importable_row_count"] == 1


def test_external_evidence_quality_accepts_cubic_meter_units():
    report = build_external_evidence_quality_report(
        {
            "evidence_rows": [
                {
                    "source_file": "drawing.pdf",
                    "page": 1,
                    "tile_id": "wall_detail",
                    "vision_pass": "wall_node",
                    "item_hint": "砖砌隔墙",
                    "spec_or_method": "100mm 新建蒸压加气砼砌块墙",
                    "suggested_unit": "m³",
                    "text": "节点图可见新建砌块墙和隔墙剖面做法",
                    "confidence": 0.8,
                }
            ]
        }
    )

    row = report["quality_rows"][0]
    assert row["quality_status"] == "accepted"
    assert "unknown_or_invalid_unit" not in row["issue_codes"]
    assert report["summary"]["filtered_importable_row_count"] == 1


def test_external_evidence_quality_drops_manual_review_rows_even_when_concrete():
    report = build_external_evidence_quality_report(
        {
            "evidence_rows": [
                {
                    "source_file": "drawing.pdf",
                    "page": 1,
                    "tile_id": "finish_detail",
                    "vision_pass": "finish_schedule",
                    "item_hint": "防潮无机涂料",
                    "spec_or_method": "防潮无机涂料饰面；未直接区分黑色/白色",
                    "suggested_unit": "㎡",
                    "text": "图纸可见防潮无机涂料饰面，但不能区分黑色或白色。",
                    "confidence": 0.88,
                    "needs_manual_review": True,
                    "reason": "局部文字未直接可见，需要人工复核。",
                }
            ]
        },
        include_review=True,
    )

    row = report["quality_rows"][0]
    assert row["quality_status"] == "review"
    assert row["import_action"] == "drop"
    assert "manual_review_requested" in row["issue_codes"]
    assert "uncertain_or_incomplete_evidence" in row["issue_codes"]
    assert report["summary"]["filtered_importable_row_count"] == 0


def test_external_evidence_quality_can_import_manual_review_rows_when_evidence_is_specific():
    report = build_external_evidence_quality_report(
        {
            "evidence_rows": [
                {
                    "source_file": "drawing.pdf",
                    "page": 1,
                    "tile_id": "finish_label",
                    "vision_pass": "finish_schedule",
                    "item_hint": "墙面瓷砖湿贴CT-04",
                    "spec_or_method": "CT04 600*1200白色墙面砖",
                    "suggested_unit": "㎡",
                    "text": "立面图可见 CT04 墙面砖作美缝标注。",
                    "confidence": 0.9,
                    "needs_manual_review": True,
                }
            ]
        },
        include_review=True,
    )

    row = report["quality_rows"][0]
    assert row["quality_status"] == "review"
    assert row["import_action"] == "import"
    assert "manual_review_requested" in row["issue_codes"]
    assert "uncertain_or_incomplete_evidence" not in row["issue_codes"]
    assert report["summary"]["filtered_importable_row_count"] == 1


def test_external_evidence_quality_drops_uncertain_text_without_manual_flag():
    report = build_external_evidence_quality_report(
        {
            "evidence_rows": [
                {
                    "source_file": "drawing.pdf",
                    "page": 1,
                    "tile_id": "detail_17",
                    "vision_pass": "finish_node",
                    "item_hint": "人造石窗台石PM-01",
                    "spec_or_method": "PM01 白色人造石；not directly visible as window sill",
                    "suggested_unit": "㎡",
                    "text": "PM01 is visible, but the window sill wording is not directly visible.",
                    "confidence": 0.82,
                }
            ]
        },
        include_review=True,
    )

    row = report["quality_rows"][0]
    assert row["quality_status"] == "review"
    assert row["import_action"] == "drop"
    assert "uncertain_or_incomplete_evidence" in row["issue_codes"]
    assert report["summary"]["filtered_importable_row_count"] == 0


def test_external_evidence_quality_cleans_prompt_placeholder_spec_and_unit():
    external_results = {
        "evidence_rows": [
            {
                "source_file": "drawing.pdf",
                "page": 1,
                "tile_id": "fixture_symbol",
                "vision_pass": "fixture_valve_schedule",
                "task_no": 125,
                "item_hint": "淋浴喷头",
                "spec_or_method": "可见规格、材质、安装方式；没有则留空",
                "suggested_unit": "个/套；不要写数量",
                "text": "图中可见淋浴喷头的平面符号",
                "confidence": 0.8,
            }
        ]
    }

    report = build_external_evidence_quality_report(external_results, include_review=True)

    row = report["quality_rows"][0]
    assert row["quality_status"] == "review"
    assert "missing_spec_or_method" in row["issue_codes"]
    assert "missing_unit" in row["issue_codes"]
    assert row["spec_or_method"] == ""
    assert row["suggested_unit"] == ""
    imported = report["filtered_external_results"]["evidence_rows"][0]
    assert imported["spec_or_method"] == ""
    assert imported["suggested_unit"] == ""


def test_external_evidence_quality_rejects_generic_finish_section_hints():
    report = build_external_evidence_quality_report(
        {
            "evidence_rows": [
                {
                    "source_file": "drawing.pdf",
                    "page": 1,
                    "tile_id": "finish_plan_top",
                    "vision_pass": "finish_schedule",
                    "task_no": 48,
                    "item_hint": "墙面",
                    "spec_or_method": "ST-001, WD-001, WD-002",
                    "suggested_unit": "㎡",
                    "text": "图中可见墙面的材料编号和分布",
                    "confidence": 0.9,
                }
            ]
        },
        include_review=True,
    )

    row = report["quality_rows"][0]
    assert row["quality_status"] == "rejected"
    assert row["import_action"] == "drop"
    assert "generic_section_item_hint" in row["issue_codes"]
    assert report["summary"]["filtered_importable_row_count"] == 0


def test_external_evidence_quality_cleans_prompt_placeholder_item_and_text():
    report = build_external_evidence_quality_report(
        {
            "evidence_rows": [
                {
                    "source_file": "drawing.pdf",
                    "page": 1,
                    "tile_id": "finish_plan_all",
                    "vision_pass": "finish_schedule",
                    "item_hint": "图中可见项目名称",
                    "spec_or_method": "可见材料编号、尺寸、材质、做法或安装方式；没有则留空",
                    "suggested_unit": "㎡/m/个/套/樘；不确定则为空",
                    "text": "摘录图中可见文字或描述可见符号/引线位置",
                    "confidence": 0.0,
                }
            ]
        },
        include_review=True,
    )

    row = report["quality_rows"][0]
    assert row["quality_status"] == "rejected"
    assert row["item_hint"] == ""
    assert row["spec_or_method"] == ""
    assert row["suggested_unit"] == ""
    assert row["text"] == ""
    assert "empty_evidence" in row["issue_codes"]


def test_external_evidence_quality_cleans_schedule_prompt_placeholders():
    report = build_external_evidence_quality_report(
        {
            "evidence_rows": [
                {
                    "source_file": "drawing.pdf",
                    "page": 1,
                    "tile_id": "schedule",
                    "vision_pass": "schedule_ocr",
                    "item_hint": "可见计划或图例证据",
                    "spec_or_method": "规格/方法",
                    "suggested_unit": "",
                    "text": "文字内容",
                    "confidence": 0.0,
                },
                {
                    "source_file": "drawing.pdf",
                    "page": 1,
                    "tile_id": "schedule",
                    "vision_pass": "schedule_ocr",
                    "item_hint": "/",
                    "spec_or_method": "/",
                    "suggested_unit": "",
                    "text": "/",
                    "confidence": 0.0,
                },
            ]
        },
        include_review=True,
    )

    assert report["summary"]["rejected_row_count"] == 2
    assert report["summary"]["filtered_importable_row_count"] == 0
    assert all(row["item_hint"] == "" for row in report["quality_rows"])
    assert all(row["quality_status"] == "rejected" for row in report["quality_rows"])


def test_external_evidence_quality_rejects_schedule_titles():
    report = build_external_evidence_quality_report(
        {
            "evidence_rows": [
                {
                    "source_file": "drawing.pdf",
                    "page": 1,
                    "tile_id": "schedule",
                    "vision_pass": "schedule_ocr",
                    "item_hint": "图纸目录",
                    "spec_or_method": "",
                    "suggested_unit": "",
                    "text": "图纸目录",
                    "confidence": 0.8,
                },
                {
                    "source_file": "drawing.pdf",
                    "page": 1,
                    "tile_id": "schedule",
                    "vision_pass": "schedule_ocr",
                    "item_hint": "日期",
                    "spec_or_method": "2023年10月15日",
                    "suggested_unit": "",
                    "text": "2023年10月15日",
                    "confidence": 0.8,
                },
            ]
        },
        include_review=True,
    )

    assert report["summary"]["filtered_importable_row_count"] == 0
    assert all(row["quality_status"] == "rejected" for row in report["quality_rows"])
    assert all("generic_section_item_hint" in row["issue_codes"] for row in report["quality_rows"])


def test_external_evidence_quality_rejects_text_only_evidence():
    report = build_external_evidence_quality_report(
        {
            "evidence_rows": [
                {
                    "source_file": "drawing.pdf",
                    "page": 1,
                    "tile_id": "finish_plan_all",
                    "vision_pass": "finish_schedule",
                    "text": "图中可见局部材料编号和引线位置，但未识别出项目名称",
                    "confidence": 0.85,
                }
            ]
        },
        include_review=True,
    )

    row = report["quality_rows"][0]
    assert row["quality_status"] == "rejected"
    assert row["import_action"] == "drop"
    assert "missing_item_hint" in row["issue_codes"]
    assert "missing_spec_or_method" in row["issue_codes"]
    assert report["summary"]["filtered_importable_row_count"] == 0


def test_external_evidence_quality_cleans_prompt_unit_option_strings():
    report = build_external_evidence_quality_report(
        {
            "evidence_rows": [
                {
                    "source_file": "drawing.pdf",
                    "page": 1,
                    "tile_id": "finish_plan",
                    "vision_pass": "finish_schedule",
                    "item_hint": "混凝土垫层",
                    "spec_or_method": "C15",
                    "suggested_unit": "?/m/?/?/?/?/?",
                    "text": "C15 混凝土垫层",
                    "confidence": 0.8,
                }
            ]
        },
        include_review=True,
    )

    row = report["quality_rows"][0]
    assert row["quality_status"] == "review"
    assert row["suggested_unit"] == ""
    assert "missing_unit" in row["issue_codes"]
    assert "unknown_or_invalid_unit" not in row["issue_codes"]
    assert row["import_action"] == "import"
    assert report["summary"]["filtered_importable_row_count"] == 1


def test_external_evidence_quality_drops_review_rows_with_invalid_unit():
    report = build_external_evidence_quality_report(
        {
            "evidence_rows": [
                {
                    "source_file": "drawing.pdf",
                    "page": 1,
                    "tile_id": "finish_plan",
                    "vision_pass": "finish_schedule",
                    "item_hint": "混凝土垫层",
                    "spec_or_method": "C15",
                    "suggested_unit": "bogus-unit",
                    "text": "C15 混凝土垫层",
                    "confidence": 0.8,
                }
            ]
        },
        include_review=True,
    )

    row = report["quality_rows"][0]
    assert row["quality_status"] == "review"
    assert row["import_action"] == "drop"
    assert "unknown_or_invalid_unit" in row["issue_codes"]
    assert report["summary"]["filtered_importable_row_count"] == 0


def test_external_evidence_quality_rejects_generic_finish_elevation_noise():
    report = build_external_evidence_quality_report(
        {
            "evidence_rows": [
                {
                    "source_file": "drawing.pdf",
                    "page": 1,
                    "tile_id": "finish_bw",
                    "vision_pass": "finish_schedule",
                    "item_hint": "材料清单或立面图证据",
                    "spec_or_method": "未指定",
                    "suggested_unit": "?/m/?/?/m",
                    "text": "材料清单/立面图",
                    "confidence": 0.0,
                },
                {
                    "source_file": "drawing.pdf",
                    "page": 1,
                    "tile_id": "finish_bw",
                    "vision_pass": "finish_schedule",
                    "item_hint": "墙面装饰材料表",
                    "spec_or_method": "墙面装饰材料表",
                    "suggested_unit": "",
                    "text": "墙面装饰材料表",
                    "confidence": 0.0,
                },
            ]
        },
        include_review=True,
    )

    assert report["summary"]["filtered_importable_row_count"] == 0
    assert all(row["quality_status"] == "rejected" for row in report["quality_rows"])
    assert all("generic_section_item_hint" in row["issue_codes"] for row in report["quality_rows"])


def test_external_evidence_quality_rejects_generic_finish_material_item_hints():
    report = build_external_evidence_quality_report(
        {
            "evidence_rows": [
                {
                    "source_file": "drawing.pdf",
                    "page": 1,
                    "tile_id": "finish_schedule_crop",
                    "vision_pass": "finish_schedule",
                    "item_hint": "墙面装饰材料",
                    "spec_or_method": "PVC板",
                    "suggested_unit": "㎡",
                    "text": "图中可以看到墙面装饰材料为 PVC 板。",
                    "confidence": 1.0,
                },
                {
                    "source_file": "drawing.pdf",
                    "page": 1,
                    "tile_id": "finish_schedule_crop",
                    "vision_pass": "finish_schedule",
                    "item_hint": "地面装饰材料",
                    "spec_or_method": "600*600",
                    "suggested_unit": "㎡",
                    "text": "图中可以看到地面装饰材料规格为 600*600。",
                    "confidence": 1.0,
                },
                {
                    "source_file": "drawing.pdf",
                    "page": 1,
                    "tile_id": "finish_schedule_crop",
                    "vision_pass": "finish_schedule",
                    "item_hint": "吊顶材料",
                    "spec_or_method": "铝扣板",
                    "suggested_unit": "㎡",
                    "text": "图中可以看到吊顶材料为铝扣板。",
                    "confidence": 1.0,
                },
            ]
        },
        include_review=True,
    )

    assert report["summary"]["filtered_importable_row_count"] == 0
    assert all(row["quality_status"] == "rejected" for row in report["quality_rows"])
    assert all(row["import_action"] == "drop" for row in report["quality_rows"])
    assert all("generic_section_item_hint" in row["issue_codes"] for row in report["quality_rows"])


def test_external_evidence_quality_rejects_quantity_estimation_text():
    report = build_external_evidence_quality_report(
        {
            "evidence_rows": [
                {
                    "source_file": "drawing.pdf",
                    "page": 1,
                    "tile_id": "fallback_image",
                    "vision_pass": "finish_schedule",
                    "item_hint": "天花板吊顶",
                    "spec_or_method": "石膏板吊顶",
                    "suggested_unit": "㎡",
                    "text": "通过测量和计算可以确定天花板的实际面积。",
                    "confidence": 1.0,
                    "reason": "通过测量来确定材料的数量。",
                }
            ]
        },
        include_review=True,
    )

    row = report["quality_rows"][0]
    assert row["quality_status"] == "rejected"
    assert row["import_action"] == "drop"
    assert "quantity_estimation_text" in row["issue_codes"]
    assert report["summary"]["filtered_importable_row_count"] == 0


def test_external_evidence_quality_writes_review_workbook(tmp_path: Path):
    report = build_external_evidence_quality_report(_external_results(), include_review=True)

    outputs = write_external_evidence_quality_outputs(report, tmp_path / "out", stem="quality")

    assert Path(outputs["json"]).exists()
    assert Path(outputs["csv"]).exists()
    assert Path(outputs["markdown"]).exists()
    assert Path(outputs["xlsx"]).exists()
    workbook = load_workbook(outputs["xlsx"])
    assert workbook.sheetnames == ["quality_summary", "quality_detail"]
    assert workbook["quality_detail"]["B2"].value == "accepted"
    workbook.close()


def test_external_evidence_quality_cli_merges_batches(tmp_path: Path, monkeypatch, capsys):
    external_1 = tmp_path / "external_1.json"
    external_2 = tmp_path / "external_2.json"
    rows = _external_results()["evidence_rows"]
    external_1.write_text(json.dumps({"evidence_rows": rows[:1]}, ensure_ascii=False), encoding="utf-8")
    external_2.write_text(json.dumps({"evidence_rows": rows[1:]}, ensure_ascii=False), encoding="utf-8")
    output_dir = tmp_path / "cli_out"
    monkeypatch.setattr(
        sys,
        "argv",
        [
            "biz2x_pdf_external_evidence_quality.py",
            "--external-results",
            str(external_1),
            "--external-results",
            str(external_2),
            "--output-dir",
            str(output_dir),
            "--timestamp",
            "fixed",
            "--include-review",
        ],
    )

    assert biz2x_pdf_external_evidence_quality.main() == 0
    payload = json.loads(capsys.readouterr().out)

    assert payload["ok"] is True
    assert payload["summary"]["input_row_count"] == 3
    assert payload["summary"]["filtered_importable_row_count"] == 2
    assert Path(payload["outputs"]["xlsx"]).exists()
