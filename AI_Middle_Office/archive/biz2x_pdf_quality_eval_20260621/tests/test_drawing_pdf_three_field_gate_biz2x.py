from __future__ import annotations

import json
import sys
from pathlib import Path

from openpyxl import load_workbook

from app.services.drawing_pdf_three_field_gate import (
    build_three_field_quality_gate,
    write_three_field_quality_gate_outputs,
)
from scripts import biz2x_pdf_three_field_gate


def _failed_review_report() -> dict[str, object]:
    return {
        "summary": {
            "answer_count": 10,
            "candidate_count": 8,
            "matched_three_fields_count": 6,
            "three_field_pass_rate": 0.6,
            "status_counts": {
                "matched_three_fields": 6,
                "missing_candidate": 2,
                "unit_conflict": 1,
                "matched_name_unit_feature_review": 1,
            },
            "quantity_status": "deferred_until_three_fields_accepted",
        },
        "review_rows": [],
    }


def _passed_review_report() -> dict[str, object]:
    return {
        "summary": {
            "answer_count": 2,
            "candidate_count": 2,
            "matched_three_fields_count": 2,
            "three_field_pass_rate": 1.0,
            "status_counts": {"matched_three_fields": 2},
        }
    }


def test_three_field_quality_gate_blocks_quantity_until_all_three_fields_pass():
    gate = build_three_field_quality_gate(_failed_review_report())

    assert gate["status"] == "failed"
    assert gate["can_enable_quantity"] is False
    assert gate["summary"]["quantity_status"] == "deferred_until_three_fields_accepted"
    failed_gates = {row["gate"] for row in gate["gate_rows"] if row["status"] == "fail"}
    assert "three_field_pass_rate" in failed_gates
    assert "missing_candidate" in failed_gates
    assert "unit_conflict" in failed_gates
    assert "feature_review" in failed_gates


def test_three_field_quality_gate_passes_clean_acceptance():
    gate = build_three_field_quality_gate(_passed_review_report())

    assert gate["status"] == "passed"
    assert gate["can_enable_quantity"] is True
    assert gate["summary"]["quantity_status"] == "ready_after_three_field_acceptance"


def test_three_field_quality_gate_writes_outputs(tmp_path: Path):
    gate = build_three_field_quality_gate(_failed_review_report())

    outputs = write_three_field_quality_gate_outputs(gate, tmp_path, stem="gate")

    assert Path(outputs["json"]).exists()
    assert Path(outputs["csv"]).exists()
    assert Path(outputs["markdown"]).exists()
    assert Path(outputs["xlsx"]).exists()
    workbook = load_workbook(outputs["xlsx"])
    assert workbook.sheetnames == ["gate_summary", "gate_checks"]
    assert workbook["gate_checks"]["A1"].value == "gate"


def test_three_field_quality_gate_cli(tmp_path: Path, monkeypatch, capsys):
    report_json = tmp_path / "review.json"
    report_json.write_text(json.dumps(_failed_review_report(), ensure_ascii=False), encoding="utf-8")
    output_dir = tmp_path / "cli_out"
    monkeypatch.setattr(
        sys,
        "argv",
        [
            "biz2x_pdf_three_field_gate.py",
            "--report-json",
            str(report_json),
            "--output-dir",
            str(output_dir),
            "--timestamp",
            "fixed",
        ],
    )

    assert biz2x_pdf_three_field_gate.main() == 0
    payload = json.loads(capsys.readouterr().out)

    assert payload["status"] == "failed"
    assert payload["can_enable_quantity"] is False
    assert Path(payload["outputs"]["xlsx"]).exists()
