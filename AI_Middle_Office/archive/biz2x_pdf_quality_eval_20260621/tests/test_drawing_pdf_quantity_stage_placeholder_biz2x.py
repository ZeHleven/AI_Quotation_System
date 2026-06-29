from __future__ import annotations

import json
import sys
from pathlib import Path

from openpyxl import load_workbook

from app.services.drawing_pdf_quantity_stage_placeholder import (
    build_quantity_stage_placeholder_report,
    write_quantity_stage_placeholder_outputs,
)
from scripts import biz2x_pdf_quantity_stage_placeholder


def _standard_bill_report(can_enable_quantity: bool = False) -> dict[str, object]:
    return {
        "summary": {
            "can_enable_quantity": can_enable_quantity,
            "export_mode": "final_candidate" if can_enable_quantity else "review_only",
        },
        "bill_rows": [
            {
                "row_no": 1,
                "export_status": "final_candidate" if can_enable_quantity else "review_only_three_field_gate_failed",
                "division": "floor",
                "standard_code": "GBT50854-2024",
                "item_code": "011102003",
                "item_name": "tile floor",
                "feature": "600x600 tile",
                "unit": "m2",
                "source_files": "drawing.pdf",
                "evidence_ids": "EVD-001",
            },
            {
                "row_no": 2,
                "export_status": "standard_unmapped_review",
                "division": "custom",
                "standard_code": "",
                "item_code": "",
                "item_name": "custom item",
                "feature": "custom feature",
                "unit": "item",
                "source_files": "drawing.pdf",
                "evidence_ids": "EVD-002",
            },
        ],
    }


def _v2_report() -> dict[str, object]:
    return {
        "human_style_rows": [
            {
                "standard_candidates": [
                    {
                        "item_code": "011102003",
                        "official_item_code": "011102003",
                        "quantity_rule": {
                            "formula_type": "area",
                            "rule_text": "measure by designed floor area",
                            "required_evidence": ["design drawing", "area boundary"],
                        },
                    }
                ]
            },
            {"standard_candidates": []},
        ]
    }


def _failed_gate() -> dict[str, object]:
    return {"can_enable_quantity": False}


def _passed_gate() -> dict[str, object]:
    return {"can_enable_quantity": True}


def test_quantity_placeholder_blocks_all_rows_when_three_field_gate_failed():
    report = build_quantity_stage_placeholder_report(
        _standard_bill_report(can_enable_quantity=False),
        v2_report=_v2_report(),
        gate_report=_failed_gate(),
    )

    assert report["safe_for_final_quantity_list"] is False
    assert report["summary"]["quantity_status"] == "blocked_until_three_field_gate_passed"
    assert report["summary"]["blocked_count"] == 2
    assert report["summary"]["gate_blocked_count"] == 2
    assert report["summary"]["quantity_filled_count"] == 0
    assert report["quantity_rows"][0]["quantity_status"] == "blocked_three_field_gate_failed"
    assert report["quantity_rows"][0]["quantity"] == ""
    assert report["quantity_rows"][0]["formula_type"] == "area"


def test_quantity_placeholder_marks_only_final_mapped_rows_ready_after_gate():
    report = build_quantity_stage_placeholder_report(
        _standard_bill_report(can_enable_quantity=True),
        v2_report=_v2_report(),
        gate_report=_passed_gate(),
    )

    assert report["summary"]["ready_placeholder_count"] == 1
    assert report["summary"]["standard_unmapped_count"] == 1
    assert report["quantity_rows"][0]["quantity_status"] == "ready_for_quantity_engine"
    assert report["quantity_rows"][1]["quantity_status"] == "blocked_standard_unmapped"


def test_quantity_placeholder_writes_workbook(tmp_path: Path):
    report = build_quantity_stage_placeholder_report(
        _standard_bill_report(can_enable_quantity=False),
        v2_report=_v2_report(),
        gate_report=_failed_gate(),
    )

    outputs = write_quantity_stage_placeholder_outputs(report, tmp_path, stem="quantity")

    assert Path(outputs["json"]).exists()
    assert Path(outputs["csv"]).exists()
    assert Path(outputs["markdown"]).exists()
    assert Path(outputs["xlsx"]).exists()
    workbook = load_workbook(outputs["xlsx"])
    assert workbook.sheetnames == ["summary", "quantity_placeholder"]
    assert workbook["quantity_placeholder"]["A1"].value == "序号"
    assert workbook["quantity_placeholder"]["B2"].value == "blocked_three_field_gate_failed"


def test_quantity_placeholder_cli(tmp_path: Path, monkeypatch, capsys):
    bill_json = tmp_path / "bill.json"
    v2_json = tmp_path / "v2.json"
    gate_json = tmp_path / "gate.json"
    bill_json.write_text(json.dumps(_standard_bill_report(False), ensure_ascii=False), encoding="utf-8")
    v2_json.write_text(json.dumps(_v2_report(), ensure_ascii=False), encoding="utf-8")
    gate_json.write_text(json.dumps(_failed_gate(), ensure_ascii=False), encoding="utf-8")
    output_dir = tmp_path / "cli_out"
    monkeypatch.setattr(
        sys,
        "argv",
        [
            "biz2x_pdf_quantity_stage_placeholder.py",
            "--standard-bill-json",
            str(bill_json),
            "--v2-json",
            str(v2_json),
            "--gate-json",
            str(gate_json),
            "--output-dir",
            str(output_dir),
            "--timestamp",
            "fixed",
        ],
    )

    assert biz2x_pdf_quantity_stage_placeholder.main() == 0
    payload = json.loads(capsys.readouterr().out)

    assert payload["safe_for_final_quantity_list"] is False
    assert payload["summary"]["blocked_count"] == 2
    assert Path(payload["outputs"]["xlsx"]).exists()
