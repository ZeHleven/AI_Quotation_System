from __future__ import annotations

import csv
import hashlib
import json
import math
import re
import struct
from collections import Counter, defaultdict
from dataclasses import dataclass
from datetime import datetime
from io import BytesIO
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from app.services.enterprise_quota_units import normalize_enterprise_quota_unit


PHASE0_VERSION = "biz2x-enterprise-quota-phase0-v0"
SUPPORTED_EXTENSIONS = {".xls", ".xlsx", ".xlsm"}
EXPECTED_SHEETS = ("企业定额", "劳务指导价", "材料价格库")

ENTERPRISE_QUOTA_COLUMNS = {
    "quota_code": "定额编码",
    "row_type": "类型",
    "item_name": "项目名称",
    "work_content": "项目特征及工作内容",
    "worker_or_subtype": "类型",
    "unit": "单位",
    "quantity": "含量",
    "unit_price": "单价",
    "labor_fee": "人工费",
    "main_material_fee": "主材费",
    "auxiliary_material_fee": "辅材费",
    "machinery_fee": "机械费",
}

LABOR_GUIDE_COLUMNS = {
    "quota_code": "定额编码",
    "item_name": "项目名称",
    "work_content": "项目特征及工作内容",
    "worker_type": "类型",
    "unit": "单位",
    "quantity": "含量",
    "guide_price": "指导价(待业务确认列名)",
}

MATERIAL_PRICE_BLOCKS = (
    {"unit_col": 6, "price_col": 7, "computed_col": 8, "label": "价格块1"},
    {"unit_col": 10, "tax_col": 11, "price_col": 12, "label": "价格块2"},
    {"unit_col": 13, "tax_col": 14, "price_col": 15, "label": "价格块3"},
    {"unit_col": 16, "tax_col": 17, "price_col": 18, "label": "价格块4"},
    {"unit_col": 19, "tax_col": 20, "price_col": 21, "label": "价格块5"},
)

_CODE_RE = re.compile(r"^[A-Za-z]{2}\d{3,}$|^\d{2}[A-Za-z]{2}\d{3,}$")
_NUMERIC_RE = re.compile(r"^-?\d+(?:\.\d+)?$")
_OLE_MAGIC = bytes.fromhex("D0 CF 11 E0 A1 B1 1A E1")
_END_OF_CHAIN = 0xFFFFFFFE
_FREE_SECTOR = 0xFFFFFFFF


class EnterpriseQuotaPhase0Error(ValueError):
    pass


@dataclass(frozen=True)
class SheetRows:
    name: str
    rows: list[list[Any]]


def preview_enterprise_quota_file(path: str | Path) -> dict[str, Any]:
    source_path = Path(path)
    if not source_path.exists():
        raise EnterpriseQuotaPhase0Error(f"文件不存在: {source_path}")
    if not source_path.is_file():
        raise EnterpriseQuotaPhase0Error(f"路径不是文件: {source_path}")
    if source_path.suffix.lower() not in SUPPORTED_EXTENSIONS:
        raise EnterpriseQuotaPhase0Error("阶段 0 仅支持 .xls/.xlsx/.xlsm 企业定额文件")

    content = source_path.read_bytes()
    sheets = read_workbook_rows(content, filename=source_path.name)
    result = analyze_enterprise_quota_rows(sheets)
    result["source"] = {
        "file_name": source_path.name,
        "file_path": str(source_path),
        "file_type": source_path.suffix.lower().lstrip("."),
        "file_size": source_path.stat().st_size,
        "sha256": hashlib.sha256(content).hexdigest(),
    }
    return result


def read_workbook_rows(content: bytes, *, filename: str | None = None) -> list[SheetRows]:
    if not content:
        raise EnterpriseQuotaPhase0Error("企业定额文件为空")
    suffix = Path(filename or "").suffix.lower()
    if suffix == ".xls" or content.startswith(_OLE_MAGIC):
        return _read_legacy_xls_rows(content)
    if suffix in {"", ".xlsx", ".xlsm"}:
        return _read_openpyxl_rows(content)
    raise EnterpriseQuotaPhase0Error("阶段 0 仅支持 .xls/.xlsx/.xlsm 企业定额文件")


def analyze_enterprise_quota_rows(sheets: list[SheetRows]) -> dict[str, Any]:
    sheet_map = {sheet.name: sheet for sheet in sheets}
    issues: list[dict[str, Any]] = []
    sheet_summaries = [_sheet_summary(sheet) for sheet in sheets]

    missing_sheets = [sheet_name for sheet_name in EXPECTED_SHEETS if sheet_name not in sheet_map]
    for sheet_name in missing_sheets:
        issues.append(_issue("error", sheet_name, None, "MISSING_SHEET", f"缺少必需 Sheet: {sheet_name}"))

    enterprise_result = _analyze_enterprise_sheet(sheet_map.get("企业定额"), issues)
    labor_result = _analyze_labor_sheet(sheet_map.get("劳务指导价"), issues)
    material_result = _analyze_material_sheet(sheet_map.get("材料价格库"), issues)

    severity_counts = Counter(issue["severity"] for issue in issues)
    summary = {
        "sheet_count": len(sheets),
        "missing_sheet_count": len(missing_sheets),
        "enterprise_quota_section_count": enterprise_result["section_count"],
        "enterprise_quota_item_count": enterprise_result["item_count"],
        "enterprise_quota_component_count": enterprise_result["component_count"],
        "labor_guide_candidate_count": labor_result["candidate_count"],
        "material_resource_candidate_count": material_result["candidate_count"],
        "error_count": severity_counts.get("error", 0),
        "warning_count": severity_counts.get("warning", 0),
    }

    return {
        "ok": summary["error_count"] == 0,
        "version": PHASE0_VERSION,
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "source": {},
        "summary": summary,
        "sheet_summaries": sheet_summaries,
        "field_mappings": {
            "企业定额": ENTERPRISE_QUOTA_COLUMNS,
            "劳务指导价": LABOR_GUIDE_COLUMNS,
            "材料价格库": {
                "resource_code": "推定列: 第2列",
                "resource_name": "推定列: 第3列",
                "manual_mapping_required": True,
                "detected_price_blocks": MATERIAL_PRICE_BLOCKS,
            },
        },
        "enterprise_quota": enterprise_result,
        "labor_guide": labor_result,
        "material_price_library": material_result,
        "issues": issues,
    }


def write_phase0_outputs(result: dict[str, Any], output_dir: str | Path, *, stem: str | None = None) -> dict[str, str]:
    target_dir = Path(output_dir)
    target_dir.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    base = stem or f"enterprise_quota_phase0_{timestamp}"

    json_path = target_dir / f"{base}.json"
    markdown_path = target_dir / f"{base}.md"
    issues_path = target_dir / f"{base}_issues.csv"

    json_path.write_text(json.dumps(result, ensure_ascii=False, indent=2), encoding="utf-8")
    markdown_path.write_text(build_phase0_markdown(result), encoding="utf-8")
    write_phase0_issues_csv(result, issues_path)

    return {
        "json": str(json_path),
        "markdown": str(markdown_path),
        "issues_csv": str(issues_path),
    }


def build_phase0_markdown(result: dict[str, Any]) -> str:
    summary = result.get("summary", {})
    source = result.get("source", {})
    enterprise = result.get("enterprise_quota", {})
    labor = result.get("labor_guide", {})
    material = result.get("material_price_library", {})
    lines = [
        "# BIZ-2x 企业定额主库 Phase 0 解析预览",
        "",
        "## 文件",
        "",
        f"- 文件名: {source.get('file_name', '')}",
        f"- 文件类型: {source.get('file_type', '')}",
        f"- SHA256: {source.get('sha256', '')}",
        "",
        "## 汇总",
        "",
        f"- Sheet 数: {summary.get('sheet_count', 0)}",
        f"- 企业定额分部: {summary.get('enterprise_quota_section_count', 0)}",
        f"- 企业定额主项: {summary.get('enterprise_quota_item_count', 0)}",
        f"- 企业定额组成明细: {summary.get('enterprise_quota_component_count', 0)}",
        f"- 劳务指导价候选行: {summary.get('labor_guide_candidate_count', 0)}",
        f"- 材料价格库候选资源: {summary.get('material_resource_candidate_count', 0)}",
        f"- 错误: {summary.get('error_count', 0)}",
        f"- 警告: {summary.get('warning_count', 0)}",
        "",
        "## Sheet 结构",
        "",
        "| Sheet | 行数 | 列数 | 非空单元格 |",
        "|---|---:|---:|---:|",
    ]
    for sheet in result.get("sheet_summaries", []):
        lines.append(
            f"| {sheet.get('name', '')} | {sheet.get('row_count', 0)} | {sheet.get('column_count', 0)} | {sheet.get('non_empty_cell_count', 0)} |"
        )

    lines.extend(
        [
            "",
            "## 企业定额 Sheet",
            "",
            f"- 表头行: {enterprise.get('header_row_index') or '未识别'}",
            f"- 分部: {enterprise.get('section_count', 0)}",
            f"- 定额主项: {enterprise.get('item_count', 0)}",
            f"- 组成明细: {enterprise.get('component_count', 0)}",
            f"- 组成类型: {json.dumps(enterprise.get('component_type_counts', {}), ensure_ascii=False)}",
            "",
            "## 劳务指导价 Sheet",
            "",
            f"- 表头行: {labor.get('header_row_index') or '未识别'}",
            f"- 候选行: {labor.get('candidate_count', 0)}",
            "",
            "## 材料价格库 Sheet",
            "",
            f"- 候选资源行: {material.get('candidate_count', 0)}",
            f"- 需要人工确认列映射: {'是' if material.get('manual_mapping_required') else '否'}",
            "",
            "## 问题清单",
            "",
            "| 级别 | Sheet | 行号 | 代码 | 说明 |",
            "|---|---|---:|---|---|",
        ]
    )
    for issue in result.get("issues", [])[:200]:
        row_display = issue.get("row_index") or ""
        lines.append(
            f"| {issue.get('severity', '')} | {issue.get('sheet', '')} | {row_display} | {issue.get('code', '')} | {_escape_md(issue.get('message', ''))} |"
        )
    if len(result.get("issues", [])) > 200:
        lines.append(f"| warning | - |  | TRUNCATED | 仅展示前 200 条，完整清单见 CSV/JSON |")
    lines.append("")
    return "\n".join(lines)


def write_phase0_issues_csv(result: dict[str, Any], path: str | Path) -> None:
    with Path(path).open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.DictWriter(
            handle,
            fieldnames=("severity", "sheet", "row_index", "code", "message", "evidence"),
        )
        writer.writeheader()
        for issue in result.get("issues", []):
            writer.writerow(
                {
                    "severity": issue.get("severity", ""),
                    "sheet": issue.get("sheet", ""),
                    "row_index": issue.get("row_index", ""),
                    "code": issue.get("code", ""),
                    "message": issue.get("message", ""),
                    "evidence": json.dumps(issue.get("evidence", {}), ensure_ascii=False),
                }
            )


def _analyze_enterprise_sheet(sheet: SheetRows | None, issues: list[dict[str, Any]]) -> dict[str, Any]:
    if sheet is None:
        return _empty_enterprise_result()

    header_index = _find_header_row(sheet.rows, ("定额编码", "项目名称", "单价"))
    if header_index is None:
        issues.append(_issue("error", sheet.name, None, "HEADER_NOT_FOUND", "未识别到企业定额表头"))
        return _empty_enterprise_result()

    sections: list[dict[str, Any]] = []
    items: list[dict[str, Any]] = []
    components: list[dict[str, Any]] = []
    item_codes: set[str] = set()
    duplicate_codes: set[str] = set()
    current_section_code = ""
    current_section_name = ""
    component_type_counts: Counter[str] = Counter()

    for row_index, row in _iter_rows_after(sheet.rows, header_index):
        if _row_is_blank(row):
            continue
        data = _enterprise_row_data(row)
        row_type = data["row_type"]
        quota_code = data["quota_code"]

        if row_type == "分部":
            current_section_code = quota_code
            current_section_name = data["item_name"]
            if not quota_code or not current_section_name:
                issues.append(_issue("warning", sheet.name, row_index, "SECTION_INCOMPLETE", "分部行缺少编码或名称", data))
            sections.append(
                {
                    "row_index": row_index,
                    "section_code": quota_code,
                    "section_name": current_section_name,
                    "source_sheet": sheet.name,
                    "raw_row": data,
                }
            )
            continue

        if row_type == "定额":
            if not quota_code:
                issues.append(_issue("warning", sheet.name, row_index, "QUOTA_CODE_MISSING", "定额主项缺少定额编码", data))
            elif quota_code in item_codes:
                duplicate_codes.add(quota_code)
                issues.append(_issue("error", sheet.name, row_index, "QUOTA_CODE_DUPLICATED", f"定额主项编码重复: {quota_code}", data))
            item_codes.add(quota_code)
            _validate_quota_item(sheet.name, row_index, data, issues)
            items.append(
                {
                    "row_index": row_index,
                    "quota_code": quota_code,
                    "section_code": current_section_code,
                    "section_name": current_section_name,
                    "item_name": data["item_name"],
                    "work_content": data["work_content"],
                    "worker_or_subtype": data["worker_or_subtype"],
                    "unit": data["unit"],
                    "quantity": data["quantity"],
                    "unit_price": data["unit_price"],
                    "labor_fee": data["labor_fee"],
                    "main_material_fee": data["main_material_fee"],
                    "auxiliary_material_fee": data["auxiliary_material_fee"],
                    "machinery_fee": data["machinery_fee"],
                    "source_sheet": sheet.name,
                    "raw_row": data,
                }
            )
            continue

        component_type_counts[row_type or "未分类"] += 1
        _validate_component(sheet.name, row_index, data, item_codes, issues)
        components.append(
            {
                "row_index": row_index,
                "parent_quota_code": quota_code,
                "component_type": row_type,
                "resource_code": data["item_name"],
                "resource_name": data["work_content"],
                "worker_or_subtype": data["worker_or_subtype"],
                "unit": data["unit"],
                "quantity": data["quantity"],
                "unit_price": data["unit_price"],
                "amount": data["component_amount"],
                "source_sheet": sheet.name,
                "raw_row": data,
            }
        )

    return {
        "header_row_index": header_index + 1,
        "section_count": len(sections),
        "item_count": len(items),
        "component_count": sum(component_type_counts.values()),
        "duplicate_quota_codes": sorted(duplicate_codes),
        "component_type_counts": dict(component_type_counts),
        "sections": sections,
        "items": items,
        "components": components,
        "section_samples": sections[:20],
        "item_samples": items[:20],
        "component_samples": components[:20],
    }


def _analyze_labor_sheet(sheet: SheetRows | None, issues: list[dict[str, Any]]) -> dict[str, Any]:
    if sheet is None:
        return {"header_row_index": None, "candidate_count": 0, "candidates": [], "samples": []}

    header_index = _find_header_row(sheet.rows, ("项目名称", "单位", "含量"))
    if header_index is None:
        issues.append(_issue("warning", sheet.name, None, "HEADER_NOT_FOUND", "未识别到劳务指导价表头"))
        return {"header_row_index": None, "candidate_count": 0, "candidates": [], "samples": []}

    candidates: list[dict[str, Any]] = []
    for row_index, row in _iter_rows_after(sheet.rows, header_index):
        if _row_is_blank(row):
            continue
        item_name = _cell(row, 2)
        work_content = _cell(row, 3)
        unit = _unit_cell(row, 6)
        quantity = _number(_cell(row, 7))
        guide_price = _number(_cell(row, 8))
        if not (item_name or work_content or guide_price is not None):
            continue
        if guide_price is None:
            issues.append(_issue("warning", sheet.name, row_index, "LABOR_PRICE_MISSING", "劳务指导价候选行缺少价格", {"item_name": item_name}))
        if unit == "":
            issues.append(_issue("warning", sheet.name, row_index, "LABOR_UNIT_MISSING", "劳务指导价候选行缺少单位", {"item_name": item_name}))
        candidate = {
            "row_index": row_index,
            "quota_code": _cell(row, 1),
            "item_name": item_name,
            "work_content": work_content,
            "worker_type": _cell(row, 5),
            "unit": unit,
            "quantity": quantity,
            "guide_price": guide_price,
            "source_sheet": sheet.name,
            "raw_row": {
                "quota_code": _cell(row, 1),
                "item_name": item_name,
                "work_content": work_content,
                "worker_type": _cell(row, 5),
                "unit": unit,
                "quantity": quantity,
                "guide_price": guide_price,
            },
        }
        candidates.append(candidate)

    return {
        "header_row_index": header_index + 1,
        "candidate_count": len(candidates),
        "candidates": candidates,
        "samples": candidates[:20],
    }


def _analyze_material_sheet(sheet: SheetRows | None, issues: list[dict[str, Any]]) -> dict[str, Any]:
    if sheet is None:
        return {"candidate_count": 0, "manual_mapping_required": True, "candidates": [], "samples": []}

    candidates: list[dict[str, Any]] = []
    block_counts: Counter[str] = Counter()
    for row_index, row in enumerate(sheet.rows, start=1):
        if _row_is_blank(row):
            continue
        resource_code = _cell(row, 2)
        resource_name = _cell(row, 3)
        prices: list[dict[str, Any]] = []
        for block in MATERIAL_PRICE_BLOCKS:
            unit = _unit_cell(row, block["unit_col"])
            price = _number(_cell(row, block["price_col"]))
            computed = _number(_cell(row, block.get("computed_col", block["price_col"])))
            tax_rate = _number(_cell(row, block.get("tax_col", 0))) if block.get("tax_col") else None
            if unit or price is not None or computed is not None:
                block_counts[block["label"]] += 1
                prices.append(
                    {
                        "block": block["label"],
                        "unit": unit,
                        "price": price,
                        "tax_rate": tax_rate,
                        "computed_price": computed,
                    }
                )
        if not prices and not (_looks_like_resource_code(resource_code) or resource_name):
            continue
        if not prices:
            issues.append(
                _issue(
                    "warning",
                    sheet.name,
                    row_index,
                    "MATERIAL_PRICE_BLOCK_EMPTY",
                    "材料候选行未识别到价格块，需要人工确认列映射",
                    {"resource_code": resource_code, "resource_name": resource_name},
                )
            )
        candidates.append(
            {
                "row_index": row_index,
                "resource_code": resource_code,
                "resource_name": resource_name,
                "price_blocks": prices,
                "source_sheet": sheet.name,
                "raw_row": {
                    "resource_code": resource_code,
                    "resource_name": resource_name,
                    "price_blocks": prices,
                },
            }
        )

    if candidates:
        issues.append(
            _issue(
                "warning",
                sheet.name,
                None,
                "MATERIAL_MAPPING_REQUIRES_CONFIRMATION",
                "材料价格库为横向多价格块结构，阶段 0 仅做候选识别，后续需人工确认列含义",
                {"detected_blocks": MATERIAL_PRICE_BLOCKS},
            )
        )

    return {
        "candidate_count": len(candidates),
        "manual_mapping_required": True,
        "detected_price_block_counts": dict(block_counts),
        "candidates": candidates,
        "samples": candidates[:20],
    }


def _validate_quota_item(sheet_name: str, row_index: int, data: dict[str, Any], issues: list[dict[str, Any]]) -> None:
    for field_name, label in (("item_name", "项目名称"), ("unit", "单位")):
        if not data[field_name]:
            issues.append(_issue("warning", sheet_name, row_index, f"QUOTA_{field_name.upper()}_MISSING", f"定额主项缺少{label}", data))
    if data["unit_price"] is None:
        issues.append(_issue("warning", sheet_name, row_index, "QUOTA_PRICE_INVALID", "定额主项单价为空或非数字", data))
        return
    split_values = [data["labor_fee"], data["main_material_fee"], data["auxiliary_material_fee"], data["machinery_fee"]]
    if all(value is not None for value in split_values):
        split_sum = sum(value or 0 for value in split_values)
        if not _money_close(split_sum, data["unit_price"]):
            issues.append(
                _issue(
                    "warning",
                    sheet_name,
                    row_index,
                    "QUOTA_PRICE_SPLIT_MISMATCH",
                    "定额单价与人工费+主材费+辅材费+机械费不一致",
                    {"unit_price": data["unit_price"], "split_sum": round(split_sum, 6), **data},
                )
            )


def _validate_component(
    sheet_name: str,
    row_index: int,
    data: dict[str, Any],
    item_codes: set[str],
    issues: list[dict[str, Any]],
) -> None:
    if not data["quota_code"]:
        issues.append(_issue("warning", sheet_name, row_index, "COMPONENT_PARENT_CODE_MISSING", "组成明细缺少父级定额编码", data))
    elif data["quota_code"] not in item_codes:
        issues.append(_issue("warning", sheet_name, row_index, "COMPONENT_PARENT_NOT_SEEN", "组成明细的父级定额编码尚未在前文出现", data))
    if not data["work_content"]:
        issues.append(_issue("warning", sheet_name, row_index, "COMPONENT_NAME_MISSING", "组成明细缺少资源名称", data))
    if not data["unit"]:
        issues.append(_issue("warning", sheet_name, row_index, "COMPONENT_UNIT_MISSING", "组成明细缺少单位", data))
    quantity = data["quantity"]
    unit_price = data["unit_price"]
    amount = data["component_amount"]
    if quantity is None or unit_price is None or amount is None:
        issues.append(_issue("warning", sheet_name, row_index, "COMPONENT_AMOUNT_INCOMPLETE", "组成明细含量、单价或金额缺失", data))
        return
    expected = quantity * unit_price
    if not _money_close(expected, amount):
        issues.append(
            _issue(
                "warning",
                sheet_name,
                row_index,
                "COMPONENT_AMOUNT_MISMATCH",
                "组成明细金额与含量×单价不一致",
                {"expected_amount": round(expected, 6), **data},
            )
        )


def _read_openpyxl_rows(content: bytes) -> list[SheetRows]:
    try:
        workbook = load_workbook(BytesIO(content), data_only=True, read_only=True)
    except Exception as exc:  # pragma: no cover - openpyxl raises implementation-specific exceptions.
        raise EnterpriseQuotaPhase0Error(f"Excel 读取失败，请确认文件格式: {exc}") from exc
    sheets: list[SheetRows] = []
    for sheet in workbook.worksheets:
        rows = [list(row) for row in sheet.iter_rows(values_only=True)]
        sheets.append(SheetRows(name=sheet.title, rows=_trim_dense_rows(rows)))
    return sheets


def _read_legacy_xls_rows(content: bytes) -> list[SheetRows]:
    reader = _LegacyXlsReader(content)
    return reader.read()


class _LegacyXlsReader:
    def __init__(self, content: bytes):
        self.content = content
        if not content.startswith(_OLE_MAGIC):
            raise EnterpriseQuotaPhase0Error("不是有效的 OLE2 .xls 文件")
        self.sector_size = 1 << self._u16(0x1E)
        self.fat = self._build_fat()

    def read(self) -> list[SheetRows]:
        workbook_stream = self._read_workbook_stream()
        records = self._read_biff_records(workbook_stream)
        sheet_refs = self._sheet_refs(records)
        shared_strings = self._shared_strings(records)
        sheets: list[SheetRows] = []
        for sheet_name, offset in sheet_refs:
            cells = self._sheet_cells(workbook_stream, offset, shared_strings)
            rows = self._cells_to_rows(cells)
            sheets.append(SheetRows(name=sheet_name, rows=rows))
        return sheets

    def _u16(self, offset: int) -> int:
        return struct.unpack_from("<H", self.content, offset)[0]

    def _u32(self, offset: int) -> int:
        return struct.unpack_from("<I", self.content, offset)[0]

    def _sector(self, sid: int) -> bytes:
        start = (sid + 1) * self.sector_size
        return self.content[start : start + self.sector_size]

    def _build_fat(self) -> list[int]:
        first_difat = self._u32(0x44)
        num_difat = self._u32(0x48)
        difat = [self._u32(0x4C + 4 * index) for index in range(109)]

        sid = first_difat
        for _ in range(num_difat):
            if sid in (_END_OF_CHAIN, _FREE_SECTOR):
                break
            sector = self._sector(sid)
            difat.extend(struct.unpack_from("<I", sector, 4 * index)[0] for index in range(self.sector_size // 4 - 1))
            sid = struct.unpack_from("<I", sector, self.sector_size - 4)[0]

        fat: list[int] = []
        for fat_sid in difat:
            if fat_sid in (_FREE_SECTOR, _END_OF_CHAIN) or fat_sid >= 0xFFFFFFF0:
                continue
            sector = self._sector(fat_sid)
            fat.extend(struct.unpack_from("<I", sector, 4 * index)[0] for index in range(self.sector_size // 4))
        return fat

    def _read_chain(self, start_sid: int, size: int | None = None) -> bytes:
        data = bytearray()
        sid = start_sid
        seen: set[int] = set()
        while sid not in (_END_OF_CHAIN, _FREE_SECTOR) and sid < len(self.fat) and sid not in seen:
            seen.add(sid)
            data.extend(self._sector(sid))
            sid = self.fat[sid]
            if size is not None and len(data) >= size:
                break
        return bytes(data[:size] if size is not None else data)

    def _read_workbook_stream(self) -> bytes:
        dir_start = self._u32(0x30)
        directory = self._read_chain(dir_start)
        for offset in range(0, len(directory), 128):
            entry = directory[offset : offset + 128]
            if len(entry) < 128:
                break
            name_len = struct.unpack_from("<H", entry, 64)[0]
            name = entry[: max(0, name_len - 2)].decode("utf-16le", "ignore") if name_len >= 2 else ""
            if name not in {"Workbook", "Book"}:
                continue
            start_sid = struct.unpack_from("<I", entry, 116)[0]
            size = struct.unpack_from("<I", entry, 120)[0]
            return self._read_chain(start_sid, size)
        raise EnterpriseQuotaPhase0Error("未找到 .xls Workbook 流")

    @staticmethod
    def _read_biff_records(workbook_stream: bytes) -> list[tuple[int, bytes, int]]:
        records: list[tuple[int, bytes, int]] = []
        offset = 0
        while offset + 4 <= len(workbook_stream):
            record_id = struct.unpack_from("<H", workbook_stream, offset)[0]
            length = struct.unpack_from("<H", workbook_stream, offset + 2)[0]
            payload = workbook_stream[offset + 4 : offset + 4 + length]
            records.append((record_id, payload, offset))
            offset += 4 + length
        return records

    @staticmethod
    def _sheet_refs(records: list[tuple[int, bytes, int]]) -> list[tuple[str, int]]:
        refs: list[tuple[str, int]] = []
        for record_id, payload, _ in records:
            if record_id != 0x0085 or len(payload) < 8:
                continue
            offset = struct.unpack_from("<I", payload, 0)[0]
            name_len = payload[6]
            flags = payload[7]
            width = 2 if flags & 1 else 1
            raw = payload[8 : 8 + name_len * width]
            name = raw.decode("utf-16le" if width == 2 else "latin1", "ignore")
            refs.append((name, offset))
        return refs

    @staticmethod
    def _shared_strings(records: list[tuple[int, bytes, int]]) -> list[str]:
        for index, (record_id, payload, _) in enumerate(records):
            if record_id != 0x00FC or len(payload) < 8:
                continue
            unique_count = struct.unpack_from("<I", payload, 4)[0]
            segments = [payload[8:]]
            cursor = index + 1
            while cursor < len(records) and records[cursor][0] == 0x003C:
                segments.append(records[cursor][1])
                cursor += 1
            reader = _BiffSegmentReader(segments)
            strings: list[str] = []
            for _ in range(unique_count):
                strings.append(reader.read_string())
            return strings
        return []

    def _sheet_cells(self, workbook_stream: bytes, offset: int, shared_strings: list[str]) -> dict[tuple[int, int], Any]:
        cells: dict[tuple[int, int], Any] = {}
        pending_formula_cell: tuple[int, int] | None = None
        cursor = offset
        while cursor + 4 <= len(workbook_stream):
            record_id = struct.unpack_from("<H", workbook_stream, cursor)[0]
            length = struct.unpack_from("<H", workbook_stream, cursor + 2)[0]
            payload = workbook_stream[cursor + 4 : cursor + 4 + length]
            cursor += 4 + length
            if record_id == 0x000A:
                break
            if record_id == 0x00FD and len(payload) >= 10:
                row = struct.unpack_from("<H", payload, 0)[0]
                col = struct.unpack_from("<H", payload, 2)[0]
                sst_index = struct.unpack_from("<I", payload, 6)[0]
                cells[(row, col)] = shared_strings[sst_index] if sst_index < len(shared_strings) else ""
            elif record_id == 0x0204 and len(payload) >= 9:
                row = struct.unpack_from("<H", payload, 0)[0]
                col = struct.unpack_from("<H", payload, 2)[0]
                cells[(row, col)] = _decode_biff_string(payload, 6)
            elif record_id == 0x0203 and len(payload) >= 14:
                row = struct.unpack_from("<H", payload, 0)[0]
                col = struct.unpack_from("<H", payload, 2)[0]
                cells[(row, col)] = struct.unpack_from("<d", payload, 6)[0]
            elif record_id == 0x027E and len(payload) >= 10:
                row = struct.unpack_from("<H", payload, 0)[0]
                col = struct.unpack_from("<H", payload, 2)[0]
                cells[(row, col)] = _decode_rk(struct.unpack_from("<I", payload, 6)[0])
            elif record_id == 0x00BD and len(payload) >= 8:
                row = struct.unpack_from("<H", payload, 0)[0]
                first_col = struct.unpack_from("<H", payload, 2)[0]
                last_col = struct.unpack_from("<H", payload, len(payload) - 2)[0]
                record_offset = 4
                for col in range(first_col, last_col + 1):
                    if record_offset + 6 > len(payload) - 2:
                        break
                    cells[(row, col)] = _decode_rk(struct.unpack_from("<I", payload, record_offset + 2)[0])
                    record_offset += 6
            elif record_id == 0x0006 and len(payload) >= 14:
                row = struct.unpack_from("<H", payload, 0)[0]
                col = struct.unpack_from("<H", payload, 2)[0]
                raw_result = payload[6:14]
                if raw_result[:2] == b"\xff\xff" and raw_result[2] == 0:
                    pending_formula_cell = (row, col)
                elif raw_result[:2] == b"\xff\xff":
                    cells[(row, col)] = ""
                else:
                    value = struct.unpack("<d", raw_result)[0]
                    if math.isfinite(value):
                        cells[(row, col)] = value
            elif record_id == 0x0207 and pending_formula_cell:
                cells[pending_formula_cell] = _decode_biff_string(payload, 0)
                pending_formula_cell = None
        return cells

    @staticmethod
    def _cells_to_rows(cells: dict[tuple[int, int], Any]) -> list[list[Any]]:
        if not cells:
            return []
        max_row = max(row for row, _ in cells)
        max_col = max(col for _, col in cells)
        rows: list[list[Any]] = []
        for row_index in range(max_row + 1):
            row = [cells.get((row_index, col_index)) for col_index in range(max_col + 1)]
            rows.append(row)
        return _trim_dense_rows(rows)


class _BiffSegmentReader:
    def __init__(self, segments: list[bytes]):
        self.segments = segments
        self.segment_index = 0
        self.offset = 0

    def read(self, size: int) -> bytes:
        data = bytearray()
        while size > 0 and self.segment_index < len(self.segments):
            segment = self.segments[self.segment_index]
            take = min(size, len(segment) - self.offset)
            data.extend(segment[self.offset : self.offset + take])
            self.offset += take
            size -= take
            if self.offset >= len(segment):
                self.segment_index += 1
                self.offset = 0
        return bytes(data)

    def read_u8(self) -> int:
        data = self.read(1)
        return data[0] if data else 0

    def read_u16(self) -> int:
        return struct.unpack("<H", self.read(2).ljust(2, b"\0"))[0]

    def read_u32(self) -> int:
        return struct.unpack("<I", self.read(4).ljust(4, b"\0"))[0]

    def read_string(self) -> str:
        char_count = self.read_u16()
        flags = self.read_u8()
        rich_text_runs = self.read_u16() if flags & 0x08 else 0
        ext_size = self.read_u32() if flags & 0x04 else 0
        char_width = 2 if flags & 1 else 1
        chunks: list[str] = []
        remaining = char_count
        while remaining > 0 and self.segment_index < len(self.segments):
            segment = self.segments[self.segment_index]
            if self.offset >= len(segment):
                self.segment_index += 1
                self.offset = 0
                if self.segment_index < len(self.segments):
                    char_width = 2 if self.read_u8() & 1 else 1
                continue
            take_chars = min(remaining, (len(segment) - self.offset) // char_width)
            chunk = self.read(take_chars * char_width)
            chunks.append(chunk.decode("utf-16le" if char_width == 2 else "latin1", "ignore"))
            remaining -= take_chars
            if remaining and self.segment_index < len(self.segments) and self.offset >= len(self.segments[self.segment_index]):
                self.segment_index += 1
                self.offset = 0
                if self.segment_index < len(self.segments):
                    char_width = 2 if self.read_u8() & 1 else 1
        if rich_text_runs:
            self.read(rich_text_runs * 4)
        if ext_size:
            self.read(ext_size)
        return "".join(chunks)


def _decode_biff_string(payload: bytes, offset: int) -> str:
    if len(payload) < offset + 3:
        return ""
    char_count = struct.unpack_from("<H", payload, offset)[0]
    flags = payload[offset + 2]
    cursor = offset + 3
    if flags & 0x08:
        cursor += 2
    if flags & 0x04:
        cursor += 4
    width = 2 if flags & 1 else 1
    raw = payload[cursor : cursor + char_count * width]
    return raw.decode("utf-16le" if width == 2 else "latin1", "ignore")


def _decode_rk(value: int) -> float:
    divide_by_100 = bool(value & 1)
    is_integer = bool(value & 2)
    if is_integer:
        number = value >> 2
        if number & (1 << 29):
            number -= 1 << 30
    else:
        raw = struct.pack("<I", 0) + struct.pack("<I", value & 0xFFFFFFFC)
        number = struct.unpack("<d", raw)[0]
    return number / 100 if divide_by_100 else float(number)


def _sheet_summary(sheet: SheetRows) -> dict[str, Any]:
    non_empty = 0
    max_col = 0
    row_count = 0
    samples: list[list[str]] = []
    for row in sheet.rows:
        if _row_is_blank(row):
            continue
        row_count += 1
        max_col = max(max_col, len(row))
        non_empty += sum(1 for value in row if _clean(value))
        if len(samples) < 8:
            samples.append([_clean(value) for value in row[:12]])
    return {
        "name": sheet.name,
        "row_count": len(sheet.rows),
        "non_empty_row_count": row_count,
        "column_count": max_col,
        "non_empty_cell_count": non_empty,
        "samples": samples,
    }


def _empty_enterprise_result() -> dict[str, Any]:
    return {
        "header_row_index": None,
        "section_count": 0,
        "item_count": 0,
        "component_count": 0,
        "duplicate_quota_codes": [],
        "component_type_counts": {},
        "section_samples": [],
        "item_samples": [],
        "component_samples": [],
    }


def _enterprise_row_data(row: list[Any]) -> dict[str, Any]:
    values = {
        "quota_code": _cell(row, 1),
        "row_type": _cell(row, 2),
        "item_name": _cell(row, 3),
        "work_content": _cell(row, 4),
        "worker_or_subtype": _cell(row, 5),
        "unit": _unit_cell(row, 6),
        "quantity": _number(_cell(row, 7)),
        "unit_price": _number(_cell(row, 8)),
        "labor_fee": _number(_cell(row, 9)),
        "main_material_fee": _number(_cell(row, 10)),
        "auxiliary_material_fee": _number(_cell(row, 11)),
        "machinery_fee": _number(_cell(row, 12)),
    }
    values["component_amount"] = next(
        (values[field] for field in ("labor_fee", "main_material_fee", "auxiliary_material_fee", "machinery_fee") if values[field] is not None),
        None,
    )
    return values


def _find_header_row(rows: list[list[Any]], required_terms: tuple[str, ...]) -> int | None:
    for index, row in enumerate(rows[:20]):
        row_text = "\t".join(_clean(value) for value in row)
        if all(term in row_text for term in required_terms):
            return index
    return None


def _iter_rows_after(rows: list[list[Any]], header_index: int):
    for zero_based_index in range(header_index + 1, len(rows)):
        yield zero_based_index + 1, rows[zero_based_index]


def _trim_dense_rows(rows: list[list[Any]]) -> list[list[Any]]:
    trimmed = [list(row) for row in rows]
    while trimmed and _row_is_blank(trimmed[-1]):
        trimmed.pop()
    max_col = 0
    for row in trimmed:
        for col_index, value in enumerate(row, start=1):
            if _clean(value):
                max_col = max(max_col, col_index)
    if max_col:
        trimmed = [row[:max_col] for row in trimmed]
    return trimmed


def _row_is_blank(row: list[Any]) -> bool:
    return all(not _clean(value) for value in row)


def _cell(row: list[Any], one_based_col: int) -> str:
    if one_based_col <= 0 or one_based_col > len(row):
        return ""
    return _clean(row[one_based_col - 1])


def _unit_cell(row: list[Any], one_based_col: int) -> str:
    return normalize_enterprise_quota_unit(_cell(row, one_based_col)) or ""


def _clean(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and math.isfinite(value) and abs(value - round(value)) < 1e-9:
        return str(int(round(value)))
    return str(value).replace("\r", " ").replace("\n", " ").strip()


def _number(value: Any) -> float | None:
    text = _clean(value)
    if not text:
        return None
    text = text.replace(",", "")
    if not _NUMERIC_RE.match(text):
        return None
    try:
        return float(text)
    except ValueError:
        return None


def _looks_like_resource_code(value: Any) -> bool:
    return bool(_CODE_RE.match(_clean(value)))


def _money_close(left: float, right: float, *, tolerance: float = 0.05) -> bool:
    return abs(left - right) <= max(tolerance, abs(right) * 0.002)


def _issue(
    severity: str,
    sheet: str,
    row_index: int | None,
    code: str,
    message: str,
    evidence: dict[str, Any] | None = None,
) -> dict[str, Any]:
    return {
        "severity": severity,
        "sheet": sheet,
        "row_index": row_index,
        "code": code,
        "message": message,
        "evidence": evidence or {},
    }


def _escape_md(value: Any) -> str:
    return _clean(value).replace("|", "\\|")
