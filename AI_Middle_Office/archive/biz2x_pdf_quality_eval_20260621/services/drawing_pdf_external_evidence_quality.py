from __future__ import annotations

import csv
import json
import re
from collections import Counter
from datetime import datetime
from pathlib import Path
from typing import Any, Iterable, Mapping, Sequence

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


PHASE = "BIZ-2x-pdf-external-evidence-quality"

QUALITY_HEADERS = [
    "row_no",
    "quality_status",
    "quality_score",
    "import_action",
    "issue_codes",
    "source_kind",
    "task_no",
    "task_nos",
    "source_file",
    "page",
    "tile_id",
    "vision_pass",
    "evidence_role",
    "discipline",
    "item_hint",
    "space",
    "material_codes",
    "spec_or_method",
    "suggested_unit",
    "text",
    "normalized_text",
    "confidence",
    "model",
    "needs_manual_review",
    "reason",
]

SUMMARY_HEADERS = ["metric", "value"]

UNKNOWN_UNITS = {
    "unknown",
    "unk",
    "n/a",
    "na",
    "none",
    "null",
    "unclear",
    "not sure",
    "not visible",
    "无法确定",
    "不详",
    "未知",
    "不清楚",
}

VALID_UNITS = {
    "㎡",
    "m2",
    "m²",
    "m3",
    "m³",
    "m^3",
    "m",
    "米",
    "立方米",
    "个",
    "只",
    "套",
    "台",
    "樘",
    "块",
    "项",
    "处",
    "set",
    "sets",
    "pc",
    "pcs",
    "piece",
    "pieces",
    "unit",
    "units",
}

PROMPT_PLACEHOLDER_TEXTS = {
    "图中可见项目名称",
    "可见规格、材质、安装方式；没有则留空",
    "可见规格、材质、安装方式;没有则留空",
    "规格、材料、做法、安装方式或构造说明",
    "材料编号、材料名称、规格、做法、安装方式或构造说明",
    "可见材料编号、尺寸、材质、做法或安装方式；没有则留空",
    "可见材料编号、尺寸、材质、做法或安装方式;没有则留空",
    "摘录图中可见文字或描述可见符号/引线位置",
    "为什么这是图中可见证据",
    "可见计划或图例证据",
    "规格/方法",
    "文字内容",
    "/",
}

GENERIC_ITEM_HINTS = {
    "removed finishes",
    "removed fixtures",
    "removed items",
    "existing finishes",
    "existing fixtures",
    "haul-away notes",
    "haul away notes",
    "demolition notes",
    "general demolition notes",
    "construction notes",
    "general notes",
    "notes",
    "fixtures",
    "finishes",
    "other items",
    "miscellaneous",
    "unclear item",
    "unknown item",
    "拆除说明",
    "清运说明",
    "图纸说明",
    "图纸目录",
    "施工说明",
    "一般说明",
    "材料清单或立面证据",
    "材料清单或立面图证据",
    "材料清单/立面图",
    "缺失的材料清单/立面",
    "墙面装饰材料表",
    "墙面装饰材料",
    "地面装饰材料",
    "天花板装饰材料",
    "吊顶材料",
    "墙面材料",
    "地面材料",
    "天花材料",
    "墙面装饰符号说明",
    "节点注释中的墙面装饰信息",
    "天花板",
    "瓷砖",
    "地板",
    "未指定",
    "其他项目",
    "未知项目",
    "不明确项目",
    "地面",
    "墙面",
    "天花",
    "吊顶",
    "门窗",
    "隔断",
    "材料",
    "材料编号",
    "日期",
    "装饰",
    "可见计划或图例证据",
    "文字内容",
}

HARD_GENERIC_ITEM_HINTS = {
    "地面",
    "墙面",
    "天花",
    "吊顶",
    "门窗",
    "隔断",
    "材料",
    "材料编号",
    "图中可见项目名称",
    "图纸目录",
    "日期",
    "装饰",
    "材料清单或立面证据",
    "材料清单或立面图证据",
    "材料清单/立面图",
    "缺失的材料清单/立面",
    "墙面装饰材料表",
    "墙面装饰材料",
    "地面装饰材料",
    "天花板装饰材料",
    "吊顶材料",
    "墙面材料",
    "地面材料",
    "天花材料",
    "墙面装饰符号说明",
    "节点注释中的墙面装饰信息",
    "天花板",
    "瓷砖",
    "地板",
    "未指定",
}

GENERIC_TEXT_PATTERNS = [
    r"\ball\b.{0,30}\b(existing|removed|finishes|fixtures)\b",
    r"\bremove(d)?\s+all\b",
    r"\bgeneral\s+(note|notes|demolition)\b",
    r"全部.{0,12}(拆除|清运|移除)",
    r"所有.{0,12}(拆除|清运|移除)",
    r"现有.{0,12}(全部|所有)",
]

SHORT_CONCRETE_TERMS = {
    "灯槽",
    "地漏",
    "马桶",
    "花洒",
    "台盆",
    "水表",
    "阀门",
    "门套",
    "踢脚线",
    "窗帘盒",
    "窗台石",
    "门槛石",
    "灯带",
    "筒灯",
    "射灯",
    "插座",
    "开关",
    "配管",
    "配线",
    "洁具",
    "镜面",
    "硬包",
    "墙布",
    "线条",
}

CONCRETE_HINT_PATTERNS = [
    r"\bDN\s*\d+",
    r"\bDe\s*\d+",
    r"\bSC\s*\d+",
    r"\bMT\s*\d+",
    r"\bWDZC[-A-Z0-9.]*",
    r"\bLED\b",
    r"\bCT[-_ ]?\d+",
    r"\bST[-_ ]?\d+",
    r"\bWD[-_ ]?\d+",
    r"\bMT[-_ ]?\d+",
    r"\d+\s*[xX*×]\s*\d+",
    r"\d+\s*(mm|cm|m|㎡|m2)",
]

UNCERTAIN_EVIDENCE_PATTERNS = [
    r"未直接",
    r"未读到",
    r"未可见",
    r"不能区分",
    r"无法区分",
    r"无法确定",
    r"不能完整证明",
    r"证据不足",
    r"需人工复核",
    r"需要人工复核",
    r"not\s+directly\s+visible",
    r"not\s+visible",
    r"cannot\s+distinguish",
    r"can\s+not\s+distinguish",
    r"unable\s+to\s+distinguish",
    r"insufficient\s+evidence",
    r"manual\s+review",
]


def build_external_evidence_quality_report(
    external_results: Mapping[str, Any],
    *,
    source_path: str | Path | None = None,
    include_review: bool = False,
) -> dict[str, Any]:
    source_rows = [_clean_source_row(row) for row in _external_source_rows(external_results)]
    quality_rows = [_score_source_row(row, row_no=index, include_review=include_review) for index, row in enumerate(source_rows, start=1)]
    accepted_count = sum(1 for row in quality_rows if row["quality_status"] == "accepted")
    review_count = sum(1 for row in quality_rows if row["quality_status"] == "review")
    rejected_count = sum(1 for row in quality_rows if row["quality_status"] == "rejected")
    filtered_rows = [
        dict(source_rows[index])
        for index, row in enumerate(quality_rows)
        if _is_importable_by_quality(row, include_review=include_review)
    ]
    summary = {
        "source_path": str(source_path or ""),
        "input_row_count": len(source_rows),
        "accepted_row_count": accepted_count,
        "review_row_count": review_count,
        "rejected_row_count": rejected_count,
        "filtered_importable_row_count": len(filtered_rows),
        "include_review": include_review,
        "status_counts": dict(Counter(str(row.get("quality_status") or "") for row in quality_rows)),
        "issue_counts": _issue_counts(quality_rows),
        "answer_columns_count_as_evidence": False,
        "quantity_status": "deferred_until_three_fields_accepted",
    }
    return {
        "ok": True,
        "phase": PHASE,
        "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "summary": summary,
        "quality_rows": quality_rows,
        "filtered_external_results": {"evidence_rows": filtered_rows},
    }


def filter_external_results_by_quality(
    external_results: Mapping[str, Any],
    *,
    include_review: bool = False,
) -> dict[str, Any]:
    report = build_external_evidence_quality_report(external_results, include_review=include_review)
    return dict(report.get("filtered_external_results") or {"evidence_rows": []})


def write_external_evidence_quality_outputs(
    report: Mapping[str, Any],
    output_dir: str | Path,
    *,
    stem: str | None = None,
) -> dict[str, str]:
    target = Path(output_dir)
    target.mkdir(parents=True, exist_ok=True)
    file_stem = stem or f"BIZ2x_PDF_external_evidence_quality_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
    json_path = target / f"{file_stem}.json"
    csv_path = target / f"{file_stem}.csv"
    markdown_path = target / f"{file_stem}.md"
    xlsx_path = target / f"{file_stem}.xlsx"
    outputs = {
        "json": str(json_path),
        "csv": str(csv_path),
        "markdown": str(markdown_path),
        "xlsx": str(xlsx_path),
    }
    payload = {**dict(report), "outputs": outputs}
    json_path.write_text(json.dumps(_json_safe(payload), ensure_ascii=False, indent=2), encoding="utf-8")
    _write_csv(csv_path, payload.get("quality_rows") or [], QUALITY_HEADERS)
    markdown_path.write_text(_build_markdown(payload), encoding="utf-8")
    _write_workbook(xlsx_path, payload)
    return outputs


def _score_source_row(row: Mapping[str, Any], *, row_no: int, include_review: bool) -> dict[str, Any]:
    item_hint = _first(row, "item_hint", "evidence_item_hint", "raw_item_name", "item_name", "项目名称", "清单项目名称")
    spec = _first(row, "spec_or_method", "evidence_spec_or_method", "feature", "项目特征", "规格/做法", "spec")
    unit = _first(row, "suggested_unit", "evidence_suggested_unit", "unit", "单位", "计量单位")
    text = _first(row, "text", "evidence_text", "normalized_text", "识别依据", "reason")
    normalized_text = _first(row, "normalized_text") or text
    confidence = _bounded_float(_first(row, "confidence", "置信度"), default=0.5)

    issues: list[str] = []
    score = 0.0

    item_generic = _is_generic_item_hint(item_hint)
    hard_generic_item = _is_hard_generic_item_hint(item_hint)
    item_concrete = _is_concrete_hint(item_hint) and not item_generic
    spec_generic = _is_generic_text(spec)
    spec_concrete = _is_concrete_hint(spec) and not spec_generic
    text_generic = _is_generic_text(text)
    text_concrete = _is_concrete_text(text) and not text_generic
    quantity_estimation_text = _mentions_quantity_estimation(text) or _mentions_quantity_estimation(_first(row, "reason"))
    manual_review_requested = _truthy(_first(row, "needs_manual_review", "需要人工复核"))
    uncertain_evidence_text = _mentions_uncertain_evidence(
        item_hint,
        spec,
        text,
        normalized_text,
        _first(row, "reason", "识别理由"),
    )
    unit_valid = _valid_unit(unit)
    trace_present = bool(_first(row, "source_file", "candidate_source_files", "PDF文件") and _first(row, "page", "evidence_pages", "页码"))
    task_trace_present = bool(_first(row, "task_no", "task_nos", "covered_task_nos"))

    if item_concrete:
        score += 2.0
    elif item_hint:
        issues.append("generic_or_weak_item_hint")
        if hard_generic_item:
            issues.append("generic_section_item_hint")
        score -= 1.5 if item_generic else 0.5
    else:
        issues.append("missing_item_hint")

    if spec_concrete:
        score += 1.5
    elif spec:
        issues.append("generic_or_weak_spec")
        score -= 0.75 if spec_generic else 0.25
    else:
        issues.append("missing_spec_or_method")

    if text_concrete:
        score += 1.0
    elif text:
        issues.append("generic_or_weak_text")
        score -= 0.75 if text_generic else 0.25
    else:
        issues.append("missing_text")

    if unit_valid:
        score += 0.75
    elif unit:
        issues.append("unknown_or_invalid_unit")
        score -= 1.0
    else:
        issues.append("missing_unit")

    if confidence >= 0.75:
        score += 0.5
    elif confidence < 0.35:
        issues.append("low_confidence")
        score -= 0.5

    if trace_present:
        score += 0.5
    else:
        issues.append("missing_source_trace")

    if task_trace_present:
        score += 0.25

    if quantity_estimation_text:
        issues.append("quantity_estimation_text")
        score -= 2.0

    if manual_review_requested:
        issues.append("manual_review_requested")
        score -= 1.0

    if uncertain_evidence_text:
        issues.append("uncertain_or_incomplete_evidence")
        score -= 1.5

    if not any((item_hint, spec, text)):
        quality_status = "rejected"
        issues.append("empty_evidence")
    elif not item_hint and not spec:
        quality_status = "rejected"
    elif hard_generic_item:
        quality_status = "rejected"
    elif quantity_estimation_text:
        quality_status = "rejected"
    elif manual_review_requested or uncertain_evidence_text:
        quality_status = "review" if score >= 1.0 and (item_concrete or spec_concrete or text_concrete) else "rejected"
    elif unit_valid and text_concrete and (item_concrete or spec_concrete) and score >= 3.0:
        quality_status = "accepted"
    elif score >= 1.0 and (item_concrete or spec_concrete or text_concrete):
        quality_status = "review"
    else:
        quality_status = "rejected"

    issue_codes = ",".join(_dedupe(issues))
    import_action = (
        "import"
        if _is_importable_by_quality(
            {"quality_status": quality_status, "issue_codes": issue_codes},
            include_review=include_review,
        )
        else "drop"
    )
    return {
        "row_no": row_no,
        "quality_status": quality_status,
        "quality_score": round(score, 2),
        "import_action": import_action,
        "issue_codes": issue_codes,
        "source_kind": _first(row, "source_kind"),
        "task_no": _first(row, "task_no"),
        "task_nos": _first(row, "task_nos", "covered_task_nos"),
        "source_file": _first(row, "source_file", "candidate_source_files", "PDF文件"),
        "page": _first(row, "page", "evidence_pages", "页码"),
        "tile_id": _first(row, "tile_id", "source_tile_id", "evidence_tiles"),
        "vision_pass": _first(row, "vision_pass", "recommended_pass", "prompt_mode"),
        "evidence_role": _first(row, "evidence_role", "role"),
        "discipline": _first(row, "discipline", "专业"),
        "item_hint": item_hint,
        "space": _first(row, "space", "部位", "空间/部位"),
        "material_codes": _cell_value(row.get("material_codes") or row.get("材料编号")),
        "spec_or_method": spec,
        "suggested_unit": unit,
        "text": text,
        "normalized_text": normalized_text,
        "confidence": confidence,
        "model": _first(row, "model"),
        "needs_manual_review": _first(row, "needs_manual_review", "需要人工复核"),
        "reason": _first(row, "reason", "识别理由"),
    }


def _is_importable_by_quality(row: Mapping[str, Any], *, include_review: bool) -> bool:
    status = str(row.get("quality_status") or "")
    if status == "accepted":
        return True
    if not include_review or status != "review":
        return False
    hard_drop_issues = {
        "unknown_or_invalid_unit",
        "generic_section_item_hint",
        "empty_evidence",
        "uncertain_or_incomplete_evidence",
    }
    issues = {issue for issue in str(row.get("issue_codes") or "").split(",") if issue}
    return not bool(issues & hard_drop_issues)


def _external_source_rows(external_results: Mapping[str, Any]) -> list[dict[str, Any]]:
    for key in ("evidence_rows", "recall_evidence", "rows"):
        rows = external_results.get(key)
        if isinstance(rows, list):
            return [dict(row) for row in rows if isinstance(row, Mapping)]
    visual_report = external_results.get("visual_evidence_report")
    if isinstance(visual_report, Mapping) and isinstance(visual_report.get("evidence_rows"), list):
        return [dict(row) for row in visual_report.get("evidence_rows") or [] if isinstance(row, Mapping)]

    source_rows: list[dict[str, Any]] = []
    call_groups = []
    for key in ("call_results", "calls", "results"):
        value = external_results.get(key)
        if isinstance(value, list):
            call_groups.extend(row for row in value if isinstance(row, Mapping))
    if not call_groups and isinstance(external_results.get("evidence_items"), list):
        call_groups.append(external_results)

    for call in call_groups:
        items = (
            call.get("evidence_rows")
            or call.get("evidence_items")
            or call.get("items")
            or call.get("drawing_items")
            or []
        )
        if not isinstance(items, list):
            continue
        call_meta = {
            "call_no": call.get("call_no"),
            "source_file": _first(call, "source_file", "PDF文件"),
            "page": _first(call, "page", "页码"),
            "tile_id": call.get("tile_id"),
            "vision_pass": _first(call, "vision_pass", "recommended_pass", "prompt_mode"),
            "model": call.get("model"),
        }
        for item in items:
            if isinstance(item, Mapping):
                source_rows.append({**call_meta, **dict(item)})
    return source_rows


def _clean_source_row(row: Mapping[str, Any]) -> dict[str, Any]:
    cleaned = dict(row)
    for key in ("item_hint", "evidence_item_hint", "raw_item_name", "item_name", "项目名称", "清单项目名称"):
        if key in cleaned:
            cleaned[key] = _clean_placeholder_text(cleaned.get(key))
    for key in ("spec_or_method", "evidence_spec_or_method", "feature", "项目特征", "规格/做法", "spec"):
        if key in cleaned:
            cleaned[key] = _clean_placeholder_text(cleaned.get(key))
    for key in ("text", "evidence_text", "normalized_text", "识别依据", "reason"):
        if key in cleaned:
            cleaned[key] = _clean_placeholder_text(cleaned.get(key))
    for key in ("suggested_unit", "evidence_suggested_unit", "unit", "单位", "计量单位"):
        if key in cleaned:
            cleaned[key] = _clean_unit_text(cleaned.get(key))
    return cleaned


def _is_hard_generic_item_hint(value: Any) -> bool:
    text = _norm(value)
    return bool(text and text in HARD_GENERIC_ITEM_HINTS)


def _is_generic_item_hint(value: Any) -> bool:
    text = _norm(value)
    if not text:
        return False
    if text in GENERIC_ITEM_HINTS:
        return True
    return _is_generic_text(text)


def _is_generic_text(value: Any) -> bool:
    text = _norm(value)
    if not text:
        return False
    if text in GENERIC_ITEM_HINTS:
        return True
    return any(re.search(pattern, text, flags=re.IGNORECASE) for pattern in GENERIC_TEXT_PATTERNS)


def _mentions_quantity_estimation(value: Any) -> bool:
    text = str(value or "").strip()
    if not text:
        return False
    patterns = [
        r"(通过|根据).{0,8}(测量|计算).{0,16}(面积|数量|工程量|总面积)",
        r"(确定|计算).{0,8}(面积|数量|工程量|总面积)",
        r"材料的数量",
        r"实际面积",
        r"总面积",
    ]
    return any(re.search(pattern, text) for pattern in patterns)


def _mentions_uncertain_evidence(*values: Any) -> bool:
    text = " ".join(str(value or "") for value in values if value is not None).strip()
    if not text:
        return False
    return any(re.search(pattern, text, flags=re.IGNORECASE) for pattern in UNCERTAIN_EVIDENCE_PATTERNS)


def _truthy(value: Any) -> bool:
    text = str(value or "").strip().lower()
    if not text:
        return False
    return text in {"1", "true", "yes", "y", "on", "是", "需要", "需"}


def _is_concrete_hint(value: Any) -> bool:
    text = str(value or "").strip()
    if not text:
        return False
    compact = _norm(text)
    if compact in GENERIC_ITEM_HINTS:
        return False
    if _contains_short_concrete_term(text):
        return True
    if any(re.search(pattern, text, flags=re.IGNORECASE) for pattern in CONCRETE_HINT_PATTERNS):
        return True
    if re.search(r"[\u4e00-\u9fff]", text) and len(re.sub(r"\s+", "", text)) >= 3:
        return True
    words = re.findall(r"[A-Za-z0-9]+", text)
    return len(words) >= 3 and not _is_generic_text(text)


def _contains_short_concrete_term(value: str) -> bool:
    text = str(value or "").strip()
    if not text:
        return False
    return any(term in text for term in SHORT_CONCRETE_TERMS)


def _is_concrete_text(value: Any) -> bool:
    text = str(value or "").strip()
    if not text:
        return False
    if _is_generic_text(text):
        return False
    if _is_concrete_hint(text):
        return True
    return len(text) >= 12


def _valid_unit(value: Any) -> bool:
    text = _norm(value)
    return bool(text and text not in UNKNOWN_UNITS and text in VALID_UNITS)


def _issue_counts(rows: Sequence[Mapping[str, Any]]) -> dict[str, int]:
    counter: Counter[str] = Counter()
    for row in rows:
        for issue in str(row.get("issue_codes") or "").split(","):
            if issue:
                counter[issue] += 1
    return dict(counter)


def _build_markdown(report: Mapping[str, Any]) -> str:
    summary = report.get("summary") or {}
    lines = [
        "# BIZ-2x PDF External Evidence Quality",
        "",
        f"- generated_at: {report.get('generated_at', '-')}",
        f"- input_row_count: {summary.get('input_row_count', 0)}",
        f"- accepted_row_count: {summary.get('accepted_row_count', 0)}",
        f"- review_row_count: {summary.get('review_row_count', 0)}",
        f"- rejected_row_count: {summary.get('rejected_row_count', 0)}",
        f"- filtered_importable_row_count: {summary.get('filtered_importable_row_count', 0)}",
        f"- include_review: {summary.get('include_review', False)}",
        "",
        "## Row Detail",
        "",
        "| row | status | score | action | source | page | tile | pass | item | unit | issues |",
        "| ---: | --- | ---: | --- | --- | --- | --- | --- | --- | --- | --- |",
    ]
    for row in (report.get("quality_rows") or [])[:160]:
        lines.append(
            "| "
            + " | ".join(
                [
                    _md(row.get("row_no")),
                    _md(row.get("quality_status")),
                    _md(row.get("quality_score")),
                    _md(row.get("import_action")),
                    _md(row.get("source_file")),
                    _md(row.get("page")),
                    _md(row.get("tile_id")),
                    _md(row.get("vision_pass")),
                    _md(row.get("item_hint")),
                    _md(row.get("suggested_unit")),
                    _md(row.get("issue_codes")),
                ]
            )
            + " |"
        )
    return "\n".join(lines) + "\n"


def _write_workbook(path: Path, report: Mapping[str, Any]) -> None:
    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "quality_summary"
    _append_rows(summary_sheet, [SUMMARY_HEADERS])
    _append_rows(summary_sheet, [[key, _cell_value(value)] for key, value in (report.get("summary") or {}).items()])
    _style_sheet(summary_sheet)

    detail_sheet = workbook.create_sheet("quality_detail")
    _append_rows(detail_sheet, [QUALITY_HEADERS])
    _append_rows(
        detail_sheet,
        [[_cell_value(row.get(header)) for header in QUALITY_HEADERS] for row in report.get("quality_rows") or []],
    )
    _style_sheet(detail_sheet)
    workbook.save(path)


def _write_csv(path: Path, rows: Iterable[Mapping[str, Any]], fieldnames: Sequence[str]) -> None:
    with path.open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        for row in rows:
            writer.writerow({field: _cell_value(row.get(field)) for field in fieldnames})


def _append_rows(sheet: Any, rows: Sequence[Sequence[Any]]) -> None:
    for row in rows:
        sheet.append(list(row))


def _style_sheet(sheet: Any) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    if sheet.max_row >= 1:
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for col_idx, column_cells in enumerate(sheet.columns, start=1):
        values = [str(cell.value or "") for cell in column_cells[:200]]
        width = min(max([len(value) for value in values] + [10]) + 2, 70)
        sheet.column_dimensions[get_column_letter(col_idx)].width = width


def _first(row: Mapping[str, Any], *keys: str) -> str:
    for key in keys:
        value = row.get(key)
        if value is None:
            continue
        text = str(value).strip()
        if text:
            return text
    return ""


def _clean_placeholder_text(value: Any) -> str:
    text = str(value or "").strip()
    if not text:
        return ""
    normalized = text.replace("；", ";")
    placeholder_norms = {item.replace("；", ";") for item in PROMPT_PLACEHOLDER_TEXTS}
    return "" if normalized in placeholder_norms else text


def _clean_unit_text(value: Any) -> str:
    text = str(value or "").strip()
    if not text:
        return ""
    if "不要写数量" in text or "不确定则为空" in text or "不确定留空" in text:
        return ""
    if _looks_like_unit_option_prompt(text):
        return ""
    return text


def _looks_like_unit_option_prompt(value: str) -> bool:
    text = str(value or "").strip()
    if not text:
        return False
    compact = re.sub(r"\s+", "", text.lower())
    slash_count = compact.count("/") + compact.count("／")
    if slash_count < 2:
        return False
    unit_hits = sum(compact.count(token) for token in ("㎡", "m2", "m²", "m³", "m3", "m", "套", "个", "樘", "项", "?"))
    return unit_hits >= 3


def _bounded_float(value: Any, *, default: float = 0.0) -> float:
    try:
        parsed = float(str(value).strip())
    except (TypeError, ValueError):
        parsed = default
    return max(0.0, min(1.0, parsed))


def _dedupe(items: Iterable[str]) -> list[str]:
    seen: set[str] = set()
    result: list[str] = []
    for item in items:
        if not item or item in seen:
            continue
        seen.add(item)
        result.append(item)
    return result


def _norm(value: Any) -> str:
    text = str(value or "").strip().lower()
    text = text.replace("；", ";").replace("，", ",")
    text = re.sub(r"\s+", " ", text)
    return text


def _cell_value(value: Any) -> Any:
    if isinstance(value, (dict, list, tuple)):
        return json.dumps(value, ensure_ascii=False)
    return value


def _json_safe(value: Any) -> Any:
    if isinstance(value, Mapping):
        return {str(key): _json_safe(item) for key, item in value.items()}
    if isinstance(value, (list, tuple)):
        return [_json_safe(item) for item in value]
    if isinstance(value, Path):
        return str(value)
    return value


def _md(value: Any) -> str:
    return str(value or "").replace("|", "\\|").replace("\n", "<br>")
