from __future__ import annotations

import hashlib
import json
import re
import uuid
from datetime import datetime
from io import BytesIO
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from sqlalchemy.orm import Session

from app.models.cost_measurement import (
    LINE_TYPE_ITEM,
    LINE_TYPE_MEASURE,
    MEASUREMENT_STATUS_DRAFT,
    PRICING_MODE_BREAKDOWN,
    PRICING_MODE_COMPOSITE,
    CostMeasurement,
    CostMeasurementEvent,
    CostMeasurementLine,
)
from app.models.enterprise_quota import EnterpriseQuotaItem, EnterpriseQuotaVersion


PARSER_VERSION = "cost_measurement_xlsx_v1"
MAX_IMPORT_BYTES = 30 * 1024 * 1024
SUPPORTED_SUFFIXES = {".xlsx", ".xlsm"}
DEFAULT_MANAGEMENT_RATE = 0.03
DEFAULT_PROFIT_RATE = 0.05
DEFAULT_TAX_RATE = 0.09
REVIEW_TOLERANCE = 0.05


class CostMeasurementImportError(ValueError):
    pass


def clean_text(value: Any, limit: int | None = None) -> str | None:
    if value is None:
        return None
    text = str(value).replace("\u3000", " ").strip()
    if not text:
        return None
    return text[:limit] if limit else text


def to_float(value: Any) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace(",", "")
    if not text:
        return None
    try:
        return float(text[:-1]) / 100 if text.endswith("%") else float(text)
    except ValueError:
        return None


def normalize_loss_rate(value: Any) -> float:
    number = to_float(value)
    if number is None or number <= 0:
        return 0.0
    return round(max(0.0, number - 1) if number >= 1 else number, 10)


def normalize_match_text(value: Any) -> str:
    text = clean_text(value) or ""
    return re.sub(r"[\W_]+", "", text).lower()


def _header_text(value: Any) -> str:
    return re.sub(r"\s+", "", clean_text(value) or "")


def _find_header(ws) -> tuple[int, dict[str, int]] | None:
    for row_index in range(1, min(ws.max_row, 40) + 1):
        mapping: dict[str, int] = {}
        for column in range(1, min(ws.max_column, 30) + 1):
            text = _header_text(ws.cell(row_index, column).value)
            if "\u9879\u76ee\u540d\u79f0" in text:
                mapping.setdefault("item_name", column)
            elif "\u9879\u76ee\u7279\u5f81" in text:
                mapping.setdefault("feature", column)
            elif text in {"\u5355\u4f4d", "\u8ba1\u91cf\u5355\u4f4d"}:
                mapping.setdefault("unit", column)
            elif "\u5de5\u7a0b\u91cf" in text:
                mapping.setdefault("quantity", column)
            elif text == "\u5e8f\u53f7":
                mapping.setdefault("sequence_no", column)
            elif "\u7efc\u5408\u5355\u4ef7" in text:
                mapping.setdefault("source_unit_price", column)
            elif "\u5408\u8ba1" in text or "\u603b\u4ef7" in text:
                mapping.setdefault("source_total_price", column)
        if not {"item_name", "unit", "quantity"}.issubset(mapping):
            continue
        for label_row in range(row_index, min(ws.max_row, row_index + 3) + 1):
            for column in range(1, min(ws.max_column, 30) + 1):
                text = _header_text(ws.cell(label_row, column).value)
                if "\u4eba\u5de5\u8d39" in text:
                    mapping.setdefault("labor_unit_price", column)
                elif "\u4e3b\u6750\u8d39" in text:
                    mapping.setdefault("main_material_unit_price", column)
                elif "\u635f\u8017" in text:
                    mapping.setdefault("material_loss_rate", column)
                elif "\u8f85\u6750" in text or "\u673a\u68b0\u8d39" in text:
                    mapping.setdefault("auxiliary_machinery_unit_price", column)
                elif "\u4e13\u4e1a\u5206\u5305" in text or "\u5206\u5305\u8d39" in text:
                    mapping.setdefault("subcontract_unit_price", column)
                elif "\u7ba1\u7406\u8d39" in text:
                    mapping.setdefault("management_rate", column)
                elif text == "\u5229\u6da6" or "\u5229\u6da6\u7387" in text:
                    mapping.setdefault("profit_rate", column)
                elif "\u7efc\u5408\u5355\u4ef7" in text:
                    mapping.setdefault("source_unit_price", column)
                elif text.startswith("\u5408\u8ba1") or "\u603b\u4ef7" in text:
                    mapping.setdefault("source_total_price", column)
        return row_index, mapping
    return None


def _rate_below_header(ws, header_row: int, column: int | None, item_column: int, unit_column: int, default: float) -> float:
    if not column:
        return default
    for row_index in range(header_row + 1, min(ws.max_row, header_row + 15) + 1):
        if clean_text(ws.cell(row_index, item_column).value) and clean_text(ws.cell(row_index, unit_column).value):
            continue
        value = to_float(ws.cell(row_index, column).value)
        if value is not None and 0 <= value < 1:
            return value
    return default


def _project_name(workbook, fallback: str) -> str:
    for ws in workbook.worksheets:
        for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 12), min_col=1, max_col=min(ws.max_column, 8)):
            for cell in row:
                text = clean_text(cell.value)
                if text and "\u5de5\u7a0b\u540d\u79f0" in text:
                    parts = re.split(r"[:\uFF1A]", text, maxsplit=1)
                    if len(parts) == 2 and parts[1].strip():
                        return parts[1].strip()[:255]
    return fallback[:255]


def _infer_tax_rate(workbook) -> float:
    for ws in workbook.worksheets:
        for row_index in range(1, min(ws.max_row, 80) + 1):
            values = [ws.cell(row_index, column).value for column in range(1, min(ws.max_column, 12) + 1)]
            if not any("\u7a0e\u91d1" in (clean_text(value) or "") for value in values):
                continue
            current = next((to_float(value) for value in reversed(values) if to_float(value)), None)
            if current is None:
                continue
            for previous_index in range(row_index - 1, max(0, row_index - 4), -1):
                previous_values = [ws.cell(previous_index, column).value for column in range(1, min(ws.max_column, 12) + 1)]
                previous = next((to_float(value) for value in reversed(previous_values) if to_float(value)), None)
                if previous:
                    ratio = current / previous
                    if 0 < ratio < 0.3:
                        return ratio
    return DEFAULT_TAX_RATE


def calculate_values(
    *,
    quantity: float,
    pricing_mode: str,
    source_unit_price: float | None,
    labor_unit_price: float,
    main_material_unit_price: float,
    material_loss_rate: float,
    auxiliary_machinery_unit_price: float,
    subcontract_unit_price: float,
    management_rate: float,
    profit_rate: float,
) -> dict[str, float]:
    quantity = max(0.0, float(quantity or 0))
    if pricing_mode == PRICING_MODE_COMPOSITE:
        direct = float(subcontract_unit_price or source_unit_price or 0)
        management = profit = 0.0
        unit_price = direct
    else:
        direct = (
            float(labor_unit_price or 0)
            + float(main_material_unit_price or 0) * (1 + max(0.0, float(material_loss_rate or 0)))
            + float(auxiliary_machinery_unit_price or 0)
            + float(subcontract_unit_price or 0)
        )
        management = direct * max(0.0, float(management_rate or 0))
        profit = (direct + management) * max(0.0, float(profit_rate or 0))
        unit_price = direct + management + profit
    return {
        "direct_unit_price": direct,
        "management_unit_price": management,
        "profit_unit_price": profit,
        "calculated_unit_price": unit_price,
        "calculated_total_price": unit_price * quantity,
        "source_variance": unit_price - float(source_unit_price or 0),
    }


def _line_warnings(line: dict[str, Any]) -> list[dict[str, Any]]:
    warnings: list[dict[str, Any]] = []
    if float(line.get("quantity") or 0) <= 0:
        warnings.append({"code": "QUANTITY_MISSING", "message": "\u5de5\u7a0b\u91cf\u4e3a\u7a7a\u6216\u4e3a 0\uff0c\u9700\u8981\u4eba\u5de5\u786e\u8ba4"})
    if line.get("source_unit_price") is None:
        warnings.append({"code": "SOURCE_UNIT_PRICE_MISSING", "message": "\u5386\u53f2\u7efc\u5408\u5355\u4ef7\u4e3a\u7a7a"})
    if line.get("pricing_mode") == PRICING_MODE_COMPOSITE:
        warnings.append({"code": "COMPOSITE_ONLY", "message": "\u4ec5\u6709\u7efc\u5408\u4ef7\uff0c\u5c1a\u672a\u62c6\u5206\u4eba\u5de5\u3001\u6750\u6599\u6216\u5206\u5305"})
    source_price = line.get("source_unit_price")
    if source_price is not None and line.get("pricing_mode") == PRICING_MODE_BREAKDOWN:
        difference = abs(float(line.get("calculated_unit_price") or 0) - float(source_price))
        if difference > max(REVIEW_TOLERANCE, abs(float(source_price)) * 0.001):
            warnings.append({"code": "SOURCE_FORMULA_VARIANCE", "message": "\u7edf\u4e00\u91cd\u7b97\u5355\u4ef7\u4e0e\u5386\u53f2 Excel \u5355\u4ef7\u4e0d\u4e00\u81f4", "difference": round(difference, 6)})
    return warnings


def parse_cost_measurement_workbook(filename: str, content: bytes) -> dict[str, Any]:
    if Path(filename or "").suffix.lower() not in SUPPORTED_SUFFIXES:
        raise CostMeasurementImportError("\u4ec5\u652f\u6301 .xlsx/.xlsm\uff0c\u8bf7\u5148\u5c06\u65e7 .xls \u53e6\u5b58\u4e3a .xlsx")
    if not content:
        raise CostMeasurementImportError("\u4e0a\u4f20\u6587\u4ef6\u4e3a\u7a7a")
    if len(content) > MAX_IMPORT_BYTES:
        raise CostMeasurementImportError("\u6587\u4ef6\u8d85\u8fc7 30MB \u9650\u5236")
    try:
        workbook = load_workbook(BytesIO(content), data_only=False)
        values_workbook = load_workbook(BytesIO(content), data_only=True)
    except Exception as exc:
        raise CostMeasurementImportError(f"\u65e0\u6cd5\u8bfb\u53d6 Excel\uff1a{exc}") from exc

    project_name = _project_name(values_workbook, Path(filename).stem)
    management_rate = DEFAULT_MANAGEMENT_RATE
    profit_rate = DEFAULT_PROFIT_RATE
    tax_rate = _infer_tax_rate(values_workbook)
    parsed_lines: list[dict[str, Any]] = []
    sheet_summaries: list[dict[str, Any]] = []
    sort_order = 0

    for ws in workbook.worksheets:
        values_ws = values_workbook[ws.title]
        header = _find_header(values_ws)
        if not header:
            continue
        header_row, mapping = header
        sheet_management = _rate_below_header(values_ws, header_row, mapping.get("management_rate"), mapping["item_name"], mapping["unit"], management_rate)
        sheet_profit = _rate_below_header(values_ws, header_row, mapping.get("profit_rate"), mapping["item_name"], mapping["unit"], profit_rate)
        management_rate, profit_rate = sheet_management, sheet_profit
        current_section: str | None = None
        sheet_line_count = 0
        for row_index in range(header_row + 1, values_ws.max_row + 1):
            item_name = clean_text(values_ws.cell(row_index, mapping["item_name"]).value, 255)
            unit = clean_text(values_ws.cell(row_index, mapping["unit"]).value, 64)
            if item_name and not unit:
                if not any(word in item_name for word in ("\u5408\u8ba1", "\u5c0f\u8ba1", "\u603b\u8ba1")):
                    current_section = item_name
                continue
            if not item_name or not unit or any(word in item_name for word in ("\u5408\u8ba1", "\u5c0f\u8ba1", "\u603b\u8ba1")):
                continue
            quantity = to_float(values_ws.cell(row_index, mapping["quantity"]).value) or 0.0
            component_values = {}
            for key in ("labor_unit_price", "main_material_unit_price", "material_loss_rate", "auxiliary_machinery_unit_price", "subcontract_unit_price", "source_unit_price", "source_total_price"):
                column = mapping.get(key)
                component_values[key] = to_float(values_ws.cell(row_index, column).value) if column else None
            has_breakdown = any(
                mapping.get(key) and values_ws.cell(row_index, mapping[key]).value is not None
                for key in ("labor_unit_price", "main_material_unit_price", "auxiliary_machinery_unit_price", "subcontract_unit_price")
            )
            pricing_mode = PRICING_MODE_BREAKDOWN if has_breakdown else PRICING_MODE_COMPOSITE
            source_unit = component_values["source_unit_price"]
            subcontract = component_values["subcontract_unit_price"] or 0.0
            if pricing_mode == PRICING_MODE_COMPOSITE:
                subcontract = source_unit or 0.0
            values = calculate_values(
                quantity=quantity,
                pricing_mode=pricing_mode,
                source_unit_price=source_unit,
                labor_unit_price=component_values["labor_unit_price"] or 0.0,
                main_material_unit_price=component_values["main_material_unit_price"] or 0.0,
                material_loss_rate=normalize_loss_rate(component_values["material_loss_rate"]),
                auxiliary_machinery_unit_price=component_values["auxiliary_machinery_unit_price"] or 0.0,
                subcontract_unit_price=subcontract,
                management_rate=sheet_management,
                profit_rate=sheet_profit,
            )
            sort_order += 1
            sequence_column = mapping.get("sequence_no")
            feature_column = mapping.get("feature")
            line = {
                "line_key": f"{ws.title}:{row_index}",
                "source_sheet": ws.title,
                "source_row_index": row_index,
                "sort_order": sort_order,
                "sequence_no": clean_text(values_ws.cell(row_index, sequence_column).value, 64) if sequence_column else None,
                "section_name": current_section,
                "item_name": item_name,
                "feature": clean_text(values_ws.cell(row_index, feature_column).value) if feature_column else None,
                "unit": unit,
                "quantity": quantity,
                "line_type": LINE_TYPE_MEASURE if "\u63aa\u65bd" in (current_section or "") or "\u63aa\u65bd" in item_name else LINE_TYPE_ITEM,
                "pricing_mode": pricing_mode,
                "price_source": "historical_excel",
                "source_unit_price": source_unit,
                "source_total_price": component_values["source_total_price"],
                "labor_unit_price": component_values["labor_unit_price"] or 0.0,
                "main_material_unit_price": component_values["main_material_unit_price"] or 0.0,
                "material_loss_rate": normalize_loss_rate(component_values["material_loss_rate"]),
                "auxiliary_machinery_unit_price": component_values["auxiliary_machinery_unit_price"] or 0.0,
                "subcontract_unit_price": subcontract,
                **values,
                "source_row": [values_ws.cell(row_index, column).value for column in range(1, min(values_ws.max_column, 30) + 1)],
            }
            line["warnings"] = _line_warnings(line)
            line["review_status"] = "required" if line["warnings"] else "ready"
            parsed_lines.append(line)
            sheet_line_count += 1
        sheet_summaries.append({"sheet_name": ws.title, "line_count": sheet_line_count})

    if not parsed_lines:
        raise CostMeasurementImportError("\u672a\u8bc6\u522b\u5230\u542b\u9879\u76ee\u540d\u79f0\u3001\u5355\u4f4d\u548c\u5de5\u7a0b\u91cf\u7684\u6709\u6548\u6e05\u5355\u884c")
    source_total = sum(float(line.get("source_total_price") or 0) for line in parsed_lines)
    calculated_total = sum(float(line.get("calculated_total_price") or 0) for line in parsed_lines)
    return {
        "parser_version": PARSER_VERSION,
        "filename": filename,
        "file_sha256": hashlib.sha256(content).hexdigest(),
        "project_name": project_name,
        "management_rate": management_rate,
        "profit_rate": profit_rate,
        "tax_rate": tax_rate,
        "line_count": len(parsed_lines),
        "review_line_count": sum(1 for line in parsed_lines if line["review_status"] == "required"),
        "sheet_summaries": sheet_summaries,
        "source_pretax_total": source_total,
        "calculated_pretax_total": calculated_total,
        "source_variance": calculated_total - source_total,
        "lines": parsed_lines,
    }


def next_measurement_code(db: Session) -> str:
    prefix = f"CM-{datetime.now().strftime('%Y%m%d')}"
    count = db.query(CostMeasurement).filter(CostMeasurement.measurement_code.like(f"{prefix}-%")).count() + 1
    return f"{prefix}-{count:03d}"


def active_quota_match_map(db: Session) -> tuple[EnterpriseQuotaVersion | None, dict[tuple[str, str], EnterpriseQuotaItem]]:
    version = db.query(EnterpriseQuotaVersion).filter(EnterpriseQuotaVersion.is_active.is_(True)).order_by(EnterpriseQuotaVersion.id.desc()).first()
    if not version:
        return None, {}
    mapping = {}
    for item in version.items:
        key = (normalize_match_text(item.item_name), normalize_match_text(item.unit))
        if key[0] and key not in mapping:
            mapping[key] = item
    return version, mapping


def serialize_measurement_line(line: CostMeasurementLine, *, include_warnings: bool = True) -> dict[str, Any]:
    keys = (
        "id", "measurement_id", "line_key", "quota_item_id", "source_sheet", "source_row_index", "sort_order",
        "sequence_no", "section_name", "item_name", "feature", "unit", "quantity", "line_type", "pricing_mode",
        "price_source", "source_unit_price", "source_total_price", "labor_unit_price", "main_material_unit_price",
        "material_loss_rate", "auxiliary_machinery_unit_price", "subcontract_unit_price", "direct_unit_price",
        "management_unit_price", "profit_unit_price", "calculated_unit_price", "calculated_total_price",
        "source_variance", "review_status",
    )
    data = {key: getattr(line, key) for key in keys}
    if include_warnings:
        data["warnings"] = _decode_json(line.warnings_json, [])
    return data


def recalculate_line(line: CostMeasurementLine, measurement: CostMeasurement) -> CostMeasurementLine:
    values = calculate_values(
        quantity=line.quantity,
        pricing_mode=line.pricing_mode,
        source_unit_price=line.source_unit_price,
        labor_unit_price=line.labor_unit_price,
        main_material_unit_price=line.main_material_unit_price,
        material_loss_rate=line.material_loss_rate,
        auxiliary_machinery_unit_price=line.auxiliary_machinery_unit_price,
        subcontract_unit_price=line.subcontract_unit_price,
        management_rate=measurement.management_rate,
        profit_rate=measurement.profit_rate,
    )
    for key, value in values.items():
        setattr(line, key, value)
    warnings = _line_warnings(serialize_measurement_line(line, include_warnings=False))
    line.warnings_json = json.dumps(warnings, ensure_ascii=False) if warnings else None
    if warnings and line.review_status not in {"reviewed", "accepted"}:
        line.review_status = "required"
    elif not warnings and line.review_status == "required":
        line.review_status = "ready"
    return line


def recalculate_measurement(db: Session, measurement: CostMeasurement) -> CostMeasurement:
    direct = management = profit = pretax = 0.0
    for line in measurement.lines:
        recalculate_line(line, measurement)
        quantity = float(line.quantity or 0)
        direct += float(line.direct_unit_price or 0) * quantity
        management += float(line.management_unit_price or 0) * quantity
        profit += float(line.profit_unit_price or 0) * quantity
        pretax += float(line.calculated_total_price or 0)
    measurement.line_count = len(measurement.lines)
    measurement.review_line_count = sum(1 for line in measurement.lines if line.review_status == "required")
    measurement.matched_quota_count = sum(1 for line in measurement.lines if line.quota_item_id)
    measurement.direct_cost = direct
    measurement.management_fee = management
    measurement.profit_fee = profit
    measurement.pretax_total = pretax
    measurement.tax_total = pretax * float(measurement.tax_rate or 0)
    measurement.grand_total = pretax + measurement.tax_total
    db.flush()
    return measurement


def write_measurement_event(db: Session, measurement: CostMeasurement, *, event_type: str, actor_user_id: int, message: str | None = None, line_id: int | None = None, payload: dict | None = None) -> CostMeasurementEvent:
    event = CostMeasurementEvent(
        measurement_id=measurement.id,
        line_id=line_id,
        event_type=event_type,
        message=clean_text(message, 2000),
        payload_json=json.dumps(payload, ensure_ascii=False, default=str) if payload else None,
        actor_user_id=actor_user_id,
    )
    db.add(event)
    return event


def create_measurement_from_import(db: Session, *, filename: str, content: bytes, name: str | None, project_name: str | None, notes: str | None, actor_user_id: int) -> CostMeasurement:
    parsed = parse_cost_measurement_workbook(filename, content)
    version, quota_map = active_quota_match_map(db)
    measurement = CostMeasurement(
        measurement_uuid=str(uuid.uuid4()),
        measurement_code=next_measurement_code(db),
        name=clean_text(name, 255) or f"{parsed['project_name']}\u6210\u672c\u6d4b\u7b97",
        project_name=clean_text(project_name, 255) or parsed["project_name"],
        status=MEASUREMENT_STATUS_DRAFT,
        source_filename=filename[:255],
        source_file_sha256=parsed["file_sha256"],
        quota_version_id=version.id if version else None,
        management_rate=parsed["management_rate"],
        profit_rate=parsed["profit_rate"],
        tax_rate=parsed["tax_rate"],
        source_summary_json=json.dumps({key: value for key, value in parsed.items() if key != "lines"}, ensure_ascii=False, default=str),
        notes=clean_text(notes),
        created_by=actor_user_id,
        updated_by=actor_user_id,
    )
    db.add(measurement)
    db.flush()
    for row in parsed["lines"]:
        quota_item = quota_map.get((normalize_match_text(row["item_name"]), normalize_match_text(row["unit"])))
        db.add(CostMeasurementLine(
            measurement_id=measurement.id,
            quota_item_id=quota_item.id if quota_item else None,
            line_key=row["line_key"],
            source_sheet=row["source_sheet"],
            source_row_index=row["source_row_index"],
            sort_order=row["sort_order"],
            sequence_no=row["sequence_no"],
            section_name=row["section_name"],
            item_name=row["item_name"],
            feature=row["feature"],
            unit=row["unit"],
            quantity=row["quantity"],
            line_type=row["line_type"],
            pricing_mode=row["pricing_mode"],
            price_source=row["price_source"],
            source_unit_price=row["source_unit_price"],
            source_total_price=row["source_total_price"],
            labor_unit_price=row["labor_unit_price"],
            main_material_unit_price=row["main_material_unit_price"],
            material_loss_rate=row["material_loss_rate"],
            auxiliary_machinery_unit_price=row["auxiliary_machinery_unit_price"],
            subcontract_unit_price=row["subcontract_unit_price"],
            direct_unit_price=row["direct_unit_price"],
            management_unit_price=row["management_unit_price"],
            profit_unit_price=row["profit_unit_price"],
            calculated_unit_price=row["calculated_unit_price"],
            calculated_total_price=row["calculated_total_price"],
            source_variance=row["source_variance"],
            review_status=row["review_status"],
            warnings_json=json.dumps(row["warnings"], ensure_ascii=False) if row["warnings"] else None,
            source_row_json=json.dumps(row["source_row"], ensure_ascii=False, default=str),
        ))
    db.flush()
    recalculate_measurement(db, measurement)
    write_measurement_event(
        db,
        measurement,
        event_type="imported",
        actor_user_id=actor_user_id,
        message=f"\u4ece {filename} \u5bfc\u5165 {measurement.line_count} \u6761\u6d4b\u7b97\u884c",
        payload={"parser_version": PARSER_VERSION, "quota_version_id": measurement.quota_version_id},
    )
    db.commit()
    db.refresh(measurement)
    return measurement


def _decode_json(value: str | None, fallback):
    if not value:
        return fallback
    try:
        return json.loads(value)
    except (TypeError, json.JSONDecodeError):
        return fallback


def format_datetime(value) -> str | None:
    return value.isoformat() if value else None


def serialize_measurement(measurement: CostMeasurement, *, include_lines: bool = False, include_events: bool = False) -> dict[str, Any]:
    keys = (
        "id", "measurement_uuid", "measurement_code", "name", "project_name", "status", "source_filename",
        "source_file_sha256", "quota_version_id", "management_rate", "profit_rate", "tax_rate", "line_count",
        "review_line_count", "matched_quota_count", "direct_cost", "management_fee", "profit_fee", "pretax_total",
        "tax_total", "grand_total", "notes", "created_by", "updated_by", "locked_by",
    )
    data = {key: getattr(measurement, key) for key in keys}
    data.update({
        "source_summary": _decode_json(measurement.source_summary_json, {}),
        "locked_at": format_datetime(measurement.locked_at),
        "created_at": format_datetime(measurement.created_at),
        "updated_at": format_datetime(measurement.updated_at),
    })
    if include_lines:
        data["lines"] = [serialize_measurement_line(line) for line in measurement.lines]
    if include_events:
        data["events"] = [{
            "id": event.id,
            "line_id": event.line_id,
            "event_type": event.event_type,
            "message": event.message,
            "payload": _decode_json(event.payload_json, {}),
            "actor_user_id": event.actor_user_id,
            "created_at": format_datetime(event.created_at),
        } for event in measurement.events]
    return data


def apply_quota_item(line: CostMeasurementLine, quota_item: EnterpriseQuotaItem, measurement: CostMeasurement) -> None:
    line.quota_item_id = quota_item.id
    line.pricing_mode = PRICING_MODE_BREAKDOWN
    line.price_source = "enterprise_quota"
    line.labor_unit_price = float(quota_item.labor_fee or 0)
    line.main_material_unit_price = float(quota_item.main_material_fee or 0)
    line.auxiliary_machinery_unit_price = float(quota_item.auxiliary_material_fee or 0) + float(quota_item.machinery_fee or 0)
    line.subcontract_unit_price = 0
    line.review_status = "ready"
    recalculate_line(line, measurement)


def build_measurement_export(measurement: CostMeasurement) -> bytes:
    workbook = Workbook()
    summary = workbook.active
    summary.title = "\u6d4b\u7b97\u6c47\u603b"
    rows = [
        ["\u6210\u672c\u6d4b\u7b97\u6c47\u603b", None],
        ["\u6d4b\u7b97\u7f16\u53f7", measurement.measurement_code],
        ["\u9879\u76ee\u540d\u79f0", measurement.project_name],
        ["\u72b6\u6001", measurement.status],
        ["\u7ba1\u7406\u8d39\u7387", measurement.management_rate],
        ["\u5229\u6da6\u7387", measurement.profit_rate],
        ["\u7a0e\u7387", measurement.tax_rate],
        ["\u76f4\u63a5\u6210\u672c", measurement.direct_cost],
        ["\u7ba1\u7406\u8d39", measurement.management_fee],
        ["\u5229\u6da6", measurement.profit_fee],
        ["\u7a0e\u524d\u5408\u8ba1", measurement.pretax_total],
        ["\u7a0e\u91d1", measurement.tax_total],
        ["\u542b\u7a0e\u5408\u8ba1", measurement.grand_total],
    ]
    for row in rows:
        summary.append(row)
    summary["A1"].font = Font(size=16, bold=True)
    summary["A1"].fill = PatternFill("solid", fgColor="DCE6F1")
    summary.column_dimensions["A"].width = 22
    summary.column_dimensions["B"].width = 24
    for row in range(5, 8):
        summary.cell(row, 2).number_format = "0.00%"
    for row in range(8, 14):
        summary.cell(row, 2).number_format = "#,##0.00"

    detail = workbook.create_sheet("\u5206\u90e8\u5206\u9879\u660e\u7ec6")
    headers = ["\u5e8f\u53f7", "\u5206\u90e8", "\u9879\u76ee\u540d\u79f0", "\u9879\u76ee\u7279\u5f81", "\u5355\u4f4d", "\u5de5\u7a0b\u91cf", "\u4eba\u5de5\u5355\u4ef7", "\u4e3b\u6750\u5355\u4ef7", "\u635f\u8017\u7387", "\u8f85\u6750\u53ca\u673a\u68b0", "\u5206\u5305\u5355\u4ef7", "\u76f4\u63a5\u8d39\u5355\u4ef7", "\u7ba1\u7406\u8d39\u5355\u4ef7", "\u5229\u6da6\u5355\u4ef7", "\u7efc\u5408\u5355\u4ef7", "\u5408\u8ba1", "\u5386\u53f2\u5355\u4ef7", "\u4ef7\u5dee", "\u4ef7\u683c\u6765\u6e90", "\u590d\u6838\u72b6\u6001", "\u4f01\u4e1a\u5b9a\u989d\u4e3b\u9879ID"]
    detail.append(headers)
    for cell in detail[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="4472C4")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for line in measurement.lines:
        detail.append([
            line.sequence_no or line.sort_order, line.section_name, line.item_name, line.feature, line.unit, line.quantity,
            line.labor_unit_price, line.main_material_unit_price, line.material_loss_rate, line.auxiliary_machinery_unit_price,
            line.subcontract_unit_price, line.direct_unit_price, line.management_unit_price, line.profit_unit_price,
            line.calculated_unit_price, line.calculated_total_price, line.source_unit_price, line.source_variance,
            line.price_source, line.review_status, line.quota_item_id,
        ])
    detail.freeze_panes = "A2"
    detail.auto_filter.ref = detail.dimensions
    widths = [10, 20, 28, 45, 9, 12, 12, 12, 10, 12, 12, 12, 12, 12, 12, 16, 12, 12, 16, 12, 16]
    for index, width in enumerate(widths, start=1):
        detail.column_dimensions[detail.cell(1, index).column_letter].width = width
    for row in range(2, detail.max_row + 1):
        detail.cell(row, 9).number_format = "0.00%"
        for column in range(6, 19):
            if column != 9:
                detail.cell(row, column).number_format = "#,##0.00"
        detail.cell(row, 4).alignment = Alignment(wrap_text=True, vertical="top")
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()
