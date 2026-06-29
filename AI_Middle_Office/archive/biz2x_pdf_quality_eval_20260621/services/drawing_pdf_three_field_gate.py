from __future__ import annotations

import csv
import json
from datetime import datetime
from pathlib import Path
from typing import Any, Iterable, Mapping, Sequence

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


PHASE = "BIZ-2x-pdf-three-field-quality-gate"

GATE_HEADERS = ["gate", "required", "actual", "status", "severity", "action"]


def build_three_field_quality_gate(
    report: Mapping[str, Any],
    *,
    min_pass_rate: float = 1.0,
    max_missing_candidate: int = 0,
    max_unit_conflict: int = 0,
    max_feature_review: int = 0,
    max_weak_match: int = 0,
) -> dict[str, Any]:
    metrics = _extract_metrics(report)
    gate_rows = [
        _gate_row(
            "answer_rows_present",
            "answer_count > 0",
            metrics["answer_count"],
            metrics["answer_count"] > 0,
            action="provide manual answer workbook",
        ),
        _gate_row(
            "candidate_rows_present",
            "candidate_count > 0",
            metrics["candidate_count"],
            metrics["candidate_count"] > 0,
            action="run PDF evidence extraction and V2 itemization",
        ),
        _gate_row(
            "three_field_pass_rate",
            f">= {min_pass_rate:.4f}",
            metrics["pass_rate"],
            metrics["pass_rate"] >= min_pass_rate,
            action="recall missing evidence and fix item/feature/unit mapping",
        ),
        _gate_row(
            "missing_candidate",
            f"<= {max_missing_candidate}",
            metrics["missing_candidate_count"],
            metrics["missing_candidate_count"] <= max_missing_candidate,
            action="run targeted visual recall for missing answer rows",
        ),
        _gate_row(
            "unit_conflict",
            f"<= {max_unit_conflict}",
            metrics["unit_conflict_count"],
            metrics["unit_conflict_count"] <= max_unit_conflict,
            action="fix unit normalization and item type rules",
        ),
        _gate_row(
            "feature_review",
            f"<= {max_feature_review}",
            metrics["feature_review_count"],
            metrics["feature_review_count"] <= max_feature_review,
            action="supplement feature evidence from schedules, legends, and node details",
        ),
        _gate_row(
            "weak_match_review",
            f"<= {max_weak_match}",
            metrics["weak_match_count"],
            metrics["weak_match_count"] <= max_weak_match,
            action="review weak item candidates before quantity stage",
        ),
    ]
    failed = [row for row in gate_rows if row["status"] != "pass"]
    can_enable_quantity = not failed
    return {
        "ok": True,
        "phase": PHASE,
        "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "status": "passed" if can_enable_quantity else "failed",
        "can_enable_quantity": can_enable_quantity,
        "summary": {
            **metrics,
            "failed_gate_count": len(failed),
            "quantity_status": "ready_after_three_field_acceptance"
            if can_enable_quantity
            else "deferred_until_three_fields_accepted",
        },
        "gate_rows": gate_rows,
    }


def write_three_field_quality_gate_outputs(
    report: Mapping[str, Any],
    output_dir: str | Path,
    *,
    stem: str | None = None,
) -> dict[str, str]:
    target = Path(output_dir)
    target.mkdir(parents=True, exist_ok=True)
    file_stem = stem or f"BIZ2x_PDF_three_field_quality_gate_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
    json_path = target / f"{file_stem}.json"
    csv_path = target / f"{file_stem}.csv"
    markdown_path = target / f"{file_stem}.md"
    xlsx_path = target / f"{file_stem}.xlsx"
    outputs = {
        "json": str(json_path),
        "csv": str(csv_path),
        "markdown": str(markdown_path),
        "xlsx": str(xlsx_path),
    }
    payload = {**dict(report), "outputs": outputs}
    json_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    _write_csv(csv_path, payload.get("gate_rows") or [], GATE_HEADERS)
    markdown_path.write_text(_build_markdown(payload), encoding="utf-8")
    _write_workbook(xlsx_path, payload)
    return outputs


def _extract_metrics(report: Mapping[str, Any]) -> dict[str, Any]:
    summary = _summary_from_report(report)
    status_counts = dict(summary.get("status_counts") or summary.get("three_field_gap_status_counts") or {})
    matched = _int(
        _first_value(
            summary,
            "matched_three_fields_count",
            "three_field_matched_count",
            "augmented_three_field_matched_count",
        ),
        0,
    )
    answer_count = _int(_first_value(summary, "answer_count", "three_field_answer_count"), 0)
    candidate_count = _int(_first_value(summary, "candidate_count", "three_field_candidate_count"), 0)
    pass_rate = _float(_first_value(summary, "three_field_pass_rate"), 0)
    if pass_rate == 0 and answer_count:
        pass_rate = round(matched / answer_count, 4)
    feature_review = _int(status_counts.get("matched_name_unit_feature_review"), 0)
    missing_candidate = _int(status_counts.get("missing_candidate"), 0)
    unit_conflict = _int(status_counts.get("unit_conflict"), 0)
    weak_match = _int(status_counts.get("weak_match_review"), 0)
    gap_count = missing_candidate + feature_review + unit_conflict + weak_match
    if not gap_count:
        gap_count = _int(_first_value(summary, "three_field_gap_count"), 0)
    return {
        "answer_count": answer_count,
        "candidate_count": candidate_count,
        "matched_three_fields_count": matched,
        "pass_rate": pass_rate,
        "gap_count": gap_count,
        "missing_candidate_count": missing_candidate,
        "unit_conflict_count": unit_conflict,
        "feature_review_count": feature_review,
        "weak_match_count": weak_match,
        "status_counts": status_counts,
    }


def _summary_from_report(report: Mapping[str, Any]) -> Mapping[str, Any]:
    if isinstance(report.get("review_summary"), Mapping):
        return report.get("review_summary") or {}
    if isinstance(report.get("augmented_summary"), Mapping):
        return report.get("augmented_summary") or {}
    if isinstance(report.get("three_field_acceptance_report"), Mapping):
        return (report.get("three_field_acceptance_report") or {}).get("summary") or {}
    if isinstance(report.get("summary"), Mapping):
        return report.get("summary") or {}
    return {}


def _gate_row(gate: str, required: str, actual: Any, passed: bool, *, action: str) -> dict[str, Any]:
    return {
        "gate": gate,
        "required": required,
        "actual": actual,
        "status": "pass" if passed else "fail",
        "severity": "blocker" if not passed else "",
        "action": "" if passed else action,
    }


def _build_markdown(report: Mapping[str, Any]) -> str:
    summary = report.get("summary") or {}
    lines = [
        "# BIZ-2x PDF Three Field Quality Gate",
        "",
        f"- generated_at: {report.get('generated_at', '-')}",
        f"- status: {report.get('status', '-')}",
        f"- can_enable_quantity: {report.get('can_enable_quantity', False)}",
        f"- pass_rate: {summary.get('pass_rate', 0)}",
        f"- quantity_status: {summary.get('quantity_status', '-')}",
        "",
        "| gate | required | actual | status | action |",
        "| --- | --- | ---: | --- | --- |",
    ]
    for row in report.get("gate_rows") or []:
        lines.append(
            "| "
            + " | ".join(
                [
                    _md(row.get("gate")),
                    _md(row.get("required")),
                    _md(row.get("actual")),
                    _md(row.get("status")),
                    _md(row.get("action")),
                ]
            )
            + " |"
        )
    return "\n".join(lines) + "\n"


def _write_workbook(path: Path, report: Mapping[str, Any]) -> None:
    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "gate_summary"
    _append_rows(summary_sheet, [["metric", "value"]])
    _append_rows(summary_sheet, [[key, _cell_value(value)] for key, value in (report.get("summary") or {}).items()])
    _append_rows(summary_sheet, [["status", report.get("status")], ["can_enable_quantity", report.get("can_enable_quantity")]])
    _style_sheet(summary_sheet)

    gate_sheet = workbook.create_sheet("gate_checks")
    _append_rows(gate_sheet, [GATE_HEADERS])
    _append_rows(gate_sheet, [[_cell_value(row.get(header)) for header in GATE_HEADERS] for row in report.get("gate_rows") or []])
    _style_sheet(gate_sheet)
    workbook.save(path)


def _write_csv(path: Path, rows: Iterable[Mapping[str, Any]], fieldnames: Sequence[str]) -> None:
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        for row in rows:
            writer.writerow({field: _cell_value(row.get(field)) for field in fieldnames})


def _append_rows(sheet: Any, rows: Sequence[Sequence[Any]]) -> None:
    for row in rows:
        sheet.append(list(row))


def _style_sheet(sheet: Any) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    if sheet.max_row >= 1:
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for col_idx, column_cells in enumerate(sheet.columns, start=1):
        values = [str(cell.value or "") for cell in column_cells[:200]]
        width = min(max([len(value) for value in values] + [10]) + 2, 60)
        sheet.column_dimensions[get_column_letter(col_idx)].width = width


def _first_value(row: Mapping[str, Any], *keys: str) -> Any:
    for key in keys:
        value = row.get(key)
        if value not in (None, ""):
            return value
    return ""


def _cell_value(value: Any) -> Any:
    if isinstance(value, (dict, list, tuple)):
        return json.dumps(value, ensure_ascii=False)
    return value


def _md(value: Any) -> str:
    return str(value or "").replace("|", "\\|").replace("\n", "<br>")


def _int(value: Any, default: int = 0) -> int:
    try:
        return int(float(str(value).strip()))
    except (TypeError, ValueError):
        return default


def _float(value: Any, default: float = 0.0) -> float:
    try:
        return float(str(value).strip())
    except (TypeError, ValueError):
        return default
