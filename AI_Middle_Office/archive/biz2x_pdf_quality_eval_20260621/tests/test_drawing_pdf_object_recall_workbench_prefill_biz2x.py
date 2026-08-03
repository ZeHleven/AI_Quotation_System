from __future__ import annotations

import json
import sys
from pathlib import Path

from openpyxl import load_workbook

from app.services.drawing_pdf_external_recall_template_status import build_external_recall_template_status
from app.services.drawing_pdf_gap_recall_importer import (
    build_gap_recall_external_import_report,
    load_external_recall_results,
)
from app.services.drawing_pdf_object_recall_workbench_prefill import (
    build_object_recall_workbench_prefill_report,
    load_object_recall_workbench_rows,
    write_object_recall_workbench_prefill_outputs,
)
from scripts import biz2x_pdf_object_recall_workbench_prefill


def test_object_recall_workbench_prefill_uses_local_evidence_not_answer_columns():
    report = build_object_recall_workbench_prefill_report(_workbench_rows(), _v2_report())

    assert report["summary"]["object_recall_task_count"] == 2
    assert report["summary"]["prefilled_row_count"] == 1
    assert report["summary"]["answer_columns_used_for_prefill"] is False
    first = report["workbench_rows"][0]
    assert first["fill_status"] == "importable"
    assert first["ready_for_import"] == "true"
    assert first["target_item_name"] == "墙面石材湿贴"
    assert first["evidence_item_hint"] == "墙面装饰板"
    assert first["evidence_spec_or_method"] == "石材"
    assert first["evidence_suggested_unit"] == "㎡"
    assert first["evidence_text"] == "石材"
    second = report["workbench_rows"][1]
    assert second["fill_status"] == "answer_only_reference"
    assert second["evidence_text"] == ""


def test_object_recall_workbench_prefill_outputs_are_importable(tmp_path: Path):
    report = build_object_recall_workbench_prefill_report(_workbench_rows(), _v2_report())

    outputs = write_object_recall_workbench_prefill_outputs(report, tmp_path / "out", stem="prefilled_workbench")

    assert Path(outputs["xlsx"]).exists()
    assert Path(outputs["prefill_status_csv"]).exists()
    workbook = load_workbook(outputs["xlsx"])
    sheet = workbook["object_recall_workbench"]
    headers = {cell.value: cell.column for cell in sheet[1]}
    assert sheet.cell(row=2, column=headers["evidence_item_hint"]).value == "墙面装饰板"
    assert sheet.cell(row=2, column=headers["evidence_text"]).value == "石材"

    external_results = load_external_recall_results(outputs["xlsx"])
    status = build_external_recall_template_status(external_results)
    imported = build_gap_recall_external_import_report(
        external_results,
        recall_plan=_recall_plan(tmp_path),
        source_name="object-prefill",
    )

    assert status["summary"]["importable_row_count"] == 1
    assert imported["summary"]["evidence_count"] == 1
    assert imported["summary"]["validation_status_counts"] == {"imported": 1, "skipped": 1}
    assert imported["evidence_rows"][0]["item_hint"] == "墙面装饰板"


def test_object_recall_workbench_prefill_cli(tmp_path: Path, monkeypatch, capsys):
    workbench_json = tmp_path / "workbench.json"
    v2_json = tmp_path / "v2.json"
    workbench_json.write_text(json.dumps({"workbench_rows": _workbench_rows()}, ensure_ascii=False), encoding="utf-8")
    v2_json.write_text(json.dumps(_v2_report(), ensure_ascii=False), encoding="utf-8")
    output_dir = tmp_path / "cli_out"
    monkeypatch.setattr(
        sys,
        "argv",
        [
            "biz2x_pdf_object_recall_workbench_prefill.py",
            "--object-workbench",
            str(workbench_json),
            "--v2-json",
            str(v2_json),
            "--output-dir",
            str(output_dir),
            "--timestamp",
            "fixed",
        ],
    )

    assert biz2x_pdf_object_recall_workbench_prefill.main() == 0
    payload = json.loads(capsys.readouterr().out)

    assert payload["ok"] is True
    assert payload["summary"]["prefilled_row_count"] == 1
    assert Path(payload["outputs"]["xlsx"]).exists()


def test_load_object_recall_workbench_rows_from_json(tmp_path: Path):
    path = tmp_path / "workbench.json"
    path.write_text(json.dumps({"workbench_rows": _workbench_rows()}, ensure_ascii=False), encoding="utf-8")

    rows = load_object_recall_workbench_rows(path)

    assert len(rows) == 2
    assert rows[0]["target_item_name"] == "墙面石材湿贴"


def _workbench_rows() -> list[dict[str, object]]:
    return [
        {
            "task_no": 1,
            "gap_priority": "P1_object_missing",
            "object_class": "finish_wall",
            "recommended_pass": "finish_schedule",
            "source_file": "drawing.pdf",
            "page": "1",
            "tile_id": "p001_whole",
            "image_exists": True,
            "target_item_name": "墙面石材湿贴",
            "target_feature": "石材湿贴",
            "target_unit": "㎡",
            "target_object_terms": "石材；墙面石材",
            "required_evidence_keywords": "石材；墙面",
            "evidence_item_hint": "",
            "evidence_spec_or_method": "",
            "evidence_suggested_unit": "",
            "evidence_text": "",
        },
        {
            "task_no": 2,
            "gap_priority": "P1_object_missing",
            "object_class": "door_window_demolition",
            "recommended_pass": "door_window_demolition",
            "source_file": "drawing.pdf",
            "page": "1",
            "tile_id": "p001_whole",
            "image_exists": True,
            "target_item_name": "拆除不锈钢玻璃门",
            "target_feature": "含门套、门扇及五金拆除",
            "target_unit": "套",
            "target_object_terms": "不锈钢玻璃门；玻璃门；门套；门扇",
            "required_evidence_keywords": "不锈钢玻璃门；拆除",
            "evidence_item_hint": "",
            "evidence_spec_or_method": "",
            "evidence_suggested_unit": "",
            "evidence_text": "",
        },
    ]


def _v2_report() -> dict[str, object]:
    return {
        "evidence_rows": [
            {
                "evidence_id": "LOCAL-STONE",
                "source_file": "drawing.pdf",
                "page": "1",
                "tile_id": "p001_whole",
                "evidence_type": "wall",
                "discipline": "decoration",
                "raw_item_name": "墙面装饰板",
                "spec_or_method": "石材",
                "suggested_unit": "㎡",
                "evidence_text": "石材",
                "confidence": 0.9,
            },
            {
                "evidence_id": "LOCAL-GENERIC-DEMOLITION",
                "source_file": "drawing.pdf",
                "page": "1",
                "tile_id": "p001_whole",
                "evidence_type": "demolition",
                "discipline": "decoration",
                "raw_item_name": "拆除",
                "spec_or_method": "拆除",
                "suggested_unit": "㎡",
                "evidence_text": "拆除",
                "confidence": 1,
            },
        ]
    }


def _recall_plan(tmp_path: Path) -> dict[str, object]:
    image = tmp_path / "page.png"
    image.write_bytes(b"fake")
    return {
        "plan_rows": [
            {
                "task_no": 1,
                "gap_no": 1,
                "answer_item_name": "墙面石材湿贴",
                "recommended_pass": "finish_schedule",
                "source_file": "drawing.pdf",
                "page": 1,
                "tile_id": "p001_whole",
                "tile_type": "whole_page_preview",
                "image_path": str(image),
            },
            {
                "task_no": 2,
                "gap_no": 2,
                "answer_item_name": "拆除不锈钢玻璃门",
                "recommended_pass": "door_window_demolition",
                "source_file": "drawing.pdf",
                "page": 1,
                "tile_id": "p001_whole",
                "tile_type": "whole_page_preview",
                "image_path": str(image),
            },
        ]
    }
