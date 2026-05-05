import base64
import hashlib
import hmac
import re
import os
import csv
import uuid
import urllib3
import logging
from datetime import datetime
from io import StringIO, BytesIO
import openpyxl
from openpyxl.utils import get_column_letter
from pathlib import Path
from typing import Optional
from fastapi import APIRouter, HTTPException, Depends, status, UploadFile, File, Form, Body
from fastapi.responses import StreamingResponse
import asyncio
from fastapi.security import OAuth2PasswordBearer
from sqlalchemy.orm import Session
import requests
import jwt
import json
from pydantic import BaseModel

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

from app.core.database import get_db, SessionLocal
from app.models.user import User
from app.models.quote_history import QuoteHistory
from app.core.security import SECRET_KEY, ALGORITHM
from app.core.config import settings
from app.core.logging import get_trace_id, reset_trace_id, set_trace_id
from app.services.model_gateway import call_glm_vision_extract, post_json_via_gateway

router = APIRouter()
logger = logging.getLogger(__name__)

# ==========================================
# 🚀 基础配置区
# ==========================================
N8N_WEBHOOK_URL_CALC = settings.n8n_webhook_url_calc
N8N_WEBHOOK_URL_PUSH = settings.n8n_webhook_url_push
oauth2_scheme = OAuth2PasswordBearer(tokenUrl="/api/v1/auth/login")

ZHIPU_API_KEY = settings.zhipu_api_key
WEBHOOK_SECRET = settings.webhook_secret
RAG_SERVICE_URL = settings.rag_service_url
RELOAD_SECRET = settings.reload_secret


def _sign_payload(body: dict) -> dict:
    """Return backward-compatible n8n auth headers plus an HMAC signature."""
    canonical_body = json.dumps(body, ensure_ascii=False, separators=(",", ":"), sort_keys=True)
    signature = hmac.new(
        WEBHOOK_SECRET.encode("utf-8"),
        canonical_body.encode("utf-8"),
        hashlib.sha256,
    ).hexdigest()
    return {
        "Content-Type": "application/json",
        "X-Webhook-Secret": WEBHOOK_SECRET,
        "X-Webhook-Signature": signature,
    }


def _sse_event(status_name: str, message: str, trace_id: Optional[str] = None, **extra):
    payload = {"status": status_name, "message": message, "trace_id": trace_id or get_trace_id()}
    payload.update(extra)
    return f"data: {json.dumps(payload, ensure_ascii=False)}\n\n"


def _build_quote_filename(username: str) -> str:
    safe_user = re.sub(r"[^A-Za-z0-9_-]+", "_", username or "user").strip("_") or "user"
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return f"quote_{timestamp}_{safe_user}.xlsx"


def _attach_quote_filename(payload: dict, username: str) -> dict:
    payload = dict(payload)
    filename = payload.get("excel_filename") or payload.get("filename") or _build_quote_filename(username)
    display_title = payload.get("display_title") or f"AI报价单-{datetime.now().strftime('%Y-%m-%d %H:%M')}"
    payload.update(
        {
            "excel_filename": filename,
            "download_filename": filename,
            "filename": filename,
            "fileName": filename,
            "file_name": filename,
            "attachment_name": filename,
            "display_title": display_title,
        }
    )
    return payload


def _cjk_len(s: str) -> int:
    """计算字符串显示宽度：CJK 字符按 2 计，其余按 1 计"""
    w = 0
    for ch in str(s):
        code = ord(ch)
        if (0x4E00 <= code <= 0x9FFF) or (0x3400 <= code <= 0x4DBF) or (0xFF01 <= code <= 0xFFEE):
            w += 2
        else:
            w += 1
    return w


def _build_excel_base64(project_details: list) -> str:
    """用 openpyxl 生成自适应列宽的 Excel，返回 base64 字符串"""
    try:
        headers = ["施工项目", "AI核准单价(元)", "项目合计(元)", "工艺备注"]
        field_map = ["project_name", "unit_price", "total_price", "notes"]

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "报价单"
        ws.append(headers)

        for item in project_details:
            ws.append([item.get(f, "") for f in field_map])

        for col_idx in range(1, len(headers) + 1):
            col_letter = get_column_letter(col_idx)
            max_w = max(_cjk_len(str(cell.value or "")) for cell in ws[col_letter])
            ws.column_dimensions[col_letter].width = max_w + 4

        buf = BytesIO()
        wb.save(buf)
        return base64.b64encode(buf.getvalue()).decode()
    except Exception:
        logger.exception("build_excel_base64_failed")
        return ""


# ==========================================
# 🛠️ 辅助功能与中间件
# ==========================================
def get_current_user(token: str = Depends(oauth2_scheme), db: Session = Depends(get_db)):
    credentials_exception = HTTPException(
        status_code=status.HTTP_401_UNAUTHORIZED, detail="认证失效", headers={"WWW-Authenticate": "Bearer"},
    )
    try:
        payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
        username: str = payload.get("sub")
        if username is None: raise credentials_exception
    except Exception:
        raise credentials_exception
    user = db.query(User).filter(User.username == username).first()
    if user is None: raise credentials_exception
    return user


async def analyze_image_with_domestic_ai(base64_image: str, mime_type: str):
    """利用国内智谱大模型 (GLM-4V) 提取业务信息"""
    return await call_glm_vision_extract(base64_image, mime_type, trace_id=get_trace_id())


# ==========================================
# 🌐 V2.0 业务核心接口 (聊天、算价、推送)
# ==========================================
@router.post("/chat", summary="AI 业务对话 (支持文本/图片/文档)")
async def process_chat(
        message: str = Form(None),
        file: UploadFile = File(None),
        current_user: User = Depends(get_current_user),
        db: Session = Depends(get_db)
):
    if current_user.quota <= 0:
        raise HTTPException(status_code=403, detail="您的 AI 调用额度已耗尽，请联系管理员充值")

    logger.info("quote_request_started", extra={"username": current_user.username, "event": "quote_request_started"})
    file_content = await file.read() if file else None
    mime_type = file.content_type if file else None
    filename = file.filename if file else None
    request_trace_id = get_trace_id()

    async def event_generator():
        trace_token = set_trace_id(request_trace_id)
        yield f": {' ' * 1024}\n\n"
        final_query = message or ""
        try:
            yield _sse_event("processing", "[API Gateway] 📡 安全握手成功，已剥离 JWT 并唤醒底层引擎...", trace_id=request_trace_id)
            await asyncio.sleep(0.5)

            if file_content:
                if "pdf" in mime_type.lower():
                    yield _sse_event("error", "❌ [格式校验] 拦截：国内引擎暂只支持图片输入，请截图重试", trace_id=request_trace_id)
                    return

                yield _sse_event("processing", f"[Vision Module] 📸 正在驱动 GLM-4V 多模态大模型扫描附件: {filename}...", trace_id=request_trace_id)

                base64_data = base64.b64encode(file_content).decode('utf-8')
                try:
                    extracted_text = await call_glm_vision_extract(
                        base64_data,
                        mime_type,
                        username=current_user.username,
                        trace_id=request_trace_id,
                    )
                except Exception as glm_err:
                    logger.exception("vision_model_failed", extra={"username": current_user.username, "event": "vision_model_failed"})
                    yield _sse_event("error", f"❌ [Vision Module] GLM-4V 调用失败: {str(glm_err)}", trace_id=request_trace_id)
                    return

                if extracted_text:
                    final_query = f"{final_query} [从文件识别到的内容]: {extracted_text}".replace('\n', '；').replace(
                        '\r', '')
                    yield _sse_event("processing", "[Vision Module] ✅ 提取完毕，已成功结构化二维图纸特征！", trace_id=request_trace_id)
                    await asyncio.sleep(0.5)
                else:
                    yield _sse_event("error", "❌ [Vision Module] GLM-4V 返回空内容，请重试", trace_id=request_trace_id)
                    return
            else:
                if final_query.strip():
                    yield _sse_event("processing", "[Text Module] 📝 识别纯文本指令，正在执行语义清洗...", trace_id=request_trace_id)
                    await asyncio.sleep(0.5)

            if not final_query.strip():
                yield _sse_event("error", "❌ 请输入业务指令或上传清单图片", trace_id=request_trace_id)
                return

            yield _sse_event("processing", "[RAG & Agent] 🔍 正在穿透企业知识库寻找刚性底价并驱动专家大脑...\n(后台算力执行中，预计静候 15~30 秒，完成后将弹出核对面板)", trace_id=request_trace_id)

            payload = {"text": {"content": final_query}, "conversationId": str(uuid.uuid4())}
            response = await post_json_via_gateway(
                provider="n8n",
                model="dify-deepseek",
                endpoint_type="quote_calc",
                url=N8N_WEBHOOK_URL_CALC,
                json_payload=payload,
                headers=_sign_payload(payload),
                timeout=180,
                username=current_user.username,
                trace_id=request_trace_id,
            )

            if response.status_code == 200:
                try:
                    calc_result = response.json()
                except Exception:
                    body_preview = response.text[:500].strip() if response.text else "<empty>"
                    logger.exception("n8n_response_parse_failed", extra={"username": current_user.username, "event": "n8n_response_parse_failed"})
                    yield _sse_event("error", f"❌ [n8n Workflow] 响应体解析失败（HTTP 200）\n实际返回内容：{body_preview}\n→ 请确认 N8N budget-calc 工作流末尾有 Respond to Webhook 节点且 Response Body 设为 JSON", trace_id=request_trace_id)
                    return
                current_user.quota -= 1
                db.commit()
                logger.info("quote_request_finished", extra={"username": current_user.username, "event": "quote_request_finished"})
                yield _sse_event("preview", "[n8n Workflow] ✅ AI 预审数据已就绪，请人工复核！", trace_id=request_trace_id, data=calc_result)
            else:
                try:
                    error_detail = response.json().get("message", "未知错误")
                except Exception:
                    error_detail = f"状态码 {response.status_code}，响应体为空"
                logger.error("n8n_workflow_failed", extra={"username": current_user.username, "event": "n8n_workflow_failed", "status_code": response.status_code})
                yield _sse_event("error", f"❌ [n8n Workflow] 中断: 底层算价引擎抛出异常 -> {error_detail}", trace_id=request_trace_id)

        except requests.exceptions.Timeout:
            logger.exception("quote_request_timeout", extra={"username": current_user.username, "event": "quote_request_timeout"})
            yield _sse_event("error", "❌ [n8n Workflow] 严重超时：请检查 Dify 模型是否拥堵挂起", trace_id=request_trace_id)
        except Exception as e:
            logger.exception("quote_request_crashed", extra={"username": current_user.username, "event": "quote_request_crashed"})
            yield _sse_event("error", f"❌ [API Gateway] 流转崩溃: {str(e)}", trace_id=request_trace_id)
        finally:
            reset_trace_id(trace_token)

    return StreamingResponse(
        event_generator(),
        media_type="text/event-stream",
        headers={"Cache-Control": "no-cache, no-transform", "Connection": "keep-alive", "X-Accel-Buffering": "no"}
    )


@router.post("/confirm_push", summary="人工审核通过后推送钉钉")
async def confirm_and_push(
        payload: dict = Body(...),
        current_user: User = Depends(get_current_user),
        db: Session = Depends(get_db)
):
    try:
        payload = _attach_quote_filename(payload, current_user.username)
        details = payload.get("project_details", [])
        excel_b64 = _build_excel_base64(details)
        logger.info("confirm_push_excel_check", extra={"excel_b64_len": len(excel_b64), "details_count": len(details)})
        if not excel_b64:
            raise HTTPException(status_code=500, detail="❌ Excel生成失败，请查看FastAPI日志中的 build_excel_base64_failed 错误")
        payload = dict(payload)
        payload["excel_base64"] = excel_b64
        logger.info("confirm_push_payload_keys", extra={"keys": list(payload.keys())})
        response = await post_json_via_gateway(
            provider="n8n",
            model="dingtalk-export",
            endpoint_type="quote_push",
            url=N8N_WEBHOOK_URL_PUSH,
            json_payload=payload,
            headers=_sign_payload(payload),
            timeout=60,
            username=current_user.username,
            trace_id=get_trace_id(),
        )
        if response.status_code == 200:
            # 写入历史记录
            try:
                details = payload.get("project_details", [])
                total = sum(float(item.get("total_price", 0)) for item in details)
                record = QuoteHistory(
                    username=current_user.username,
                    total_amount=round(total, 2),
                    item_count=len(details),
                    payload_json=json.dumps(payload, ensure_ascii=False),
                )
                db.add(record)
                db.commit()
            except Exception:
                pass  # 历史写入失败不影响主流程
            return {"message": "✅ 最终报价单已成功投递至钉钉群！"}
        else:
            raise HTTPException(status_code=500, detail="底层推送流水线异常")
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ==========================================
# 📊 V3.0 管理员后台 API (物料库增删改查与同步)
# ==========================================
# 数据模型：增加 is_draft 字段，用于标记新导入的待审数据
class MaterialItem(BaseModel):
    id: str
    item_name: str
    unit_price: float
    unit: str
    notes: str
    is_draft: Optional[bool] = False


DATA_FILE = settings.materials_file
MATERIAL_AUDIT_DIR = DATA_FILE.parent / "materials_audit"
SNAPSHOT_ID_PATTERN = re.compile(r"^\d{8}_\d{6}_[0-9a-f]{8}$")


def load_data():
    if not DATA_FILE.exists():
        return []
    with DATA_FILE.open("r", encoding="utf-8") as f:
        return json.load(f)


def save_data(data):
    DATA_FILE.parent.mkdir(parents=True, exist_ok=True)
    with DATA_FILE.open("w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


def _snapshot_metadata(snapshot: dict) -> dict:
    return {
        "snapshot_id": snapshot.get("snapshot_id"),
        "created_at": snapshot.get("created_at"),
        "username": snapshot.get("username"),
        "action": snapshot.get("action"),
        "reason": snapshot.get("reason"),
        "item_count": snapshot.get("item_count", 0),
    }


def _safe_snapshot_path(snapshot_id: str) -> Path:
    if not SNAPSHOT_ID_PATTERN.match(snapshot_id or ""):
        raise HTTPException(status_code=400, detail="Invalid snapshot id")
    return MATERIAL_AUDIT_DIR / f"{snapshot_id}.json"


def create_material_snapshot(username: str, action: str, reason: str = "", data: Optional[list] = None) -> dict:
    snapshot_data = load_data() if data is None else data
    snapshot_id = f"{datetime.now().strftime('%Y%m%d_%H%M%S')}_{uuid.uuid4().hex[:8]}"
    snapshot = {
        "snapshot_id": snapshot_id,
        "created_at": datetime.now().isoformat(timespec="seconds"),
        "username": username,
        "action": action,
        "reason": reason,
        "item_count": len(snapshot_data),
        "data": snapshot_data,
    }
    MATERIAL_AUDIT_DIR.mkdir(parents=True, exist_ok=True)
    with _safe_snapshot_path(snapshot_id).open("w", encoding="utf-8") as f:
        json.dump(snapshot, f, ensure_ascii=False, indent=2)
    return _snapshot_metadata(snapshot)


def list_material_snapshots(limit: int = 30) -> list[dict]:
    if not MATERIAL_AUDIT_DIR.exists():
        return []
    snapshots = []
    for path in sorted(MATERIAL_AUDIT_DIR.glob("*.json"), reverse=True):
        try:
            with path.open("r", encoding="utf-8") as f:
                snapshots.append(_snapshot_metadata(json.load(f)))
        except Exception:
            logger.exception("material_snapshot_read_failed", extra={"path": str(path)})
        if len(snapshots) >= limit:
            break
    return snapshots


def load_material_snapshot(snapshot_id: str) -> dict:
    path = _safe_snapshot_path(snapshot_id)
    if not path.exists():
        raise HTTPException(status_code=404, detail="Snapshot not found")
    with path.open("r", encoding="utf-8") as f:
        snapshot = json.load(f)
    if not isinstance(snapshot.get("data"), list):
        raise HTTPException(status_code=400, detail="Snapshot data is invalid")
    return snapshot


def require_admin(current_user: User = Depends(get_current_user)):
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="权限不足，仅管理员可操作")
    return current_user


@router.get("/admin/materials", summary="获取所有物料清单")
async def get_materials(current_user: User = Depends(require_admin)):
    return {"code": 200, "data": load_data()}


@router.post("/admin/materials", summary="保存/覆盖整个物料清单")
async def save_materials(items: list[MaterialItem], current_user: User = Depends(require_admin)):
    before_data = load_data()
    snapshot = create_material_snapshot(
        username=current_user.username,
        action="save_before_overwrite",
        reason="Auto snapshot before saving materials",
        data=before_data,
    )
    data = [item.dict() if hasattr(item, 'dict') else item.model_dump() for item in items]
    save_data(data)
    return {"code": 200, "message": "保存成功", "snapshot": snapshot}


@router.get("/admin/materials/audit", summary="查看知识库变更快照")
async def get_material_audit(limit: int = 30, current_user: User = Depends(require_admin)):
    limit = max(1, min(limit, 100))
    return {"code": 200, "data": list_material_snapshots(limit)}


@router.post("/admin/materials/rollback/{snapshot_id}", summary="回滚知识库到指定快照")
async def rollback_materials(snapshot_id: str, current_user: User = Depends(require_admin)):
    snapshot = load_material_snapshot(snapshot_id)
    current_snapshot = create_material_snapshot(
        username=current_user.username,
        action="rollback_before_restore",
        reason=f"Auto snapshot before rollback to {snapshot_id}",
    )
    save_data(snapshot["data"])
    return {
        "code": 200,
        "message": "Rollback completed",
        "restored_snapshot": _snapshot_metadata(snapshot),
        "current_snapshot": current_snapshot,
    }


# 🚀🚀 新增功能：CSV 历史记录提炼引擎
@router.post("/admin/upload_csv", summary="解析并提炼历史成交记录")
async def upload_csv_history(file: UploadFile = File(...), current_user: User = Depends(require_admin)):
    """接收新的 CSV 成交记录，过滤重复项，提炼出异动/新项目"""
    if not file.filename.endswith('.csv'):
        raise HTTPException(status_code=400, detail="目前仅支持 CSV 格式历史记录导入")

    content = await file.read()

    # 🚀 防爆盾 1：拦截“假 CSV”（检测 Excel 文件的 PK 压缩包头）
    if content.startswith(b'PK'):
        raise HTTPException(status_code=400,
                            detail="解析拦截：检测到您直接修改了 Excel 文件的后缀名。请在 Excel 中打开文件，点击【文件 -> 另存为 -> CSV (逗号分隔)】后再上传！")

    # 兼容多种常见的中文编码
    try:
        text = content.decode('utf-8-sig')
    except UnicodeDecodeError:
        try:
            text = content.decode('gbk')
        except UnicodeDecodeError:
            text = content.decode('utf-8', errors='ignore')

    try:
        # 🚀 防爆盾 2：加入 newline='' 防止真实的 CSV 内部包含换行符导致解析崩溃
        reader = csv.DictReader(StringIO(text, newline=''))
        rows = list(reader)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"CSV 格式异常或损坏: {str(e)}")

    existing_data = load_data()
    existing_items = {str(item['item_name']).strip(): float(item['unit_price']) for item in existing_data}

    new_drafts = []

    for row in rows:
        # 🚀 兼容补充：加入了你表格里的专属表头 “AI核准单价(元)”
        item_name = row.get('施工项目') or row.get('项目名称') or row.get('item_name') or ""
        unit_price_str = row.get('AI核准单价(元)') or row.get('综合单价(元)') or row.get('单价') or row.get(
            'unit_price') or "0"
        unit = row.get('单位') or row.get('unit') or "项"
        notes = row.get('规格/工艺/材料说明') or row.get('备注说明') or row.get('备注') or row.get('notes') or ""

        item_name = str(item_name).strip()
        # 过滤汇总噪音和空行
        if not item_name or "汇总" in item_name or "结算" in item_name:
            continue

        try:
            price_val = float(unit_price_str)
        except ValueError:
            price_val = 0.0

        is_new = False

        # 核心逻辑 1：如果标准库里没有这个项目 -> 全新工艺，拉出待审
        if item_name not in existing_items:
            is_new = True
        else:
            # 核心逻辑 2：如果标准库里有，但这笔成交价格偏差超过 20% -> 异动待审
            old_price = existing_items[item_name]
            if old_price > 0 and abs(price_val - old_price) / old_price > 0.2:
                is_new = True
                item_name = f"{item_name} (历史异动待审)"

        if is_new:
            # 去重，防止同一个文件里的历史项目自己复读机
            if not any(d['item_name'] == item_name for d in new_drafts):
                new_drafts.append({
                    "id": f"draft_{uuid.uuid4().hex[:8]}",
                    "item_name": item_name,
                    "unit_price": price_val,
                    "unit": unit,
                    "notes": notes,
                    "is_draft": True  # 打上红色高亮草稿烙印
                })

    return {"code": 200, "message": "解析成功", "data": new_drafts}


@router.post("/admin/sync_milvus", summary="同步数据至 Milvus（委托 RAG 服务执行，零停机蓝绿切换）")
async def sync_to_milvus(current_user: User = Depends(require_admin)):
    """将本地 JSON 数据 POST 至 CentOS RAG 服务的 /admin/reload，由其完成向量化和蓝绿切换。
    Windows 端不再加载大模型，速度更快、内存占用更低。"""
    data = load_data()
    if not data:
        raise HTTPException(status_code=400, detail="本地知识库为空，无法同步")

    try:
        payload = {"materials": data, "secret": RELOAD_SECRET}
        response = await asyncio.to_thread(
            requests.post,
            f"{RAG_SERVICE_URL}/admin/reload",
            json=payload,
            timeout=120,
        )
        if response.status_code == 200:
            eval_report_id = None
            if settings.rag_eval_enabled:
                from app.services.rag_evaluator import trigger_eval_background
                eval_report_id = trigger_eval_background(current_user.username, SessionLocal)
            return {
                "code": 200,
                "message": response.json().get("message", "同步完成"),
                "eval_triggered": settings.rag_eval_enabled,
                "eval_report_id": eval_report_id,
            }
        else:
            detail = response.json().get("detail", f"状态码 {response.status_code}")
            raise HTTPException(status_code=500, detail=f"RAG 服务返回错误: {detail}")
    except requests.exceptions.Timeout:
        raise HTTPException(status_code=504, detail="RAG 服务超时，请检查 CentOS 容器状态")
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ==========================================
# 👥 用户配额管理接口
# ==========================================
class QuotaUpdate(BaseModel):
    quota: int


@router.get("/admin/users", summary="获取所有用户列表")
async def list_users(db: Session = Depends(get_db), current_user: User = Depends(require_admin)):
    users = db.query(User).order_by(User.id).all()
    return {"code": 200, "data": [
        {"id": u.id, "username": u.username, "role": u.role, "quota": u.quota, "is_active": u.is_active}
        for u in users
    ]}


# ==========================================
# 📋 报价历史记录接口
# ==========================================
@router.get("/history", summary="查询报价历史（本人；admin 可查全部）")
async def get_history(
    page: int = 1,
    page_size: int = 20,
    username: Optional[str] = None,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    query = db.query(QuoteHistory)
    if current_user.role != "admin":
        query = query.filter(QuoteHistory.username == current_user.username)
    elif username:
        query = query.filter(QuoteHistory.username == username)

    total = query.count()
    records = query.order_by(QuoteHistory.created_at.desc()) \
                   .offset((page - 1) * page_size).limit(page_size).all()

    return {"code": 200, "total": total, "data": [
        {
            "id": r.id,
            "username": r.username,
            "created_at": r.created_at.strftime("%Y-%m-%d %H:%M:%S") if r.created_at else "",
            "total_amount": r.total_amount,
            "item_count": r.item_count,
            "payload_json": r.payload_json,
        } for r in records
    ]}


@router.patch("/admin/users/{user_id}/quota", summary="设置指定用户的 AI 调用额度")
async def set_user_quota(
    user_id: int,
    body: QuotaUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin)
):
    if body.quota < 0:
        raise HTTPException(status_code=400, detail="额度不能为负数")
    user = db.query(User).filter(User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="用户不存在")
    user.quota = body.quota
    db.commit()
    return {"code": 200, "message": f"已将 {user.username} 的额度设置为 {body.quota} 次"}
