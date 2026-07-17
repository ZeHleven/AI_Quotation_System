from __future__ import annotations

import json
import re
from io import BytesIO
from collections import Counter
from copy import deepcopy
from datetime import datetime, timezone
from typing import Any
from zipfile import ZIP_DEFLATED, ZipFile
from xml.sax.saxutils import escape

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from app.core.config import settings
from app.models.bidding import BidParseRun, BidProject, BidProjectFile, TenderBusinessObject, TenderRequirement, TenderRisk
from app.services.bidding_parser import dumps_json, loads_json
from app.services.bidding_risk_cards import RISK_TYPE_LABELS
from app.services.model_gateway import post_json_via_gateway


TENDER_ANALYSIS_SCHEMA_VERSION = "biz4a_tender_analysis_schema_v1.0"
TENDER_ANALYSIS_PREVIEW_VERSION = "biz4a_tender_analysis_preview_v1.0"
TENDER_ANALYSIS_SUMMARY_LLM_PROMPT_VERSION = "biz4a_summary_semantic_extract_v2"
TENDER_ANALYSIS_IMPORTANT_INFO_PROMPT_VERSION = "biz4a_important_info_llm_extract_v1"
TENDER_ANALYSIS_RISK_CLAUSE_PROMPT_VERSION = "biz4a_risk_clause_llm_extract_v1"


IMPORTANT_INFO_SECTION_SCHEMA: tuple[dict[str, Any], ...] = (
    {
        "section_key": "project_overview",
        "title": "一、项目概况",
        "fields": (
            ("tender_project_name", "招标工程名称"),
            ("project_location", "工程地点"),
            ("tenderer", "招标单位"),
            ("general_contractor", "总承包单位"),
            ("scale_features", "工程规模及特征"),
            ("tender_scope", "招标范围"),
        ),
    },
    {
        "section_key": "qa_site_deadline",
        "title": "二、答疑、踏勘与截标",
        "fields": (
            ("qa_deadline", "答疑截至时间"),
            ("clarification_reply", "澄清回复"),
            ("qa_contact_email", "答疑联系人/邮箱"),
            ("site_visit_arrangement", "踏勘安排"),
            ("site_visit_contact", "踏勘联系人"),
            ("submission_deadline", "截标时间"),
            ("submission_place", "递交地点"),
            ("submission_contact", "递交联系人"),
        ),
    },
    {
        "section_key": "pricing_contracting",
        "title": "三、报价方式与工程承包方式",
        "fields": (
            ("contract_form", "合同形式"),
            ("unit_price_includes", "综合单价包含"),
            ("quantity_rule", "工程量规则"),
            ("lump_sum_items", "包干项目"),
            ("pricing_risk", "报价风险"),
        ),
    },
    {
        "section_key": "bond_guarantee",
        "title": "四、保证金与担保",
        "fields": (
            ("bid_bond_amount", "投标担保金额"),
            ("payment_time", "缴交时间"),
            ("bond_form", "形式"),
            ("receiving_account", "收款账户"),
            ("guarantee_requirement", "保函要求"),
            ("missing_consequence", "未提交后果"),
            ("forfeiture_situations", "没收情形"),
            ("performance_bond", "履约担保"),
        ),
    },
    {
        "section_key": "bid_document_requirements",
        "title": "五、标书制作要求",
        "fields": (
            ("bidder_qualification", "投标人资质"),
            ("document_copies", "投标文件份数"),
            ("commercial_bid_content", "商务标内容"),
            ("technical_bid_content", "技术标内容"),
            ("format_requirement", "格式要求"),
            ("signing_requirement", "签署要求"),
            ("pricing_requirement", "报价要求"),
        ),
    },
    {
        "section_key": "sealing_requirements",
        "title": "六、封标要求",
        "fields": (("sealing_requirements", "封标要求"),),
    },
    {
        "section_key": "evaluation_rejection",
        "title": "七、评标标准与废标条件",
        "fields": (
            ("evaluation_principle", "评标原则"),
            ("winner_determination", "确定中标人"),
            ("insufficient_bidders", "投标人数不足"),
            ("bid_rejection_conditions", "废标条件"),
        ),
    },
    {
        "section_key": "construction_period",
        "title": "八、工期",
        "fields": (
            ("tentative_start", "暂定开工"),
            ("actual_start", "实际开工"),
            ("contract_period", "合同工期"),
            ("clarification_items", "需澄清项"),
        ),
    },
    {
        "section_key": "payment_terms",
        "title": "九、付款方式",
        "fields": (
            ("advance_payment", "预付款"),
            ("performance_guarantee", "履约保函"),
            ("progress_claim", "进度款申报"),
            ("progress_payment", "进度款支付"),
            ("payment_cap", "支付上限"),
            ("settlement_payment", "结算款"),
            ("preliminaries_fee", "开办费"),
        ),
    },
    {
        "section_key": "pre_bid_clarifications",
        "title": "十、回标前优先澄清清单",
        "fields": (("priority_clarifications", "回标前优先澄清清单"),),
    },
)

IMPORTANT_INFO_FIELD_KEYS = {
    str(field_key)
    for section in IMPORTANT_INFO_SECTION_SCHEMA
    for field_key, _ in section["fields"]
}

TENDER_ANALYSIS_REVIEW_STATUSES = {
    "pending": "待复核",
    "confirmed": "已确认",
    "needs_revision": "需修正",
    "ignored": "已忽略",
    "to_clarify": "需答疑",
    "to_quote_allowance": "报价预留",
    "mixed": "多状态",
}

TENDER_ANALYSIS_TABLE_LABELS = {
    "summary": "结构化信息摘要表",
    "scoring": "评分细则表",
    "risk_clause": "风险条款清单",
}

SUMMARY_COMBINE_ITEM_KEYS = {
    "project_overview",
    "bid_bond",
    "contact_person",
    "site_visit",
    "bid_document_requirements",
    "sealing_requirements",
    "scoring_weight",
    "payment_terms",
    "pre_bid_clarifications",
}

SUMMARY_KEYWORDS: dict[str, tuple[str, ...]] = {
    "project_overview": (
        "项目概况",
        "项目地点",
        "工程概况",
        "招标工程名称",
        "工程名称",
        "建设地点",
        "工程地点",
        "工程规模",
        "工程特征",
        "工程范围",
        "招标范围",
        "承包范围",
        "施工范围",
        "发包范围",
        "装修范围",
        "工程内容",
    ),
    "qa_deadline": (
        "答疑截至时间",
        "答疑截止时间",
        "答疑截止",
        "质疑截止",
        "澄清截止",
        "提问截止",
        "疑问提交",
        "问题提交",
        "答疑联系人",
        "答疑邮箱",
        "澄清",
        "补遗",
        "答疑",
        "质疑",
    ),
    "pricing_method": (
        "工程承包方式",
        "承包方式",
        "合同形式",
        "合同计价方式",
        "报价方式",
        "报价口径",
        "报价形式",
        "固定单价包干",
        "固定综合单价",
        "固定总价",
        "总价包干",
        "单价包干",
        "综合单价",
        "工程量清单",
    ),
    "bid_bond": (
        "投标保证金",
        "投标担保",
        "投标保函",
        "保证金",
        "保函",
        "缴纳形式",
        "缴交形式",
        "提交形式",
        "银行转账",
        "银行汇款",
        "开户行",
        "账户名",
        "户名",
        "账号",
        "递交时间",
    ),
    "contact_person": (
        "招投标联系人",
        "答疑联系人",
        "踏勘现场联系人",
        "现场踏勘联系人",
        "踏勘联系人",
        "递交联系人",
        "招标联系人",
        "代理联系人",
        "联系人",
        "联系方式",
        "联系电话",
        "手机",
        "邮箱",
    ),
    "site_visit": (
        "踏勘时间",
        "踏勘地点",
        "踏勘联系人",
        "现场踏勘联系人",
        "现场踏勘",
        "踏勘",
        "自行踏勘",
        "统一组织踏勘",
        "现场勘察",
        "勘察现场",
    ),
    "bid_document_requirements": (
        "投标书的编制",
        "投标文件要求",
        "投标文件的组成",
        "投标文件份数",
        "投标人资质",
        "资质要求",
        "资格证明文件",
        "商务标",
        "技术标",
        "经济标",
        "报价标",
        "标书",
        "投标文件",
        "正本",
        "副本",
        "电子文档",
        "签字",
        "盖章",
        "装订",
    ),
    "sealing_requirements": (
        "投标文件的密封和标记",
        "密封和标记",
        "密封",
        "封标",
        "封套",
        "密封袋",
        "外封套",
        "内封套",
        "封口",
        "骑缝",
        "标识",
        "标记",
        "注明工程名称",
        "正本",
        "副本",
    ),
    "submission_deadline": (
        "投标文件送交截止日期",
        "投标文件送交截止时间",
        "投标文件递交截止",
        "递交投标文件截止",
        "投标截止",
        "截标",
        "递交截止",
        "开标时间",
        "送达截止",
        "递交投标文件",
    ),
    "scoring_weight": (
        "评标标准",
        "评标办法",
        "废标条件",
        "无效标",
        "否决投标",
        "确定中标人",
        "定标原则",
        "中标候选人",
        "综合评定最优",
        "合理低价",
        "最低价",
        "评分权重",
        "权重",
        "技术分",
        "商务分",
        "报价分",
        "评标",
        "评分",
    ),
    "construction_period": ("合同工期", "工期要求", "计划工期", "总工期", "绝对工期", "日历天", "开工令", "开工", "竣工"),
    "payment_terms": (
        "合同价款的支付",
        "合同价款支付",
        "付款方式",
        "支付方式",
        "付款条件",
        "支付节点",
        "支付比例",
        "工程款",
        "进度款",
        "结算款",
        "预付款",
        "质保金",
        "履约保函",
        "审计后",
    ),
    "pre_bid_clarifications": (
        "重点提醒",
        "需澄清",
        "未列明",
        "未填写",
        "空白",
        "不一致",
        "冲突",
        "补遗",
        "回标前",
        "答疑",
        "下浮率",
        "另行提供",
        "待定",
    ),
}

SUMMARY_SEMANTIC_ITEM_DEFINITIONS: dict[str, dict[str, Any]] = {
    "project_overview": {
        "target": "项目概况。提取招标工程名称、工程地点、招标单位、总承包单位、工程规模及特征、招标范围；范围要区分不同楼栋/区域，排除仅为封面或目录的泛化标题。",
        "keywords": SUMMARY_KEYWORDS["project_overview"],
    },
    "qa_deadline": {
        "target": "答疑与澄清安排。提取答疑截至时间、澄清回复时间、答疑联系人/邮箱；若文件写了规则但具体日期、邮箱或联系人为空，要明确写未填写/需澄清，不得用普通收发联系人替代。",
        "keywords": SUMMARY_KEYWORDS["qa_deadline"],
    },
    "pricing_method": {
        "target": "报价方式与工程承包方式。提取合同形式/工程承包方式、综合单价包含范围、工程量调整/计量规则、包干项目口径、漏填/漏报报价风险；优先招标文件的工程承包方式原文。",
        "keywords": SUMMARY_KEYWORDS["pricing_method"],
    },
    "bid_bond": {
        "target": "保证金与担保。提取投标担保金额、缴交时间、缴交形式、收款账户、保函有效期/原件要求、未提交后果、没收情形、履约担保比例。",
        "keywords": SUMMARY_KEYWORDS["bid_bond"],
    },
    "contact_person": {
        "target": "联系人。按角色提取招投标联系人、答疑联系人、踏勘/现场预约联系人、递交联系人及联系方式；普通项目电话只能在角色明确时使用。",
        "keywords": SUMMARY_KEYWORDS["contact_person"],
    },
    "site_visit": {
        "target": "踏勘安排。提取是否统一组织、踏勘时间要求、踏勘地点、踏勘联系人/预约方式、现场进入要求；未列明联系人时要标注未列明。",
        "keywords": SUMMARY_KEYWORDS["site_visit"],
    },
    "bid_document_requirements": {
        "target": "标书制作要求。提取投标人资质、商务标/技术标/电子标份数，商务标内容，技术标内容，必须使用的格式、装订要求、签署盖章要求、报价填报要求；不要只返回'按招标文件要求'。",
        "keywords": SUMMARY_KEYWORDS["bid_document_requirements"],
    },
    "sealing_requirements": {
        "target": "封标要求。提取商务标、技术标、电子标的内外密封袋拆分方式，密封袋标识，工程名称/投标人名称标注要求，封口贴封条和骑缝盖章要求。",
        "keywords": SUMMARY_KEYWORDS["sealing_requirements"],
    },
    "submission_deadline": {
        "target": "截标与递交。提取具体投标截止/截标/递交截止日期时间、递交地点、递交联系人；如日期时间空白，要明确空白并列为需澄清，不得猜测。",
        "keywords": SUMMARY_KEYWORDS["submission_deadline"],
    },
    "scoring_weight": {
        "target": "评标标准与废标条件。提取评标原则/办法、确定中标人口径、投标人数不足处理、废标/无效标条件；若无评分分值，不要编造分值。",
        "keywords": SUMMARY_KEYWORDS["scoring_weight"],
    },
    "construction_period": {
        "target": "工期。提取各区域/楼栋暂定开工时间、实际开工依据、合同工期、是否含周末节假日；如果投标书格式工期与合同工期冲突，要明确标为需澄清。",
        "keywords": SUMMARY_KEYWORDS["construction_period"],
    },
    "payment_terms": {
        "target": "付款方式。提取预付款、履约保函对付款的影响、进度款申报时间、进度款支付比例/时限/最低支付额、累计支付上限、结算款、质保金、开办费等。",
        "keywords": SUMMARY_KEYWORDS["payment_terms"],
    },
    "pre_bid_clarifications": {
        "target": "回标前优先澄清清单。汇总文件中的空白项、冲突项和会影响报价/封标/工期/递交的未决事项，例如截标时间空白、答疑邮箱缺失、踏勘联系人未列明、工期前后不一致、下浮率空白、清单或电子文件是否另行提供。",
        "keywords": SUMMARY_KEYWORDS["pre_bid_clarifications"],
    },
}

SCORING_KEYWORDS = ("评标", "评分", "得分", "满分", "权重", "技术分", "商务分", "报价分", "分值")
SCORING_DETAIL_KEYWORDS = (
    "评分标准",
    "评分细则",
    "评分项",
    "评分因素",
    "评分内容",
    "评审标准",
    "评审因素",
    "评审内容",
    "分值",
    "满分",
    "技术标",
    "商务标",
    "报价分",
    "技术分",
    "商务分",
    "资信",
)
SCORING_BID_CONTEXT_KEYWORDS = (
    "评标",
    "评审",
    "投标",
    "技术标",
    "商务标",
    "报价",
    "资信",
    "资格",
    "综合评估",
)
SCORING_PROCESS_NOISE_KEYWORDS = (
    "监理",
    "月度考核",
    "现场安全文明",
    "安全文明大检查",
    "检查报告",
    "评分检查表",
    "总包考评",
    "罚款",
    "经济处罚",
    "停工整改",
)
PRICING_RISK_TYPES = {
    "fixed_total_price",
    "omission_liability",
    "no_price_adjustment",
    "advance_funding",
    "delayed_payment",
    "material_brand_constraint",
    "design_or_drawing_unclear",
}
CLARIFICATION_RISK_TYPES = {
    "site_condition",
    "design_or_drawing_unclear",
    "material_brand_constraint",
    "claim_time_limit",
}

RISK_LEVEL_ORDER = {"low": 1, "medium": 2, "high": 3}

PACKAGE_TYPE_LABELS = {
    "business": "商务标",
    "technical": "技术标",
    "pricing": "报价",
    "contract": "合同",
    "mixed": "综合",
    "unknown": "待确认",
}

RISK_LEVEL_LABELS = {
    "high": "高",
    "medium": "中",
    "low": "低",
}

REVIEW_CATEGORY_LABELS = {
    "summary_missing": "摘要缺失",
    "manual_review": "人工复核",
    "low_confidence": "低置信度",
    "high_risk": "高风险",
    "pricing": "影响报价",
    "clarification": "需答疑",
}

TENDER_ANALYSIS_SUMMARY_ITEM_CATALOG: tuple[dict[str, Any], ...] = (
    {
        "item_key": "project_overview",
        "item_name": "项目概况",
        "category": "项目基础",
        "value_type": "text",
        "required": True,
        "downstream": ["项目台账", "商务标基础信息", "技术标项目概况"],
    },
    {
        "item_key": "qa_deadline",
        "item_name": "答疑时间",
        "category": "投标时间",
        "value_type": "datetime_or_text",
        "required": False,
        "downstream": ["投标计划", "答疑清单"],
    },
    {
        "item_key": "pricing_method",
        "item_name": "报价方式",
        "category": "报价规则",
        "value_type": "enum_or_text",
        "required": True,
        "downstream": ["报价系统", "商务标报价说明"],
    },
    {
        "item_key": "bid_bond",
        "item_name": "保证金",
        "category": "投标规则",
        "value_type": "amount_or_text",
        "required": False,
        "downstream": ["商务标附件清单", "资金计划"],
    },
    {
        "item_key": "contact_person",
        "item_name": "联系人",
        "category": "项目基础",
        "value_type": "text",
        "required": False,
        "downstream": ["项目台账", "投标沟通"],
    },
    {
        "item_key": "site_visit",
        "item_name": "踏勘时间地点",
        "category": "投标时间",
        "value_type": "datetime_place_or_text",
        "required": False,
        "downstream": ["投标计划", "技术标现场条件"],
    },
    {
        "item_key": "bid_document_requirements",
        "item_name": "标书制作要求",
        "category": "文件要求",
        "value_type": "text",
        "required": True,
        "downstream": ["商务标工作台", "技术标工作台", "文件清单"],
    },
    {
        "item_key": "sealing_requirements",
        "item_name": "封标要求",
        "category": "文件要求",
        "value_type": "text",
        "required": True,
        "downstream": ["商务标递交检查", "废标风险检查"],
    },
    {
        "item_key": "submission_deadline",
        "item_name": "截标时间",
        "category": "投标时间",
        "value_type": "datetime_or_text",
        "required": True,
        "downstream": ["投标计划", "废标风险检查"],
    },
    {
        "item_key": "scoring_weight",
        "item_name": "评标标准",
        "category": "评分规则",
        "value_type": "percentage_or_text",
        "required": True,
        "downstream": ["评分细则表", "商务/技术策略"],
    },
    {
        "item_key": "construction_period",
        "item_name": "工期",
        "category": "合同/履约",
        "value_type": "duration_or_text",
        "required": True,
        "downstream": ["技术标进度计划", "报价工期成本", "合同风险"],
    },
    {
        "item_key": "payment_terms",
        "item_name": "付款方式",
        "category": "合同/报价",
        "value_type": "text",
        "required": True,
        "downstream": ["报价系统", "合同风险", "商务标响应"],
    },
    {
        "item_key": "pre_bid_clarifications",
        "item_name": "回标前优先澄清清单",
        "category": "待复核",
        "value_type": "list_or_text",
        "required": False,
        "downstream": ["待复核队列", "答疑清单", "投标计划"],
    },
)

COMMON_EVIDENCE_FIELDS: tuple[dict[str, Any], ...] = (
    {"key": "source_file", "label": "来源文件", "type": "string", "required": False, "export": True},
    {"key": "source_location", "label": "页码/章节", "type": "string", "required": False, "export": True},
    {"key": "evidence_text", "label": "原文依据", "type": "text", "required": True, "export": True},
    {"key": "confidence", "label": "置信度", "type": "number", "required": True, "export": True},
    {
        "key": "review_status",
        "label": "复核状态",
        "type": "enum",
        "enum": sorted(TENDER_ANALYSIS_REVIEW_STATUSES),
        "required": True,
        "export": True,
    },
    {"key": "review_note", "label": "复核备注", "type": "text", "required": False, "export": True},
)

TENDER_ANALYSIS_TABLE_SCHEMAS: dict[str, dict[str, Any]] = {
    "summary": {
        "table_key": "summary",
        "table_label": TENDER_ANALYSIS_TABLE_LABELS["summary"],
        "sheet_name": "结构化信息摘要表",
        "purpose": "把招标文件中的投标管理关键字段整理成可复核摘要，用于投标计划、商务标基础信息和报价/技术分流。",
        "primary_key_fields": ["item_key"],
        "source_scope": ["招标公告", "投标须知", "投标文件组成", "报价须知", "合同主要条款", "评标办法"],
        "downstream_consumers": ["项目台账", "响应矩阵", "商务标工作台", "技术标工作台", "报价系统"],
        "item_catalog": TENDER_ANALYSIS_SUMMARY_ITEM_CATALOG,
        "fields": (
            {"key": "item_key", "label": "信息项编码", "type": "string", "required": True, "export": False},
            {"key": "category", "label": "信息类别", "type": "string", "required": True, "export": True},
            {"key": "item_name", "label": "信息项", "type": "string", "required": True, "export": True},
            {"key": "extracted_value", "label": "提取值", "type": "text", "required": True, "export": True},
            {"key": "normalized_value", "label": "标准化值", "type": "string", "required": False, "export": True},
            {"key": "value_type", "label": "值类型", "type": "string", "required": True, "export": False},
            {"key": "is_required", "label": "是否关键项", "type": "boolean", "required": True, "export": True},
            {"key": "downstream", "label": "下游用途", "type": "list", "required": False, "export": True},
            *COMMON_EVIDENCE_FIELDS,
        ),
    },
    "scoring": {
        "table_key": "scoring",
        "table_label": TENDER_ANALYSIS_TABLE_LABELS["scoring"],
        "sheet_name": "评分细则表",
        "purpose": "拆解评分项、满分值和得分口径，指导商务标/技术标写作重点、资料补齐和得分差距分析。",
        "primary_key_fields": ["scoring_item_key"],
        "source_scope": ["评标办法", "评分标准", "技术标评分", "商务标评分", "报价评分"],
        "downstream_consumers": ["商务标工作台", "技术标工作台", "企业资料库检索", "投标策略复核"],
        "fields": (
            {"key": "scoring_item_key", "label": "评分项编码", "type": "string", "required": True, "export": False},
            {"key": "package_type", "label": "所属标书", "type": "enum", "enum": ["business", "technical", "pricing", "mixed", "unknown"], "required": True, "export": True},
            {"key": "scoring_item", "label": "评分项", "type": "string", "required": True, "export": True},
            {"key": "full_score", "label": "满分值", "type": "number_or_text", "required": False, "export": True},
            {"key": "scoring_weight", "label": "评分权重", "type": "percentage_or_text", "required": False, "export": True},
            {"key": "scoring_standard", "label": "评分标准说明", "type": "text", "required": True, "export": True},
            {"key": "related_bid_section", "label": "关联标书章节", "type": "string", "required": False, "export": True},
            {"key": "estimated_score", "label": "我方预估得分", "type": "number_or_text", "required": False, "export": True},
            {"key": "gap_analysis", "label": "差距分析", "type": "text", "required": False, "export": True},
            {"key": "suggested_action", "label": "建议动作", "type": "text", "required": False, "export": True},
            {"key": "owner_role", "label": "责任角色", "type": "enum", "enum": ["经营", "预算", "技术", "法务"], "required": True, "export": True},
            *COMMON_EVIDENCE_FIELDS,
        ),
    },
    "risk_clause": {
        "table_key": "risk_clause",
        "table_label": TENDER_ANALYSIS_TABLE_LABELS["risk_clause"],
        "sheet_name": "风险条款清单",
        "purpose": "把不利于施工方或影响报价/履约/废标的条款整理为风险清单，支撑法务复核、报价预留和答疑。",
        "primary_key_fields": ["risk_clause_key"],
        "source_scope": ["合同条款", "报价须知", "技术要求", "废标条款", "付款结算", "工期违约", "材料品牌"],
        "downstream_consumers": ["风险复核", "报价系统", "商务标偏离/响应", "答疑清单", "履约策划"],
        "risk_level_enum": ["high", "medium", "low"],
        "fields": (
            {"key": "risk_clause_key", "label": "风险条款编码", "type": "string", "required": True, "export": False},
            {"key": "clause_text", "label": "条款原文", "type": "text", "required": True, "export": True},
            {"key": "clause_section", "label": "所在章节", "type": "string", "required": False, "export": True},
            {"key": "risk_category", "label": "风险类型", "type": "string", "required": True, "export": True},
            {"key": "risk_level", "label": "风险等级", "type": "enum", "enum": ["high", "medium", "low"], "required": True, "export": True},
            {"key": "risk_description", "label": "风险说明", "type": "text", "required": True, "export": True},
            {"key": "suggested_response", "label": "建议应对方式", "type": "text", "required": True, "export": True},
            {"key": "owner_role", "label": "责任角色", "type": "enum", "enum": ["经营", "预算", "技术", "法务"], "required": True, "export": True},
            {"key": "need_clarification", "label": "是否需答疑", "type": "boolean", "required": True, "export": True},
            {"key": "affects_pricing", "label": "是否影响报价", "type": "boolean", "required": True, "export": True},
            {"key": "related_bid_package", "label": "关联标书/文件", "type": "enum", "enum": ["business", "technical", "pricing", "contract", "unknown"], "required": True, "export": True},
            *COMMON_EVIDENCE_FIELDS,
        ),
    },
}


def get_tender_analysis_schema() -> dict[str, Any]:
    tables = [deepcopy(TENDER_ANALYSIS_TABLE_SCHEMAS[key]) for key in ("summary", "scoring", "risk_clause")]
    for table in tables:
        fields = list(table["fields"])
        table["fields"] = fields
        table["export_fields"] = [
            {
                "key": field["key"],
                "label": field["label"],
                "type": field["type"],
                "required": bool(field.get("required")),
            }
            for field in fields
            if field.get("export")
        ]
    return {
        "schema_version": TENDER_ANALYSIS_SCHEMA_VERSION,
        "task_name": "招标文件分析",
        "task_goal": "输出结构化信息摘要表、评分细则表和风险条款清单三张可复核、可导出的成果表。",
        "review_statuses": deepcopy(TENDER_ANALYSIS_REVIEW_STATUSES),
        "business_object_policy": {
            "frontstage": "默认仅展示进入三张成果表或待复核队列的对象。",
            "backstage": "原业务对象保留为候选证据层，供规则收敛、LLM复核和下游追溯使用。",
        },
        "tables": tables,
    }


def get_tender_analysis_export_fields(table_key: str) -> list[dict[str, Any]]:
    table = TENDER_ANALYSIS_TABLE_SCHEMAS.get(table_key)
    if not table:
        raise KeyError(f"Unknown tender analysis table: {table_key}")
    return [
        {
            "key": field["key"],
            "label": field["label"],
            "type": field["type"],
            "required": bool(field.get("required")),
        }
        for field in table["fields"]
        if field.get("export")
    ]


def get_tender_analysis_table_schema(table_key: str) -> dict[str, Any]:
    table = TENDER_ANALYSIS_TABLE_SCHEMAS.get(table_key)
    if not table:
        raise KeyError(f"Unknown tender analysis table: {table_key}")
    result = deepcopy(table)
    result["fields"] = list(result["fields"])
    result["export_fields"] = get_tender_analysis_export_fields(table_key)
    return result


async def build_tender_analysis_preview_with_semantic_summary(
    db: Session,
    project: BidProject,
    run: BidParseRun,
    *,
    username: str | None = None,
    trace_id: str | None = None,
) -> dict[str, Any]:
    important_info = await _important_info_with_cache(
        db,
        project,
        run,
        username=username,
        trace_id=trace_id or run.run_uuid,
    )
    return build_tender_analysis_preview(
        db,
        project,
        run,
        important_info=important_info,
    )


async def analyze_tender_risk_clause_with_llm(
    db: Session,
    project: BidProject,
    run: BidParseRun,
    *,
    username: str | None = None,
    trace_id: str | None = None,
    force: bool = False,
) -> dict[str, Any]:
    return await _risk_clause_llm_with_cache(
        db,
        project,
        run,
        username=username,
        trace_id=trace_id or run.run_uuid,
        force=force,
    )


def get_cached_tender_risk_clause_llm(run: BidParseRun) -> dict[str, Any]:
    model = _bidding_summary_llm_model()
    cached = _cached_risk_clause_llm(run, model=model)
    if not cached:
        return _empty_risk_clause_result(
            status="not_generated",
            metadata={
                "status": "not_generated",
                "prompt_version": TENDER_ANALYSIS_RISK_CLAUSE_PROMPT_VERSION,
                "model": model,
            },
        )
    metadata = dict(cached.get("metadata") or {})
    metadata.update({"status": "cached", "model": model})
    return {
        "status": "cached",
        "metadata": metadata,
        "basic_info": cached.get("basic_info") or {},
        "priority_attention": cached.get("priority_attention") or [],
        "risks": cached.get("risks") or [],
        "overall_note": cached.get("overall_note") or "",
    }


def build_tender_analysis_preview(
    db: Session,
    project: BidProject,
    run: BidParseRun,
    *,
    summary_semantic_items: dict[str, dict[str, Any]] | None = None,
    summary_semantic_metadata: dict[str, Any] | None = None,
    important_info: dict[str, Any] | None = None,
) -> dict[str, Any]:
    requirements, risks, business_objects = _load_tender_analysis_inputs(db, run)
    summary_segment_sources = _summary_file_segment_sources(db, project, run)
    scoring_segment_sources = _collect_scoring_segment_sources(db, project, run)
    summary_items = _build_summary_items(
        project,
        requirements,
        risks,
        business_objects,
        segment_sources=summary_segment_sources,
        semantic_items=summary_semantic_items,
    )
    scoring_items = _build_scoring_items(requirements, business_objects, scoring_segment_sources)
    risk_clause_items = _build_risk_clause_items(risks)
    review_queue = _build_review_queue(summary_items, scoring_items, risk_clause_items)
    quality_summary = _build_quality_summary(summary_items, scoring_items, risk_clause_items, review_queue, run)
    semantic_metadata = dict(summary_semantic_metadata or {})
    semantic_metadata.setdefault("enabled", False)
    preview = {
        "preview_version": TENDER_ANALYSIS_PREVIEW_VERSION,
        "schema_version": TENDER_ANALYSIS_SCHEMA_VERSION,
        "project_uuid": project.project_uuid,
        "project_name": project.project_name,
        "run_uuid": run.run_uuid,
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "business_object_policy": {
            "hidden_by_default": True,
            "raw_business_object_count": len(business_objects),
            "used_as_candidate_count": _used_business_object_count(summary_items, scoring_items),
            "frontstage_note": "业务对象仅作为后台候选证据层；前台默认展示三张成果表和待复核项。",
        },
        "source_counts": {
            "requirement_count": len(requirements),
            "risk_count": len(risks),
            "business_object_count": len(business_objects),
            "summary_segment_source_count": len(summary_segment_sources),
            "scoring_segment_source_count": len(scoring_segment_sources),
            "summary_semantic_candidate_count": int(semantic_metadata.get("candidate_count") or 0),
        },
        "summary_semantic": semantic_metadata,
        "quality_summary": quality_summary,
        "review_queue": review_queue,
        "tables": {
            "summary": {
                "schema": get_tender_analysis_table_schema("summary"),
                "items": summary_items,
                "export_fields": get_tender_analysis_export_fields("summary"),
            },
            "scoring": {
                "schema": get_tender_analysis_table_schema("scoring"),
                "items": scoring_items,
                "export_fields": get_tender_analysis_export_fields("scoring"),
            },
            "risk_clause": {
                "schema": get_tender_analysis_table_schema("risk_clause"),
                "items": risk_clause_items,
                "export_fields": get_tender_analysis_export_fields("risk_clause"),
            },
        },
    }
    if important_info is not None:
        preview["important_info"] = important_info
    return preview


def _load_tender_analysis_inputs(
    db: Session,
    run: BidParseRun,
) -> tuple[list[TenderRequirement], list[TenderRisk], list[TenderBusinessObject]]:
    requirements = (
        db.query(TenderRequirement)
        .filter(TenderRequirement.parse_run_id == run.id, TenderRequirement.status == "active")
        .order_by(TenderRequirement.id.asc())
        .all()
    )
    risks = db.query(TenderRisk).filter(TenderRisk.parse_run_id == run.id).order_by(TenderRisk.id.asc()).all()
    business_objects = (
        db.query(TenderBusinessObject)
        .filter(TenderBusinessObject.parse_run_id == run.id, TenderBusinessObject.status == "active")
        .order_by(TenderBusinessObject.id.asc())
        .all()
    )
    return requirements, risks, business_objects


def build_tender_analysis_export_workbook(preview: dict[str, Any]) -> bytes:
    workbook = Workbook()
    default_sheet = workbook.active
    workbook.remove(default_sheet)
    tables = preview.get("tables") or {}
    summary_rows = _simple_summary_export_rows((tables.get("summary") or {}).get("items") or [])
    scoring_rows = _simple_scoring_export_rows((tables.get("scoring") or {}).get("items") or [])
    risk_rows = _simple_risk_clause_export_rows((tables.get("risk_clause") or {}).get("items") or [])
    review_rows = _simple_review_queue_export_rows(preview.get("review_queue") or [])
    _append_export_overview_sheet(
        workbook,
        preview,
        summary_rows=summary_rows,
        scoring_rows=scoring_rows,
        risk_rows=risk_rows,
        review_rows=review_rows,
    )
    _append_export_sheet(
        workbook,
        sheet_name="结构化信息摘要表",
        fields=(
            {"key": "category", "label": "类别"},
            {"key": "item_name", "label": "信息项"},
            {"key": "key_content", "label": "重点内容"},
            {"key": "review_label", "label": "状态"},
            {"key": "source_ref", "label": "来源"},
        ),
        rows=summary_rows,
    )
    _append_export_sheet(
        workbook,
        sheet_name="评分细则表",
        fields=(
            {"key": "package_label", "label": "所属"},
            {"key": "scoring_item", "label": "评分项"},
            {"key": "score_text", "label": "分值/权重"},
            {"key": "main_point", "label": "评分要点"},
            {"key": "owner_role", "label": "负责人"},
            {"key": "review_label", "label": "状态"},
            {"key": "source_ref", "label": "来源"},
        ),
        rows=scoring_rows,
    )
    _append_export_sheet(
        workbook,
        sheet_name="风险条款清单",
        fields=(
            {"key": "risk_level_label", "label": "等级"},
            {"key": "item_type", "label": "类型"},
            {"key": "risk_title", "label": "风险事项"},
            {"key": "clause_point", "label": "条款要点"},
            {"key": "business_impact", "label": "影响"},
            {"key": "suggested_response", "label": "建议动作"},
            {"key": "owner_role", "label": "负责人"},
            {"key": "source_ref", "label": "来源"},
        ),
        rows=risk_rows,
    )
    _append_export_sheet(
        workbook,
        sheet_name="待复核项",
        fields=(
            {"key": "title", "label": "复核事项"},
            {"key": "reason_text", "label": "原因"},
            {"key": "table_label", "label": "对应表"},
            {"key": "item_count", "label": "数量"},
            {"key": "source_ref", "label": "来源"},
        ),
        rows=review_rows,
    )
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def build_tender_analysis_export_document(preview: dict[str, Any]) -> bytes:
    if "important_info" in preview:
        important_info = preview.get("important_info") if isinstance(preview.get("important_info"), dict) else {}
        if _important_info_has_sections(important_info):
            return _build_important_info_export_document(preview, important_info)
        return _build_important_info_unavailable_document(preview, important_info)

    tables = preview.get("tables") or {}
    summary_items = list((tables.get("summary") or {}).get("items") or [])
    summary_by_key = {str(item.get("item_key") or ""): item for item in summary_items if isinstance(item, dict)}
    scoring_rows = _simple_scoring_export_rows((tables.get("scoring") or {}).get("items") or [])
    risk_rows = _simple_risk_clause_export_rows((tables.get("risk_clause") or {}).get("items") or [])
    review_rows = _simple_review_queue_export_rows(preview.get("review_queue") or [])

    doc = _TenderAnalysisDocxBuilder()
    project_name = str(preview.get("project_name") or "招标文件").strip()
    doc.add_title(project_name)
    doc.add_subtitle("投标重要有效信息提取表")
    doc.add_paragraph(
        f"资料来源：{_docx_primary_source(summary_items, scoring_rows, risk_rows)}；整理角色：投标预算员。",
        style="Meta",
    )
    doc.add_callout("重点提醒", _docx_key_reminder(summary_by_key, review_rows))

    doc.add_heading("一、项目概况")
    doc.add_kv_table(_docx_summary_kv_rows(summary_by_key.get("project_overview"), "项目概况"))

    doc.add_heading("二、答疑、踏勘与截标")
    timing_rows: list[tuple[str, str]] = []
    timing_rows.extend(_docx_summary_kv_rows(summary_by_key.get("qa_deadline"), "答疑安排"))
    timing_rows.extend(_docx_summary_kv_rows(summary_by_key.get("site_visit"), "踏勘安排"))
    timing_rows.extend(_docx_summary_kv_rows(summary_by_key.get("submission_deadline"), "截标/递交"))
    timing_rows.extend(_docx_summary_kv_rows(summary_by_key.get("contact_person"), "联系人"))
    doc.add_kv_table(_docx_dedupe_kv_rows(timing_rows) or [("答疑、踏勘与截标", "未识别明确要求，需人工复核。")])

    doc.add_heading("三、报价方式与工程承包方式")
    doc.add_kv_table(_docx_summary_kv_rows(summary_by_key.get("pricing_method"), "报价方式与工程承包方式"))

    doc.add_heading("四、保证金与担保")
    doc.add_kv_table(_docx_summary_kv_rows(summary_by_key.get("bid_bond"), "保证金与担保"))

    doc.add_heading("五、标书制作要求")
    doc.add_labeled_paragraphs(summary_by_key.get("bid_document_requirements"), fallback_label="标书制作要求")

    doc.add_heading("六、封标要求")
    doc.add_labeled_paragraphs(summary_by_key.get("sealing_requirements"), fallback_label="封标要求")

    doc.add_heading("七、评标标准与废标条件")
    scoring_summary_rows = _docx_summary_kv_rows(summary_by_key.get("scoring_weight"), "评标标准")
    if _docx_has_scoring_detail(scoring_rows):
        scoring_summary_rows.extend(
            (row.get("scoring_item") or "评分项", _docx_join_nonempty([row.get("score_text"), row.get("main_point")]))
            for row in scoring_rows
            if row.get("scoring_item") and row.get("scoring_item") != "未识别明确评分细则"
        )
    doc.add_kv_table(_docx_dedupe_kv_rows(scoring_summary_rows) or [("评标标准", "未识别明确评标标准，需人工复核。")])

    doc.add_heading("八、工期")
    doc.add_kv_table(_docx_summary_kv_rows(summary_by_key.get("construction_period"), "工期"))

    doc.add_heading("九、付款方式")
    doc.add_kv_table(_docx_summary_kv_rows(summary_by_key.get("payment_terms"), "付款方式"))

    doc.add_heading("十、回标前优先澄清清单")
    clarification_lines = _docx_clarification_lines(summary_by_key.get("pre_bid_clarifications"), review_rows)
    if clarification_lines:
        for line in clarification_lines:
            doc.add_paragraph(line, style="Body")
    else:
        doc.add_paragraph("当前未识别到明确空白、冲突或需补遗确认事项。", style="Body")

    if risk_rows:
        doc.add_heading("十一、风险条款清单")
        doc.add_table(
            [["等级", "类型", "风险事项", "条款要点", "建议动作"]]
            + [
                [
                    row.get("risk_level_label") or "-",
                    row.get("item_type") or "-",
                    row.get("risk_title") or "-",
                    row.get("clause_point") or "-",
                    row.get("suggested_response") or "-",
                ]
                for row in risk_rows
            ],
            widths=(900, 900, 1900, 3500, 2160),
            header=True,
        )

    return doc.to_bytes()


def build_tender_risk_clause_export_document(project: BidProject, risk_clause: dict[str, Any]) -> bytes:
    if not _risk_clause_llm_has_risks(risk_clause):
        return _build_risk_clause_unavailable_document(project, risk_clause)
    return _build_risk_clause_llm_export_document(project, risk_clause)


def _risk_clause_llm_has_risks(risk_clause: dict[str, Any]) -> bool:
    risks = risk_clause.get("risks") if isinstance(risk_clause, dict) else None
    return isinstance(risks, list) and any(isinstance(risk, dict) for risk in risks)


def _build_risk_clause_llm_export_document(project: BidProject, risk_clause: dict[str, Any]) -> bytes:
    doc = _TenderAnalysisDocxBuilder()
    project_name = str(project.project_name or "招标文件").strip()
    basic_info = risk_clause.get("basic_info") if isinstance(risk_clause.get("basic_info"), dict) else {}
    risks = [risk for risk in risk_clause.get("risks") or [] if isinstance(risk, dict)]
    priority_attention = [item for item in risk_clause.get("priority_attention") or [] if isinstance(item, dict)]

    doc.add_title("合同风险评审")
    doc.add_subtitle(f"{project_name}风险条款清单")
    doc.add_paragraph("依据招标文件提取：条款原文、所在章节、风险等级、风险说明、建议应对方式", style="Meta")
    doc.add_kv_table(
        [
            ("依据文件", str(basic_info.get("source_files") or "招标文件原文片段")),
            ("适用场景", str(basic_info.get("applicable_scenario") or "投标评审、合同谈判、项目履约交底、签证与索赔管控")),
            ("风险分布", str(basic_info.get("risk_distribution") or _risk_clause_distribution_text(risks))),
            ("生成说明", str(basic_info.get("generation_note") or "风险等级为商务/履约初评，建议由法务、商务、项目经理共同复核。")),
        ]
    )

    doc.add_heading("一、优先关注事项")
    doc.add_table(
        [["类别", "建议"]]
        + [
            [item.get("category") or "-", item.get("suggestion") or "-"]
            for item in priority_attention
        ],
        widths=(1700, 7660),
        header=True,
    )

    doc.add_heading("二、风险清单概览")
    doc.add_table(
        [["序号", "等级", "所在章节", "风险说明"]]
        + [
            [
                str(risk.get("risk_id") or f"R-{index:02d}").replace("R-", ""),
                _risk_level_export_label(risk.get("risk_level")),
                risk.get("source_location") or "-",
                risk.get("risk_explanation") or "-",
            ]
            for index, risk in enumerate(risks, start=1)
        ],
        widths=(700, 800, 2300, 5560),
        header=True,
    )

    doc.add_heading("三、风险条款明细")
    for risk in risks:
        doc.add_paragraph(
            f"{risk.get('risk_id') or '-'}｜{_risk_level_export_label(risk.get('risk_level'))}｜{risk.get('source_location') or '-'}",
            style="Heading2",
        )
        doc.add_kv_table(
            [
                ("所在章节", risk.get("source_location") or "-"),
                ("风险等级", _risk_level_export_label(risk.get("risk_level"))),
                ("条款原文", risk.get("clause_original") or "-"),
                ("风险说明", risk.get("risk_explanation") or "-"),
                ("建议应对方式", risk.get("suggested_response") or "-"),
            ]
        )
    overall_note = _clean_llm_business_value(risk_clause.get("overall_note"), limit=None)
    if overall_note:
        doc.add_heading("整理说明")
        doc.add_paragraph(overall_note, style="Body")
    return doc.to_bytes()


def _build_risk_clause_unavailable_document(project: BidProject, risk_clause: dict[str, Any]) -> bytes:
    doc = _TenderAnalysisDocxBuilder()
    project_name = str(project.project_name or "招标文件").strip()
    metadata = risk_clause.get("metadata") if isinstance(risk_clause, dict) else {}
    status_text = str((metadata or {}).get("status") or risk_clause.get("status") or "not_generated")
    reason = str((metadata or {}).get("skip_reason") or (metadata or {}).get("error") or "LLM风险条款分析结果为空").strip()
    doc.add_title("合同风险评审")
    doc.add_subtitle(f"{project_name}风险条款清单")
    doc.add_callout("未生成", f"LLM风险条款清单未生成，状态：{status_text}。原因：{reason}。请先在系统中点击“风险分析”。")
    return doc.to_bytes()


def _risk_level_export_label(value: Any) -> str:
    return RISK_LEVEL_LABELS.get(str(value or ""), str(value or "-"))


def _important_info_has_sections(important_info: dict[str, Any]) -> bool:
    sections = important_info.get("sections") if isinstance(important_info, dict) else None
    return bool(isinstance(sections, list) and any(isinstance(section, dict) and section.get("items") for section in sections))


def _build_important_info_export_document(preview: dict[str, Any], important_info: dict[str, Any]) -> bytes:
    doc = _TenderAnalysisDocxBuilder()
    project_name = str(preview.get("project_name") or "招标文件").strip()
    doc.add_title(project_name)
    doc.add_subtitle("投标重要有效信息提取表")
    doc.add_paragraph(
        f"资料来源：{_important_info_primary_source(important_info)}；整理方式：LLM依据招标文件原文证据整理；整理角色：投标预算员。",
        style="Meta",
    )
    reminder = "；".join(_important_info_clarification_lines(important_info)[:8])
    doc.add_callout("重点提醒", reminder or "当前未识别到明确空白、冲突或需补遗确认事项。")

    section_by_key = {
        str(section.get("section_key") or ""): section
        for section in important_info.get("sections") or []
        if isinstance(section, dict)
    }
    for section_spec in IMPORTANT_INFO_SECTION_SCHEMA:
        section_key = str(section_spec["section_key"])
        section = section_by_key.get(section_key) or {}
        item_by_key = {
            str(item.get("field_key") or ""): item
            for item in section.get("items") or []
            if isinstance(item, dict)
        }
        doc.add_heading(str(section_spec["title"]))
        if section_key == "pre_bid_clarifications":
            lines = _important_info_clarification_lines(important_info)
            if lines:
                for line in lines:
                    doc.add_paragraph(line, style="Body")
            else:
                doc.add_paragraph("当前未识别到明确空白、冲突或需补遗确认事项。", style="Body")
            continue
        rows = [
            (field_name, _important_info_item_display_value(item_by_key.get(field_key)))
            for field_key, field_name in section_spec["fields"]
        ]
        doc.add_kv_table(rows)

    overall_note = _clean_llm_business_value(important_info.get("overall_note"), limit=None)
    if overall_note:
        doc.add_heading("整理说明")
        doc.add_paragraph(overall_note, style="Body")
    return doc.to_bytes()


def _build_important_info_unavailable_document(preview: dict[str, Any], important_info: dict[str, Any]) -> bytes:
    doc = _TenderAnalysisDocxBuilder()
    project_name = str(preview.get("project_name") or "招标文件").strip()
    metadata = important_info.get("metadata") if isinstance(important_info, dict) else {}
    status_text = str((metadata or {}).get("status") or important_info.get("status") or "not_generated")
    reason = str((metadata or {}).get("skip_reason") or (metadata or {}).get("error") or "LLM结构化提取结果为空").strip()
    doc.add_title(project_name)
    doc.add_subtitle("投标重要有效信息提取表")
    doc.add_callout("未生成", f"LLM结构化提取未生成，状态：{status_text}。原因：{reason}。本次未使用旧规则兜底生成 Word，避免输出答非所问内容。")
    doc.add_paragraph("请启用招标文件 LLM 识别后重新导出，或在系统内补充模型配置后重新生成。", style="Body")
    return doc.to_bytes()


def _important_info_item_display_value(item: dict[str, Any] | None) -> str:
    if not isinstance(item, dict):
        return "未明确，需人工复核。"
    value = _clean_llm_business_value(item.get("value"), limit=None)
    status = str(item.get("status") or "").lower()
    note = _clean_llm_business_value(item.get("note"), limit=None)
    if value:
        return value
    if status == "unclear":
        return note or "原文未填写、存在矛盾或需补遗确认。"
    if status == "not_found":
        return note or "未识别到明确原文依据，需人工复核。"
    return note or "未明确，需人工复核。"


def _important_info_primary_source(important_info: dict[str, Any]) -> str:
    sources: list[str] = []
    for section in important_info.get("sections") or []:
        if not isinstance(section, dict):
            continue
        for item in section.get("items") or []:
            if not isinstance(item, dict):
                continue
            sources.extend(_clean_string_list(item.get("source_file")))
    return "；".join(_unique_text(sources)[:3]) or "招标文件原文片段"


def _important_info_clarification_lines(important_info: dict[str, Any]) -> list[str]:
    lines: list[str] = []
    for raw in important_info.get("priority_clarifications") or []:
        if not isinstance(raw, dict):
            continue
        item = _clean_llm_business_value(raw.get("item"), limit=None)
        reason = _clean_llm_business_value(raw.get("reason"), limit=None)
        if item and reason:
            lines.append(f"{item}：{reason}")
        elif item:
            lines.append(item)
        elif reason:
            lines.append(reason)

    section_by_key = {
        str(section.get("section_key") or ""): section
        for section in important_info.get("sections") or []
        if isinstance(section, dict)
    }
    clarification_section = section_by_key.get("pre_bid_clarifications") or {}
    for item in clarification_section.get("items") or []:
        if not isinstance(item, dict):
            continue
        value = _clean_llm_business_value(item.get("value"), limit=None)
        if value:
            lines.extend(part.strip(" ；;。") for part in re.split(r"[；;\n]+|(?<=。)", value) if part.strip(" ；;。"))

    if lines:
        return _unique_text(lines)

    for section in important_info.get("sections") or []:
        if not isinstance(section, dict):
            continue
        for item in section.get("items") or []:
            if not isinstance(item, dict):
                continue
            text = _important_info_item_display_value(item)
            if item.get("status") == "unclear" or _has_any(text, ("未填写", "空白", "不一致", "冲突", "需澄清", "未明确")):
                lines.append(f"{item.get('field_name') or '需澄清项'}：{text}")
    return _unique_text(lines)


def _append_export_overview_sheet(
    workbook: Workbook,
    preview: dict[str, Any],
    *,
    summary_rows: list[dict[str, Any]],
    scoring_rows: list[dict[str, Any]],
    risk_rows: list[dict[str, Any]],
    review_rows: list[dict[str, Any]],
) -> None:
    quality = preview.get("quality_summary") or {}
    summary_total = int(quality.get("summary_item_count") or len(summary_rows) or 0)
    summary_extracted = int(quality.get("summary_extracted_count") or 0)
    high_risk_count = len([row for row in risk_rows if row.get("risk_level") == "high"])
    overview_rows = [
        {"item": "项目名称", "value": preview.get("project_name") or "-"},
        {"item": "导出时间", "value": preview.get("generated_at") or datetime.now().isoformat(timespec="seconds")},
        {"item": "解析版本", "value": preview.get("run_uuid") or "-"},
        {"item": "摘要信息", "value": f"已识别 {summary_extracted}/{summary_total} 项"},
        {"item": "评分细则", "value": f"{len(scoring_rows)} 项"},
        {"item": "风险条款", "value": f"{len(risk_rows)} 项，其中高风险 {high_risk_count} 项"},
        {"item": "待复核项", "value": f"{len(review_rows)} 项"},
        {"item": "交付说明", "value": "本成果包仅展示重点信息；详细原文、置信度和后台对象请在系统页面复核。"},
    ]
    _append_export_sheet(
        workbook,
        sheet_name="成果概览",
        fields=(
            {"key": "item", "label": "项目"},
            {"key": "value", "label": "内容"},
        ),
        rows=overview_rows,
    )


def _simple_summary_export_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    simple_rows: list[dict[str, Any]] = []
    for row in rows:
        content = str(row.get("normalized_value") or row.get("extracted_value") or "").strip()
        is_required = bool(row.get("is_required"))
        if not content and not is_required:
            continue
        review_label = "待补充" if is_required and not content else _review_status_export_label(row.get("review_status"))
        simple_rows.append(
            {
                "category": row.get("category") or "-",
                "item_name": row.get("item_name") or "-",
                "key_content": _clip(content or "未识别", 180),
                "review_label": review_label,
                "source_ref": _source_ref(row),
                "_tone": "warning" if review_label in {"待补充", "需修正", "待复核"} else "",
            }
        )
    return simple_rows


def _simple_scoring_export_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    leaf_rows = _leaf_table_rows(rows)
    if not leaf_rows:
        return [
            {
                "package_label": "待确认",
                "scoring_item": "未识别明确评分细则",
                "score_text": "-",
                "main_point": "当前解析结果未发现明确的评分项、分值或权重，需人工确认是否存在单独评分表或补遗文件。",
                "owner_role": "经营",
                "review_label": "待补充",
                "source_ref": "-",
                "_tone": "warning",
            }
        ]
    simple_rows: list[dict[str, Any]] = []
    for row in leaf_rows:
        score_text = _score_text(row)
        simple_rows.append(
            {
                "package_label": PACKAGE_TYPE_LABELS.get(str(row.get("package_type") or "unknown"), "待确认"),
                "scoring_item": row.get("scoring_item") or "评分项",
                "score_text": score_text or "-",
                "main_point": _clip(row.get("scoring_standard") or row.get("evidence_text") or "", 180),
                "owner_role": row.get("owner_role") or "-",
                "review_label": _review_status_export_label(row.get("review_status")),
                "source_ref": _source_ref(row),
                "_tone": "warning" if row.get("review_status") != "confirmed" else "",
            }
        )
    return simple_rows


def _simple_risk_clause_export_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    blocks: list[dict[str, Any]] = []
    for row in rows:
        parent_row = _simple_risk_clause_export_row(
            row,
            item_type="汇总" if row.get("children") else "单项",
        )
        child_rows = [
            _simple_risk_clause_export_row(child, item_type="细分")
            for child in row.get("children") or []
            if isinstance(child, dict)
        ]
        blocks.append(
            {
                "sort_level": -RISK_LEVEL_ORDER.get(str(parent_row.get("risk_level")), 2),
                "sort_title": str(parent_row.get("risk_title") or ""),
                "rows": [parent_row, *child_rows],
            }
        )
    simple_rows: list[dict[str, Any]] = []
    for block in sorted(blocks, key=lambda item: (item["sort_level"], item["sort_title"])):
        simple_rows.extend(block["rows"])
    return simple_rows


def _simple_risk_clause_export_row(row: dict[str, Any], *, item_type: str) -> dict[str, Any]:
    impact_flags = []
    if row.get("affects_pricing"):
        impact_flags.append("影响报价")
    if row.get("need_clarification"):
        impact_flags.append("建议答疑")
    if item_type == "汇总" and row.get("risk_count") and int(row.get("risk_count") or 0) > 1:
        impact_flags.append(f"已合并 {row.get('risk_count')} 条")
    risk_level = str(row.get("risk_level") or "medium")
    risk_title = row.get("risk_title") or RISK_TYPE_LABELS.get(
        str(row.get("risk_category") or ""),
        row.get("risk_category") or "风险事项",
    )
    if item_type == "细分":
        risk_title = f"  - {risk_title}"
    return {
        "risk_level": risk_level,
        "risk_level_label": RISK_LEVEL_LABELS.get(risk_level, risk_level),
        "item_type": item_type,
        "risk_title": risk_title,
        "clause_point": _readable_risk_clause_point(row),
        "business_impact": "；".join(impact_flags) or "需复核",
        "suggested_response": _clip(row.get("suggested_response") or "", 180),
        "owner_role": row.get("owner_role") or "-",
        "source_ref": _source_ref(row),
        "_tone": "danger" if risk_level == "high" else "warning" if risk_level == "medium" else "",
    }


def _readable_risk_clause_point(row: dict[str, Any]) -> str:
    raw_text = str(row.get("clause_text") or row.get("evidence_text") or "").strip()
    if not raw_text:
        return "-"
    risk_text = " ".join(
        str(item or "")
        for item in [row.get("risk_title"), row.get("risk_category"), row.get("risk_description")]
    )
    keywords = _risk_point_keywords(risk_text)
    clauses = _docx_split_business_clauses(raw_text)
    useful_clauses = [
        clause
        for clause in clauses
        if len(clause) >= 8 and not _risk_clause_is_low_value_noise(clause, risk_text)
    ]
    for clause in useful_clauses:
        if _has_any(clause, keywords):
            return _docx_clean_business_value(clause, limit=150)
    if useful_clauses:
        return _docx_clean_business_value(useful_clauses[0], limit=150)
    return _docx_clean_business_value(raw_text, limit=150)


def _risk_point_keywords(risk_text: str) -> tuple[str, ...]:
    text = str(risk_text or "")
    keywords = ["不予调整", "不作调整", "包干", "漏项", "不另行支付", "无预付款", "付款", "质保金", "违约金", "延误", "索赔", "签证", "密封", "废标", "否决", "材料", "品牌", "图纸", "现场", "保函"]
    if _has_any(text, ("付款", "垫资", "预付款")):
        keywords.extend(["无预付款", "进度款", "支付", "付款"])
    if _has_any(text, ("工期", "延误")):
        keywords.extend(["延误", "违约金", "日历天"])
    if _has_any(text, ("固定", "包干", "单价")):
        keywords.extend(["包干", "不予调整", "不作调整", "单价"])
    if _has_any(text, ("废标", "无效", "否决")):
        keywords.extend(["废标", "无效标", "否决投标"])
    return tuple(dict.fromkeys(keywords))


def _risk_clause_is_low_value_noise(clause: str, risk_text: str) -> bool:
    if _has_any(clause, ("支持性表格", "常用工程用表", "格式：", "工程竣工结算资料清单")) and not _has_any(risk_text, ("资料", "签证")):
        return True
    if _has_any(clause, ("术语和定义", "指各承建商", "指承建商", "工程进度款：指", "工程结算款：指", "资金计划：指")) and not _has_any(
        clause,
        ("无预付款", "支付至", "审核值", "15个工作日", "质保金", "违约金", "废标", "不予调整", "不另行支付"),
    ):
        return True
    if len(re.sub(r"\s+", "", clause)) < 8:
        return True
    return False


def _simple_review_queue_export_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    simple_rows: list[dict[str, Any]] = []
    for row in rows:
        simple_rows.append(
            {
                "title": row.get("title") or "待复核事项",
                "reason_text": _excel_cell_value(row.get("reasons")) or _review_status_export_label(row.get("review_status")),
                "table_label": row.get("table_label") or "-",
                "item_count": row.get("item_count") or 1,
                "source_ref": _source_ref(row),
                "_tone": "warning",
            }
        )
    return simple_rows


def _score_text(row: dict[str, Any]) -> str:
    parts = []
    if row.get("full_score"):
        parts.append(str(row.get("full_score")))
    if row.get("scoring_weight"):
        parts.append(f"权重 {row.get('scoring_weight')}")
    return " / ".join(parts)


def _source_ref(row: dict[str, Any]) -> str:
    source_file = str(row.get("source_file") or "").strip()
    source_location = str(row.get("source_location") or "").strip()
    if source_file and source_location:
        return f"{source_file} / {source_location}"
    return source_file or source_location or "-"


def _review_status_export_label(status: Any) -> str:
    value = str(status or "pending")
    return TENDER_ANALYSIS_REVIEW_STATUSES.get(value, value)


def _append_export_sheet(
    workbook: Workbook,
    *,
    sheet_name: str,
    fields: Any,
    rows: list[dict[str, Any]],
) -> None:
    sheet = workbook.create_sheet(title=sheet_name[:31])
    field_list = [dict(field) for field in fields]
    header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin_fill = PatternFill(fill_type="solid", fgColor="F8FAFC")
    tone_fills = {
        "danger": PatternFill(fill_type="solid", fgColor="FEE2E2"),
        "warning": PatternFill(fill_type="solid", fgColor="FEF3C7"),
    }
    for column_index, field in enumerate(field_list, start=1):
        cell = sheet.cell(row=1, column=column_index, value=field.get("label") or field.get("key"))
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row_index, row in enumerate(rows, start=2):
        row_fill = tone_fills.get(str(row.get("_tone") or ""))
        for column_index, field in enumerate(field_list, start=1):
            value = _excel_cell_value(row.get(field.get("key")))
            cell = sheet.cell(row=row_index, column=column_index, value=value)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if row_fill:
                cell.fill = row_fill
            elif row_index % 2 == 0:
                cell.fill = thin_fill
            if field.get("key") == "confidence" and isinstance(value, (int, float)):
                cell.number_format = "0.00"
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for column_index, field in enumerate(field_list, start=1):
        key = str(field.get("key") or "")
        label = str(field.get("label") or key)
        values = [label]
        for row in rows[:80]:
            values.append(str(_excel_cell_value(row.get(key)) or ""))
        width = max(len(item) for item in values) + 3 if values else 12
        if key in {"key_content", "main_point", "clause_point", "suggested_response", "source_ref"}:
            width = min(max(width, 28), 48)
        else:
            width = min(max(width, 10), 32)
        sheet.column_dimensions[get_column_letter(column_index)].width = width
    for row_index in range(1, min(len(rows) + 1, 200) + 1):
        sheet.row_dimensions[row_index].height = 24 if row_index == 1 else 42


def _excel_cell_value(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "是" if value else "否"
    if isinstance(value, (int, float)):
        return value
    if isinstance(value, list):
        return "；".join(str(item) for item in value if item is not None)
    if isinstance(value, dict):
        return "；".join(f"{key}:{item}" for key, item in value.items() if item is not None)
    return str(value)


class _TenderAnalysisDocxBuilder:
    def __init__(self) -> None:
        self.blocks: list[str] = []
        self.media: list[dict[str, Any]] = []
        self._image_counter = 0
        self.header_text = ""
        self.footer_text = ""

    def set_page_header_footer(self, *, header_text: Any = "", footer_text: Any = "") -> None:
        self.header_text = str(header_text or "").strip()
        self.footer_text = str(footer_text or "").strip()

    def add_title(self, text: str) -> None:
        self.blocks.append(_docx_paragraph(text, style="DocTitle"))

    def add_subtitle(self, text: str) -> None:
        self.blocks.append(_docx_paragraph(text, style="DocSubtitle"))

    def add_heading(self, text: str, *, level: int = 1, page_break_before: bool = False) -> None:
        style = "Heading1" if level <= 1 else "Heading2"
        self.blocks.append(_docx_paragraph(text, style=style, page_break_before=page_break_before))

    def add_toc_field(self, *, levels: str = "1-1", fallback_lines: list[str] | None = None) -> None:
        self.blocks.append(_docx_toc_field(levels=levels, fallback_lines=fallback_lines))

    def add_paragraph(self, text: Any, *, style: str = "Body") -> None:
        content = str(text or "").strip()
        if content:
            self.blocks.append(_docx_paragraph(content, style=style))

    def add_page_break(self) -> None:
        self.blocks.append('<w:p><w:r><w:br w:type="page"/></w:r></w:p>')

    def add_callout(self, title: str, body: str) -> None:
        paragraphs = [
            _docx_paragraph(title, style="CalloutTitle", in_table=True),
            _docx_paragraph(body or "无", style="CalloutBody", in_table=True),
        ]
        self.blocks.append(
            _docx_table(
                [[paragraphs]],
                widths=(9360,),
                header=False,
                first_col_shading=False,
                cell_shading="FFF7D6",
                border_color="D6B656",
            )
        )

    def add_labeled_paragraphs(self, item: dict[str, Any] | None, *, fallback_label: str) -> None:
        rows = _docx_summary_kv_rows(item, fallback_label)
        for label, value in rows:
            text = f"{label}：{value}" if label != fallback_label or len(rows) > 1 else value
            self.add_paragraph(text, style="Body")

    def add_kv_table(self, rows: list[tuple[str, str]]) -> None:
        self.blocks.append(_docx_table([[label, value] for label, value in rows], widths=(2500, 6860), first_col_shading=True))

    def add_table(self, rows: list[list[Any]], *, widths: tuple[int, ...], header: bool = False) -> None:
        self.blocks.append(_docx_table(rows, widths=widths, header=header, first_col_shading=False))

    def add_image(
        self,
        image_bytes: bytes,
        *,
        extension: str,
        filename: str,
        width_px: int | None = None,
        height_px: int | None = None,
        max_width_emu: int = 5_943_600,
        max_height_emu: int = 8_229_600,
    ) -> bool:
        ext = _docx_image_extension(extension)
        if not image_bytes or ext not in {"png", "jpg", "jpeg"}:
            return False
        self._image_counter += 1
        rel_id = f"rIdImage{self._image_counter}"
        target = f"media/image{self._image_counter}.{ext}"
        cx, cy = _docx_image_extent(
            width_px=width_px,
            height_px=height_px,
            max_width_emu=max_width_emu,
            max_height_emu=max_height_emu,
        )
        self.media.append(
            {
                "rel_id": rel_id,
                "target": target,
                "content_type": _docx_image_content_type(ext),
                "content": image_bytes,
            }
        )
        self.blocks.append(_docx_image_paragraph(rel_id=rel_id, filename=filename, cx=cx, cy=cy, doc_pr_id=self._image_counter))
        return True

    def to_bytes(self) -> bytes:
        body = "".join(self.blocks)
        section_references = ""
        header_xml = _docx_header_xml(self.header_text) if self.header_text else ""
        footer_xml = _docx_footer_xml(self.footer_text) if self.footer_text else ""
        if header_xml:
            section_references += '<w:headerReference w:type="default" r:id="rIdHeader1"/>'
        if footer_xml:
            section_references += '<w:footerReference w:type="default" r:id="rIdFooter1"/>'
        section = (
            '<w:sectPr>'
            f"{section_references}"
            '<w:pgSz w:w="12240" w:h="15840"/>'
            '<w:pgMar w:top="1440" w:right="1440" w:bottom="1440" w:left="1440" w:header="708" w:footer="708" w:gutter="0"/>'
            "</w:sectPr>"
        )
        document_xml = (
            '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
            '<w:document xmlns:w="http://schemas.openxmlformats.org/wordprocessingml/2006/main" '
            'xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships" '
            'xmlns:wp="http://schemas.openxmlformats.org/drawingml/2006/wordprocessingDrawing" '
            'xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main" '
            'xmlns:pic="http://schemas.openxmlformats.org/drawingml/2006/picture">'
            f"<w:body>{body}{section}</w:body>"
            "</w:document>"
        )
        return _docx_package(document_xml, media=self.media, header_xml=header_xml, footer_xml=footer_xml)


def _docx_primary_source(*row_groups: Any) -> str:
    sources: list[str] = []
    for rows in row_groups:
        if not isinstance(rows, list):
            continue
        for row in rows:
            if not isinstance(row, dict):
                continue
            source = str(row.get("source_file") or row.get("source_ref") or "").strip()
            if not source or source == "-":
                continue
            source = source.split("/", 1)[0].strip()
            sources.append(source)
    return _join_unique(sources, limit=2) or "招标文件"


def _docx_key_reminder(summary_by_key: dict[str, dict[str, Any]], review_rows: list[dict[str, Any]]) -> str:
    clarification = _docx_summary_value(summary_by_key.get("pre_bid_clarifications"))
    if clarification:
        return clarification
    review_lines = [
        str(row.get("title") or "").strip()
        for row in review_rows
        if _has_any(str(row.get("reason_text") or row.get("title") or ""), ("需答疑", "待补充", "缺失", "空白", "冲突"))
    ][:5]
    if review_lines:
        return "；".join(review_lines)
    return "当前未识别到明确空白、冲突或需补遗确认事项。"


def _docx_summary_value(item: dict[str, Any] | None) -> str:
    if not isinstance(item, dict):
        return ""
    return _clip(str(item.get("extracted_value") or item.get("normalized_value") or "").strip(), 1200)


def _docx_summary_kv_rows(item: dict[str, Any] | None, fallback_label: str) -> list[tuple[str, str]]:
    content = _docx_summary_value(item)
    if not content:
        if item and item.get("is_required"):
            return [(fallback_label, "未识别明确内容，需人工复核。")]
        return []
    business_rows = _docx_business_summary_rows(item, content, fallback_label=fallback_label)
    if business_rows:
        return business_rows
    rows = _docx_labeled_rows_from_text(content)
    if rows:
        return rows
    return [(fallback_label, content)]


def _docx_business_summary_rows(item: dict[str, Any] | None, content: str, *, fallback_label: str) -> list[tuple[str, str]]:
    item_key = str((item or {}).get("item_key") or "")
    if item_key == "project_overview":
        return _docx_project_overview_rows(content)
    if item_key == "pricing_method":
        return _docx_pricing_rows(content)
    if item_key == "bid_bond":
        return _docx_bid_bond_rows(content)
    if item_key == "bid_document_requirements":
        return _docx_document_requirement_rows(content)
    if item_key == "sealing_requirements":
        return _docx_sealing_rows(content)
    if item_key == "scoring_weight":
        return _docx_scoring_summary_rows(content)
    if item_key == "construction_period":
        return _docx_construction_period_rows(content)
    if item_key == "payment_terms":
        rows = _docx_payment_rows(content)
        if rows:
            return rows
        if _docx_is_definition_like_payment_text(content):
            return [(fallback_label, "未识别到明确付款比例、支付节点或质保金条款，需人工复核合同价款支付章节。")]
    return []


def _docx_project_overview_rows(content: str) -> list[tuple[str, str]]:
    label_rows = _docx_labeled_rows_from_text(content)
    rows: list[tuple[str, str]] = []
    for label, aliases in (
        ("招标工程名称", ("招标工程名称", "工程名称")),
        ("工程地点", ("工程地点", "建设地点", "项目地点")),
        ("招标单位", ("招标单位", "招标人", "建设单位")),
        ("总承包单位", ("总承包单位", "总承包", "总包单位")),
        ("工程规模及特征", ("工程规模及特征", "工程规模", "工程特征")),
    ):
        value = _docx_value_for_labels(label_rows, aliases)
        if value and not _docx_is_placeholder_value(value):
            rows.append((label, _docx_clean_business_value(value, limit=260)))
    scope_value = _docx_value_for_labels(label_rows, ("招标范围", "工程范围", "施工范围", "承包范围"))
    scope_clauses = _docx_matching_clauses(
        content,
        ("商业街区", "6#楼", "32F", "32层", "办公区", "公共走道", "电梯厅", "会议室", "外摆区", "商铺"),
        reject=("具体信息未提供", "详见《协议书》", "详见《技术要求》"),
        limit=3,
    )
    scope_parts = []
    if scope_value and not _docx_is_placeholder_value(scope_value):
        scope_parts.append(scope_value)
    scope_parts.extend(scope_clauses)
    scope_text = _join_unique([_docx_clean_business_value(part, limit=260) for part in scope_parts], limit=3)
    if scope_text:
        rows.append(("招标范围", scope_text))
    return _docx_dedupe_kv_rows(rows)


def _docx_pricing_rows(content: str) -> list[tuple[str, str]]:
    label_rows = _docx_labeled_rows_from_text(content)
    rows = []
    for label, aliases, keywords in (
        ("承包方式", ("工程承包方式", "承包方式", "合同形式"), ("综合单价包干", "固定单价", "总价包干", "包干合同")),
        ("综合单价包含范围", ("综合单价包含", "综合单价包含范围"), ("人工", "材料", "税金", "二次运输", "夜间施工", "检测试验", "临时设施")),
        ("工程量调整/计量规则", ("工程量规则", "工程量调整", "计量规则"), ("暂定", "竣工图", "重新量度", "单价不调整")),
        ("包干项目口径", ("包干项目", "包干项目口径"), ("项", "包干", "结算不调整")),
        ("漏报风险", ("报价风险", "漏报风险"), ("漏填", "未填", "视为已含", "不另行支付")),
    ):
        value = _docx_value_for_labels(label_rows, aliases) or _docx_first_clause(content, keywords)
        if value:
            rows.append((label, _docx_clean_business_value(value, limit=240)))
    return _docx_dedupe_kv_rows(rows)


def _docx_bid_bond_rows(content: str) -> list[tuple[str, str]]:
    label_rows = _docx_labeled_rows_from_text(content)
    rows = []
    for label, aliases, keywords in (
        ("投标担保金额", ("投标担保金额", "投标保证金", "保证金与担保"), ("RMB", "人民币", "元", "保函")),
        ("缴交时间", ("缴交时间", "缴纳时间", "提交时间", "递交时间"), ("递交投标文件", "同时提交", "提交")),
        ("缴交形式", ("形式", "缴交形式", "缴纳形式", "提交形式"), ("现金", "支票", "银行保函", "保函")),
        ("收款账户", ("收款账户", "账户", "户名", "账号"), ("开户", "账户", "账号", "户名")),
        ("保函要求", ("保函要求", "保函有效期", "原件要求"), ("有效期", "原件", "投标有效期")),
        ("未提交后果", ("未提交后果", "不提交后果"), ("不予受理", "无效标", "废标")),
        ("没收情形", ("没收情形",), ("撤回投标", "拒签合同", "不提交履约担保", "没收")),
        ("履约担保", ("履约担保", "履约担保比例"), ("10%", "中标价", "合同价款")),
    ):
        value = _docx_value_for_labels(label_rows, aliases) or _docx_first_clause(content, keywords)
        if value:
            rows.append((label, _docx_clean_business_value(value, limit=260)))
    return _docx_dedupe_kv_rows(rows)


def _docx_document_requirement_rows(content: str) -> list[tuple[str, str]]:
    label_rows = _docx_labeled_rows_from_text(content)
    rows = []
    for label, aliases, keywords in (
        ("投标人资质", ("投标人资质", "资质要求"), ("建筑装修装饰", "专业承包", "资质")),
        ("投标文件份数", ("投标文件份数", "份数"), ("正本", "副本", "电子标书", "U盘")),
        ("商务标内容", ("商务标内容", "商务标"), ("工程量清单", "投标书", "保函", "澄清问卷")),
        ("技术标内容", ("技术标内容", "技术标"), ("营业执照", "施工组织设计", "项目班子", "进度计划", "材料品牌")),
        ("格式/装订要求", ("格式要求", "装订"), ("提供格式", "装订成册", "不得活页")),
        ("签署盖章要求", ("签署要求", "签署盖章要求"), ("签字", "盖章", "授权委托书")),
        ("报价填报要求", ("报价要求", "报价填报要求"), ("逐项填报", "不得选择性报价", "漏填")),
    ):
        value = _docx_value_for_labels(label_rows, aliases) or _docx_first_clause(content, keywords)
        if value:
            rows.append((label, _docx_clean_business_value(value, limit=360)))
    return _docx_dedupe_kv_rows(rows)


def _docx_sealing_rows(content: str) -> list[tuple[str, str]]:
    label_rows = _docx_labeled_rows_from_text(content)
    rows = []
    for label, aliases, keywords in (
        ("分别密封", ("分别密封", "密封袋拆分"), ("商务标", "技术标", "电子标书", "密封袋")),
        ("密封袋标识", ("标识", "密封袋标识", "标记"), ("商务标部分", "技术标部分", "工程名称", "投标人名称")),
        ("外袋组织", ("外袋组织", "外密封袋"), ("商务电子", "技术电子", "统一装入", "外密封袋")),
        ("封口要求", ("封口", "封口要求"), ("封条", "骑缝", "公章")),
    ):
        value = _docx_value_for_labels(label_rows, aliases) or _docx_first_clause(content, keywords)
        if value:
            rows.append((label, _docx_clean_business_value(value, limit=260)))
    return _docx_dedupe_kv_rows(rows)


def _docx_scoring_summary_rows(content: str) -> list[tuple[str, str]]:
    label_rows = _docx_labeled_rows_from_text(content)
    rows = []
    for label, aliases, keywords in (
        ("评标原则", ("评标原则", "评标标准", "评标办法"), ("综合评标", "综合评定", "评标法")),
        ("确定中标人", ("确定中标人", "定标原则"), ("自行确定", "不承诺最低价", "不中标")),
        ("投标人数不足处理", ("投标人数不足", "投标人数不足处理"), ("少于3", "暂停开标", "招标失败")),
        ("废标条件", ("废标条件", "无效标", "否决投标"), ("未按要求密封", "未签字", "授权委托书", "逾期", "串标")),
    ):
        value = _docx_value_for_labels(label_rows, aliases) or _docx_first_clause(content, keywords)
        if value:
            rows.append((label, _docx_clean_business_value(value, limit=360)))
    return _docx_dedupe_kv_rows(rows)


def _docx_construction_period_rows(content: str) -> list[tuple[str, str]]:
    label_rows = _docx_labeled_rows_from_text(content)
    rows = []
    for label, aliases, keywords in (
        ("商业街区暂定开工", ("商业街区暂定开工",), ("商业街区", "暂定开工")),
        ("6#楼32F办公区暂定开工", ("6#楼32F办公区暂定开工", "6#楼32层办公区暂定开工"), ("6#楼", "32F", "暂定开工")),
        ("实际开工", ("实际开工", "实际开工依据"), ("开工令", "甲方开工")),
        ("合同工期", ("合同工期", "工期"), ("45天", "60天", "日历天", "节假日")),
        ("需澄清项", ("需澄清项",), ("334日历天", "不一致", "冲突", "需澄清")),
    ):
        value = _docx_value_for_labels(label_rows, aliases) or _docx_first_clause(content, keywords)
        if value and not _docx_is_placeholder_value(value):
            rows.append((label, _docx_clean_business_value(value, limit=260)))
    return _docx_dedupe_kv_rows(rows)


def _docx_payment_rows(content: str) -> list[tuple[str, str]]:
    label_rows = _docx_labeled_rows_from_text(content)
    rows = []
    for label, aliases, keywords in (
        ("预付款", ("预付款",), ("无预付款", "预付款")),
        ("履约保函", ("履约保函",), ("履约保函", "未提交前", "顺延付款")),
        ("进度款申报", ("进度款申报",), ("每月25日前", "工程计量报表", "报送")),
        ("进度款支付", ("进度款支付",), ("审核值", "80%", "15个工作日", "当月10万元")),
        ("支付上限", ("支付上限",), ("累计", "80%", "暂停支付")),
        ("结算款/质保金", ("结算款", "质保金"), ("97%", "质保金", "质保期")),
        ("开办费", ("开办费",), ("开办费", "第二", "第四", "第六")),
    ):
        value = _docx_value_for_labels(label_rows, aliases) or _docx_first_clause(content, keywords, reject=("术语和定义", "指各承建商", "指承建商"))
        if value:
            rows.append((label, _docx_clean_business_value(value, limit=300)))
    return _docx_dedupe_kv_rows(rows)


def _docx_value_for_labels(rows: list[tuple[str, str]], labels: tuple[str, ...]) -> str:
    for label, value in rows:
        if any(alias and (alias == label or alias in label or label in alias) for alias in labels):
            return value
    return ""


def _docx_first_clause(content: str, keywords: tuple[str, ...], *, reject: tuple[str, ...] = ()) -> str:
    for clause in _docx_matching_clauses(content, keywords, reject=reject, limit=1):
        return clause
    return ""


def _docx_matching_clauses(
    content: str,
    keywords: tuple[str, ...],
    *,
    reject: tuple[str, ...] = (),
    limit: int = 3,
) -> list[str]:
    matches: list[str] = []
    for clause in _docx_split_business_clauses(content):
        if reject and _has_any(clause, reject):
            continue
        if _has_any(clause, keywords):
            matches.append(clause)
        if len(matches) >= limit:
            break
    return matches


def _docx_split_business_clauses(content: str) -> list[str]:
    normalized = re.sub(r"\s+", " ", str(content or "")).strip()
    if not normalized:
        return []
    pieces = [part.strip(" ；;。") for part in re.split(r"[；;\n]+|(?<=。)", normalized) if part.strip(" ；;。")]
    clauses: list[str] = []
    for piece in pieces:
        if len(piece) <= 360:
            clauses.append(piece)
            continue
        subparts = [
            part.strip(" ，,。")
            for part in re.split(r"[。]|(?<=[，,])(?=(?:商业街区|6#楼|本工程|工程|甲方|乙方|丙方|承包人|发包人|招标人|投标人))", piece)
            if part.strip(" ，,。")
        ]
        clauses.extend(subparts or [piece])
    return clauses


def _docx_clean_business_value(value: Any, *, limit: int = 260) -> str:
    text = re.sub(r"\s+", " ", str(value or "")).strip(" ；;。")
    text = re.sub(r"^(?:\d+[、.．]\s*)", "", text)
    text = re.sub(r"^项号\s*内\s*容\s*规\s*定\s*", "", text)
    text = re.sub(r"东莞香港中心项目\s*商业街区及6#楼32F办公区装修专业分包工程\s*", "", text)
    text = text.strip(" ；;。")
    return _clip(text, limit)


def _docx_is_placeholder_value(value: str) -> bool:
    text = str(value or "")
    return _has_any(text, ("具体信息未提供", "具体内容未提供", "未在本摘要中提供", "详见《协议书》", "详见《技术要求》", "未明确"))


def _docx_is_definition_like_payment_text(content: str) -> bool:
    text = str(content or "")
    definition_like = _has_any(text, ("术语和定义", "指各承建商", "指承建商", "资金计划：指", "工程进度款：指", "工程结算款：指"))
    actual_payment = _has_any(text, ("无预付款", "支付至", "审核值", "15个工作日", "质保金", "结算总价", "合同价款的支付"))
    return definition_like and not actual_payment


def _docx_labeled_rows_from_text(text: str) -> list[tuple[str, str]]:
    normalized = re.sub(r"\s+", " ", str(text or "")).strip()
    if not normalized:
        return []
    parts = [part.strip(" ；;。") for part in re.split(r"[；;\n]+", normalized) if part.strip(" ；;。")]
    rows: list[tuple[str, str]] = []
    carry_label = ""
    for part in parts:
        match = re.match(r"^([^：:]{2,28})[：:]\s*(.+)$", part)
        if not match:
            if carry_label and rows:
                rows[-1] = (rows[-1][0], f"{rows[-1][1]}；{part}")
            continue
        label = match.group(1).strip()
        value = match.group(2).strip()
        if not label or not value:
            continue
        rows.append((label, value))
        carry_label = label
    return _docx_dedupe_kv_rows(rows)


def _docx_dedupe_kv_rows(rows: Any) -> list[tuple[str, str]]:
    seen: set[tuple[str, str]] = set()
    result: list[tuple[str, str]] = []
    for row in rows or []:
        if not row:
            continue
        label, value = row
        clean_label = _clip(str(label or "").strip(), 36)
        clean_value = _clip(str(value or "").strip(), 1000)
        if not clean_label or not clean_value:
            continue
        key = (clean_label, clean_value)
        if key in seen:
            continue
        seen.add(key)
        result.append(key)
    return result


def _docx_has_scoring_detail(scoring_rows: list[dict[str, Any]]) -> bool:
    return any(row.get("scoring_item") and row.get("scoring_item") != "未识别明确评分细则" for row in scoring_rows)


def _docx_clarification_lines(item: dict[str, Any] | None, review_rows: list[dict[str, Any]]) -> list[str]:
    value = _docx_summary_value(item)
    rows = _docx_labeled_rows_from_text(value)
    if rows:
        return [f"{label}：{content}" for label, content in rows]
    if value:
        return [part.strip(" 。；;") for part in re.split(r"[；;\n]+", value) if part.strip(" 。；;")]
    return [
        str(row.get("title") or "").strip()
        for row in review_rows
        if _has_any(str(row.get("reason_text") or row.get("title") or ""), ("需答疑", "待补充", "缺失", "空白", "冲突"))
    ][:10]


def _docx_join_nonempty(values: list[Any]) -> str:
    return "；".join(str(item).strip() for item in values if str(item or "").strip()) or "-"


def _docx_paragraph(
    text: Any,
    *,
    style: str = "Body",
    in_table: bool = False,
    page_break_before: bool = False,
) -> str:
    style_xml = f'<w:pStyle w:val="{style}"/>'
    outline_xml = ""
    if style == "Heading1":
        outline_xml = '<w:outlineLvl w:val="0"/>'
    elif style == "Heading2":
        outline_xml = '<w:outlineLvl w:val="1"/>'
    flow_xml = '<w:keepNext/><w:keepLines/>' if style in {"Heading1", "Heading2"} else '<w:widowControl/>'
    page_break_xml = '<w:pageBreakBefore/>' if page_break_before else ""
    spacing = "" if in_table else '<w:spacing w:before="0" w:after="120" w:line="280" w:lineRule="auto"/>'
    return f"<w:p><w:pPr>{style_xml}{outline_xml}{flow_xml}{page_break_xml}{spacing}</w:pPr>{_docx_run(text)}</w:p>"


def _docx_run(text: Any) -> str:
    value = str(text or "")
    if not value:
        return "<w:r><w:t></w:t></w:r>"
    parts = value.splitlines()
    chunks: list[str] = []
    for index, part in enumerate(parts):
        if index:
            chunks.append("<w:br/>")
        chunks.append(f'<w:t xml:space="preserve">{escape(part)}</w:t>')
    return "<w:r>" + "".join(chunks) + "</w:r>"


def _docx_toc_field(*, levels: str, fallback_lines: list[str] | None = None) -> str:
    safe_levels = escape(str(levels or "1-1"))
    instr = f'TOC \\o "{safe_levels}" \\h \\z \\u'
    clean_fallback = [
        str(line or "").strip()
        for line in (fallback_lines or [])
        if str(line or "").strip()
    ]
    if clean_fallback:
        paragraphs: list[str] = []
        first = clean_fallback[0]
        paragraphs.append(
            "<w:p>"
            '<w:pPr><w:pStyle w:val="Body"/><w:spacing w:before="0" w:after="120" w:line="280" w:lineRule="auto"/></w:pPr>'
            '<w:r><w:fldChar w:fldCharType="begin" w:dirty="true"/></w:r>'
            f'<w:r><w:instrText xml:space="preserve"> {escape(instr)} </w:instrText></w:r>'
            '<w:r><w:fldChar w:fldCharType="separate"/></w:r>'
            f'<w:r><w:t xml:space="preserve">{escape(first)}</w:t></w:r>'
            "</w:p>"
        )
        for line in clean_fallback[1:]:
            paragraphs.append(
                "<w:p>"
                '<w:pPr><w:pStyle w:val="Body"/><w:spacing w:before="0" w:after="120" w:line="280" w:lineRule="auto"/></w:pPr>'
                f'<w:r><w:t xml:space="preserve">{escape(line)}</w:t></w:r>'
                "</w:p>"
            )
        paragraphs.append('<w:p><w:r><w:fldChar w:fldCharType="end"/></w:r></w:p>')
        return "".join(paragraphs)
    return (
        "<w:p>"
        '<w:pPr><w:pStyle w:val="Body"/><w:spacing w:before="0" w:after="120" w:line="280" w:lineRule="auto"/></w:pPr>'
        '<w:r><w:fldChar w:fldCharType="begin" w:dirty="true"/></w:r>'
        f'<w:r><w:instrText xml:space="preserve"> {escape(instr)} </w:instrText></w:r>'
        '<w:r><w:fldChar w:fldCharType="separate"/></w:r>'
        '<w:r><w:t></w:t></w:r>'
        '<w:r><w:fldChar w:fldCharType="end"/></w:r>'
        "</w:p>"
    )


def _docx_table(
    rows: list[list[Any]],
    *,
    widths: tuple[int, ...],
    header: bool = False,
    first_col_shading: bool = False,
    cell_shading: str | None = None,
    border_color: str = "C7CDD6",
) -> str:
    grid = "".join(f'<w:gridCol w:w="{width}"/>' for width in widths)
    table_rows: list[str] = []
    for row_index, row in enumerate(rows):
        cells: list[str] = []
        for col_index, raw_value in enumerate(row):
            width = widths[min(col_index, len(widths) - 1)]
            fill = cell_shading
            if header and row_index == 0:
                fill = "1F4E78"
            elif first_col_shading and col_index == 0:
                fill = "EEF3F8"
            shading = f'<w:shd w:fill="{fill}"/>' if fill else ""
            cell_paragraphs = raw_value if isinstance(raw_value, list) else [_docx_paragraph(raw_value, style="TableHeader" if header and row_index == 0 else "TableText", in_table=True)]
            cells.append(
                '<w:tc>'
                f'<w:tcPr><w:tcW w:w="{width}" w:type="dxa"/>{shading}<w:vAlign w:val="center"/></w:tcPr>'
                f'{"".join(cell_paragraphs)}'
                "</w:tc>"
            )
        row_properties = '<w:trPr><w:cantSplit/>'
        if header and row_index == 0:
            row_properties += '<w:tblHeader/>'
        row_properties += '</w:trPr>'
        table_rows.append("<w:tr>" + row_properties + "".join(cells) + "</w:tr>")
    return (
        "<w:tbl>"
        "<w:tblPr>"
        '<w:tblW w:w="9360" w:type="dxa"/>'
        '<w:tblInd w:w="0" w:type="dxa"/>'
        '<w:tblBorders>'
        f'<w:top w:val="single" w:sz="6" w:space="0" w:color="{border_color}"/>'
        f'<w:left w:val="single" w:sz="6" w:space="0" w:color="{border_color}"/>'
        f'<w:bottom w:val="single" w:sz="6" w:space="0" w:color="{border_color}"/>'
        f'<w:right w:val="single" w:sz="6" w:space="0" w:color="{border_color}"/>'
        f'<w:insideH w:val="single" w:sz="6" w:space="0" w:color="{border_color}"/>'
        f'<w:insideV w:val="single" w:sz="6" w:space="0" w:color="{border_color}"/>'
        "</w:tblBorders>"
        '<w:tblCellMar><w:top w:w="80" w:type="dxa"/><w:left w:w="120" w:type="dxa"/><w:bottom w:w="80" w:type="dxa"/><w:right w:w="120" w:type="dxa"/></w:tblCellMar>'
        "</w:tblPr>"
        f"<w:tblGrid>{grid}</w:tblGrid>"
        f"{''.join(table_rows)}"
        "</w:tbl>"
        '<w:p><w:pPr><w:spacing w:after="120"/></w:pPr></w:p>'
    )


def _docx_package(
    document_xml: str,
    *,
    media: list[dict[str, Any]] | None = None,
    header_xml: str = "",
    footer_xml: str = "",
) -> bytes:
    media = media or []
    buffer = BytesIO()
    with ZipFile(buffer, "w", ZIP_DEFLATED) as archive:
        archive.writestr("[Content_Types].xml", _docx_content_types_xml(media, has_header=bool(header_xml), has_footer=bool(footer_xml)))
        archive.writestr("_rels/.rels", _docx_root_rels_xml())
        archive.writestr("word/_rels/document.xml.rels", _docx_document_rels_xml(media, has_header=bool(header_xml), has_footer=bool(footer_xml)))
        archive.writestr("word/document.xml", document_xml)
        archive.writestr("word/styles.xml", _docx_styles_xml())
        archive.writestr("word/settings.xml", _docx_settings_xml())
        if header_xml:
            archive.writestr("word/header1.xml", header_xml)
        if footer_xml:
            archive.writestr("word/footer1.xml", footer_xml)
        archive.writestr("docProps/app.xml", _docx_app_xml())
        archive.writestr("docProps/core.xml", _docx_core_xml())
        for item in media:
            archive.writestr(f"word/{item['target']}", item["content"])
    return buffer.getvalue()


def _docx_content_types_xml(
    media: list[dict[str, Any]] | None = None,
    *,
    has_header: bool = False,
    has_footer: bool = False,
) -> str:
    media_defaults = ""
    seen_extensions: set[str] = set()
    for item in media or []:
        target = str(item.get("target") or "")
        extension = target.rsplit(".", 1)[-1].lower() if "." in target else ""
        content_type = str(item.get("content_type") or "")
        if not extension or not content_type or extension in seen_extensions:
            continue
        seen_extensions.add(extension)
        media_defaults += f'<Default Extension="{escape(extension)}" ContentType="{escape(content_type)}"/>'
    header_override = (
        '<Override PartName="/word/header1.xml" ContentType="application/vnd.openxmlformats-officedocument.wordprocessingml.header+xml"/>'
        if has_header
        else ""
    )
    footer_override = (
        '<Override PartName="/word/footer1.xml" ContentType="application/vnd.openxmlformats-officedocument.wordprocessingml.footer+xml"/>'
        if has_footer
        else ""
    )
    return (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">'
        '<Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>'
        '<Default Extension="xml" ContentType="application/xml"/>'
        f"{media_defaults}"
        '<Override PartName="/word/document.xml" ContentType="application/vnd.openxmlformats-officedocument.wordprocessingml.document.main+xml"/>'
        '<Override PartName="/word/styles.xml" ContentType="application/vnd.openxmlformats-officedocument.wordprocessingml.styles+xml"/>'
        '<Override PartName="/word/settings.xml" ContentType="application/vnd.openxmlformats-officedocument.wordprocessingml.settings+xml"/>'
        f"{header_override}"
        f"{footer_override}"
        '<Override PartName="/docProps/core.xml" ContentType="application/vnd.openxmlformats-package.core-properties+xml"/>'
        '<Override PartName="/docProps/app.xml" ContentType="application/vnd.openxmlformats-officedocument.extended-properties+xml"/>'
        "</Types>"
    )


def _docx_root_rels_xml() -> str:
    return (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
        '<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" Target="word/document.xml"/>'
        '<Relationship Id="rId2" Type="http://schemas.openxmlformats.org/package/2006/relationships/metadata/core-properties" Target="docProps/core.xml"/>'
        '<Relationship Id="rId3" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/extended-properties" Target="docProps/app.xml"/>'
        "</Relationships>"
    )


def _docx_document_rels_xml(media: list[dict[str, Any]], *, has_header: bool = False, has_footer: bool = False) -> str:
    relationships = [
        '<Relationship Id="rIdStyles" '
        'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/styles" '
        'Target="styles.xml"/>',
        '<Relationship Id="rIdSettings" '
        'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/settings" '
        'Target="settings.xml"/>'
    ]
    if has_header:
        relationships.append(
            '<Relationship Id="rIdHeader1" '
            'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/header" '
            'Target="header1.xml"/>'
        )
    if has_footer:
        relationships.append(
            '<Relationship Id="rIdFooter1" '
            'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/footer" '
            'Target="footer1.xml"/>'
        )
    for item in media:
        relationships.append(
            f'<Relationship Id="{escape(str(item["rel_id"]))}" '
            'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/image" '
            f'Target="{escape(str(item["target"]))}"/>'
        )
    return (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
        f"{''.join(relationships)}"
        "</Relationships>"
    )


def _docx_header_xml(text: Any) -> str:
    return (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<w:hdr xmlns:w="http://schemas.openxmlformats.org/wordprocessingml/2006/main">'
        '<w:p>'
        '<w:pPr><w:jc w:val="center"/><w:pBdr><w:bottom w:val="single" w:sz="4" w:space="2" w:color="D9D9D9"/></w:pBdr></w:pPr>'
        f"{_docx_header_footer_run(text)}"
        "</w:p>"
        "</w:hdr>"
    )


def _docx_footer_xml(text: Any) -> str:
    prefix = str(text or "").strip() or "技术标部分"
    return (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<w:ftr xmlns:w="http://schemas.openxmlformats.org/wordprocessingml/2006/main">'
        '<w:p>'
        '<w:pPr><w:jc w:val="center"/><w:pBdr><w:top w:val="single" w:sz="4" w:space="2" w:color="D9D9D9"/></w:pBdr></w:pPr>'
        f"{_docx_header_footer_run(prefix + '  第 ')}"
        f"{_docx_field_runs('PAGE', '1')}"
        f"{_docx_header_footer_run(' 页 / 共 ')}"
        f"{_docx_field_runs('NUMPAGES', '1')}"
        f"{_docx_header_footer_run(' 页')}"
        "</w:p>"
        "</w:ftr>"
    )


def _docx_header_footer_run(text: Any) -> str:
    return (
        "<w:r>"
        '<w:rPr><w:rFonts w:ascii="Microsoft YaHei" w:hAnsi="Microsoft YaHei" w:eastAsia="Microsoft YaHei"/>'
        '<w:color w:val="666666"/><w:sz w:val="18"/><w:szCs w:val="18"/></w:rPr>'
        f'<w:t xml:space="preserve">{escape(str(text or ""))}</w:t>'
        "</w:r>"
    )


def _docx_field_runs(instruction: str, fallback: str) -> str:
    safe_instruction = escape(str(instruction or "").strip())
    return (
        '<w:r><w:fldChar w:fldCharType="begin"/></w:r>'
        f'<w:r><w:instrText xml:space="preserve"> {safe_instruction} </w:instrText></w:r>'
        '<w:r><w:fldChar w:fldCharType="separate"/></w:r>'
        f"{_docx_header_footer_run(fallback)}"
        '<w:r><w:fldChar w:fldCharType="end"/></w:r>'
    )


def _docx_image_extension(value: str) -> str:
    ext = re.sub(r"[^A-Za-z0-9]+", "", str(value or "").lower().strip("."))
    if ext == "jpeg":
        return "jpg"
    return ext


def _docx_image_content_type(extension: str) -> str:
    ext = _docx_image_extension(extension)
    if ext == "png":
        return "image/png"
    if ext in {"jpg", "jpeg"}:
        return "image/jpeg"
    return "application/octet-stream"


def _docx_image_extent(
    *,
    width_px: int | None,
    height_px: int | None,
    max_width_emu: int,
    max_height_emu: int,
) -> tuple[int, int]:
    emu_per_px = 9525
    width = max(int(width_px or 1200) * emu_per_px, 1)
    height = max(int(height_px or 800) * emu_per_px, 1)
    scale = min(max_width_emu / width, max_height_emu / height, 1.0)
    return max(1, int(width * scale)), max(1, int(height * scale))


def _docx_image_paragraph(*, rel_id: str, filename: str, cx: int, cy: int, doc_pr_id: int) -> str:
    name = escape(str(filename or f"image-{doc_pr_id}"))
    return (
        '<w:p><w:pPr><w:jc w:val="center"/><w:spacing w:before="80" w:after="80"/></w:pPr><w:r><w:drawing>'
        '<wp:inline distT="0" distB="0" distL="0" distR="0">'
        f'<wp:extent cx="{cx}" cy="{cy}"/>'
        f'<wp:docPr id="{doc_pr_id}" name="{name}"/>'
        '<wp:cNvGraphicFramePr><a:graphicFrameLocks noChangeAspect="1"/></wp:cNvGraphicFramePr>'
        '<a:graphic><a:graphicData uri="http://schemas.openxmlformats.org/drawingml/2006/picture">'
        '<pic:pic>'
        '<pic:nvPicPr>'
        f'<pic:cNvPr id="{doc_pr_id}" name="{name}"/>'
        '<pic:cNvPicPr/>'
        '</pic:nvPicPr>'
        '<pic:blipFill>'
        f'<a:blip r:embed="{escape(rel_id)}"/>'
        '<a:stretch><a:fillRect/></a:stretch>'
        '</pic:blipFill>'
        '<pic:spPr>'
        f'<a:xfrm><a:off x="0" y="0"/><a:ext cx="{cx}" cy="{cy}"/></a:xfrm>'
        '<a:prstGeom prst="rect"><a:avLst/></a:prstGeom>'
        '</pic:spPr>'
        '</pic:pic>'
        '</a:graphicData></a:graphic>'
        '</wp:inline>'
        '</w:drawing></w:r></w:p>'
    )


def _docx_styles_xml() -> str:
    def style(
        style_id: str,
        name: str,
        size: int,
        *,
        bold: bool = False,
        color: str = "000000",
        align: str | None = None,
        after: int = 120,
        outline_level: int | None = None,
    ) -> str:
        jc = f'<w:jc w:val="{align}"/>' if align else ""
        outline = f'<w:outlineLvl w:val="{outline_level}"/>' if outline_level is not None else ""
        bold_xml = "<w:b/>" if bold else ""
        return (
            f'<w:style w:type="paragraph" w:styleId="{style_id}">'
            f'<w:name w:val="{name}"/>'
            f'<w:pPr>{jc}{outline}<w:spacing w:after="{after}" w:line="280" w:lineRule="auto"/></w:pPr>'
            '<w:rPr>'
            '<w:rFonts w:ascii="Microsoft YaHei" w:hAnsi="Microsoft YaHei" w:eastAsia="Microsoft YaHei"/>'
            f"{bold_xml}<w:color w:val=\"{color}\"/><w:sz w:val=\"{size}\"/><w:szCs w:val=\"{size}\"/>"
            "</w:rPr>"
            "</w:style>"
        )

    return (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<w:styles xmlns:w="http://schemas.openxmlformats.org/wordprocessingml/2006/main">'
        '<w:docDefaults><w:rPrDefault><w:rPr><w:rFonts w:ascii="Microsoft YaHei" w:hAnsi="Microsoft YaHei" w:eastAsia="Microsoft YaHei"/><w:sz w:val="22"/><w:szCs w:val="22"/></w:rPr></w:rPrDefault></w:docDefaults>'
        + style("Normal", "Normal", 22, after=120)
        + style("DocTitle", "Document Title", 32, bold=True, align="center", after=60)
        + style("DocSubtitle", "Document Subtitle", 28, bold=True, align="center", after=160)
        + style("Meta", "Metadata", 20, color="555555", align="center", after=160)
        + style("Heading1", "Heading 1", 26, bold=True, color="1F4E78", after=120, outline_level=0)
        + style("Heading2", "Heading 2", 23, bold=True, color="244062", after=80, outline_level=1)
        + style("Body", "Body", 22, after=100)
        + style("CalloutTitle", "Callout Title", 22, bold=True, color="7A5A00", after=40)
        + style("CalloutBody", "Callout Body", 22, color="4A3B00", after=40)
        + style("TableText", "Table Text", 20, after=40)
        + style("TableHeader", "Table Header", 20, bold=True, color="FFFFFF", align="center", after=40)
        + "</w:styles>"
    )


def _docx_settings_xml() -> str:
    return (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<w:settings xmlns:w="http://schemas.openxmlformats.org/wordprocessingml/2006/main">'
        '<w:updateFields w:val="true"/>'
        "</w:settings>"
    )


def _docx_app_xml() -> str:
    return (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Properties xmlns="http://schemas.openxmlformats.org/officeDocument/2006/extended-properties" '
        'xmlns:vt="http://schemas.openxmlformats.org/officeDocument/2006/docPropsVTypes">'
        "<Application>AI Middle Office</Application>"
        "</Properties>"
    )


def _docx_core_xml() -> str:
    now = datetime.now(timezone.utc).isoformat(timespec="seconds").replace("+00:00", "Z")
    return (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<cp:coreProperties xmlns:cp="http://schemas.openxmlformats.org/package/2006/metadata/core-properties" '
        'xmlns:dc="http://purl.org/dc/elements/1.1/" '
        'xmlns:dcterms="http://purl.org/dc/terms/" '
        'xmlns:dcmitype="http://purl.org/dc/dcmitype/" '
        'xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance">'
        "<dc:title>招标文件分析成果表</dc:title>"
        "<dc:creator>AI Middle Office</dc:creator>"
        f'<dcterms:created xsi:type="dcterms:W3CDTF">{now}</dcterms:created>'
        f'<dcterms:modified xsi:type="dcterms:W3CDTF">{now}</dcterms:modified>'
        "</cp:coreProperties>"
    )


def _flatten_export_rows(table_key: str, rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    flattened: list[dict[str, Any]] = []

    def visit(row: dict[str, Any], depth: int) -> None:
        export_row = dict(row)
        children = export_row.pop("children", None) or []
        if depth > 0:
            prefix = "  " * depth + "- "
            if table_key == "scoring" and export_row.get("scoring_item"):
                export_row["scoring_item"] = prefix + str(export_row.get("scoring_item"))
            if table_key == "risk_clause" and export_row.get("risk_title"):
                export_row["risk_title"] = prefix + str(export_row.get("risk_title"))
            if table_key == "risk_clause" and export_row.get("clause_text"):
                export_row["clause_text"] = prefix + str(export_row.get("clause_text"))
        flattened.append(export_row)
        for child in children:
            if isinstance(child, dict):
                visit(child, depth + 1)

    for row in rows:
        visit(row, 0)
    return flattened


async def _important_info_with_cache(
    db: Session,
    project: BidProject,
    run: BidParseRun,
    *,
    username: str | None,
    trace_id: str | None,
) -> dict[str, Any]:
    provider = (settings.bidding_llm_provider or "deepseek").strip().lower()
    model = _bidding_summary_llm_model()
    metadata: dict[str, Any] = {
        "enabled": bool(settings.feature_bidding_llm_review),
        "provider": provider,
        "model": model,
        "prompt_version": TENDER_ANALYSIS_IMPORTANT_INFO_PROMPT_VERSION,
        "status": "disabled",
    }
    if not settings.feature_bidding_llm_review:
        return _empty_important_info_result(status="disabled", metadata=metadata)
    if provider != "deepseek":
        metadata.update({"status": "skipped", "skip_reason": "bidding_llm_provider_not_deepseek"})
        return _empty_important_info_result(status="skipped", metadata=metadata)
    if not (settings.deepseek_api_key or "").strip():
        metadata.update({"status": "skipped", "skip_reason": "deepseek_api_key_missing"})
        return _empty_important_info_result(status="skipped", metadata=metadata)

    cached = _cached_important_info(run, model=model)
    if cached:
        cached_metadata = dict(cached.get("metadata") or {})
        cached_metadata.update({"status": "cached", "enabled": True, "provider": provider, "model": model})
        return {
            "status": "cached",
            "metadata": cached_metadata,
            "sections": cached.get("sections") or [],
            "priority_clarifications": cached.get("priority_clarifications") or [],
            "overall_note": cached.get("overall_note") or "",
        }

    context = _build_important_info_context(db, project, run)
    metadata["chunk_count"] = len(context.get("document_chunks") or [])
    metadata["context_truncated"] = bool(context.get("context_truncated"))
    if not context.get("document_chunks"):
        metadata.update({"status": "no_source"})
        return _empty_important_info_result(status="no_source", metadata=metadata)

    try:
        raw_payload = await _call_deepseek_tender_important_info_extraction(
            context,
            username=username,
            trace_id=trace_id or run.run_uuid,
        )
        cleaned = _clean_important_info_payload(raw_payload, context)
        metadata.update(
            {
                "status": "completed",
                "generated_at": datetime.now().isoformat(timespec="seconds"),
                "found_count": _important_info_found_count(cleaned.get("sections") or []),
                "field_count": len(IMPORTANT_INFO_FIELD_KEYS),
            }
        )
        cleaned["status"] = "completed"
        cleaned["metadata"] = metadata
        _store_important_info_cache(db, run, result=cleaned, metadata=metadata)
        return cleaned
    except Exception as exc:
        metadata.update({"status": "error", "error": (str(exc) or exc.__class__.__name__)[:300]})
        return _empty_important_info_result(status="error", metadata=metadata)


def _empty_important_info_result(*, status: str, metadata: dict[str, Any] | None = None) -> dict[str, Any]:
    return {
        "status": status,
        "metadata": metadata or {"status": status, "prompt_version": TENDER_ANALYSIS_IMPORTANT_INFO_PROMPT_VERSION},
        "sections": [],
        "priority_clarifications": [],
        "overall_note": "",
    }


def _cached_important_info(run: BidParseRun, *, model: str) -> dict[str, Any] | None:
    summary = loads_json(run.summary_json, {}) if run.summary_json else {}
    cached = summary.get("tender_analysis_important_info_llm")
    if not isinstance(cached, dict):
        return None
    if cached.get("prompt_version") != TENDER_ANALYSIS_IMPORTANT_INFO_PROMPT_VERSION:
        return None
    if cached.get("model") != model:
        return None
    if not isinstance(cached.get("sections"), list):
        return None
    return cached


def _store_important_info_cache(
    db: Session,
    run: BidParseRun,
    *,
    result: dict[str, Any],
    metadata: dict[str, Any],
) -> None:
    summary = loads_json(run.summary_json, {}) if run.summary_json else {}
    summary["tender_analysis_important_info_llm"] = {
        "prompt_version": TENDER_ANALYSIS_IMPORTANT_INFO_PROMPT_VERSION,
        "model": metadata.get("model"),
        "metadata": metadata,
        "sections": result.get("sections") or [],
        "priority_clarifications": result.get("priority_clarifications") or [],
        "overall_note": result.get("overall_note") or "",
    }
    run.summary_json = dumps_json(summary)
    db.add(run)
    db.commit()
    db.refresh(run)


def _build_important_info_context(db: Session, project: BidProject, run: BidParseRun) -> dict[str, Any]:
    chunks, truncated = _important_info_document_chunks(db, project, run)
    return {
        "prompt_version": TENDER_ANALYSIS_IMPORTANT_INFO_PROMPT_VERSION,
        "task": "bidding_tender_important_information_extraction",
        "project": {
            "project_uuid": project.project_uuid,
            "project_name": project.project_name,
            "tenderer_name": project.tenderer_name,
            "tender_agency": project.tender_agency,
            "project_location": project.project_location,
            "project_type": project.project_type,
        },
        "run_uuid": run.run_uuid,
        "output_schema": [
            {
                "section_key": section["section_key"],
                "title": section["title"],
                "fields": [
                    {"field_key": field_key, "field_name": field_name}
                    for field_key, field_name in section["fields"]
                ],
            }
            for section in IMPORTANT_INFO_SECTION_SCHEMA
        ],
        "document_chunks": chunks,
        "context_truncated": truncated,
    }


def _important_info_document_chunks(db: Session, project: BidProject, run: BidParseRun) -> tuple[list[dict[str, Any]], bool]:
    input_file_uuids = [str(item) for item in loads_json(run.input_file_ids_json, []) or [] if item]
    query = db.query(BidProjectFile).filter(BidProjectFile.project_id == project.id, BidProjectFile.parser_status == "parsed")
    if input_file_uuids:
        query = query.filter(BidProjectFile.file_uuid.in_(input_file_uuids))
    files = query.order_by(BidProjectFile.id.asc()).all()

    raw_chunks: list[dict[str, Any]] = []
    seen: set[str] = set()
    for file_obj in files:
        segments = loads_json(file_obj.segments_json, []) if file_obj.segments_json else []
        if isinstance(segments, list) and segments:
            for index, segment in enumerate(segments, start=1):
                if not isinstance(segment, dict):
                    continue
                text = re.sub(r"\s+", " ", str(segment.get("text") or "")).strip()
                if len(text) < 6:
                    continue
                key = _stable_text_key(text)
                if key in seen:
                    continue
                seen.add(key)
                raw_chunks.append(
                    {
                        "source_file": segment.get("source_file") or file_obj.original_filename,
                        "source_location": segment.get("source_location") or f"段落{index}",
                        "document_section": segment.get("document_section") or segment.get("document_section_label") or "",
                        "text": text,
                    }
                )
            continue
        for index, text in enumerate(_split_text_for_important_info(file_obj.extracted_text), start=1):
            key = _stable_text_key(text)
            if key in seen:
                continue
            seen.add(key)
            raw_chunks.append(
                {
                    "source_file": file_obj.original_filename,
                    "source_location": f"文本片段{index}",
                    "document_section": "",
                    "text": text,
                }
            )

    chunks: list[dict[str, Any]] = []
    for item in raw_chunks:
        text = re.sub(r"\s+", " ", str(item.get("text") or "")).strip()
        if not text:
            continue
        chunks.append(
            {
                "evidence_id": f"D{len(chunks) + 1:03d}",
                "source_file": item.get("source_file"),
                "source_location": item.get("source_location"),
                "document_section": item.get("document_section"),
                "text": text,
            }
        )
    return chunks, False


def _split_text_for_important_info(text: Any) -> list[str]:
    value = str(text or "").strip()
    if not value:
        return []
    return [part.strip() for part in re.split(r"\n{2,}|(?<=。)\s*", value) if part.strip()]


IMPORTANT_INFO_SYSTEM_PROMPT = """你是装饰工程投标预算员，正在阅读一份招标文件。
任务：只根据输入的 document_chunks 原文证据，识别并整理“投标重要有效信息提取表”。

重要要求：
1. output_schema 只是最终表格结构，不是关键词规则；请理解原文含义后填写。
2. 每个字段都必须返回一个 item。原文明确写了就 status=found；原文空白、未填写、前后矛盾或需要补遗确认就 status=unclear；确实没有看到证据就 status=not_found。
3. value 要写成投标预算员可直接使用的整理口径，具体、完整、可读，不要整段复制原文。
4. 不得因为出现字段关键词就硬套；例如“工程进度款”的术语定义不是付款方式，投标书格式里的联系人不一定是答疑联系人。
5. 不得编造原文没有的信息；没有明确日期、邮箱、联系人、比例、金额时，要写“未填写/未明确/需澄清”。
6. source_evidence_ids 只能引用 document_chunks 中的 evidence_id。
7. 回标前优先澄清清单要汇总空白项、矛盾项、影响报价或废标风险的待确认事项。
8. 输出严格 JSON，不要 Markdown，不要代码块。

返回格式：
{
  "sections": [
    {
      "section_key": "project_overview",
      "items": [
        {
          "field_key": "tender_project_name",
          "field_name": "招标工程名称",
          "status": "found|unclear|not_found",
          "value": "整理后的中文结果",
          "source_evidence_ids": ["D001"],
          "confidence": 0.0,
          "note": "一句话说明依据或需澄清原因"
        }
      ]
    }
  ],
  "priority_clarifications": [
    {
      "item": "需澄清事项",
      "reason": "为什么需要回标前确认",
      "source_evidence_ids": ["D001"]
    }
  ],
  "overall_note": "可选，整体说明"
}
"""


async def _call_deepseek_tender_important_info_extraction(
    context: dict[str, Any],
    *,
    username: str | None,
    trace_id: str | None,
) -> dict[str, Any]:
    model = _bidding_summary_llm_model()
    payload = {
        "model": model,
        "messages": [
            {"role": "system", "content": IMPORTANT_INFO_SYSTEM_PROMPT},
            {"role": "user", "content": json.dumps(context, ensure_ascii=False)},
        ],
        "temperature": 0.05,
        "response_format": {"type": "json_object"},
    }
    response = await post_json_via_gateway(
        provider="deepseek",
        model=model,
        endpoint_type="bidding_tender_important_info_extract",
        url=settings.deepseek_chat_url,
        json_payload=payload,
        headers={
            "Authorization": f"Bearer {settings.deepseek_api_key.strip()}",
            "Content-Type": "application/json",
        },
        timeout=settings.bidding_llm_timeout_seconds,
        username=username,
        trace_id=trace_id,
    )
    if not 200 <= response.status_code < 300:
        raise RuntimeError(f"HTTP {response.status_code}")
    content = response.json().get("choices", [{}])[0].get("message", {}).get("content", "")
    return _extract_json_object(content)


def _clean_important_info_payload(payload: dict[str, Any], context: dict[str, Any]) -> dict[str, Any]:
    chunks_by_id = {
        str(chunk.get("evidence_id")): chunk
        for chunk in context.get("document_chunks") or []
        if isinstance(chunk, dict) and chunk.get("evidence_id")
    }
    raw_sections = payload.get("sections") if isinstance(payload.get("sections"), list) else []
    raw_by_section = {
        str(section.get("section_key") or ""): section
        for section in raw_sections
        if isinstance(section, dict)
    }

    cleaned_sections: list[dict[str, Any]] = []
    for section_spec in IMPORTANT_INFO_SECTION_SCHEMA:
        section_key = str(section_spec["section_key"])
        raw_section = raw_by_section.get(section_key) or {}
        raw_items = raw_section.get("items") if isinstance(raw_section.get("items"), list) else []
        raw_by_field = {
            str(item.get("field_key") or ""): item
            for item in raw_items
            if isinstance(item, dict)
        }
        items: list[dict[str, Any]] = []
        for field_key, field_name in section_spec["fields"]:
            raw_item = raw_by_field.get(field_key) or {}
            status = str(raw_item.get("status") or "not_found").strip().lower()
            if status not in {"found", "unclear", "not_found"}:
                status = "not_found"
            value = _clean_llm_business_value(raw_item.get("value"), limit=None)
            source_ids = [
                source_id
                for source_id in _clean_string_list(raw_item.get("source_evidence_ids"))
                if source_id in chunks_by_id
            ][:5]
            if status == "found" and not value:
                status = "not_found"
            if status == "found" and not source_ids:
                status = "unclear"
            selected_chunks = [chunks_by_id[source_id] for source_id in source_ids]
            note = _clean_llm_business_value(raw_item.get("note"), limit=None)
            if not raw_item:
                note = "LLM未返回该字段，需人工复核原文是否未明确。"
            items.append(
                {
                    "field_key": field_key,
                    "field_name": field_name,
                    "status": status,
                    "value": value if status in {"found", "unclear"} else "",
                    "source_evidence_ids": source_ids,
                    "source_file": _join_unique([chunk.get("source_file") for chunk in selected_chunks], limit=3),
                    "source_location": _join_unique([chunk.get("source_location") for chunk in selected_chunks], limit=5),
                    "evidence_text": _join_unique([chunk.get("text") for chunk in selected_chunks], limit=2),
                    "confidence": _safe_confidence(raw_item.get("confidence")),
                    "note": note,
                }
            )
        cleaned_sections.append(
            {
                "section_key": section_key,
                "title": section_spec["title"],
                "items": items,
            }
        )

    priority_clarifications = []
    raw_clarifications = payload.get("priority_clarifications") if isinstance(payload.get("priority_clarifications"), list) else []
    for raw in raw_clarifications:
        if not isinstance(raw, dict):
            continue
        source_ids = [
            source_id
            for source_id in _clean_string_list(raw.get("source_evidence_ids"))
            if source_id in chunks_by_id
        ][:5]
        item = _clean_llm_business_value(raw.get("item"), limit=None)
        reason = _clean_llm_business_value(raw.get("reason"), limit=None)
        if not item and not reason:
            continue
        priority_clarifications.append(
            {
                "item": item,
                "reason": reason,
                "source_evidence_ids": source_ids,
            }
        )
    return {
        "sections": cleaned_sections,
        "priority_clarifications": priority_clarifications,
        "overall_note": _clean_llm_business_value(payload.get("overall_note"), limit=None),
    }


def _clean_llm_business_value(value: Any, *, limit: int | None = 600) -> str:
    text = re.sub(r"\s+", " ", str(value or "")).strip(" ；;。")
    text = re.sub(r"^[-*•\s]+", "", text).strip()
    text = re.sub(r"^\d+[、.．]\s*", "", text).strip()
    if text in {"无", "暂无", "N/A", "null", "None"}:
        return ""
    if limit is None or limit <= 0:
        return text
    return _clip(text, limit)


def _important_info_found_count(sections: list[dict[str, Any]]) -> int:
    count = 0
    for section in sections:
        for item in section.get("items") or []:
            if isinstance(item, dict) and item.get("status") == "found" and item.get("value"):
                count += 1
    return count


async def _risk_clause_llm_with_cache(
    db: Session,
    project: BidProject,
    run: BidParseRun,
    *,
    username: str | None,
    trace_id: str | None,
    force: bool = False,
) -> dict[str, Any]:
    provider = (settings.bidding_llm_provider or "deepseek").strip().lower()
    model = _bidding_summary_llm_model()
    metadata: dict[str, Any] = {
        "enabled": bool(settings.feature_bidding_llm_review),
        "provider": provider,
        "model": model,
        "prompt_version": TENDER_ANALYSIS_RISK_CLAUSE_PROMPT_VERSION,
        "status": "disabled",
    }
    if not settings.feature_bidding_llm_review:
        return _empty_risk_clause_result(status="disabled", metadata=metadata)
    if provider != "deepseek":
        metadata.update({"status": "skipped", "skip_reason": "bidding_llm_provider_not_deepseek"})
        return _empty_risk_clause_result(status="skipped", metadata=metadata)
    if not (settings.deepseek_api_key or "").strip():
        metadata.update({"status": "skipped", "skip_reason": "deepseek_api_key_missing"})
        return _empty_risk_clause_result(status="skipped", metadata=metadata)

    cached = None if force else _cached_risk_clause_llm(run, model=model)
    if cached:
        cached_metadata = dict(cached.get("metadata") or {})
        cached_metadata.update({"status": "cached", "enabled": True, "provider": provider, "model": model})
        return {
            "status": "cached",
            "metadata": cached_metadata,
            "basic_info": cached.get("basic_info") or {},
            "priority_attention": cached.get("priority_attention") or [],
            "risks": cached.get("risks") or [],
            "overall_note": cached.get("overall_note") or "",
        }

    context = _build_risk_clause_llm_context(db, project, run)
    metadata["chunk_count"] = len(context.get("document_chunks") or [])
    metadata["context_truncated"] = bool(context.get("context_truncated"))
    if not context.get("document_chunks"):
        metadata.update({"status": "no_source"})
        return _empty_risk_clause_result(status="no_source", metadata=metadata)

    try:
        raw_payload = await _call_deepseek_tender_risk_clause_extraction(
            context,
            username=username,
            trace_id=trace_id or run.run_uuid,
        )
        cleaned = _clean_risk_clause_payload(raw_payload, context)
        metadata.update(
            {
                "status": "completed",
                "generated_at": datetime.now().isoformat(timespec="seconds"),
                "risk_count": len(cleaned.get("risks") or []),
                "high_count": _risk_clause_level_count(cleaned.get("risks") or [], "high"),
                "medium_count": _risk_clause_level_count(cleaned.get("risks") or [], "medium"),
                "low_count": _risk_clause_level_count(cleaned.get("risks") or [], "low"),
            }
        )
        cleaned["status"] = "completed"
        cleaned["metadata"] = metadata
        _store_risk_clause_llm_cache(db, run, result=cleaned, metadata=metadata)
        return cleaned
    except Exception as exc:
        metadata.update({"status": "error", "error": (str(exc) or exc.__class__.__name__)[:300]})
        return _empty_risk_clause_result(status="error", metadata=metadata)


def _empty_risk_clause_result(*, status: str, metadata: dict[str, Any] | None = None) -> dict[str, Any]:
    return {
        "status": status,
        "metadata": metadata or {"status": status, "prompt_version": TENDER_ANALYSIS_RISK_CLAUSE_PROMPT_VERSION},
        "basic_info": {},
        "priority_attention": [],
        "risks": [],
        "overall_note": "",
    }


def _cached_risk_clause_llm(run: BidParseRun, *, model: str) -> dict[str, Any] | None:
    summary = loads_json(run.summary_json, {}) if run.summary_json else {}
    cached = summary.get("tender_analysis_risk_clause_llm")
    if not isinstance(cached, dict):
        return None
    if cached.get("prompt_version") != TENDER_ANALYSIS_RISK_CLAUSE_PROMPT_VERSION:
        return None
    if cached.get("model") != model:
        return None
    if not isinstance(cached.get("risks"), list):
        return None
    return cached


def _store_risk_clause_llm_cache(
    db: Session,
    run: BidParseRun,
    *,
    result: dict[str, Any],
    metadata: dict[str, Any],
) -> None:
    summary = loads_json(run.summary_json, {}) if run.summary_json else {}
    summary["tender_analysis_risk_clause_llm"] = {
        "prompt_version": TENDER_ANALYSIS_RISK_CLAUSE_PROMPT_VERSION,
        "model": metadata.get("model"),
        "metadata": metadata,
        "basic_info": result.get("basic_info") or {},
        "priority_attention": result.get("priority_attention") or [],
        "risks": result.get("risks") or [],
        "overall_note": result.get("overall_note") or "",
    }
    run.summary_json = dumps_json(summary)
    db.add(run)
    db.commit()
    db.refresh(run)


def _build_risk_clause_llm_context(db: Session, project: BidProject, run: BidParseRun) -> dict[str, Any]:
    chunks, truncated = _important_info_document_chunks(db, project, run)
    return {
        "prompt_version": TENDER_ANALYSIS_RISK_CLAUSE_PROMPT_VERSION,
        "task": "bidding_tender_risk_clause_extraction",
        "project": {
            "project_uuid": project.project_uuid,
            "project_name": project.project_name,
            "tenderer_name": project.tenderer_name,
            "tender_agency": project.tender_agency,
            "project_location": project.project_location,
            "project_type": project.project_type,
        },
        "run_uuid": run.run_uuid,
        "output_schema": {
            "basic_info": ["依据文件", "适用场景", "风险分布", "生成说明"],
            "priority_attention": [
                {"category": "优先谈判", "suggestion": "需回标前谈判或澄清的重大商务/合同风险"},
                {"category": "优先管控", "suggestion": "中标后必须落地的履约、商务、技术、资料管控动作"},
                {"category": "关键证据", "suggestion": "支撑签证索赔、顺延、结算、抗辩所需的证据类型"},
            ],
            "risk_overview_fields": ["序号", "等级", "所在章节", "风险说明"],
            "risk_detail_fields": ["所在章节", "风险等级", "条款原文", "风险说明", "建议应对方式"],
        },
        "document_chunks": chunks,
        "context_truncated": truncated,
    }


RISK_CLAUSE_SYSTEM_PROMPT = """你是装饰工程投标阶段的合同风险评审专家，正在阅读一份招标文件。
任务：只根据输入的 document_chunks 原文证据，识别并整理“风险条款清单”。这是一条独立工作流，不要依赖结构化信息摘要表，不要沿用关键词规则。

识别范围：
1. 重点识别会影响报价、利润、工期、付款、结算、签证索赔、废标合规、质量安全、材料采购、履约担保、合同解释权、现场条件、资料提交和保修责任的条款。
2. 同一风险可合并为一条，但必须保留最关键的条款原文和所在章节。
3. 不要把普通流程说明误判为风险；只有责任明显加重、权利受限、费用/工期/付款不确定、违约责任或废标后果明确时才列入。

输出要求：
1. 条款原文必须摘录关键原文，不要只写概括；风险说明和建议应对方式要写成投标预算员、商务、法务、项目经理可直接使用的中文。
2. 每条风险必须给出 risk_level：high、medium、low。高风险优先给合同价款不可调整、漏项视为已含、无预付款/付款滞后、重罚、单方解除、废标、见索即付保函、未定价先施工、签证索赔失权等。
3. source_evidence_ids 只能引用 document_chunks 中的 evidence_id。
4. 不得编造原文没有的条款；无法确认的不要输出为风险。
5. priority_attention 必须包含三类：优先谈判、优先管控、关键证据，每类至少一条建议。
6. 输出严格 JSON，不要 Markdown，不要代码块。

返回格式：
{
  "basic_info": {
    "applicable_scenario": "投标评审、合同谈判、项目履约交底、签证与索赔管控",
    "generation_note": "风险等级为商务/履约初评，建议由法务、商务、项目经理共同复核。"
  },
  "priority_attention": [
    {"category": "优先谈判", "suggestion": "建议内容"},
    {"category": "优先管控", "suggestion": "建议内容"},
    {"category": "关键证据", "suggestion": "建议内容"}
  ],
  "risks": [
    {
      "risk_id": "R-01",
      "risk_level": "high|medium|low",
      "source_location": "章节/条款号",
      "clause_original": "关键条款原文摘录",
      "risk_explanation": "风险说明",
      "suggested_response": "建议应对方式",
      "source_evidence_ids": ["D001"],
      "confidence": 0.0
    }
  ],
  "overall_note": "可选，整体说明"
}
"""


async def _call_deepseek_tender_risk_clause_extraction(
    context: dict[str, Any],
    *,
    username: str | None,
    trace_id: str | None,
) -> dict[str, Any]:
    model = _bidding_summary_llm_model()
    payload = {
        "model": model,
        "messages": [
            {"role": "system", "content": RISK_CLAUSE_SYSTEM_PROMPT},
            {"role": "user", "content": json.dumps(context, ensure_ascii=False)},
        ],
        "temperature": 0.05,
        "response_format": {"type": "json_object"},
    }
    response = await post_json_via_gateway(
        provider="deepseek",
        model=model,
        endpoint_type="bidding_tender_risk_clause_extract",
        url=settings.deepseek_chat_url,
        json_payload=payload,
        headers={
            "Authorization": f"Bearer {settings.deepseek_api_key.strip()}",
            "Content-Type": "application/json",
        },
        timeout=settings.bidding_llm_timeout_seconds,
        username=username,
        trace_id=trace_id,
    )
    if not 200 <= response.status_code < 300:
        raise RuntimeError(f"HTTP {response.status_code}")
    content = response.json().get("choices", [{}])[0].get("message", {}).get("content", "")
    return _extract_json_object(content)


def _clean_risk_clause_payload(payload: dict[str, Any], context: dict[str, Any]) -> dict[str, Any]:
    chunks_by_id = {
        str(chunk.get("evidence_id")): chunk
        for chunk in context.get("document_chunks") or []
        if isinstance(chunk, dict) and chunk.get("evidence_id")
    }
    all_source_files = _join_unique(
        [chunk.get("source_file") for chunk in context.get("document_chunks") or [] if isinstance(chunk, dict)],
        limit=8,
    )
    basic_info_raw = payload.get("basic_info") if isinstance(payload.get("basic_info"), dict) else {}
    priority_attention = _clean_risk_clause_priority_attention(payload.get("priority_attention"))

    risks: list[dict[str, Any]] = []
    raw_risks = payload.get("risks") if isinstance(payload.get("risks"), list) else []
    for raw in raw_risks:
        if not isinstance(raw, dict):
            continue
        source_ids = [
            source_id
            for source_id in _clean_string_list(raw.get("source_evidence_ids"))
            if source_id in chunks_by_id
        ][:6]
        selected_chunks = [chunks_by_id[source_id] for source_id in source_ids]
        clause_original = _clean_llm_business_value(raw.get("clause_original"), limit=None)
        risk_explanation = _clean_llm_business_value(raw.get("risk_explanation"), limit=None)
        suggested_response = _clean_llm_business_value(raw.get("suggested_response"), limit=None)
        if not clause_original or not risk_explanation:
            continue
        risk_level = str(raw.get("risk_level") or "medium").strip().lower()
        if risk_level not in {"high", "medium", "low"}:
            risk_level = "medium"
        source_file = _join_unique([chunk.get("source_file") for chunk in selected_chunks], limit=3)
        source_location = _join_unique([chunk.get("source_location") for chunk in selected_chunks], limit=5)
        if not source_location:
            source_location = _clean_llm_business_value(raw.get("source_location"), limit=120)
        if not source_file:
            source_file = _clean_llm_business_value(raw.get("source_file"), limit=200)
        risks.append(
            {
                "risk_id": f"R-{len(risks) + 1:02d}",
                "risk_level": risk_level,
                "risk_level_label": RISK_LEVEL_LABELS.get(risk_level, risk_level),
                "source_file": source_file,
                "source_location": source_location,
                "clause_original": clause_original,
                "risk_explanation": risk_explanation,
                "suggested_response": suggested_response,
                "source_evidence_ids": source_ids,
                "evidence_text": _join_unique([chunk.get("text") for chunk in selected_chunks], limit=2),
                "confidence": _safe_confidence(raw.get("confidence")),
            }
        )

    risks = sorted(risks, key=lambda row: (-RISK_LEVEL_ORDER.get(str(row.get("risk_level")), 2), row.get("risk_id") or ""))
    for index, risk in enumerate(risks, start=1):
        risk["risk_id"] = f"R-{index:02d}"

    basic_info = {
        "source_files": all_source_files,
        "applicable_scenario": _clean_llm_business_value(
            basic_info_raw.get("applicable_scenario"),
            limit=None,
        )
        or "投标评审、合同谈判、项目履约交底、签证与索赔管控",
        "risk_distribution": _risk_clause_distribution_text(risks),
        "generation_note": _clean_llm_business_value(basic_info_raw.get("generation_note"), limit=None)
        or "风险等级为商务/履约初评，建议由法务、商务、项目经理共同复核。",
    }
    return {
        "basic_info": basic_info,
        "priority_attention": priority_attention,
        "risks": risks,
        "overall_note": _clean_llm_business_value(payload.get("overall_note"), limit=None),
    }


def _clean_risk_clause_priority_attention(value: Any) -> list[dict[str, str]]:
    by_category: dict[str, str] = {}
    raw_items = value if isinstance(value, list) else []
    for raw in raw_items:
        if not isinstance(raw, dict):
            continue
        category = _clean_llm_business_value(raw.get("category"), limit=20)
        suggestion = _clean_llm_business_value(raw.get("suggestion"), limit=None)
        if category in {"优先谈判", "优先管控", "关键证据"} and suggestion:
            by_category[category] = suggestion
    defaults = {
        "优先谈判": "优先谈判影响合同价款调整、范围变化、未定价先施工、违约责任和履约担保索赔的关键条款。",
        "优先管控": "中标后建立报价完整性、工期节点、签证时效、材料送审、进度款申报和安全质量闭环台账。",
        "关键证据": "全过程留存书面指令、会议纪要、影像资料、现场量测、材料下料/采购确认、隐蔽验收和进度影响分析。",
    }
    return [{"category": category, "suggestion": by_category.get(category) or suggestion} for category, suggestion in defaults.items()]


def _risk_clause_level_count(risks: list[dict[str, Any]], level: str) -> int:
    return sum(1 for risk in risks if risk.get("risk_level") == level)


def _risk_clause_distribution_text(risks: list[dict[str, Any]]) -> str:
    high = _risk_clause_level_count(risks, "high")
    medium = _risk_clause_level_count(risks, "medium")
    low = _risk_clause_level_count(risks, "low")
    return f"高风险 {high} 条；中风险 {medium} 条；低风险 {low} 条"


async def _semantic_summary_items_with_cache(
    db: Session,
    project: BidProject,
    run: BidParseRun,
    *,
    requirements: list[TenderRequirement],
    risks: list[TenderRisk],
    business_objects: list[TenderBusinessObject],
    username: str | None,
    trace_id: str | None,
) -> dict[str, Any]:
    provider = (settings.bidding_llm_provider or "deepseek").strip().lower()
    model = _bidding_summary_llm_model()
    metadata: dict[str, Any] = {
        "enabled": bool(settings.feature_bidding_llm_review),
        "provider": provider,
        "model": model,
        "prompt_version": TENDER_ANALYSIS_SUMMARY_LLM_PROMPT_VERSION,
        "status": "disabled",
    }
    if not settings.feature_bidding_llm_review:
        return {"items": {}, "metadata": metadata}
    if provider != "deepseek":
        metadata.update({"status": "skipped", "skip_reason": "bidding_llm_provider_not_deepseek"})
        return {"items": {}, "metadata": metadata}
    if not (settings.deepseek_api_key or "").strip():
        metadata.update({"status": "skipped", "skip_reason": "deepseek_api_key_missing"})
        return {"items": {}, "metadata": metadata}

    cached = _cached_semantic_summary(run, model=model)
    if cached:
        cached_metadata = dict(cached.get("metadata") or {})
        cached_metadata.update({"status": "cached", "enabled": True, "provider": provider, "model": model})
        return {"items": cached.get("items") or {}, "metadata": cached_metadata}

    context = _build_summary_semantic_context(db, project, run, requirements, risks, business_objects)
    candidate_count = sum(len(items) for items in (context.get("candidate_snippets") or {}).values())
    metadata["candidate_count"] = candidate_count
    if candidate_count <= 0:
        metadata.update({"status": "no_candidates"})
        return {"items": {}, "metadata": metadata}

    try:
        raw_payload = await _call_deepseek_tender_summary_extraction(
            context,
            username=username,
            trace_id=trace_id or run.run_uuid,
        )
        items = _clean_semantic_summary_payload(raw_payload, context)
        metadata.update(
            {
                "status": "completed",
                "item_count": len(items),
                "found_count": len([item for item in items.values() if item.get("status") == "found"]),
                "generated_at": datetime.now().isoformat(timespec="seconds"),
            }
        )
        _store_semantic_summary_cache(db, run, items=items, metadata=metadata)
        return {"items": items, "metadata": metadata}
    except Exception as exc:
        metadata.update({"status": "error", "error": (str(exc) or exc.__class__.__name__)[:300]})
        return {"items": {}, "metadata": metadata}


def _bidding_summary_llm_model() -> str:
    return (settings.bidding_llm_model or settings.deepseek_model or "deepseek-v4-pro").strip()


def _cached_semantic_summary(run: BidParseRun, *, model: str) -> dict[str, Any] | None:
    summary = loads_json(run.summary_json, {}) if run.summary_json else {}
    cached = summary.get("tender_analysis_summary_semantic")
    if not isinstance(cached, dict):
        return None
    if cached.get("prompt_version") != TENDER_ANALYSIS_SUMMARY_LLM_PROMPT_VERSION:
        return None
    if cached.get("model") != model:
        return None
    items = cached.get("items")
    if not isinstance(items, dict):
        return None
    return cached


def _store_semantic_summary_cache(
    db: Session,
    run: BidParseRun,
    *,
    items: dict[str, dict[str, Any]],
    metadata: dict[str, Any],
) -> None:
    summary = loads_json(run.summary_json, {}) if run.summary_json else {}
    summary["tender_analysis_summary_semantic"] = {
        "prompt_version": TENDER_ANALYSIS_SUMMARY_LLM_PROMPT_VERSION,
        "model": metadata.get("model"),
        "metadata": metadata,
        "items": items,
    }
    run.summary_json = dumps_json(summary)
    db.add(run)
    db.commit()
    db.refresh(run)


def _build_summary_semantic_context(
    db: Session,
    project: BidProject,
    run: BidParseRun,
    requirements: list[TenderRequirement],
    risks: list[TenderRisk],
    business_objects: list[TenderBusinessObject],
) -> dict[str, Any]:
    source_pool = _summary_semantic_source_pool(db, project, run, requirements, risks, business_objects)
    candidate_snippets: dict[str, list[dict[str, Any]]] = {}
    for catalog in TENDER_ANALYSIS_SUMMARY_ITEM_CATALOG:
        item_key = str(catalog["item_key"])
        definition = SUMMARY_SEMANTIC_ITEM_DEFINITIONS.get(item_key) or {}
        keywords = tuple(definition.get("keywords") or SUMMARY_KEYWORDS.get(item_key, ()))
        scored: list[tuple[int, int, dict[str, Any]]] = []
        for index, source in enumerate(source_pool):
            score = _summary_semantic_source_score(source, item_key=item_key, keywords=keywords)
            if score <= 0:
                continue
            scored.append((score, index, source))
        scored.sort(key=lambda item: (-item[0], item[1]))
        snippets: list[dict[str, Any]] = []
        seen_text: set[str] = set()
        for score, _, source in scored:
            text = _clip(source.get("text"), 700)
            text_key = _stable_text_key(text)
            if not text or text_key in seen_text:
                continue
            seen_text.add(text_key)
            snippets.append(
                {
                    "evidence_id": f"{item_key}:E{len(snippets) + 1}",
                    "source_file": source.get("source_file"),
                    "source_location": source.get("source_location"),
                    "source_kind": source.get("source_kind"),
                    "document_section": source.get("document_section"),
                    "score_hint": score,
                    "text": text,
                }
            )
            if len(snippets) >= 6:
                break
        candidate_snippets[item_key] = snippets
    return {
        "prompt_version": TENDER_ANALYSIS_SUMMARY_LLM_PROMPT_VERSION,
        "task": "bidding_tender_summary_semantic_extraction",
        "project": {
            "project_uuid": project.project_uuid,
            "project_name": project.project_name,
            "tenderer_name": project.tenderer_name,
            "tender_agency": project.tender_agency,
            "project_location": project.project_location,
            "project_type": project.project_type,
        },
        "run_uuid": run.run_uuid,
        "item_definitions": [
            {
                "item_key": str(catalog["item_key"]),
                "item_name": catalog["item_name"],
                "category": catalog["category"],
                "required": bool(catalog["required"]),
                "target": (SUMMARY_SEMANTIC_ITEM_DEFINITIONS.get(str(catalog["item_key"])) or {}).get("target"),
            }
            for catalog in TENDER_ANALYSIS_SUMMARY_ITEM_CATALOG
        ],
        "candidate_snippets": candidate_snippets,
    }


def _summary_semantic_source_pool(
    db: Session,
    project: BidProject,
    run: BidParseRun,
    requirements: list[TenderRequirement],
    risks: list[TenderRisk],
    business_objects: list[TenderBusinessObject],
) -> list[dict[str, Any]]:
    sources: list[dict[str, Any]] = []
    project_text = _project_summary_value(project, "project_overview")
    if project_text:
        sources.append(
            {
                "source_kind": "project",
                "source_file": "项目基础信息",
                "source_location": "-",
                "document_section": "project",
                "text": project_text,
                "item_key_hint": "project_overview",
            }
        )
    for item in requirements:
        sources.append(
            {
                "source_kind": "requirement",
                "source_file": item.source_file,
                "source_location": item.source_location,
                "document_section": item.requirement_type,
                "text": item.parsed_requirement or item.original_text,
                "item_key_hint": None,
            }
        )
    for item in risks:
        sources.append(
            {
                "source_kind": "risk",
                "source_file": item.source_file,
                "source_location": item.source_location,
                "document_section": item.impact_area,
                "text": " ".join(part for part in [item.original_text, item.risk_explanation, item.suggested_action] if part),
                "item_key_hint": None,
            }
        )
    for item in business_objects:
        normalized = loads_json(item.normalized_json, {}) if item.normalized_json else {}
        sources.append(
            {
                "source_kind": "business_object",
                "source_file": item.source_file,
                "source_location": item.source_location,
                "document_section": item.document_section,
                "text": " ".join(part for part in [item.title, item.normalized_value, item.original_text] if part),
                "item_key_hint": normalized.get("analysis_item_key"),
            }
        )
    sources.extend(_summary_file_segment_sources(db, project, run))
    return [source for source in sources if str(source.get("text") or "").strip()]


def _summary_file_segment_sources(db: Session, project: BidProject, run: BidParseRun) -> list[dict[str, Any]]:
    input_file_uuids = [str(item) for item in loads_json(run.input_file_ids_json, []) or [] if item]
    query = db.query(BidProjectFile).filter(BidProjectFile.project_id == project.id, BidProjectFile.parser_status == "parsed")
    if input_file_uuids:
        query = query.filter(BidProjectFile.file_uuid.in_(input_file_uuids))
    files = query.order_by(BidProjectFile.id.asc()).all()
    sources: list[dict[str, Any]] = []
    seen: set[str] = set()
    for file_obj in files:
        segments = loads_json(file_obj.segments_json, []) if file_obj.segments_json else []
        if not isinstance(segments, list):
            continue
        for index, segment in enumerate(segments, start=1):
            if not isinstance(segment, dict):
                continue
            text = re.sub(r"\s+", " ", str(segment.get("text") or "")).strip()
            if len(text) < 8:
                continue
            text_key = _stable_text_key(text)
            if text_key in seen:
                continue
            seen.add(text_key)
            source_location = segment.get("source_location") or f"段落{index}"
            document_section = segment.get("document_section") or segment.get("document_section_label")
            sources.append(
                {
                    "source_kind": "file_segment",
                    "source_id": f"{file_obj.file_uuid}:{index}",
                    "source_file": segment.get("source_file") or file_obj.original_filename,
                    "source_location": source_location,
                    "document_section": document_section,
                    "text": text,
                    "confidence": _summary_segment_confidence(
                        text,
                        {"source_location": source_location, "document_section": document_section},
                    ),
                    "item_key_hint": None,
                }
            )
            if len(sources) >= 220:
                break
        if len(sources) >= 220:
            break
    return sources


def _summary_semantic_source_score(source: dict[str, Any], *, item_key: str, keywords: tuple[str, ...]) -> int:
    text = str(source.get("text") or "")
    section = str(source.get("document_section") or "")
    compact = f"{section} {text}"
    score = 0
    if source.get("item_key_hint") == item_key:
        score += 80
    for keyword in keywords:
        if keyword and keyword in compact:
            score += 16 if len(keyword) >= 4 else 10
    section_bonus = {
        "project_overview": ("overview", "project", "招标公告", "工程概况", "项目概况"),
        "qa_deadline": ("qa", "clarification", "答疑", "澄清", "质疑"),
        "pricing_method": ("pricing", "报价", "承包方式", "商务"),
        "bid_bond": ("bond", "保证金", "保函"),
        "contact_person": ("contact", "联系人", "联系方式"),
        "site_visit": ("site_visit", "踏勘", "现场"),
        "bid_document_requirements": ("bid_document", "投标文件", "投标书的编制", "文件要求"),
        "sealing_requirements": ("sealing", "密封", "标记", "封标"),
        "submission_deadline": ("deadline", "截标", "投标截止", "递交截止"),
        "scoring_weight": ("evaluation", "评标", "评分", "废标", "中标"),
        "construction_period": ("schedule", "工期", "合同工期"),
        "payment_terms": ("payment", "付款", "支付", "合同价款"),
        "pre_bid_clarifications": ("clarification", "澄清", "答疑", "空白", "未填写", "不一致", "重点提醒"),
    }.get(item_key, ())
    if any(str(term).lower() in compact.lower() for term in section_bonus):
        score += 22
    if item_key == "pricing_method" and "工程承包方式" in text:
        score += 60
    if item_key == "payment_terms" and "合同价款" in text and "支付" in text:
        score += 60
    if item_key == "scoring_weight" and _has_any(text, ("废标条件", "确定中标人", "中标候选人")):
        score += 60
    if item_key == "submission_deadline" and re.search(r"20\d{2}[-年/.]\d{1,2}[-月/.]\d{1,2}", text):
        score += 24
    if item_key == "pre_bid_clarifications" and _has_any(text, ("未填写", "空白", "不一致", "需澄清", "补遗")):
        score += 60
    return score


SUMMARY_SEMANTIC_SYSTEM_PROMPT = """你是装饰工程投标中台的招标文件摘要抽取助手。
任务：从给定候选原文片段中，为固定的结构化信息摘要表抽取真正有效的信息。
交付口径：参考投标预算员整理的“投标重要有效信息提取表”，按业务分组输出可直接复核的信息，而不是关键词摘抄。
必须遵守：
1. 只能基于 candidate_snippets 中的原文证据抽取，不得编造、补全或使用常识推断。
2. 不要因为片段里出现关键词就直接选中，要按 item_definitions.target 判断是否真是该信息项。
3. 如果没有明确证据，status 必须为 not_found，extracted_value 和 normalized_value 留空。
4. 每个 summary_items 项必须返回 item_key、status、extracted_value、normalized_value、source_evidence_ids、confidence、reason。
5. source_evidence_ids 只能使用输入里的 evidence_id，例如 project_overview:E1。
6. 对空白、未填写、前后不一致、需补遗确认的信息，不要丢弃；应在对应信息项或 pre_bid_clarifications 中明确写出“未填写/需澄清/存在冲突”。
7. 对标书制作、封标、付款、保证金、工期、评标/废标等长条款，要压缩成业务人员可复核的要点清单，不要整段复制。
8. 输出必须是严格 JSON，不要 Markdown，不要代码块。
返回格式：
{
  "summary_items": [
    {
      "item_key": "project_overview",
      "status": "found|not_found",
      "extracted_value": "面向业务人员的简洁中文结果",
      "normalized_value": "可为空；如能标准化时间/金额/工期则标准化",
      "source_evidence_ids": ["project_overview:E1"],
      "confidence": 0.0,
      "reason": "一句话说明为什么选这条证据，或为什么未找到"
    }
  ]
}
"""


async def _call_deepseek_tender_summary_extraction(
    context: dict[str, Any],
    *,
    username: str | None,
    trace_id: str | None,
) -> dict[str, Any]:
    model = _bidding_summary_llm_model()
    payload = {
        "model": model,
        "messages": [
            {"role": "system", "content": SUMMARY_SEMANTIC_SYSTEM_PROMPT},
            {"role": "user", "content": json.dumps(context, ensure_ascii=False)},
        ],
        "temperature": 0.1,
        "response_format": {"type": "json_object"},
    }
    response = await post_json_via_gateway(
        provider="deepseek",
        model=model,
        endpoint_type="bidding_tender_summary_semantic_extract",
        url=settings.deepseek_chat_url,
        json_payload=payload,
        headers={
            "Authorization": f"Bearer {settings.deepseek_api_key.strip()}",
            "Content-Type": "application/json",
        },
        timeout=settings.bidding_llm_timeout_seconds,
        username=username,
        trace_id=trace_id,
    )
    if not 200 <= response.status_code < 300:
        raise RuntimeError(f"HTTP {response.status_code}")
    content = response.json().get("choices", [{}])[0].get("message", {}).get("content", "")
    return _extract_json_object(content)


def _clean_semantic_summary_payload(payload: dict[str, Any], context: dict[str, Any]) -> dict[str, dict[str, Any]]:
    raw_items = payload.get("summary_items") if isinstance(payload.get("summary_items"), list) else []
    allowed_item_keys = {str(item["item_key"]) for item in TENDER_ANALYSIS_SUMMARY_ITEM_CATALOG}
    snippets_by_id: dict[str, dict[str, Any]] = {}
    allowed_ids_by_key: dict[str, set[str]] = {}
    for item_key, snippets in (context.get("candidate_snippets") or {}).items():
        allowed_ids_by_key[str(item_key)] = set()
        for snippet in snippets or []:
            if not isinstance(snippet, dict):
                continue
            evidence_id = str(snippet.get("evidence_id") or "")
            if not evidence_id:
                continue
            snippets_by_id[evidence_id] = snippet
            allowed_ids_by_key[str(item_key)].add(evidence_id)
    cleaned: dict[str, dict[str, Any]] = {}
    for raw in raw_items:
        if not isinstance(raw, dict):
            continue
        item_key = str(raw.get("item_key") or "").strip()
        if item_key not in allowed_item_keys:
            continue
        status = str(raw.get("status") or "not_found").strip().lower()
        if status not in {"found", "not_found"}:
            status = "not_found"
        extracted_value = _clip(raw.get("extracted_value") or "", 360)
        if status == "found" and not extracted_value:
            status = "not_found"
        source_ids = [
            source_id
            for source_id in _clean_string_list(raw.get("source_evidence_ids"))
            if source_id in allowed_ids_by_key.get(item_key, set())
        ][:3]
        selected_snippets = [snippets_by_id[source_id] for source_id in source_ids if source_id in snippets_by_id]
        cleaned[item_key] = {
            "item_key": item_key,
            "status": status,
            "extracted_value": extracted_value if status == "found" else "",
            "normalized_value": _clip(raw.get("normalized_value") or extracted_value, 360) if status == "found" else "",
            "source_evidence_ids": source_ids,
            "source_file": _join_unique([snippet.get("source_file") for snippet in selected_snippets], limit=2),
            "source_location": _join_unique([snippet.get("source_location") for snippet in selected_snippets], limit=4),
            "evidence_text": _join_unique([snippet.get("text") for snippet in selected_snippets], limit=2),
            "confidence": _safe_confidence(raw.get("confidence")),
            "reason": _clip(raw.get("reason") or "", 220),
        }
    return cleaned


def _extract_json_object(content: Any) -> dict[str, Any]:
    text = str(content or "").strip()
    if not text:
        raise ValueError("EMPTY_LLM_CONTENT")
    text = re.sub(r"^```(?:json)?\s*", "", text)
    text = re.sub(r"\s*```$", "", text).strip()
    try:
        value = json.loads(text)
    except json.JSONDecodeError:
        start = text.find("{")
        end = text.rfind("}")
        if start < 0 or end <= start:
            raise
        value = json.loads(text[start : end + 1])
    if not isinstance(value, dict):
        raise ValueError("LLM_JSON_NOT_OBJECT")
    return value


def _clean_string_list(value: Any) -> list[str]:
    if value is None:
        return []
    if isinstance(value, str):
        raw_items = [value]
    elif isinstance(value, (list, tuple, set)):
        raw_items = list(value)
    else:
        raw_items = [value]
    return _unique_text([str(item).strip() for item in raw_items if str(item).strip()])


def _safe_confidence(value: Any) -> float:
    try:
        number = float(value)
    except (TypeError, ValueError):
        return 0.0
    return round(max(0.0, min(number, 1.0)), 4)


def _build_summary_items(
    project: BidProject,
    requirements: list[TenderRequirement],
    risks: list[TenderRisk],
    business_objects: list[TenderBusinessObject],
    *,
    segment_sources: list[dict[str, Any]] | None = None,
    semantic_items: dict[str, dict[str, Any]] | None = None,
) -> list[dict[str, Any]]:
    candidates = _analysis_candidates(requirements, risks, business_objects, segment_sources=segment_sources)
    semantic_by_key = semantic_items or {}
    rows: list[dict[str, Any]] = []
    for catalog in TENDER_ANALYSIS_SUMMARY_ITEM_CATALOG:
        item_key = str(catalog["item_key"])
        semantic_item = semantic_by_key.get(item_key)
        rule_candidate = _rule_summary_candidate(project, candidates, item_key)
        semantic_found = bool(
            semantic_item
            and semantic_item.get("status") == "found"
            and str(semantic_item.get("extracted_value") or "").strip()
        )
        if semantic_found:
            candidate = _summary_candidate_from_semantic_item(semantic_item)
            extracted_value = _clip(semantic_item.get("extracted_value"), 360) if semantic_item.get("status") == "found" else ""
            normalized_value = _clip(semantic_item.get("normalized_value") or extracted_value, 360) if extracted_value else ""
            extraction_method = "llm_semantic_summary"
        else:
            candidate = rule_candidate
            extracted_value = _clip(candidate.get("text"), 240) if candidate else ""
            normalized_value = _normalize_summary_value(item_key, extracted_value)
            extraction_method = "rule_after_semantic_not_found" if semantic_item else "rule"
        confidence = round(float(candidate.get("confidence") or 0.0), 4) if candidate else 0.0
        evidence_fields = _evidence_fields(candidate, missing_text=f"未从当前解析结果中识别到：{catalog['item_name']}")
        if semantic_found:
            if item_key == "pre_bid_clarifications" and extracted_value:
                evidence_fields["review_status"] = "to_clarify"
            else:
                evidence_fields["review_status"] = "confirmed" if confidence >= 0.82 and extracted_value else "pending"
            evidence_fields["review_note"] = semantic_item.get("reason") or evidence_fields.get("review_note")
        elif semantic_item and extracted_value:
            evidence_fields["review_note"] = _join_unique(
                [
                    "语义摘要未找到明确结果，已使用规则兜底原文证据。",
                    semantic_item.get("reason"),
                    evidence_fields.get("review_note"),
                ],
                limit=3,
            )
        rows.append(
            {
                "row_key": f"summary:{item_key}",
                "table_key": "summary",
                "item_key": item_key,
                "category": catalog["category"],
                "item_name": catalog["item_name"],
                "extracted_value": extracted_value,
                "normalized_value": normalized_value,
                "value_type": catalog["value_type"],
                "is_required": bool(catalog["required"]),
                "downstream": list(catalog["downstream"]),
                "extraction_method": extraction_method,
                **evidence_fields,
            }
        )
    return rows


def _rule_summary_candidate(project: BidProject, candidates: list[dict[str, Any]], item_key: str) -> dict[str, Any] | None:
    project_value = _project_summary_value(project, item_key)
    file_candidate = _best_candidate_for_keywords(
        candidates,
        SUMMARY_KEYWORDS.get(item_key, ()),
        preferred_item_key=item_key,
    )
    if item_key in SUMMARY_COMBINE_ITEM_KEYS:
        file_candidate = _combined_candidate_for_keywords(
            candidates,
            SUMMARY_KEYWORDS.get(item_key, ()),
            preferred_item_key=item_key,
            limit=8 if item_key == "project_overview" else 5 if item_key in {"bid_document_requirements", "scoring_weight", "payment_terms"} else 4,
        ) or file_candidate
    if not project_value:
        return file_candidate
    project_candidate = {
        "text": project_value,
        "source_file": "项目基础信息",
        "source_location": "-",
        "confidence": 0.88,
        "source_kind": "project",
        "source_id": project.project_uuid,
        "business_object_uuid": None,
    }
    if item_key == "project_overview" and file_candidate:
        candidate = dict(file_candidate)
        candidate["text"] = _join_unique([project_value, file_candidate.get("text")], limit=2)
        candidate["confidence"] = max(float(file_candidate.get("confidence") or 0.0), 0.82)
        return candidate
    return project_candidate


def _summary_candidate_from_semantic_item(item: dict[str, Any]) -> dict[str, Any] | None:
    if not isinstance(item, dict):
        return None
    text = str(item.get("evidence_text") or item.get("extracted_value") or "").strip()
    if not text and item.get("status") == "not_found":
        text = str(item.get("reason") or "").strip()
    return {
        "text": text,
        "source_file": item.get("source_file"),
        "source_location": item.get("source_location"),
        "confidence": item.get("confidence") or 0.0,
        "source_kind": "llm_semantic_summary",
        "source_id": item.get("source_evidence_ids") or item.get("source_id"),
        "business_object_uuid": None,
    }


def _build_scoring_items(
    requirements: list[TenderRequirement],
    business_objects: list[TenderBusinessObject],
    segment_sources: list[dict[str, Any]] | None = None,
) -> list[dict[str, Any]]:
    sources: list[dict[str, Any]] = []
    for item in requirements:
        text = item.parsed_requirement or item.original_text
        if item.requirement_type == "evaluation" or _has_any(text, SCORING_KEYWORDS):
            sources.append(
                {
                    "source_kind": "requirement",
                    "source_id": item.id,
                    "source_file": item.source_file,
                    "source_location": item.source_location,
                    "text": text,
                    "title": _derive_scoring_item(text),
                    "confidence": item.confidence,
                    "owner_role": item.owner_role or "经营",
                    "business_object_uuid": None,
                }
            )
    for item in business_objects:
        text = " ".join(part for part in [item.title, item.normalized_value, item.original_text] if part)
        if item.object_subtype == "evaluation" or item.object_type == "bid_rule" and _has_any(text, SCORING_KEYWORDS):
            sources.append(
                {
                    "source_kind": "business_object",
                    "source_id": item.object_uuid,
                    "source_file": item.source_file,
                    "source_location": item.source_location,
                    "text": text,
                    "title": item.title,
                    "confidence": item.confidence,
                    "owner_role": item.owner_role or "经营",
                    "business_object_uuid": item.object_uuid,
                }
            )
    sources.extend(segment_sources or [])
    rows: list[dict[str, Any]] = []
    seen_source_keys: set[str] = set()
    seen_item_keys: set[str] = set()
    seen_group_keys: set[str] = set()
    for source in sources:
        text = str(source["text"])
        source_key = _stable_text_key(text)
        if source_key in seen_source_keys:
            continue
        seen_source_keys.add(source_key)
        for row in _split_scoring_source(source):
            children = [dict(child) for child in row.get("children") or []]
            if children:
                filtered_children = []
                child_keys = []
                for child in children:
                    item_key = _scoring_dedupe_key(child)
                    if item_key in seen_item_keys:
                        continue
                    seen_item_keys.add(item_key)
                    child_keys.append(item_key)
                    filtered_children.append(child)
                if not filtered_children:
                    continue
                group_key = "|".join(child_keys)
                if group_key in seen_group_keys:
                    continue
                seen_group_keys.add(group_key)
                row = dict(row)
                row["children"] = filtered_children
                row["child_count"] = len(filtered_children)
                rows.append(row)
                continue
            item_key = _scoring_dedupe_key(row)
            if item_key in seen_item_keys:
                continue
            seen_item_keys.add(item_key)
            rows.append(row)
    return _number_scoring_rows(rows)


def _collect_scoring_segment_sources(db: Session, project: BidProject, run: BidParseRun) -> list[dict[str, Any]]:
    input_file_uuids = [str(item) for item in loads_json(run.input_file_ids_json, []) or [] if item]
    query = db.query(BidProjectFile).filter(BidProjectFile.project_id == project.id, BidProjectFile.parser_status == "parsed")
    if input_file_uuids:
        query = query.filter(BidProjectFile.file_uuid.in_(input_file_uuids))
    files = query.order_by(BidProjectFile.id.asc()).all()
    sources: list[dict[str, Any]] = []
    seen: set[str] = set()
    for file_obj in files:
        segments = loads_json(file_obj.segments_json, []) if file_obj.segments_json else []
        if not isinstance(segments, list):
            continue
        for index, segment in enumerate(segments, start=1):
            if not isinstance(segment, dict):
                continue
            text = _normalize_scoring_source_text(str(segment.get("text") or ""))
            if not _is_scoring_detail_source(text, str(segment.get("document_section") or "")):
                continue
            key = _stable_text_key(text)
            if key in seen:
                continue
            seen.add(key)
            sources.append(
                {
                    "source_kind": "segment",
                    "source_id": f"{file_obj.file_uuid}:{index}",
                    "source_file": segment.get("source_file") or file_obj.original_filename,
                    "source_location": segment.get("source_location") or f"段落{index}",
                    "text": text,
                    "title": _derive_scoring_item(text),
                    "confidence": _scoring_segment_confidence(text, str(segment.get("document_section") or "")),
                    "owner_role": "经营",
                    "business_object_uuid": None,
                }
            )
    return sources[:80]


def _is_scoring_detail_source(text: str, document_section: str = "") -> bool:
    value = str(text or "")
    if len(value) < 12:
        return False
    has_score_value = bool(
        re.search(r"\d+(?:\.\d+)?\s*分", value)
        or re.search(r"(?:分值|满分|权重|占比)\s*[:：| ]+\s*\d+(?:\.\d+)?", value)
        or re.search(r"[|｜]\s*\d+(?:\.\d+)?\s*[|｜]", value)
    )
    if not has_score_value:
        return False
    has_detail_keyword = _has_any(value, SCORING_DETAIL_KEYWORDS)
    has_bid_context = _has_any(value, SCORING_BID_CONTEXT_KEYWORDS) or document_section == "evaluation"
    if not (has_detail_keyword and has_bid_context):
        return False
    process_noise = _has_any(value, SCORING_PROCESS_NOISE_KEYWORDS)
    strong_bid_context = _has_any(value, ("评标", "评审", "投标", "技术标", "商务标", "报价分", "技术分", "商务分", "资信"))
    if process_noise and not strong_bid_context:
        return False
    return True


def _scoring_segment_confidence(text: str, document_section: str = "") -> float:
    confidence = 0.72
    if document_section == "evaluation":
        confidence += 0.08
    if _has_any(text, ("评分标准", "评分细则", "评审标准", "分值", "满分")):
        confidence += 0.08
    if re.search(r"\d+(?:\.\d+)?\s*分", text):
        confidence += 0.04
    return round(min(confidence, 0.9), 4)


def _split_scoring_source(source: dict[str, Any]) -> list[dict[str, Any]]:
    text = str(source.get("text") or "")
    clauses = _split_scoring_clauses(text)
    child_rows: list[dict[str, Any]] = []
    current_package = "unknown"
    for clause in clauses:
        parsed = _parse_scoring_clause(clause)
        if not parsed:
            continue
        title = parsed["title"]
        package_type = _scoring_package_type(title)
        if package_type == "unknown":
            package_type = _scoring_package_type(clause)
        if package_type == "unknown" and current_package != "unknown":
            package_type = current_package
        elif package_type != "unknown":
            current_package = package_type
        child_rows.append(
            _scoring_row(
                source,
                title=title,
                text=clause,
                package_type=package_type,
                full_score=parsed["full_score"],
                split_from_parent=True,
            )
        )
    if len(child_rows) >= 2:
        group_title = _derive_scoring_group_title(source)
        group_package = _scoring_group_package_type(child_rows)
        for child in child_rows:
            child["parent_scoring_item"] = group_title
        group_row = _scoring_row(
            source,
            title=group_title,
            text=text,
            package_type=group_package,
            full_score="",
            split_from_parent=False,
        )
        group_row.update(
            {
                "is_scoring_group": True,
                "child_count": len(child_rows),
                "children": child_rows,
                "review_status": "pending",
                "review_note": f"已按分值拆分为{len(child_rows)}个评分项，需复核是否存在父子分值重复或漏拆。",
            }
        )
        return [group_row]
    if child_rows:
        for child in child_rows:
            child["split_from_parent"] = False
        return child_rows
    package_type = _scoring_package_type(text)
    return [
        _scoring_row(
            source,
            title=str(source.get("title") or _derive_scoring_item(text)),
            text=text,
            package_type=package_type,
            full_score=None,
            split_from_parent=False,
        )
    ]


def _scoring_row(
    source: dict[str, Any],
    *,
    title: str,
    text: str,
    package_type: str,
    full_score: str | None,
    split_from_parent: bool,
) -> dict[str, Any]:
    evidence_source = dict(source)
    evidence_source["text"] = text
    score_value = _extract_score_value(text) if full_score is None else full_score
    return {
        "row_key": "",
        "table_key": "scoring",
        "scoring_item_key": "",
        "package_type": package_type,
        "scoring_item": title or "评分项",
        "full_score": score_value,
        "scoring_weight": _extract_weight_value(text),
        "scoring_standard": _clip(text, 500),
        "related_bid_section": _related_section_for_package(package_type),
        "estimated_score": "",
        "gap_analysis": _default_gap_analysis(package_type),
        "suggested_action": _default_scoring_action(package_type),
        "owner_role": _owner_role_for_package(package_type, source.get("owner_role")),
        "is_scoring_group": False,
        "split_from_parent": split_from_parent,
        "parent_scoring_item": "",
        **_evidence_fields(evidence_source),
    }


def _split_scoring_clauses(text: str) -> list[str]:
    prepared = re.sub(r"\s+", " ", str(text or "")).strip()
    if not prepared:
        return []
    prepared = re.sub(r"[，,]\s*(?:其中|包括|包含|分为|由)\s*[:：]?", "；", prepared)
    prepared = re.sub(r"(?:其中|包括|包含)\s*[:：]", "", prepared)
    parts = re.split(r"[；;。\n\r]+", prepared)
    clauses: list[str] = []
    for part in parts:
        part = part.strip(" ，,；;")
        if not part:
            continue
        for subpart in re.split(r"[，,]\s*(?=[^，,；;。]{1,80}\d+(?:\.\d+)?\s*分)", part):
            subpart = subpart.strip(" ，,；;")
            if not subpart:
                continue
            clauses.extend(_split_compact_scoring_clause(subpart))
    return [
        clause
        for clause in clauses
        if re.search(r"\d+(?:\.\d+)?\s*分", clause) or _parse_table_score_clause(clause)
    ]


def _normalize_scoring_source_text(text: str) -> str:
    value = re.sub(r"\s+", " ", str(text or "")).strip()
    if not value:
        return ""
    value = value.replace("｜", "|")
    parts = [part.strip() for part in value.split("|")]
    if len(parts) >= 3 and _has_any(value, ("分值", "评分标准", "评审标准", "评分项", "评审项")):
        normalized_parts = []
        for part in parts:
            if re.fullmatch(r"\d+(?:\.\d+)?", part):
                normalized_parts.append(f"{part}分")
            else:
                normalized_parts.append(part)
        value = " | ".join(normalized_parts)
    return value


def _split_compact_scoring_clause(text: str) -> list[str]:
    matches = list(re.finditer(r"\d+(?:\.\d+)?\s*分", text))
    if len(matches) <= 1:
        return [text]
    clauses: list[str] = []
    start = 0
    for match in matches:
        clause = text[start : match.end()].strip(" ，,；;")
        if clause:
            clauses.append(clause)
        start = match.end()
    if start < len(text) and clauses:
        clauses[-1] = f"{clauses[-1]}{text[start:]}"
    return clauses


def _parse_scoring_clause(clause: str) -> dict[str, str] | None:
    score_match = re.search(r"(\d+(?:\.\d+)?)\s*分", clause)
    table_match = None if score_match else _parse_table_score_clause(clause)
    if not score_match and not table_match:
        return None
    if score_match:
        title = _clean_scoring_title(clause[: score_match.start()])
        score = score_match.group(1)
    else:
        title = _clean_scoring_title(table_match["title"])
        score = table_match["score"]
    if not title:
        title = _clean_scoring_title(_derive_scoring_item(clause))
    if not title:
        title = "评分项"
    return {"title": title, "full_score": f"{score}分"}


def _parse_table_score_clause(clause: str) -> dict[str, str] | None:
    value = re.sub(r"\s+", " ", str(clause or "")).strip()
    if "|" not in value:
        return None
    parts = [part.strip() for part in value.replace("｜", "|").split("|") if part.strip()]
    if len(parts) < 2:
        return None
    if _has_any(value, ("序号", "评分项", "评审项", "分值", "评分标准", "评审标准")) and not re.search(r"\d+(?:\.\d+)?", value):
        return None
    for index, part in enumerate(parts):
        match = re.fullmatch(r"(\d+(?:\.\d+)?)\s*(?:分)?", part)
        if not match:
            continue
        previous_parts = [item for item in parts[:index] if not re.fullmatch(r"\d+(?:\.\d+)?", item)]
        title = previous_parts[-1] if previous_parts else (parts[0] if index > 0 else "")
        title = re.sub(r"^第?[一二三四五六七八九十\d]+[、.．)）]?\s*", "", title).strip()
        if title and not _has_any(title, ("序号", "分值", "满分", "权重")):
            return {"title": title, "score": match.group(1)}
    return None


def _clean_scoring_title(value: str) -> str:
    title = re.sub(r"\s+", " ", str(value or "")).strip()
    title = re.sub(r"^第?[一二三四五六七八九十\d]+[、.．)）]\s*", "", title)
    title = re.sub(r"^(?:评分办法|评标办法|评分标准|评审办法|评审标准|评分权重|评标细则)\s*[:：]?", "", title)
    if "：" in title or ":" in title:
        title = re.split(r"[:：]", title)[-1]
    if "|" in title or "｜" in title:
        parts = [part.strip() for part in title.replace("｜", "|").split("|") if part.strip()]
        parts = [part for part in parts if not _has_any(part, ("序号", "评分项", "评审项", "分值", "满分", "权重"))]
        title = parts[-1] if parts else title
    title = re.sub(r"^(?:其中|包括|包含|分为|由|按|对|根据)\s*", "", title)
    title = title.strip(" ，,；;。:：")
    title = re.sub(r"(?:评分|得分|分值|满分|权重|占比)$", "", title)
    title = title.strip(" ，,；;。:：")
    return _clip(title, 60)


def _derive_scoring_group_title(source: dict[str, Any]) -> str:
    title = _clean_scoring_title(str(source.get("title") or ""))
    if title and title != "评分项":
        return title
    derived = _derive_scoring_item(str(source.get("text") or ""))
    return derived if derived and derived != "评分项" else "评分细则汇总"


def _scoring_group_package_type(children: list[dict[str, Any]]) -> str:
    packages = {str(child.get("package_type") or "unknown") for child in children}
    packages.discard("unknown")
    if len(packages) == 1:
        return next(iter(packages))
    if len(packages) > 1:
        return "mixed"
    return "unknown"


def _scoring_dedupe_key(row: dict[str, Any]) -> str:
    title = _normalize_scoring_title_key(str(row.get("scoring_item") or ""))
    score = str(row.get("full_score") or "")
    package_type = str(row.get("package_type") or "unknown")
    if title and score:
        return f"{package_type}:{title}:{score}"
    return f"{package_type}:{title}:{_stable_text_key(str(row.get('scoring_standard') or ''))[:80]}"


def _normalize_scoring_title_key(title: str) -> str:
    value = re.sub(r"(评分办法|评标办法|评分标准|评审办法|评审标准|评分权重|评分|得分|分值|满分|权重|标|部分|内容|项目)", "", title)
    value = re.sub(r"[^0-9A-Za-z\u4e00-\u9fff]", "", value)
    return value[:40]


def _number_scoring_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    counter = 1

    def assign(row: dict[str, Any], parent_row_key: str = "") -> None:
        nonlocal counter
        row["scoring_item_key"] = f"SC-{counter:03d}"
        row["row_key"] = f"{parent_row_key}:child:{counter}" if parent_row_key else f"scoring:{counter}"
        counter += 1
        for child in row.get("children") or []:
            if isinstance(child, dict):
                assign(child, str(row["row_key"]))

    for row in rows:
        assign(row)
    return rows


def _build_risk_clause_items(risks: list[TenderRisk]) -> list[dict[str, Any]]:
    grouped: dict[str, list[tuple[int, TenderRisk]]] = {}
    for index, risk in enumerate(risks, start=1):
        grouped.setdefault(_risk_clause_group_key(risk), []).append((index, risk))
    rows: list[dict[str, Any]] = []
    for group_index, group in enumerate(grouped.values(), start=1):
        child_rows = [_risk_clause_row(risk, index, row_key_prefix="risk_clause_child") for index, risk in group]
        if len(child_rows) == 1:
            row = dict(child_rows[0])
            row["risk_clause_key"] = f"RC-{group_index:03d}"
            rows.append(row)
            continue
        first = group[0][1]
        risk_title = _risk_topic_title(first)
        risk_level = max((row.get("risk_level") or "medium" for row in child_rows), key=lambda level: RISK_LEVEL_ORDER.get(str(level), 2))
        review_statuses = _unique_text([str(row.get("review_status") or "pending") for row in child_rows])
        row = {
            "row_key": f"risk_clause_group:{_risk_clause_group_key(first)}",
            "table_key": "risk_clause",
            "risk_clause_key": f"RC-{group_index:03d}",
            "risk_title": f"{risk_title}（已合并{len(child_rows)}条）",
            "risk_count": len(child_rows),
            "is_grouped": True,
            "clause_text": _clip(child_rows[0].get("clause_text"), 500),
            "clause_section": _join_unique([row.get("clause_section") for row in child_rows], limit=4),
            "risk_category": first.risk_type,
            "risk_level": risk_level,
            "risk_description": _join_unique([row.get("risk_description") for row in child_rows], limit=3),
            "suggested_response": _join_unique([row.get("suggested_response") for row in child_rows], limit=3),
            "owner_role": _majority_text([row.get("owner_role") for row in child_rows]) or _risk_owner_role(first),
            "need_clarification": any(bool(row.get("need_clarification")) for row in child_rows),
            "affects_pricing": any(bool(row.get("affects_pricing")) for row in child_rows),
            "related_bid_package": _majority_text([row.get("related_bid_package") for row in child_rows]) or _risk_related_package(first),
            "source_file": _join_unique([row.get("source_file") for row in child_rows], limit=2),
            "source_location": _join_unique([row.get("source_location") for row in child_rows], limit=6),
            "evidence_text": _join_unique([row.get("evidence_text") for row in child_rows], limit=3),
            "confidence": round(sum(float(row.get("confidence") or 0.0) for row in child_rows) / len(child_rows), 4),
            "review_status": review_statuses[0] if len(review_statuses) == 1 else "mixed",
            "review_note": _join_unique([row.get("review_note") for row in child_rows], limit=3),
            "children": child_rows,
        }
        rows.append(row)
    return rows


def _risk_clause_row(risk: TenderRisk, index: int, *, row_key_prefix: str = "risk_clause") -> dict[str, Any]:
    affects_pricing = _risk_affects_pricing(risk)
    need_clarification = _risk_needs_clarification(risk)
    related_package = _risk_related_package(risk)
    return {
        "row_key": f"{row_key_prefix}:{risk.risk_uuid}",
        "table_key": "risk_clause",
        "risk_clause_key": f"RC-{index:03d}",
        "risk_title": _risk_topic_title(risk),
        "risk_count": 1,
        "is_grouped": False,
        "clause_text": _clip(risk.original_text, 500),
        "clause_section": risk.impact_area or "",
        "risk_category": risk.risk_type,
        "risk_level": risk.risk_level,
        "risk_description": risk.risk_explanation,
        "suggested_response": risk.suggested_action or "",
        "owner_role": _risk_owner_role(risk),
        "need_clarification": need_clarification,
        "affects_pricing": affects_pricing,
        "related_bid_package": related_package,
        "source_file": risk.source_file,
        "source_location": risk.source_location,
        "evidence_text": risk.original_text,
        "confidence": risk.confidence,
        "review_status": risk.review_status or "pending",
        "review_note": risk.reviewer_note,
    }


def _analysis_candidates(
    requirements: list[TenderRequirement],
    risks: list[TenderRisk],
    business_objects: list[TenderBusinessObject],
    *,
    segment_sources: list[dict[str, Any]] | None = None,
) -> list[dict[str, Any]]:
    candidates: list[dict[str, Any]] = []
    for item in requirements:
        text = _join_unique([item.original_text, item.parsed_requirement], limit=2)
        candidates.append(
            {
                "source_kind": "requirement",
                "source_id": item.id,
                "source_file": item.source_file,
                "source_location": item.source_location,
                "text": text,
                "confidence": item.confidence,
                "requirement_type": item.requirement_type,
                "business_object_uuid": None,
            }
        )
    for source in segment_sources or []:
        text = str(source.get("text") or "")
        candidates.append(
            {
                "source_kind": source.get("source_kind") or "file_segment",
                "source_id": source.get("source_id"),
                "source_file": source.get("source_file"),
                "source_location": source.get("source_location"),
                "document_section": source.get("document_section"),
                "text": text,
                "confidence": source.get("confidence") or _summary_segment_confidence(text, source),
                "item_key_hint": source.get("item_key_hint"),
                "business_object_uuid": None,
            }
        )
    for item in risks:
        candidates.append(
            {
                "source_kind": "risk",
                "source_id": item.risk_uuid,
                "source_file": item.source_file,
                "source_location": item.source_location,
                "text": " ".join(part for part in [item.original_text, item.risk_explanation, item.suggested_action] if part),
                "confidence": item.confidence,
                "risk_type": item.risk_type,
                "business_object_uuid": None,
            }
        )
    for item in business_objects:
        normalized = loads_json(item.normalized_json, {})
        candidates.append(
            {
                "source_kind": "business_object",
                "source_id": item.object_uuid,
                "source_file": item.source_file,
                "source_location": item.source_location,
                "text": " ".join(part for part in [item.title, item.normalized_value, item.original_text] if part),
                "confidence": item.confidence,
                "object_type": item.object_type,
                "object_subtype": item.object_subtype,
                "item_key_hint": normalized.get("analysis_item_key"),
                "business_object_uuid": item.object_uuid,
            }
        )
    return candidates


def _summary_segment_confidence(text: str, source: dict[str, Any]) -> float:
    confidence = 0.62
    location = str(source.get("source_location") or "")
    section = str(source.get("document_section") or "")
    if "表" in location or "|" in text:
        confidence += 0.1
    if section and section != "other":
        confidence += 0.05
    if re.search(r"20\d{2}[-年/.]\d{1,2}[-月/.]\d{1,2}", text):
        confidence += 0.04
    if re.search(r"(?:人民币|￥|RMB)?\s*\d+(?:\.\d+)?\s*(?:元|万元|%)", text):
        confidence += 0.04
    return round(min(confidence, 0.82), 4)


def _best_candidate_for_keywords(
    candidates: list[dict[str, Any]],
    keywords: tuple[str, ...],
    *,
    preferred_item_key: str,
) -> dict[str, Any] | None:
    best: dict[str, Any] | None = None
    best_score = 0.0
    for candidate in candidates:
        if not _summary_candidate_is_field_relevant(candidate, preferred_item_key):
            continue
        score = _summary_candidate_score(candidate, keywords, preferred_item_key=preferred_item_key)
        if score > best_score:
            best_score = score
            best = candidate
    return best if best_score >= 1.0 else None


def _combined_candidate_for_keywords(
    candidates: list[dict[str, Any]],
    keywords: tuple[str, ...],
    *,
    preferred_item_key: str,
    limit: int = 4,
) -> dict[str, Any] | None:
    scored: list[tuple[float, int, dict[str, Any]]] = []
    for index, candidate in enumerate(candidates):
        if not _summary_candidate_is_field_relevant(candidate, preferred_item_key):
            continue
        score = _summary_candidate_score(candidate, keywords, preferred_item_key=preferred_item_key)
        if score >= 1.35:
            scored.append((score, index, candidate))
    scored.sort(key=lambda item: (-item[0], item[1]))
    selected: list[dict[str, Any]] = []
    seen: set[str] = set()
    for _, _, candidate in scored:
        text = str(candidate.get("text") or "")
        text_key = _stable_text_key(text)
        if not text or text_key in seen:
            continue
        seen.add(text_key)
        selected.append(candidate)
        if len(selected) >= limit:
            break
    if not selected:
        return None
    if len(selected) == 1:
        return selected[0]
    confidence_values = [float(item.get("confidence") or 0.0) for item in selected]
    return {
        "source_kind": "combined_rule",
        "source_id": _join_unique([item.get("source_id") for item in selected], limit=limit),
        "source_file": _join_unique([item.get("source_file") for item in selected], limit=limit),
        "source_location": _join_unique([item.get("source_location") for item in selected], limit=limit),
        "text": _join_unique([item.get("text") for item in selected], limit=limit),
        "confidence": round(max(confidence_values), 4) if confidence_values else 0.0,
        "business_object_uuid": None,
    }


def _summary_candidate_score(candidate: dict[str, Any], keywords: tuple[str, ...], *, preferred_item_key: str) -> float:
    text = str(candidate.get("text") or "")
    compact = f"{candidate.get('document_section') or ''} {text}"
    score = 0.0
    if candidate.get("item_key_hint") == preferred_item_key:
        score += 3.0
    for keyword in keywords:
        if keyword and keyword in compact:
            score += 1.6 if len(keyword) >= 4 else 1.0
    score += _summary_candidate_item_bonus(preferred_item_key, text)
    score -= _summary_candidate_placeholder_penalty(preferred_item_key, text)
    if candidate.get("source_kind") == "file_segment":
        score += 0.2
    score += min(float(candidate.get("confidence") or 0.0), 1.0) * 0.25
    return score


def _summary_candidate_placeholder_penalty(item_key: str, text: str) -> float:
    value = str(text or "")
    if item_key == "project_overview" and _project_overview_is_placeholder(value):
        return 3.2
    if _has_any(value, ("具体信息未提供", "未在本摘要中提供", "未提供具体信息")):
        return 2.0
    return 0.0


def _summary_candidate_is_field_relevant(candidate: dict[str, Any], item_key: str) -> bool:
    text = re.sub(r"\s+", " ", str(candidate.get("text") or "")).strip()
    if not text:
        return False
    if candidate.get("item_key_hint") == item_key:
        return True
    if _is_generic_parser_hint(text):
        return False
    if item_key == "project_overview":
        if _project_overview_is_placeholder(text):
            return True
        return _has_any(text, ("招标工程名称", "工程名称", "工程地点", "建设地点", "招标单位", "招标人", "总承包单位", "工程规模", "工程特征", "招标范围", "工程范围", "施工范围", "承包范围")) or _project_overview_has_substantive_detail(text)
    if item_key == "qa_deadline":
        has_qa = _has_any(text, ("答疑", "澄清", "质疑", "提问", "疑问", "补遗"))
        has_arrangement = _has_any(text, ("截止", "截至", "前", "提交", "回复", "联系人", "邮箱", "电邮", "未填写", "空白")) or _has_date_or_time(text)
        return has_qa and has_arrangement
    if item_key == "pricing_method":
        if _has_any(text, ("工程承包方式", "承包方式", "合同形式", "合同计价方式", "固定单价包干", "固定总价包干", "综合单价包干", "总价包干", "单价包干")):
            return True
        return _has_any(text, ("工程量清单", "综合单价", "报价方式", "报价形式")) and _has_any(text, ("报价", "单价", "合价", "计量", "结算", "包干", "暂定数量", "不调整"))
    if item_key == "bid_bond":
        has_bond = _has_any(text, ("投标保证金", "投标担保", "投标保函", "履约担保", "保证金", "保函"))
        has_detail = _has_any(text, ("缴纳", "缴交", "提交", "递交", "形式", "现金", "支票", "银行", "账户", "账号", "开户", "没收", "未提交", "无效标", "合同价款", "中标价")) or _has_amount(text)
        return has_bond and has_detail
    if item_key == "contact_person":
        has_role = _has_any(text, ("招投标联系人", "答疑联系人", "踏勘联系人", "现场联系人", "递交联系人", "招标联系人", "代理联系人", "联系人"))
        has_contact = bool(re.search(r"1[3-9]\d{9}", text)) or bool(re.search(r"[\w.+-]+@[\w.-]+", text)) or _has_any(text, ("电话", "手机", "邮箱", "电邮"))
        return has_role and has_contact
    if item_key == "site_visit":
        return _has_any(text, ("踏勘", "现场勘察", "勘察现场")) and _has_any(text, ("时间", "地点", "联系人", "自行", "统一", "组织", "预约", "费用", "现场"))
    if item_key == "bid_document_requirements":
        has_doc = _has_any(text, ("投标文件", "投标书", "商务标", "技术标", "电子标书", "电子文档", "标书", "资质"))
        has_requirement = _has_any(text, ("份数", "正本", "副本", "内容", "组成", "编制", "要求", "格式", "装订", "签字", "盖章", "报价", "U盘", "资格证明"))
        return has_doc and has_requirement
    if item_key == "sealing_requirements":
        has_seal = _has_any(text, ("密封袋", "封套", "封口", "封条", "骑缝", "封标", "密封包装", "密封和标记", "投标文件的密封"))
        has_packaging = _has_any(text, ("商务标", "技术标", "电子标书", "电子文档", "外密封", "内密封", "工程名称", "投标人名称", "注明", "标识", "标记", "公章", "正本", "副本"))
        if _has_any(text, ("废标", "无效标", "否决投标")) and not _has_any(text, ("密封袋", "封套", "封口", "封条", "骑缝", "外密封", "内密封")):
            return False
        return has_seal and has_packaging
    if item_key == "submission_deadline":
        has_submission = _has_any(text, ("投标文件送交截止", "投标文件递交截止", "递交投标文件截止", "投标截止时间", "投标截止日期", "截标", "开标时间", "递交截止", "送达截止"))
        if _has_any(text, ("答疑", "澄清", "质疑", "补遗")) and not _has_any(text, ("递交", "截标", "开标", "送交", "送达")):
            return False
        return has_submission and (_has_date_or_time(text) or _has_any(text, ("空白", "未填写", "地点", "联系人")))
    if item_key == "scoring_weight":
        return _has_any(text, ("评标", "评分", "废标", "无效标", "否决投标", "确定中标人", "中标候选人", "定标")) and _has_any(text, ("办法", "标准", "原则", "条件", "确定", "认定", "综合", "分", "权重", "候选", "最低价"))
    if item_key == "construction_period":
        if _has_any(text, ("违约金", "延误")) and not _has_any(text, ("合同工期", "计划工期", "总工期", "开工", "竣工", "暂定开工")):
            return False
        return _has_any(text, ("合同工期", "计划工期", "总工期", "绝对工期", "暂定开工", "开工日期", "竣工日期", "日历天")) and (_has_duration(text) or _has_date_or_time(text) or _has_any(text, ("开工令", "节假日")))
    if item_key == "payment_terms":
        has_payment = _has_any(text, ("合同价款的支付", "付款方式", "支付方式", "付款条件", "工程款", "进度款", "预付款", "结算款", "质保金", "支付至"))
        has_detail = _has_any(text, ("无预付款", "支付至", "支付比例", "每月", "申报", "申请", "审核", "工作日", "扣除", "累计", "上限", "结算", "质保", "履约保函", "开办费")) or bool(re.search(r"\d+(?:\.\d+)?\s*%", text))
        actual_payment = _has_any(text, ("无预付款", "支付至", "审核值", "15个工作日", "质保金", "结算总价", "合同价款的支付", "累计期中支付"))
        if _has_any(text, ("术语和定义", "指各承建商", "指承建商", "工程进度款：指", "工程结算款：指", "资金计划：指")) and not actual_payment:
            return False
        if _has_any(text, ("支持性表格", "常用工程用表", "格式")) and not _has_any(text, ("支付至", "无预付款", "审核值", "质保金", "合同价款的支付")):
            return False
        return has_payment and has_detail
    if item_key == "pre_bid_clarifications":
        return _has_any(text, ("未填写", "未列明", "空白", "不一致", "冲突", "需澄清", "待定", "另行提供", "需确认", "未明确"))
    return True


def _is_generic_parser_hint(text: str) -> bool:
    return bool(re.fullmatch(r"识别到[^，。；;]{2,30}，请在后续响应矩阵中确认是否满足并绑定证明材料。?", text.strip()))


def _has_date_or_time(text: str) -> bool:
    return bool(re.search(r"20\d{2}[-年/.]\d{1,2}[-月/.]\d{1,2}(?:日)?(?:\s*\d{1,2}[:：]\d{2})?", text)) or bool(re.search(r"\d{1,2}[:：]\d{2}", text))


def _has_amount(text: str) -> bool:
    return bool(re.search(r"(?:RMB|人民币|￥)?\s*\d+(?:,\d{3})*(?:\.\d+)?\s*(?:元|万元|%)", text, flags=re.IGNORECASE))


def _has_duration(text: str) -> bool:
    return bool(re.search(r"\d+(?:\.\d+)?\s*(?:日历天|天|个月|月)", text))


def _project_overview_is_placeholder(text: str) -> bool:
    value = str(text or "")
    has_reference = _has_any(value, ("详见", "参见", "见《协议书》", "见协议书", "见技术要求", "详见《协议书》", "详见《技术要求》"))
    has_missing = _has_any(value, ("具体信息未提供", "具体内容未提供", "未在本摘要中提供", "未明确", "未提供具体信息"))
    return has_reference and (has_missing or len(value) <= 140)


def _project_overview_has_substantive_detail(text: str) -> bool:
    value = str(text or "")
    has_party = _has_any(value, ("东莞市港心房地产开发有限公司", "中建八局", "总承包", "发包人", "招标人", "建设单位"))
    has_area = bool(re.search(r"\d+(?:,\d{3})*(?:\.\d+)?\s*(?:平方米|㎡|m2|M2)", value, flags=re.IGNORECASE))
    has_building_area = _has_any(value, ("商业街区", "6#楼", "32F", "32层", "办公区", "公共走道", "外摆区", "商铺", "卫生间", "休息区", "电梯厅", "会议室", "独立办公室", "水吧", "咖啡吧", "接待室", "前台", "多功能区", "储藏间"))
    has_scope_work = _has_any(value, ("天花", "地面", "墙面", "装饰", "机电", "给排水", "白蚁防治", "精保洁", "防水", "环境治理", "墙体隔断", "推拉门", "窗帘", "不含空调", "消防改造", "办公家具", "软装"))
    has_location = _has_any(value, ("广东省东莞市", "东莞市南城", "东莞国际商务区", "纬三路"))
    return has_party or has_area or (has_building_area and has_scope_work) or has_location


def _summary_candidate_item_bonus(item_key: str, text: str) -> float:
    if item_key == "project_overview":
        bonus = 0.0
        if _has_any(text, ("招标工程名称", "工程地点", "工程规模", "招标范围", "工程范围")):
            bonus += 1.8
        if _project_overview_has_substantive_detail(text):
            bonus += 2.6
        if _has_any(text, ("商业街区", "6#楼", "32F", "办公区")) and _has_any(text, ("约", "平方米", "㎡")):
            bonus += 1.2
        if _has_any(text, ("招标单位", "招标人", "建设单位", "总承包单位")):
            bonus += 1.2
        return bonus
    if item_key == "qa_deadline" and _has_any(text, ("答疑截止", "答疑截至", "质疑截止", "澄清截止", "答疑联系人", "答疑邮箱")):
        return 1.8
    if item_key == "pricing_method" and _has_any(text, ("工程承包方式", "合同形式", "固定单价包干", "固定总价包干", "总价包干")):
        return 2.0
    if item_key == "bid_bond" and _has_any(text, ("投标保证金", "投标担保", "投标保函")):
        return 2.0
    if item_key == "contact_person" and _has_any(text, ("招投标联系人", "答疑联系人", "踏勘联系人", "递交联系人", "联系电话")):
        return 1.4
    if item_key == "site_visit" and _has_any(text, ("现场踏勘", "踏勘时间", "踏勘地点", "自行踏勘")):
        return 1.8
    if item_key == "bid_document_requirements" and _has_any(text, ("投标文件份数", "投标书的编制", "投标文件的组成", "正本", "副本")):
        return 1.8
    if item_key == "sealing_requirements" and _has_any(text, ("投标文件的密封和标记", "密封袋", "封套", "骑缝", "封标")):
        return 1.8
    if item_key == "submission_deadline" and _has_any(text, ("投标文件送交截止", "投标截止", "递交截止", "截标")):
        return 2.0
    if item_key == "scoring_weight" and _has_any(text, ("评标标准", "废标条件", "确定中标人", "否决投标", "无效标")):
        return 1.8
    if item_key == "construction_period" and _has_any(text, ("合同工期", "绝对工期", "日历天")):
        return 1.6
    if item_key == "payment_terms" and _has_any(text, ("合同价款的支付", "付款方式", "进度款", "质保金", "支付至")):
        return 2.0
    if item_key == "pre_bid_clarifications" and _has_any(text, ("未填写", "未列明", "空白", "不一致", "需澄清", "冲突")):
        return 2.0
    return 0.0


def _project_summary_value(project: BidProject, item_key: str) -> str:
    if item_key == "project_overview":
        parts = [
            project.project_name,
            f"招标人：{project.tenderer_name}" if project.tenderer_name else None,
            f"代理机构：{project.tender_agency}" if project.tender_agency else None,
            f"地点：{project.project_location}" if project.project_location else None,
            f"类型：{project.project_type}" if project.project_type else None,
        ]
        return "；".join(part for part in parts if part)
    if item_key == "submission_deadline" and project.tender_deadline_at:
        return project.tender_deadline_at.strftime("%Y-%m-%d %H:%M:%S")
    if item_key == "qa_deadline" and project.inquiry_deadline_at:
        return project.inquiry_deadline_at.strftime("%Y-%m-%d %H:%M:%S")
    return ""


def _evidence_fields(candidate: dict[str, Any] | None, *, missing_text: str = "") -> dict[str, Any]:
    if not candidate:
        return {
            "source_file": None,
            "source_location": None,
            "evidence_text": missing_text,
            "confidence": 0.0,
            "review_status": "needs_revision",
            "review_note": "招标文件当前解析结果未识别到明确原文证据，需人工确认是否未列明或补充答疑。",
            "source_kind": None,
            "source_id": None,
            "business_object_uuid": None,
        }
    confidence = round(float(candidate.get("confidence") or 0.0), 4)
    return {
        "source_file": candidate.get("source_file"),
        "source_location": candidate.get("source_location"),
        "evidence_text": _clip(candidate.get("text"), 500),
        "confidence": confidence,
        "review_status": "pending" if confidence < 0.9 else "confirmed",
        "review_note": None if confidence >= 0.6 else "低置信度，需人工确认。",
        "source_kind": candidate.get("source_kind"),
        "source_id": candidate.get("source_id"),
        "business_object_uuid": candidate.get("business_object_uuid"),
    }


def _build_review_queue(
    summary_items: list[dict[str, Any]],
    scoring_items: list[dict[str, Any]],
    risk_clause_items: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    queue_by_key: dict[str, dict[str, Any]] = {}
    for table_key, items in (("summary", summary_items), ("scoring", scoring_items), ("risk_clause", risk_clause_items)):
        for item in items:
            reasons: list[str] = []
            categories: list[str] = []
            if item.get("review_status") != "confirmed":
                reasons.append(TENDER_ANALYSIS_REVIEW_STATUSES.get(item.get("review_status"), item.get("review_status") or "待复核"))
                categories.append("manual_review")
            if float(item.get("confidence") or 0.0) < 0.65:
                reasons.append("低置信度")
                categories.append("low_confidence")
            if table_key == "summary" and item.get("is_required") and not item.get("extracted_value"):
                reasons.append("关键摘要项缺失")
                categories.append("summary_missing")
            if table_key == "summary" and item.get("item_key") == "pre_bid_clarifications" and item.get("extracted_value"):
                reasons.append("需答疑")
                categories.append("clarification")
            if table_key == "risk_clause" and item.get("risk_level") == "high":
                reasons.append("高风险条款")
                categories.append("high_risk")
            if table_key == "risk_clause" and item.get("affects_pricing"):
                reasons.append("影响报价")
                categories.append("pricing")
            if table_key == "risk_clause" and item.get("need_clarification"):
                reasons.append("需答疑")
                categories.append("clarification")
            if not reasons:
                continue
            title = _review_queue_title(table_key, item)
            category = _primary_review_category(categories)
            group_key = "|".join(
                [
                    table_key,
                    category,
                    title,
                    str(item.get("risk_category") or ""),
                    str(item.get("item_key") or item.get("scoring_item_key") or ""),
                ]
            )
            child = {
                "table_key": table_key,
                "table_label": TENDER_ANALYSIS_TABLE_LABELS[table_key],
                "row_key": item.get("row_key"),
                "title": title,
                "review_status": item.get("review_status"),
                "confidence": item.get("confidence"),
                "reasons": _unique_text(reasons),
                "review_category": category,
                "review_category_label": REVIEW_CATEGORY_LABELS.get(category, "待复核"),
                "source_file": item.get("source_file"),
                "source_location": item.get("source_location"),
            }
            queue_children = [
                {
                    "table_key": table_key,
                    "table_label": TENDER_ANALYSIS_TABLE_LABELS[table_key],
                    "row_key": subitem.get("row_key"),
                    "title": _review_queue_title(table_key, subitem),
                    "review_status": subitem.get("review_status"),
                    "confidence": subitem.get("confidence"),
                    "reasons": _unique_text(reasons),
                    "review_category": category,
                    "review_category_label": REVIEW_CATEGORY_LABELS.get(category, "待复核"),
                    "source_file": subitem.get("source_file"),
                    "source_location": subitem.get("source_location"),
                    "item_count": 1,
                }
                for subitem in item.get("children") or []
            ]
            existing = queue_by_key.get(group_key)
            item_count = int(item.get("risk_count") or item.get("child_count") or len(item.get("children") or []) or 1)
            if existing is None:
                queue_by_key[group_key] = {
                    **child,
                    "item_count": item_count,
                    "children": queue_children,
                }
                continue
            existing["item_count"] = int(existing.get("item_count") or 1) + item_count
            existing["reasons"] = _unique_text(list(existing.get("reasons") or []) + reasons)
            existing["source_file"] = _join_unique([existing.get("source_file"), item.get("source_file")], limit=2)
            existing["source_location"] = _join_unique([existing.get("source_location"), item.get("source_location")], limit=6)
            existing["confidence"] = min(float(existing.get("confidence") or 1.0), float(item.get("confidence") or 1.0))
            existing.setdefault("children", []).append(child)
            existing["children"].extend(queue_children)
    queue = list(queue_by_key.values())
    for row in queue:
        if int(row.get("item_count") or 1) > 1 and "（" not in str(row.get("title") or ""):
            row["title"] = f"{row['title']}（{row['item_count']}项）"
        if not row.get("children"):
            row.pop("children", None)
    return sorted(queue, key=_review_queue_sort_key)[:80]


def _build_quality_summary(
    summary_items: list[dict[str, Any]],
    scoring_items: list[dict[str, Any]],
    risk_clause_items: list[dict[str, Any]],
    review_queue: list[dict[str, Any]],
    run: BidParseRun,
) -> dict[str, Any]:
    summary_missing = [item for item in summary_items if item.get("is_required") and not item.get("extracted_value")]
    risk_counter = Counter(item.get("risk_level") for item in risk_clause_items)
    scoring_leaf_items = _leaf_table_rows(scoring_items)
    package_counter = Counter(item.get("package_type") for item in scoring_leaf_items)
    run_summary = loads_json(run.summary_json, {})
    return {
        "summary_item_count": len(summary_items),
        "summary_extracted_count": len([item for item in summary_items if item.get("extracted_value")]),
        "summary_required_missing_count": len(summary_missing),
        "scoring_item_count": len(scoring_leaf_items),
        "scoring_by_package": dict(package_counter),
        "risk_clause_count": len(risk_clause_items),
        "risk_by_level": dict(risk_counter),
        "risk_affects_pricing_count": len([item for item in risk_clause_items if item.get("affects_pricing")]),
        "risk_need_clarification_count": len([item for item in risk_clause_items if item.get("need_clarification")]),
        "review_queue_count": len(review_queue),
        "document_structure": run_summary.get("document_structure") or {},
    }


def _leaf_table_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    leaf_rows: list[dict[str, Any]] = []
    for row in rows:
        children = [child for child in row.get("children") or [] if isinstance(child, dict)]
        if not children:
            leaf_rows.append(row)
            continue
        leaf_rows.extend(_leaf_table_rows(children))
    return leaf_rows


def _used_business_object_count(*tables: list[dict[str, Any]]) -> int:
    object_uuids: set[str] = set()
    for rows in tables:
        for row in rows:
            value = row.get("business_object_uuid")
            if value:
                object_uuids.add(str(value))
    return len(object_uuids)


def _risk_clause_group_key(risk: TenderRisk) -> str:
    topic = _risk_topic_key(risk)
    if topic != "generic":
        return f"{risk.risk_type}:{topic}"
    compact = re.sub(r"\d+(?:\.\d+)?", "#", _stable_text_key(risk.original_text))
    return f"{risk.risk_type}:{compact[:90]}"


def _risk_topic_key(risk: TenderRisk) -> str:
    text = " ".join(part for part in [risk.risk_type, risk.impact_area, risk.original_text, risk.risk_explanation] if part)
    if risk.risk_type in {"fixed_total_price", "no_price_adjustment"}:
        if _has_any(text, ("综合单价", "单价")):
            return "unit_price_no_adjustment"
        return "fixed_total_contract"
    if risk.risk_type == "omission_liability":
        return "missing_item_included"
    if risk.risk_type == "liquidated_damages":
        if _has_any(text, ("工期", "延期", "延误", "竣工", "节点")):
            return "schedule_delay_penalty"
        if _has_any(text, ("质量", "材料", "验收", "品牌", "样品")):
            return "quality_material_penalty"
        if _has_any(text, ("人员", "项目经理", "更换", "到岗")):
            return "personnel_penalty"
        return "liquidated_damages"
    if risk.risk_type == "claim_time_limit":
        if _has_any(text, ("反索赔", "扣除", "扣款")):
            return "counterclaim_deduction"
        return "claim_document_requirement"
    if risk.risk_type == "material_brand_constraint":
        if _has_any(text, ("甲限", "指定", "品牌")):
            return "owner_limited_brand"
        return "sample_or_approval"
    if risk.risk_type == "design_or_drawing_unclear":
        return "drawing_or_scope_unclear"
    if risk.risk_type == "bid_rejection":
        return "bid_rejection_clause"
    if risk.risk_type == "anonymous_bid":
        return "anonymous_bid_identity"
    return "generic"


def _risk_topic_title(risk: TenderRisk) -> str:
    topic = _risk_topic_key(risk)
    labels = {
        "fixed_total_contract": "固定总价/总价包干边界",
        "unit_price_no_adjustment": "综合单价不调整",
        "missing_item_included": "清单漏项责任",
        "schedule_delay_penalty": "工期延误违约责任",
        "quality_material_penalty": "质量/材料违约责任",
        "personnel_penalty": "人员到岗/更换违约责任",
        "liquidated_damages": "违约金条款",
        "claim_document_requirement": "签证索赔资料要求",
        "counterclaim_deduction": "反索赔/扣款机制",
        "owner_limited_brand": "甲限/指定品牌报价约束",
        "sample_or_approval": "样品封样/认质认价约束",
        "drawing_or_scope_unclear": "图纸或范围边界不清",
        "bid_rejection_clause": "废标条件",
        "anonymous_bid_identity": "技术标暗标身份信息风险",
    }
    return labels.get(topic) or RISK_TYPE_LABELS.get(risk.risk_type) or risk.risk_type or "风险条款"


def _review_queue_title(table_key: str, item: dict[str, Any]) -> str:
    if table_key == "summary":
        value = item.get("item_name") or "摘要信息项"
        if item.get("is_required") and not item.get("extracted_value"):
            return f"{value}缺失"
        return str(value)
    if table_key == "scoring":
        return str(item.get("scoring_item") or "评分细则")
    if table_key == "risk_clause":
        return str(item.get("risk_title") or RISK_TYPE_LABELS.get(str(item.get("risk_category"))) or item.get("risk_category") or "风险条款")
    return str(item.get("title") or item.get("row_key") or "待复核项")


def _primary_review_category(categories: list[str]) -> str:
    order = ["summary_missing", "high_risk", "pricing", "clarification", "low_confidence", "manual_review"]
    for item in order:
        if item in categories:
            return item
    return "manual_review"


def _review_queue_sort_key(row: dict[str, Any]) -> tuple[int, int, str]:
    category_order = {
        "summary_missing": 0,
        "high_risk": 1,
        "pricing": 2,
        "clarification": 3,
        "low_confidence": 4,
        "manual_review": 5,
    }
    table_order = {"summary": 0, "risk_clause": 1, "scoring": 2}
    return (
        category_order.get(str(row.get("review_category")), 9),
        table_order.get(str(row.get("table_key")), 9),
        str(row.get("title") or ""),
    )


def _join_unique(values: list[Any], *, limit: int = 5) -> str:
    unique = _unique_text([str(value) for value in values if value])
    if len(unique) <= limit:
        return "；".join(unique)
    return "；".join(unique[:limit]) + f"；等{len(unique)}项"


def _majority_text(values: list[Any]) -> str:
    cleaned = [str(value) for value in values if value]
    if not cleaned:
        return ""
    return Counter(cleaned).most_common(1)[0][0]


def _normalize_summary_value(item_key: str, value: str) -> str:
    if not value:
        return ""
    if item_key in {"submission_deadline", "qa_deadline"}:
        match = re.search(r"20\d{2}[-年/.]\d{1,2}[-月/.]\d{1,2}(?:日)?(?:\s*\d{1,2}[:：]\d{2})?", value)
        return match.group(0) if match else value
    if item_key == "construction_period":
        match = re.search(r"\d+(?:\.\d+)?\s*(?:日历天|天|个月|月)", value)
        return match.group(0) if match else value
    if item_key == "bid_bond":
        match = re.search(r"(?:RMB|人民币|￥)?\s*\d+(?:\.\d+)?\s*(?:元|万元|%)", value)
        return match.group(0) if match else value
    return value


def _derive_scoring_item(text: str) -> str:
    compact = re.sub(r"\s+", " ", str(text or "")).strip()
    if not compact:
        return "评分项"
    for separator in ("：", ":", "。", "；", ";"):
        if separator in compact:
            compact = compact.split(separator, 1)[0]
            break
    return _clip(compact, 80)


def _scoring_package_type(text: str) -> str:
    value = str(text or "")
    if _has_any(value, ("报价", "价格", "投标报价", "商务报价")):
        return "pricing"
    if _has_any(value, ("技术", "施工组织", "施工方案", "质量", "安全", "进度", "项目经理", "项目团队", "机械设备", "施工部署")):
        return "technical"
    if _has_any(value, ("商务", "资信", "资格", "资质", "信誉", "财务", "服务承诺", "企业业绩", "类似业绩")):
        return "business"
    return "unknown"


def _extract_score_value(text: str) -> str:
    value = str(text or "")
    match = re.search(r"(?:满分|分值|得分)?\s*(\d+(?:\.\d+)?)\s*分", value)
    return f"{match.group(1)}分" if match else ""


def _extract_weight_value(text: str) -> str:
    value = str(text or "")
    match = re.search(r"(?:权重|占比)?\s*(\d+(?:\.\d+)?)\s*%", value)
    return f"{match.group(1)}%" if match else ""


def _related_section_for_package(package_type: str) -> str:
    return {
        "business": "商务标",
        "technical": "技术标",
        "pricing": "报价文件/商务标报价部分",
        "mixed": "商务标/技术标",
    }.get(package_type, "待人工确认")


def _default_gap_analysis(package_type: str) -> str:
    if package_type == "technical":
        return "需结合企业资料库、施工经验和项目团队材料评估可得分。"
    if package_type == "pricing":
        return "需结合报价系统、成本库和报价策略评估可得分。"
    if package_type == "business":
        return "需结合企业资质、业绩、授权和商务附件完整性评估可得分。"
    return "评分归属待人工确认。"


def _default_scoring_action(package_type: str) -> str:
    if package_type == "technical":
        return "转技术标工作台，绑定企业资料库和施工经验模板。"
    if package_type == "pricing":
        return "转报价系统，复核报价得分策略和成本边界。"
    if package_type == "business":
        return "转商务标工作台，补齐资质、业绩和商务响应材料。"
    return "人工确认评分项归属后再分派。"


def _owner_role_for_package(package_type: str, fallback: Any) -> str:
    if package_type == "technical":
        return "技术"
    if package_type == "pricing":
        return "预算"
    return str(fallback or "经营")


def _risk_affects_pricing(risk: TenderRisk) -> bool:
    text = " ".join(part for part in [risk.risk_type, risk.impact_area, risk.original_text, risk.risk_explanation] if part)
    return risk.risk_type in PRICING_RISK_TYPES or _has_any(text, ("报价", "价格", "费用", "总价", "单价", "付款", "结算", "质保金", "材料"))


def _risk_needs_clarification(risk: TenderRisk) -> bool:
    text = " ".join(part for part in [risk.risk_type, risk.original_text, risk.risk_explanation, risk.suggested_action] if part)
    return risk.review_status == "to_clarify" or risk.risk_type in CLARIFICATION_RISK_TYPES or _has_any(text, ("答疑", "澄清", "确认", "不明确", "暂定"))


def _risk_related_package(risk: TenderRisk) -> str:
    if _risk_affects_pricing(risk):
        return "pricing"
    text = " ".join(part for part in [risk.impact_area, risk.original_text, risk.risk_explanation] if part)
    if _has_any(text, ("技术标", "暗标", "施工", "质量", "安全", "材料")):
        return "technical"
    if _has_any(text, ("合同", "付款", "违约", "结算", "质保", "签证", "索赔")):
        return "contract"
    return "business"


def _risk_owner_role(risk: TenderRisk) -> str:
    if _risk_affects_pricing(risk):
        return "预算"
    if risk.risk_type in {"liquidated_damages", "claim_time_limit", "delayed_payment", "advance_funding"}:
        return "法务"
    if _risk_related_package(risk) == "technical":
        return "技术"
    return "经营"


def _has_any(text: str, keywords: tuple[str, ...]) -> bool:
    return any(keyword in str(text or "") for keyword in keywords)


def _stable_text_key(text: str) -> str:
    compact = re.sub(r"\s+", "", str(text or ""))
    return compact[:160]


def _clip(value: Any, limit: int = 160) -> str:
    text = re.sub(r"\s+", " ", str(value or "")).strip()
    if len(text) <= limit:
        return text
    return text[: limit - 1].rstrip() + "…"


def _unique_text(items: list[str]) -> list[str]:
    seen: set[str] = set()
    result: list[str] = []
    for item in items:
        text = str(item or "").strip()
        if not text or text in seen:
            continue
        seen.add(text)
        result.append(text)
    return result
