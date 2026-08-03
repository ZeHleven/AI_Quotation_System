from __future__ import annotations

import json
from io import BytesIO
from xml.etree import ElementTree
from zipfile import ZipFile

import pytest
from openpyxl import Workbook

from app.services.budget_projects import (
    BUDGET_SHEET_ROLE_BILL,
    BUDGET_SHEET_ROLE_MATERIAL_REFERENCE,
    BUDGET_SHEET_ROLE_METADATA,
    BUDGET_SHEET_ROLE_SUMMARY_ANALYSIS,
    _apply_workbook_semantics,
    _batch_counts,
    _ensure_budget_mapping_columns,
    _mapping_by_sheet,
    _sanitize_automatic_quantity_mappings,
    _standard_row_model,
    budget_preview_quote_rows,
    standardize_budget_workbook_bytes,
)
from app.services.requirement_standardizer import standardize_requirement_excel_bytes


_SHEET_XML_NAMESPACE = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"


def _set_formula_cached_values(
    content: bytes,
    patches: dict[tuple[int, str], tuple[object, bool]],
) -> bytes:
    """Add cached values while preserving formula text in a synthetic xlsx.

    openpyxl intentionally does not calculate formulas.  The real Liasheng
    workbook carries both formula text and Excel-cached results, so this helper
    patches the generated worksheet XML to reproduce that contract exactly.
    """

    ElementTree.register_namespace("", _SHEET_XML_NAMESPACE)
    namespace = {"x": _SHEET_XML_NAMESPACE}
    with ZipFile(BytesIO(content), "r") as source:
        members = [(info, source.read(info.filename)) for info in source.infolist()]

    patched_members: list[tuple[object, bytes]] = []
    for info, data in members:
        matching = {
            coordinate: (cached_value, is_error)
            for (sheet_index, coordinate), (cached_value, is_error) in patches.items()
            if info.filename == f"xl/worksheets/sheet{sheet_index}.xml"
        }
        if matching:
            root = ElementTree.fromstring(data)
            for coordinate, (cached_value, is_error) in matching.items():
                cell = root.find(f".//x:c[@r='{coordinate}']", namespace)
                assert cell is not None, (info.filename, coordinate)
                cached = cell.find("x:v", namespace)
                if cached is None:
                    cached = ElementTree.SubElement(cell, f"{{{_SHEET_XML_NAMESPACE}}}v")
                cached.text = str(cached_value)
                if is_error:
                    cell.set("t", "e")
                else:
                    cell.attrib.pop("t", None)
            data = ElementTree.tostring(root, encoding="utf-8", xml_declaration=True)
        patched_members.append((info, data))

    output = BytesIO()
    with ZipFile(output, "w") as target:
        for info, data in patched_members:
            target.writestr(info, data)
    return output.getvalue()


def _liansheng_like_workbook() -> bytes:
    """Build a reduced 8-sheet fixture with Liasheng's structural semantics.

    The real acceptance denominator is 198 bill rows and 99 material-reference
    rows.  This fixture scales that to 2 decor + 2 MEP + 1 measure bill rows and
    3 material-reference rows while keeping the same column roles.
    """

    workbook = Workbook()

    cover = workbook.active
    cover.title = "封面"
    cover["A1"] = "联昇集团办公楼装饰工程"
    cover["A3"] = "工程量清单"

    instructions = workbook.create_sheet("报价说明")
    instructions["A1"] = "报价说明"
    instructions["A2"] = "本表仅用于预算项目导入回归测试"

    summary = workbook.create_sheet("报价汇总")
    summary.append(["序号", "工程名称", "金额"])
    summary.append([1, "装饰工程", "=SUM(装饰清单!Q4:Q5)"])

    decor = workbook.create_sheet("装饰清单")
    decor["A1"] = "装饰工程清单"
    decor.append(
        [
            "序号",
            "分部分项工程",
            "项目特征",
            "单位",
            "主材",
            "工程量",
            None,
            None,
            None,
            None,
            "工程量小计",
            "人工费",
            "主材费",
            "辅材/机械费",
            "管理费+利润",
            "不含税单价",
            "不含税合价",
            "备注",
        ]
    )
    decor.merge_cells("F2:J2")
    decor.append(
        [
            "序号",
            "分部分项工程",
            "项目特征",
            "单位",
            "主材",
            "首层",
            "二层",
            "三层",
            "四层",
            "五层",
            "工程量小计",
            "人工费",
            "主材费",
            "辅材/机械费",
            "管理费+利润",
            "不含税单价",
            "不含税合价",
            "备注",
        ]
    )
    decor.append(
        [
            1,
            "新建120-200mm轻质砖墙",
            "120mm AAC轻质砖",
            "㎡",
            "轻质砖",
            0,
            64.02,
            0,
            0,
            0,
            "=SUM(F4:J4)",
            20,
            30,
            5,
            3,
            58,
            3713.16,
            "名称中的范围不是工程量",
        ]
    )
    decor.append(
        [
            2,
            "地面找平",
            "20mm水泥砂浆找平",
            "㎡",
            "水泥砂浆",
            0,
            900,
            990.32,
            0,
            0,
            "=SUM(F5:J5)",
            12,
            18,
            4,
            2,
            36,
            68051.52,
            "",
        ]
    )

    mep = workbook.create_sheet("机电清单")
    mep["A1"] = "机电工程清单"
    mep.append(
        [
            "序号",
            "分部分项工程",
            "项目特征",
            "单位",
            "工程量",
            None,
            None,
            None,
            None,
            None,
            "人工费",
            "主材费",
            "辅材/机械费",
            "管理费+利润",
            "不含税单价",
            "不含税合价",
            "备注",
        ]
    )
    mep.merge_cells("E2:I2")
    mep.append(
        [
            "序号",
            "分部分项工程",
            "项目特征",
            "单位",
            "首层",
            "二层",
            "三层",
            "四层",
            "五层",
            "工程量小计",
            "人工费",
            "主材费",
            "辅材/机械费",
            "管理费+利润",
            "不含税单价",
            "不含税合价",
            "备注",
        ]
    )
    mep.append(
        [
            1,
            "电力电缆WDZB-YJY-5*25mm2",
            "低烟无卤阻燃铜芯电缆",
            "m",
            0,
            296.9,
            0,
            0,
            0,
            "=SUM(E4:I4)",
            8,
            35,
            2,
            3,
            48,
            14251.2,
            "",
        ]
    )
    mep.append(
        [
            2,
            "电力电缆WDZB-YJY-5*25mm2",
            None,
            None,
            0,
            42,
            0,
            0,
            0,
            "=SUM(E5:I5)",
            3,
            7,
            1,
            1,
            12,
            504,
            "缺单位但总工程量有效",
        ]
    )

    measures = workbook.create_sheet("措施费")
    measures.append(["措施项目清单"])
    measures.append([])
    measures.append(["序号", "项目名称", "项目特征", "单位", "工程量", "综合单价", "合价", "备注"])
    measures.append(["序号", "项目名称", "项目特征", "单位", "工程量", "综合单价", "合价", "备注"])
    measures.append([1, "脚手架措施", "室内满堂脚手架", "项", "=7000+213", 2, 14426, ""])

    cost_analysis = workbook.create_sheet("标前成本分析")
    cost_analysis.append(["标前成本分析"])
    cost_analysis.append(["人工费", "材料费", "合计"])
    cost_analysis.append([100, 200, "=SUM(A3:B3)"])

    materials = workbook.create_sheet("主材表")
    materials.append(["主要材料参考表"])
    materials.append([])
    materials.append(
        [
            "序号",
            "材料名称",
            "规格型号",
            "参考样板",
            "品牌厂家",
            "工程量",
            "单位",
            "主材单价",
            "备注",
            "辅助列",
            "供应商含税报价",
        ]
    )
    materials.append([1, "饰面板", "9mm", "=1+1", "甲厂", 8, "㎡", 88, "", "", 100])
    materials.append([2, "轻钢龙骨", "50系列", "", "乙厂", "=#REF!", "m", 15, "断链数量", "", 17])
    materials.append([3, "乳胶漆", "白色", "", "丙厂", 0, "kg", 22, "明确零数量", "", 25])

    output = BytesIO()
    workbook.save(output)
    workbook.close()
    return _set_formula_cached_values(
        output.getvalue(),
        {
            (4, "K4"): (64.02, False),
            (4, "K5"): (1890.32, False),
            (5, "J4"): (296.9, False),
            (5, "J5"): (42, False),
            (6, "E5"): (7213, False),
            (8, "D4"): (2, False),
            (8, "F5"): ("#REF!", True),
        },
    )


@pytest.fixture(scope="module")
def liansheng_preview_contract():
    content = _liansheng_like_workbook()
    preview = standardize_requirement_excel_bytes(content, filename="联昇缩小回归样本.xlsx")
    preview = _apply_workbook_semantics(preview, content)
    preview = _ensure_budget_mapping_columns(preview)
    preview = _sanitize_automatic_quantity_mappings(preview)
    mappings = _mapping_by_sheet(preview)
    rows = [
        _standard_row_model(
            batch_id=1,
            row=row,
            sort_order=index,
            mapping_revision=0,
            sheet_mapping=mappings[row["source_sheet"]],
        )
        for index, row in enumerate(preview["rows"])
    ]
    return preview, mappings, rows


def test_liansheng_sheet_roles_and_formal_row_denominator(liansheng_preview_contract):
    _preview, mappings, rows = liansheng_preview_contract
    assert {name: mapping["sheet_role"] for name, mapping in mappings.items()} == {
        "封面": BUDGET_SHEET_ROLE_METADATA,
        "报价说明": BUDGET_SHEET_ROLE_METADATA,
        "报价汇总": BUDGET_SHEET_ROLE_SUMMARY_ANALYSIS,
        "装饰清单": BUDGET_SHEET_ROLE_BILL,
        "机电清单": BUDGET_SHEET_ROLE_BILL,
        "措施费": BUDGET_SHEET_ROLE_BILL,
        "标前成本分析": BUDGET_SHEET_ROLE_SUMMARY_ANALYSIS,
        "主材表": BUDGET_SHEET_ROLE_MATERIAL_REFERENCE,
    }

    counts = _batch_counts(rows)
    assert counts["standard_item_count"] == 5
    assert counts["valid_quantity_count"] == 5
    assert sum(row.sheet_role == BUDGET_SHEET_ROLE_MATERIAL_REFERENCE for row in rows if row.row_type == "reference_row") == 5
    assert all(row.sheet_role == BUDGET_SHEET_ROLE_BILL for row in rows if row.is_standard_item)
    assert not any(row.is_standard_item for row in rows if row.sheet_role != BUDGET_SHEET_ROLE_BILL)

    repeated = [row for row in rows if row.row_type == "repeated_header"]
    assert {(row.source_sheet, row.raw_row_index) for row in repeated} >= {("措施费", 4)}
    assert all(not row.is_standard_item for row in repeated)
    second_level_headers = [
        row
        for row in rows
        if (row.source_sheet, row.raw_row_index) in {("装饰清单", 3), ("机电清单", 3)}
    ]
    assert len(second_level_headers) == 2
    assert all(not row.is_standard_item for row in second_level_headers)


def test_liansheng_total_quantity_wins_and_explicit_spec_keeps_full_name(liansheng_preview_contract):
    _preview, mappings, rows = liansheng_preview_contract
    assert mappings["装饰清单"]["preferred_quantity_column"] == "K"
    assert mappings["机电清单"]["preferred_quantity_column"] == "J"
    assert mappings["措施费"]["preferred_quantity_column"] == "E"

    formal = {(row.source_sheet, row.raw_row_index): row for row in rows if row.is_standard_item}
    decor = formal[("装饰清单", 4)]
    assert decor.item_name == "新建120-200mm轻质砖墙"
    assert decor.spec == "120mm AAC轻质砖"
    assert decor.raw_quantity == "64.02"
    assert float(decor.calculation_quantity) == 64.02
    assert json.loads(decor.quantity_source_json)["budget"]["column"] == "K"

    cable = formal[("机电清单", 4)]
    assert cable.item_name == "电力电缆WDZB-YJY-5*25mm2"
    assert cable.spec == "低烟无卤阻燃铜芯电缆"
    assert cable.raw_quantity == "296.9"
    assert float(cable.calculation_quantity) == 296.9
    assert json.loads(cable.quantity_source_json)["budget"]["column"] == "J"

    missing_unit = formal[("机电清单", 5)]
    assert missing_unit.item_name == "电力电缆WDZB-YJY-5*25mm2"
    assert missing_unit.spec in {None, ""}
    assert missing_unit.raw_quantity == "42"
    assert float(missing_unit.calculation_quantity) == 42
    assert missing_unit.quantity_status == "valid"
    assert missing_unit.unit is None
    assert "MISSING_UNIT" in json.loads(missing_unit.warnings_json)


def test_liansheng_chat_quote_rows_reuse_budget_quantity_semantics():
    content = _liansheng_like_workbook()
    preview = standardize_budget_workbook_bytes(
        content,
        filename="联昇缩小回归样本.xlsx",
    )
    rows = budget_preview_quote_rows(preview)

    assert len(rows) == 5
    assert {row["source_sheet"] for row in rows} == {"装饰清单", "机电清单", "措施费"}
    assert not any(row["source_sheet"] == "主材表" for row in rows)
    assert all(row["quantity"] is not None for row in rows)

    formal = {
        (row["source_sheet"], row["raw_row_index"]): row
        for row in rows
    }
    assert formal[("装饰清单", 4)]["quantity"] == 64.02
    assert formal[("装饰清单", 5)]["quantity"] == 1890.32
    assert formal[("机电清单", 4)]["quantity"] == 296.9
    assert formal[("机电清单", 5)]["quantity"] == 42.0
    assert formal[("措施费", 5)]["quantity"] == 7213.0
    assert formal[("装饰清单", 4)]["quantity_source"]["budget"]["column"] == "K"
    assert formal[("机电清单", 4)]["quantity_source"]["budget"]["column"] == "J"
    assert formal[("措施费", 5)]["quantity_source"]["budget"]["column"] == "E"


def test_liansheng_locked_columns_and_broken_formula_evidence(liansheng_preview_contract):
    preview, mappings, rows = liansheng_preview_contract
    assert mappings["装饰清单"]["budget_locked_ignore_reasons"] == {
        "F": "LAYER_QUANTITY_COLUMN",
        "G": "LAYER_QUANTITY_COLUMN",
        "H": "LAYER_QUANTITY_COLUMN",
        "I": "LAYER_QUANTITY_COLUMN",
        "J": "LAYER_QUANTITY_COLUMN",
        "L": "PRICE_AMOUNT_COLUMN",
        "M": "PRICE_AMOUNT_COLUMN",
        "N": "PRICE_AMOUNT_COLUMN",
        "O": "PRICE_AMOUNT_COLUMN",
        "P": "PRICE_AMOUNT_COLUMN",
        "Q": "PRICE_AMOUNT_COLUMN",
    }
    assert mappings["机电清单"]["budget_locked_ignore_reasons"] == {
        "E": "LAYER_QUANTITY_COLUMN",
        "F": "LAYER_QUANTITY_COLUMN",
        "G": "LAYER_QUANTITY_COLUMN",
        "H": "LAYER_QUANTITY_COLUMN",
        "I": "LAYER_QUANTITY_COLUMN",
        "K": "PRICE_AMOUNT_COLUMN",
        "L": "PRICE_AMOUNT_COLUMN",
        "M": "PRICE_AMOUNT_COLUMN",
        "N": "PRICE_AMOUNT_COLUMN",
        "O": "PRICE_AMOUNT_COLUMN",
        "P": "PRICE_AMOUNT_COLUMN",
    }
    assert mappings["措施费"]["budget_locked_ignore_reasons"] == {
        "F": "PRICE_AMOUNT_COLUMN",
        "G": "PRICE_AMOUNT_COLUMN",
    }
    material_locks = mappings["主材表"]["budget_locked_ignore_reasons"]
    assert material_locks["D"] == "FORMULA_COLUMN"
    assert material_locks["H"] == "PRICE_AMOUNT_COLUMN"
    assert material_locks["K"] == "PRICE_AMOUNT_COLUMN"
    assert "F" not in material_locks

    source_row = next(
        row
        for row in preview["rows"]
        if row["source_sheet"] == "主材表" and row["raw_row_index"] == 5
    )
    formula_cell = next(cell for cell in source_row["raw_cells"] if cell["column"] == "F")
    assert (source_row["source_sheet"], source_row["raw_row_index"], formula_cell["column"]) == (
        "主材表",
        5,
        "F",
    )
    assert formula_cell["raw_formula"] == "=#REF!"
    # The standardizer deliberately normalizes cached Excel error tokens to an
    # empty usable value; raw formula + explicit error code retain the evidence.
    assert formula_cell["cached_value"] == ""
    assert formula_cell["formula_error"] == "BROKEN_FORMULA_REF"
    assert "BROKEN_FORMULA_REF" in source_row["warnings"]

    persisted = next(
        row for row in rows if row.source_sheet == "主材表" and row.raw_row_index == 5
    )
    persisted_cell = next(
        cell for cell in json.loads(persisted.raw_cells_json) if cell["column"] == "F"
    )
    assert persisted_cell["raw_formula"] == "=#REF!"
    assert persisted_cell["cached_value"] == ""
    assert persisted_cell["formula_error"] == "BROKEN_FORMULA_REF"
