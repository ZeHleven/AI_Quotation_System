from __future__ import annotations

import json
import sys
from pathlib import Path

from openpyxl import load_workbook

from app.services.drawing_pdf_standard_bill_export import (
    build_standard_bill_preview_report,
    write_standard_bill_preview_outputs,
)
from scripts import biz2x_pdf_standard_bill_preview


def _v2_report() -> dict[str, object]:
    return {
        "summary": {
            "quantity_status": "deferred_until_three_fields_accepted",
        },
        "human_style_rows": [
            {
                "row_id": "PDFV2-000001",
                "division": "floor",
                "item_name": "tile floor",
                "feature": "600x600 tile",
                "unit": "m2",
                "quantity": "",
                "standard_code": "GBT50854-2024",
                "standard_item_code": "011102003",
                "standard_item_name": "block floor",
                "standard_match_score": 80.0,
                "evidence_ids": "EVD-001",
                "source_files": "drawing.pdf",
                "review_note": "quantity deferred",
            },
            {
                "row_id": "PDFV2-000002",
                "division": "custom",
                "item_name": "unmapped custom item",
                "feature": "custom feature",
                "unit": "item",
                "quantity": "",
                "standard_code": "",
                "standard_item_code": "",
                "standard_item_name": "",
                "standard_match_score": "",
                "evidence_ids": "EVD-002",
                "source_files": "drawing.pdf",
                "review_note": "",
            },
        ],
        "three_field_acceptance_report": {
            "comparison_rows": [
                {
                    "status": "matched_three_fields",
                    "candidate_row_no": 1,
                    "issue": "passed",
                },
                {
                    "status": "missing_candidate",
                    "candidate_row_no": 2,
                    "issue": "missing in answer",
                },
            ]
        },
    }


def _failed_gate() -> dict[str, object]:
    return {
        "can_enable_quantity": False,
        "summary": {
            "quantity_status": "deferred_until_three_fields_accepted",
        },
    }


def _passed_gate() -> dict[str, object]:
    return {
        "can_enable_quantity": True,
        "summary": {
            "quantity_status": "ready_after_three_field_acceptance",
        },
    }


def test_standard_bill_preview_is_review_only_when_three_field_gate_failed():
    report = build_standard_bill_preview_report(_v2_report(), gate_report=_failed_gate())

    assert report["safe_for_final_standard_bill"] is False
    assert report["summary"]["export_mode"] == "review_only"
    assert report["summary"]["standard_mapped_count"] == 1
    assert report["summary"]["standard_unmapped_count"] == 1
    assert report["bill_rows"][0]["export_status"] == "review_only_three_field_gate_failed"
    assert report["bill_rows"][0]["quantity"] == ""
    assert "三字段门禁未通过" in report["bill_rows"][0]["review_note"]


def test_standard_bill_preview_marks_final_candidate_only_after_gate_and_mapping():
    report = build_standard_bill_preview_report(_v2_report(), gate_report=_passed_gate())

    assert report["safe_for_final_standard_bill"] is False
    assert report["summary"]["final_candidate_count"] == 1
    assert report["bill_rows"][0]["export_status"] == "final_candidate"
    assert report["bill_rows"][1]["export_status"] == "standard_unmapped_review"


def test_standard_bill_preview_writes_workbook(tmp_path: Path):
    report = build_standard_bill_preview_report(_v2_report(), gate_report=_failed_gate())

    outputs = write_standard_bill_preview_outputs(report, tmp_path, stem="bill")

    assert Path(outputs["json"]).exists()
    assert Path(outputs["csv"]).exists()
    assert Path(outputs["markdown"]).exists()
    assert Path(outputs["xlsx"]).exists()
    workbook = load_workbook(outputs["xlsx"])
    assert workbook.sheetnames == ["summary", "standard_bill_preview"]
    assert workbook["standard_bill_preview"]["A1"].value == "序号"
    assert workbook["standard_bill_preview"]["F2"].value == "011102003"


def test_standard_bill_preview_cli(tmp_path: Path, monkeypatch, capsys):
    v2_json = tmp_path / "v2.json"
    gate_json = tmp_path / "gate.json"
    v2_json.write_text(json.dumps(_v2_report(), ensure_ascii=False), encoding="utf-8")
    gate_json.write_text(json.dumps(_failed_gate(), ensure_ascii=False), encoding="utf-8")
    output_dir = tmp_path / "cli_out"
    monkeypatch.setattr(
        sys,
        "argv",
        [
            "biz2x_pdf_standard_bill_preview.py",
            "--v2-json",
            str(v2_json),
            "--gate-json",
            str(gate_json),
            "--output-dir",
            str(output_dir),
            "--timestamp",
            "fixed",
        ],
    )

    assert biz2x_pdf_standard_bill_preview.main() == 0
    payload = json.loads(capsys.readouterr().out)

    assert payload["safe_for_final_standard_bill"] is False
    assert payload["summary"]["standard_bill_row_count"] == 2
    assert Path(payload["outputs"]["xlsx"]).exists()
