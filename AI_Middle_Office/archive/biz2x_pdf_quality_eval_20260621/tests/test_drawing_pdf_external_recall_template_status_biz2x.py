from __future__ import annotations

import json
import sys
from pathlib import Path

from openpyxl import load_workbook

from app.services.drawing_pdf_external_recall_template import (
    build_external_recall_template,
    write_external_recall_template_outputs,
)
from app.services.drawing_pdf_external_recall_template_status import (
    build_external_recall_template_status,
    build_external_recall_template_status_from_path,
    write_external_recall_template_status_outputs,
)
from app.services.drawing_pdf_gap_recall_importer import load_external_recall_results
from scripts import biz2x_pdf_external_recall_template_status


def _recall_plan(tmp_path: Path) -> dict[str, object]:
    image = tmp_path / "page.png"
    image.write_bytes(b"fake image bytes")
    return {
        "plan_rows": [
            {
                "task_no": 1,
                "gap_no": 11,
                "gap_priority": "P1_missing_core",
                "gap_type": "missing_candidate",
                "recommended_pass": "door_window_demolition",
                "source_file": "drawing.pdf",
                "page": 1,
                "tile_id": "p001_whole",
                "tile_type": "whole_page_preview",
                "image_path": str(image),
                "answer_item_name": "door demolition",
                "answer_feature": "remove stainless glass door",
                "answer_unit": "set",
            },
            {
                "task_no": 2,
                "gap_no": 12,
                "gap_priority": "P2_missing_mep",
                "gap_type": "missing_candidate",
                "recommended_pass": "fixture_valve_schedule",
                "source_file": "drawing.pdf",
                "page": 1,
                "tile_id": "p001_whole",
                "tile_type": "whole_page_preview",
                "image_path": str(image),
                "answer_item_name": "",
                "answer_feature": "",
                "answer_unit": "",
            },
        ]
    }


def _template_xlsx(tmp_path: Path) -> str:
    report = build_external_recall_template(_recall_plan(tmp_path))
    outputs = write_external_recall_template_outputs(report, tmp_path / "template", stem="template")
    return outputs["xlsx"]


def test_external_recall_template_status_treats_answer_columns_as_reference_only(tmp_path: Path):
    payload = load_external_recall_results(_template_xlsx(tmp_path))
    report = build_external_recall_template_status(payload)

    assert report["summary"]["input_row_count"] == 2
    assert report["summary"]["importable_row_count"] == 0
    assert report["summary"]["answer_only_count"] == 1
    assert report["summary"]["blank_task_count"] == 1
    assert report["summary"]["answer_columns_count_as_evidence"] is False
    assert report["summary"]["ready_for_external_import"] is False
    assert report["summary"]["missing_field_counts"] == {
        "item_hint": 2,
        "spec_or_method": 2,
        "suggested_unit": 2,
        "text": 2,
    }
    assert report["summary"]["pass_group_count"] == 2
    assert report["summary"]["source_page_group_count"] == 1
    assert report["status_rows"][0]["fill_status"] == "answer_only_reference"
    assert report["status_rows"][1]["fill_status"] == "blank_task"
    by_pass = {row["group_key"]: row for row in report["pass_summary_rows"]}
    assert by_pass["door_window_demolition"]["answer_only_count"] == 1
    assert by_pass["fixture_valve_schedule"]["blank_task_count"] == 1


def test_external_recall_template_status_detects_filled_rows(tmp_path: Path):
    xlsx_path = _template_xlsx(tmp_path)
    workbook = load_workbook(xlsx_path)
    sheet = workbook["external_recall_template"]
    headers = {cell.value: cell.column for cell in sheet[1]}
    sheet.cell(row=2, column=headers["item_hint"]).value = "door demolition"
    sheet.cell(row=2, column=headers["spec_or_method"]).value = "remove stainless glass door"
    sheet.cell(row=2, column=headers["text"]).value = "door schedule note: remove stainless glass door"
    sheet.cell(row=3, column=headers["text"]).value = "water meter DN25"
    workbook.save(xlsx_path)
    workbook.close()

    report = build_external_recall_template_status_from_path(xlsx_path)

    assert report["summary"]["importable_row_count"] == 2
    assert report["summary"]["strong_importable_row_count"] == 1
    assert report["summary"]["weak_importable_row_count"] == 1
    assert report["summary"]["ready_for_external_import"] is True
    assert report["summary"]["missing_field_counts"] == {
        "item_hint": 1,
        "spec_or_method": 1,
        "suggested_unit": 2,
        "text": 0,
    }
    assert report["status_rows"][0]["fill_status"] == "importable"
    assert report["status_rows"][1]["fill_status"] == "importable_weak"


def test_external_recall_template_status_writes_outputs(tmp_path: Path):
    report = build_external_recall_template_status_from_path(_template_xlsx(tmp_path))

    outputs = write_external_recall_template_status_outputs(report, tmp_path / "out", stem="status")

    assert Path(outputs["json"]).exists()
    assert Path(outputs["csv"]).exists()
    assert Path(outputs["markdown"]).exists()
    assert Path(outputs["xlsx"]).exists()
    workbook = load_workbook(outputs["xlsx"])
    assert workbook.sheetnames == ["summary", "template_status", "pass_summary", "source_page_summary"]
    assert workbook["template_status"]["B2"].value == "answer_only_reference"
    assert workbook["pass_summary"]["A1"].value == "group_key"


def test_external_recall_template_status_cli(tmp_path: Path, monkeypatch, capsys):
    xlsx_path = _template_xlsx(tmp_path)
    output_dir = tmp_path / "cli_out"
    monkeypatch.setattr(
        sys,
        "argv",
        [
            "biz2x_pdf_external_recall_template_status.py",
            "--external-template",
            xlsx_path,
            "--output-dir",
            str(output_dir),
            "--timestamp",
            "fixed",
        ],
    )

    assert biz2x_pdf_external_recall_template_status.main() == 0
    payload = json.loads(capsys.readouterr().out)

    assert payload["ok"] is True
    assert payload["summary"]["input_row_count"] == 2
    assert payload["summary"]["ready_for_external_import"] is False
    assert Path(payload["outputs"]["xlsx"]).exists()
