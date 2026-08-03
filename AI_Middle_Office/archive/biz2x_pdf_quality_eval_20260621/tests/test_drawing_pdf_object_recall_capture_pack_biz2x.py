from __future__ import annotations

import json
import sys
from pathlib import Path

from openpyxl import load_workbook

from app.services.drawing_pdf_object_recall_capture_pack import (
    build_object_recall_capture_pack,
    write_object_recall_capture_pack_outputs,
)
from scripts import biz2x_pdf_object_recall_capture_pack


def test_object_recall_capture_pack_is_answer_blind():
    report = build_object_recall_capture_pack(_workbench_rows())

    assert report["summary"]["source_workbench_row_count"] == 3
    assert report["summary"]["capture_source_row_count"] == 2
    assert report["summary"]["capture_call_count"] == 1
    assert report["summary"]["target_fields_in_prompt"] is False
    row = report["capture_rows"][0]
    assert row["task_nos"] == "1;2"
    assert row["target_fields_in_prompt"] is False
    assert "Do not write unknown" in row["prompt_text"]
    assert "Removed Finishes" in row["prompt_text"]
    serialized = json.dumps(report, ensure_ascii=False)
    assert "SECRET_TARGET_WALL" not in serialized
    assert "SECRET_TARGET_DOOR" not in serialized
    assert "secret feature" not in serialized


def test_object_recall_capture_pack_outputs_prompt_and_template(tmp_path: Path):
    report = build_object_recall_capture_pack(_workbench_rows())

    outputs = write_object_recall_capture_pack_outputs(report, tmp_path / "out", stem="capture")

    assert Path(outputs["json"]).exists()
    assert Path(outputs["capture_csv"]).exists()
    assert Path(outputs["evidence_template_csv"]).exists()
    assert Path(outputs["xlsx"]).exists()
    prompt_files = sorted(Path(outputs["prompt_dir"]).glob("*.txt"))
    assert len(prompt_files) == 1
    assert "SECRET_TARGET_WALL" not in prompt_files[0].read_text(encoding="utf-8")

    workbook = load_workbook(outputs["xlsx"])
    assert workbook.sheetnames == ["summary", "capture_tasks", "blank_evidence_template"]
    capture_sheet = workbook["capture_tasks"]
    capture_headers = {cell.value: cell.column for cell in capture_sheet[1]}
    assert capture_sheet.cell(row=2, column=capture_headers["task_nos"]).value == "1;2"
    assert capture_sheet.cell(row=2, column=capture_headers["target_fields_in_prompt"]).value is False

    template_sheet = workbook["blank_evidence_template"]
    template_headers = {cell.value: cell.column for cell in template_sheet[1]}
    assert template_sheet.cell(row=2, column=template_headers["call_no"]).value == 1
    assert template_sheet.cell(row=2, column=template_headers["evidence_text"]).value is None


def test_object_recall_capture_pack_cli(tmp_path: Path, monkeypatch, capsys):
    workbench_json = tmp_path / "workbench.json"
    workbench_json.write_text(json.dumps({"workbench_rows": _workbench_rows()}, ensure_ascii=False), encoding="utf-8")
    output_dir = tmp_path / "cli_out"
    monkeypatch.setattr(
        sys,
        "argv",
        [
            "biz2x_pdf_object_recall_capture_pack.py",
            "--object-workbench",
            str(workbench_json),
            "--output-dir",
            str(output_dir),
            "--timestamp",
            "fixed",
        ],
    )

    assert biz2x_pdf_object_recall_capture_pack.main() == 0
    payload = json.loads(capsys.readouterr().out)

    assert payload["ok"] is True
    assert payload["summary"]["capture_call_count"] == 1
    assert payload["summary"]["target_fields_in_prompt"] is False
    assert Path(payload["outputs"]["xlsx"]).exists()


def _workbench_rows() -> list[dict[str, object]]:
    return [
        {
            "task_no": 1,
            "object_class": "finish_wall",
            "recommended_pass": "finish_schedule",
            "source_file": "drawing.pdf",
            "page": "1",
            "tile_id": "p001_whole",
            "image_path": "page.png",
            "image_exists": True,
            "target_item_name": "SECRET_TARGET_WALL",
            "target_feature": "secret feature wall",
            "target_unit": "m2",
            "fill_status": "answer_only_reference",
            "ready_for_import": "false",
        },
        {
            "task_no": 2,
            "object_class": "door_window_demolition",
            "recommended_pass": "finish_schedule",
            "source_file": "drawing.pdf",
            "page": "1",
            "tile_id": "p001_whole",
            "image_path": "page.png",
            "image_exists": True,
            "target_item_name": "SECRET_TARGET_DOOR",
            "target_feature": "secret feature door",
            "target_unit": "set",
            "fill_status": "answer_only_reference",
            "ready_for_import": "false",
        },
        {
            "task_no": 3,
            "object_class": "finish_floor",
            "recommended_pass": "finish_schedule",
            "source_file": "drawing.pdf",
            "page": "1",
            "tile_id": "p001_whole",
            "image_path": "page.png",
            "image_exists": True,
            "target_item_name": "SECRET_ALREADY_READY",
            "target_feature": "secret already ready",
            "target_unit": "m2",
            "fill_status": "importable",
            "ready_for_import": "true",
            "evidence_text": "visible floor finish",
        },
    ]
