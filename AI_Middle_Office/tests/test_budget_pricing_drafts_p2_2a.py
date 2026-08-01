from __future__ import annotations

import asyncio
import json
import importlib.util
from decimal import Decimal
from io import BytesIO
from pathlib import Path
from types import SimpleNamespace
from uuid import uuid4

import pytest
from openpyxl import load_workbook
from fastapi import FastAPI, HTTPException
from fastapi.testclient import TestClient
from sqlalchemy import create_engine, event
from sqlalchemy import Column, Integer, MetaData, Table, inspect
from alembic.migration import MigrationContext
from alembic.operations import Operations
from sqlalchemy.orm import sessionmaker
from sqlalchemy.pool import StaticPool
from starlette.requests import Request

import app.main  # noqa: F401 - register complete metadata before create_all
from app.core.config import settings
from app.core.database import Base
from app.core.database import get_db
from app.dependencies import get_current_user
from app.api.v1 import budget_pricing as budget_pricing_api
from app.api.v1 import users as users_api
from app.models.account import Account, AccountBudgetProject, AccountMembership
from app.models.budget_project import (
    BudgetProjectImportBatch,
    BudgetProjectImportRevision,
    BudgetProjectProfile,
)
from app.models.budget_pricing import (
    BudgetProjectPricingRun,
    BudgetProjectPricingRunDraftSnapshot,
)
from app.models.budget_pricing_draft import (
    PRICING_MODE_ACCOUNT_STRICT,
    PRICING_MODE_ENTERPRISE_AI,
    BudgetProjectPricingDraft,
    BudgetProjectPricingDraftEvent,
    BudgetProjectPricingDraftLine,
    BudgetProjectPricingDraftQuoteJob,
    BudgetProjectPricingDraftQuoteJobLine,
)
from app.models.budget_project_quota import (
    BudgetProjectQuotaEvent,
    BudgetProjectQuotaResource,
    BudgetProjectQuotaSnapshot,
)
from app.schemas.budget_pricing import (
    BudgetPricingDraftQuoteJobCreate,
    BudgetProjectQuotaEnterpriseSyncIn,
)
from app.models.enterprise_quota import (
    EnterpriseCostResource,
    EnterpriseQuotaComponent,
    EnterpriseQuotaItem,
    EnterpriseQuotaVersion,
)
from app.models.project_progress import Project
from app.models.user import User, UserRole
from app.services.account_tenancy import AccountTenancyError, resolve_current_account
from app.services.budget_pricing import (
    BudgetPricingError,
    activate_budget_pricing_run,
    archive_budget_pricing_run,
    create_budget_pricing_run,
)
from app.services.budget_pricing_drafts import (
    capture_budget_pricing_run_draft_snapshot,
    create_or_rebuild_budget_pricing_draft,
    get_current_budget_pricing_draft,
    patch_budget_pricing_draft_line,
    patch_budget_pricing_draft_line_construction_note,
    refresh_budget_pricing_draft_summary,
    restore_budget_pricing_draft_from_run_snapshot,
    serialize_budget_pricing_draft_line,
    update_budget_pricing_draft_totals_config,
)
from app.services.budget_pricing_resource_details import (
    build_budget_procurement_statistics,
    build_budget_pricing_resource_details,
)
from app.services.budget_pricing_statistics_export import (
    render_budget_pricing_statistics_export,
)
from app.services.budget_project_quotas import (
    create_project_quota_resource,
    delete_project_quota_resource,
    materialize_project_quota,
    sync_project_quota_to_enterprise,
    update_project_quota_resource,
)
from app.services.budget_pricing_ai_estimates import (
    build_budget_pricing_ai_estimate_input,
    estimate_budget_pricing_draft_line,
)
from app.services import budget_pricing_draft_quote_jobs as quote_job_service
from app.services.budget_pricing_draft_quote_jobs import (
    create_budget_pricing_draft_quote_job,
    run_budget_pricing_draft_quote_job,
)
from app.services.budget_projects import accessible_budget_profile_query, get_budget_profile


MIGRATION_PATH = (
    Path(__file__).resolve().parents[1]
    / "alembic"
    / "versions"
    / "20260716_0052_add_account_pricing_drafts.py"
)
MIGRATION_0058_PATH = (
    Path(__file__).resolve().parents[1]
    / "alembic"
    / "versions"
    / "20260717_0058_add_budget_pricing_draft_quote_jobs.py"
)
MIGRATION_0059_PATH = (
    Path(__file__).resolve().parents[1]
    / "alembic"
    / "versions"
    / "20260720_0059_allow_pricing_drafts_per_mode.py"
)
MIGRATION_0060_PATH = (
    Path(__file__).resolve().parents[1]
    / "alembic"
    / "versions"
    / "20260720_0060_add_pricing_draft_breakdown_json.py"
)
MIGRATION_0076_PATH = (
    Path(__file__).resolve().parents[1]
    / "alembic"
    / "versions"
    / "20260731_0076_add_project_quota_resource_workbench.py"
)
MIGRATION_0081_PATH = (
    Path(__file__).resolve().parents[1]
    / "alembic"
    / "versions"
    / "20260801_0081_add_pricing_run_draft_snapshots.py"
)


@pytest.fixture()
def db():
    engine = create_engine(
        "sqlite://",
        connect_args={"check_same_thread": False},
        poolclass=StaticPool,
    )
    Base.metadata.create_all(engine)
    session = sessionmaker(bind=engine)()
    try:
        yield session, engine
    finally:
        session.close()
        Base.metadata.drop_all(engine)
        engine.dispose()


def _formal_rows() -> list[dict]:
    rows = []
    for order, (name, unit, quantity) in enumerate(
        (("石材地面", "㎡", "2"), ("账户专属陌生项目", "项", "3")),
        start=1,
    ):
        rows.append(
            {
                "row_key": f"bill:{order}",
                "source_sheet": "清单",
                "sheet_role": "bill",
                "raw_row_index": order + 1,
                "sort_order": order,
                "mapping_revision": 0,
                "row_type": "item",
                "is_standard_item": True,
                "quantity_status": "valid",
                "standard_row": {
                    "item_name": name,
                    "spec": None,
                    "unit": unit,
                    "calculation_quantity": quantity,
                    "quantity_status": "valid",
                    "warnings": [],
                },
            }
        )
    return rows


def test_ai_estimate_input_includes_reference_context_and_effective_quantity():
    draft = SimpleNamespace(
        id=1,
        draft_uuid="draft-1",
        revision=2,
        pricing_mode=PRICING_MODE_ENTERPRISE_AI,
        source_snapshot_json=json.dumps(
            {
                "ai_reference_context": {
                    "version": "budget-reference-context-v1",
                    "metadata": [{"source_sheet": "编制说明", "source_row": 1, "text": "材料品牌按甲方要求"}],
                    "calculation_rules": [{"source_sheet": "计算规则", "source_row": 2, "text": "门窗洞口按规则扣除"}],
                }
            }
        ),
    )
    line = SimpleNamespace(
        id=2,
        line_uuid="line-2",
        line_revision=3,
        source_sheet="装饰部分清单（户内）",
        source_raw_row_index=10,
        source_row_key="室内:10",
        item_name="墙面抹灰",
        spec="20mm厚",
        unit="㎡",
        calculation_quantity=Decimal("12"),
        quantity_status="valid",
        match_status="unmatched",
        pricing_status="pending_match",
        source_row_snapshot_json=json.dumps(
            {
                "standard_row": {
                    "budget_summary_multiplier": "88",
                    "budget_summary_multiplier_sources": [{"summary_scope": [{"label": "栋号", "value": "5#"}]}],
                }
            }
        ),
        match_evidence_json="{}",
        selected_source_snapshot_json=None,
    )

    payload = build_budget_pricing_ai_estimate_input(draft, line)

    assert payload["summary_context"]["summary_multiplier"] == "88.000000"
    assert payload["summary_context"]["effective_calculation_quantity"] == "1056.000000"
    assert payload["reference_context"]["metadata"][0]["text"] == "材料品牌按甲方要求"
    assert payload["reference_context"]["calculation_rules"][0]["text"] == "门窗洞口按规则扣除"


def _seed_account_project(db, *, suffix: str):
    user = User(
        username=f"admin-{suffix}",
        hashed_password="x",
        role="admin",
        role_version=1,
        quota=10,
        is_active=True,
    )
    db.add(user)
    db.flush()
    account = Account(
        account_uuid=str(uuid4()),
        account_code=f"account-{suffix}",
        account_name=f"账号 {suffix}",
        status="active",
        created_by=user.id,
    )
    db.add(account)
    db.flush()
    db.add(
        AccountMembership(
            account_id=account.id,
            user_id=user.id,
            member_role="owner",
            status="active",
            is_default=True,
            created_by=user.id,
        )
    )
    project = Project(
        project_code=f"P-{suffix}",
        name=f"项目 {suffix}",
        status="planning",
        risk_level="normal",
        progress_percent=0,
        project_manager_id=user.id,
        created_by=user.id,
    )
    db.add(project)
    db.flush()
    profile = BudgetProjectProfile(
        project_id=project.id,
        workspace_status="active",
        created_by=user.id,
        updated_by=user.id,
    )
    db.add(profile)
    db.add(
        AccountBudgetProject(
            account_id=account.id,
            project_id=project.id,
            created_by=user.id,
        )
    )
    db.flush()

    rows = _formal_rows()
    batch = BudgetProjectImportBatch(
        batch_uuid=str(uuid4()),
        project_id=project.id,
        source_filename=f"{suffix}.xlsx",
        source_file_sha256=(suffix[0] * 64),
        source_file_size=1,
        source_storage_mode="metadata_only",
        parser_version="test",
        status="active",
        remap_revision=0,
        sheet_count=1,
        total_output_row_count=2,
        standard_item_count=2,
        valid_quantity_count=2,
        invalid_quantity_count=0,
        original_preview_json="{}",
        current_preview_json="{}",
        issues_json="[]",
        created_by=user.id,
        updated_by=user.id,
    )
    db.add(batch)
    db.flush()
    revision = BudgetProjectImportRevision(
        revision_uuid=str(uuid4()),
        batch_id=batch.id,
        revision_number=1,
        revision_kind="initial",
        snapshot_sha256=(suffix[-1] * 64),
        preview_json="{}",
        sheet_mappings_json="[]",
        standard_rows_json=json.dumps(rows, ensure_ascii=False),
        summary_json=json.dumps({"standard_item_count": 2}),
        created_by=user.id,
    )
    db.add(revision)
    db.flush()
    batch.current_revision_id = revision.id
    batch.confirmed_revision_id = revision.id
    profile.active_import_batch_id = batch.id
    profile.active_import_revision_id = revision.id
    db.flush()
    return user, account, project, profile, batch, revision


def _seed_enterprise_quota(db, user: User):
    version = EnterpriseQuotaVersion(
        version_code="quota-active-v1",
        version_name="企业定额 active",
        status="active",
        is_active=True,
        source_file_sha256="q" * 64,
        created_by=user.id,
    )
    db.add(version)
    db.flush()
    item = EnterpriseQuotaItem(
        version_id=version.id,
        quota_code="Q-001",
        item_name="石材地面",
        unit="㎡",
        unit_price=10,
        labor_fee=2,
        main_material_fee=6,
        auxiliary_material_fee=1,
        machinery_fee=1,
        sort_order=1,
    )
    db.add(item)
    db.flush()
    return version, item


def test_account_first_access_and_missing_membership_fail_closed(db):
    session, _ = db
    user_a, account_a, project_a, _, _, _ = _seed_account_project(session, suffix="a1")
    _, account_b, project_b, _, _, _ = _seed_account_project(session, suffix="b2")
    outsider = User(
        username="no-account",
        hashed_password="x",
        role="admin",
        role_version=1,
        is_active=True,
    )
    session.add(outsider)
    session.commit()

    old = settings.feature_budget_pricing_drafts
    object.__setattr__(settings, "feature_budget_pricing_drafts", True)
    try:
        visible = accessible_budget_profile_query(session, user_a).all()
        assert [row.project_id for row in visible] == [project_a.id]
        assert all(row.project_id != project_b.id for row in visible)
        assert resolve_current_account(session, user_a).id == account_a.id
        assert account_a.id != account_b.id
        with pytest.raises(HTTPException) as cross_account:
            get_budget_profile(session, project_b.id, user_a)
        assert cross_account.value.status_code == 404
        with pytest.raises(HTTPException) as missing:
            accessible_budget_profile_query(session, outsider).all()
        assert missing.value.status_code == 403
        assert missing.value.detail["code"] == "ACCOUNT_MEMBERSHIP_REQUIRED"
        with pytest.raises(AccountTenancyError):
            resolve_current_account(session, outsider)
    finally:
        object.__setattr__(settings, "feature_budget_pricing_drafts", old)


def test_admin_user_creation_commits_user_roles_and_membership_atomically(db, monkeypatch):
    session, engine = db
    operator, account, _, _, _, _ = _seed_account_project(session, suffix="atomic-ok")
    session.commit()
    monkeypatch.setattr(users_api, "get_password_hash", lambda value: f"hashed:{value}")
    request = Request(
        {
            "type": "http",
            "method": "POST",
            "path": "/api/v1/admin/users",
            "headers": [],
            "query_string": b"",
            "client": ("testclient", 50000),
            "server": ("testserver", 80),
            "scheme": "http",
        }
    )
    old = settings.feature_budget_pricing_drafts
    object.__setattr__(settings, "feature_budget_pricing_drafts", True)
    try:
        result = asyncio.run(
            users_api.create_user_by_admin(
                users_api.AdminUserCreate(
                    username="atomic-created",
                    password="secret1",
                    quota=5,
                    roles=["staff", "quote_user"],
                    note="原子创建",
                ),
                request,
                session,
                operator,
            )
        )
        assert result["data"]["username"] == "atomic-created"

        audit = sessionmaker(bind=engine)()
        try:
            created = audit.query(User).filter(User.username == "atomic-created").one()
            assert {item.role for item in created.role_assignments} == {"staff", "quote_user"}
            assert resolve_current_account(audit, created).id == account.id
        finally:
            audit.close()

        no_account_operator = User(
            username="operator-without-account",
            hashed_password="x",
            role="admin",
            role_version=1,
            is_active=True,
        )
        session.add(no_account_operator)
        session.commit()
        with pytest.raises(HTTPException) as failed:
            asyncio.run(
                users_api.create_user_by_admin(
                    users_api.AdminUserCreate(
                        username="must-rollback",
                        password="secret2",
                        quota=5,
                        roles=["staff"],
                        note="应回滚",
                    ),
                    request,
                    session,
                    no_account_operator,
                )
            )
        assert failed.value.status_code == 403
        audit = sessionmaker(bind=engine)()
        try:
            assert audit.query(User).filter(User.username == "must-rollback").first() is None
        finally:
            audit.close()
    finally:
        object.__setattr__(settings, "feature_budget_pricing_drafts", old)


def test_enterprise_ai_matches_and_keeps_unmatched_price_null(db):
    session, engine = db
    user, _, _, profile, batch, revision = _seed_account_project(session, suffix="c3")
    quota, _ = _seed_enterprise_quota(session, user)
    session.commit()
    run_count_before = session.query(BudgetProjectPricingRun).count()

    enterprise_draft = create_or_rebuild_budget_pricing_draft(
        session,
        profile,
        user,
        pricing_mode=PRICING_MODE_ENTERPRISE_AI,
        source_import_batch_id=batch.id,
        source_import_revision_id=revision.id,
        expected_active_quota_version_id=quota.id,
    )
    session.flush()
    enterprise_id = enterprise_draft.id
    enterprise_revision = enterprise_draft.revision
    enterprise_lines = (
        session.query(BudgetProjectPricingDraftLine)
        .filter(BudgetProjectPricingDraftLine.draft_id == enterprise_id)
        .order_by(BudgetProjectPricingDraftLine.source_sort_order)
        .all()
    )
    assert len(enterprise_lines) == 2
    assert enterprise_lines[0].effective_unit_price == Decimal("10.000000")
    assert enterprise_lines[0].line_total == Decimal("20.000000")
    assert enterprise_lines[1].base_unit_price is None
    assert enterprise_lines[1].effective_unit_price is None
    enterprise_summary = json.loads(enterprise_draft.summary_json)
    assert enterprise_summary["account_quota_matched_count"] == 0
    assert enterprise_summary["enterprise_quota_matched_count"] == 1
    assert enterprise_summary["attention_count"] == 1
    assert enterprise_summary["llm_auto_estimation_connected"] is False
    assert session.query(BudgetProjectPricingRun).count() == run_count_before


def test_pricing_run_activation_restores_full_draft_snapshot_and_archive_updates_pointer(db):
    session, _ = db
    user, _, _, profile, batch, revision = _seed_account_project(
        session,
        suffix="version-state",
    )
    quota, _ = _seed_enterprise_quota(session, user)
    draft = create_or_rebuild_budget_pricing_draft(
        session,
        profile,
        user,
        pricing_mode=PRICING_MODE_ENTERPRISE_AI,
        source_import_batch_id=batch.id,
        source_import_revision_id=revision.id,
        expected_active_quota_version_id=quota.id,
    )
    session.flush()
    draft_lines_before_snapshot = (
        session.query(BudgetProjectPricingDraftLine)
        .filter(BudgetProjectPricingDraftLine.draft_id == draft.id)
        .order_by(BudgetProjectPricingDraftLine.source_sort_order)
        .all()
    )
    first_draft_line, second_draft_line = draft_lines_before_snapshot
    first_draft_line.manual_unit_price = Decimal("99")
    first_draft_line.effective_unit_price = Decimal("99")
    first_draft_line.line_total = Decimal("198")
    first_draft_line.price_source = "manual"
    first_draft_line.pricing_breakdown_json = json.dumps(
        {"composite_unit_price": "99.000000", "remark": "保留施工备注"},
        ensure_ascii=False,
    )
    second_draft_line.ai_estimated_unit_price = Decimal("88")
    second_draft_line.ai_estimate_snapshot_json = json.dumps(
        {"provider": "deepseek", "unit_price": "88.000000"},
        ensure_ascii=False,
    )
    second_draft_line.effective_unit_price = Decimal("88")
    second_draft_line.line_total = Decimal("264")
    second_draft_line.amount_included = True
    second_draft_line.pricing_status = "priced"
    second_draft_line.price_source = "ai_estimate"
    original_revision = draft.revision

    run = create_budget_pricing_run(
        session,
        profile,
        user,
        source_import_batch_id=batch.id,
        source_import_revision_id=revision.id,
        expected_quota_version_id=quota.id,
    )
    session.flush()
    snapshot = capture_budget_pricing_run_draft_snapshot(
        session,
        profile,
        user,
        run,
    )
    activate_budget_pricing_run(session, profile, user, run.id)
    activated_draft = restore_budget_pricing_draft_from_run_snapshot(
        session,
        profile,
        user,
        run,
    )
    session.flush()

    session.refresh(profile)
    session.refresh(run)
    assert profile.active_pricing_run_id == run.id
    assert run.status == "confirmed"
    assert snapshot.row_count == 2
    assert session.query(BudgetProjectPricingRunDraftSnapshot).count() == 1
    assert activated_draft.revision == original_revision + 1
    activated_lines = (
        session.query(BudgetProjectPricingDraftLine)
        .filter(BudgetProjectPricingDraftLine.draft_id == activated_draft.id)
        .order_by(BudgetProjectPricingDraftLine.source_sort_order)
        .all()
    )
    assert activated_lines[0].effective_unit_price == Decimal("99.000000")
    assert activated_lines[0].manual_unit_price == Decimal("99.000000")
    assert json.loads(activated_lines[0].pricing_breakdown_json)["remark"] == "保留施工备注"
    assert activated_lines[1].ai_estimated_unit_price == Decimal("88.000000")
    assert json.loads(activated_lines[1].ai_estimate_snapshot_json)["provider"] == "deepseek"
    assert activated_lines[0].match_evidence_json is not None
    assert any(
        event.event_type == "pricing_run_snapshot_activated"
        for event in activated_draft.events
    )

    archive_budget_pricing_run(session, profile, user, run.id)
    session.flush()
    session.refresh(profile)
    session.refresh(run)
    assert profile.active_pricing_run_id is None
    assert run.status == "superseded"

    activate_budget_pricing_run(session, profile, user, run.id)
    restored_again = restore_budget_pricing_draft_from_run_snapshot(
        session,
        profile,
        user,
        run,
    )
    session.flush()
    session.refresh(profile)
    session.refresh(run)
    assert profile.active_pricing_run_id == run.id
    assert run.status == "confirmed"
    assert restored_again.lines[0].manual_unit_price == Decimal("99.000000")


def test_pricing_run_without_full_snapshot_cannot_replace_current_draft(db):
    session, _ = db
    user, _, _, profile, batch, revision = _seed_account_project(
        session,
        suffix="version-no-snapshot",
    )
    quota, _ = _seed_enterprise_quota(session, user)
    draft = create_or_rebuild_budget_pricing_draft(
        session,
        profile,
        user,
        pricing_mode=PRICING_MODE_ENTERPRISE_AI,
        source_import_batch_id=batch.id,
        source_import_revision_id=revision.id,
        expected_active_quota_version_id=quota.id,
    )
    first_line = draft.lines[0]
    first_line.manual_unit_price = Decimal("77")
    first_line.effective_unit_price = Decimal("77")
    run = create_budget_pricing_run(
        session,
        profile,
        user,
        source_import_batch_id=batch.id,
        source_import_revision_id=revision.id,
        expected_quota_version_id=quota.id,
    )
    session.flush()

    with pytest.raises(
        BudgetPricingError,
        match="BUDGET_PRICING_RUN_DRAFT_SNAPSHOT_REQUIRED",
    ):
        restore_budget_pricing_draft_from_run_snapshot(
            session,
            profile,
            user,
            run,
        )

    remaining_lines = (
        session.query(BudgetProjectPricingDraftLine)
        .filter(BudgetProjectPricingDraftLine.draft_id == draft.id)
        .order_by(BudgetProjectPricingDraftLine.source_sort_order)
        .all()
    )
    assert len(remaining_lines) == 2
    assert remaining_lines[0].manual_unit_price == Decimal("77.000000")


def test_pricing_draft_uses_budget_summary_multiplier_for_totals_and_manual_prices(db):
    session, _ = db
    user, _, _, profile, batch, revision = _seed_account_project(session, suffix="summary-mult")
    quota, _ = _seed_enterprise_quota(session, user)
    rows = json.loads(revision.standard_rows_json)
    rows[0]["standard_row"]["budget_summary_multiplier"] = "44"
    revision.standard_rows_json = json.dumps(rows, ensure_ascii=False)
    batch.valid_quantity_count = 2
    session.commit()

    draft = create_or_rebuild_budget_pricing_draft(
        session,
        profile,
        user,
        pricing_mode=PRICING_MODE_ENTERPRISE_AI,
        source_import_batch_id=batch.id,
        source_import_revision_id=revision.id,
        expected_active_quota_version_id=quota.id,
    )
    session.flush()
    line = (
        session.query(BudgetProjectPricingDraftLine)
        .filter(BudgetProjectPricingDraftLine.draft_id == draft.id)
        .order_by(BudgetProjectPricingDraftLine.source_sort_order)
        .first()
    )
    assert line.calculation_quantity == Decimal("2.000000")
    assert line.line_total == Decimal("880.000000")

    patch_budget_pricing_draft_line(
        session,
        profile,
        user,
        pricing_mode=PRICING_MODE_ENTERPRISE_AI,
        line_identifier=line.id,
        expected_revision=draft.revision,
        expected_line_revision=line.line_revision,
        manual_unit_price=Decimal("12"),
    )
    session.flush()
    assert line.line_total == Decimal("1056.000000")
    serialized = serialize_budget_pricing_draft_line(line)
    assert serialized["summary_multiplier"] == "44.000000"
    assert serialized["effective_calculation_quantity"] == "88.000000"


def test_pricing_draft_resource_details_use_enterprise_library_fields_and_quote_quantity(db):
    session, _ = db
    user, _, _, profile, batch, revision = _seed_account_project(
        session,
        suffix="resource-details",
    )
    version, quota_item = _seed_enterprise_quota(session, user)
    quota_item.work_content = "基层处理、石材铺贴及成品保护"
    quota_item.specification = "20mm厚"
    quota_item.brand = "主库指定品牌"
    resources_and_components = (
        (
            EnterpriseCostResource(
                version_id=version.id,
                resource_code="RG-001",
                resource_name="石材铺贴工",
                resource_type="人工",
                library_kind="labor",
                work_content="基层处理、铺贴",
                calculation_rule="按完成面积计算",
                unit="工日",
                default_quantity=0.2,
                price=10,
                sort_order=1,
            ),
            "labor",
            0.2,
            10,
        ),
        (
            EnterpriseCostResource(
                version_id=version.id,
                resource_code="CL-001",
                resource_name="石材",
                resource_type="主材",
                library_kind="material",
                category="石材",
                specification="20mm",
                brand="测试品牌",
                unit="㎡",
                default_quantity=2,
                price=3,
                sort_order=2,
            ),
            "main_material",
            2,
            3,
        ),
        (
            EnterpriseCostResource(
                version_id=version.id,
                resource_code="CL-002",
                resource_name="水泥砂浆",
                resource_type="辅材",
                library_kind="material",
                category="砂浆",
                specification="1:3",
                unit="kg",
                default_quantity=0.5,
                price=2,
                sort_order=3,
            ),
            "auxiliary_material",
            0.5,
            2,
        ),
    )
    for sort_order, (resource, fee_bucket, content, price) in enumerate(
        resources_and_components,
        start=1,
    ):
        session.add(resource)
        session.flush()
        session.add(
            EnterpriseQuotaComponent(
                version_id=version.id,
                quota_item_id=quota_item.id,
                resource_id=resource.id,
                parent_quota_code=quota_item.quota_code,
                component_type=resource.resource_type,
                resource_code=resource.resource_code,
                resource_name=resource.resource_name,
                unit=resource.unit,
                quantity=content,
                unit_price=price,
                amount=content * price,
                fee_bucket=fee_bucket,
                sort_order=sort_order,
            )
        )
    session.commit()

    draft = create_or_rebuild_budget_pricing_draft(
        session,
        profile,
        user,
        pricing_mode=PRICING_MODE_ENTERPRISE_AI,
        source_import_batch_id=batch.id,
        source_import_revision_id=revision.id,
        expected_active_quota_version_id=version.id,
    )
    session.flush()

    labor = build_budget_pricing_resource_details(session, draft, bucket="labor")
    main = build_budget_pricing_resource_details(session, draft, bucket="main_material")
    auxiliary = build_budget_pricing_resource_details(
        session,
        draft,
        bucket="auxiliary_material",
    )

    assert labor["headers"] == [
        "编码",
        "类型",
        "项目名称",
        "工作内容",
        "计算规则",
        "单位",
        "数量",
        "不含税人工单价",
        "人工总价",
    ]
    assert main["headers"] == [
        "分类",
        "材料编码",
        "类型",
        "材料名称",
        "规格",
        "品牌",
        "单位",
        "数量",
        "除税单价",
        "总价",
    ]
    assert labor["rows"][0]["resource_name"] == "石材铺贴工"
    assert labor["rows"][0]["quantity"] == "0.400000"
    assert labor["rows"][0]["amount"] == "4.000000"
    assert labor["total_amount"] == "4.000000"
    assert main["rows"][0]["category"] == "石材"
    assert "project_feature_work_content" not in main["rows"][0]
    assert main["rows"][0]["specification"] == "20mm"
    assert main["rows"][0]["brand"] == "测试品牌"
    assert main["rows"][0]["quantity"] == "4.000000"
    assert main["rows"][0]["amount"] == "12.000000"
    assert main["total_amount"] == "12.000000"
    assert auxiliary["rows"][0]["quantity"] == "1.000000"
    assert auxiliary["total_amount"] == "2.000000"
    assert all(result["enterprise_resource_row_count"] == 1 for result in (labor, main, auxiliary))
    assert all(result["derived_row_count"] == 0 for result in (labor, main, auxiliary))

    procurement = build_budget_procurement_statistics(session, draft)
    assert procurement["material_kind_count"] == 2
    assert procurement["labor_trade_count"] == 1
    assert procurement["unresolved_line_count"] == 0
    assert {row["resource_code"] for row in procurement["material_rows"]} == {
        "CL-001",
        "CL-002",
    }
    assert {
        (row["resource_code"], row["quantity"])
        for row in procurement["material_rows"]
    } == {
        ("CL-001", "4.000000"),
        ("CL-002", "1.000000"),
    }
    assert procurement["labor_rows"][0]["resource_code"] == "RG-001"
    assert procurement["labor_rows"][0]["quantity"] == "0.400000"
    assert procurement["material_unit_totals"] == [
        {"unit": "kg", "quantity": "1.000000"},
        {"unit": "㎡", "quantity": "4.000000"},
    ]
    assert procurement["labor_unit_totals"] == [
        {"unit": "工日", "quantity": "0.400000"},
    ]
    assert all(
        "project_feature_work_content" not in row
        for row in procurement["material_rows"]
    )

    export = render_budget_pricing_statistics_export(
        session,
        draft,
        project_name=profile.project.name,
    )
    workbook = load_workbook(BytesIO(export.content), read_only=True, data_only=True)
    assert workbook.sheetnames == ["统计汇总", "主材明细", "辅材明细", "人工明细"]
    assert export.filename.endswith("_报价统计_R1.xlsx")
    assert export.content_type == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    for sheet_name in ("主材明细", "辅材明细"):
        headers = [cell.value for cell in next(workbook[sheet_name].iter_rows(max_row=1))]
        assert headers == [
            "分类",
            "材料编码",
            "类型",
            "材料名称",
            "规格",
            "品牌",
            "单位",
            "数量",
            "除税单价",
            "总价",
        ]
        assert "项目特征及工作内容" not in headers

    selected_export = render_budget_pricing_statistics_export(
        session,
        draft,
        project_name=profile.project.name,
        sections=("labor", "main_material", "labor"),
    )
    selected_workbook = load_workbook(
        BytesIO(selected_export.content),
        read_only=True,
        data_only=True,
    )
    assert selected_workbook.sheetnames == ["主材明细", "人工明细"]
    assert selected_export.filename.endswith("_报价统计_R1.xlsx")

    single_export = render_budget_pricing_statistics_export(
        session,
        draft,
        project_name=profile.project.name,
        sections=("auxiliary_material",),
    )
    single_workbook = load_workbook(
        BytesIO(single_export.content),
        read_only=True,
        data_only=True,
    )
    assert single_workbook.sheetnames == ["辅材明细"]
    assert single_export.filename.endswith("_辅材明细_R1.xlsx")

    with pytest.raises(ValueError, match="至少选择一项"):
        render_budget_pricing_statistics_export(
            session,
            draft,
            project_name=profile.project.name,
            sections=(),
        )


def test_pricing_draft_http_api_is_account_scoped_and_rejects_account_id(db):
    session, _ = db
    user_a, _, project_a, _, batch_a, revision_a = _seed_account_project(
        session,
        suffix="http-a",
    )
    _, _, project_b, _, _, _ = _seed_account_project(session, suffix="http-b")
    session.commit()

    test_app = FastAPI()
    test_app.include_router(budget_pricing_api.router, prefix="/api/v1")

    def override_user():
        return user_a

    def override_db():
        yield session

    test_app.dependency_overrides[get_current_user] = override_user
    test_app.dependency_overrides[get_db] = override_db
    previous = (
        settings.feature_budget_projects,
        settings.feature_budget_pricing,
        settings.feature_budget_pricing_drafts,
    )
    object.__setattr__(settings, "feature_budget_projects", True)
    object.__setattr__(settings, "feature_budget_pricing", True)
    object.__setattr__(settings, "feature_budget_pricing_drafts", True)
    try:
        with TestClient(test_app) as client:
            current = client.get(
                f"/api/v1/admin/budget-projects/{project_b.id}/pricing-draft/current"
            )
            lines = client.get(
                f"/api/v1/admin/budget-projects/{project_b.id}/pricing-draft/lines"
            )
            resource_details = client.get(
                f"/api/v1/admin/budget-projects/{project_b.id}/pricing-draft/resource-details",
                params={"bucket": "labor"},
            )
            statistics_export = client.get(
                f"/api/v1/admin/budget-projects/{project_b.id}/pricing-draft/statistics-export"
            )
            procurement = client.get(
                f"/api/v1/admin/budget-projects/{project_b.id}/pricing-draft/procurement-statistics"
            )
            patch = client.patch(
                f"/api/v1/admin/budget-projects/{project_b.id}/pricing-draft/lines/1",
                json={
                    "expected_revision": 1,
                    "expected_line_revision": 1,
                    "manual_unit_price": "12.000000",
                },
            )
            note_patch = client.patch(
                f"/api/v1/admin/budget-projects/{project_b.id}/pricing-draft/lines/1/construction-note",
                json={
                    "expected_revision": 1,
                    "expected_line_revision": 1,
                    "remark": "施工提示",
                },
            )
            for response in (
                current,
                lines,
                resource_details,
                statistics_export,
                procurement,
                patch,
                note_patch,
            ):
                assert response.status_code == 404, response.text
                assert response.json()["detail"] == "BUDGET_PROJECT_NOT_FOUND"

            injected = client.post(
                f"/api/v1/admin/budget-projects/{project_a.id}/pricing-draft",
                json={
                    "pricing_mode": "account_strict",
                    "source_import_batch_id": batch_a.id,
                    "source_import_revision_id": revision_a.id,
                    "account_id": 999999,
                },
            )
            assert injected.status_code == 422, injected.text
            assert any(
                item.get("loc", [])[-1:] == ["account_id"]
                and item.get("type") == "extra_forbidden"
                for item in injected.json()["detail"]
            )
    finally:
        (
            budget_projects_enabled,
            budget_pricing_enabled,
            pricing_drafts_enabled,
        ) = previous
        object.__setattr__(settings, "feature_budget_projects", budget_projects_enabled)
        object.__setattr__(settings, "feature_budget_pricing", budget_pricing_enabled)
        object.__setattr__(settings, "feature_budget_pricing_drafts", pricing_drafts_enabled)


def test_modes_keep_separate_drafts_manual_patch_clear_and_optimistic_locks(db):
    session, engine = db
    user, _, _, profile, batch, revision = _seed_account_project(
        session,
        suffix="switch-c3",
    )
    quota, _ = _seed_enterprise_quota(session, user)
    session.commit()
    run_count_before = session.query(BudgetProjectPricingRun).count()
    enterprise_draft = create_or_rebuild_budget_pricing_draft(
        session,
        profile,
        user,
        pricing_mode=PRICING_MODE_ENTERPRISE_AI,
        source_import_batch_id=batch.id,
        source_import_revision_id=revision.id,
        expected_active_quota_version_id=quota.id,
    )
    session.flush()
    enterprise_id = enterprise_draft.id
    enterprise_revision = enterprise_draft.revision

    statements: list[str] = []

    def capture_sql(_conn, _cursor, statement, _parameters, _context, _executemany):
        statements.append(statement.lower())

    event.listen(engine, "before_cursor_execute", capture_sql)
    try:
        account_draft = create_or_rebuild_budget_pricing_draft(
            session,
            profile,
            user,
            pricing_mode=PRICING_MODE_ACCOUNT_STRICT,
            source_import_batch_id=batch.id,
            source_import_revision_id=revision.id,
            expected_revision=None,
        )
        session.flush()
    finally:
        event.remove(engine, "before_cursor_execute", capture_sql)
    assert account_draft.id != enterprise_id
    assert account_draft.revision == 1
    assert not any("enterprise_quota_versions" in statement for statement in statements)
    assert not any("enterprise_quota_items" in statement for statement in statements)

    retained_enterprise_draft = get_current_budget_pricing_draft(
        session,
        profile,
        user,
        pricing_mode=PRICING_MODE_ENTERPRISE_AI,
    )
    retained_account_draft = get_current_budget_pricing_draft(
        session,
        profile,
        user,
        pricing_mode=PRICING_MODE_ACCOUNT_STRICT,
    )
    assert retained_enterprise_draft.id == enterprise_id
    assert retained_enterprise_draft.revision == enterprise_revision
    assert retained_account_draft.id == account_draft.id

    enterprise_lines = (
        session.query(BudgetProjectPricingDraftLine)
        .filter(BudgetProjectPricingDraftLine.draft_id == enterprise_id)
        .order_by(BudgetProjectPricingDraftLine.source_sort_order)
        .all()
    )
    assert len(enterprise_lines) == 2
    assert enterprise_lines[0].base_unit_price == Decimal("10.000000")
    assert enterprise_lines[0].effective_unit_price == Decimal("10.000000")
    assert enterprise_lines[0].line_total == Decimal("20.000000")

    account_lines = (
        session.query(BudgetProjectPricingDraftLine)
        .filter(BudgetProjectPricingDraftLine.draft_id == account_draft.id)
        .order_by(BudgetProjectPricingDraftLine.source_sort_order)
        .all()
    )
    assert len(account_lines) == 2
    assert all(line.base_unit_price is None for line in account_lines)
    assert all(line.effective_unit_price is None for line in account_lines)
    assert all(line.line_total is None for line in account_lines)
    assert session.query(BudgetProjectPricingRun).count() == run_count_before

    line = account_lines[0]
    price_draft_revision = account_draft.revision
    old_line_revision = line.line_revision
    account_draft, line = patch_budget_pricing_draft_line(
        session,
        profile,
        user,
        pricing_mode=PRICING_MODE_ACCOUNT_STRICT,
        line_identifier=line.id,
        expected_revision=price_draft_revision,
        expected_line_revision=old_line_revision,
        manual_unit_price=Decimal("25.123456"),
        reason="人工核价",
    )
    assert account_draft.revision == price_draft_revision + 1
    assert line.line_revision == old_line_revision + 1
    assert line.effective_unit_price == Decimal("25.123456")
    assert line.line_total == Decimal("50.246912")
    assert line.price_source == "manual"

    with pytest.raises(BudgetPricingError) as stale_draft:
        patch_budget_pricing_draft_line(
            session,
            profile,
            user,
            pricing_mode=PRICING_MODE_ACCOUNT_STRICT,
            line_identifier=line.id,
            expected_revision=price_draft_revision,
            expected_line_revision=line.line_revision,
            manual_unit_price=Decimal("30"),
        )
    assert stale_draft.value.code == "BUDGET_PRICING_DRAFT_REVISION_CONFLICT"

    with pytest.raises(BudgetPricingError) as stale_line:
        patch_budget_pricing_draft_line(
            session,
            profile,
            user,
            pricing_mode=PRICING_MODE_ACCOUNT_STRICT,
            line_identifier=line.id,
            expected_revision=account_draft.revision,
            expected_line_revision=old_line_revision,
            manual_unit_price=Decimal("30"),
        )
    assert stale_line.value.code == "BUDGET_PRICING_DRAFT_LINE_REVISION_CONFLICT"

    account_draft, line = patch_budget_pricing_draft_line(
        session,
        profile,
        user,
        pricing_mode=PRICING_MODE_ACCOUNT_STRICT,
        line_identifier=line.id,
        expected_revision=account_draft.revision,
        expected_line_revision=line.line_revision,
        manual_unit_price=Decimal("99"),
        pricing_breakdown={
            "labor_unit_cost": "8.5",
            "main_material_unit_cost": "9.25",
            "auxiliary_material_unit_cost": "1.25",
            "management_unit_cost": "0.5",
            "tax_amount": "0.95",
            "loss_rate": "0.03",
            "material_supply_mode": "乙供",
            "remark": "拆分费用试算",
        },
        reason="拆分费用",
    )
    assert line.manual_unit_price == Decimal("19.500000")
    assert line.effective_unit_price == Decimal("19.500000")
    assert line.line_total == Decimal("39.000000")
    assert line.price_source == "manual_breakdown"
    assert json.loads(line.pricing_breakdown_json)["composite_unit_price"] == "19.500000"
    assert json.loads(line.pricing_breakdown_json)["tax_amount"] == "1.755000"
    summary = json.loads(account_draft.summary_json)
    assert summary["totals"]["labor_total"] == "17.000000"
    assert summary["totals"]["main_material_total"] == "18.500000"
    assert summary["totals"]["auxiliary_material_total"] == "2.500000"
    assert summary["totals"]["tax_total"] == "3.510000"
    assert summary["totals"]["tax_included_total"] == "42.510000"

    price_before_note = line.effective_unit_price
    total_before_note = line.line_total
    price_source_before_note = line.price_source
    manual_price_before_note = line.manual_unit_price
    draft_revision_before_note = account_draft.revision
    line_revision_before_note = line.line_revision
    account_draft, line = patch_budget_pricing_draft_line_construction_note(
        session,
        profile,
        user,
        pricing_mode=PRICING_MODE_ACCOUNT_STRICT,
        line_identifier=line.id,
        expected_revision=account_draft.revision,
        expected_line_revision=line.line_revision,
        remark="先做成品保护；垃圾外运另计。",
        reason="施工提示人工补充",
    )
    assert json.loads(line.pricing_breakdown_json)["remark"] == "先做成品保护；垃圾外运另计。"
    assert line.effective_unit_price == price_before_note
    assert line.line_total == total_before_note
    assert line.price_source == price_source_before_note
    assert line.manual_unit_price == manual_price_before_note
    assert account_draft.revision == draft_revision_before_note + 1
    assert line.line_revision == line_revision_before_note + 1
    note_event = (
        session.query(BudgetProjectPricingDraftEvent)
        .filter(
            BudgetProjectPricingDraftEvent.draft_id == account_draft.id,
            BudgetProjectPricingDraftEvent.event_type == "construction_note_updated",
        )
        .order_by(BudgetProjectPricingDraftEvent.id.desc())
        .first()
    )
    assert note_event is not None
    assert note_event.actor_id == user.id
    assert note_event.to_revision == account_draft.revision
    assert json.loads(note_event.event_json)["line_id"] == line.id
    assert json.loads(note_event.event_json)["previous_remark"] == "拆分费用试算"
    assert json.loads(note_event.event_json)["remark"] == "先做成品保护；垃圾外运另计。"

    account_draft = update_budget_pricing_draft_totals_config(
        session,
        profile,
        user,
        pricing_mode=PRICING_MODE_ACCOUNT_STRICT,
        expected_revision=account_draft.revision,
        config_patch={"measures_rate": "2", "management_rate": "3", "other_fee": "10", "quote_adjustment_percent": "5"},
        reason="调整费率",
    )
    summary = json.loads(account_draft.summary_json)
    assert summary["totals_config"]["quote_adjustment_percent"] == "5.000000"
    assert summary["totals"]["measures_fee"] == "0.760000"
    assert summary["totals"]["management_fee"] == "1.140000"
    assert summary["totals"]["quote_amount"] == "53.445000"

    line.pricing_breakdown_json = None
    session.flush()
    summary = refresh_budget_pricing_draft_summary(session, account_draft)
    serialized_line = serialize_budget_pricing_draft_line(line)
    assert serialized_line["pricing_breakdown"]["source"] == "derived_breakdown"
    assert serialized_line["pricing_breakdown"]["tax_amount"] == "1.755000"
    assert summary["totals"]["tax_excluded_total"] == "39.000000"
    assert Decimal(summary["totals"]["labor_total"]) > 0

    account_draft, line = patch_budget_pricing_draft_line(
        session,
        profile,
        user,
        pricing_mode=PRICING_MODE_ACCOUNT_STRICT,
        line_identifier=line.id,
        expected_revision=account_draft.revision,
        expected_line_revision=line.line_revision,
        manual_unit_price=None,
        reason="清空人工价",
    )
    assert line.manual_unit_price is None
    assert line.effective_unit_price is None
    assert line.line_total is None
    retained_enterprise_draft = get_current_budget_pricing_draft(
        session,
        profile,
        user,
        pricing_mode=PRICING_MODE_ENTERPRISE_AI,
    )
    assert retained_enterprise_draft.id == enterprise_id
    assert retained_enterprise_draft.revision == enterprise_revision
    assert session.query(BudgetProjectPricingRun).count() == run_count_before


def test_p2_2a_draft_source_has_no_legacy_or_model_orchestration_dependencies():
    source = (
        Path(__file__).resolve().parents[1]
        / "app"
        / "services"
        / "budget_pricing_drafts.py"
    ).read_text(encoding="utf-8").lower()
    for forbidden in ("cost_items", "model_gateway", "n8n", "create_budget_pricing_run("):
        assert forbidden not in source


def test_p2_2c1_manual_ai_estimate_enters_draft_and_manual_price_still_wins(db):
    session, _ = db
    old_provider = settings.budget_pricing_ai_provider
    object.__setattr__(settings, "budget_pricing_ai_provider", "rule")
    user, _, _, profile, batch, revision = _seed_account_project(session, suffix="ai-c1")
    try:
        _seed_enterprise_quota(session, user)
        session.commit()
        run_count_before = session.query(BudgetProjectPricingRun).count()

        draft = create_or_rebuild_budget_pricing_draft(
            session,
            profile,
            user,
            pricing_mode=PRICING_MODE_ACCOUNT_STRICT,
            source_import_batch_id=batch.id,
            source_import_revision_id=revision.id,
            expected_revision=None,
            reason="account mode missing price",
        )
        session.flush()
        line = (
            session.query(BudgetProjectPricingDraftLine)
            .filter(BudgetProjectPricingDraftLine.draft_id == draft.id)
            .order_by(BudgetProjectPricingDraftLine.source_sort_order)
            .first()
        )
        assert line.base_unit_price is None
        assert line.effective_unit_price is None

        draft, line = asyncio.run(
            estimate_budget_pricing_draft_line(
                session,
                profile,
                user,
                line_identifier=line.id,
                expected_revision=draft.revision,
                expected_line_revision=line.line_revision,
                reason="manual ai estimate",
            )
        )
        assert line.ai_estimated_unit_price is not None
        assert line.manual_unit_price is None
        assert line.effective_unit_price == line.ai_estimated_unit_price
        assert line.price_source == "ai_estimate"
        assert line.line_total is not None
        pricing_breakdown = json.loads(line.pricing_breakdown_json)
        assert pricing_breakdown["tax_rate"] == "0.090000"
        assert Decimal(pricing_breakdown["tax_amount"]) == (line.ai_estimated_unit_price * Decimal("0.09")).quantize(Decimal("0.000001"))
        assert "labor_unit_cost" in pricing_breakdown
        assert "main_material_unit_cost" in pricing_breakdown
        assert "auxiliary_material_unit_cost" in pricing_breakdown
        assert "main_material_without_loss" in pricing_breakdown
        assert "loss_rate" in pricing_breakdown
        estimate_snapshot = json.loads(line.ai_estimate_snapshot_json)
        assert estimate_snapshot["estimate"]["provider"] == "rule"
        assert estimate_snapshot["estimate"]["pricing_breakdown"]["tax_rate"] == "0.090000"
        assert estimate_snapshot["input"]["pricing_mode"] == PRICING_MODE_ACCOUNT_STRICT
        assert json.loads(draft.summary_json)["ai_estimate_count"] == 1
        assert session.query(BudgetProjectPricingRun).count() == run_count_before

        draft, line = patch_budget_pricing_draft_line(
            session,
            profile,
            user,
            line_identifier=line.id,
            expected_revision=draft.revision,
            expected_line_revision=line.line_revision,
            manual_unit_price=Decimal("33.333333"),
            reason="manual overrides ai",
        )
        assert line.price_source == "manual"
        assert line.effective_unit_price == Decimal("33.333333")
        draft, line = patch_budget_pricing_draft_line(
            session,
            profile,
            user,
            line_identifier=line.id,
            expected_revision=draft.revision,
            expected_line_revision=line.line_revision,
            manual_unit_price=None,
            reason="clear manual restores ai",
        )
        assert line.manual_unit_price is None
        assert line.price_source == "ai_estimate"
        assert line.effective_unit_price == line.ai_estimated_unit_price
    finally:
        object.__setattr__(settings, "budget_pricing_ai_provider", old_provider)


def test_p2_2c1_ai_estimate_does_not_override_enterprise_base_price(db):
    session, _ = db
    user, _, _, profile, batch, revision = _seed_account_project(session, suffix="ai-base")
    quota, _ = _seed_enterprise_quota(session, user)
    session.commit()
    draft = create_or_rebuild_budget_pricing_draft(
        session,
        profile,
        user,
        pricing_mode=PRICING_MODE_ENTERPRISE_AI,
        source_import_batch_id=batch.id,
        source_import_revision_id=revision.id,
        expected_active_quota_version_id=quota.id,
        reason="enterprise base price",
    )
    line = (
        session.query(BudgetProjectPricingDraftLine)
        .filter(BudgetProjectPricingDraftLine.draft_id == draft.id)
        .order_by(BudgetProjectPricingDraftLine.source_sort_order)
        .first()
    )
    assert line.base_unit_price == Decimal("10.000000")
    with pytest.raises(BudgetPricingError) as blocked:
        asyncio.run(
            estimate_budget_pricing_draft_line(
                session,
                profile,
                user,
                line_identifier=line.id,
                expected_revision=draft.revision,
                expected_line_revision=line.line_revision,
            )
        )
    assert blocked.value.code == "BUDGET_PRICING_AI_ESTIMATE_BASE_PRICE_EXISTS"


def test_p2_2c2_quote_job_defaults_are_conservative_for_large_imports():
    payload = BudgetPricingDraftQuoteJobCreate(
        source_import_batch_id=1,
        source_import_revision_id=1,
    )

    assert payload.ai_concurrency == 1
    assert payload.ai_batch_size == 3


def test_p2_2c2_one_click_quote_job_prices_unmatched_lines_in_background(db, monkeypatch):
    session, engine = db
    user, _, _, profile, batch, revision = _seed_account_project(session, suffix="quote-job")
    quota, _ = _seed_enterprise_quota(session, user)
    session.commit()
    run_count_before = session.query(BudgetProjectPricingRun).count()
    monkeypatch.setattr(quote_job_service, "SessionLocal", sessionmaker(bind=engine, autoflush=False))

    batch_calls: list[int] = []

    async def fake_generate_batch(snapshots, *, current_user):
        batch_calls.append(len(snapshots))
        results = {}
        for snapshot in snapshots:
            assert snapshot["pricing_mode"] == PRICING_MODE_ENTERPRISE_AI
            assert snapshot["item_name"]
            results[snapshot["line_uuid"]] = {
                "unit_price": "22.500000",
                "confidence": 0.86,
                "basis": "mocked DeepSeek batch estimate for background job",
                "risks": ["manual review required"],
                "mode": "deepseek",
                "provider": "deepseek",
                "model": "deepseek-test",
                "prompt_version": "test",
                "engine_version": "test",
                "batch_mode": True,
            }
        return results

    monkeypatch.setattr(quote_job_service, "generate_budget_pricing_ai_estimate_batch", fake_generate_batch)

    job = create_budget_pricing_draft_quote_job(
        session,
        profile,
        user,
        BudgetPricingDraftQuoteJobCreate(
            pricing_mode=PRICING_MODE_ENTERPRISE_AI,
            source_import_batch_id=batch.id,
            source_import_revision_id=revision.id,
            expected_active_quota_version_id=quota.id,
            ai_concurrency=1,
            ai_batch_size=12,
            reason="one click quote job test",
        ),
    )
    assert job.total_line_count == 2
    assert job.enterprise_priced_count == 1
    assert job.ai_total_count == 1
    assert job.ai_completed_count == 0
    session.commit()

    asyncio.run(run_budget_pricing_draft_quote_job(job.id))
    session.expire_all()
    finished = session.query(BudgetProjectPricingDraftQuoteJob).filter(BudgetProjectPricingDraftQuoteJob.id == job.id).one()
    assert finished.status == "succeeded"
    assert finished.progress_percent == 100
    assert finished.enterprise_priced_count == 1
    assert finished.ai_total_count == 1
    assert finished.ai_completed_count == 1
    assert finished.ai_failed_count == 0
    assert batch_calls == [1]
    job_lines = (
        session.query(BudgetProjectPricingDraftQuoteJobLine)
        .filter(BudgetProjectPricingDraftQuoteJobLine.job_id == job.id)
        .order_by(BudgetProjectPricingDraftQuoteJobLine.source_sort_order)
        .all()
    )
    assert [line.status for line in job_lines] == ["enterprise_matched", "ai_succeeded"]

    draft_lines = (
        session.query(BudgetProjectPricingDraftLine)
        .filter(BudgetProjectPricingDraftLine.draft_id == finished.draft_id)
        .order_by(BudgetProjectPricingDraftLine.source_sort_order)
        .all()
    )
    assert draft_lines[0].base_unit_price == Decimal("10.000000")
    assert draft_lines[0].price_source == "enterprise_quota"
    assert draft_lines[1].base_unit_price is None
    assert draft_lines[1].ai_estimated_unit_price == Decimal("22.500000")
    assert draft_lines[1].effective_unit_price == Decimal("22.500000")
    assert draft_lines[1].price_source == "ai_estimate"
    pricing_breakdown = json.loads(draft_lines[1].pricing_breakdown_json)
    assert pricing_breakdown["tax_rate"] == "0.090000"
    assert pricing_breakdown["tax_amount"] == "2.025000"
    assert "labor_unit_cost" in pricing_breakdown
    assert "main_material_unit_cost" in pricing_breakdown
    assert "auxiliary_material_unit_cost" in pricing_breakdown
    snapshot = json.loads(draft_lines[1].ai_estimate_snapshot_json)
    assert snapshot["estimate"]["provider"] == "deepseek"
    assert snapshot["estimate"]["pricing_breakdown"]["tax_amount"] == "2.025000"
    assert session.query(BudgetProjectPricingRun).count() == run_count_before


def test_0052_migration_backfills_memberships_and_budget_project_bindings():
    spec = importlib.util.spec_from_file_location("budget_pricing_0052", MIGRATION_PATH)
    assert spec and spec.loader
    migration = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(migration)
    assert migration.revision == "20260716_0052"
    assert migration.down_revision == "20260716_0051"

    metadata = MetaData()
    users = Table("users", metadata, Column("id", Integer, primary_key=True))
    projects = Table("projects", metadata, Column("id", Integer, primary_key=True))
    profiles = Table(
        "budget_project_profiles",
        metadata,
        Column("id", Integer, primary_key=True),
        Column("project_id", Integer, nullable=False),
        Column("created_by", Integer, nullable=False),
    )
    for name in (
        "budget_project_import_batches",
        "budget_project_import_revisions",
        "enterprise_quota_versions",
        "enterprise_quota_items",
    ):
        Table(name, metadata, Column("id", Integer, primary_key=True))

    engine = create_engine("sqlite://")
    with engine.begin() as connection:
        metadata.create_all(connection)
        connection.execute(users.insert(), [{"id": 1}, {"id": 2}])
        connection.execute(projects.insert(), [{"id": 10}, {"id": 11}])
        connection.execute(
            profiles.insert(),
            [
                {"id": 100, "project_id": 10, "created_by": 1},
                {"id": 101, "project_id": 11, "created_by": 2},
            ],
        )
        context = MigrationContext.configure(connection)
        operations = Operations(context)
        original_op = migration.op
        migration.op = operations
        try:
            migration.upgrade()
            database = inspect(connection)
            assert {
                "accounts",
                "account_memberships",
                "account_budget_projects",
                "budget_project_pricing_drafts",
                "budget_project_pricing_draft_lines",
                "budget_project_pricing_draft_events",
            }.issubset(database.get_table_names())
            assert connection.exec_driver_sql("SELECT COUNT(*) FROM accounts").scalar_one() == 1
            assert connection.exec_driver_sql("SELECT COUNT(*) FROM account_memberships").scalar_one() == 2
            assert connection.exec_driver_sql("SELECT COUNT(*) FROM account_budget_projects").scalar_one() == 2
            assert connection.exec_driver_sql(
                "SELECT COUNT(*) FROM account_memberships WHERE status='active' AND is_default=1"
            ).scalar_one() == 2
            assert connection.exec_driver_sql(
                "SELECT COUNT(*) FROM account_budget_projects WHERE project_id IN (10, 11)"
            ).scalar_one() == 2
            migration.downgrade()
            database = inspect(connection)
            assert not {
                "accounts",
                "account_memberships",
                "account_budget_projects",
                "budget_project_pricing_drafts",
                "budget_project_pricing_draft_lines",
                "budget_project_pricing_draft_events",
            }.intersection(database.get_table_names())
        finally:
            migration.op = original_op


def test_alembic_metadata_registers_account_and_pricing_draft_models():
    env_source = (
        Path(__file__).resolve().parents[1] / "alembic" / "env.py"
    ).read_text(encoding="utf-8")
    assert "    account," in env_source
    assert "    budget_pricing_draft," in env_source
    assert {
        "accounts",
        "account_memberships",
        "account_budget_projects",
        "budget_project_pricing_drafts",
        "budget_project_pricing_draft_lines",
        "budget_project_pricing_draft_events",
        "budget_project_pricing_draft_quote_jobs",
        "budget_project_pricing_draft_quote_job_lines",
    }.issubset(Base.metadata.tables)


def test_p2_2c1_migration_declares_ai_estimate_columns():
    migration = (
        Path(__file__).resolve().parents[1]
        / "alembic"
        / "versions"
        / "20260716_0057_add_budget_pricing_ai_estimate_fields.py"
    )
    text = migration.read_text(encoding="utf-8")
    assert 'revision: str = "20260716_0057"' in text
    assert 'down_revision: Union[str, None] = "20260716_0056"' in text
    assert '"ai_estimated_unit_price"' in text
    assert '"ai_estimate_snapshot_json"' in text
    assert "ai_estimated_unit_price" in BudgetProjectPricingDraftLine.__table__.c
    assert "ai_estimate_snapshot_json" in BudgetProjectPricingDraftLine.__table__.c


def test_p2_2c2_migration_declares_background_quote_job_tables():
    text = MIGRATION_0058_PATH.read_text(encoding="utf-8")
    assert 'revision: str = "20260717_0058"' in text
    assert 'down_revision: Union[str, None] = "20260716_0057"' in text
    assert "budget_project_pricing_draft_quote_jobs" in text
    assert "budget_project_pricing_draft_quote_job_lines" in text
    assert "budget_project_pricing_draft_quote_jobs" in Base.metadata.tables
    assert "budget_project_pricing_draft_quote_job_lines" in Base.metadata.tables


def test_pricing_draft_unique_constraint_allows_one_draft_per_mode():
    text = MIGRATION_0059_PATH.read_text(encoding="utf-8")
    assert 'revision: str = "20260720_0059"' in text
    assert 'down_revision: Union[str, None] = "20260717_0058"' in text
    assert '"uq_budget_pricing_drafts_account_project"' in text
    assert '"uq_budget_pricing_drafts_account_project_mode"' in text
    assert '["account_id", "project_id", "pricing_mode"]' in text
    unique_constraints = {
        constraint.name: tuple(constraint.columns.keys())
        for constraint in BudgetProjectPricingDraft.__table__.constraints
        if constraint.__class__.__name__ == "UniqueConstraint"
    }
    assert unique_constraints["uq_budget_pricing_drafts_account_project_mode"] == (
        "account_id",
        "project_id",
        "pricing_mode",
    )


def test_pricing_draft_breakdown_migration_declares_json_column():
    text = MIGRATION_0060_PATH.read_text(encoding="utf-8")
    assert 'revision: str = "20260720_0060"' in text
    assert 'down_revision: Union[str, None] = "20260720_0059"' in text
    assert '"pricing_breakdown_json"' in text
    assert "pricing_breakdown_json" in BudgetProjectPricingDraftLine.__table__.c


def _seed_project_quota_components(session, version, quota_item):
    rows = (
        ("labor", "RG-101", "安装工", "工日", Decimal("1"), Decimal("2")),
        ("main_material", "CL-101", "项目主材", "㎡", Decimal("2"), Decimal("4")),
    )
    components = []
    for sort_order, (bucket, code, name, unit, quantity, unit_price) in enumerate(rows, start=1):
        resource = EnterpriseCostResource(
            version_id=version.id,
            resource_code=code,
            resource_name=name,
            resource_type=bucket,
            library_kind="labor" if bucket == "labor" else "material",
            category="项目测试",
            work_content="项目工料机专项测试",
            calculation_rule="按含量计算",
            unit=unit,
            default_quantity=float(quantity),
            price=float(unit_price),
            computed_price=float(unit_price),
            sort_order=sort_order,
        )
        session.add(resource)
        session.flush()
        component = EnterpriseQuotaComponent(
            version_id=version.id,
            quota_item_id=quota_item.id,
            resource_id=resource.id,
            parent_quota_code=quota_item.quota_code,
            component_type="人工" if bucket == "labor" else "主材",
            resource_code=code,
            resource_name=name,
            unit=unit,
            quantity=float(quantity),
            unit_price=float(unit_price),
            amount=float(quantity * unit_price),
            fee_bucket=bucket,
            formula_library_kind=resource.library_kind,
            formula_link_status="linked",
            sort_order=sort_order,
        )
        session.add(component)
        components.append(component)
    quota_item.labor_fee = 2
    quota_item.main_material_fee = 8
    quota_item.auxiliary_material_fee = 0
    quota_item.machinery_fee = 0
    quota_item.unit_price = 10
    session.flush()
    return components


def test_project_quota_resource_crud_recalculates_project_only(db):
    session, _ = db
    user, _, _, profile, batch, revision = _seed_account_project(
        session,
        suffix="project-quota-crud",
    )
    version, quota_item = _seed_enterprise_quota(session, user)
    enterprise_components = _seed_project_quota_components(session, version, quota_item)
    session.commit()
    draft = create_or_rebuild_budget_pricing_draft(
        session,
        profile,
        user,
        pricing_mode=PRICING_MODE_ENTERPRISE_AI,
        source_import_batch_id=batch.id,
        source_import_revision_id=revision.id,
        expected_active_quota_version_id=version.id,
    )
    session.flush()
    line = draft.lines[0]

    snapshot = materialize_project_quota(
        session,
        profile,
        user,
        pricing_mode=PRICING_MODE_ENTERPRISE_AI,
        line_identifier=line.id,
    )
    session.flush()
    assert snapshot.source_enterprise_quota_item_id == quota_item.id
    assert len(snapshot.resources) == 2
    assert snapshot.unit_price == Decimal("10.000000")

    labor = next(resource for resource in snapshot.resources if resource.fee_bucket == "labor")
    snapshot = update_project_quota_resource(
        session,
        profile,
        user,
        line_identifier=line.id,
        resource_identifier=labor.resource_uuid,
        expected_snapshot_revision=snapshot.revision,
        expected_resource_revision=labor.revision,
        payload={"quantity": "1.5", "unit_price": "3"},
        reason="项目人工含量复核",
    )
    session.flush()
    assert snapshot.labor_fee == Decimal("4.500000")
    assert snapshot.unit_price == Decimal("12.500000")
    session.refresh(line)
    assert line.manual_unit_price == Decimal("12.500000")
    assert line.effective_unit_price == Decimal("12.500000")

    snapshot = create_project_quota_resource(
        session,
        profile,
        user,
        line_identifier=line.id,
        expected_snapshot_revision=snapshot.revision,
        payload={
            "fee_bucket": "machinery",
            "resource_name": "小型切割机",
            "component_type": "机械",
            "unit": "台班",
            "quantity": "1",
            "unit_price": "4",
        },
        reason="补充机械消耗",
    )
    session.flush()
    assert snapshot.machinery_fee == Decimal("4.000000")
    assert snapshot.unit_price == Decimal("16.500000")

    main_material = next(
        resource for resource in snapshot.resources if resource.fee_bucket == "main_material"
    )
    snapshot = delete_project_quota_resource(
        session,
        profile,
        user,
        line_identifier=line.id,
        resource_identifier=main_material.resource_uuid,
        expected_snapshot_revision=snapshot.revision,
        expected_resource_revision=main_material.revision,
        reason="项目改用甲供材料",
    )
    session.flush()
    assert len(snapshot.resources) == 2
    assert snapshot.main_material_fee == Decimal("0.000000")
    assert snapshot.unit_price == Decimal("8.500000")
    assert session.query(BudgetProjectQuotaEvent).filter_by(snapshot_id=snapshot.id).count() == 4

    # Project edits never mutate the active enterprise master.
    session.refresh(quota_item)
    for component in enterprise_components:
        session.refresh(component)
    assert quota_item.unit_price == 10
    assert len(enterprise_components) == 2
    assert [component.amount for component in enterprise_components] == [2.0, 8.0]


def test_project_quota_enterprise_sync_creates_versioned_draft(db):
    session, _ = db
    user, _, _, profile, batch, revision = _seed_account_project(
        session,
        suffix="project-quota-sync",
    )
    version, quota_item = _seed_enterprise_quota(session, user)
    _seed_project_quota_components(session, version, quota_item)
    session.commit()
    draft = create_or_rebuild_budget_pricing_draft(
        session,
        profile,
        user,
        pricing_mode=PRICING_MODE_ENTERPRISE_AI,
        source_import_batch_id=batch.id,
        source_import_revision_id=revision.id,
        expected_active_quota_version_id=version.id,
    )
    session.flush()
    line = draft.lines[0]
    snapshot = materialize_project_quota(
        session,
        profile,
        user,
        pricing_mode=PRICING_MODE_ENTERPRISE_AI,
        line_identifier=line.id,
    )
    machinery = create_project_quota_resource(
        session,
        profile,
        user,
        line_identifier=line.id,
        expected_snapshot_revision=snapshot.revision,
        payload={
            "fee_bucket": "machinery",
            "resource_name": "同步测试机械",
            "unit": "台班",
            "quantity": "1",
            "unit_price": "5",
        },
        reason="项目复核补充机械",
    )

    result = sync_project_quota_to_enterprise(
        session,
        profile,
        user,
        line_identifier=line.id,
        expected_snapshot_revision=machinery.revision,
        reason="成本核定后同步项目工料机",
    )
    session.flush()
    target_version_id = result["enterprise_version"]["id"]
    target_version = session.query(EnterpriseQuotaVersion).filter_by(id=target_version_id).one()
    target_item = (
        session.query(EnterpriseQuotaItem)
        .filter_by(version_id=target_version_id, quota_code=quota_item.quota_code)
        .one()
    )

    assert target_version.status == "draft"
    assert target_version.is_active is False
    assert result["requires_activation"] is True
    assert target_item.unit_price == 15.0
    assert (
        session.query(EnterpriseQuotaComponent)
        .filter_by(version_id=target_version_id, quota_item_id=target_item.id)
        .count()
        == 3
    )
    session.refresh(version)
    session.refresh(quota_item)
    assert version.status == "active"
    assert version.is_active is True
    assert quota_item.unit_price == 10


def test_project_quota_enterprise_sync_requires_cost_approver(db):
    session, _ = db
    user, _, project, _, _, _ = _seed_account_project(
        session,
        suffix="project-quota-permission",
    )
    user.role = "none"
    session.add(
        UserRole(
            user_id=user.id,
            role="cost_editor",
            created_by=user.id,
            note="仅允许编辑项目成本，不允许核定企业定额",
        )
    )
    session.add(
        UserRole(
            user_id=user.id,
            role="project_manager",
            created_by=user.id,
            note="允许访问本人预算项目",
        )
    )
    session.commit()
    session.expire(user, ["role_assignments"])

    old_budget = settings.feature_budget_projects
    old_pricing = settings.feature_budget_pricing
    old_drafts = settings.feature_budget_pricing_drafts
    old_cost_db = settings.feature_cost_db
    object.__setattr__(settings, "feature_budget_projects", True)
    object.__setattr__(settings, "feature_budget_pricing", True)
    object.__setattr__(settings, "feature_budget_pricing_drafts", True)
    object.__setattr__(settings, "feature_cost_db", True)
    try:
        with pytest.raises(HTTPException) as denied:
            asyncio.run(
                budget_pricing_api.sync_project_quota_to_enterprise_endpoint(
                    project.id,
                    "missing-line",
                    BudgetProjectQuotaEnterpriseSyncIn(
                        expected_snapshot_revision=1,
                        sync_to_enterprise=True,
                        reason="尝试同步企业定额",
                    ),
                    user,
                    session,
                )
            )
        assert denied.value.status_code == 403
        assert denied.value.detail["code"] == "ENTERPRISE_QUOTA_SYNC_PERMISSION_DENIED"
    finally:
        object.__setattr__(settings, "feature_budget_projects", old_budget)
        object.__setattr__(settings, "feature_budget_pricing", old_pricing)
        object.__setattr__(settings, "feature_budget_pricing_drafts", old_drafts)
        object.__setattr__(settings, "feature_cost_db", old_cost_db)


def test_project_quota_migration_and_metadata_contract():
    text = MIGRATION_0076_PATH.read_text(encoding="utf-8")
    assert 'revision: str = "20260731_0076"' in text
    assert 'down_revision: Union[str, None] = "20260730_0075"' in text
    assert {
        "budget_project_quota_snapshots",
        "budget_project_quota_resources",
        "budget_project_quota_events",
    }.issubset(Base.metadata.tables)
    unique_constraints = {
        constraint.name: tuple(constraint.columns.keys())
        for constraint in BudgetProjectQuotaSnapshot.__table__.constraints
        if constraint.__class__.__name__ == "UniqueConstraint"
    }
    assert unique_constraints["uq_budget_project_quota_snapshots_draft_line"] == (
        "draft_line_id",
    )
    assert "source_enterprise_component_id" in BudgetProjectQuotaResource.__table__.c


def test_pricing_version_full_draft_snapshot_migration_contract():
    text = MIGRATION_0081_PATH.read_text(encoding="utf-8")

    assert 'revision: str = "20260801_0081"' in text
    assert 'down_revision: Union[str, None] = "20260731_0080"' in text
    assert "budget_project_pricing_run_draft_snapshots" in text
    table = Base.metadata.tables["budget_project_pricing_run_draft_snapshots"]
    assert {
        "run_id",
        "account_id",
        "project_id",
        "source_draft_revision",
        "snapshot_sha256",
        "snapshot_json",
    }.issubset(table.c.keys())
