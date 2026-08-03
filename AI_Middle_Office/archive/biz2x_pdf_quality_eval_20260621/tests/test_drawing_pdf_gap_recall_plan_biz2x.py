from __future__ import annotations

import json
import sys
from pathlib import Path

from openpyxl import load_workbook

from app.services.drawing_pdf_gap_recall_plan import (
    build_gap_recall_plan,
    write_gap_recall_plan_outputs,
)
from scripts import biz2x_pdf_gap_recall_plan


def _source_report(tmp_path: Path) -> Path:
    whole = tmp_path / "whole.png"
    grid = tmp_path / "grid.png"
    whole.write_bytes(b"whole")
    grid.write_bytes(b"grid")
    path = tmp_path / "source.json"
    path.write_text(
        json.dumps(
            {
                "tile_report": {
                    "tile_rows": [
                        {
                            "source_file": "drawing.pdf",
                            "page": 1,
                            "tile_id": "p001_whole",
                            "tile_type": "whole_page_preview",
                            "image_path": str(whole),
                        },
                        {
                            "source_file": "drawing.pdf",
                            "page": 1,
                            "tile_id": "p001_g03_r02_c02",
                            "tile_type": "grid",
                            "image_path": str(grid),
                        },
                    ]
                }
            },
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )
    return path


def _gap_pack(source_report_name: str) -> dict[str, object]:
    return {
        "manifest_rows": [
            {
                "gap_no": 1,
                "gap_priority": "P1_missing_core",
                "gap_type": "missing_candidate",
                "section": "拆除工程",
                "answer_item_name": "拆除不锈钢玻璃门",
                "answer_feature": "含门套、门扇及五金拆除并清运",
                "answer_unit": "套",
                "candidate_item_name": "拆除地面",
                "candidate_unit": "㎡",
                "source_file": "drawing.pdf",
                "page": 1,
                "tile_id": "p001_g03_r02_c02",
                "source_report_file": source_report_name,
                "evidence_id": "PDFEV-000001",
                "suggested_next_action": "补拆除对象证据",
            },
            {
                "gap_no": 2,
                "gap_priority": "P2_missing_mep",
                "gap_type": "missing_candidate",
                "section": "三、给排水部分",
                "answer_item_name": "水表供货及安装",
                "answer_feature": "DN25",
                "answer_unit": "组",
                "candidate_item_name": "管道安装",
                "candidate_unit": "m",
                "source_file": "drawing.pdf",
                "page": 1,
                "tile_id": "p001_g03_r02_c02",
                "source_report_file": source_report_name,
                "evidence_id": "PDFEV-000002",
                "suggested_next_action": "补洁具五金表",
            },
            {
                "gap_no": 2,
                "gap_priority": "P2_missing_mep",
                "gap_type": "missing_candidate",
                "section": "三、给排水部分",
                "answer_item_name": "水表供货及安装",
                "source_file": "drawing.pdf",
                "page": 1,
                "tile_id": "p001_g03_r02_c02",
                "source_report_file": source_report_name,
            },
            {
                "gap_no": 3,
                "gap_priority": "P3_review",
                "gap_type": "weak_match_review",
                "section": "其他",
                "answer_item_name": "人工复核项",
            },
        ]
    }


def test_gap_recall_plan_assigns_passes_and_prefers_whole_page(tmp_path: Path):
    source_report = _source_report(tmp_path)
    plan = build_gap_recall_plan(_gap_pack(source_report.name), source_report_dir=tmp_path)

    assert plan["summary"]["gap_count"] == 2
    assert plan["summary"]["plan_task_count"] == 2
    by_gap = {row["gap_no"]: row for row in plan["plan_rows"]}
    assert by_gap[1]["recommended_pass"] == "door_window_demolition"
    assert by_gap[1]["tile_id"] == "p001_whole"
    assert by_gap[1]["tile_type"] == "whole_page_preview"
    assert by_gap[2]["recommended_pass"] == "fixture_valve_schedule"
    assert by_gap[2]["tile_id"] == "p001_whole"


def test_gap_recall_plan_writes_outputs(tmp_path: Path):
    source_report = _source_report(tmp_path)
    plan = build_gap_recall_plan(_gap_pack(source_report.name), source_report_dir=tmp_path)

    outputs = write_gap_recall_plan_outputs(plan, tmp_path / "out", stem="plan")

    assert Path(outputs["json"]).exists()
    assert Path(outputs["csv"]).exists()
    assert Path(outputs["markdown"]).exists()
    assert Path(outputs["xlsx"]).exists()
    workbook = load_workbook(outputs["xlsx"])
    assert workbook.sheetnames == ["recall_summary", "recall_plan"]
    assert workbook["recall_plan"]["I2"].value in {"door_window_demolition", "fixture_valve_schedule"}


def test_gap_recall_plan_cli_writes_from_gap_pack_json(tmp_path: Path, monkeypatch, capsys):
    source_report = _source_report(tmp_path)
    gap_pack_json = tmp_path / "gap_pack.json"
    gap_pack_json.write_text(json.dumps(_gap_pack(source_report.name), ensure_ascii=False), encoding="utf-8")
    output_dir = tmp_path / "cli_out"
    monkeypatch.setattr(
        sys,
        "argv",
        [
            "biz2x_pdf_gap_recall_plan.py",
            "--gap-pack-json",
            str(gap_pack_json),
            "--source-report-dir",
            str(tmp_path),
            "--output-dir",
            str(output_dir),
            "--priority-prefixes",
            "P1",
            "--timestamp",
            "fixed",
        ],
    )

    assert biz2x_pdf_gap_recall_plan.main() == 0
    payload = json.loads(capsys.readouterr().out)

    assert payload["ok"] is True
    assert payload["summary"]["gap_count"] == 1
    assert Path(payload["outputs"]["xlsx"]).exists()


def test_gap_recall_plan_cli_parses_priority_prefixes():
    assert biz2x_pdf_gap_recall_plan._parse_priority_prefixes("P1,P2;P3") == ["P1", "P2", "P3"]
    assert biz2x_pdf_gap_recall_plan._parse_priority_prefixes("all") == [""]
