"""Automatic parser for historical priced workbooks.

The product contract exposes fixed system fields.  Source columns are detected
from known aliases; users are never asked to map columns manually.
"""

from __future__ import annotations

import hashlib
import json
import re
import unicodedata
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from io import BytesIO
from pathlib import Path
from typing import Any
from uuid import uuid4

from openpyxl import load_workbook


PARSER_VERSION = "pricing-archive-parser-v1.1"
SUPPORTED_EXTENSIONS = {".xlsx", ".xlsm"}
_Q6 = Decimal("0.000001")

_ALIASES: dict[str, tuple[str, ...]] = {
    "item_code": ("项目编码", "清单编码", "定额编号", "定额编码", "编号", "编码"),
    "item_name": ("项目名称", "清单名称", "定额名称", "材料名称", "施工项目", "项目", "名称"),
    "specification": ("项目特征", "规格型号", "规格", "特征描述", "项目描述", "工作内容"),
    "unit": ("计量单位", "单位"),
    "quantity": ("工程量", "数量", "定额工程量"),
    "unit_price": (
        "税前综合单价",
        "不含税综合单价",
        "含税综合单价",
        "综合单价",
        "报价单价",
        "结算单价",
        "合同单价",
        "单价",
    ),
    "total_price": ("合价", "综合合价", "总价", "金额", "报价合计"),
}

_SUMMARY_MARKERS = (
    "合计",
    "小计",
    "总计",
    "本页合计",
    "分部合计",
    "措施项目",
    "规费",
    "税金",
)


class PricingArchiveParseError(RuntimeError):
    def __init__(self, code: str, *, context: dict[str, Any] | None = None):
        super().__init__(code)
        self.code = code
        self.context = context or {}

    @property
    def detail(self) -> dict[str, Any]:
        return {"code": self.code, **self.context}


@dataclass(frozen=True)
class ParsedArchiveLine:
    line_uuid: str
    source_sheet: str
    source_row_index: int
    sort_order: int
    item_code: str | None
    item_name: str
    specification: str | None
    unit: str | None
    quantity: Decimal | None
    unit_price: Decimal
    total_price: Decimal | None
    normalized_code: str | None
    normalized_name: str
    normalized_spec: str | None
    normalized_unit: str | None
    price_derivation: str
    fingerprint: str
    raw_text: str
    raw_row_json: str


@dataclass(frozen=True)
class ParsedArchiveWorkbook:
    lines: tuple[ParsedArchiveLine, ...]
    summary: dict[str, Any]
    issues: tuple[dict[str, Any], ...]


@dataclass(frozen=True)
class ParsedDemandWorkbook:
    lines: tuple[dict[str, Any], ...]
    summary: dict[str, Any]
    issues: tuple[dict[str, Any], ...]


def clean_text(value: Any, *, limit: int | None = None) -> str | None:
    if value is None:
        return None
    text = re.sub(r"\s+", " ", str(value)).strip()
    if not text:
        return None
    return text[:limit] if limit else text


def normalize_text(value: Any) -> str:
    text = unicodedata.normalize("NFKC", clean_text(value) or "").lower()
    text = text.replace("×", "x").replace("*", "x")
    return re.sub(r"[\W_]+", "", text, flags=re.UNICODE)


def normalize_unit(value: Any) -> str | None:
    text = unicodedata.normalize("NFKC", clean_text(value) or "").replace(" ", "").lower()
    if not text:
        return None
    aliases = {
        "㎡": "m2",
        "m²": "m2",
        "m2": "m2",
        "平方米": "m2",
        "平米": "m2",
        "平方": "m2",
        "㎥": "m3",
        "m³": "m3",
        "m3": "m3",
        "立方米": "m3",
        "立米": "m3",
        "延米": "m",
        "米": "m",
        "m": "m",
        "个": "item",
        "项": "item",
        "套": "set",
        "台": "set",
        "樘": "item",
        "块": "piece",
        "片": "piece",
        "只": "piece",
    }
    return aliases.get(text, text)


def parse_decimal(value: Any) -> Decimal | None:
    if value is None or isinstance(value, bool):
        return None
    text = unicodedata.normalize("NFKC", str(value)).strip()
    if not text or text in {"-", "--", "/", "—"}:
        return None
    text = text.replace(",", "").replace("￥", "").replace("¥", "").replace("元", "").strip()
    if text.endswith("%"):
        text = text[:-1].strip()
    try:
        result = Decimal(text)
    except (InvalidOperation, TypeError, ValueError):
        return None
    if not result.is_finite():
        return None
    return result.quantize(_Q6, rounding=ROUND_HALF_UP)


def _header_key(value: Any) -> str:
    return normalize_text(value)


def _match_field(header: Any) -> tuple[str | None, int]:
    key = _header_key(header)
    if not key:
        return None, 0
    best_field: str | None = None
    best_score = 0
    for field, aliases in _ALIASES.items():
        for priority, alias in enumerate(aliases):
            alias_key = normalize_text(alias)
            if key == alias_key:
                score = 1000 - priority
            elif len(alias_key) >= 2 and alias_key in key:
                score = 500 - priority
            else:
                continue
            if score > best_score:
                best_field, best_score = field, score
    return best_field, best_score


def _detect_header(
    rows: list[tuple[Any, ...]],
    *,
    require_price: bool = True,
) -> tuple[int, dict[str, int], dict[str, str]] | None:
    candidates: list[tuple[int, int, dict[str, int], dict[str, str]]] = []
    for row_offset, row in enumerate(rows[:80]):
        mapping: dict[str, int] = {}
        labels: dict[str, str] = {}
        field_scores: dict[str, int] = {}
        for column_index, value in enumerate(row):
            field, score = _match_field(value)
            if field is None or score <= field_scores.get(field, -1):
                continue
            mapping[field] = column_index
            field_scores[field] = score
            labels[field] = clean_text(value, limit=100) or ""
        if "item_name" not in mapping:
            continue
        if require_price and "unit_price" not in mapping and "total_price" not in mapping:
            continue
        coverage = len(mapping)
        price_bonus = 10 if "unit_price" in mapping else 0
        candidates.append((coverage * 100 + price_bonus, row_offset, mapping, labels))
    if not candidates:
        return None
    candidates.sort(key=lambda item: (-item[0], item[1]))
    _, row_offset, mapping, labels = candidates[0]
    return row_offset, mapping, labels


def _cell(row: tuple[Any, ...], mapping: dict[str, int], field: str) -> Any:
    index = mapping.get(field)
    return row[index] if index is not None and index < len(row) else None


def _is_summary_row(item_name: str, row: tuple[Any, ...]) -> bool:
    normalized = item_name.replace(" ", "")
    if normalized in _SUMMARY_MARKERS:
        return True
    populated = sum(value not in (None, "") for value in row)
    return populated <= 3 and any(normalized.startswith(marker) for marker in _SUMMARY_MARKERS)


def _is_non_item_structure_row(
    *,
    item_code: str | None,
    specification: str | None,
    unit: str | None,
    quantity: Decimal | None,
    has_positive_price: bool = False,
) -> bool:
    """Identify headings/serial rows that cannot be priced as bill items."""

    return bool(
        not item_code
        and not specification
        and not unit
        and quantity is None
        and not has_positive_price
    )


def _fingerprint(*values: Any) -> str:
    payload = "|".join(normalize_text(value) for value in values)
    return hashlib.sha256(payload.encode("utf-8")).hexdigest()


def _json_dump(value: Any) -> str:
    return json.dumps(value, ensure_ascii=False, separators=(",", ":"), default=str)


def parse_priced_workbook(content: bytes, filename: str) -> ParsedArchiveWorkbook:
    suffix = Path(filename or "").suffix.lower()
    if suffix not in SUPPORTED_EXTENSIONS:
        raise PricingArchiveParseError(
            "PRICING_ARCHIVE_FILE_TYPE_UNSUPPORTED",
            context={"supported_extensions": sorted(SUPPORTED_EXTENSIONS)},
        )
    try:
        workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:
        raise PricingArchiveParseError("PRICING_ARCHIVE_WORKBOOK_INVALID") from exc

    parsed: list[ParsedArchiveLine] = []
    issues: list[dict[str, Any]] = []
    rejected = 0
    skipped_non_item_rows = 0
    sort_order = 0
    sheet_summaries: list[dict[str, Any]] = []
    try:
        for worksheet in workbook.worksheets:
            rows = list(worksheet.iter_rows(values_only=True))
            detected = _detect_header(rows)
            if detected is None:
                if any(any(value not in (None, "") for value in row) for row in rows):
                    issues.append(
                        {
                            "code": "HEADER_NOT_RECOGNIZED",
                            "sheet": worksheet.title,
                            "message": "未同时识别到项目名称和报价字段，已跳过该工作表。",
                        }
                    )
                continue
            header_offset, mapping, labels = detected
            accepted_on_sheet = 0
            rejected_on_sheet = 0
            skipped_on_sheet = 0
            for source_row_index, row in enumerate(rows[header_offset + 1 :], start=header_offset + 2):
                item_name = clean_text(_cell(row, mapping, "item_name"), limit=500)
                if not item_name:
                    continue
                if _is_summary_row(item_name, row):
                    continue
                item_code = clean_text(_cell(row, mapping, "item_code"), limit=128)
                specification = clean_text(_cell(row, mapping, "specification"), limit=1000)
                unit = clean_text(_cell(row, mapping, "unit"), limit=64)
                quantity = parse_decimal(_cell(row, mapping, "quantity"))
                unit_price = parse_decimal(_cell(row, mapping, "unit_price"))
                total_price = parse_decimal(_cell(row, mapping, "total_price"))
                price_derivation = "source_unit_price"
                if (unit_price is None or unit_price <= 0) and quantity and quantity > 0 and total_price and total_price > 0:
                    unit_price = (total_price / quantity).quantize(_Q6, rounding=ROUND_HALF_UP)
                    price_derivation = "derived_from_total"
                if _is_non_item_structure_row(
                    item_code=item_code,
                    specification=specification,
                    unit=unit,
                    quantity=quantity,
                    has_positive_price=bool(unit_price is not None and unit_price > 0),
                ):
                    skipped_non_item_rows += 1
                    skipped_on_sheet += 1
                    continue
                if unit_price is None or unit_price <= 0:
                    rejected += 1
                    rejected_on_sheet += 1
                    if rejected_on_sheet <= 20:
                        issues.append(
                            {
                                "code": "POSITIVE_UNIT_PRICE_REQUIRED",
                                "sheet": worksheet.title,
                                "row": source_row_index,
                                "item_name": item_name,
                                "message": "该行没有可用的正数单价，未进入检索索引。",
                            }
                        )
                    continue
                raw_values = [value for value in row if value not in (None, "")]
                raw_text = " | ".join(clean_text(value) or "" for value in raw_values)[:12000]
                raw_payload = {
                    str(index + 1): value
                    for index, value in enumerate(row)
                    if value not in (None, "")
                }
                sort_order += 1
                parsed.append(
                    ParsedArchiveLine(
                        line_uuid=str(uuid4()),
                        source_sheet=worksheet.title[:128],
                        source_row_index=source_row_index,
                        sort_order=sort_order,
                        item_code=item_code,
                        item_name=item_name,
                        specification=specification,
                        unit=unit,
                        quantity=quantity,
                        unit_price=unit_price,
                        total_price=total_price,
                        normalized_code=normalize_text(item_code) or None,
                        normalized_name=normalize_text(item_name),
                        normalized_spec=normalize_text(specification) or None,
                        normalized_unit=normalize_unit(unit),
                        price_derivation=price_derivation,
                        fingerprint=_fingerprint(item_code, item_name, specification, unit, unit_price),
                        raw_text=raw_text,
                        raw_row_json=_json_dump(raw_payload),
                    )
                )
                accepted_on_sheet += 1
            sheet_summaries.append(
                {
                    "sheet": worksheet.title,
                    "header_row": header_offset + 1,
                    "detected_columns": labels,
                    "indexed_rows": accepted_on_sheet,
                    "rejected_rows": rejected_on_sheet,
                    "skipped_non_item_rows": skipped_on_sheet,
                }
            )
    finally:
        workbook.close()

    if not parsed:
        raise PricingArchiveParseError(
            "PRICING_ARCHIVE_NO_PRICED_LINES",
            context={"issues": issues[:50]},
        )
    summary = {
        "parser_version": PARSER_VERSION,
        "sheet_count": len(workbook.sheetnames),
        "parsed_sheet_count": len(sheet_summaries),
        "indexed_row_count": len(parsed),
        "rejected_row_count": rejected,
        "skipped_non_item_row_count": skipped_non_item_rows,
        "sheets": sheet_summaries,
        "mapping_mode": "automatic_fixed_system_fields",
    }
    return ParsedArchiveWorkbook(lines=tuple(parsed), summary=summary, issues=tuple(issues[:200]))


def parse_demand_workbook(content: bytes, filename: str) -> ParsedDemandWorkbook:
    """Parse a current demand workbook into the fixed pricing-agent row schema."""

    suffix = Path(filename or "").suffix.lower()
    if suffix not in SUPPORTED_EXTENSIONS:
        raise PricingArchiveParseError(
            "PRICING_AGENT_DEMAND_FILE_TYPE_UNSUPPORTED",
            context={"supported_extensions": sorted(SUPPORTED_EXTENSIONS)},
        )
    try:
        workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:
        raise PricingArchiveParseError("PRICING_AGENT_DEMAND_WORKBOOK_INVALID") from exc

    lines: list[dict[str, Any]] = []
    issues: list[dict[str, Any]] = []
    sheets: list[dict[str, Any]] = []
    skipped_non_item_rows = 0
    try:
        for worksheet in workbook.worksheets:
            rows = list(worksheet.iter_rows(values_only=True))
            detected = _detect_header(rows, require_price=False)
            if detected is None:
                if any(any(value not in (None, "") for value in row) for row in rows):
                    issues.append(
                        {
                            "code": "DEMAND_HEADER_NOT_RECOGNIZED",
                            "sheet": worksheet.title,
                            "message": "未识别到项目名称字段，已跳过该工作表。",
                        }
                    )
                continue
            header_offset, mapping, labels = detected
            sheet_count = 0
            skipped_on_sheet = 0
            for source_row_index, row in enumerate(rows[header_offset + 1 :], start=header_offset + 2):
                item_name = clean_text(_cell(row, mapping, "item_name"), limit=500)
                if not item_name or _is_summary_row(item_name, row):
                    continue
                quantity = parse_decimal(_cell(row, mapping, "quantity"))
                if quantity is not None and quantity <= 0:
                    quantity = None
                item_code = clean_text(_cell(row, mapping, "item_code"), limit=128)
                specification = clean_text(_cell(row, mapping, "specification"), limit=1000)
                unit = clean_text(_cell(row, mapping, "unit"), limit=64)
                if _is_non_item_structure_row(
                    item_code=item_code,
                    specification=specification,
                    unit=unit,
                    quantity=quantity,
                ):
                    skipped_non_item_rows += 1
                    skipped_on_sheet += 1
                    continue
                lines.append(
                    {
                        "row_key": f"{worksheet.title}:{source_row_index}",
                        "source_sheet": worksheet.title[:128],
                        "source_row_index": source_row_index,
                        "item_code": item_code,
                        "item_name": item_name,
                        "specification": specification,
                        "quantity": format(quantity, "f") if quantity is not None else None,
                        "unit": unit,
                    }
                )
                sheet_count += 1
            sheets.append(
                {
                    "sheet": worksheet.title,
                    "header_row": header_offset + 1,
                    "detected_columns": labels,
                    "line_count": sheet_count,
                    "skipped_non_item_rows": skipped_on_sheet,
                }
            )
    finally:
        workbook.close()

    if not lines:
        raise PricingArchiveParseError(
            "PRICING_AGENT_DEMAND_NO_LINES",
            context={"issues": issues[:50]},
        )
    return ParsedDemandWorkbook(
        lines=tuple(lines),
        summary={
            "parser_version": PARSER_VERSION,
            "line_count": len(lines),
            "skipped_non_item_row_count": skipped_non_item_rows,
            "parsed_sheet_count": len(sheets),
            "sheets": sheets,
            "mapping_mode": "automatic_fixed_system_fields",
        },
        issues=tuple(issues[:200]),
    )
