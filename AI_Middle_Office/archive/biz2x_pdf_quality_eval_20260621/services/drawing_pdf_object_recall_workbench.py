from __future__ import annotations

import csv
import json
import re
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any, Iterable, Mapping, Sequence

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


PHASE = "BIZ-2x-pdf-object-recall-workbench"

WORKBENCH_HEADERS = [
    "task_no",
    "gap_priority",
    "object_class",
    "recommended_pass",
    "source_file",
    "page",
    "tile_id",
    "image_link",
    "image_path",
    "image_exists",
    "image_source",
    "target_item_name",
    "target_feature",
    "target_unit",
    "target_object_terms",
    "required_evidence_keywords",
    "review_instruction",
    "current_candidate_item_name",
    "current_candidate_unit",
    "current_candidate_evidence_ids",
    "evidence_item_hint",
    "evidence_spec_or_method",
    "evidence_suggested_unit",
    "evidence_text",
    "ready_for_import",
    "fill_status",
    "fill_hint",
]

GROUP_HEADERS = [
    "group_key",
    "task_count",
    "image_link_count",
    "missing_image_count",
    "importable_row_count",
    "answer_only_count",
]


def build_object_recall_workbench(
    object_recall_pack: Mapping[str, Any],
    *,
    recall_plans: Sequence[Mapping[str, Any]] | None = None,
    image_roots: Sequence[str | Path] | None = None,
    fallback_images: Mapping[str, str | Path] | None = None,
    task_images: Mapping[str, str | Path] | None = None,
) -> dict[str, Any]:
    recall_rows = [dict(row) for row in object_recall_pack.get("recall_rows") or [] if isinstance(row, Mapping)]
    resolver = _ImageResolver(
        recall_plans=recall_plans or [],
        image_roots=image_roots or [],
        fallback_images=fallback_images or {},
        task_images=task_images or {},
    )
    workbench_rows: list[dict[str, Any]] = []
    for row in recall_rows:
        image = resolver.resolve(row)
        trace = _row_trace(row, image)
        item_hint = _first(row, "evidence_item_hint")
        spec = _first(row, "evidence_spec_or_method")
        unit = _first(row, "evidence_suggested_unit")
        text = _first(row, "evidence_text")
        fill_status = _fill_status(item_hint=item_hint, spec=spec, unit=unit, text=text, row=row)
        workbench_rows.append(
            {
                "task_no": row.get("task_no", ""),
                "gap_priority": row.get("gap_priority", ""),
                "object_class": row.get("object_class", ""),
                "recommended_pass": row.get("recommended_pass", ""),
                "source_file": trace["source_file"],
                "page": trace["page"],
                "tile_id": trace["tile_id"],
                "image_link": "open_image" if image.get("image_exists") else "",
                "image_path": image.get("image_path", ""),
                "image_exists": image.get("image_exists", False),
                "image_source": image.get("image_source", ""),
                "target_item_name": row.get("target_item_name", ""),
                "target_feature": row.get("target_feature", ""),
                "target_unit": row.get("target_unit", ""),
                "target_object_terms": row.get("target_object_terms", ""),
                "required_evidence_keywords": row.get("required_evidence_keywords", ""),
                "review_instruction": row.get("review_instruction", ""),
                "current_candidate_item_name": row.get("current_candidate_item_name", ""),
                "current_candidate_unit": row.get("current_candidate_unit", ""),
                "current_candidate_evidence_ids": row.get("current_candidate_evidence_ids", ""),
                "evidence_item_hint": item_hint,
                "evidence_spec_or_method": spec,
                "evidence_suggested_unit": unit,
                "evidence_text": text,
                "ready_for_import": "true" if fill_status == "importable" else "false",
                "fill_status": fill_status,
                "fill_hint": _fill_hint(fill_status),
            }
        )

    class_summary = _aggregate(workbench_rows, group_fields=("object_class",))
    pass_summary = _aggregate(workbench_rows, group_fields=("recommended_pass",))
    source_page_summary = _aggregate(workbench_rows, group_fields=("source_file", "page"))
    summary = {
        "object_recall_task_count": len(workbench_rows),
        "image_link_count": sum(1 for row in workbench_rows if row.get("image_exists")),
        "missing_image_count": sum(1 for row in workbench_rows if not row.get("image_exists")),
        "importable_row_count": sum(1 for row in workbench_rows if row.get("fill_status") == "importable"),
        "answer_only_count": sum(1 for row in workbench_rows if row.get("fill_status") == "answer_only_reference"),
        "blank_task_count": sum(1 for row in workbench_rows if row.get("fill_status") == "blank_task"),
        "object_class_counts": dict(Counter(str(row.get("object_class") or "") for row in workbench_rows)),
        "recommended_pass_counts": dict(Counter(str(row.get("recommended_pass") or "") for row in workbench_rows)),
        "image_source_counts": dict(Counter(str(row.get("image_source") or "") for row in workbench_rows)),
        "answer_columns_count_as_evidence": False,
        "quantity_status": "deferred_until_three_fields_accepted",
    }
    return {
        "ok": True,
        "phase": PHASE,
        "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "safe_to_import_without_evidence": False,
        "answer_columns_count_as_evidence": False,
        "summary": summary,
        "class_summary_rows": class_summary,
        "pass_summary_rows": pass_summary,
        "source_page_summary_rows": source_page_summary,
        "workbench_rows": workbench_rows,
    }


def write_object_recall_workbench_outputs(
    report: Mapping[str, Any],
    output_dir: str | Path,
    *,
    stem: str | None = None,
) -> dict[str, str]:
    target = Path(output_dir)
    target.mkdir(parents=True, exist_ok=True)
    file_stem = stem or f"BIZ2x_PDF_object_recall_workbench_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
    json_path = target / f"{file_stem}.json"
    csv_path = target / f"{file_stem}.csv"
    markdown_path = target / f"{file_stem}.md"
    xlsx_path = target / f"{file_stem}.xlsx"
    outputs = {
        "json": str(json_path),
        "csv": str(csv_path),
        "markdown": str(markdown_path),
        "xlsx": str(xlsx_path),
    }
    payload = {**dict(report), "outputs": outputs}
    json_path.write_text(json.dumps(_json_safe(payload), ensure_ascii=False, indent=2), encoding="utf-8")
    _write_csv(csv_path, payload.get("workbench_rows") or [], WORKBENCH_HEADERS)
    markdown_path.write_text(_build_markdown(payload), encoding="utf-8")
    _write_workbook(xlsx_path, payload)
    return outputs


class _ImageResolver:
    def __init__(
        self,
        *,
        recall_plans: Sequence[Mapping[str, Any]],
        image_roots: Sequence[str | Path],
        fallback_images: Mapping[str, str | Path],
        task_images: Mapping[str, str | Path],
    ) -> None:
        self.by_task_no: dict[str, str] = {}
        self.by_target: dict[str, str] = {}
        self.by_evidence_id: dict[str, str] = {}
        self.by_source_page_tile: dict[tuple[str, str, str], str] = {}
        self.image_by_evidence_id: dict[str, str] = {}
        self.fallback_images: dict[str, str] = {}
        self.root_images: list[tuple[str, str]] = []
        for key, value in fallback_images.items():
            path = str(value or "").strip()
            if not path:
                continue
            self.fallback_images[str(key or "").strip()] = path
            self.fallback_images[_compact(key)] = path
        for key, value in task_images.items():
            path = str(value or "").strip()
            if not path:
                continue
            for task_no in _split_values(key):
                self.by_task_no.setdefault(task_no, path)
        for plan in recall_plans:
            self._load_plan(plan)
        for root in image_roots:
            self._load_image_root(Path(root))

    def resolve(self, row: Mapping[str, Any]) -> dict[str, Any]:
        candidates = [
            self._by_task_no(row),
            self._by_target(row),
            self._by_evidence_id(row),
            self._by_source_page_tile(row),
            self._image_by_evidence_id(row),
            self._image_root_source_page_whole(row),
            self._fallback_image(row),
        ]
        for image_source, path in candidates:
            if path:
                image_path = str(path)
                return {
                    "image_path": image_path,
                    "image_exists": Path(image_path).exists(),
                    "image_source": image_source,
                }
        return {"image_path": "", "image_exists": False, "image_source": ""}

    def _by_task_no(self, row: Mapping[str, Any]) -> tuple[str, str]:
        task_no = _first(row, "task_no")
        if task_no in self.by_task_no:
            return f"task_image:{task_no}", self.by_task_no[task_no]
        return "", ""

    def _load_plan(self, plan: Mapping[str, Any]) -> None:
        for row in plan.get("plan_rows") or []:
            if not isinstance(row, Mapping):
                continue
            image_path = _first(row, "image_path")
            if not image_path:
                continue
            target_key = _compact(_first(row, "answer_item_name", "target_item_name"))
            evidence_id = _first(row, "evidence_id", "current_candidate_evidence_ids")
            source_key = _source_page_tile_key(
                _first(row, "source_file"),
                _first(row, "page"),
                _first(row, "tile_id"),
            )
            if target_key:
                self.by_target.setdefault(target_key, image_path)
            if evidence_id:
                for value in _split_values(evidence_id):
                    self.by_evidence_id.setdefault(value, image_path)
            if source_key:
                self.by_source_page_tile.setdefault(source_key, image_path)

    def _load_image_root(self, root: Path) -> None:
        if not root.exists():
            return
        for path in root.rglob("*"):
            if path.suffix.lower() not in {".png", ".jpg", ".jpeg"}:
                continue
            name = path.name
            self.root_images.append((_compact(path.stem), str(path)))
            for evidence_id in _evidence_ids_from_text(name):
                self.image_by_evidence_id.setdefault(evidence_id, str(path))

    def _by_target(self, row: Mapping[str, Any]) -> tuple[str, str]:
        return "recall_plan_target", self.by_target.get(_compact(row.get("target_item_name")), "")

    def _by_evidence_id(self, row: Mapping[str, Any]) -> tuple[str, str]:
        for evidence_id in _split_values(row.get("current_candidate_evidence_ids")):
            if evidence_id in self.by_evidence_id:
                return "recall_plan_evidence_id", self.by_evidence_id[evidence_id]
        return "", ""

    def _by_source_page_tile(self, row: Mapping[str, Any]) -> tuple[str, str]:
        key = _source_page_tile_key(
            _first(row, "candidate_source_files"),
            _first(row, "evidence_pages"),
            _first(row, "evidence_tiles"),
        )
        return "recall_plan_source_page_tile", self.by_source_page_tile.get(key, "")

    def _image_by_evidence_id(self, row: Mapping[str, Any]) -> tuple[str, str]:
        for evidence_id in _split_values(row.get("current_candidate_evidence_ids")):
            if evidence_id in self.image_by_evidence_id:
                return "image_root_evidence_id", self.image_by_evidence_id[evidence_id]
        return "", ""

    def _image_root_source_page_whole(self, row: Mapping[str, Any]) -> tuple[str, str]:
        tile_text = _first(row, "evidence_tiles")
        if "whole" not in tile_text.lower():
            return "", ""
        page = _first(row, "evidence_pages")
        for source_file in _split_values(_first(row, "candidate_source_files")):
            source_key = _compact_file_stem(source_file)
            if not source_key:
                continue
            for image_name, image_path in self.root_images:
                if source_key in image_name and _image_name_matches_page(image_name, page):
                    return "image_root_source_page_whole", image_path
        return "", ""

    def _fallback_image(self, row: Mapping[str, Any]) -> tuple[str, str]:
        for key in (
            _first(row, "recommended_pass"),
            _first(row, "object_class"),
            "default",
        ):
            if not key:
                continue
            path = self.fallback_images.get(key) or self.fallback_images.get(_compact(key))
            if path:
                return f"fallback_image:{key}", path
        return "", ""


def _fill_status(*, item_hint: str, spec: str, unit: str, text: str, row: Mapping[str, Any]) -> str:
    evidence_count = sum(1 for value in (item_hint, spec, unit, text) if value)
    if (item_hint or spec) and text:
        return "importable"
    if evidence_count:
        return "importable_weak"
    if any(_first(row, key) for key in ("target_item_name", "target_feature", "target_unit")):
        return "answer_only_reference"
    return "blank_task"


def _row_trace(row: Mapping[str, Any], image: Mapping[str, Any]) -> dict[str, str]:
    source_file = _first(row, "candidate_source_files")
    page = _first(row, "evidence_pages")
    tile_id = _first(row, "evidence_tiles")
    image_source = _first(image, "image_source")
    image_path = _first(image, "image_path")
    if not source_file and image_path and (
        image_source.startswith("fallback_image:") or image_source.startswith("task_image:")
    ):
        return {
            "source_file": image_path,
            "page": "fallback" if image_source.startswith("fallback_image:") else "task_image",
            "tile_id": image_source,
        }
    return {
        "source_file": source_file,
        "page": page,
        "tile_id": tile_id,
    }


def _fill_hint(fill_status: str) -> str:
    if fill_status == "importable":
        return "证据字段已具备，可回灌验收。"
    if fill_status == "importable_weak":
        return "补齐 evidence_text，并尽量填写对象名称/做法/单位。"
    if fill_status == "answer_only_reference":
        return "当前只有目标答案参考；必须从图纸图片中填写真实 evidence_* 字段。"
    return "补充图纸证据字段。"


def _aggregate(rows: Sequence[Mapping[str, Any]], *, group_fields: Sequence[str]) -> list[dict[str, Any]]:
    grouped: dict[str, list[Mapping[str, Any]]] = defaultdict(list)
    for row in rows:
        key = " | ".join(str(row.get(field) or "") for field in group_fields)
        grouped[key].append(row)
    result: list[dict[str, Any]] = []
    for key, group_rows in sorted(grouped.items(), key=lambda item: item[0]):
        status_counts = Counter(str(row.get("fill_status") or "") for row in group_rows)
        result.append(
            {
                "group_key": key,
                "task_count": len(group_rows),
                "image_link_count": sum(1 for row in group_rows if row.get("image_exists")),
                "missing_image_count": sum(1 for row in group_rows if not row.get("image_exists")),
                "importable_row_count": status_counts.get("importable", 0),
                "answer_only_count": status_counts.get("answer_only_reference", 0),
            }
        )
    return result


def _source_page_tile_key(source_file: Any, page: Any, tile_id: Any) -> tuple[str, str, str]:
    return (_compact_file(source_file), str(page or "").strip(), str(tile_id or "").strip())


def _compact_file(value: Any) -> str:
    return _compact(Path(str(value or "")).name)


def _compact_file_stem(value: Any) -> str:
    return _compact(Path(str(value or "")).stem)


def _compact(value: Any) -> str:
    return "".join(ch.lower() for ch in str(value or "") if ch.isalnum())


def _image_name_matches_page(image_name: str, page: Any) -> bool:
    page_text = str(page or "").strip()
    if not page_text:
        return True
    first_page = _split_values(page_text)[0] if _split_values(page_text) else page_text
    try:
        page_no = int(float(first_page))
    except ValueError:
        return True
    return f"p{page_no:03d}" in image_name or image_name.endswith(str(page_no))


def _split_values(value: Any) -> list[str]:
    if value is None:
        return []
    if isinstance(value, (list, tuple, set)):
        return [str(item).strip() for item in value if str(item).strip()]
    text = str(value or "").strip()
    if not text:
        return []
    return [item.strip() for item in text.replace("；", ";").replace(",", ";").split(";") if item.strip()]


def _evidence_ids_from_text(value: str) -> list[str]:
    return re.findall(r"R\d+-PDFEV-\d+", str(value or ""))


def _first(row: Mapping[str, Any], *keys: str) -> str:
    for key in keys:
        value = row.get(key)
        if value is None:
            continue
        text = str(value).strip()
        if text:
            return text
    return ""


def _build_markdown(report: Mapping[str, Any]) -> str:
    summary = report.get("summary") or {}
    lines = [
        "# BIZ-2x PDF Object Recall Workbench",
        "",
        f"- generated_at: {report.get('generated_at', '-')}",
        f"- tasks: {summary.get('object_recall_task_count', 0)}",
        f"- image_link_count: {summary.get('image_link_count', 0)}",
        f"- importable_row_count: {summary.get('importable_row_count', 0)}",
        f"- answer_columns_count_as_evidence: {summary.get('answer_columns_count_as_evidence', False)}",
        "",
        "| task | class | target | unit | image | status | hint |",
        "| ---: | --- | --- | --- | --- | --- | --- |",
    ]
    for row in (report.get("workbench_rows") or [])[:120]:
        lines.append(
            "| "
            + " | ".join(
                [
                    _md(row.get("task_no")),
                    _md(row.get("object_class")),
                    _md(row.get("target_item_name")),
                    _md(row.get("target_unit")),
                    _md(row.get("image_link")),
                    _md(row.get("fill_status")),
                    _md(row.get("fill_hint")),
                ]
            )
            + " |"
        )
    return "\n".join(lines) + "\n"


def _write_workbook(path: Path, report: Mapping[str, Any]) -> None:
    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "summary"
    _append_rows(summary_sheet, [["metric", "value"]])
    _append_rows(summary_sheet, [[key, _cell_value(value)] for key, value in (report.get("summary") or {}).items()])
    _style_sheet(summary_sheet)

    readme = workbook.create_sheet("readme")
    _append_rows(
        readme,
        [
            ["rule", "description"],
            ["填写范围", "只填写 evidence_item_hint / evidence_spec_or_method / evidence_suggested_unit / evidence_text。"],
            ["证据来源", "必须来自 image_link 指向的图纸图片或同一图纸页的真实文字。"],
            ["禁止事项", "target_* 是人工答案参考，不能当作证据导入。"],
            ["工程量", "当前不验收工程量，quantity 继续锁定。"],
        ],
    )
    _style_sheet(readme)

    workbench = workbook.create_sheet("object_recall_workbench")
    _append_rows(workbench, [WORKBENCH_HEADERS])
    _append_rows(workbench, [[_cell_value(row.get(header)) for header in WORKBENCH_HEADERS] for row in report.get("workbench_rows") or []])
    _style_sheet(workbench)
    workbench.freeze_panes = "A2"
    _add_image_hyperlinks(workbench, report.get("workbench_rows") or [], image_link_column=8)

    class_sheet = workbook.create_sheet("class_summary")
    _append_rows(class_sheet, [GROUP_HEADERS])
    _append_rows(class_sheet, [[_cell_value(row.get(header)) for header in GROUP_HEADERS] for row in report.get("class_summary_rows") or []])
    _style_sheet(class_sheet)

    pass_sheet = workbook.create_sheet("pass_summary")
    _append_rows(pass_sheet, [GROUP_HEADERS])
    _append_rows(pass_sheet, [[_cell_value(row.get(header)) for header in GROUP_HEADERS] for row in report.get("pass_summary_rows") or []])
    _style_sheet(pass_sheet)

    source_sheet = workbook.create_sheet("source_page_summary")
    _append_rows(source_sheet, [GROUP_HEADERS])
    _append_rows(source_sheet, [[_cell_value(row.get(header)) for header in GROUP_HEADERS] for row in report.get("source_page_summary_rows") or []])
    _style_sheet(source_sheet)
    workbook.save(path)


def _add_image_hyperlinks(sheet: Any, rows: Sequence[Mapping[str, Any]], *, image_link_column: int) -> None:
    for idx, row in enumerate(rows, start=2):
        path = str(row.get("image_path") or "").strip()
        if not path:
            continue
        cell = sheet.cell(row=idx, column=image_link_column)
        cell.value = "open_image"
        cell.hyperlink = path
        cell.style = "Hyperlink"


def _write_csv(path: Path, rows: Iterable[Mapping[str, Any]], fieldnames: Sequence[str]) -> None:
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        for row in rows:
            writer.writerow({field: _cell_value(row.get(field)) for field in fieldnames})


def _append_rows(sheet: Any, rows: Iterable[Sequence[Any]]) -> None:
    for row in rows:
        sheet.append(list(row))


def _style_sheet(sheet: Any) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    if sheet.max_row >= 1:
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for col_idx, column_cells in enumerate(sheet.columns, start=1):
        values = [str(cell.value or "") for cell in column_cells[:200]]
        width = min(max([len(value) for value in values] + [10]) + 2, 75)
        sheet.column_dimensions[get_column_letter(col_idx)].width = width


def _cell_value(value: Any) -> Any:
    if isinstance(value, (dict, list, tuple)):
        return json.dumps(value, ensure_ascii=False)
    return value


def _json_safe(value: Any) -> Any:
    if isinstance(value, Mapping):
        return {str(key): _json_safe(item) for key, item in value.items()}
    if isinstance(value, (list, tuple)):
        return [_json_safe(item) for item in value]
    if isinstance(value, Path):
        return str(value)
    return value


def _md(value: Any) -> str:
    return str(value or "").replace("|", "\\|").replace("\n", "<br>")
