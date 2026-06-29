from __future__ import annotations

import json
import sys
from pathlib import Path

from openpyxl import load_workbook

from app.services.drawing_pdf_closed_loop_stage_report import (
    build_closed_loop_stage_report,
    write_closed_loop_stage_report_outputs,
)
from scripts import biz2x_pdf_closed_loop_stage_report


def test_closed_loop_stage_report_keeps_quantity_locked_until_three_field_gate_passes(tmp_path: Path):
    review = _review_report()
    gate = _failed_gate()
    object_pack = _object_pack()
    object_status = _object_status()
    standard_bill = _standard_bill()
    quantity = _quantity_placeholder()

    report = build_closed_loop_stage_report(
        review_report=review,
        gate_report=gate,
        object_recall_report=object_pack,
        template_status_report=object_status,
        standard_bill_report=standard_bill,
        quantity_report=quantity,
    )

    assert report["summary"]["current_stage_no"] == 2
    assert report["summary"]["can_enable_quantity"] is False
    assert report["summary"]["missing_candidate_count"] == 69
    assert report["summary"]["object_recall_task_count"] == 69
    assert report["summary"]["object_importable_row_count"] == 0
    assert report["stage_rows"][1]["status"] == "failed"
    assert report["stage_rows"][2]["status"] == "needs_evidence"
    assert report["stage_rows"][6]["status"] == "blocked"
    assert "三字段" in report["stage_rows"][6]["blocker"]


def test_closed_loop_stage_report_marks_gate_passed_but_quantity_engine_placeholder():
    report = build_closed_loop_stage_report(
        review_report=_passed_review_report(),
        gate_report=_passed_gate(),
        object_recall_report={"summary": {"object_recall_task_count": 0}},
        standard_bill_report={"summary": {"standard_bill_row_count": 1, "mapped_row_count": 1, "unmapped_row_count": 0}},
        quantity_report={"summary": {"quantity_status": "placeholder_only_ready_for_quantity_engine", "quantity_filled_count": 0}},
    )

    assert report["summary"]["can_enable_quantity"] is True
    assert report["stage_rows"][1]["status"] == "complete"
    assert report["stage_rows"][5]["status"] == "complete"
    assert report["stage_rows"][6]["status"] == "blocked"
    assert "工程量引擎尚未启用" in report["stage_rows"][6]["blocker"]


def test_closed_loop_stage_report_uses_object_workbench_summary_over_general_template_status():
    report = build_closed_loop_stage_report(
        review_report=_review_report(),
        gate_report=_failed_gate(),
        object_recall_report=_object_pack(),
        object_workbench_report={
            "summary": {
                "object_recall_task_count": 69,
                "importable_row_count": 0,
                "answer_only_count": 69,
            }
        },
        template_status_report={
            "summary": {
                "source_path": "outputs/source_page_prefill.xlsx",
                "input_row_count": 108,
                "importable_row_count": 91,
                "answer_only_count": 17,
                "ready_for_external_import": True,
            }
        },
        standard_bill_report=_standard_bill(),
        quantity_report=_quantity_placeholder(),
    )

    assert report["summary"]["object_importable_row_count"] == 0
    assert report["summary"]["object_answer_only_count"] == 69
    assert report["stage_rows"][2]["key_metric"] == "tasks=69; importable=0; answer_only=69"
    assert "对象召回任务 69 行" in report["stage_rows"][2]["blocker"]


def test_closed_loop_stage_report_writes_outputs(tmp_path: Path):
    review_xlsx = tmp_path / "review.xlsx"
    review_xlsx.write_text("fake", encoding="utf-8")
    report = build_closed_loop_stage_report(
        review_report=_review_report(),
        gate_report=_failed_gate(),
        object_recall_report=_object_pack(),
        template_status_report=_object_status(),
        standard_bill_report=_standard_bill(),
        quantity_report=_quantity_placeholder(),
        artifacts={"review.xlsx": str(review_xlsx)},
    )

    outputs = write_closed_loop_stage_report_outputs(report, tmp_path / "out", stem="closed_loop")

    assert Path(outputs["json"]).exists()
    assert Path(outputs["stage_csv"]).exists()
    assert Path(outputs["artifact_csv"]).exists()
    assert Path(outputs["markdown"]).exists()
    assert Path(outputs["xlsx"]).exists()
    workbook = load_workbook(outputs["xlsx"])
    assert workbook.sheetnames == ["closed_loop_summary", "stage_status", "artifacts"]
    assert workbook["stage_status"]["B2"].value == "PDF输入与候选列项"


def test_closed_loop_stage_report_cli(tmp_path: Path, monkeypatch, capsys):
    review_json = tmp_path / "review.json"
    gate_json = tmp_path / "gate.json"
    object_json = tmp_path / "object.json"
    status_json = tmp_path / "status.json"
    quantity_json = tmp_path / "quantity.json"
    standard_json = tmp_path / "standard.json"
    review_json.write_text(json.dumps(_review_report(), ensure_ascii=False), encoding="utf-8")
    gate_json.write_text(json.dumps(_failed_gate(), ensure_ascii=False), encoding="utf-8")
    object_json.write_text(json.dumps(_object_pack(), ensure_ascii=False), encoding="utf-8")
    status_json.write_text(json.dumps(_object_status(), ensure_ascii=False), encoding="utf-8")
    quantity_json.write_text(json.dumps(_quantity_placeholder(), ensure_ascii=False), encoding="utf-8")
    standard_json.write_text(json.dumps(_standard_bill(), ensure_ascii=False), encoding="utf-8")
    output_dir = tmp_path / "cli_out"
    monkeypatch.setattr(
        sys,
        "argv",
        [
            "biz2x_pdf_closed_loop_stage_report.py",
            "--review-json",
            str(review_json),
            "--gate-json",
            str(gate_json),
            "--object-recall-json",
            str(object_json),
            "--template-status-json",
            str(status_json),
            "--quantity-json",
            str(quantity_json),
            "--standard-bill-json",
            str(standard_json),
            "--output-dir",
            str(output_dir),
            "--timestamp",
            "fixed",
        ],
    )

    assert biz2x_pdf_closed_loop_stage_report.main() == 0
    payload = json.loads(capsys.readouterr().out)

    assert payload["ok"] is True
    assert payload["summary"]["current_stage_no"] == 2
    assert payload["summary"]["can_enable_quantity"] is False
    assert Path(payload["outputs"]["xlsx"]).exists()


def _review_report() -> dict[str, object]:
    return {
        "summary": {
            "answer_count": 127,
            "candidate_count": 92,
            "matched_three_fields_count": 19,
            "three_field_pass_rate": 0.1496,
            "status_counts": {
                "missing_candidate": 69,
                "matched_three_fields": 19,
                "matched_name_unit_feature_review": 34,
                "unit_conflict": 2,
                "weak_match_review": 3,
            },
        }
    }


def _passed_review_report() -> dict[str, object]:
    return {
        "summary": {
            "answer_count": 1,
            "candidate_count": 1,
            "matched_three_fields_count": 1,
            "three_field_pass_rate": 1,
            "status_counts": {"matched_three_fields": 1},
        }
    }


def _failed_gate() -> dict[str, object]:
    return {
        "can_enable_quantity": False,
        "summary": {
            "failed_gate_count": 4,
            "quantity_status": "deferred_until_three_fields_accepted",
            "status_counts": _review_report()["summary"]["status_counts"],
        },
    }


def _passed_gate() -> dict[str, object]:
    return {
        "can_enable_quantity": True,
        "summary": {
            "failed_gate_count": 0,
            "quantity_status": "ready_after_three_field_acceptance",
            "status_counts": {"matched_three_fields": 1},
        },
    }


def _object_pack() -> dict[str, object]:
    return {"summary": {"object_recall_task_count": 69, "object_class_counts": {"finish_wall": 22}}}


def _object_status() -> dict[str, object]:
    return {
        "summary": {
            "source_path": "outputs/object_recall_pack/object_recall.xlsx",
            "input_row_count": 69,
            "importable_row_count": 0,
            "answer_only_count": 69,
            "ready_for_external_import": False,
        }
    }


def _standard_bill() -> dict[str, object]:
    return {"summary": {"standard_bill_row_count": 92, "standard_mapped_count": 69, "standard_unmapped_count": 23}}


def _quantity_placeholder() -> dict[str, object]:
    return {
        "summary": {
            "quantity_status": "blocked_until_three_field_gate_passed",
            "quantity_row_count": 92,
            "quantity_filled_count": 0,
            "can_enable_quantity": False,
        }
    }
