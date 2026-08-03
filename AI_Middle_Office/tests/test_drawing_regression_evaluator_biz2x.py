from __future__ import annotations

from openpyxl import load_workbook

from app.services.drawing_regression_evaluator import (
    build_dwg_regression_report,
    write_dwg_regression_outputs,
)


def test_biz2x_a7_dwg_regression_report_summarizes_readiness_and_blocks(tmp_path):
    listing_report = {
        "__source_filename": "BIZ2x_DWG上传列项_20260615_130000.json",
        "summary": {"dwg_file_count": 5},
        "project_recognition_summary": {
            "source_signal_count": 84,
            "matched_signal_count": 35,
            "recognized_project_count": 2,
            "unique_standard_item_count": 2,
        },
        "project_region_binding_summary": {"binding_ready_project_count": 1},
        "room_boundary_summary": {"room_boundary_count": 1},
        "project_rows": [
            {
                "识别项目编号": "P-001",
                "图纸项目名称": "石膏板饰面吊顶",
                "标准项目编码": "011302001",
                "项目名称": "平面吊顶天棚",
                "单位": "㎡",
            },
            {
                "识别项目编号": "P-002",
                "图纸项目名称": "洗手间墙面防水",
                "标准项目编码": "010904002",
                "项目名称": "墙面防水",
                "单位": "㎡",
            },
        ],
        "special_quantity_trace_rows": [
            {
                "专项算量编号": "BIZ2xSQ-00001",
                "项目名称": "平面吊顶天棚",
                "trace状态": "special_quantity_trace_ready_for_manual_review",
                "是否可复核": "是",
                "标准规则执行状态": "standard_rule_execution_ready_for_manual_review",
                "建议工程量": 16.8,
                "建议单位": "㎡",
            },
            {
                "专项算量编号": "BIZ2xSQ-00002",
                "项目名称": "墙面防水",
                "trace状态": "special_quantity_trace_blocked",
                "是否可复核": "否",
                "标准规则执行状态": "blocked_missing_room_perimeter",
                "建议工程量": "",
                "阻断原因": "缺少净周长候选，不能计算墙面防水面积",
            },
        ],
    }

    report = build_dwg_regression_report(
        [listing_report],
        reference_rows_by_sample={
            "BIZ2x_DWG上传列项_20260615_130000.json": [
                {"项目名称": "平面吊顶天棚", "单位": "㎡", "工程量": "16.8"}
            ]
        },
    )

    assert report["ok"] is True
    assert report["summary"]["sample_count"] == 1
    assert report["summary"]["recognized_project_count"] == 2
    assert report["summary"]["ready_special_trace_count"] == 1
    assert report["summary"]["blocked_special_trace_count"] == 1
    assert report["sample_rows"][0]["最终生成准备度"] == "可进入专项 trace 复核"
    assert report["issue_rows"][0]["专项算量编号"] == "BIZ2xSQ-00002"
    assert "净周长" in report["issue_rows"][0]["建议处理"]
    assert report["reference_compare_rows"][0]["识别状态"] == "已识别"

    outputs = write_dwg_regression_outputs(report, tmp_path, stem="regression")
    workbook = load_workbook(outputs["xlsx"])
    assert {"样例汇总", "问题清单", "参考清单对比", "统计摘要"}.issubset(set(workbook.sheetnames))
    assert workbook["样例汇总"]["A1"].value == "样例编号"
