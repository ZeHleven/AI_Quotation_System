from __future__ import annotations

import csv
import json
import re
from collections import Counter
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.services.drawing_project_lexicon import load_project_lexicon, match_source_signal_to_lexicon
from app.services.drawing_project_standard_mapping import (
    MAPPING_STATUS_OUT_OF_SCOPE,
    find_standard_mapping_for_lexicon_entry,
    load_project_standard_mapping,
)
from app.services.drawing_project_scope_review import (
    find_scope_review_for_mapping_entry,
    load_project_scope_review,
)


PROJECT_ROW_HEADERS = [
    "识别项目编号",
    "图纸项目名称",
    "标准项目编码",
    "项目名称",
    "项目特征",
    "单位",
    "工程量",
    "工程量状态",
    "来源文件",
    "来源类型",
    "识别证据",
    "匹配置信度",
    "匹配理由",
    "来源线索数",
]
FOUR_FIELD_HEADERS = ["项目名称", "项目特征", "单位", "工程量"]


def build_drawing_project_recognition_report(
    standard_match_report: dict[str, Any],
    *,
    project_lexicon: dict[str, Any] | None = None,
    project_standard_mapping: dict[str, Any] | None = None,
    project_scope_review: dict[str, Any] | None = None,
) -> dict[str, Any]:
    raw_rows = _build_raw_project_rows(standard_match_report)
    lexicon = project_lexicon if project_lexicon is not None else load_project_lexicon()
    standard_mapping = (
        project_standard_mapping
        if project_standard_mapping is not None
        else load_project_standard_mapping()
    )
    scope_review = project_scope_review if project_scope_review is not None else load_project_scope_review()
    lexicon_rows = _build_lexicon_project_rows(standard_match_report, raw_rows, lexicon, standard_mapping, scope_review)
    raw_rows = [*raw_rows, *lexicon_rows]
    project_rows = _merge_project_rows(raw_rows)
    source_type_counts = Counter(row["来源类型"] for row in project_rows)
    quantity_status_counts = Counter(row["工程量状态"] for row in project_rows)
    return {
        "ok": True,
        "phase": "BIZ-2x-project-recognition-standard-draft",
        "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "safe_for_final_quantity_list": False,
        "summary": {
            "source_signal_count": (standard_match_report.get("summary") or {}).get("source_signal_count", 0),
            "matched_signal_count": (standard_match_report.get("summary") or {}).get("matched_signal_count", 0),
            "raw_project_signal_count": len(raw_rows),
            "recognized_project_count": len(project_rows),
            "standard_project_count": sum(1 for row in project_rows if row.get("标准项目编码")),
            "lexicon_supplemental_project_count": len(lexicon_rows),
            "lexicon_entry_count": (lexicon.get("summary") or {}).get("lexicon_entry_count", 0),
            "standard_mapping_entry_count": (standard_mapping.get("summary") or {}).get("mapping_entry_count", 0),
            "standard_mapping_status_counts": (standard_mapping.get("summary") or {}).get("mapping_status_counts", {}),
            "scope_review_entry_count": (scope_review.get("summary") or {}).get("scope_review_entry_count", 0),
            "scope_review_action_counts": (scope_review.get("summary") or {}).get("review_action_counts", {}),
            "quantity_ready_count": 0,
            "quantity_pending_count": len(project_rows),
            "unique_standard_item_count": len({row["标准项目编码"] for row in project_rows if row["标准项目编码"]}),
            "source_type_counts": dict(source_type_counts.most_common()),
            "quantity_status_counts": dict(quantity_status_counts.most_common()),
            "final_generation_status": "project_recognition_ready_quantity_pending_region_binding",
            "next_step": "bind_project_to_cad_region_or_boundary_then_apply_standard_quantity_rule",
        },
        "project_rows": project_rows,
        "draft_four_field_rows": [
            {header: row[header] for header in FOUR_FIELD_HEADERS}
            for row in project_rows
        ],
        "notes": [
            "本报告只确认图纸项目识别和 GB/T 标准清单映射，不把工程量空白伪装成最终结果。",
            "项目特征字段名来自 active GB/T 标准库；候选值只来自图纸证据，缺失项保留为待补充。",
            "工程量必须在后续 CAD 区域/边界绑定完成后，按标准库工程量计算规则生成。",
        ],
    }


def write_drawing_project_recognition_outputs(
    report: dict[str, Any],
    output_dir: str | Path,
    *,
    stem: str | None = None,
) -> dict[str, str]:
    target_dir = Path(output_dir)
    target_dir.mkdir(parents=True, exist_ok=True)
    file_stem = stem or f"BIZ2x_图纸项目识别_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
    json_path = target_dir / f"{file_stem}.json"
    markdown_path = target_dir / f"{file_stem}.md"
    csv_path = target_dir / f"{file_stem}_项目识别清单.csv"
    xlsx_path = target_dir / f"{file_stem}_标准列项草稿四字段.xlsx"

    json_path.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    markdown_path.write_text(build_drawing_project_recognition_markdown(report), encoding="utf-8")
    _write_csv(csv_path, report.get("project_rows") or [], PROJECT_ROW_HEADERS)
    _write_project_workbook(xlsx_path, report)
    return {
        "json": str(json_path),
        "markdown": str(markdown_path),
        "project_csv": str(csv_path),
        "draft_four_field_xlsx": str(xlsx_path),
    }


def build_drawing_project_recognition_markdown(report: dict[str, Any]) -> str:
    summary = report.get("summary") or {}
    lines = [
        "# BIZ-2x 图纸项目识别与标准列项草稿",
        "",
        f"- 生成时间：{report.get('generated_at', '-')}",
        f"- 图纸线索数：{summary.get('source_signal_count', 0)}",
        f"- 已匹配线索数：{summary.get('matched_signal_count', 0)}",
        f"- 合并后图纸项目数：{summary.get('recognized_project_count', 0)}",
        f"- 工程量可生成数：{summary.get('quantity_ready_count', 0)}",
        f"- 工程量待区域绑定数：{summary.get('quantity_pending_count', 0)}",
        "",
        "## 图纸项目识别结果",
        "",
        "| 编号 | 图纸项目 | 标准项目 | 单位 | 工程量状态 | 来源 |",
        "| --- | --- | --- | --- | --- | --- |",
    ]
    for row in (report.get("project_rows") or [])[:120]:
        lines.append(
            "| "
            + " | ".join(
                [
                    _md(row.get("识别项目编号")),
                    _md(row.get("图纸项目名称")),
                    _md(row.get("项目名称")),
                    _md(row.get("单位")),
                    _md(row.get("工程量状态")),
                    _md(row.get("来源类型")),
                ]
            )
            + " |"
        )
    lines.extend(
        [
            "",
            "## 边界",
            "",
            "- 这不是最终工程量清单，工程量列必须等待 CAD 区域/边界绑定。",
            "- 后续只允许把有来源区域、计算公式和标准规则 trace 的行推进到最终四字段 Excel。",
        ]
    )
    return "\n".join(lines) + "\n"


def _build_raw_project_rows(standard_match_report: dict[str, Any]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for group in standard_match_report.get("candidate_groups") or []:
        signal = group.get("source_signal") or {}
        candidates = list(group.get("standard_candidates") or [])
        if not candidates:
            continue
        candidate = _select_project_candidate(signal, candidates)
        if not _should_include_project_signal(signal, candidate):
            continue
        feature_text = _project_feature_text(candidate)
        unit = _preferred_unit(candidate.get("unit_options") or [])
        rows.append(
            {
                "图纸项目名称": _clean_text(signal.get("source_name")),
                "标准项目编码": _clean_text(candidate.get("item_code")),
                "项目名称": _clean_text(candidate.get("item_name")),
                "项目特征": feature_text,
                "单位": unit,
                "工程量": "",
                "工程量状态": "待 CAD 区域/边界绑定后按标准规则计算",
                "来源文件": _clean_text(signal.get("source_file")),
                "来源类型": _clean_text(signal.get("source_kind_label")) or _clean_text(signal.get("source_kind")),
                "识别证据": _clean_text(signal.get("evidence_text")) or _clean_text(signal.get("raw_row_text")),
                "匹配置信度": f"{float(candidate.get('match_confidence') or 0):.2f}",
                "匹配理由": "；".join(_clean_text(item) for item in candidate.get("match_reasons") or [] if _clean_text(item)),
                "_feature_signature": _normalize(feature_text),
                "_recognition_source": "standard_match",
                "_source_signal_key": _source_signal_key(signal),
            }
        )
    return rows


def _build_lexicon_project_rows(
    standard_match_report: dict[str, Any],
    existing_raw_rows: list[dict[str, Any]],
    lexicon: dict[str, Any],
    standard_mapping: dict[str, Any],
    scope_review: dict[str, Any],
) -> list[dict[str, Any]]:
    if not lexicon.get("entries"):
        return []
    signals = list(standard_match_report.get("source_signals") or [])
    if not signals:
        return []
    existing_keys = {
        (
            _normalize(row.get("标准项目编码")),
            _normalize(row.get("图纸项目名称")),
            _normalize(row.get("项目名称")),
            row.get("_source_signal_key", ""),
        )
        for row in existing_raw_rows
    }
    rows: list[dict[str, Any]] = []
    seen: set[tuple[str, str, str]] = set()
    for signal in signals:
        signal_key = _source_signal_key(signal)
        for match in match_source_signal_to_lexicon(signal, lexicon):
            entry = match["entry"]
            mapping_entry = find_standard_mapping_for_lexicon_entry(entry, standard_mapping)
            if mapping_entry.get("mapping_status") == MAPPING_STATUS_OUT_OF_SCOPE:
                continue
            scope_decision = find_scope_review_for_mapping_entry(mapping_entry, scope_review)
            if scope_decision and not bool(scope_decision.get("recognition_allowed", True)):
                continue
            if scope_decision:
                min_score = float(scope_decision.get("recognition_min_score") or 0)
                if float(match.get("score") or 0) < min_score:
                    continue
            if not _should_include_lexicon_match(signal, entry, mapping_entry, scope_decision, match):
                continue
            source_name = _clean_text(signal.get("source_name")) or entry.get("manual_item_name", "")
            project_name = _lexicon_project_name(entry, mapping_entry)
            standard_code = _lexicon_standard_code(entry, mapping_entry)
            dedupe_key = (
                standard_code or _normalize(project_name),
                _normalize(entry.get("manual_item_name")),
                signal_key,
            )
            if dedupe_key in seen:
                continue
            seen.add(dedupe_key)
            existing_key = (
                _normalize(standard_code),
                _normalize(source_name),
                _normalize(project_name),
                signal_key,
            )
            if existing_key in existing_keys:
                continue
            feature_text = _lexicon_feature_text(entry, signal, mapping_entry)
            feature_signature = feature_text
            mapping_id = _clean_text(mapping_entry.get("mapping_id")) if mapping_entry else ""
            if mapping_id:
                feature_signature = f"{feature_text}；R2映射编号：{mapping_id}"
            elif _clean_text(entry.get("entry_id")):
                feature_signature = f"{feature_text}；词库编号：{entry.get('entry_id')}"
            rows.append(
                {
                    "图纸项目名称": source_name,
                    "标准项目编码": standard_code,
                    "项目名称": project_name,
                    "项目特征": feature_text,
                    "单位": _lexicon_unit(entry, mapping_entry),
                    "工程量": "",
                    "工程量状态": _lexicon_quantity_status(entry, mapping_entry, scope_decision),
                    "来源文件": _clean_text(signal.get("source_file")),
                    "来源类型": "样例答案词库补充候选",
                    "识别证据": _clean_text(signal.get("evidence_text")) or _clean_text(signal.get("raw_row_text")) or source_name,
                    "匹配置信度": f"{float(match.get('score') or 0):.2f}",
                    "匹配理由": _lexicon_match_reason(match, entry, mapping_entry, scope_decision),
                    "_feature_signature": _normalize(feature_signature),
                    "_recognition_source": "sample_answer_lexicon",
                    "_source_signal_key": signal_key,
                }
            )
    return rows


PROJECT_NAME_KEYWORDS = (
    "吊顶",
    "天棚",
    "天花",
    "地面",
    "楼地面",
    "地板",
    "墙面",
    "踢脚",
    "踢脚线",
    "防水",
    "门",
    "窗",
    "窗帘盒",
    "隔墙",
    "隔断",
    "涂料",
    "乳胶漆",
    "油漆",
    "裱糊",
    "墙纸",
)

NON_PROJECT_SIGNAL_PATTERNS = (
    r"^[·•]",
    r"^[A-ZＡ-Ｚ]\s",
    r"^[A-ZＡ-Ｚ][.、]",
    r"^\(?\d+\)?[.、]",
    r"^\d+(\.\d+)+[.、]?",
)

NON_PROJECT_SIGNAL_TERMS = (
    "必须",
    "采用",
    "应",
    "不得",
    "不能",
    "需",
    "要求",
    "详见",
    "间距",
    "调整",
    "增设",
    "安装",
    "设置",
    "选用",
    "处理",
    "符合",
    "注：",
    "用料做法",
    "工序",
    "打底",
    "扫毛",
    "压光",
    "夯实",
    "本图为",
)

MATERIAL_FRAGMENT_TERMS = (
    "龙骨",
    "吊筋",
    "夹板",
    "岩棉",
    "玻璃棉",
    "石膏板层",
    "砂浆",
    "水泥",
    "石膏板",
    "骨架",
    "饰面层",
    "面板",
)

TITLE_ONLY_TERMS = (
    "详图",
    "尺寸图",
    "剖面图",
    "节点图",
    "大样图",
    "索引图",
)

INSTRUCTION_NOTE_TERMS = (
    "必须",
    "应",
    "不得",
    "不能",
    "根据工程情况",
    "注意",
    "注：",
    "本图为",
    "参见",
    "详见",
    "防水附加层",
    "阴阳角",
)


def _should_include_project_signal(signal: dict[str, Any], candidate: dict[str, Any]) -> bool:
    source_name = _clean_text(signal.get("source_name"))
    evidence_text = _clean_text(signal.get("evidence_text")) or _clean_text(signal.get("raw_row_text"))
    source_kind = _clean_text(signal.get("source_kind"))
    candidate_name = _clean_text(candidate.get("item_name"))
    text = source_name or evidence_text
    normalized = _normalize(text)
    if not normalized or len(normalized) < 2:
        return False
    if len(normalized) > 42:
        return False
    if any(re.search(pattern, text) for pattern in NON_PROJECT_SIGNAL_PATTERNS):
        return False
    if _looks_like_title_or_note(text):
        return False
    if any(term in text for term in NON_PROJECT_SIGNAL_TERMS):
        return False
    if "拆除" in text and "拆除" not in candidate_name:
        return False
    if "天棚" in candidate_name and any(term in text for term in ("隔墙", "墙面", "原墙")) and not any(
        term in text for term in ("天棚", "吊顶", "天花")
    ):
        return False
    has_project_keyword = any(term in text for term in PROJECT_NAME_KEYWORDS)
    if not has_project_keyword:
        return False
    if "刮腻子" in candidate_name and "刮腻子" not in source_name:
        return False
    if any(term in text for term in MATERIAL_FRAGMENT_TERMS) and not any(
        term in text for term in ("吊顶", "天棚", "天花", "墙面", "隔墙", "隔断", "踢脚", "踢脚线", "涂料", "乳胶漆", "油漆")
    ):
        return False
    if source_kind in {"construction_method", "material"} and len(normalized) > 24:
        return False
    return True


def _should_include_lexicon_match(
    signal: dict[str, Any],
    entry: dict[str, Any],
    mapping_entry: dict[str, Any] | None,
    scope_decision: dict[str, Any] | None,
    match: dict[str, Any],
) -> bool:
    source_name = _clean_text(signal.get("source_name"))
    evidence_text = _clean_text(signal.get("evidence_text")) or _clean_text(signal.get("raw_row_text"))
    source_text = "；".join(part for part in [source_name, evidence_text] if part)
    normalized = _normalize(source_text)
    if not normalized:
        return False
    if _looks_like_title_or_note(source_text):
        return False
    score = float(match.get("score") or 0)
    manual_item = _clean_text(entry.get("manual_item_name"))
    manual_hit = manual_item and _normalize(manual_item) in normalized
    if _is_instruction_note(source_text) and not manual_hit:
        return False

    mapping = mapping_entry or {}
    category = _clean_text(mapping.get("category")) or _clean_text(entry.get("category"))
    standard_code = _lexicon_standard_code(entry, mapping)
    if "拆除" in source_text and "拆除" not in manual_item and category != "demolition":
        return False
    if category == "waterproof" or standard_code in {"010904002", "010903002"}:
        if _looks_like_coating_or_board_not_waterproof(source_text):
            return False
    if category == "floor" or standard_code in {"011102001", "011102003"}:
        if standard_code == "011102001" and not any(term in source_text for term in ("石材", "大理石", "门槛石", "ST-")):
            return False
        if _looks_like_floor_base_layer_only(source_text):
            return False
    if category == "sanitary" and any(term in source_text for term in ("防水附加层", "阴阳角", "根部")):
        return False
    if category == "ceiling" and "应根据工程情况" in source_text and not manual_hit:
        return False
    if scope_decision and _clean_text(scope_decision.get("review_action")) == "hold_other_specialty_until_scope_confirmed":
        return score >= float(scope_decision.get("recognition_min_score") or 0.82)
    return True


def _looks_like_title_or_note(text: str) -> bool:
    cleaned = _clean_text(text)
    if not cleaned:
        return False
    if "本图为" in cleaned:
        return True
    if any(cleaned.endswith(term) for term in TITLE_ONLY_TERMS):
        return True
    if any(term in cleaned for term in ("图纸目录", "设计说明", "材料表")) and len(_normalize(cleaned)) <= 18:
        return True
    return False


def _is_instruction_note(text: str) -> bool:
    return any(term in text for term in INSTRUCTION_NOTE_TERMS)


def _looks_like_coating_or_board_not_waterproof(text: str) -> bool:
    has_real_waterproof = any(term in text for term in ("涂膜防水", "聚氨酯", "防水层", "防水保护层", "地面防水", "墙面防水"))
    if has_real_waterproof:
        return False
    return any(term in text for term in ("防水石膏板", "防潮无机涂料", "无机涂料", "乳胶漆"))


def _looks_like_floor_base_layer_only(text: str) -> bool:
    has_floor_finish = any(
        term in text
        for term in ("地砖", "瓷砖", "石材", "大理石", "门槛石", "地板", "美缝", "CT-", "ST-", "块料")
    )
    if has_floor_finish:
        return False
    return any(term in text for term in ("水泥砂浆", "找平层", "结合层", "细石混凝土"))


def _select_project_candidate(signal: dict[str, Any], candidates: list[dict[str, Any]]) -> dict[str, Any]:
    source_name = _clean_text(signal.get("source_name"))
    evidence_text = _clean_text(signal.get("evidence_text"))
    source_text = source_name or evidence_text
    if not candidates:
        return {}
    scored: list[tuple[float, dict[str, Any]]] = []
    for index, candidate in enumerate(candidates):
        score = _project_candidate_priority(source_text, candidate)
        try:
            confidence = float(candidate.get("match_confidence") or 0)
        except (TypeError, ValueError):
            confidence = 0.0
        scored.append((score + confidence - index * 0.01, candidate))
    scored.sort(key=lambda item: item[0], reverse=True)
    return scored[0][1]


def _project_candidate_priority(source_text: str, candidate: dict[str, Any]) -> float:
    item_name = _clean_text(candidate.get("item_name"))
    score = 0.0
    if any(term in source_text for term in ("乳胶漆", "涂料", "油漆", "喷刷")):
        if any(term in item_name for term in ("涂料", "油漆", "喷刷", "裱糊")):
            score += 3.0
        if "吊顶" in item_name or "天棚" in item_name:
            score -= 1.0
    if "防水" in source_text and "防水" in item_name:
        score += 3.0
    if "踢脚" in source_text and "踢脚" in item_name:
        score += 3.0
    if any(term in source_text for term in ("地砖", "地面", "楼地面", "地板")) and any(
        term in item_name for term in ("楼地面", "地面", "地板")
    ):
        score += 2.5
    if any(term in source_text for term in ("吊顶", "天棚", "天花")) and any(term in item_name for term in ("吊顶", "天棚")):
        score += 2.5
    if any(term in source_text for term in ("墙纸", "裱糊")) and any(term in item_name for term in ("裱糊", "墙纸")):
        score += 2.5
    if item_name and item_name.replace(" | ", "") in source_text:
        score += 1.0
    return score


def _merge_project_rows(raw_rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    merged: dict[tuple[str, str], dict[str, Any]] = {}
    for row in raw_rows:
        key = (row["标准项目编码"], row["_feature_signature"])
        current = merged.get(key)
        if current is None:
            current = {
                **{header: row.get(header, "") for header in PROJECT_ROW_HEADERS},
                "_source_names": [],
                "_source_files": [],
                "_source_types": [],
                "_evidences": [],
                "_match_reasons": [],
                "_max_confidence": 0.0,
                "_source_count": 0,
            }
            merged[key] = current
        current["_source_count"] += 1
        _append_unique(current["_source_names"], row.get("图纸项目名称"))
        _append_unique(current["_source_files"], row.get("来源文件"))
        _append_unique(current["_source_types"], row.get("来源类型"))
        _append_unique(current["_evidences"], row.get("识别证据"), limit=5)
        for reason in str(row.get("匹配理由") or "").split("；"):
            _append_unique(current["_match_reasons"], reason, limit=8)
        try:
            current["_max_confidence"] = max(current["_max_confidence"], float(row.get("匹配置信度") or 0))
        except ValueError:
            pass

    result: list[dict[str, Any]] = []
    for index, row in enumerate(
        sorted(merged.values(), key=lambda item: (-item["_max_confidence"], item["项目名称"], item["项目特征"])),
        start=1,
    ):
        row["识别项目编号"] = f"BIZ2xP-{index:04d}"
        row["图纸项目名称"] = "；".join(row["_source_names"][:5])
        row["来源文件"] = "；".join(row["_source_files"][:5])
        row["来源类型"] = "；".join(row["_source_types"][:5])
        row["识别证据"] = "；".join(row["_evidences"][:5])
        row["匹配理由"] = "；".join(row["_match_reasons"][:8])
        row["匹配置信度"] = f"{row['_max_confidence']:.2f}"
        row["来源线索数"] = row["_source_count"]
        result.append({header: row.get(header, "") for header in PROJECT_ROW_HEADERS})
    return result


def _project_feature_text(candidate: dict[str, Any]) -> str:
    features = list(candidate.get("feature_fill_candidates") or [])
    if not features:
        return "待按标准项目特征字段补充"
    parts: list[str] = []
    has_named_field = False
    for feature in features:
        field_name = _clean_text(feature.get("field_name"))
        if not field_name:
            continue
        has_named_field = True
        value = _clean_text(feature.get("candidate_value")) or "待补充"
        parts.append(f"{field_name}：{value}")
    if parts:
        return "；".join(parts)
    if not has_named_field:
        return "无项目特征字段（标准表为空）"
    return "待按标准项目特征字段补充"


def _preferred_unit(units: list[Any]) -> str:
    cleaned = [_clean_text(unit) for unit in units if _clean_text(unit)]
    for preferred in ("㎡", "m²", "m2", "平方米", "m", "米", "个", "项"):
        if preferred in cleaned:
            return "㎡" if preferred in {"m²", "m2", "平方米"} else preferred
    return cleaned[0] if cleaned else ""


def _lexicon_project_name(entry: dict[str, Any], mapping_entry: dict[str, Any] | None = None) -> str:
    mapping = mapping_entry or {}
    return (
        _clean_text(mapping.get("standard_item_name"))
        or _clean_text(entry.get("standard_item_name"))
        or _clean_text(mapping.get("manual_item_name"))
        or _clean_text(entry.get("manual_item_name"))
    )


def _lexicon_standard_code(entry: dict[str, Any], mapping_entry: dict[str, Any] | None = None) -> str:
    mapping = mapping_entry or {}
    return _clean_text(mapping.get("standard_item_code")) or _clean_text(entry.get("standard_item_code"))


def _lexicon_unit(entry: dict[str, Any], mapping_entry: dict[str, Any] | None = None) -> str:
    mapping = mapping_entry or {}
    standard_units = list(mapping.get("standard_unit_options") or entry.get("standard_unit_options") or [])
    if _lexicon_standard_code(entry, mapping):
        return _preferred_unit(standard_units)
    return _clean_text(mapping.get("manual_unit")) or _clean_text(entry.get("manual_unit")) or _preferred_unit(standard_units)


def _lexicon_quantity_status(
    entry: dict[str, Any],
    mapping_entry: dict[str, Any] | None = None,
    scope_decision: dict[str, Any] | None = None,
) -> str:
    scope = scope_decision or {}
    if _clean_text(scope.get("final_quantity_status")):
        return _clean_text(scope.get("final_quantity_status"))
    mapping = mapping_entry or {}
    if _clean_text(mapping.get("quantity_status")):
        return _clean_text(mapping.get("quantity_status"))
    if _lexicon_standard_code(entry, mapping):
        return "待 CAD 区域/边界绑定后按标准规则计算"
    return "待 R2 标准映射确认，不进入最终算量"


def _lexicon_feature_text(
    entry: dict[str, Any],
    signal: dict[str, Any],
    mapping_entry: dict[str, Any] | None = None,
) -> str:
    mapping = mapping_entry or {}
    if _clean_text(mapping.get("feature_text_template")):
        return _clean_text(mapping.get("feature_text_template"))
    feature_fields = [
        _clean_text(field)
        for field in (mapping.get("standard_feature_fields") or entry.get("standard_feature_fields") or [])
        if _clean_text(field)
    ]
    if feature_fields:
        parts = []
        for field in feature_fields:
            parts.append(f"{field}：{_lexicon_feature_value(field, entry, signal)}")
        return "；".join(parts)
    if _lexicon_standard_code(entry, mapping):
        return "无项目特征字段（标准表为空）"
    category = _clean_text(entry.get("category_label")) or "待映射项目"
    item_name = _clean_text(entry.get("manual_item_name"))
    return f"补充清单类别：{category}；识别项目：{item_name}；标准映射状态：待 R2 确认"


def _lexicon_match_reason(
    match: dict[str, Any],
    entry: dict[str, Any],
    mapping_entry: dict[str, Any] | None = None,
    scope_decision: dict[str, Any] | None = None,
) -> str:
    reasons = [_clean_text(reason) for reason in match.get("reasons") or [] if _clean_text(reason)]
    mapping = mapping_entry or {}
    scope = scope_decision or {}
    if _clean_text(mapping.get("mapping_status")):
        reasons.append(f"R2映射状态：{mapping.get('mapping_status')}")
    if _clean_text(mapping.get("mapping_reason")):
        reasons.append(_clean_text(mapping.get("mapping_reason")))
    if _clean_text(scope.get("review_action")):
        reasons.append(f"R2-2复核动作：{scope.get('review_action')}")
    if _clean_text(scope.get("false_positive_guard")):
        reasons.append(_clean_text(scope.get("false_positive_guard")))
    if not reasons and _clean_text(entry.get("mapping_reason")):
        reasons.append(_clean_text(entry.get("mapping_reason")))
    return "；".join(_dedupe(reasons))


def _lexicon_feature_value(field_name: str, entry: dict[str, Any], signal: dict[str, Any]) -> str:
    field_norm = _normalize(field_name)
    source_text = "；".join(
        part
        for part in [
            _clean_text(signal.get("source_name")),
            _clean_text(signal.get("source_spec_or_method")),
            _clean_text(entry.get("manual_item_name")),
            _clean_text(entry.get("manual_feature")),
        ]
        if part
    )
    material_codes = list(entry.get("material_codes") or [])
    category = _clean_text(entry.get("category"))
    if any(key in field_norm for key in ("材料", "品种", "规格", "材质", "面层", "基层", "防水层", "涂料")):
        if material_codes:
            return f"{_clean_text(entry.get('manual_item_name'))}（材料编号：{'、'.join(material_codes)}）"
        return _short_feature_value(source_text)
    if any(key in field_norm for key in ("部位", "位置")):
        return _clean_text(signal.get("source_name")) or "待补充"
    if any(key in field_norm for key in ("形式", "类型")):
        if category == "ceiling":
            return _clean_text(entry.get("manual_item_name")) or "待补充"
        if category == "wall":
            return _clean_text(entry.get("manual_item_name")) or "待补充"
    if any(key in field_norm for key in ("高度", "厚度", "遍数", "强度等级", "做法")):
        return _short_feature_value(source_text) or "待补充"
    return _short_feature_value(source_text) or "待补充"


def _short_feature_value(value: str, limit: int = 120) -> str:
    text = _clean_text(value)
    if not text:
        return "待补充"
    return text[:limit]


def _source_signal_key(signal: dict[str, Any]) -> str:
    return "|".join(
        [
            _clean_text(signal.get("source_file")),
            _clean_text(signal.get("source_kind")),
            str(signal.get("source_row_number") or ""),
            _normalize(signal.get("source_name")),
            _normalize(signal.get("source_spec_or_method")),
        ]
    )


def _write_project_workbook(path: Path, report: dict[str, Any]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "四字段草稿"
    ws.append(FOUR_FIELD_HEADERS)
    for row in report.get("draft_four_field_rows") or []:
        ws.append([row.get(header, "") for header in FOUR_FIELD_HEADERS])
    evidence_ws = wb.create_sheet("识别证据")
    evidence_ws.append(PROJECT_ROW_HEADERS)
    for row in report.get("project_rows") or []:
        evidence_ws.append([row.get(header, "") for header in PROJECT_ROW_HEADERS])
    for sheet in wb.worksheets:
        _style_sheet(sheet)
    wb.save(path)


def _style_sheet(sheet: Any) -> None:
    fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    widths = {
        "A": 22,
        "B": 42,
        "C": 16,
        "D": 18,
        "E": 60,
        "F": 12,
        "G": 12,
        "H": 34,
        "I": 34,
        "J": 20,
        "K": 60,
        "L": 12,
        "M": 44,
        "N": 12,
    }
    for index in range(1, sheet.max_column + 1):
        letter = get_column_letter(index)
        sheet.column_dimensions[letter].width = widths.get(letter, 18)
    sheet.freeze_panes = "A2"


def _write_csv(path: Path, rows: list[dict[str, Any]], headers: list[str]) -> None:
    with path.open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.DictWriter(handle, fieldnames=headers)
        writer.writeheader()
        writer.writerows([{header: row.get(header, "") for header in headers} for row in rows])


def _append_unique(target: list[str], value: Any, *, limit: int | None = None) -> None:
    cleaned = _clean_text(value)
    if not cleaned or cleaned in target:
        return
    if limit is not None and len(target) >= limit:
        return
    target.append(cleaned)


def _normalize(value: Any) -> str:
    return re.sub(r"\s+", "", _clean_text(value)).lower()


def _clean_text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).replace("\r", "\n")
    text = re.sub(r"[ \t]+", " ", text)
    text = re.sub(r"\n{2,}", "\n", text)
    return text.strip()


def _dedupe(values: list[str]) -> list[str]:
    result: list[str] = []
    seen: set[str] = set()
    for value in values:
        cleaned = _clean_text(value)
        key = _normalize(cleaned)
        if not cleaned or key in seen:
            continue
        seen.add(key)
        result.append(cleaned)
    return result


def _md(value: Any) -> str:
    return str(value or "").replace("|", "\\|").replace("\n", "<br>")
