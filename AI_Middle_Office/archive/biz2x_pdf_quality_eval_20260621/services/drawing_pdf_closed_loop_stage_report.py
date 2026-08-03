from __future__ import annotations

import csv
import json
from datetime import datetime
from pathlib import Path
from typing import Any, Iterable, Mapping, Sequence

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


PHASE = "BIZ-2x-pdf-closed-loop-stage-report"

STAGE_HEADERS = [
    "stage_no",
    "stage_name",
    "status",
    "key_metric",
    "blocker",
    "next_action",
    "evidence",
]

ARTIFACT_HEADERS = ["artifact", "path", "exists", "role"]


def build_closed_loop_stage_report(
    *,
    v2_report: Mapping[str, Any] | None = None,
    template_status_report: Mapping[str, Any] | None = None,
    import_report: Mapping[str, Any] | None = None,
    evaluation_report: Mapping[str, Any] | None = None,
    review_report: Mapping[str, Any] | None = None,
    object_recall_report: Mapping[str, Any] | None = None,
    object_workbench_report: Mapping[str, Any] | None = None,
    gate_report: Mapping[str, Any] | None = None,
    standard_bill_report: Mapping[str, Any] | None = None,
    quantity_report: Mapping[str, Any] | None = None,
    pipeline_report: Mapping[str, Any] | None = None,
    artifacts: Mapping[str, str] | None = None,
) -> dict[str, Any]:
    pipeline_report = pipeline_report or {}
    review_summary = _first_mapping(
        (review_report or {}).get("summary"),
        pipeline_report.get("review_summary"),
        (v2_report or {}).get("summary"),
        (evaluation_report or {}).get("augmented_summary"),
    )
    gate_summary = _first_mapping((gate_report or {}).get("summary"), pipeline_report.get("gate_summary"))
    object_summary = _first_mapping(
        (object_recall_report or {}).get("summary"),
        pipeline_report.get("object_recall_summary"),
    )
    template_summary = _first_mapping(
        (template_status_report or {}).get("summary"),
        pipeline_report.get("template_status_summary"),
    )
    object_template_summary = _first_mapping(
        (object_workbench_report or {}).get("summary"),
        pipeline_report.get("object_workbench_summary"),
        template_summary if _is_object_recall_template_status(template_summary) else {},
    )
    import_summary = _first_mapping((import_report or {}).get("summary"), pipeline_report.get("import_summary"))
    standard_summary = _first_mapping(
        (standard_bill_report or {}).get("summary"),
        pipeline_report.get("standard_bill_summary"),
    )
    quantity_summary = _first_mapping(
        (quantity_report or {}).get("summary"),
        pipeline_report.get("quantity_placeholder_summary"),
    )
    eval_summary = _first_mapping((evaluation_report or {}).get("summary"), pipeline_report.get("evaluation_summary"))

    can_enable_quantity = _bool(
        _first_value(
            gate_report or {},
            "can_enable_quantity",
            default=pipeline_report.get("can_enable_quantity"),
        )
    )
    if not gate_report and gate_summary:
        can_enable_quantity = _bool(gate_summary.get("can_enable_quantity"), default=can_enable_quantity)

    stage_rows = [
        _stage_pdf_candidate(v2_report, eval_summary, review_summary, import_summary),
        _stage_three_field(review_summary, can_enable_quantity),
        _stage_object_recall(object_summary, object_template_summary),
        _stage_standard_bill(standard_summary, can_enable_quantity),
        _stage_excel_outputs(artifacts or {}, pipeline_report),
        _stage_gate(gate_summary, can_enable_quantity),
        _stage_quantity(quantity_summary, can_enable_quantity),
    ]
    blocker_rows = [row for row in stage_rows if row["status"] in {"blocked", "needs_evidence", "failed", "review_only"}]
    artifact_rows = _artifact_rows(artifacts or {}, pipeline_report)
    summary = {
        "stage_count": len(stage_rows),
        "complete_stage_count": sum(1 for row in stage_rows if row["status"] == "complete"),
        "blocked_stage_count": len(blocker_rows),
        "current_stage_no": _current_stage_no(stage_rows),
        "current_blocker": blocker_rows[0]["blocker"] if blocker_rows else "",
        "next_action": blocker_rows[0]["next_action"] if blocker_rows else "三字段闭环已通过，可进入工程量阶段设计验收。",
        "can_enable_quantity": can_enable_quantity,
        "quantity_status": quantity_summary.get("quantity_status")
        or gate_summary.get("quantity_status")
        or ("ready_after_three_field_acceptance" if can_enable_quantity else "deferred_until_three_fields_accepted"),
        "answer_count": _int(_first_value(review_summary, "answer_count", "three_field_answer_count"), 0),
        "candidate_count": _int(_first_value(review_summary, "candidate_count", "three_field_candidate_count"), 0),
        "matched_three_fields_count": _int(
            _first_value(review_summary, "matched_three_fields_count", "three_field_matched_count"),
            0,
        ),
        "missing_candidate_count": _status_count(review_summary, "missing_candidate"),
        "feature_review_count": _status_count(review_summary, "matched_name_unit_feature_review"),
        "unit_conflict_count": _status_count(review_summary, "unit_conflict"),
        "weak_match_count": _status_count(review_summary, "weak_match_review"),
        "object_recall_task_count": _int(object_summary.get("object_recall_task_count"), 0),
        "object_importable_row_count": _int(object_template_summary.get("importable_row_count"), 0),
        "object_answer_only_count": _int(object_template_summary.get("answer_only_count"), 0),
        "standard_bill_row_count": _int(standard_summary.get("standard_bill_row_count"), 0),
        "standard_mapped_count": _int(_first_value(standard_summary, "mapped_row_count", "standard_mapped_count"), 0),
        "standard_unmapped_count": _int(_first_value(standard_summary, "unmapped_row_count", "standard_unmapped_count"), 0),
        "quantity_filled_count": _int(quantity_summary.get("quantity_filled_count"), 0),
    }
    return {
        "ok": True,
        "phase": PHASE,
        "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "summary": summary,
        "stage_rows": stage_rows,
        "artifact_rows": artifact_rows,
    }


def write_closed_loop_stage_report_outputs(
    report: Mapping[str, Any],
    output_dir: str | Path,
    *,
    stem: str | None = None,
) -> dict[str, str]:
    target = Path(output_dir)
    target.mkdir(parents=True, exist_ok=True)
    file_stem = stem or f"BIZ2x_PDF_closed_loop_stage_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
    json_path = target / f"{file_stem}.json"
    csv_path = target / f"{file_stem}_stages.csv"
    artifact_csv_path = target / f"{file_stem}_artifacts.csv"
    markdown_path = target / f"{file_stem}.md"
    xlsx_path = target / f"{file_stem}.xlsx"
    outputs = {
        "json": str(json_path),
        "stage_csv": str(csv_path),
        "artifact_csv": str(artifact_csv_path),
        "markdown": str(markdown_path),
        "xlsx": str(xlsx_path),
    }
    payload = {**dict(report), "outputs": outputs}
    json_path.write_text(json.dumps(_json_safe(payload), ensure_ascii=False, indent=2), encoding="utf-8")
    _write_csv(csv_path, payload.get("stage_rows") or [], STAGE_HEADERS)
    _write_csv(artifact_csv_path, payload.get("artifact_rows") or [], ARTIFACT_HEADERS)
    markdown_path.write_text(_build_markdown(payload), encoding="utf-8")
    _write_workbook(xlsx_path, payload)
    return outputs


def load_json_report(path: str | Path | None) -> dict[str, Any]:
    if not path:
        return {}
    source = Path(path)
    if not source.exists():
        return {}
    return json.loads(source.read_text(encoding="utf-8"))


def _stage_pdf_candidate(
    v2_report: Mapping[str, Any] | None,
    eval_summary: Mapping[str, Any],
    review_summary: Mapping[str, Any],
    import_summary: Mapping[str, Any],
) -> dict[str, Any]:
    summary = (v2_report or {}).get("summary") or {}
    evidence_count = _int(summary.get("evidence_count"), 0)
    if not evidence_count:
        evidence_count = _int(eval_summary.get("recall_evidence_count"), 0) + _int(import_summary.get("evidence_count"), 0)
    candidate_count = _int(_first_value(review_summary, "candidate_count", "three_field_candidate_count"), 0)
    status = "complete" if candidate_count > 0 else "blocked"
    return _stage_row(
        1,
        "PDF输入与候选列项",
        status,
        f"evidence={evidence_count}; candidates={candidate_count}",
        "" if status == "complete" else "系统候选行为 0，无法进入三字段验收。",
        "" if status == "complete" else "先运行 PDF V2 识图或导入外部/对象召回证据。",
        "V2 summary / import summary",
    )


def _stage_three_field(review_summary: Mapping[str, Any], can_enable_quantity: bool) -> dict[str, Any]:
    answer_count = _int(_first_value(review_summary, "answer_count", "three_field_answer_count"), 0)
    candidate_count = _int(_first_value(review_summary, "candidate_count", "three_field_candidate_count"), 0)
    matched = _int(_first_value(review_summary, "matched_three_fields_count", "three_field_matched_count"), 0)
    missing = _status_count(review_summary, "missing_candidate")
    feature_review = _status_count(review_summary, "matched_name_unit_feature_review")
    unit_conflict = _status_count(review_summary, "unit_conflict")
    weak = _status_count(review_summary, "weak_match_review")
    status = "complete" if can_enable_quantity else "failed"
    blocker = ""
    action = ""
    if status != "complete":
        blocker = f"三字段未全通过：missing={missing}, feature_review={feature_review}, unit_conflict={unit_conflict}, weak={weak}"
        action = "优先补 missing_candidate 的对象证据，再处理单位冲突和项目特征复核。"
    return _stage_row(
        2,
        "三字段验收",
        status,
        f"matched={matched}/{answer_count}; candidates={candidate_count}",
        blocker,
        action,
        "three_field_review / gate",
    )


def _stage_object_recall(object_summary: Mapping[str, Any], template_summary: Mapping[str, Any]) -> dict[str, Any]:
    task_count = _int(object_summary.get("object_recall_task_count"), 0)
    importable = _int(template_summary.get("importable_row_count"), 0)
    answer_only = _int(template_summary.get("answer_only_count"), 0)
    has_object_status = bool(template_summary)
    if task_count == 0:
        status = "complete"
        blocker = ""
        action = ""
    elif importable > 0:
        status = "needs_evidence"
        blocker = f"对象召回工作台已有 {importable} 行可导入证据，尚未回灌验收。"
        action = "运行 external_recall_acceptance_pipeline 导入填写后的对象召回工作台。"
    elif not has_object_status:
        status = "needs_evidence"
        blocker = f"对象召回任务 {task_count} 行已生成，等待填写真实图纸证据。"
        action = "打开 object_recall_workbench Excel，填写 evidence_item_hint/spec_or_method/suggested_unit/text 四类证据字段。"
    else:
        status = "needs_evidence"
        blocker = f"对象召回任务 {task_count} 行，当前 {answer_only} 行只有答案参考，没有真实图纸证据。"
        action = "填写 evidence_item_hint/spec_or_method/suggested_unit/text 四类证据字段。"
    return _stage_row(
        3,
        "对象级补召回",
        status,
        f"tasks={task_count}; importable={importable}; answer_only={answer_only}",
        blocker,
        action,
        "object_recall_workbench / template_status",
    )


def _stage_standard_bill(standard_summary: Mapping[str, Any], can_enable_quantity: bool) -> dict[str, Any]:
    rows = _int(standard_summary.get("standard_bill_row_count"), 0)
    mapped = _int(_first_value(standard_summary, "mapped_row_count", "standard_mapped_count"), 0)
    unmapped = _int(_first_value(standard_summary, "unmapped_row_count", "standard_unmapped_count"), 0)
    if rows == 0:
        status = "blocked"
        blocker = "没有国标清单预览行。"
        action = "先完成三字段候选列项与标准库映射。"
    elif not can_enable_quantity:
        status = "review_only"
        blocker = "三字段门禁未通过，国标清单只能作为复核预览。"
        action = "先通过三字段验收，再确认最终国标清单行。"
    elif unmapped:
        status = "failed"
        blocker = f"仍有 {unmapped} 行未映射国标。"
        action = "补标准库映射或人工确认项目编码。"
    else:
        status = "complete"
        blocker = ""
        action = ""
    return _stage_row(
        4,
        "国标清单格式映射",
        status,
        f"rows={rows}; mapped={mapped}; unmapped={unmapped}",
        blocker,
        action,
        "standard_bill_preview",
    )


def _stage_excel_outputs(artifacts: Mapping[str, str], pipeline_report: Mapping[str, Any]) -> dict[str, Any]:
    artifact_rows = _artifact_rows(artifacts, pipeline_report)
    xlsx_count = sum(1 for row in artifact_rows if str(row.get("path") or "").lower().endswith(".xlsx") and row.get("exists"))
    status = "complete" if xlsx_count >= 3 else "blocked"
    return _stage_row(
        5,
        "Excel验收输出",
        status,
        f"xlsx_outputs={xlsx_count}",
        "" if status == "complete" else "缺少三字段、国标预览或工程量占位 Excel 输出。",
        "" if status == "complete" else "重新运行验收流水线并检查各阶段 xlsx 产物。",
        "outputs",
    )


def _stage_gate(gate_summary: Mapping[str, Any], can_enable_quantity: bool) -> dict[str, Any]:
    failed = _int(gate_summary.get("failed_gate_count"), 0)
    status = "complete" if can_enable_quantity else "failed"
    return _stage_row(
        6,
        "三字段质量门禁",
        status,
        f"failed_gate_count={failed}; can_enable_quantity={can_enable_quantity}",
        "" if status == "complete" else "三字段门禁未通过，工程量必须继续锁定。",
        "" if status == "complete" else "处理 gate_checks 中的 blocker 后再复验。",
        "three_field_gate",
    )


def _stage_quantity(quantity_summary: Mapping[str, Any], can_enable_quantity: bool) -> dict[str, Any]:
    filled = _int(quantity_summary.get("quantity_filled_count"), 0)
    quantity_status = str(quantity_summary.get("quantity_status") or "")
    if can_enable_quantity and quantity_status.startswith("placeholder_only"):
        status = "blocked"
        blocker = "三字段已通过但工程量引擎尚未启用；当前仍是占位。"
        action = "进入工程量规则/图纸尺寸提取阶段。"
    elif can_enable_quantity:
        status = "complete"
        blocker = ""
        action = ""
    else:
        status = "blocked"
        blocker = "三字段未通过，工程量阶段按规则锁定。"
        action = "继续三字段补召回，不验收工程量。"
    return _stage_row(
        7,
        "工程量识别",
        status,
        f"quantity_status={quantity_status or '-'}; filled={filled}",
        blocker,
        action,
        "quantity_stage_placeholder",
    )


def _stage_row(
    stage_no: int,
    stage_name: str,
    status: str,
    key_metric: str,
    blocker: str,
    next_action: str,
    evidence: str,
) -> dict[str, Any]:
    return {
        "stage_no": stage_no,
        "stage_name": stage_name,
        "status": status,
        "key_metric": key_metric,
        "blocker": blocker,
        "next_action": next_action,
        "evidence": evidence,
    }


def _artifact_rows(artifacts: Mapping[str, str], pipeline_report: Mapping[str, Any]) -> list[dict[str, Any]]:
    merged: dict[str, str] = {str(key): str(value) for key, value in artifacts.items() if value}
    for output_key in (
        "template_status_outputs",
        "import_outputs",
        "eval_outputs",
        "review_outputs",
        "object_recall_outputs",
        "gate_outputs",
        "standard_bill_outputs",
        "quantity_placeholder_outputs",
    ):
        outputs = pipeline_report.get(output_key)
        if not isinstance(outputs, Mapping):
            continue
        for kind, path in outputs.items():
            merged[f"{output_key}.{kind}"] = str(path)
    rows: list[dict[str, Any]] = []
    for key, value in sorted(merged.items()):
        rows.append({"artifact": key, "path": value, "exists": Path(value).exists(), "role": _artifact_role(key)})
    return rows


def _artifact_role(key: str) -> str:
    if "review" in key:
        return "三字段验收"
    if "object_recall" in key:
        return "对象补召回"
    if "standard_bill" in key or "stdbill" in key:
        return "国标清单预览"
    if "quantity" in key or "qty" in key:
        return "工程量占位"
    if "gate" in key:
        return "质量门禁"
    return "中间产物"


def _is_object_recall_template_status(summary: Mapping[str, Any]) -> bool:
    source_path = str(summary.get("source_path") or "").replace("\\", "/").lower()
    if "object_recall" in source_path or "object_pack" in source_path:
        return True
    status_counts = summary.get("status_counts") or {}
    return bool(summary.get("input_row_count")) and bool(status_counts) and str(summary.get("source_kind") or "") == "object_recall_pack"


def _current_stage_no(stage_rows: Sequence[Mapping[str, Any]]) -> int:
    for row in stage_rows:
        if row.get("status") != "complete":
            return _int(row.get("stage_no"), 0)
    return _int(stage_rows[-1].get("stage_no"), 0) if stage_rows else 0


def _status_count(summary: Mapping[str, Any], status: str) -> int:
    counts = summary.get("status_counts") or summary.get("three_field_gap_status_counts") or {}
    if not isinstance(counts, Mapping):
        return 0
    return _int(counts.get(status), 0)


def _first_mapping(*values: Any) -> dict[str, Any]:
    for value in values:
        if isinstance(value, Mapping) and value:
            return dict(value)
    return {}


def _first_value(row: Mapping[str, Any], *keys: str, default: Any = "") -> Any:
    for key in keys:
        value = row.get(key)
        if value not in (None, ""):
            return value
    return default


def _bool(value: Any, *, default: bool = False) -> bool:
    if value in (None, ""):
        return default
    if isinstance(value, bool):
        return value
    return str(value).strip().lower() in {"1", "true", "yes", "y", "passed", "pass"}


def _int(value: Any, default: int = 0) -> int:
    try:
        return int(float(str(value).strip()))
    except (TypeError, ValueError):
        return default


def _build_markdown(report: Mapping[str, Any]) -> str:
    summary = report.get("summary") or {}
    lines = [
        "# BIZ-2x PDF Closed Loop Stage Report",
        "",
        f"- generated_at: {report.get('generated_at', '-')}",
        f"- current_stage_no: {summary.get('current_stage_no', 0)}",
        f"- can_enable_quantity: {summary.get('can_enable_quantity', False)}",
        f"- current_blocker: {_md(summary.get('current_blocker'))}",
        f"- next_action: {_md(summary.get('next_action'))}",
        "",
        "| stage | name | status | metric | blocker | next action |",
        "| ---: | --- | --- | --- | --- | --- |",
    ]
    for row in report.get("stage_rows") or []:
        lines.append(
            "| "
            + " | ".join(
                [
                    _md(row.get("stage_no")),
                    _md(row.get("stage_name")),
                    _md(row.get("status")),
                    _md(row.get("key_metric")),
                    _md(row.get("blocker")),
                    _md(row.get("next_action")),
                ]
            )
            + " |"
        )
    return "\n".join(lines) + "\n"


def _write_workbook(path: Path, report: Mapping[str, Any]) -> None:
    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "closed_loop_summary"
    _append_rows(summary_sheet, [["metric", "value"]])
    _append_rows(summary_sheet, [[key, _cell_value(value)] for key, value in (report.get("summary") or {}).items()])
    _style_sheet(summary_sheet)

    stage_sheet = workbook.create_sheet("stage_status")
    _append_rows(stage_sheet, [STAGE_HEADERS])
    _append_rows(stage_sheet, [[_cell_value(row.get(header)) for header in STAGE_HEADERS] for row in report.get("stage_rows") or []])
    _style_sheet(stage_sheet)

    artifact_sheet = workbook.create_sheet("artifacts")
    _append_rows(artifact_sheet, [ARTIFACT_HEADERS])
    _append_rows(
        artifact_sheet,
        [[_cell_value(row.get(header)) for header in ARTIFACT_HEADERS] for row in report.get("artifact_rows") or []],
    )
    _style_sheet(artifact_sheet)
    workbook.save(path)


def _write_csv(path: Path, rows: Iterable[Mapping[str, Any]], fieldnames: Sequence[str]) -> None:
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        for row in rows:
            writer.writerow({field: _cell_value(row.get(field)) for field in fieldnames})


def _append_rows(sheet: Any, rows: Iterable[Sequence[Any]]) -> None:
    for row in rows:
        sheet.append(list(row))


def _style_sheet(sheet: Any) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    if sheet.max_row >= 1:
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for col_idx, column_cells in enumerate(sheet.columns, start=1):
        values = [str(cell.value or "") for cell in column_cells[:200]]
        width = min(max([len(value) for value in values] + [10]) + 2, 80)
        sheet.column_dimensions[get_column_letter(col_idx)].width = width


def _cell_value(value: Any) -> Any:
    if isinstance(value, (dict, list, tuple)):
        return json.dumps(value, ensure_ascii=False)
    return value


def _json_safe(value: Any) -> Any:
    if isinstance(value, Mapping):
        return {str(key): _json_safe(item) for key, item in value.items()}
    if isinstance(value, (list, tuple)):
        return [_json_safe(item) for item in value]
    if isinstance(value, Path):
        return str(value)
    return value


def _md(value: Any) -> str:
    return str(value or "").replace("|", "\\|").replace("\n", "<br>")
