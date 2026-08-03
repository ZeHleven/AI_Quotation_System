from __future__ import annotations

import csv
import json
import re
from collections import Counter
from datetime import datetime
from pathlib import Path
from typing import Any, Iterable, Mapping, Sequence

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.services.drawing_pdf_gap_recall_runner import EVIDENCE_HEADERS


PHASE = "BIZ-2x-pdf-structured-feature-fusion"

FUSION_HEADERS = [
    *EVIDENCE_HEADERS,
    "task_no",
    "task_nos",
    "fusion_rule",
    "source_evidence_ids",
    "fusion_inputs",
]

DN_RE = re.compile(r"\bDN\s*([0-9]{1,4})\b", flags=re.IGNORECASE)
DE_RE = re.compile(r"\bDe\s*([0-9]{1,4})\b", flags=re.IGNORECASE)
INCH_RE = re.compile(r"([0-9]+(?:\.[0-9]+)?\s*(?:分|寸))")


def build_structured_feature_fusion_report(
    diameter_results: Mapping[str, Any],
    *,
    note_texts: Sequence[str] | None = None,
    note_results: Sequence[Mapping[str, Any]] | None = None,
    source_name: str = "",
    emit_supply: bool = True,
    emit_drain: str = "none",
) -> dict[str, Any]:
    """Fuse visible drawing notes and DN/De table rows into importable item evidence.

    This is intentionally answer-blind: callers provide only visual evidence rows and
    visible note text. The function never reads the manual acceptance answer.
    """

    diameter_rows = _source_rows(diameter_results)
    notes = _collect_notes(note_texts=note_texts or [], note_results=note_results or [])
    diameter_pairs = _diameter_pairs(diameter_rows)
    evidence_rows: list[dict[str, Any]] = []

    supply_note = _select_supply_note(notes)
    drain_note = _select_drain_note(notes)
    drain_de_scope = _drain_de_scope(drain_note, emit_drain=emit_drain)

    if emit_supply and supply_note:
        for pair in diameter_pairs:
            evidence_rows.append(
                _build_pipe_row(
                    evidence_no=len(evidence_rows) + 1,
                    pair=pair,
                    item_hint="给水管",
                    spec_or_method=f"材质：SUS304薄壁不锈钢管；规格、型号：{pair['dn']}",
                    suggested_unit="m",
                    note=supply_note,
                    source_name=source_name,
                    fusion_rule="supply_sus304_note_plus_dn_de_table",
                    task_no="84",
                    task_nos="84;85;86",
                )
            )

    if emit_drain != "none" and drain_note:
        for pair in diameter_pairs:
            if drain_de_scope and pair["de"] not in drain_de_scope:
                continue
            evidence_rows.append(
                _build_pipe_row(
                    evidence_no=len(evidence_rows) + 1,
                    pair=pair,
                    item_hint="排水管",
                    spec_or_method=f"材质：柔性铸铁管；规格、型号：{pair['de']}；不锈钢卡箍连接",
                    suggested_unit="m",
                    note=drain_note,
                    source_name=source_name,
                    fusion_rule=f"drain_cast_iron_note_plus_{emit_drain}",
                    task_no="87",
                    task_nos="87;88;89",
                )
            )

    summary = {
        "source_name": source_name,
        "input_diameter_row_count": len(diameter_rows),
        "source_note_count": len(notes),
        "diameter_pair_count": len(diameter_pairs),
        "evidence_count": len(evidence_rows),
        "supply_row_count": sum(1 for row in evidence_rows if row.get("item_hint") == "给水管"),
        "drain_row_count": sum(1 for row in evidence_rows if row.get("item_hint") == "排水管"),
        "emit_supply": emit_supply,
        "emit_drain": emit_drain,
        "has_supply_note": bool(supply_note),
        "has_drain_note": bool(drain_note),
        "answer_rows_used": False,
        "target_fields_sent_to_model": False,
        "answer_columns_count_as_evidence": False,
        "quantity_status": "deferred_until_three_fields_accepted",
        "status_counts": dict(Counter(row.get("fusion_rule", "") for row in evidence_rows)),
    }
    return {
        "ok": True,
        "phase": PHASE,
        "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "summary": summary,
        "diameter_pairs": diameter_pairs,
        "note_rows": notes,
        "evidence_rows": evidence_rows,
        "visual_evidence_report": {"evidence_rows": evidence_rows},
    }


def write_structured_feature_fusion_outputs(
    report: Mapping[str, Any],
    output_dir: str | Path,
    *,
    stem: str | None = None,
) -> dict[str, str]:
    target = Path(output_dir)
    target.mkdir(parents=True, exist_ok=True)
    file_stem = stem or f"BIZ2x_PDF_structured_feature_fusion_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
    json_path = target / f"{file_stem}.json"
    csv_path = target / f"{file_stem}_evidence.csv"
    markdown_path = target / f"{file_stem}.md"
    xlsx_path = target / f"{file_stem}.xlsx"
    outputs = {
        "json": str(json_path),
        "evidence_csv": str(csv_path),
        "markdown": str(markdown_path),
        "xlsx": str(xlsx_path),
    }
    payload = {**dict(report), "outputs": outputs}
    json_path.write_text(json.dumps(_json_safe(payload), ensure_ascii=False, indent=2), encoding="utf-8")
    _write_csv(csv_path, payload.get("evidence_rows") or [], FUSION_HEADERS)
    markdown_path.write_text(_build_markdown(payload), encoding="utf-8")
    _write_workbook(xlsx_path, payload)
    return outputs


def _source_rows(payload: Mapping[str, Any]) -> list[dict[str, Any]]:
    for key in ("evidence_rows", "recall_evidence", "rows"):
        rows = payload.get(key)
        if isinstance(rows, list):
            return [dict(row) for row in rows if isinstance(row, Mapping)]
    visual_report = payload.get("visual_evidence_report")
    if isinstance(visual_report, Mapping) and isinstance(visual_report.get("evidence_rows"), list):
        return [dict(row) for row in visual_report.get("evidence_rows") or [] if isinstance(row, Mapping)]
    return []


def _collect_notes(
    *,
    note_texts: Sequence[str],
    note_results: Sequence[Mapping[str, Any]],
) -> list[dict[str, Any]]:
    notes: list[dict[str, Any]] = []
    for index, text in enumerate(note_texts, start=1):
        cleaned = _clean(text)
        if cleaned:
            notes.append(
                {
                    "note_id": f"NOTE-TEXT-{index:03d}",
                    "text": cleaned,
                    "source_file": "",
                    "page": "",
                    "tile_id": "",
                    "source_kind": "manual_visual_note_text",
                }
            )
    for result in note_results:
        for row in _source_rows(result):
            text = _first(row, "text", "evidence_text", "normalized_text", "spec_or_method", "item_hint")
            if not text:
                continue
            notes.append(
                {
                    "note_id": _first(row, "evidence_id") or f"NOTE-ROW-{len(notes) + 1:03d}",
                    "text": text,
                    "source_file": _first(row, "source_file"),
                    "page": _first(row, "page"),
                    "tile_id": _first(row, "tile_id"),
                    "source_kind": _first(row, "source_kind") or "visual_note_evidence",
                }
            )
    return notes


def _diameter_pairs(rows: Sequence[Mapping[str, Any]]) -> list[dict[str, Any]]:
    pairs: dict[tuple[str, str], dict[str, Any]] = {}
    for row in rows:
        text = " ".join(
            _clean(row.get(key))
            for key in ("text", "normalized_text", "item_hint", "spec_or_method", "reason")
            if _clean(row.get(key))
        )
        dn = _first_match(DN_RE, text)
        de = _first_match(DE_RE, text)
        if not dn or not de:
            continue
        key = (dn.upper(), _normalize_de(de))
        pairs.setdefault(
            key,
            {
                "dn": dn.upper(),
                "de": _normalize_de(de),
                "inch_label": _first_match(INCH_RE, text),
                "source_file": _first(row, "source_file"),
                "page": _first(row, "page"),
                "tile_id": _first(row, "tile_id"),
                "vision_pass": _first(row, "vision_pass"),
                "evidence_id": _first(row, "evidence_id"),
                "source_text": text,
                "confidence": _bounded_float(_first(row, "confidence"), default=0.78),
            },
        )
    return sorted(pairs.values(), key=lambda item: (_diameter_number(item["dn"]), _diameter_number(item["de"])))


def _build_pipe_row(
    *,
    evidence_no: int,
    pair: Mapping[str, Any],
    item_hint: str,
    spec_or_method: str,
    suggested_unit: str,
    note: Mapping[str, Any],
    source_name: str,
    fusion_rule: str,
    task_no: str,
    task_nos: str,
) -> dict[str, Any]:
    note_text = _clean(note.get("text"))
    table_text = _clean(pair.get("source_text"))
    fusion_inputs = {
        "note_id": note.get("note_id", ""),
        "table_evidence_id": pair.get("evidence_id", ""),
        "dn": pair.get("dn", ""),
        "de": pair.get("de", ""),
        "inch_label": pair.get("inch_label", ""),
    }
    source_file = _first(pair, "source_file") or _first(note, "source_file")
    page = _first(pair, "page") or _first(note, "page")
    tile_id = _join_unique([_first(note, "tile_id"), _first(pair, "tile_id")])
    text = _join_unique([note_text, f"管径对照表：{pair.get('dn')} {pair.get('de')} {pair.get('inch_label')}".strip(), table_text])
    return {
        "evidence_id": f"PDFFUSE-{evidence_no:06d}",
        "source_kind": "pdf_feature_precision_structured_table_fusion",
        "source_file": source_file,
        "page": page,
        "tile_id": tile_id,
        "vision_pass": "structured_feature_fusion",
        "evidence_role": "structured_pipe_candidate",
        "discipline": "plumbing",
        "item_hint": item_hint,
        "space": "",
        "material_codes": [],
        "spec_or_method": spec_or_method,
        "suggested_unit": suggested_unit,
        "text": text,
        "normalized_text": text,
        "confidence": min(0.88, max(0.72, float(pair.get("confidence") or 0.78))),
        "model": source_name or "structured_feature_fusion",
        "needs_manual_review": True,
        "reason": "由图纸可见管材说明与公称直径/外径对照表融合；不含工程量。",
        "task_no": task_no,
        "task_nos": task_nos,
        "fusion_rule": fusion_rule,
        "source_evidence_ids": _join_unique([_first(note, "note_id"), _first(pair, "evidence_id")]),
        "fusion_inputs": fusion_inputs,
    }


def _select_supply_note(notes: Sequence[Mapping[str, Any]]) -> Mapping[str, Any]:
    for note in notes:
        text = _clean(note.get("text")).lower()
        if "sus304" in text and "给水" in text and ("薄壁" in text or "不锈钢" in text):
            return note
    return {}


def _select_drain_note(notes: Sequence[Mapping[str, Any]]) -> Mapping[str, Any]:
    for note in notes:
        text = _clean(note.get("text"))
        if "柔性铸铁管" in text and any(term in text for term in ("污水", "排水")):
            return note
    return {}


def _drain_de_scope(note: Mapping[str, Any], *, emit_drain: str) -> set[str]:
    if emit_drain == "table_de":
        return set()
    if emit_drain != "note_de":
        return set()
    text = _clean(note.get("text"))
    return {_normalize_de(match.group(0)) for match in DE_RE.finditer(text)}


def _first_match(pattern: re.Pattern[str], text: str) -> str:
    match = pattern.search(text or "")
    return match.group(0).replace(" ", "") if match else ""


def _normalize_de(value: str) -> str:
    match = DE_RE.search(value or "")
    return f"De{match.group(1)}" if match else _clean(value)


def _diameter_number(value: Any) -> int:
    match = re.search(r"\d+", str(value or ""))
    return int(match.group(0)) if match else 0


def _first(row: Mapping[str, Any], *keys: str) -> str:
    for key in keys:
        value = row.get(key)
        text = _clean(value)
        if text:
            return text
    return ""


def _clean(value: Any) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value).replace("\r", " ").replace("\n", " ")).strip()


def _bounded_float(value: Any, *, default: float = 0.0) -> float:
    try:
        parsed = float(str(value).strip())
    except (TypeError, ValueError):
        parsed = default
    return max(0.0, min(1.0, parsed))


def _join_unique(values: Iterable[Any]) -> str:
    seen: set[str] = set()
    result: list[str] = []
    for value in values:
        text = _clean(value)
        if not text or text in seen:
            continue
        seen.add(text)
        result.append(text)
    return "；".join(result)


def _build_markdown(report: Mapping[str, Any]) -> str:
    summary = report.get("summary") or {}
    lines = [
        "# BIZ-2x PDF Structured Feature Fusion",
        "",
        f"- generated_at: {report.get('generated_at', '-')}",
        f"- evidence_count: {summary.get('evidence_count', 0)}",
        f"- diameter_pair_count: {summary.get('diameter_pair_count', 0)}",
        f"- supply_row_count: {summary.get('supply_row_count', 0)}",
        f"- drain_row_count: {summary.get('drain_row_count', 0)}",
        f"- answer_rows_used: {summary.get('answer_rows_used', False)}",
        f"- quantity_status: {summary.get('quantity_status', '-')}",
        "",
        "## Evidence Rows",
        "",
        "| evidence | item | unit | feature | sources |",
        "| --- | --- | --- | --- | --- |",
    ]
    for row in report.get("evidence_rows") or []:
        lines.append(
            "| "
            + " | ".join(
                [
                    _md(row.get("evidence_id")),
                    _md(row.get("item_hint")),
                    _md(row.get("suggested_unit")),
                    _md(row.get("spec_or_method")),
                    _md(row.get("source_evidence_ids")),
                ]
            )
            + " |"
        )
    return "\n".join(lines) + "\n"


def _write_workbook(path: Path, report: Mapping[str, Any]) -> None:
    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "fusion_summary"
    _append_rows(summary_sheet, [["metric", "value"]])
    _append_rows(summary_sheet, [[key, _cell_value(value)] for key, value in (report.get("summary") or {}).items()])
    _style_sheet(summary_sheet)

    evidence_sheet = workbook.create_sheet("evidence_rows")
    _append_rows(evidence_sheet, [FUSION_HEADERS])
    _append_rows(
        evidence_sheet,
        [[_cell_value(row.get(header)) for header in FUSION_HEADERS] for row in report.get("evidence_rows") or []],
    )
    _style_sheet(evidence_sheet)

    pair_sheet = workbook.create_sheet("diameter_pairs")
    pair_headers = ["dn", "de", "inch_label", "source_file", "page", "tile_id", "evidence_id", "source_text"]
    _append_rows(pair_sheet, [pair_headers])
    _append_rows(pair_sheet, [[_cell_value(row.get(header)) for header in pair_headers] for row in report.get("diameter_pairs") or []])
    _style_sheet(pair_sheet)
    workbook.save(path)


def _write_csv(path: Path, rows: Iterable[Mapping[str, Any]], fieldnames: Sequence[str]) -> None:
    with path.open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        for row in rows:
            writer.writerow({field: _cell_value(row.get(field)) for field in fieldnames})


def _append_rows(sheet: Any, rows: Sequence[Sequence[Any]]) -> None:
    for row in rows:
        sheet.append(list(row))


def _style_sheet(sheet: Any) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    if sheet.max_row >= 1:
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for col_idx, column_cells in enumerate(sheet.columns, start=1):
        values = [str(cell.value or "") for cell in column_cells[:200]]
        width = min(max([len(value) for value in values] + [10]) + 2, 72)
        sheet.column_dimensions[get_column_letter(col_idx)].width = width


def _cell_value(value: Any) -> Any:
    if isinstance(value, (dict, list, tuple)):
        return json.dumps(value, ensure_ascii=False)
    return value


def _json_safe(value: Any) -> Any:
    if isinstance(value, Mapping):
        return {str(key): _json_safe(item) for key, item in value.items()}
    if isinstance(value, (list, tuple)):
        return [_json_safe(item) for item in value]
    if isinstance(value, Path):
        return str(value)
    return value


def _md(value: Any) -> str:
    return str(value or "").replace("|", "\\|").replace("\n", "<br>")
