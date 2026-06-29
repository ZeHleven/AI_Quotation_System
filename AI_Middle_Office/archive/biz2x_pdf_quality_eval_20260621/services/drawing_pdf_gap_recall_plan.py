from __future__ import annotations

import csv
import json
import re
from collections import Counter
from datetime import datetime
from pathlib import Path
from typing import Any, Iterable, Mapping, Sequence

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


PHASE = "BIZ-2x-pdf-gap-driven-visual-recall-plan"

PLAN_HEADERS = [
    "task_no",
    "gap_no",
    "gap_priority",
    "gap_type",
    "section",
    "answer_item_name",
    "answer_feature",
    "answer_unit",
    "recommended_pass",
    "secondary_passes",
    "recall_reason",
    "source_file",
    "page",
    "tile_id",
    "tile_type",
    "image_path",
    "evidence_id",
    "current_candidate_item_name",
    "current_candidate_unit",
    "suggested_next_action",
    "rerun_priority_score",
]


def build_gap_recall_plan(
    gap_pack: Mapping[str, Any],
    *,
    source_report_dir: str | Path | None = None,
    priority_prefixes: Sequence[str] | None = None,
    max_gaps: int | None = None,
) -> dict[str, Any]:
    prefixes = tuple(priority_prefixes or ("P1", "P2"))
    source_dir = Path(source_report_dir) if source_report_dir else None
    source_reports = _load_source_reports(gap_pack, source_report_dir=source_dir)
    gap_rows = _unique_gap_rows(gap_pack.get("manifest_rows") or [])

    selected_gaps: list[Mapping[str, Any]] = []
    for gap in gap_rows:
        priority = str(gap.get("gap_priority") or "")
        if prefixes and not priority.startswith(prefixes):
            continue
        selected_gaps.append(gap)
        if max_gaps is not None and len(selected_gaps) >= max_gaps:
            break

    plan_rows: list[dict[str, Any]] = []
    for gap in selected_gaps:
        pass_info = _assign_visual_pass(gap)
        tile = _select_recall_tile(gap, pass_info["recommended_pass"], source_reports)
        plan_rows.append(
            {
                "task_no": len(plan_rows) + 1,
                "gap_no": gap.get("gap_no", ""),
                "gap_priority": gap.get("gap_priority", ""),
                "gap_type": gap.get("gap_type", ""),
                "section": gap.get("section", ""),
                "answer_item_name": gap.get("answer_item_name", ""),
                "answer_feature": gap.get("answer_feature", ""),
                "answer_unit": gap.get("answer_unit", ""),
                "recommended_pass": pass_info["recommended_pass"],
                "secondary_passes": "；".join(pass_info["secondary_passes"]),
                "recall_reason": pass_info["reason"],
                "source_file": tile.get("source_file") or gap.get("source_file", ""),
                "page": tile.get("page") or gap.get("page", ""),
                "tile_id": tile.get("tile_id") or gap.get("tile_id", ""),
                "tile_type": tile.get("tile_type", ""),
                "image_path": tile.get("image_path") or gap.get("tile_image_path", ""),
                "evidence_id": gap.get("evidence_id", ""),
                "current_candidate_item_name": gap.get("candidate_item_name", ""),
                "current_candidate_unit": gap.get("candidate_unit", ""),
                "suggested_next_action": gap.get("suggested_next_action", ""),
                "rerun_priority_score": _priority_score(gap, pass_info, tile),
            }
        )

    plan_rows.sort(
        key=lambda row: (
            -int(row.get("rerun_priority_score") or 0),
            str(row.get("recommended_pass") or ""),
            int(_float(row.get("gap_no"), 0)),
        )
    )
    for index, row in enumerate(plan_rows, start=1):
        row["task_no"] = index

    return {
        "ok": True,
        "phase": PHASE,
        "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "summary": _build_summary(plan_rows, selected_gaps, prefixes),
        "plan_rows": plan_rows,
    }


def write_gap_recall_plan_outputs(
    plan: Mapping[str, Any],
    output_dir: str | Path,
    *,
    stem: str | None = None,
) -> dict[str, str]:
    target = Path(output_dir)
    target.mkdir(parents=True, exist_ok=True)
    file_stem = stem or f"BIZ2x_PDF_gap_recall_plan_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
    json_path = target / f"{file_stem}.json"
    csv_path = target / f"{file_stem}.csv"
    markdown_path = target / f"{file_stem}.md"
    xlsx_path = target / f"{file_stem}.xlsx"
    outputs = {
        "json": str(json_path),
        "csv": str(csv_path),
        "markdown": str(markdown_path),
        "xlsx": str(xlsx_path),
    }
    report = {**dict(plan), "outputs": outputs}
    json_path.write_text(json.dumps(_json_safe(report), ensure_ascii=False, indent=2), encoding="utf-8")
    _write_csv(csv_path, report.get("plan_rows") or [], PLAN_HEADERS)
    markdown_path.write_text(_build_markdown(report), encoding="utf-8")
    _write_workbook(xlsx_path, report)
    return outputs


def _unique_gap_rows(rows: Iterable[Mapping[str, Any]]) -> list[Mapping[str, Any]]:
    seen: set[Any] = set()
    result: list[Mapping[str, Any]] = []
    for row in rows:
        gap_no = row.get("gap_no")
        if gap_no in seen:
            continue
        seen.add(gap_no)
        result.append(row)
    return result


def _assign_visual_pass(gap: Mapping[str, Any]) -> dict[str, Any]:
    text = _normalize(" ".join([gap.get("section", ""), gap.get("answer_item_name", ""), gap.get("answer_feature", "")]))
    if any(term in text for term in ("阀门", "水表", "地漏", "台盆", "马桶", "花洒", "龙头", "洁具", "小便器", "浴缸", "洗涤盆")):
        return {
            "recommended_pass": "fixture_valve_schedule",
            "secondary_passes": ["plumbing_fixture", "table_legend"],
            "reason": "给排水洁具、阀门、水表、地漏缺口，需要逐行读取洁具表/阀门表/图例。",
        }
    if any(term in text for term in ("不锈钢玻璃门", "实木门", "铝合金门", "门套", "门扇", "门窗", "售卖窗口", "窗台石", "淋浴隔断", "台阶拆除", "洗手台拆除")):
        return {
            "recommended_pass": "door_window_demolition",
            "secondary_passes": ["demolition_node", "node_detail", "table_legend"],
            "reason": "门窗、洞口、售卖窗口或拆除对象缺口，需要从门窗表、拆除说明、节点详图召回。",
        }
    if "拆除" in text:
        return {
            "recommended_pass": "demolition_node",
            "secondary_passes": ["door_window_demolition", "node_detail"],
            "reason": "拆除对象不够具体，需要专门识别拆除说明和节点。",
        }
    if any(term in text for term in ("电缆", "配线", "配管", "配电箱", "灯具", "筒灯", "射灯", "格栅灯", "开关", "插座")):
        return {
            "recommended_pass": "electrical_mep",
            "secondary_passes": ["table_legend", "node_detail"],
            "reason": "电气项目缺少型号、回路、功率、规格或具体设备证据。",
        }
    if any(term in text for term in ("墙面", "地面", "天花", "天棚", "吊顶", "瓷砖", "石材", "墙布", "硬包", "防水", "灯槽", "窗帘盒")):
        return {
            "recommended_pass": "finish_schedule",
            "secondary_passes": ["node_detail", "table_legend"],
            "reason": "装饰面层或节点做法缺口，需要读取材料表、做法表和节点详图。",
        }
    return {
        "recommended_pass": "table_legend",
        "secondary_passes": ["node_detail", "general"],
        "reason": "缺口类别不明确，优先从表格/图例和节点做法补证据。",
    }


def _select_recall_tile(
    gap: Mapping[str, Any],
    recommended_pass: str,
    source_reports: Mapping[str, Mapping[str, Any]],
) -> dict[str, Any]:
    source_report_file = str(gap.get("source_report_file") or "")
    source_report = source_reports.get(source_report_file) or {}
    tile_rows = (source_report.get("tile_report") or {}).get("tile_rows") or []
    source_file = str(gap.get("source_file") or "")
    page = str(gap.get("page") or "")
    if recommended_pass in {"finish_schedule", "fixture_valve_schedule", "door_window_demolition", "table_legend", "node_detail"}:
        whole = _find_tile(tile_rows, source_file=source_file, page=page, tile_type="whole_page_preview")
        if whole:
            return dict(whole)
    exact = _find_tile(tile_rows, source_file=source_file, page=page, tile_id=str(gap.get("tile_id") or ""))
    if exact:
        return dict(exact)
    grid = _find_tile(tile_rows, source_file=source_file, page=page, tile_type="grid")
    if grid:
        return dict(grid)
    return {}


def _find_tile(
    tile_rows: Sequence[Mapping[str, Any]],
    *,
    source_file: str,
    page: str,
    tile_type: str | None = None,
    tile_id: str | None = None,
) -> Mapping[str, Any] | None:
    for row in tile_rows:
        if source_file and str(row.get("source_file") or "") != source_file:
            continue
        if page and str(row.get("page") or "") != page:
            continue
        if tile_type and str(row.get("tile_type") or "") != tile_type:
            continue
        if tile_id and str(row.get("tile_id") or "") != tile_id:
            continue
        return row
    return None


def _load_source_reports(
    gap_pack: Mapping[str, Any],
    *,
    source_report_dir: Path | None,
) -> dict[str, Mapping[str, Any]]:
    file_names = sorted(
        {
            str(row.get("source_report_file") or "")
            for row in gap_pack.get("manifest_rows") or []
            if isinstance(row, Mapping) and row.get("source_report_file")
        }
    )
    reports: dict[str, Mapping[str, Any]] = {}
    for file_name in file_names:
        path = _find_source_report_path(file_name, source_report_dir=source_report_dir)
        if not path:
            continue
        try:
            reports[file_name] = json.loads(path.read_text(encoding="utf-8"))
        except (OSError, json.JSONDecodeError):
            continue
    return reports


def _find_source_report_path(file_name: str, *, source_report_dir: Path | None) -> Path | None:
    if not file_name:
        return None
    path = Path(file_name)
    if path.exists():
        return path
    if source_report_dir:
        candidate = source_report_dir / file_name
        if candidate.exists():
            return candidate
    return None


def _priority_score(gap: Mapping[str, Any], pass_info: Mapping[str, Any], tile: Mapping[str, Any]) -> int:
    priority = str(gap.get("gap_priority") or "")
    score = 100 if priority.startswith("P1") else 60 if priority.startswith("P2") else 20
    if str(gap.get("gap_type") or "") == "missing_candidate":
        score += 20
    if pass_info.get("recommended_pass") in {"door_window_demolition", "fixture_valve_schedule"}:
        score += 15
    if tile.get("tile_type") == "whole_page_preview":
        score += 5
    if not tile:
        score -= 10
    return score


def _build_summary(
    plan_rows: Sequence[Mapping[str, Any]],
    selected_gaps: Sequence[Mapping[str, Any]],
    prefixes: Sequence[str],
) -> dict[str, Any]:
    return {
        "gap_count": len(selected_gaps),
        "plan_task_count": len(plan_rows),
        "priority_prefixes": list(prefixes),
        "pass_counts": dict(Counter(str(row.get("recommended_pass") or "") for row in plan_rows)),
        "priority_counts": dict(Counter(str(row.get("gap_priority") or "") for row in plan_rows)),
        "source_file_counts": dict(Counter(str(row.get("source_file") or "") for row in plan_rows)),
        "missing_image_task_count": sum(1 for row in plan_rows if not row.get("image_path")),
    }


def _build_markdown(report: Mapping[str, Any]) -> str:
    summary = report.get("summary") or {}
    lines = [
        "# BIZ-2x PDF Gap-Driven Visual Recall Plan",
        "",
        f"- generated_at: {report.get('generated_at', '-')}",
        f"- gaps: {summary.get('gap_count', 0)}",
        f"- tasks: {summary.get('plan_task_count', 0)}",
        f"- missing images: {summary.get('missing_image_task_count', 0)}",
        "",
        "## Pass Counts",
        "",
    ]
    for name, count in (summary.get("pass_counts") or {}).items():
        lines.append(f"- {name}: {count}")
    lines.extend(["", "## Top Tasks", "", "| priority | pass | item | source | tile | reason |", "| --- | --- | --- | --- | --- | --- |"])
    for row in (report.get("plan_rows") or [])[:80]:
        lines.append(
            "| "
            + " | ".join(
                [
                    _md(row.get("gap_priority")),
                    _md(row.get("recommended_pass")),
                    _md(row.get("answer_item_name")),
                    _md(row.get("source_file")),
                    _md(row.get("tile_id")),
                    _md(row.get("recall_reason")),
                ]
            )
            + " |"
        )
    return "\n".join(lines) + "\n"


def _write_workbook(path: Path, report: Mapping[str, Any]) -> None:
    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "recall_summary"
    _append_rows(summary_sheet, [["metric", "value"]])
    summary = report.get("summary") or {}
    for key, value in summary.items():
        _append_rows(summary_sheet, [[key, json.dumps(value, ensure_ascii=False) if isinstance(value, (dict, list)) else value]])
    _style_sheet(summary_sheet)

    plan_sheet = workbook.create_sheet("recall_plan")
    _append_rows(plan_sheet, [PLAN_HEADERS])
    _append_rows(plan_sheet, [[_cell_value(row.get(header)) for header in PLAN_HEADERS] for row in report.get("plan_rows") or []])
    _style_sheet(plan_sheet)
    workbook.save(path)


def _append_rows(sheet: Any, rows: Iterable[Sequence[Any]]) -> None:
    for row in rows:
        sheet.append(list(row))


def _style_sheet(sheet: Any) -> None:
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for col_idx, column_cells in enumerate(sheet.columns, start=1):
        values = [str(cell.value or "") for cell in column_cells[:80]]
        width = min(max([len(value) for value in values] + [10]) + 2, 80)
        sheet.column_dimensions[get_column_letter(col_idx)].width = width
    sheet.freeze_panes = "A2"


def _write_csv(path: Path, rows: Iterable[Mapping[str, Any]], fieldnames: Sequence[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        for row in rows:
            writer.writerow({key: _cell_value(row.get(key)) for key in fieldnames})


def _cell_value(value: Any) -> Any:
    if isinstance(value, (dict, list, tuple)):
        return json.dumps(value, ensure_ascii=False)
    return value


def _json_safe(value: Any) -> Any:
    if isinstance(value, Mapping):
        return {str(key): _json_safe(item) for key, item in value.items()}
    if isinstance(value, (list, tuple)):
        return [_json_safe(item) for item in value]
    if isinstance(value, Path):
        return str(value)
    return value


def _normalize(value: Any) -> str:
    return re.sub(r"\s+", "", str(value or "").lower())


def _md(value: Any) -> str:
    return str(value or "").replace("|", "\\|").replace("\n", "<br>")


def _float(value: Any, default: float = 0.0) -> float:
    try:
        return float(str(value).strip())
    except (TypeError, ValueError):
        return default
