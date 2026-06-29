from __future__ import annotations

import csv
import json
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any, Iterable, Mapping, Sequence

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.services.drawing_pdf_gap_recall_runner import CALL_HEADERS, EVIDENCE_HEADERS, run_gap_recall_plan


PHASE = "BIZ-2x-pdf-gap-recall-external-import"

VALIDATION_HEADERS = ["row_no", "status", "evidence_id", "source_file", "page", "tile_id", "vision_pass", "issue"]


def load_external_recall_results(path: str | Path) -> dict[str, Any]:
    source = Path(path)
    if source.suffix.lower() == ".csv":
        with source.open("r", encoding="utf-8-sig", newline="") as handle:
            return {"evidence_rows": list(csv.DictReader(handle))}
    if source.suffix.lower() in {".xlsx", ".xlsm"}:
        return {"evidence_rows": _load_workbook_rows(source)}
    return json.loads(source.read_text(encoding="utf-8"))


def build_gap_recall_external_import_report(
    external_results: Mapping[str, Any],
    *,
    recall_plan: Mapping[str, Any] | None = None,
    source_name: str = "",
) -> dict[str, Any]:
    dry_run = run_gap_recall_plan(recall_plan or {"plan_rows": []}, execute=False)
    call_rows = [dict(row) for row in dry_run.get("call_rows") or []]
    call_lookup = _call_lookup(call_rows)
    imported_sources = _external_source_rows(external_results)

    evidence_rows: list[dict[str, Any]] = []
    validation_rows: list[dict[str, Any]] = []
    evidence_count_by_call: dict[int, int] = defaultdict(int)
    unassigned_count = 0

    for row_no, source_row in enumerate(imported_sources, start=1):
        call = _match_call(source_row, call_lookup)
        evidence = _normalize_evidence_row(
            source_row,
            call=call,
            evidence_no=len(evidence_rows) + 1,
            source_name=source_name,
        )
        if not evidence:
            validation_rows.append(
                {
                    "row_no": row_no,
                    "status": "skipped",
                    "evidence_id": "",
                    "source_file": _first(source_row, "source_file", "candidate_source_files", "PDF文件"),
                    "page": _first(source_row, "page", "evidence_pages", "页码"),
                    "tile_id": _first(source_row, "tile_id", "source_tile_id", "evidence_tiles"),
                    "vision_pass": _first(source_row, "vision_pass", "recommended_pass", "prompt_mode"),
                    "issue": "missing evidence text/item/spec",
                }
            )
            continue
        evidence_rows.append(evidence)
        call_no = _int((call or {}).get("call_no"), 0)
        if call_no:
            evidence_count_by_call[call_no] += 1
            status = "imported"
            issue = ""
        else:
            unassigned_count += 1
            status = "imported_unmatched_call"
            issue = "no matching recall-plan call; evidence kept for downstream review"
        validation_rows.append(
            {
                "row_no": row_no,
                "status": status,
                "evidence_id": evidence.get("evidence_id", ""),
                "source_file": evidence.get("source_file", ""),
                "page": evidence.get("page", ""),
                "tile_id": evidence.get("tile_id", ""),
                "vision_pass": evidence.get("vision_pass", ""),
                "issue": issue,
            }
        )

    for call_row in call_rows:
        call_no = _int(call_row.get("call_no"), 0)
        count = evidence_count_by_call.get(call_no, 0)
        call_row["status"] = "external_imported" if count else "external_no_evidence"
        call_row["reason"] = f"imported_{count}_evidence_rows"
        call_row["raw_content"] = source_name

    summary = {
        "external_import": True,
        "source_name": source_name,
        "plan_task_count": len((recall_plan or {}).get("plan_rows") or []),
        "unique_visual_call_count": len(call_rows),
        "call_count": len(call_rows),
        "input_row_count": len(imported_sources),
        "evidence_count": len(evidence_rows),
        "skipped_input_row_count": sum(1 for row in validation_rows if row.get("status") == "skipped"),
        "unassigned_evidence_count": unassigned_count,
        "status_counts": dict(Counter(str(row.get("status") or "") for row in call_rows)),
        "validation_status_counts": dict(Counter(str(row.get("status") or "") for row in validation_rows)),
        "pass_counts": dict(Counter(str(row.get("vision_pass") or "") for row in evidence_rows)),
    }
    return {
        "ok": True,
        "phase": PHASE,
        "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "summary": summary,
        "call_rows": call_rows,
        "evidence_rows": evidence_rows,
        "validation_rows": validation_rows,
        "visual_evidence_report": {"evidence_rows": evidence_rows},
    }


def write_gap_recall_external_import_outputs(
    report: Mapping[str, Any],
    output_dir: str | Path,
    *,
    stem: str | None = None,
) -> dict[str, str]:
    target = Path(output_dir)
    target.mkdir(parents=True, exist_ok=True)
    file_stem = stem or f"BIZ2x_PDF_gap_recall_external_import_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
    json_path = target / f"{file_stem}.json"
    call_csv = target / f"{file_stem}_calls.csv"
    evidence_csv = target / f"{file_stem}_evidence.csv"
    validation_csv = target / f"{file_stem}_validation.csv"
    markdown_path = target / f"{file_stem}.md"
    xlsx_path = target / f"{file_stem}.xlsx"
    outputs = {
        "json": str(json_path),
        "call_csv": str(call_csv),
        "evidence_csv": str(evidence_csv),
        "validation_csv": str(validation_csv),
        "markdown": str(markdown_path),
        "xlsx": str(xlsx_path),
    }
    payload = {**dict(report), "outputs": outputs}
    json_path.write_text(json.dumps(_json_safe(payload), ensure_ascii=False, indent=2), encoding="utf-8")
    _write_csv(call_csv, report.get("call_rows") or [], CALL_HEADERS)
    _write_csv(evidence_csv, report.get("evidence_rows") or [], EVIDENCE_HEADERS)
    _write_csv(validation_csv, report.get("validation_rows") or [], VALIDATION_HEADERS)
    markdown_path.write_text(_build_markdown(payload), encoding="utf-8")
    _write_workbook(xlsx_path, payload)
    return outputs


def _load_workbook_rows(path: Path) -> list[dict[str, Any]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet_name = _select_external_recall_sheet(workbook.sheetnames)
        sheet = workbook[sheet_name]
        row_iter = sheet.iter_rows(values_only=True)
        try:
            header_row = next(row_iter)
        except StopIteration:
            return []
        headers = [str(value or "").strip() for value in header_row]
        rows: list[dict[str, Any]] = []
        for values in row_iter:
            row = {
                header: _cell_value_for_import(value)
                for header, value in zip(headers, values)
                if header
            }
            if any(str(value or "").strip() for value in row.values()):
                rows.append(row)
        return rows
    finally:
        workbook.close()


def _select_external_recall_sheet(sheet_names: Sequence[str]) -> str:
    preferred = (
        "external_recall_template",
        "object_recall_workbench",
        "object_recall_tasks",
        "recall_evidence",
        "evidence_rows",
    )
    for name in preferred:
        if name in sheet_names:
            return name
    return sheet_names[0]


def _external_source_rows(external_results: Mapping[str, Any]) -> list[dict[str, Any]]:
    for key in ("evidence_rows", "recall_evidence", "rows"):
        rows = external_results.get(key)
        if isinstance(rows, list):
            return [dict(row) for row in rows if isinstance(row, Mapping)]
    visual_report = external_results.get("visual_evidence_report")
    if isinstance(visual_report, Mapping) and isinstance(visual_report.get("evidence_rows"), list):
        return [dict(row) for row in visual_report.get("evidence_rows") or [] if isinstance(row, Mapping)]

    source_rows: list[dict[str, Any]] = []
    call_groups = []
    for key in ("call_results", "calls", "results"):
        value = external_results.get(key)
        if isinstance(value, list):
            call_groups.extend(row for row in value if isinstance(row, Mapping))
    if not call_groups and isinstance(external_results.get("evidence_items"), list):
        call_groups.append(external_results)

    for call in call_groups:
        items = (
            call.get("evidence_rows")
            or call.get("evidence_items")
            or call.get("items")
            or call.get("drawing_items")
            or []
        )
        if not isinstance(items, list):
            continue
        call_meta = {
            "call_no": call.get("call_no"),
            "source_file": _first(call, "source_file", "PDF文件"),
            "page": _first(call, "page", "页码"),
            "tile_id": call.get("tile_id"),
            "vision_pass": _first(call, "vision_pass", "recommended_pass", "prompt_mode"),
            "model": call.get("model"),
        }
        for item in items:
            if isinstance(item, Mapping):
                source_rows.append({**call_meta, **dict(item)})
    return source_rows


def _normalize_evidence_row(
    row: Mapping[str, Any],
    *,
    call: Mapping[str, Any] | None,
    evidence_no: int,
    source_name: str,
) -> dict[str, Any] | None:
    item_hint = _first(
        row,
        "item_hint",
        "evidence_item_hint",
        "raw_item_name",
        "item_name",
        "项目名称",
        "清单项目名称",
    )
    spec = _first(row, "spec_or_method", "evidence_spec_or_method", "feature", "项目特征", "规格/做法", "spec")
    text = _first(row, "text", "evidence_text", "normalized_text", "识别依据", "reason")
    if not text:
        text = " ".join(part for part in (item_hint, spec) if part).strip()
    if not any((text, item_hint, spec)):
        return None
    call = call or {}
    return {
        "evidence_id": _first(row, "evidence_id") or f"PDFGAP-EXT-{evidence_no:06d}",
        "source_kind": _first(row, "source_kind") or "pdf_gap_recall_external_import",
        "source_file": _first(row, "source_file", "candidate_source_files", "PDF文件") or call.get("source_file", ""),
        "page": _first(row, "page", "evidence_pages", "页码") or call.get("page", ""),
        "tile_id": _first(row, "tile_id", "source_tile_id", "evidence_tiles") or call.get("tile_id", ""),
        "vision_pass": _first(row, "vision_pass", "recommended_pass", "prompt_mode") or call.get("recommended_pass", ""),
        "evidence_role": _first(row, "evidence_role", "role") or "external_recall_note",
        "discipline": _first(row, "discipline", "专业") or "unknown",
        "item_hint": item_hint,
        "space": _first(row, "space", "部位", "空间/部位"),
        "material_codes": _normalize_string_list(row.get("material_codes") or row.get("材料编号")),
        "spec_or_method": spec,
        "suggested_unit": _first(row, "suggested_unit", "evidence_suggested_unit", "unit", "单位", "计量单位"),
        "text": text,
        "normalized_text": _first(row, "normalized_text") or text,
        "confidence": _bounded_float(_first(row, "confidence", "置信度"), default=0.5),
        "model": _first(row, "model") or source_name,
        "needs_manual_review": _boolish(_first(row, "needs_manual_review", "需要人工复核"), default=True),
        "reason": _first(row, "reason", "识别理由"),
    }


def _call_lookup(call_rows: Sequence[Mapping[str, Any]]) -> dict[str, Any]:
    by_no: dict[int, Mapping[str, Any]] = {}
    by_task_no: dict[int, Mapping[str, Any]] = {}
    by_key: dict[tuple[str, str, str, str], Mapping[str, Any]] = {}
    for row in call_rows:
        call_no = _int(row.get("call_no"), 0)
        if call_no:
            by_no[call_no] = row
        for task_no in _split_ints(row.get("covered_task_nos")):
            by_task_no.setdefault(task_no, row)
        by_key[_call_key(row)] = row
    return {"by_no": by_no, "by_task_no": by_task_no, "by_key": by_key}


def _match_call(row: Mapping[str, Any], lookup: Mapping[str, Any]) -> Mapping[str, Any] | None:
    call_no = _int(row.get("call_no"), 0)
    if call_no and call_no in lookup["by_no"]:
        return lookup["by_no"][call_no]
    key = (
        str(_first(row, "vision_pass", "recommended_pass", "prompt_mode") or ""),
        str(_first(row, "source_file", "candidate_source_files", "PDF文件") or ""),
        str(_first(row, "page", "evidence_pages", "页码") or ""),
        str(_first(row, "tile_id", "source_tile_id", "evidence_tiles") or ""),
    )
    call = lookup["by_key"].get(key) or lookup["by_key"].get(_call_key(row))
    if call:
        return call
    task_no = _int(row.get("task_no"), 0)
    if task_no and task_no in lookup["by_task_no"]:
        return lookup["by_task_no"][task_no]
    return None


def _call_key(row: Mapping[str, Any]) -> tuple[str, str, str, str]:
    return (
        _compact_key(_first(row, "recommended_pass", "vision_pass", "prompt_mode")),
        _source_file_key(_first(row, "source_file", "candidate_source_files")),
        _compact_key(_first(row, "page", "evidence_pages")),
        _compact_key(_first(row, "tile_id", "source_tile_id", "evidence_tiles")),
    )


def _build_markdown(report: Mapping[str, Any]) -> str:
    summary = report.get("summary") or {}
    lines = [
        "# BIZ-2x PDF Gap Recall External Import",
        "",
        f"- generated_at: {report.get('generated_at', '-')}",
        f"- source_name: {summary.get('source_name', '-')}",
        f"- input_row_count: {summary.get('input_row_count', 0)}",
        f"- evidence_count: {summary.get('evidence_count', 0)}",
        f"- skipped_input_row_count: {summary.get('skipped_input_row_count', 0)}",
        f"- unassigned_evidence_count: {summary.get('unassigned_evidence_count', 0)}",
        "",
        "## Validation",
        "",
        "| row | status | evidence | source | page | tile | pass | issue |",
        "| ---: | --- | --- | --- | --- | --- | --- | --- |",
    ]
    for row in report.get("validation_rows") or []:
        lines.append(
            "| "
            + " | ".join(
                [
                    _md(row.get("row_no")),
                    _md(row.get("status")),
                    _md(row.get("evidence_id")),
                    _md(row.get("source_file")),
                    _md(row.get("page")),
                    _md(row.get("tile_id")),
                    _md(row.get("vision_pass")),
                    _md(row.get("issue")),
                ]
            )
            + " |"
        )
    return "\n".join(lines) + "\n"


def _write_workbook(path: Path, report: Mapping[str, Any]) -> None:
    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "import_summary"
    _append_rows(summary_sheet, [["metric", "value"]])
    _append_rows(summary_sheet, [[key, _cell_value(value)] for key, value in (report.get("summary") or {}).items()])
    _style_sheet(summary_sheet)

    call_sheet = workbook.create_sheet("visual_calls")
    _append_rows(call_sheet, [CALL_HEADERS])
    _append_rows(call_sheet, [[_cell_value(row.get(header)) for header in CALL_HEADERS] for row in report.get("call_rows") or []])
    _style_sheet(call_sheet)

    evidence_sheet = workbook.create_sheet("recall_evidence")
    _append_rows(evidence_sheet, [EVIDENCE_HEADERS])
    _append_rows(evidence_sheet, [[_cell_value(row.get(header)) for header in EVIDENCE_HEADERS] for row in report.get("evidence_rows") or []])
    _style_sheet(evidence_sheet)

    validation_sheet = workbook.create_sheet("validation")
    _append_rows(validation_sheet, [VALIDATION_HEADERS])
    _append_rows(validation_sheet, [[_cell_value(row.get(header)) for header in VALIDATION_HEADERS] for row in report.get("validation_rows") or []])
    _style_sheet(validation_sheet)
    workbook.save(path)


def _write_csv(path: Path, rows: Iterable[Mapping[str, Any]], fieldnames: Sequence[str]) -> None:
    with path.open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        for row in rows:
            writer.writerow({key: _cell_value(row.get(key)) for key in fieldnames})


def _append_rows(sheet: Any, rows: Sequence[Sequence[Any]]) -> None:
    for row in rows:
        sheet.append(list(row))


def _style_sheet(sheet: Any) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    if sheet.max_row >= 1:
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for col_idx, column_cells in enumerate(sheet.columns, start=1):
        values = [str(cell.value or "") for cell in column_cells[:200]]
        width = min(max([len(value) for value in values] + [10]) + 2, 60)
        sheet.column_dimensions[get_column_letter(col_idx)].width = width


def _normalize_string_list(value: Any) -> list[str]:
    if value is None:
        return []
    if isinstance(value, str):
        text = value.strip()
        if text.startswith("[") and text.endswith("]"):
            try:
                parsed = json.loads(text)
                if isinstance(parsed, list):
                    return [str(item).strip() for item in parsed if str(item).strip()]
            except json.JSONDecodeError:
                pass
        return [item.strip() for item in text.replace(";", ",").split(",") if item.strip()]
    if isinstance(value, (list, tuple, set)):
        return [str(item).strip() for item in value if str(item).strip()]
    return [str(value).strip()] if str(value).strip() else []


def _bounded_float(value: Any, *, default: float = 0.0) -> float:
    try:
        parsed = float(str(value).strip())
    except (TypeError, ValueError):
        parsed = default
    return max(0.0, min(1.0, parsed))


def _boolish(value: Any, *, default: bool = False) -> bool:
    if value is None or value == "":
        return default
    if isinstance(value, bool):
        return value
    return str(value).strip().lower() in {"1", "true", "yes", "y", "是", "需要"}


def _first(row: Mapping[str, Any], *keys: str) -> str:
    for key in keys:
        value = row.get(key)
        if value is None:
            continue
        text = str(value).strip()
        if text:
            return text
    return ""


def _int(value: Any, default: int = 0) -> int:
    try:
        return int(float(str(value).strip()))
    except (TypeError, ValueError):
        return default


def _compact_key(value: Any) -> str:
    return str(value or "").strip().lower()


def _source_file_key(value: Any) -> str:
    text = str(value or "").strip().replace("\\", "/")
    if not text:
        return ""
    first_value = text.replace("；", ";").replace("，", ",").split(";")[0].split(",")[0].strip()
    return Path(first_value).name.lower() or first_value.lower()


def _split_ints(value: Any) -> list[int]:
    text = str(value or "").strip()
    if not text:
        return []
    items: list[int] = []
    for part in text.replace("；", ",").replace(";", ",").split(","):
        number = _int(part, 0)
        if number:
            items.append(number)
    return items


def _cell_value(value: Any) -> Any:
    if isinstance(value, (dict, list, tuple)):
        return json.dumps(value, ensure_ascii=False)
    return value


def _cell_value_for_import(value: Any) -> Any:
    if value is None:
        return ""
    return value


def _json_safe(value: Any) -> Any:
    if isinstance(value, Mapping):
        return {str(key): _json_safe(item) for key, item in value.items()}
    if isinstance(value, (list, tuple)):
        return [_json_safe(item) for item in value]
    if isinstance(value, Path):
        return str(value)
    return value


def _md(value: Any) -> str:
    return str(value or "").replace("|", "\\|").replace("\n", "<br>")
