from __future__ import annotations

import argparse
import json
from pathlib import Path
from typing import Any, Iterable, Mapping

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


BACKEND_ROOT = Path(__file__).resolve().parents[1]
REPO_ROOT = BACKEND_ROOT.parent
DEFAULT_REPORT_DIR = REPO_ROOT / "reports" / "biz2t" / "20260528_current"
DEFAULT_RAW_JSON = DEFAULT_REPORT_DIR / "cost_governance_raw.json"


RISK_LABELS = {
    "high": "高风险：试运行前必须处理或说明",
    "medium": "中风险：试运行前建议复核",
    "low": "低风险：可在试运行中观察",
}

ISSUE_TYPE_LABELS = {
    "missing_named_reference_price": "缺少业务可解释来源价",
    "same_name_multi_spec": "同名多规格，需要确认规格是否清晰",
    "missing_spec_on_multi_name": "同名多规格中规格缺失",
    "missing_spec": "规格/特征缺失或不够清晰",
    "unit_needs_review": "单位需要人工确认",
    "similar_active_items": "相似 active 条目，需要确认是否合理共存",
    "exact_active_duplicate": "重复 active 条目",
    "invalid_main_price": "主参考价为空或为 0",
    "missing_unit": "单位缺失",
    "same_name_mixed_units": "同名条目单位不一致",
    "rag_sync_latest_not_success": "RAG 最近同步未成功",
    "rag_sync_count_mismatch": "RAG 同步数量不一致",
    "rag_sync_not_checked": "未检查 RAG 同步状态",
    "draft_missing_price_or_unit": "draft 缺少价格或单位",
    "ai_suggested_draft_review": "AI 建议 draft 需要人工审核",
    "draft_duplicate_with_active": "draft 与 active 疑似重复",
}

STATUS_LABELS = {
    "active": "active：已启用，参与报价",
    "draft": "draft：草稿，不参与报价",
    "archived": "archived：已归档，不参与报价",
}

SOURCE_LABELS = {
    "imported": "导入",
    "manual": "人工",
    "ai_suggested": "AI 建议",
}

ACTION_HEADERS = [
    "问题编号",
    "风险等级",
    "是否阻断试运行",
    "问题类型",
    "成本条目ID",
    "施工项目",
    "规格/特征",
    "单位",
    "当前主参考价",
    "状态",
    "来源",
    "报价引用次数",
    "最近引用时间",
    "为什么提示",
    "建议处理",
    "负责人",
    "处理结论",
    "备注",
]


def _load_pack(raw_json_path: Path) -> dict[str, Any]:
    return json.loads(raw_json_path.read_text(encoding="utf-8"))


def _map_issue_type(value: Any) -> str:
    parts = [part.strip() for part in str(value or "").split(";") if part.strip()]
    if not parts:
        return ""
    return "；".join(ISSUE_TYPE_LABELS.get(part, part) for part in parts)


def _yes_no(value: Any) -> str:
    return "是" if str(value or "").lower() == "yes" else "否"


def _risk_label(value: Any) -> str:
    return RISK_LABELS.get(str(value or ""), str(value or ""))


def _status_label(value: Any) -> str:
    return STATUS_LABELS.get(str(value or ""), str(value or ""))


def _source_label(value: Any) -> str:
    return SOURCE_LABELS.get(str(value or ""), str(value or ""))


def _explain_action(action: Mapping[str, Any]) -> str:
    issue_type = str(action.get("issue_type") or "")
    explanations: list[str] = []
    if "missing_named_reference_price" in issue_type:
        explanations.append("该 active 条目已有主参考价，但缺少对甲价、劳务价或班组价等可解释来源价。")
    if "same_name_multi_spec" in issue_type:
        explanations.append("同名条目存在多个规格，需确认规格文字能否让业务员正确选择。")
    if "missing_spec" in issue_type:
        explanations.append("规格/特征不够清楚，后续报价依据解释可能不足。")
    if "unit_needs_review" in issue_type or "same_name_mixed_units" in issue_type:
        explanations.append("单位写法或单位口径需要人工确认，避免报价数量乘错。")
    if "similar_active_items" in issue_type or "exact_active_duplicate" in issue_type:
        explanations.append("存在相似或重复 active 条目，需要确认是否合理共存。")
    if "rag_sync" in issue_type:
        explanations.append("RAG 同步状态可能影响 AI 检索到的成本知识是否最新。")
    if "draft" in issue_type:
        explanations.append("该 draft 不参与报价，但需要人工判断后续是否保留、补全或归档。")
    quote_usage = int(action.get("quote_usage_count") or 0)
    if quote_usage:
        explanations.append(f"该条目已被报价引用 {quote_usage} 次，所以风险等级会提高。")
    return " ".join(explanations) or "系统按成本库体检规则提示该条需要人工复核。"


def _action_to_business_row(action: Mapping[str, Any]) -> list[Any]:
    return [
        action.get("issue_id", ""),
        _risk_label(action.get("risk_level")),
        _yes_no(action.get("trial_blocker")),
        _map_issue_type(action.get("issue_type")),
        action.get("cost_item_id", ""),
        action.get("item_name", ""),
        action.get("spec", ""),
        action.get("unit", ""),
        action.get("price", ""),
        _status_label(action.get("status")),
        _source_label(action.get("source")),
        action.get("quote_usage_count", 0),
        action.get("latest_quote_used_at", ""),
        _explain_action(action),
        action.get("suggested_action", ""),
        action.get("owner", ""),
        action.get("done", ""),
        action.get("notes", ""),
    ]


def _actions_by_risk(pack: Mapping[str, Any], risk: str) -> list[Mapping[str, Any]]:
    return [action for action in pack.get("actions", []) if action.get("risk_level") == risk]


def _append_key_value_rows(sheet, rows: Iterable[tuple[Any, Any]]) -> None:
    for key, value in rows:
        sheet.append([key, value])


def _style_sheet(sheet, *, freeze: str | None = "A2") -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    if freeze:
        sheet.freeze_panes = freeze
    if sheet.max_row >= 1 and sheet.max_column >= 1:
        sheet.auto_filter.ref = sheet.dimensions
    for column_cells in sheet.columns:
        letter = get_column_letter(column_cells[0].column)
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        sheet.column_dimensions[letter].width = min(max(max_length + 2, 12), 42)
    sheet.row_dimensions[1].height = 28


def _write_actions_sheet(workbook: Workbook, title: str, actions: list[Mapping[str, Any]]) -> None:
    sheet = workbook.create_sheet(title)
    sheet.append(ACTION_HEADERS)
    for action in actions:
        sheet.append(_action_to_business_row(action))
    _style_sheet(sheet)
    for col in ("F", "G", "N", "O", "R"):
        sheet.column_dimensions[col].width = 42


def write_business_xlsx(pack: Mapping[str, Any], output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    summary = pack.get("summary") or {}
    readiness = pack.get("trial_readiness") or {}
    status_counts = summary.get("status_counts") or {}
    source_counts = summary.get("source_counts") or {}
    risk_counts = summary.get("risk_counts") or {}
    sync = summary.get("latest_rag_sync") or {}

    guide = workbook.active
    guide.title = "验收说明"
    guide.append(["验收问题", "在哪里看 / 怎么验"])
    guide_rows = [
        ("风险分类在哪里看", "打开本 Excel 的“高风险”“中风险”“低风险”三个 Sheet。高风险是试运行前必须处理或说明的项；中风险是试运行前建议复核的项；低风险可以在试运行中观察。"),
        ("报告可读性怎么验", "先看“总览”Sheet，再看风险 Sheet。中文字段、自动换行和较宽列宽应能让项目名、规格、原因、建议处理看完整。"),
        ("只读边界怎么验", "本业务可读版只读取 cost_governance_raw.json 生成，不连接数据库。严格复核时，可在生成前后对比成本库总数、active、archived、draft 数量，应保持不变。"),
        ("不影响报价怎么验", "本报告位于 reports 目录，不被报价接口读取。打开或重新生成报告后，普通报价流程不应新增报价行、不改总价、不改变预审逻辑。"),
        ("先看哪张表", "先看“总览”，再看“高风险”。当前 5 条高风险已登记为 accepted_risk，可结合 BIZ-2t-2 复核报告一起看。"),
    ]
    for row in guide_rows:
        guide.append(list(row))
    _style_sheet(guide)
    guide.column_dimensions["A"].width = 22
    guide.column_dimensions["B"].width = 90

    overview = workbook.create_sheet("总览")
    overview.append(["指标", "当前值"])
    _append_key_value_rows(
        overview,
        [
            ("生成时间", pack.get("generated_at", "")),
            ("成本条目总数", summary.get("total_count", 0)),
            ("active 条目数", status_counts.get("active", 0)),
            ("archived 条目数", status_counts.get("archived", 0)),
            ("draft 条目数", status_counts.get("draft", 0)),
            ("导入来源", source_counts.get("imported", 0)),
            ("人工来源", source_counts.get("manual", 0)),
            ("AI 建议来源", source_counts.get("ai_suggested", 0)),
            ("被报价引用过的 active", summary.get("quote_used_active_count", 0)),
            ("治理动作总数", summary.get("action_count", 0)),
            ("高风险", risk_counts.get("high", 0)),
            ("中风险", risk_counts.get("medium", 0)),
            ("低风险", risk_counts.get("low", 0)),
            ("试运行阻断项", summary.get("trial_blocker_count", 0)),
            ("试运行建议", readiness.get("recommendation", "")),
            ("最近 RAG 同步状态", sync.get("status", "")),
            ("最近 RAG 同步数量", f"{sync.get('synced_count', 0)} / {sync.get('requested_count', 0)}"),
            ("最近 RAG 同步时间", sync.get("started_at", "")),
        ],
    )
    _style_sheet(overview, freeze=None)
    overview.column_dimensions["A"].width = 26
    overview.column_dimensions["B"].width = 70

    _write_actions_sheet(workbook, "高风险", _actions_by_risk(pack, "high"))
    _write_actions_sheet(workbook, "中风险", _actions_by_risk(pack, "medium"))
    _write_actions_sheet(workbook, "低风险", _actions_by_risk(pack, "low"))
    _write_actions_sheet(workbook, "全部明细", list(pack.get("actions", [])))

    workbook.save(output_path)


def _markdown_table(rows: list[Mapping[str, Any]]) -> list[str]:
    lines = [
        "| 问题编号 | 风险 | 成本ID | 施工项目 | 单位 | 当前价 | 报价引用 | 为什么提示 | 建议处理 |",
        "|---|---|---:|---|---|---:|---:|---|---|",
    ]
    for action in rows:
        lines.append(
            "| {issue_id} | {risk} | {cost_id} | {name} | {unit} | {price} | {usage} | {reason} | {suggestion} |".format(
                issue_id=action.get("issue_id", ""),
                risk=_risk_label(action.get("risk_level")),
                cost_id=action.get("cost_item_id", ""),
                name=str(action.get("item_name", "")).replace("|", "/"),
                unit=action.get("unit", ""),
                price=action.get("price", ""),
                usage=action.get("quote_usage_count", 0),
                reason=_explain_action(action).replace("|", "/"),
                suggestion=str(action.get("suggested_action", "")).replace("|", "/"),
            )
        )
    return lines


def write_business_markdown(pack: Mapping[str, Any], output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    summary = pack.get("summary") or {}
    readiness = pack.get("trial_readiness") or {}
    status_counts = summary.get("status_counts") or {}
    risk_counts = summary.get("risk_counts") or {}
    sync = summary.get("latest_rag_sync") or {}
    high_actions = _actions_by_risk(pack, "high")
    medium_actions = _actions_by_risk(pack, "medium")
    low_actions = _actions_by_risk(pack, "low")

    lines = [
        "# BIZ-2k-1 成本库治理业务可读报告",
        "",
        "> 来源：`cost_governance_raw.json`。本报告只读生成，不连接数据库，不修改成本库，不影响报价流程。",
        "",
        "## 1. 先看哪里",
        "",
        "- 打开 `cost_governance_business_readable.xlsx`。",
        "- 先看 `总览` Sheet，确认成本库总数和风险数量。",
        "- 再看 `高风险` Sheet，确认试运行前必须处理或说明的项目。",
        "- `中风险` Sheet 用于试运行前复核；`低风险` Sheet 用于试运行中观察。",
        "",
        "## 2. 当前结论",
        "",
        f"- 成本条目总数：{summary.get('total_count', 0)}",
        f"- active：{status_counts.get('active', 0)}",
        f"- archived：{status_counts.get('archived', 0)}",
        f"- draft：{status_counts.get('draft', 0)}",
        f"- 治理动作总数：{summary.get('action_count', 0)}",
        f"- 高风险：{risk_counts.get('high', 0)}",
        f"- 中风险：{risk_counts.get('medium', 0)}",
        f"- 低风险：{risk_counts.get('low', 0)}",
        f"- 试运行阻断项：{summary.get('trial_blocker_count', 0)}",
        f"- 试运行建议：`{readiness.get('recommendation', '')}`",
        f"- 最近 RAG 同步：{sync.get('status', '')}，{sync.get('synced_count', 0)} / {sync.get('requested_count', 0)}",
        "",
        "## 3. 风险分类怎么理解",
        "",
        "| 风险 | 含义 | 验收口径 |",
        "|---|---|---|",
        "| 高风险 | 试运行前必须处理或写明接受风险 | 看 `高风险` Sheet，每条都要有处理结论或已知风险说明 |",
        "| 中风险 | 试运行前建议复核 | 看 `中风险` Sheet，确认不会影响首批样例 |",
        "| 低风险 | 可试运行中观察 | 看 `低风险` Sheet，不要求一次性处理完 |",
        "",
        "## 4. 高风险明细",
        "",
        *_markdown_table(high_actions),
        "",
        "## 5. 中低风险看法",
        "",
        f"- 中风险共 {len(medium_actions)} 条，主要看规格是否清楚、单位是否合理、相似 active 是否合理共存。",
        f"- 低风险共 {len(low_actions)} 条，主要作为后续标准化优化清单，不阻断当前功能验收。",
        "",
        "## 6. 只读边界怎么验",
        "",
        "1. 打开后台成本数据库，记录当前总数、active、archived、draft 数量。",
        "2. 打开或重新生成本业务可读报告。",
        "3. 再次查看成本数据库数量。",
        "4. 若数量和状态没有变化，即可确认报告生成不新增、不删除、不改价、不改状态。",
        "",
        "本次 BIZ-2k-1 脚本只读取 `cost_governance_raw.json`，连数据库查询都不做，因此不会写库。",
        "",
        "## 7. 不影响报价怎么验",
        "",
        "1. 打开一条普通报价任务或发起一个小型测试报价。",
        "2. 只打开本报告或重新生成本报告。",
        "3. 回到报价预审，确认没有新增报价行、没有自动改总价、没有改变成本库匹配依据。",
        "",
        "本报告位于 `reports/` 目录，不被报价 API、预审逻辑、N8N/Dify 或 RAG 同步读取。",
        "",
        "## 8. 本次补强解决的验收问题",
        "",
        "- 字段改成中文。",
        "- Excel 分成总览、高风险、中风险、低风险和全部明细。",
        "- 长文本自动换行，列宽放大。",
        "- 明确写出风险分类在哪里看。",
        "- 明确写出只读边界和不影响报价的验收方法。",
        "",
    ]
    output_path.write_text("\n".join(lines), encoding="utf-8")


def _parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Generate BIZ-2k-1 business-readable governance acceptance pack.")
    parser.add_argument("--raw-json", default=str(DEFAULT_RAW_JSON), help="Path to cost_governance_raw.json.")
    parser.add_argument("--output-dir", default=str(DEFAULT_REPORT_DIR), help="Output directory.")
    return parser.parse_args()


def main() -> int:
    args = _parse_args()
    raw_json_path = Path(args.raw_json).resolve()
    output_dir = Path(args.output_dir).resolve()
    pack = _load_pack(raw_json_path)

    markdown_path = output_dir / "cost_governance_business_readable.md"
    xlsx_path = output_dir / "cost_governance_business_readable.xlsx"
    write_business_markdown(pack, markdown_path)
    write_business_xlsx(pack, xlsx_path)

    print(
        json.dumps(
            {
                "status": "ok",
                "raw_json": str(raw_json_path),
                "outputs": {
                    "business_markdown": str(markdown_path),
                    "business_xlsx": str(xlsx_path),
                },
            },
            ensure_ascii=False,
            indent=2,
        )
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
