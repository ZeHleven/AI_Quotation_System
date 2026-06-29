from __future__ import annotations

import asyncio
import base64
import csv
import json
import re
from collections import Counter
from datetime import datetime
from pathlib import Path
from typing import Any, Awaitable, Callable, Mapping, Sequence

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.core.config import settings
from app.services.drawing_pdf_gap_recall_runner import EVIDENCE_HEADERS
from app.services.model_gateway import call_glm_drawing_tile_extract


PHASE = "BIZ-2x-pdf-object-recall-capture-runner"

CALL_HEADERS = [
    "call_no",
    "status",
    "recommended_pass",
    "source_file",
    "page",
    "tile_id",
    "image_path",
    "task_nos",
    "object_classes",
    "task_count",
    "reason",
    "raw_content",
]

CAPTURE_EVIDENCE_HEADERS = [
    "evidence_id",
    "source_kind",
    "task_no",
    "task_nos",
    *[header for header in EVIDENCE_HEADERS if header not in {"evidence_id", "source_kind"}],
]

VisionClient = Callable[..., Awaitable[Mapping[str, Any]]]


def run_object_recall_capture_pack(
    capture_pack: Mapping[str, Any],
    *,
    execute: bool = False,
    max_calls: int | None = None,
    start_call_no: int = 1,
    end_call_no: int | None = None,
    vision_client: VisionClient | None = None,
    trace_id: str | None = None,
) -> dict[str, Any]:
    all_capture_rows = [dict(row) for row in capture_pack.get("capture_rows") or [] if isinstance(row, Mapping)]
    capture_rows = _slice_capture_rows(all_capture_rows, start_call_no=start_call_no, end_call_no=end_call_no)
    if max_calls is not None:
        capture_rows = capture_rows[: max(max_calls, 0)]

    call_rows: list[dict[str, Any]] = []
    evidence_rows: list[dict[str, Any]] = []
    client = vision_client or call_glm_drawing_tile_extract
    for index, row in enumerate(capture_rows, start=1):
        call_no = _int(row.get("capture_no"), index)
        call_row = {
            "call_no": call_no,
            "status": "planned_dry_run",
            "recommended_pass": row.get("recommended_pass", ""),
            "source_file": row.get("source_file", ""),
            "page": row.get("page", ""),
            "tile_id": row.get("tile_id", ""),
            "image_path": row.get("image_path", ""),
            "task_nos": row.get("task_nos", ""),
            "object_classes": row.get("object_classes", ""),
            "task_count": row.get("task_count", ""),
            "reason": "dry_run_no_external_model_call",
            "raw_content": "",
        }
        if execute:
            _execute_capture_call(
                row,
                call_row=call_row,
                evidence_rows=evidence_rows,
                vision_client=client,
                trace_id=trace_id,
            )
        call_rows.append(call_row)

    summary = {
        "execute": execute,
        "source_capture_call_count": len(all_capture_rows),
        "capture_call_count": len(call_rows),
        "start_call_no": start_call_no,
        "end_call_no": end_call_no or "",
        "evidence_count": len(evidence_rows),
        "status_counts": dict(Counter(str(row.get("status") or "") for row in call_rows)),
        "pass_counts": dict(Counter(str(row.get("recommended_pass") or "") for row in call_rows)),
        "missing_image_call_count": sum(1 for row in call_rows if row.get("status") == "skipped_missing_image"),
        "target_fields_sent_to_model": False,
        "answer_columns_count_as_evidence": False,
        "quantity_status": "deferred_until_three_fields_accepted",
    }
    return {
        "ok": True,
        "phase": PHASE,
        "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "summary": summary,
        "call_rows": call_rows,
        "evidence_rows": evidence_rows,
        "visual_evidence_report": {"evidence_rows": evidence_rows},
    }


def write_object_recall_capture_run_outputs(
    report: Mapping[str, Any],
    output_dir: str | Path,
    *,
    stem: str | None = None,
) -> dict[str, str]:
    target = Path(output_dir)
    target.mkdir(parents=True, exist_ok=True)
    file_stem = stem or f"BIZ2x_PDF_object_recall_capture_run_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
    json_path = target / f"{file_stem}.json"
    call_csv = target / f"{file_stem}_calls.csv"
    evidence_csv = target / f"{file_stem}_evidence.csv"
    markdown_path = target / f"{file_stem}.md"
    xlsx_path = target / f"{file_stem}.xlsx"
    outputs = {
        "json": str(json_path),
        "call_csv": str(call_csv),
        "evidence_csv": str(evidence_csv),
        "markdown": str(markdown_path),
        "xlsx": str(xlsx_path),
    }
    payload = {**dict(report), "outputs": outputs}
    json_path.write_text(json.dumps(_json_safe(payload), ensure_ascii=False, indent=2), encoding="utf-8")
    _write_csv(call_csv, payload.get("call_rows") or [], CALL_HEADERS)
    _write_csv(evidence_csv, payload.get("evidence_rows") or [], CAPTURE_EVIDENCE_HEADERS)
    markdown_path.write_text(_build_markdown(payload), encoding="utf-8")
    _write_workbook(xlsx_path, payload)
    return outputs


def _slice_capture_rows(
    rows: Sequence[Mapping[str, Any]],
    *,
    start_call_no: int,
    end_call_no: int | None,
) -> list[dict[str, Any]]:
    start = max(1, _int(start_call_no, 1))
    end = _int(end_call_no, 0) if end_call_no is not None else 0
    selected: list[dict[str, Any]] = []
    for index, row in enumerate(rows, start=1):
        call_no = _int(row.get("capture_no"), index)
        if call_no < start:
            continue
        if end and call_no > end:
            continue
        selected.append(dict(row))
    return selected


def _execute_capture_call(
    row: Mapping[str, Any],
    *,
    call_row: dict[str, Any],
    evidence_rows: list[dict[str, Any]],
    vision_client: VisionClient,
    trace_id: str | None,
) -> None:
    image_path = Path(str(row.get("image_path") or ""))
    if not image_path.exists() or not image_path.is_file():
        call_row["status"] = "skipped_missing_image"
        call_row["reason"] = "image_path_missing"
        return
    try:
        model_result = asyncio.run(
            vision_client(
                base64.b64encode(image_path.read_bytes()).decode("ascii"),
                _image_mime_type(image_path),
                tile_context=_sanitized_tile_context(row),
                prompt_mode=str(row.get("recommended_pass") or "general"),
                prompt_override=str(row.get("prompt_text") or "") or None,
                trace_id=trace_id,
            )
        )
    except Exception as exc:  # pragma: no cover - concrete gateway/network errors vary
        call_row["status"] = "error"
        call_row["reason"] = _exception_reason(exc)
        return

    before_count = len(evidence_rows)
    for item in model_result.get("evidence_items") or []:
        text = _evidence_text(item)
        if not text:
            continue
        evidence_rows.append(
            {
                "evidence_id": f"PDFCAP-{len(evidence_rows) + 1:06d}",
                "source_kind": "pdf_object_recall_capture_llm",
                "task_no": _first_task_no(row.get("task_nos")),
                "task_nos": row.get("task_nos", ""),
                "source_file": row.get("source_file", ""),
                "page": row.get("page", ""),
                "tile_id": row.get("tile_id", ""),
                "vision_pass": row.get("recommended_pass", ""),
                "evidence_role": str(item.get("evidence_role") or "").strip() or _fallback_evidence_role(item),
                "discipline": str(item.get("discipline") or "").strip() or "unknown",
                "item_hint": _clean_item_hint(item.get("item_hint") or _first_present(item, _PUBLIC_DIAMETER_KEYS)),
                "space": str(item.get("space") or "").strip(),
                "material_codes": _normalize_string_list(item.get("material_codes")),
                "spec_or_method": _clean_text_value(
                    item.get("spec_or_method") or _first_present(item, _PLASTIC_OUTSIDE_DIAMETER_KEYS)
                ),
                "suggested_unit": _clean_unit(item.get("suggested_unit")),
                "text": text,
                "normalized_text": str(item.get("normalized_text") or "").strip() or text,
                "confidence": _bounded_float(item.get("confidence")),
                "model": settings.glm_vision_model,
                "needs_manual_review": bool(item.get("needs_manual_review", True)),
                "reason": str(item.get("reason") or "").strip(),
            }
        )
    call_row["status"] = "success"
    call_row["reason"] = f"extracted_{len(evidence_rows) - before_count}_items"
    call_row["raw_content"] = str(model_result.get("raw_content") or "")[:500]


_PUBLIC_DIAMETER_KEYS = ("public_diameter", "public diameter", "公称直径")
_PLASTIC_OUTSIDE_DIAMETER_KEYS = (
    "plastic_pipe_outside_diameter",
    "plastic pipe outside diameter",
    "塑料管外径",
)
_INCH_LABEL_KEYS = ("inch_label", "inch label", "英寸")


def _evidence_text(item: Mapping[str, Any]) -> str:
    explicit = str(item.get("text") or item.get("evidence_text") or "").strip()
    if explicit:
        return explicit
    row_parts = [
        _first_present(item, _PUBLIC_DIAMETER_KEYS),
        _first_present(item, _PLASTIC_OUTSIDE_DIAMETER_KEYS),
        _first_present(item, _INCH_LABEL_KEYS),
    ]
    if any(row_parts):
        return " | ".join(row_parts).strip()
    return ""


def _fallback_evidence_role(item: Mapping[str, Any]) -> str:
    if _first_present(item, _PUBLIC_DIAMETER_KEYS) or _first_present(item, _PLASTIC_OUTSIDE_DIAMETER_KEYS):
        return "table_row"
    return "unknown_note"


def _first_present(item: Mapping[str, Any], keys: Sequence[str]) -> str:
    for key in keys:
        value = str(item.get(key) or "").strip()
        if value:
            return value
    return ""


def _sanitized_tile_context(row: Mapping[str, Any]) -> dict[str, Any]:
    return {
        "source_file": row.get("source_file", ""),
        "page": row.get("page", ""),
        "tile_id": row.get("tile_id", ""),
        "vision_pass": row.get("recommended_pass", ""),
        "task_nos": row.get("task_nos", ""),
        "object_classes": row.get("object_classes", ""),
        "capture_pack_answer_blind": True,
    }


def _image_mime_type(path: Path) -> str:
    suffix = path.suffix.lower()
    if suffix in {".jpg", ".jpeg"}:
        return "image/jpeg"
    if suffix == ".webp":
        return "image/webp"
    return "image/png"


def _normalize_string_list(value: Any) -> list[str]:
    if isinstance(value, str):
        return [item.strip() for item in value.replace("；", ";").replace("，", ",").split(",") if item.strip()]
    if isinstance(value, Sequence):
        return [str(item).strip() for item in value if str(item).strip()]
    return []


def _clean_item_hint(value: Any) -> str:
    text = _clean_text_value(value)
    if text.lower() in {
        "removed finishes",
        "removed fixtures",
        "haul-away notes",
        "construction note",
        "demolition note",
        "unknown",
        "n/a",
        "na",
    }:
        return ""
    return text


def _first_task_no(value: Any) -> str:
    text = str(value or "")
    for part in text.replace("；", ";").replace(",", ";").split(";"):
        part = part.strip()
        if part.isdigit():
            return part
    return ""


def _clean_unit(value: Any) -> str:
    text = _clean_text_value(value)
    normalized = text.lower()
    if normalized in {"unknown", "n/a", "na", "unclear", "none", "null"}:
        return ""
    if "不要写数量" in text or "不确定则为空" in text:
        return ""
    if _looks_like_unit_option_prompt(text):
        return ""
    return text


def _looks_like_unit_option_prompt(value: str) -> bool:
    text = str(value or "").strip()
    if not text:
        return False
    compact = re.sub(r"\s+", "", text.lower())
    slash_count = compact.count("/") + compact.count("／")
    if slash_count < 2:
        return False
    unit_hits = sum(compact.count(token) for token in ("㎡", "m2", "m²", "m3", "m³", "m", "套", "个", "樘", "项", "?"))
    return unit_hits >= 3


def _clean_text_value(value: Any) -> str:
    text = str(value or "").strip()
    if not text:
        return ""
    placeholders = (
        "可见规格、材质、安装方式；没有则留空",
        "可见规格、材质、安装方式;没有则留空",
        "规格、材料、做法、安装方式或构造说明",
        "材料编号、材料名称、规格、做法、安装方式或构造说明",
    )
    if text in placeholders:
        return ""
    return text


def _exception_reason(exc: Exception) -> str:
    message = str(exc).strip() or repr(exc)
    return f"{exc.__class__.__name__}: {message}"[:300]


def _bounded_float(value: Any) -> float:
    try:
        parsed = float(value)
    except (TypeError, ValueError):
        return 0.0
    return max(0.0, min(1.0, parsed))


def _int(value: Any, default: int = 0) -> int:
    try:
        return int(value)
    except (TypeError, ValueError):
        return default


def _build_markdown(report: Mapping[str, Any]) -> str:
    summary = report.get("summary") or {}
    lines = [
        "# BIZ-2x PDF Object Recall Capture Run",
        "",
        f"- generated_at: {report.get('generated_at', '-')}",
        f"- execute: {summary.get('execute', False)}",
        f"- capture calls: {summary.get('capture_call_count', 0)}",
        f"- evidence rows: {summary.get('evidence_count', 0)}",
        f"- target_fields_sent_to_model: {summary.get('target_fields_sent_to_model', False)}",
        f"- quantity_status: {summary.get('quantity_status', '')}",
        "",
        "## Status Counts",
        "",
    ]
    for status, count in (summary.get("status_counts") or {}).items():
        lines.append(f"- {status}: {count}")
    return "\n".join(lines) + "\n"


def _write_workbook(path: Path, report: Mapping[str, Any]) -> None:
    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "run_summary"
    _append_rows(summary_sheet, [["metric", "value"]])
    for key, value in (report.get("summary") or {}).items():
        _append_rows(summary_sheet, [[key, json.dumps(value, ensure_ascii=False) if isinstance(value, (dict, list)) else value]])
    _style_sheet(summary_sheet)

    call_sheet = workbook.create_sheet("capture_calls")
    _append_rows(call_sheet, [CALL_HEADERS])
    _append_rows(call_sheet, [[_cell_value(row.get(header)) for header in CALL_HEADERS] for row in report.get("call_rows") or []])
    _style_sheet(call_sheet)

    evidence_sheet = workbook.create_sheet("recall_evidence")
    _append_rows(evidence_sheet, [CAPTURE_EVIDENCE_HEADERS])
    _append_rows(
        evidence_sheet,
        [[_cell_value(row.get(header)) for header in CAPTURE_EVIDENCE_HEADERS] for row in report.get("evidence_rows") or []],
    )
    _style_sheet(evidence_sheet)
    workbook.save(path)


def _append_rows(sheet: Any, rows: Sequence[Sequence[Any]]) -> None:
    for row in rows:
        sheet.append(list(row))


def _style_sheet(sheet: Any) -> None:
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for col_idx, column_cells in enumerate(sheet.columns, start=1):
        values = [str(cell.value or "") for cell in column_cells[:80]]
        width = min(max([len(value) for value in values] + [10]) + 2, 80)
        sheet.column_dimensions[get_column_letter(col_idx)].width = width
    sheet.freeze_panes = "A2"


def _write_csv(path: Path, rows: Sequence[Mapping[str, Any]], fieldnames: Sequence[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        for row in rows:
            writer.writerow({key: _cell_value(row.get(key)) for key in fieldnames})


def _cell_value(value: Any) -> Any:
    if isinstance(value, (dict, list, tuple)):
        return json.dumps(value, ensure_ascii=False)
    return value


def _json_safe(value: Any) -> Any:
    if isinstance(value, Mapping):
        return {str(key): _json_safe(item) for key, item in value.items()}
    if isinstance(value, (list, tuple)):
        return [_json_safe(item) for item in value]
    if isinstance(value, Path):
        return str(value)
    return value
