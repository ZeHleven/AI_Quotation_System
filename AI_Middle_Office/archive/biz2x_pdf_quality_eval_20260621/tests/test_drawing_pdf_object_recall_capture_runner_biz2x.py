from __future__ import annotations

import asyncio
import json
import sys
from pathlib import Path

from openpyxl import load_workbook

from app.services.drawing_pdf_object_recall_capture_runner import (
    run_object_recall_capture_pack,
    write_object_recall_capture_run_outputs,
)
from scripts import biz2x_pdf_object_recall_capture_run


def test_capture_runner_dry_run_keeps_answer_blind_context(tmp_path: Path):
    pack = _capture_pack(tmp_path)

    report = run_object_recall_capture_pack(pack, execute=False)

    assert report["summary"]["execute"] is False
    assert report["summary"]["capture_call_count"] == 1
    assert report["summary"]["evidence_count"] == 0
    assert report["summary"]["target_fields_sent_to_model"] is False
    serialized = json.dumps(report, ensure_ascii=False)
    assert "SECRET_TARGET" not in serialized


def test_capture_runner_can_slice_call_range(tmp_path: Path):
    pack = _capture_pack(tmp_path)
    second_image = tmp_path / "tile2.png"
    second_image.write_bytes(b"fake image 2")
    pack["capture_rows"].append(
        {
            "capture_no": 2,
            "recommended_pass": "finish_schedule",
            "source_file": "drawing.pdf",
            "page": 1,
            "tile_id": "p001_g02",
            "image_path": str(second_image),
            "task_nos": "3",
            "object_classes": "finish_wall",
            "task_count": 1,
            "prompt_text": "answer blind prompt 2",
        }
    )

    report = run_object_recall_capture_pack(pack, execute=False, start_call_no=2, end_call_no=2)

    assert report["summary"]["source_capture_call_count"] == 2
    assert report["summary"]["capture_call_count"] == 1
    assert report["summary"]["start_call_no"] == 2
    assert report["summary"]["end_call_no"] == 2
    assert report["call_rows"][0]["call_no"] == 2
    assert report["call_rows"][0]["recommended_pass"] == "finish_schedule"


def test_capture_runner_execute_uses_prompt_override_without_target_fields(tmp_path: Path):
    pack = _capture_pack(tmp_path)
    calls = []

    async def fake_vision_client(base64_image, mime_type, **kwargs):
        calls.append({"base64_image": base64_image, "mime_type": mime_type, "kwargs": kwargs})
        return {
            "raw_content": '{"evidence_items":[]}',
            "evidence_items": [
                {
                    "evidence_role": "construction_note",
                    "discipline": "decoration",
                    "item_hint": "visible door demolition",
                    "spec_or_method": "remove frame and hardware",
                    "suggested_unit": "set",
                    "text": "visible note: remove door frame and hardware",
                    "confidence": 0.82,
                    "needs_manual_review": True,
                    "reason": "visible note",
                }
            ],
        }

    report = run_object_recall_capture_pack(pack, execute=True, vision_client=fake_vision_client)

    assert report["summary"]["execute"] is True
    assert report["summary"]["evidence_count"] == 1
    assert report["call_rows"][0]["status"] == "success"
    assert report["evidence_rows"][0]["evidence_id"] == "PDFCAP-000001"
    assert report["evidence_rows"][0]["source_kind"] == "pdf_object_recall_capture_llm"
    assert report["evidence_rows"][0]["task_no"] == "1"
    assert report["evidence_rows"][0]["task_nos"] == "1;2"
    kwargs = calls[0]["kwargs"]
    assert kwargs["prompt_override"] == "answer blind prompt"
    assert kwargs["tile_context"]["capture_pack_answer_blind"] is True
    assert "target_item_name" not in kwargs["tile_context"]
    assert "SECRET_TARGET" not in json.dumps(kwargs, ensure_ascii=False)


def test_capture_runner_cleans_generic_item_hint_and_unknown_unit(tmp_path: Path):
    async def weak_client(base64_image, mime_type, **kwargs):
        return {
            "evidence_items": [
                {
                    "evidence_role": "construction_note",
                    "discipline": "decoration",
                    "item_hint": "Removed Fixtures",
                    "suggested_unit": "unknown",
                    "text": "visible note: remove existing fixtures",
                    "confidence": 0.4,
                    "needs_manual_review": True,
                }
            ]
        }

    report = run_object_recall_capture_pack(_capture_pack(tmp_path), execute=True, vision_client=weak_client)

    evidence = report["evidence_rows"][0]
    assert evidence["item_hint"] == ""
    assert evidence["suggested_unit"] == ""
    assert evidence["text"] == "visible note: remove existing fixtures"


def test_capture_runner_cleans_prompt_placeholder_spec_and_unit(tmp_path: Path):
    async def placeholder_client(base64_image, mime_type, **kwargs):
        return {
            "evidence_items": [
                {
                    "evidence_role": "sanitary_fixture_or_accessory",
                    "discipline": "plumbing",
                    "item_hint": "淋浴喷头",
                    "spec_or_method": "可见规格、材质、安装方式；没有则留空",
                    "suggested_unit": "个/套；不要写数量",
                    "text": "图中可见淋浴喷头符号",
                    "confidence": 0.8,
                    "needs_manual_review": True,
                }
            ]
        }

    report = run_object_recall_capture_pack(_capture_pack(tmp_path), execute=True, vision_client=placeholder_client)

    evidence = report["evidence_rows"][0]
    assert evidence["item_hint"] == "淋浴喷头"
    assert evidence["spec_or_method"] == ""
    assert evidence["suggested_unit"] == ""
    assert evidence["text"] == "图中可见淋浴喷头符号"


def test_capture_runner_cleans_prompt_unit_option_strings(tmp_path: Path):
    async def unit_options_client(base64_image, mime_type, **kwargs):
        return {
            "evidence_items": [
                {
                    "evidence_role": "finish_schedule_or_elevation_evidence",
                    "discipline": "decoration",
                    "item_hint": "墙面墙砖作美缝",
                    "spec_or_method": "墙面墙砖作美缝",
                    "suggested_unit": "?/m/?/?/m",
                    "text": "墙面墙砖作美缝",
                    "confidence": 0.0,
                    "needs_manual_review": True,
                }
            ]
        }

    report = run_object_recall_capture_pack(_capture_pack(tmp_path), execute=True, vision_client=unit_options_client)

    evidence = report["evidence_rows"][0]
    assert evidence["item_hint"] == "墙面墙砖作美缝"
    assert evidence["suggested_unit"] == ""
    assert evidence["text"] == "墙面墙砖作美缝"


def test_capture_runner_keeps_structured_table_rows_without_text(tmp_path: Path):
    async def table_client(base64_image, mime_type, **kwargs):
        return {
            "evidence_items": [
                {
                    "public_diameter": "DN40",
                    "plastic_pipe_outside_diameter": "De50",
                    "inch_label": "",
                    "confidence": 0.95,
                    "needs_manual_review": False,
                },
                {
                    "public diameter": "DN50",
                    "plastic pipe outside diameter": "De63",
                    "inch label": "",
                    "confidence": 0.95,
                    "needs_manual_review": False,
                },
            ]
        }

    report = run_object_recall_capture_pack(_capture_pack(tmp_path), execute=True, vision_client=table_client)

    assert report["summary"]["evidence_count"] == 2
    first = report["evidence_rows"][0]
    assert first["evidence_role"] == "table_row"
    assert first["item_hint"] == "DN40"
    assert first["spec_or_method"] == "De50"
    assert first["suggested_unit"] == ""
    assert first["text"] == "DN40 | De50 |"
    second = report["evidence_rows"][1]
    assert second["text"] == "DN50 | De63 |"


def test_capture_runner_outputs_are_external_import_compatible(tmp_path: Path):
    report = run_object_recall_capture_pack(_capture_pack(tmp_path), execute=True, vision_client=_fake_client)

    outputs = write_object_recall_capture_run_outputs(report, tmp_path / "out", stem="capture_run")

    assert Path(outputs["json"]).exists()
    assert Path(outputs["evidence_csv"]).exists()
    assert Path(outputs["xlsx"]).exists()
    workbook = load_workbook(outputs["xlsx"])
    assert workbook.sheetnames == ["run_summary", "capture_calls", "recall_evidence"]
    evidence_sheet = workbook["recall_evidence"]
    headers = {cell.value: cell.column for cell in evidence_sheet[1]}
    assert evidence_sheet.cell(row=2, column=headers["evidence_id"]).value == "PDFCAP-000001"
    assert evidence_sheet.cell(row=2, column=headers["item_hint"]).value == "visible door demolition"


def test_capture_runner_cli_dry_run(tmp_path: Path, monkeypatch, capsys):
    pack_json = tmp_path / "capture_pack.json"
    pack_json.write_text(json.dumps(_capture_pack(tmp_path), ensure_ascii=False), encoding="utf-8")
    output_dir = tmp_path / "cli_out"
    monkeypatch.setattr(
        sys,
        "argv",
        [
            "biz2x_pdf_object_recall_capture_run.py",
            "--capture-pack-json",
            str(pack_json),
            "--output-dir",
            str(output_dir),
            "--timestamp",
            "fixed",
        ],
    )

    assert biz2x_pdf_object_recall_capture_run.main() == 0
    payload = json.loads(capsys.readouterr().out)

    assert payload["ok"] is True
    assert payload["summary"]["execute"] is False
    assert payload["summary"]["capture_call_count"] == 1
    assert Path(payload["outputs"]["xlsx"]).exists()


async def _fake_client(base64_image, mime_type, **kwargs):
    return {
        "raw_content": '{"evidence_items":[]}',
        "evidence_items": [
            {
                "evidence_role": "construction_note",
                "discipline": "decoration",
                "item_hint": "visible door demolition",
                "spec_or_method": "remove frame and hardware",
                "suggested_unit": "set",
                "text": "visible note: remove door frame and hardware",
                "confidence": 0.82,
                "needs_manual_review": True,
                "reason": "visible note",
            }
        ],
    }


def _capture_pack(tmp_path: Path) -> dict[str, object]:
    image = tmp_path / "tile.png"
    image.write_bytes(b"fake image")
    return {
        "capture_rows": [
            {
                "capture_no": 1,
                "recommended_pass": "door_window_demolition",
                "source_file": "drawing.pdf",
                "page": 1,
                "tile_id": "p001_whole",
                "image_path": str(image),
                "task_nos": "1;2",
                "object_classes": "door_window_demolition",
                "task_count": 2,
                "prompt_text": "answer blind prompt",
                "target_item_name": "SECRET_TARGET",
            }
        ]
    }
