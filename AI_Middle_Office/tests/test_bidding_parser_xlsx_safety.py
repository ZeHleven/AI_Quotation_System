from __future__ import annotations

import re
import zipfile
from io import BytesIO

from openpyxl import Workbook

from app.services.bidding_parser import (
    BIDDING_PARSER_VERSION,
    extract_tender_text,
)


def _workbook_with_inflated_dimension() -> bytes:
    workbook = Workbook()
    current = workbook.active
    current.title = "二期清单"
    current.append(["项目名称", "二期公区精装修工程"])
    current.append(["项目编码", "项目名称", "单位", "工程量"])
    current.append(["011102003001", "块料楼地面", "m2", 125.5])

    conflicting = workbook.create_sheet("Sheet1")
    conflicting.append(["一期零星工程综合单价表"])
    conflicting.append(["不应混入二期项目的旧期次内容"])

    wide = workbook.create_sheet("超宽列测试")
    wide["A1"] = "正常列内容"
    wide.cell(row=2, column=300, value="列上限之外的内容")

    duplicate = workbook.create_sheet("二期清单副本")
    duplicate.append(["项目名称", "二期公区精装修工程"])
    duplicate.append(["项目编码", "项目名称", "单位", "工程量"])
    duplicate.append(["011102003001", "块料楼地面", "m2", 125.5])

    buffer = BytesIO()
    workbook.save(buffer)
    source = BytesIO(buffer.getvalue())
    target = BytesIO()
    with zipfile.ZipFile(source) as archive, zipfile.ZipFile(
        target,
        "w",
        compression=zipfile.ZIP_DEFLATED,
    ) as output:
        for item in archive.infolist():
            payload = archive.read(item.filename)
            if item.filename == "xl/worksheets/sheet1.xml":
                payload = re.sub(
                    rb'<dimension ref="[^"]+"',
                    b'<dimension ref="A1:XFD1048576"',
                    payload,
                    count=1,
                )
            output.writestr(item, payload)
    return target.getvalue()


def test_xlsx_parser_uses_effective_cells_and_quarantines_conflicts() -> None:
    result = extract_tender_text(
        _workbook_with_inflated_dimension(),
        "某项目二期招标清单.xlsx",
        (
            "application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        ),
    )

    assert result["parser_version"] == BIDDING_PARSER_VERSION
    assert "二期公区精装修工程" in result["text"]
    assert "二期清单 第3行" in result["text"]
    assert "不应混入二期项目的旧期次内容" not in result["text"]
    assert "列上限之外的内容" not in result["text"]

    diagnostics = result["parse_diagnostics"]
    by_name = {
        item["sheet_name"]: item
        for item in diagnostics["sheets"]
    }
    assert diagnostics["schema_version"] == "tender_xlsx_scan_v1"
    assert diagnostics["quarantined_sheet_count"] == 2
    assert (
        "bloated_declared_dimension"
        in by_name["二期清单"]["warning_codes"]
    )
    assert by_name["二期清单"]["effective_range"] == {
        "min_row": 1,
        "max_row": 3,
        "min_column": 1,
        "max_column": 4,
    }
    assert by_name["Sheet1"]["status"] == "quarantined"
    assert (
        by_name["Sheet1"]["quarantine_reason"]
        == "project_phase_mismatch"
    )
    assert (
        by_name["二期清单副本"]["quarantine_reason"]
        == "duplicate_sheet_content"
    )
    assert (
        by_name["超宽列测试"]["ignored_value_cell_count"] == 1
    )
    assert (
        "column_limit_applied"
        in by_name["超宽列测试"]["warning_codes"]
    )
