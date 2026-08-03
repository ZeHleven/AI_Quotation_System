from __future__ import annotations

import argparse
from collections import defaultdict
from dataclasses import dataclass
from difflib import SequenceMatcher
from pathlib import Path
import math
import re
import statistics
import struct
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


DEFAULT_INPUT = Path(
    r"C:\Users\12521\Documents\WXWork\1688855982836930\Cache\File\2026-06\旗胜劳务成本收集(20260604).xls"
)
DEFAULT_OUTPUT = Path(
    "AI_Middle_Office/data/exports/旗胜劳务成本清洗复核包_20260604.xlsx"
)

REVIEW_COLUMNS_BY_SHEET = {
    "疑似重复待复核": [
        "人工复核结论",
        "处理方式",
        "合并目标清洗ID",
        "人工修正名称",
        "人工修正项目特征",
        "复核说明",
        "复核人",
        "复核日期",
    ],
    "价格异常待复核": [
        "人工复核结论",
        "人工确认单价",
        "是否纳入主表",
        "处理方式",
        "复核说明",
        "复核人",
        "复核日期",
    ],
    "项目特征包含关系保留项": [
        "人工复核结论",
        "处理方式",
        "合并目标清洗ID",
        "人工修正项目特征",
        "复核说明",
        "复核人",
        "复核日期",
    ],
}

REVIEW_VALIDATIONS_BY_SHEET = {
    "疑似重复待复核": {
        "人工复核结论": ["合并", "不合并", "待补资料"],
        "处理方式": ["合并到条目A", "合并到条目B", "保留两项", "重新命名后保留"],
    },
    "价格异常待复核": {
        "人工复核结论": ["价格正确", "修正单价", "剔除", "待补资料"],
        "是否纳入主表": ["是", "否"],
        "处理方式": ["按原价保留", "按确认单价修正", "剔除该条", "待补资料后再定"],
    },
    "项目特征包含关系保留项": {
        "人工复核结论": ["确认保留两项", "仍需合并", "待补资料"],
        "处理方式": ["保留两项", "合并到短项", "合并到长项", "重新命名后保留"],
    },
}

FREE_SECTOR = 0xFFFFFFFF
END_OF_CHAIN = 0xFFFFFFFE


def _u16(data: bytes, offset: int) -> int:
    return struct.unpack_from("<H", data, offset)[0]


def _u32(data: bytes, offset: int) -> int:
    return struct.unpack_from("<I", data, offset)[0]


def _number_or_none(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, bool):
        return float(value)
    if isinstance(value, (int, float)) and math.isfinite(float(value)):
        return float(value)
    text = str(value).strip()
    if not text:
        return None
    text = text.replace(",", "").replace("，", "")
    try:
        parsed = float(text)
    except ValueError:
        return None
    if not math.isfinite(parsed):
        return None
    return parsed


def _text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _clean_header(value: Any) -> str:
    text = _text(value).replace("\n", "").replace("\r", "")
    text = re.sub(r"\s+", "", text)
    if text.endswith("E") and "辅材及机械费" in text:
        text = text[:-1]
    return text


def normalize_text(value: Any) -> str:
    text = _text(value)
    text = text.replace("\r\n", "\n").replace("\r", "\n")
    text = re.sub(r"[ \t\u3000]+", " ", text)
    text = re.sub(r"\n+", "\n", text)
    return text.strip()


def normalize_key_text(value: Any) -> str:
    text = normalize_text(value).lower()
    replacements = {
        "（": "(",
        "）": ")",
        "，": ",",
        "。": ".",
        "：": ":",
        "；": ";",
        "、": ",",
        "×": "*",
        "㎡": "m2",
        "ｍ": "m",
    }
    for source, target in replacements.items():
        text = text.replace(source, target)
    text = re.sub(r"\s+", "", text)
    return text


def display_number(value: float | None, digits: int = 4) -> float | None:
    if value is None:
        return None
    rounded = round(float(value), digits)
    if abs(rounded - round(rounded)) < 10 ** -digits:
        return int(round(rounded))
    return rounded


class _SstReader:
    def __init__(self, segments: list[bytes]) -> None:
        self.segments = segments
        self.segment_index = 0
        self.position = 0

    def _advance(self) -> None:
        while (
            self.segment_index < len(self.segments)
            and self.position >= len(self.segments[self.segment_index])
        ):
            self.segment_index += 1
            self.position = 0

    def read(self, length: int) -> bytes:
        output = bytearray()
        while length > 0:
            self._advance()
            if self.segment_index >= len(self.segments):
                break
            segment = self.segments[self.segment_index]
            take = min(length, len(segment) - self.position)
            output += segment[self.position : self.position + take]
            self.position += take
            length -= take
        return bytes(output)

    def read_u8(self) -> int:
        raw = self.read(1)
        return raw[0] if raw else 0

    def read_u16(self) -> int:
        return struct.unpack("<H", self.read(2).ljust(2, b"\0"))[0]

    def read_u32(self) -> int:
        return struct.unpack("<I", self.read(4).ljust(4, b"\0"))[0]

    def read_chars(self, count: int, is_16_bit: bool) -> str:
        pieces: list[str] = []
        remaining = count
        current_is_16_bit = is_16_bit
        while remaining > 0:
            self._advance()
            if self.segment_index >= len(self.segments):
                break
            segment = self.segments[self.segment_index]
            bytes_per_char = 2 if current_is_16_bit else 1
            available = len(segment) - self.position
            can_read = min(remaining, available // bytes_per_char)
            raw = segment[self.position : self.position + can_read * bytes_per_char]
            self.position += can_read * bytes_per_char
            remaining -= can_read
            if raw:
                pieces.append(
                    raw.decode("utf-16le" if current_is_16_bit else "latin1", errors="replace")
                )
            if remaining > 0:
                self._advance()
                if self.segment_index >= len(self.segments):
                    break
                flags = self.read_u8()
                current_is_16_bit = bool(flags & 1)
        return "".join(pieces)


def _parse_sst(segments: list[bytes] | None) -> list[str]:
    if not segments:
        return []
    reader = _SstReader(segments)
    reader.read_u32()
    unique_count = reader.read_u32()
    strings: list[str] = []
    for _ in range(unique_count):
        char_count = reader.read_u16()
        flags = reader.read_u8()
        is_16_bit = bool(flags & 1)
        has_ext = bool(flags & 4)
        has_rich = bool(flags & 8)
        rich_runs = reader.read_u16() if has_rich else 0
        ext_size = reader.read_u32() if has_ext else 0
        strings.append(reader.read_chars(char_count, is_16_bit))
        if has_rich:
            reader.read(4 * rich_runs)
        if has_ext:
            reader.read(ext_size)
    return strings


def _decode_rk(raw: int) -> int | float:
    if raw & 0x02:
        value = raw >> 2
        if value & 0x20000000:
            value -= 0x40000000
        decoded = float(value)
    else:
        decoded = struct.unpack("<d", struct.pack("<Q", (raw & 0xFFFFFFFC) << 32))[0]
    if raw & 0x01:
        decoded /= 100.0
    if abs(decoded - round(decoded)) < 1e-12:
        return int(round(decoded))
    return decoded


def _formula_result(raw: bytes) -> Any:
    if len(raw) < 8:
        return None
    if raw[6:8] == b"\xff\xff":
        marker = raw[0]
        if marker == 0:
            return ""
        if marker == 1:
            return bool(raw[2])
        if marker == 2:
            return "#ERR"
        return None
    value = struct.unpack("<d", raw)[0]
    if math.isfinite(value) and abs(value - round(value)) < 1e-12:
        return int(round(value))
    return value


def _decode_inline_label(record: bytes) -> str:
    if len(record) < 9:
        return ""
    char_count = _u16(record, 6)
    flags = record[8]
    is_16_bit = bool(flags & 1)
    start = 9
    raw = record[start : start + char_count * (2 if is_16_bit else 1)]
    return raw.decode("utf-16le" if is_16_bit else "latin1", errors="replace")


def read_xls_sheets(path: Path) -> dict[str, list[list[Any]]]:
    content = path.read_bytes()
    if content[:8] != bytes.fromhex("D0CF11E0A1B11AE1"):
        raise ValueError("Only OLE/BIFF .xls files are supported by this cleaner.")

    sector_size = 1 << _u16(content, 30)
    mini_sector_size = 1 << _u16(content, 32)
    fat_count = _u32(content, 44)
    first_dir = _u32(content, 48)
    mini_cutoff = _u32(content, 56)
    first_mini_fat = _u32(content, 60)
    mini_fat_count = _u32(content, 64)
    first_difat = _u32(content, 68)
    difat_count = _u32(content, 72)

    def sector_bytes(sector_id: int) -> bytes:
        offset = (sector_id + 1) * sector_size
        return content[offset : offset + sector_size]

    difat = [_u32(content, 76 + index * 4) for index in range(109)]
    difat = [value for value in difat if value not in (FREE_SECTOR, END_OF_CHAIN)]
    sector_id = first_difat
    for _ in range(difat_count):
        if sector_id in (FREE_SECTOR, END_OF_CHAIN):
            break
        sector = sector_bytes(sector_id)
        difat.extend(
            _u32(sector, index * 4)
            for index in range(sector_size // 4 - 1)
            if _u32(sector, index * 4) not in (FREE_SECTOR, END_OF_CHAIN)
        )
        sector_id = _u32(sector, sector_size - 4)

    fat: list[int] = []
    for fat_sector_id in difat[:fat_count]:
        sector = sector_bytes(fat_sector_id)
        fat.extend(_u32(sector, index * 4) for index in range(sector_size // 4))

    def read_chain(start: int, size: int | None = None) -> bytes:
        pieces: list[bytes] = []
        sector_id = start
        seen: set[int] = set()
        while (
            sector_id not in (FREE_SECTOR, END_OF_CHAIN)
            and sector_id < len(fat)
            and sector_id not in seen
        ):
            seen.add(sector_id)
            pieces.append(sector_bytes(sector_id))
            sector_id = fat[sector_id]
        data = b"".join(pieces)
        return data[:size] if size is not None else data

    directory_stream = read_chain(first_dir)
    entries: list[dict[str, Any]] = []
    for offset in range(0, len(directory_stream), 128):
        entry = directory_stream[offset : offset + 128]
        if len(entry) < 128:
            continue
        name_length = _u16(entry, 64)
        name = (
            entry[: max(name_length - 2, 0)].decode("utf-16le", errors="replace")
            if name_length >= 2
            else ""
        )
        entries.append(
            {
                "name": name,
                "type": entry[66],
                "start": _u32(entry, 116),
                "size": _u32(entry, 120),
            }
        )

    root = next((entry for entry in entries if entry["type"] == 5), None)
    workbook_entry = next(
        (entry for entry in entries if entry["name"] in {"Workbook", "Book"}), None
    )
    if not workbook_entry:
        raise ValueError("Workbook stream was not found in .xls file.")

    if workbook_entry["size"] < mini_cutoff and root is not None:
        mini_fat_stream = (
            read_chain(first_mini_fat, mini_fat_count * sector_size)
            if first_mini_fat not in (FREE_SECTOR, END_OF_CHAIN)
            else b""
        )
        mini_fat = [
            _u32(mini_fat_stream, index * 4) for index in range(len(mini_fat_stream) // 4)
        ]
        mini_stream = read_chain(root["start"], root["size"])

        def read_mini_chain(start: int, size: int | None = None) -> bytes:
            pieces: list[bytes] = []
            sector_id = start
            seen: set[int] = set()
            while (
                sector_id not in (FREE_SECTOR, END_OF_CHAIN)
                and sector_id < len(mini_fat)
                and sector_id not in seen
            ):
                seen.add(sector_id)
                offset = sector_id * mini_sector_size
                pieces.append(mini_stream[offset : offset + mini_sector_size])
                sector_id = mini_fat[sector_id]
            data = b"".join(pieces)
            return data[:size] if size is not None else data

        workbook = read_mini_chain(workbook_entry["start"], workbook_entry["size"])
    else:
        workbook = read_chain(workbook_entry["start"], workbook_entry["size"])

    records: list[tuple[int, int, bytes]] = []
    position = 0
    while position + 4 <= len(workbook):
        record_id = _u16(workbook, position)
        record_length = _u16(workbook, position + 2)
        records.append((position, record_id, workbook[position + 4 : position + 4 + record_length]))
        position += 4 + record_length

    bounds: list[dict[str, Any]] = []
    sst_segments: list[bytes] | None = None
    for index, (position, record_id, record) in enumerate(records):
        if record_id == 0x0085 and len(record) >= 8:
            name_length = record[6]
            flags = record[7]
            raw = record[8:]
            name = (
                raw[: name_length * 2].decode("utf-16le", errors="replace")
                if flags & 1
                else raw[:name_length].decode("latin1", errors="replace")
            )
            bounds.append({"offset": _u32(record, 0), "name": name})
        elif record_id == 0x00FC:
            segments = [record]
            next_index = index + 1
            while next_index < len(records) and records[next_index][1] == 0x003C:
                segments.append(records[next_index][2])
                next_index += 1
            sst_segments = segments

    shared_strings = _parse_sst(sst_segments)
    sheets: dict[str, list[list[Any]]] = {}
    for index, bound in enumerate(bounds):
        start = int(bound["offset"])
        end = int(bounds[index + 1]["offset"]) if index + 1 < len(bounds) else len(workbook)
        cells: dict[tuple[int, int], Any] = {}
        position = start
        while position + 4 <= end and position + 4 <= len(workbook):
            record_id = _u16(workbook, position)
            record_length = _u16(workbook, position + 2)
            record = workbook[position + 4 : position + 4 + record_length]

            if record_id == 0x00FD and len(record) >= 10:
                row = _u16(record, 0)
                column = _u16(record, 2)
                string_index = _u32(record, 6)
                cells[(row, column)] = (
                    shared_strings[string_index]
                    if string_index < len(shared_strings)
                    else f"<SST {string_index}>"
                )
            elif record_id == 0x0203 and len(record) >= 14:
                row = _u16(record, 0)
                column = _u16(record, 2)
                value = struct.unpack_from("<d", record, 6)[0]
                cells[(row, column)] = (
                    int(round(value))
                    if math.isfinite(value) and abs(value - round(value)) < 1e-12
                    else value
                )
            elif record_id == 0x027E and len(record) >= 10:
                cells[(_u16(record, 0), _u16(record, 2))] = _decode_rk(_u32(record, 6))
            elif record_id == 0x00BD and len(record) >= 8:
                row = _u16(record, 0)
                first_column = _u16(record, 2)
                last_column = _u16(record, len(record) - 2)
                offset = 4
                for column in range(first_column, last_column + 1):
                    if offset + 6 <= len(record) - 2:
                        cells[(row, column)] = _decode_rk(_u32(record, offset + 2))
                    offset += 6
            elif record_id == 0x0204 and len(record) >= 8:
                cells[(_u16(record, 0), _u16(record, 2))] = _decode_inline_label(record)
            elif record_id == 0x0006 and len(record) >= 14:
                value = _formula_result(record[6:14])
                if value not in (None, ""):
                    cells[(_u16(record, 0), _u16(record, 2))] = value
            elif record_id == 0x0205 and len(record) >= 8:
                cells[(_u16(record, 0), _u16(record, 2))] = (
                    bool(record[6]) if record[7] == 0 else f"#ERR{record[6]}"
                )

            position += 4 + record_length
            if record_id == 0x000A:
                break

        if not cells:
            sheets[str(bound["name"])] = []
            continue

        min_row = min(row for row, _ in cells)
        max_row = max(row for row, _ in cells)
        min_column = min(column for _, column in cells)
        max_column = max(column for _, column in cells)
        rows: list[list[Any]] = []
        for row in range(min_row, max_row + 1):
            rows.append([cells.get((row, column)) for column in range(min_column, max_column + 1)])
        sheets[str(bound["name"])] = rows
    return sheets


@dataclass(frozen=True)
class SourceRow:
    source_id: str
    trade: str
    source_sheet: str
    source_row: int
    project_name: str
    sequence: str
    item_name: str
    spec: str
    unit: str
    quantity: float | None
    price_type: str
    unit_price: float | None
    total_price: float | None
    remark: str
    labor_fee: float | None
    main_material_fee: float | None
    main_material_loss_rate: float | None
    auxiliary_machine_fee: float | None
    group_key: tuple[str, str, str, str, str]
    issues: tuple[str, ...]


def _find_column(headers: list[str], candidates: list[str]) -> int | None:
    for candidate in candidates:
        clean_candidate = _clean_header(candidate)
        for index, header in enumerate(headers):
            if clean_candidate == header:
                return index
    for candidate in candidates:
        clean_candidate = _clean_header(candidate)
        for index, header in enumerate(headers):
            if clean_candidate and clean_candidate in header:
                return index
    return None


def _is_project_row(row: list[Any]) -> bool:
    non_empty = [value for value in row if normalize_text(value)]
    if len(non_empty) != 1:
        return False
    first = normalize_text(row[0] if row else "")
    if not first:
        return False
    if _number_or_none(first) is not None:
        return False
    return True


def extract_source_rows(sheets: dict[str, list[list[Any]]]) -> list[SourceRow]:
    source_rows: list[SourceRow] = []
    source_counter = 1
    for sheet_name, rows in sheets.items():
        if not rows:
            continue
        headers = [_clean_header(value) for value in rows[0]]
        sequence_col = _find_column(headers, ["序号"])
        item_col = _find_column(headers, ["分部分项工程", "项目名称", "施工项目"])
        spec_col = _find_column(headers, ["项目特征", "规格", "规格/特征"])
        unit_col = _find_column(headers, ["单位"])
        quantity_col = _find_column(headers, ["工程量", "数量"])
        unit_price_col = _find_column(headers, ["不含税综合单价（元）", "综合单价（元）", "单价"])
        total_price_col = _find_column(headers, ["不含税合价（元）", "合价（元）", "总价"])
        remark_col = _find_column(headers, ["备注"])
        labor_col = _find_column(headers, ["人工费（元）", "人工费"])
        main_material_col = _find_column(headers, ["主材费（元）", "主材费"])
        loss_col = _find_column(headers, ["主材损耗率（%）", "主材损耗率"])
        auxiliary_col = _find_column(headers, ["辅材及机械费（元）", "辅材及机械费"])

        unit_price_header = headers[unit_price_col] if unit_price_col is not None else ""
        price_type = "不含税综合单价" if "不含税" in unit_price_header else "综合单价"
        current_project = ""

        for row_index, row in enumerate(rows[1:], start=2):
            if _is_project_row(row):
                current_project = normalize_text(row[0])
                continue

            sequence = normalize_text(row[sequence_col]) if sequence_col is not None else ""
            item_name = normalize_text(row[item_col]) if item_col is not None else ""
            spec = normalize_text(row[spec_col]) if spec_col is not None else ""
            unit = normalize_text(row[unit_col]) if unit_col is not None else ""
            quantity = _number_or_none(row[quantity_col]) if quantity_col is not None else None
            unit_price = _number_or_none(row[unit_price_col]) if unit_price_col is not None else None
            total_price = _number_or_none(row[total_price_col]) if total_price_col is not None else None
            remark = normalize_text(row[remark_col]) if remark_col is not None else ""

            if not item_name:
                continue
            if _number_or_none(sequence) is None and unit_price is None and quantity is None:
                continue

            issues: list[str] = []
            if not unit:
                issues.append("缺少单位")
            if unit_price is None:
                issues.append("缺少单价")
            elif unit_price == 0:
                issues.append("单价为0")
            if not spec:
                issues.append("缺少项目特征")
            if not current_project:
                issues.append("缺少工程项目分组")

            group_key = (
                sheet_name,
                normalize_key_text(item_name),
                normalize_key_text(spec),
                normalize_key_text(unit),
                price_type,
            )
            source_rows.append(
                SourceRow(
                    source_id=f"S{source_counter:05d}",
                    trade=sheet_name,
                    source_sheet=sheet_name,
                    source_row=row_index,
                    project_name=current_project,
                    sequence=sequence,
                    item_name=item_name,
                    spec=spec,
                    unit=unit,
                    quantity=quantity,
                    price_type=price_type,
                    unit_price=unit_price,
                    total_price=total_price,
                    remark=remark,
                    labor_fee=_number_or_none(row[labor_col]) if labor_col is not None else None,
                    main_material_fee=(
                        _number_or_none(row[main_material_col]) if main_material_col is not None else None
                    ),
                    main_material_loss_rate=(
                        _number_or_none(row[loss_col]) if loss_col is not None else None
                    ),
                    auxiliary_machine_fee=(
                        _number_or_none(row[auxiliary_col]) if auxiliary_col is not None else None
                    ),
                    group_key=group_key,
                    issues=tuple(issues),
                )
            )
            source_counter += 1
    return source_rows


def _avg(values: list[float | None]) -> float | None:
    clean = [float(value) for value in values if value is not None and math.isfinite(float(value))]
    return statistics.fmean(clean) if clean else None


def _price_warning(min_price: float | None, max_price: float | None) -> tuple[str, float | None]:
    if min_price is None or max_price is None:
        return "缺少单价，需复核", None
    if min_price <= 0 and max_price > 0:
        return "价格波动>100%（含0价）", None
    if min_price == 0 and max_price == 0:
        return "单价为0，需复核", 0
    if min_price <= 0:
        return "", None
    ratio = (max_price - min_price) / min_price
    if ratio > 1:
        return "价格波动>100%，需复核", ratio
    if ratio > 0.5:
        return "价格波动>50%，需复核", ratio
    return "", ratio


def build_clean_rows(source_rows: list[SourceRow]) -> tuple[list[dict[str, Any]], dict[tuple[str, str, str, str, str], str]]:
    groups: dict[tuple[str, str, str, str, str], list[SourceRow]] = defaultdict(list)
    for row in source_rows:
        groups[row.group_key].append(row)

    clean_rows: list[dict[str, Any]] = []
    group_ids: dict[tuple[str, str, str, str, str], str] = {}
    for index, (group_key, rows) in enumerate(sorted(groups.items(), key=lambda item: item[0]), start=1):
        group_id = f"C{index:05d}"
        group_ids[group_key] = group_id
        prices = [row.unit_price for row in rows if row.unit_price is not None]
        min_price = min(prices) if prices else None
        max_price = max(prices) if prices else None
        warning, ratio = _price_warning(min_price, max_price)
        projects = sorted({row.project_name for row in rows if row.project_name})
        remarks = sorted({row.remark for row in rows if row.remark})
        source_refs = "; ".join(f"{row.source_sheet}!R{row.source_row}" for row in rows[:20])
        clean_rows.append(
            {
                "清洗ID": group_id,
                "工种": rows[0].trade,
                "分部分项工程": rows[0].item_name,
                "项目特征": rows[0].spec,
                "单位": rows[0].unit,
                "价格口径": rows[0].price_type,
                "平均单价（简单平均）": display_number(_avg([row.unit_price for row in rows]), 4),
                "样本数": len(rows),
                "最低单价": display_number(min_price, 4),
                "最高单价": display_number(max_price, 4),
                "价格差异率": display_number(ratio, 4),
                "价格波动标记": warning,
                "来源工程项目数": len(projects),
                "来源工程项目": "；".join(projects[:12]) + ("；..." if len(projects) > 12 else ""),
                "平均人工费": display_number(_avg([row.labor_fee for row in rows]), 4),
                "平均主材费": display_number(_avg([row.main_material_fee for row in rows]), 4),
                "平均主材损耗率": display_number(_avg([row.main_material_loss_rate for row in rows]), 4),
                "平均辅材及机械费": display_number(_avg([row.auxiliary_machine_fee for row in rows]), 4),
                "代表备注": "；".join(remarks[:5]),
                "来源行": source_refs,
                "维护建议": warning or ("可作为候选成本条目复核" if len(rows) > 1 else "单样本，建议人工确认后维护"),
                "归并键": "|".join(group_key),
            }
        )
    return clean_rows, group_ids


def source_detail_rows(source_rows: list[SourceRow], group_ids: dict[tuple[str, str, str, str, str], str]) -> list[dict[str, Any]]:
    details: list[dict[str, Any]] = []
    for row in source_rows:
        details.append(
            {
                "来源ID": row.source_id,
                "清洗ID": group_ids[row.group_key],
                "工种": row.trade,
                "来源Sheet": row.source_sheet,
                "来源行号": row.source_row,
                "工程项目": row.project_name,
                "序号": row.sequence,
                "分部分项工程": row.item_name,
                "项目特征": row.spec,
                "单位": row.unit,
                "工程量": display_number(row.quantity, 4),
                "价格口径": row.price_type,
                "单价": display_number(row.unit_price, 4),
                "合价": display_number(row.total_price, 4),
                "备注": row.remark,
                "人工费": display_number(row.labor_fee, 4),
                "主材费": display_number(row.main_material_fee, 4),
                "主材损耗率": display_number(row.main_material_loss_rate, 4),
                "辅材及机械费": display_number(row.auxiliary_machine_fee, 4),
                "是否纳入均价": "是" if row.unit_price is not None else "否",
                "记录问题": "；".join(row.issues),
                "归并键": "|".join(row.group_key),
            }
        )
    return details


def price_exception_rows(clean_rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    return [
        row
        for row in clean_rows
        if row.get("价格波动标记")
        or (row.get("样本数") == 1 and row.get("价格口径") and row.get("平均单价（简单平均）") in (0, None))
    ]


def _similarity(a: str, b: str) -> float:
    if not a or not b:
        return 0.0
    ratio = SequenceMatcher(None, a, b).ratio()
    tokens_a = set(re.findall(r"[\w\u4e00-\u9fff]+", a))
    tokens_b = set(re.findall(r"[\w\u4e00-\u9fff]+", b))
    jaccard = len(tokens_a & tokens_b) / len(tokens_a | tokens_b) if tokens_a | tokens_b else 0.0
    return max(ratio, jaccard)


def build_review_relations(clean_rows: list[dict[str, Any]]) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    by_scope: dict[tuple[str, str, str], list[dict[str, Any]]] = defaultdict(list)
    for row in clean_rows:
        by_scope[(row["工种"], row["价格口径"], row["单位"])].append(row)

    duplicates: list[dict[str, Any]] = []
    contains: list[dict[str, Any]] = []
    for (_, _, _), rows in by_scope.items():
        rows_sorted = sorted(rows, key=lambda row: str(row["清洗ID"]))
        for i, left in enumerate(rows_sorted):
            left_name_key = normalize_key_text(left["分部分项工程"])
            left_spec_key = normalize_key_text(left["项目特征"])
            for right in rows_sorted[i + 1 :]:
                right_name_key = normalize_key_text(right["分部分项工程"])
                right_spec_key = normalize_key_text(right["项目特征"])
                if not left_spec_key or not right_spec_key:
                    continue
                same_name = left_name_key == right_name_key
                name_similarity = _similarity(left_name_key, right_name_key)
                spec_similarity = _similarity(left_spec_key, right_spec_key)
                contains_relation = (
                    same_name
                    and left_spec_key != right_spec_key
                    and (left_spec_key in right_spec_key or right_spec_key in left_spec_key)
                )
                if contains_relation:
                    short, long = (
                        (left, right)
                        if len(left_spec_key) <= len(right_spec_key)
                        else (right, left)
                    )
                    contains.append(
                        {
                            "工种": short["工种"],
                            "价格口径": short["价格口径"],
                            "单位": short["单位"],
                            "短项ID": short["清洗ID"],
                            "长项ID": long["清洗ID"],
                            "分部分项工程": short["分部分项工程"],
                            "短项目特征": short["项目特征"],
                            "长项目特征": long["项目特征"],
                            "短项平均单价": short["平均单价（简单平均）"],
                            "长项平均单价": long["平均单价（简单平均）"],
                            "规则说明": "项目特征存在包含关系，按规则保留两项，不自动合并",
                        }
                    )
                    continue
                if same_name and spec_similarity >= 0.9:
                    duplicates.append(
                        {
                            "工种": left["工种"],
                            "价格口径": left["价格口径"],
                            "单位": left["单位"],
                            "条目A_ID": left["清洗ID"],
                            "条目B_ID": right["清洗ID"],
                            "条目A名称": left["分部分项工程"],
                            "条目B名称": right["分部分项工程"],
                            "条目A项目特征": left["项目特征"],
                            "条目B项目特征": right["项目特征"],
                            "A平均单价": left["平均单价（简单平均）"],
                            "B平均单价": right["平均单价（简单平均）"],
                            "名称相似度": display_number(name_similarity, 4),
                            "项目特征相似度": display_number(spec_similarity, 4),
                            "复核建议": "项目特征高度相似但不完全一致，建议人工确认是否可合并",
                        }
                    )
                elif name_similarity >= 0.92 and spec_similarity >= 0.88:
                    duplicates.append(
                        {
                            "工种": left["工种"],
                            "价格口径": left["价格口径"],
                            "单位": left["单位"],
                            "条目A_ID": left["清洗ID"],
                            "条目B_ID": right["清洗ID"],
                            "条目A名称": left["分部分项工程"],
                            "条目B名称": right["分部分项工程"],
                            "条目A项目特征": left["项目特征"],
                            "条目B项目特征": right["项目特征"],
                            "A平均单价": left["平均单价（简单平均）"],
                            "B平均单价": right["平均单价（简单平均）"],
                            "名称相似度": display_number(name_similarity, 4),
                            "项目特征相似度": display_number(spec_similarity, 4),
                            "复核建议": "名称和项目特征均较相似，建议人工确认是否重复",
                        }
                    )
    return duplicates, contains


def _write_sheet(workbook: Workbook, name: str, rows: list[dict[str, Any]]) -> None:
    sheet = workbook.create_sheet(name)
    review_columns = REVIEW_COLUMNS_BY_SHEET.get(name, [])
    if not rows:
        sheet.append(["说明", *review_columns])
        sheet.append(["本页暂无记录"])
        sheet.freeze_panes = "A2"
        return

    headers = list(rows[0].keys())
    for column in review_columns:
        if column not in headers:
            headers.append(column)
    sheet.append(headers)
    for row in rows:
        sheet.append([row.get(header) for header in headers])

    header_fill = PatternFill("solid", fgColor="1F4E78")
    review_header_fill = PatternFill("solid", fgColor="548235")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in sheet[1]:
        cell.fill = review_header_fill if cell.value in review_columns else header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    warning_fill = PatternFill("solid", fgColor="FFF2CC")
    danger_fill = PatternFill("solid", fgColor="FCE4D6")
    header_index = {cell.value: cell.column for cell in sheet[1]}
    warning_col = header_index.get("价格波动标记") or header_index.get("记录问题")
    for row_cells in sheet.iter_rows(min_row=2):
        for cell in row_cells:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        if warning_col:
            warning_value = sheet.cell(row=row_cells[0].row, column=warning_col).value
            if warning_value:
                fill = danger_fill if "100%" in str(warning_value) else warning_fill
                for cell in row_cells:
                    cell.fill = fill

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions

    header_index = {cell.value: cell.column for cell in sheet[1]}
    for header, options in REVIEW_VALIDATIONS_BY_SHEET.get(name, {}).items():
        column_index = header_index.get(header)
        if not column_index:
            continue
        formula = '"' + ",".join(options) + '"'
        validation = DataValidation(type="list", formula1=formula, allow_blank=True)
        validation.error = "请从下拉选项中选择复核结论。"
        validation.errorTitle = "无效选项"
        validation.prompt = "请选择人工复核判断。"
        validation.promptTitle = header
        sheet.add_data_validation(validation)
        column_letter = get_column_letter(column_index)
        validation.add(f"{column_letter}2:{column_letter}{max(sheet.max_row, 2000)}")

    max_width_by_header = {
        "项目特征": 58,
        "条目A项目特征": 48,
        "条目B项目特征": 48,
        "短项目特征": 48,
        "长项目特征": 58,
        "来源工程项目": 42,
        "工程项目": 42,
        "备注": 36,
        "代表备注": 36,
        "人工修正项目特征": 48,
        "复核说明": 40,
        "归并键": 42,
    }
    for column in range(1, sheet.max_column + 1):
        letter = get_column_letter(column)
        header = sheet.cell(row=1, column=column).value
        width = 12
        for cell in sheet[letter][:200]:
            value = "" if cell.value is None else str(cell.value)
            width = max(width, min(max_width_by_header.get(header, 26), len(value) * 1.3 + 2))
        sheet.column_dimensions[letter].width = max_width_by_header.get(header, width)


def _write_review_guide_sheet(workbook: Workbook) -> None:
    sheet = workbook.create_sheet("复核指南")
    rows = [
        ["旗胜劳务成本清洗复核指南", "", "", ""],
        ["适用范围", "本文件用于业务部门人工复核劳务成本清洗结果；复核对象主要是疑似重复、价格异常、项目特征包含关系三类记录。", "", ""],
        ["总体原则", "只在同一个工种 Sheet 内判断是否合并；不同价格口径（综合单价/不含税综合单价）分开维护；项目特征包含另一个项目特征时默认不合并。", "", ""],
        ["复核顺序", "1. 先看价格异常待复核；2. 再看项目特征包含关系保留项；3. 最后看疑似重复待复核；4. 必要时回到来源明细查原始工程项目和行号。", "", ""],
        ["", "", "", ""],
        ["Sheet", "业务员要判断什么", "需要填写的字段", "填写口径"],
        [
            "价格异常待复核",
            "判断单价为 0、缺价、或同类价格波动大的记录是否真实可用。",
            "人工复核结论、人工确认单价、是否纳入主表、处理方式、复核说明、复核人、复核日期",
            "价格正确则选“价格正确/按原价保留”；需要调整则选“修正单价/按确认单价修正”并填写人工确认单价；无依据或明显错误则选“剔除”或“待补资料”。",
        ],
        [
            "项目特征包含关系保留项",
            "判断短项目特征和长项目特征是否确实是不同施工内容。",
            "人工复核结论、处理方式、合并目标清洗ID、人工修正项目特征、复核说明、复核人、复核日期",
            "若长短特征对应不同做法、材料、施工范围或价格明显不同，选“确认保留两项/保留两项”；只有业务确认完全同项时才选“仍需合并”。",
        ],
        [
            "疑似重复待复核",
            "判断两个相似条目是否应该合并成同一个成本条目。",
            "人工复核结论、处理方式、合并目标清洗ID、人工修正名称、人工修正项目特征、复核说明、复核人、复核日期",
            "名称、项目特征、单位、施工范围基本一致时可选“合并”；只要做法、规格、材料、部位、计量口径不同，选“不合并/保留两项”。",
        ],
        ["", "", "", ""],
        ["判断规则", "怎么判断", "建议结论", "备注"],
        [
            "同名同特征同单位",
            "同工种、同价格口径下，名称、项目特征、单位都一致。",
            "可合并",
            "系统已自动归并到清洗主表，一般无需在复核页处理。",
        ],
        [
            "同名但特征多/少一段",
            "例如一个只写“灯槽”，另一个写“灯槽+基层+石膏板”。",
            "通常保留两项",
            "这是包含关系，不代表同一种成本项，除非业务确认施工内容完全一致。",
        ],
        [
            "同名但材料规格不同",
            "例如 9.5mm 石膏板、12mm 阻燃板、不同厚度铝单板。",
            "不合并",
            "材料规格不同通常会导致价格不同，应分别维护。",
        ],
        [
            "同名但单位不同",
            "例如 m、m2、项、个。",
            "不合并",
            "计量口径不同，不能直接平均。",
        ],
        [
            "价格波动超过 50%",
            "同一清洗条目内最高价/最低价差异较大。",
            "需复核",
            "若有特殊工程条件，在复核说明写明原因；若明显录错，填写人工确认单价。",
        ],
        [
            "单价为 0 或缺价",
            "原始行没有有效单价。",
            "修正/剔除/待补资料",
            "不能直接作为有效成本价使用。",
        ],
        ["", "", "", ""],
        ["字段填写说明", "字段", "怎么填", "示例"],
        ["通用", "人工复核结论", "从下拉选项选择最终判断。", "合并、不合并、价格正确、修正单价、确认保留两项"],
        ["通用", "处理方式", "说明后续怎么处理这条记录。", "保留两项、合并到条目A、按确认单价修正、剔除该条"],
        ["合并类", "合并目标清洗ID", "需要合并时填写保留的目标清洗ID。", "C00123"],
        ["价格类", "人工确认单价", "价格异常需要修正时填写最终确认单价。", "58.5"],
        ["价格类", "是否纳入主表", "确认这条是否进入后续成本主表。", "是/否"],
        ["修正类", "人工修正名称/人工修正项目特征", "业务认为名称或特征需要规范化时填写。", "石膏板吊顶；轻钢龙骨+阻燃板基层+石膏板面层"],
        ["通用", "复核说明", "写清判断依据，便于后续追溯。", "规格不同，单价差异明显，保留两项"],
        ["通用", "复核人/复核日期", "填写复核人姓名和日期。", "张三；2026-06-04"],
        ["", "", "", ""],
        ["复核完成标准", "三个复核 Sheet 中需要业务判断的行，均已填写人工复核结论；涉及修正价格的行已填写人工确认单价；涉及合并的行已填写合并目标清洗ID；复核说明能解释主要判断依据。", "", ""],
        ["后续处理", "复核完成后，由数据维护人员根据人工结论更新清洗主表，再进入成本库 draft 待审核；不要直接把未复核异常项导入 active 成本库。", "", ""],
    ]
    for row in rows:
        sheet.append(row)

    title_fill = PatternFill("solid", fgColor="1F4E78")
    section_fill = PatternFill("solid", fgColor="D9EAF7")
    table_header_fill = PatternFill("solid", fgColor="548235")
    title_font = Font(color="FFFFFF", bold=True, size=15)
    section_font = Font(bold=True)
    header_font = Font(color="FFFFFF", bold=True)

    sheet.merge_cells("A1:D1")
    sheet["A1"].fill = title_fill
    sheet["A1"].font = title_font
    sheet["A1"].alignment = Alignment(horizontal="center", vertical="center")

    table_header_rows = {6, 11, 19}
    for row in range(2, sheet.max_row + 1):
        first = sheet.cell(row=row, column=1).value
        if row in table_header_rows:
            for cell in sheet[row]:
                cell.fill = table_header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        elif first and sheet.cell(row=row, column=2).value and not sheet.cell(row=row, column=3).value:
            for cell in sheet[row]:
                cell.fill = section_fill
                cell.font = section_font

    widths = [18, 48, 42, 58]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    sheet.freeze_panes = "A6"


def write_workbook(
    output_path: Path,
    clean_rows: list[dict[str, Any]],
    detail_rows: list[dict[str, Any]],
    duplicate_rows: list[dict[str, Any]],
    exception_rows: list[dict[str, Any]],
    contains_rows: list[dict[str, Any]],
) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    workbook.remove(workbook.active)
    _write_review_guide_sheet(workbook)
    _write_sheet(workbook, "清洗主表", clean_rows)
    _write_sheet(workbook, "来源明细", detail_rows)
    _write_sheet(workbook, "疑似重复待复核", duplicate_rows)
    _write_sheet(workbook, "价格异常待复核", exception_rows)
    _write_sheet(workbook, "项目特征包含关系保留项", contains_rows)
    workbook.save(output_path)


def verify_workbook(path: Path) -> dict[str, Any]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    return {
        "path": str(path),
        "sheets": {
            sheet_name: {
                "rows": workbook[sheet_name].max_row,
                "cols": workbook[sheet_name].max_column,
            }
            for sheet_name in workbook.sheetnames
        },
    }


def main() -> None:
    parser = argparse.ArgumentParser(description="Clean labor cost .xls workbook into a reviewable .xlsx pack.")
    parser.add_argument("--input", type=Path, default=DEFAULT_INPUT)
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT)
    args = parser.parse_args()

    sheets = read_xls_sheets(args.input)
    source_rows = extract_source_rows(sheets)
    clean_rows, group_ids = build_clean_rows(source_rows)
    details = source_detail_rows(source_rows, group_ids)
    duplicates, contains = build_review_relations(clean_rows)
    exceptions = price_exception_rows(clean_rows)
    write_workbook(args.output, clean_rows, details, duplicates, exceptions, contains)
    verification = verify_workbook(args.output)
    print(
        {
            "source_rows": len(source_rows),
            "clean_rows": len(clean_rows),
            "duplicate_review_rows": len(duplicates),
            "price_exception_rows": len(exceptions),
            "contains_review_rows": len(contains),
            "verification": verification,
        }
    )


if __name__ == "__main__":
    main()
