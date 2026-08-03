from __future__ import annotations

import csv
import json
from datetime import datetime
from pathlib import Path
from typing import Any, Iterable, Mapping, Sequence

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


PHASE = "BIZ-2x-pdf-three-field-human-review"

REVIEW_HEADERS = [
    "review_no",
    "review_result",
    "status",
    "answer_sheet",
    "answer_row_no",
    "answer_section",
    "answer_item_code",
    "answer_item_name",
    "answer_feature",
    "answer_unit",
    "candidate_row_no",
    "candidate_item_name",
    "candidate_feature",
    "candidate_unit",
    "name_match",
    "feature_match",
    "unit_match",
    "name_score",
    "feature_score",
    "unit_score",
    "overall_score",
    "candidate_source_files",
    "evidence_pages",
    "evidence_tiles",
    "evidence_ids",
    "standard_code",
    "standard_item_code",
    "standard_item_name",
    "issue",
    "manual_review_note",
]

SUMMARY_HEADERS = ["metric", "value"]


def build_three_field_human_review_report(v2_report: Mapping[str, Any]) -> dict[str, Any]:
    acceptance = dict(v2_report.get("three_field_acceptance_report") or {})
    if not acceptance:
        raise ValueError("v2_report does not contain three_field_acceptance_report")

    comparison_rows = list(acceptance.get("comparison_rows") or [])
    answer_rows = _answer_rows_by_key(acceptance.get("answer_rows") or [])
    candidate_rows = _candidate_rows_by_no(acceptance.get("candidate_rows") or [])
    human_rows = list(v2_report.get("human_style_rows") or [])
    evidence_index = _evidence_index(v2_report.get("evidence_rows") or [])

    review_rows = []
    for index, row in enumerate(comparison_rows, start=1):
        answer = answer_rows.get(_answer_key(row))
        candidate_no = _int(row.get("candidate_row_no"), 0)
        candidate = candidate_rows.get(candidate_no, {})
        human_row = _human_row_for_candidate(candidate_no, human_rows, candidate)
        evidence_ids = _evidence_ids(human_row, candidate)
        evidence_locations = [evidence_index.get(evidence_id, {}) for evidence_id in evidence_ids]
        review_rows.append(
            {
                "review_no": index,
                "review_result": _review_result(row),
                "status": row.get("status", ""),
                "answer_sheet": row.get("answer_sheet", ""),
                "answer_row_no": row.get("answer_row_no", ""),
                "answer_section": (answer or {}).get("section", ""),
                "answer_item_code": (answer or {}).get("item_code", ""),
                "answer_item_name": row.get("answer_item_name", ""),
                "answer_feature": row.get("answer_feature", ""),
                "answer_unit": row.get("answer_unit", ""),
                "candidate_row_no": row.get("candidate_row_no", ""),
                "candidate_item_name": row.get("candidate_item_name", ""),
                "candidate_feature": row.get("candidate_feature", ""),
                "candidate_unit": row.get("candidate_unit", ""),
                "name_match": _match_label(row.get("name_score"), pass_threshold=0.78, review_threshold=0.55),
                "feature_match": _match_label(row.get("feature_score"), pass_threshold=0.55, review_threshold=0.35),
                "unit_match": "yes" if _float(row.get("unit_score"), 0) >= 1 else "no",
                "name_score": row.get("name_score", ""),
                "feature_score": row.get("feature_score", ""),
                "unit_score": row.get("unit_score", ""),
                "overall_score": row.get("overall_score", ""),
                "candidate_source_files": _first_non_empty(
                    human_row.get("source_files"),
                    candidate.get("source"),
                    row.get("candidate_source"),
                ),
                "evidence_pages": _join_unique(location.get("page", "") for location in evidence_locations),
                "evidence_tiles": _join_unique(location.get("tile_id", "") for location in evidence_locations),
                "evidence_ids": ", ".join(evidence_ids),
                "standard_code": human_row.get("standard_code", ""),
                "standard_item_code": human_row.get("standard_item_code", ""),
                "standard_item_name": human_row.get("standard_item_name", ""),
                "issue": row.get("issue", ""),
                "manual_review_note": "",
            }
        )

    acceptance_summary = dict(acceptance.get("summary") or {})
    v2_summary = dict(v2_report.get("summary") or {})
    status_counts: dict[str, int] = {}
    for row in review_rows:
        status = str(row.get("status") or "")
        status_counts[status] = status_counts.get(status, 0) + 1

    return {
        "ok": True,
        "phase": PHASE,
        "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "quantity_acceptance_enabled": False,
        "summary": {
            "answer_count": acceptance_summary.get("answer_count", len(acceptance.get("answer_rows") or [])),
            "candidate_count": acceptance_summary.get("candidate_count", len(acceptance.get("candidate_rows") or [])),
            "matched_three_fields_count": acceptance_summary.get("matched_three_fields_count", 0),
            "three_field_pass_rate": acceptance_summary.get("three_field_pass_rate", 0),
            "review_row_count": len(review_rows),
            "status_counts": status_counts,
            "quantity_status": v2_summary.get("quantity_status", "deferred_until_three_fields_accepted"),
        },
        "review_rows": review_rows,
    }


def write_three_field_human_review_outputs(
    report: Mapping[str, Any],
    output_dir: str | Path,
    *,
    stem: str | None = None,
) -> dict[str, str]:
    target = Path(output_dir)
    target.mkdir(parents=True, exist_ok=True)
    file_stem = stem or f"BIZ2x_PDF_three_field_human_review_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
    json_path = target / f"{file_stem}.json"
    csv_path = target / f"{file_stem}.csv"
    markdown_path = target / f"{file_stem}.md"
    xlsx_path = target / f"{file_stem}.xlsx"

    outputs = {
        "json": str(json_path),
        "csv": str(csv_path),
        "markdown": str(markdown_path),
        "xlsx": str(xlsx_path),
    }
    payload = {**dict(report), "outputs": outputs}
    json_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    _write_csv(csv_path, payload.get("review_rows") or [], REVIEW_HEADERS)
    markdown_path.write_text(_build_markdown(payload), encoding="utf-8")
    _write_workbook(xlsx_path, payload)
    return outputs


def _answer_rows_by_key(rows: Sequence[Mapping[str, Any]]) -> dict[tuple[str, str], Mapping[str, Any]]:
    result = {}
    for row in rows:
        result[(str(row.get("sheet_name") or ""), str(row.get("row_no") or ""))] = row
    return result


def _answer_key(row: Mapping[str, Any]) -> tuple[str, str]:
    return str(row.get("answer_sheet") or ""), str(row.get("answer_row_no") or "")


def _candidate_rows_by_no(rows: Sequence[Mapping[str, Any]]) -> dict[int, Mapping[str, Any]]:
    result = {}
    for row in rows:
        row_no = _int(row.get("row_no"), 0)
        if row_no:
            result[row_no] = row
    return result


def _human_row_for_candidate(
    candidate_no: int,
    human_rows: Sequence[Mapping[str, Any]],
    candidate: Mapping[str, Any],
) -> dict[str, Any]:
    if candidate_no > 0 and candidate_no <= len(human_rows):
        return dict(human_rows[candidate_no - 1])
    raw = str(candidate.get("raw") or "").strip()
    if raw:
        try:
            parsed = json.loads(raw)
            if isinstance(parsed, Mapping):
                return dict(parsed)
        except json.JSONDecodeError:
            return {}
    return {}


def _evidence_index(rows: Sequence[Mapping[str, Any]]) -> dict[str, dict[str, Any]]:
    result = {}
    for row in rows:
        evidence_id = str(row.get("evidence_id") or "").strip()
        if not evidence_id:
            continue
        result[evidence_id] = {
            "source_file": row.get("source_file", ""),
            "page": row.get("page", ""),
            "tile_id": row.get("tile_id", ""),
        }
    return result


def _evidence_ids(*rows: Mapping[str, Any]) -> list[str]:
    values: list[Any] = []
    for row in rows:
        values.extend(
            [
                row.get("evidence_ids"),
                (row.get("normalized_evidence") or {}).get("evidence_ids")
                if isinstance(row.get("normalized_evidence"), Mapping)
                else "",
            ]
        )
    result = []
    seen = set()
    for value in values:
        for item in _split_value(value):
            if item not in seen:
                seen.add(item)
                result.append(item)
    return result


def _split_value(value: Any) -> list[str]:
    if value is None:
        return []
    if isinstance(value, (list, tuple, set)):
        values = value
    else:
        values = str(value).replace(";", ",").split(",")
    return [str(item).strip() for item in values if str(item).strip()]


def _review_result(row: Mapping[str, Any]) -> str:
    status = str(row.get("status") or "")
    if status == "matched_three_fields":
        return "pass"
    if status == "missing_candidate":
        return "missing_candidate"
    if status == "unit_conflict":
        return "unit_conflict"
    if status == "matched_name_unit_feature_review":
        return "feature_review"
    return "manual_review"


def _match_label(value: Any, *, pass_threshold: float, review_threshold: float) -> str:
    score = _float(value, 0)
    if score >= pass_threshold:
        return "yes"
    if score >= review_threshold:
        return "review"
    return "no"


def _join_unique(values: Iterable[Any]) -> str:
    result = []
    seen = set()
    for value in values:
        text = str(value or "").strip()
        if not text or text in seen:
            continue
        seen.add(text)
        result.append(text)
    return ", ".join(result)


def _first_non_empty(*values: Any) -> str:
    for value in values:
        text = str(value or "").strip()
        if text:
            return text
    return ""


def _write_csv(path: Path, rows: Iterable[Mapping[str, Any]], fieldnames: Sequence[str]) -> None:
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        for row in rows:
            writer.writerow({field: _cell_value(row.get(field)) for field in fieldnames})


def _build_markdown(report: Mapping[str, Any]) -> str:
    summary = report.get("summary") or {}
    lines = [
        "# BIZ-2x PDF Three Field Human Review",
        "",
        f"- generated_at: {report.get('generated_at', '-')}",
        f"- answer_count: {summary.get('answer_count', 0)}",
        f"- candidate_count: {summary.get('candidate_count', 0)}",
        f"- matched_three_fields_count: {summary.get('matched_three_fields_count', 0)}",
        f"- quantity_status: {summary.get('quantity_status', '-')}",
        "",
        "## Review Rows",
        "",
        "| no | result | answer item | candidate item | name | feature | unit | evidence pages | issue |",
        "| ---: | --- | --- | --- | --- | --- | --- | --- | --- |",
    ]
    for row in (report.get("review_rows") or [])[:60]:
        lines.append(
            "| "
            + " | ".join(
                [
                    _md(row.get("review_no")),
                    _md(row.get("review_result")),
                    _md(row.get("answer_item_name")),
                    _md(row.get("candidate_item_name")),
                    _md(row.get("name_match")),
                    _md(row.get("feature_match")),
                    _md(row.get("unit_match")),
                    _md(row.get("evidence_pages")),
                    _md(row.get("issue")),
                ]
            )
            + " |"
        )
    return "\n".join(lines) + "\n"


def _write_workbook(path: Path, report: Mapping[str, Any]) -> None:
    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "summary"
    summary_rows = [[key, _cell_value(value)] for key, value in (report.get("summary") or {}).items()]
    _append_rows(summary_sheet, [SUMMARY_HEADERS, *summary_rows])
    _style_sheet(summary_sheet)

    review_sheet = workbook.create_sheet("three_field_review")
    _append_rows(
        review_sheet,
        [REVIEW_HEADERS, *[[_cell_value(row.get(header)) for header in REVIEW_HEADERS] for row in report.get("review_rows") or []]],
    )
    _style_sheet(review_sheet)
    workbook.save(path)


def _append_rows(sheet: Any, rows: Sequence[Sequence[Any]]) -> None:
    for row in rows:
        sheet.append(list(row))


def _style_sheet(sheet: Any) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    if sheet.max_row >= 1:
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for col_idx, column_cells in enumerate(sheet.columns, start=1):
        values = [str(cell.value or "") for cell in column_cells[:200]]
        width = min(max([len(value) for value in values] + [10]) + 2, 60)
        sheet.column_dimensions[get_column_letter(col_idx)].width = width


def _cell_value(value: Any) -> Any:
    if isinstance(value, (dict, list, tuple)):
        return json.dumps(value, ensure_ascii=False)
    return value


def _md(value: Any) -> str:
    return str(value or "").replace("|", "\\|").replace("\n", "<br>")


def _int(value: Any, default: int = 0) -> int:
    try:
        return int(float(str(value).strip()))
    except (TypeError, ValueError):
        return default


def _float(value: Any, default: float = 0.0) -> float:
    try:
        return float(str(value).strip())
    except (TypeError, ValueError):
        return default
