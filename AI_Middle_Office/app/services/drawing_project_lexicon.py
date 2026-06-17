from __future__ import annotations

import csv
import json
import re
import unicodedata
from collections import Counter
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from app.services.quantity_standard_library import QuantityStandardItem, QuantityStandardLibrary


BACKEND_ROOT = Path(__file__).resolve().parents[2]
WORKSPACE_ROOT = BACKEND_ROOT.parent
DEFAULT_PROJECT_LEXICON_PATH = BACKEND_ROOT / "data" / "drawing_project_lexicons" / "biz2x_sample_answer_lexicon.json"
DEFAULT_SAMPLE_ANSWER_XLSX = (
    WORKSPACE_ROOT
    / "AI识图练习图纸"
    / "AI识图练习图纸"
    / "信达公司职工食堂装修改造工程量清单.xlsx"
)

ITEM_ALIASES = ("项目名称", "清单项目名称", "项目", "工程内容", "名称")
FEATURE_ALIASES = ("项目特征", "项目特征描述", "特征描述", "特征", "规格", "规格/特征")
UNIT_ALIASES = ("计量单位", "单位")
QUANTITY_ALIASES = ("工程量", "数量", "工程数量")
CODE_ALIASES = ("项目编码", "编码", "清单编码")

LEXICON_CSV_HEADERS = [
    "词条编号",
    "人工行号",
    "分类",
    "人工项目名称",
    "识别强词",
    "材料编号",
    "标准范围",
    "标准项目编码",
    "标准项目名称",
    "单位",
    "工程量",
    "项目特征字段",
]

PROJECT_SIGNAL_TERMS = (
    "拆除",
    "铲除",
    "地砖",
    "地面",
    "地板",
    "门槛石",
    "美缝",
    "石材",
    "墙面",
    "墙砖",
    "墙布",
    "硬包",
    "木饰面",
    "隔墙",
    "隔断",
    "吊顶",
    "天花",
    "天棚",
    "灯槽",
    "踢脚",
    "线条",
    "窗帘盒",
    "窗台",
    "挡水条",
    "门套",
    "玻璃门",
    "实木门",
    "铝合金门",
    "洁具",
    "台盆",
    "马桶",
    "龙头",
    "地漏",
    "灯具",
    "开关",
    "插座",
    "管线",
    "保护",
    "运输",
    "保洁",
    "砌筑",
    "回填",
)

CATEGORY_LABELS = {
    "demolition": "拆除类",
    "floor": "地面类",
    "wall": "墙面/隔断类",
    "ceiling": "吊顶/天棚类",
    "waterproof": "防水防潮类",
    "baseboard": "踢脚线类",
    "linear_finish": "线性收口类",
    "door_window": "门窗/门套类",
    "sanitary": "洁具给排水类",
    "lighting_electrical": "灯具电气类",
    "measure": "措施/保洁运输类",
    "masonry_backfill": "砌筑回填类",
    "other": "其它待确认类",
}

SUPPLEMENTAL_CATEGORIES = {
    "demolition",
    "sanitary",
    "lighting_electrical",
    "measure",
}

MATERIAL_CODE_RE = re.compile(r"(?<![A-Z0-9])([A-Z]{2,4}\s*[-－]\s*\d{1,3})(?![A-Z0-9])", re.IGNORECASE)
TEXT_SPLITTER_RE = re.compile(r"[\s\-_—、，。；;:：|/\\()（）\[\]{}《》<>\"'“”‘’]+")


def extract_sample_answer_rows(path: str | Path) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    workbook = load_workbook(path, data_only=True)
    rows: list[dict[str, Any]] = []
    sheet_summaries: list[dict[str, Any]] = []
    for worksheet in workbook.worksheets:
        values = list(worksheet.iter_rows(values_only=True))
        header = _find_header(values)
        summary = {
            "sheet_name": worksheet.title,
            "max_row": worksheet.max_row,
            "max_column": worksheet.max_column,
            "header_found": bool(header),
            "header_row": header["row_index"] if header else "",
            "parsed_row_count": 0,
        }
        if not header:
            sheet_summaries.append(summary)
            continue
        for row_index, row in enumerate(values[header["row_index"] + 1 :], start=header["row_index"] + 2):
            raw_values = [_clean_text(value) for value in row]
            if not any(raw_values):
                continue
            item_name = _value_at(raw_values, header.get("item"))
            feature = _value_at(raw_values, header.get("feature"))
            unit = _value_at(raw_values, header.get("unit"))
            quantity = _value_at(raw_values, header.get("quantity"))
            item_code = _value_at(raw_values, header.get("code"))
            if not _looks_like_bill_row(item_name, feature, unit, quantity, raw_values):
                continue
            rows.append(
                {
                    "row_no": row_index,
                    "sheet_name": worksheet.title,
                    "item_code": item_code,
                    "item_name": item_name,
                    "feature": feature,
                    "unit": unit,
                    "quantity": quantity,
                    "raw_text": "；".join(value for value in raw_values if value),
                }
            )
            summary["parsed_row_count"] += 1
        sheet_summaries.append(summary)
    return rows, sheet_summaries


def build_project_lexicon_from_answer_rows(
    answer_rows: list[dict[str, Any]],
    *,
    library: QuantityStandardLibrary | None = None,
) -> dict[str, Any]:
    entries: list[dict[str, Any]] = []
    seen: set[tuple[str, str, str]] = set()
    for row in answer_rows:
        item_name = _clean_text(row.get("item_name"))
        feature = _clean_text(row.get("feature"))
        unit = _clean_text(row.get("unit"))
        if not item_name:
            continue
        category = classify_project_category(item_name, feature)
        material_codes = extract_material_codes(" ".join([item_name, feature, row.get("raw_text", "")]))
        standard = infer_standard_mapping(item_name, feature, category, library)
        scope = "gbt50854_candidate" if standard else "supplemental_or_scope_pending"
        if category in SUPPLEMENTAL_CATEGORIES and not standard:
            scope = "supplemental_or_other_specialty"
        strong_terms, weak_terms = build_recognition_terms(item_name, feature, category, material_codes)
        key = (_normalize(item_name), unit, "；".join(material_codes))
        if key in seen:
            continue
        seen.add(key)
        entries.append(
            {
                "entry_id": f"BIZ2xLEX-{len(entries) + 1:04d}",
                "source_row_no": row.get("row_no", ""),
                "source_sheet_name": row.get("sheet_name", ""),
                "source_item_code": row.get("item_code", ""),
                "manual_item_name": item_name,
                "manual_feature": feature,
                "manual_unit": unit,
                "manual_quantity": _clean_text(row.get("quantity")),
                "category": category,
                "category_label": CATEGORY_LABELS.get(category, CATEGORY_LABELS["other"]),
                "material_codes": material_codes,
                "strong_terms": strong_terms,
                "weak_terms": weak_terms,
                "standard_scope": scope,
                "standard_item_code": standard.get("item_code", ""),
                "standard_item_name": standard.get("item_name", ""),
                "standard_chapter_name": standard.get("chapter_name", ""),
                "standard_unit_options": standard.get("unit_options", []),
                "standard_feature_fields": standard.get("feature_fields", []),
                "standard_quantity_rule_text": standard.get("quantity_rule_text", ""),
                "mapping_reason": standard.get("mapping_reason", ""),
                "recognition_priority": _recognition_priority(category),
            }
        )

    category_counts = Counter(entry["category"] for entry in entries)
    scope_counts = Counter(entry["standard_scope"] for entry in entries)
    return {
        "ok": True,
        "phase": "BIZ-2x-R1-1-sample-answer-project-lexicon",
        "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "source": {
            "manual_answer_row_count": len(answer_rows),
            "policy": "answer_driven_regression_lexicon_for_dwg_project_recognition",
        },
        "summary": {
            "manual_answer_row_count": len(answer_rows),
            "lexicon_entry_count": len(entries),
            "category_counts": dict(category_counts.most_common()),
            "standard_scope_counts": dict(scope_counts.most_common()),
            "gbt_candidate_count": sum(1 for entry in entries if entry.get("standard_item_code")),
            "supplemental_or_pending_count": sum(1 for entry in entries if not entry.get("standard_item_code")),
        },
        "entries": entries,
        "notes": [
            "本词库来自样例人工四字段清单，用于 R1 项目识别覆盖增强。",
            "有标准项目编码的词条可进入 GB/T 字段口径候选；无标准编码的词条只作为补充/待映射候选，不直接生成最终工程量。",
            "工程量仍必须在 CAD 区域/构件绑定后按标准库工程量规则计算。",
        ],
    }


def load_project_lexicon(path: str | Path | None = None) -> dict[str, Any]:
    target = Path(path or DEFAULT_PROJECT_LEXICON_PATH)
    if not target.exists():
        return {"ok": False, "summary": {"lexicon_entry_count": 0}, "entries": []}
    return json.loads(target.read_text(encoding="utf-8"))


def match_source_signal_to_lexicon(
    signal: dict[str, Any],
    lexicon: dict[str, Any],
    *,
    limit: int = 3,
    min_score: float = 0.58,
) -> list[dict[str, Any]]:
    text = _clean_text(signal.get("evidence_text")) or _clean_text(signal.get("raw_row_text")) or _clean_text(signal.get("source_name"))
    source_name = _clean_text(signal.get("source_name"))
    if not text and not source_name:
        return []
    scored: list[dict[str, Any]] = []
    for entry in lexicon.get("entries") or []:
        score, reasons = _lexicon_match_score(entry, source_name, text)
        if score < min_score:
            continue
        scored.append({"entry": entry, "score": round(score, 3), "reasons": reasons})
    scored.sort(key=lambda item: (item["score"], item["entry"].get("recognition_priority", 0)), reverse=True)
    return scored[:limit]


def classify_project_category(item_name: str, feature: str = "") -> str:
    text = f"{item_name} {feature}"
    if any(term in text for term in ("拆除", "铲除")):
        return "demolition"
    if any(term in text for term in ("吊顶", "天花", "天棚", "灯槽", "矿棉板", "铝扣板", "石膏板天花")):
        return "ceiling"
    if "防水石膏板" not in text and any(term in text for term in ("防水", "防潮")):
        return "waterproof"
    if "踢脚" in text:
        return "baseboard"
    if any(term in text for term in ("窗帘盒", "窗台", "挡水条", "金属线条")):
        return "linear_finish"
    if any(term in text for term in ("门套", "实木门", "木门", "玻璃门", "铝合金门", "售卖窗口", "窗")):
        return "door_window"
    if any(term in text for term in ("台盆", "马桶", "龙头", "厕纸架", "地漏", "洗手台", "洁具")):
        return "sanitary"
    if any(term in text for term in ("灯具", "开关", "插座", "管线", "配电", "电气")):
        return "lighting_electrical"
    if any(term in text for term in ("地面", "地砖", "地板", "门槛石", "美缝", "保护层")):
        return "floor"
    if any(term in text for term in ("墙面", "墙砖", "墙布", "硬包", "木饰面", "包柱", "隔墙", "隔断", "玻璃隔墙", "石材湿贴")):
        return "wall"
    if any(term in text for term in ("保洁", "保护", "运输", "成品保护")):
        return "measure"
    if any(term in text for term in ("砌筑", "回填")):
        return "masonry_backfill"
    return "other"


def extract_material_codes(text: str) -> list[str]:
    values = []
    for match in MATERIAL_CODE_RE.finditer(unicodedata.normalize("NFKC", text or "")):
        values.append(match.group(0).upper().replace("－", "-").replace(" ", ""))
    return _dedupe(values)


def build_recognition_terms(
    item_name: str,
    feature: str,
    category: str,
    material_codes: list[str],
) -> tuple[list[str], list[str]]:
    cleaned_name = _clean_text(item_name)
    name_without_code = _clean_text(MATERIAL_CODE_RE.sub("", unicodedata.normalize("NFKC", cleaned_name)))
    strong = [cleaned_name, name_without_code, *material_codes]
    weak: list[str] = []
    for token in TEXT_SPLITTER_RE.split(f"{name_without_code} {feature}"):
        token = _clean_text(token)
        if len(_normalize(token)) >= 2 and token not in {"供货", "安装", "处理", "成品", "施工", "项目"}:
            weak.append(token)
    if category == "demolition":
        weak.extend(["拆除", name_without_code.replace("拆除", ""), f"{name_without_code.replace('拆除', '')}拆除"])
    if category == "floor":
        weak.extend(["地面", "地砖", "楼地面", "门槛石", "石材地面", "瓷砖地面"])
    if category == "wall":
        weak.extend(["墙面", "墙砖", "隔墙", "隔断", "墙面石材", "墙面瓷砖"])
    if category == "ceiling":
        weak.extend(["吊顶", "天花", "天棚", "灯槽", "石膏板吊顶", "铝扣板吊顶"])
    if category == "sanitary":
        weak.extend(["洁具", "台盆", "马桶", "龙头", "地漏"])
    return _dedupe([item for item in strong if _clean_text(item)]), _dedupe([item for item in weak if _clean_text(item)])


def infer_standard_mapping(
    item_name: str,
    feature: str,
    category: str,
    library: QuantityStandardLibrary | None,
) -> dict[str, Any]:
    if library is None:
        return {}
    text = f"{item_name} {feature}"
    preferred_names: list[str] = []
    reason = ""
    if category == "floor":
        if "石材" in text or "门槛石" in text:
            preferred_names = ["石材楼地面", "石材零星项目"]
            reason = "样例地面石材/门槛石按石材楼地面或石材零星项目候选"
        elif any(term in text for term in ("瓷砖", "地砖", "块料", "美缝")):
            preferred_names = ["块料楼地面"]
            reason = "样例瓷砖/地砖/美缝地面按块料楼地面候选"
    elif category == "wall":
        if any(term in text for term in ("墙砖", "瓷砖", "块料")):
            preferred_names = ["块料墙、柱面"]
            reason = "样例墙砖/瓷砖墙面按块料墙、柱面候选"
        elif any(term in text for term in ("石材", "人造石")):
            preferred_names = ["石材墙、柱面", "石材零星项目"]
            reason = "样例石材/人造石墙面按石材墙、柱面候选"
        elif any(term in text for term in ("隔墙", "隔断", "玻璃隔墙", "淋浴隔断")):
            preferred_names = ["成品隔断", "轻质隔断", "轻质隔墙"]
            reason = "样例隔墙/隔断按隔断类标准候选"
        elif any(term in text for term in ("墙布", "硬包")):
            preferred_names = ["墙纸裱糊", "织锦缎裱糊"]
            reason = "样例墙布/硬包按裱糊类候选，需 R2 复核"
        elif "涂料" in text:
            preferred_names = ["墙面喷刷涂料"]
            reason = "样例墙面涂料按墙面喷刷涂料候选"
    elif category == "ceiling":
        if "涂料" in text or "乳胶漆" in text or "无机涂料" in text:
            preferred_names = ["天棚喷刷涂料"]
            reason = "样例天棚涂料按天棚喷刷涂料候选"
        elif "造型" in text or "灯槽" in text:
            preferred_names = ["艺术造型吊顶天棚", "艺术造型 | 吊顶天棚"]
            reason = "样例造型吊顶/灯槽按艺术造型吊顶天棚候选"
        else:
            preferred_names = ["平面吊顶天棚", "平面吊顶 | 天棚", "格栅吊顶"]
            reason = "样例普通吊顶/天花按平面吊顶天棚候选"
    elif category == "waterproof":
        if "墙" in text:
            preferred_names = ["墙面涂膜防水"]
            reason = "样例墙面防水按墙面涂膜防水候选"
        else:
            preferred_names = ["楼(地)面涂膜防水", "楼(地) 面涂膜防水"]
            reason = "样例地面防水按楼地面涂膜防水候选"
    elif category == "baseboard":
        preferred_names = ["金属踢脚线", "块料踢脚线"]
        reason = "样例踢脚线按踢脚线类标准候选"
    elif category == "linear_finish":
        if "窗帘盒" in text:
            preferred_names = ["窗帘盒"]
            reason = "样例窗帘盒按窗帘盒候选"
        elif "窗台" in text:
            preferred_names = ["窗台板"]
            reason = "样例窗台石/窗台板按窗台板候选"
    elif category == "door_window":
        if "实木" in text or "木门" in text:
            preferred_names = ["木质门", "木门油漆"]
            reason = "样例实木门按木质门候选，需 R2 复核成品门口径"
        elif any(term in text for term in ("玻璃门", "铝合金门", "不锈钢")):
            preferred_names = ["金属(塑钢)门", "全玻自由门", "金属门油漆"]
            reason = "样例金属/玻璃门按门类候选，需 R2 复核成品门口径"
    elif category == "masonry_backfill":
        if "砌筑" in text:
            preferred_names = ["零星砌砖", "砖砌体"]
            reason = "样例零星砌筑按砌体类候选"
        elif "回填" in text:
            preferred_names = ["回填方"]
            reason = "样例陶粒回填按回填类候选，需 R2 复核材料口径"
    item = _find_standard_item(library, preferred_names)
    if not item:
        return {}
    return _standard_item_payload(item, reason)


def write_project_lexicon_outputs(
    lexicon: dict[str, Any],
    output_dir: str | Path,
    *,
    stem: str | None = None,
) -> dict[str, str]:
    directory = Path(output_dir)
    directory.mkdir(parents=True, exist_ok=True)
    file_stem = stem or f"BIZ2x_R1_项目识别词库_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
    json_path = directory / f"{file_stem}.json"
    csv_path = directory / f"{file_stem}.csv"
    markdown_path = directory / f"{file_stem}.md"
    json_path.write_text(json.dumps(lexicon, ensure_ascii=False, indent=2), encoding="utf-8")
    _write_lexicon_csv(csv_path, lexicon.get("entries") or [])
    markdown_path.write_text(build_project_lexicon_markdown(lexicon), encoding="utf-8")
    return {"json": str(json_path), "csv": str(csv_path), "markdown": str(markdown_path)}


def build_project_lexicon_markdown(lexicon: dict[str, Any]) -> str:
    summary = lexicon.get("summary") or {}
    lines = [
        "# BIZ-2x R1-1 项目识别词库",
        "",
        f"- 生成时间：{lexicon.get('generated_at', '-')}",
        f"- 人工清单行数：{summary.get('manual_answer_row_count', 0)}",
        f"- 词库条目数：{summary.get('lexicon_entry_count', 0)}",
        f"- GB/T 候选映射数：{summary.get('gbt_candidate_count', 0)}",
        f"- 补充/待映射数：{summary.get('supplemental_or_pending_count', 0)}",
        f"- 分类分布：{summary.get('category_counts', {})}",
        "",
        "## 词库节选",
        "",
        "| 编号 | 分类 | 人工项目 | 标准映射 | 单位 | 强识别词 |",
        "| --- | --- | --- | --- | --- | --- |",
    ]
    for entry in (lexicon.get("entries") or [])[:120]:
        standard = entry.get("standard_item_name") or entry.get("standard_scope", "")
        lines.append(
            "| "
            + " | ".join(
                [
                    _md(entry.get("entry_id")),
                    _md(entry.get("category_label")),
                    _md(entry.get("manual_item_name")),
                    _md(standard),
                    _md(entry.get("manual_unit")),
                    _md("、".join(entry.get("strong_terms") or [])),
                ]
            )
            + " |"
        )
    lines.extend(
        [
            "",
            "## 使用边界",
            "",
            "- 本词库只增强 DWG 项目识别覆盖，不单独生成最终工程量。",
            "- 无标准项目编码的词条只进入补充/待映射候选，后续必须经过 R2 标准映射和 R3/R4 算量证据确认。",
        ]
    )
    return "\n".join(lines) + "\n"


def _lexicon_match_score(entry: dict[str, Any], source_name: str, evidence_text: str) -> tuple[float, list[str]]:
    source_norm = _normalize(" ".join([source_name, evidence_text]))
    source_raw = " ".join([source_name, evidence_text])
    if not source_norm:
        return 0.0, []
    score = 0.0
    reasons: list[str] = []
    for term in entry.get("strong_terms") or []:
        term_norm = _normalize(term)
        if not term_norm:
            continue
        if term_norm in source_norm:
            score = max(score, 0.86)
            reasons.append(f"命中强识别词：{term}")
    material_hits = [code for code in entry.get("material_codes") or [] if code and code in source_raw.upper()]
    if material_hits:
        score = max(score, 0.82)
        reasons.append("命中材料编号：" + "、".join(material_hits))
    weak_hits = []
    for term in entry.get("weak_terms") or []:
        term_norm = _normalize(term)
        if len(term_norm) < 2:
            continue
        if term_norm in source_norm:
            weak_hits.append(term)
    if weak_hits:
        score = max(score, min(0.78, 0.46 + len(set(weak_hits)) * 0.08))
        reasons.append("命中弱识别词：" + "、".join(_dedupe(weak_hits)[:5]))
    if entry.get("category") in {"demolition", "sanitary", "lighting_electrical"} and any(
        term in source_raw for term in ("拆除", "台盆", "马桶", "灯具", "开关", "地漏", "龙头")
    ):
        score += 0.05
    return min(score, 0.98), _dedupe(reasons)


def _find_standard_item(library: QuantityStandardLibrary, preferred_names: list[str]) -> QuantityStandardItem | None:
    if not preferred_names:
        return None
    active_items = [item for item in library.items if item.status == "active"]
    for name in preferred_names:
        name_norm = _normalize(name)
        for item in active_items:
            item_norm = _normalize(item.item_name)
            if name_norm and (name_norm == item_norm or name_norm in item_norm or item_norm in name_norm):
                return item
    return None


def _standard_item_payload(item: QuantityStandardItem, reason: str) -> dict[str, Any]:
    return {
        "item_code": item.item_code,
        "item_name": item.item_name,
        "chapter_name": item.chapter_name,
        "unit_options": list(item.unit_options),
        "feature_fields": item.feature_names,
        "quantity_rule_text": _clean_text(item.quantity_rule.get("rule_text")),
        "quantity_formula_type": _clean_text(item.quantity_rule.get("formula_type")),
        "mapping_reason": reason,
    }


def _find_header(values: list[tuple[Any, ...]]) -> dict[str, int] | None:
    best: dict[str, int] | None = None
    best_score = 0
    for row_index, row in enumerate(values[:80]):
        cells = [_clean_text(value) for value in row]
        mapping = {
            "item": _find_alias_column(cells, ITEM_ALIASES),
            "feature": _find_alias_column(cells, FEATURE_ALIASES),
            "unit": _find_alias_column(cells, UNIT_ALIASES),
            "quantity": _find_alias_column(cells, QUANTITY_ALIASES),
            "code": _find_alias_column(cells, CODE_ALIASES),
        }
        score = sum(1 for key in ("item", "feature", "unit", "quantity") if mapping[key] is not None)
        if score > best_score:
            best_score = score
            best = {"row_index": row_index, **{key: value for key, value in mapping.items() if value is not None}}
        if score >= 4:
            return best
    if best_score >= 3 and best:
        return best
    return None


def _find_alias_column(cells: list[str], aliases: tuple[str, ...]) -> int | None:
    for index, cell in enumerate(cells):
        normalized = _normalize_header(cell)
        if not normalized:
            continue
        for alias in aliases:
            if _normalize_header(alias) == normalized or _normalize_header(alias) in normalized:
                return index
    return None


def _looks_like_bill_row(item_name: str, feature: str, unit: str, quantity: str, raw_values: list[str]) -> bool:
    raw_text = "".join(raw_values)
    if not item_name and not feature:
        return False
    if any(term in raw_text for term in ("合计", "小计", "总价", "综合单价", "暂列金额", "措施项目")) and not unit:
        return False
    if item_name in {"项目名称", "清单项目名称", "名称"}:
        return False
    return bool(item_name and (unit or quantity or feature))


def _write_lexicon_csv(path: Path, entries: list[dict[str, Any]]) -> None:
    with path.open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.DictWriter(handle, fieldnames=LEXICON_CSV_HEADERS)
        writer.writeheader()
        for entry in entries:
            writer.writerow(
                {
                    "词条编号": entry.get("entry_id", ""),
                    "人工行号": entry.get("source_row_no", ""),
                    "分类": entry.get("category_label", ""),
                    "人工项目名称": entry.get("manual_item_name", ""),
                    "识别强词": "；".join(entry.get("strong_terms") or []),
                    "材料编号": "；".join(entry.get("material_codes") or []),
                    "标准范围": entry.get("standard_scope", ""),
                    "标准项目编码": entry.get("standard_item_code", ""),
                    "标准项目名称": entry.get("standard_item_name", ""),
                    "单位": entry.get("manual_unit", ""),
                    "工程量": entry.get("manual_quantity", ""),
                    "项目特征字段": "；".join(entry.get("standard_feature_fields") or []),
                }
            )


def _recognition_priority(category: str) -> int:
    priorities = {
        "demolition": 90,
        "floor": 86,
        "wall": 84,
        "ceiling": 84,
        "waterproof": 82,
        "baseboard": 78,
        "linear_finish": 76,
        "door_window": 72,
        "sanitary": 68,
        "lighting_electrical": 66,
        "measure": 58,
    }
    return priorities.get(category, 50)


def _value_at(values: list[str], index: int | None) -> str:
    if index is None or index >= len(values):
        return ""
    return values[index]


def _normalize_header(value: str) -> str:
    return re.sub(r"[\s:：()（）\[\]【】、，,；;|]+", "", _clean_text(value).lower())


def _normalize(value: Any) -> str:
    text = unicodedata.normalize("NFKC", _clean_text(value)).lower()
    return TEXT_SPLITTER_RE.sub("", text)


def _clean_text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).replace("\r", "\n")
    text = re.sub(r"[ \t]+", " ", text)
    text = re.sub(r"\n{2,}", "\n", text)
    return text.strip()


def _dedupe(values: list[str]) -> list[str]:
    result: list[str] = []
    seen: set[str] = set()
    for value in values:
        cleaned = _clean_text(value)
        key = _normalize(cleaned)
        if not cleaned or key in seen:
            continue
        seen.add(key)
        result.append(cleaned)
    return result


def _md(value: Any) -> str:
    return str(value if value is not None else "").replace("|", "\\|").replace("\n", "<br>")
