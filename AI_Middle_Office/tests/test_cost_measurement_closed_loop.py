import uuid
from io import BytesIO

from openpyxl import Workbook, load_workbook

from app.core.config import settings
from app.core.database import SessionLocal
from app.core.security import get_password_hash
from app.models.cost_item import COST_STATUS_DRAFT, CostItem
from app.models.user import User, UserRole
from app.services.cost_measurement import CostMeasurementImportError, parse_cost_measurement_workbook


PASSWORD = "secret123"


def _set_flag(name: str, value):
    old = getattr(settings, name)
    object.__setattr__(settings, name, value)
    return old


def _user(role: str = "admin") -> User:
    db = SessionLocal()
    try:
        user = User(
            username=f"cost_measure_{role}_{uuid.uuid4().hex[:8]}",
            hashed_password=get_password_hash(PASSWORD),
            role="admin" if role in {"admin", "system_admin"} else "none",
            role_version=1,
            quota=20,
        )
        db.add(user)
        db.flush()
        db.add(UserRole(user_id=user.id, role=role, created_by=None, note="cost measurement test"))
        db.commit()
        db.refresh(user)
        return user
    finally:
        db.close()


def _headers(client, role: str = "admin") -> dict:
    user = _user(role)
    response = client.post("/api/v1/auth/login", data={"username": user.username, "password": PASSWORD})
    assert response.status_code == 200
    return {"Authorization": f"Bearer {response.json()['access_token']}"}


def _workbook_bytes() -> bytes:
    workbook = Workbook()
    summary = workbook.active
    summary.title = "\u6c47\u603b\u8868"
    summary.append(["\u9879\u76ee\u603b\u4ef7", 1000])
    summary.append(["\u7a0e\u91d1", 90])

    sheet = workbook.create_sheet("\u88c5\u4fee\u5de5\u7a0b\u91cf\u6e05\u5355")
    sheet.append(["\u88c5\u4fee\u90e8\u5206\u5de5\u7a0b\u91cf\u62a5\u4ef7\u6e05\u5355"])
    sheet.append(["\u5de5\u7a0b\u540d\u79f0\uff1a\u6210\u672c\u6d4b\u7b97\u6d4b\u8bd5\u9879\u76ee"])
    sheet.append([
        "\u5e8f\u53f7", "\u9879\u76ee\u540d\u79f0", "\u9879\u76ee\u7279\u5f81", "\u5355\u4f4d", "\u5de5\u7a0b\u91cf",
        "\u4eba\u5de5\u8d39", "\u4e3b\u6750\u8d39", "\u635f\u8017\u7cfb\u6570", "\u8f85\u6750\u53ca\u673a\u68b0\u8d39",
        "\u7ba1\u7406\u8d39", "\u5229\u6da6", "\u7efc\u5408\u5355\u4ef7", "\u5408\u8ba1",
    ])
    sheet.append([None, None, None, None, None, None, None, None, None, None, None, None, None])
    sheet.append([None, "\u4e00\u3001\u88c5\u4fee\u90e8\u5206", None, None, None, None, None, None, None, 0.03, 0.05, None, None])
    sheet.append([1, "\u6807\u51c6\u7ec4\u4ef7\u9879", "\u6d4b\u8bd5\u505a\u6cd5", "m2", 2, 10, 20, 1.05, 5, 1.08, 1.854, 38.934, 77.868])
    sheet.append([2, "\u5386\u53f2\u516c\u5f0f\u5f02\u5e38\u9879", "\u6d4b\u8bd5\u505a\u6cd5", "m2", 1, 10, 20, 1.05, 5, 2, 3, 60, 60])
    sheet.append([None, "\u63aa\u65bd\u9879\u76ee", None, None, None, None, None, None, None, None, None, None, None])
    sheet.append([3, "\u6210\u54c1\u4fdd\u62a4", None, "\u9879", 1, None, None, None, None, None, None, 100, 100])

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def _workbook_with_duplicate_bytes() -> bytes:
    workbook = load_workbook(BytesIO(_workbook_bytes()))
    sheet = workbook["\u88c5\u4fee\u5de5\u7a0b\u91cf\u6e05\u5355"]
    sheet["B6"] = "\u91cd\u590d\u5019\u9009\u9879"
    sheet["B7"] = "\u5386\u53f2\u516c\u5f0f\u5f02\u5e38\u9879-\u91cd\u590d\u6d4b\u8bd5"
    sheet["B9"] = "\u6210\u54c1\u4fdd\u62a4-\u91cd\u590d\u6d4b\u8bd5"
    sheet.append([4, "\u91cd\u590d\u5019\u9009\u9879", "\u6d4b\u8bd5\u505a\u6cd5", "m2", 1, 11, 20, 1.05, 5, 1.11, 1.9055, 40.0155, 40.0155])
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def test_parser_normalizes_rates_and_flags_formula_variance():
    parsed = parse_cost_measurement_workbook("measurement.xlsx", _workbook_bytes())

    assert parsed["project_name"] == "\u6210\u672c\u6d4b\u7b97\u6d4b\u8bd5\u9879\u76ee"
    assert parsed["line_count"] == 3
    assert parsed["management_rate"] == 0.03
    assert parsed["profit_rate"] == 0.05
    assert parsed["tax_rate"] == 0.09
    assert parsed["lines"][0]["material_loss_rate"] == 0.05
    assert round(parsed["lines"][0]["calculated_unit_price"], 3) == 38.934
    assert parsed["lines"][0]["review_status"] == "ready"
    assert {row["code"] for row in parsed["lines"][1]["warnings"]} == {"SOURCE_FORMULA_VARIANCE"}
    assert {row["code"] for row in parsed["lines"][2]["warnings"]} == {"COMPOSITE_ONLY"}


def test_parser_rejects_legacy_xls():
    try:
        parse_cost_measurement_workbook("legacy.xls", b"legacy")
    except CostMeasurementImportError as exc:
        assert ".xlsx" in str(exc)
    else:
        raise AssertionError("legacy .xls should be rejected")


def test_cost_measurement_api_import_lock_and_export(client):
    admin_headers = _headers(client, "admin")
    viewer_headers = _headers(client, "cost_viewer")
    old_flag = _set_flag("feature_cost_measurement", True)
    try:
        preview = client.post(
            "/api/v1/admin/cost-measurements/import-preview",
            headers=viewer_headers,
            files={"file": ("measurement.xlsx", _workbook_bytes(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
        assert preview.status_code == 200, preview.text
        assert preview.json()["data"]["line_count"] == 3

        imported = client.post(
            "/api/v1/admin/cost-measurements/import",
            headers=admin_headers,
            data={"name": "\u6d4b\u8bd5\u6210\u672c\u6d4b\u7b97"},
            files={"file": ("measurement.xlsx", _workbook_bytes(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
        assert imported.status_code == 200, imported.text
        measurement = imported.json()["data"]
        assert measurement["status"] == "draft"
        assert measurement["line_count"] == 3
        assert measurement["review_line_count"] == 2

        draft_preview_blocked = client.post(
            f"/api/v1/admin/cost-measurements/{measurement['id']}/cost-drafts/preview",
            headers=admin_headers,
            json={},
        )
        assert draft_preview_blocked.status_code == 409
        assert draft_preview_blocked.json()["detail"] == "MEASUREMENT_MUST_BE_LOCKED"

        blocked = client.post(
            f"/api/v1/admin/cost-measurements/{measurement['id']}/lock",
            headers=admin_headers,
            json={"note": ""},
        )
        assert blocked.status_code == 409
        assert blocked.json()["detail"] == "MEASUREMENT_REVIEW_NOTE_REQUIRED"

        locked = client.post(
            f"/api/v1/admin/cost-measurements/{measurement['id']}/lock",
            headers=admin_headers,
            json={"note": "\u5df2\u4eba\u5de5\u590d\u6838\u5386\u53f2\u5f02\u5e38\u548c\u63aa\u65bd\u9879"},
        )
        assert locked.status_code == 200, locked.text
        assert locked.json()["data"]["status"] == "locked"

        exported = client.get(
            f"/api/v1/admin/cost-measurements/{measurement['id']}/export",
            headers=admin_headers,
        )
        assert exported.status_code == 200
        assert exported.content[:2] == b"PK"
    finally:
        _set_flag("feature_cost_measurement", old_flag)


def test_locked_measurement_can_create_cost_db_drafts_without_overwrite(client):
    admin_headers = _headers(client, "admin")
    old_measurement_flag = _set_flag("feature_cost_measurement", True)
    old_cost_db_flag = _set_flag("feature_cost_db", True)
    try:
        imported = client.post(
            "/api/v1/admin/cost-measurements/import",
            headers=admin_headers,
            data={"name": "\u5386\u53f2\u6d4b\u7b97\u6c89\u6dc0\u6d4b\u8bd5"},
            files={"file": ("measurement.xlsx", _workbook_bytes(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
        assert imported.status_code == 200, imported.text
        measurement = imported.json()["data"]
        for line in measurement["lines"]:
            if line["review_status"] != "required":
                continue
            reviewed = client.patch(
                f"/api/v1/admin/cost-measurements/{measurement['id']}/lines/{line['id']}",
                headers=admin_headers,
                json={"review_status": "reviewed"},
            )
            assert reviewed.status_code == 200, reviewed.text

        locked = client.post(
            f"/api/v1/admin/cost-measurements/{measurement['id']}/lock",
            headers=admin_headers,
            json={"note": "\u5df2\u9010\u884c\u590d\u6838\u5e76\u786e\u8ba4\u5386\u53f2\u4ef7\u683c\u53e3\u5f84"},
        )
        assert locked.status_code == 200, locked.text
        locked_data = locked.json()["data"]
        assert locked_data["status"] == "locked"
        assert locked_data["review_line_count"] == 0

        preview = client.post(
            f"/api/v1/admin/cost-measurements/{measurement['id']}/cost-drafts/preview",
            headers=admin_headers,
            json={},
        )
        assert preview.status_code == 200, preview.text
        preview_data = preview.json()["data"]
        assert preview_data["summary"]["selected_line_count"] == 3
        assert preview_data["summary"]["eligible_count"] == 3
        assert preview_data["summary"]["blocked_count"] == 0
        first = preview_data["candidates"][0]
        assert first["candidate_status"] == "ready"
        assert round(first["cost_item_payload"]["price"], 3) == 38.934
        assert round(first["cost_item_payload"]["client_direct_fee"], 3) == 36.0
        assert round(first["cost_item_payload"]["client_main_material_price"], 3) == 21.0
        assert round(first["cost_item_payload"]["client_management_profit"], 3) == 2.934

        line_ids = [row["line_id"] for row in preview_data["candidates"]]
        created = client.post(
            f"/api/v1/admin/cost-measurements/{measurement['id']}/cost-drafts",
            headers=admin_headers,
            json={"line_ids": line_ids, "note": "\u4f5c\u4e3a\u9996\u6279\u5386\u53f2\u6210\u672c\u5e93\u5019\u9009"},
        )
        assert created.status_code == 200, created.text
        result = created.json()["data"]
        assert result["created_count"] == 3
        assert result["skipped_count"] == 0

        db = SessionLocal()
        try:
            rows = db.query(CostItem).filter(CostItem.id.in_(result["created_items"][index]["cost_item_id"] for index in range(3))).all()
            assert len(rows) == 3
            assert all(row.status == COST_STATUS_DRAFT for row in rows)
            assert all(row.source == "imported" for row in rows)
            standard = next(row for row in rows if row.item_name == "\u6807\u51c6\u7ec4\u4ef7\u9879")
            assert round(standard.price, 3) == 38.934
            assert round(standard.client_main_material_price, 3) == 21.0
            assert "measurement_code:" in (standard.notes or "")
            composite = next(row for row in rows if row.item_name == "\u6210\u54c1\u4fdd\u62a4")
            assert composite.subcontract_composite_price is None
            assert "subcontract_is_composite_proxy: true" in (composite.notes or "")
        finally:
            db.close()

        repeated = client.post(
            f"/api/v1/admin/cost-measurements/{measurement['id']}/cost-drafts",
            headers=admin_headers,
            json={"line_ids": line_ids, "note": "\u91cd\u590d\u63d0\u4ea4\u5e94\u8df3\u8fc7"},
        )
        assert repeated.status_code == 200, repeated.text
        repeated_result = repeated.json()["data"]
        assert repeated_result["created_count"] == 0
        assert repeated_result["skipped_count"] == 3
        assert {row["reason_code"] for row in repeated_result["skipped"]} == {"DRAFT_DUPLICATE"}

        detail = client.get(
            f"/api/v1/admin/cost-measurements/{measurement['id']}",
            headers=admin_headers,
        )
        assert detail.status_code == 200
        event_types = [row["event_type"] for row in detail.json()["data"]["events"]]
        assert "cost_drafts_created" in event_types
    finally:
        _set_flag("feature_cost_measurement", old_measurement_flag)
        _set_flag("feature_cost_db", old_cost_db_flag)


def test_duplicate_measurement_candidates_require_one_choice(client):
    admin_headers = _headers(client, "admin")
    old_measurement_flag = _set_flag("feature_cost_measurement", True)
    old_cost_db_flag = _set_flag("feature_cost_db", True)
    try:
        imported = client.post(
            "/api/v1/admin/cost-measurements/import",
            headers=admin_headers,
            data={"name": "\u91cd\u590d\u5019\u9009\u6d4b\u8bd5"},
            files={"file": ("measurement-duplicate.xlsx", _workbook_with_duplicate_bytes(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
        assert imported.status_code == 200, imported.text
        measurement = imported.json()["data"]
        for line in measurement["lines"]:
            if line["review_status"] == "required":
                response = client.patch(
                    f"/api/v1/admin/cost-measurements/{measurement['id']}/lines/{line['id']}",
                    headers=admin_headers,
                    json={"review_status": "reviewed"},
                )
                assert response.status_code == 200, response.text
        locked = client.post(
            f"/api/v1/admin/cost-measurements/{measurement['id']}/lock",
            headers=admin_headers,
            json={"note": "\u5df2\u786e\u8ba4\u91cd\u590d\u5019\u9009\u9700\u4eba\u5de5\u4e8c\u9009\u4e00"},
        )
        assert locked.status_code == 200, locked.text

        preview = client.post(
            f"/api/v1/admin/cost-measurements/{measurement['id']}/cost-drafts/preview",
            headers=admin_headers,
            json={},
        )
        assert preview.status_code == 200, preview.text
        data = preview.json()["data"]
        assert data["summary"]["selected_line_count"] == 4
        assert data["summary"]["eligible_count"] == 3
        assert data["summary"]["within_measurement_duplicate_count"] == 1
        duplicate_group = [row for row in data["candidates"] if row["item_name"] == "\u91cd\u590d\u5019\u9009\u9879"]
        assert len(duplicate_group) == 2
        assert duplicate_group[1]["candidate_status"] == "duplicate_within_measurement"
        assert duplicate_group[1]["can_create"] is True

        created = client.post(
            f"/api/v1/admin/cost-measurements/{measurement['id']}/cost-drafts",
            headers=admin_headers,
            json={"line_ids": [row["line_id"] for row in duplicate_group]},
        )
        assert created.status_code == 200, created.text
        result = created.json()["data"]
        assert result["created_count"] == 1
        assert result["skipped_count"] == 1
        assert result["skipped"][0]["reason_code"] == "DUPLICATE_WITHIN_MEASUREMENT"
    finally:
        _set_flag("feature_cost_measurement", old_measurement_flag)
        _set_flag("feature_cost_db", old_cost_db_flag)
