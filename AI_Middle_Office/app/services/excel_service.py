import base64
import logging
import re
from io import BytesIO

import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


logger = logging.getLogger(__name__)


def _cjk_len(s: str) -> int:
    """计算字符串显示宽度：CJK 字符按 2 计，其余按 1 计"""
    width = 0
    for ch in str(s):
        code = ord(ch)
        if (0x4E00 <= code <= 0x9FFF) or (0x3400 <= code <= 0x4DBF) or (0xFF01 <= code <= 0xFFEE):
            width += 2
        else:
            width += 1
    return width


def _parse_amount(value) -> float:
    if value in (None, ""):
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).replace(",", "")
    match = re.search(r"-?\d+(?:\.\d+)?", text)
    if not match:
        return 0.0
    try:
        return float(match.group(0))
    except ValueError:
        return 0.0


def build_excel_base64(project_details: list) -> str:
    """用 openpyxl 生成自适应列宽的 Excel，返回 base64 字符串"""
    try:
        headers = ["施工项目", "AI核准单价(元)", "项目合计(元)", "工艺备注"]
        field_map = ["project_name", "unit_price", "total_price", "notes"]

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "报价单"
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True)

        total_amount = 0.0
        for item in project_details:
            ws.append([item.get(f, "") for f in field_map])
            total_amount += _parse_amount(item.get("total_price"))

        total_row = ws.max_row + 1
        ws.append(["总合计报价金额", "", round(total_amount, 2), ""])
        for cell in ws[total_row]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="FFF2CC")
        ws.cell(row=total_row, column=3).number_format = '#,##0.00'

        for col_idx in range(1, len(headers) + 1):
            col_letter = get_column_letter(col_idx)
            max_w = max(_cjk_len(str(cell.value or "")) for cell in ws[col_letter])
            ws.column_dimensions[col_letter].width = max_w + 4

        buf = BytesIO()
        wb.save(buf)
        return base64.b64encode(buf.getvalue()).decode()
    except Exception:
        logger.exception("build_excel_base64_failed")
        return ""
