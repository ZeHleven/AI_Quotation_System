from __future__ import annotations

from datetime import datetime
from io import BytesIO
from pathlib import Path
from typing import Any

from openpyxl import Workbook


LEGACY_EXCEL_EXTENSION = ".xls"


class LegacyExcelConversionError(ValueError):
    pass


def is_legacy_excel_filename(filename: str | None) -> bool:
    return Path(filename or "").suffix.lower() == LEGACY_EXCEL_EXTENSION


def normalize_excel_workbook_bytes(
    content: bytes,
    *,
    filename: str | None,
) -> bytes:
    """Return OpenXML workbook bytes, converting legacy BIFF ``.xls`` files.

    The quotation parsers only consume cell values, sheet names and merged-cell
    structure.  Converting those structures once lets the existing, hardened
    ``openpyxl`` parsing path remain the single source of truth for both old and
    new Excel formats.
    """

    if not is_legacy_excel_filename(filename):
        return content
    if not content:
        raise LegacyExcelConversionError("旧版 Excel 文件为空，请重新上传")

    try:
        import xlrd
    except ImportError as exc:  # pragma: no cover - guarded by production dependency checks.
        raise LegacyExcelConversionError("服务器缺少旧版 Excel 解析组件 xlrd") from exc

    try:
        source = xlrd.open_workbook(file_contents=content, on_demand=True)
    except Exception as exc:
        raise LegacyExcelConversionError(f"旧版 Excel 文件读取失败: {exc}") from exc

    target = Workbook()
    target.remove(target.active)
    try:
        for source_sheet in source.sheets():
            target_sheet = target.create_sheet(title=source_sheet.name)
            for row_index in range(source_sheet.nrows):
                target_sheet.append(
                    [
                        _legacy_cell_value(source, source_sheet.cell(row_index, column_index), xlrd)
                        for column_index in range(source_sheet.ncols)
                    ]
                )
            for row_start, row_end, column_start, column_end in source_sheet.merged_cells:
                if row_end <= row_start or column_end <= column_start:
                    continue
                target_sheet.merge_cells(
                    start_row=row_start + 1,
                    end_row=row_end,
                    start_column=column_start + 1,
                    end_column=column_end,
                )
    finally:
        source.release_resources()

    output = BytesIO()
    target.save(output)
    target.close()
    return output.getvalue()


def _legacy_cell_value(source: Any, cell: Any, xlrd_module: Any) -> Any:
    cell_type = int(cell.ctype)
    value = cell.value
    if cell_type in {xlrd_module.XL_CELL_EMPTY, xlrd_module.XL_CELL_BLANK}:
        return None
    if cell_type == xlrd_module.XL_CELL_DATE:
        try:
            return xlrd_module.xldate.xldate_as_datetime(value, source.datemode)
        except Exception:
            return value
    if cell_type == xlrd_module.XL_CELL_BOOLEAN:
        return bool(value)
    if cell_type == xlrd_module.XL_CELL_ERROR:
        return xlrd_module.biffh.error_text_from_code.get(int(value), "#VALUE!")
    if cell_type == xlrd_module.XL_CELL_NUMBER:
        number = float(value)
        return int(number) if number.is_integer() else number
    if isinstance(value, datetime):
        return value
    return value
