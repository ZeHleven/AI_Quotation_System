from __future__ import annotations

import csv
import json
from collections import Counter
from datetime import datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any, Mapping

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


PHASE = "BIZ-2x-dwg-regression-evaluation"
READY_SPECIAL_TRACE_STATUS = "special_quantity_trace_ready_for_manual_review"
READY_STANDARD_RULE_STATUS = "standard_rule_execution_ready_for_manual_review"

SAMPLE_HEADERS = [
    "样例编号",
    "结果文件",
    "DWG文件数",
    "图纸线索数",
    "标准匹配线索数",
    "识别项目数",
    "唯一标准项目数",
    "项目-区域可绑定数",
    "房间边界数",
    "专项trace总数",
    "专项trace可复核数",
    "专项trace阻断数",
    "最终生成准备度",
    "主要阻断原因",
    "建议下一步",
]

ISSUE_HEADERS = [
    "样例编号",
    "结果文件",
    "问题类型",
    "项目名称",
    "专项算量编号",
    "问题说明",
    "建议处理",
]

REFERENCE_HEADERS = [
    "样例编号",
    "结果文件",
    "参考项目名称",
    "参考单位",
    "参考工程量",
    "识别状态",
    "匹配到的项目名称",
    "最终清单工程量",
    "工程量误差率",
    "说明",
]


def build_dwg_regression_report(
    listing_reports: list[Mapping[str, Any]],
    *,
    reference_rows_by_sample: Mapping[str, list[Mapping[str, Any]]] | list[Mapping[str, Any]] | None = None,
) -> dict[str, Any]:
    sample_rows: list[dict[str, Any]] = []
    issue_rows: list[dict[str, Any]] = []
    reference_compare_rows: list[dict[str, Any]] = []
    aggregate_status_counter: Counter[str] = Counter()
    aggregate_block_counter: Counter[str] = Counter()

    for index, raw_report in enumerate(listing_reports, start=1):
        report = dict(raw_report)
        source_name = _sample_source_name(report, index)
        sample_id = f"BIZ2xREG-{index:03d}"
        analysis = _analyze_listing_report(report)
        sample_row = {
            "样例编号": sample_id,
            "结果文件": source_name,
            "DWG文件数": analysis["dwg_file_count"],
            "图纸线索数": analysis["source_signal_count"],
            "标准匹配线索数": analysis["matched_signal_count"],
            "识别项目数": analysis["recognized_project_count"],
            "唯一标准项目数": analysis["unique_standard_item_count"],
            "项目-区域可绑定数": analysis["region_ready_count"],
            "房间边界数": analysis["room_boundary_count"],
            "专项trace总数": analysis["special_trace_count"],
            "专项trace可复核数": analysis["ready_special_trace_count"],
            "专项trace阻断数": analysis["blocked_special_trace_count"],
            "最终生成准备度": analysis["readiness_status"],
            "主要阻断原因": _join_counter(analysis["block_reason_counter"]),
            "建议下一步": analysis["next_action"],
        }
        sample_rows.append(sample_row)
        aggregate_status_counter[analysis["readiness_status"]] += 1
        aggregate_block_counter.update(analysis["block_reason_counter"])
        issue_rows.extend(_build_issue_rows(sample_id, source_name, report, analysis))

        reference_rows = _reference_rows_for_sample(reference_rows_by_sample, source_name)
        if reference_rows:
            reference_compare_rows.extend(
                _compare_reference_rows(
                    sample_id=sample_id,
                    source_name=source_name,
                    reference_rows=reference_rows,
                    report=report,
                )
            )

    summary = _build_summary(sample_rows, issue_rows, reference_compare_rows, aggregate_status_counter, aggregate_block_counter)
    return {
        "ok": bool(sample_rows),
        "phase": PHASE,
        "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "safe_for_trial": summary["sample_count"] >= 3 and summary["ready_special_trace_count"] > 0,
        "summary": summary,
        "sample_rows": sample_rows,
        "issue_rows": issue_rows,
        "reference_compare_rows": reference_compare_rows,
    }


def write_dwg_regression_outputs(
    report: Mapping[str, Any],
    output_dir: str | Path,
    *,
    stem: str | None = None,
) -> dict[str, str]:
    target_dir = Path(output_dir)
    target_dir.mkdir(parents=True, exist_ok=True)
    file_stem = stem or f"BIZ2x_DWG回归评估报告_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
    json_path = target_dir / f"{file_stem}.json"
    markdown_path = target_dir / f"{file_stem}.md"
    sample_csv_path = target_dir / f"{file_stem}_样例汇总.csv"
    issue_csv_path = target_dir / f"{file_stem}_问题清单.csv"
    reference_csv_path = target_dir / f"{file_stem}_参考清单对比.csv"
    workbook_path = target_dir / f"{file_stem}.xlsx"

    json_path.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    markdown_path.write_text(build_dwg_regression_markdown(report), encoding="utf-8")
    _write_csv(sample_csv_path, list(report.get("sample_rows") or []), SAMPLE_HEADERS)
    _write_csv(issue_csv_path, list(report.get("issue_rows") or []), ISSUE_HEADERS)
    _write_csv(reference_csv_path, list(report.get("reference_compare_rows") or []), REFERENCE_HEADERS)
    _write_workbook(workbook_path, report)
    return {
        "json": str(json_path),
        "markdown": str(markdown_path),
        "sample_csv": str(sample_csv_path),
        "issue_csv": str(issue_csv_path),
        "reference_compare_csv": str(reference_csv_path),
        "xlsx": str(workbook_path),
    }


def build_dwg_regression_markdown(report: Mapping[str, Any]) -> str:
    summary = report.get("summary") or {}
    lines = [
        "# BIZ-2x DWG 识图回归评估报告",
        "",
        f"- 生成时间：{report.get('generated_at', '-')}",
        f"- 样例数：{summary.get('sample_count', 0)}",
        f"- DWG 文件数：{summary.get('dwg_file_count', 0)}",
        f"- 识别项目数：{summary.get('recognized_project_count', 0)}",
        f"- 专项 trace 总数：{summary.get('special_trace_count', 0)}",
        f"- 专项 trace 可复核数：{summary.get('ready_special_trace_count', 0)}",
        f"- 专项 trace 阻断数：{summary.get('blocked_special_trace_count', 0)}",
        f"- 已具备试运行样例数：{summary.get('ready_sample_count', 0)}",
        "",
        "## 样例汇总",
        "",
        "| 样例 | 识别项目 | 可复核专项 trace | 阻断专项 trace | 准备度 | 主要阻断原因 |",
        "| --- | ---: | ---: | ---: | --- | --- |",
    ]
    for row in report.get("sample_rows") or []:
        lines.append(
            "| "
            + " | ".join(
                [
                    _md(row.get("样例编号")),
                    str(row.get("识别项目数", 0)),
                    str(row.get("专项trace可复核数", 0)),
                    str(row.get("专项trace阻断数", 0)),
                    _md(row.get("最终生成准备度")),
                    _md(row.get("主要阻断原因")),
                ]
            )
            + " |"
        )
    lines.extend(["", "## 使用口径", ""])
    lines.extend(
        [
            "- 本报告只评估 DWG 识图链路，不把未复核 trace 写入最终清单。",
            "- 专项 trace 可复核，表示它已经有项目、区域/房间、CAD 几何建议量和标准规则执行证据，下一步仍需业务复核。",
            "- 阻断项用于指导下一轮几何规则增强，例如缺区域、缺净周长、缺高度、标准规则要求展开面积但当前只有水平面积。",
            "- 若上传人工清单参考行，报告会补充参考清单对比；没有参考清单时只输出系统链路覆盖情况。",
        ]
    )
    if report.get("issue_rows"):
        lines.extend(["", "## 前 20 条问题", ""])
        for issue in list(report.get("issue_rows") or [])[:20]:
            lines.append(f"- {issue.get('样例编号')} {issue.get('项目名称') or issue.get('专项算量编号')}: {issue.get('问题说明')}")
    return "\n".join(lines) + "\n"


def _analyze_listing_report(report: Mapping[str, Any]) -> dict[str, Any]:
    summary = report.get("summary") or {}
    project_rows = list(report.get("project_rows") or [])
    special_rows = list(report.get("special_quantity_trace_rows") or [])
    ready_special_rows = [row for row in special_rows if _is_ready_special_trace(row)]
    blocked_special_rows = [row for row in special_rows if not _is_ready_special_trace(row)]
    block_reason_counter = Counter(_block_reason(row) for row in blocked_special_rows)
    block_reason_counter.pop("", None)

    project_summary = report.get("project_recognition_summary") or {}
    region_summary = report.get("project_region_binding_summary") or {}
    room_summary = report.get("room_boundary_summary") or {}

    recognized_project_count = _int(project_summary.get("recognized_project_count"), len(project_rows))
    special_trace_count = len(special_rows)
    ready_count = len(ready_special_rows)
    blocked_count = len(blocked_special_rows)
    status, next_action = _readiness_status_and_action(recognized_project_count, special_trace_count, ready_count)

    return {
        "dwg_file_count": _int(summary.get("dwg_file_count")),
        "source_signal_count": _int(project_summary.get("source_signal_count"), _int(summary.get("source_signal_count"))),
        "matched_signal_count": _int(project_summary.get("matched_signal_count"), _int(summary.get("matched_signal_count"))),
        "recognized_project_count": recognized_project_count,
        "unique_standard_item_count": _int(
            project_summary.get("unique_standard_item_count"),
            len({_value(row, "标准项目编码", "standard_item_code") for row in project_rows if _value(row, "标准项目编码", "standard_item_code")}),
        ),
        "region_ready_count": _int(region_summary.get("binding_ready_project_count")),
        "room_boundary_count": _int(room_summary.get("room_boundary_count")),
        "special_trace_count": special_trace_count,
        "ready_special_trace_count": ready_count,
        "blocked_special_trace_count": blocked_count,
        "readiness_status": status,
        "next_action": next_action,
        "block_reason_counter": block_reason_counter,
        "ready_special_rows": ready_special_rows,
        "blocked_special_rows": blocked_special_rows,
    }


def _build_summary(
    sample_rows: list[dict[str, Any]],
    issue_rows: list[dict[str, Any]],
    reference_compare_rows: list[dict[str, Any]],
    status_counter: Counter[str],
    block_counter: Counter[str],
) -> dict[str, Any]:
    totals = {
        "sample_count": len(sample_rows),
        "dwg_file_count": sum(_int(row.get("DWG文件数")) for row in sample_rows),
        "source_signal_count": sum(_int(row.get("图纸线索数")) for row in sample_rows),
        "matched_signal_count": sum(_int(row.get("标准匹配线索数")) for row in sample_rows),
        "recognized_project_count": sum(_int(row.get("识别项目数")) for row in sample_rows),
        "unique_standard_item_count_total": sum(_int(row.get("唯一标准项目数")) for row in sample_rows),
        "region_ready_count": sum(_int(row.get("项目-区域可绑定数")) for row in sample_rows),
        "room_boundary_count": sum(_int(row.get("房间边界数")) for row in sample_rows),
        "special_trace_count": sum(_int(row.get("专项trace总数")) for row in sample_rows),
        "ready_special_trace_count": sum(_int(row.get("专项trace可复核数")) for row in sample_rows),
        "blocked_special_trace_count": sum(_int(row.get("专项trace阻断数")) for row in sample_rows),
        "ready_sample_count": sum(1 for row in sample_rows if row.get("最终生成准备度") == "可进入专项 trace 复核"),
        "issue_count": len(issue_rows),
        "reference_compare_count": len(reference_compare_rows),
        "reference_matched_count": sum(1 for row in reference_compare_rows if row.get("识别状态") in {"已识别", "已生成最终清单"}),
        "status_counts": dict(status_counter.most_common()),
        "top_block_reasons": dict(block_counter.most_common(10)),
    }
    if totals["source_signal_count"]:
        totals["standard_match_rate"] = round(totals["matched_signal_count"] / totals["source_signal_count"], 4)
    else:
        totals["standard_match_rate"] = 0
    if totals["special_trace_count"]:
        totals["ready_special_trace_rate"] = round(totals["ready_special_trace_count"] / totals["special_trace_count"], 4)
    else:
        totals["ready_special_trace_rate"] = 0
    if totals["reference_compare_count"]:
        totals["reference_match_rate"] = round(totals["reference_matched_count"] / totals["reference_compare_count"], 4)
    else:
        totals["reference_match_rate"] = 0
    return totals


def _build_issue_rows(
    sample_id: str,
    source_name: str,
    report: Mapping[str, Any],
    analysis: Mapping[str, Any],
) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    if not analysis["recognized_project_count"]:
        rows.append(
            {
                "样例编号": sample_id,
                "结果文件": source_name,
                "问题类型": "未识别到标准项目",
                "项目名称": "",
                "专项算量编号": "",
                "问题说明": "当前 DWG 结果没有可进入标准列项的项目。",
                "建议处理": "检查文字提取、材料表/做法表重建和平立面项目标注识别。",
            }
        )
    elif not analysis["special_trace_count"]:
        rows.append(
            {
                "样例编号": sample_id,
                "结果文件": source_name,
                "问题类型": "未形成专项算量 trace",
                "项目名称": "",
                "专项算量编号": "",
                "问题说明": "已识别项目，但没有形成可检查的专项算量 trace。",
                "建议处理": "补强项目-区域绑定、房间边界和专项算量类型映射。",
            }
        )
    for trace in analysis["blocked_special_rows"]:
        rows.append(
            {
                "样例编号": sample_id,
                "结果文件": source_name,
                "问题类型": "专项算量阻断",
                "项目名称": _value(trace, "项目名称", "project_name"),
                "专项算量编号": _value(trace, "专项算量编号", "special_quantity_id"),
                "问题说明": _block_reason(trace) or "专项 trace 未满足可复核条件。",
                "建议处理": _blocked_trace_action(trace),
            }
        )
    return rows


def _compare_reference_rows(
    *,
    sample_id: str,
    source_name: str,
    reference_rows: list[Mapping[str, Any]],
    report: Mapping[str, Any],
) -> list[dict[str, Any]]:
    project_rows = list(report.get("project_rows") or [])
    final_rows = _final_rows(report)
    rows: list[dict[str, Any]] = []
    for reference in reference_rows:
        reference_name = _value(reference, "项目名称", "project_name", "name")
        reference_unit = _value(reference, "单位", "unit")
        reference_quantity = _value(reference, "工程量", "quantity")
        matched_project = _match_project(reference_name, project_rows)
        matched_final = _match_final_row(reference_name, final_rows)
        status = "未识别"
        note = "未在 DWG 项目识别结果中找到相近项目。"
        final_quantity = ""
        quantity_error_rate = ""
        if matched_project:
            status = "已识别"
            note = "已识别到项目，仍需检查专项 trace 是否可复核。"
        if matched_final:
            status = "已生成最终清单"
            final_quantity = _value(matched_final, "工程量", "quantity")
            quantity_error_rate = _quantity_error_rate(reference_quantity, final_quantity)
            note = "已生成最终四字段清单，可结合误差率复核。"
        rows.append(
            {
                "样例编号": sample_id,
                "结果文件": source_name,
                "参考项目名称": reference_name,
                "参考单位": reference_unit,
                "参考工程量": reference_quantity,
                "识别状态": status,
                "匹配到的项目名称": _value(matched_final or matched_project or {}, "项目名称", "project_name"),
                "最终清单工程量": final_quantity,
                "工程量误差率": quantity_error_rate,
                "说明": note,
            }
        )
    return rows


def _is_ready_special_trace(row: Mapping[str, Any]) -> bool:
    return (
        _value(row, "trace状态", "trace_status") == READY_SPECIAL_TRACE_STATUS
        and _yes(_value(row, "是否可复核", "ready_for_manual_review"))
        and _value(row, "标准规则执行状态", "standard_rule_execution_status") == READY_STANDARD_RULE_STATUS
        and _positive_decimal(_value(row, "建议工程量", "suggested_quantity")) is not None
    )


def _readiness_status_and_action(project_count: int, special_count: int, ready_count: int) -> tuple[str, str]:
    if ready_count:
        return "可进入专项 trace 复核", "填写专项 trace 确认工作簿，核验通过后调用最终四字段生成。"
    if special_count:
        return "已有专项 trace 但被阻断", "按问题清单补几何证据、净周长/高度/展开面积或标准规则模板。"
    if project_count:
        return "已识别项目但未形成算量 trace", "优先补项目-区域绑定、房间边界和专项类型映射。"
    return "未识别到可报价项目", "先检查 DWG 文本/表格提取和标准项目匹配。"


def _blocked_trace_action(trace: Mapping[str, Any]) -> str:
    text = f"{_block_reason(trace)} {_value(trace, '未解决事项', 'unresolved_requirements')}"
    if "净周长" in text or "周长" in text:
        return "补房间边界、门洞宽度和净周长扣减证据。"
    if "高度" in text:
        return "补墙面/防水高度来源，优先从设计说明、材料做法和图纸标注提取。"
    if "展开面积" in text:
        return "补墙面展开面积规则，不能用水平投影面积直接替代。"
    if "区域" in text or "房间" in text:
        return "补项目与房间/闭合区域绑定。"
    if "标准规则" in text:
        return "补标准库规则模板映射或规则执行器支持。"
    return "人工复核阻断原因，归入下一轮几何规则增强。"


def _sample_source_name(report: Mapping[str, Any], index: int) -> str:
    for key in ("source_filename", "result_filename", "listing_result_json", "__source_filename"):
        value = _clean_text(report.get(key))
        if value:
            return Path(value).name
    inputs = report.get("inputs") or {}
    upload_dir = _clean_text(inputs.get("upload_dir"))
    if upload_dir:
        return Path(upload_dir).name
    return f"sample_{index:03d}"


def _reference_rows_for_sample(
    reference_rows_by_sample: Mapping[str, list[Mapping[str, Any]]] | list[Mapping[str, Any]] | None,
    source_name: str,
) -> list[Mapping[str, Any]]:
    if not reference_rows_by_sample:
        return []
    if isinstance(reference_rows_by_sample, list):
        return reference_rows_by_sample
    candidates = [source_name, Path(source_name).name, Path(source_name).stem]
    for key in candidates:
        rows = reference_rows_by_sample.get(key)
        if rows:
            return list(rows)
    return []


def _final_rows(report: Mapping[str, Any]) -> list[Mapping[str, Any]]:
    validation = report.get("confirmation_validation") or {}
    if validation.get("final_rows"):
        return list(validation.get("final_rows") or [])
    if report.get("final_rows"):
        return list(report.get("final_rows") or [])
    return []


def _match_project(project_name: str, project_rows: list[Mapping[str, Any]]) -> Mapping[str, Any] | None:
    target = _normalize(project_name)
    if not target:
        return None
    scored: list[tuple[int, Mapping[str, Any]]] = []
    for row in project_rows:
        candidate = _normalize(_value(row, "项目名称", "project_name"))
        drawing_name = _normalize(_value(row, "图纸项目名称", "drawing_project_name"))
        score = _name_score(target, candidate, drawing_name)
        if score:
            scored.append((score, row))
    return sorted(scored, key=lambda item: item[0], reverse=True)[0][1] if scored else None


def _match_final_row(project_name: str, final_rows: list[Mapping[str, Any]]) -> Mapping[str, Any] | None:
    target = _normalize(project_name)
    if not target:
        return None
    scored: list[tuple[int, Mapping[str, Any]]] = []
    for row in final_rows:
        score = _name_score(target, _normalize(_value(row, "项目名称", "project_name")), "")
        if score:
            scored.append((score, row))
    return sorted(scored, key=lambda item: item[0], reverse=True)[0][1] if scored else None


def _name_score(target: str, candidate: str, alternate: str) -> int:
    if not candidate and not alternate:
        return 0
    if target == candidate or target == alternate:
        return 100
    if target in candidate or candidate in target:
        return 80
    if alternate and (target in alternate or alternate in target):
        return 70
    shared = set(_tokens(target)) & set(_tokens(candidate + alternate))
    return len(shared) * 10 if len(shared) >= 2 else 0


def _tokens(value: str) -> list[str]:
    known = ("吊顶", "天棚", "地面", "楼地面", "墙面", "防水", "踢脚线", "窗帘盒", "涂料", "乳胶漆", "石膏板", "饰面")
    return [term for term in known if term in value]


def _quantity_error_rate(reference_quantity: Any, final_quantity: Any) -> str:
    reference = _positive_decimal(reference_quantity)
    final = _positive_decimal(final_quantity)
    if reference is None or final is None:
        return ""
    if reference == 0:
        return ""
    return f"{abs(final - reference) / reference:.2%}"


def _block_reason(row: Mapping[str, Any]) -> str:
    return (
        _value(row, "阻断原因", "block_reason")
        or _value(row, "待补量原因", "quantity_block_reason")
        or _value(row, "未解决事项", "unresolved_requirements")
        or _value(row, "标准规则执行状态", "standard_rule_execution_status")
    )


def _value(row: Mapping[str, Any], *keys: str) -> str:
    for key in keys:
        if key in row and row.get(key) not in (None, ""):
            return _clean_text(row.get(key))
    return ""


def _clean_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _normalize(value: Any) -> str:
    text = _clean_text(value)
    for token in (" ", "\n", "\t", "，", ",", "；", ";", "、", "|", "（", "）", "(", ")", "工程"):
        text = text.replace(token, "")
    return text.lower()


def _yes(value: Any) -> bool:
    return _clean_text(value) in {"是", "Y", "y", "yes", "YES", "True", "true", "1"}


def _int(value: Any, default: int = 0) -> int:
    try:
        return int(value)
    except (TypeError, ValueError):
        return default


def _positive_decimal(value: Any) -> Decimal | None:
    text = _clean_text(value).replace(",", "")
    if not text:
        return None
    for unit in ("㎡", "m2", "m²", "m", "米", "个", "套"):
        text = text.replace(unit, "")
    try:
        number = Decimal(text)
    except (InvalidOperation, ValueError):
        return None
    return number if number > 0 else None


def _join_counter(counter: Counter[str]) -> str:
    return "；".join(f"{key}({value})" for key, value in counter.most_common(5))


def _md(value: Any) -> str:
    return _clean_text(value).replace("|", "\\|").replace("\n", "<br>")


def _write_csv(path: Path, rows: list[Mapping[str, Any]], headers: list[str]) -> None:
    with path.open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.DictWriter(handle, fieldnames=headers, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


def _write_workbook(path: Path, report: Mapping[str, Any]) -> None:
    workbook = Workbook()
    sample_sheet = workbook.active
    sample_sheet.title = "样例汇总"
    _write_sheet(sample_sheet, SAMPLE_HEADERS, list(report.get("sample_rows") or []))
    issue_sheet = workbook.create_sheet("问题清单")
    _write_sheet(issue_sheet, ISSUE_HEADERS, list(report.get("issue_rows") or []))
    reference_sheet = workbook.create_sheet("参考清单对比")
    _write_sheet(reference_sheet, REFERENCE_HEADERS, list(report.get("reference_compare_rows") or []))
    summary_sheet = workbook.create_sheet("统计摘要")
    summary_rows = [{"指标": key, "数值": json.dumps(value, ensure_ascii=False) if isinstance(value, dict) else value} for key, value in (report.get("summary") or {}).items()]
    _write_sheet(summary_sheet, ["指标", "数值"], summary_rows)
    workbook.save(path)


def _write_sheet(sheet, headers: list[str], rows: list[Mapping[str, Any]]) -> None:
    sheet.append(headers)
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in rows:
        sheet.append([row.get(header, "") for header in headers])
    for column_index, header in enumerate(headers, start=1):
        width = min(max(len(header) + 4, 14), 42)
        for cell in sheet.iter_cols(min_col=column_index, max_col=column_index, min_row=2, values_only=True):
            for value in cell[:200]:
                width = min(max(width, min(len(_clean_text(value)) + 2, 42)), 42)
        sheet.column_dimensions[get_column_letter(column_index)].width = width
    sheet.freeze_panes = "A2"
