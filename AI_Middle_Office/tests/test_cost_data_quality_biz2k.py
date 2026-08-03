from __future__ import annotations

from datetime import datetime, timezone
from types import SimpleNamespace

from openpyxl import load_workbook

from app.models.cost_item import COST_STATUS_ACTIVE, COST_STATUS_DRAFT, CostItem
from app.services.cost_data_quality import (
    analyze_cost_items_quality,
    build_demo_regression_pack,
    build_markdown_report,
    write_issues_csv,
    write_xlsx_report,
)


def _item(
    item_id: int,
    item_name: str,
    *,
    spec: str | None = "常规",
    unit: str = "㎡",
    price: float | None = 10.0,
    status: str = COST_STATUS_ACTIVE,
    notes: str | None = "测试备注",
    client_tax_excluded_price: float | None = 12.0,
    subcontract_composite_price: float | None = 10.0,
    crew_benchmark_price: float | None = 9.0,
) -> CostItem:
    return CostItem(
        id=item_id,
        category="测试分类",
        subcategory="测试子类",
        item_name=item_name,
        spec=spec,
        unit=unit,
        price=price,
        client_tax_excluded_price=client_tax_excluded_price,
        subcontract_composite_price=subcontract_composite_price,
        crew_benchmark_price=crew_benchmark_price,
        price_type="combined",
        status=status,
        source="manual",
        notes=notes,
    )


def test_analyze_cost_items_quality_reports_active_data_risks():
    items = [
        _item(1, "墙面乳胶漆", spec="一底两面", unit="㎡", price=18),
        _item(2, "墙面乳胶漆", spec="一底两面", unit="㎡", price=18),
        _item(3, "墙面乳胶漆", spec="", unit="m2", price=20, notes=""),
        _item(4, "墙面乳胶漆修补", spec="一底两面", unit="㎡", price=17),
        _item(
            5,
            "特殊工艺",
            spec="",
            unit="车",
            price=0,
            client_tax_excluded_price=None,
            subcontract_composite_price=None,
            crew_benchmark_price=None,
        ),
        _item(6, "draft 不应纳入体检", status=COST_STATUS_DRAFT, price=0),
    ]
    sync_runs = [
        SimpleNamespace(
            id=9,
            source="cost_items.active",
            status="success",
            requested_count=3,
            synced_count=3,
            started_at=datetime(2026, 5, 25, tzinfo=timezone.utc),
            finished_at=datetime(2026, 5, 25, tzinfo=timezone.utc),
            message="ok",
            error=None,
        )
    ]

    result = analyze_cost_items_quality(items, sync_runs=sync_runs, generated_at=datetime(2026, 5, 25, tzinfo=timezone.utc))
    categories = result["category_counts"]

    assert result["scope"] == "cost_items.active"
    assert result["input_item_count"] == 6
    assert result["active_count"] == 5
    assert categories["exact_active_duplicate"] == 1
    assert categories["same_name_multi_spec"] == 1
    assert categories["missing_spec_on_multi_name"] == 1
    assert categories["invalid_main_price"] == 1
    assert categories["missing_named_reference_price"] == 1
    assert categories["unit_needs_review"] == 1
    assert categories["rag_sync_count_mismatch"] == 1
    assert "similar_active_items" in categories
    assert result["severity_counts"]["high"] >= 2
    assert any(case["case_type"] == "same_name_multi_spec_switch" for case in result["demo_cases"])
    assert all(item["status"] == COST_STATUS_ACTIVE for item in result["active_items"])


def test_quality_report_builders_write_markdown_csv_and_xlsx(tmp_path):
    items = [
        _item(11, "地面找平", spec="30mm 内", unit="㎡", price=25),
        _item(12, "地面找平", spec="50mm 内", unit="㎡", price=35),
    ]
    sync_runs = [
        SimpleNamespace(
            id=10,
            source="cost_items.active",
            status="success",
            requested_count=2,
            synced_count=2,
            started_at=datetime(2026, 5, 25, tzinfo=timezone.utc),
            finished_at=datetime(2026, 5, 25, tzinfo=timezone.utc),
            message="ok",
            error=None,
        )
    ]
    result = analyze_cost_items_quality(items, sync_runs=sync_runs, generated_at=datetime(2026, 5, 25, tzinfo=timezone.utc))

    markdown = build_markdown_report(result)
    demo_pack = build_demo_regression_pack(result)
    csv_path = tmp_path / "issues.csv"
    xlsx_path = tmp_path / "quality.xlsx"

    write_issues_csv(result, csv_path)
    write_xlsx_report(result, xlsx_path)

    assert "BIZ-2k 成本库数据质量体检报告" in markdown
    assert "演示回归包" in demo_pack
    assert csv_path.read_text(encoding="utf-8-sig").startswith("severity,category")
    workbook = load_workbook(xlsx_path)
    assert set(workbook.sheetnames) == {"Summary", "Issues", "Demo Cases", "Active Items"}
    assert workbook["Summary"]["A1"].value == "字段"
