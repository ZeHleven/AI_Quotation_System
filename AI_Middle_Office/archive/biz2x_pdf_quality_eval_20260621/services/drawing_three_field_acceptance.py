from __future__ import annotations

import csv
import json
import re
from dataclasses import asdict, dataclass
from datetime import datetime
from difflib import SequenceMatcher
from pathlib import Path
from typing import Any, Iterable, Mapping, Sequence

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


PHASE = "BIZ-2x-pdf-three-field-acceptance"

ITEM_ALIASES = ("项目名称", "清单项目名称", "项目", "工程内容", "名称")
FEATURE_ALIASES = ("项目特征", "项目特征描述", "特征描述", "特征", "规格", "规格/特征", "做法", "规格型号")
UNIT_ALIASES = ("计量单位", "单位", "报价单位")
QUANTITY_ALIASES = ("工程量", "数量", "工程数量")
CODE_ALIASES = ("项目编码", "编码", "清单编码", "国标编码")

ANSWER_HEADERS = ["sheet_name", "row_no", "section", "seq", "item_code", "item_name", "feature", "unit", "quantity"]
CANDIDATE_HEADERS = ["source", "row_no", "item_name", "feature", "unit", "quantity", "raw"]
COMPARISON_HEADERS = [
    "status",
    "answer_sheet",
    "answer_row_no",
    "answer_item_name",
    "answer_feature",
    "answer_unit",
    "candidate_source",
    "candidate_item_name",
    "candidate_feature",
    "candidate_unit",
    "name_score",
    "feature_score",
    "unit_score",
    "overall_score",
    "issue",
]


@dataclass(frozen=True)
class ThreeFieldAnswerRow:
    sheet_name: str
    row_no: int
    section: str
    seq: str
    item_code: str
    item_name: str
    feature: str
    unit: str
    quantity: str = ""


@dataclass(frozen=True)
class ThreeFieldCandidateRow:
    source: str
    row_no: int
    item_name: str
    feature: str
    unit: str
    quantity: str = ""
    raw: str = ""


def load_answer_rows_from_workbook(
    path: str | Path,
    *,
    sheet_names: Sequence[str] | None = None,
) -> tuple[list[ThreeFieldAnswerRow], list[dict[str, Any]]]:
    workbook = load_workbook(path, data_only=True, read_only=True)
    wanted = set(sheet_names or [])
    rows: list[ThreeFieldAnswerRow] = []
    sheet_summaries: list[dict[str, Any]] = []

    for worksheet in workbook.worksheets:
        if wanted and worksheet.title not in wanted:
            continue
        values = list(worksheet.iter_rows(values_only=True))
        header = _find_header(values)
        summary = {
            "sheet_name": worksheet.title,
            "max_row": worksheet.max_row,
            "max_column": worksheet.max_column,
            "header_found": bool(header),
            "header_row": (header or {}).get("row_index", ""),
            "parsed_row_count": 0,
        }
        if not header:
            sheet_summaries.append(summary)
            continue

        section = ""
        for row_index, row in enumerate(values[header["row_index"] + 1 :], start=header["row_index"] + 2):
            raw_values = [_clean_text(value) for value in row]
            if not any(raw_values):
                continue
            seq = _value_at(raw_values, header.get("seq"))
            item_name = _value_at(raw_values, header.get("item"))
            feature = _value_at(raw_values, header.get("feature"))
            unit = _value_at(raw_values, header.get("unit"))
            quantity = _value_at(raw_values, header.get("quantity"))
            item_code = _value_at(raw_values, header.get("code"))
            if item_name and not seq and not feature and not unit:
                section = item_name
                continue
            if not _looks_like_item_row(item_name=item_name, feature=feature, unit=unit, raw_values=raw_values):
                continue
            rows.append(
                ThreeFieldAnswerRow(
                    sheet_name=worksheet.title,
                    row_no=row_index,
                    section=section,
                    seq=seq,
                    item_code=item_code,
                    item_name=item_name,
                    feature=feature,
                    unit=unit,
                    quantity=quantity,
                )
            )
            summary["parsed_row_count"] += 1
        sheet_summaries.append(summary)
    return rows, sheet_summaries


def load_candidate_rows_from_report(report_or_rows: Mapping[str, Any] | Sequence[Mapping[str, Any]]) -> list[ThreeFieldCandidateRow]:
    if isinstance(report_or_rows, Mapping):
        rows = _candidate_source_rows(report_or_rows)
    else:
        rows = list(report_or_rows)
    candidates: list[ThreeFieldCandidateRow] = []
    for index, row in enumerate(rows, start=1):
        item_name = _first(row, "项目名称", "project_name", "item_name", "图纸项目名称", "standard_item_name")
        feature = _first(row, "项目特征", "project_feature", "feature", "features", "规格/做法", "规格/特征", "spec", "notes")
        unit = _first(row, "单位", "计量单位", "建议单位", "unit", "suggested_unit", "标准单位")
        quantity = _first(row, "工程量", "quantity", "suggested_quantity")
        if not item_name:
            continue
        candidates.append(
            ThreeFieldCandidateRow(
                source=_first(row, "source", "来源", "PDF文件", "source_file") or "candidate",
                row_no=_int(_first(row, "row_no", "行号", "index", "序号"), index),
                item_name=item_name,
                feature=feature,
                unit=unit,
                quantity=quantity,
                raw=json.dumps(row, ensure_ascii=False, default=str)[:1000],
            )
        )
    return candidates


def build_three_field_acceptance_report(
    *,
    answer_rows: Sequence[ThreeFieldAnswerRow],
    candidate_rows: Sequence[ThreeFieldCandidateRow],
    source_name: str = "",
    sheet_summaries: Sequence[Mapping[str, Any]] | None = None,
) -> dict[str, Any]:
    comparisons, extra_candidates = compare_three_fields(answer_rows, candidate_rows)
    summary = _build_summary(answer_rows, candidate_rows, comparisons, extra_candidates)
    return {
        "ok": True,
        "phase": PHASE,
        "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "source_name": source_name,
        "summary": summary,
        "sheet_summaries": list(sheet_summaries or []),
        "answer_rows": [asdict(row) for row in answer_rows],
        "candidate_rows": [asdict(row) for row in candidate_rows],
        "comparison_rows": comparisons,
        "extra_candidate_rows": extra_candidates,
    }


def compare_three_fields(
    answer_rows: Sequence[ThreeFieldAnswerRow],
    candidate_rows: Sequence[ThreeFieldCandidateRow],
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    blocked_candidates: set[int] = set()
    referenced_candidates: set[int] = set()
    comparison_rows: list[dict[str, Any]] = []
    for answer in answer_rows:
        best_index = -1
        best_scores = (0.0, 0.0, 0.0, 0.0)
        best_rank = (0, 0, 0.0)
        for index, candidate in enumerate(candidate_rows):
            if index in blocked_candidates:
                continue
            scores = _score_answer_candidate(answer, candidate)
            status, _ = _comparison_status(answer, candidate, scores)
            status_rank = _candidate_status_rank(status)
            rank = (
                status_rank,
                _candidate_specificity_rank(answer, candidate) if status_rank > 0 else 0,
                scores[3] if status_rank > 0 else 0.0,
            )
            if rank > best_rank:
                best_scores = scores
                best_index = index
                best_rank = rank
        candidate = candidate_rows[best_index] if best_index >= 0 else None
        status, issue = _comparison_status(answer, candidate, best_scores)
        if candidate and status != "missing_candidate":
            referenced_candidates.add(best_index)
            if status in {"matched_three_fields", "unit_conflict"}:
                blocked_candidates.add(best_index)
        comparison_rows.append(_comparison_row(answer, candidate, best_scores, status, issue))

    extra_candidates = [
        {
            "source": candidate.source,
            "row_no": candidate.row_no,
            "item_name": candidate.item_name,
            "feature": candidate.feature,
            "unit": candidate.unit,
            "quantity": candidate.quantity,
            "issue": "系统识别了该项，但人工答案中未找到足够相近的三字段行。",
        }
        for index, candidate in enumerate(candidate_rows)
        if index not in referenced_candidates
    ]
    return comparison_rows, extra_candidates


def _candidate_status_rank(status: str) -> int:
    if status == "matched_three_fields":
        return 4
    if status == "matched_name_unit_feature_review":
        return 3
    if status == "unit_conflict":
        return 2
    if status == "weak_match_review":
        return 1
    return 0


def _candidate_specificity_rank(answer: ThreeFieldAnswerRow, candidate: ThreeFieldCandidateRow) -> int:
    answer_tokens = _specific_feature_tokens(" ".join([answer.item_name, answer.feature]))
    if not answer_tokens:
        return 0
    candidate_tokens = _specific_feature_tokens(" ".join([candidate.item_name, candidate.feature]))
    exact_hits = len(answer_tokens & candidate_tokens)
    if not candidate_tokens:
        return 0
    wrong_family_hits = sum(
        1
        for token in candidate_tokens
        if token not in answer_tokens and any(_specific_token_family(token) == _specific_token_family(answer_token) for answer_token in answer_tokens)
    )
    return exact_hits * 3 - wrong_family_hits


def write_three_field_acceptance_outputs(
    report: Mapping[str, Any],
    output_dir: str | Path,
    *,
    stem: str | None = None,
) -> dict[str, str]:
    target = Path(output_dir)
    target.mkdir(parents=True, exist_ok=True)
    file_stem = stem or f"BIZ2x_PDF三字段验收_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
    json_path = target / f"{file_stem}.json"
    md_path = target / f"{file_stem}.md"
    answer_csv = target / f"{file_stem}_人工答案三字段.csv"
    candidate_csv = target / f"{file_stem}_系统候选三字段.csv"
    comparison_csv = target / f"{file_stem}_三字段对比.csv"
    extra_csv = target / f"{file_stem}_系统多识别.csv"
    xlsx_path = target / f"{file_stem}.xlsx"

    json_path.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    md_path.write_text(build_three_field_acceptance_markdown(report), encoding="utf-8")
    _write_csv(answer_csv, list(report.get("answer_rows") or []), ANSWER_HEADERS)
    _write_csv(candidate_csv, list(report.get("candidate_rows") or []), CANDIDATE_HEADERS)
    _write_csv(comparison_csv, list(report.get("comparison_rows") or []), COMPARISON_HEADERS)
    _write_csv(extra_csv, list(report.get("extra_candidate_rows") or []), ["source", "row_no", "item_name", "feature", "unit", "quantity", "issue"])
    _write_workbook(xlsx_path, report)
    return {
        "json": str(json_path),
        "markdown": str(md_path),
        "answer_csv": str(answer_csv),
        "candidate_csv": str(candidate_csv),
        "comparison_csv": str(comparison_csv),
        "extra_csv": str(extra_csv),
        "xlsx": str(xlsx_path),
    }


def build_three_field_acceptance_markdown(report: Mapping[str, Any]) -> str:
    summary = report.get("summary") or {}
    status_counts = summary.get("status_counts") or {}
    lines = [
        "# BIZ-2x PDF 三字段验收报告",
        "",
        f"- 生成时间：{report.get('generated_at', '-')}",
        f"- 样本：{report.get('source_name') or '-'}",
        f"- 人工答案行数：{summary.get('answer_count', 0)}",
        f"- 系统候选行数：{summary.get('candidate_count', 0)}",
        f"- 三字段全匹配：{summary.get('matched_three_fields_count', 0)}",
        f"- 项目名称通过率：{summary.get('name_pass_rate', 0):.2%}",
        f"- 项目特征通过率：{summary.get('feature_pass_rate', 0):.2%}",
        f"- 单位通过率：{summary.get('unit_pass_rate', 0):.2%}",
        "",
        "## 状态分布",
        "",
    ]
    for status, count in status_counts.items():
        lines.append(f"- {status}: {count}")
    lines.extend(
        [
            "",
            "## 验收口径",
            "",
            "- 第一阶段只验收项目名称、项目特征、单位，不验收工程量。",
            "- `matched_three_fields` 表示名称、特征、单位均达到当前阈值。",
            "- `matched_name_unit_feature_review` 表示名称和单位基本通过，但项目特征需要补图纸证据或标准库字段。",
            "- `unit_conflict` 表示项目可疑似匹配，但计量单位和人工答案不一致。",
            "- `missing_candidate` 表示人工答案中有该项，系统识别候选中没有足够相近项。",
            "",
            "## 前 30 条未全匹配问题",
            "",
            "| 状态 | 答案项目 | 答案单位 | 候选项目 | 候选单位 | 名称分 | 特征分 | 问题 |",
            "| --- | --- | --- | --- | --- | ---: | ---: | --- |",
        ]
    )
    problem_rows = [
        row for row in report.get("comparison_rows") or [] if row.get("status") != "matched_three_fields"
    ][:30]
    for row in problem_rows:
        lines.append(
            "| "
            + " | ".join(
                [
                    _md(row.get("status")),
                    _md(row.get("answer_item_name")),
                    _md(row.get("answer_unit")),
                    _md(row.get("candidate_item_name")),
                    _md(row.get("candidate_unit")),
                    str(row.get("name_score", "")),
                    str(row.get("feature_score", "")),
                    _md(row.get("issue")),
                ]
            )
            + " |"
        )
    return "\n".join(lines) + "\n"


def _candidate_source_rows(report: Mapping[str, Any]) -> list[Mapping[str, Any]]:
    for key in (
        "quantity_list_rows",
        "base_quantity_list_rows",
        "final_rows",
        "project_rows",
        "pdf_direct_item_rows",
        "item_rows",
        "standard_mapping_rows",
    ):
        rows = report.get(key)
        if isinstance(rows, list) and rows:
            return rows
    return []


def _find_header(values: list[tuple[Any, ...]]) -> dict[str, int] | None:
    best: dict[str, int] | None = None
    best_score = 0
    for row_index, row in enumerate(values[:80]):
        cells = [_clean_text(value) for value in row]
        mapping = {
            "seq": _find_alias_column(cells, ("序号",)),
            "code": _find_alias_column(cells, CODE_ALIASES),
            "item": _find_alias_column(cells, ITEM_ALIASES),
            "feature": _find_alias_column(cells, FEATURE_ALIASES),
            "unit": _find_alias_column(cells, UNIT_ALIASES),
            "quantity": _find_alias_column(cells, QUANTITY_ALIASES),
        }
        score = sum(1 for key in ("item", "feature", "unit") if mapping[key] is not None)
        if score > best_score:
            best_score = score
            best = {"row_index": row_index, **{key: value for key, value in mapping.items() if value is not None}}
        if score >= 3:
            return best
    return best if best_score >= 3 else None


def _find_alias_column(cells: Sequence[str], aliases: Sequence[str]) -> int | None:
    normalized_aliases = {_normalize_header(alias) for alias in aliases}
    for index, cell in enumerate(cells):
        normalized = _normalize_header(cell)
        if not normalized:
            continue
        if normalized in normalized_aliases:
            return index
    for index, cell in enumerate(cells):
        normalized = _normalize_header(cell)
        if not normalized:
            continue
        if any(alias and len(alias) >= 3 and alias in normalized for alias in normalized_aliases):
            return index
    return None


def _looks_like_item_row(*, item_name: str, feature: str, unit: str, raw_values: Sequence[str]) -> bool:
    if not item_name or item_name in {"项目名称", "清单项目名称", "名称"}:
        return False
    raw_text = "".join(raw_values)
    if any(term in raw_text for term in ("合计", "小计", "综合单价", "报价单位", "报价日期")) and not unit:
        return False
    return bool(feature or unit)


def _score_answer_candidate(answer: ThreeFieldAnswerRow, candidate: ThreeFieldCandidateRow) -> tuple[float, float, float, float]:
    if _object_category_conflict(answer, candidate):
        return 0.0, 0.0, 0.0, 0.0
    name_score = _text_score(answer.item_name, candidate.item_name)
    feature_score = _feature_score(answer.feature, candidate.feature)
    if feature_score < 0.95 and _candidate_feature_contains_answer(answer, candidate, name_score=name_score):
        feature_score = 0.95
    unit_score = 1.0 if _normalize_unit(answer.unit) and _normalize_unit(answer.unit) == _normalize_unit(candidate.unit) else 0.0
    overall = round(0.55 * name_score + 0.30 * feature_score + 0.15 * unit_score, 4)
    return round(name_score, 4), round(feature_score, 4), round(unit_score, 4), overall


def _candidate_feature_contains_answer(
    answer: ThreeFieldAnswerRow,
    candidate: ThreeFieldCandidateRow,
    *,
    name_score: float,
) -> bool:
    if name_score < 0.78:
        return False
    answer_feature = _feature_containment_key(answer.feature)
    candidate_feature = _feature_containment_key(candidate.feature)
    if not answer_feature or not candidate_feature or answer_feature not in candidate_feature:
        return False
    if len(answer_feature) >= 6:
        return True
    answer_item = _normalize_text(answer.item_name)
    candidate_item = _normalize_text(candidate.item_name)
    return bool(
        len(answer_feature) >= 3
        and answer_item
        and candidate_item
        and (answer_item == candidate_item or answer_item in candidate_item or candidate_item in answer_item)
    )


def _feature_containment_key(value: Any) -> str:
    text = _strip_feature_score_noise(value)
    text = re.sub(r"(^|[\s,;:，；：])\d+[、.．)]\s*", r"\1", text)
    return _normalize_text(text)


def _comparison_status(
    answer: ThreeFieldAnswerRow,
    candidate: ThreeFieldCandidateRow | None,
    scores: tuple[float, float, float, float],
) -> tuple[str, str]:
    if candidate is None or scores[3] < 0.45 or scores[0] < 0.55:
        return "missing_candidate", "人工答案存在该项目，但系统候选中未找到足够相近的项目名称。"
    name_score, feature_score, unit_score, _ = scores
    if _item_name_conflict(answer, candidate):
        return "matched_name_unit_feature_review", "项目名称存在平级/造型等关键列项冲突，需人工复核。"
    if _feature_specificity_conflict(answer, candidate):
        return "matched_name_unit_feature_review", "项目特征存在灯型、线缆规格或拆除对象等关键细分差异，需人工复核。"
    if name_score >= 0.78 and feature_score >= 0.55 and unit_score == 1:
        return "matched_three_fields", "项目名称、项目特征、单位均通过当前验收阈值。"
    if unit_score == 0 and name_score >= 0.70:
        return "unit_conflict", f"疑似同一项目，但单位不一致：人工为 {answer.unit or '-'}，系统为 {candidate.unit or '-'}。"
    if name_score >= 0.72 and unit_score == 1:
        return "matched_name_unit_feature_review", "项目名称和单位基本匹配，项目特征还需要补充或修正。"
    return "weak_match_review", "系统候选和人工答案有弱相似，需要人工判断是否同一清单项。"


def _item_name_conflict(answer: ThreeFieldAnswerRow, candidate: ThreeFieldCandidateRow) -> bool:
    answer_text = _normalize_text(" ".join([answer.item_name, answer.feature]))
    candidate_name = _normalize_text(candidate.item_name)
    candidate_text = _normalize_text(" ".join([candidate.item_name, candidate.feature]))
    answer_is_shaped = any(term in answer_text for term in ("造型", "跌级", "二级"))
    candidate_is_shaped = any(term in candidate_text for term in ("造型", "跌级", "二级"))
    answer_is_flat = "平级" in answer_text
    candidate_is_flat_name = "平级" in candidate_name
    if answer_is_shaped and candidate_is_flat_name and not candidate_is_shaped:
        return True
    if answer_is_flat and candidate_is_shaped and "平级" not in candidate_text:
        return True
    return False


def _feature_specificity_conflict(answer: ThreeFieldAnswerRow, candidate: ThreeFieldCandidateRow) -> bool:
    answer_text = _normalize_text(" ".join([answer.item_name, answer.feature]))
    candidate_text = _normalize_text(" ".join([candidate.item_name, candidate.feature]))

    lamp_terms = ("单眼格栅灯", "双眼格栅灯", "格栅灯", "筒灯", "射灯", "灯盘", "艺术灯", "吊灯", "灯带")
    for term in lamp_terms:
        if term in answer_text and term not in candidate_text:
            return True

    if "冷热" in answer_text and (
        "单冷" in candidate_text or not ("冷热" in candidate_text or ("冷水" in candidate_text and "热水" in candidate_text))
    ):
        return True

    demolition_terms = (
        "售卖窗口",
        "不锈钢玻璃门",
        "玻璃门",
        "实木门",
        "铝合金门",
        "矿棉板",
        "条形扣板",
        "石膏板天花",
        "马桶",
        "管线",
        "台阶",
        "隔墙",
    )
    if "拆除" in answer_text and "拆除" in candidate_text:
        for term in demolition_terms:
            if term in answer_text and term not in candidate_text:
                return True

    answer_compact = re.sub(r"[^a-z0-9]+", "", _clean_text(answer.feature).lower())
    candidate_compact = re.sub(r"[^a-z0-9]+", "", _clean_text(candidate.feature).lower())
    for pattern in (
        r"wdzcbyj\d+(?:\d+)?",
        r"wdzcyjy\d+(?:\d+)?(?:x\d+)?",
        r"wdzcyjv\d+(?:\d+)?(?:x\d+)?",
    ):
        for spec in re.findall(pattern, answer_compact):
            if spec and spec not in candidate_compact:
                return True
    answer_tokens = _specific_feature_tokens(answer_text)
    if answer_tokens:
        candidate_tokens = _specific_feature_tokens(candidate_text)
        for token in answer_tokens:
            if token not in candidate_tokens:
                return True
    return False


def _specific_feature_tokens(value: Any) -> set[str]:
    text = _normalize_text(value)
    if not text:
        return set()
    tokens: set[str] = set()
    ascii_text = text.replace("×", "x")
    compact = re.sub(r"[^a-z0-9+*.]+", "", ascii_text)
    for pattern in (
        r"(?:sc|mt|jdg)\d+",
        r"dn\d+",
        r"de\d+",
        r"(?:wdzc|wdzn|nh)?(?:yjy|yjv|byj|bv)\d[a-z0-9+*.]*",
    ):
        tokens.update(match.group(0).lower() for match in re.finditer(pattern, compact))
    material_patterns = (
        ("sus304", ("sus304",)),
        ("不锈钢", ("不锈钢",)),
        ("ppr", ("ppr",)),
        ("pvc-u", ("pvc-u", "pvcu")),
        ("pvc", ("pvc",)),
        ("铜质", ("铜质",)),
        ("柔性铸铁", ("柔性铸铁",)),
        ("铸铁", ("铸铁",)),
        ("镀锌钢管", ("镀锌钢管",)),
        ("金属波纹软管", ("金属波纹软管", "波纹软管")),
    )
    for token, aliases in material_patterns:
        if any(alias in text for alias in aliases):
            tokens.add(token)
    return tokens


def _specific_token_family(token: str) -> str:
    lowered = token.lower()
    for prefix in ("sc", "mt", "jdg", "dn", "de"):
        if lowered.startswith(prefix):
            return prefix
    for family in ("wdzcbyj", "wdzcyjy", "wdzcyjv", "wdznyjy", "wdznyjv", "nhyjy", "nhyjv", "byj", "bv", "yjy", "yjv"):
        if lowered.startswith(family):
            return family
    return lowered


OBJECT_CATEGORY_TERMS: tuple[tuple[str, tuple[str, ...]], ...] = (
    ("distribution_box", ("配电箱",)),
    ("water_meter", ("水表",)),
    ("valve", ("阀门", "截止阀", "止回阀")),
    ("floor_drain", ("地漏",)),
    ("toilet", ("马桶", "坐便器")),
    ("basin", ("台盆", "洗手盆", "洗脸盆")),
    ("faucet", ("龙头", "水龙头")),
    ("paper_holder", ("厕纸架", "纸巾架")),
    ("mirror", ("梳妆镜", "镜子", "镜面", "清镜", "镜面墙", "镜墙")),
    ("shower", ("花洒", "淋浴")),
    ("water_heater", ("电热水器", "热水器")),
    ("door_window", ("拆除门", "成品门", "玻璃门", "实木门", "铝合金门", "不锈钢门", "地弹门", "门套", "门扇")),
    ("floor_finish", ("地砖", "木地板", "楼地面", "地面")),
    ("ceiling_finish", ("矿棉板", "条形扣板", "石膏板天花", "天花", "吊顶")),
    ("wall_partition", ("隔墙", "墙体")),
    ("stair", ("台阶",)),
    ("pipe_line", ("管线",)),
    ("water_pipe", ("给水管", "不锈钢管", "dn40", "dn20", "dn15")),
    ("drain_pipe", ("排水管", "铸铁管", "de110", "de63")),
    ("conduit", ("配管", "sc40", "sc50", "mt20", "mt25")),
    ("wiring", ("配线", "wdzcbyj", "电线")),
    ("cable", ("电缆", "wdzcyjy", "wdzcyjv")),
    ("socket", ("插座",)),
    ("switch", ("开关",)),
    ("lamp_trough", ("灯槽",)),
    ("light_strip", ("灯带",)),
    ("downlight", ("筒灯",)),
    ("spotlight", ("射灯",)),
    ("grille_light", ("格栅灯",)),
    ("light_panel", ("灯盘",)),
    ("curtain_box", ("窗帘盒",)),
    ("skirting", ("踢脚线",)),
    ("door_set", ("门套",)),
)


def _object_category_conflict(answer: ThreeFieldAnswerRow, candidate: ThreeFieldCandidateRow) -> bool:
    answer_categories = _object_categories(" ".join([answer.item_name, answer.feature]))
    candidate_categories = _object_categories(" ".join([candidate.item_name, candidate.feature]))
    if not answer_categories or not candidate_categories:
        return False
    return answer_categories.isdisjoint(candidate_categories)


def _object_categories(value: Any) -> set[str]:
    text = _normalize_text(value)
    if not text:
        return set()
    categories: set[str] = set()
    for category, terms in OBJECT_CATEGORY_TERMS:
        if any(_normalize_text(term) in text for term in terms):
            categories.add(category)
    return categories


def _comparison_row(
    answer: ThreeFieldAnswerRow,
    candidate: ThreeFieldCandidateRow | None,
    scores: tuple[float, float, float, float],
    status: str,
    issue: str,
) -> dict[str, Any]:
    return {
        "status": status,
        "answer_sheet": answer.sheet_name,
        "answer_row_no": answer.row_no,
        "answer_item_name": answer.item_name,
        "answer_feature": answer.feature,
        "answer_unit": answer.unit,
        "candidate_row_no": candidate.row_no if candidate else "",
        "candidate_source": candidate.source if candidate else "",
        "candidate_item_name": candidate.item_name if candidate else "",
        "candidate_feature": candidate.feature if candidate else "",
        "candidate_unit": candidate.unit if candidate else "",
        "name_score": scores[0],
        "feature_score": scores[1],
        "unit_score": scores[2],
        "overall_score": scores[3],
        "issue": issue,
    }


def _build_summary(
    answer_rows: Sequence[ThreeFieldAnswerRow],
    candidate_rows: Sequence[ThreeFieldCandidateRow],
    comparison_rows: Sequence[Mapping[str, Any]],
    extra_candidates: Sequence[Mapping[str, Any]],
) -> dict[str, Any]:
    answer_count = len(answer_rows)
    status_counts: dict[str, int] = {}
    for row in comparison_rows:
        status = str(row.get("status") or "")
        status_counts[status] = status_counts.get(status, 0) + 1
    name_pass = sum(1 for row in comparison_rows if _float(row.get("name_score")) >= 0.78)
    feature_pass = sum(1 for row in comparison_rows if _float(row.get("feature_score")) >= 0.55)
    unit_pass = sum(1 for row in comparison_rows if _float(row.get("unit_score")) == 1.0)
    matched_three = status_counts.get("matched_three_fields", 0)
    return {
        "answer_count": answer_count,
        "candidate_count": len(candidate_rows),
        "extra_candidate_count": len(extra_candidates),
        "matched_three_fields_count": matched_three,
        "status_counts": status_counts,
        "name_pass_rate": round(name_pass / answer_count, 4) if answer_count else 0,
        "feature_pass_rate": round(feature_pass / answer_count, 4) if answer_count else 0,
        "unit_pass_rate": round(unit_pass / answer_count, 4) if answer_count else 0,
        "three_field_pass_rate": round(matched_three / answer_count, 4) if answer_count else 0,
    }


def _text_score(left: str, right: str) -> float:
    left_norm = _normalize_text(left)
    right_norm = _normalize_text(right)
    if not left_norm or not right_norm:
        return 0.0
    ratio = SequenceMatcher(None, left_norm, right_norm).ratio()
    if left_norm in right_norm or right_norm in left_norm:
        ratio = max(ratio, 0.82)
    left_core = _name_containment_key(left)
    right_core = _name_containment_key(right)
    if (
        left_core
        and right_core
        and min(len(left_core), len(right_core)) >= 4
        and (left_core in right_core or right_core in left_core)
    ):
        ratio = max(ratio, 0.82)
    left_tokens = set(_meaningful_tokens(left_norm))
    right_tokens = set(_meaningful_tokens(right_norm))
    if left_tokens and right_tokens:
        overlap = len(left_tokens & right_tokens) / max(len(left_tokens | right_tokens), 1)
        ratio = max(ratio, overlap)
    return min(ratio, 1.0)


def _name_containment_key(value: Any) -> str:
    text = _clean_text(value)
    text = re.sub(r"（[^）]*）", "", text)
    text = re.sub(r"\([^)]*\)", "", text)
    return _normalize_text(text)


def _feature_score(left: str, right: str) -> float:
    left_norm = _normalize_text(_strip_feature_score_noise(left))
    right_norm = _normalize_text(_strip_feature_score_noise(right))
    if not left_norm and not right_norm:
        return 1.0
    if not left_norm or not right_norm:
        return 0.0
    ratio = SequenceMatcher(None, left_norm[:400], right_norm[:400]).ratio()
    left_tokens = set(_meaningful_tokens(left_norm))
    right_tokens = set(_meaningful_tokens(right_norm))
    if left_tokens and right_tokens:
        containment = len(left_tokens & right_tokens) / max(min(len(left_tokens), len(right_tokens)), 1)
        ratio = max(ratio, containment)
    return min(ratio, 1.0)


def _strip_feature_score_noise(value: Any) -> str:
    text = _clean_text(value)
    text = re.sub(r"\d+[.、]\s*报价范围[:：]?.*?需人工复核[。.]?", "", text, flags=re.IGNORECASE | re.DOTALL)
    text = re.sub(r"报价范围[:：]?.*?需人工复核[。.]?", "", text, flags=re.IGNORECASE | re.DOTALL)
    text = _strip_empty_feature_labels(text)
    return text


def _strip_empty_feature_labels(value: str) -> str:
    labels = "名称|型号|规格|材质|类型|品牌|颜色|尺寸|部位|区域"
    text = str(value or "")
    previous = None
    while previous != text:
        previous = text
        text = re.sub(
            rf"(^|[\s；;，,])\d+[、.．)]\s*(?:{labels})(?:、(?:{labels}))*\s*[:：]\s*(?=(?:\d+[、.．)])|$|[\s；;，,])",
            r"\1",
            text,
        )
    return _clean_text(text)


def _meaningful_tokens(text: str) -> list[str]:
    known_terms = (
        "拆除",
        "地砖",
        "瓷砖",
        "石材",
        "门槛石",
        "防水",
        "聚氨酯",
        "涂膜防水",
        "保护层",
        "成品保护",
        "基层",
        "找平层",
        "结合层",
        "水泥砂浆",
        "界面剂",
        "轻集料混凝土",
        "墙面",
        "吊顶",
        "天花",
        "轻钢龙骨",
        "铝扣板",
        "阻燃板",
        "踢脚线",
        "乳胶漆",
        "石膏板",
        "防水石膏板",
        "墙布",
        "粘结胶",
        "硬包",
        "木方",
        "配电箱",
        "电缆",
        "配管",
        "配线",
        "插座",
        "开关",
        "86型",
        "灯具",
        "led",
        "筒灯",
        "射灯",
        "格栅灯",
        "灯盘",
        "灯带",
        "色温",
        "给水管",
        "排水管",
        "阀门",
        "水表",
        "地漏",
        "马桶",
        "台盆",
        "洗脸盆",
        "梳妆镜",
        "镜面",
        "厕纸架",
        "纸巾架",
        "花洒",
        "龙头",
        "窗帘盒",
        "不锈钢",
        "玻璃",
        "木门",
        "铝合金",
    )
    tokens = [term for term in known_terms if term in text]
    tokens.extend(re.findall(r"[a-z]{1,8}-?\d{1,6}", text))
    tokens.extend(re.findall(r"\d+(?:\.\d+)?\s*(?:kw|lm|w|v|a|k)(?=\D|$)", text))
    tokens.extend(re.findall(r"\d+型", text))
    tokens.extend(re.findall(r"\d+(?:x|×|\*)\d+(?:x|×|\*)?\d*", text))
    return tokens


def _normalize_text(value: Any) -> str:
    text = _clean_text(value).lower()
    replacements = {
        "㎡": "m2",
        "m²": "m2",
        "平方米": "m2",
        "m3": "m3",
        "m³": "m3",
        "立方米": "m3",
        "清境": "清镜",
        "×": "x",
        "*": "x",
        "：": ":",
    }
    for source, target in replacements.items():
        text = text.replace(source, target)
    return re.sub(r"[\s,，;；:：、。.\-_/\\()（）【】\[\]<>《》]+", "", text)


def _normalize_unit(value: Any) -> str:
    text = _clean_text(value).lower()
    mapping = {
        "㎡": "m2",
        "m²": "m2",
        "m2": "m2",
        "平方米": "m2",
        "m³": "m3",
        "m3": "m3",
        "立方米": "m3",
        "米": "m",
        "m": "m",
        "套": "套",
        "个": "个",
        "樘": "樘",
        "台": "台",
    }
    return mapping.get(text, text)


def _normalize_header(value: Any) -> str:
    return re.sub(r"[\s,，;；:：、。.\-_/\\()（）【】\[\]]+", "", _clean_text(value).lower())


def _clean_text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).replace("\r", " ").replace("\n", " ").strip()
    return re.sub(r"\s+", " ", text)


def _value_at(values: Sequence[str], index: int | None) -> str:
    if index is None or index < 0 or index >= len(values):
        return ""
    return values[index]


def _first(row: Mapping[str, Any], *keys: str) -> str:
    for key in keys:
        if key in row and row.get(key) not in (None, ""):
            return _clean_text(row.get(key))
    return ""


def _int(value: Any, default: int = 0) -> int:
    try:
        if value in (None, ""):
            return default
        return int(float(str(value)))
    except (TypeError, ValueError):
        return default


def _float(value: Any, default: float = 0.0) -> float:
    try:
        if value in (None, ""):
            return default
        return float(str(value))
    except (TypeError, ValueError):
        return default


def _write_csv(path: Path, rows: Iterable[Mapping[str, Any]], fieldnames: Sequence[str]) -> None:
    with path.open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.DictWriter(handle, fieldnames=list(fieldnames))
        writer.writeheader()
        for row in rows:
            writer.writerow({field: row.get(field, "") for field in fieldnames})


def _write_workbook(path: Path, report: Mapping[str, Any]) -> None:
    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "验收概览"
    summary = report.get("summary") or {}
    summary_rows = [
        ["指标", "数值"],
        ["人工答案行数", summary.get("answer_count", 0)],
        ["系统候选行数", summary.get("candidate_count", 0)],
        ["三字段全匹配", summary.get("matched_three_fields_count", 0)],
        ["三字段通过率", summary.get("three_field_pass_rate", 0)],
        ["项目名称通过率", summary.get("name_pass_rate", 0)],
        ["项目特征通过率", summary.get("feature_pass_rate", 0)],
        ["单位通过率", summary.get("unit_pass_rate", 0)],
        ["系统多识别", summary.get("extra_candidate_count", 0)],
    ]
    _append_rows(summary_sheet, summary_rows)
    _style_sheet(summary_sheet)

    comparison_sheet = workbook.create_sheet("三字段对比")
    _append_rows(comparison_sheet, [COMPARISON_HEADERS, *[[row.get(header, "") for header in COMPARISON_HEADERS] for row in report.get("comparison_rows") or []]])
    _style_sheet(comparison_sheet)

    answer_sheet = workbook.create_sheet("人工答案")
    _append_rows(answer_sheet, [ANSWER_HEADERS, *[[row.get(header, "") for header in ANSWER_HEADERS] for row in report.get("answer_rows") or []]])
    _style_sheet(answer_sheet)

    candidate_sheet = workbook.create_sheet("系统候选")
    _append_rows(candidate_sheet, [CANDIDATE_HEADERS, *[[row.get(header, "") for header in CANDIDATE_HEADERS] for row in report.get("candidate_rows") or []]])
    _style_sheet(candidate_sheet)
    workbook.save(path)


def _append_rows(sheet: Any, rows: Sequence[Sequence[Any]]) -> None:
    for row in rows:
        sheet.append(list(row))


def _style_sheet(sheet: Any) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for col_idx, column_cells in enumerate(sheet.columns, start=1):
        max_len = max(len(str(cell.value or "")) for cell in column_cells[:200])
        width = min(max(max_len + 2, 10), 42)
        sheet.column_dimensions[get_column_letter(col_idx)].width = width


def _md(value: Any) -> str:
    return str(value if value is not None else "").replace("|", "\\|").replace("\n", " ")
