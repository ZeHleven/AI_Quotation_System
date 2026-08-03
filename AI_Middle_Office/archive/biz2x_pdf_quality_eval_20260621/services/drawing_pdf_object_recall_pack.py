from __future__ import annotations

import csv
import json
import re
from collections import Counter
from datetime import datetime
from pathlib import Path
from typing import Any, Iterable, Mapping, Sequence

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


PHASE = "BIZ-2x-pdf-object-level-recall-pack"

DEFAULT_STATUSES = ("missing_candidate",)

OBJECT_RECALL_HEADERS = [
    "task_no",
    "review_no",
    "status",
    "gap_priority",
    "answer_sheet",
    "answer_row_no",
    "answer_section",
    "target_item_name",
    "target_feature",
    "target_unit",
    "object_class",
    "recommended_pass",
    "target_object_terms",
    "required_evidence_keywords",
    "current_candidate_item_name",
    "current_candidate_unit",
    "current_candidate_evidence_ids",
    "candidate_source_files",
    "evidence_pages",
    "evidence_tiles",
    "review_instruction",
    "evidence_item_hint",
    "evidence_spec_or_method",
    "evidence_suggested_unit",
    "evidence_text",
    "needs_manual_review",
    "ready_for_import",
    "block_reason",
]


def build_object_recall_pack(
    review_report: Mapping[str, Any],
    *,
    statuses: Sequence[str] | None = None,
    max_rows: int | None = None,
) -> dict[str, Any]:
    wanted_statuses = tuple(statuses or DEFAULT_STATUSES)
    review_rows = [dict(row) for row in review_report.get("review_rows") or [] if isinstance(row, Mapping)]
    selected_rows = [
        row for row in review_rows if not wanted_statuses or str(row.get("status") or "") in wanted_statuses
    ]
    if max_rows is not None:
        selected_rows = selected_rows[:max_rows]

    recall_rows: list[dict[str, Any]] = []
    for row in selected_rows:
        object_class = _object_class(row)
        terms = _object_terms(row)
        recommended_pass = _recommended_pass(object_class, row)
        recall_rows.append(
            {
                "task_no": len(recall_rows) + 1,
                "review_no": row.get("review_no", ""),
                "status": row.get("status", ""),
                "gap_priority": _gap_priority(row, object_class),
                "answer_sheet": row.get("answer_sheet", ""),
                "answer_row_no": row.get("answer_row_no", ""),
                "answer_section": row.get("answer_section", ""),
                "target_item_name": row.get("answer_item_name", ""),
                "target_feature": row.get("answer_feature", ""),
                "target_unit": row.get("answer_unit", ""),
                "object_class": object_class,
                "recommended_pass": recommended_pass,
                "target_object_terms": "；".join(terms),
                "required_evidence_keywords": "；".join(_required_keywords(object_class, terms)),
                "current_candidate_item_name": row.get("candidate_item_name", ""),
                "current_candidate_unit": row.get("candidate_unit", ""),
                "current_candidate_evidence_ids": row.get("evidence_ids", ""),
                "candidate_source_files": row.get("candidate_source_files", ""),
                "evidence_pages": row.get("evidence_pages", ""),
                "evidence_tiles": row.get("evidence_tiles", ""),
                "review_instruction": _review_instruction(object_class, row, terms),
                "evidence_item_hint": "",
                "evidence_spec_or_method": "",
                "evidence_suggested_unit": "",
                "evidence_text": "",
                "needs_manual_review": "true",
                "ready_for_import": "false",
                "block_reason": "object evidence fields are intentionally blank; target answer fields are not evidence",
            }
        )

    return {
        "ok": True,
        "phase": PHASE,
        "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "safe_to_import_without_evidence": False,
        "answer_columns_count_as_evidence": False,
        "summary": _summary(recall_rows, selected_rows, wanted_statuses),
        "recall_rows": recall_rows,
    }


def write_object_recall_pack_outputs(
    pack: Mapping[str, Any],
    output_dir: str | Path,
    *,
    stem: str | None = None,
) -> dict[str, str]:
    target = Path(output_dir)
    target.mkdir(parents=True, exist_ok=True)
    file_stem = stem or f"BIZ2x_PDF_object_recall_pack_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
    json_path = target / f"{file_stem}.json"
    csv_path = target / f"{file_stem}.csv"
    markdown_path = target / f"{file_stem}.md"
    xlsx_path = target / f"{file_stem}.xlsx"
    outputs = {
        "json": str(json_path),
        "csv": str(csv_path),
        "markdown": str(markdown_path),
        "xlsx": str(xlsx_path),
    }
    payload = {**dict(pack), "outputs": outputs}
    json_path.write_text(json.dumps(_json_safe(payload), ensure_ascii=False, indent=2), encoding="utf-8")
    _write_csv(csv_path, payload.get("recall_rows") or [], OBJECT_RECALL_HEADERS)
    markdown_path.write_text(_build_markdown(payload), encoding="utf-8")
    _write_workbook(xlsx_path, payload)
    return outputs


def _summary(
    recall_rows: Sequence[Mapping[str, Any]],
    selected_rows: Sequence[Mapping[str, Any]],
    wanted_statuses: Sequence[str],
) -> dict[str, Any]:
    return {
        "selected_review_row_count": len(selected_rows),
        "object_recall_task_count": len(recall_rows),
        "statuses": list(wanted_statuses),
        "status_counts": dict(Counter(str(row.get("status") or "") for row in recall_rows)),
        "object_class_counts": dict(Counter(str(row.get("object_class") or "") for row in recall_rows)),
        "recommended_pass_counts": dict(Counter(str(row.get("recommended_pass") or "") for row in recall_rows)),
        "ready_for_import_count": 0,
        "quantity_status": "deferred_until_three_fields_accepted",
    }


def _object_class(row: Mapping[str, Any]) -> str:
    text = _compact(" ".join([row.get("answer_section", ""), row.get("answer_item_name", ""), row.get("answer_feature", "")]))
    if "拆除" in text and any(term in text for term in ("门", "窗", "门套", "门扇", "售卖窗口", "隔断")):
        return "door_window_demolition"
    if "拆除" in text and any(term in text for term in ("马桶", "台盆", "洗手台", "洁具", "地漏", "阀", "水表")):
        return "fixture_demolition"
    if any(term in text for term in ("阀", "水表", "地漏", "马桶", "台盆", "洗脸盆", "小便器", "大便器", "花洒", "龙头", "给水", "排水")):
        return "fixture_valve_schedule"
    if any(term in text for term in ("电缆", "配线", "配管", "桥架", "灯具", "筒灯", "射灯", "格栅灯", "灯带", "开关", "插座", "配电箱", "电热水器", "热水器")):
        return "electrical_mep"
    if "拆除" in text:
        return "demolition_node"
    if any(term in text for term in ("地面", "地砖", "瓷砖", "石材", "防水", "保护层", "挡水条", "美缝")):
        return "finish_floor"
    if any(term in text for term in ("天花", "天棚", "吊顶", "灯槽", "窗帘盒", "矿棉板", "扣板", "涂料")):
        return "finish_ceiling"
    if any(term in text for term in ("墙面", "隔墙", "抹灰", "墙布", "硬包", "回填", "砌筑")):
        return "finish_wall"
    return "general_object"


def _recommended_pass(object_class: str, row: Mapping[str, Any]) -> str:
    if object_class in {"door_window_demolition", "fixture_demolition", "demolition_node"}:
        return "demolition_node" if object_class != "door_window_demolition" else "door_window_demolition"
    if object_class == "fixture_valve_schedule":
        return "fixture_valve_schedule"
    if object_class == "electrical_mep":
        return "electrical_mep"
    if object_class.startswith("finish_"):
        return "finish_schedule"
    return "table_legend"


def _object_terms(row: Mapping[str, Any]) -> list[str]:
    text = " ".join([str(row.get("answer_item_name") or ""), str(row.get("answer_feature") or "")])
    known_terms = (
        "不锈钢玻璃门",
        "玻璃门",
        "单开实木门",
        "双开实木门",
        "实木门",
        "铝合金门",
        "售卖窗口",
        "门套",
        "门扇",
        "矿棉板天花",
        "条形扣板天花",
        "石膏板天花",
        "台阶",
        "洗手台",
        "马桶",
        "地砖",
        "石材地面",
        "挡水条",
        "美缝",
        "灯槽",
        "无机涂料",
        "砖砌隔墙",
        "墙面抹灰",
        "陶粒回填",
        "墙面瓷砖",
        "配管",
        "配线",
        "电缆",
        "灯具",
        "开关",
        "插座",
        "阀门",
        "地漏",
        "水表",
        "台盆",
        "龙头",
        "洁具",
    )
    terms = [term for term in known_terms if term in text]
    if not terms:
        name = re.sub(r"^(拆除|成品|新做|安装)", "", str(row.get("answer_item_name") or "").strip())
        name = re.sub(r"(拆除|安装|供货|制作)$", "", name).strip()
        if name:
            terms.append(name)
    return _unique(terms)


def _required_keywords(object_class: str, terms: Sequence[str]) -> list[str]:
    keywords = list(terms)
    if object_class in {"door_window_demolition", "fixture_demolition", "demolition_node"}:
        keywords.extend(["拆除", "清运", "原有"])
    if object_class.startswith("finish_"):
        keywords.extend(["材料", "规格", "做法", "节点"])
    if object_class in {"fixture_valve_schedule", "electrical_mep"}:
        keywords.extend(["规格", "型号", "图例", "系统"])
    return _unique(keywords)


def _gap_priority(row: Mapping[str, Any], object_class: str) -> str:
    status = str(row.get("status") or "")
    if status == "missing_candidate" and object_class in {"door_window_demolition", "fixture_demolition", "electrical_mep"}:
        return "P1_object_missing"
    if status == "missing_candidate":
        return "P2_object_missing"
    if status == "unit_conflict":
        return "P2_unit_conflict"
    if status == "weak_match_review":
        return "P2_weak_object_review"
    return "P3_feature_detail"


def _review_instruction(object_class: str, row: Mapping[str, Any], terms: Sequence[str]) -> str:
    target = "、".join(terms) or str(row.get("answer_item_name") or "")
    if object_class == "door_window_demolition":
        return f"在图纸、门窗表、拆除说明或节点中找到“{target}”对应的真实文字证据；不得只填写“拆除”。"
    if object_class in {"fixture_demolition", "fixture_valve_schedule"}:
        return f"在洁具表、给排水图例、节点或说明中找到“{target}”对应的规格/单位证据。"
    if object_class == "electrical_mep":
        return f"在电气图例、系统图或材料表中找到“{target}”对应的型号、规格或点位证据。"
    if object_class.startswith("finish_"):
        return f"在材料表、做法表或节点中找到“{target}”对应的材料/做法证据。"
    return f"在图纸表格、图例或节点中找到“{target}”对应的对象级证据。"


def _compact(value: Any) -> str:
    return re.sub(r"[\s,，。；;:：、\-_/\\()（）\[\]【】<>《》|]+", "", str(value or "").lower())


def _unique(values: Iterable[Any]) -> list[str]:
    seen: set[str] = set()
    result: list[str] = []
    for value in values:
        text = str(value or "").strip()
        if not text:
            continue
        key = _compact(text)
        if key in seen:
            continue
        seen.add(key)
        result.append(text)
    return result


def _build_markdown(report: Mapping[str, Any]) -> str:
    summary = report.get("summary") or {}
    lines = [
        "# BIZ-2x PDF Object-Level Recall Pack",
        "",
        f"- generated_at: {report.get('generated_at', '-')}",
        f"- tasks: {summary.get('object_recall_task_count', 0)}",
        f"- safe_to_import_without_evidence: {report.get('safe_to_import_without_evidence')}",
        "",
        "## Object Classes",
        "",
    ]
    for key, count in (summary.get("object_class_counts") or {}).items():
        lines.append(f"- {key}: {count}")
    lines.extend(
        [
            "",
            "## Top Tasks",
            "",
            "| priority | class | target | unit | current candidate | instruction |",
            "| --- | --- | --- | --- | --- | --- |",
        ]
    )
    for row in (report.get("recall_rows") or [])[:80]:
        lines.append(
            "| "
            + " | ".join(
                [
                    _md(row.get("gap_priority")),
                    _md(row.get("object_class")),
                    _md(row.get("target_item_name")),
                    _md(row.get("target_unit")),
                    _md(row.get("current_candidate_item_name")),
                    _md(row.get("review_instruction")),
                ]
            )
            + " |"
        )
    return "\n".join(lines) + "\n"


def _write_workbook(path: Path, report: Mapping[str, Any]) -> None:
    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "object_recall_summary"
    _append_rows(summary_sheet, [["metric", "value"]])
    for key, value in (report.get("summary") or {}).items():
        _append_rows(summary_sheet, [[key, json.dumps(value, ensure_ascii=False) if isinstance(value, (dict, list)) else value]])
    _style_sheet(summary_sheet)

    recall_sheet = workbook.create_sheet("object_recall_tasks")
    _append_rows(recall_sheet, [OBJECT_RECALL_HEADERS])
    _append_rows(
        recall_sheet,
        [[_cell_value(row.get(header)) for header in OBJECT_RECALL_HEADERS] for row in report.get("recall_rows") or []],
    )
    _style_sheet(recall_sheet)
    recall_sheet.freeze_panes = "A2"
    workbook.save(path)


def _write_csv(path: Path, rows: Iterable[Mapping[str, Any]], fieldnames: Sequence[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        for row in rows:
            writer.writerow({key: _cell_value(row.get(key)) for key in fieldnames})


def _append_rows(sheet: Any, rows: Iterable[Sequence[Any]]) -> None:
    for row in rows:
        sheet.append(list(row))


def _style_sheet(sheet: Any) -> None:
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for col_idx, column_cells in enumerate(sheet.columns, start=1):
        values = [str(cell.value or "") for cell in column_cells[:80]]
        width = min(max([len(value) for value in values] + [10]) + 2, 80)
        sheet.column_dimensions[get_column_letter(col_idx)].width = width


def _cell_value(value: Any) -> Any:
    if isinstance(value, (dict, list, tuple)):
        return json.dumps(value, ensure_ascii=False)
    return value


def _json_safe(value: Any) -> Any:
    if isinstance(value, Mapping):
        return {str(key): _json_safe(item) for key, item in value.items()}
    if isinstance(value, (list, tuple)):
        return [_json_safe(item) for item in value]
    return value


def _md(value: Any) -> str:
    return str(value or "").replace("|", "\\|").replace("\n", "<br>")
