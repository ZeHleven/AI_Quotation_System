from __future__ import annotations

import csv
import json
import shutil
from collections import defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.core.config import BASE_DIR
from app.services.drawing_standard_matcher import (
    DEFAULT_ACTIVE_STANDARD_LIBRARY_PATH,
    match_drawing_fields_to_standard,
    write_standard_match_outputs,
)
from app.services.drawing_project_recognizer import (
    build_drawing_project_recognition_report,
    write_drawing_project_recognition_outputs,
)
from app.services.drawing_project_geometry_binder import (
    build_project_geometry_binding_report,
    write_project_geometry_binding_outputs,
)
from app.services.drawing_project_region_binder import (
    build_project_region_binding_report,
    write_project_region_binding_outputs,
)
from app.services.drawing_project_material_binder import (
    build_project_material_binding_report,
    write_project_material_binding_outputs,
)
from app.services.drawing_floor_paving_locator import (
    build_floor_paving_locator_report,
    write_floor_paving_locator_outputs,
)
from app.services.drawing_floor_layer_rescanner import (
    build_floor_layer_rescan_report,
    write_floor_layer_rescan_outputs,
)
from app.services.drawing_floor_region_reconstructor import (
    build_floor_region_reconstruction_report,
    write_floor_region_reconstruction_outputs,
)
from app.services.drawing_dynamic_itemization import (
    build_dynamic_itemization_report_runtime,
    write_dynamic_itemization_outputs,
)
from app.services.drawing_special_quantity_calculator import (
    build_special_quantity_calculation_report,
    write_special_quantity_calculation_outputs,
)
from app.services.drawing_special_trace_finalizer import build_special_trace_confirmation_pack
from app.services import drawing_quantity_confirmation as quantity_confirmation
from app.services.dwg_oda_converter import convert_dwg_directory_to_dxf_with_oda, write_oda_conversion_outputs
from app.services.dwg_preview_probe import probe_converter_tools
from app.services.dxf_geometry_probe import build_geometry_probe_report, parse_dxf_geometry_file, write_geometry_probe_outputs
from app.services.dxf_layer_block_mapper import build_layer_block_mapping_report, write_layer_block_mapping_outputs
from app.services.dxf_quantity_suggester import build_low_risk_quantity_suggestion_report, write_quantity_suggestion_outputs
from app.services.dxf_region_label_binder import build_region_label_binding_report, write_region_label_binding_outputs
from app.services.dxf_room_boundary_analyzer import build_room_boundary_analysis_report, write_room_boundary_analysis_outputs
from app.services.dxf_scale_unit_probe import (
    build_manual_scale_unit_confirmation_report,
    build_scale_unit_probe_report,
    write_manual_scale_unit_confirmation_outputs,
    write_scale_unit_probe_outputs,
)
from app.services.dxf_standard_rule_binder import build_standard_rule_binding_report, write_standard_rule_binding_outputs
from app.services.dxf_table_field_convergence import (
    append_drawing_annotation_rows,
    converge_table_fields,
    write_field_convergence_outputs,
)
from app.services.dxf_table_reconstructor import reconstruct_dxf_tables, write_table_reconstruction_outputs
from app.services.dxf_text_extractor import (
    DEFAULT_TEXT_RECORD_LIMIT,
    build_dxf_extraction_report,
    parse_dxf_file,
    write_dxf_extraction_outputs,
)
from app.services.dxf_trace_review_pack import build_trace_review_pack, write_trace_review_outputs
from app.services.quantity_standard_library import load_quantity_standard_library


PHASE = "BIZ-2x-dwg-upload-item-listing"
DEFAULT_ODA_EXE = BASE_DIR.parent / "tools" / "oda" / "extracted" / "ODAFileConverter.exe"
ITEM_LIST_HEADERS = [
    "序号",
    "标准项目编码",
    "项目名称",
    "章节",
    "单位",
    "匹配置信度",
    "来源文件",
    "图纸识别名称",
    "图纸识别规格或做法",
    "项目特征字段",
    "工程量计算规则",
    "工程量状态",
    "逐条绑定状态",
    "绑定建议编号",
    "绑定置信度",
    "绑定说明",
    "系统建议工程量",
    "建议单位",
    "建议量状态",
    "标准规则Trace状态",
    "算量证据",
    "来源证据",
]
QUANTITY_LIST_HEADERS = ["项目名称", "项目特征", "单位", "工程量"]
PENDING_QUANTITY_LABEL = "待算量"


class DwgItemListingError(ValueError):
    pass


def run_dwg_item_listing(
    *,
    upload_dir: str | Path,
    output_dir: str | Path,
    oda_executable: str | Path | None = None,
    standard_file: str | Path = DEFAULT_ACTIVE_STANDARD_LIBRARY_PATH,
    timestamp: str | None = None,
    limit_per_source: int = 3,
    min_confidence: float = 0.45,
    text_record_limit: int = DEFAULT_TEXT_RECORD_LIMIT,
) -> dict[str, Any]:
    run_timestamp = timestamp or datetime.now().strftime("%Y%m%d_%H%M%S")
    source_dir = Path(upload_dir)
    target_dir = Path(output_dir)
    target_dir.mkdir(parents=True, exist_ok=True)
    dwg_files = _collect_dwg_files(source_dir)
    if not dwg_files:
        raise DwgItemListingError("没有找到可识别的 .dwg 文件")

    executable = Path(oda_executable) if oda_executable else find_oda_executable()
    if not executable:
        raise DwgItemListingError("未找到 ODAFileConverter.exe，无法把 DWG 转为 DXF")

    dxf_dir = target_dir / f"dxf_{run_timestamp}"
    conversion = convert_dwg_directory_to_dxf_with_oda(
        source_dir,
        dxf_dir,
        executable,
        recursive=False,
        audit=True,
    )
    conversion_outputs = write_oda_conversion_outputs(
        conversion,
        target_dir,
        stem=f"BIZ2x_DWG转DXF_{run_timestamp}",
    )
    if conversion.status not in {"converted", "partial_converted"} or not conversion.output_files:
        raise DwgItemListingError(f"DWG 转 DXF 未成功：{conversion.message}")

    parsed_files = [parse_dxf_file(path, text_record_limit=text_record_limit) for path in conversion.output_files]
    text_report = build_dxf_extraction_report(parsed_files)
    text_outputs = write_dxf_extraction_outputs(
        parsed_files,
        target_dir,
        stem=f"BIZ2x_DXF文字提取_{run_timestamp}",
    )
    table_report = reconstruct_dxf_tables(parsed_files)
    table_outputs = write_table_reconstruction_outputs(
        table_report,
        target_dir,
        stem=f"BIZ2x_DXF表格重建_{run_timestamp}",
    )
    field_report = append_drawing_annotation_rows(converge_table_fields(table_report), parsed_files)
    field_outputs = write_field_convergence_outputs(
        field_report,
        target_dir,
        stem=f"BIZ2x_DXF字段收敛_{run_timestamp}",
    )
    dynamic_itemization_report = build_dynamic_itemization_report_runtime(field_report)
    dynamic_itemization_outputs = write_dynamic_itemization_outputs(
        dynamic_itemization_report,
        target_dir,
        stem=f"BIZ2x_DWG_R0_R9_dynamic_itemization_{run_timestamp}",
    )
    library = load_quantity_standard_library(standard_file)
    match_report = match_drawing_fields_to_standard(
        field_report,
        library,
        limit_per_source=limit_per_source,
        min_confidence=min_confidence,
    )
    match_report["inputs"] = {
        "upload_dir": str(source_dir.resolve()),
        "dxf_dir": str(dxf_dir.resolve()),
        "standard_file": str(Path(standard_file).resolve()),
    }
    match_outputs = write_standard_match_outputs(
        match_report,
        target_dir,
        stem=f"BIZ2x_DWG列项标准匹配_{run_timestamp}",
    )
    project_report = build_drawing_project_recognition_report(match_report)
    project_outputs = write_drawing_project_recognition_outputs(
        project_report,
        target_dir,
        stem=f"BIZ2x_DWG图纸项目识别_{run_timestamp}",
    )

    geometry_quantity = build_geometry_quantity_context(
        dxf_files=conversion.output_files,
        parsed_text_files=parsed_files,
        text_report=text_report,
        match_report=match_report,
        library=library,
        output_dir=target_dir,
        timestamp=run_timestamp,
    )
    project_geometry_binding_report = build_project_geometry_binding_report(
        project_report=project_report,
        geometry_report=geometry_quantity.get("geometry_report") or {},
        unit_conversion=geometry_quantity.get("unit_conversion") or {},
        region_label_report=geometry_quantity.get("region_label_report") or {},
    )
    project_geometry_binding_outputs = write_project_geometry_binding_outputs(
        project_geometry_binding_report,
        target_dir,
        stem=f"BIZ2x_DWG项目几何绑定_{run_timestamp}",
    )
    project_region_binding_report = build_project_region_binding_report(
        project_report=project_report,
        region_label_report=geometry_quantity.get("region_label_report") or {},
    )
    project_region_binding_outputs = write_project_region_binding_outputs(
        project_region_binding_report,
        target_dir,
        stem=f"BIZ2x_DWG项目区域绑定_{run_timestamp}",
    )
    project_material_binding_report = build_project_material_binding_report(
        project_report=project_report,
        region_label_report=geometry_quantity.get("region_label_report") or {},
        geometry_report=geometry_quantity.get("geometry_report") or {},
        unit_conversion=geometry_quantity.get("unit_conversion") or {},
        field_report=field_report,
    )
    project_material_binding_outputs = write_project_material_binding_outputs(
        project_material_binding_report,
        target_dir,
        stem=f"BIZ2x_DWG材料编号CAD证据绑定_{run_timestamp}",
    )
    floor_paving_locator_report = build_floor_paving_locator_report(
        project_material_binding_report=project_material_binding_report,
        field_report=field_report,
        geometry_report=geometry_quantity.get("geometry_report") or {},
        region_label_report=geometry_quantity.get("region_label_report") or {},
        unit_conversion=geometry_quantity.get("unit_conversion") or {},
    )
    floor_paving_locator_outputs = write_floor_paving_locator_outputs(
        floor_paving_locator_report,
        target_dir,
        stem=f"BIZ2x_DWG地面铺装有效区域定位_{run_timestamp}",
    )
    floor_layer_rescan_report = build_floor_layer_rescan_report(
        dxf_files=conversion.output_files,
        floor_paving_locator_report=floor_paving_locator_report,
        unit_conversion=geometry_quantity.get("unit_conversion") or {},
    )
    floor_layer_rescan_outputs = write_floor_layer_rescan_outputs(
        floor_layer_rescan_report,
        target_dir,
        stem=f"BIZ2x_DWG地面图层定向重扫_{run_timestamp}",
    )
    floor_region_reconstruction_report = build_floor_region_reconstruction_report(
        floor_layer_rescan_report=floor_layer_rescan_report,
        room_boundary_report=geometry_quantity.get("room_boundary_report") or {},
        area_to_square_meter_factor=(geometry_quantity.get("unit_conversion") or {}).get("area_to_square_meter_factor"),
    )
    floor_region_reconstruction_outputs = write_floor_region_reconstruction_outputs(
        floor_region_reconstruction_report,
        target_dir,
        stem=f"BIZ2x_DWG地面线段闭合区域重构_{run_timestamp}",
    )
    special_quantity_report = build_special_quantity_calculation_report(
        project_report=project_report,
        project_region_binding_report=project_region_binding_report,
        room_boundary_report=geometry_quantity.get("room_boundary_report") or {},
        standard_match_report=match_report,
    )
    special_quantity_outputs = write_special_quantity_calculation_outputs(
        special_quantity_report,
        target_dir,
        stem=f"BIZ2x_DWG专项算量trace_{run_timestamp}",
    )
    special_trace_confirmation_pack = build_special_trace_confirmation_pack(special_quantity_report)
    special_trace_confirmation_outputs = quantity_confirmation.write_confirmation_outputs(
        special_trace_confirmation_pack,
        target_dir,
        stem=f"BIZ2x_DWG专项trace确认工作簿_{run_timestamp}",
    )
    item_rows = build_item_listing_rows(match_report, geometry_quantity.get("standard_rule_binding_report"))
    quantity_trace_rows = build_quantity_trace_rows(geometry_quantity.get("standard_rule_binding_report"))
    line_quantity_candidate_rows = build_line_quantity_candidate_rows(item_rows)
    quantity_list_rows = build_quantity_list_rows(
        project_report.get("project_rows", []),
        special_quantity_report,
    )
    quantity_list_outputs = write_quantity_list_outputs(
        quantity_list_rows,
        target_dir,
        stem=f"BIZ2x_DWG识图四字段清单_{run_timestamp}",
    )
    item_outputs = write_item_listing_outputs(
        {
            "ok": True,
            "phase": PHASE,
            "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "summary": build_item_listing_summary(
                dwg_file_count=len(dwg_files),
                dxf_file_count=len(conversion.output_files),
                text_report=text_report,
                field_report=field_report,
                match_report=match_report,
                item_rows=item_rows,
                geometry_quantity=geometry_quantity,
                project_report=project_report,
                project_geometry_binding_report=project_geometry_binding_report,
                project_region_binding_report=project_region_binding_report,
                project_material_binding_report=project_material_binding_report,
                floor_paving_locator_report=floor_paving_locator_report,
                floor_layer_rescan_report=floor_layer_rescan_report,
                floor_region_reconstruction_report=floor_region_reconstruction_report,
                special_quantity_report=special_quantity_report,
            ),
            "item_rows": item_rows,
            "quantity_list_rows": quantity_list_rows,
            "project_rows": project_report.get("project_rows", []),
            "project_recognition_summary": project_report.get("summary", {}),
            "project_geometry_binding_rows": project_geometry_binding_report.get("binding_rows", []),
            "project_geometry_candidate_rows": project_geometry_binding_report.get("candidate_rows", []),
            "project_geometry_binding_summary": project_geometry_binding_report.get("summary", {}),
            "project_region_binding_rows": project_region_binding_report.get("binding_rows", []),
            "project_region_candidate_rows": project_region_binding_report.get("candidate_rows", []),
            "project_region_binding_summary": project_region_binding_report.get("summary", {}),
            "project_material_binding_rows": project_material_binding_report.get("project_binding_rows", []),
            "project_material_index_rows": project_material_binding_report.get("material_index_rows", []),
            "project_material_inheritance_rows": project_material_binding_report.get("material_inheritance_rows", []),
            "project_material_binding_summary": project_material_binding_report.get("summary", {}),
            "floor_paving_project_rows": floor_paving_locator_report.get("floor_project_rows", []),
            "floor_paving_geometry_rows": floor_paving_locator_report.get("floor_geometry_rows", []),
            "floor_paving_material_text_rows": floor_paving_locator_report.get("floor_material_text_rows", []),
            "floor_paving_summary": floor_paving_locator_report.get("summary", {}),
            "floor_layer_segment_rows": floor_layer_rescan_report.get("floor_segment_rows", []),
            "floor_layer_package_rows": floor_layer_rescan_report.get("floor_package_rows", []),
            "floor_layer_rescan_summary": floor_layer_rescan_report.get("summary", {}),
            "floor_closed_region_rows": floor_region_reconstruction_report.get("closed_region_rows", []),
            "floor_region_project_rows": floor_region_reconstruction_report.get("project_region_rows", []),
            "floor_region_reconstruct_summary": floor_region_reconstruction_report.get("summary", {}),
            "room_boundary_rows": geometry_quantity.get("room_boundary_report", {}).get("room_rows", []),
            "room_opening_candidate_rows": geometry_quantity.get("room_boundary_report", {}).get("opening_candidate_rows", []),
            "room_boundary_summary": geometry_quantity.get("room_boundary_report", {}).get("summary", {}),
            "special_quantity_trace_rows": special_quantity_report.get("special_quantity_trace_rows", []),
            "special_quantity_summary": special_quantity_report.get("summary", {}),
            "quantity_trace_rows": quantity_trace_rows,
            "line_quantity_candidate_rows": line_quantity_candidate_rows,
            "dynamic_itemization_summary": dynamic_itemization_report.get("summary", {}),
            "dynamic_itemization_stage_results": dynamic_itemization_report.get("stage_results", []),
            "dynamic_itemization_decision_rows": dynamic_itemization_report.get("itemization_decisions", []),
            "geometry_quantity_summary": geometry_quantity.get("summary", {}),
        },
        target_dir,
        stem=f"BIZ2x_DWG上传列项_{run_timestamp}",
    )
    combined_outputs = {
        "item_list_json": item_outputs["json"],
        "item_list_xlsx": item_outputs["xlsx"],
        "item_list_csv": item_outputs["csv"],
        "item_list_markdown": item_outputs["markdown"],
        "quantity_list_xlsx": quantity_list_outputs["xlsx"],
        "quantity_list_csv": quantity_list_outputs["csv"],
        "conversion_json": conversion_outputs["json"],
        "conversion_markdown": conversion_outputs["markdown"],
        "text_json": text_outputs["json"],
        "text_markdown": text_outputs["markdown"],
        "text_csv": text_outputs["csv"],
        "table_json": table_outputs["json"],
        "table_markdown": table_outputs["markdown"],
        "field_json": field_outputs["json"],
        "field_markdown": field_outputs["markdown"],
        "field_material_csv": field_outputs["material_method_csv"],
        "field_annotation_csv": field_outputs["drawing_annotation_csv"],
        "standard_match_json": match_outputs["json"],
        "standard_match_markdown": match_outputs["markdown"],
        "standard_match_csv": match_outputs["standard_match_csv"],
        "feature_fill_csv": match_outputs["feature_fill_csv"],
        "project_recognition_json": project_outputs["json"],
        "project_recognition_markdown": project_outputs["markdown"],
        "project_recognition_csv": project_outputs["project_csv"],
        "project_draft_four_field_xlsx": project_outputs["draft_four_field_xlsx"],
        "project_geometry_binding_json": project_geometry_binding_outputs["json"],
        "project_geometry_binding_markdown": project_geometry_binding_outputs["markdown"],
        "project_geometry_binding_csv": project_geometry_binding_outputs["binding_csv"],
        "project_geometry_candidate_csv": project_geometry_binding_outputs["candidate_csv"],
        "project_region_binding_json": project_region_binding_outputs["json"],
        "project_region_binding_markdown": project_region_binding_outputs["markdown"],
        "project_region_binding_csv": project_region_binding_outputs["binding_csv"],
        "project_region_candidate_csv": project_region_binding_outputs["candidate_csv"],
        "project_material_binding_json": project_material_binding_outputs["json"],
        "project_material_binding_markdown": project_material_binding_outputs["markdown"],
        "project_material_binding_csv": project_material_binding_outputs["project_binding_csv"],
        "project_material_index_csv": project_material_binding_outputs["material_index_csv"],
        "project_material_table_csv": project_material_binding_outputs["material_table_csv"],
        "project_material_inheritance_csv": project_material_binding_outputs["material_inheritance_csv"],
        "floor_paving_locator_json": floor_paving_locator_outputs["json"],
        "floor_paving_locator_markdown": floor_paving_locator_outputs["markdown"],
        "floor_paving_project_csv": floor_paving_locator_outputs["project_csv"],
        "floor_paving_geometry_csv": floor_paving_locator_outputs["geometry_csv"],
        "floor_paving_material_text_csv": floor_paving_locator_outputs["text_csv"],
        "floor_layer_rescan_json": floor_layer_rescan_outputs["json"],
        "floor_layer_rescan_markdown": floor_layer_rescan_outputs["markdown"],
        "floor_layer_segment_csv": floor_layer_rescan_outputs["segment_csv"],
        "floor_layer_package_csv": floor_layer_rescan_outputs["package_csv"],
        "floor_region_reconstruct_json": floor_region_reconstruction_outputs["json"],
        "floor_region_reconstruct_markdown": floor_region_reconstruction_outputs["markdown"],
        "floor_region_closed_csv": floor_region_reconstruction_outputs["closed_region_csv"],
        "floor_region_project_csv": floor_region_reconstruction_outputs["project_region_csv"],
        "special_quantity_json": special_quantity_outputs["json"],
        "special_quantity_markdown": special_quantity_outputs["markdown"],
        "special_quantity_trace_csv": special_quantity_outputs["trace_csv"],
        "special_trace_confirmation_json": special_trace_confirmation_outputs["json"],
        "special_trace_confirmation_markdown": special_trace_confirmation_outputs["markdown"],
        "special_trace_confirmation_csv": special_trace_confirmation_outputs["confirmation_csv"],
        "special_trace_confirmation_xlsx": special_trace_confirmation_outputs["confirmation_xlsx"],
        "dynamic_itemization_json": dynamic_itemization_outputs["json"],
        "dynamic_itemization_markdown": dynamic_itemization_outputs["markdown"],
        "dynamic_itemization_csv": dynamic_itemization_outputs["itemization_decision_csv"],
        "dynamic_itemization_confirmation_xlsx": dynamic_itemization_outputs.get("confirmation_confirmation_xlsx", ""),
        "dynamic_itemization_confirmation_json": dynamic_itemization_outputs.get("confirmation_json", ""),
        **geometry_quantity.get("outputs", {}),
    }
    issues = build_item_listing_issues(conversion.as_dict(), match_report, item_rows, geometry_quantity)
    Path(item_outputs["json"]).write_text(
        json.dumps(
            {
                "ok": True,
                "phase": PHASE,
                "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                "summary": build_item_listing_summary(
                    dwg_file_count=len(dwg_files),
                    dxf_file_count=len(conversion.output_files),
                    text_report=text_report,
                    field_report=field_report,
                    match_report=match_report,
                    item_rows=item_rows,
                    geometry_quantity=geometry_quantity,
                    project_report=project_report,
                    project_geometry_binding_report=project_geometry_binding_report,
                    project_region_binding_report=project_region_binding_report,
                    project_material_binding_report=project_material_binding_report,
                    floor_paving_locator_report=floor_paving_locator_report,
                    floor_layer_rescan_report=floor_layer_rescan_report,
                    floor_region_reconstruction_report=floor_region_reconstruction_report,
                    special_quantity_report=special_quantity_report,
                ),
                "item_rows": item_rows,
                "quantity_list_rows": quantity_list_rows,
                "project_rows": project_report.get("project_rows", []),
                "project_recognition_summary": project_report.get("summary", {}),
                "project_geometry_binding_rows": project_geometry_binding_report.get("binding_rows", []),
                "project_geometry_candidate_rows": project_geometry_binding_report.get("candidate_rows", []),
                "project_geometry_binding_summary": project_geometry_binding_report.get("summary", {}),
                "project_region_binding_rows": project_region_binding_report.get("binding_rows", []),
                "project_region_candidate_rows": project_region_binding_report.get("candidate_rows", []),
                "project_region_binding_summary": project_region_binding_report.get("summary", {}),
                "project_material_binding_rows": project_material_binding_report.get("project_binding_rows", []),
                "project_material_index_rows": project_material_binding_report.get("material_index_rows", []),
                "project_material_inheritance_rows": project_material_binding_report.get("material_inheritance_rows", []),
                "project_material_binding_summary": project_material_binding_report.get("summary", {}),
                "floor_paving_project_rows": floor_paving_locator_report.get("floor_project_rows", []),
                "floor_paving_geometry_rows": floor_paving_locator_report.get("floor_geometry_rows", []),
                "floor_paving_material_text_rows": floor_paving_locator_report.get("floor_material_text_rows", []),
                "floor_paving_summary": floor_paving_locator_report.get("summary", {}),
                "floor_layer_segment_rows": floor_layer_rescan_report.get("floor_segment_rows", []),
                "floor_layer_package_rows": floor_layer_rescan_report.get("floor_package_rows", []),
                "floor_layer_rescan_summary": floor_layer_rescan_report.get("summary", {}),
                "floor_closed_region_rows": floor_region_reconstruction_report.get("closed_region_rows", []),
                "floor_region_project_rows": floor_region_reconstruction_report.get("project_region_rows", []),
                "floor_region_reconstruct_summary": floor_region_reconstruction_report.get("summary", {}),
                "room_boundary_rows": geometry_quantity.get("room_boundary_report", {}).get("room_rows", []),
                "room_opening_candidate_rows": geometry_quantity.get("room_boundary_report", {}).get("opening_candidate_rows", []),
                "room_boundary_summary": geometry_quantity.get("room_boundary_report", {}).get("summary", {}),
                "special_quantity_trace_rows": special_quantity_report.get("special_quantity_trace_rows", []),
                "special_quantity_summary": special_quantity_report.get("summary", {}),
                "quantity_trace_rows": quantity_trace_rows,
                "line_quantity_candidate_rows": line_quantity_candidate_rows,
                "dynamic_itemization_summary": dynamic_itemization_report.get("summary", {}),
                "dynamic_itemization_stage_results": dynamic_itemization_report.get("stage_results", []),
                "dynamic_itemization_decision_rows": dynamic_itemization_report.get("itemization_decisions", []),
                "geometry_quantity_summary": geometry_quantity.get("summary", {}),
                "outputs": combined_outputs,
                "issues": issues,
            },
            ensure_ascii=False,
            indent=2,
        ),
        encoding="utf-8",
    )

    return {
        "ok": True,
        "phase": PHASE,
        "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "summary": build_item_listing_summary(
            dwg_file_count=len(dwg_files),
            dxf_file_count=len(conversion.output_files),
            text_report=text_report,
            field_report=field_report,
            match_report=match_report,
            item_rows=item_rows,
            geometry_quantity=geometry_quantity,
            project_report=project_report,
            project_geometry_binding_report=project_geometry_binding_report,
            project_region_binding_report=project_region_binding_report,
            project_material_binding_report=project_material_binding_report,
            floor_paving_locator_report=floor_paving_locator_report,
            floor_layer_rescan_report=floor_layer_rescan_report,
            floor_region_reconstruction_report=floor_region_reconstruction_report,
            special_quantity_report=special_quantity_report,
        ),
        "inputs": {
            "upload_dir": str(source_dir.resolve()),
            "oda_executable": str(executable.resolve()),
            "standard_file": str(Path(standard_file).resolve()),
        },
        "reports": {
            "conversion": conversion.as_dict(),
            "text_summary": text_report.get("summary", {}),
            "field_summary": field_report.get("summary", {}),
            "match_summary": match_report.get("summary", {}),
            "project_recognition_summary": project_report.get("summary", {}),
            "project_geometry_binding_summary": project_geometry_binding_report.get("summary", {}),
            "project_region_binding_summary": project_region_binding_report.get("summary", {}),
            "project_material_binding_summary": project_material_binding_report.get("summary", {}),
            "floor_paving_summary": floor_paving_locator_report.get("summary", {}),
            "floor_layer_rescan_summary": floor_layer_rescan_report.get("summary", {}),
            "floor_region_reconstruct_summary": floor_region_reconstruction_report.get("summary", {}),
            "room_boundary_summary": geometry_quantity.get("room_boundary_report", {}).get("summary", {}),
            "special_quantity_summary": special_quantity_report.get("summary", {}),
            "dynamic_itemization_summary": dynamic_itemization_report.get("summary", {}),
            "geometry_quantity_summary": geometry_quantity.get("summary", {}),
        },
        "project_rows": project_report.get("project_rows", []),
        "quantity_list_rows": quantity_list_rows,
        "project_recognition_summary": project_report.get("summary", {}),
        "project_geometry_binding_rows": project_geometry_binding_report.get("binding_rows", []),
        "project_geometry_candidate_rows": project_geometry_binding_report.get("candidate_rows", []),
        "project_geometry_binding_summary": project_geometry_binding_report.get("summary", {}),
        "project_region_binding_rows": project_region_binding_report.get("binding_rows", []),
        "project_region_candidate_rows": project_region_binding_report.get("candidate_rows", []),
        "project_region_binding_summary": project_region_binding_report.get("summary", {}),
        "project_material_binding_rows": project_material_binding_report.get("project_binding_rows", []),
        "project_material_index_rows": project_material_binding_report.get("material_index_rows", []),
        "project_material_inheritance_rows": project_material_binding_report.get("material_inheritance_rows", []),
        "project_material_binding_summary": project_material_binding_report.get("summary", {}),
        "floor_paving_project_rows": floor_paving_locator_report.get("floor_project_rows", []),
        "floor_paving_geometry_rows": floor_paving_locator_report.get("floor_geometry_rows", []),
        "floor_paving_material_text_rows": floor_paving_locator_report.get("floor_material_text_rows", []),
        "floor_paving_summary": floor_paving_locator_report.get("summary", {}),
        "floor_layer_segment_rows": floor_layer_rescan_report.get("floor_segment_rows", []),
        "floor_layer_package_rows": floor_layer_rescan_report.get("floor_package_rows", []),
        "floor_layer_rescan_summary": floor_layer_rescan_report.get("summary", {}),
        "floor_closed_region_rows": floor_region_reconstruction_report.get("closed_region_rows", []),
        "floor_region_project_rows": floor_region_reconstruction_report.get("project_region_rows", []),
        "floor_region_reconstruct_summary": floor_region_reconstruction_report.get("summary", {}),
        "room_boundary_rows": geometry_quantity.get("room_boundary_report", {}).get("room_rows", []),
        "room_opening_candidate_rows": geometry_quantity.get("room_boundary_report", {}).get("opening_candidate_rows", []),
        "room_boundary_summary": geometry_quantity.get("room_boundary_report", {}).get("summary", {}),
        "special_quantity_trace_rows": special_quantity_report.get("special_quantity_trace_rows", []),
        "special_quantity_summary": special_quantity_report.get("summary", {}),
        "item_rows": item_rows,
        "quantity_trace_rows": quantity_trace_rows,
        "line_quantity_candidate_rows": line_quantity_candidate_rows,
        "dynamic_itemization_summary": dynamic_itemization_report.get("summary", {}),
        "dynamic_itemization_stage_results": dynamic_itemization_report.get("stage_results", []),
        "dynamic_itemization_decision_rows": dynamic_itemization_report.get("itemization_decisions", []),
        "geometry_quantity_summary": geometry_quantity.get("summary", {}),
        "outputs": combined_outputs,
        "issues": issues,
    }


def find_oda_executable() -> Path | None:
    if DEFAULT_ODA_EXE.exists():
        return DEFAULT_ODA_EXE
    tools = probe_converter_tools(extra_search_paths=[DEFAULT_ODA_EXE.parent], include_system_tools=True)
    for tool in tools:
        if tool.tool_id == "oda_file_converter" and tool.available and tool.executable:
            return Path(tool.executable)
    return None


def _collect_dwg_files(source_dir: Path) -> list[Path]:
    seen: set[str] = set()
    files: list[Path] = []
    for path in [*sorted(source_dir.glob("*.dwg")), *sorted(source_dir.glob("*.DWG"))]:
        key = str(path.resolve()).lower()
        if key in seen:
            continue
        seen.add(key)
        files.append(path)
    return files


def build_geometry_quantity_context(
    *,
    dxf_files: list[Path],
    parsed_text_files: list[Any] | None = None,
    text_report: dict[str, Any],
    match_report: dict[str, Any],
    library: Any,
    output_dir: str | Path,
    timestamp: str,
) -> dict[str, Any]:
    try:
        parsed_geometry = [parse_dxf_geometry_file(path) for path in dxf_files]
        geometry_report = build_geometry_probe_report(parsed_geometry)
        geometry_outputs = write_geometry_probe_outputs(
            parsed_geometry,
            output_dir,
            stem=f"BIZ2x_DWG上传_CAD几何图元探测_{timestamp}",
        )
        scale_unit_report = build_scale_unit_probe_report(text_report=text_report, geometry_report=geometry_report)
        scale_unit_outputs = write_scale_unit_probe_outputs(
            scale_unit_report,
            output_dir,
            stem=f"BIZ2x_DWG上传_图框比例单位校验_{timestamp}",
        )
        manual_confirmation_report = build_manual_scale_unit_confirmation_report(
            scale_unit_report=scale_unit_report,
            geometry_report=geometry_report,
            drawing_unit="mm",
            model_space_scale="1:1",
            title_block_scale_usage="plot_scale_only_not_quantity_multiplier",
            title_block_scale_varies_by_drawing=True,
            allow_geometry_quantity_for_all_files=True,
            confirmation_note="试运行默认口径：绘图单位按 mm，模型空间按真实尺寸 1:1，标题栏比例仅作出图比例，不参与工程量换算。",
        )
        manual_confirmation_outputs = write_manual_scale_unit_confirmation_outputs(
            manual_confirmation_report,
            output_dir,
            stem=f"BIZ2x_DWG上传_比例单位确认配置_{timestamp}",
        )
        mapping_report = build_layer_block_mapping_report(
            geometry_report=geometry_report,
            confirmation_report=manual_confirmation_report,
        )
        mapping_outputs = write_layer_block_mapping_outputs(
            mapping_report,
            output_dir,
            stem=f"BIZ2x_DWG上传_低风险图层块名映射_{timestamp}",
        )
        quantity_suggestion_report = build_low_risk_quantity_suggestion_report(
            geometry_report=geometry_report,
            mapping_report=mapping_report,
        )
        quantity_suggestion_outputs = write_quantity_suggestion_outputs(
            quantity_suggestion_report,
            output_dir,
            stem=f"BIZ2x_DWG上传_低风险几何建议量_{timestamp}",
        )
        binding_report = build_standard_rule_binding_report(
            quantity_suggestion_report=quantity_suggestion_report,
            standard_match_report=match_report,
            library=library,
        )
        binding_outputs = write_standard_rule_binding_outputs(
            binding_report,
            output_dir,
            stem=f"BIZ2x_DWG上传_标准规则绑定trace_{timestamp}",
        )
        trace_review_pack = build_trace_review_pack(binding_report)
        trace_review_outputs = write_trace_review_outputs(
            trace_review_pack,
            output_dir,
            stem=f"BIZ2x_DWG上传_标准规则trace复核包_{timestamp}",
        )
        manual_confirmation = manual_confirmation_report.get("manual_confirmation") or {}
        unit_to_meter_factor = float(manual_confirmation.get("unit_to_meter_factor") or 0.001)
        unit_conversion = {
            "drawing_unit": manual_confirmation.get("drawing_unit", "mm"),
            "unit_to_meter_factor": unit_to_meter_factor,
            "area_to_square_meter_factor": unit_to_meter_factor * unit_to_meter_factor,
            "source": "manual_scale_unit_confirmation",
        }
        region_label_report = build_region_label_binding_report(
            geometry_report=geometry_report,
            parsed_text_files=parsed_text_files,
            text_report=text_report,
            unit_conversion=unit_conversion,
        )
        region_label_outputs = write_region_label_binding_outputs(
            region_label_report,
            output_dir,
            stem=f"BIZ2x_DWG上传_CAD区域文字绑定_{timestamp}",
        )
        room_boundary_report = build_room_boundary_analysis_report(
            region_label_report=region_label_report,
            geometry_report=geometry_report,
            unit_conversion=unit_conversion,
        )
        room_boundary_outputs = write_room_boundary_analysis_outputs(
            room_boundary_report,
            output_dir,
            stem=f"BIZ2x_DWG上传_房间边界净周长_{timestamp}",
        )
        return {
            "ok": True,
            "summary": {
                "geometry_entity_count": geometry_report.get("summary", {}).get("geometry_entity_count", 0),
                "geometry_area_candidate_count": geometry_report.get("summary", {}).get("area_candidate_count", 0),
                "geometry_length_candidate_count": geometry_report.get("summary", {}).get("length_candidate_count", 0),
                "geometry_count_candidate_count": geometry_report.get("summary", {}).get("count_candidate_count", 0),
                "allowed_geometry_group_count": mapping_report.get("summary", {}).get("allowed_group_count", 0),
                "geometry_suggestion_count": quantity_suggestion_report.get("summary", {}).get("suggestion_count", 0),
                "ready_geometry_suggestion_count": quantity_suggestion_report.get("summary", {}).get("ready_for_manual_review_count", 0),
                "standard_rule_trace_count": binding_report.get("summary", {}).get("standard_rule_trace_count", 0),
                "compatible_standard_rule_trace_count": binding_report.get("summary", {}).get("compatible_standard_rule_trace_count", 0),
                "trace_review_row_count": trace_review_pack.get("summary", {}).get("trace_review_row_count", 0),
                "auto_adopt_trace_count": trace_review_pack.get("summary", {}).get("auto_action_counts", {}).get("建议采用", 0),
                "region_label_candidate_count": region_label_report.get("summary", {}).get("region_candidate_count", 0),
                "region_labeled_candidate_count": region_label_report.get("summary", {}).get("labeled_region_count", 0),
                "region_room_labeled_count": region_label_report.get("summary", {}).get("room_labeled_region_count", 0),
                "region_project_labeled_count": region_label_report.get("summary", {}).get("project_labeled_region_count", 0),
                "room_boundary_count": room_boundary_report.get("summary", {}).get("room_boundary_count", 0),
                "room_opening_candidate_count": room_boundary_report.get("summary", {}).get("opening_candidate_count", 0),
                "room_with_net_perimeter_candidate_count": room_boundary_report.get("summary", {}).get("room_with_net_perimeter_candidate_count", 0),
                "room_net_perimeter_blocked_count": room_boundary_report.get("summary", {}).get("net_perimeter_blocked_count", 0),
            },
            "geometry_report": geometry_report,
            "unit_conversion": unit_conversion,
            "region_label_report": region_label_report,
            "room_boundary_report": room_boundary_report,
            "standard_rule_binding_report": binding_report,
            "trace_review_pack": trace_review_pack,
            "outputs": {
                "geometry_json": geometry_outputs["json"],
                "geometry_markdown": geometry_outputs["markdown"],
                "geometry_candidate_csv": geometry_outputs["geometry_candidate_csv"],
                "scale_unit_json": scale_unit_outputs["json"],
                "scale_unit_markdown": scale_unit_outputs["markdown"],
                "scale_unit_evidence_csv": scale_unit_outputs["evidence_csv"],
                "manual_scale_json": manual_confirmation_outputs["json"],
                "manual_scale_markdown": manual_confirmation_outputs["markdown"],
                "manual_scale_confirmation_csv": manual_confirmation_outputs["confirmation_csv"],
                "region_label_json": region_label_outputs["json"],
                "region_label_markdown": region_label_outputs["markdown"],
                "region_label_csv": region_label_outputs["region_label_csv"],
                "room_boundary_json": room_boundary_outputs["json"],
                "room_boundary_markdown": room_boundary_outputs["markdown"],
                "room_boundary_csv": room_boundary_outputs["room_csv"],
                "room_opening_candidate_csv": room_boundary_outputs["opening_csv"],
                "layer_mapping_json": mapping_outputs["json"],
                "layer_mapping_markdown": mapping_outputs["markdown"],
                "layer_mapping_csv": mapping_outputs["mapping_csv"],
                "quantity_suggestion_json": quantity_suggestion_outputs["json"],
                "quantity_suggestion_markdown": quantity_suggestion_outputs["markdown"],
                "quantity_suggestion_csv": quantity_suggestion_outputs["suggestion_csv"],
                "standard_rule_binding_json": binding_outputs["json"],
                "standard_rule_binding_markdown": binding_outputs["markdown"],
                "standard_rule_binding_csv": binding_outputs["binding_csv"],
                "standard_rule_trace_csv": binding_outputs["trace_csv"],
                "trace_review_json": trace_review_outputs["json"],
                "trace_review_markdown": trace_review_outputs["markdown"],
                "trace_review_csv": trace_review_outputs["trace_review_csv"],
                "trace_review_xlsx": trace_review_outputs["trace_review_xlsx"],
            },
            "issues": [],
        }
    except Exception as exc:
        return {
            "ok": False,
            "summary": {"geometry_quantity_status": "geometry_quantity_failed"},
            "standard_rule_binding_report": {},
            "outputs": {},
            "issues": [{"级别": "warning", "说明": f"几何建议量生成失败，已保留列项候选：{exc}"}],
        }


def build_item_listing_rows(
    match_report: dict[str, Any],
    standard_rule_binding_report: dict[str, Any] | None = None,
) -> list[dict[str, Any]]:
    candidates = list(match_report.get("standard_item_candidates") or [])
    best_by_signal: dict[str, dict[str, Any]] = {}
    for candidate in candidates:
        key = str(candidate.get("candidate_key") or "")
        current = best_by_signal.get(key)
        if current is None or float(candidate.get("match_confidence") or 0) > float(current.get("match_confidence") or 0):
            best_by_signal[key] = candidate

    trace_index = _quantity_traces_by_item_code(standard_rule_binding_report or {})
    rows: list[dict[str, Any]] = []
    for index, candidate in enumerate(best_by_signal.values(), start=1):
        code = str(candidate.get("standard_item_code") or "")
        traces = trace_index.get(code, [])
        trace_summary = _bind_quantity_trace_to_candidate(candidate, traces)
        candidate_options = build_line_quantity_candidate_options(candidate, traces)
        rows.append(
            {
                "序号": index,
                "标准项目编码": code,
                "项目名称": candidate.get("standard_item_name", ""),
                "章节": candidate.get("chapter_name", ""),
                "单位": "、".join(candidate.get("unit_options") or []),
                "匹配置信度": round(float(candidate.get("match_confidence") or 0), 2),
                "来源文件": candidate.get("source_file", ""),
                "图纸识别名称": candidate.get("source_name", ""),
                "图纸识别规格或做法": candidate.get("source_spec_or_method", ""),
                "项目特征字段": "；".join(candidate.get("feature_fields") or []),
                "工程量计算规则": candidate.get("quantity_rule_text", ""),
                "工程量状态": trace_summary["工程量状态"] or _quantity_status_label(candidate.get("quantity_evidence_status", "")),
                "逐条绑定状态": trace_summary["逐条绑定状态"],
                "绑定建议编号": trace_summary["绑定建议编号"],
                "绑定置信度": trace_summary["绑定置信度"],
                "绑定说明": trace_summary["绑定说明"],
                "系统建议工程量": trace_summary["系统建议工程量"],
                "建议单位": trace_summary["建议单位"],
                "建议量状态": trace_summary["建议量状态"],
                "标准规则Trace状态": trace_summary["标准规则Trace状态"],
                "算量证据": trace_summary["算量证据"],
                "来源证据": candidate.get("evidence_text", ""),
                "匹配理由": "；".join(candidate.get("match_reasons") or []),
                "CAD候选数量": len(candidate_options),
                "默认选择建议编号": trace_summary["绑定建议编号"] if trace_summary["逐条绑定状态"] == "已逐条绑定CAD建议量，需复核" else "",
                "CAD候选列表": candidate_options,
            }
        )
    return rows


def build_quantity_list_rows(
    project_rows: list[dict[str, Any]],
    special_quantity_report: dict[str, Any] | None = None,
) -> list[dict[str, Any]]:
    trace_by_project = {
        str(row.get("识别项目编号") or ""): row
        for row in (special_quantity_report or {}).get("special_quantity_trace_rows", [])
        if row.get("识别项目编号")
    }
    rows: list[dict[str, Any]] = []
    for project in project_rows:
        trace = trace_by_project.get(str(project.get("识别项目编号") or "")) or {}
        quantity = _display_quantity(project, trace)
        unit = trace.get("建议单位") if _has_ready_quantity(trace) else project.get("单位", "")
        rows.append(
            {
                "项目名称": project.get("项目名称", ""),
                "项目特征": project.get("项目特征", ""),
                "单位": unit or project.get("单位", ""),
                "工程量": quantity,
            }
        )
    return rows


def write_quantity_list_outputs(rows: list[dict[str, Any]], output_dir: str | Path, *, stem: str) -> dict[str, str]:
    directory = Path(output_dir)
    directory.mkdir(parents=True, exist_ok=True)
    csv_path = directory / f"{stem}.csv"
    xlsx_path = directory / f"{stem}.xlsx"
    _write_four_field_csv(csv_path, rows)
    _write_four_field_workbook(xlsx_path, rows)
    return {"csv": str(csv_path), "xlsx": str(xlsx_path)}


def _has_ready_quantity(trace: dict[str, Any]) -> bool:
    value = trace.get("建议工程量")
    if value in (None, ""):
        return False
    return str(trace.get("是否可复核") or "") == "是"


def _display_quantity(project: dict[str, Any], trace: dict[str, Any]) -> str:
    if _has_ready_quantity(trace):
        return str(trace.get("建议工程量"))
    project_quantity = project.get("工程量")
    if project_quantity not in (None, ""):
        return str(project_quantity)
    return PENDING_QUANTITY_LABEL


def build_item_listing_summary(
    *,
    dwg_file_count: int,
    dxf_file_count: int,
    text_report: dict[str, Any],
    field_report: dict[str, Any],
    match_report: dict[str, Any],
    item_rows: list[dict[str, Any]],
    geometry_quantity: dict[str, Any] | None = None,
    project_report: dict[str, Any] | None = None,
    project_geometry_binding_report: dict[str, Any] | None = None,
    project_region_binding_report: dict[str, Any] | None = None,
    project_material_binding_report: dict[str, Any] | None = None,
    floor_paving_locator_report: dict[str, Any] | None = None,
    floor_layer_rescan_report: dict[str, Any] | None = None,
    floor_region_reconstruction_report: dict[str, Any] | None = None,
    special_quantity_report: dict[str, Any] | None = None,
) -> dict[str, Any]:
    unique_codes = sorted({str(row.get("标准项目编码") or "") for row in item_rows if row.get("标准项目编码")})
    by_name: dict[str, int] = defaultdict(int)
    for row in item_rows:
        by_name[str(row.get("项目名称") or "")] += 1
    text_summary = text_report.get("summary", {})
    field_summary = field_report.get("summary", {})
    match_summary = match_report.get("summary", {})
    geometry_summary = (geometry_quantity or {}).get("summary", {})
    project_summary = (project_report or {}).get("summary", {})
    project_binding_summary = (project_geometry_binding_report or {}).get("summary", {})
    project_region_summary = (project_region_binding_report or {}).get("summary", {})
    project_material_summary = (project_material_binding_report or {}).get("summary", {})
    floor_paving_summary = (floor_paving_locator_report or {}).get("summary", {})
    floor_layer_summary = (floor_layer_rescan_report or {}).get("summary", {})
    floor_region_summary = (floor_region_reconstruction_report or {}).get("summary", {})
    special_quantity_summary = (special_quantity_report or {}).get("summary", {})
    suggested_item_count = sum(1 for row in item_rows if row.get("系统建议工程量"))
    bound_item_count = sum(1 for row in item_rows if row.get("逐条绑定状态") == "已逐条绑定CAD建议量，需复核")
    ambiguous_item_count = sum(1 for row in item_rows if row.get("逐条绑定状态") in {"多个CAD候选需选择", "有同类CAD候选，需人工选择"})
    candidate_option_count = sum(len(row.get("CAD候选列表") or []) for row in item_rows)
    candidate_item_count = sum(1 for row in item_rows if row.get("CAD候选列表"))
    return {
        "dwg_file_count": dwg_file_count,
        "dxf_file_count": dxf_file_count,
        "text_entity_count": text_summary.get("total_text_entity_count", text_summary.get("text_entity_count", 0)),
        "stored_text_record_count": text_summary.get("stored_text_record_count", 0),
        "material_method_row_count": field_summary.get("material_method_row_count", 0),
        "source_signal_count": match_summary.get("source_signal_count", 0),
        "matched_signal_count": match_summary.get("matched_signal_count", 0),
        "recognized_project_count": project_summary.get("recognized_project_count", 0),
        "project_quantity_pending_count": project_summary.get("quantity_pending_count", 0),
        "project_geometry_binding_ready_count": project_binding_summary.get("binding_ready_project_count", 0),
        "project_geometry_ambiguous_count": project_binding_summary.get("ambiguous_project_count", 0),
        "project_geometry_unbound_count": project_binding_summary.get("unbound_project_count", 0),
        "project_geometry_candidate_option_count": project_binding_summary.get("candidate_option_count", 0),
        "project_region_binding_ready_count": project_region_summary.get("binding_ready_project_count", 0),
        "project_region_ambiguous_count": project_region_summary.get("ambiguous_project_count", 0),
        "project_region_unbound_count": project_region_summary.get("unbound_project_count", 0),
        "project_region_candidate_option_count": project_region_summary.get("candidate_option_count", 0),
        "project_material_code_count": project_material_summary.get("material_code_count", 0),
        "project_with_material_code_count": project_material_summary.get("project_with_material_code_count", 0),
        "project_material_region_bound_count": project_material_summary.get("material_region_bound_project_count", 0),
        "project_material_geometry_bound_count": project_material_summary.get("material_geometry_bound_project_count", 0),
        "project_material_inheritance_candidate_count": project_material_summary.get("material_inheritance_candidate_count", 0),
        "project_material_inheritance_project_count": project_material_summary.get("material_inheritance_project_count", 0),
        "project_material_inherited_region_candidate_project_count": project_material_summary.get("material_inherited_region_candidate_project_count", 0),
        "project_material_legend_risk_candidate_count": project_material_summary.get("material_legend_risk_candidate_count", 0),
        "project_material_unbound_count": project_material_summary.get("unbound_project_count", 0),
        "floor_paving_material_project_count": floor_paving_summary.get("floor_material_project_count", 0),
        "floor_paving_material_text_evidence_count": floor_paving_summary.get("floor_material_text_evidence_count", 0),
        "floor_paving_layer_area_candidate_count": floor_paving_summary.get("floor_layer_area_candidate_count", 0),
        "floor_paving_effective_area_candidate_count": floor_paving_summary.get("effective_floor_area_candidate_count", 0),
        "floor_paving_bound_candidate_count": floor_paving_summary.get("floor_project_bound_candidate_count", 0),
        "floor_paving_sample_missing_count": floor_paving_summary.get("floor_project_sample_missing_count", 0),
        "floor_layer_segment_count": floor_layer_summary.get("floor_segment_count", 0),
        "floor_layer_package_count": floor_layer_summary.get("floor_package_count", 0),
        "floor_layer_ready_package_count": floor_layer_summary.get("ready_floor_package_count", 0),
        "floor_layer_small_or_layout_package_count": floor_layer_summary.get("small_or_layout_floor_package_count", 0),
        "floor_layer_unbound_package_count": floor_layer_summary.get("unbound_floor_package_count", 0),
        "floor_closed_region_candidate_count": floor_region_summary.get("closed_region_candidate_count", 0),
        "floor_ready_closed_region_candidate_count": floor_region_summary.get("ready_closed_region_candidate_count", 0),
        "floor_region_project_ready_count": floor_region_summary.get("project_ready_closed_region_count", 0),
        "floor_region_project_blocked_count": floor_region_summary.get("project_blocked_count", 0),
        "floor_region_project_blocked_no_closed_region_count": floor_region_summary.get("project_blocked_no_closed_region_count", 0),
        "floor_region_project_blocked_small_region_count": floor_region_summary.get("project_blocked_small_region_count", 0),
        "special_quantity_trace_count": special_quantity_summary.get("special_quantity_trace_count", 0),
        "special_quantity_ready_for_review_count": special_quantity_summary.get("ready_for_manual_review_count", 0),
        "special_quantity_blocked_trace_count": special_quantity_summary.get("blocked_trace_count", 0),
        "item_row_count": len(item_rows),
        "unique_standard_item_count": len(unique_codes),
        "unique_standard_item_codes": unique_codes,
        "item_name_counts": dict(sorted(by_name.items())),
        "suggested_quantity_item_row_count": suggested_item_count,
        "line_bound_quantity_item_row_count": bound_item_count,
        "ambiguous_quantity_item_row_count": ambiguous_item_count,
        "line_quantity_candidate_item_row_count": candidate_item_count,
        "line_quantity_candidate_option_count": candidate_option_count,
        **geometry_summary,
        "final_generation_status": _final_generation_status(
            bound_item_count=bound_item_count,
            ambiguous_item_count=ambiguous_item_count,
        ),
    }


def build_item_listing_issues(
    conversion: dict[str, Any],
    match_report: dict[str, Any],
    item_rows: list[dict[str, Any]],
    geometry_quantity: dict[str, Any] | None = None,
) -> list[dict[str, Any]]:
    issues: list[dict[str, Any]] = []
    if conversion.get("status") == "partial_converted":
        issues.append({"级别": "warning", "说明": "部分 DWG 未能转成 DXF，请检查转换报告。"})
    if not item_rows:
        issues.append({"级别": "error", "说明": "未识别出可匹配标准库的项目，请检查图纸文字、材料表或构造做法表。"})
    source_count = int((match_report.get("summary") or {}).get("source_signal_count") or 0)
    matched_count = int((match_report.get("summary") or {}).get("matched_signal_count") or 0)
    if source_count and matched_count < source_count:
        issues.append(
            {
                "级别": "info",
                "说明": f"有 {source_count - matched_count} 条图纸线索未匹配到标准项目，建议查看标准匹配明细。",
            }
        )
    if geometry_quantity and not geometry_quantity.get("ok"):
        issues.extend(list(geometry_quantity.get("issues") or []))
    elif geometry_quantity:
        summary = geometry_quantity.get("summary") or {}
        ready_count = int(summary.get("compatible_standard_rule_trace_count") or 0)
        if ready_count:
            issues.append({"级别": "info", "说明": f"已生成 {ready_count} 条可复核标准规则工程量 trace，需业务复核后才能进入最终清单。"})
        else:
            issues.append({"级别": "info", "说明": "本次未形成可复核标准规则工程量 trace，列项仍需几何算量增强或人工补量。"})
        bound_count = sum(1 for row in item_rows if row.get("逐条绑定状态") == "已逐条绑定CAD建议量，需复核")
        ambiguous_count = sum(1 for row in item_rows if row.get("逐条绑定状态") in {"多个CAD候选需选择", "有同类CAD候选，需人工选择"})
        if ambiguous_count:
            issues.append({"级别": "info", "说明": f"逐条绑定初判：{bound_count} 行可直接绑定，{ambiguous_count} 行需要在候选 CAD 建议量中人工选择。"})
    return issues


def write_item_listing_outputs(report: dict[str, Any], output_dir: str | Path, *, stem: str) -> dict[str, str]:
    directory = Path(output_dir)
    directory.mkdir(parents=True, exist_ok=True)
    json_path = directory / f"{stem}.json"
    csv_path = directory / f"{stem}_列项候选.csv"
    xlsx_path = directory / f"{stem}_列项候选.xlsx"
    md_path = directory / f"{stem}.md"

    json_path.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    _write_item_csv(csv_path, list(report.get("item_rows") or []))
    write_item_listing_workbook(list(report.get("item_rows") or []), xlsx_path)
    md_path.write_text(build_item_listing_markdown(report), encoding="utf-8")
    return {"json": str(json_path), "csv": str(csv_path), "xlsx": str(xlsx_path), "markdown": str(md_path)}


def write_item_listing_workbook(rows: list[dict[str, Any]], path: str | Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "DWG列项候选"
    sheet.append(ITEM_LIST_HEADERS)
    for row in rows:
        sheet.append([row.get(header, "") for header in ITEM_LIST_HEADERS])
    _style_sheet(sheet)
    workbook.save(path)


def _write_four_field_csv(path: Path, rows: list[dict[str, Any]]) -> None:
    with path.open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.DictWriter(handle, fieldnames=QUANTITY_LIST_HEADERS)
        writer.writeheader()
        for row in rows:
            writer.writerow({header: row.get(header, "") for header in QUANTITY_LIST_HEADERS})


def _write_four_field_workbook(path: Path, rows: list[dict[str, Any]]) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "识图四字段清单"
    sheet.append(QUANTITY_LIST_HEADERS)
    for row in rows:
        sheet.append([row.get(header, "") for header in QUANTITY_LIST_HEADERS])
    _style_sheet(sheet)
    workbook.save(path)


def build_item_listing_markdown(report: dict[str, Any]) -> str:
    summary = report.get("summary", {})
    lines = [
        "# DWG 上传列项结果",
        "",
        f"- 生成时间：{report.get('generated_at', '-')}",
        f"- DWG 文件数：{summary.get('dwg_file_count', 0)}",
        f"- DXF 文件数：{summary.get('dxf_file_count', 0)}",
        f"- 图纸文字实体：{summary.get('text_entity_count', 0)}",
        f"- 图纸线索：{summary.get('source_signal_count', 0)}",
        f"- 已匹配线索：{summary.get('matched_signal_count', 0)}",
        f"- 列项候选：{summary.get('item_row_count', 0)}",
        f"- 唯一标准项目：{summary.get('unique_standard_item_count', 0)}",
        f"- 已逐条绑定建议量的列项：{summary.get('line_bound_quantity_item_row_count', 0)}",
        f"- 需人工选择 CAD 候选的列项：{summary.get('ambiguous_quantity_item_row_count', 0)}",
        f"- 可复核标准规则 trace：{summary.get('compatible_standard_rule_trace_count', 0)}",
        "",
        "## 列项候选",
        "",
        "| 序号 | 标准项目编码 | 项目名称 | 单位 | 置信度 | 绑定状态 | 建议工程量 | 工程量状态 | 来源 |",
        "| ---: | --- | --- | --- | ---: | --- | --- | --- | --- |",
    ]
    for row in list(report.get("item_rows") or [])[:100]:
        lines.append(
            "| "
            + " | ".join(
                [
                    _md(row.get("序号", "")),
                    _md(row.get("标准项目编码", "")),
                    _md(row.get("项目名称", "")),
                    _md(row.get("单位", "")),
                    _md(row.get("匹配置信度", "")),
                    _md(row.get("逐条绑定状态", "")),
                    _md(row.get("系统建议工程量", "")),
                    _md(row.get("工程量状态", "")),
                    _md(row.get("图纸识别名称", "")),
                ]
            )
            + " |"
        )
    return "\n".join(lines) + "\n"


def copy_uploaded_dwg(source: Path, target_dir: Path, index: int) -> Path:
    target_dir.mkdir(parents=True, exist_ok=True)
    suffix = source.suffix.lower()
    if suffix != ".dwg":
        raise DwgItemListingError("仅支持 .dwg 图纸文件")
    target = target_dir / f"{index:02d}_{_safe_filename(source.stem)}.dwg"
    shutil.copyfile(source, target)
    return target


def _write_item_csv(path: Path, rows: list[dict[str, Any]]) -> None:
    with path.open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.DictWriter(handle, fieldnames=ITEM_LIST_HEADERS)
        writer.writeheader()
        for row in rows:
            writer.writerow({header: row.get(header, "") for header in ITEM_LIST_HEADERS})


def _style_sheet(sheet) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    widths = {
        "A": 8,
        "B": 16,
        "C": 24,
        "D": 22,
        "E": 12,
        "F": 12,
        "G": 24,
        "H": 24,
        "I": 42,
        "J": 44,
        "K": 46,
        "L": 24,
        "M": 24,
        "N": 18,
        "O": 12,
        "P": 42,
        "Q": 20,
        "R": 12,
        "S": 30,
        "T": 32,
        "U": 60,
        "V": 60,
    }
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:{get_column_letter(sheet.max_column)}{sheet.max_row}"


def _quantity_status_label(value: Any) -> str:
    text = str(value or "")
    if text == "missing_quantity_measurement_needs_manual_review":
        return "待几何算量或人工补量"
    return text or "待确认"


def _final_generation_status(*, bound_item_count: int, ambiguous_item_count: int) -> str:
    if bound_item_count:
        return "line_item_quantity_bound_review_required"
    if ambiguous_item_count:
        return "line_item_quantity_candidates_need_selection"
    return "item_listing_ready_quantity_pending"


def build_quantity_trace_rows(standard_rule_binding_report: dict[str, Any] | None) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for trace in (standard_rule_binding_report or {}).get("standard_rule_traces", []):
        source_trace = (trace.get("calculation_trace") or {}).get("source_calculation_trace") or {}
        rows.append(
            {
                "建议编号": trace.get("suggestion_key", ""),
                "标准项目编码": trace.get("item_code", ""),
                "标准项目名称": trace.get("item_name", ""),
                "trace状态": trace.get("trace_status", ""),
                "是否可复核": "是" if trace.get("ready_for_manual_review") else "否",
                "建议工程量": trace.get("standard_rule_suggested_quantity") or "",
                "建议单位": trace.get("suggested_unit", ""),
                "几何建议量": trace.get("geometry_quantity", ""),
                "几何单位": trace.get("geometry_unit", ""),
                "工程量计算规则": trace.get("quantity_rule_text", ""),
                "CAD公式": source_trace.get("formula", "") or (trace.get("calculation_trace") or {}).get("geometry_formula", ""),
                "CAD来源图元行号": "、".join(str(item) for item in source_trace.get("sample_line_numbers") or []),
                "未解决事项": "；".join(trace.get("unresolved_requirements") or []),
                "阻断原因": trace.get("block_reason", ""),
            }
        )
    return rows


def build_line_quantity_candidate_rows(item_rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for item in item_rows:
        for option in item.get("CAD候选列表") or []:
            rows.append(
                {
                    "列项序号": item.get("序号", ""),
                    "列项标准项目编码": item.get("标准项目编码", ""),
                    "列项项目名称": item.get("项目名称", ""),
                    "列项图纸线索": item.get("图纸识别名称", ""),
                    "列项规格或做法": item.get("图纸识别规格或做法", ""),
                    **option,
                }
            )
    return rows


def build_line_quantity_candidate_options(candidate: dict[str, Any], traces: list[dict[str, Any]]) -> list[dict[str, Any]]:
    scored_options = []
    for trace in traces:
        scored = _score_trace_for_candidate(candidate, trace)
        scored_options.append(scored)
    scored_options.sort(
        key=lambda item: (
            bool(item["trace"].get("ready_for_manual_review") and item["trace"].get("standard_rule_suggested_quantity")),
            item["score"],
        ),
        reverse=True,
    )
    best_score = scored_options[0]["score"] if scored_options else 0
    options: list[dict[str, Any]] = []
    for item in scored_options[:8]:
        trace = item["trace"]
        options.append(_line_quantity_candidate_option(candidate, trace, item["score"], item["reasons"], best_score))
    return options


def _line_quantity_candidate_option(
    candidate: dict[str, Any],
    trace: dict[str, Any],
    score: float,
    reasons: list[str],
    best_score: float,
) -> dict[str, Any]:
    source_trace = (trace.get("calculation_trace") or {}).get("source_calculation_trace") or {}
    trace_parts = _trace_source_parts(trace)
    ready = bool(trace.get("ready_for_manual_review") and trace.get("standard_rule_suggested_quantity"))
    if not ready:
        action = "不建议采纳"
        action_reason = "该 CAD trace 暂未满足标准规则可复核条件"
    elif score >= 6 and score >= best_score - 1.5:
        action = "建议优先核验"
        action_reason = "候选与列项来源线索较接近，仍需人工确认是否对应本行施工范围"
    else:
        action = "需人工选择"
        action_reason = "同标准项目下存在可复核 CAD 候选，但列项与 CAD 范围对应关系不足以自动绑定"
    return {
        "建议编号": str(trace.get("suggestion_key") or ""),
        "标准项目编码": str(trace.get("item_code") or candidate.get("standard_item_code") or ""),
        "标准项目名称": str(trace.get("item_name") or candidate.get("standard_item_name") or ""),
        "建议工程量": trace.get("standard_rule_suggested_quantity") or "",
        "建议单位": str(trace.get("suggested_unit") or ""),
        "trace状态": str(trace.get("trace_status") or ""),
        "是否可复核": "是" if trace.get("ready_for_manual_review") else "否",
        "绑定评分": score,
        "绑定置信度": _binding_confidence_label(score),
        "推荐动作": action,
        "推荐原因": "；".join(reasons) or action_reason,
        "推荐说明": action_reason,
        "CAD公式": source_trace.get("formula", "") or (trace.get("calculation_trace") or {}).get("geometry_formula", ""),
        "CAD来源图元行号": "、".join(str(item) for item in source_trace.get("sample_line_numbers") or []),
        "CAD来源": " / ".join(part for part in [trace_parts["source_file"], trace_parts["candidate_type"], trace_parts["layer"], trace_parts["block_name"]] if part),
        "未解决事项": "；".join(trace.get("unresolved_requirements") or []),
        "阻断原因": str(trace.get("block_reason") or ""),
        "算量证据": _trace_evidence(trace),
    }


def _quantity_traces_by_item_code(standard_rule_binding_report: dict[str, Any]) -> dict[str, list[dict[str, Any]]]:
    indexed: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for trace in standard_rule_binding_report.get("standard_rule_traces", []):
        code = str(trace.get("item_code") or "")
        if code:
            indexed[code].append(trace)
    return indexed


def _bind_quantity_trace_to_candidate(candidate: dict[str, Any], traces: list[dict[str, Any]]) -> dict[str, str]:
    if not traces:
        return {
            "工程量状态": "",
            "逐条绑定状态": "未找到CAD标准规则建议量",
            "绑定建议编号": "",
            "绑定置信度": "",
            "绑定说明": "当前列项没有可关联的 CAD 几何建议量或标准规则 trace",
            "系统建议工程量": "",
            "建议单位": "",
            "建议量状态": "暂无CAD标准规则建议量",
            "标准规则Trace状态": "",
            "算量证据": "",
        }
    ready = [trace for trace in traces if trace.get("ready_for_manual_review") and trace.get("standard_rule_suggested_quantity")]
    status_counts: dict[str, int] = defaultdict(int)
    for trace in traces:
        status_counts[str(trace.get("trace_status") or "unknown")] += 1
    if ready:
        scored = sorted((_score_trace_for_candidate(candidate, trace) for trace in ready), key=lambda item: item["score"], reverse=True)
        best = scored[0]
        close_count = sum(1 for item in scored if item["score"] >= max(0, best["score"] - 1.5))
        if best["score"] >= 6 and close_count == 1:
            trace = best["trace"]
            return {
                "工程量状态": "有CAD几何建议量，需复核",
                "逐条绑定状态": "已逐条绑定CAD建议量，需复核",
                "绑定建议编号": str(trace.get("suggestion_key") or ""),
                "绑定置信度": _binding_confidence_label(best["score"]),
                "绑定说明": "；".join(best["reasons"]),
                "系统建议工程量": _format_quantity(trace.get("standard_rule_suggested_quantity"), trace.get("suggested_unit")),
                "建议单位": str(trace.get("suggested_unit") or ""),
                "建议量状态": "已按来源线索绑定1条可复核建议量，未作为最终工程量",
                "标准规则Trace状态": str(trace.get("trace_status") or ""),
                "算量证据": _trace_evidence(trace),
            }
        top_quantities = [_format_quantity(item["trace"].get("standard_rule_suggested_quantity"), item["trace"].get("suggested_unit")) for item in scored[:5]]
        top_keys = [str(item["trace"].get("suggestion_key") or "") for item in scored[:5]]
        status = "多个CAD候选需选择" if len(ready) > 1 else "有同类CAD候选，需人工选择"
        note = "；".join(best["reasons"]) if best["reasons"] else "只有同标准项目候选，缺少足够的图纸线索/图层线索证明逐条对应"
        return {
            "工程量状态": "有CAD候选量，需选择",
            "逐条绑定状态": status,
            "绑定建议编号": "；".join(top_keys),
            "绑定置信度": _binding_confidence_label(best["score"]),
            "绑定说明": f"{note}；同标准项目共有{len(ready)}条可复核CAD候选",
            "系统建议工程量": "",
            "建议单位": "",
            "建议量状态": f"候选建议量：{'；'.join(top_quantities)}；未逐条绑定，需人工选择",
            "标准规则Trace状态": "；".join(f"{key}:{value}" for key, value in sorted(status_counts.items())),
            "算量证据": "；".join(_trace_evidence(item["trace"]) for item in scored[:3] if _trace_evidence(item["trace"])),
        }
    return {
        "工程量状态": "有CAD trace但暂不可采用",
        "逐条绑定状态": "有CAD trace但规则不匹配或阻断",
        "绑定建议编号": "",
        "绑定置信度": "",
        "绑定说明": "存在同标准项目 trace，但规则类型、证据或阻断条件不满足可复核建议量要求",
        "系统建议工程量": "",
        "建议单位": "",
        "建议量状态": "存在标准规则trace，但需人工选择项目/补证据/处理阻断",
        "标准规则Trace状态": "；".join(f"{key}:{value}" for key, value in sorted(status_counts.items())),
        "算量证据": "；".join(_trace_evidence(trace) for trace in traces[:3] if _trace_evidence(trace)),
    }


def _format_quantity(quantity: Any, unit: Any) -> str:
    if quantity in {"", None}:
        return ""
    return f"{quantity}{unit or ''}"


def _score_trace_for_candidate(candidate: dict[str, Any], trace: dict[str, Any]) -> dict[str, Any]:
    score = 0.0
    reasons: list[str] = []
    candidate_text = _normalize_binding_text(
        " ".join(
            [
                str(candidate.get("standard_item_name") or ""),
                str(candidate.get("source_name") or ""),
                str(candidate.get("source_spec_or_method") or ""),
                str(candidate.get("evidence_text") or ""),
                " ".join(candidate.get("match_reasons") or []),
            ]
        )
    )
    trace_parts = _trace_source_parts(trace)
    trace_text = _normalize_binding_text(
        " ".join(
            [
                str(trace.get("item_name") or ""),
                trace_parts["source_file"],
                trace_parts["candidate_type"],
                trace_parts["layer"],
                trace_parts["block_name"],
                trace_parts["business_hint"],
                trace_parts["matched_reason"],
            ]
        )
    )

    if _same_base_file(str(candidate.get("source_file") or ""), trace_parts["source_file"]):
        score += 3.0
        reasons.append("同一DXF来源文件")

    shared_terms = [term for term in _BINDING_KEY_TERMS if term in candidate_text and term in trace_text]
    if shared_terms:
        score += min(6.0, len(shared_terms) * 1.5)
        reasons.append("关键词一致：" + "、".join(shared_terms[:4]))

    category_terms = _category_terms_for_standard_code(str(candidate.get("standard_item_code") or ""))
    matched_category_terms = [term for term in category_terms if term in trace_text]
    if matched_category_terms:
        score += 2.0
        reasons.append("CAD图层/业务提示符合标准项目类型：" + "、".join(matched_category_terms[:3]))

    if str(trace.get("quantity_formula_type") or "") and str(trace.get("quantity_formula_type") or "") in str(candidate.get("quantity_rule_text") or ""):
        score += 0.5

    risk_terms = _risk_terms_for_line_binding(candidate_text, trace_parts, trace)
    if risk_terms:
        score -= sum(value for _, value in risk_terms)
        reasons.extend(reason for reason, _ in risk_terms)

    return {"trace": trace, "score": round(score, 2), "reasons": reasons}


_BINDING_KEY_TERMS = (
    "地面",
    "楼地面",
    "地台",
    "地砖",
    "玻化砖",
    "防水",
    "聚氨酯",
    "吊顶",
    "天棚",
    "顶面",
    "天花",
    "造型",
    "石膏板",
    "龙骨",
    "软膜",
    "窗帘",
    "窗帘盒",
    "踢脚",
    "窗台",
    "线脚",
    "线条",
)


def _category_terms_for_standard_code(code: str) -> tuple[str, ...]:
    if code in {"011102003", "010904002"}:
        return ("地面", "楼地面", "地台", "防水")
    if code in {"011302001", "011302003", "011404002"}:
        return ("吊顶", "天棚", "顶面", "天花", "造型")
    if code == "011105006":
        return ("踢脚",)
    if code == "010810002":
        return ("窗帘", "窗帘盒")
    if code == "010809001":
        return ("窗台",)
    return ()


def _risk_terms_for_line_binding(candidate_text: str, trace_parts: dict[str, str], trace: dict[str, Any]) -> list[tuple[str, float]]:
    risks: list[tuple[str, float]] = []
    if any(term in candidate_text for term in ("材料说明", "说明", "详图", "节点", "大样")):
        risks.append(("列项来源偏说明/节点，需防止把做法说明误绑定为施工区域", 2.0))
    source_file = trace_parts["source_file"]
    if any(term in source_file for term in ("通用节点", "节点", "大样")):
        risks.append(("CAD候选来源为节点/大样图，需人工确认是否为真实施工区域", 2.0))
    quantity = _float_or_none(trace.get("standard_rule_suggested_quantity"))
    if quantity is not None and quantity < 1:
        risks.append(("CAD建议量小于1，可能是节点/局部构造候选", 1.5))
    return risks


def _trace_source_parts(trace: dict[str, Any]) -> dict[str, str]:
    calculation_trace = trace.get("calculation_trace") or {}
    source_trace = calculation_trace.get("source_calculation_trace") or {}
    source_key = str(calculation_trace.get("geometry_source_key") or source_trace.get("source_key") or "")
    parts = source_key.split("|")
    return {
        "source_key": source_key,
        "source_file": parts[0] if len(parts) > 0 else "",
        "candidate_type": parts[1] if len(parts) > 1 else "",
        "layer": parts[2] if len(parts) > 2 else "",
        "block_name": parts[3] if len(parts) > 3 else "",
        "business_hint": str(source_trace.get("mapping_business_hint") or ""),
        "matched_reason": str(source_trace.get("matched_reason") or ""),
    }


def _normalize_binding_text(value: str) -> str:
    return value.replace(" ", "").replace("|", "").replace("\n", "")


def _same_base_file(left: str, right: str) -> bool:
    if not left or not right:
        return False
    return Path(left).stem == Path(right).stem


def _binding_confidence_label(score: float) -> str:
    if score >= 8:
        return f"高({score})"
    if score >= 5:
        return f"中({score})"
    if score > 0:
        return f"低({score})"
    return f"不足({score})"


def _trace_evidence(trace: dict[str, Any]) -> str:
    source_trace = (trace.get("calculation_trace") or {}).get("source_calculation_trace") or {}
    formula = source_trace.get("formula", "") or (trace.get("calculation_trace") or {}).get("geometry_formula", "")
    line_numbers = "、".join(str(item) for item in source_trace.get("sample_line_numbers") or [])
    parts = []
    if formula:
        parts.append(f"公式：{formula}")
    if line_numbers:
        parts.append(f"CAD行号：{line_numbers}")
    return "，".join(parts)


def _float_or_none(value: Any) -> float | None:
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _safe_filename(value: str) -> str:
    cleaned = "".join(ch if ch.isalnum() or ch in "-_." or "\u4e00" <= ch <= "\u9fff" else "_" for ch in value)
    return cleaned.strip("._")[:80] or "drawing"


def _md(value: Any) -> str:
    return str(value).replace("|", "\\|").replace("\n", "<br>")
